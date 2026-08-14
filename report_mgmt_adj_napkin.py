"""
Management Adjustment Worksheet (the "Mgmt Adj Napkin" report).

Standalone Excel workbook that mirrors the master template's
``Mgmt Adj Napkin`` tab. Each pool gets a 15-row block with a Total
row whose **Management Adj** cell (column I) is user-editable; the
Adjusted-Loss-Rate / Adjusted-Allowance / Variance columns recompute
in-cell so the analyst can experiment with management adjustments
and see the impact relative to the baseline (column H).

Public entry point:

    compose_mgmt_adj_napkin(client_name, snapshot_date, df, config,
                            grades, hist) -> (wb, fname)

This module follows the same composer signature used by
``report_tct.compose_tct`` and ``report_vizo.compose_vizo_main`` so it
can be dispatched from ``generate_report.generate_report`` without any
new data-loading code.
"""
from __future__ import annotations

from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Reuse helpers from report_tct so baseline math matches the standard
# reports exactly (life-loss, distribution factors, env-factor scoring,
# mgmt-adj resolution).
from report_tct import (
    DIST_FACTORS,
    DQ_RANGES,
    ES_RANGES,
    NCC_RANGES,
    _all_grades,
    _brr_grade_labels,
    _brr_pools_set,
    _build_pool_use_default_map,
    _dist_factor,
    _econ_stress,
    _is_brr_pool,
    _is_hidden,
    _load_admin_default_mgmt_adj,
    _merge_pool_orders,
    _ncc,
    _other_allowance_considerations,
    _resolve_mgmt_adj_total,
    _sort_pools,
)


def _env_score(value, ranges):
    """Range-table score (matches the nested helper in
    ``report_tct._sheet_env_factor`` / ``_sheet_acl_reserve`` — value
    is already in 0-100 percentage units, no auto-multiplication)."""
    for lo, hi, s in ranges:
        if lo <= value < hi:
            return s
    return 0


# ── Styling ────────────────────────────────────────────────────────
FNT_TITLE = Font(name='Arial', bold=True, size=14)
FNT_SUB   = Font(name='Arial', bold=True, size=12)
FNT_HDR   = Font(name='Arial', bold=True, size=11, color='FFFFFF')
FNT_LBL   = Font(name='Arial', bold=True, size=10)
FNT_FORMULA_LBL = Font(name='Arial', italic=True, size=9, color='555555')
FNT_BOLD  = Font(name='Arial', bold=True, size=10)
FNT_NORM  = Font(name='Arial', size=10)
FNT_INPUT = Font(name='Arial', bold=True, size=10, color='1B4F72')
FNT_NOTE  = Font(name='Arial', italic=True, size=9, color='555555')

FILL_HDR    = PatternFill('solid', fgColor='2C3E50')
FILL_POOL   = PatternFill('solid', fgColor='D6EAF8')
FILL_TOTAL  = PatternFill('solid', fgColor='FCF3CF')
FILL_INPUT  = PatternFill('solid', fgColor='FFF8DC')   # editable cell
FILL_BASE   = PatternFill('solid', fgColor='EAF2F8')   # baseline (read-only)
FILL_ADJ    = PatternFill('solid', fgColor='E8F8F5')   # adjusted (formula)
FILL_VAR    = PatternFill('solid', fgColor='FDEDEC')   # variance highlight

THIN = Border(
    left=Side('thin', color='BDC3C7'),
    right=Side('thin', color='BDC3C7'),
    top=Side('thin', color='BDC3C7'),
    bottom=Side('thin', color='BDC3C7'),
)

ACCT  = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
ACCT2 = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
PCT   = '0.00%'
PCT4  = '0.0000%'

# ── Block layout (constants) ──────────────────────────────────────
# Each pool block is 15 rows: [letter row, formula-label row,
# header row, 11 grade rows, total row].
BLOCK_LETTER_OFFSET   = 0
BLOCK_FORMULA_OFFSET  = 1
BLOCK_HEADER_OFFSET   = 2
BLOCK_FIRST_GRADE_OFF = 3
BLOCK_NUM_GRADES      = 11
BLOCK_TOTAL_OFFSET    = 14   # 3 + 11
BLOCK_HEIGHT          = 15
BLOCK_GAP             = 1    # blank row between blocks

# Column letter labels exposed in the block (A-O = 1-15).
COL_LABELS_AO = list("ABCDEFGHIJKLMNO")
HEADER_TEXTS = {
    'A': 'Pool / Grade',
    'B': 'Loan Loss Calculation Balance',
    'C': 'Aggregated Pool Loss Rate',
    'D': 'Distribution Factor',
    'E': 'ACL Base Loss Rate',
    'F': 'Allowance before Environmental',
    'G': 'Base Environmental',
    'H': 'Base Total ACL Allowance',
    'I': 'Management Adj',
    'J': 'Distributed Management Adjustment',
    'K': 'Adjusted Loss Rate',
    'L': 'Adjusted Balance Before Environmental',
    'M': 'Adjusted Environmental',
    'N': 'Adjusted Total ACL Allowance',
    'O': 'Variance from Base',
}
FORMULA_LABELS = {
    'E': 'C*D',
    'F': 'B*E',
    'J': 'D*I',
    'K': 'E+J',
    'L': 'B*K',
    'N': 'L+M',
    'O': 'N-H',
}


def _snap_display(snap: str) -> str:
    """Render '2025-12-31' as 'December 31, 2025'."""
    try:
        return datetime.strptime(snap, '%Y-%m-%d').strftime('%B %d, %Y')
    except (ValueError, TypeError):
        return str(snap)


# ── Per-pool baseline computation ─────────────────────────────────
def _pool_life_loss_acl(pools, hist, config):
    """Match _sheet_acl_reserve's life_loss computation (annual-grade-avg
    style with per-pool ACL months window). Falls back to avg_balances
    when WARM annual grade data is absent.
    """
    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    avg_bals = hist.get('avg_balances', {}) if hist else {}
    years = hist.get('years', []) if hist else []
    imp = hist.get('impaired', {}) if hist else {}
    acl_months_map = imp.get('acl_months', {})
    warm_net_co = imp.get('warm_net_co', {})
    hbd = imp.get('hist_bal_data', {})

    snap_year = int(config.get('_snap_year', 0)) or None
    snap_month = int(config.get('_snap_month', 0)) or None

    annual_grade_avg = {}
    for pk, pdata in hbd.items():
        dates = pdata.get('dates', [])
        grades_data = pdata.get('grades', {})
        annual_grade_avg[pk] = {}
        for gk, vals in grades_data.items():
            if gk.upper().startswith('HIDE'):
                continue
            yr_sums, yr_cnts = {}, {}
            for i, d in enumerate(dates):
                if i < len(vals) and vals[i] > 0:
                    yr_sums[d.year] = yr_sums.get(d.year, 0) + vals[i]
                    yr_cnts[d.year] = yr_cnts.get(d.year, 0) + 1
            for y in yr_sums:
                annual_grade_avg[pk].setdefault(y, {})
                annual_grade_avg[pk][y][gk] = yr_sums[y] / yr_cnts[y]

    out = {}
    for pool in pools:
        pool_acl = acl_months_map.get(pool, 36)
        if snap_year and snap_month:
            abs_first = (snap_year * 12 + snap_month) - pool_acl + 1
            pe = (abs_first - 1) // 12
        else:
            pe = 0
        pa = annual_grade_avg.get(pool, {})
        yr_tots = []
        for y in years:
            if y < pe:
                continue
            yt = sum(pa.get(y, {}).values())
            if not yt:
                yt = avg_bals.get(y, {}).get(pool, 0)
            if yt:
                yr_tots.append(yt)
        avg_tot = sum(yr_tots) / len(yr_tots) if yr_tots else 0
        pool_stripped = pool.strip()
        net_match = warm_net_co.get(pool_stripped, warm_net_co.get(pool, None))
        if net_match is not None:
            total_net = net_match
        else:
            total_net = 0
            for y in years:
                if y < pe:
                    continue
                total_net += abs(co_data.get(y, {}).get(pool, 0) or 0) \
                             - abs(rc_data.get(y, {}).get(pool, 0) or 0)
        out[pool] = total_net / avg_tot if avg_tot > 0 else 0
    return out


def _compute_env_factor(pool, pdf, grades, config, hist, dq_var_map,
                        econ_stress, ncc_r, dq_r, es_r, is_rr):
    """Mirror of the env-factor math in report_tct._sheet_env_factor."""
    if is_rr and not pdf.empty:
        try:
            _, _, ncc_pct = _ncc(pdf, grades, config)
        except Exception:
            ncc_pct = 0.0
    else:
        ncc_pct = 0.0
    dq_v = dq_var_map.get(pool, 0) or 0
    ncc_score = _env_score(ncc_pct * 100, ncc_r) / 100.0
    dq_score = _env_score(dq_v * 100, dq_r) / 100.0
    es_score = _env_score(econ_stress, es_r) / 100.0
    return ncc_score + dq_score + es_score


def _compute_dq_var(pools, hist, config, snap_year, snap_month):
    """Same DQ variance math as _sheet_acl_reserve."""
    dq_pct = hist.get('dq_pct', {}) if hist else {}
    if not dq_pct:
        dq_pct = (hist.get('impaired') or {}).get('warm_dq_pct', {})
    dq_years = sorted(dq_pct.keys())
    imp = hist.get('impaired', {}) if hist else {}
    acl_months_map = imp.get('acl_months', {})
    out = {}
    for pool in pools:
        pool_acl = acl_months_map.get(pool, 36)
        abs_first = (snap_year * 12 + snap_month) - pool_acl + 1
        earliest = (abs_first - 1) // 12
        rates = [dq_pct.get(y, {}).get(pool, 0) for y in dq_years if y >= earliest]
        if len(rates) >= 2:
            avg = sum(rates) / len(rates)
            out[pool] = rates[-1] - avg
        else:
            out[pool] = 0
    return out


# ── Block writers ─────────────────────────────────────────────────
def _write_block_letter_row(ws, row):
    """Row 1 of block: 'A','B','C',...,'O'."""
    for ci, letter in enumerate(COL_LABELS_AO, start=1):
        c = ws.cell(row=row, column=ci, value=letter)
        c.font = FNT_LBL
        c.alignment = Alignment(horizontal='center')
        c.fill = FILL_HDR
        c.font = FNT_HDR
        c.border = THIN


def _write_block_formula_row(ws, row):
    """Row 2 of block: formula labels under E,F,J,K,L,N,O."""
    for letter, lbl in FORMULA_LABELS.items():
        ci = ord(letter) - ord('A') + 1
        c = ws.cell(row=row, column=ci, value=lbl)
        c.font = FNT_FORMULA_LBL
        c.alignment = Alignment(horizontal='center')


def _write_block_header_row(ws, row, pool_name, rr_flag):
    """Row 3 of block: pool name in A; column header text in B-O.

    Retained for backwards-compat / single-block usage; the main composer
    now emits column headers once globally and a single pool-name row per
    block via :func:`_write_pool_name_row`.
    """
    a = ws.cell(row=row, column=1, value=pool_name)
    a.font = FNT_TITLE
    a.fill = FILL_POOL
    a.alignment = Alignment(horizontal='left', vertical='center')
    a.border = THIN
    for letter, text in HEADER_TEXTS.items():
        if letter == 'A':
            continue
        ci = ord(letter) - ord('A') + 1
        c = ws.cell(row=row, column=ci, value=text)
        c.font = FNT_SUB
        c.fill = FILL_POOL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = THIN


def _write_global_column_header_row(ws, row):
    """Single global column-header row used at the top of the worksheet.

    Writes the column-header text in B-O (and a 'Pool' label in A). This
    replaces the per-block repetition of the same text.
    """
    a = ws.cell(row=row, column=1, value='Pool')
    a.font = FNT_TITLE
    a.fill = FILL_POOL
    a.alignment = Alignment(horizontal='left', vertical='center')
    a.border = THIN
    for letter, text in HEADER_TEXTS.items():
        if letter == 'A':
            continue
        ci = ord(letter) - ord('A') + 1
        c = ws.cell(row=row, column=ci, value=text)
        c.font = FNT_SUB
        c.fill = FILL_POOL
        c.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        c.border = THIN


def _write_pool_name_row(ws, row, pool_name):
    """Per-pool name row: A=pool name, borders/fill across A-O."""
    a = ws.cell(row=row, column=1, value=pool_name)
    a.font = FNT_TITLE
    a.fill = FILL_POOL
    a.alignment = Alignment(horizontal='left', vertical='center')
    a.border = THIN
    for ci in range(2, 16):  # B..O
        c = ws.cell(row=row, column=ci)
        c.fill = FILL_POOL
        c.border = THIN


def _write_grade_row(ws, row, total_row, grade_label, balance, dist):
    """Per-grade row: A=label, B=balance(static), D=dist(static),
    formulas for E/F/J/K/L."""
    ws.cell(row=row, column=1, value=grade_label).font = FNT_NORM
    ws.cell(row=row, column=1).border = THIN

    bcell = ws.cell(row=row, column=2, value=balance)
    bcell.font = FNT_NORM
    bcell.fill = FILL_BASE
    bcell.number_format = ACCT
    bcell.border = THIN

    dcell = ws.cell(row=row, column=4, value=dist)
    dcell.font = FNT_NORM
    dcell.fill = FILL_BASE
    dcell.number_format = PCT4
    dcell.border = THIN

    # Formulas (E,F,J,K,L)
    tr = total_row
    # ACL Base Loss Rate is floored at 0 to match the Display Hist Bal and
    # ACL Env by Pool Mgmt Adj tabs (max(0, pool_ll * dist)).
    ecell = ws.cell(row=row, column=5, value=f"=MAX(0,D{row}*C{tr})")
    ecell.number_format = PCT4
    ecell.fill = FILL_BASE
    ecell.font = FNT_NORM
    ecell.border = THIN

    fcell = ws.cell(row=row, column=6, value=f"=B{row}*E{row}")
    fcell.number_format = ACCT
    fcell.fill = FILL_BASE
    fcell.font = FNT_NORM
    fcell.border = THIN

    jcell = ws.cell(row=row, column=10, value=f"=D{row}*I{tr}")
    jcell.number_format = PCT4
    jcell.fill = FILL_ADJ
    jcell.font = FNT_NORM
    jcell.border = THIN

    kcell = ws.cell(row=row, column=11, value=f"=E{row}+J{row}")
    kcell.number_format = PCT4
    kcell.fill = FILL_ADJ
    kcell.font = FNT_NORM
    kcell.border = THIN

    lcell = ws.cell(row=row, column=12, value=f"=K{row}*B{row}")
    lcell.number_format = ACCT
    lcell.fill = FILL_ADJ
    lcell.font = FNT_NORM
    lcell.border = THIN

    # Empty cells get borders for visual cleanliness
    for ci in (3, 7, 8, 9, 13, 14, 15):
        cc = ws.cell(row=row, column=ci)
        cc.border = THIN


def _write_total_row(ws, row, first_grade_row, last_grade_row,
                     total_balance, life_loss, env_factor, mgmt_adj, is_rr):
    """Per-pool Total row: B=balance(static), C=life_loss(static),
    F=allowance-before formula, G=env*F, H=F+G, I=mgmt_adj(EDITABLE),
    L=adjusted balance, M=L*env, N=L+M, O=N-H. The environmental factor
    is inlined as a constant in the G and M formulas."""
    ws.cell(row=row, column=1, value="Total").font = FNT_BOLD
    ws.cell(row=row, column=1).fill = FILL_TOTAL
    ws.cell(row=row, column=1).border = THIN

    b = ws.cell(row=row, column=2, value=total_balance)
    b.font = FNT_BOLD
    b.fill = FILL_TOTAL
    b.number_format = ACCT
    b.border = THIN

    c = ws.cell(row=row, column=3, value=life_loss)
    c.font = FNT_BOLD
    c.fill = FILL_TOTAL
    c.number_format = PCT4
    c.border = THIN

    # F formula: risk-rated pools sum their grade rows; non-risk-rated
    # pools use C*B directly (no grade detail). We know is_rr at build
    # time, so the branch is chosen here rather than via an in-cell IF.
    if is_rr:
        f_formula = f"=SUM(F{first_grade_row}:F{last_grade_row})"
        l_formula = f"=SUM(L{first_grade_row}:L{last_grade_row})"
    else:
        # NRR: total balance × loss rate (no grade rows contribute).
        # ACL Base Loss Rate floored at 0 to match the Display Hist Bal and
        # ACL Env by Pool Mgmt Adj tabs (max(0, pool_ll)).
        f_formula = f"=MAX(0,C{row})*B{row}"
        l_formula = f"=(I{row}+MAX(0,C{row}))*B{row}"

    fcell = ws.cell(row=row, column=6, value=f_formula)
    fcell.font = FNT_BOLD
    fcell.fill = FILL_TOTAL
    fcell.number_format = ACCT
    fcell.border = THIN

    # Environmental factor inlined as a constant so no hidden helper column
    # is needed.
    env = float(env_factor or 0)
    gcell = ws.cell(row=row, column=7, value=f"=F{row}*{env!r}")
    gcell.font = FNT_BOLD
    gcell.fill = FILL_TOTAL
    gcell.number_format = ACCT
    gcell.border = THIN

    hcell = ws.cell(row=row, column=8, value=f"=F{row}+G{row}")
    hcell.font = FNT_BOLD
    hcell.fill = FILL_TOTAL
    hcell.number_format = ACCT
    hcell.border = THIN

    # I — USER EDITABLE: pool-level mgmt adj
    icell = ws.cell(row=row, column=9, value=mgmt_adj)
    icell.font = FNT_INPUT
    icell.fill = FILL_INPUT
    icell.number_format = PCT4
    icell.border = Border(
        left=Side('medium', color='1B4F72'),
        right=Side('medium', color='1B4F72'),
        top=Side('medium', color='1B4F72'),
        bottom=Side('medium', color='1B4F72'),
    )

    lcell = ws.cell(row=row, column=12, value=l_formula)
    lcell.font = FNT_BOLD
    lcell.fill = FILL_TOTAL
    lcell.number_format = ACCT
    lcell.border = THIN

    mcell = ws.cell(row=row, column=13, value=f"=L{row}*{env!r}")
    mcell.font = FNT_BOLD
    mcell.fill = FILL_TOTAL
    mcell.number_format = ACCT
    mcell.border = THIN

    ncell = ws.cell(row=row, column=14, value=f"=L{row}+M{row}")
    ncell.font = FNT_BOLD
    ncell.fill = FILL_TOTAL
    ncell.number_format = ACCT
    ncell.border = THIN

    ocell = ws.cell(row=row, column=15, value=f"=N{row}-H{row}")
    ocell.font = FNT_BOLD
    ocell.fill = FILL_VAR
    ocell.number_format = ACCT
    ocell.border = THIN


def _write_grand_totals(ws, row, pool_total_rows):
    """Grand totals row summing across all pool Total rows."""
    if not pool_total_rows:
        return
    refs_b = ",".join(f"B{r}" for r in pool_total_rows)
    refs_f = ",".join(f"F{r}" for r in pool_total_rows)
    refs_g = ",".join(f"G{r}" for r in pool_total_rows)
    refs_h = ",".join(f"H{r}" for r in pool_total_rows)
    refs_l = ",".join(f"L{r}" for r in pool_total_rows)
    refs_m = ",".join(f"M{r}" for r in pool_total_rows)
    refs_n = ",".join(f"N{r}" for r in pool_total_rows)
    refs_o = ",".join(f"O{r}" for r in pool_total_rows)

    ws.cell(row=row, column=1, value="Grand Total").font = FNT_TITLE
    ws.cell(row=row, column=1).fill = FILL_TOTAL
    ws.cell(row=row, column=1).border = THIN

    for ci, formula, fmt in [
        (2, f"=SUM({refs_b})", ACCT),
        (6, f"=SUM({refs_f})", ACCT),
        (7, f"=SUM({refs_g})", ACCT),
        (8, f"=SUM({refs_h})", ACCT),
        (12, f"=SUM({refs_l})", ACCT),
        (13, f"=SUM({refs_m})", ACCT),
        (14, f"=SUM({refs_n})", ACCT),
        (15, f"=SUM({refs_o})", ACCT),
    ]:
        c = ws.cell(row=row, column=ci, value=formula)
        c.font = FNT_TITLE
        c.fill = FILL_TOTAL
        c.number_format = fmt
        c.border = THIN


# ── Public composer ───────────────────────────────────────────────
def compose_mgmt_adj_napkin(client_name, snapshot_date, df, config, grades, hist):
    """Build the Management Adjustment Worksheet workbook.

    Returns ``(wb, fname)`` so it can be saved by the standard
    ``generate_report.generate_report`` dispatcher alongside the TCT /
    Vizo / Vizo-Supplemental outputs.
    """
    cu = config.get('credit_union', client_name)
    no_score = config.get('no_score_label', 'Not Reported')
    snap_year = int(snapshot_date[:4])
    snap_month = int(snapshot_date[5:7])
    config_for_calc = dict(config)
    config_for_calc['_snap_year'] = snap_year
    config_for_calc['_snap_month'] = snap_month

    # Mgmt adj resolution helpers (firm-default + per-pool overrides).
    mgmt_adj_by_pool = config.get('mgmt_adj_by_pool') or {}
    pool_use_default = _build_pool_use_default_map(config)
    admin_default = _load_admin_default_mgmt_adj()

    # Pool lists (matching report_tct ordering rules).
    db_pools = set(df['loan_pool'].unique()) if 'loan_pool' in df.columns else set()
    nrr = set(config.get('not_risk_rated', []))
    imp = hist.get('impaired', {}) if hist else {}
    risk_rated_flags = imp.get('risk_rated', {})
    warm_order = config.get('pool_order') or imp.get('pool_order', [])
    hist_bal_pools = list((imp.get('hist_bal_data') or {}).keys())
    acl_pools_data = imp.get('acl_pools', {}) or {}
    # Specifically Identified (impaired) balances per pool/grade, used to
    # derive the Loan Loss Calculation Balance for column B.
    spec_id_by_pool = imp.get('spec_id_by_pool', {}) or {}
    pools = _merge_pool_orders(
        list(db_pools),
        warm_order,
        extra=[list(acl_pools_data.keys()), hist_bal_pools],
    )
    if not pools:
        pools = _sort_pools(list(db_pools), config)

    # Drop sentinel pools (Exclude / Ignore) from the report.
    pools = [p for p in pools if str(p).strip().lower() not in ('exclude', 'ignore')]

    # Baseline math.
    life_loss = _pool_life_loss_acl(pools, hist, config_for_calc)
    dq_var_map = _compute_dq_var(pools, hist, config_for_calc, snap_year, snap_month)
    econ_stress = _econ_stress(config) if hist else 0.0

    # Env-Range overrides from WARM (same precedence as report_tct).
    er = imp.get('env_ranges', {}) or {}
    ncc_r = er.get('ncc') or NCC_RANGES
    dq_r  = er.get('dq')  or DQ_RANGES
    es_r  = er.get('es')  or ES_RANGES

    # Grade list — visible grades for each pool. BRR pools use BRR labels.
    visible_gl = [g for g in _all_grades(grades, no_score) if not _is_hidden(g)]
    brr_labels = _brr_grade_labels(config, no_score)
    brr_pool_lcs = _brr_pools_set(config) if brr_labels else set()

    # Workbook setup.
    wb = Workbook()
    ws = wb.active
    ws.title = "Mgmt Adj Napkin"

    # Column widths sized so each header word fits without truncation
    # when wrapped across multiple lines (longest tokens: Distribution=12,
    # Environmental=13, Distributed=11, Aggregated=10, Adjustment=10).
    widths = [22, 16, 14, 14, 14, 18, 16, 18, 14, 16, 14, 20, 16, 18, 14]
    for ci, w in enumerate(widths, start=1):
        if w > 0:
            ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Title block (rows 1-3) ──────────────────────────────────
    ws['A1'] = cu
    ws['A1'].font = Font(name='Arial', bold=True, size=16, color='1B4F72')
    ws['A2'] = "Management Adjustment Worksheet"
    ws['A2'].font = FNT_TITLE
    ws['A3'] = f"For Period Ending {_snap_display(snapshot_date)}"
    ws['A3'].font = FNT_SUB

    # User instructions box (rows 1-3 columns J onwards).
    ws['J1'] = "How to use this worksheet"
    ws['J1'].font = FNT_SUB
    ws['J2'] = (
        "Edit the yellow Management Adj cell (column I, Total row) for any"
        " pool. The Adjusted columns (J-O) recalculate automatically; column O"
        " shows the variance versus the baseline (column H)."
    )
    ws['J2'].font = FNT_NOTE
    ws['J2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[2].height = 40

    # ── Global header rows ─────────────────────────────────────
    # Letter row + formula-label row + column-header row appear once at
    # the top; per-pool blocks below carry only the pool name + grades.
    GLOBAL_LETTER_ROW  = 5
    GLOBAL_FORMULA_ROW = 6
    GLOBAL_HEADER_ROW  = 7
    _write_block_letter_row(ws, GLOBAL_LETTER_ROW)
    _write_block_formula_row(ws, GLOBAL_FORMULA_ROW)
    _write_global_column_header_row(ws, GLOBAL_HEADER_ROW)
    # Header text wraps to up to 3 lines (e.g. "Adjusted Balance Before
    # Environmental") — bump height so wrapped text isn't clipped.
    ws.row_dimensions[GLOBAL_HEADER_ROW].height = 60

    # ── Per-pool blocks ────────────────────────────────────────
    block_top = GLOBAL_HEADER_ROW + 1 + BLOCK_GAP  # leave a blank row gap
    pool_total_rows: list[int] = []

    for pool in pools:
        pdf = df[df['loan_pool'] == pool] if 'loan_pool' in df.columns else df.iloc[0:0]
        is_rr = (
            risk_rated_flags.get(pool, pool not in nrr)
            if risk_rated_flags else (pool not in nrr)
        )

        env_factor = _compute_env_factor(
            pool, pdf, grades, config, hist, dq_var_map,
            econ_stress, ncc_r, dq_r, es_r, is_rr,
        )

        # Per-pool baseline mgmt adj (Step 16 manual; admin default
        # applies only when pool's life_loss == 0 per report_tct rule).
        pool_ll = life_loss.get(pool, 0)
        mgmt_adj = _resolve_mgmt_adj_total(
            pool, pool_use_default, mgmt_adj_by_pool,
            admin_default, base_rate=pool_ll,
        )

        # Grade-row balances (only meaningful for risk-rated pools).
        # Non-risk-rated pools render as a single Total row with no
        # grade breakout — their Total-row formulas use C*B directly
        # and don't reference any grade rows.
        if is_rr:
            pool_grade_labels = (
                brr_labels if _is_brr_pool(pool, brr_pool_lcs)
                else visible_gl
            )
        else:
            pool_grade_labels = []

        # Block rows — dynamic height with a single pool-name row at the
        # top, followed by one row per grade (RR pools only) and a Total
        # row. Letter/formula/column headers are emitted once globally
        # above the first block, not per pool.
        num_grades = len(pool_grade_labels)
        name_row    = block_top
        first_grade = block_top + 1
        total_row   = first_grade + num_grades
        last_grade  = total_row - 1

        _write_pool_name_row(ws, name_row, pool)

        # Try WARM-sourced per-grade balance first (matches main report).
        pool_lc = pool.strip().lower()
        warm_pool = next(
            (v for k, v in acl_pools_data.items() if k.strip().lower() == pool_lc),
            None,
        )
        warm_grades = warm_pool['grades'] if warm_pool else {}
        bal_detail = imp.get('pool_bal_detail', {}).get(pool, {})

        # Render exactly len(pool_grade_labels) grade rows — no padding.
        # Column B carries the Loan Loss Calculation Balance (Balance with
        # Specifically Identified / impaired amounts removed) so it matches
        # the ACL Env by Pool Mgmt Adj tab's "Loan Loss Calc Balance" column.
        pool_grade_balance_sum = 0.0
        pool_grade_spec_id_sum = 0.0
        for gi, g in enumerate(pool_grade_labels):
            row_n = first_grade + gi
            wg = warm_grades.get(g, {})
            gd = bal_detail.get(g, {})

            # Balance precedence: pool_bal_detail > WARM grades > df.
            bst = gd.get('balance_sheet_total') if gd else None
            if bst is not None:
                balance = bst
            elif wg:
                balance = wg.get('balance', 0)
            elif not pdf.empty:
                g_df = pdf[pdf.get('current_grade') == g]
                balance = float(g_df.get('current_balance', pd.Series(dtype=float)).sum())
            else:
                balance = 0.0
            # Specific Identification precedence mirrors the ACL Env by Pool
            # tab: WARM grade spec_id first, else Impaired Loans detail.
            specific_id = wg.get('spec_id', 0) if wg else 0
            if not specific_id and pool in spec_id_by_pool:
                specific_id = spec_id_by_pool[pool].get(g, 0)
            calc_bal = (balance or 0) - (specific_id or 0)
            pool_grade_balance_sum += balance or 0
            pool_grade_spec_id_sum += specific_id or 0
            dist = (
                _dist_factor(len(DIST_FACTORS) - 1)
                if g == no_score else _dist_factor(gi)
            )
            _write_grade_row(ws, row_n, total_row, g, calc_bal, dist)

        # Pool total balance: prefer pool_bal_detail Total; else WARM total;
        # else sum of df.current_balance for the pool.
        ptd = bal_detail.get('Total', {}) if bal_detail else {}
        if _is_brr_pool(pool, brr_pool_lcs):
            # BRR pools: derive the Total from the rendered per-grade rows;
            # pool_bal_detail's Total is keyed to FICO grades only.
            total_balance = pool_grade_balance_sum
        elif ptd and ptd.get('balance_sheet_total') is not None:
            total_balance = ptd['balance_sheet_total']
        elif warm_pool and warm_pool.get('total'):
            total_balance = warm_pool['total'].get('balance', 0)
        elif not pdf.empty:
            total_balance = float(pdf.get('current_balance', pd.Series(dtype=float)).sum())
        else:
            total_balance = 0.0

        # Total Specific Identification (impaired) — mirror the ACL Env by
        # Pool tab so column B holds the Loan Loss Calculation Balance.
        if _is_brr_pool(pool, brr_pool_lcs):
            total_spec_id = pool_grade_spec_id_sum
        else:
            total_spec_id = (
                warm_pool['total'].get('spec_id', 0)
                if warm_pool and warm_pool.get('total') else 0
            )
            if not total_spec_id and pool in spec_id_by_pool:
                total_spec_id = sum(spec_id_by_pool[pool].values())
        total_calc_bal = (total_balance or 0) - (total_spec_id or 0)

        _write_total_row(
            ws, total_row, first_grade, last_grade,
            total_calc_bal, pool_ll, env_factor, mgmt_adj, is_rr,
        )
        pool_total_rows.append(total_row)

        # Modest height bump on the per-pool name row.
        ws.row_dimensions[name_row].height = 22

        block_top = total_row + 1 + BLOCK_GAP

    # ── Grand totals ────────────────────────────────────────────
    grand_row = block_top + 1
    _write_grand_totals(ws, grand_row, pool_total_rows)

    # ── CECL Adjustment summary (mirrors "ACL Env by Pool Mgmt Adj") ──
    # These lines sit below the Grand Total in column N (Adjusted Total ACL
    # Allowance) so the analyst sees how the adjusted pooled allowance rolls
    # up into the Total Allowance Needed and the resulting overfunded /
    # underfunded Adjustment. The Specifically Identified Allowance and ACL
    # Balance are static values pulled from the ACL Env by Pool Mgmt Adj tab
    # (via hist['impaired']); Total Allowance Needed and Adjustment are live
    # formulas referencing the Grand Total's column N cell, so they recompute
    # as the Management Adj cells (column I) are edited.
    summary_last_row = grand_row
    if pool_total_rows:
        # Specifically Identified Allowance — same precedence the ACL Env by
        # Pool Mgmt Adj tab uses (report_vizo/report_tct._sheet_acl_reserve).
        _acl_sum = imp.get('acl_summary') or {}
        _acl_imp = imp.get('acl_impaired') or {}
        if 'total_spec_allow' in _acl_sum:
            spec_id_allow = _acl_sum.get('total_spec_allow', 0)
        elif _acl_imp:
            spec_id_allow = sum(_acl_imp.values())
        else:
            spec_id_allow = imp.get('total_spec_id', 0)
        oac_total = sum(o['amount'] for o in _other_allowance_considerations(config))
        acl_bal = imp.get('acl_balance', config.get('acl_balance', 0))

        spec_row   = grand_row + 2
        needed_row = spec_row + 1
        balbal_row = needed_row + 1
        adj_row    = balbal_row + 1
        summary_last_row = adj_row

        needed_formula = f"=N{grand_row}+N{spec_row}" + (
            f"+{oac_total}" if oac_total else "")

        summary_lines = [
            (spec_row, "Total Specifically Identified Allowance",
             spec_id_allow, FILL_TOTAL),
            (needed_row, "Total Allowance Needed",
             needed_formula, FILL_TOTAL),
            (balbal_row,
             f"Allowance for Credit Loss Balance as of {_snap_display(snapshot_date)}",
             acl_bal, FILL_BASE),
            (adj_row, "Adjustment (Overfunded)",
             f"=N{needed_row}-N{balbal_row}", FILL_VAR),
        ]
        for srow, label, value, fill in summary_lines:
            lc = ws.cell(row=srow, column=1, value=label)
            lc.font = FNT_BOLD
            lc.fill = fill
            lc.border = THIN
            for ci in range(2, 14):  # fill/border across A-M for a clean band
                mc = ws.cell(row=srow, column=ci)
                mc.fill = fill
                mc.border = THIN
            vc = ws.cell(row=srow, column=14, value=value)
            vc.font = FNT_BOLD
            vc.fill = fill
            vc.number_format = ACCT
            vc.border = THIN
            ws.cell(row=srow, column=15).border = THIN

    # ── Footer note ─────────────────────────────────────────────
    note_row = summary_last_row + 3
    ws.cell(row=note_row, column=1, value=(
        "Note: Yellow cells in column I are the editable management"
        " adjustments. Baseline figures (columns A-H) are static snapshots"
        " from the period-end report and are not linked to any other tab."
        " Each pool's environmental factor is applied directly within the"
        " G and M formulas."
    )).font = FNT_NOTE
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=15)
    ws.row_dimensions[note_row].height = 36

    # Freeze rows above the first pool block so the title and global
    # column headers stay visible while scrolling.
    ws.freeze_panes = 'A8'

    # Page setup: landscape, fit-to-width.
    ws.page_setup.orientation = 'landscape'
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)
    ws.print_title_rows = '1:7'

    # File name follows the user's requested convention.
    safe_cu = cu.replace('/', '-').replace('\\', '-')
    snap_prefix = snapshot_date[:7]   # "YYYY-MM"
    fname = f"{snap_prefix} Management Adjustment Worksheet- {safe_cu}.xlsx"
    return wb, fname
