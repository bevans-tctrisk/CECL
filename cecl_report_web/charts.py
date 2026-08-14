"""Read embedded workbook charts and re-render them as inline SVG.

openpyxl exposes each worksheet's charts (``ws._charts``) with their type,
title, and data/category references. We resolve those references to real
values and draw dependency-free, vector SVG — deterministic (good for the
visual-regression harness) and crisp in the PDF. Excel-default palette so
the charts read as "the same chart".
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter

# Excel 2016+ default series palette.
PALETTE = ["#4472C4", "#ED7D31", "#A5A5A5", "#FFC000", "#5B9BD5", "#70AD47"]


def _fill_hex(spPr) -> str | None:
    """Return the 6-hex solid fill of a graphicalProperties/spPr, or None."""
    if spPr is None:
        return None
    sf = getattr(spPr, "solidFill", None)
    if sf is None:
        return None
    v = getattr(sf, "value", None) or getattr(sf, "srgbClr", None)
    if isinstance(v, str) and len(v) == 6:
        return "#" + v
    rgb = getattr(v, "rgb", None) if v is not None else None
    if isinstance(rgb, str) and len(rgb) >= 6:
        return "#" + rgb[-6:]
    return None


def _stroke_hex(spPr) -> str | None:
    """Return the 6-hex line/stroke color of a graphicalProperties/spPr."""
    if spPr is None:
        return None
    ln = getattr(spPr, "ln", None)
    if ln is None:
        return None
    sf = getattr(ln, "solidFill", None)
    if sf is None:
        return None
    v = getattr(sf, "value", None) or getattr(sf, "srgbClr", None)
    if isinstance(v, str) and len(v) == 6:
        return "#" + v
    rgb = getattr(v, "rgb", None) if v is not None else None
    if isinstance(rgb, str) and len(rgb) >= 6:
        return "#" + rgb[-6:]
    return None


def _lighten(hex_color: str, frac: float) -> str:
    try:
        h = hex_color.lstrip("#")
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    except (ValueError, IndexError):
        return hex_color
    f = max(0.0, min(1.0, frac))
    r = round(r + (255 - r) * f)
    g = round(g + (255 - g) * f)
    b = round(b + (255 - b) * f)
    return f"#{r:02X}{g:02X}{b:02X}"

_REF_RE = re.compile(r"^(?:'([^']+)'|([^!]+))!(.+)$")


def _split_ref(ref: str) -> tuple[str | None, str]:
    """'Sheet'!$A$1:$B$2 -> ('Sheet', '$A$1:$B$2')."""
    m = _REF_RE.match(ref or "")
    if not m:
        return None, ref
    sheet = m.group(1) or m.group(2)
    return sheet, m.group(3)


def _resolve(wb, ref: str | None) -> list[Any]:
    """Resolve a cell/range reference to a flat list of values."""
    if not ref:
        return []
    sheet, a1 = _split_ref(ref)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    a1 = a1.replace("$", "")
    try:
        min_c, min_r, max_c, max_r = range_boundaries(a1)
    except Exception:  # noqa: BLE001
        return []
    out: list[Any] = []
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            out.append(ws.cell(r, c).value)
    return out


def _resolve_scalar(wb, ref: str | None) -> Any:
    vals = _resolve(wb, ref)
    return vals[0] if vals else None


def _ref_numfmt(wb, ref: str | None) -> str | None:
    """Number format of the first cell of a value range (used for data labels)."""
    if not ref:
        return None
    sheet, a1 = _split_ref(ref)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    a1 = a1.replace("$", "")
    try:
        min_c, min_r, _, _ = range_boundaries(a1)
        return ws.cell(min_r, min_c).number_format
    except Exception:  # noqa: BLE001
        return None


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _chart_title(ch) -> str | None:
    try:
        if ch.title and ch.title.tx and ch.title.tx.rich:
            return "".join(
                (r.t or "") for p in ch.title.tx.rich.p for r in (p.r or [])
            ).strip() or None
    except Exception:  # noqa: BLE001
        pass
    return None


def read_chart_specs(report_path: str | Path, sheet: str) -> list[dict]:
    """Return normalized chart specs for one worksheet."""
    wb = load_workbook(report_path)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    specs: list[dict] = []
    for ch in getattr(ws, "_charts", []) or []:
        ctype = type(ch).__name__
        series: list[dict] = []
        for s in getattr(ch, "series", []) or []:
            name = None
            try:
                if s.tx and s.tx.strRef:
                    name = _resolve_scalar(wb, s.tx.strRef.f)
                elif s.tx and s.tx.v:
                    name = s.tx.v
            except Exception:  # noqa: BLE001
                name = None
            vref = None
            try:
                vref = s.val.numRef.f if (s.val and s.val.numRef) else None
            except Exception:  # noqa: BLE001
                vref = None
            cref = None
            try:
                if s.cat and s.cat.numRef:
                    cref = s.cat.numRef.f
                elif s.cat and s.cat.strRef:
                    cref = s.cat.strRef.f
            except Exception:  # noqa: BLE001
                cref = None
            sp = getattr(s, "graphicalProperties", None)
            _fill = _fill_hex(sp)
            _line = _stroke_hex(sp)
            series.append({
                "name": name,
                "values": [_num(v) for v in _resolve(wb, vref)],
                "cats": [("" if v is None else str(v)) for v in _resolve(wb, cref)],
                "color": _fill or _line,
                "filled": _fill is not None,
                "point_colors": [
                    (_fill_hex(getattr(dp, "graphicalProperties", None))
                     or _stroke_hex(getattr(dp, "graphicalProperties", None)))
                    for dp in (getattr(s, "data_points", None) or [])
                ],
                "label_fmt": (getattr(getattr(s, "dLbls", None), "numFmt", None)
                              or getattr(getattr(ch, "dLbls", None), "numFmt", None)
                              or _ref_numfmt(wb, vref)),
                "show_labels": bool(getattr(s, "dLbls", None)
                                    or getattr(ch, "dLbls", None)),
            })
        specs.append({
            "type": ctype,
            "bar_dir": getattr(ch, "type", None),      # 'col' | 'bar'
            "grouping": getattr(ch, "grouping", None),  # 'clustered' | 'stacked'
            "title": _chart_title(ch),
            "series": series,
        })
    return specs


# ── SVG rendering ────────────────────────────────────────────────────

import html as _html

_W, _H = 340, 240


def _esc(s) -> str:
    return _html.escape("" if s is None else str(s))


def _wrap(inner: str, title: str | None, w: int = _W, h: int = _H) -> str:
    t = (f'<text x="{w/2:.0f}" y="16" text-anchor="middle" '
         f'font-size="12" font-weight="700" fill="#000">{_esc(title)}</text>'
         if title else "")
    return (
        f'<svg class="chart" xmlns="http://www.w3.org/2000/svg" '
        f'viewBox="0 0 {w} {h}" width="{w}" height="{h}">'
        f'<rect x="0" y="0" width="{w}" height="{h}" fill="#fff"/>{t}{inner}</svg>'
    )


def _legend(items: list[tuple[str, str]], x: int, y: int,
            *, horizontal: bool = False) -> str:
    """items = [(label, color)]. Vertical (default) or horizontal legend."""
    out = []
    if horizontal:
        cx = x
        for label, color in items:
            out.append(
                f'<rect x="{cx}" y="{y}" width="10" height="10" fill="{color}"/>'
                f'<text x="{cx + 14}" y="{y + 9}" font-size="9" fill="#000">'
                f'{_esc(label)}</text>')
            cx += 22 + len(str(label)) * 5.2
        return "".join(out)
    for i, (label, color) in enumerate(items):
        yy = y + i * 16
        out.append(
            f'<rect x="{x}" y="{yy}" width="10" height="10" fill="{color}"/>'
            f'<text x="{x + 14}" y="{yy + 9}" font-size="10" fill="#000">'
            f'{_esc(label)}</text>')
    return "".join(out)


def _svg_bar(series: list[dict], cats: list[str], title: str | None,
             *, horizontal: bool, stacked: bool) -> str:
    from .format import excel_format

    top, right = 30, 14
    bottom = 56 if (not horizontal and len(cats) > 16) else 40
    left = 150 if horizontal else 40
    pw, ph = _W - left - right, _H - top - bottom
    nseries = max(1, len(series))
    multi = nseries > 1

    def _fmt(val, s):
        fmt = s.get("label_fmt") or "0%"
        try:
            return excel_format(val, fmt)
        except Exception:  # noqa: BLE001
            return f"{val*100:.0f}%"

    def _bar_color(s, i, n, si):
        pts = s.get("point_colors") or []
        base = s.get("color") or PALETTE[si % len(PALETTE)]
        distinct = {p for p in pts if p}
        if len(distinct) > 1:
            # genuinely per-point colors (e.g. green/maroon/teal points)
            return pts[i] if i < len(pts) and pts[i] else base
        if pts and s.get("filled", True) and n > 1:
            # single color across points -> reproduce Excel's lumMod gradient
            return _lighten(base, (i / (n - 1)) * 0.68)
        return base

    def _rect_attrs(s, color):
        if s.get("filled", True):
            return f'fill="{color}"'
        return f'fill="none" stroke="{color}" stroke-width="1.4"'

    all_vals = [v for s in series for v in s["values"]]
    ncat = len(cats)
    if stacked:
        sums = [sum(abs(s["values"][i]) for s in series if i < len(s["values"]))
                for i in range(ncat)]
        vmax = max(sums + [0.0]) or 1.0
    else:
        vmax = max([abs(v) for v in all_vals] + [0.0]) or 1.0

    parts: list[str] = []

    if horizontal:
        row_h = ph / max(1, ncat)
        bar_h = min(row_h * 0.62, 18)
        for i, cat in enumerate(cats):
            cy = top + i * row_h + (row_h - bar_h) / 2
            if stacked:
                acc = 0.0
                for si, s in enumerate(series):
                    val = s["values"][i] if i < len(s["values"]) else 0.0
                    w = (abs(val) / vmax) * pw
                    if w > 0.5:
                        parts.append(
                            f'<rect x="{left + acc:.1f}" y="{cy:.1f}" '
                            f'width="{w:.1f}" height="{bar_h:.1f}" '
                            f'{_rect_attrs(s, _bar_color(s, i, ncat, si))}/>')
                        if s.get("show_labels") and abs(val) > 1e-9:
                            parts.append(
                                f'<text x="{left + acc + w/2:.1f}" '
                                f'y="{cy + bar_h/2 + 3:.1f}" text-anchor="middle" '
                                f'font-size="8" fill="#fff">{_esc(_fmt(val, s))}</text>')
                    acc += w
            else:
                s = series[0]
                val = s["values"][i] if i < len(s["values"]) else 0.0
                w = (abs(val) / vmax) * (pw - 26)  # leave room for the end label
                parts.append(
                    f'<rect x="{left:.1f}" y="{cy:.1f}" width="{max(0, w):.1f}" '
                    f'height="{bar_h:.1f}" {_rect_attrs(s, _bar_color(s, i, ncat, 0))}/>')
                if s.get("show_labels"):
                    parts.append(
                        f'<text x="{left + w + 3:.1f}" y="{cy + bar_h/2 + 3:.1f}" '
                        f'font-size="8" fill="#000">{_esc(_fmt(val, s))}</text>')
            parts.append(
                f'<text x="{left - 5}" y="{top + i * row_h + row_h/2 + 3:.1f}" '
                f'text-anchor="end" font-size="8" fill="#000">{_esc(cat)}</text>')
    else:
        group_w = pw / max(1, ncat)
        # thin crowded x-axis labels (e.g. 40-month time series)
        lbl_step = max(1, (ncat + 11) // 12) if ncat > 16 else 1
        rot = ncat > 16

        def _catlabel(i, cat):
            if i % lbl_step:
                return ""
            cx = left + i * group_w + group_w / 2
            txt = _esc(str(cat)[:7] if len(str(cat)) > 7 else cat)
            if rot:
                yy = top + ph + 8
                return (f'<text x="{cx:.1f}" y="{yy:.1f}" text-anchor="end" '
                        f'font-size="7" fill="#000" '
                        f'transform="rotate(-55 {cx:.1f} {yy:.1f})">{txt}</text>')
            return (f'<text x="{cx:.1f}" y="{top + ph + 12}" text-anchor="middle" '
                    f'font-size="8" fill="#000">{txt}</text>')

        if stacked:
            bw = min(group_w * 0.8, 40)
            for i, cat in enumerate(cats):
                gx = left + i * group_w + (group_w - bw) / 2
                acc = 0.0
                for si, s in enumerate(series):
                    val = s["values"][i] if i < len(s["values"]) else 0.0
                    bh = (abs(val) / vmax) * ph
                    y = top + ph - acc - bh
                    parts.append(
                        f'<rect x="{gx:.1f}" y="{y:.1f}" width="{bw:.1f}" '
                        f'height="{bh:.1f}" {_rect_attrs(s, _bar_color(s, i, ncat, si))}/>')
                    acc += bh
                parts.append(_catlabel(i, cat))
        else:
            bw = min(group_w * 0.72 / nseries, 40)
            for i, cat in enumerate(cats):
                gx = left + i * group_w + (group_w - bw * nseries) / 2
                for si, s in enumerate(series):
                    val = s["values"][i] if i < len(s["values"]) else 0.0
                    bh = (abs(val) / vmax) * ph
                    y = top + ph - bh
                    parts.append(
                        f'<rect x="{gx + si * bw:.1f}" y="{y:.1f}" width="{bw:.1f}" '
                        f'height="{bh:.1f}" {_rect_attrs(s, _bar_color(s, i, ncat, si))}/>')
                    if s.get("show_labels") and nseries == 1:
                        lx, ly = gx + bw / 2, y + 12
                        parts.append(
                            f'<text x="{lx:.1f}" y="{ly:.1f}" text-anchor="middle" '
                            f'font-size="8" fill="#fff" transform="rotate(-90 {lx:.1f} {ly:.1f})">'
                            f'{_esc(_fmt(val, s))}</text>')
                parts.append(_catlabel(i, cat))

    if multi:
        parts.append(_legend(
            [(s["name"] or f"Series {i+1}",
              s.get("color") or PALETTE[i % len(PALETTE)])
             for i, s in enumerate(series)],
            left, top + ph + (34 if (not horizontal and ncat > 16) else 22),
            horizontal=True))
    return _wrap("".join(parts), title)


def _svg_pie(values: list[float], cats: list[str], title: str | None,
             *, doughnut: bool, colors: list[str] | None = None) -> str:
    import math
    cx, cy, rad = 100, 130, 78
    inner = rad * 0.55 if doughnut else 0
    total = sum(v for v in values if v > 0)

    def _col(i):
        if colors and i < len(colors) and colors[i]:
            return colors[i]
        return PALETTE[i % len(PALETTE)]

    parts: list[str] = []
    if total <= 0:
        parts.append(f'<circle cx="{cx}" cy="{cy}" r="{rad}" fill="none" '
                     f'stroke="#c9c9c9"/>'
                     f'<text x="{cx}" y="{cy}" text-anchor="middle" '
                     f'font-size="10" fill="#888">No data</text>')
    else:
        ang = -math.pi / 2
        for i, v in enumerate(values):
            if v <= 0:
                continue
            frac = v / total
            a2 = ang + frac * 2 * math.pi
            large = 1 if frac > 0.5 else 0
            x1, y1 = cx + rad * math.cos(ang), cy + rad * math.sin(ang)
            x2, y2 = cx + rad * math.cos(a2), cy + rad * math.sin(a2)
            color = _col(i)
            if frac >= 0.999:
                parts.append(f'<circle cx="{cx}" cy="{cy}" r="{rad}" '
                             f'fill="{color}"/>')
            else:
                parts.append(
                    f'<path d="M{cx},{cy} L{x1:.1f},{y1:.1f} '
                    f'A{rad},{rad} 0 {large} 1 {x2:.1f},{y2:.1f} Z" '
                    f'fill="{color}"/>')
            ang = a2
        if doughnut and inner:
            parts.append(f'<circle cx="{cx}" cy="{cy}" r="{inner:.0f}" '
                         f'fill="#fff"/>')
    parts.append(_legend(
        [(f"{cats[i] if i < len(cats) else ''} "
          f"({(values[i]/total*100 if total else 0):.0f}%)",
          _col(i))
         for i in range(len(values))],
        195, 100))
    return _wrap("".join(parts), title)


def render_chart_svg(spec: dict) -> str:
    """Dispatch a normalized chart spec to the right SVG renderer."""
    ctype = spec.get("type", "")
    series = spec.get("series", [])
    title = spec.get("title")
    if "Pie" in ctype or "Doughnut" in ctype:
        s0 = series[0] if series else {"values": [], "cats": []}
        return _svg_pie(s0.get("values", []), s0.get("cats", []), title,
                        doughnut="Doughnut" in ctype,
                        colors=s0.get("point_colors") or None)
    # Bar/column chart.
    cats = series[0]["cats"] if series else []
    horizontal = spec.get("bar_dir") == "bar"
    stacked = spec.get("grouping") == "stacked"
    return _svg_bar(series, cats, title, horizontal=horizontal, stacked=stacked)


def render_charts_for_sheet(report_path: str | Path, sheet: str) -> list[str]:
    """Return a list of SVG strings for every chart on a worksheet."""
    return [render_chart_svg(s) for s in read_chart_specs(report_path, sheet)]

