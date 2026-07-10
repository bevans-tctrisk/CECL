"""Adapter: read a generated report workbook into the ReportModel.

Phase-0 data source. Populating the model from the already-generated
xlsx guarantees the PDF numbers match the Excel report exactly, and it
doubles as the "extract a page" capability. Later phases can add a
compute-time populator with the same model as output — templates don't
change.

Scanning is label-anchored (find the row by its text, not a fixed
coordinate) so it survives minor row-position drift between report
versions.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .model import (
    AclEnvPage,
    AclPoolRow,
    AdjustmentRow,
    CoverPage,
    GradeMigrationRow,
    GridCell,
    GridPage,
    ImprDeterPage,
    KeyValueRow,
    MatrixCell,
    MatrixRow,
    PoolMigrationRow,
    RiskChangeMatrix,
    RiskChangePage,
)
from .format import excel_format


def load_cover(path, sheet: str | None = None) -> "CoverPage":
    """Read the Vizo cover sheet faithfully: title, CU, date, disclaimer
    paragraph, footer, and the two embedded logos (as data URIs)."""
    import base64
    import datetime as _dt

    wb = load_workbook(path)  # NOT read_only — need ws._images
    if sheet is None:
        sheet = next((s for s in wb.sheetnames if "cover" in s.lower()),
                     wb.sheetnames[0])
    ws = wb[sheet]

    def _txt(coord):
        v = ws[coord].value
        return None if v is None else str(v).strip()

    title = _txt("A14") or "CECL Credit Migration Report"
    cu = _txt("A16") or ""
    dv = ws["A17"].value
    if isinstance(dv, (_dt.datetime, _dt.date)):
        date_text = f"{dv.month}/{dv.day}/{dv.year}"
    else:
        date_text = None if dv is None else str(dv).strip()
    para = ws["B21"].value
    paragraph = None if para is None else str(para)
    footer = _txt("B44")

    logos = []
    for im in getattr(ws, "_images", []):
        try:
            data = im._data()
            row = im.anchor._from.row
            fmt = getattr(im, "format", "png") or "png"
            logos.append((row, f"data:image/{fmt};base64,"
                          + base64.b64encode(data).decode("ascii")))
        except Exception:  # noqa: BLE001
            continue
    logos.sort(key=lambda x: x[0])
    top_logo = logos[0][1] if logos else None
    bottom_logo = logos[-1][1] if len(logos) > 1 else None
    wb.close()

    return CoverPage(
        credit_union=cu, period_ending=date_text or "", title=title,
        date_text=date_text, paragraph=paragraph, footer=footer,
        top_logo=top_logo, bottom_logo=bottom_logo)


_IMPR_DETER_TAB_HINTS = ("impr deter", "improved", "deteriorated")
_CECL_LABELS = (
    "Total Specifically Identified Allowance",
    "Total Allowance Needed",
    "Allowance for Credit Loss Balance",
    "Adjustment (Overfunded)",
)


def _find_tab(wb, hints: tuple[str, ...]) -> str | None:
    for sn in wb.sheetnames:
        low = sn.strip().lower()
        if any(h in low for h in hints):
            return sn
    return None


def _num(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _cell_str(v) -> str:
    return "" if v is None else str(v).strip()


def load_impr_deter(report_path: str | Path) -> ImprDeterPage:
    """Read the 'Impr Deter' tab of a Vizo Model workbook into the model."""
    wb = load_workbook(report_path, data_only=True)
    tab = _find_tab(wb, _IMPR_DETER_TAB_HINTS)
    if tab is None:
        raise ValueError(
            f"No Improved/Deteriorated tab found in {Path(report_path).name} "
            f"(sheets: {wb.sheetnames})"
        )
    ws = wb[tab]

    # ── Header lines: first 4 populated rows of column A ──
    heading: list[str] = []
    cu = ""
    period = ""
    for r in range(1, 6):
        v = ws.cell(r, 1).value
        if v is None or v == "":
            continue
        if r == 1:
            cu = _cell_str(v)
        elif hasattr(v, "strftime"):
            period = f"{v.month}/{v.day}/{v.year}"
            heading.append(period)
        else:
            heading.append(_cell_str(v))

    # ── CECL Adjustment box: label in a left column, value ~3 cols right ──
    cecl: list[KeyValueRow] = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            label = _cell_str(ws.cell(r, c).value)
            if not label:
                continue
            for want in _CECL_LABELS:
                if label.startswith(want) or want.split(" as of")[0] in label:
                    # value: first numeric cell to the right on this row
                    val = None
                    for cc in range(c + 1, ws.max_column + 1):
                        n = _num(ws.cell(r, cc).value)
                        if n is not None:
                            val = n
                            break
                    cecl.append(KeyValueRow(label=label, value=val))
                    break

    # ── Improved/Deteriorated by pool: anchor on a "Loan Type" header ──
    by_pool: list[PoolMigrationRow] = []
    anchor = _find_label(ws, "Loan Type")
    if anchor:
        hr, hc = anchor
        r = hr + 1
        while r <= ws.max_row:
            pool = _cell_str(ws.cell(r, hc).value)
            if not pool:
                break
            by_pool.append(PoolMigrationRow(
                pool=pool,
                improved=_num(ws.cell(r, hc + 1).value),
                deteriorated=_num(ws.cell(r, hc + 2).value),
                net_change=_num(ws.cell(r, hc + 3).value),
            ))
            r += 1

    # ── Improved/Deteriorated by grade: anchor on a "Grade" header ──
    by_grade: list[GradeMigrationRow] = []
    ganchor = _find_label(ws, "Grade")
    if ganchor:
        hr, hc = ganchor
        r = hr + 1
        while r <= ws.max_row:
            grade = _cell_str(ws.cell(r, hc).value)
            if not grade:
                break
            by_grade.append(GradeMigrationRow(
                grade=grade,
                balance=_num(ws.cell(r, hc + 1).value),
                improved=_num(ws.cell(r, hc + 3).value),
                deteriorated=_num(ws.cell(r, hc + 4).value),
            ))
            r += 1

    return ImprDeterPage(
        credit_union=cu,
        period_ending=period,
        heading_lines=heading,
        cecl_adjustment=cecl,
        by_pool=by_pool,
        by_grade=by_grade,
    )


def _find_label(ws, text: str) -> tuple[int, int] | None:
    """Return (row, col) of the first cell whose stripped value == text."""
    low = text.strip().lower()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if _cell_str(ws.cell(r, c).value).lower() == low:
                return (r, c)
    return None


# ── Risk Change matrices ─────────────────────────────────────────────

_RISK_TAB_HINTS = ("risk change",)
# Fill colors (last 6 hex) -> cell state.
_FILL_STATE = {
    "0D4D5E": "header",
    "829901": "improved",
    "873A3A": "deteriorated",
}


def _fill_state(styled_cell) -> str:
    try:
        fill = styled_cell.fill
        if fill and fill.patternType:
            rgb = fill.fgColor.rgb
            if isinstance(rgb, str):
                return _FILL_STATE.get(rgb[-6:].upper(), "plain")
    except Exception:  # noqa: BLE001
        pass
    return "plain"


def _read_matrix(wsv, wsf, hr: int, *, is_pct: bool) -> RiskChangeMatrix:
    """Read one matrix whose corner ('$/% Current Grade') header is at
    row ``hr``, column 1. Column B (2) is the score range; the original-grade
    buckets run from C (3) up to the 'Grand Total' column (detected), which is
    the row total. The Deteriorated/Improved/Unchanged side panel is part of the
    workbook but NOT the printed report, so it is omitted.
    """
    corner = _cell_str(wsv.cell(hr, 1).value)
    # Find the Grand Total column (its header may sit on row hr or the merged
    # row above). Grade buckets are everything between C and it.
    gt_col = None
    for c in range(3, 25):
        h = (_cell_str(wsv.cell(hr, c).value)
             or _cell_str(wsv.cell(hr - 1, c).value))
        if h and h.lower() == "grand total":
            gt_col = c
            break
    if gt_col is None:
        gt_col = 12
    nbuckets = gt_col - 3
    col_headers = [
        _cell_str(wsv.cell(hr, c).value) for c in range(3, gt_col)
        if _cell_str(wsv.cell(hr, c).value)
    ]
    rows: list[MatrixRow] = []
    r = hr + 1
    while r <= wsv.max_row:
        label = _cell_str(wsv.cell(r, 1).value)
        if not label or label in ("Balance Adjustment", "Total in Portfolio"):
            break
        rng = _cell_str(wsv.cell(r, 2).value)
        is_total_row = label.lower() == "grand total"
        cells = [
            MatrixCell(
                value=_num(wsv.cell(r, c).value),
                state=_fill_state(wsf.cell(r, c)),
                is_pct=is_pct,
                bold=is_total_row,
            )
            for c in range(3, 3 + nbuckets)
        ]
        total = MatrixCell(
            value=_num(wsv.cell(r, gt_col).value), is_pct=is_pct, bold=True,
        )
        rows.append(MatrixRow(label=label, range_label=rng, cells=cells,
                              total=total, side=[]))
        r += 1
    return RiskChangeMatrix(
        corner=corner, col_headers=col_headers, rows=rows,
        side_headers=[], is_pct=is_pct,
    )


def load_risk_change(report_path: str | Path) -> RiskChangePage:
    """Read the 'Risk Change Total' tab into the model (both matrices)."""
    wbv = load_workbook(report_path, data_only=True)
    wbf = load_workbook(report_path, data_only=False)
    tab = _find_tab(wbv, _RISK_TAB_HINTS)
    if tab is None:
        raise ValueError(
            f"No Risk Change tab in {Path(report_path).name} "
            f"(sheets: {wbv.sheetnames})"
        )
    wsv, wsf = wbv[tab], wbf[tab]

    heading: list[str] = []
    cu = _cell_str(wsv.cell(1, 1).value)
    for r in range(2, 5):
        v = _cell_str(wsv.cell(r, 1).value)
        if v:
            heading.append(v)

    matrices: list[RiskChangeMatrix] = []
    for r in range(1, wsv.max_row + 1):
        label = _cell_str(wsv.cell(r, 1).value)
        if label in ("$ Current Grade", "% Current Grade"):
            matrices.append(_read_matrix(
                wsv, wsf, r, is_pct=label.startswith("%")))

    summary: list[KeyValueRow] = []
    for name in ("Balance Adjustment", "Total in Portfolio"):
        loc = _find_label(wsv, name)
        if loc:
            r, _c = loc
            # value sits in the Grand Total column — take the rightmost numeric.
            val = 0.0
            for c in range(wsv.max_column, 2, -1):
                cv = wsv.cell(r, c).value
                if isinstance(cv, (int, float)):
                    val = float(cv)
                    break
            summary.append(KeyValueRow(label=name, value=val))

    return RiskChangePage(
        credit_union=cu, heading_lines=heading,
        matrices=matrices, summary=summary,
    )


# ── ACL Env by Pool ──────────────────────────────────────────────────

_ACL_TAB_HINTS = ("acl env",)
_ACL_ADJ_LABELS = (
    "Total Specifically Identified Allowance",
    "Total Allowance Needed",
    "Allowance for Credit Loss Balance",
    "Adjustment (Overfunded)",
)


def _bold(wsf, r: int, c: int) -> bool:
    try:
        return bool(wsf.cell(r, c).font.bold)
    except Exception:  # noqa: BLE001
        return False


def load_acl_env(report_path: str | Path) -> AclEnvPage:
    """Read the 'ACL Env by Pool Mgmt Adj' tab into the model."""
    wbv = load_workbook(report_path, data_only=True)
    wbf = load_workbook(report_path, data_only=False)
    tab = _find_tab(wbv, _ACL_TAB_HINTS)
    if tab is None:
        raise ValueError(
            f"No ACL Env tab in {Path(report_path).name} "
            f"(sheets: {wbv.sheetnames})"
        )
    wsv, wsf = wbv[tab], wbf[tab]

    cu = _cell_str(wsv.cell(1, 1).value)
    heading = [_cell_str(wsv.cell(r, 1).value)
               for r in (2, 3) if _cell_str(wsv.cell(r, 1).value)]

    hdr = _find_label(wsv, "Current Grade")
    hr = hdr[0] if hdr else 5
    col_headers = [
        _cell_str(wsv.cell(hr, c).value).replace("\n", " ").strip()
        for c in range(1, 12) if _cell_str(wsv.cell(hr, c).value)
    ]

    def _row_vals(r: int) -> dict:
        g = lambda c: _num(wsv.cell(r, c).value)  # noqa: E731
        return dict(
            balance=g(2), specific_id=g(3), llc_balance=g(4),
            base_loss_rate=g(5), mgmt_adj=g(6), allowance_factor=g(7),
            allowance_before_env=g(8), env_factor=g(9),
            env_allowance=g(10), total_allowance=g(11),
        )

    pool_rows: list[AclPoolRow] = []
    pooled_totals: AclPoolRow | None = None
    impaired_rows: list[AdjustmentRow] = []
    adjustment_rows: list[AdjustmentRow] = []
    current_pool: str | None = None
    section = "pools"

    for r in range(hr + 1, wsv.max_row + 1):
        a = _cell_str(wsv.cell(r, 1).value)
        if not a:
            continue
        if a == "Pooled Totals":
            pooled_totals = AclPoolRow(pool=a, is_total=True, **_row_vals(r))
            section = "await_impaired"
        elif a == "Impaired Loans":
            section = "impaired"
        elif section == "pools":
            has_balance = _num(wsv.cell(r, 2).value) is not None
            if a == "Total":
                # Pool total row (bold), labelled "Total" as in Excel; the
                # pool name sits on the header row above it.
                pool_rows.append(AclPoolRow(
                    pool=a, kind="total", **_row_vals(r)))
            elif not has_balance:
                # A pool header row (name only, no Balance) — grade-rated
                # and single-line pools both start here.
                current_pool = a
                pool_rows.append(AclPoolRow(pool=a, kind="header"))
            else:
                # A per-grade detail row under the current pool.
                pool_rows.append(AclPoolRow(pool=a, kind="grade", **_row_vals(r)))
        elif section == "impaired":
            val = _num(wsv.cell(r, 11).value)  # column K
            bold = _bold(wsf, r, 1)
            row = AdjustmentRow(label=a, value=val, bold=bold)
            if any(a.startswith(lbl) for lbl in _ACL_ADJ_LABELS):
                adjustment_rows.append(row)
            else:
                impaired_rows.append(row)

    return AclEnvPage(
        credit_union=cu, heading_lines=heading, col_headers=col_headers,
        pool_rows=pool_rows, pooled_totals=pooled_totals,
        impaired_rows=impaired_rows, adjustment_rows=adjustment_rows,
    )


# ── Generic faithful grid (any labelled-grid tab) ────────────────────

def _hex6(color) -> str | None:
    """Return a 6-hex RGB string for an openpyxl color, or None.

    Skips black (default text) and theme/indexed colors we can't resolve
    to an RGB reliably.
    """
    try:
        rgb = color.rgb
    except Exception:  # noqa: BLE001
        return None
    if not isinstance(rgb, str) or len(rgb) < 6:
        return None
    h = rgb[-6:].upper()
    return None if h in ("000000",) else h


def _align(styled_cell, value) -> str:
    h = None
    try:
        h = styled_cell.alignment.horizontal
    except Exception:  # noqa: BLE001
        pass
    if h in ("left", "center", "right"):
        return h
    # Excel's "General": numbers right, text left.
    return "right" if isinstance(value, (int, float)) else "left"


def load_grid(report_path: str | Path, sheet: str,
              *, landscape: bool = False) -> GridPage:
    """Read an entire tab into a styled :class:`GridPage`.

    Every cell keeps its formatted text (per its Excel number format),
    background fill, font color, bold, and alignment, plus merged-cell
    spans — so labelled-grid tabs render faithfully with no bespoke
    model. Blank rows are preserved as spacers.
    """
    wbv = load_workbook(report_path, data_only=True)
    wbf = load_workbook(report_path, data_only=False)
    if sheet not in wbv.sheetnames:
        raise ValueError(f"Sheet {sheet!r} not in {Path(report_path).name}")
    wsv, wsf = wbv[sheet], wbf[sheet]

    # Effective extent: trim trailing empty columns/rows.
    max_r, max_c = 0, 0
    for r in range(1, wsv.max_row + 1):
        for c in range(1, wsv.max_column + 1):
            if wsv.cell(r, c).value not in (None, ""):
                max_r = max(max_r, r)
                max_c = max(max_c, c)
    if max_r == 0:
        return GridPage(credit_union="", sheet_name=sheet, rows=[])

    # Merged-cell handling: anchor -> (rowspan, colspan); covered cells skip.
    anchor_span: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for mr in wsf.merged_cells.ranges:
        anchor_span[(mr.min_row, mr.min_col)] = (
            mr.max_row - mr.min_row + 1, mr.max_col - mr.min_col + 1)
        for rr in range(mr.min_row, mr.max_row + 1):
            for cc in range(mr.min_col, mr.max_col + 1):
                if (rr, cc) != (mr.min_row, mr.min_col):
                    covered.add((rr, cc))

    cu = _cell_str(wsv.cell(1, 1).value)
    rows: list[list[GridCell]] = []
    for r in range(1, max_r + 1):
        row_cells: list[GridCell] = []
        for c in range(1, max_c + 1):
            if (r, c) in covered:
                continue
            v = wsv.cell(r, c).value
            sc = wsf.cell(r, c)
            fill = None
            try:
                if sc.fill and sc.fill.patternType:
                    fill = _hex6(sc.fill.fgColor)
            except Exception:  # noqa: BLE001
                fill = None
            f = sc.font
            rowspan, colspan = anchor_span.get((r, c), (1, 1))
            text = excel_format(v, sc.number_format)
            wrap = False
            try:
                wrap = bool(sc.alignment.wrap_text)
            except Exception:  # noqa: BLE001
                wrap = False
            wrap = wrap or ("\n" in text)
            row_cells.append(GridCell(
                text=text,
                align=_align(sc, v),
                bold=bool(f and f.bold),
                italic=bool(f and f.italic),
                fill=fill,
                color=_hex6(f.color) if f and f.color else None,
                size=float(f.size) if f and f.size else None,
                colspan=colspan,
                rowspan=rowspan,
                wrap=wrap,
            ))
        rows.append(row_cells)

    return GridPage(credit_union=cu, sheet_name=sheet, rows=rows,
                    landscape=landscape)



