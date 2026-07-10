"""Assemble a whole report: workbook -> all tabs -> one HTML doc / PDF.

Each worksheet is dispatched to its renderer — the bespoke templates for
the four special tabs (cover, Impr Deter, Risk Change, ACL Env) and the
generic faithful grid for everything else. Sheet order is preserved from
the workbook so the assembled document matches the report's tab order.

Two outputs:
  * :func:`render_report_html` — one scrollable HTML doc (browser preview).
  * :func:`render_report_pdf`  — per-tab PDFs (each in its correct
    orientation) merged into a single file, so wide tabs stay landscape.
"""

from __future__ import annotations

import io
from pathlib import Path

from . import from_workbook as fw
from . import render as R
from . import charts as C
from .model import CoverPage

# Wide tabs render landscape; everything else portrait. A grid tab also
# flips to landscape automatically when it has many columns (below).
_LANDSCAPE_HINTS = ("risk change", "acl env", "hist bal", "co-recov", "co recov")
_WIDE_COL_THRESHOLD = 12


def find_report(ws_root: str | Path, credit_union: str, snapshot: str,
                *, supplemental: bool = False) -> Path | None:
    """Locate the generated Vizo workbook for a CU/period."""
    safe_cu = credit_union.replace(" ", "_").replace("/", "-")
    kind = "CECL_Supplemental" if supplemental else "CECL_Migration"
    name = f"{snapshot}_{kind}_{safe_cu}_Vizo_Model.xlsx"
    path = Path(ws_root) / "Reports" / name
    if path.exists():
        return path
    # Fall back to a looser glob (handles minor CU-name punctuation drift).
    hits = sorted((Path(ws_root) / "Reports").glob(
        f"{snapshot}_{kind}_*Vizo_Model.xlsx"))
    return hits[0] if hits else None


def _fragment(full_html: str) -> str:
    """Extract the <body> inner HTML from a rendered page."""
    try:
        return full_html.split("<body>", 1)[1].rsplit("</body>", 1)[0]
    except IndexError:
        return full_html


def _render_sheet(report_path: Path, sheet: str, cu: str, snap: str,
                  *, supplemental: bool) -> tuple[str, bool]:
    """Return (full_html, landscape) for one worksheet."""
    low = sheet.strip().lower()
    landscape = any(h in low for h in _LANDSCAPE_HINTS)
    try:
        # Any embedded charts on this sheet, re-rendered as inline SVG.
        try:
            svgs = C.render_charts_for_sheet(report_path, sheet)
        except Exception:  # noqa: BLE001
            svgs = []
        if "cover" in low:
            cover = fw.load_cover(report_path, sheet)
            cover.credit_union = cover.credit_union or cu
            if supplemental:
                cover.subtitle = "Supplemental Reports"
            return R.render_html("cover.html", cover=cover), False
        if "impr deter" in low or ("improved" in low and "deteriorated" in low):
            page = fw.load_impr_deter(report_path)
            return R.render_html("impr_deter.html", page=page, charts=svgs), False
        if "risk change" in low:
            page = fw.load_risk_change(report_path)
            return R.render_html("risk_change.html", page=page, charts=svgs), True
        if "acl env" in low:
            page = fw.load_acl_env(report_path)
            return R.render_html("acl_env.html", page=page, charts=svgs), True
        # Generic faithful grid for every other tab.
        gpage = fw.load_grid(report_path, sheet, landscape=landscape)
        ncols = max((len(r) for r in gpage.rows), default=0)
        landscape = landscape or ncols > _WIDE_COL_THRESHOLD
        return R.render_html("grid.html", page=gpage, charts=svgs), landscape
    except Exception as exc:  # noqa: BLE001
        # Never let one bad tab kill the whole preview.
        return (f'<section class="page"><p style="color:#b00">'
                f'Could not render tab “{sheet}”: {exc}</p></section>'), landscape


def build_pages(report_path: str | Path, cu: str, snap: str,
                *, supplemental: bool = False) -> list[dict]:
    """Render every worksheet to a fragment + orientation, in tab order."""
    from openpyxl import load_workbook
    wb = load_workbook(report_path, read_only=True)
    pages: list[dict] = []
    for sheet in wb.sheetnames:
        full, landscape = _render_sheet(
            Path(report_path), sheet, cu, snap, supplemental=supplemental)
        pages.append({
            "sheet": sheet,
            "fragment": _fragment(full),
            "full": full,
            "landscape": landscape,
        })
    wb.close()
    return pages


def render_report_html(report_path: str | Path, cu: str, snap: str,
                       *, supplemental: bool = False) -> str:
    """One self-contained HTML doc with every tab (browser preview)."""
    pages = build_pages(report_path, cu, snap, supplemental=supplemental)
    css = R.load_css()
    body = "\n".join(p["fragment"] for p in pages)
    return (
        "<!DOCTYPE html><html lang=\"en\"><head><meta charset=\"utf-8\">"
        f"<title>{cu} — CECL Report {snap}</title>"
        f"<style>{css}</style></head><body>{body}</body></html>"
    )


def render_report_pdf(report_path: str | Path, cu: str, snap: str,
                      *, supplemental: bool = False) -> bytes:
    """Merge per-tab PDFs (each in its own orientation) into one file."""
    from pypdf import PdfWriter, PdfReader

    pages = build_pages(report_path, cu, snap, supplemental=supplemental)
    writer = PdfWriter()
    for p in pages:
        pdf_bytes = R.render_pdf(p["full"], landscape=p["landscape"])
        reader = PdfReader(io.BytesIO(pdf_bytes))
        for pg in reader.pages:
            writer.add_page(pg)
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()
