"""Period-over-period Change Analysis sheet for TCT and Vizo CECL reports.

Adds a final "Change Analysis" tab that compares the current report to the
most recent prior report for the same credit union / report format:

  * ACL total allowance variance for every pool
  * Impaired-loan reserve variance by impairment type
  * Variance in the pooled totals (Total Allowance Needed, ACL balance,
    over/under-funded adjustment)
  * Plain-language expert commentary on every pool (and the impaired book)
    whose reserve moved materially, explaining the primary driver.

Both ``report_tct.compose_tct`` and ``report_vizo.compose_vizo_main`` call
``append_change_analysis`` right before returning the workbook.
"""
import os
import re
import glob
import math

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

ACL_SHEET = "ACL Env by Pool Mgmt Adj"
SHEET_NAME = "Change Analysis"

# Non-pool labels that appear in column A of the ACL tab.
_SECTION_WORDS = {
    "current grade", "current risk rating", "impaired loans", "allowance",
    "amount at risk", "allowance %", "other provision considerations",
    "total specifically identified allowance", "total allowance needed",
    "allowance & provision for credit loss reserve analysis",
    "allowance & provision for loan loss reserve analysis",
}

# ── styling ──────────────────────────────────────────────────────────────
_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
F_TITLE = Font(name="Arial", size=14, bold=True)
F_SUB = Font(name="Arial", size=10, italic=True, color="595959")
F_HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_CELL = Font(name="Arial", size=10)
F_BOLD = Font(name="Arial", size=10, bold=True)
F_NOTE = Font(name="Arial", size=10)
FILL_HDR = PatternFill("solid", fgColor="305496")
FILL_TOT = PatternFill("solid", fgColor="D9E1F2")
FILL_FLAG = PatternFill("solid", fgColor="FCE4D6")
ACCT = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
PCT = '0.0%'
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right")
CENTER = Alignment(horizontal="center")


def _s(v):
    return str(v).strip() if v is not None else ""


def _num(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


# ── parsing ──────────────────────────────────────────────────────────────
def _parse_acl_sheet(ws):
    """Parse an ACL Env by Pool Mgmt Adj worksheet into a dict:
    {pools:{name:{balance,spec_id,total_allow}}, order:[...],
     impaired:{cat:allow}, totals:{...}}."""
    pools, order, impaired, totals = {}, [], {}, {}
    last_header = None
    phase = "pools"
    in_impaired = False
    for row in ws.iter_rows(values_only=True):
        a = _s(row[0])
        if not a:
            continue
        al = a.lower()
        if al.startswith("pooled total"):
            totals["pooled_balance"] = _num(row[1])
            totals["pooled_total_allow"] = _num(row[10])
            phase = "post"
            last_header = None
            continue
        if al == "impaired loans":
            in_impaired = True
            continue
        if al.startswith("total specifically identified"):
            totals["total_spec_allow"] = _num(row[10])
            in_impaired = False
            continue
        if al.startswith("total allowance needed"):
            totals["total_allow_needed"] = _num(row[10])
            continue
        if al.startswith("allowance for credit loss"):
            totals["acl_balance"] = _num(row[10])
            continue
        if al.startswith("adjustment"):
            totals["adjustment"] = _num(row[10])
            continue
        if in_impaired:
            if a.upper().startswith("HIDE"):
                continue
            val = _num(row[10]) or _num(row[9])
            impaired[a] = val
            continue
        if phase != "pools":
            continue
        if a == "Total" and last_header:
            pools[last_header] = {
                "balance": _num(row[1]),
                "spec_id": _num(row[2]),
                "total_allow": _num(row[10]),
            }
            if last_header not in order:
                order.append(last_header)
            last_header = None
            continue
        # Pool header candidate: label with an empty second column.
        if (row[1] in (None, "")) and al not in _SECTION_WORDS \
                and not al.startswith("total") and not a.upper().startswith("HIDE"):
            last_header = a
    return {"pools": pools, "order": order, "impaired": impaired, "totals": totals}


# ── locating the prior report ────────────────────────────────────────────
def _find_prior_report(rpt_dir, safe_cu, suffix, snap):
    """Return (path, snap_date) of the most recent report for this CU/format
    dated strictly before ``snap``, or (None, None)."""
    pattern = os.path.join(rpt_dir, f"*_CECL_Migration_{safe_cu}_{suffix}.xlsx")
    rx = re.compile(r"(\d{4}-\d{2}-\d{2})_CECL_Migration_"
                    + re.escape(safe_cu) + rf"_{re.escape(suffix)}\.xlsx$")
    best, best_date = None, None
    for path in glob.glob(pattern):
        m = rx.search(os.path.basename(path))
        if not m:
            continue
        d = m.group(1)
        if d >= snap:
            continue
        if best_date is None or d > best_date:
            best, best_date = path, d
    return best, best_date


# ── expert commentary ────────────────────────────────────────────────────
def _join(parts):
    parts = [p for p in parts if p]
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0]
    return ", ".join(parts[:-1]) + " and " + parts[-1]


def _is_significant(r):
    d = abs(r["delta"])
    if r["prior"] == 0 and r["cur"] != 0:
        return abs(r["cur"]) >= 1000
    if r["cur"] == 0 and r["prior"] != 0:
        return abs(r["prior"]) >= 1000
    if d >= 25000:
        return True
    if r["prior"] and abs(r["delta"] / r["prior"]) >= 0.15 and d >= 5000:
        return True
    return False


def _explain_pool(r, specific=False):
    pool, d, pa, ca = r["pool"], r["delta"], r["prior"], r["cur"]
    if pa == 0 and ca != 0:
        return (f"{pool} is newly reserved this period, adding "
                f"${ca:,.0f} to the required allowance.")
    if ca == 0 and pa != 0:
        return (f"{pool}'s required allowance was eliminated (down "
                f"${pa:,.0f}); the pool no longer carries a modeled reserve.")
    dirw = "increased" if d > 0 else "decreased"
    pct = f"{r['pct']:+.1%}" if r["pct"] is not None else "n/a"
    dbal = r["cur_bal"] - r["prior_bal"]
    dspec = r["cur_spec"] - r["prior_spec"]
    cur_calc = r["cur_bal"] - r["cur_spec"]
    prior_calc = r["prior_bal"] - r["prior_spec"]
    cur_rate = ca / cur_calc if cur_calc else 0.0
    prior_rate = pa / prior_calc if prior_calc else 0.0
    bal_component = prior_rate * (cur_calc - prior_calc)
    rate_component = d - bal_component

    drivers = []
    if abs(rate_component) >= abs(bal_component) and abs(rate_component) >= 0.2 * abs(d):
        if specific:
            # Commercial / specific-reserve pool: the reserve is set per loan
            # from the analyst's allowance schedule, not a modeled loss rate.
            drivers.append(f"{'higher' if d > 0 else 'lower'} specific reserves "
                           f"assigned to individual loans in this pool")
        elif cur_rate > prior_rate:
            drivers.append(f"a higher effective loss rate ({prior_rate:.2%} to "
                           f"{cur_rate:.2%}), reflecting credit-quality "
                           f"deterioration or heavier loss experience")
        else:
            drivers.append(f"a lower effective loss rate ({prior_rate:.2%} to "
                           f"{cur_rate:.2%}), reflecting improved credit quality "
                           f"or lighter loss experience")
    if abs(bal_component) >= 0.2 * abs(d) and abs(dbal) >= 1000:
        drivers.append(f"a ${abs(dbal):,.0f} "
                       f"{'increase' if dbal > 0 else 'decline'} in pool balance")
    if abs(dspec) >= max(10000, 0.15 * abs(d)):
        drivers.append(f"${abs(dspec):,.0f} "
                       f"{'more' if dspec > 0 else 'less'} in "
                       f"specifically-identified (impaired) balances")
    if not drivers:
        drivers.append("modest shifts in balance and loss rates")
    return (f"{pool} ACL {dirw} ${abs(d):,.0f} ({pct}), driven primarily by "
            f"{_join(drivers[:2])}.")


def _explain_impaired(cur_imp, prior_imp, cur_tot, prior_tot):
    d = cur_tot - prior_tot
    if abs(d) < 5000:
        return None
    cats = set(cur_imp) | set(prior_imp)
    moves = sorted(((c, cur_imp.get(c, 0) - prior_imp.get(c, 0)) for c in cats),
                   key=lambda x: -abs(x[1]))
    top = [f"{c} ({'+' if m > 0 else '-'}${abs(m):,.0f})"
           for c, m in moves[:2] if abs(m) >= 1000]
    lead = (f"Impaired-loan reserves {'rose' if d > 0 else 'fell'} "
            f"${abs(d):,.0f} to ${cur_tot:,.0f}")
    if top:
        return lead + ", concentrated in " + _join(top) + "."
    return lead + "."


# ── sheet builder ────────────────────────────────────────────────────────
def _hdr_row(ws, r, labels, widths=None):
    for c, lbl in enumerate(labels, start=1):
        cell = ws.cell(row=r, column=c, value=lbl)
        cell.font = F_HDR
        cell.fill = FILL_HDR
        cell.border = BORDER
        cell.alignment = CENTER if c == 1 else RIGHT


def append_change_analysis(wb, cu, snap, config, suffix):
    """Append the Change Analysis sheet. ``suffix`` = 'TCT_Model' or
    'Vizo_Model'. Never raises — on any problem it writes a short note."""
    try:
        cur = _parse_acl_sheet(wb[ACL_SHEET])
    except Exception as exc:  # noqa: BLE001
        cur = None

    ws = wb.create_sheet(SHEET_NAME)
    ws.sheet_view.showGridLines = False
    for col, w in (("A", 34), ("B", 18), ("C", 18), ("D", 16), ("E", 12)):
        ws.column_dimensions[col].width = w

    ws.cell(row=1, column=1, value=cu).font = F_TITLE
    ws.cell(row=2, column=1,
            value="Change Analysis — Period over Period").font = F_BOLD

    base = os.environ.get('CECL_WORKSPACE_ROOT') \
        or os.path.dirname(os.path.abspath(__file__))
    rpt_dir = os.path.join(base, 'Reports')
    safe_cu = cu.replace(' ', '_').replace('/', '-')
    prior_path, prior_snap = _find_prior_report(rpt_dir, safe_cu, suffix, snap)

    if not cur or not prior_path:
        msg = ("No prior report is available for comparison — this is the "
               "earliest report on file for this credit union."
               if not prior_path else
               "The current ACL detail could not be parsed for comparison.")
        ws.cell(row=3, column=1, value=f"Current period: {snap}").font = F_SUB
        ws.cell(row=5, column=1, value=msg).font = F_NOTE
        return

    ws.cell(row=3, column=1,
            value=f"Current: {snap}    |    Prior: {prior_snap}").font = F_SUB

    try:
        from openpyxl import load_workbook
        pw = load_workbook(prior_path, read_only=True, data_only=True)
        prior = _parse_acl_sheet(pw[ACL_SHEET])
        pw.close()
    except Exception as exc:  # noqa: BLE001
        ws.cell(row=5, column=1,
                value=f"Prior report could not be read: {exc}").font = F_NOTE
        return

    # Build per-pool rows.
    all_pools = list(cur["order"]) + [p for p in prior["order"]
                                      if p not in cur["order"]]
    rows = []
    for p in all_pools:
        c = cur["pools"].get(p, {})
        pr = prior["pools"].get(p, {})
        ca, pa = c.get("total_allow", 0.0), pr.get("total_allow", 0.0)
        rows.append({
            "pool": p, "cur": ca, "prior": pa, "delta": ca - pa,
            "pct": (ca - pa) / pa if pa else None,
            "cur_bal": c.get("balance", 0.0), "prior_bal": pr.get("balance", 0.0),
            "cur_spec": c.get("spec_id", 0.0), "prior_spec": pr.get("spec_id", 0.0),
        })

    r = 5
    ws.cell(row=r, column=1, value="Pool ACL Allowance Variance").font = F_BOLD
    r += 1
    _hdr_row(ws, r, ["Pool", "Current", "Prior", "$ Change", "% Change"])
    r += 1
    for row in rows:
        flag = _is_significant(row)
        ws.cell(row=r, column=1, value=row["pool"]).font = F_CELL
        ws.cell(row=r, column=2, value=row["cur"]).number_format = ACCT
        ws.cell(row=r, column=3, value=row["prior"]).number_format = ACCT
        ws.cell(row=r, column=4, value=row["delta"]).number_format = ACCT
        pc = ws.cell(row=r, column=5,
                     value=(row["pct"] if row["pct"] is not None else None))
        pc.number_format = PCT
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c > 1:
                cell.font = F_CELL
            if flag:
                cell.fill = FILL_FLAG
        r += 1
    # Pooled totals row.
    ctot = cur["totals"].get("pooled_total_allow", 0.0)
    ptot = prior["totals"].get("pooled_total_allow", 0.0)
    for c, val, fmt in ((1, "Pooled Total Allowance", None), (2, ctot, ACCT),
                        (3, ptot, ACCT), (4, ctot - ptot, ACCT),
                        (5, (ctot - ptot) / ptot if ptot else None, PCT)):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = F_BOLD
        cell.fill = FILL_TOT
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt
    r += 2

    # Impaired variance.
    ws.cell(row=r, column=1, value="Impaired Loan Reserve Variance").font = F_BOLD
    r += 1
    _hdr_row(ws, r, ["Impairment Type", "Current", "Prior", "$ Change", ""])
    r += 1
    cats = list(cur["impaired"].keys()) + [c for c in prior["impaired"]
                                           if c not in cur["impaired"]]
    for cat in cats:
        cv = cur["impaired"].get(cat, 0.0)
        pv = prior["impaired"].get(cat, 0.0)
        ws.cell(row=r, column=1, value=cat).font = F_CELL
        ws.cell(row=r, column=2, value=cv).number_format = ACCT
        ws.cell(row=r, column=3, value=pv).number_format = ACCT
        ws.cell(row=r, column=4, value=cv - pv).number_format = ACCT
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
            if c > 1:
                ws.cell(row=r, column=c).font = F_CELL
        r += 1
    cimp = cur["totals"].get("total_spec_allow",
                             sum(cur["impaired"].values()))
    pimp = prior["totals"].get("total_spec_allow",
                               sum(prior["impaired"].values()))
    for c, val, fmt in ((1, "Total Specifically Identified", None),
                        (2, cimp, ACCT), (3, pimp, ACCT), (4, cimp - pimp, ACCT)):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = F_BOLD
        cell.fill = FILL_TOT
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt
    r += 2

    # Summary totals.
    ws.cell(row=r, column=1, value="Summary").font = F_BOLD
    r += 1
    _hdr_row(ws, r, ["Metric", "Current", "Prior", "$ Change", ""])
    r += 1
    for label, key in (("Total Allowance Needed", "total_allow_needed"),
                       ("Allowance for Credit Loss Balance", "acl_balance"),
                       ("Adjustment (Over)/Under-funded", "adjustment")):
        cv = cur["totals"].get(key, 0.0)
        pv = prior["totals"].get(key, 0.0)
        ws.cell(row=r, column=1, value=label).font = F_CELL
        ws.cell(row=r, column=2, value=cv).number_format = ACCT
        ws.cell(row=r, column=3, value=pv).number_format = ACCT
        ws.cell(row=r, column=4, value=cv - pv).number_format = ACCT
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
            if c > 1:
                ws.cell(row=r, column=c).font = F_CELL
        r += 1
    r += 1

    # Expert commentary.
    ws.cell(row=r, column=1, value="Analysis of Significant Changes").font = F_BOLD
    r += 1
    notes = []
    specific_pools = {str(p).strip().lower()
                      for p in (config.get('warm_allowance_pools') or [])}
    for row in sorted(rows, key=lambda x: -abs(x["delta"])):
        if _is_significant(row):
            notes.append(_explain_pool(
                row, specific=row["pool"].strip().lower() in specific_pools))
    imp_note = _explain_impaired(cur["impaired"], prior["impaired"], cimp, pimp)
    if imp_note:
        notes.append(imp_note)
    # Overall total.
    tot_d = cur["totals"].get("total_allow_needed", 0.0) \
        - prior["totals"].get("total_allow_needed", 0.0)
    if abs(tot_d) >= 1000:
        notes.insert(0, f"Total Allowance Needed "
                        f"{'increased' if tot_d > 0 else 'decreased'} "
                        f"${abs(tot_d):,.0f} versus the prior report, to "
                        f"${cur['totals'].get('total_allow_needed', 0.0):,.0f}.")
    if not notes:
        notes.append("No pool reserve moved materially versus the prior "
                     "report; changes were within normal quarter-to-quarter "
                     "variation.")
    for n in notes:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        cell = ws.cell(row=r, column=1, value="•  " + n)
        cell.font = F_NOTE
        cell.alignment = LEFT_WRAP
        ws.row_dimensions[r].height = 14 * max(1, math.ceil(len(n) / 92)) + 4
        r += 1
