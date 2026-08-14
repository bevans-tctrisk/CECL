import os, sys
sys.path.insert(0, r'C:\dev\CECL')
import fitz  # PyMuPDF
from cecl_report_web import assembly

WP = r'Z:\Shared\TCT Files\CECL - CM Files\Web Printing'
XLSX = os.path.join(WP, '2026-06-30_CECL_Migration_Mountain_CU_Vizo_Model.xlsx')
REF_PDF = os.path.join(WP, '2026-06-30_CECL_Migration_Mountain_CU_Vizo_Model.pdf')
OUT = os.path.join(WP, '_compare')
MINE = os.path.join(OUT, 'mine'); REF = os.path.join(OUT, 'ref')
os.makedirs(MINE, exist_ok=True); os.makedirs(REF, exist_ok=True)

CU, SNAP = 'Mountain CU', '2026-06-30'

# sheet names in source
from openpyxl import load_workbook
wb = load_workbook(XLSX, read_only=True)
print('SOURCE sheets (%d):' % len(wb.sheetnames))
for s in wb.sheetnames:
    print('   ', s)
wb.close()

# render my PDF
print('\nrendering my PDF via cecl_report_web ...')
pdf_bytes = assembly.render_report_pdf(XLSX, CU, SNAP)
MY_PDF = os.path.join(OUT, 'mine.pdf')
open(MY_PDF, 'wb').write(pdf_bytes)
html = assembly.render_report_html(XLSX, CU, SNAP)
open(os.path.join(OUT, 'mine.html'), 'w', encoding='utf-8').write(html)
print('  wrote', MY_PDF, '(%d bytes)' % len(pdf_bytes))

def rasterize(pdf_path, outdir, tag):
    doc = fitz.open(pdf_path)
    print('%s: %d pages' % (tag, doc.page_count))
    for i, pg in enumerate(doc):
        pix = pg.get_pixmap(dpi=110)
        pix.save(os.path.join(outdir, '%02d.png' % i))
    doc.close()
    return

rasterize(REF_PDF, REF, 'REFERENCE PDF')
rasterize(MY_PDF, MINE, 'MY PDF')
print('\ncompare folders:\n  ', MINE, '\n  ', REF)
