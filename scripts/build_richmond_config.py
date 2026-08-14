import os, yaml
os.environ.setdefault('CECL_WORKSPACE_ROOT', r'Z:\Shared\TCT Files\CECL - CM Files')
WS = os.environ['CECL_WORKSPACE_ROOT']
SHORT = 'credit_union_of_richmond'
CLIENT_FOLDER = r'Z:\Shared\Clients\Vizo Financial Corporate CU\Client Access\Clients\Credit Union of Richmond 66929\Portfolio Management\CECL Migration IDLR'

# ---- pool map (from reference file) ----
import openpyxl
ref = openpyxl.load_workbook(CLIENT_FOLDER + r'\CU of Richmond Loan Type Codes with Pools.xlsx', data_only=True)['Assets']
pool_map = {}
for r in range(2, ref.max_row + 1):
    code = ref.cell(r, 2).value
    pool = ref.cell(r, 4).value
    if code is not None and pool:
        c = str(code).strip()
        p = str(pool).strip()
        pool_map[c] = 'Ignore' if p == 'Investments' else p
# also normalized numeric variants (map_pool_codes normalizes "08"->"8")
for c in list(pool_map):
    if c.isdigit():
        pool_map.setdefault(str(int(c)), pool_map[c])
# static-pool identity keys (loan_pool_code_static value must be a pool_map key)
pool_map['Credit Cards'] = 'Credit Cards'

POOLS = ['New Auto', 'Used Auto', 'CUDL Auto Loans', 'First Mortgage',
         'Second Mortgages and HELOC', 'Other Secured', 'Other Unsecured', 'Credit Cards']
RE_POOLS = {'First Mortgage', 'Second Mortgages and HELOC'}
# Pools with ~0 scored balance (no FICO) -> single-line, not grade-rated
NOT_RISK_RATED = {'Credit Cards'}

def pool_entry(name):
    return {'name': name, 'risk_rated': name not in NOT_RISK_RATED,
            'acl_months': 84 if name in RE_POOLS else 36,
            'excluded': False,
            'use_default_mgmt_adj': name in RE_POOLS, 'brr': False}

pools = [pool_entry(p) for p in POOLS] + [
    {'name': 'Ignore', 'risk_rated': False, 'acl_months': None, 'excluded': False,
     'use_default_mgmt_adj': False, 'brr': False}]

grades = [
    {'label': 'A+', 'min_score': 770, 'max_score': 900, 'reserve_rate': 0.0011},
    {'label': 'A',  'min_score': 720, 'max_score': 769, 'reserve_rate': 0.0025},
    {'label': 'B',  'min_score': 680, 'max_score': 719, 'reserve_rate': 0.0050},
    {'label': 'C',  'min_score': 601, 'max_score': 679, 'reserve_rate': 0.0116},
    {'label': 'D',  'min_score': 561, 'max_score': 600, 'reserve_rate': 0.0250},
    {'label': 'E',  'min_score': 350, 'max_score': 560, 'reserve_rate': 0.0500},
]

aires_cm = {
    'member_number': 'Account Number', 'current_balance': 'LOAN BALANCE',
    'loan_pool_code': 'LOAN TYPE', 'original_fico_score': 'CREDIT SCORE',
    'current_fico_score': 'CREDIT SCORE', 'open_date': 'LOAN ORIGINAL DATE',
    'original_loan_amount': 'LOAN ORIGINAL AMNT', 'interest_rate': 'Annual Rate 1',
    'total_available_credit': 'CREDIT LIMIT',
}
aires_ma = {'mode': 'delimiter', 'suffix_length': 0, 'delimiter': 'L'}

AIRES_PAT = r'(?i)^(?!credit\s*card)(?!airesv2)(?!.*\bshares?\b)(?=.*aires).*\.(xlsx|xls)$'
AIRESV2_PAT = r'(?i)^airesv2\b.*\.(xlsx|xls)$'
CC_PAT = r'(?i)^credit\s*cards?\b.*\.(xlsx|xls)$'
NS_PAT = r'(?i)^negative\s+shares?\b.*\.(xlsx|xls)$'

# v2 AIRES layout (Feb 2026+): named columns interleaved with blank tab cols
aires_v2_cm = {
    'member_number': 'Account Number', 'current_balance': 'Current Balance',
    'loan_pool_code': 'Loan Type Code', 'original_fico_score': 'FICO',
    'current_fico_score': 'FICO', 'open_date': 'Origination Date',
    'original_loan_amount': 'Original Balance', 'interest_rate': 'Int Rate',
    'total_available_credit': 'Credit Limit', 'days_delinquent': 'Days Delinquent',
}
# v2 loan-type codes use a different scheme than the legacy AIRES 'LOAN TYPE'.
# Derived empirically (member+balance match to prior old-format months); see
# scripts/derive_richmond_v2_codes.py. Applied only to the v2 extract.
aires_v2_pool_map = {
    '3': 'Used Auto', '99': 'Used Auto',
    '1': 'New Auto',
    '10': 'Other Unsecured', '50': 'Other Unsecured',
    '22': 'First Mortgage', '23': 'First Mortgage',
    '20': 'Second Mortgages and HELOC', '52': 'Second Mortgages and HELOC',
    '6': 'Other Secured', '8': 'Other Secured', '9': 'Other Secured',
    'Credit Cards': 'Credit Cards',
}

config = {
    'credit_union': 'Credit Union of Richmond',
    'charter_number': '66929',
    'report_period': '2025-12',
    'file_pattern': AIRES_PAT,
    'date_pattern': r'(\d{1,2})-(\d{1,2})-(20\d{2})',
    'date_format': 'MMDDYYYY',
    'account_suffix_length': 0,
    'member_account': aires_ma,
    'has_header': True,
    'column_mappings': aires_cm,
    'balance_format': {'remove_chars': ['$'], 'accounting_negatives': True},
    'pool_code_split': '',
    'pool_map': pool_map,
    'default_pool': 'Ignore',
    'credit_grades': grades,
    'no_score_label': 'Not Reported',
    'reports': {'tct': False, 'vizo': True, 'vizo_supp': True, 'impdet': True},
    'economic_data': {'state': 'Virginia', 'county': 'Richmond City',
                      'unemployment_rate': 0.031, 'foreclosures': 1500,
                      'bankruptcies': 4000, 'population': 226000},
    'mgmt_adj': {'ltv_baseline': 0.9, 'probability_factor': 0.35},
    # Derive current credit scores from each member's most-recently-opened
    # loan's origination score (no separate bureau pull) -> credit migration.
    'credit_pull': {'prefer_original_for_new_loans': True},
    # Pull the booked ACL (allowance) balance from NCUA 5300 (field AAS0048).
    'acl': {'use_5300_fallback': True},
    'loan_data_extracts': [
        {'label': 'Credit Cards', 'file_pattern': CC_PAT,
         'column_mappings': {'member_number': 'Account Identifier -PK',
                             'current_balance': 'Account Balance Total Amount',
                             'loan_pool_code_static': 'Credit Cards',
                             'total_available_credit': 'Credit Line Amount',
                             'days_delinquent': 'Account Cycle Delinquent Count'},
         'member_account': {'mode': 'fixed_suffix', 'suffix_length': 0, 'delimiter': '-'},
         'has_header': True},
        {'label': 'Negative Shares', 'file_pattern': NS_PAT,
         'column_mappings': {'member_number': 'Acct_Type',
                             'current_balance': 'Current Balance',
                             'loan_pool_code': 'Loan Type Code'},
         'member_account': {'mode': 'fixed_suffix', 'suffix_length': 3, 'delimiter': '-'},
         'has_header': True},
        {'label': 'AIRES v2', 'file_pattern': AIRESV2_PAT,
         'column_mappings': aires_v2_cm,
         'member_account': {'mode': 'fixed_suffix', 'suffix_length': 3, 'delimiter': 'L'},
         'pool_map': aires_v2_pool_map, 'has_header': True},
        {'label': 'AIRES', 'file_pattern': AIRES_PAT,
         'column_mappings': aires_cm, 'member_account': aires_ma, 'has_header': True},
    ],
    'pools': pools,
    'pool_order': POOLS + ['Ignore'],
    'not_risk_rated': sorted(NOT_RISK_RATED),
    'acl_months_by_pool': {p: (84 if p in RE_POOLS else 36) for p in POOLS},
    'historical_file_formats': {
        'chargeoff': {'has_header': True, 'skip_rows': 0, 'account_col': 0,
                      'code_col': 1, 'amount_col': 4, 'date_col': 5, 'member_col': 0,
                      'member_account': {'mode': 'delimiter', 'suffix_length': 0, 'delimiter': 'L'}},
        'recovery': {'has_header': True, 'skip_rows': 0, 'account_col': 0,
                     'code_col': 1, 'amount_col': 3, 'date_col': 5, 'member_col': 0,
                     'member_account': {'mode': 'delimiter', 'suffix_length': 0, 'delimiter': 'L'}},
    },
    'data_directory': os.path.join(WS, 'Raw_Uploads', SHORT),
    'loan_source_folder': CLIENT_FOLDER,
    'monthly_balance': {'source': 'single'},
}

os.makedirs(config['data_directory'], exist_ok=True)
out = os.path.join(WS, 'client_configs', SHORT + '.yaml')
with open(out, 'w', encoding='utf-8') as f:
    yaml.safe_dump(config, f, sort_keys=False, default_flow_style=False, allow_unicode=True)
print('WROTE', out)
print('pools:', POOLS)
print('pool_map codes:', len(pool_map))
# validate parse
c2 = yaml.safe_load(open(out, encoding='utf-8'))
print('reparse OK; extracts:', [e['label'] for e in c2['loan_data_extracts']])
