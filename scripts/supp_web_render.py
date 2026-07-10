import sys, os
sys.path.insert(0, r'C:\dev\CECL')
import fitz
from openpyxl import load_workbook
from cecl_report_web import assembly

XLSX = r'Z:\Shared\TCT Files\CECL - CM Files\Reports\2026-06-30_CECL_Supplemental_Mountain_CU_Vizo_Model.xlsx'
OUT = r'Z:\Shared\TCT Files\CECL - CM Files\Web Printing\_compare_supp'
os.makedirs(OUT, exist_ok=True)

wb = load_workbook(XLSX, read_only=True)
print('SUPPLEMENTAL sheets (%d):' % len(wb.sheetnames))
for s in wb.sheetnames:
    print('   ', s)
wb.close()

print('\nrendering supplemental PDF ...')
pdf = assembly.render_report_pdf(XLSX, 'Mountain CU', '2026-06-30', supplemental=True)
p = os.path.join(OUT, 'supp_mine.pdf'); open(p, 'wb').write(pdf)
doc = fitz.open(p)
print('pages:', doc.page_count)
for i in range(doc.page_count):
    doc[i].get_pixmap(dpi=100).save(os.path.join(OUT, '%02d.png' % i))
doc.close(); print('rasterized ->', OUT)
