"""Diagnostic: compute HLOOKUP-driven per-pool totals for Greensboro Dec vs Mar."""
import openpyxl

WB_M = r'Z:\Shared\TCT Files\CECL - CM Files\Generated_Reports\greensboro_cu\2026-03\2026-03_CECL_SCALE_greensboro_cu_Vizo.xlsx'
WB_D = r'Z:\Shared\TCT Files\CECL - CM Files\Generated_Reports\greensboro_cu\2025-12\2025-12_CECL_SCALE_greensboro_cu.xlsx'

wb_m = openpyxl.load_workbook(WB_M, data_only=False)
wb_d = openpyxl.load_workbook(WB_D, data_only=False)
hd_d = wb_d['Historical Data']
hd_m = wb_m['Historical Data']


def compute_sumifs(ws, col_letter):
    col_idx = openpyxl.utils.column_index_from_string(col_letter)
    sums = {}
    for r in range(9, 35):
        label = ws.cell(row=r, column=3).value
        val = ws.cell(row=r, column=col_idx).value
        if label and isinstance(val, (int, float)):
            sums[label] = sums.get(label, 0) + val
    return sums


print('=== MAR workbook ===')
ay = compute_sumifs(hd_m, 'AY')
az = compute_sumifs(hd_m, 'AZ')
all_pools = set(list(ay.keys()) + list(az.keys()))
for p in sorted(all_pools):
    a = ay.get(p, 0)
    b = az.get(p, 0)
    print(f'{str(p)[:38]:38s}  AY={a:>14,.0f}  AZ={b:>14,.0f}  delta={b - a:>+12,.0f}')
print(f'TOTAL AY={sum(ay.values()):>14,.0f}  AZ={sum(az.values()):>14,.0f}')

print()
print('=== Per HLOOKUP row 162-174 (SUMIF target = C162-C174) ===')
for r in range(162, 175):
    label = hd_m.cell(row=r, column=3).value
    ay_v = ay.get(label, 0)
    az_v = az.get(label, 0)
    print(f'r{r}  {str(label)[:32]:32s}  AY={ay_v:>14,.0f}  AZ={az_v:>14,.0f}  delta={az_v - ay_v:>+12,.0f}')

# Also check what the env factor / loss rate columns look like
print()
print('=== Calc tab AE6:AE24 vs DEC ===')
for r in range(5, 25):
    d_v = wb_d['Calc tab'].cell(row=r, column=31).value
    m_v = wb_m['Calc tab'].cell(row=r, column=31).value
    if d_v or m_v:
        same = '=' if d_v == m_v else 'DIFF'
        print(f'AE{r}  {same}  DEC={d_v!r}  MAR={m_v!r}')

# Now look at the "Scale Calculation-Vizo" tab — the main display
print()
print('=== Scale Calculation-Vizo tab full formula dump (rows 7-30) ===')
ws_m = wb_m['Scale Calculation-Vizo']
ws_d = wb_d['Scale Calculation-Vizo']
for r in range(1, min(ws_m.max_row + 1, 35)):
    for c in range(1, min(ws_m.max_column + 1, 18)):
        m_v = ws_m.cell(row=r, column=c).value
        d_v = ws_d.cell(row=r, column=c).value
        if m_v is not None or d_v is not None:
            cl = openpyxl.utils.get_column_letter(c)
            same = '=' if m_v == d_v else 'DIFF'
            print(f'{cl}{r}  {same}  DEC={str(d_v)[:55]!r}  MAR={str(m_v)[:55]!r}')
