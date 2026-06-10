"""Prototype the pool-extractor for 'Balance Sheets <CU> CU' workbooks.

Run from c:\\Dev\\CECL with C:\\cecl-venv\\Scripts\\python.exe.
"""
from __future__ import annotations

import re
from pathlib import Path
import openpyxl

# Sub-category code: "NNN-NN" in col A (e.g. "708-01", "702-03").
_SUBCAT_CODE_RX = re.compile(r"^\d{3}-\d{2}$")

# Grand-total row: col B is exactly "Total Loans" (no continuation).
_GRAND_TOTAL_RX = re.compile(r"^\s*total\s+loans\s*$", re.IGNORECASE)


def _cell_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _is_bold(cell) -> bool:
    try:
        return bool(cell.font and cell.font.bold)
    except Exception:
        return False


def extract_pools_from_balance_workbook(xlsx_path: str | Path) -> dict:
    """Walk a 'Balance Sheets <CU>' workbook and pull pools + sub-categories.

    Convention learned from Destinations CU:
      * Col A holds either an INTEGER (pool number) or a "NNN-NN" code
        (sub-category NCUA loan code) or is empty.
      * Col B holds the label (pool name OR sub-category name).
      * BOLD col B with INTEGER col A => POOL row.
      * BOLD col B with EMPTY col A and leading whitespace in the label
        => continuation of previous pool's name.
      * "NNN-NN" col A => sub-category row; flows into the NEXT pool row.
      * "Total Loans" row (B only, no children) => grand total. Skip.

    Returns::
        {
            "ok": True,
            "sheet": str,
            "pools": [
                {"name": "Unsecured Credit Card Loans",
                 "subcategories": [
                     {"code": "708-01", "label": "Mastercard Loans"}
                 ]},
                ...
            ],
            "subcategory_to_pool": {
                "Mastercard Loans": "Unsecured Credit Card Loans",
                "Signature Loans": "All Other Unsecured Loans/LOC",
                ...
            },
            "code_to_pool": {
                "708-01": "Unsecured Credit Card Loans",
                ...
            },
        }
    """
    path = Path(xlsx_path)
    if not path.exists():
        return {"ok": False, "error": f"file not found: {path}"}
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"openpyxl failed: {exc}"}

    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    pools: list[dict] = []
    subcat_buf: list[dict] = []
    pending_pool_name: str | None = None

    for r in range(1, (ws.max_row or 0) + 1):
        a_cell = ws.cell(row=r, column=1)
        b_cell = ws.cell(row=r, column=2)
        a_raw = a_cell.value
        b_raw = b_cell.value

        a_txt = _cell_text(a_raw)
        b_txt_raw = "" if b_raw is None else str(b_raw)  # KEEP leading WS for continuation
        b_txt = b_txt_raw.strip()

        if not a_txt and not b_txt:
            continue

        # Grand-total row -> stop processing
        if not a_txt and _GRAND_TOTAL_RX.match(b_txt):
            break

        # Sub-category row (NCUA loan code in col A)
        if _SUBCAT_CODE_RX.match(a_txt):
            if b_txt:
                subcat_buf.append({"code": a_txt, "label": b_txt})
            continue

        # Pool row: BOLD col B, with integer in col A OR (empty col A AND
        # leading whitespace in label = continuation)
        if _is_bold(b_cell) and b_txt:
            is_continuation = (
                not a_txt
                and (b_txt_raw != b_txt_raw.lstrip())  # had leading whitespace
            )
            is_integer_a = False
            if a_txt:
                try:
                    int(float(a_txt))
                    is_integer_a = True
                except (ValueError, TypeError):
                    is_integer_a = False

            if is_continuation and pending_pool_name:
                # Append to the in-progress pool name
                pending_pool_name = f"{pending_pool_name} {b_txt}".strip()
                # Re-stamp the just-added pool's name
                if pools and pools[-1].get("_open"):
                    pools[-1]["name"] = pending_pool_name
                continue

            if is_integer_a:
                # Close any previously-open pool (no more continuations)
                if pools and pools[-1].get("_open"):
                    pools[-1]["_open"] = False
                # Start a new pool
                pool = {
                    "name": b_txt,
                    "subcategories": list(subcat_buf),
                    "_open": True,
                }
                pools.append(pool)
                pending_pool_name = b_txt
                subcat_buf = []
                continue

        # Anything else: drop quietly (header rows, blank labels, formulas)
        # Keep accumulating sub-categories.

    # Strip the bookkeeping flag
    for p in pools:
        p.pop("_open", None)

    # Build the flat lookup dicts
    subcat_to_pool: dict[str, str] = {}
    code_to_pool: dict[str, str] = {}
    for p in pools:
        for sc in p.get("subcategories") or []:
            subcat_to_pool[sc["label"]] = p["name"]
            code_to_pool[sc["code"]] = p["name"]

    return {
        "ok": True,
        "sheet": sheet_name,
        "pools": pools,
        "subcategory_to_pool": subcat_to_pool,
        "code_to_pool": code_to_pool,
    }


if __name__ == "__main__":
    p = r"Z:\Shared\Clients\Vizo Financial Corporate CU\Client Access\Clients\Destinations CU 66333\Portfolio Management\CECL Migration IDLR\Balance Sheets Destinations CU.xlsx"
    result = extract_pools_from_balance_workbook(p)
    print("OK:", result["ok"])
    if not result["ok"]:
        print("ERROR:", result.get("error"))
        raise SystemExit(1)
    print(f"Sheet: {result['sheet']!r}")
    print(f"Pools detected: {len(result['pools'])}")
    print()
    for i, pool in enumerate(result["pools"], 1):
        print(f"  Pool {i}: {pool['name']!r}")
        for sc in pool["subcategories"]:
            print(f"     {sc['code']}  {sc['label']}")
        print()
    print(f"Sub-cat -> pool mapping: {len(result['subcategory_to_pool'])} entries")
    print(f"Code -> pool mapping:    {len(result['code_to_pool'])} entries")
