"""Add May & June 2026 (and refresh April 2026) to Lanco's monthly loan
balance workbook from the CU's GL Balance Sheets file
('Balance Sheets - April to June.xlsx', one sheet per month).

Uses the SAME methodology as scripts/build_lanco_monthly_bal.py (GL account
-> pool map, 702100 MBL Mortgage split RE/Business by the Dec-2025 Aires
commercial-mtg ratio, participation-sold contras netted into their pool),
EXTENDED to cover 3 loan accounts the original map dropped:
  702075 COMMERCIAL BLANKET MTG TERM  -> Business Loans  (new in Apr-2026, $1.568M)
  703150 TAKE BACK USED POWERSPORT    -> Consumer Secured (like other take-back veh)
  704104 CASH ADV & BAL CONSOLIDATION -> Credit Cards      (immaterial)
Each written month reconciles pool_sum == GL NET LOANS + CECL + ODP to the penny.
"""
import os, calendar, shutil, datetime
import openpyxl, pandas as pd

CLIENT = (r'Z:\Shared\Clients\Vizo Financial Corporate CU\Client Access\Clients'
          r'\Lanco FCU 16657\Portfolio Management\CECL Migration IDLR')
NEWF = CLIENT + r'\2026\2026\June 2026\Balance Sheets - April to June.xlsx'
WB_PATH = r'Z:\Shared\TCT Files\CECL - CM Files\Raw_Uploads\lanco_fcu\Monthly Loan Balances by Type.xlsx'

def num(v):
    if v is None: return 0.0
    if isinstance(v,(int,float)): return float(v)
    s=str(v).strip().replace(',','').replace('$','')
    neg=(s.startswith('<') and s.endswith('>')) or (s.startswith('(') and s.endswith(')'))
    s=s.strip('<>()')
    try: return -float(s) if neg else float(s)
    except ValueError: return 0.0

# 702100 RE/Business split ratio from Dec-2025 Aires (build-script method; matches existing columns)
a = pd.read_excel(CLIENT + r'\2025\Dec 2025\Aires Loan Data 12-31-2025 WO Charge-offs.xlsx', header=0)
a.columns = [str(c).strip() for c in a.columns]
bal = pd.to_numeric(a['CURRENT BAL'], errors='coerce').fillna(0.0)
code = a['Loan Type Code'].astype(str).str.strip()
re_cm = bal[code.isin({'100','101'})].sum(); bus_cm = bal[code.isin({'110','111','112','113','114','115'})].sum()
RE_FRAC = re_cm/(re_cm+bus_cm) if (re_cm+bus_cm) else 0.6
print(f"702100 split RE_FRAC (Dec-2025 Aires) = {RE_FRAC:.4f}")

GL_POOL = {
    701000:'Consumer Unsecured',701001:'Consumer Secured',
    701100:'Real Estate',701110:'Real Estate',701120:'Real Estate',
    701130:'Real Estate',701140:'Real Estate',701150:'Real Estate',
    701200:'Consumer Secured',701300:'Consumer Secured',701400:'Consumer Secured',
    701450:'Consumer Indirect',701500:'Consumer Indirect',
    701600:'Consumer Unsecured',701700:'Consumer Unsecured',
    701900:'Consumer Unsecured',701950:'Consumer Unsecured',
    702000:'Business Loans',702025:'Business Loans',702050:'Business Loans',
    702075:'Business Loans',                    # extended
    702110:'Real Estate',
    702111:'Business Loans',702112:'Business Loans',702114:'Business Loans',
    702119:'Business Loans',702120:'Business Loans',702130:'Business Loans',
    702140:'Business Loans',
    703000:'Business Loans',703141:'Consumer Unsecured',703143:'Real Estate',
    703144:'Consumer Secured',703145:'Consumer Unsecured',
    703146:'Consumer Secured',703148:'Consumer Secured',
    703150:'Consumer Secured',                  # extended
    704101:'Credit Cards',704102:'Credit Cards',
    704104:'Credit Cards',                      # extended
    704110:'Student Loans',704120:'Student Loans',
    704140:'Real Estate',704150:'Real Estate',704160:'Real Estate',704170:'Real Estate',
}
POOLS = ['Real Estate','Consumer Secured','Consumer Indirect','Consumer Unsecured',
         'Credit Cards','Student Loans','Business Loans']

def sheet_to_date(sn):
    p = sn.split('.')
    if len(p) == 2: m, y = int(p[0]), int(p[1])
    else: m, y = int(p[0]), 2000+int(p[2])
    return '%04d-%02d-%02d' % (y, m, calendar.monthrange(y, m)[1])

nwb = openpyxl.load_workbook(NEWF, data_only=True)
data = {}
for sn in nwb.sheetnames:
    d = sheet_to_date(sn); ws = nwb[sn]
    by = {p:0.0 for p in POOLS}; net=cecl=odp=0.0
    for row in ws.iter_rows(values_only=True):
        try: acc = int(str(row[0]).strip())
        except (ValueError,TypeError): continue
        v = num(row[2])
        if acc==720999: net=v; continue
        if acc==719000: cecl=v; continue
        if acc==720000: odp=v; continue
        if v==0: continue
        if acc==702100:
            by['Real Estate'] += v*RE_FRAC; by['Business Loans'] += v*(1-RE_FRAC)
        elif acc in GL_POOL:
            by[GL_POOL[acc]] += v
    gross = net - cecl - odp
    assert abs(sum(by.values()) - gross) < 0.01, (d, sum(by.values()), gross)
    data[d] = {p: round(by[p],2) for p in POOLS}
    print(f"  {d}: reconciled to GL NET LOANS (${gross:,.2f})")

# ---- update the workbook ----
shutil.copyfile(WB_PATH, WB_PATH + '.bak.phase_junebs_' + datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
wb = openpyxl.load_workbook(WB_PATH)
ws = wb['Sheet1']
# map pool name -> row
pool_row = {ws.cell(r,1).value: r for r in range(2, ws.max_row+1)}
# map existing date header -> col
date_col = {str(ws.cell(1,c).value)[:10]: c for c in range(2, ws.max_column+1)}
next_col = ws.max_column + 1
for d in ['2026-04-30','2026-05-31','2026-06-30']:
    if d in date_col:
        c = date_col[d]; action='refreshed'
    else:
        c = next_col; next_col += 1
        ws.cell(1, c).value = d
        action='added'
    for p in POOLS:
        ws.cell(pool_row[p], c).value = data[d][p]
    print(f"  {action} column {d} (col {c})")
wb.save(WB_PATH)
print("SAVED", WB_PATH)
# show final headers tail + June column
hdr = [str(ws.cell(1,c).value)[:10] for c in range(2, ws.max_column+1)]
print("final months tail:", hdr[-5:])
print("June 2026 by pool:")
for p in POOLS:
    print(f"   {p:<20} {data['2026-06-30'][p]:>16,.2f}")
