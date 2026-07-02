"""Verify Nucor Spec ID column is grade-specific after Turn 8 fix."""
from openpyxl import load_workbook
p = r'Z:\Shared\TCT Files\CECL - CM Files\Reports\2026-06-30_CECL_Migration_Nucor_Emp_CU_Vizo_Model.xlsx'
wb = load_workbook(p, data_only=True)
ws = wb['ACL Env by Pool Mgmt Adj']
print(f"Sheet dims: {ws.dimensions}")
# Find the Unsecured Loans pool section
for row in range(1, ws.max_row + 1):
    a = ws.cell(row=row, column=1).value
    b = ws.cell(row=row, column=2).value
    c = ws.cell(row=row, column=3).value
    if a and 'unsecured' in str(a).lower():
        # Dump 12 rows starting here
        for r2 in range(row, min(row + 15, ws.max_row + 1)):
            vals = [ws.cell(row=r2, column=col).value for col in range(1, 6)]
            print(f"R{r2}: {vals}")
        print("---")
        break
# Also look for any 'Specific' header rows to sanity-check
for row in range(1, ws.max_row + 1):
    c = ws.cell(row=row, column=3).value
    if c and 'specific' in str(c).lower():
        print(f"Spec ID header at R{row}: {c!r}")
        break
