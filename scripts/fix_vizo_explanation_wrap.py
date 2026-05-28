"""Fix wrap-text and row-height on the Vizo Explanation-of-ACL-Calc tab.

Issue:
- Rows 19 and 23 contain long paragraphs but wrap_text=False, so the
  text spills into adjacent columns / gets visually truncated when
  printed or viewed in Vizo's standard column layout.
- Rows 21 and 25 already have wrap_text=True but their row heights
  (40.15 and 38.45) are too short to show every wrapped line — the
  bottom row of letters gets clipped.

Fix:
- Set wrap_text=True + vertical='top' on A19/A21/A23/A25.
- Recompute row height based on character count vs. column-A width so
  every wrapped line is fully visible, with a small padding line.

Applies to every template under cecl_ui/data/scale/templates/.
"""
from __future__ import annotations

import math
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

TEMPLATES_DIR = Path(r"C:\Dev\CECL\cecl_ui\data\scale\templates")
SHEET = "Explanation of ACL Calc-Vizo"
TARGET_ROWS = (19, 23, 25)
COLUMN = "A"

# Empirical: at column-A width ~124 with 11pt default font, ~110 chars
# fit on one wrapped line in Excel's display. Using a slightly
# pessimistic 105 plus a small padding gives every line a bit of
# breathing room without ballooning the row height.
CHARS_PER_LINE = 105
LINE_HEIGHT = 15.75   # one line at 11pt default
PADDING_LINES = 0.3   # extra breathing room so descenders aren't clipped


def compute_height(text: str) -> float:
    lines = max(1, math.ceil(len(text) / CHARS_PER_LINE))
    return round((lines + PADDING_LINES) * LINE_HEIGHT, 2)


def fix_template(path: Path, *, dry_run: bool = False) -> None:
    wb = load_workbook(path, data_only=False)
    if SHEET not in wb.sheetnames:
        print(f"  SKIP (no '{SHEET}' tab): {path.name}")
        wb.close()
        return

    ws = wb[SHEET]
    changed: list[str] = []
    for row in TARGET_ROWS:
        cell = ws[f"{COLUMN}{row}"]
        text = "" if cell.value is None else str(cell.value)
        if not text.strip():
            continue

        # Preserve existing horizontal alignment; set wrap + top-align.
        existing = cell.alignment or Alignment()
        new_align = Alignment(
            horizontal=existing.horizontal,
            vertical="top",
            wrap_text=True,
            indent=existing.indent,
            shrink_to_fit=False,
            text_rotation=existing.text_rotation,
        )
        before_wrap = bool(existing.wrap_text)
        cell.alignment = new_align

        needed = compute_height(text)
        rd = ws.row_dimensions[row]
        current = rd.height or 0
        # Only grow the row -- never shrink (the analyst may have
        # manually made it taller, and we don't want to undo that).
        if needed > current + 0.1:
            rd.height = needed
            grew = f"{current:.2f}->{needed:.2f}"
        else:
            grew = f"unchanged ({current:.2f}, needed {needed:.2f})"

        changed.append(
            f"    A{row}: wrap {before_wrap}->True, height {grew}, "
            f"chars={len(text)}"
        )

    if not changed:
        print(f"  NO-OP: {path.name}")
        wb.close()
        return

    print(f"  {path.name}")
    for line in changed:
        print(line)

    if dry_run:
        wb.close()
        return

    # Backup once per run so the change is reversible.
    backup = path.with_suffix(path.suffix + ".bak_vizo_wrap")
    if not backup.exists():
        shutil.copy2(path, backup)
    wb.save(path)
    wb.close()


def main() -> int:
    dry_run = "--dry-run" in sys.argv
    if not TEMPLATES_DIR.is_dir():
        print(f"Templates dir not found: {TEMPLATES_DIR}")
        return 2
    templates = sorted(TEMPLATES_DIR.glob("*_CECL_SCALE_template.xlsx"))
    if not templates:
        print(f"No templates in {TEMPLATES_DIR}")
        return 1
    print(f"{'DRY-RUN: ' if dry_run else ''}Patching {len(templates)} templates")
    for t in templates:
        fix_template(t, dry_run=dry_run)
    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
