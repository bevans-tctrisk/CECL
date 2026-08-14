"""Produce a cleaned-up version of the CU's new pool mapping file that:
 - resolves the loan-code vs negative-share-code collisions (separate sheets),
 - merges the two 'Used Auto' pool-name variants,
 - adds the loan codes that were missing from the CU's file,
 - documents every change on a Notes sheet.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import OrderedDict

CLIENT = r'Z:\Shared\Clients\Vizo Financial Corporate CU\Client Access\Clients\Mountain CU 68531\Portfolio Management\CECL Migration IDRL'
SRC = os.path.join(CLIENT, '2026', 'Jun 2026', 'Mountain CU Loan Type Codes with Pools NEW.xlsx')
OUT = os.path.join(CLIENT, '2026', 'Jun 2026', 'Mountain CU Loan Type Codes with Pools NEW - CLEANED.xlsx')

USED_MERGE = 'Consumer Auto Loan-Used Auto'
MERGE = {'Consumer Auto Loan-Used Auto': USED_MERGE, 'Consumer Auto Loan-Used': USED_MERGE}
SHARE_SECTIONS = {'Negative Shares', 'Removed from file'}

src = openpyxl.load_workbook(SRC, data_only=True)['Assets']
loan_rows, share_rows = [], []
for i, row in enumerate(src.iter_rows(values_only=True)):
    if i == 0:
        continue
    code, desc, pool = row[1], row[2], row[3]
    if code is None or pool is None:
        continue
    p = str(pool).strip()
    rec = (str(code).strip(), '' if desc is None else str(desc).strip())
    if p in SHARE_SECTIONS:
        share_rows.append(rec + (p,))
    else:
        loan_rows.append(rec + (MERGE.get(p, p),))

# codes missing from the CU file (found in the June loan extract), assigned by current pool
ADDED = [
    ('183', 'Real Estate loan (added; was mapped First Mortgage) - CONFIRM', 'Real Estate'),
    ('168', 'Unsecured loan (added; was mapped Unsecured) - CONFIRM', 'Consumer Unsecured'),
    ('93',  'Real Estate loan (added; was Other Real Estate) - CONFIRM', 'Real Estate'),
    ('200', 'Excluded (added; was Ignore) - CONFIRM', 'Exclude'),
]

wb = openpyxl.Workbook()
# --- Sheet 1: Assets (loan codes) ---
ws = wb.active; ws.title = 'Assets'
hdr_font = Font(bold=True, color='FFFFFF'); hdr_fill = PatternFill('solid', fgColor='4472C4')
ws.append(['Loan Type Code', 'Description', 'Pool'])
for c in ws[1]:
    c.font = hdr_font; c.fill = hdr_fill
# sort by pool then code
def sortkey(r):
    return (r[2], r[0])
for code, desc, pool in sorted(loan_rows + ADDED, key=sortkey):
    ws.append([code, desc, pool])
add_fill = PatternFill('solid', fgColor='FFF2CC')
added_codes = {a[0] for a in ADDED}
for r in ws.iter_rows(min_row=2):
    if str(r[0].value) in added_codes:
        for c in r:
            c.fill = add_fill
ws.column_dimensions['A'].width = 16; ws.column_dimensions['B'].width = 46; ws.column_dimensions['C'].width = 34

# --- Sheet 2: Negative Share Codes (separate) ---
ws2 = wb.create_sheet('Negative Share Codes')
ws2.append(['Share Type Code', 'Description', 'Pool'])
for c in ws2[1]:
    c.font = hdr_font; c.fill = hdr_fill
note = ws2.cell(2, 1, 'NOTE: these codes come from the negative-shares source, NOT the loan file. '
                      'They numerically collide with loan codes (10/20/30/60/70/110), so they are kept '
                      'SEPARATE and are excluded from the loan-file mapping above.')
ws2.merge_cells('A2:C2'); note.font = Font(italic=True, color='C00000'); note.alignment = Alignment(wrap_text=True)
ws2.row_dimensions[2].height = 42
for code, desc, pool in sorted(share_rows, key=lambda r: (r[2], r[0])):
    ws2.append([code, desc, pool])
ws2.column_dimensions['A'].width = 16; ws2.column_dimensions['B'].width = 46; ws2.column_dimensions['C'].width = 34

# --- Sheet 3: Notes / Changes ---
ws3 = wb.create_sheet('Notes - Changes')
ws3.column_dimensions['A'].width = 110
notes = [
 'CLEANED-UP MOUNTAIN CU LOAN TYPE / POOL MAPPING - prepared 2026-07-10',
 '',
 'Changes made to the CU-provided "Mountain CU Loan Type Codes with Pools NEW.xlsx":',
 '',
 '1. NEGATIVE-SHARE CODE COLLISION (most important): the original file listed negative-share',
 '   codes (0, 10, 20, 30, 60, 70, 110, 210, 260) in the same list as loan codes. Several of',
 '   these numerically match real LOAN codes (e.g. 20 = Used Vehicle = ~$41.5M). Applied as a',
 '   single flat list, the loans would be mis-classified as Negative Shares. The share codes are',
 '   now on a SEPARATE sheet ("Negative Share Codes") and are NOT used for the loan file.',
 '',
 '2. DUPLICATE POOL NAME: the file had two nearly identical used-auto pools -',
 '   "Consumer Auto Loan-Used Auto" (only code 0020) and "Consumer Auto Loan-Used" (12 codes).',
 '   These were merged into ONE pool: "Consumer Auto Loan-Used Auto".',
 '',
 '3. CODES MISSING FROM THE FILE (highlighted yellow on the Assets sheet): 4 loan codes that',
 '   carry balances in the June 2026 loan file were not in the mapping (~$2.06M total):',
 '     183 -> Real Estate (~$1.73M; was First Mortgage)',
 '     168 -> Consumer Unsecured (~$303K; was Unsecured)',
 '      93 -> Real Estate (~$12.6K; was Other Real Estate)',
 '     200 -> Exclude (~$13.3K; was Ignore)',
 '   Please CONFIRM these four assignments.',
 '',
 '4. NON-LOAN-FILE POOLS: Member Business Loans, Loan Participations, and Collateral-in-Process',
 '   are driven by the GL/balance sheet (not numeric loan codes). Loan Participations was renamed',
 '   to "CRE Loan Participations" to match the new file. MBL and Collateral were retained.',
 '',
 '5. "Removed from file" (code 9000 = charged off) stays excluded from the pools.',
]
for n in notes:
    ws3.append([n])
ws3['A1'].font = Font(bold=True, size=12)

wb.save(OUT)
print('WROTE', OUT)
print('loan codes:', len(loan_rows) + len(ADDED), '| share codes:', len(share_rows))
