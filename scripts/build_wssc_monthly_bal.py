"""Generate monthly loan-balance-by-pool workbook from imported DB snapshots for WSSC FCU."""
import os
os.environ.setdefault("CECL_WORKSPACE_ROOT", r"Z:\Shared\TCT Files\CECL - CM Files")
WS = os.environ["CECL_WORKSPACE_ROOT"]
import pandas as pd
from sqlalchemy import create_engine, text
from cecl_credentials import get_database_url
from openpyxl import Workbook

CU = "WSSC FCU"
OUT = (r"Z:\Shared\TCT Files\CECL - CM Files\Raw_Uploads\wssc_fcu"
       r"\Monthly Loan Balances by Type.xlsx")

eng = create_engine(get_database_url())
with eng.connect() as c:
    df = pd.read_sql(text(
        "SELECT snapshot_date, loan_pool, SUM(current_balance) AS bal "
        "FROM monthly_loan_data WHERE credit_union=:cu "
        "GROUP BY snapshot_date, loan_pool ORDER BY snapshot_date"), c,
        params={"cu": CU})
df["snapshot_date"] = pd.to_datetime(df["snapshot_date"])
df = df[~df["loan_pool"].isin(["Exclude", "Ignore"])]
pivot = df.pivot_table(index="loan_pool", columns="snapshot_date",
                       values="bal", aggfunc="sum", fill_value=0).sort_index()
dates = list(pivot.columns)
print("pools:", list(pivot.index))
print("months:", len(dates), dates[0].date(), "->", dates[-1].date())

wb = Workbook(); ws = wb.active; ws.title = "Loan Balances by Pool"
ws.cell(row=1, column=1, value="Pool")
for j, d in enumerate(dates):
    ws.cell(row=1, column=2 + j, value=d.to_pydatetime())
r = 2
for pool, row in pivot.iterrows():
    ws.cell(row=r, column=1, value=str(pool))
    for j, d in enumerate(dates):
        ws.cell(row=r, column=2 + j, value=float(row[d]))
    r += 1
ws.cell(row=r, column=1, value="ACL Balance")
for j, d in enumerate(dates):
    ws.cell(row=r, column=2 + j, value=0.0)
wb.save(OUT)
print("Saved:", OUT)

# Record honest provenance on the wizard draft so Step 4 shows that these
# balances were DERIVED from the imported loan extracts (not uploaded).
try:
    from cecl_ui.services import wizard_drafts
    from cecl_ui.routes.setup import build_derived_loan_extract_provenance
    _state = wizard_drafts.load_draft(WS, "wssc_fcu", model="migration")
    if _state is not None:
        _state["hist_balance_provenance"] = build_derived_loan_extract_provenance(CU)
        wizard_drafts.save_draft(
            WS, _state,
            active_step=_state.get("_active_step") or "historical",
            model="migration",
        )
        print("Recorded hist_balance_provenance on draft.")
except Exception as _e:  # noqa: BLE001
    print("provenance record warn:", _e)

