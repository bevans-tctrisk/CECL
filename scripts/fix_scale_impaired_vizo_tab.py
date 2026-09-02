"""Fix the SCALE " Impaired Loans-Vizo" tab so impaired loans compute like the
" Impaired Loans ASC 310-10" tab.

Root cause
----------
The Vizo tab's data rows (28-412) RE-COMPUTED the calc columns from *mirror*
cells, e.g. ``N=IFERROR(E30+G30,"")``. Those mirror cells return an empty
STRING ("") for a blank source cell (blank "Balance at Other Lender" /
collateral). In Excel ``number + ""`` is a #VALUE! error, so ``N`` (Total
Loans) collapsed to "" -> ``O`` (LTV) errored -> ``Q`` (Loss Given Default) and
``R`` (Provision Amount) became 0. Only ``S`` (Bal Removed = E) survived. The
310-10 tab computes fine because its E/G/H are real blank cells (blank -> 0 in
arithmetic).

Fix
---
Make the Vizo data rows MIRROR the 310-10 tab's already-computed columns
directly (the header row 27 already establishes this mapping), guarded by the
source Impairment Type so empty rows stay blank. This guarantees the Vizo tab
equals the 310-10 tab and avoids a divergent copy of the calc logic.

Vizo col  ->  310-10 source col
  M (Member #-Suffix)                 K
  N (Total Loans)                     L
  O (LTV)                             M
  P (Percent applied to Provision)    O
  Q (Loss Given Default)              N
  R (Provision Amount)                P
  S (Bal Removed from Pools)          Q

Applied to EVERY SCALE template (so any period resolves to a fixed template)
and can also patch a specific generated output workbook, passed as CLI args.
"""
import glob
import os
import shutil
import sys
from datetime import datetime

import openpyxl

TEMPLATE_DIR = r"C:\Dev\CECL\cecl_ui\data\scale\templates"
VIZO = " Impaired Loans-Vizo"
SRC = " Impaired Loans ASC 310-10"
FIRST_ROW = 28
LAST_ROW = 412

# Vizo column letter -> 310-10 source column letter
MIRROR = {
    "M": "K",
    "N": "L",
    "O": "M",
    "P": "O",
    "Q": "N",
    "R": "P",
    "S": "Q",
}


def patch_vizo(path: str, backup_tag: str = "impairedvizo") -> int:
    """Rewrite the Vizo tab calc columns of ``path`` to mirror 310-10.

    Returns the number of cells changed (0 if no Vizo sheet). Creates a
    timestamped backup next to the file first.
    """
    wb = openpyxl.load_workbook(path)
    if VIZO not in wb.sheetnames or SRC not in wb.sheetnames:
        print(f"  SKIP {os.path.basename(path)} (missing Vizo/310-10 sheet)")
        return 0
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    shutil.copyfile(path, f"{path}.bak.{backup_tag}_{ts}")
    ws = wb[VIZO]
    changed = 0
    for r in range(FIRST_ROW, LAST_ROW + 1):
        guard = f"'{SRC}'!A{r}"
        for vcol, scol in MIRROR.items():
            ws[f"{vcol}{r}"] = f'=IF({guard}="","",\'{SRC}\'!{scol}{r})'
            changed += 1
    # Force Excel to recalc on open so mirrored values populate even
    # though openpyxl writes no cached values.
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:  # noqa: BLE001
        pass
    wb.save(path)
    print(f"  patched {changed} cells in {os.path.basename(path)}")
    return changed


def main(extra_paths):
    templates = sorted(glob.glob(os.path.join(TEMPLATE_DIR,
                                              "*_CECL_SCALE_template.xlsx")))
    print("=== templates ===")
    for p in templates:
        patch_vizo(p)
    if extra_paths:
        print("=== extra workbooks ===")
        for p in extra_paths:
            patch_vizo(p)


if __name__ == "__main__":
    main(sys.argv[1:])
