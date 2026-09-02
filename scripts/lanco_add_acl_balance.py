"""Add an 'ACL Balance' row to Lanco's monthly loan balance workbook, sourced
from the GL Balance Sheets file: ACL = 719000 CURRENT EXPECTED CREDIT LOSSES
+ 720000 ALLOWANCE FOR ODP (both are accounting-negative '<...>' contras;
stored as a positive balance). Populates the Apr/May/Jun 2026 columns the
'Balance Sheets - April to June.xlsx' file provides. The report engine's
wizard loader auto-detects the 'ACL Balance' label row and the 5300 ACL
fallback will no longer override these explicit values.
"""
import os, calendar, shutil, datetime
import openpyxl

CLIENT = (r'Z:\Shared\Clients\Vizo Financial Corporate CU\Client Access\Clients'
          r'\Lanco FCU 16657\Portfolio Management\CECL Migration IDLR')
NEWF = CLIENT + r'\2026\2026\June 2026\Balance Sheets - April to June.xlsx'
WB_PATH = r'Z:\Shared\TCT Files\CECL - CM Files\Raw_Uploads\lanco_fcu\Monthly Loan Balances by Type.xlsx'

def num(v):
    if v is None: return 0.0
    if isinstance(v,(int,float)): return float(v)
    s=str(v).strip().replace(',','').replace('$','')
    neg=(s.startswith('<') and s.endswith('>')) or (s.startswith('(') and s.endswith(')'))
    s=s.strip('<>()')
    try: return -float(s) if neg else float(s)
    except ValueError: return 0.0

def sheet_to_date(sn):
    p=sn.split('.')
    m,y=(int(p[0]),int(p[1])) if len(p)==2 else (int(p[0]),2000+int(p[2]))
    return '%04d-%02d-%02d'%(y,m,calendar.monthrange(y,m)[1])

# --- extract ACL (719000 + 720000, as positive) per month from balance sheet ---
nwb = openpyxl.load_workbook(NEWF, data_only=True)
acl_by_date = {}
for sn in nwb.sheetnames:
    d = sheet_to_date(sn); ws = nwb[sn]
    cecl = odp = 0.0
    for row in ws.iter_rows(values_only=True):
        try: acc = int(str(row[0]).strip())
        except (ValueError,TypeError): continue
        if acc == 719000: cecl = num(row[2])
        elif acc == 720000: odp = num(row[2])
    acl = abs(cecl) + abs(odp)
    acl_by_date[d] = round(acl, 2)
    print(f"  {d}: ACL = 719000 ${abs(cecl):,.2f} + 720000 ${abs(odp):,.2f} = ${acl:,.2f}")

# --- add 'ACL Balance' row to the workbook ---
shutil.copyfile(WB_PATH, WB_PATH + '.bak.phase_acl_' + datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
wb = openpyxl.load_workbook(WB_PATH)
ws = wb['Sheet1']
date_col = {str(ws.cell(1,c).value)[:10]: c for c in range(2, ws.max_column+1)}

# find or create the ACL Balance row
acl_row = None
for r in range(2, ws.max_row+1):
    if str(ws.cell(r,1).value).strip().lower() in ('acl balance','alll balance'):
        acl_row = r; break
if acl_row is None:
    acl_row = ws.max_row + 1
    ws.cell(acl_row, 1).value = 'ACL Balance'
    print(f"  added 'ACL Balance' row at row {acl_row}")
else:
    print(f"  updating existing ACL row at row {acl_row}")

for d, acl in acl_by_date.items():
    c = date_col.get(d)
    if c:
        ws.cell(acl_row, c).value = acl
        print(f"     set {d} (col {c}) = ${acl:,.2f}")
    else:
        print(f"     WARNING: {d} column not in workbook")
wb.save(WB_PATH)
print("SAVED", WB_PATH)
