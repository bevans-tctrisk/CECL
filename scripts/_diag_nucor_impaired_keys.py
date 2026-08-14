"""Confirm the impaired parser member-key mismatch for Nucor (Turn 8)."""
import os, sys
os.environ.setdefault('CECL_WORKSPACE_ROOT', r'Z:\Shared\TCT Files\CECL - CM Files')
sys.path.insert(0, r'c:\Dev\CECL')
from cecl_credentials import get_database_url
os.environ['DATABASE_URL'] = get_database_url()
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

eng = create_engine(os.environ['DATABASE_URL'])
# Load impaired workbook
p = r'Z:\Shared\TCT Files\CECL - CM Files\Raw_Uploads\nucor_emp_cu\impaired loans july 2026.xlsx'
wb = load_workbook(p, data_only=True)
ws = wb[' Impaired Loans']

# Load grade_lookup from DB
with eng.connect() as c:
    rows = c.execute(
        text("SELECT member_number, current_fico_score FROM monthly_loan_data "
             "WHERE credit_union=:cu AND snapshot_date=:d"),
        {"cu": "Nucor Emp CU", "d": "2026-06-30"},
    ).fetchall()
db_members = {r[0]: r[1] for r in rows}
print(f"DB members: {len(db_members)}")

# Iterate impaired rows
suffix_len = 0  # Nucor YAML has account_suffix_length: 0
tried_keys_hit = 0
tried_keys_miss = 0
sample_miss = []
for row in range(30, ws.max_row + 1):
    imp_type = ws.cell(row=row, column=1).value
    if not imp_type:
        continue
    member = ws.cell(row=row, column=2).value
    suffix = ws.cell(row=row, column=3).value
    if member is None or suffix is None:
        continue
    try:
        mi = int(float(member))
        si = int(float(suffix))
    except (TypeError, ValueError):
        continue
    # Current parser logic
    key1 = f"{mi}{si:0{suffix_len}d}"  # "105571" for member=10557 suffix=1
    # Candidate variants
    keys = [
        key1,
        f"{mi}{si:02d}",  # 105701
        f"{mi}{si:03d}",  # 10557001
        f"{mi:09d}{si:02d}",  # 00001055701 (9+2)
        f"{mi:08d}{si:03d}",  # 0001055701 (8+3)
    ]
    hits = [k for k in keys if k in db_members]
    if hits:
        tried_keys_hit += 1
    else:
        tried_keys_miss += 1
        if len(sample_miss) < 5:
            sample_miss.append((mi, si, keys))

print(f"Hit: {tried_keys_hit}, Miss: {tried_keys_miss}")
print("Sample miss:")
for m, s, k in sample_miss:
    print(f"  member={m} suf={s} keys tried: {k}")
