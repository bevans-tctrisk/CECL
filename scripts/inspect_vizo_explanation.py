"""Inspect rows 18-26 on the Vizo Explanation-of-ACL-Calc tab."""
from pathlib import Path
from openpyxl import load_workbook

TPL = Path(r"C:\Dev\CECL\cecl_ui\data\scale\templates\2026_03_CECL_SCALE_template.xlsx")
SHEET = "Explanation of ACL Calc-Vizo"

wb = load_workbook(TPL, data_only=False)
ws = wb[SHEET]
print("dims:", ws.dimensions, "max_row:", ws.max_row, "max_col:", ws.max_column)

print("\nmerged ranges touching rows 18-26:")
for r in ws.merged_cells.ranges:
    if r.min_row <= 26 and r.max_row >= 18:
        print(" ", r)

print("\ncolumn widths:")
for k, v in sorted(ws.column_dimensions.items()):
    if v.width:
        print(f"  {k}: {v.width}")

for row_idx in range(18, 27):
    rd = ws.row_dimensions[row_idx]
    print(f"\n--- row {row_idx}  height={rd.height}  customHeight={rd.customHeight} ---")
    for cell in ws[row_idx]:
        if cell.value is not None:
            wt = cell.alignment.wrap_text if cell.alignment else None
            v = str(cell.value)
            short = v[:160] + ("..." if len(v) > 160 else "")
            print(f"  {cell.coordinate}: wrap={wt}  len={len(v)}  v={short!r}")
wb.close()
