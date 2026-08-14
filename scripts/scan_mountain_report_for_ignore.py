"""Scan the Mountain CU March 2026 Vizo report for any remaining 'Ignore' pool cells."""
import sys
from pathlib import Path
from openpyxl import load_workbook

REPORT = Path(r"Z:\Shared\TCT Files\CECL - CM Files\Reports\2026-03-31_CECL_Migration_Mountain_CU_Vizo_Model.xlsx")

if not REPORT.exists():
    print(f"NOT FOUND: {REPORT}")
    sys.exit(1)

wb = load_workbook(REPORT, read_only=True, data_only=True)
hits = []
for ws in wb.worksheets:
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and v.strip().lower() == "ignore":
                hits.append((ws.title, v))
                break  # one per row is enough to flag
print(f"Sheets scanned: {len(wb.sheetnames)}")
print(f"'Ignore' cells found: {len(hits)}")
for sheet, val in hits[:20]:
    print(f"  {sheet}: {val!r}")
if not hits:
    print("PASS: No 'Ignore' pool cells in the report.")
