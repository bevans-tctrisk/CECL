"""Phase SCALE.prior-leak — one-shot remediation for Greensboro CU Mar 2026.

The reported bug: Executive Summary-Vizo "prior" column shows the SAME
numbers as the "current" column. Root cause: the carry-history workflow
that built Mar from Dec left ``Historical Data!AY2..AY5`` populated
with formulas like ``='Scale Calculation'!U29..U33``. Those formulas
pointed at the prior-period column when the Dec workbook was created,
but after Mar's column shift they now point at the current-period
column instead, leaking current values into the prior slot.

Behaviour matches the carry-history shape in ``runner.py``:
``seed_new_historical_data_column`` writes the same formulas into the
new (current) column, but never clears or rewrites the AY column.
``_inject_prior_acl`` is supposed to overwrite AY2..AY5 with literal
values pulled from the prior workbook, but it silently no-ops when the
prior workbook's cached formula values are all None (no Excel pass yet
and ``pywin32`` missing for headless recalc).

This script overwrites AY2..AY5 with literal 0.0 on Greensboro's Mar
output workbook so the "prior column = current column" leak stops
immediately. After the wizard's Python env gets pywin32 (or the user
opens the Dec workbook in Excel + Save to populate cached values), the
analyst can re-run the Mar report and the real prior values will be
injected. Idempotent. Backs the workbook up to a ``.bak.phase_prior_leak``
sibling.

Run from the repo root:

    python scripts/fix_greensboro_mar_prior_acl.py
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK = Path(
    r"Z:\Shared\TCT Files\CECL - CM Files\Generated_Reports"
    r"\greensboro_cu\2026-03\2026-03_CECL_SCALE_greensboro_cu_Vizo.xlsx"
)
SHEET = "Historical Data"
CELLS = ("AY2", "AY3", "AY4", "AY5")


def main() -> int:
    if not WORKBOOK.exists():
        print(f"ERROR: workbook not found at {WORKBOOK}")
        return 2
    backup = WORKBOOK.with_suffix(WORKBOOK.suffix + ".bak.phase_prior_leak")
    if not backup.exists():
        shutil.copy2(WORKBOOK, backup)
        print(f"Backed up to {backup}")
    else:
        print(f"Backup already present at {backup} (leaving as-is)")

    wb = load_workbook(WORKBOOK)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: sheet '{SHEET}' not found; sheets are {wb.sheetnames}")
        return 2
    ws = wb[SHEET]
    print(f"Before:")
    for c in CELLS:
        cell = ws[c]
        print(f"  {c}: value={cell.value!r}")
    for c in CELLS:
        ws[c] = 0.0
    wb.save(WORKBOOK)
    print(f"After (saved):")
    wb2 = load_workbook(WORKBOOK, data_only=False)
    ws2 = wb2[SHEET]
    for c in CELLS:
        print(f"  {c}: value={ws2[c].value!r}")
    print(
        "\nDone. Prior column on Executive Summary-Vizo will now read $0 "
        "instead of duplicating the current column. To populate real "
        "prior values: (a) install pywin32 in the wizard env, OR (b) "
        "open the Dec workbook in Excel + Save so cached values "
        "populate, then re-run Mar from the wizard."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
