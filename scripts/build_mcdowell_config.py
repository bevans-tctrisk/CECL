"""Build client_configs/mcdowell_cornerstone_cu.yaml from scratch.

McDowell Cornerstone CU (charter 60149), Vizo client. Two loan extracts:
 - AIRES core loan file (main portfolio, alpha loan-type codes, single FICO column)
 - Buy Side Combined Report (LoanStreet purchased participations -> single-line pool)
 - Negative Shares -> Ignore.
Single FICO column (no separate current-score pull) -> original-score credit method
(credit_pull.prefer_original_for_new_loans) per user direction. ACL booked balance
pulled from NCUA 5300 (acl.use_5300_fallback).
"""
import os, yaml
os.environ.setdefault('CECL_WORKSPACE_ROOT', r'Z:\Shared\TCT Files\CECL - CM Files')
WS = os.environ['CECL_WORKSPACE_ROOT']
SHORT = 'mcdowell_cornerstone_cu'
CLIENT_FOLDER = (r'Z:\Shared\Clients\Vizo Financial Corporate CU\Client Access\Clients'
                 r'\McDowell Cornerstone CU 60149\Portfolio Management\CECL Migration IDLR')

# ---- pool map (from reference file 'Assets' sheet: code | desc | pool) ----
import openpyxl
ref = openpyxl.load_workbook(
    os.path.join(CLIENT_FOLDER, 'Credit Union C Loan Type Codes with Pools.xlsx'),
    data_only=True)['Assets']
pool_map = {}
for r in range(2, ref.max_row + 1):
    code = ref.cell(r, 1).value
    pool = ref.cell(r, 3).value
    if code is None or not pool:
        continue
    c = str(code).strip()
    p = str(pool).strip()
    if p == 'Negative Shares':
        p = 'Ignore'
    pool_map[c] = p
# CO/recovery + buy-side use upper-case 'LOAN STREET'
pool_map['Loan Street'] = 'Loan Participations'
pool_map['LOAN STREET'] = 'Loan Participations'
pool_map['Loan Participations'] = 'Loan Participations'   # static-pool identity key
pool_map['Neg Shares'] = 'Ignore'

POOLS = ['Real Estate', 'Consumer Secured', 'Consumer Unsecured', 'Loan Participations']
RE_POOLS = {'Real Estate'}
# Participations reported as a single line (no grade breakout).
NOT_RISK_RATED = {'Loan Participations'}

def pool_entry(name):
    return {'name': name, 'risk_rated': name not in NOT_RISK_RATED,
            'acl_months': 84 if name in RE_POOLS else 36,
            'excluded': False,
            'use_default_mgmt_adj': name in RE_POOLS, 'brr': False}

pools = [pool_entry(p) for p in POOLS] + [
    {'name': 'Ignore', 'risk_rated': False, 'acl_months': None, 'excluded': False,
     'use_default_mgmt_adj': False, 'brr': False}]

# FICO bands from 'Credit Union C FICO Score Ranges.docx' (cross-checked vs the
# core 'Loan Risk Grade' column): A+ 730+, A 680-729, B 640-679, C 600-639,
# D 550-599, E <=549. Reserve rates are PLACEHOLDERS (tune with CU model).
grades = [
    {'label': 'A+', 'min_score': 730, 'max_score': 900, 'reserve_rate': 0.0011},
    {'label': 'A',  'min_score': 680, 'max_score': 729, 'reserve_rate': 0.0025},
    {'label': 'B',  'min_score': 640, 'max_score': 679, 'reserve_rate': 0.0050},
    {'label': 'C',  'min_score': 600, 'max_score': 639, 'reserve_rate': 0.0116},
    {'label': 'D',  'min_score': 550, 'max_score': 599, 'reserve_rate': 0.0250},
    {'label': 'E',  'min_score': 350, 'max_score': 549, 'reserve_rate': 0.0500},
]

# AIRES core loan file (header names stable across the 36/39-col month variants)
aires_cm = {
    'member_number': 'Account', 'current_balance': 'Current Balance',
    'loan_pool_code': 'Loan Type Code', 'original_fico_score': 'FICO',
    'current_fico_score': 'FICO', 'open_date': 'Original Date',
    'original_loan_amount': 'Original Loan Amount', 'interest_rate': 'Interest Rate',
    'total_available_credit': 'Credit Limit', 'days_delinquent': 'Days Delinquent',
}
aires_ma = {'mode': 'fixed_suffix', 'suffix_length': 4}

# Buy Side Combined Report (LoanStreet participations). 'Loan Level' is the first
# sheet so it is read by default. Balance = the CU's OWNED principal.
buyside_cm = {
    'member_number': 'LoanStreet ID',
    'current_balance': 'Principal Owned at Period End',
    'loan_pool_code_static': 'Loan Participations',
    'original_fico_score': 'Credit Score', 'current_fico_score': 'Credit Score',
    'open_date': 'Origination Date', 'original_loan_amount': 'Original Principal',
    'interest_rate': 'Current Interest Rate', 'days_delinquent': 'Days Delinquent',
}

AIRES_PAT = r'(?i)^aires\b.*\.(xlsx|xls)$'
BUYSIDE_PAT = r'(?i)^buy\s*side\b.*\.(xlsx|xls)$'
NS_PAT = r'(?i)^negative\s+shares?\b.*\.(xlsx|xls)$'

config = {
    'credit_union': 'McDowell Cornerstone CU',
    'charter_number': '60149',
    'report_period': '2026-03',
    'file_pattern': AIRES_PAT,
    'date_pattern': r'(\d{2})-(\d{2})-(\d{4})',
    'date_format': 'MMDDYYYY',
    'account_suffix_length': 4,
    'member_account': aires_ma,
    'has_header': True,
    'column_mappings': aires_cm,
    'balance_format': {'remove_chars': ['$'], 'accounting_negatives': True},
    'pool_code_split': '',
    'pool_map': pool_map,
    'default_pool': 'Ignore',
    'credit_grades': grades,
    'no_score_label': 'Not Reported',
    'reports': {'tct': False, 'vizo': True, 'vizo_supp': True, 'impdet': True},
    'economic_data': {'state': 'North Carolina', 'county': 'McDowell County',
                      'unemployment_rate': 0.037, 'foreclosures': 60,
                      'bankruptcies': 150, 'population': 45000},
    'mgmt_adj': {'ltv_baseline': 0.9, 'probability_factor': 0.35},
    # Single FICO column (no separate bureau pull): derive each member's CURRENT
    # score from their most-recently-opened loan's origination score -> migration.
    'credit_pull': {'prefer_original_for_new_loans': True},
    # Booked ACL (allowance) balance from NCUA 5300 (field AAS0048).
    'acl': {'use_5300_fallback': True},
    'loan_data_extracts': [
        {'label': 'Buy Side', 'file_pattern': BUYSIDE_PAT,
         'column_mappings': buyside_cm,
         'member_account': {'mode': 'fixed_suffix', 'suffix_length': 0},
         'has_header': True},
        {'label': 'Negative Shares', 'file_pattern': NS_PAT,
         'column_mappings': {'member_number': 'ACCOUNT',
                             'current_balance': 'BALANCE',
                             'loan_pool_code': 'Loan Type Code'},
         'member_account': {'mode': 'fixed_suffix', 'suffix_length': 0},
         'has_header': True},
        {'label': 'AIRES', 'file_pattern': AIRES_PAT,
         'column_mappings': aires_cm, 'member_account': aires_ma, 'has_header': True},
    ],
    'pools': pools,
    'pool_order': POOLS + ['Ignore'],
    'not_risk_rated': sorted(NOT_RISK_RATED),
    'acl_months_by_pool': {p: (84 if p in RE_POOLS else 36) for p in POOLS},
    'historical_file_formats': {
        'chargeoff': {'has_header': True, 'skip_rows': 0, 'account_col': 0,
                      'code_col': 1, 'amount_col': 3, 'date_col': 4, 'member_col': 0,
                      'strict_columns': True,
                      'member_account': {'mode': 'fixed_suffix', 'suffix_length': 0}},
        'recovery': {'has_header': True, 'skip_rows': 0, 'account_col': 0,
                     'code_col': 1, 'amount_col': 2, 'date_col': 4, 'member_col': 0,
                     'strict_columns': True,
                     'member_account': {'mode': 'fixed_suffix', 'suffix_length': 0}},
    },
    'data_directory': os.path.join(WS, 'Raw_Uploads', SHORT),
    'loan_source_folder': CLIENT_FOLDER,
    'monthly_balance': {'source': 'single'},
}

os.makedirs(config['data_directory'], exist_ok=True)
out = os.path.join(WS, 'client_configs', SHORT + '.yaml')
with open(out, 'w', encoding='utf-8') as f:
    yaml.safe_dump(config, f, sort_keys=False, default_flow_style=False, allow_unicode=True)
print('WROTE', out)
print('pools:', POOLS)
print('pool_map codes:', len(pool_map))
c2 = yaml.safe_load(open(out, encoding='utf-8'))
print('reparse OK; extracts:', [e['label'] for e in c2['loan_data_extracts']])
