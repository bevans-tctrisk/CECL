"""Generate a monthly loan-balance-by-pool workbook from the imported DB
snapshots, so the report's Loss Factor tab has real historical balances.
Also writes an ACL Balance row (zeros — CU supplied no ACL history)."""
import os
os.environ.setdefault("CECL_WORKSPACE_ROOT", r"Z:\Shared\TCT Files\CECL - CM Files")
import pandas as pd
from sqlalchemy import create_engine, text
from cecl_credentials import get_database_url
from openpyxl import Workbook

CU = "Curis CU-Palmetto Health CU"
OUT = (r"Z:\Shared\TCT Files\CECL - CM Files\Raw_Uploads\curis_palmetto_health_cu"
       r"\Monthly Loan Balances by Type.xlsx")

eng = create_engine(get_database_url())
with eng.connect() as c:
    df = pd.read_sql(text(
        "SELECT snapshot_date, loan_pool, SUM(current_balance) AS bal "
        "FROM monthly_loan_data WHERE credit_union=:cu "
        "GROUP BY snapshot_date, loan_pool ORDER BY snapshot_date"), c,
        params={"cu": CU})

df["snapshot_date"] = pd.to_datetime(df["snapshot_date"])
# Drop the Exclude/Ignore bucket from the balance trend.
df = df[~df["loan_pool"].isin(["Exclude", "Ignore"])]
pivot = df.pivot_table(index="loan_pool", columns="snapshot_date",
                       values="bal", aggfunc="sum", fill_value=0)
pivot = pivot.sort_index()
dates = list(pivot.columns)
print("pools:", list(pivot.index))
print("months:", len(dates), dates[0].date(), "->", dates[-1].date())

wb = Workbook()
ws = wb.active
ws.title = "Loan Balances by Pool"
# Header row: col A label, then one datetime per month.
ws.cell(row=1, column=1, value="Pool")
for j, d in enumerate(dates):
    ws.cell(row=1, column=2 + j, value=d.to_pydatetime())
# Pool rows.
r = 2
for pool, row in pivot.iterrows():
    ws.cell(row=r, column=1, value=str(pool))
    for j, d in enumerate(dates):
        ws.cell(row=r, column=2 + j, value=float(row[d]))
    r += 1
# ACL Balance row (zeros — no ACL history supplied by CU).
ws.cell(row=r, column=1, value="ACL Balance")
for j, d in enumerate(dates):
    ws.cell(row=r, column=2 + j, value=0.0)

wb.save(OUT)
print("Saved:", OUT)
