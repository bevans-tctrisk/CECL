import os, yaml, calendar
import openpyxl, pandas as pd
os.environ['CECL_WORKSPACE_ROOT'] = r'Z:\Shared\TCT Files\CECL - CM Files'
WS = os.environ['CECL_WORKSPACE_ROOT']
SHORT = 'wnc_community_cu'
cfg = yaml.safe_load(open(os.path.join(WS, 'client_configs', SHORT + '.yaml'), encoding='utf-8'))
pm = {str(k).strip(): v for k, v in cfg['pool_map'].items()}

bs = r'Z:\Shared\Clients\Vizo Financial Corporate CU\Client Access\Clients\WNC Community CU 66854\Portfolio Management\CECL Migration IDLR\36 Month Balance Sheet.xlsx'
wb = openpyxl.load_workbook(bs, data_only=True)

def sheet_to_date(sn):
    m, y = sn.split('-'); m = int(m); y = 2000 + int(y)
    return '%04d-%02d-%02d' % (y, m, calendar.monthrange(y, m)[1])

POOLS = ['Real Estate', 'Consumer Secured', 'Consumer Unsecured']
data = {}  # date -> {pool: bal}
for sn in wb.sheetnames:
    try:
        d = sheet_to_date(sn)
    except Exception:
        continue
    ws = wb[sn]
    by = {}
    for r in range(1, ws.max_row + 1):
        code = ws.cell(r, 2).value      # col B
        bal = ws.cell(r, 4).value       # col D
        if code is None or bal in (None, ''):
            continue
        pool = pm.get(str(code).strip())
        if pool and pool not in ('Ignore', 'Exclude'):
            try:
                by[pool] = by.get(pool, 0.0) + float(bal)
            except (TypeError, ValueError):
                pass
    if by:
        data[d] = by

dates = sorted(data)
rows = []
for p in POOLS:
    row = {'Pool': p}
    for d in dates:
        row[d] = round(data[d].get(p, 0.0), 2)
    rows.append(row)
wide = pd.DataFrame(rows, columns=['Pool'] + dates)
out_dir = os.path.join(WS, 'Raw_Uploads', SHORT)
os.makedirs(out_dir, exist_ok=True)
out = os.path.join(out_dir, 'Monthly Loan Balances by Type.xlsx')
with pd.ExcelWriter(out, engine='openpyxl') as xw:
    wide.to_excel(xw, sheet_name='Sheet1', index=False)
print('WROTE', out)
print('months:', len(dates), dates[0], '->', dates[-1])
print('Dec-2025 total:', round(sum(data.get('2025-12-31', {}).values()), 2), '(GL should be 51,854,893.43)')
for p in POOLS:
    print('   %-22s Dec25 %.2f' % (p, data.get('2025-12-31', {}).get(p, 0)))

# wire into config
cfg['monthly_balance'] = {'source': 'single', 'sheet': 'Sheet1', 'header_row': 1,
                          'pool_name_col': 'A', 'first_date_col': 'B',
                          'filename': 'Monthly Loan Balances by Type.xlsx', 'saved_path': out,
                          'pool_map': {p: p for p in POOLS}}
with open(os.path.join(WS, 'client_configs', SHORT + '.yaml'), 'w', encoding='utf-8') as f:
    yaml.safe_dump(cfg, f, sort_keys=False, allow_unicode=True)
print('config monthly_balance wired')
