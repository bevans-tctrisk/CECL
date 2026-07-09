"""Stage all WNC CU charge-off/recovery files into Raw_Uploads, then run the
NCUA 5300 CO/recovery + distributed-balance backfills (charter 66854) so the
single-line pools get a real loss rate over their acl_months windows
(Real Estate 84mo, Consumer 36mo)."""
import os, shutil, glob
os.environ['CECL_WORKSPACE_ROOT'] = r'Z:\Shared\TCT Files\CECL - CM Files'
import cecl_credentials
os.environ['DATABASE_URL'] = cecl_credentials.get_database_url()
import sqlalchemy as sa
from cecl_ui.services import solr_5300_co_backfill as co, solr_5300_recov_backfill as rec, solr_5300_backfill as bf

WS = os.environ['CECL_WORKSPACE_ROOT']
CLIENT = r'Z:\Shared\Clients\Vizo Financial Corporate CU\Client Access\Clients\WNC Community CU 66854\Portfolio Management\CECL Migration IDLR'
DST = os.path.join(WS, 'Raw_Uploads', 'wnc_community_cu')
CU = 'WNC Community CU'; CHARTER = 66854
SOLR = 'http://searchserver1.tctrisk.com:8983/solr'; CORE = 'ncua'
TARGET = '2026-03-31'; MONTHS = 90

# 1. stage all CU CO/recovery files (2020-2026)
n = 0
for f in glob.glob(os.path.join(CLIENT, '**', '*hargeoffs and Recoveries*.xls*'), recursive=True):
    shutil.copy2(f, os.path.join(DST, os.path.basename(f))); n += 1
print('staged CO files:', n)

# 2. 5300 CO + recovery backfill
r1 = co.backfill_missing_chargeoff_quarters(CU, CHARTER, SOLR, CORE, TARGET, MONTHS, overwrite=True)
print('5300 CO: ok=%s rows=%s filled=%s' % (r1.get('ok'), r1.get('rows_written'), len(r1.get('months_filled') or [])))
r2 = rec.backfill_missing_recovery_quarters(CU, CHARTER, SOLR, CORE, TARGET, MONTHS, overwrite=True)
print('5300 REC: ok=%s rows=%s filled=%s' % (r2.get('ok'), r2.get('rows_written'), len(r2.get('months_filled') or [])))

# 3. 5300 distributed balance backfill (RE 84mo). pool_distribution from Dec 2025 pool mix.
import openpyxl
eng = sa.create_engine(cecl_credentials.get_database_url())
with eng.connect() as cx:
    rows = cx.execute(sa.text(
        "SELECT loan_pool, SUM(current_balance) b FROM monthly_loan_data WHERE credit_union=:c "
        "AND snapshot_date='2025-12-31' AND loan_pool NOT IN ('Ignore','Exclude') GROUP BY loan_pool"),
        {'c': CU}).fetchall()
    by = {r.loan_pool: float(r.b or 0) for r in rows if (r.b or 0) > 0}
    tot = sum(by.values())
    dist = {p: v / tot for p, v in by.items()}
# existing = months already covered by the 36-month balance file (2023-07..2026-06);
# 5300 fills only the earlier RE-window years (back to ~2019).
mbf = os.path.join(DST, 'Monthly Loan Balances by Type.xlsx')
ws = openpyxl.load_workbook(mbf, data_only=True)['Sheet1']
existing = set()
for c in range(2, ws.max_column + 1):
    v = ws.cell(1, c).value
    if v:
        existing.add(str(v)[:10])
r3 = bf.backfill_missing_quarters_distributed(CU, CHARTER, SOLR, CORE, TARGET, MONTHS,
        pool_distribution=dist, existing_dates=existing, source_period_iso='2025-12-31')
print('5300 BAL: ok=%s rows=%s filled=%s' % (r3.get('ok'), r3.get('rows_written'), len(r3.get('months_filled') or [])))
# purge any in-range distributed rows (idempotence guard: >= earliest balance-file month)
with eng.begin() as cx:
    lo = min(existing)
    d = cx.execute(sa.text("DELETE FROM loan_code_history WHERE cu=:c AND source LIKE '5300-distributed:%' "
                           "AND as_of_date >= :lo"), {'c': CU, 'lo': lo})
    print('purged in-range distributed rows (>= %s):' % lo, d.rowcount)
