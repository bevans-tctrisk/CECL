"""Backfill WSSC pre-2023 historical balances from NCUA 5300 (distributed mode):
distribute each quarter's 5300 total loan balance across the CU's pools using the
pool mix from the earliest clean month (2023-07), writing loan_code_history rows
(source '5300-distributed:<qe>'). Real months 2023-07..2025-12 are passed as
'existing' so 5300 only fills the earlier quarters (back to ~2018)."""
import os, datetime
os.environ.setdefault("CECL_WORKSPACE_ROOT", r"Z:\Shared\TCT Files\CECL - CM Files")
from openpyxl import load_workbook
from cecl_ui.services import solr_5300_backfill

CU = "WSSC FCU"
CHARTER = 16268
SOLR_URL = "http://searchserver1.tctrisk.com:8983/solr"
CORE = "ncua"
PERIOD = "2025-12-31"
MONTHS = 84
BALFILE = r"Z:\Shared\TCT Files\CECL - CM Files\Raw_Uploads\wssc_fcu\Monthly Loan Balances by Type.xlsx"

# Read the balance file: pools x month columns.
wb = load_workbook(BALFILE, data_only=True); ws = wb.active
hdr = [c.value for c in ws[1]]
date_cols = {j: (v.date().isoformat()) for j, v in enumerate(hdr) if isinstance(v, datetime.datetime)}
existing_dates = set(date_cols.values())               # 30 real month-ends
earliest_j = min(date_cols, key=lambda j: date_cols[j])
earliest_iso = date_cols[earliest_j]
by_pool = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    pool = r[0]
    if pool in (None, "ACL Balance"):
        continue
    by_pool[str(pool)] = float(r[earliest_j] or 0)
wb.close()
total = sum(by_pool.values())
pool_distribution = {p: (b / total) for p, b in by_pool.items() if b > 0}
print(f"Distribution period: {earliest_iso}  total=${total:,.0f}")
for p, ratio in sorted(pool_distribution.items(), key=lambda kv: -kv[1]):
    print(f"  {p:26} {ratio*100:5.1f}%  (${by_pool[p]:,.0f})")
print(f"\nExisting (skip) months: {len(existing_dates)}  ({min(existing_dates)}..{max(existing_dates)})")

res = solr_5300_backfill.backfill_missing_quarters_distributed(
    CU, CHARTER, SOLR_URL, CORE, PERIOD, MONTHS,
    pool_distribution=pool_distribution,
    existing_dates=existing_dates,
    source_period_iso=earliest_iso,
)
print("\n=== backfill result ===")
print("ok:", res.get("ok"), "| error:", res.get("error"))
print("rows_written:", res.get("rows_written"), "| stale_removed:", res.get("stale_rows_removed"))
mf = res.get("months_filled") or []
print("months_filled:", len(mf))
if mf:
    print("  first:", mf[0].get("period") if isinstance(mf[0], dict) else mf[0],
          " last:", mf[-1].get("period") if isinstance(mf[-1], dict) else mf[-1])
print("months_no_data:", res.get("months_no_data"))
print("months_skipped:", len(res.get("months_skipped") or []))
