"""Rebuild Mountain CU with the CU's NEW granular pool breakout.

- Loan pool_map from "Mountain CU Loan Type Codes with Pools NEW.xlsx" (loan
  sections only; the 'Negative Shares' + 'Removed from file' sections are a
  DIFFERENT code system that numerically collides with loan codes, so they are
  excluded from the loan map). The two 'Used Auto' pool-name variants are merged.
- Codes missing from the new file are assigned by their current pool
  (183->Real Estate, 168->Consumer Unsecured, 93->Real Estate, 200->Exclude).
- Member Business Loans + Collateral-in-Process kept; Loan Participations ->
  'CRE Loan Participations'.
- Historical pool balances (GL) are SPLIT into the new pools by the fixed
  June-2026 loan-extract percentages (CU-approved method), written to a new wide
  balance file. monthly_file_pattern (old GL labels) is removed.
"""
import os, re, shutil, copy, yaml
import openpyxl
from collections import defaultdict, OrderedDict

WS = r'Z:\Shared\TCT Files\CECL - CM Files'
CFG = os.path.join(WS, 'client_configs', 'mountain_cu.yaml')
CLIENT = r'Z:\Shared\Clients\Vizo Financial Corporate CU\Client Access\Clients\Mountain CU 68531\Portfolio Management\CECL Migration IDRL'
NEWMAP_F = os.path.join(CLIENT, '2026', 'Jun 2026', 'Mountain CU Loan Type Codes with Pools NEW.xlsx')
HIST_F = os.path.join(CLIENT, 'Historical Balance Sheet - Mountain CU - 68531.xlsx')
JUNE_BS = os.path.join(CLIENT, '2026', 'Jun 2026', 'Balance Sheet 6-30-26.xlsx')
DD = os.path.join(WS, 'Raw_Uploads', 'mountain_cu')
LOANF = os.path.join(DD, 'Loan File - 6-30-26 .xlsx')
NEW_HIST_OUT = os.path.join(DD, 'Historical Balance Sheet - Mountain CU - GRANULAR.xlsx')

USED_MERGE_NAME = 'Consumer Auto Loan-Used Auto'
EXCLUDE_SECTIONS = {'Negative Shares', 'Removed from file'}
MERGE_NAMES = {'Consumer Auto Loan-Used Auto': USED_MERGE_NAME,
               'Consumer Auto Loan-Used': USED_MERGE_NAME}
MANUAL = {'183': 'Real Estate', '168': 'Consumer Unsecured',
          '93': 'Real Estate', '200': 'Exclude', '9000': 'Exclude'}

def norm(c):
    s = str(c).strip()
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s

# ---- 1. Build loan pool_map from NEW file (loan sections only) ----
wb = openpyxl.load_workbook(NEWMAP_F, data_only=True)['Assets']
loanmap = {}
for i, row in enumerate(wb.iter_rows(values_only=True)):
    if i == 0:
        continue
    code, pool = row[1], row[3]
    if code is None or pool is None:
        continue
    p = str(pool).strip()
    if p in EXCLUDE_SECTIONS:
        continue
    p = MERGE_NAMES.get(p, p)
    raw = str(code).strip()
    loanmap[raw] = p
    n = norm(code)
    loanmap.setdefault(n, p)
    if n.isdigit():
        loanmap[n] = p
        loanmap[n.zfill(4)] = p

# current config (for special-pool codes + grades + everything else)
cfg = yaml.safe_load(open(CFG, encoding='utf-8'))
curmap = cfg.get('pool_map', {})

# keep Member Business Loans + Collateral codes; rename Loan Participations pool
for k, v in curmap.items():
    if v == 'Member Business Loans':
        loanmap[k] = 'Member Business Loans'
    elif v == 'Collateral in Process of Liquidation':
        loanmap[k] = 'Collateral in Process of Liquidation'
    elif v == 'Loan Participations':
        loanmap[k] = 'CRE Loan Participations'

# manual assignments for codes missing from the new file
for c, p in MANUAL.items():
    loanmap[c] = p
    if c.isdigit():
        loanmap[c.zfill(4)] = p

NEW_POOLS = ['Consumer Auto Loan-New Auto', USED_MERGE_NAME,
             'Consumer Indirect Auto-New Auto', 'Consumer Indirect Auto-Used Auto',
             'Consumer Secured', 'Consumer Unsecured', 'Credit Cards', 'Real Estate',
             'CRE Loan Participations', 'Member Business Loans',
             'Collateral in Process of Liquidation']
RE_POOLS = {'Real Estate', 'CRE Loan Participations', 'Member Business Loans'}
NOT_RR = {'CRE Loan Participations', 'Member Business Loans',
          'Collateral in Process of Liquidation', 'Ignore'}

# ---- 2. Crosswalk fractions from June extract (old pool -> new pool) ----
curmap_n = {norm(k): v for k, v in curmap.items()}
loanmap_n = {norm(k): v for k, v in loanmap.items()}
wb2 = openpyxl.load_workbook(LOANF, read_only=True, data_only=True)
ws2 = wb2[wb2.sheetnames[0]]
it = ws2.iter_rows(values_only=True)
hdr = [str(h).strip() if h is not None else '' for h in next(it)]
ci_code, ci_bal, ci_co = hdr.index('Loan Type Code'), hdr.index('Current Loan Balance'), hdr.index('C/O Amount')
cross = defaultdict(lambda: defaultdict(float))
for r in it:
    if r[ci_code] is None and r[ci_bal] is None:
        continue
    try:
        bal = float(r[ci_bal] or 0)
    except (TypeError, ValueError):
        bal = 0.0
    try:
        co = float(r[ci_co] or 0)
    except (TypeError, ValueError):
        co = 0.0
    if co != 0:
        continue
    n = norm(r[ci_code]) if r[ci_code] is not None else '(blank)'
    old = curmap_n.get(n, 'Ignore')
    new = loanmap_n.get(n)
    if new in (None, 'Exclude', 'Ignore'):
        continue
    cross[old][new] += bal

# fractions; hardcode 1:1 for pools with no/È extract signal
frac = {}
for old, d in cross.items():
    tot = sum(d.values())
    if tot > 0:
        frac[old] = {new: b / tot for new, b in d.items()}
frac.setdefault('Member Business Loans', {'Member Business Loans': 1.0})
frac.setdefault('Loan Participations', {'CRE Loan Participations': 1.0})
frac.setdefault('Collateral in Process of Liquidation', {'Collateral in Process of Liquidation': 1.0})
frac.setdefault('Credit Cards', {'Credit Cards': 1.0})
frac.setdefault('First Mortgage', {'Real Estate': 1.0})
frac.setdefault('Other Real Estate', {'Real Estate': 1.0})

print('=== crosswalk fractions (old -> new) ===')
for old in sorted(frac):
    print(' ', old, {k: round(v, 4) for k, v in frac[old].items()})

# ---- 3. Build new wide balance history ----
def read_hist():
    wb = openpyxl.load_workbook(HIST_F, read_only=True, data_only=True)['Sheet1']
    rows = list(wb.iter_rows(values_only=True))
    hdr = rows[0]
    dates = []
    for j in range(1, len(hdr)):
        dates.append((j, hdr[j]))
    series = {}
    for r in rows[1:]:
        if r[0]:
            series[str(r[0]).strip()] = r
    return dates, series

dates, series = read_hist()
# June old-pool balances from the single-month balance sheet (label col C=2, bal col H=7, header row 8)
def read_june():
    wb = openpyxl.load_workbook(JUNE_BS, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = {}
    for r in ws.iter_rows(min_row=1, values_only=True):
        if len(r) < 8:
            continue
        lbl, val = r[2], r[7]
        if lbl and isinstance(val, (int, float)):
            out[str(lbl).strip()] = float(val)
    return out
june = read_june()

OLD_LOAN_POOLS = ['Vehicle', 'First Mortgage', 'Unsecured', 'Other Real Estate',
                  'Other Secured', 'Credit Cards', 'Member Business Loans',
                  'Loan Participations', 'Collateral in Process of Liquidation']

# assemble old-pool monthly series (value per (pool, date)); add June column
col_dates = []
for j, d in dates:
    if d is None:
        continue
    # only keep populated months (Vehicle row non-null)
    if series['Vehicle'][j] is None:
        continue
    col_dates.append((j, d))
# build new-pool wide matrix
import datetime as dt
new_series = {p: [] for p in NEW_POOLS}
out_dates = []
for j, d in col_dates:
    out_dates.append(d)
    contrib = defaultdict(float)
    for old in OLD_LOAN_POOLS:
        val = series.get(old, [None] * 99)[j]
        if val is None:
            continue
        for new, f in frac.get(old, {}).items():
            contrib[new] += float(val) * f
    for p in NEW_POOLS:
        new_series[p].append(contrib.get(p, 0.0))
# June column
out_dates.append(dt.datetime(2026, 6, 30))
contrib = defaultdict(float)
for old in OLD_LOAN_POOLS:
    val = june.get(old)
    if val is None:
        continue
    for new, f in frac.get(old, {}).items():
        contrib[new] += float(val) * f
for p in NEW_POOLS:
    new_series[p].append(contrib.get(p, 0.0))

# write new wide file
owb = openpyxl.Workbook(); osh = owb.active; osh.title = 'Sheet1'
osh.cell(1, 1, 'Pool')
for k, d in enumerate(out_dates):
    osh.cell(1, 2 + k, d.strftime('%Y-%m-%d'))
for ri, p in enumerate(NEW_POOLS):
    osh.cell(2 + ri, 1, p)
    for k, v in enumerate(new_series[p]):
        osh.cell(2 + ri, 2 + k, round(v, 2))
owb.save(NEW_HIST_OUT)
print('\nWROTE new balance history:', NEW_HIST_OUT, '(%d months)' % len(out_dates))

# validation: new-pool sum vs old-loan-pool sum per month (last 3 + June)
print('\n=== validation: new-total vs old-total per month ===')
for idx in [0, len(out_dates)//2, len(out_dates)-2, len(out_dates)-1]:
    newt = sum(new_series[p][idx] for p in NEW_POOLS)
    if idx == len(out_dates)-1:
        oldt = sum(june.get(o, 0) for o in OLD_LOAN_POOLS)
    else:
        j = col_dates[idx][0]
        oldt = sum(float(series[o][j] or 0) for o in OLD_LOAN_POOLS)
    print('  {}  new=${:,.2f}  old=${:,.2f}  diff=${:,.2f}'.format(
        out_dates[idx].strftime('%Y-%m-%d'), newt, oldt, newt - oldt))

# ---- 4. Rebuild config ----
shutil.copy2(CFG, CFG + '.bak.newpools')
new_cfg = copy.deepcopy(cfg)
new_cfg['pool_map'] = loanmap
new_cfg['default_pool'] = 'Ignore'
def pool_entry(name):
    return {'name': name, 'risk_rated': name not in NOT_RR,
            'acl_months': 84 if name in RE_POOLS else 36, 'excluded': False,
            'use_default_mgmt_adj': name in {'Real Estate'}, 'brr': False}
new_cfg['pools'] = [pool_entry(p) for p in NEW_POOLS] + [
    {'name': 'Ignore', 'risk_rated': False, 'acl_months': None, 'excluded': False,
     'use_default_mgmt_adj': False, 'brr': False}]
new_cfg['pool_order'] = NEW_POOLS + ['Ignore']
new_cfg['not_risk_rated'] = sorted(NOT_RR - {'Ignore'})
new_cfg['acl_months_by_pool'] = {p: (84 if p in RE_POOLS else 36) for p in NEW_POOLS}
mb = dict(new_cfg['monthly_balance'])
mb.pop('monthly_file_pattern', None)
mb.pop('monthly_sheet', None); mb.pop('monthly_label_col', None)
mb.pop('monthly_balance_col', None); mb.pop('monthly_header_row', None)
mb.pop('monthly_strict_pool_map', None)
mb['source'] = 'single'; mb['sheet'] = 'Sheet1'; mb['header_row'] = 1
mb['pool_name_col'] = 'A'; mb['first_date_col'] = 'B'
mb['filename'] = os.path.basename(NEW_HIST_OUT); mb['saved_path'] = NEW_HIST_OUT
mb['pool_map'] = {p: p for p in NEW_POOLS}
new_cfg['monthly_balance'] = mb

with open(CFG, 'w', encoding='utf-8') as f:
    yaml.safe_dump(new_cfg, f, sort_keys=False, allow_unicode=True)
print('\nWROTE config (backup .bak.newpools). New pools:', NEW_POOLS)
print('reparse OK:', bool(yaml.safe_load(open(CFG, encoding='utf-8'))))
