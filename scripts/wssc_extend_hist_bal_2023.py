"""Rebuild WSSC 'Monthly Loan Balances by Type.xlsx' as a clean 30-month history
(2023-07 .. 2025-12) for the Display Hist Bal trend.

Source per month (chosen for reliability):
  * 2023-07 .. 2024-01  -> '07-2023 - 01-2024 Loan Totals.xlsx' (CU balance
                           summary; complete incl. CUMA mortgages). This also
                           replaces the anomalous DB 2024-01 ($31.8M, inflated).
  * 2024-02, 2025-05    -> INTERPOLATED per pool between good neighbours. These
                           two DB snapshots are known-bad (2024-02 missing the
                           ~$4.3M CUMA mortgages; 2025-05 a partial import of
                           only ~218 loans / $587K from a '(1)'-suffix file).
  * all other months    -> DB snapshots (complete).

Backs up the existing workbook first. Prints a monthly-total sanity table.
"""
import os, shutil, datetime
os.environ.setdefault("CECL_WORKSPACE_ROOT", r"Z:\Shared\TCT Files\CECL - CM Files")
import pandas as pd
from sqlalchemy import create_engine, text
from cecl_credentials import get_database_url
from openpyxl import load_workbook, Workbook
import yaml

CU = "WSSC FCU"
WSR = r"Z:\Shared\TCT Files\CECL - CM Files"
OUT = os.path.join(WSR, "Raw_Uploads", "wssc_fcu", "Monthly Loan Balances by Type.xlsx")
TOTALS = (r"Z:\Shared\Clients\Vizo Financial Corporate CU\Client Access\Clients"
          r"\WSSC FCU 16268\Portfolio Management\CECL-Migration IDLR"
          r"\07-2023 - 01-2024 Loan Totals.xlsx")
STAMP = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
# DB (2024-01..2025-12) is now clean after re-importing the bad months, so no
# interpolation is needed and Loan Totals is only used for the 2023 months.
BAD = set()

cfg = yaml.safe_load(open(os.path.join(WSR, "client_configs", "wssc_fcu.yaml"), encoding="utf-8"))
pool_map = {str(k).strip(): v for k, v in (cfg.get("pool_map") or {}).items()}
default_pool = cfg.get("default_pool", "Ignore")
DROP = {"Ignore", "Exclude"}
CODE_OVERRIDE = {"CUMA": "All Other Real Estate"}  # aggregate mortgage row -> AORE

def map_code(code):
    c = str(code).strip()
    return CODE_OVERRIDE.get(c) or pool_map.get(c, default_pool)

# ---- DB snapshots by pool ----
eng = create_engine(get_database_url())
db = pd.read_sql(text(
    "SELECT snapshot_date, loan_pool, SUM(current_balance) AS bal "
    "FROM monthly_loan_data WHERE credit_union=:cu GROUP BY snapshot_date, loan_pool"),
    eng, params={"cu": CU})
db["snapshot_date"] = pd.to_datetime(db["snapshot_date"])
db = db[~db["loan_pool"].isin(DROP)]
db_bal = {(r.snapshot_date, r.loan_pool): float(r.bal) for r in db.itertuples()}

# ---- Loan Totals -> early months by pool ----
wb = load_workbook(TOTALS, read_only=True, data_only=True); ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True)); wb.close()
hdr = rows[0]
# Only take the 2023 month columns from Loan Totals; 2024-01 onward comes from
# the (now clean) DB snapshots.
early_cols = {ci: pd.Timestamp(h) + pd.offsets.MonthEnd(0)
             for ci, h in enumerate(hdr)
             if isinstance(h, datetime.datetime) and h.year == 2023}
early = {}
for r in rows[1:]:
    if not r or r[0] in (None, "", "Type"):
        continue
    pool = map_code(r[0])
    if pool in DROP:
        continue
    for ci, ts in early_cols.items():
        v = r[ci] if ci < len(r) else None
        if isinstance(v, (int, float)):
            early[(ts, pool)] = early.get((ts, pool), 0.0) + float(v)
early_dates = sorted(set(early_cols.values()))          # 2023-07 .. 2024-01

# ---- Master month list & pools ----
db_dates = sorted(db["snapshot_date"].unique())
all_dates = sorted(set(early_dates) | set(pd.Timestamp(d) for d in db_dates))
pools = sorted(set(p for (_, p) in early) | set(db["loan_pool"].unique()))

def raw_bal(pool, ts):
    """Balance from the preferred source, or None for a bad/absent month."""
    if ts in early_dates:
        return early.get((ts, pool), 0.0)
    if ts in BAD:
        return None
    return db_bal.get((ts, pool), 0.0)

def bal_for(pool, ts):
    v = raw_bal(pool, ts)
    if v is not None:
        return v
    i = all_dates.index(ts)
    lo = next((raw_bal(pool, all_dates[j]) for j in range(i - 1, -1, -1)
               if raw_bal(pool, all_dates[j]) is not None), None)
    hi = next((raw_bal(pool, all_dates[j]) for j in range(i + 1, len(all_dates))
               if raw_bal(pool, all_dates[j]) is not None), None)
    if lo is not None and hi is not None:
        return (lo + hi) / 2.0
    return lo if lo is not None else (hi or 0.0)

# ---- Write workbook ----
if os.path.exists(OUT):
    shutil.copy2(OUT, OUT + f".bak.phase_extend2023clean_{STAMP}")
    print("Backup ->", os.path.basename(OUT) + f".bak.phase_extend2023clean_{STAMP}")
out_wb = Workbook(); out_ws = out_wb.active; out_ws.title = "Loan Balances by Pool"
out_ws.cell(row=1, column=1, value="Pool")
for j, d in enumerate(all_dates):
    out_ws.cell(row=1, column=2 + j, value=pd.Timestamp(d).to_pydatetime())
r = 2
for pool in pools:
    out_ws.cell(row=r, column=1, value=str(pool))
    for j, d in enumerate(all_dates):
        out_ws.cell(row=r, column=2 + j, value=round(bal_for(pool, pd.Timestamp(d)), 2))
    r += 1
out_ws.cell(row=r, column=1, value="ACL Balance")
for j, d in enumerate(all_dates):
    out_ws.cell(row=r, column=2 + j, value=0.0)
out_wb.save(OUT)

print(f"\nSaved: {OUT}")
print(f"Range: {all_dates[0].date()} -> {all_dates[-1].date()} "
      f"({len(all_dates)} months, {len(pools)} pools)\n")
print("Monthly total (sanity):")
for d in all_dates:
    tot = sum(bal_for(p, pd.Timestamp(d)) for p in pools)
    tag = "  <- interpolated" if d in BAD else ("  <- Loan Totals" if d in early_dates else "")
    print(f"  {d.date()}  ${tot:>14,.0f}{tag}")
