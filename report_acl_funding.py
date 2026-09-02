"""
ACL Funding Worksheet (between-report allowance funding schedule).

Standalone Excel workbook — a companion to the CECL Migration report,
delivered like the Management Adjustment Worksheet. It lays out three
months (the report month plus the two following) so the credit union can
track how much to fund the Allowance for Credit Losses between quarterly
reports.

Columns (one row per month):
  Month | Total Allowance Needed | Allowance as of Date (Current ACL
  Balance) | Adjustment (Needed - Current) | Actual Adjustment (CU entry)
  | Adjusted Balance | Difference from Allowance Needed

The report month's Total Allowance Needed and ACL Balance are pulled from
the report calculation; the off months are left for the credit union to
complete. Adjustment / Adjusted Balance / Difference are live Excel
formulas.

Public entry point:

    compose_acl_funding(client_name, snapshot_date, df, config,
                        grades, hist) -> (wb, fname)

Follows the same composer signature as ``report_vizo.compose_vizo_main``
and ``report_mgmt_adj_napkin.compose_mgmt_adj_napkin`` so it can be
dispatched from ``generate_report.generate_report`` with no new
data-loading code.
"""
from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# Reuse the shared ACL totals helper so Total Allowance Needed and the ACL
# Balance match the ACL Env by Pool Mgmt Adj tab exactly.
from report_vizo import _compute_acl_totals

# ── Styling ────────────────────────────────────────────────────────
HDR_FILL   = PatternFill('solid', fgColor='0D4D5E')   # Vizo accent1 teal
SUB_FILL   = PatternFill('solid', fgColor='2E7D7B')
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')   # CU enters
PULL_FILL  = PatternFill('solid', fgColor='E2EFDA')   # from report
CALC_FILL  = PatternFill('solid', fgColor='DEEAF6')   # calculated

FNT_TITLE = Font(name='Calibri', bold=True, size=16, color='FFFFFF')
FNT_SUB   = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
FNT_HDR   = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
FNT_MONTH = Font(name='Calibri', bold=True, size=12)
FNT_CELL  = Font(name='Calibri', size=12)
FNT_LEG   = Font(name='Calibri', bold=True, size=10)
FNT_NOTE  = Font(name='Calibri', size=9, italic=True, color='595959')

_thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
ACCT2 = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'


def compose_acl_funding(client_name, snapshot_date, df, config, grades, hist):
    """Build the ACL Funding Worksheet workbook.

    Returns ``(wb, fname)`` so it can be saved by the standard
    ``generate_report.generate_report`` dispatcher alongside the TCT /
    Vizo / Vizo-Supplemental / Management Adjustment outputs.
    """
    cu = config.get('credit_union', client_name)

    acl = _compute_acl_totals(df, grades, config, hist, snapshot_date)
    total_allow_needed = acl.get('total_allowance_needed', 0) or 0
    acl_bal = acl.get('acl_balance', 0) or 0

    # Report month + the two following months.
    try:
        base = datetime.strptime(str(snapshot_date)[:10], '%Y-%m-%d')
    except ValueError:
        base = datetime.today()
    months = []
    for i in range(3):
        mm = base.month + i
        yy = base.year + (mm - 1) // 12
        mm = (mm - 1) % 12 + 1
        months.append(datetime(yy, mm, 1))

    wb = Workbook()
    ws = wb.active
    ws.title = "ACL Funding Worksheet"
    ws.sheet_view.showGridLines = False

    cols = {
        'A': ('Month', 22),
        'B': ('Total Allowance\nNeeded', 18),
        'C': ('Allowance as of Date\n(Current ACL Balance)', 22),
        'D': ('Adjustment\n(Needed \u2212 Current)', 18),
        'E': ('Actual Adjustment\n(CU Entry)', 18),
        'F': ('Adjusted\nBalance', 18),
        'G': ('Difference from\nAllowance Needed', 18),
    }
    for letter, (_lbl, width) in cols.items():
        ws.column_dimensions[letter].width = width

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center')
    right = Alignment(horizontal='right', vertical='center')

    # ── Title block ──
    ws.merge_cells('A1:G1')
    ws['A1'] = 'CECL Allowance Funding Worksheet'
    ws['A1'].font = FNT_TITLE
    ws['A1'].alignment = center
    ws['A1'].fill = HDR_FILL
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:G2')
    ws['A2'] = (f'{cu}  \u2022  CECL Migration Model  \u2022  '
                f'Between-Report Allowance Funding Schedule')
    ws['A2'].font = FNT_SUB
    ws['A2'].alignment = center
    ws['A2'].fill = SUB_FILL
    ws.row_dimensions[2].height = 20

    ws.merge_cells('A3:G3')
    ws['A3'] = (f'Report period {snapshot_date}.  {months[0].strftime("%B %Y")} '
                f'figures are pulled from the CECL report; the following months '
                f'are estimated between quarterly reports.')
    ws['A3'].font = FNT_NOTE
    ws['A3'].alignment = center
    ws.row_dimensions[3].height = 16

    # ── Header row ──
    hdr_row = 5
    for letter, (label, _w) in cols.items():
        c = ws[f'{letter}{hdr_row}']
        c.value = label
        c.font = FNT_HDR
        c.fill = HDR_FILL
        c.alignment = center
        c.border = BORDER
    ws.row_dimensions[hdr_row].height = 34

    # ── Data rows ──
    first_data = hdr_row + 1
    for i, mdt in enumerate(months):
        r = first_data + i
        ws.row_dimensions[r].height = 22

        a = ws[f'A{r}']
        a.value = mdt.strftime('%B %Y')
        a.font = FNT_MONTH
        a.alignment = left
        a.border = BORDER

        b = ws[f'B{r}']
        if i == 0:
            b.value = total_allow_needed
            b.fill = PULL_FILL
        else:
            b.value = f'=B{first_data}'
            b.fill = CALC_FILL
        b.number_format = ACCT2
        b.alignment = right
        b.border = BORDER
        b.font = FNT_CELL

        c = ws[f'C{r}']
        if i == 0:
            c.value = acl_bal
            c.fill = PULL_FILL
        else:
            c.value = None
            c.fill = INPUT_FILL
        c.number_format = ACCT2
        c.alignment = right
        c.border = BORDER
        c.font = FNT_CELL

        d = ws[f'D{r}']
        d.value = f'=B{r}-C{r}'
        d.number_format = ACCT2
        d.alignment = right
        d.border = BORDER
        d.fill = CALC_FILL
        d.font = FNT_CELL

        e = ws[f'E{r}']
        e.value = None
        e.fill = INPUT_FILL
        e.number_format = ACCT2
        e.alignment = right
        e.border = BORDER
        e.font = FNT_CELL

        f = ws[f'F{r}']
        f.value = f'=C{r}+E{r}'
        f.number_format = ACCT2
        f.alignment = right
        f.border = BORDER
        f.fill = CALC_FILL
        f.font = FNT_CELL

        g = ws[f'G{r}']
        g.value = f'=F{r}-B{r}'
        g.number_format = ACCT2
        g.alignment = right
        g.border = BORDER
        g.fill = CALC_FILL
        g.font = FNT_CELL

    # ── Legend ──
    legend_row = first_data + len(months) + 1
    ws[f'A{legend_row}'] = 'Legend:'
    ws[f'A{legend_row}'].font = FNT_LEG
    legends = [
        (PULL_FILL, 'Pulled from the CECL report'),
        (INPUT_FILL, 'Entered by the credit union'),
        (CALC_FILL, 'Calculated automatically'),
    ]
    for j, (fill, text) in enumerate(legends):
        rr = legend_row + 1 + j
        key = ws[f'A{rr}']
        key.fill = fill
        key.border = BORDER
        ws.merge_cells(f'B{rr}:D{rr}')
        lbl = ws[f'B{rr}']
        lbl.value = text
        lbl.font = FNT_NOTE
        lbl.alignment = left

    # ── Notes ──
    note_row = legend_row + len(legends) + 2
    notes = [
        'Total Allowance Needed is held constant across all three months (per the report).',
        'For the off months, enter the Current ACL Balance (Allowance as of that date).',
        'Adjustment = Total Allowance Needed \u2212 Current ACL Balance (the funding gap).',
        'Enter the Actual Adjustment the credit union books for the month.',
        'Adjusted Balance = Current ACL Balance + Actual Adjustment.',
        'Difference from Allowance Needed = Adjusted Balance \u2212 Total Allowance Needed '
        '(negative means still underfunded).',
    ]
    for k, text in enumerate(notes):
        r = note_row + k
        ws.merge_cells(f'A{r}:G{r}')
        cell = ws[f'A{r}']
        cell.value = f'\u2022  {text}'
        cell.font = FNT_NOTE
        cell.alignment = left

    last_row = note_row + len(notes) - 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.4,
                                  header=0.3, footer=0.3)
    ws.print_area = f'A1:G{last_row}'

    # File name follows the Management Adjustment Worksheet convention.
    safe_cu = cu.replace('/', '-').replace('\\', '-')
    snap_prefix = snapshot_date[:7]   # "YYYY-MM"
    fname = f"{snap_prefix} ACL Funding Worksheet - {safe_cu}.xlsx"
    return wb, fname
