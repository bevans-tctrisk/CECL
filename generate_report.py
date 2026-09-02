"""
CECL Report Generator - Three Report Types

Generates TCT-format and Vizo-format CECL reports.

Report Types:
  tct      - Full CECL-Migration-WARM report (Franklin Trust style, single file)
  vizo     - CECL Credit Migration Report (Credit Union B style, main report)
  vizo_supp - CECL Supplemental Report (Credit Union B style, supplemental)

Usage:
    python generate_report.py --client franklin --date 2025-12-31
    python generate_report.py --client franklin --reports tct
    python generate_report.py --client franklin --reports tct vizo vizo_supp
    python generate_report.py --all --date 2025-12-31
    python generate_report.py --list
"""
import os, re, argparse, glob
from import_data import _detect_text_encoding, _sniff_delimiter
from datetime import datetime, date
import numpy as np
import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text
from dotenv import load_dotenv
from cecl_engine import (
    calculate_cecl, risk_change_matrix, pool_summary,
    migration_summary_by_pool, grade_distribution,
    years_on_books, principal_paid, build_grade_order,
)
from report_tct import compose_tct as compose_tct_new
from report_vizo import compose_vizo_main as compose_vizo_main_new, compose_vizo_supp as compose_vizo_supp_new, patch_impdet_charts, patch_drawing_onecell_to_twocell, patch_dq_pie_zero_labels, patch_remove_chart_borders_and_axis_lines
from fetch_econ_data import fetch_economic_data
from cecl_audit_log import get_audit_logger, log_report_generation, log_session_start, log_session_end
from cecl_credentials import get_database_url

load_dotenv()
# When the code clone lives on a local SSD (e.g. C:\Dev\CECL) and the
# analyst data (client_configs/, Reports/, Raw_Uploads/, ...) lives on a
# shared drive (e.g. Egnyte), CECL_WORKSPACE_ROOT points at the data root.
# Falls back to the historical 'code dir == data dir' layout when unset.
BASE = os.environ.get('CECL_WORKSPACE_ROOT') or os.path.dirname(os.path.abspath(__file__))
CFG_DIR = os.path.join(BASE, 'client_configs')
RPT_DIR = os.path.join(BASE, 'Reports')
engine = create_engine(get_database_url())


def resolve_path(path_value, base=BASE):
    """Resolve configured paths: keep absolute paths, join relative paths to base."""
    if not path_value:
        return ''
    return path_value if os.path.isabs(path_value) else os.path.join(base, path_value)

# ── Styling Constants ──────────────────────────────────────────────
TITLE_FONT = Font(name='Calibri', bold=True, size=18, color='1B4F72')
HDR_FONT = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
SUB_FONT = Font(name='Calibri', bold=True, size=11)
NORM = Font(name='Calibri', size=10)
MONEY = '#,##0'
MONEY2 = '#,##0.00'
PCT = '0.00%'
PCT4 = '0.0000%'
HDR_FILL = PatternFill('solid', fgColor='1B4F72')
ALT_FILL = PatternFill('solid', fgColor='D6EAF8')
IMP_FILL = PatternFill('solid', fgColor='D5F5E3')
DET_FILL = PatternFill('solid', fgColor='FADBD8')
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
TEAL_FILL = PatternFill('solid', fgColor='1A5276')
DARK_FILL = PatternFill('solid', fgColor='2C3E50')

# ── Environmental Factor Score Tables ──────────────────────────────
# Net Credit Change scoring (same in TCT and Vizo)
NCC_RANGES = [
    (-999, -18.00, 7), (-18.00, -16.00, 6), (-16.00, -14.00, 5),
    (-14.00, -11.00, 4), (-11.00, -8.00, 3), (-8.00, -6.00, 2),
    (-6.00, -4.00, 1), (-4.00, 4.00, 0), (4.00, 6.00, -1),
    (6.00, 8.00, -2), (8.00, 9.00, -3), (9.00, 11.00, -4),
    (11.00, 13.00, -5), (13.00, 15.00, -6), (15.00, 999, -7),
]
# Delinquency scoring
DQ_RANGES = [
    (5.00, 999, 20), (4.00, 5.00, 17), (3.00, 4.00, 12),
    (2.50, 3.00, 8), (2.00, 2.50, 4), (1.50, 2.00, 2.5),
    (1.00, 1.50, 1.5), (0.50, 1.00, 0.75), (-0.50, 0.50, 0),
    (-1.00, -0.50, -0.75), (-1.50, -1.00, -1.5), (-2.00, -1.50, -2.5),
    (-2.50, -2.00, -4), (-3.00, -2.50, -8), (-4.00, -3.00, -12),
    (-5.00, -4.00, -17), (-999, -5.00, -20),
]
# Economic Stress scoring
ES_RANGES = [
    (25.00, 999, 10), (24.00, 25.00, 8), (22.00, 24.00, 7),
    (20.00, 22.00, 6), (18.00, 20.00, 5), (16.00, 18.00, 4),
    (14.00, 16.00, 3.5), (12.00, 14.00, 3), (10.00, 12.00, 2),
    (8.00, 10.00, 1), (6.00, 8.00, 0), (4.00, 6.00, 0),
    (2.00, 4.00, -1), (0.00, 2.00, -2),
]
# Standard TCT Distribution Factors per grade position
DIST_FACTORS = [10.52, 22.93, 45.15, 116.10, 141.17, 152.04, 160.21]


def load_workbook_resilient(path, **kwargs):
    """load_workbook that survives transient SMB read errors on shared drives.

    Reading an .xlsx directly off a network path (e.g. Z:) can raise
    OSError [Errno 22] mid-stream. Retry, then fall back to reading a local
    temp copy. Not used with read_only (the temp copy is deleted on return).
    """
    import shutil as _shutil
    import tempfile as _tempfile
    from openpyxl import load_workbook as _load
    last_exc = None
    for _ in range(3):
        try:
            return _load(path, **kwargs)
        except OSError as exc:
            last_exc = exc
    tmp = None
    try:
        fd, tmp = _tempfile.mkstemp(suffix=os.path.splitext(path)[1] or '.xlsx')
        os.close(fd)
        _shutil.copyfile(path, tmp)
        return _load(tmp, **kwargs)
    except OSError:
        raise last_exc
    finally:
        if tmp and os.path.isfile(tmp):
            try:
                os.remove(tmp)
            except OSError:
                pass


def score_from_ranges(value, ranges):
    """Look up a score from a range table."""
    v = value * 100 if abs(value) < 1 else value  # handle both 0.05 and 5.0
    for lo, hi, score in ranges:
        if lo <= v < hi:
            return score
    return 0


# ── Data Loading ───────────────────────────────────────────────────
def load_config(client):
    with open(os.path.join(CFG_DIR, f'{client}.yaml'), 'r', encoding='utf-8') as f:
        cfg = yaml.safe_load(f)
    # Normalize the 'Ignore' sentinel: any pool_map value (or default_pool)
    # set to 'Ignore' is rewritten to the existing 'Exclude' sentinel, which
    # is already dropped everywhere downstream (per-pool sheets,
    # _apply_excluded_pools, HIDE/Exclude prefix filters).
    pm = cfg.get('pool_map') or {}
    if any(v == 'Ignore' for v in pm.values()):
        cfg['pool_map'] = {
            k: ('Exclude' if v == 'Ignore' else v) for k, v in pm.items()
        }
    if cfg.get('default_pool') == 'Ignore':
        cfg['default_pool'] = 'Exclude'
    # Also strip any pool literally named "Ignore" or "Exclude" from the
    # pool registries (pools/pool_order/not_risk_rated/risk_rated). The
    # wizard's Loan Pools step lets users define a sentinel pool called
    # "Ignore" meaning "drop these loan codes"; downstream pool enumerators
    # iterate cfg['pools']/cfg['pool_order'] directly and would render an
    # empty "Ignore" row/column on every per-pool sheet without this.
    _SENTINELS = {'ignore', 'exclude'}
    pools_list = cfg.get('pools')
    if isinstance(pools_list, list):
        cfg['pools'] = [
            p for p in pools_list
            if not (
                (isinstance(p, dict)
                 and str(p.get('name', '')).strip().lower() in _SENTINELS)
                or (isinstance(p, str) and p.strip().lower() in _SENTINELS)
            )
        ]
    for _key in ('risk_rated', 'not_risk_rated', 'pool_order'):
        _val = cfg.get(_key)
        if isinstance(_val, list):
            cfg[_key] = [
                p for p in _val
                if not (isinstance(p, str) and p.strip().lower() in _SENTINELS)
            ]
    # Honor excluded_pools by remapping any pool_map value matching an
    # excluded pool name to the existing 'Exclude' sentinel. All downstream
    # filters (HIDE/Exclude prefix checks throughout report_tct/generate_report)
    # already drop 'Exclude' rows from balances, charge-offs, recoveries, and
    # per-pool sheets, so this single rewrite excludes them from the entire
    # analysis without touching every individual filter site.
    excl = set((cfg.get('excluded_pools') or []))
    if excl:
        pm = cfg.get('pool_map') or {}
        cfg['pool_map'] = {
            k: ('Exclude' if v in excl else v) for k, v in pm.items()
        }
        if cfg.get('default_pool') in excl:
            cfg['default_pool'] = 'Exclude'
        # Also remove excluded pools from the pool registries so the report
        # engine's pool enumerators (which iterate ``cfg['pools']`` /
        # ``risk_rated`` / ``not_risk_rated`` directly) don't render empty
        # columns/rows/sheets for them. Without this, _apply_excluded_pools
        # drops the *loan rows* but the *pool name* still shows up as a
        # zero-balance bucket on sheets like "ACL Env by Pool Mgmt Adj".
        pools_list = cfg.get('pools')
        if isinstance(pools_list, list):
            cfg['pools'] = [
                p for p in pools_list
                if not (
                    (isinstance(p, dict) and p.get('name') in excl)
                    or (isinstance(p, str) and p in excl)
                )
            ]
        for key in ('risk_rated', 'not_risk_rated', 'pool_order'):
            val = cfg.get(key)
            if isinstance(val, list):
                cfg[key] = [p for p in val if p not in excl]
    return cfg

def list_clients():
    return sorted(os.path.splitext(f)[0] for f in os.listdir(CFG_DIR)
                  if f.endswith('.yaml') and not f.startswith('_'))

def load_loans(cu, snap=None, config=None):
    if snap:
        df = pd.read_sql(text("SELECT * FROM monthly_loan_data WHERE credit_union=:c AND snapshot_date=:s"),
                         engine, params={"c": cu, "s": snap})
    else:
        df = pd.read_sql(text("SELECT * FROM monthly_loan_data WHERE credit_union=:c"), engine, params={"c": cu})
    # Older databases may predate the Business Risk Rating column. Surface
    # it as an all-None column so downstream BRR-aware code can rely on
    # the field existing without doing its own column-check.
    if 'business_risk_rating' not in df.columns:
        df['business_risk_rating'] = None
    return _apply_excluded_pools(df, config)


def _apply_excluded_pools(df, config):
    """Drop any loan rows whose ``loan_pool`` is listed in
    ``cfg['excluded_pools']`` (and any rows already tagged with the legacy
    'Exclude' sentinel from import-time pool_map remapping).

    Existing DB rows were categorized at import time using the original
    pool_map, so configs that excluded pools *after* import would otherwise
    still include those rows in every per-pool sheet that just iterates
    ``df['loan_pool'].unique()``. Dropping them here makes the runtime
    config authoritative across the whole report.
    """
    if df is None or df.empty or 'loan_pool' not in df.columns:
        return df
    excl = set((config.get('excluded_pools') or [])) if config else set()
    # 'Exclude' is the canonical sentinel; 'Ignore' is the legacy/wizard
    # synonym (load_config rewrites Ignore -> Exclude in pool_map and
    # default_pool, but historical DB rows imported under default_pool='Ignore'
    # still carry loan_pool='Ignore' verbatim). Treat both as drops.
    excl.add('Exclude')
    excl.add('Ignore')
    mask = df['loan_pool'].isin(excl)
    if not mask.any():
        return df
    return df.loc[~mask].copy()

def _load_prior_brr_lookup(cu, snap, brr_pools):
    """Return ``{member_number_str: prior_business_risk_rating}`` for
    BRR-flagged pool loans at the most recent snapshot strictly before
    ``snap`` for ``cu``. Empty dict on the very first report (no prior
    snapshot) or when no BRR pools are configured.

    Used by the per-loan BRR migration logic so the Risk Change tab can
    show quarter-over-quarter rating movement for business pools.
    """
    if not brr_pools or snap is None:
        return {}
    try:
        with engine.connect() as c:
            row = c.execute(
                text(
                    "SELECT MAX(snapshot_date) FROM monthly_loan_data "
                    "WHERE credit_union=:cu AND snapshot_date < :snap"
                ),
                {"cu": cu, "snap": snap},
            ).fetchone()
            prior = row[0] if row and row[0] else None
            if not prior:
                return {}
            rows = c.execute(
                text(
                    "SELECT member_number, business_risk_rating "
                    "FROM monthly_loan_data "
                    "WHERE credit_union=:cu AND snapshot_date=:prior "
                    "AND loan_pool = ANY(:pools) "
                    "AND business_risk_rating IS NOT NULL"
                ),
                {
                    "cu": cu,
                    "prior": str(prior),
                    "pools": list(brr_pools),
                },
            ).fetchall()
            print(f"    Prior BRR snapshot: {prior} ({len(rows)} loan(s))")
    except Exception as exc:
        print(f"    [warn] prior-BRR lookup failed: {exc}")
        return {}
    lookup = {}
    for member, brr in rows:
        if member is None:
            continue
        key = str(member).strip()
        if not key:
            continue
        # Last write wins on duplicate member_number — should be rare
        # since member_number is the full account string.
        lookup[key] = brr
    return lookup


def latest_date(cu):
    with engine.connect() as c:
        r = c.execute(text("SELECT MAX(snapshot_date) FROM monthly_loan_data WHERE credit_union=:c"), {"c": cu}).fetchone()
    return str(r[0]) if r and r[0] else None

def all_dates(cu):
    with engine.connect() as c:
        rows = c.execute(text("SELECT DISTINCT snapshot_date FROM monthly_loan_data WHERE credit_union=:c ORDER BY snapshot_date DESC"), {"c": cu}).fetchall()
    return [str(r[0]) for r in rows]


# ── Historical Data Loading ───────────────────────────────────────

def _find_quarter_folders(data_dir):
    """Find all quarterly data folders under the data directory.
    Returns list of (folder_path, quarter_label) sorted by date."""
    quarters = []
    for root, dirs, files in os.walk(data_dir):
        folder = os.path.basename(root)
        # Match patterns like 2024-03, 2024-06, 2022-10, 2022-12, 2023-03, etc.
        m = re.match(r'^(\d{4})-(\d{2})$', folder)
        if m:
            quarters.append((root, f"{m.group(1)}-{m.group(2)}"))
    return sorted(quarters, key=lambda x: x[1])


def _read_csv_any(path, **kw):
    """``pd.read_csv`` that survives core-system exports.

    DEXA / TCT report exports are frequently UTF-16 (LE) with a BOM and
    tab-delimited despite carrying a ``.csv`` extension. A plain read rejects
    those with ``0xff at position 0``; every caller here swallows the error
    and simply loses the file, so a credit union silently reports no
    charge-offs (Utah lost two quarters this way).

    Only a file carrying a BOM takes the detected path, so ordinary UTF-8
    comma CSVs continue to parse exactly as before.
    """
    enc = _detect_text_encoding(path)
    if not enc:
        return pd.read_csv(path, **kw)
    kw.setdefault('sep', _sniff_delimiter(path))
    return pd.read_csv(path, encoding=enc, engine='python', **kw)


def _read_data_file(filepath):
    """Read an Excel or CSV file, returning a DataFrame with no header.
    For Excel files with multiple sheets, concatenates all sheets that share
    the maximum column count (so multi-sheet quarterly files are fully read).
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.csv':
        return _read_csv_any(filepath, header=None)
    xl = pd.ExcelFile(filepath)
    parts = []
    for s in xl.sheet_names:
        d = pd.read_excel(xl, sheet_name=s, header=None)
        if not d.empty:
            parts.append(d)
    if not parts:
        return pd.DataFrame()
    if len(parts) == 1:
        return parts[0]
    target_cols = max(d.shape[1] for d in parts)
    parts = [d for d in parts if d.shape[1] == target_cols]
    # Drop sheets whose content exactly duplicates an earlier sheet. Some
    # CUs keep a working copy as a second tab (e.g. a "…File 201 (2)" sheet
    # alongside "…File 2018.01"); concatenating both would double-count every
    # charge-off / recovery row. Dedupe by (shape, content-hash) so only
    # byte-identical sheets are dropped — genuinely different tabs are kept.
    unique_parts = []
    seen_keys = set()
    for d in parts:
        try:
            content_hash = int(pd.util.hash_pandas_object(
                d.fillna(""), index=False).sum())
        except Exception:  # noqa: BLE001
            content_hash = None
        key = (d.shape, content_hash)
        if content_hash is not None and key in seen_keys:
            continue
        seen_keys.add(key)
        unique_parts.append(d)
    parts = unique_parts
    if len(parts) == 1:
        return parts[0]
    return pd.concat(parts, ignore_index=True)


def _looks_like_loan_code(val):
    """Return True if val looks like a 1-5 char loan code (has at least one letter, not pure digits)."""
    if not isinstance(val, str):
        return False
    s = val.strip()
    if not (1 <= len(s) <= 5):
        return False
    if s.isdigit():
        return False
    return any(ch.isalpha() for ch in s)


def _coerce_mixed_dates(values):
    """Parse a column that mixes ``datetime`` objects and Excel serial
    numbers (e.g. 41091 -> 2012-06-30) into a tz-naive Timestamp Series.

    Honolulu's CO / Recovery tracking workbooks have both formats
    interleaved on the same column, which ``pd.to_datetime`` alone
    misinterprets (it treats raw ints as nanoseconds and produces 1970).
    """
    s = pd.Series(values)
    # Identify cells that are numeric in the Excel-serial range
    # (~1927..~2119) up-front and route them through the Excel-origin
    # parser. Everything else goes through the standard parser.
    nums = pd.to_numeric(s, errors='coerce')
    serial_mask = nums.between(10000, 80000) & ~s.apply(
        lambda v: isinstance(v, (pd.Timestamp, datetime)))
    out = pd.to_datetime(s.mask(serial_mask), errors='coerce')
    if serial_mask.any():
        converted = pd.to_datetime(
            nums[serial_mask], unit='D', origin='1899-12-30', errors='coerce')
        out.loc[converted.index] = converted
    return out


def _is_numeric_or_date(v):
    """True if ``v`` is a non-null number, datetime, or numeric-looking
    string. Used by the account-column filter so that rows whose
    account_col is actually a date (e.g. recovery files with no
    account number) are not dropped.
    """
    if pd.isna(v):
        return False
    if isinstance(v, (pd.Timestamp, datetime)):
        return True
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        return v.replace('-', '').replace(' ', '').isdigit()
    return False


def _resolve_hist_cols_by_header(header_row, kind):
    """Locate charge-off / recovery columns by HEADER TEXT.

    Some CUs drift column positions between months (one month inserts or
    drops a blank column), which silently mis-reads a fixed-index
    ``historical_file_formats`` config. The header labels stay consistent,
    so they are the reliable anchor. Returns a dict with any of
    ``{account, code, amount, date}`` whose label was confidently matched;
    callers keep their configured index for anything not found. ``kind`` is
    ``'co'`` (charge-off amount) or ``'rc'`` (recovery amount).
    """
    labels = [str(v).strip().lower() if pd.notna(v) else '' for v in header_row]
    out = {}
    for i, lab in enumerate(labels):
        if not lab:
            continue
        if 'account' not in out and ('account' in lab or 'acct' in lab):
            out['account'] = i
        if ('loan type code' in lab or 'loan code' in lab
                or (lab.endswith('code') and 'account' not in lab)):
            out['code'] = i
        if 'date' in lab and 'date' not in out:
            out['date'] = i
        if kind == 'co':
            if ('charge off amount' in lab or 'chargeoff amount' in lab
                    or 'charge-off amount' in lab):
                out['amount'] = i
        else:
            if 'recover' in lab:
                out['amount'] = i
    return out


def _parse_chargeoff_file(filepath, parse_config=None):
    """Parse a charge-off file (varying formats). Returns DataFrame with [code, amount, date]."""
    df = _read_data_file(filepath)
    if df.empty:
        return pd.DataFrame(columns=['code', 'amount', 'date'])

    if parse_config:
        cfg_df = df.copy()
        if parse_config.get('has_header', False):
            cfg_df = cfg_df.iloc[1:]

        skip_rows = int(parse_config.get('skip_rows', 0) or 0)
        if skip_rows > 0:
            cfg_df = cfg_df.iloc[skip_rows:]

        account_col = parse_config.get('account_col', 0)
        code_col = parse_config.get('code_col')
        amount_col = parse_config.get('amount_col')
        date_col = parse_config.get('date_col')

        # Trust consistent header TEXT over fixed indices when the file has
        # a header row — protects against month-to-month column drift.
        # Skipped when ``strict_columns`` is set: multi-format entries wire
        # explicit indices for combined files whose CO and recovery amount
        # columns differ, and the header heuristic (which keys off the word
        # "amount") can't tell those two columns apart — it would misroute.
        if (parse_config.get('has_header', False) and len(df) > 0
                and not parse_config.get('strict_columns')):
            _hdr = _resolve_hist_cols_by_header(df.iloc[0], 'co')
            if 'amount' in _hdr and 'code' in _hdr:
                account_col = _hdr.get('account', account_col)
                code_col = _hdr['code']
                amount_col = _hdr['amount']
                if 'date' in _hdr:
                    date_col = _hdr['date']

        # Files with no loan-code column (e.g. a credit-card CO file where
        # every row is implicitly the same pool) set ``code_static`` instead;
        # each parsed row is stamped with it and mapped through pool_map.
        code_static = parse_config.get('code_static')

        ncols = cfg_df.shape[1]
        code_valid = code_col is not None and 0 <= int(code_col) < ncols
        amount_valid = amount_col is not None and 0 <= int(amount_col) < ncols
        date_valid = date_col is not None and 0 <= int(date_col) < ncols
        account_valid = account_col is not None and 0 <= int(account_col) < ncols

        if amount_valid and (code_valid or code_static not in (None, '')):
            if account_valid:
                acct_series = cfg_df.iloc[:, account_col]
                acct_numeric = acct_series.apply(_is_numeric_or_date)
                cfg_df = cfg_df[acct_numeric]

            if not cfg_df.empty:
                code_vals = (cfg_df.iloc[:, code_col].values if code_valid
                             else [code_static] * len(cfg_df))
                result = pd.DataFrame({
                    'code': code_vals,
                    'amount': pd.to_numeric(cfg_df.iloc[:, amount_col], errors='coerce').values,
                })
                if date_valid:
                    result['date'] = _coerce_mixed_dates(
                        cfg_df.iloc[:, date_col].values).values
                else:
                    result['date'] = pd.NaT
                result = result.dropna(subset=['amount'])
                if not result.empty:
                    return result
            # strict_columns: the explicit amount_col is authoritative. An
            # empty result means "no charge-offs in this file", NOT a cue to
            # fall through to auto-detect (which in a combined Charge-Off /
            # Recovery workbook would wrongly grab the recovery-amount column).
            if parse_config.get('strict_columns'):
                return pd.DataFrame(columns=['code', 'amount', 'date'])

    # Check if first row is a header.  Require col 0 to NOT be numeric (a real
    # header has 'Account Number'-style text in col 0; a data row starts with
    # an account number).  This prevents trailing comment cells like
    # "...use Col D for charge off amounts." from triggering a false header.
    first_row = df.iloc[0]
    first_a = first_row.iloc[0] if len(first_row) > 0 else None
    col0_is_numeric = isinstance(first_a, (int, float)) and pd.notna(first_a)
    first_vals = [str(v).lower() if pd.notna(v) else '' for v in first_row]
    has_header = (not col0_is_numeric) and any(
        'account' in v or 'charge' in v or 'loan' in v or 'acct' in v
        for v in first_vals
    )

    # Use header keywords to identify columns when header is present
    code_col = amount_col = date_col = None
    if has_header:
        for c, hdr in enumerate(first_vals):
            # Skip the FICO score column (contains 'sc'/'score' but is not a loan code)
            if 'fico' in hdr or 'score' in hdr:
                continue
            if any(k in hdr for k in ('security', 'code')) and 'account' not in hdr and 'sub' not in hdr:
                code_col = c
            elif any(k in hdr for k in ('amount', 'chg off am', 'chargeoff am', 'pymt', 'principal')):
                amount_col = c
            elif 'date' in hdr or 'effective' in hdr:
                date_col = c
        df = df.iloc[1:]

    # Drop rows with NaN in the first column (empty/total rows)
    df = df.dropna(subset=[0])
    # Drop total rows (where account is not a number)
    # Handle accounts with hyphens like '65670-029'
    acct_numeric = df[0].apply(lambda v: pd.notna(v) and (
        str(v).replace('-', '').replace(' ', '').isdigit() if isinstance(v, str)
        else isinstance(v, (int, float)) and not pd.isna(v)))
    df = df[acct_numeric]

    if df.empty:
        return pd.DataFrame(columns=['code', 'amount', 'date'])

    # Identify columns by heuristic if header-based detection missed any
    ncols = df.shape[1]

    if code_col is None or amount_col is None or date_col is None:
        for c in range(ncols):
            sample = df[c].dropna()
            if sample.empty:
                continue
            first_val = sample.iloc[0]
            if code_col is None and _looks_like_loan_code(first_val):
                code_col = c
            elif date_col is None and isinstance(first_val, (pd.Timestamp, datetime)):
                date_col = c
            elif isinstance(first_val, (int, float)) and c > 1 and amount_col is None:
                amount_col = c

    if code_col is None or amount_col is None:
        # Fallback: try common layouts
        if ncols >= 5:
            if code_col is None:
                code_col = ncols - 1  # last column is often the code
            # Check if col 3 is numeric (amount) or datetime (date)
            s3 = df[3].dropna().iloc[0] if len(df[3].dropna()) > 0 else None
            if isinstance(s3, (pd.Timestamp, datetime)):
                if date_col is None:
                    date_col = 3
                if amount_col is None:
                    amount_col = 4
            else:
                if amount_col is None:
                    amount_col = 3
                if date_col is None:
                    date_col = 4
        elif ncols == 4:
            # Common 4-column layout: Account, Date, Amount, Code
            if code_col is None:
                code_col = 3
            if amount_col is None:
                amount_col = 2
            if date_col is None:
                date_col = 1

    result = pd.DataFrame({
        'code': df[code_col].values if code_col is not None else '',
        'amount': pd.to_numeric(df[amount_col], errors='coerce').values if amount_col is not None else 0,
    })
    if date_col is not None:
        result['date'] = pd.to_datetime(df[date_col].values, errors='coerce')
    else:
        result['date'] = pd.NaT

    return result.dropna(subset=['amount'])


def _parse_recovery_file(filepath, parse_config=None):
    """Parse a recovery file (varying formats). Returns DataFrame with [code, amount, date]."""
    df = _read_data_file(filepath)
    if df.empty:
        return pd.DataFrame(columns=['code', 'amount', 'date'])

    if parse_config:
        cfg_df = df.copy()
        if parse_config.get('has_header', False):
            cfg_df = cfg_df.iloc[1:]

        skip_rows = int(parse_config.get('skip_rows', 0) or 0)
        if skip_rows > 0:
            cfg_df = cfg_df.iloc[skip_rows:]

        account_col = parse_config.get('account_col', 0)
        code_col = parse_config.get('code_col')
        amount_col = parse_config.get('amount_col')
        date_col = parse_config.get('date_col')

        # Trust consistent header TEXT over fixed indices when the file has
        # a header row — protects against month-to-month column drift.
        # Skipped when ``strict_columns`` is set (see _parse_chargeoff_file).
        if (parse_config.get('has_header', False) and len(df) > 0
                and not parse_config.get('strict_columns')):
            _hdr = _resolve_hist_cols_by_header(df.iloc[0], 'rc')
            if 'amount' in _hdr and 'code' in _hdr:
                account_col = _hdr.get('account', account_col)
                code_col = _hdr['code']
                amount_col = _hdr['amount']
                if 'date' in _hdr:
                    date_col = _hdr['date']

        # See _parse_chargeoff_file: files with no loan-code column stamp
        # every row with ``code_static``.
        code_static = parse_config.get('code_static')

        ncols = cfg_df.shape[1]
        code_valid = code_col is not None and 0 <= int(code_col) < ncols
        amount_valid = amount_col is not None and 0 <= int(amount_col) < ncols
        date_valid = date_col is not None and 0 <= int(date_col) < ncols
        account_valid = account_col is not None and 0 <= int(account_col) < ncols

        if amount_valid and (code_valid or code_static not in (None, '')):
            if account_valid:
                acct_series = cfg_df.iloc[:, account_col]
                acct_numeric = acct_series.apply(_is_numeric_or_date)
                cfg_df = cfg_df[acct_numeric]

            if not cfg_df.empty:
                code_vals = (cfg_df.iloc[:, code_col].values if code_valid
                             else [code_static] * len(cfg_df))
                result = pd.DataFrame({
                    'code': code_vals,
                    'amount': pd.to_numeric(cfg_df.iloc[:, amount_col], errors='coerce').values,
                })
                if date_valid:
                    result['date'] = _coerce_mixed_dates(
                        cfg_df.iloc[:, date_col].values).values
                else:
                    result['date'] = pd.NaT
                result = result.dropna(subset=['amount'])
                if not result.empty:
                    return result
            # strict_columns: the explicit amount_col is authoritative. An
            # empty result means "no recoveries in this file", NOT a cue to
            # fall through to auto-detect (which in a combined Charge-Off /
            # Recovery workbook would wrongly grab the charge-off-amount column).
            if parse_config.get('strict_columns'):
                return pd.DataFrame(columns=['code', 'amount', 'date'])

    # Check if first row is a header (see _parse_chargeoff_file for rationale)
    first_row = df.iloc[0]
    first_a = first_row.iloc[0] if len(first_row) > 0 else None
    col0_is_numeric = isinstance(first_a, (int, float)) and pd.notna(first_a)
    first_vals = [str(v).lower() if pd.notna(v) else '' for v in first_row]
    has_header = (not col0_is_numeric) and any(
        'account' in v or 'recov' in v or 'loan' in v or 'acct' in v
        for v in first_vals
    )

    # Use header keywords to identify columns when header is present
    code_col = amount_col = date_col = None
    if has_header:
        for c, hdr in enumerate(first_vals):
            # Skip the FICO score column (contains 'sc'/'score' but is not a loan code)
            if 'fico' in hdr or 'score' in hdr:
                continue
            if any(k in hdr for k in ('security', 'code')) and 'account' not in hdr and 'sub' not in hdr:
                code_col = c
            elif any(k in hdr for k in ('amount', 'pymt', 'payment', 'principal')):
                amount_col = c
            elif 'date' in hdr or 'effective' in hdr:
                date_col = c
        df = df.iloc[1:]

    df = df.dropna(subset=[0])
    # Handle accounts with hyphens like '51930-27'
    acct_numeric = df[0].apply(lambda v: pd.notna(v) and (
        str(v).replace('-', '').replace(' ', '').isdigit() if isinstance(v, str)
        else isinstance(v, (int, float)) and not pd.isna(v)))
    df = df[acct_numeric]

    if df.empty:
        return pd.DataFrame(columns=['code', 'amount', 'date'])

    ncols = df.shape[1]

    # Heuristic fallback if header-based detection missed columns
    if code_col is None or amount_col is None or date_col is None:
        for c in range(ncols):
            sample = df[c].dropna()
            if sample.empty:
                continue
            first_val = sample.iloc[0]
            if code_col is None and _looks_like_loan_code(first_val):
                code_col = c
            elif date_col is None and isinstance(first_val, (pd.Timestamp, datetime)):
                date_col = c
            elif isinstance(first_val, (int, float)) and c > 1 and amount_col is None:
                amount_col = c

    if code_col is None or amount_col is None:
        if ncols >= 5:
            if code_col is None:
                code_col = ncols - 1
            s3 = df[3].dropna().iloc[0] if len(df[3].dropna()) > 0 else None
            if isinstance(s3, (pd.Timestamp, datetime)):
                if date_col is None:
                    date_col = 3
                if amount_col is None:
                    amount_col = 4
            else:
                if amount_col is None:
                    amount_col = 3
                if date_col is None:
                    date_col = 4
        elif ncols == 4:
            if code_col is None:
                code_col = 3
            if amount_col is None:
                amount_col = 2
            if date_col is None:
                date_col = 1

    result = pd.DataFrame({
        'code': df[code_col].values if code_col is not None else '',
        'amount': pd.to_numeric(df[amount_col], errors='coerce').values if amount_col is not None else 0,
    })
    if date_col is not None:
        result['date'] = pd.to_datetime(df[date_col].values, errors='coerce')
    else:
        result['date'] = pd.NaT

    return result.dropna(subset=['amount'])


def load_chargeoff_recovery_history(config):
    """Load all historical charge-off and recovery data.
    Returns dict: {'chargeoffs': {year: {pool: amount}}, 'recoveries': {year: {pool: amount}}}"""
    data_dir = resolve_path(config.get('data_directory', ''))
    if not data_dir or not os.path.isdir(data_dir):
        return {'chargeoffs': {}, 'recoveries': {}, 'years': []}

    historical_parse_cfg = config.get('historical_file_formats', {})
    chargeoff_parse_cfg = historical_parse_cfg.get('chargeoff')
    recovery_parse_cfg = historical_parse_cfg.get('recovery')
    pool_map = config.get('pool_map', {})
    quarters = _find_quarter_folders(data_dir)

    chargeoffs = {}  # {year: {pool: amount}}
    recoveries = {}
    co_monthly = {}   # {(year, month): {pool: amount}}
    rc_monthly = {}   # {(year, month): {pool: amount}}

    # --- Detect cumulative charge-off files (Ontario-style) ---
    # These have sheets "C-Offs 3 Years" and "Recoveries 3 Years" with all data
    cumulative_files = []
    for root, dirs, files in os.walk(data_dir):
        for f in files:
            fl = f.lower()
            if ('charge' in fl and 'off' in fl) and fl.endswith('.xlsx'):
                filepath = os.path.join(root, f)
                try:
                    xl = pd.ExcelFile(filepath)
                    if 'C-Offs 3 Years' in xl.sheet_names:
                        cumulative_files.append(filepath)
                except Exception:
                    pass

    if cumulative_files:
        # Use the largest cumulative file (most recent/complete data)
        cumulative_files.sort(key=os.path.getsize, reverse=True)
        filepath = cumulative_files[0]
        print(f"    Using cumulative charge-off file: {os.path.basename(filepath)}")

        def _extract_pool_code(raw_code, pool_map):
            """Extract pool from Ontario-style loan codes like '99 / Sig', 'Visa', 99, 'LP #1'."""
            code = str(raw_code).strip()
            # Try the raw value first (handles 'Visa', 'LP #1', etc.)
            pool = pool_map.get(code) or pool_map.get(code.upper()) or pool_map.get(code.lower())
            if pool:
                return pool
            # Try numeric extraction: "99 / Sig" -> "99", "11 / New Car" -> "11"
            if ' / ' in code:
                num_part = code.split(' / ')[0].strip()
                pool = pool_map.get(num_part) or pool_map.get(num_part.upper())
                if pool:
                    return pool
            # Try as-is for pure integers
            try:
                int_code = str(int(float(code)))
                pool = pool_map.get(int_code)
                if pool:
                    return pool
            except (ValueError, TypeError):
                pass
            return None

        # Parse charge-offs sheet
        try:
            df = pd.read_excel(filepath, sheet_name='C-Offs 3 Years', header=None)
            # Find header row (contains "Charge off Amount")
            hdr_idx = 0
            data_start = 0
            for i in range(min(10, len(df))):
                row_vals = [str(v).lower() if pd.notna(v) else '' for v in df.iloc[i]]
                if any('charge off amount' in v for v in row_vals):
                    hdr_idx = i
                    data_start = i + 1
                    break
            # Skip blank rows after header
            while data_start < len(df) and df.iloc[data_start].isna().all():
                data_start += 1

            # Find column positions from header
            hdr = df.iloc[hdr_idx]
            # Determine which columns have data - skip leading NaN columns
            first_data_col = 0
            for c in range(df.shape[1]):
                if pd.notna(hdr.iloc[c]):
                    first_data_col = c
                    break

            # Columns relative to first_data_col: Member#, Suffix, Code, Amount, Date, [FICO]
            code_col = first_data_col + 2
            amount_col = first_data_col + 3
            date_col = first_data_col + 4

            for i in range(data_start, len(df)):
                raw_code = df.iloc[i, code_col]
                if pd.isna(raw_code):
                    continue
                code_str = str(raw_code).strip()
                if code_str.upper() in ('TOTAL', ''):
                    continue
                pool = _extract_pool_code(raw_code, pool_map)
                amount = pd.to_numeric(df.iloc[i, amount_col], errors='coerce')
                date_val = pd.to_datetime(df.iloc[i, date_col], errors='coerce')
                if pool and pd.notna(amount) and pd.notna(date_val) and 2000 <= date_val.year <= 2099:
                    yr = date_val.year
                    chargeoffs.setdefault(yr, {})
                    chargeoffs[yr][pool] = chargeoffs[yr].get(pool, 0) + amount
                    ym = (yr, date_val.month)
                    co_monthly.setdefault(ym, {})
                    co_monthly[ym][pool] = co_monthly[ym].get(pool, 0) + amount
        except Exception as e:
            print(f"    Warning: Could not parse charge-offs from {filepath}: {e}")

        # Parse recoveries sheet
        try:
            xl = pd.ExcelFile(filepath)
            if 'Recoveries 3 Years' in xl.sheet_names:
                df = pd.read_excel(filepath, sheet_name='Recoveries 3 Years', header=None)
                hdr_idx = 0
                data_start = 0
                for i in range(min(10, len(df))):
                    row_vals = [str(v).lower() if pd.notna(v) else '' for v in df.iloc[i]]
                    if any('recovery amount' in v for v in row_vals):
                        hdr_idx = i
                        data_start = i + 1
                        break
                while data_start < len(df) and df.iloc[data_start].isna().all():
                    data_start += 1

                hdr = df.iloc[hdr_idx]
                first_data_col = 0
                for c in range(df.shape[1]):
                    if pd.notna(hdr.iloc[c]):
                        first_data_col = c
                        break

                code_col = first_data_col + 2
                amount_col = first_data_col + 3
                date_col = first_data_col + 4

                for i in range(data_start, len(df)):
                    raw_code = df.iloc[i, code_col]
                    if pd.isna(raw_code):
                        continue
                    code_str = str(raw_code).strip()
                    if code_str.lower() in ('total', ''):
                        continue
                    pool = _extract_pool_code(raw_code, pool_map)
                    amount = pd.to_numeric(df.iloc[i, amount_col], errors='coerce')
                    date_val = pd.to_datetime(df.iloc[i, date_col], errors='coerce')
                    if pool and pd.notna(amount) and pd.notna(date_val) and 2000 <= date_val.year <= 2099:
                        yr = date_val.year
                        recoveries.setdefault(yr, {})
                        recoveries[yr][pool] = recoveries[yr].get(pool, 0) + amount
                        ym = (yr, date_val.month)
                        rc_monthly.setdefault(ym, {})
                        rc_monthly[ym][pool] = rc_monthly[ym].get(pool, 0) + amount
        except Exception as e:
            print(f"    Warning: Could not parse recoveries from {filepath}: {e}")

    else:
        # --- Franklin-style: per-quarter charge-off/recovery files ---
        # Build a string-keyed pool map that handles numeric codes
        str_pool_map = {str(k).strip(): v for k, v in pool_map.items()}

        # ------------------------------------------------------------------
        # Combined-file mode detection.
        #
        # Some CUs (e.g. Shuford FCU) ship a SINGLE workbook per quarter
        # containing both charge-off and recovery rows side-by-side: same
        # account / code / date columns, just different amount columns
        # (e.g. column 8 = chargeoff amount, column 9 = recovery amount).
        # When ``historical_file_formats`` describes both halves AND the
        # account/code/date column wiring is identical, every matched
        # file should be parsed via BOTH parsers — otherwise the half
        # whose token ('charge'/'off' vs 'recov') doesn't appear in the
        # filename gets silently dropped (the original logic explicitly
        # excludes 'recov' files from the CO branch and vice-versa).
        # ------------------------------------------------------------------
        def _shared_locator(co_cfg, rc_cfg):
            if not (co_cfg and rc_cfg):
                return False
            keys = ('account_col', 'code_col', 'date_col',
                    'has_header', 'skip_rows')
            for k in keys:
                if co_cfg.get(k) != rc_cfg.get(k):
                    return False
            # Distinct amount columns are the whole point — refuse to
            # collapse if the YAML accidentally points both halves at
            # the same column.
            return co_cfg.get('amount_col') != rc_cfg.get('amount_col')

        combined_mode = _shared_locator(chargeoff_parse_cfg, recovery_parse_cfg)
        if combined_mode:
            print(
                "    Combined CO+Recovery file mode: every matched file "
                "will be parsed for both halves (account/code/date "
                "columns shared; amount columns differ)."
            )
        # ------------------------------------------------------------------
        # Multi-format mode.
        #
        # When the config supplies ``historical_file_formats.formats`` (a
        # list of named formats, each with its own ``file_pattern`` plus
        # ``chargeoff`` / ``recovery`` column wiring), every file is routed
        # to the FIRST format whose pattern matches its name and parsed with
        # that format's columns. This lets ONE credit union mix several
        # CO/recovery layouts — e.g. a consumer-loan file, a credit-card file
        # (no code column → ``code_static``) and an overdraft file — which a
        # single top-level chargeoff/recovery block cannot express. Files
        # matching no format are skipped. Absent a formats list, behaviour is
        # unchanged (legacy single chargeoff/recovery config).
        # ------------------------------------------------------------------
        _co_recov_formats = historical_parse_cfg.get('formats') or []
        multi_format = bool(_co_recov_formats)
        _compiled_formats = []
        for _fmt in _co_recov_formats:
            _pat = _fmt.get('file_pattern') or ''
            try:
                _rx = re.compile(_pat, re.IGNORECASE) if _pat else None
            except re.error:
                _rx = None
            _compiled_formats.append((_rx, _fmt))
        if multi_format:
            print(f"    Multi-format CO/Recovery mode: "
                  f"{len(_compiled_formats)} format(s) configured.")

        def _resolve_file_format(fname):
            """Return (co_cfg, rc_cfg, combined) for *fname*.

            Multi-format: first format whose file_pattern matches, else
            (None, None, False) so the file is skipped. Legacy: the single
            top-level chargeoff/recovery configs.
            """
            if multi_format:
                for _rx, _fmt in _compiled_formats:
                    if _rx is not None and _rx.search(fname):
                        _co = _fmt.get('chargeoff')
                        _rc = _fmt.get('recovery')
                        return _co, _rc, _shared_locator(_co, _rc)
                return None, None, False
            return chargeoff_parse_cfg, recovery_parse_cfg, combined_mode

        # Track files already fed into each side so a filename matching
        # both tokens isn't double-counted. Tuple of (kind, abs path).
        _processed_co: set[str] = set()
        _processed_rc: set[str] = set()

        def _lookup_pool(raw_code):
            """Look up pool from a code value (numeric or text)."""
            code = str(raw_code).strip()
            # Try as-is
            pool = str_pool_map.get(code) or str_pool_map.get(code.upper()) or str_pool_map.get(code.lower())
            if pool:
                return pool
            # Try integer conversion (handles 28.0 -> "28")
            try:
                int_code = str(int(float(code)))
                pool = str_pool_map.get(int_code)
                if pool:
                    return pool
            except (ValueError, TypeError):
                pass
            # Try matching against pool names (for text codes like "visa" -> pool "VISA")
            code_upper = code.upper()
            for v in set(str_pool_map.values()):
                if v.upper() == code_upper:
                    return v
            return None

        # Fallback for flat folders (no YYYY-MM quarter subfolders):
        # treat ``data_dir`` itself as a single bucket so wizard-style
        # CECL setups (Raw_Uploads/<short>/*.xlsx) get picked up. The
        # per-row date column (configured in
        # ``historical_file_formats``) supplies the actual year/month;
        # ``qlabel`` is only used as a fallback when row-level date
        # parsing fails.
        if not quarters:
            from datetime import datetime as _dt
            quarters = [(data_dir, f"{_dt.today().year}-12")]
            print(f"    No YYYY-MM quarter subfolders under {data_dir}; "
                  f"scanning top-level for charge-off / recovery files.")

        # Filename-date fallback: many CUs ship per-month CO/Recovery
        # files like "CECL Charge Off 12312025.xlsx" or "Charge Off
        # 12-31-22.xlsx" where individual recovery ROWS may have NULL
        # date cells (e.g. Shuford's combined files use ChargeOffDateS
        # for both halves; recovery rows leave it blank). When that
        # happens, fall back to the date encoded in the filename
        # before defaulting to today's year. Reuses the same fallback
        # layouts used by import_data.extract_snapshot_date.
        try:
            from import_data import _try_common_date_layouts as _file_iso
        except Exception:  # noqa: BLE001
            _file_iso = None  # type: ignore
        try:
            from import_data import extract_snapshot_date as _cfg_file_iso
        except Exception:  # noqa: BLE001
            _cfg_file_iso = None  # type: ignore

        def _file_period(fname: str):
            """Return (year, month) parsed from filename, or None.

            Try the config-agnostic layout matcher first; when it can't
            resolve a date (e.g. month-name filenames like "... APRIL
            2026.xlsx"), fall back to ``extract_snapshot_date`` which honors
            the CU's configured ``date_pattern`` (month names, 2-digit
            years, etc.). Without this fallback, month-name CO/recovery
            files whose in-file date column is blank all fail filename
            parsing and default to month 12 — piling every month into a
            spurious December bucket.
            """
            iso = _file_iso(fname) if _file_iso else None
            if (not iso or len(iso) < 7) and _cfg_file_iso is not None:
                try:
                    iso = _cfg_file_iso(fname, config)
                except Exception:  # noqa: BLE001
                    iso = None
            # Last resort: a spelled-out / abbreviated month name adjacent to a
            # 2- or 4-digit year in the filename (e.g. "MAY26", "Aug25",
            # "Jan2025", "June 25"). Common for CUs whose combined CO/recovery
            # files are named by period in a non-ISO layout AND whose recovery
            # ROWS carry no date cell. Without this every such dateless row
            # defaults to today's year -> cross-year leakage (all recoveries
            # piled into a phantom December bucket; a prior year's recoveries
            # vanish to $0).
            if not iso or len(iso) < 7:
                _mn = re.search(
                    r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*'
                    r'[\s_\-\.]*((?:19|20)?\d{2})(?!\d)',
                    fname, re.IGNORECASE)
                if _mn:
                    _mo = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5,
                           'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10,
                           'nov': 11, 'dec': 12}[_mn.group(1)[:3].lower()]
                    _yr = int(_mn.group(2))
                    if _yr < 100:
                        _yr += 2000
                    if 2000 <= _yr <= 2099:
                        return _yr, _mo
            # Also handle COMPACT NUMERIC month+year tokens with no month name:
            # "MMYY" / "MMYYYY" / "MM-YY" (e.g. "...0226" -> Feb 2026,
            # "...122022" -> Dec 2022, "...07-23" -> Jul 2023). Common for
            # per-month charge-off/recovery files whose recovery ROWS carry no
            # date cell. A valid 01-12 month plus word boundaries are required
            # so a longer digit run (account #s, YYYYMMDD) doesn't false-match;
            # runs only after the ISO / config / month-name parsers all fail.
            if not iso or len(iso) < 7:
                _num = re.search(
                    r'\b(0[1-9]|1[0-2])[-_\. ]?((?:19|20)?\d{2})\b', fname)
                if _num:
                    _mo = int(_num.group(1))
                    _yr = int(_num.group(2))
                    if _yr < 100:
                        _yr += 2000
                    if 2000 <= _yr <= 2099:
                        return _yr, _mo
            if not iso or len(iso) < 7:
                return None
            try:
                return int(iso[0:4]), int(iso[5:7])
            except (TypeError, ValueError):
                return None

        for folder, qlabel in quarters:
            year = int(qlabel[:4])

            for f in os.listdir(folder):
                fl = f.lower()
                filepath = os.path.join(folder, f)
                # Resolve the per-file CO/recovery column config. In
                # multi-format mode this routes by file_pattern; otherwise it
                # returns the single legacy config. A file matching no format
                # yields (None, None) and is skipped.
                _co_cfg, _rc_cfg, _file_combined = _resolve_file_format(f)
                # Determine which side(s) this file should feed.
                # In combined mode any file matching EITHER token is
                # parsed for BOTH halves; in legacy mode the filename
                # tokens decide and the original 'proposed'/'3yr'/'recov'
                # exclusions still apply.
                want_co = False
                want_rc = False
                if (_co_cfg is not None or _rc_cfg is not None) and (
                        fl.endswith('.xlsx') or fl.endswith('.xls')
                        or fl.endswith('.csv')):
                    if _file_combined:
                        _skip = ('proposed' in fl or '3yr' in fl)
                        # Legacy combined mode gates purely on the filename
                        # (no per-format file_pattern), so it must still
                        # require a charge-off / recovery indicator in the
                        # name. Otherwise it vacuums every unrelated workbook
                        # that shares a flat Raw_Uploads folder — AIRES loan
                        # extracts, balance sheets, credit pulls, Vizo models
                        # — and sums their balance / member-number columns
                        # into billions of dollars of phantom recoveries.
                        # In multi-format mode the format's file_pattern has
                        # already selected the file, so don't second-guess it.
                        if not multi_format and not (
                                ('charge' in fl and 'off' in fl)
                                or 'recov' in fl):
                            _skip = True
                        if not _skip:
                            want_co = bool(_co_cfg)
                            want_rc = bool(_rc_cfg)
                    else:
                        if (_co_cfg and ('charge' in fl and 'off' in fl)
                                and 'proposed' not in fl
                                and '3yr' not in fl
                                and 'recov' not in fl):
                            want_co = True
                        if _rc_cfg and 'recov' in fl and '3yr' not in fl:
                            want_rc = True
                        # A format that defines only one side (and whose
                        # file_pattern already selected this file) applies
                        # that side even when the filename lacks the token.
                        if multi_format and not (want_co or want_rc):
                            if _co_cfg and not _rc_cfg:
                                want_co = True
                            elif _rc_cfg and not _co_cfg:
                                want_rc = True

                if want_co and filepath not in _processed_co:
                    _processed_co.add(filepath)
                    # Per-file filename date (for rows with NULL date cells).
                    _fp = _file_period(f)
                    file_default_year = _fp[0] if _fp else year
                    file_default_month = _fp[1] if _fp else (
                        int(qlabel[5:7]) if len(qlabel) >= 7 else 12
                    )
                    try:
                        df = _parse_chargeoff_file(filepath, parse_config=_co_cfg)
                        for _, row in df.iterrows():
                            pool = _lookup_pool(row['code'])
                            if pool and pd.notna(row['amount']):
                                row_year = file_default_year
                                row_month = file_default_month
                                if pd.notna(row.get('date')):
                                    try:
                                        dt = pd.to_datetime(row['date'])
                                        y = int(dt.year)
                                        if 2000 <= y <= 2099:
                                            row_year = y
                                            row_month = int(dt.month)
                                    except Exception:
                                        pass
                                chargeoffs.setdefault(row_year, {})
                                chargeoffs[row_year][pool] = chargeoffs[row_year].get(pool, 0) + row['amount']
                                ym = (row_year, row_month)
                                co_monthly.setdefault(ym, {})
                                co_monthly[ym][pool] = co_monthly[ym].get(pool, 0) + row['amount']
                    except Exception as e:
                        print(f"    Warning: Could not parse charge-offs from {filepath}: {e}")

                if want_rc and filepath not in _processed_rc:
                    _processed_rc.add(filepath)
                    # Per-file filename date (for rows with NULL date cells).
                    _fp = _file_period(f)
                    file_default_year = _fp[0] if _fp else year
                    file_default_month = _fp[1] if _fp else (
                        int(qlabel[5:7]) if len(qlabel) >= 7 else 12
                    )
                    try:
                        df = _parse_recovery_file(filepath, parse_config=_rc_cfg)
                        for _, row in df.iterrows():
                            pool = _lookup_pool(row['code'])
                            if pool and pd.notna(row['amount']):
                                row_year = file_default_year
                                row_month = file_default_month
                                if pd.notna(row.get('date')):
                                    try:
                                        dt = pd.to_datetime(row['date'])
                                        y = int(dt.year)
                                        if 2000 <= y <= 2099:
                                            row_year = y
                                            row_month = int(dt.month)
                                    except Exception:
                                        pass
                                recoveries.setdefault(row_year, {})
                                recoveries[row_year][pool] = recoveries[row_year].get(pool, 0) + row['amount']
                                ym = (row_year, row_month)
                                rc_monthly.setdefault(ym, {})
                                rc_monthly[ym][pool] = rc_monthly[ym].get(pool, 0) + row['amount']
                    except Exception as e:
                        print(f"    Warning: Could not parse recoveries from {filepath}: {e}")

        # Also check for 3yr file (covers 2019-2022 Q3)
        for root, dirs, files in os.walk(data_dir):
            for f in files:
                if '3yr' in f.lower() and f.endswith('.xlsx'):
                    filepath = os.path.join(root, f)
                    try:
                        df = _parse_chargeoff_file(filepath)
                        for _, row in df.iterrows():
                            code = str(row['code']).strip().upper()
                            pool = pool_map.get(code, pool_map.get(code.lower(), None))
                            if pool and pd.notna(row['amount']) and pd.notna(row['date']):
                                yr = row['date'].year
                                chargeoffs.setdefault(yr, {})
                                chargeoffs[yr][pool] = chargeoffs[yr].get(pool, 0) + row['amount']
                                ym = (yr, int(row['date'].month))
                                co_monthly.setdefault(ym, {})
                                co_monthly[ym][pool] = co_monthly[ym].get(pool, 0) + row['amount']
                    except Exception as e:
                        print(f"    Warning: Could not parse 3yr file {filepath}: {e}")

        # Also check the standalone recovery file in 2022-10
        for root, dirs, files in os.walk(data_dir):
            for f in files:
                if f.lower() == 'recovery.xlsx':
                    filepath = os.path.join(root, f)
                    try:
                        df = _parse_recovery_file(filepath)
                        for _, row in df.iterrows():
                            code = str(row['code']).strip().upper()
                            pool = pool_map.get(code, pool_map.get(code.lower(), None))
                            if pool and pd.notna(row['amount']) and pd.notna(row['date']):
                                yr = row['date'].year
                                recoveries.setdefault(yr, {})
                                recoveries[yr][pool] = recoveries[yr].get(pool, 0) + row['amount']
                                ym = (yr, int(row['date'].month))
                                rc_monthly.setdefault(ym, {})
                                rc_monthly[ym][pool] = rc_monthly[ym].get(pool, 0) + row['amount']
                    except Exception as e:
                        print(f"    Warning: Could not parse recovery file {filepath}: {e}")

    all_years = sorted(set(list(chargeoffs.keys()) + list(recoveries.keys())))
    return {
        'chargeoffs': chargeoffs, 'recoveries': recoveries, 'years': all_years,
        'co_monthly': co_monthly, 'rc_monthly': rc_monthly,
    }


def _col_letter_to_idx(letter):
    """Convert an Excel column letter (A, B, ..., Z, AA, AB, ...) to a 0-based
    index. Accepts already-numeric values as well."""
    if letter is None:
        return 0
    if isinstance(letter, (int, float)):
        return int(letter)
    s = str(letter).strip().upper()
    if not s:
        return 0
    if s.isdigit():
        return int(s)
    n = 0
    for ch in s:
        if not ('A' <= ch <= 'Z'):
            return 0
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n - 1


def _load_monthly_balances_manual(mb_cfg):
    """Build (df, alll_by_date) from a wizard-entered pool × month grid.

    The optional ``alll`` block (``{YYYY-MM-DD: balance}``) carries the ACL /
    ALLL balance per month — e.g. the "ALLL Balance" row of a Monthly
    Balances by Pool file. Values are stored as absolute amounts keyed by a
    month-end timestamp, matching the per_month / per_year loaders so the
    downstream ACL-balance consumer picks up the snapshot month.
    """
    entries = mb_cfg.get('entries') or {}
    records = []
    for pool, row in entries.items():
        if not pool or not isinstance(row, dict):
            continue
        for d, v in row.items():
            try:
                dt = pd.to_datetime(d, errors='coerce')
            except Exception:
                continue
            if pd.isna(dt):
                continue
            try:
                bal = float(v)
            except (TypeError, ValueError):
                continue
            records.append({'pool': str(pool).strip(),
                            'date': dt,
                            'balance': bal})
    alll_by_date: dict = {}
    for d, v in (mb_cfg.get('alll') or {}).items():
        dt = pd.to_datetime(d, errors='coerce')
        if pd.isna(dt):
            continue
        try:
            alll_by_date[dt] = abs(float(v))
        except (TypeError, ValueError):
            continue
    return pd.DataFrame(records, columns=['pool', 'date', 'balance']), alll_by_date


def _pm_effective_balance_idx(df, start_row, label_idx, configured_idx):
    """Return the 0-based column index to read balances from for THIS file.

    Normally the configured ``balance_col``. But when a workbook's title or
    header wraps into an extra leading column, the balances shift one column
    to the right, leaving the configured column empty for that file. In that
    case fall back to the densest numeric column at/after the label column so
    the period still loads. Only overrides when the configured column has NO
    numeric values for this file, so files that already parse are never
    affected.
    """
    try:
        ncols = int(df.shape[1])
    except Exception:  # noqa: BLE001
        return configured_idx
    if configured_idx is None:
        configured_idx = 0

    def _numeric_count(c):
        if c is None or c < 0 or c >= ncols:
            return 0
        n = 0
        for i in range(max(0, start_row), df.shape[0]):
            lbl = (df.iat[i, label_idx]
                   if (label_idx is not None and 0 <= label_idx < ncols)
                   else None)
            if lbl is None or str(lbl).strip() == '':
                continue
            if _coerce_balance(df.iat[i, c]) is not None:
                n += 1
        return n

    if 0 <= configured_idx < ncols and _numeric_count(configured_idx) > 0:
        return configured_idx
    # Configured column is empty for this file — auto-detect the densest
    # numeric column to the right of the label column.
    best_c, best_n = configured_idx, 0
    lo = (label_idx + 1) if label_idx is not None else 0
    for c in range(max(0, lo), ncols):
        n = _numeric_count(c)
        if n > best_n:
            best_n, best_c = n, c
    if best_n > 0 and best_c != configured_idx:
        print(f"    Monthly-balance: configured balance column was empty for "
              f"this file; using detected column {best_c} instead.")
        return best_c
    return configured_idx


def _load_monthly_balances_per_month(mb_cfg, acl_cfg=None):
    """Read one balance-sheet style file per month and emit (pool, date, bal)
    rows. Each file is opened on ``layout.sheet`` (or the first sheet) and
    the label/balance columns are pulled from the configured letters,
    skipping ``header_row`` rows. Labels are mapped to wizard pool names via
    ``pool_map`` (case-insensitive, falls back to the raw label).

    When ``acl_cfg`` is provided and has ``source == 'monthly_file'`` with an
    ``acl.row`` (1-based) set, the same balance column is also read at that
    row from each per-month file and returned as the ``alll_by_date`` dict
    so the report engine's ACL Balance lookup picks it up. An optional
    ``acl.col`` letter override is honored when present.

    Supported file types: .xlsx / .xls / .csv (via pandas) and .pdf (via
    pdfplumber's table extraction). For PDFs the ``sheet`` field is
    interpreted as a 1-based page number; blank means scan every page and
    concatenate rows.
    """
    layout = mb_cfg.get('layout') or {}
    sheet = (layout.get('sheet') or '').strip()
    header_row = int(layout.get('header_row') or 1)
    label_idx = _col_letter_to_idx(layout.get('label_col') or 'A')
    balance_idx = _col_letter_to_idx(layout.get('balance_col') or 'B')
    raw_map = mb_cfg.get('pool_map') or {}
    pool_map = {str(k).strip().lower(): str(v).strip()
                for k, v in raw_map.items() if str(k).strip() and str(v).strip()}

    # Resolve ACL extraction settings. Only active when the wizard's ACL
    # source is "monthly_file" and a row number was captured.
    acl_cfg = acl_cfg or {}
    acl_src = (acl_cfg.get('source') or '').strip().lower()
    acl_row_1b = acl_cfg.get('row')
    try:
        acl_row_1b = int(acl_row_1b) if acl_row_1b not in (None, '') else 0
    except (TypeError, ValueError):
        acl_row_1b = 0
    acl_col_override = acl_cfg.get('col') or ''
    acl_col_idx = balance_idx
    if isinstance(acl_col_override, str) and acl_col_override.strip():
        _ovr = _col_letter_to_idx(acl_col_override)
        if _ovr is not None:
            acl_col_idx = _ovr
    extract_acl = bool(acl_src == 'monthly_file' and acl_row_1b > 0)
    alll_by_date: dict = {}

    records = []
    for entry in (mb_cfg.get('files') or []):
        period = (entry.get('period') or '').strip()
        path = (entry.get('saved_path') or '').strip()
        if not path:
            # Fall back to looking up filename inside data_directory if it
            # was copied there post-save.
            continue
        try:
            dt = pd.to_datetime(period, errors='coerce')
        except Exception:
            continue
        if pd.isna(dt) or not os.path.isfile(path):
            continue
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == '.pdf':
                df = _read_pdf_balance_table(path, sheet)
            elif ext == '.csv':
                df = _read_csv_any(path, header=None, dtype=str)
            else:
                if sheet:
                    try:
                        df = pd.read_excel(path, sheet_name=sheet, header=None)
                    except Exception:
                        # Some CUs (e.g. NOVA) name each monthly workbook's
                        # sheet after the document number, so the configured
                        # ``sheet`` only matches one file. Fall back to the
                        # first worksheet so every period still loads.
                        df = pd.read_excel(path, sheet_name=0, header=None)
                else:
                    df = pd.read_excel(path, header=None)
        except Exception as e:
            print(f"    Warning: Could not read monthly-balance file {path}: {e}")
            continue
        if df is None or df.empty or df.shape[1] <= max(label_idx, balance_idx):
            continue
        # Determine the effective balance column for THIS file (bulletproof
        # against a wrapped title/header pushing balances one column right,
        # which leaves the configured column empty for that file only).
        eff_balance_idx = _pm_effective_balance_idx(
            df, header_row, label_idx, balance_idx)
        # Pull the ACL value for this period from the configured row+col
        # (defaults to balance_col). The wizard's "monthly_file" source
        # captures one row number that applies across every period file.
        if extract_acl:
            _ridx0 = acl_row_1b - 1
            if 0 <= _ridx0 < df.shape[0] and acl_col_idx < df.shape[1]:
                _av = _coerce_balance(df.iat[_ridx0, acl_col_idx])
                if _av is not None:
                    alll_by_date[dt] = abs(_av)
        for i in range(header_row, df.shape[0]):
            label = df.iat[i, label_idx]
            bal = df.iat[i, eff_balance_idx]
            if pd.isna(label) or str(label).strip() == '':
                continue
            if pd.isna(bal):
                continue
            bal_f = _coerce_balance(bal)
            if bal_f is None:
                continue
            key = str(label).strip().lower()
            pool = pool_map.get(key, str(label).strip())
            if not pool:
                continue
            records.append({'pool': pool, 'date': dt, 'balance': bal_f})
    if extract_acl and alll_by_date:
        print(f"    ACL Balance extracted from per_month files (row {acl_row_1b}): {len(alll_by_date)} period(s)")
    return pd.DataFrame(records, columns=['pool', 'date', 'balance']), alll_by_date


def _load_monthly_balances_per_year(mb_cfg, acl_cfg=None):
    """Read one annual balance-sheet workbook per calendar year and emit
    (pool, date, bal) rows. Each file is opened on ``layout.sheet`` (or
    the first/best sheet) and the per-file header row is re-scanned to
    pull month-end columns. Labels are mapped to wizard pool names via
    ``pool_map`` (case-insensitive, falls back to the raw label).

    When ``acl_cfg`` is provided with ``source == 'monthly_file'`` and a
    1-based ``row`` (and optional ``col`` letter override), the same row
    is read from each period column of each yearly workbook and returned
    as the ``alll_by_date`` dict keyed by month-end timestamp.

    Delegates the heavy lifting to
    ``cecl_ui.services.monthly_bal_parser.pool_balances_for_per_year_files``
    so the wizard's auto-detect logic and the runtime importer always
    agree on column/row interpretation.
    """
    try:
        from cecl_ui.services.monthly_bal_parser import (
            pool_balances_for_per_year_files,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"    Warning: per_year loader unavailable: {exc}")
        return pd.DataFrame(columns=['pool', 'date', 'balance']), {}

    layout = mb_cfg.get('layout') or {}
    raw_map = mb_cfg.get('pool_map') or {}
    pool_map = {str(k).strip().lower(): str(v).strip()
                for k, v in raw_map.items()
                if str(k).strip() and str(v).strip()}
    year_files = mb_cfg.get('files') or []

    result = pool_balances_for_per_year_files(year_files, layout, pool_map)
    by_period = result.get('by_period') or {}
    records = []
    for period_iso, payload in by_period.items():
        try:
            dt = pd.to_datetime(period_iso, errors='coerce')
        except Exception:
            continue
        if pd.isna(dt):
            continue
        for pool, bal in (payload.get('by_pool') or {}).items():
            try:
                bal_f = float(bal)
            except (TypeError, ValueError):
                continue
            records.append({'pool': pool, 'date': dt, 'balance': bal_f})
    if result.get('error'):
        print(f"    Warning: per_year loader: {result['error']}")

    # ------------------------------------------------------------------
    # ACL Balance extraction (mirrors per_month flow).
    # ------------------------------------------------------------------
    alll_by_date: dict = {}
    acl_cfg = acl_cfg or {}
    acl_src = (acl_cfg.get('source') or '').strip().lower()
    acl_row_1b = acl_cfg.get('row')
    try:
        acl_row_1b = int(acl_row_1b) if acl_row_1b is not None else 0
    except (TypeError, ValueError):
        acl_row_1b = 0
    acl_col_override = acl_cfg.get('col') or ''
    extract_acl = bool(acl_src == 'monthly_file' and acl_row_1b > 0)
    if extract_acl and year_files:
        from openpyxl import load_workbook as _lwb
        from openpyxl.utils import column_index_from_string as _col2idx
        # Re-detect per-file period columns (mirrors what
        # pool_balances_for_per_year_files does internally), then read
        # the acl row at each period column.
        try:
            from cecl_ui.services.monthly_bal_parser import (
                analyse_per_year_file as _analyse_py,
            )
        except Exception:
            _analyse_py = None
        # Allow an explicit column letter to override the per-file
        # detection (useful when user wants "always col C" semantics).
        forced_col_idx = None
        if acl_col_override:
            try:
                forced_col_idx = _col2idx(str(acl_col_override).upper())
            except Exception:
                forced_col_idx = None
        for yf in year_files:
            path = yf.get('saved_path') or yf.get('path') or ''
            if not path:
                continue
            try:
                wb = _lwb(path, read_only=True, data_only=True)
            except Exception as e:
                print(f"    Warning: ACL extract failed to open {path}: {e}")
                continue
            try:
                sheet_name = (layout.get('sheet') or '').strip()
                if sheet_name and sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                elif _analyse_py:
                    info = _analyse_py(path) or {}
                    sn = info.get('sheet')
                    ws = wb[sn] if sn and sn in wb.sheetnames else wb.active
                else:
                    ws = wb.active
                # Determine period columns: re-detect per file when
                # available; fall back to layout's period_columns.
                period_cols = []
                if _analyse_py:
                    info = _analyse_py(path) or {}
                    period_cols = info.get('period_columns') or []
                if not period_cols:
                    period_cols = layout.get('period_columns') or []
                for pc in period_cols:
                    try:
                        if forced_col_idx:
                            col_idx = forced_col_idx
                        else:
                            raw_col = pc.get('col')
                            if isinstance(raw_col, str):
                                col_idx = _col2idx(raw_col.strip().upper())
                            else:
                                col_idx = int(raw_col)
                        period_iso = pc.get('period_iso') or pc.get('period')
                        if not period_iso:
                            continue
                        cell_val = ws.cell(row=acl_row_1b, column=col_idx).value
                        av = _coerce_balance(cell_val)
                        if av is None:
                            continue
                        dt = pd.to_datetime(period_iso, errors='coerce')
                        if pd.isna(dt):
                            continue
                        alll_by_date[dt] = abs(float(av))
                    except Exception:
                        continue
            finally:
                try:
                    wb.close()
                except Exception:
                    pass
        if alll_by_date:
            print(f"    ACL Balance extracted from per_year files (row {acl_row_1b}): {len(alll_by_date)} period(s)")
    return pd.DataFrame(records, columns=['pool', 'date', 'balance']), alll_by_date


def _coerce_balance(v):
    """Convert a cell value to float, stripping $, commas, parens for
    negatives, and whitespace. Returns None on failure."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            f = float(v)
        except (TypeError, ValueError):
            return None
        if pd.isna(f):
            return None
        return f
    s = str(v).strip()
    if not s:
        return None
    neg = False
    if s.startswith('(') and s.endswith(')'):
        neg = True
        s = s[1:-1]
    s = s.replace('$', '').replace(',', '').replace(' ', '')
    if s in ('', '-', '–', '—'):
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return -f if neg else f


def _read_pdf_balance_table(path, sheet):
    """Open a PDF via pdfplumber, extract tables on the requested page (or
    all pages when ``sheet`` is blank), and return a single pandas
    DataFrame whose rows are the concatenated table rows. Returns an empty
    DataFrame if pdfplumber is unavailable or no tables are detected."""
    try:
        import pdfplumber  # type: ignore
    except ImportError:
        print(f"    Warning: pdfplumber not installed; cannot read {path}")
        return pd.DataFrame()

    page_filter = None
    sht = (sheet or '').strip()
    if sht:
        try:
            page_filter = int(sht) - 1  # convert 1-based to 0-based
        except ValueError:
            page_filter = None

    all_rows = []
    max_width = 0
    try:
        with pdfplumber.open(path) as pdf:
            pages = pdf.pages
            if page_filter is not None and 0 <= page_filter < len(pages):
                pages = [pages[page_filter]]
            for page in pages:
                try:
                    tables = page.extract_tables() or []
                except Exception:
                    tables = []
                for tbl in tables:
                    for row in tbl:
                        if not row:
                            continue
                        cleaned = [('' if c is None else str(c).strip())
                                   for c in row]
                        if not any(cleaned):
                            continue
                        all_rows.append(cleaned)
                        if len(cleaned) > max_width:
                            max_width = len(cleaned)
    except Exception as e:
        print(f"    Warning: pdfplumber failed on {path}: {e}")
        return pd.DataFrame()

    if not all_rows:
        return pd.DataFrame()
    # Normalize ragged rows so iat[i, col] is always valid.
    normalized = [r + [''] * (max_width - len(r)) for r in all_rows]
    return pd.DataFrame(normalized)


def _merge_acl_history(alll_by_date: dict, config: dict) -> dict:
    """Merge wizard-entered/manually-extracted ACL history from
    ``cfg["acl"]["history"]`` into the ALLL-by-date map loaded from the
    monthly balances file. YAML history values take precedence on key
    collision (they are the user's audited / latest source of truth).
    """
    acl_cfg = (config or {}).get('acl') or {}
    hist_map = acl_cfg.get('history') or {}
    if not hist_map:
        return alll_by_date
    out = dict(alll_by_date or {})
    n = 0
    for k, v in hist_map.items():
        try:
            d = pd.Timestamp(k)
            out[d] = abs(float(v))
            n += 1
        except (ValueError, TypeError):
            continue
    if n:
        print(f"    Merged {n} ACL history entries from YAML cfg['acl']['history']")
    return out


def _load_monthly_balances_from_wizard(config, mb_cfg=None, with_labels=False):
    """Load monthly balances using wizard-provided cfg['monthly_balance']
    metadata (``saved_path`` + ``sheet`` + ``pool_name_col`` +
    ``first_date_col`` + ``header_row``) plus optional
    ``cfg['balance_title_map']`` for label→pool translation and
    ``cfg['acl']['row']`` / ``cfg['acl']['label']`` for locating the
    ACL row.

    Returns ``(df, alll_by_date)`` or ``(None, None)`` if the wizard
    metadata is missing / file is unreadable.
    """
    mb_cfg = mb_cfg if mb_cfg is not None else ((config or {}).get('monthly_balance') or {})
    saved_path = mb_cfg.get('saved_path')
    if not saved_path or not os.path.isfile(saved_path):
        return None, None

    sheet = mb_cfg.get('sheet')
    header_row = mb_cfg.get('header_row')  # 1-based
    pool_col_letter = mb_cfg.get('pool_name_col')
    date_col_letter = mb_cfg.get('first_date_col')

    def _col_letter_to_idx(letter):
        if not letter or not isinstance(letter, str):
            return None
        letter = letter.strip().upper()
        if not letter.isalpha():
            return None
        n = 0
        for ch in letter:
            n = n * 26 + (ord(ch) - ord('A') + 1)
        return n - 1  # 0-based

    pool_col = _col_letter_to_idx(pool_col_letter)
    date_start_col = _col_letter_to_idx(date_col_letter)

    try:
        if sheet:
            df_raw = pd.read_excel(saved_path, sheet_name=sheet, header=None)
        else:
            df_raw = pd.read_excel(saved_path, header=None)
    except Exception as e:
        print(f"    Warning: could not read monthly_balance saved_path "
              f"{saved_path}: {e}")
        return None, None

    # Resolve header row: prefer wizard's value, else autodetect (same
    # rules as legacy loader — first row with ≥3 datetime cells).
    hdr_idx = None
    if header_row is not None:
        try:
            hdr_idx = max(0, int(header_row) - 1)
        except (TypeError, ValueError):
            hdr_idx = None
    if hdr_idx is None:
        for i in range(min(8, len(df_raw))):
            row = df_raw.iloc[i]
            dt_count = sum(1 for v in row
                           if isinstance(v, (pd.Timestamp, datetime)))
            if dt_count >= 3:
                hdr_idx = i
                break
    if hdr_idx is None:
        return None, None

    # Resolve column anchors: prefer wizard letters, else autodetect.
    if date_start_col is None or pool_col is None:
        hdr_row = df_raw.iloc[hdr_idx]
        for c in range(df_raw.shape[1]):
            val = hdr_row.iloc[c]
            if (date_start_col is None
                    and isinstance(val, (pd.Timestamp, datetime))):
                date_start_col = c
            if (pool_col is None and isinstance(val, str)
                    and 'pool' in val.lower()):
                pool_col = c
        if date_start_col is None:
            return None, None
        if pool_col is None:
            pool_col = max(0, date_start_col - 1)

    dates = pd.to_datetime(
        df_raw.iloc[hdr_idx, date_start_col:].values, errors='coerce')

    # Optional label→pool translation map; when present, only rows whose
    # label appears in the map are included (translated to the pool
    # name). When absent, fall through to the legacy "use label as-is
    # with a few hard-coded skips" behavior.
    #
    # The wizard writes its mapping at ``monthly_balance.pool_map``;
    # legacy/manual configs put it at top-level ``balance_title_map``.
    # Honor both — wizard's nested map wins when present so that
    # downstream consumers (build_hist_bal_from_monthly,
    # _compute_balance_adjustments, etc.) see the user's pool names
    # instead of raw workbook labels.
    title_map = dict((config or {}).get('balance_title_map') or {})
    nested_map = mb_cfg.get('pool_map') or {}
    if nested_map:
        # Drop empty/None mappings (the wizard stores 'ignore' as '').
        for k, v in nested_map.items():
            if v:
                title_map[k] = v
    use_title_map = bool(title_map)

    # ACL row resolution: cfg['acl']['row'] is a 1-based row number on
    # the same sheet; cfg['acl']['label'] is an alternate text match.
    acl_cfg = (config or {}).get('acl') or {}
    acl_row_1based = acl_cfg.get('row')
    acl_label = (acl_cfg.get('label') or '').strip().lower()
    try:
        acl_row_idx = (int(acl_row_1based) - 1) if acl_row_1based else None
    except (TypeError, ValueError):
        acl_row_idx = None

    records = []
    alll_by_date: dict = {}
    # Track rows per source label so we can drop redundant aggregate
    # rows (Phase 9.39 — Destinations CU): when multiple workbook
    # labels collapse to the same canonical pool via pool_map (e.g.
    # 'Mastercard Loans' and 'Unsecured Credit Card Loans' both -> 
    # 'Unsecured Credit Card Loans'), naively summing them
    # double-counts at the months where both are populated. The
    # typical Vizo IDLR balance-sheet pattern has detailed monthly
    # sub-pool rows alongside NCUA-canonical aggregate roll-up rows
    # that carry values only at quarter-ends. Collect per-label rows
    # here; the subset-dedup pass below drops the aggregate.
    by_label: dict = {}

    # Extract ACL row first (by explicit row index when given).
    if acl_row_idx is not None and 0 <= acl_row_idx < df_raw.shape[0]:
        for j in range(len(dates)):
            if pd.notna(dates[j]):
                aval = df_raw.iloc[acl_row_idx, date_start_col + j]
                if pd.notna(aval):
                    try:
                        alll_by_date[dates[j]] = abs(float(aval))
                    except (ValueError, TypeError):
                        pass

    for i in range(hdr_idx + 1, df_raw.shape[0]):
        if i == acl_row_idx:
            continue  # already handled above
        raw_label = df_raw.iloc[i, pool_col]
        if pd.isna(raw_label) or str(raw_label).strip() == '':
            continue
        label = str(raw_label).strip()
        label_lc = label.lower()

        # ACL row by label match (when no explicit row given).
        if acl_row_idx is None and (
                (acl_label and label_lc == acl_label)
                or label in ('ALLL Balance', 'ACL Balance')):
            for j in range(len(dates)):
                if pd.notna(dates[j]):
                    aval = df_raw.iloc[i, date_start_col + j]
                    if pd.notna(aval):
                        try:
                            alll_by_date[dates[j]] = abs(float(aval))
                        except (ValueError, TypeError):
                            pass
            continue

        # Pool resolution.
        if use_title_map:
            pool_name = title_map.get(label)
            if not pool_name:
                continue  # label not opted-in; skip silently
        else:
            if label in ('Total', 'Total Loans'):
                continue
            if len(label) > 35 or label.startswith('In ') or label.startswith('Before'):
                continue
            pool_name = label

        key = (pool_name, label)
        entry = by_label.setdefault(
            key, {'pool': pool_name, 'label': label, 'rows': []})
        for j in range(len(dates)):
            if pd.notna(dates[j]):
                bal = df_raw.iloc[i, date_start_col + j]
                if pd.notna(bal):
                    try:
                        entry['rows'].append((dates[j], float(bal)))
                    except (ValueError, TypeError):
                        pass

    # Subset-dedup: when several source labels map to the same pool,
    # drop labels whose value-date set is a strict subset of another
    # label's date set in the same pool. Eliminates the common Vizo
    # IDLR balance-sheet shape where an NCUA-canonical aggregate row
    # (values only at Mar/Jun/Sep/Dec) duplicates a detailed monthly
    # sub-pool row that's already mapped to the same canonical pool.
    by_pool: dict = {}
    for entry in by_label.values():
        by_pool.setdefault(entry['pool'], []).append(entry)

    dropped_labels = []
    for pool_name, entries in list(by_pool.items()):
        if len(entries) < 2:
            continue
        date_sets = [(e, {d for d, _ in e['rows']}) for e in entries]
        survivors = []
        for e_i, ds_i in date_sets:
            is_subset = False
            for e_j, ds_j in date_sets:
                if e_i is e_j:
                    continue
                # Strict subset (and strictly smaller).
                if ds_i and ds_i <= ds_j and len(ds_i) < len(ds_j):
                    is_subset = True
                    break
            if is_subset:
                dropped_labels.append((e_i['label'], pool_name, len(ds_i)))
            else:
                survivors.append(e_i)
        by_pool[pool_name] = survivors

    if dropped_labels:
        print("    Dropped redundant balance labels (date set is a "
              "subset of another label mapped to the same pool — "
              "likely NCUA-canonical roll-up duplicating detailed "
              "monthly rows):")
        for lab, pool_name, n in dropped_labels:
            print(f"      - {lab!r} -> {pool_name!r} ({n} month(s))")

    # Emit records from surviving entries.
    for entries in by_pool.values():
        for entry in entries:
            for dt, bal in entry['rows']:
                rec = {
                    'pool': entry['pool'],
                    'date': dt,
                    'balance': bal,
                }
                if with_labels:
                    rec['label'] = entry['label']
                records.append(rec)

    out_df = pd.DataFrame(records)
    if out_df.empty and not alll_by_date:
        return None, None
    print(f"    Loaded monthly balances from wizard saved_path "
          f"({os.path.basename(saved_path)}): "
          f"{len(out_df)} rows, {len(alll_by_date)} ACL dates")
    return out_df, alll_by_date


_SUPP_MONTH_NAMES = {
    'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
    'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7,
    'aug': 8, 'august': 8, 'sep': 9, 'sept': 9, 'september': 9,
    'oct': 10, 'october': 10, 'nov': 11, 'november': 11,
    'dec': 12, 'december': 12,
}


def _period_from_month_name_in(name):
    """Return 'YYYY-MM-DD' (month-end) for a filename that carries a
    spelled-out month + 4-digit year (e.g. 'Balance Sheet June 2026.xlsx'),
    else None. Complements the numeric filename date parsing so CUs that
    name their balance sheets by month name still resolve a period."""
    import re as _re
    import calendar as _cal
    stem = os.path.splitext(os.path.basename(str(name)))[0]
    m = _re.search(r'(?<![A-Za-z])([A-Za-z]{3,9})[\s\-_]+(20\d{2})(?!\d)', stem)
    if not m:
        return None
    mo = _SUPP_MONTH_NAMES.get(m.group(1).strip().lower())
    if not mo:
        return None
    y = int(m.group(2))
    last = _cal.monthrange(y, mo)[1]
    return f"{y:04d}-{mo:02d}-{last:02d}"


def _discover_supplemental_monthly_balances(config, with_labels=False):
    """Discover per-month 'snapshot' balance files (e.g.
    ``June 2026 Loans by Type.xlsx``) in the data directory and return
    ``[{pool, date, balance}, ...]`` rows.

    Enabled only when ``config['monthly_balance']['monthly_file_pattern']``
    is set — a regex (case-insensitive) matched against each candidate
    filename. Each matched file is a single-month snapshot: a label column
    (``monthly_label_col``, default ``A``) and a balance column
    (``monthly_balance_col``, default ``B``), with the reporting period
    taken from the *filename* (month name + year, e.g. "June 2026").

    Lets a CU that switched from a wide historical workbook to per-month
    drops keep feeding the report without re-running setup: these rows are
    merged over the base balances so the snapshot wins for its own month.
    ``monthly_header_row`` (default 1) is the number of leading rows to
    skip before the first data row. The wizard-configured
    ``monthly_balance.filename`` (the historical wide workbook) is always
    excluded so a shared "Loans by Type" naming doesn't re-ingest it.
    """
    import re as _re

    mb = config.get('monthly_balance') or {}
    pat = str(mb.get('monthly_file_pattern') or '').strip()
    if not pat:
        return []
    data_dir = resolve_path(config.get('data_directory', ''))
    if not data_dir or not os.path.isdir(data_dir):
        return []
    try:
        rx = _re.compile(pat, _re.IGNORECASE)
    except _re.error as exc:
        print(f"    Warning: invalid monthly_file_pattern {pat!r}: {exc}")
        return []

    label_idx = _col_letter_to_idx(mb.get('monthly_label_col') or 'A')
    balance_idx = _col_letter_to_idx(mb.get('monthly_balance_col') or 'B')
    header_row = int(mb.get('monthly_header_row') or 1)
    # Optional section anchor: when set, each file is scanned for a row
    # containing this text (case-insensitive) and extraction begins on the
    # row BELOW it. Lets a combined balance sheet (e.g. a Deposit Account
    # Summary above a Loan Account Summary) be parsed without leaking the
    # non-loan section into pool balances -- robust to the upper section
    # changing length month to month.
    start_marker = str(mb.get('monthly_start_marker') or '').strip().lower()
    sheet = str(mb.get('monthly_sheet') or '').strip()
    base_name = str(mb.get('filename') or '').strip().lower()
    raw_map = mb.get('pool_map') or {}
    pool_map = {str(k).strip().lower(): str(v).strip()
                for k, v in raw_map.items()
                if str(k).strip() and str(v).strip()}
    split = str(config.get('pool_code_split') or '/').strip()
    # When the snapshot file is a formatted balance sheet that interleaves
    # non-loan rows (Cash, Bonds, Allowance, subtotals) in the same balance
    # column as the loan pools, ``monthly_strict_pool_map`` drops any label
    # not present in ``pool_map`` instead of falling back to using the raw
    # label as its own pool name (which would ingest those non-loan rows as
    # phantom pools). Opt-in so existing per-month single-pool drops keep
    # their permissive label passthrough.
    strict_map = bool(mb.get('monthly_strict_pool_map'))

    try:
        from import_data import extract_snapshot_date as _esd
    except Exception:  # noqa: BLE001
        _esd = None

    # Collect matching candidates (exclude the historical wide workbook).
    candidates = []
    for root, _dirs, files in os.walk(data_dir):
        for f in files:
            if not f.lower().endswith(('.xlsx', '.xlsm', '.xls')):
                continue
            if f.startswith('~$'):
                continue
            if base_name and f.strip().lower() == base_name:
                continue
            if not rx.search(f):
                continue
            candidates.append(os.path.join(root, f))
    # Newest-first so if two files map to the same month the latest wins.
    candidates.sort(key=os.path.getmtime, reverse=True)

    records = []
    seen_periods = set()
    for path in candidates:
        fn = os.path.basename(path)
        iso = None
        if _esd is not None:
            iso = _esd(fn, {'date_pattern': r'(20\d{2})-(\d{2})',
                            'date_format': 'YYYY-MM'})
        if not iso:
            # Fall back to a spelled-out month name in the filename
            # (e.g. 'Balance Sheet June 2026.xlsx').
            iso = _period_from_month_name_in(fn)
        if not iso:
            continue
        period = iso[:7]  # YYYY-MM
        if period in seen_periods:
            continue
        dt = pd.to_datetime(iso, errors='coerce')
        if pd.isna(dt):
            continue
        try:
            if sheet:
                try:
                    df = pd.read_excel(path, sheet_name=sheet, header=None)
                except Exception:
                    df = pd.read_excel(path, sheet_name=0, header=None)
            else:
                df = pd.read_excel(path, header=None)
        except Exception as exc:  # noqa: BLE001
            print(f"    Warning: could not read monthly-balance file {fn}: {exc}")
            continue
        if df is None or df.empty or df.shape[1] <= max(label_idx, balance_idx):
            continue
        # Resolve the first data row. Default to the configured header
        # offset; when a section marker is set, anchor below the marker
        # row so anything above it (e.g. a Deposit Account Summary) is
        # ignored regardless of its length.
        start_i = header_row
        if start_marker:
            for _mi in range(df.shape[0]):
                _hit = False
                for _c in range(df.shape[1]):
                    _v = df.iat[_mi, _c]
                    if isinstance(_v, str):
                        _vs = _v.strip().lower()
                        if _vs == start_marker or _vs.startswith(start_marker):
                            _hit = True
                            break
                if _hit:
                    start_i = _mi + 1
                    break
        seen_periods.add(period)
        for i in range(start_i, df.shape[0]):
            label = df.iat[i, label_idx]
            bal = df.iat[i, balance_idx]
            if pd.isna(label) or str(label).strip() == '':
                continue
            bal_f = _coerce_balance(bal)
            if bal_f is None:
                continue
            key = str(label).strip().lower()
            pool = pool_map.get(key)
            # Fall back to splitting composite codes (e.g. "IU/PL" -> "IU").
            if not pool and split and split in key:
                for part in key.split(split):
                    pool = pool_map.get(part.strip())
                    if pool:
                        break
            if not pool:
                if strict_map:
                    continue  # non-loan / subtotal row — skip
                pool = str(label).strip()
            rec = {'pool': pool, 'date': dt, 'balance': bal_f}
            if with_labels:
                rec['label'] = str(label).strip()
            records.append(rec)

    if seen_periods:
        print(f"    Supplemental monthly-balance snapshot file(s): "
              f"{len(seen_periods)} period(s) matched "
              f"({', '.join(sorted(seen_periods))}).")
    return records


def _apply_supplemental_monthly_balances(base_df, config, with_labels=False):
    """Merge per-month snapshot balance rows over ``base_df``.

    The snapshot rows win for any (year, month) they cover — the matching
    months are dropped from ``base_df`` first, then the snapshot rows are
    appended. Returns ``base_df`` unchanged when nothing is discovered.
    """
    try:
        supp = _discover_supplemental_monthly_balances(
            config, with_labels=with_labels)
    except Exception as exc:  # noqa: BLE001
        print(f"    Warning: supplemental monthly-balance discovery failed: {exc}")
        return base_df
    if not supp:
        return base_df
    _cols = ['pool', 'date', 'balance'] + (['label'] if with_labels else [])
    supp_df = pd.DataFrame(supp, columns=_cols)
    if base_df is None or base_df.empty:
        return supp_df
    try:
        supp_months = set(supp_df['date'].dt.to_period('M'))
        keep = ~base_df['date'].dt.to_period('M').isin(supp_months)
        merged = pd.concat([base_df[keep], supp_df], ignore_index=True)
        return merged
    except Exception as exc:  # noqa: BLE001
        print(f"    Warning: could not merge supplemental balances: {exc}")
        return base_df


def _apply_supplemental_wide_balances(base_df, config, with_labels=False):
    """Merge one or more *wide* (multi-month) supplemental balance files
    over ``base_df``, newer months winning.

    Enabled via ``config['monthly_balance']['supplemental_wide']`` — a dict
    or list of dicts, each describing an additional wide balance workbook in
    the same layout family as the main ``saved_path`` file (pool label
    column + a header row of month-end dates spread across columns). Use
    this when a CU starts delivering a fresh workbook (e.g.
    ``Loan Balance Sheet.xlsx`` covering only the latest few months)
    alongside the original historical workbook: the history stays intact and
    the newer file supplies the recent month(s).

    Each spec inherits the main block's ``sheet`` / ``pool_name_col`` /
    ``first_date_col`` / ``pool_map`` unless it overrides them, and must name
    the file via ``saved_path`` (absolute) or ``filename`` (resolved inside
    ``data_directory``). ``header_row`` typically differs from the main file
    and should be set per spec.
    """
    mb = (config or {}).get('monthly_balance') or {}
    specs = mb.get('supplemental_wide') or []
    if isinstance(specs, dict):
        specs = [specs]
    if not specs:
        return base_df
    data_dir = resolve_path(config.get('data_directory', ''))
    frames = []
    for spec in specs:
        if not isinstance(spec, dict):
            continue
        # Inherit the main monthly_balance metadata; the spec overrides.
        merged_cfg = {k: v for k, v in mb.items() if k != 'supplemental_wide'}
        for k, v in spec.items():
            if v is not None:
                merged_cfg[k] = v
        # Resolve the file from the SPEC (not the inherited saved_path):
        # spec.saved_path (absolute) wins, else spec.filename within data_dir.
        path = str(spec.get('saved_path') or '').strip()
        if not path or not os.path.isfile(path):
            fname = str(spec.get('filename') or '').strip()
            path = ''
            if fname and data_dir and os.path.isdir(data_dir):
                for root, _d, files in os.walk(data_dir):
                    if fname in files:
                        path = os.path.join(root, fname)
                        break
        if not path or not os.path.isfile(path):
            print(f"    Supplemental wide balance file not found: "
                  f"{spec.get('saved_path') or spec.get('filename')}")
            continue
        merged_cfg['saved_path'] = path
        merged_cfg['filename'] = os.path.basename(path)
        df, _alll = _load_monthly_balances_from_wizard(
            config, mb_cfg=merged_cfg, with_labels=with_labels)
        if df is not None and not df.empty:
            frames.append(df)
    if not frames:
        return base_df
    supp_df = pd.concat(frames, ignore_index=True)
    if base_df is None or base_df.empty:
        return supp_df
    try:
        supp_months = set(supp_df['date'].dt.to_period('M'))
        keep = ~base_df['date'].dt.to_period('M').isin(supp_months)
        merged = pd.concat([base_df[keep], supp_df], ignore_index=True)
        print(f"    Merged supplemental wide balance file(s): "
              f"{len(supp_months)} month(s) override base "
              f"({', '.join(str(p) for p in sorted(supp_months))}).")
        return merged
    except Exception as exc:  # noqa: BLE001
        print(f"    Warning: could not merge supplemental wide balances: {exc}")
        return base_df


def load_monthly_balances(config, with_labels=False):
    """Load monthly loan balances by pool from the most recent file available.
    Returns (DataFrame with columns [pool, date, balance],
            dict mapping date -> ALLL balance (absolute value)).

    When ``with_labels`` is True the returned DataFrame also carries a
    ``label`` column holding the raw balance-file row label each pool
    balance came from (used by the run-flow Loan Comparison page to show
    which monthly-balance titles roll into each pool). Off by default so
    the report engine's DataFrame shape is unchanged.
    """
    def _collapse(df):
        # Collapse to ONE row per (pool, date) by summing balance. Balance
        # files (and per-month snapshots like WSSC's worksheet) can map
        # several raw labels into the same pool for the same month; several
        # downstream consumers (build_hist_bal_from_monthly,
        # _compute_balance_adjustments) key on (pool, date) and would
        # otherwise keep only the FIRST/LAST label's balance and silently
        # drop the rest. Skipped when ``with_labels`` so the Loan Comparison
        # breakdown keeps its per-label rows.
        if with_labels or df is None or getattr(df, 'empty', True):
            return df
        if not {'pool', 'date', 'balance'}.issubset(df.columns):
            return df
        try:
            return (df.groupby(['pool', 'date'], as_index=False)['balance']
                    .sum())
        except Exception:  # noqa: BLE001
            return df

    # New (May 2026): the wizard can declare three delivery modes in
    # ``config["monthly_balance"]``. Honor per_month / manual modes first;
    # fall through to the legacy data_directory scan when no block is set
    # or source == "single".
    mb_cfg = config.get('monthly_balance') or {}
    mb_source = (mb_cfg.get('source') or '').strip().lower()
    if mb_source == 'manual':
        return _load_monthly_balances_manual(mb_cfg)
    if mb_source == 'per_month':
        df, alll = _load_monthly_balances_per_month(mb_cfg, acl_cfg=config.get('acl'))
        if not df.empty:
            df = _apply_supplemental_monthly_balances(df, config, with_labels=with_labels)
            df = _apply_supplemental_wide_balances(df, config, with_labels=with_labels)
            return _collapse(df), _merge_acl_history(alll, config)
        # If per_month failed (no files / unreadable), fall through to the
        # legacy scan so the user at least gets the historical context.
    if mb_source == 'per_year':
        df, alll = _load_monthly_balances_per_year(mb_cfg, acl_cfg=config.get('acl'))
        if not df.empty:
            df = _apply_supplemental_monthly_balances(df, config, with_labels=with_labels)
            df = _apply_supplemental_wide_balances(df, config, with_labels=with_labels)
            return _collapse(df), _merge_acl_history(alll, config)

    # Preferred path for "single" mode: use the wizard's saved_path +
    # sheet metadata directly. Honors cfg['balance_title_map'] (label
    # → pool translation) and cfg['acl']['row'/'label'] for ACL row
    # discovery. Falls through to the legacy data_directory scan when
    # the wizard metadata is missing or the file can't be read.
    wiz_df, wiz_alll = _load_monthly_balances_from_wizard(config, with_labels=with_labels)
    if wiz_df is not None:
        wiz_df = _apply_supplemental_monthly_balances(wiz_df, config, with_labels=with_labels)
        wiz_df = _apply_supplemental_wide_balances(wiz_df, config, with_labels=with_labels)
        return _collapse(wiz_df), _merge_acl_history(wiz_alll or {}, config)

    # No wizard base workbook, but the CU may deliver ONLY per-month snapshot
    # files (``monthly_file_pattern``) or a fresh wide workbook
    # (``supplemental_wide``) with no historical base. Apply those discovery
    # mechanisms against an empty base so the snapshot month still loads.
    if mb_cfg.get('monthly_file_pattern') or mb_cfg.get('supplemental_wide'):
        _empty_cols = ['pool', 'date', 'balance'] + (['label'] if with_labels else [])
        base = pd.DataFrame(columns=_empty_cols)
        base = _apply_supplemental_monthly_balances(base, config, with_labels=with_labels)
        base = _apply_supplemental_wide_balances(base, config, with_labels=with_labels)
        if base is not None and not base.empty:
            return _collapse(base), _merge_acl_history(wiz_alll or {}, config)

    data_dir = resolve_path(config.get('data_directory', ''))
    if not data_dir or not os.path.isdir(data_dir):
        return (pd.DataFrame(columns=['pool', 'date', 'balance']),
                _merge_acl_history({}, config))

    # Find balance files - match various naming conventions
    balance_files = []
    for root, dirs, files in os.walk(data_dir):
        for f in files:
            fl = f.lower()
            if not fl.endswith('.xlsx'):
                continue
            # Match: "monthly*balance*", "Loan Balances*", "*BalancesGrades*", "*LoanDataBalances*", "*cecl book*"
            if (('monthly' in fl and 'balance' in fl) or
                ('loan' in fl and 'balance' in fl) or
                ('balancesgrades' in fl) or
                ('loandatabalances' in fl) or
                ('cecl book' in fl or 'cecl_book' in fl)):
                balance_files.append(os.path.join(root, f))

    if not balance_files:
        return (pd.DataFrame(columns=['pool', 'date', 'balance']),
                _merge_acl_history({}, config))

    # Use the most recently modified file
    balance_files.sort(key=os.path.getmtime, reverse=True)
    filepath = balance_files[0]

    try:
        # Try to find a balance sheet (various naming conventions)
        xl = pd.ExcelFile(filepath)
        balance_sheet = None
        for sname in xl.sheet_names:
            sl = sname.lower()
            if 'balances by pool' in sl or ('loan' in sl and 'balance' in sl):
                balance_sheet = sname
                break
        if balance_sheet:
            df = pd.read_excel(filepath, sheet_name=balance_sheet, header=None)
        else:
            df = pd.read_excel(filepath, header=None)

        # Detect layout: find the header row with dates
        hdr_row_idx = None
        pool_col = None
        date_start_col = None
        for i in range(min(5, len(df))):
            row = df.iloc[i]
            # Look for a row that has datetime values
            date_count = sum(1 for v in row if isinstance(v, (pd.Timestamp, datetime)))
            if date_count >= 3:
                hdr_row_idx = i
                # Find where dates start and where pool names are
                for c in range(df.shape[1]):
                    val = row.iloc[c]
                    if isinstance(val, (pd.Timestamp, datetime)) and date_start_col is None:
                        date_start_col = c
                    if isinstance(val, str) and 'pool' in val.lower() and pool_col is None:
                        pool_col = c
                break

        if hdr_row_idx is None:
            # Fallback: assume row 1 has dates
            hdr_row_idx = 1
            date_start_col = 1
            pool_col = 0

        # If pool column wasn't labeled "Pool", it's the column just before dates
        if pool_col is None:
            pool_col = max(0, date_start_col - 1)

        dates = pd.to_datetime(df.iloc[hdr_row_idx, date_start_col:].values, errors='coerce')

        records = []
        alll_by_date = {}
        for i in range(hdr_row_idx + 1, df.shape[0]):
            pool_name = df.iloc[i, pool_col]
            if pd.isna(pool_name) or str(pool_name).strip() == '':
                continue
            pool_name = str(pool_name).strip()
            # Skip notes/metadata rows
            if len(pool_name) > 35 or pool_name.startswith('In ') or pool_name.startswith('Before'):
                continue
            if pool_name in ('ALLL Balance', 'Total', 'Total Loans', 'ACL Balance'):
                if pool_name in ('ALLL Balance', 'ACL Balance'):
                    for j in range(len(dates)):
                        if pd.notna(dates[j]):
                            aval = df.iloc[i, date_start_col + j]
                            if pd.notna(aval):
                                try:
                                    alll_by_date[dates[j]] = abs(float(aval))
                                except (ValueError, TypeError):
                                    pass
                continue
            for j in range(len(dates)):
                if pd.notna(dates[j]):
                    bal = df.iloc[i, date_start_col + j]
                    if pd.notna(bal):
                        try:
                            records.append({
                                'pool': pool_name,
                                'date': dates[j],
                                'balance': float(bal)
                            })
                        except (ValueError, TypeError):
                            pass
        return pd.DataFrame(records), _merge_acl_history(alll_by_date, config)
    except Exception as e:
        print(f"    Warning: Could not parse monthly balances {filepath}: {e}")
        return (pd.DataFrame(columns=['pool', 'date', 'balance']),
                _merge_acl_history({}, config))


def _compute_balance_adjustments(df, hist, config, snapshot_date):
    """Compare loan-file balances with monthly-balance-file totals per pool.

    If a difference exists, populate hist['impaired'] with balance_adjustments,
    total_balance_adjustment, and total_in_portfolio so the Pool_Balance Adjust
    sheet and migration matrix reflect the gap.
    """
    snap_dt = pd.Timestamp(snapshot_date)
    snap_ym = snap_dt.to_period('M')

    # Primary source: monthly balance file on disk.
    monthly_bals = {}
    monthly_df = hist.get('monthly_balances')
    if monthly_df is not None and not monthly_df.empty:
        mb = monthly_df.copy()
        mb['ym'] = mb['date'].dt.to_period('M')
        month_data = mb[mb['ym'] == snap_ym]
        for _, row in month_data.iterrows():
            pool = str(row['pool']).strip()
            monthly_bals[pool] = float(row['balance'])

    # A per-pool monthly balance FILE covering the snapshot month is the only
    # authoritative "monthly balance by pool/type" source for the current
    # period. When it is absent, CUs without a monthly balance by pool/type
    # (e.g. no monthly balance file yet for the quarter) should NOT have the
    # current loan file compared against the most recent prior month's balances
    # (the WARM/hist-bal fallbacks below). That comparison produces a
    # misleading adjustment (e.g. Dec balances vs Mar loans). Instead, leave the
    # balance-adjustment data unset so the Balance Adjustment section renders as
    # zeros. WARM CUs that already carry per-pool balances from the Risk Change
    # Data Entry tab (pool_bal_detail) are left on their existing code path.
    if not monthly_bals:
        _imp_existing = hist.get('impaired', {}) or {}
        if not _imp_existing.get('pool_bal_detail'):
            print("    Balance adjustments: no monthly balance by pool/type for "
                  f"{snap_ym} - section left at zero (no prior-month comparison)")
            return

    # Fallback: WARM template's per-pool monthly balance series. Use the
    # snapshot month's value (or the most recent value at/before snapshot).
    _SKIP_POOLS = {'grand total', 'total', 'exclude', 'excluded'}
    if not monthly_bals:
        # Prefer the pre-extension snapshot captured at load time; the live
        # hist_bal_data 'total' for snapshot month may have been overwritten
        # with loan-extract sums by extend_hist_bal_with_db.
        warm_snap = (hist.get('impaired') or {}).get('warm_snapshot_balances') or {}
        for pool, val in warm_snap.items():
            pname = str(pool).strip()
            if pname.lower() in _SKIP_POOLS:
                continue
            try:
                monthly_bals[pname] = float(val)
            except (TypeError, ValueError):
                continue

    if not monthly_bals:
        hbd = (hist.get('impaired') or {}).get('hist_bal_data') or {}
        for pool, pdata in hbd.items():
            pname = str(pool).strip()
            if pname.lower() in _SKIP_POOLS:
                continue
            dates = pdata.get('dates') or []
            tots = pdata.get('total') or []
            best_idx = None
            for i, d in enumerate(dates):
                try:
                    d_ym = pd.Timestamp(d).to_period('M')
                except Exception:
                    continue
                if d_ym <= snap_ym:
                    best_idx = i
                if d_ym == snap_ym:
                    break
            if best_idx is not None and best_idx < len(tots):
                try:
                    monthly_bals[pname] = float(tots[best_idx])
                except (TypeError, ValueError):
                    continue

    if not monthly_bals:
        return

    # Loan-file balances by pool
    loan_bals = df.groupby('loan_pool')['current_balance'].sum().to_dict()

    # Map monthly pool names to loan-file pool names (fuzzy match)
    pool_order = config.get('pool_order', [])

    def _match_pool(monthly_name):
        """Return the loan-file pool name that best matches a monthly name."""
        mn = monthly_name.lower().strip()
        # Direct match
        for lp in loan_bals:
            if lp.lower() == mn:
                return lp
        # Match against pool_order (canonical names)
        for pn in pool_order:
            if pn.lower() == mn:
                return pn
        # Prefix match (e.g. "Re-write(RW TM)" -> "Re-write")
        for lp in list(loan_bals.keys()) + pool_order:
            if mn.startswith(lp.lower()) or lp.lower().startswith(mn[:6]):
                return lp
        return None

    balance_adjustments = {}  # {pool: total_adj}
    total_adj = 0.0
    grand_loan = 0.0

    all_pools = set(loan_bals.keys())
    matched_loan_pools = set()
    nrr_set = set(config.get('not_risk_rated', []) or [])
    # Configured loan pools — used to drop balance-sheet line items that
    # aren't loan pools (e.g. "ACH Clearing", "Accrued Interest"). Anything
    # not in this set AND not already a loan-extract pool is ignored.
    configured_pools = set(pool_order)
    configured_pools.update(p.get('name') for p in (config.get('pools') or [])
                            if p and p.get('name'))
    configured_pools.update(nrr_set)
    skipped_non_pool: list[str] = []

    for mp, mb_val in monthly_bals.items():
        lp = _match_pool(mp)
        if lp is None:
            # Monthly-balance pool with no loan-extract counterpart (e.g.
            # non-risk-rated pools like Loan Participation, Repo/Foreclosed).
            # Treat the full monthly balance as an adjustment so it shows up
            # in the per-pool "Loans Not Risk Rated and Adjustments" row —
            # but ONLY when the name is a configured loan pool. Balance-sheet
            # line items like ACH Clearing have no place on the ACL tabs.
            mp_clean = str(mp).strip()
            if configured_pools and mp_clean not in configured_pools:
                # Try case-insensitive match before giving up.
                lc = mp_clean.lower()
                hit = next((c for c in configured_pools
                            if str(c).strip().lower() == lc), None)
                if hit is None:
                    if mb_val and abs(mb_val) > 0.005:
                        skipped_non_pool.append(f"{mp_clean} (${mb_val:,.2f})")
                    continue
                mp_clean = hit
            if mb_val and abs(mb_val) > 0.005:
                balance_adjustments[mp_clean] = round(float(mb_val), 2)
                total_adj += float(mb_val)
            continue
        matched_loan_pools.add(lp)
        lb = loan_bals.get(lp, 0)
        diff = mb_val - lb
        if abs(diff) > 0.005:  # ignore sub-penny rounding
            balance_adjustments[lp] = round(diff, 2)
            total_adj += diff

    # Include loan pools with no monthly match (balance goes unreported)
    # These don't need adjustments — they're just in the loan file

    total_adj = round(total_adj, 2)
    grand_loan = sum(loan_bals.values())
    total_in_portfolio = round(grand_loan + total_adj, 2)

    # Always build pool_bal_detail (even when balance_adjustments is empty)
    # so the Vizo "Pool_Balance Adjust" sheet has loan-file balances to
    # display. Without this, the sheet renders pool headers + zero rows
    # for any CU whose monthly balance file matches the loan extract at
    # the snapshot date (no adjustment needed).
    imp = hist.setdefault('impaired', {})
    if balance_adjustments:
        imp['balance_adjustments'] = balance_adjustments
        imp['total_balance_adjustment'] = total_adj
        imp['total_in_portfolio'] = total_in_portfolio

    # Build pool_bal_detail for the Vizo Balance Adjust sheet.
    # Per-grade detail uses loan-file balances; adjustment is pool-level.
    grades_cfg = config.get('credit_grades', config.get('grades', []))
    no_score = config.get('no_score_label', 'Not Reported')
    grade_labels = [g['label'] for g in grades_cfg] + [no_score]
    pool_bal_detail = {}
    for pool in set(list(loan_bals.keys()) + list(balance_adjustments.keys())):
        pool_df = df[df['loan_pool'] == pool]
        adj = balance_adjustments.get(pool, 0)
        gd = {}
        pool_loan_total = 0
        grade_bals = {}
        for g in grade_labels:
            g_bal = pool_df[pool_df['current_grade'] == g]['current_balance'].sum() \
                    if not pool_df.empty else 0
            grade_bals[g] = g_bal
            pool_loan_total += g_bal
        # Distribute pool adjustment across grades proportionally
        adj_remaining = adj
        for i, g in enumerate(grade_labels):
            g_bal = grade_bals[g]
            if pool_loan_total and adj:
                if i == len(grade_labels) - 1:
                    g_adj = adj_remaining  # last grade gets remainder to avoid rounding drift
                else:
                    g_adj = round(adj * g_bal / pool_loan_total, 2)
                    adj_remaining = round(adj_remaining - g_adj, 2)
            else:
                g_adj = 0
            gd[g] = {
                'loan_report_bal': g_bal,
                'bal_adj': g_adj,
                'balance_sheet_total': g_bal + g_adj,
            }
        gd['Total'] = {
            'loan_report_bal': pool_loan_total,
            'bal_adj': adj,
            'balance_sheet_total': pool_loan_total + adj,
        }
        pool_bal_detail[pool] = gd
    imp['pool_bal_detail'] = pool_bal_detail

    # Also store in config for the TCT Pool_Balance Adjust detail sheet
    config['balance_adjustments'] = {pool: {'_pool_total': adj}
                                     for pool, adj in balance_adjustments.items()}

    n_adj = len(balance_adjustments)
    n_pools_detail = len(pool_bal_detail)
    if n_adj:
        print(f"    Balance adjustments: {n_adj} pools, "
              f"total adj: ${total_adj:,.2f}, "
              f"total in portfolio: ${total_in_portfolio:,.2f}")
    else:
        print(f"    Balance adjustments: none (loan-file matches monthly); "
              f"built pool_bal_detail for {n_pools_detail} pool(s) so the "
              f"Vizo Pool_Balance Adjust tab can render loan-file balances")
    if skipped_non_pool:
        print(f"    Skipped {len(skipped_non_pool)} balance-sheet line "
              f"item(s) not mapped to any loan pool: "
              f"{', '.join(skipped_non_pool[:8])}"
              f"{'...' if len(skipped_non_pool) > 8 else ''}")


def load_delinquency_history(config):
    """Load delinquency data from all available quarterly files.
    Returns dict: {quarter_label: {pool: dq_balance}}"""
    data_dir = resolve_path(config.get('data_directory', ''))
    if not data_dir or not os.path.isdir(data_dir):
        return {}

    pool_map = config.get('pool_map', {})
    quarters = _find_quarter_folders(data_dir)
    dq_data = {}  # {quarter_label: {pool: dq_balance}}

    for folder, qlabel in quarters:
        for f in os.listdir(folder):
            fl = f.lower()
            if 'delq' in fl and fl.endswith('.xlsx'):
                filepath = os.path.join(folder, f)
                try:
                    df = pd.read_excel(filepath, header=None)
                    # Check for header row
                    first_vals = [str(v).lower() if pd.notna(v) else '' for v in df.iloc[0]]
                    if any('account' in v or 'delinq' in v for v in first_vals):
                        df = df.iloc[1:]
                    df = df.dropna(subset=[0])
                    df = df[pd.to_numeric(df[0], errors='coerce').notna()]

                    if df.empty:
                        continue

                    # Format: [account, suffix, code, balance, days]
                    # code col = 2, balance col = 3
                    for _, row in df.iterrows():
                        code = str(row[2]).strip().upper() if pd.notna(row[2]) else ''
                        pool = pool_map.get(code, pool_map.get(code.lower(), None))
                        bal = pd.to_numeric(row[3], errors='coerce')
                        if pool and pd.notna(bal):
                            dq_data.setdefault(qlabel, {})
                            dq_data[qlabel][pool] = dq_data[qlabel].get(pool, 0) + bal
                except Exception as e:
                    print(f"    Warning: Could not parse {filepath}: {e}")

    return dq_data


# ─────────────────────────────────────────────────────────────────────
# Manual-WARM template loader
#
# Some "prior reports" in Reports/ are not previously-generated TCTs but
# the manual CECL-Migration-WARM workbook the credit-union staff used to
# produce, imported as a baseline during new-CU setup.  Those workbooks
# carry the *same* historical data we want, but in differently-named
# sheets and with a one-column offset in the historical-balance grid:
#
#   Generated TCT sheet        Manual WARM sheet
#   ─────────────────────────  ─────────────────────────────────
#   > Detail_HIst Balances     HIst Bal Data        (col B blank, dates start col C)
#   Display CO-Recov-DQ        Display CO-Recov -DQ (note the stray space)
#   >Detail_Charge off Hist    Charge off History
#   ACL Env by Pool Mgmt Adj   ACL Env by Pool Mgmt Adj  (same name)
#
# These helpers detect a manual-WARM workbook and extract the same
# result dict shape that load_prior_tct_hist_bal produces from a
# generated TCT.  The WARM hist-bal grid is column-mapped (we record the
# actual column index of every date instead of assuming contiguous
# columns) so a leading blank column does not shift values.
# ─────────────────────────────────────────────────────────────────────

# Grade labels in HIst Bal Data that are not real grades and must be
# excluded from the synthesised hist_bal_data result.
_WARM_HIDDEN_GRADE_PREFIXES = ('hide', 'hide-', 'minimum', 'maximum',
                               'max for', 'min for')


def _is_warm_template_workbook(sheet_names):
    """Return True if the workbook looks like a manual WARM (vs. one of our
    generated TCT outputs)."""
    sn = set(sheet_names)
    has_warm_hist = 'HIst Bal Data' in sn
    has_tct_hist = '> Detail_HIst Balances' in sn
    return has_warm_hist and not has_tct_hist


def _warm_resolve_sheet(sheet_names, *candidates):
    """Return the first matching sheet name from candidates, or None."""
    sn = set(sheet_names)
    for c in candidates:
        if c in sn:
            return c
    return None


def _read_sheet_rows(ws):
    """Materialise a worksheet to a list[list] for fast random access."""
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _warm_parse_hist_bal(rows):
    """Parse a manual-WARM 'HIst Bal Data' sheet.

    Returns (hist_bal_data, pool_order, risk_rated). Each pool block is:
        Row N:   pool name in col A
        Row N+1: 'Current Grade' OR 'Current Risk Rating' in col A;
                 dates in cols C..  (col B blank)
        Row N+2..: grade label in col A, balance values in same cols as dates
        Row M:   'Total' in col A, balance values
        Row M+1: blank spacer

    Pools with a 'Current Risk Rating' header use BRR labels
    (Highest-Excellent / Good / Acceptable / Minimum / Watch /
    Substandard-Loss / Not Reported) instead of FICO grades.
    """
    hist_bal_data = {}
    pool_order = []
    risk_rated = {}

    r = 0
    n = len(rows)
    while r < n:
        row = rows[r] or []
        a = row[0] if row else None
        if a is None or str(a).strip() == '':
            r += 1
            continue
        a_s = str(a).strip()
        # Skip header / metadata rows
        low = a_s.lower()
        if low in ('current grade', 'current risk rating', 'total',
                   'balance', 'grand total', 'excluded', 'exclude') \
           or low.startswith(('for period', 'loss factor', 'allowance',
                              'charge off', 'tongass', 'siskiyou')) \
           or any(low.startswith(p) for p in ('hide', 'minimum', 'maximum',
                                              'max for')):
            r += 1
            continue

        # Need a 'Current Grade' or 'Current Risk Rating' row immediately
        # below to qualify as a pool. BRR-flagged pools (e.g. Commercial)
        # use the Risk Rating header with BRR labels in the rows below.
        if r + 1 >= n:
            break
        next_a = (rows[r + 1] or [None])[0]
        next_a_s = str(next_a).strip() if next_a is not None else ''
        if next_a_s not in ('Current Grade', 'Current Risk Rating'):
            r += 1
            continue

        pool_name = a_s
        hdr_row = rows[r + 1]

        # Capture date columns by their actual indices
        date_cols = []  # list[(col_idx, pd.Timestamp)]
        for ci in range(1, len(hdr_row)):
            v = hdr_row[ci]
            if v is None:
                continue
            try:
                date_cols.append((ci, pd.Timestamp(v)))
            except Exception:
                continue

        if not date_cols:
            r += 2
            continue

        dates = [d for _, d in date_cols]
        pool_grades = {}
        pool_total = []
        gr = r + 2
        while gr < n:
            grow = rows[gr] or []
            ga = grow[0] if grow else None
            if ga is None or str(ga).strip() == '':
                break
            glabel = str(ga).strip()
            glow = glabel.lower()

            vals = []
            for ci, _ in date_cols:
                v = grow[ci] if ci < len(grow) else 0
                try:
                    vals.append(float(v) if v is not None else 0.0)
                except (ValueError, TypeError):
                    vals.append(0.0)

            if glow == 'total':
                pool_total = vals
                gr += 1
                break
            # Filter out 'Hide-*', 'Minimum', etc.
            if any(glow.startswith(p) for p in _WARM_HIDDEN_GRADE_PREFIXES):
                gr += 1
                continue
            pool_grades[glabel] = vals
            gr += 1

        if pool_grades or pool_total:
            pool_order.append(pool_name)
            risk_rated[pool_name] = bool(pool_grades)
            # Ensure pool_total length always matches dates so downstream
            # code that does `pool_total[idx] = ...` is safe even when the
            # WARM block omitted a 'Total' row. Reconstruct from grade
            # rows when missing; pad with zeros as a last resort.
            if len(pool_total) != len(dates):
                if pool_grades:
                    rebuilt = []
                    for ci in range(len(dates)):
                        s = 0.0
                        for vals in pool_grades.values():
                            if ci < len(vals):
                                s += vals[ci]
                        rebuilt.append(s)
                    pool_total = rebuilt
                else:
                    pool_total = [0.0] * len(dates)
            hist_bal_data[pool_name] = {
                'dates': dates,
                'grades': pool_grades,
                'total': pool_total,
            }
        r = gr if gr > r + 2 else r + 2

    return hist_bal_data, pool_order, risk_rated


def _load_hist_from_warm_template(wb, snap):
    """Build the same result dict as load_prior_tct_hist_bal from a manual
    WARM workbook (HIst Bal Data + Display CO-Recov -DQ + Charge off History
    + ACL Env by Pool Mgmt Adj).
    """
    result = {}
    sheets = wb.sheetnames

    # ── Hist balances (synthetic '> Detail_HIst Balances') ──
    hb_sheet = _warm_resolve_sheet(sheets, 'HIst Bal Data')
    if hb_sheet:
        rows = _read_sheet_rows(wb[hb_sheet])
        hbd, pord, rrated = _warm_parse_hist_bal(rows)
        if hbd:
            result['hist_bal_data'] = hbd
            result['pool_order'] = pord
            result['risk_rated'] = rrated
            try:
                n_dates = max(len(d.get('dates', [])) for d in hbd.values())
            except ValueError:
                n_dates = 0
            print(f"    WARM template hist bal: {len(hbd)} pools, "
                  f"{n_dates} months")

    # ── CO/RC/Net/DQ year totals (Display CO-Recov -DQ) ──
    co_sheet = _warm_resolve_sheet(sheets, 'Display CO-Recov -DQ',
                                   'Display CO-Recov-DQ')
    if co_sheet:
        co_rows = _read_sheet_rows(wb[co_sheet])
        # Reuse the same section parser used for generated TCTs by
        # shelling out to a local copy (kept inline to avoid threading
        # the helper through module scope).
        warm_co, warm_co_tot, _ = _warm_parse_co_section(co_rows, 'Charge offs')
        warm_rc, warm_rc_tot, _ = _warm_parse_co_section(co_rows, 'Recoveries')
        warm_net, warm_net_tot, _ = _warm_parse_co_section(co_rows,
                                                            'Net Charge offs')
        warm_dq, _, _ = _warm_parse_co_section(co_rows, 'DQ %')
        # Recoveries are stored negative in WARM — flip to positive.
        for yr in warm_rc:
            for p in warm_rc[yr]:
                warm_rc[yr][p] = abs(warm_rc[yr][p])
        warm_rc_tot = {p: abs(v) for p, v in warm_rc_tot.items()}

        if warm_co:
            result['warm_co'] = warm_co
            result['warm_rc'] = warm_rc
            result['warm_net'] = warm_net
            result['warm_co_totals'] = warm_co_tot
            result['warm_rc_totals'] = warm_rc_tot
            result['warm_net_co'] = warm_net_tot
            if warm_dq:
                result['warm_dq_pct'] = warm_dq
            n_pools = len(set(p for yr in warm_co.values() for p in yr))
            n_years = len(warm_co)
            print(f"    WARM template CO/RC: {n_pools} pools, {n_years} years"
                  f" (CO totals: {len(warm_co_tot)} pools)")

    # ── Monthly CO / Recovery detail (Charge off History) ──
    mo_sheet = _warm_resolve_sheet(sheets, 'Charge off History',
                                   '>Detail_Charge off Hist')
    if mo_sheet:
        mo_rows = _read_sheet_rows(wb[mo_sheet])
        warm_co_mo = _warm_parse_monthly_co(mo_rows, 'Charge offs')
        warm_rc_mo = _warm_parse_monthly_co(mo_rows, 'Recoveries')
        # Recoveries again may be negative
        for ym in warm_rc_mo:
            for p in warm_rc_mo[ym]:
                warm_rc_mo[ym][p] = abs(warm_rc_mo[ym][p])
        if warm_co_mo:
            result['warm_co_monthly'] = warm_co_mo
            print(f"    WARM template monthly CO: {len(warm_co_mo)} months")
        if warm_rc_mo:
            result['warm_rc_monthly'] = warm_rc_mo

    return result


def _warm_parse_co_section(rows, start_label):
    """Parse a CO/RC/Net/DQ section of a Display CO-Recov-DQ-style sheet.

    Same shape & tolerances as the inline _parse_co_section inside
    load_prior_tct_hist_bal — kept separate so the WARM-template loader can
    be invoked outside that function.
    """
    import re as _re
    year_data = {}
    totals = {}
    header_row = None
    pool_start = None
    col_years = []
    acl_col = None

    for ri, row in enumerate(rows):
        c0 = str((row[0] if row else '') or '').strip().lower()
        if c0 == start_label.lower():
            header_row = ri
            for ci in range(1, len(row)):
                val = row[ci]
                if val is None:
                    continue
                sv = str(val).strip()
                m = _re.match(r'(?:YTD\s+)?(\d{4})', sv)
                if m:
                    col_years.append((ci, int(m.group(1))))
                elif 'acl' in sv.lower() or 'net charge' in sv.lower():
                    acl_col = ci
            pool_start = ri + 1
            break

    if header_row is None:
        return year_data, totals, []

    for ri in range(pool_start, len(rows)):
        row = rows[ri] or []
        pool_name = str((row[0] if row else '') or '').strip()
        if not pool_name:
            break
        pl = pool_name.lower()
        if any(kw in pl for kw in ['recoveries', 'net charge', 'dq %',
                                    'charge offs']):
            break
        # Skip 'Hide-*' / 'Exclude' rows
        if pl.startswith(('hide', 'exclude')):
            continue

        for ci, yr in col_years:
            val = row[ci] if ci < len(row) else None
            if val is not None and val != 0:
                try:
                    fval = float(val)
                except (ValueError, TypeError):
                    continue
                year_data.setdefault(yr, {})[pool_name] = fval

        if acl_col is not None and acl_col < len(row):
            aval = row[acl_col]
            if aval is not None:
                try:
                    totals[pool_name] = float(aval)
                except (ValueError, TypeError):
                    pass

    return year_data, totals, [yr for _, yr in col_years]


def _warm_parse_monthly_co(rows, start_label):
    """Parse the WARM 'Charge off History' sheet's monthly section.

    Date headers are on the same row as the section label (e.g. row 9 for
    Charge offs, with col B = 2019-01-31, col C = 2019-02-28, ...).  Pool
    rows follow until a blank or 'Recoveries' label.
    """
    import datetime as _dt
    monthly = {}

    header_row = None
    for ri, row in enumerate(rows):
        c0 = str((row[0] if row else '') or '').strip().lower()
        if c0 == start_label.lower():
            header_row = ri
            break
    if header_row is None:
        return monthly

    date_cols = []
    for ci in range(1, len(rows[header_row])):
        v = rows[header_row][ci]
        if isinstance(v, _dt.datetime):
            date_cols.append((ci, v))

    if not date_cols:
        return monthly

    for ri in range(header_row + 1, len(rows)):
        row = rows[ri] or []
        pool_name = str((row[0] if row else '') or '').strip()
        if not pool_name:
            break
        pl = pool_name.lower()
        if pl in ('recoveries', 'charge offs', 'net charge offs', 'dq %') \
           or pl.startswith('total'):
            break
        if pl.startswith(('hide', 'exclude')):
            continue
        for ci, dt_val in date_cols:
            val = row[ci] if ci < len(row) else None
            if val is not None and val != 0:
                try:
                    fval = float(val)
                except (ValueError, TypeError):
                    continue
                ym = (dt_val.year, dt_val.month)
                monthly.setdefault(ym, {})[pool_name] = fval

    return monthly


def load_prior_tct_hist_bal(config, snap):
    """Load hist_bal_data from the most recent prior TCT report.

    When no WARM file exists for the current snapshot (i.e. the TCT report IS
    the new WARM replacement), this function reads the previous TCT report's
    '> Detail_HIst Balances' sheet so historical months carry forward.

    Returns dict compatible with hist['impaired'] keys:
      'hist_bal_data': {pool: {dates, grades, total}},
      'pool_order': [...],
      'acl_months': {pool: n},
      'risk_rated': {pool: bool},
    or {} if no prior report found.
    """
    from openpyxl import load_workbook as _load_wb

    cu = config['credit_union']
    safe_cu = cu.replace(' ', '_').replace('/', '-')

    # Search Reports/ for prior TCT files for this credit union
    rpt_dir = os.path.join(BASE, 'Reports')
    if not os.path.isdir(rpt_dir):
        return {}

    pattern = re.compile(
        rf'\d{{4}}-\d{{2}}-\d{{2}}_CECL_Migration_{re.escape(safe_cu)}_TCT_Model\.xlsx$',
        re.IGNORECASE,
    )
    candidates = []
    for root, dirs, files in os.walk(rpt_dir):
        for f in files:
            if f.startswith('~$'):
                continue
            if pattern.match(f):
                # Extract date from filename
                date_str = f[:10]
                # Allow equal-date matches only when the file lives in the
                # dedicated WARM-baseline subdir (so a WARM workbook
                # uploaded as the baseline for the current snapshot still
                # supplies historical hist_bal data). Exclude same-date
                # top-level files because those are this run's own output.
                in_warm_baselines = '_warm_baselines' in os.path.normpath(
                    os.path.relpath(root, rpt_dir)
                ).split(os.sep)
                if date_str < snap or (date_str == snap and in_warm_baselines):
                    candidates.append((date_str, in_warm_baselines,
                                       os.path.join(root, f)))

    if not candidates:
        return {}

    # Sort by date desc; for ties, prefer WARM-baseline (in_warm_baselines=True).
    candidates.sort(key=lambda c: (c[0], c[1]), reverse=True)
    prior_date, _, prior_path = candidates[0]
    print(f"    Loading prior TCT hist balances from: {os.path.basename(prior_path)}")

    try:
        wb = _load_wb(prior_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"    Warning: Could not open prior report: {e}")
        return {}

    # Branch: manual-WARM workbook (uploaded as baseline during new-CU
    # setup) uses different sheet names than our generated TCTs.  Use the
    # WARM-template loader for everything except the ACL Env sheet, which
    # has the same name in both formats and is parsed below.
    if _is_warm_template_workbook(wb.sheetnames):
        print(f"    Detected manual-WARM workbook layout in "
              f"{os.path.basename(prior_path)}; using WARM template loader.")
        result = _load_hist_from_warm_template(wb, snap)

        # Continue into the existing 'ACL Env by Pool Mgmt Adj' parser
        # below so management adjustments + env factors carry forward.
        acl_rows_data = []
        if 'ACL Env by Pool Mgmt Adj' in wb.sheetnames:
            ws_acl = wb['ACL Env by Pool Mgmt Adj']
            for row in ws_acl.iter_rows(min_row=1, max_row=ws_acl.max_row,
                                        max_col=ws_acl.max_column,
                                        values_only=True):
                acl_rows_data.append(list(row))
        wb.close()

        if acl_rows_data:
            prior_mgmt_adj = {}
            prior_env_factor = {}
            current_pool = None
            for row in acl_rows_data:
                a_val = row[0] if row else None
                if a_val is None:
                    continue
                label = str(a_val).strip()
                if not label:
                    continue
                e_val = row[4] if len(row) > 4 else None
                f_val = row[5] if len(row) > 5 else None
                i_val = row[8] if len(row) > 8 else None
                if label in ('Current Grade', 'Current Risk Rating'):
                    continue
                if label == 'Total':
                    if current_pool and i_val is not None:
                        try:
                            prior_env_factor[current_pool] = float(i_val)
                        except (ValueError, TypeError):
                            pass
                    current_pool = None
                    continue
                if current_pool and e_val is not None:
                    try:
                        mgmt = float(f_val) if f_val is not None else 0.0
                    except (ValueError, TypeError):
                        mgmt = 0.0
                    if mgmt != 0:
                        prior_mgmt_adj.setdefault(current_pool, {})[label] = mgmt
                    continue
                if label not in ('Current Grade', 'Current Risk Rating',
                                 'Total') and e_val is None:
                    current_pool = label
            if prior_mgmt_adj or prior_env_factor:
                result['prior_mgmt_adj'] = prior_mgmt_adj
                result['prior_env_factor'] = prior_env_factor
                print(f"    Prior ACL adjustments (WARM): "
                      f"{len(prior_mgmt_adj)} pools with mgmt adj, "
                      f"{len(prior_env_factor)} pools with env factor")
        return result

    if '> Detail_HIst Balances' not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb['> Detail_HIst Balances']

    # Parse structure:
    # Row 5+: pool blocks, each has:
    #   pool_name row  (col A = name)
    #   "Current Grade" row  (col A = "Current Grade", cols B+ = dates)
    #   grade rows  (col A = grade label, cols B+ = values)
    #   "Total" row  (col A = "Total", cols B+ = values)
    #   blank spacer row

    result = {}
    hist_bal_data = {}
    pool_order = []
    acl_months = {}
    risk_rated = {}

    # Read all cell values into memory for fast random access
    rows_data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            max_col=ws.max_column, values_only=True):
        rows_data.append(list(row))

    # Also read CO-Recov-DQ sheet if present
    co_rows_data = []
    if 'Display CO-Recov-DQ' in wb.sheetnames:
        ws_co = wb['Display CO-Recov-DQ']
        for row in ws_co.iter_rows(min_row=1, max_row=ws_co.max_row,
                                   max_col=ws_co.max_column, values_only=True):
            co_rows_data.append(list(row))

    # Read monthly CO detail sheet if present
    co_monthly_rows = []
    if '>Detail_Charge off Hist' in wb.sheetnames:
        ws_co_det = wb['>Detail_Charge off Hist']
        for row in ws_co_det.iter_rows(min_row=1, max_row=ws_co_det.max_row,
                                       max_col=ws_co_det.max_column, values_only=True):
            co_monthly_rows.append(list(row))

    # Read ACL sheet to carry forward management adjustments & env factors
    acl_rows_data = []
    if 'ACL Env by Pool Mgmt Adj' in wb.sheetnames:
        ws_acl = wb['ACL Env by Pool Mgmt Adj']
        for row in ws_acl.iter_rows(min_row=1, max_row=ws_acl.max_row,
                                    max_col=ws_acl.max_column, values_only=True):
            acl_rows_data.append(list(row))

    wb.close()

    if not rows_data:
        return {}

    max_col = max(len(r) for r in rows_data)
    r = 4  # 0-indexed, start at row 5 (Excel row 5)
    while r < len(rows_data):
        row_vals = rows_data[r]
        a_val = row_vals[0] if row_vals else None

        if a_val is None or str(a_val).strip() == '':
            r += 1
            continue

        pool_name = str(a_val).strip()
        if pool_name in ('Current Grade', 'Current Risk Rating', 'Total',
                         'Balance', '% of Loans',
                         'WARM\nMonths', 'Loss Factor Historical Detail'):
            r += 1
            continue
        # Skip header rows
        if pool_name.startswith('For Quarter') or pool_name == cu:
            r += 1
            continue

        # This should be a pool name; next row should be "Current Grade" or "Balance"
        if r + 1 >= len(rows_data):
            break
        next_a = rows_data[r + 1][0] if rows_data[r + 1] else None
        next_label = str(next_a).strip() if next_a else ''

        # 'Current Risk Rating' header is used for BRR-flagged pools in
        # both manual WARM workbooks and our generated TCT reports.
        if next_label not in ('Current Grade', 'Current Risk Rating',
                               'Balance'):
            r += 1
            continue

        pool_order.append(pool_name)
        is_rr = (next_label in ('Current Grade', 'Current Risk Rating'))
        risk_rated[pool_name] = is_rr

        # Read dates from the header row (row r+1)
        hdr_row = rows_data[r + 1]
        dates = []
        date_start_col = 1  # 0-indexed col B
        for ci in range(date_start_col, len(hdr_row)):
            v = hdr_row[ci]
            if v is None:
                continue
            if isinstance(v, str):
                vs = v.strip()
                if 'WARM' in vs:
                    # ACL months column — read value from first data row
                    for gr in range(r + 2, min(r + 20, len(rows_data))):
                        acl_val = rows_data[gr][ci] if ci < len(rows_data[gr]) else None
                        if acl_val is not None:
                            try:
                                acl_months[pool_name] = int(acl_val)
                            except (ValueError, TypeError):
                                pass
                            break
                    break
                # Skip non-date string headers like "% of Loans"
                continue
            try:
                dates.append(pd.Timestamp(v))
            except Exception:
                continue

        if not is_rr:
            # Non-risk-rated: just has Balance header + Total row
            r += 2  # skip to Total row
            total_row = rows_data[r] if r < len(rows_data) else []
            total_vals = []
            for ci in range(date_start_col, date_start_col + len(dates)):
                v = total_row[ci] if ci < len(total_row) else 0
                try:
                    total_vals.append(float(v) if v is not None else 0.0)
                except (ValueError, TypeError):
                    total_vals.append(0.0)
            hist_bal_data[pool_name] = {
                'dates': dates,
                'grades': {},
                'total': total_vals,
            }
            r += 1
            continue

        # Risk-rated pool: read grade rows
        pool_grades = {}
        pool_total = []
        gr_idx = r + 2  # first grade row
        while gr_idx < len(rows_data):
            ga = rows_data[gr_idx][0] if rows_data[gr_idx] else None
            if ga is None or str(ga).strip() == '':
                break
            glabel = str(ga).strip()
            vals = []
            for ci in range(date_start_col, date_start_col + len(dates)):
                v = rows_data[gr_idx][ci] if ci < len(rows_data[gr_idx]) else 0
                try:
                    vals.append(float(v) if v is not None else 0.0)
                except (ValueError, TypeError):
                    vals.append(0.0)
            if glabel == 'Total':
                pool_total = vals
                gr_idx += 1
                break
            else:
                pool_grades[glabel] = vals
            gr_idx += 1

        hist_bal_data[pool_name] = {
            'dates': dates,
            'grades': pool_grades,
            'total': pool_total,
        }
        r = gr_idx
        continue

    if hist_bal_data:
        result['hist_bal_data'] = hist_bal_data
        result['pool_order'] = pool_order
        if acl_months:
            result['acl_months'] = acl_months
        if risk_rated:
            result['risk_rated'] = risk_rated
        n_dates = max(len(d.get('dates', [])) for d in hist_bal_data.values())
        print(f"    Prior TCT hist bal: {len(hist_bal_data)} pools, {n_dates} months (from {prior_date})")

    # ── Parse CO/Recovery data from prior report's Display CO-Recov-DQ ──
    if co_rows_data:
        warm_co, warm_rc, warm_net = {}, {}, {}
        warm_co_totals, warm_rc_totals, warm_net_co = {}, {}, {}
        warm_dq_pct = {}

        def _parse_co_section(rows, start_label):
            """Parse a CO/RC/Net/DQ section from the Display CO-Recov-DQ sheet.

            Returns (year_data, totals, acl_col_years).
            year_data = {year(int): {pool: value}}
            totals = {pool: total}
            acl_col_years = list of year ints from header row
            """
            import re
            year_data = {}
            totals = {}
            header_row = None
            pool_start = None
            col_years = []
            acl_col = None  # column index of the ACL totals column

            for ri, row in enumerate(rows):
                c0 = str(row[0] or '').strip().lower()
                if c0 == start_label.lower():
                    header_row = ri
                    # Parse year columns from header
                    for ci in range(1, len(row)):
                        val = row[ci]
                        if val is None:
                            continue
                        sv = str(val).strip()
                        # Match year (e.g. "2019") or "YTD 2026"
                        m = re.match(r'(?:YTD\s+)?(\d{4})', sv)
                        if m:
                            col_years.append((ci, int(m.group(1))))
                        elif 'acl' in sv.lower() or 'net charge' in sv.lower():
                            acl_col = ci
                    pool_start = ri + 1
                    break

            if header_row is None:
                return year_data, totals, []

            # Read pool rows until blank or next section header
            for ri in range(pool_start, len(rows)):
                row = rows[ri]
                pool_name = str(row[0] or '').strip()
                if not pool_name:
                    break
                # Stop at next section header
                pl = pool_name.lower()
                if any(kw in pl for kw in ['recoveries', 'net charge', 'dq %',
                                            'charge offs']):
                    break

                for ci, yr in col_years:
                    val = row[ci] if ci < len(row) else None
                    if val is not None and val != 0:
                        try:
                            fval = float(val)
                        except (ValueError, TypeError):
                            continue
                        year_data.setdefault(yr, {})[pool_name] = fval

                # ACL total column
                if acl_col is not None and acl_col < len(row):
                    aval = row[acl_col]
                    if aval is not None:
                        try:
                            totals[pool_name] = float(aval)
                        except (ValueError, TypeError):
                            pass

            return year_data, totals, [yr for _, yr in col_years]

        # Parse Charge offs section
        co_data, co_tots, co_years = _parse_co_section(co_rows_data,
                                                        'Charge offs')
        if co_data:
            warm_co = co_data
            warm_co_totals = co_tots

        # Parse Recoveries section — values may be negative in prior WARM
        rc_data, rc_tots, _ = _parse_co_section(co_rows_data, 'Recoveries')
        if rc_data:
            # Ensure recovery values are positive (WARM stores them negative)
            for yr in rc_data:
                for pool in rc_data[yr]:
                    rc_data[yr][pool] = abs(rc_data[yr][pool])
            for pool in rc_tots:
                rc_tots[pool] = abs(rc_tots[pool])
            warm_rc = rc_data
            warm_rc_totals = rc_tots

        # Parse Net Charge offs section
        net_data, net_tots, _ = _parse_co_section(co_rows_data,
                                                   'Net Charge offs')
        if net_data:
            warm_net = net_data
            warm_net_co = net_tots

        # Parse DQ % section
        dq_data, _, _ = _parse_co_section(co_rows_data, 'DQ %')
        if dq_data:
            warm_dq_pct = dq_data

        if warm_co:
            result['warm_co'] = warm_co
            result['warm_rc'] = warm_rc
            result['warm_net'] = warm_net
            result['warm_co_totals'] = warm_co_totals
            result['warm_rc_totals'] = warm_rc_totals
            result['warm_net_co'] = warm_net_co
            if warm_dq_pct:
                result['warm_dq_pct'] = warm_dq_pct
            n_pools = len(set(p for yr in warm_co.values() for p in yr))
            n_years = len(warm_co)
            print(f"    Prior TCT CO/RC: {n_pools} pools, {n_years} years"
                  f" (CO totals: {len(warm_co_totals)} pools)")

    # ── Parse monthly CO detail from prior report ──
    if co_monthly_rows:
        import datetime as _dt
        warm_co_monthly = {}
        warm_rc_monthly = {}

        def _parse_monthly_section(rows, start_label, pool_list):
            """Parse a monthly CO or RC section.

            Returns {(year, month): {pool: amount}}.
            """
            monthly = {}
            header_row = None
            date_cols = []  # [(col_idx, datetime)]

            for ri, row in enumerate(rows):
                c0 = str(row[0] or '').strip().lower()
                if c0 == start_label.lower():
                    header_row = ri
                    # Header is one row above the pool rows; dates in cols B+
                    # Actually the header with dates is this row itself
                    # But sometimes the dates are in an earlier header row
                    break

            if header_row is None:
                return monthly

            # Find the date header row (row 5 for CO, check the previous
            # occurrence of dates)
            # Dates are in the same row as the section label or in a prior row
            # Check if this row has dates
            for ci in range(1, len(rows[header_row])):
                val = rows[header_row][ci]
                if isinstance(val, _dt.datetime):
                    date_cols.append((ci, val))

            # If no dates in header row, check above
            if not date_cols:
                # The dates are typically in row 5 (index 4)
                for ri in range(header_row - 1, -1, -1):
                    for ci in range(1, len(rows[ri])):
                        val = rows[ri][ci]
                        if isinstance(val, _dt.datetime):
                            date_cols.append((ci, val))
                    if date_cols:
                        break

            if not date_cols:
                return monthly

            # Read pool rows
            for pi, pool in enumerate(pool_list):
                ri = header_row + 1 + pi
                if ri >= len(rows):
                    break
                row = rows[ri]
                row_label = str(row[0] or '').strip()
                # Verify pool name matches (or just read in order)
                for ci, dt_val in date_cols:
                    val = row[ci] if ci < len(row) else None
                    if val is not None and val != 0:
                        try:
                            fval = float(val)
                        except (ValueError, TypeError):
                            continue
                        ym = (dt_val.year, dt_val.month)
                        monthly.setdefault(ym, {})[pool] = fval

            return monthly

        # Get pool list from CO section (row labels)
        monthly_pools = []
        for ri, row in enumerate(co_monthly_rows):
            c0 = str(row[0] or '').strip().lower()
            if c0 == 'charge offs':
                # Read pool names from subsequent rows until "Total"
                for pi in range(ri + 1, len(co_monthly_rows)):
                    prow = co_monthly_rows[pi]
                    pname = str(prow[0] or '').strip()
                    if not pname or pname.lower().startswith('total'):
                        break
                    monthly_pools.append(pname)
                break

        if monthly_pools:
            warm_co_monthly = _parse_monthly_section(
                co_monthly_rows, 'Charge offs', monthly_pools)
            warm_rc_monthly = _parse_monthly_section(
                co_monthly_rows, 'Recoveries', monthly_pools)

            if warm_co_monthly:
                result['warm_co_monthly'] = warm_co_monthly
            if warm_rc_monthly:
                result['warm_rc_monthly'] = warm_rc_monthly
            n_mo = len(warm_co_monthly)
            print(f"    Prior TCT monthly CO: {n_mo} months, "
                  f"{len(monthly_pools)} pools")

    # ── Parse management adjustments & env factors from prior ACL sheet ──
    if acl_rows_data:
        prior_mgmt_adj = {}   # {pool: {grade: float}}
        prior_env_factor = {}  # {pool: float}
        current_pool = None
        for row in acl_rows_data:
            a_val = row[0] if row else None
            if a_val is None:
                continue
            label = str(a_val).strip()
            if not label:
                continue
            # Detect pool header: next row-ish will have 'Current Grade'
            # Grade rows have data in columns B-H; pool headers don't have col E data
            e_val = row[4] if len(row) > 4 else None
            f_val = row[5] if len(row) > 5 else None
            i_val = row[8] if len(row) > 8 else None
            if label in ('Current Grade', 'Current Risk Rating'):
                continue
            if label == 'Total':
                # Pool total row — read env factor (col I, index 8)
                if current_pool and i_val is not None:
                    try:
                        prior_env_factor[current_pool] = float(i_val)
                    except (ValueError, TypeError):
                        pass
                current_pool = None
                continue
            # If this row has a base_rate in col E but no 'Total' label,
            # it's a grade data row
            if current_pool and e_val is not None:
                try:
                    mgmt = float(f_val) if f_val is not None else 0.0
                except (ValueError, TypeError):
                    mgmt = 0.0
                if mgmt != 0:
                    prior_mgmt_adj.setdefault(current_pool, {})[label] = mgmt
                continue
            # Otherwise this might be a pool header
            if label not in ('Current Grade', 'Current Risk Rating',
                             'Total') and e_val is None:
                current_pool = label

        if prior_mgmt_adj or prior_env_factor:
            result['prior_mgmt_adj'] = prior_mgmt_adj
            result['prior_env_factor'] = prior_env_factor
            n_pools_ma = len(prior_mgmt_adj)
            n_pools_ef = len(prior_env_factor)
            print(f"    Prior ACL adjustments: {n_pools_ma} pools with mgmt adj, "
                  f"{n_pools_ef} pools with env factor")

    return result


def build_hist_bal_from_monthly(monthly_balances, df, snap, grades, config):
    """Build a fresh hist_bal_data dict from monthly pool balances + current snapshot.

    Used when no WARM file and no prior TCT report exist for the credit union.
    Pool-level monthly totals come from the monthly balances workbook; grade-
    level distribution is allocated proportionally using the current snapshot's
    grade mix per pool.

    Returns dict shaped like load_prior_tct_hist_bal output:
      {'hist_bal_data': {...}, 'pool_order': [...], 'risk_rated': {...}}
    or {} if no monthly balance data is available.
    """
    if monthly_balances is None or monthly_balances.empty:
        return {}

    no_score = config.get('no_score_label', 'Not Reported')
    all_gl = [g['label'] for g in grades] + [no_score]
    snap_ts = pd.Timestamp(snap) + pd.offsets.MonthEnd(0)

    # Map monthly-file pool names to DB pool names (case-insensitive, strip parens)
    db_pools = list(df['loan_pool'].dropna().unique())
    pool_norm = {}
    for p in db_pools:
        pool_norm[p.strip().lower()] = p
        clean = re.sub(r'\s*\(.*\)\s*$', '', str(p)).strip()
        if clean.lower() != str(p).strip().lower():
            pool_norm[clean.lower()] = p

    hist_bal_data = {}
    pool_order = []
    risk_rated = {}

    for pool_key, grp in monthly_balances.groupby('pool'):
        pk = str(pool_key).strip()
        mapped = pool_norm.get(pk.lower())
        if not mapped:
            clean = re.sub(r'\s*\(.*\)\s*$', '', pk).strip()
            mapped = pool_norm.get(clean.lower())
        if not mapped:
            # Pool exists in monthly file but not in current DB snapshot — keep
            # it as a non-risk-rated pool so its history still displays.
            mapped = pk
            is_rr = False
        else:
            is_rr = True

        # Build current-snapshot grade percentages for this pool
        pcts = {}
        if is_rr:
            pdf = df[df['loan_pool'] == mapped]
            ptotal = pdf['current_balance'].sum()
            if ptotal > 0:
                for g in all_gl:
                    gbal = pdf[pdf['current_grade'] == g]['current_balance'].sum()
                    pcts[g] = gbal / ptotal
            else:
                is_rr = False

        # Sorted, unique month-end dates from monthly file
        sorted_grp = grp.sort_values('date')
        dates = []
        totals = []
        seen = set()
        for _, row in sorted_grp.iterrows():
            dt = pd.Timestamp(row['date']) + pd.offsets.MonthEnd(0)
            if dt in seen:
                continue
            seen.add(dt)
            dates.append(dt)
            totals.append(float(row['balance']))

        # Ensure current snapshot date is the last entry (use DB total if present)
        if is_rr:
            db_total = float(df[df['loan_pool'] == mapped]['current_balance'].sum())
        else:
            db_total = None
        if snap_ts not in seen:
            dates.append(snap_ts)
            totals.append(db_total if db_total is not None else 0.0)
        elif db_total is not None:
            idx = dates.index(snap_ts)
            totals[idx] = db_total

        if not dates:
            continue

        if is_rr and pcts:
            # Use exact grade balances for the snap month from the DB; allocate
            # earlier months proportionally with the same percentages.
            grade_vals = {g: [] for g in all_gl}
            pdf = df[df['loan_pool'] == mapped]
            for di, dt in enumerate(dates):
                pool_total = totals[di]
                if dt == snap_ts:
                    for g in all_gl:
                        grade_vals[g].append(
                            float(pdf[pdf['current_grade'] == g]['current_balance'].sum())
                        )
                else:
                    for g in all_gl:
                        grade_vals[g].append(pool_total * pcts.get(g, 0.0))
        else:
            grade_vals = {}

        hist_bal_data[mapped] = {
            'dates': dates,
            'grades': grade_vals,
            'total': totals,
        }
        pool_order.append(mapped)
        risk_rated[mapped] = bool(grade_vals)

    # Add pools that are in DB snapshot but not in monthly file (single-point entry)
    for pool in db_pools:
        if pool in hist_bal_data:
            continue
        pdf = df[df['loan_pool'] == pool]
        ptotal = float(pdf['current_balance'].sum())
        grade_vals = {}
        for g in all_gl:
            grade_vals[g] = [float(pdf[pdf['current_grade'] == g]['current_balance'].sum())]
        hist_bal_data[pool] = {
            'dates': [snap_ts],
            'grades': grade_vals,
            'total': [ptotal],
        }
        pool_order.append(pool)
        risk_rated[pool] = True

    if not hist_bal_data:
        return {}

    # Reorder pool_order to follow config['pool_order'] (with config-listed pools
    # first in their declared order, then any extra pools alphabetically). This
    # ensures every TCT sheet that consults impaired['pool_order'] uses the
    # same canonical order as the rest of the report.
    cfg_order = config.get('pool_order', []) or []
    nrr = set(config.get('not_risk_rated', []) or [])
    order_idx = {name: i for i, name in enumerate(cfg_order)}
    fallback = len(cfg_order)
    rr_pools = [p for p in pool_order if p not in nrr]
    nrr_pools = [p for p in pool_order if p in nrr]
    rr_pools.sort(key=lambda p: (order_idx.get(p, fallback), str(p)))
    nrr_pools.sort(key=lambda p: (order_idx.get(p, fallback), str(p)))
    pool_order = rr_pools + nrr_pools

    return {
        'hist_bal_data': hist_bal_data,
        'pool_order': pool_order,
        'risk_rated': risk_rated,
    }


def _grade_pct_from_last_month(pdata):
    """Return grade percentage distribution from the most recent month with data.

    Looks backwards through the pool's date list for the last month where
    grades sum to a nonzero value.  Returns {grade_label: fraction} or {}
    if the pool has no grade data at all.
    """
    grades = pdata.get('grades', {})
    if not grades:
        return {}
    n = len(pdata['dates'])
    for i in range(n - 1, -1, -1):
        total = sum(vals[i] for vals in grades.values())
        if total > 0:
            return {g: vals[i] / total for g, vals in grades.items()}
    return {}


def extend_hist_bal_with_monthly(hist_bal_data, monthly_balances):
    """Extend hist_bal_data with pool-level monthly balance records.

    Adds any (pool, month-end) that isn't already in hist_bal_data — both
    AFTER the prior report's coverage (going-forward extension to current
    snapshot) and BEFORE it (back-filling years from the 5300 backfill in
    ``loan_code_history`` that landed in hist['monthly_balances'] but
    weren't in the prior TCT report when it was generated). Grade-level
    values are distributed proportionally using the most recent month's
    grade percentages from the prior report.
    """
    if monthly_balances is None or monthly_balances.empty:
        return

    # Pre-compute grade percentage distributions per pool (before adding
    # new months) so back-fill and forward-fill both use the prior
    # report's most-recent-month grade mix as the proxy.
    grade_pcts = {}
    for pool, pdata in hist_bal_data.items():
        grade_pcts[pool] = _grade_pct_from_last_month(pdata)

    # Pool name normalization map (handle trailing spaces, parentheticals)
    pool_norm = {}
    for p in hist_bal_data:
        pool_norm[p.strip().lower()] = p
        # Also handle parenthetical variants e.g. "Re-write(RW TM)" → "Re-write"
        clean = re.sub(r'\s*\(.*\)\s*$', '', p).strip()
        if clean.lower() != p.strip().lower():
            pool_norm[clean.lower()] = p

    for pool_key, grp in monthly_balances.groupby('pool'):
        pk = str(pool_key).strip()
        norm_key = pk.lower()
        mapped = pool_norm.get(norm_key)
        if not mapped:
            clean = re.sub(r'\s*\(.*\)\s*$', '', pk).strip()
            mapped = pool_norm.get(clean.lower())
        if not mapped:
            continue

        pdata = hist_bal_data[mapped]
        pcts = grade_pcts.get(mapped, {})
        existing_dates_set = set(
            pd.Timestamp(d) for d in pdata.get('dates', [])
        )

        added_any = False
        for _, row in grp.sort_values('date').iterrows():
            dt = pd.Timestamp(row['date']) + pd.offsets.MonthEnd(0)
            if dt in existing_dates_set:
                continue
            existing_dates_set.add(dt)
            pool_total = float(row['balance'])
            pdata['dates'].append(dt)
            pdata['total'].append(pool_total)
            for g, vals in pdata.get('grades', {}).items():
                vals.append(pool_total * pcts.get(g, 0.0))
            added_any = True

        if added_any:
            # Re-sort all arrays by date so chronological order is
            # preserved (back-filled months land before existing ones).
            order = sorted(
                range(len(pdata['dates'])),
                key=lambda i: pd.Timestamp(pdata['dates'][i]),
            )
            pdata['dates'] = [pdata['dates'][i] for i in order]
            pdata['total'] = [pdata['total'][i] for i in order]
            for g, vals in pdata.get('grades', {}).items():
                pdata['grades'][g] = [vals[i] for i in order]


def extend_hist_bal_with_db(hist_bal_data, df, snap, grades, config):
    """Extend hist_bal_data with new months from the current DB snapshot.

    Computes grade-level balances for each pool from `df` and appends them
    as new monthly columns *after* whatever the prior report already contains.
    Only adds months not yet present.
    """
    no_score = config.get('no_score_label', 'Not Reported')
    snap_ts = pd.Timestamp(snap)

    for pool, pdata in hist_bal_data.items():
        existing_dates = [pd.Timestamp(d) for d in pdata.get('dates', [])]
        pdf = df[df['loan_pool'] == pool]
        pgrades = pdata.get('grades', {})

        if snap_ts in existing_dates:
            # Already present (e.g. from monthly file) — update grade values in-place
            idx = existing_dates.index(snap_ts)
            if pgrades:
                for g in list(pgrades.keys()):
                    bal = pdf[pdf['current_grade'] == g]['current_balance'].sum()
                    pgrades[g][idx] = bal
            pdata['total'][idx] = pdf['current_balance'].sum()
            continue

        if pgrades:
            # Risk-rated: compute per-grade balances
            for g in list(pgrades.keys()):
                bal = pdf[pdf['current_grade'] == g]['current_balance'].sum()
                pgrades[g].append(bal)
        pool_total = pdf['current_balance'].sum()
        pdata['total'].append(pool_total)
        pdata['dates'].append(snap_ts)

    # Handle pools in DB that aren't in hist_bal_data yet (new pools)
    all_gl = [g['label'] for g in grades]
    all_gl.append(no_score)
    for pool in df['loan_pool'].unique():
        if pool in hist_bal_data:
            continue
        pdf = df[df['loan_pool'] == pool]
        pool_grades = {}
        for g in all_gl:
            pool_grades[g] = [pdf[pdf['current_grade'] == g]['current_balance'].sum()]
        hist_bal_data[pool] = {
            'dates': [snap_ts],
            'grades': pool_grades,
            'total': [pdf['current_balance'].sum()],
        }


def _apply_co_recovery_overrides(co_rec, config):
    """Force analyst-supplied summary CO/recovery values over the merged
    file/DB history for PAST months only (as-of month <= ``cutoff``).

    Future months are left untouched so ongoing monthly-file ingestion still
    governs new quarters. Config schema (``co_recovery_overrides``)::

        co_recovery_overrides:
          cutoff: 'YYYY-MM'        # inclusive upper bound; omit => no bound
          source: 'free text'      # provenance note (logged only)
          chargeoffs: {pool: {'YYYY-MM': amount, ...}, ...}
          recoveries: {pool: {'YYYY-MM': amount, ...}, ...}

    Only pools named in the override are recomputed; other pools (e.g. the
    MBL/CRE 5300 annual backfill) are left as-is.
    """
    ov = config.get('co_recovery_overrides')
    if not ov:
        return None
    cutoff = str(ov.get('cutoff') or '')
    n_cells = 0
    for annual_key, monthly_key, grid in (
        ('chargeoffs', 'co_monthly', ov.get('chargeoffs') or {}),
        ('recoveries', 'rc_monthly', ov.get('recoveries') or {}),
    ):
        monthly = co_rec.setdefault(monthly_key, {})
        annual = co_rec.setdefault(annual_key, {})
        touched = {}
        for pool, months in grid.items():
            for ymstr, amt in (months or {}).items():
                ym = str(ymstr)[:7]
                if cutoff and ym > cutoff:
                    continue
                try:
                    y, m = int(ym[:4]), int(ym[5:7])
                except (ValueError, IndexError):
                    continue
                monthly.setdefault((y, m), {})[pool] = float(amt or 0)
                touched.setdefault(pool, set()).add(y)
                n_cells += 1
        for pool, years in touched.items():
            for y in years:
                annual.setdefault(y, {})[pool] = sum(
                    bp.get(pool, 0.0)
                    for (yy, _mm), bp in monthly.items()
                    if yy == y)
    if n_cells:
        co_rec['years'] = sorted(set(co_rec.get('years', [])) | set(co_rec.get('chargeoffs', {})) | set(co_rec.get('recoveries', {})))
        print(f"    Applied {n_cells} CO/recovery override cell(s) (cutoff {cutoff or 'none'}; source: {ov.get('source', 'summary override')}).")
        return None


def load_historical_data(config):
    """Load all historical data for a client. Returns a dict with all historical DataFrames."""
    print("  Loading historical data...")
    co_rec = load_chargeoff_recovery_history(config)
    # Merge the wizard's DB tables (loan_code_chargeoff_history,
    # loan_code_recovery_history — which hold the NCUA 5300 backfill plus
    # any wizard-aggregated workbook rows) with the CU's own file-derived
    # history.
    #
    # SOURCE PRECEDENCE: the credit union's OWN charge-off / recovery files
    # win. The DB (5300) backfill only fills the YEARS the files don't cover
    # ("back the missing months with the 5300, but the credit union's files
    # should be used first"). Gating at the year level — rather than per
    # (year, pool) cell — keeps a single year's numbers from mixing the CU's
    # loan-code pool taxonomy with the NCUA 5300 taxonomy. When the CU ships
    # NO files (DB-only setups, e.g. monthly-summary or pure 5300 CUs), the
    # file result is empty, so every year falls through to the DB unchanged.
    db_corc = _load_co_rc_history_from_db(config)
    if db_corc.get('years'):
        # ``co_recovery_month_level_fill: true`` opts a CU into MONTH-level
        # precedence: the CU's own files win for every (year, month) they
        # cover and the DB fills only the months the files DON'T cover, with
        # annual totals recomputed from the merged monthly grid. This is for
        # CUs whose monthly files start partway through a year (so year-level
        # gating would drop the earlier months) AND whose DB history is
        # itself monthly-granular (a rebuilt monthly history), so there is no
        # quarterly-vs-monthly double count. The default remains YEAR-level
        # gating, which safely avoids mixing CU monthly data with quarterly
        # 5300 rows inside a single year.
        if config.get('co_recovery_month_level_fill'):
            # A (year, month) only counts as "covered by the CU's files" when
            # the file reports NON-ZERO data for it (see the year-level note
            # below). Cumulative recovery-tracking exports otherwise emit
            # all-zero month cells that would wrongly suppress the DB history.
            def _nonzero_keys(section):
                out = set()
                for k, bp in (section or {}).items():
                    if isinstance(bp, dict) and any(
                            abs(float(v or 0)) > 0.005 for v in bp.values()):
                        out.add(k)
                return out
            _file_co_months = _nonzero_keys(co_rec.get('co_monthly'))
            _file_rc_months = _nonzero_keys(co_rec.get('rc_monthly'))
            co_m = co_rec.setdefault('co_monthly', {})
            for ym, by_pool in db_corc['co_monthly'].items():
                if ym not in _file_co_months:
                    co_m[ym] = dict(by_pool)
            rc_m = co_rec.setdefault('rc_monthly', {})
            for ym, by_pool in db_corc['rc_monthly'].items():
                if ym not in _file_rc_months:
                    rc_m[ym] = dict(by_pool)

            def _annual_from_monthly(monthly):
                out: dict = {}
                for (yr, _mo), by_pool in monthly.items():
                    y = out.setdefault(yr, {})
                    for pool, amt in by_pool.items():
                        y[pool] = y.get(pool, 0.0) + amt
                return out

            co_rec['chargeoffs'] = _annual_from_monthly(co_m)
            co_rec['recoveries'] = _annual_from_monthly(rc_m)
            print(f"    CO/Recovery source precedence (MONTH-level): CU files "
                  f"win for {len(_file_co_months)} CO month(s) / "
                  f"{len(_file_rc_months)} recovery month(s); DB fills the rest.")
        else:
            # A year only counts as "covered by the CU's files" when the file
            # actually reports NON-ZERO charge-offs / recoveries for it. Some
            # CU charge-off/recovery exports are cumulative recovery-tracking
            # files that list previously charged-off loans with a BLANK
            # charge-off amount, so naive ``if bp`` truthiness marks every
            # historical year as covered (a non-empty dict of all-zero pools)
            # and wrongly suppresses the DB's real charge-off history.
            def _nonzero_years(section):
                out = set()
                for yr, bp in (section or {}).items():
                    if isinstance(bp, dict) and any(
                            abs(float(v or 0)) > 0.005 for v in bp.values()):
                        out.add(yr)
                return out

            _file_co_years = _nonzero_years(co_rec.get('chargeoffs'))
            _file_rc_years = _nonzero_years(co_rec.get('recoveries'))

            # ``co_recovery_db_fills_covered_years`` opts a CU into (year, pool)
            # CELL-level fill for the DB backfill: within a year the file DOES
            # report, the DB still fills the pools the file left absent / at
            # zero. This is for CUs whose CO/recovery export is a cumulative
            # recovery-tracking file that only carries a charge-off AMOUNT in
            # the period of charge-off, so a single stray amount would
            # otherwise let one pool's value stand in for the whole year. Only
            # safe when the DB history is in the CU's own pool taxonomy, so it
            # is guarded by a subset check and left OFF by default (whole-year
            # gating) to avoid mixing the CU taxonomy with NCUA 5300 categories.
            _cell_level = False
            if config.get('co_recovery_db_fills_covered_years'):
                _cfg_pools = set()
                for _p in (config.get('pool_order') or []):
                    if _p:
                        _cfg_pools.add(str(_p).strip().lower())
                for _p in (config.get('pools') or []):
                    _n = _p.get('name') if isinstance(_p, dict) else _p
                    if _n:
                        _cfg_pools.add(str(_n).strip().lower())
                _db_pools = set()
                for _sec in (db_corc.get('chargeoffs'), db_corc.get('recoveries')):
                    for _bp in (_sec or {}).values():
                        if isinstance(_bp, dict):
                            _db_pools.update(str(p).strip().lower() for p in _bp)
                _cell_level = bool(_cfg_pools) and _db_pools.issubset(_cfg_pools)

            def _fill(target, source_db, covered_years, year_key):
                for key, by_pool in (source_db or {}).items():
                    yr = key[0] if year_key else key
                    if yr not in covered_years:
                        target[key] = dict(by_pool)
                    elif _cell_level:
                        cur = target.setdefault(key, {})
                        for pool, amt in by_pool.items():
                            if amt and abs(float(cur.get(pool, 0) or 0)) <= 0.005:
                                cur[pool] = amt

            co = co_rec.setdefault('chargeoffs', {})
            _fill(co, db_corc.get('chargeoffs'), _file_co_years, False)
            rc = co_rec.setdefault('recoveries', {})
            _fill(rc, db_corc.get('recoveries'), _file_rc_years, False)
            co_m = co_rec.setdefault('co_monthly', {})
            _fill(co_m, db_corc.get('co_monthly'), _file_co_years, True)
            rc_m = co_rec.setdefault('rc_monthly', {})
            _fill(rc_m, db_corc.get('rc_monthly'), _file_rc_years, True)
            if _file_co_years or _file_rc_years:
                _mode = "cell-level" if _cell_level else "year-level"
                print(f"    CO/Recovery source precedence ({_mode}): CU files win "
                      f"for CO year(s) {sorted(_file_co_years)} and recovery "
                      f"year(s) {sorted(_file_rc_years)}; DB backfill fills the rest.")
        co_rec['years'] = sorted(
            set(co_rec.get('years', []))
            | set(co_rec['chargeoffs'])
            | set(co_rec['recoveries'])
        )
    _apply_co_recovery_overrides(co_rec, config)
    balances, alll_by_date = load_monthly_balances(config)
    dq = load_delinquency_history(config)

    # Drop balance-sheet line items that aren't loan pools (e.g. "ACH
    # Clearing", "ATM Machine", "Vizo Financial Corp ..."). The monthly
    # balance file is often a full balance sheet; we only want rows that
    # map to a configured loan pool. Anything else would pollute
    # hist_bal_data, avg_balances, and downstream Vizo/TCT tabs.
    if not balances.empty:
        nrr_set = set(config.get('not_risk_rated', []) or [])
        configured_pools = set(config.get('pool_order', []) or [])
        configured_pools.update(
            p.get('name') for p in (config.get('pools') or [])
            if p and p.get('name')
        )
        configured_pools.update(nrr_set)
        if configured_pools:
            configured_lc = {str(p).strip().lower(): str(p).strip()
                             for p in configured_pools if p}
            pool_norm = balances['pool'].astype(str).str.strip().str.lower()
            keep_mask = pool_norm.isin(configured_lc.keys())
            dropped = sorted(set(
                balances.loc[~keep_mask, 'pool'].astype(str).str.strip()
            ))
            if dropped:
                print(f"    Dropped {len(dropped)} non-loan balance-sheet "
                      f"line item(s) from monthly balances: "
                      f"{', '.join(dropped[:5])}"
                      f"{'...' if len(dropped) > 5 else ''}")
            balances = balances.loc[keep_mask].copy()
            # Canonicalize pool names to the configured spelling so downstream
            # exact-match lookups (e.g. set membership in pool_order) work.
            balances['pool'] = pool_norm[keep_mask].map(configured_lc)

    # Collapse same-pool/same-date rows by summing balances. When several
    # source labels (e.g. "Home Equity Fixed Rate Loans", "HELOC",
    # "Mortgage Loans", "Mortgage Portfolio", "Mobile Home Loans") all map
    # to a single configured pool (e.g. "Real Estate") via
    # cfg['monthly_balance']['pool_map'], the per_month loader emits one row
    # per source label. Without this collapse, avg_balances would average
    # sub-pool balances (instead of summing them) and
    # build_hist_bal_from_monthly would dedupe by date and silently drop
    # all but the first sub-pool's value, producing wildly understated
    # historical balances on the Display Hist Bal tab.
    if not balances.empty:
        before = len(balances)
        balances = (
            balances.groupby(['pool', 'date'], as_index=False)['balance']
            .sum()
            .sort_values(['pool', 'date'])
            .reset_index(drop=True)
        )
        after = len(balances)
        if after < before:
            print(f"    Collapsed {before - after} duplicate "
                  f"(pool, date) row(s) into {after} aggregated row(s) "
                  f"(multi-label pools summed)")

    # Extend monthly balances with rows from ``loan_code_history`` (5300
    # backfill, both per-loan-code and distributed modes). Adds (pool,
    # month-end) cells that aren't already present in the workbook —
    # the workbook remains authoritative for any month it covers. This
    # lets the supplemental Detail_HIst Balances tab show historical
    # columns back into the 5300-backfill years for pools whose
    # Life-of-Loan window exceeds the workbook's coverage.
    db_monthly = _load_balance_history_monthly_from_db(config)
    if not db_monthly.empty:
        if balances.empty:
            balances = db_monthly.sort_values(['pool', 'date']).reset_index(drop=True)
            print(f"    Seeded monthly balances from 5300 history: "
                  f"{len(balances)} row(s)")
        else:
            existing_keys = set(zip(balances['pool'].astype(str),
                                    balances['date']))
            mask = [(p, d) not in existing_keys
                    for p, d in zip(db_monthly['pool'].astype(str),
                                    db_monthly['date'])]
            new_rows = db_monthly.loc[mask].copy()
            if not new_rows.empty:
                balances = (
                    pd.concat([balances, new_rows], ignore_index=True)
                    .sort_values(['pool', 'date'])
                    .reset_index(drop=True)
                )
                print(f"    Extended monthly balances with {len(new_rows)} "
                      f"5300-history row(s) (pools/months not in workbook)")

    # Compute annual average balances per pool from monthly data
    avg_balances = {}  # {year: {pool: avg_balance}}
    if not balances.empty:
        balances['year'] = balances['date'].dt.year
        for (year, pool), grp in balances.groupby(['year', 'pool']):
            avg_balances.setdefault(int(year), {})
            avg_balances[int(year)][pool] = grp['balance'].mean()

    # Overlay annual balances from the wizard's 5300 backfill (DB table
    # ``loan_code_history``). Only fills (year, pool) cells NOT already
    # populated by the per_month / monthly_balances file — per_month
    # remains authoritative wherever it covers a year. This lets the
    # Vizo "Display Hist Bal" tab fill the early-year columns (e.g.
    # 2019-2022 for an 84-month Life-of-Loan Real Estate pool when the
    # per_month file only starts in 2023).
    db_bal = _load_balance_history_from_db(config)
    if db_bal:
        added_cells = 0
        added_years = set()
        for yr, by_pool in db_bal.items():
            existing = avg_balances.setdefault(yr, {})
            for pool, bal in by_pool.items():
                if not existing.get(pool):
                    existing[pool] = bal
                    added_cells += 1
                    added_years.add(yr)
        if added_cells:
            yrs = sorted(added_years)
            print(f"    Filled {added_cells} missing avg-balance cell(s) "
                  f"from 5300 backfill ({yrs[0]}-{yrs[-1]})")

    # Compute delinquency % per pool per year
    dq_pct = {}  # {year: {pool: dq_pct}}
    for qlabel, pools in dq.items():
        year = int(qlabel[:4])
        for pool, dq_bal in pools.items():
            # Get total balance for that pool at that time
            total = avg_balances.get(year, {}).get(pool, 0)
            if total > 0:
                pct = dq_bal / total
                dq_pct.setdefault(year, {})
                # Average across quarters within a year
                if pool in dq_pct[year]:
                    dq_pct[year][pool] = (dq_pct[year][pool] + pct) / 2
                else:
                    dq_pct[year][pool] = pct

    hist = {
        'chargeoffs': co_rec['chargeoffs'],
        'recoveries': co_rec['recoveries'],
        'years': co_rec['years'],
        'co_monthly': co_rec.get('co_monthly', {}),
        'rc_monthly': co_rec.get('rc_monthly', {}),
        'monthly_balances': balances,
        'avg_balances': avg_balances,
        'delinquency': dq,
        'dq_pct': dq_pct,
        'alll_by_date': alll_by_date,
    }

    # Extend hist['years'] to include any year covered only by the
    # balance-history overlay (5300 distributed backfill writes balance
    # rows but not CO/RC rows, so years 2018-2022 typically appear in
    # avg_balances but not in co_rec['years']). Without this, the
    # Display HIst Bal year axis would only show CO/RC-covered years.
    extra_years = {
        y for y in (avg_balances or {}).keys()
        if isinstance(y, int) and y not in set(hist['years'])
    }
    if extra_years:
        hist['years'] = sorted(set(hist['years']) | extra_years)

    # Print summary
    if co_rec['years']:
        print(f"    Charge-off/recovery years: {co_rec['years'][0]}-{co_rec['years'][-1]}")
        total_co = sum(sum(p.values()) for p in co_rec['chargeoffs'].values())
        total_rc = sum(sum(p.values()) for p in co_rec['recoveries'].values())
        print(f"    Total charge-offs: ${total_co:,.2f}  Recoveries: ${total_rc:,.2f}")
    if not balances.empty:
        print(f"    Monthly balance records: {len(balances)} ({balances['date'].min().strftime('%Y-%m')} to {balances['date'].max().strftime('%Y-%m')})")
    if dq:
        print(f"    Delinquency quarters: {len(dq)}")

    return hist


def _find_prior_tct_report(config, snap):
    """Return path to the most recent prior TCT report (snap_date < snap), or None."""
    cu = config['credit_union']
    safe_cu = cu.replace(' ', '_').replace('/', '-')
    rpt_dir = os.path.join(BASE, 'Reports')
    if not os.path.isdir(rpt_dir):
        return None
    pattern = re.compile(
        rf'(\d{{4}}-\d{{2}}-\d{{2}})_CECL_Migration_{re.escape(safe_cu)}_TCT_Model\.xlsx$',
        re.IGNORECASE,
    )
    candidates = []
    for root, dirs, files in os.walk(rpt_dir):
        for f in files:
            if f.startswith('~$'):
                continue
            m = pattern.match(f)
            if not m:
                continue
            d = m.group(1)
            # Allow equal-date matches only when the file lives in the
            # dedicated WARM-baseline subdir; same-date top-level files
            # are this run's own output.
            in_warm_baselines = '_warm_baselines' in os.path.normpath(
                os.path.relpath(root, rpt_dir)
            ).split(os.sep)
            if d < snap or (d == snap and in_warm_baselines):
                candidates.append((d, in_warm_baselines,
                                   os.path.join(root, f)))
    if not candidates:
        return None
    candidates.sort(key=lambda c: (c[0], c[1]), reverse=True)
    return candidates[0][2]


def _load_acl_months_from_tct(filepath):
    """Read WARM Months column from a prior TCT report's '> Detail_HIst Balances' sheet."""
    from openpyxl import load_workbook as _load_wb
    try:
        wb = _load_wb(filepath, read_only=True, data_only=True)
    except Exception:
        return {}
    if '> Detail_HIst Balances' not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb['> Detail_HIst Balances']
    rows = [list(r) for r in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                          max_col=ws.max_column, values_only=True)]
    wb.close()

    acl_months = {}
    r = 4
    while r < len(rows):
        a = rows[r][0] if rows[r] else None
        if a is None or str(a).strip() == '':
            r += 1
            continue
        pool_name = str(a).strip()
        if pool_name in ('Current Grade', 'Current Risk Rating', 'Total',
                         'Balance', '% of Loans',
                         'WARM\nMonths', 'Loss Factor Historical Detail'):
            r += 1
            continue
        if r + 1 >= len(rows):
            break
        next_a = rows[r + 1][0] if rows[r + 1] else None
        next_label = str(next_a).strip() if next_a else ''
        if next_label not in ('Current Grade', 'Current Risk Rating',
                              'Balance'):
            r += 1
            continue
        # Find WARM column in header row r+1
        hdr = rows[r + 1]
        warm_ci = None
        for ci, v in enumerate(hdr):
            if isinstance(v, str) and 'WARM' in v:
                warm_ci = ci
                break
        if warm_ci is not None:
            for gr in range(r + 2, min(r + 20, len(rows))):
                v = rows[gr][warm_ci] if warm_ci < len(rows[gr]) else None
                if v is not None:
                    try:
                        acl_months[pool_name] = int(v)
                    except (ValueError, TypeError):
                        pass
                    break
        r += 2
    return acl_months


def _find_prior_warm_xlsx(config, snap):
    """Return path to the most recent prior CECL-Migration-WARM xlsx (snap_prefix < snap).

    Searches data_directory and fallback_report_folder. Skips ~$ temp files,
    DNU prefixes, and any non-.xlsx files (e.g. PDFs).
    """
    data_dir = config.get('data_directory', '')
    if not data_dir:
        return None
    if not os.path.isabs(data_dir):
        data_dir = os.path.join(BASE, data_dir)

    cu = config['credit_union']
    snap_prefix = snap[:7] if snap else ''

    search_dirs = [data_dir]
    fb_folder = config.get('credit_pull', {}).get('fallback_report_folder', '')
    if fb_folder and fb_folder != data_dir:
        if not os.path.isabs(fb_folder):
            fb_folder = os.path.join(BASE, fb_folder)
        search_dirs.append(fb_folder)

    pattern = re.compile(r'^(\d{4}-\d{2})(?:-\d{2})?\s+CECL-Migration-WARM.*\.xlsx$',
                         re.IGNORECASE)
    candidates = []
    for sdir in search_dirs:
        if not os.path.isdir(sdir):
            continue
        for root, dirs, files in os.walk(sdir):
            for f in files:
                if f.startswith('~$') or f.upper().startswith('DNU'):
                    continue
                m = pattern.match(f)
                if not m:
                    continue
                pfx = m.group(1)
                if pfx < snap_prefix and cu.lower().split()[0] in f.lower():
                    candidates.append((pfx, os.path.join(root, f)))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


def _load_acl_months_from_warm_xlsx(filepath):
    """Read the BS CO DQ Data Enter tab and return {pool: acl_months_int}."""
    try:
        bs_df = pd.read_excel(filepath, sheet_name='BS CO DQ Data Enter', header=None)
    except (ValueError, KeyError, FileNotFoundError):
        return {}
    acl_months = {}
    if bs_df.shape[1] <= 6:
        return {}
    for idx in range(4, bs_df.shape[0]):
        pool_name = bs_df.iloc[idx, 0]
        months_val = bs_df.iloc[idx, 6]
        if pd.isna(pool_name) or str(pool_name).strip() == '':
            continue
        pn = str(pool_name).strip()
        if pn.upper().startswith(('HIDE', 'EXCLUDE', 'GRAND TOTAL')):
            continue
        if pd.isna(months_val):
            break
        try:
            acl_months[pn] = int(months_val)
        except (ValueError, TypeError):
            continue
    return acl_months


def _parse_display_co_recov_dq(found):
    """Parse the 'Display CO-Recov -DQ' tab and return a dict of warm_* keys.

    Reads per-year Charge offs / Recoveries / Net Charge offs / DQ % sections
    plus column-J totals (warm_co_totals, warm_rc_totals, warm_net_co).

    Returns dict with any of: warm_co, warm_rc, warm_net, warm_co_totals,
    warm_rc_totals, warm_dq_pct, warm_net_co. Returns {} if tab not present.

    Used by both load_impaired_data (legacy CECL-Migration-WARM file) and
    load_impaired_from_tct_baseline (Reports/_warm_baselines/*_TCT_Model.xlsx).
    """
    out = {}
    try:
        dq_display_df = pd.read_excel(found, sheet_name='Display CO-Recov -DQ',
                                      header=None)
    except (ValueError, KeyError, FileNotFoundError):
        return out

    # Row 3 (0-indexed 2) has year numbers in cols 1..N
    dq_years = []
    for c in range(1, dq_display_df.shape[1]):
        v = dq_display_df.iloc[2, c]
        if pd.notna(v):
            try:
                dq_years.append(int(v))
            except (ValueError, TypeError):
                break
        else:
            break

    # Find the DQ % section header row
    dq_start = None
    for idx in range(len(dq_display_df)):
        val = dq_display_df.iloc[idx, 0]
        if pd.notna(val) and str(val).strip().upper().startswith('DQ'):
            dq_start = idx + 1
            break

    def _parse_section(section_label, exact_start=False):
        """Return {year: {pool: value}} for a section identified by label."""
        sec_start = None
        for idx2 in range(len(dq_display_df)):
            val2 = dq_display_df.iloc[idx2, 0]
            if pd.notna(val2):
                cell_text = str(val2).strip().lower()
                if exact_start:
                    if cell_text.startswith(section_label.lower()):
                        sec_start = idx2 + 1
                        break
                else:
                    if section_label.lower() in cell_text:
                        sec_start = idx2 + 1
                        break
        if sec_start is None or not dq_years:
            return {}
        result = {}
        for idx2 in range(sec_start, min(sec_start + 30, len(dq_display_df))):
            pool_name = dq_display_df.iloc[idx2, 0]
            if pd.isna(pool_name) or str(pool_name).strip() == '':
                break
            pn = str(pool_name).strip()
            if pn.upper().startswith(('HIDE', 'EXCLUDE')):
                continue
            for ci, yr in enumerate(dq_years):
                col_idx = 1 + ci
                if col_idx >= dq_display_df.shape[1]:
                    break
                v = dq_display_df.iloc[idx2, col_idx]
                if pd.notna(v):
                    try:
                        result.setdefault(yr, {})[pn] = float(v)
                    except (ValueError, TypeError):
                        pass
        return result

    def _parse_section_totals(section_label, exact_start=False):
        """Return {pool: acl_total} for a section's total column (col J)."""
        sec_start = None
        for idx2 in range(len(dq_display_df)):
            val2 = dq_display_df.iloc[idx2, 0]
            if pd.notna(val2):
                cell_text = str(val2).strip().lower()
                if exact_start:
                    if cell_text.startswith(section_label.lower()):
                        sec_start = idx2 + 1
                        break
                else:
                    if section_label.lower() in cell_text:
                        sec_start = idx2 + 1
                        break
        if sec_start is None:
            return {}
        total_col = 9
        result = {}
        for idx2 in range(sec_start, min(sec_start + 30, len(dq_display_df))):
            pool_name = dq_display_df.iloc[idx2, 0]
            if pd.isna(pool_name) or str(pool_name).strip() == '':
                break
            pn = str(pool_name).strip()
            if pn.upper().startswith(('HIDE', 'EXCLUDE')):
                continue
            v = (dq_display_df.iloc[idx2, total_col]
                 if dq_display_df.shape[1] > total_col else None)
            if pd.notna(v):
                try:
                    result[pn] = float(v)
                except (ValueError, TypeError):
                    pass
        return result

    warm_co = _parse_section('charge offs', exact_start=True)
    warm_rc = _parse_section('recoveries', exact_start=True)
    warm_net = _parse_section('net charge offs')
    warm_co_totals = _parse_section_totals('charge offs', exact_start=True)
    warm_rc_totals = _parse_section_totals('recoveries', exact_start=True)

    if warm_co:
        out['warm_co'] = warm_co
        out['warm_rc'] = warm_rc
        out['warm_net'] = warm_net
        out['warm_co_totals'] = warm_co_totals
        out['warm_rc_totals'] = warm_rc_totals
        print(f"    WARM CO/RC data: {len(warm_co)} years, "
              f"CO pool-years: {sum(len(v) for v in warm_co.values())}, "
              f"RC pool-years: {sum(len(v) for v in warm_rc.values())}")

    if dq_start and dq_years:
        warm_dq_pct = {}
        for idx in range(dq_start, min(dq_start + 30, len(dq_display_df))):
            pool_name = dq_display_df.iloc[idx, 0]
            if pd.isna(pool_name) or str(pool_name).strip() == '':
                break
            pn = str(pool_name).strip()
            if pn.upper().startswith(('HIDE', 'EXCLUDE')):
                continue
            for ci, yr in enumerate(dq_years):
                col_idx = 1 + ci
                if col_idx >= dq_display_df.shape[1]:
                    break
                v = dq_display_df.iloc[idx, col_idx]
                if pd.notna(v):
                    try:
                        warm_dq_pct.setdefault(yr, {})[pn] = float(v)
                    except (ValueError, TypeError):
                        pass
        if warm_dq_pct:
            out['warm_dq_pct'] = warm_dq_pct
            print(f"    WARM DQ% data: {len(warm_dq_pct)} years, "
                  f"{sum(len(v) for v in warm_dq_pct.values())} pool-year entries")

    # Net Chargeoff totals per pool from CO-Recov-DQ (col J total)
    net_co_start = None
    for idx in range(len(dq_display_df)):
        val = dq_display_df.iloc[idx, 0]
        if (pd.notna(val) and 'net' in str(val).strip().lower()
                and 'charge' in str(val).strip().lower()):
            net_co_start = idx + 1
            break
    if net_co_start:
        warm_net_co = {}
        total_col = 9
        for idx in range(net_co_start, min(net_co_start + 30, len(dq_display_df))):
            pool_name = dq_display_df.iloc[idx, 0]
            if pd.isna(pool_name) or str(pool_name).strip() == '':
                break
            pn = str(pool_name).strip()
            if pn.upper().startswith(('HIDE', 'EXCLUDE')):
                continue
            v = (dq_display_df.iloc[idx, total_col]
                 if dq_display_df.shape[1] > total_col else None)
            if pd.notna(v):
                try:
                    warm_net_co[pn] = float(v)
                except (ValueError, TypeError):
                    pass
        if warm_net_co:
            out['warm_net_co'] = warm_net_co
            print(f"    WARM Net CO data: {len(warm_net_co)} pools")

    return out


def load_impaired_data(config, snap):
    """Load impaired-loan summary from the existing CECL-Migration-WARM working file.

    Reads the 'Impaired Loans' tab, column L (category) and P (Sum of Provision Amount)
    from the summary pivot at rows 5-10 (skipping rows labelled 'HIDE').

    Returns dict with:
      'items': {category: provision_amount, ...},
      'total_spec_id': float  (sum of all provision amounts)
    or empty dict if file/tab not found.
    """
    data_dir = config.get('data_directory', '')
    if not data_dir:
        return {}
    # Resolve data_dir (may be absolute or relative)
    if not os.path.isabs(data_dir):
        data_dir = os.path.join(BASE, data_dir)

    cu = config['credit_union']
    safe_cu = cu.replace(' ', '_').replace('/', '-')
    # Build expected filename:  e.g. "2025-12 CECL-Migration-WARM - Franklin Trust FCU.xlsx"
    # snap is like "2025-12-31" — extract YYYY-MM
    snap_prefix = snap[:7] if snap else ''

    # Search for the file in data_dir, then fallback_report_folder
    # NOTE: fallback_report_folder is often a SHARED temp dir holding WARM
    # uploads for multiple credit unions (e.g. cecl_ui_warm). All match paths
    # below MUST verify the filename contains *this* CU's name so we don't
    # accidentally load another CU's WARM workbook. Without this guard,
    # `2025-12 CECL-Migration-WARM - Bridgeton Onized FCU.xlsx` would be
    # loaded as Utah Community FCU's WARM (alphabetical first-match win).
    target_name = f"{snap_prefix} CECL-Migration-WARM - {cu}.xlsx"
    # Accept both space- and underscore-separated filenames (the wizard
    # rewrites spaces to underscores on save).
    target_name_alt = target_name.replace(' ', '_')
    search_dirs = [data_dir]
    fb_folder = config.get('credit_pull', {}).get('fallback_report_folder', '')
    if fb_folder and fb_folder != data_dir:
        if not os.path.isabs(fb_folder):
            fb_folder = os.path.join(BASE, fb_folder)
        search_dirs.append(fb_folder)

    def _cu_in_filename(fname: str) -> bool:
        """Return True iff *fname* references this CU (space/underscore tolerant)."""
        if not cu:
            return True
        norm_f = fname.lower().replace(' ', '_')
        # Full CU name with underscores
        if safe_cu.lower() in norm_f:
            return True
        # Allow first-token fallback (e.g. "Utah" matches "Utah_Community_FCU")
        first = cu.lower().split()[0] if cu.strip() else ''
        if first and len(first) >= 4 and first in norm_f:
            return True
        return False

    found = None
    for sdir in search_dirs:
        if not os.path.isdir(sdir):
            continue
        for root, dirs, files in os.walk(sdir):
            for f in files:
                if f.startswith('~$') or f.upper().startswith('DNU'):
                    continue
                if f == target_name or f == target_name_alt:
                    found = os.path.join(root, f)
                    break
            if found:
                break
        if found:
            break

    # Fallback: search by pattern, but REQUIRE this CU's name in the filename
    # so a shared fallback_report_folder doesn't pull in another CU's WARM.
    if not found:
        pattern = re.compile(rf'^{re.escape(snap_prefix)}.*CECL-Migration-WARM.*\.xlsx$', re.IGNORECASE)
        for sdir in search_dirs:
            if not os.path.isdir(sdir):
                continue
            for root, dirs, files in os.walk(sdir):
                for f in files:
                    if f.startswith('~$') or f.upper().startswith('DNU'):
                        continue
                    if pattern.match(f) and _cu_in_filename(f):
                        found = os.path.join(root, f)
                        break
                if found:
                    break
            if found:
                break

    if not found:
        print(f"    No CECL-Migration-WARM file found for {snap_prefix}")
        return {}

    print(f"    Loading impaired loan data from: {os.path.basename(found)}")
    try:
        imp_df = pd.read_excel(found, sheet_name='Impaired Loans', header=None)
    except (ValueError, KeyError):
        print(f"    'Impaired Loans' tab not found in {os.path.basename(found)}")
        return {}

    # The summary pivot is at rows 4-10 (0-indexed: 3-9)
    # Column L = index 11, Column P = index 15
    # Row 20 has the "Total" row; stop before it
    items = {}
    total = 0.0
    for idx in range(4, min(30, len(imp_df))):  # rows 5-30 (1-indexed)
        cat = imp_df.iloc[idx, 11] if imp_df.shape[1] > 11 else None
        prov = imp_df.iloc[idx, 15] if imp_df.shape[1] > 15 else 0
        if cat is None or pd.isna(cat) or str(cat).strip() == '':
            continue
        cat_str = str(cat).strip()
        if cat_str.upper() in ('HIDE', 'TOTAL', 'CALCULATION'):
            continue
        prov_val = 0.0
        try:
            prov_val = float(prov) if pd.notna(prov) else 0.0
        except (ValueError, TypeError):
            continue  # skip non-numeric rows
        items[cat_str] = prov_val
        total += prov_val

    print(f"    Impaired loan categories: {len(items)}, Total: ${total:,.2f}")
    result = {'items': items, 'total_spec_id': total}

    # ── Extract per-pool per-grade "Balance Removed" from detail rows ──
    # Detail rows start after the "Data Entry" marker row.
    # Col Q (16) = Balance Removed, Col R (17) = Loan Pool, Col S (18) = Credit Grade
    spec_id_by_pool = {}   # {pool: {grade: balance_removed, ...}, ...}
    for idx in range(30, len(imp_df)):
        a_val = imp_df.iloc[idx, 0] if imp_df.shape[1] > 0 else None
        if a_val is not None and not pd.isna(a_val):
            lbl = str(a_val).strip()
            if lbl in ('Data Entry', 'Impairment Type'):
                continue
            # This should be a detail row with impairment type in col A
            bal_removed = imp_df.iloc[idx, 16] if imp_df.shape[1] > 16 else 0
            pool_name = imp_df.iloc[idx, 17] if imp_df.shape[1] > 17 else None
            grade = imp_df.iloc[idx, 18] if imp_df.shape[1] > 18 else None
            if pool_name is None or pd.isna(pool_name):
                continue
            pool_str = str(pool_name).strip()
            grade_str = str(grade).strip() if grade is not None and not pd.isna(grade) else ''
            try:
                bal_val = float(bal_removed) if pd.notna(bal_removed) else 0.0
            except (ValueError, TypeError):
                bal_val = 0.0
            if bal_val > 0 and pool_str:
                if pool_str not in spec_id_by_pool:
                    spec_id_by_pool[pool_str] = {}
                spec_id_by_pool[pool_str][grade_str] = (
                    spec_id_by_pool[pool_str].get(grade_str, 0) + bal_val
                )
    if spec_id_by_pool:
        result['spec_id_by_pool'] = spec_id_by_pool
        n_pools = len(spec_id_by_pool)
        total_removed = sum(sum(g.values()) for g in spec_id_by_pool.values())
        print(f"    Specific ID by pool: {n_pools} pools, "
              f"Total removed: ${total_removed:,.2f}")

    # Load Improved/Deteriorated by grade from "Executive Summary (3)" tab
    try:
        es3_df = pd.read_excel(found, sheet_name='Executive Summary (3)', header=None)
        imp_grades = {}   # {grade: balance}
        det_grades = {}   # {grade: balance}

        def _find_section_grades(df, section_keyword):
            """Dynamically find grade/balance rows after a section header."""
            grades = {}
            # Search col C (idx 2) for section header containing the keyword
            start_idx = None
            for idx in range(len(df)):
                val = df.iloc[idx, 2] if df.shape[1] > 2 else None
                if val is None or pd.isna(val):
                    # Also check col B (idx 1) for merged header cells
                    val = df.iloc[idx, 1] if df.shape[1] > 1 else None
                if val is not None and not pd.isna(val):
                    s = str(val).strip().lower()
                    if section_keyword in s and 'summary' in s:
                        start_idx = idx
                        break
            if start_idx is None:
                return grades
            # Find the "Grade" header row after the section header
            for idx in range(start_idx, min(start_idx + 5, len(df))):
                val = df.iloc[idx, 2] if df.shape[1] > 2 else None
                if val is not None and not pd.isna(val) and str(val).strip().lower() == 'grade':
                    # Read grade rows starting from next row
                    for gi in range(idx + 1, min(idx + 15, len(df))):
                        g_val = df.iloc[gi, 2] if df.shape[1] > 2 else None
                        b_val = df.iloc[gi, 3] if df.shape[1] > 3 else 0
                        if g_val is None or pd.isna(g_val):
                            continue
                        g = str(g_val).strip()
                        if g.lower().startswith('total'):
                            break
                        if g.lower().startswith('hide'):
                            continue
                        try:
                            grades[g] = float(b_val) if pd.notna(b_val) else 0.0
                        except (ValueError, TypeError):
                            grades[g] = 0.0
                    break
            return grades

        imp_grades = _find_section_grades(es3_df, 'improved')
        det_grades = _find_section_grades(es3_df, 'deteriorated')
        result['exec_summary_3'] = {'improved': imp_grades, 'deteriorated': det_grades}
        print(f"    Executive Summary (3): {len(imp_grades)} improved grades, {len(det_grades)} deteriorated grades")
    except (ValueError, KeyError):
        print(f"    'Executive Summary (3)' tab not found in {os.path.basename(found)}")

    # Also read Pooled Totals and ACL Balance from "ACL Env by Pool Mgmt Adj" tab
    try:
        acl_df = pd.read_excel(found, sheet_name='ACL Env by Pool Mgmt Adj', header=None)
    except (ValueError, KeyError):
        # Don't bail: a WARM can carry 'DQ Data Entry'/'CO Data Entry' tabs
        # without an ACL Env tab.  Continue with an empty frame so the ACL
        # parsing below no-ops and the DQ/CO migration blocks still run.
        # See docs/pdf_migration/04_blank_charts.md.
        print(f"    'ACL Env by Pool Mgmt Adj' tab not found")
        acl_df = pd.DataFrame()

    # Search column A for key labels, read value from column K (index 10)
    for idx in range(len(acl_df)):
        label = acl_df.iloc[idx, 0]  # column A
        if pd.isna(label):
            continue
        lbl = str(label).strip()
        k_val = acl_df.iloc[idx, 10] if acl_df.shape[1] > 10 else None  # column K
        if lbl.startswith('Pooled Totals'):
            try:
                result['pooled_total_allowance'] = float(k_val) if pd.notna(k_val) else 0.0
                print(f"    Pooled Total Allowance: ${result['pooled_total_allowance']:,.2f}")
            except (ValueError, TypeError):
                pass
        elif lbl.startswith('Allowance for Credit Loss Balance'):
            try:
                result['acl_balance'] = float(k_val) if pd.notna(k_val) else 0.0
                print(f"    ACL Balance: ${result['acl_balance']:,.2f}")
            except (ValueError, TypeError):
                pass

    # ── Parse the full per-pool per-grade ACL data from the same tab ──
    acl_pools = {}   # {pool_name: {'grades': {grade: {balance, spec_id, calc_bal, base_rate, mgmt_adj, factor, allow_before}}, 'total': {...}}}
    acl_impaired = {}  # {label: allowance}
    acl_summary = {}   # pooled_total_spec_id, total_spec_allow, total_allow_needed, acl_bal, adjustment
    current_pool = None
    current_grades = {}
    in_impaired = False   # inside the 'Impaired Loans' breakout section
    for idx in range(len(acl_df)):
        a_val = acl_df.iloc[idx, 0]
        if pd.isna(a_val):
            continue
        label = str(a_val).strip()

        # Pool header row: next row has "Current Grade" or "Current Risk Rating" header
        if idx + 1 < len(acl_df):
            next_a = acl_df.iloc[idx + 1, 0]
            if pd.notna(next_a) and str(next_a).strip() in (
                'Current Grade', 'Current Risk Rating'
            ):
                # Save previous pool
                if current_pool and current_grades:
                    acl_pools[current_pool]['grades'] = current_grades
                current_pool = label
                acl_pools[current_pool] = {'grades': {}, 'total': {}}
                current_grades = {}
                continue

        # Grade data row (inside a pool block): A=grade, B=balance, ...
        if current_pool and label not in ('Current Grade',
                                           'Current Risk Rating', 'Total'):
            b = acl_df.iloc[idx, 1] if acl_df.shape[1] > 1 else 0
            c = acl_df.iloc[idx, 2] if acl_df.shape[1] > 2 else 0
            d = acl_df.iloc[idx, 3] if acl_df.shape[1] > 3 else 0
            e = acl_df.iloc[idx, 4] if acl_df.shape[1] > 4 else 0
            f = acl_df.iloc[idx, 5] if acl_df.shape[1] > 5 else 0
            g = acl_df.iloc[idx, 6] if acl_df.shape[1] > 6 else 0
            h = acl_df.iloc[idx, 7] if acl_df.shape[1] > 7 else 0
            current_grades[label] = {
                'balance': float(b) if pd.notna(b) else 0.0,
                'spec_id': float(c) if pd.notna(c) else 0.0,
                'calc_bal': float(d) if pd.notna(d) else 0.0,
                'base_rate': float(e) if pd.notna(e) else 0.0,
                'mgmt_adj': float(f) if pd.notna(f) else 0.0,
                'factor': float(g) if pd.notna(g) else 0.0,
                'allow_before': float(h) if pd.notna(h) else 0.0,
            }

        # Total row for current pool
        if current_pool and label == 'Total':
            e = acl_df.iloc[idx, 4] if acl_df.shape[1] > 4 else 0
            f_val = acl_df.iloc[idx, 5] if acl_df.shape[1] > 5 else 0
            g_val = acl_df.iloc[idx, 6] if acl_df.shape[1] > 6 else 0
            h = acl_df.iloc[idx, 7] if acl_df.shape[1] > 7 else 0
            i_val = acl_df.iloc[idx, 8] if acl_df.shape[1] > 8 else 0
            j = acl_df.iloc[idx, 9] if acl_df.shape[1] > 9 else 0
            k = acl_df.iloc[idx, 10] if acl_df.shape[1] > 10 else 0
            acl_pools[current_pool]['total'] = {
                'balance': float(acl_df.iloc[idx, 1]) if pd.notna(acl_df.iloc[idx, 1]) else 0.0,
                'spec_id': float(acl_df.iloc[idx, 2]) if pd.notna(acl_df.iloc[idx, 2]) else 0.0,
                'base_rate': float(e) if pd.notna(e) else 0.0,
                'mgmt_adj': float(f_val) if pd.notna(f_val) else 0.0,
                'factor': float(g_val) if pd.notna(g_val) else 0.0,
                'allow_before': float(h) if pd.notna(h) else 0.0,
                'env_factor': float(i_val) if pd.notna(i_val) else 0.0,
                'env_allow': float(j) if pd.notna(j) else 0.0,
                'total_allow': float(k) if pd.notna(k) else 0.0,
            }
            if current_grades:
                acl_pools[current_pool]['grades'] = current_grades
            current_grades = {}
            current_pool = None

        # Pooled Totals row
        if label.startswith('Pooled Totals'):
            acl_summary['pooled_balance'] = float(acl_df.iloc[idx, 1]) if pd.notna(acl_df.iloc[idx, 1]) else 0.0
            acl_summary['pooled_spec_id'] = float(acl_df.iloc[idx, 2]) if pd.notna(acl_df.iloc[idx, 2]) else 0.0
            acl_summary['pooled_allow_before'] = float(acl_df.iloc[idx, 7]) if pd.notna(acl_df.iloc[idx, 7]) else 0.0
            acl_summary['pooled_env_allow'] = float(acl_df.iloc[idx, 9]) if pd.notna(acl_df.iloc[idx, 9]) else 0.0
            acl_summary['pooled_total_allow'] = float(acl_df.iloc[idx, 10]) if pd.notna(acl_df.iloc[idx, 10]) else 0.0

        # Impaired Loans section. The breakout categories vary by CU:
        # standard TCT names (Delinquent Loans / Known Losses / ...) or the
        # CU's own codes (e.g. Erie: DQ / DQ90 / REPO / BK / BKNOTDQ). They
        # render under an 'Impaired Loans' header with the per-category
        # allowance in the 'Allowance' column (index 9; some layouts use
        # index 10). Track the section so pool / other-provision rows aren't
        # mis-captured, and read whichever allowance column is populated.
        if label == 'Impaired Loans':
            in_impaired = True
        elif (label.startswith('Total Specifically Identified')
              or label.startswith('Pooled Totals')
              or label.startswith('Total Other Provision')):
            in_impaired = False
        elif in_impaired and label and not label.upper().startswith('HIDE') \
                and label.lower() not in ('amount at risk', 'allowance',
                                          'allowance %'):
            _v9 = acl_df.iloc[idx, 9] if acl_df.shape[1] > 9 else None
            _v10 = acl_df.iloc[idx, 10] if acl_df.shape[1] > 10 else None
            _v = _v9 if pd.notna(_v9) else _v10
            if pd.notna(_v):
                acl_impaired[label] = float(_v)

        if label == 'Total Specifically Identified Allowance':
            k = acl_df.iloc[idx, 10] if acl_df.shape[1] > 10 else 0
            acl_summary['total_spec_allow'] = float(k) if pd.notna(k) else 0.0
        if label == 'Total Allowance Needed':
            k = acl_df.iloc[idx, 10] if acl_df.shape[1] > 10 else 0
            acl_summary['total_allow_needed'] = float(k) if pd.notna(k) else 0.0
        if label.startswith('Allowance for Credit Loss Balance'):
            k = acl_df.iloc[idx, 10] if acl_df.shape[1] > 10 else 0
            acl_summary['acl_balance'] = float(k) if pd.notna(k) else 0.0
        if label.startswith('Adjustment'):
            k = acl_df.iloc[idx, 10] if acl_df.shape[1] > 10 else 0
            acl_summary['adjustment'] = float(k) if pd.notna(k) else 0.0

    # Filter out HIDE pools and Exclude
    acl_pools = {k: v for k, v in acl_pools.items()
                 if not k.upper().startswith('HIDE') and k != 'Exclude'}
    result['acl_pools'] = acl_pools
    result['pool_order'] = list(acl_pools.keys())
    result['acl_impaired'] = acl_impaired
    result['acl_summary'] = acl_summary
    print(f"    ACL per-pool data: {len(acl_pools)} pools, "
          f"{len(acl_impaired)} impaired categories")

    # Helper: read ALL pool blocks from a DQ/CO Data Entry tab.
    # Each block has a "Loan Status" header in col P with pool name in col A.
    # Returns (grand_total_dict, per_pool_dict).
    def _read_migration_blocks(sheet_df):
        by_pool = {}
        grand_total = {}
        for idx in range(len(sheet_df)):
            cell_p = sheet_df.iloc[idx, 15] if sheet_df.shape[1] > 15 else None
            if not (pd.notna(cell_p) and str(cell_p).strip() == 'Loan Status'):
                continue
            pool_raw = sheet_df.iloc[idx, 0] if pd.notna(sheet_df.iloc[idx, 0]) else ''
            pool_key = str(pool_raw).strip()
            block = {}
            for di in range(1, 5):
                ri = idx + di
                if ri >= len(sheet_df):
                    break
                status = sheet_df.iloc[ri, 15] if sheet_df.shape[1] > 15 else None
                balance = sheet_df.iloc[ri, 16] if sheet_df.shape[1] > 16 else 0
                pct = sheet_df.iloc[ri, 17] if sheet_df.shape[1] > 17 else 0
                if pd.notna(status):
                    s = str(status).strip()
                    try:
                        block[s] = {
                            'balance': float(balance) if pd.notna(balance) else 0.0,
                            'pct': float(pct) if pd.notna(pct) else 0.0,
                        }
                    except (ValueError, TypeError):
                        pass
            if not block:
                continue
            if pool_key.lower().startswith('grand total'):
                grand_total = block
            elif pool_key.lower().startswith(('hide', 'exclude', 'credit grade', 'risk rated')):
                continue
            else:
                by_pool[pool_key] = block
        return grand_total, by_pool

    # Load DQ by migration status from "DQ Data Entry" tab (all pools + grand total)
    try:
        dq_df = pd.read_excel(found, sheet_name='DQ Data Entry', header=None)
        dq_grand, dq_by_pool = _read_migration_blocks(dq_df)
        if dq_grand:
            result['dq_by_status'] = dq_grand
            total_dq = sum(v['balance'] for v in dq_grand.values())
            print(f"    DQ by migration status: {len(dq_grand)} categories, Total DQ: ${total_dq:,.2f}")
        if dq_by_pool:
            result['dq_by_pool'] = dq_by_pool
            print(f"    DQ per-pool data: {len(dq_by_pool)} pools")
    except (ValueError, KeyError):
        print(f"    'DQ Data Entry' tab not found in {os.path.basename(found)}")

    # Load CO by migration status from "CO Data Entry" tab (all pools + grand total)
    try:
        co_df = pd.read_excel(found, sheet_name='CO Data Entry', header=None)
        co_grand, co_by_pool = _read_migration_blocks(co_df)
        if co_grand:
            result['co_by_status'] = co_grand
            total_co = sum(v['balance'] for v in co_grand.values())
            print(f"    CO by migration status: {len(co_grand)} categories, Total CO: ${total_co:,.2f}")
        if co_by_pool:
            result['co_by_pool'] = co_by_pool
            print(f"    CO per-pool data: {len(co_by_pool)} pools")
    except (ValueError, KeyError):
        print(f"    'CO Data Entry' tab not found in {os.path.basename(found)}")

    # ── Economic Stress Data from "BS CO DQ Data Enter" tab ──
    try:
        bs_df = pd.read_excel(found, sheet_name='BS CO DQ Data Enter', header=None)
        # Row 5 (0-indexed) has headers: L=State, M=County, N=Unemp%, O=FC/Person, P=BK, Q=Population
        # Row 6 (0-indexed) has values
        if bs_df.shape[0] > 6 and bs_df.shape[1] > 16:
            state = bs_df.iloc[6, 11] if pd.notna(bs_df.iloc[6, 11]) else ''
            county = bs_df.iloc[6, 12] if pd.notna(bs_df.iloc[6, 12]) else ''
            unemp = float(bs_df.iloc[6, 13]) if pd.notna(bs_df.iloc[6, 13]) else 0.0
            fc = bs_df.iloc[6, 14] if pd.notna(bs_df.iloc[6, 14]) else 0
            bk = bs_df.iloc[6, 15] if pd.notna(bs_df.iloc[6, 15]) else 0
            pop = bs_df.iloc[6, 16] if pd.notna(bs_df.iloc[6, 16]) else 1
            try:
                fc = int(fc)
            except (ValueError, TypeError):
                fc = 0
            try:
                bk = int(bk)
            except (ValueError, TypeError):
                bk = 0
            try:
                pop = int(pop)
            except (ValueError, TypeError):
                pop = 1
            result['economic_data'] = {
                'state': str(state).strip(),
                'county': str(county).strip(),
                'unemployment_rate': unemp,
                'foreclosures': fc,
                'bankruptcies': bk,
                'population': pop,
            }
            print(f"    Economic stress data: {result['economic_data']['state']}, "
                  f"{result['economic_data']['county']}")

        # ── Risk Rated flag per pool (col B = "Risk Rated Yes/No") ──
        risk_rated = {}
        for idx in range(4, bs_df.shape[0]):
            pool_name = bs_df.iloc[idx, 0]
            rr_val = bs_df.iloc[idx, 1] if bs_df.shape[1] > 1 else None
            if pd.isna(pool_name) or str(pool_name).strip() == '':
                continue
            pn = str(pool_name).strip()
            if pn.upper().startswith(('HIDE', 'EXCLUDE', 'GRAND TOTAL')):
                continue
            risk_rated[pn] = str(rr_val).strip().lower() == 'yes' if pd.notna(rr_val) else True
        if risk_rated:
            result['risk_rated'] = risk_rated
            nr_count = sum(1 for v in risk_rated.values() if not v)
            print(f"    Risk rated flags: {len(risk_rated)} pools ({nr_count} not risk-rated)")

        # ── ACL Months (Life of Loan) per pool ──
        # Row 3 (0-indexed) has header with "ACL Months" at col 6
        # Rows 4+ have pool name (col 0) and ACL months (col 6)
        # Stop at 'Exclude' or 'Grand Total' sentinel rows
        acl_months = {}
        if bs_df.shape[1] > 6:
            for idx in range(4, bs_df.shape[0]):
                pool_name = bs_df.iloc[idx, 0]
                months_val = bs_df.iloc[idx, 6]
                if pd.isna(pool_name) or str(pool_name).strip() == '':
                    continue
                pn = str(pool_name).strip()
                if pn.upper().startswith(('HIDE', 'EXCLUDE', 'GRAND TOTAL')):
                    continue
                if pd.isna(months_val):
                    break  # reached end of ACL months section
                try:
                    acl_months[pn] = int(months_val)
                except (ValueError, TypeError):
                    acl_months[pn] = 36
        if acl_months:
            result['acl_months'] = acl_months
            print(f"    ACL months (life of loan): {len(acl_months)} pools")
    except (ValueError, KeyError):
        print(f"    'BS CO DQ Data Enter' tab not found in {os.path.basename(found)}")

    # ── Environmental Factor Ranges from "Envir Fact Ranges" tab ──
    try:
        ef_df = pd.read_excel(found, sheet_name='Envir Fact Ranges', header=None)
        # Row 6 (0-indexed) is header row: col 1=Range, 2=Score, 3=Range, 4=Score, 5=Range, 6=Score
        # NCC ranges: cols 1-2 starting row 7, DQ: cols 3-4, ES: cols 5-6

        def _parse_range_label(lbl):
            """Parse range labels like '>5.00%', '4.00% to 4.99%', '<-5.00%'."""
            lbl = lbl.replace('%', '').strip()
            if lbl.startswith('>') or lbl.startswith('>='):
                val = float(lbl.lstrip('>= '))
                return (val, 999)
            elif lbl.startswith('<') or lbl.startswith('<='):
                val = float(lbl.lstrip('<= '))
                return (-999, val)
            elif ' to ' in lbl:
                parts = lbl.split(' to ')
                return (float(parts[0].strip()), float(parts[1].strip()) + 0.01)
            return (0, 0)

        def _read_range_col(col_lbl, col_sc):
            """Read label+score columns and return list of (lo, hi, score) + labels."""
            rows = []
            labels = []
            for i in range(7, ef_df.shape[0]):
                sc = ef_df.iloc[i, col_sc] if pd.notna(ef_df.iloc[i, col_sc]) else None
                lbl = str(ef_df.iloc[i, col_lbl]).strip() if pd.notna(ef_df.iloc[i, col_lbl]) else ''
                if sc is not None and lbl:
                    lo, hi = _parse_range_label(lbl)
                    rows.append((lo, hi, round(float(sc) * 100, 2)))
                    labels.append(lbl)
            return rows, labels

        parsed_ncc, lbl_ncc = _read_range_col(1, 2)   # Net Credit Change
        parsed_dq,  lbl_dq  = _read_range_col(3, 4)   # Delinquency
        parsed_es,  lbl_es  = _read_range_col(5, 6)   # Economic Stress

        if parsed_ncc and parsed_dq and parsed_es:
            result['env_ranges'] = {
                'ncc': parsed_ncc,
                'dq': parsed_dq,
                'es': parsed_es,
                'ncc_labels': lbl_ncc,
                'dq_labels': lbl_dq,
                'es_labels': lbl_es,
            }
            print(f"    Env factor ranges: NCC={len(parsed_ncc)}, "
                  f"DQ={len(parsed_dq)}, ES={len(parsed_es)}")
    except (ValueError, KeyError):
        print(f"    'Envir Fact Ranges' tab not found in {os.path.basename(found)}")

    # ── Balance Adjustments per pool from "Risk Change Data Entry" tab ──
    # Col A (0) = Grade, Col M (12) = Loan Report Balance, Col O (14) = Balance Sheet Total,
    # Col P (15) = Bal Adjustment, Col Q (16) = Specific Identification
    # Pool name in col N (13); Total rows have col A = "Total"
    try:
        rc_df = pd.read_excel(found, sheet_name='Risk Change Data Entry', header=None)
        bal_adj = {}       # {pool_name: adjustment_amount}
        pool_bal_detail = {}  # {pool_name: {grade: {loan_report_bal, bal_adj, balance_sheet_total, specific_id}}}
        total_bal_adj = 0.0
        total_in_portfolio = 0.0
        current_pool = None
        current_grades = {}  # grade detail accumulator for current pool
        skip_labels = {'', 'Loan Pool', '% of Loan Balance', 'Grand Total ', 'Grand Total'}

        def _safe_float(val):
            try:
                return float(val) if pd.notna(val) else 0.0
            except (ValueError, TypeError):
                return 0.0

        for idx in range(len(rc_df)):
            n_val = rc_df.iloc[idx, 13] if rc_df.shape[1] > 13 else None
            a_val = rc_df.iloc[idx, 0] if pd.notna(rc_df.iloc[idx, 0]) else None

            # Pool name row: col N has a string that matches pool name (not header labels)
            if pd.notna(n_val) and isinstance(n_val, str):
                nstr = n_val.strip()
                if nstr not in skip_labels:
                    current_pool = nstr
                    current_grades = {}

            if a_val is None or not current_pool:
                continue
            a_str = str(a_val).strip()

            # Skip header row and hidden grades
            if a_str in ('Current Grade', '') or a_str.upper().startswith('HIDE'):
                continue

            # Read per-row values: M=12, O=14, P=15, Q=16
            m_val = _safe_float(rc_df.iloc[idx, 12]) if rc_df.shape[1] > 12 else 0.0
            o_val = _safe_float(rc_df.iloc[idx, 14]) if rc_df.shape[1] > 14 else 0.0
            p_val = _safe_float(rc_df.iloc[idx, 15]) if rc_df.shape[1] > 15 else 0.0
            q_val = _safe_float(rc_df.iloc[idx, 16]) if rc_df.shape[1] > 16 else 0.0

            if a_str == 'Total':
                if not current_pool.upper().startswith('HIDE'):
                    bal_adj[current_pool] = p_val
                    total_bal_adj += p_val
                    total_in_portfolio += o_val
                    # Store grade detail plus total row
                    current_grades['Total'] = {
                        'loan_report_bal': m_val, 'bal_adj': p_val,
                        'balance_sheet_total': o_val, 'specific_id': q_val,
                    }
                    pool_bal_detail[current_pool] = current_grades
                current_pool = None
                current_grades = {}
            else:
                # Regular grade row
                current_grades[a_str] = {
                    'loan_report_bal': m_val, 'bal_adj': p_val,
                    'balance_sheet_total': o_val, 'specific_id': q_val,
                }

        result['balance_adjustments'] = bal_adj
        result['pool_bal_detail'] = pool_bal_detail
        result['total_balance_adjustment'] = round(total_bal_adj, 2)
        result['total_in_portfolio'] = round(total_in_portfolio, 2)
        if abs(total_bal_adj) > 0.01:
            print(f"    Balance adjustments: {sum(1 for v in bal_adj.values() if abs(v) > 0.01)} pools, "
                  f"Total: ${total_bal_adj:,.2f}")
    except (ValueError, KeyError):
        print(f"    'Risk Change Data Entry' tab not found in {os.path.basename(found)}")

    # ── Historical grade-level balances from "HIst Bal Data" tab ──
    try:
        hb_df = pd.read_excel(found, sheet_name='HIst Bal Data', header=None)
        # Row layout per pool block (15 rows):
        #   pool name | blank | ...
        #   "Current Grade" | blank | date1 | date2 | ...
        #   grade_label | blank | val1 | val2 | ...  (11 grades)
        #   "Total" | blank | val1 | val2 | ...
        #   blank row
        # Header rows 1-5 have metadata; dates are in row 5 (idx 4), col C onwards

        # Read dates from row 5 (index 4)
        hist_dates = []
        for c in range(2, hb_df.shape[1]):
            v = hb_df.iloc[4, c] if 4 < len(hb_df) else None
            if pd.notna(v):
                try:
                    hist_dates.append(pd.Timestamp(v))
                except Exception:
                    pass

        hist_bal_data = {}  # {pool: {dates: [...], grades: {grade: [vals]}, total: [vals]}}
        idx = 5  # start scanning after header rows
        while idx < len(hb_df):
            # Look for pool name row: col A has text, next row has "Current Grade"
            a_val = hb_df.iloc[idx, 0] if pd.notna(hb_df.iloc[idx, 0]) else None
            if a_val is None:
                idx += 1
                continue
            pool_name = str(a_val).strip()
            if pool_name in ('', 'Current Grade', 'Current Risk Rating',
                             'Total'):
                idx += 1
                continue
            # Check next row is "Current Grade" or "Current Risk Rating"
            # (BRR-flagged pools use Risk Rating with BRR labels)
            if idx + 1 < len(hb_df):
                next_a = hb_df.iloc[idx + 1, 0]
                if pd.notna(next_a) and str(next_a).strip() in (
                    'Current Grade', 'Current Risk Rating'
                ):
                    # This is a pool header; read grade rows
                    pool_grades = {}
                    pool_total = []
                    gr_idx = idx + 2  # first grade row
                    while gr_idx < len(hb_df):
                        ga = hb_df.iloc[gr_idx, 0]
                        if pd.isna(ga) or str(ga).strip() == '':
                            break
                        glabel = str(ga).strip()
                        # Skip Hide-* rows (FICO Hide-F/G/H/I AND BRR
                        # Hide-RF/RG/RH/RI)
                        if glabel.lower().startswith('hide'):
                            gr_idx += 1
                            continue
                        vals = []
                        for c in range(2, 2 + len(hist_dates)):
                            v = hb_df.iloc[gr_idx, c] if c < hb_df.shape[1] else 0
                            try:
                                vals.append(float(v) if pd.notna(v) else 0.0)
                            except (ValueError, TypeError):
                                vals.append(0.0)
                        if glabel == 'Total':
                            pool_total = vals
                            gr_idx += 1
                            break
                        else:
                            pool_grades[glabel] = vals
                        gr_idx += 1
                    if not pool_name.upper().startswith('HIDE'):
                        hist_bal_data[pool_name] = {
                            'dates': hist_dates,
                            'grades': pool_grades,
                            'total': pool_total,
                        }
                    idx = gr_idx
                    continue
            idx += 1

        result['hist_bal_data'] = hist_bal_data
        if hist_bal_data:
            print(f"    HIst Bal Data: {len(hist_bal_data)} pools, {len(hist_dates)} months")
    except (ValueError, KeyError):
        print(f"    'HIst Bal Data' tab not found in {os.path.basename(found)}")

    # ── DQ % / CO / RC per year from "Display CO-Recov -DQ" tab ──
    result.update(_parse_display_co_recov_dq(found))

    # ── Monthly CO/RC from WARM "Charge off History" tab ──
    try:
        co_hist_df = pd.read_excel(found, sheet_name='Charge off History', header=None)
        # Row 8 (0-indexed): section header "Charge offs" with dates in cols 2+
        # Rows 9-19: pool CO values (may include HIDE/Exclude rows after visible pools)
        # Row 33: section header "Recoveries" with dates in cols 2+
        # Rows 34-44: pool RC values (negative in WARM)
        co_dates = []
        for c in range(2, co_hist_df.shape[1]):
            v = co_hist_df.iloc[8, c]
            if pd.notna(v):
                try:
                    co_dates.append((c, pd.Timestamp(v)))
                except Exception:
                    pass

        def _parse_co_hist_section(start_idx):
            """Parse pool rows below a section header, returning {(yr,mo): {pool: val}}."""
            out = {}
            for ri in range(start_idx + 1, min(start_idx + 25, len(co_hist_df))):
                pn_raw = co_hist_df.iloc[ri, 0]
                if pd.isna(pn_raw):
                    continue
                pn = str(pn_raw).strip()
                if pn.upper().startswith(('HIDE', 'EXCLUDE', 'TOTAL')):
                    continue
                if pn == '':
                    break
                for ci, dt in co_dates:
                    v = co_hist_df.iloc[ri, ci]
                    val = float(v) if pd.notna(v) else 0.0
                    if val != 0:
                        ym = (dt.year, dt.month)
                        out.setdefault(ym, {})[pn] = val
            return out

        # Find section start rows
        co_start = None
        rc_start = None
        for idx in range(len(co_hist_df)):
            val = co_hist_df.iloc[idx, 0]
            if pd.notna(val):
                txt = str(val).strip().lower()
                if txt == 'charge offs':
                    co_start = idx
                elif txt == 'recoveries':
                    rc_start = idx

        if co_start is not None:
            result['warm_co_monthly'] = _parse_co_hist_section(co_start)
        if rc_start is not None:
            result['warm_rc_monthly'] = _parse_co_hist_section(rc_start)

        n_co = sum(len(v) for v in result.get('warm_co_monthly', {}).values())
        n_rc = sum(len(v) for v in result.get('warm_rc_monthly', {}).values())
        if n_co or n_rc:
            print(f"    WARM Charge off History: "
                  f"{len(result.get('warm_co_monthly', {}))} CO months ({n_co} entries), "
                  f"{len(result.get('warm_rc_monthly', {}))} RC months ({n_rc} entries)")
    except (ValueError, KeyError):
        pass  # tab not found, silently skip

    return result


def load_impaired_from_tct_baseline(config, snap):
    """Load impaired-loan data from the previously-generated TCT model baseline.

    Used when the source CECL-Migration-WARM file has no 'Impaired Loans' tab
    (because the TCT model is replacing it). Reads from
    Reports/_warm_baselines/<snap>_CECL_Migration_<CU>_TCT_Model.xlsx:

      - 'Impaired Loans' tab: pivot summary at rows 4-9 with
          col A=Impairment Type, col B=Provision Percentage,
          col L=Impairment Type, col N=Sum of Loss Given Default
        Allowance per category = Provision Pct * LGD.
      - 'Impaired Loans Pivot' tab: pool x grade pivot of Balance Removed.

    Returns dict with 'acl_impaired', 'spec_id_by_pool', 'total_spec_id'
    or {} if file/tabs not found.
    """
    cu = config['credit_union']
    safe_cu = cu.replace(' ', '_').replace('/', '-')
    baseline_dir = os.path.join(BASE, 'Reports', '_warm_baselines')
    if not os.path.isdir(baseline_dir):
        return {}
    target = f"{snap}_CECL_Migration_{safe_cu}_TCT_Model.xlsx"
    found = None
    for f in os.listdir(baseline_dir):
        if f == target and not f.startswith('~$'):
            found = os.path.join(baseline_dir, f)
            break
    if not found:
        # Fallback: pattern match for any TCT_Model baseline for this CU,
        # preferring the closest-dated baseline (snap exact, then most recent
        # prior, then earliest later if none prior exists). Provides historical
        # Display CO-Recov -DQ / Impaired tabs when no exact-snap baseline
        # exists.
        any_pat = re.compile(
            r'^(\d{4}-\d{2}-\d{2})_CECL_Migration_.*_TCT_Model\.xlsx$',
            re.IGNORECASE)
        prior_candidates = []  # (date, path) where date <= snap
        later_candidates = []  # (date, path) where date > snap
        for f in os.listdir(baseline_dir):
            if f.startswith('~$'):
                continue
            m = any_pat.match(f)
            if not m or safe_cu.lower() not in f.lower():
                continue
            d = m.group(1)
            path = os.path.join(baseline_dir, f)
            if d <= snap:
                prior_candidates.append((d, path))
            else:
                later_candidates.append((d, path))
        if prior_candidates:
            prior_candidates.sort(reverse=True)
            found = prior_candidates[0][1]
        elif later_candidates:
            later_candidates.sort()
            found = later_candidates[0][1]
    if not found:
        return {}

    print(f"    Loading impaired data from TCT baseline: {os.path.basename(found)}")
    try:
        from openpyxl import load_workbook as _lw
        wb = _lw(found, data_only=True, read_only=True)
    except Exception as e:
        print(f"    Warning: could not open baseline: {e}")
        return {}

    result = {}

    # ── Parse 'Impaired Loans' summary pivot for allowance per category ──
    acl_impaired = {}
    if 'Impaired Loans' in wb.sheetnames:
        ws = wb['Impaired Loans']
        # Iterate first ~25 rows; collect category->prov pct from cols A/B,
        # category->LGD from cols L/N.
        prov_pct = {}   # {category: pct}
        lgd_by_cat = {} # {category: lgd}
        for row in ws.iter_rows(min_row=1, max_row=25, max_col=14, values_only=True):
            if not row:
                continue
            a = row[0] if len(row) > 0 else None
            b = row[1] if len(row) > 1 else None
            l = row[11] if len(row) > 11 else None
            n = row[13] if len(row) > 13 else None
            if a and isinstance(a, str):
                a_str = a.strip()
                if a_str and a_str.upper() not in ('HIDE', 'IMPAIRMENT TYPE',
                                                   'TOTAL', 'CALCULATION',
                                                   'DATA ENTRY'):
                    try:
                        if b is not None:
                            prov_pct[a_str] = float(b)
                    except (ValueError, TypeError):
                        pass
            if l and isinstance(l, str):
                l_str = l.strip()
                if l_str and l_str.upper() not in ('HIDE', 'IMPAIRMENT TYPE',
                                                   'TOTAL'):
                    try:
                        if n is not None:
                            lgd_by_cat[l_str] = float(n)
                    except (ValueError, TypeError):
                        pass
        for cat, pct in prov_pct.items():
            lgd = lgd_by_cat.get(cat, 0.0)
            acl_impaired[cat] = pct * lgd
        if acl_impaired:
            result['acl_impaired'] = acl_impaired
            total_allow = sum(acl_impaired.values())
            print(f"    Impaired allowance: {len(acl_impaired)} categories, "
                  f"Total allow: ${total_allow:,.2f}")

    # ── Parse 'Impaired Loans Pivot' for spec_id_by_pool ──
    spec_id_by_pool = {}
    if 'Impaired Loans Pivot' in wb.sheetnames:
        ws = wb['Impaired Loans Pivot']
        # Header row at row 4: col A "Row Labels", cols B..(N-1) pool names,
        # last col "Grand Total".
        header_row = None
        for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=10,
                                                values_only=True), start=1):
            if not row:
                continue
            a = row[0]
            if isinstance(a, str) and a.strip().lower() == 'row labels':
                header_row = ridx
                pool_cols = []
                for ci, val in enumerate(row[1:], start=2):
                    if not val:
                        continue
                    s = str(val).strip()
                    if s.lower() in ('grand total', 'column labels'):
                        continue
                    pool_cols.append((ci, s))
                break

        if header_row is not None:
            for row in ws.iter_rows(min_row=header_row + 1,
                                    max_row=ws.max_row, values_only=True):
                if not row:
                    continue
                grade = row[0]
                if not grade:
                    continue
                g_str = str(grade).strip()
                if not g_str or g_str.lower() in ('grand total',):
                    continue
                if g_str.lower().startswith('hide'):
                    continue
                for ci, pname in pool_cols:
                    val = row[ci - 1] if ci - 1 < len(row) else None
                    if val is None:
                        continue
                    try:
                        v = float(val)
                    except (ValueError, TypeError):
                        continue
                    if v <= 0:
                        continue
                    if pname.lower() == 'exclude':
                        continue
                    spec_id_by_pool.setdefault(pname, {})
                    spec_id_by_pool[pname][g_str] = (
                        spec_id_by_pool[pname].get(g_str, 0.0) + v
                    )
        if spec_id_by_pool:
            result['spec_id_by_pool'] = spec_id_by_pool
            n = len(spec_id_by_pool)
            tot = sum(sum(g.values()) for g in spec_id_by_pool.values())
            print(f"    Spec ID by pool: {n} pools, "
                  f"Total removed: ${tot:,.2f}")

    if 'acl_impaired' in result or 'spec_id_by_pool' in result:
        result['total_spec_id'] = sum(result.get('acl_impaired', {}).values())

    # ── DQ % / CO / RC per year from "Display CO-Recov -DQ" tab ──
    # When the source CECL-Migration-WARM file is absent (modern setups have
    # only a TCT_Model baseline in Reports/_warm_baselines/), this is the only
    # path that populates warm_dq_pct, warm_co, warm_rc, warm_net, etc.
    result.update(_parse_display_co_recov_dq(found))

    try:
        wb.close()
    except Exception:
        pass
    return result


def load_wizard_impaired(config):
    """Load impaired-loan data persisted by the setup wizard.

    Reads ``cfg["impaired_loans"]["data_rows"]`` (written by
    ``cecl_ui.services.config_service.build_yaml_from_wizard``) where
    each row carries the wizard-resolved ``loan_pool``, ``credit_grade``,
    ``balance_removed``, ``provision_amount``, and ``impairment_type``.

    Returns ``None`` when no wizard rows are present, otherwise a dict
    with the same keys as :func:`load_standalone_impaired`:
      'acl_impaired'     : {impairment_type: total_provision_amount}
      'spec_id_by_pool'  : {pool: {grade: balance_removed, ...}}
      'total_spec_id'    : sum of all provision_amounts
    """
    imp_cfg = config.get('impaired_loans') or {}
    rows = imp_cfg.get('data_rows') or []
    if not rows:
        return None

    acl_impaired: dict[str, float] = {}
    spec_id_by_pool: dict[str, dict[str, float]] = {}
    n_rows = 0
    for r in rows:
        if not isinstance(r, dict):
            continue
        try:
            bal = float(r.get('balance_removed') or 0.0)
        except (TypeError, ValueError):
            bal = 0.0
        if bal <= 0:
            continue
        pool = str(r.get('loan_pool') or '').strip()
        if not pool:
            continue
        grade = str(r.get('credit_grade') or '').strip()
        spec_id_by_pool.setdefault(pool, {})
        spec_id_by_pool[pool][grade] = (
            spec_id_by_pool[pool].get(grade, 0.0) + bal
        )
        imp_type = str(r.get('impairment_type') or '').strip() or 'Other'
        prov = r.get('provision_amount')
        try:
            prov_val = float(prov) if prov is not None else 0.0
        except (TypeError, ValueError):
            prov_val = 0.0
        acl_impaired[imp_type] = acl_impaired.get(imp_type, 0.0) + prov_val
        n_rows += 1

    if not spec_id_by_pool:
        return None

    result = {
        'acl_impaired': acl_impaired,
        'spec_id_by_pool': spec_id_by_pool,
        'total_spec_id': sum(acl_impaired.values()),
    }
    n_pools = len(spec_id_by_pool)
    tot_removed = sum(sum(g.values()) for g in spec_id_by_pool.values())
    print(f"    Wizard impaired loans: {n_rows} row(s), "
          f"{n_pools} pool(s), ${tot_removed:,.2f} balance removed, "
          f"${result['total_spec_id']:,.2f} provision")
    return result


def _resolve_pool_ci(code, pool_map_ci, default_pool):
    """Case-insensitive pool_map lookup. ``pool_map_ci`` is the pool_map
    with keys lower-cased and stripped. Returns the mapped pool name or
    ``default_pool`` (or '' when default is also empty/ignore). Returns
    None when the result should be skipped (empty / 'ignore' / 'exclude').
    """
    if code is None:
        return None
    key = str(code).strip().lower()
    if not key:
        return None
    pool = pool_map_ci.get(key)
    if not pool:
        # Try with collapsed whitespace.
        key2 = ' '.join(key.split())
        pool = pool_map_ci.get(key2)
    if not pool:
        pool = (default_pool or '').strip()
    if not pool or pool.lower() in ('ignore', 'exclude'):
        return None
    return pool


# ── NCUA 5300 canonical-name → pool inference ─────────────────────────
# The wizard's 5300 backfill populates loan_code_history /
# loan_code_chargeoff_history / loan_code_recovery_history with NCUA
# canonical names (e.g. 'First Liens', 'Used Vehicles', 'Unsecured
# Credit Card'). A CU's pool_map usually contains only the CU's own
# local loan codes ('17', '40', 'GL-12', ...) and not these NCUA
# names. Without a fallback, every 5300-derived row collapses into
# ``default_pool`` (typically 'Other/Uncategorized') and the older
# years on the Vizo "Display Hist Bal" tab show no Real Estate /
# Consumer breakdown.
#
# Below maps each known NCUA canonical name to a generic category
# tag. Each tag has a list of keyword fragments — we pick whichever
# configured pool name contains one of those keywords. This way the
# inference adapts to per-CU pool naming (Real Estate vs Mortgage,
# Consumer Secured vs Auto, etc.) without per-CU YAML edits.
_NCUA_CANONICAL_TAG = {
    '1st mortgage real estate':         'real_estate',
    'first liens':                      'real_estate',
    'other real estate':                'real_estate',
    'junior liens':                     'real_estate',
    'other real estate (other)':        'real_estate',
    'commercial real estate':           'commercial',
    'new vehicles':                     'consumer_secured',
    'used vehicles':                    'consumer_secured',
    'leases receivable':                'consumer_secured',
    'unsecured credit card':            'consumer_unsecured',
    'all other unsecured':              'consumer_unsecured',
    'payday alternative loans':         'consumer_unsecured',
    'non-federally guaranteed student': 'consumer_unsecured',
    'all other':                        'consumer_unsecured',
    'all other loans':                  'consumer_unsecured',
    'commercial (non-re)':              'commercial',
}
# Tag -> ordered list of keyword fragments to scan in configured pool
# names. First pool whose lower-cased name contains any keyword wins.
_NCUA_TAG_KEYWORDS = {
    'real_estate':        ('real estate', 'mortgage', 'realestate',
                           'home equity', 'heloc', 'real-estate'),
    'consumer_secured':   ('auto', 'vehicle', 'consumer secured',
                           'secured', 'lease'),
    'consumer_unsecured': ('consumer unsecured', 'unsecured', 'credit card',
                           'signature'),
    'commercial':         ('commercial', 'business', 'mbl'),
}


def _build_ncua_canonical_pool_lookup(config):
    """Return ``{canonical_name_lower: pool_name}`` for NCUA canonical
    names auto-classified into the CU's configured pools. Used as a
    fallback for the 5300-history loaders so older-year rows land in
    real pools rather than collapsing into ``default_pool``. Returns
    an empty dict when no configured pool matches a tag's keywords.
    """
    pool_names = []
    for p in (config.get('pools') or []):
        n = (p.get('name') or '').strip()
        if n:
            pool_names.append(n)
    for n in (config.get('pool_order') or []):
        n = (n or '').strip()
        if n and n not in pool_names:
            pool_names.append(n)
    if not pool_names:
        return {}
    excluded = {(s or '').strip().lower()
                for s in (config.get('excluded_pools') or [])}
    excluded.add('exclude')
    excluded.add('ignore')
    pool_lc = [(n.lower(), n) for n in pool_names
               if n.lower() not in excluded]

    def _pool_for_tag(tag, hint_keywords=()):
        # Try canonical-specific hint keywords FIRST (e.g. 'new' / 'used'
        # / 'first' / 'junior') so a CU with finer-grained pools like
        # New Auto + Used Auto + New Indirect Auto + Used Indirect Auto
        # gets per-canonical-name routing rather than collapsing every
        # vehicle DQ row into whichever pool first matches 'auto'.
        for hint in hint_keywords:
            for kw in _NCUA_TAG_KEYWORDS.get(tag, ()):
                for n_lc, n in pool_lc:
                    if hint in n_lc and kw in n_lc:
                        return n
        for kw in _NCUA_TAG_KEYWORDS.get(tag, ()):
            for n_lc, n in pool_lc:
                if kw in n_lc:
                    return n
        return None

    def _hints_for(canonical_lc):
        # Strip noise words so 'new vehicles' yields ('new',) and
        # 'used vehicles' yields ('used',). Multi-word hints listed
        # before single-word so 'first liens' tries 'first lien' first.
        out = []
        for w in ('first lien', 'junior lien', 'new', 'used', 'first',
                  'junior', 'farm', 'multifamily', 'agricultural',
                  'construction', 'student', 'credit card', 'payday',
                  'commercial', 'lease'):
            if w in canonical_lc:
                out.append(w)
        return tuple(out)

    lookup: dict[str, str] = {}
    for canonical, tag in _NCUA_CANONICAL_TAG.items():
        pool = _pool_for_tag(tag, _hints_for(canonical))
        if pool:
            lookup[canonical] = pool
    return lookup


def _resolve_pool_with_ncua(code, pool_map_ci, ncua_lookup, default_pool):
    """Like :func:`_resolve_pool_ci` but consults the NCUA canonical
    inference map BEFORE falling back to ``default_pool``. The user's
    explicit ``pool_map`` always wins.
    """
    if code is None:
        return None
    key = str(code).strip().lower()
    if not key:
        return None
    pool = pool_map_ci.get(key) or pool_map_ci.get(' '.join(key.split()))
    if not pool and ncua_lookup:
        pool = ncua_lookup.get(key) or ncua_lookup.get(' '.join(key.split()))
    if not pool:
        pool = (default_pool or '').strip()
    if not pool or pool.lower() in ('ignore', 'exclude'):
        return None
    return pool


def _overlay_warm_history_into_hist(hist, snap):
    """Fold WARM-template historical CO / Recoveries / DQ% / hist balances
    from ``hist['impaired']`` into the top-level ``hist`` keys (chargeoffs,
    recoveries, co_monthly, rc_monthly, dq_pct, avg_balances, years).

    The WARM workbook is the analyst's authoritative source for historical
    data — when present, its multi-year history should drive the Display
    HIst Bal / Display CO-Recov-DQ year axes even when there are no
    file-based or DB-backfilled rows for the CU.

    Precedence: WARM cells WIN for any (year, pool) cell they cover.
    File / DB cells outside that coverage (e.g. fresh current-quarter CO
    data parsed from a file) are preserved. This mirrors the prior-TCT
    fallback path in :func:`generate_report` which fully replaces the
    historical year axis with the prior report's WARM totals while
    splicing in the current snapshot year from raw-file parsing.

    Years beyond ``snap`` are dropped (same clamp behaviour as the
    report-period trim a few blocks earlier).
    """
    imp = hist.get('impaired') if hist else None
    if not imp:
        return

    snap_year = None
    snap_month = None
    try:
        snap_year = int(str(snap)[:4])
        snap_month = int(str(snap)[5:7])
    except (TypeError, ValueError):
        pass

    def _yr_ok(y):
        try:
            yi = int(y)
        except (TypeError, ValueError):
            return False
        return snap_year is None or yi <= snap_year

    overlay_co = 0
    overlay_rc = 0
    overlay_dq = 0
    overlay_bal_cells = 0

    warm_co = imp.get('warm_co') or {}
    if warm_co:
        co = hist.setdefault('chargeoffs', {})
        for yr, by_pool in warm_co.items():
            if not _yr_ok(yr):
                continue
            target = co.setdefault(int(yr), {})
            for pool, amt in (by_pool or {}).items():
                target[pool] = amt
                overlay_co += 1

    warm_rc = imp.get('warm_rc') or {}
    if warm_rc:
        rc = hist.setdefault('recoveries', {})
        for yr, by_pool in warm_rc.items():
            if not _yr_ok(yr):
                continue
            target = rc.setdefault(int(yr), {})
            for pool, amt in (by_pool or {}).items():
                target[pool] = amt
                overlay_rc += 1

    warm_dq = imp.get('warm_dq_pct') or {}
    if warm_dq:
        dq_pct = hist.setdefault('dq_pct', {})
        for yr, by_pool in warm_dq.items():
            if not _yr_ok(yr):
                continue
            target = dq_pct.setdefault(int(yr), {})
            for pool, pct in (by_pool or {}).items():
                target[pool] = pct
                overlay_dq += 1

    warm_co_m = imp.get('warm_co_monthly') or {}
    if warm_co_m:
        com = hist.setdefault('co_monthly', {})
        for ym, by_pool in warm_co_m.items():
            try:
                yy = int(ym[0])
                mm = int(ym[1])
            except (TypeError, ValueError, IndexError):
                continue
            if snap_year is not None and (
                yy > snap_year
                or (yy == snap_year and snap_month is not None and mm > snap_month)
            ):
                continue
            com.setdefault((yy, mm), {}).update(by_pool or {})

    warm_rc_m = imp.get('warm_rc_monthly') or {}
    if warm_rc_m:
        rcm = hist.setdefault('rc_monthly', {})
        for ym, by_pool in warm_rc_m.items():
            try:
                yy = int(ym[0])
                mm = int(ym[1])
            except (TypeError, ValueError, IndexError):
                continue
            if snap_year is not None and (
                yy > snap_year
                or (yy == snap_year and snap_month is not None and mm > snap_month)
            ):
                continue
            rcm.setdefault((yy, mm), {}).update(by_pool or {})

    # Annual average balances per pool from WARM hist_bal_data.
    # WARM cells fill (year, pool) slots not already populated by the
    # per_month / monthly_balances file path, so live monthly data
    # uploaded by the analyst remains authoritative for any year it
    # covers.
    hbd = imp.get('hist_bal_data') or {}
    if hbd:
        avg = hist.setdefault('avg_balances', {})
        for pool, pdata in hbd.items():
            dates = (pdata or {}).get('dates') or []
            totals = (pdata or {}).get('total') or []
            if not dates:
                continue
            yr_sums, yr_cnts = {}, {}
            for i, d in enumerate(dates):
                if i >= len(totals):
                    continue
                v = totals[i]
                if not v:
                    continue
                try:
                    yr = int(d.year)
                except AttributeError:
                    try:
                        yr = int(str(d)[:4])
                    except (TypeError, ValueError):
                        continue
                if snap_year is not None and yr > snap_year:
                    continue
                yr_sums[yr] = yr_sums.get(yr, 0) + v
                yr_cnts[yr] = yr_cnts.get(yr, 0) + 1
            for yr, ssum in yr_sums.items():
                ya = avg.setdefault(yr, {})
                if not ya.get(pool):
                    ya[pool] = ssum / yr_cnts[yr]
                    overlay_bal_cells += 1

    # Recompute years union (drop any > snap_year).
    yrs = set()
    for k in ('chargeoffs', 'recoveries', 'avg_balances', 'dq_pct'):
        d = hist.get(k) or {}
        for y in d.keys():
            if isinstance(y, int) and _yr_ok(y):
                yrs.add(y)
    if yrs:
        hist['years'] = sorted(yrs)

    if overlay_co or overlay_rc or overlay_dq or overlay_bal_cells:
        n_yrs = len(hist.get('years') or [])
        print(f"    Overlay from WARM file: {overlay_co} CO cell(s), "
              f"{overlay_rc} Rc cell(s), {overlay_dq} DQ cell(s), "
              f"{overlay_bal_cells} avg-balance cell(s); "
              f"hist['years'] now covers {n_yrs} year(s)")


def _load_co_rc_history_from_db(config):
    """Aggregate ``loan_code_chargeoff_history`` and
    ``loan_code_recovery_history`` rows into annual + monthly per-pool
    totals, mirroring the shape returned by
    :func:`load_chargeoff_recovery_history`.

    Returns ``{'chargeoffs', 'recoveries', 'co_monthly', 'rc_monthly',
    'years'}`` — empty dicts when nothing is available (missing tables,
    no rows, DB unavailable).

    Codes are mapped to pools via ``cfg['pool_map']`` (case-insensitive)
    with fallback to ``cfg['default_pool']``. Rows whose mapped pool is
    empty / 'ignore' / 'exclude' are dropped.
    """
    empty = {'chargeoffs': {}, 'recoveries': {},
             'co_monthly': {}, 'rc_monthly': {}, 'years': []}
    cu = (config.get('credit_union') or '').strip()
    if not cu:
        return empty
    raw_map = config.get('pool_map') or {}
    pool_map_ci = {str(k).strip().lower(): str(v).strip()
                   for k, v in raw_map.items()
                   if str(k).strip() and str(v).strip()}
    # Identity fallback for stored codes that are ALREADY pool names.
    # The monthly CO/Recovery aggregator (monthly_co_recov_aggregator) stores
    # each row's ``loan_code`` as the pool-MAPPED name (e.g. raw "Unsecured"
    # is saved as "Personal"), whereas the 5300 backfill stores raw NCUA
    # category codes. At report time we re-resolve every ``loan_code`` through
    # pool_map, which is keyed by RAW codes — so a stored pool name only
    # round-trips when it happens to equal its own source key (e.g.
    # "Credit Cards" -> "CREDIT CARDS"). Names like "Personal" or
    # "1st/2nd Lien Mortgage" are pool_map VALUES, not keys, so they resolved
    # to nothing and (with default_pool="Ignore") were dropped entirely.
    # Register every known pool name as an identity mapping so an
    # already-mapped code resolves to itself. ``setdefault`` keeps any real
    # raw-code mapping authoritative.
    for _v in raw_map.values():
        _vn = str(_v).strip()
        if _vn and _vn.lower() not in ('ignore', 'exclude'):
            pool_map_ci.setdefault(_vn.lower(), _vn)
    for _p in (config.get('pools') or []):
        _pn = _p.get('name') if isinstance(_p, dict) else _p
        _pn = str(_pn or '').strip()
        if _pn and _pn.lower() not in ('ignore', 'exclude'):
            pool_map_ci.setdefault(_pn.lower(), _pn)
    default_pool = config.get('default_pool') or ''
    try:
        from cecl_credentials import get_database_url
        from sqlalchemy import create_engine, text as _sql_text
    except Exception:  # noqa: BLE001
        return empty
    try:
        eng = create_engine(get_database_url())
    except Exception as exc:  # noqa: BLE001
        print(f"    CO/Recovery DB skipped: {type(exc).__name__}: {exc}")
        return empty

    def _read(table, amt_col):
        try:
            with eng.begin() as conn:
                rows = conn.execute(
                    _sql_text(
                        f"SELECT as_of_date, loan_code, {amt_col} "
                        f"FROM {table} WHERE cu = :cu"
                    ),
                    {"cu": cu},
                ).fetchall()
        except Exception as exc:  # noqa: BLE001
            print(f"    {table} read skipped: {type(exc).__name__}: {exc}")
            return []
        return rows

    co_rows = _read('loan_code_chargeoff_history', 'chargeoff_amount')
    rc_rows = _read('loan_code_recovery_history', 'recovery_amount')
    if not co_rows and not rc_rows:
        return empty

    ncua_lookup = _build_ncua_canonical_pool_lookup(config)

    def _aggregate(rows):
        annual: dict[int, dict[str, float]] = {}
        monthly: dict[tuple[int, int], dict[str, float]] = {}
        for r in rows:
            d = r[0]
            try:
                yr = int(d.year)
                mo = int(d.month)
            except Exception:
                continue
            if not (2000 <= yr <= 2099):
                continue
            pool = _resolve_pool_with_ncua(
                r[1], pool_map_ci, ncua_lookup, default_pool,
            )
            if not pool:
                continue
            try:
                amt = float(r[2] or 0.0)
            except (TypeError, ValueError):
                continue
            if not amt:
                continue
            annual.setdefault(yr, {})
            annual[yr][pool] = annual[yr].get(pool, 0.0) + amt
            ym = (yr, mo)
            monthly.setdefault(ym, {})
            monthly[ym][pool] = monthly[ym].get(pool, 0.0) + amt
        return annual, monthly

    co_annual, co_monthly = _aggregate(co_rows)
    rc_annual, rc_monthly = _aggregate(rc_rows)
    years = sorted(set(co_annual) | set(rc_annual))
    n_co_cells = sum(len(v) for v in co_annual.values())
    n_rc_cells = sum(len(v) for v in rc_annual.values())
    if n_co_cells or n_rc_cells:
        print(f"    Loaded CO/Recovery history from DB: "
              f"{len(co_rows)} CO row(s) -> {n_co_cells} pool-year cell(s); "
              f"{len(rc_rows)} recovery row(s) -> {n_rc_cells} pool-year cell(s).")
    return {
        'chargeoffs': co_annual,
        'recoveries': rc_annual,
        'co_monthly': co_monthly,
        'rc_monthly': rc_monthly,
        'years': years,
    }


def _load_balance_history_from_db(config):
    """Aggregate ``loan_code_history.total_balance`` into per-year/per-pool
    average balances, suitable for overlaying onto
    ``hist['avg_balances']`` for years that aren't covered by the per_month
    file or single-snapshot ``monthly_balances`` data.

    The wizard's 5300 backfill writes one row per (cu, as_of_date,
    loan_code) into ``loan_code_history``, covering several years of
    quarterly history before the per_month / WARM coverage begins.
    Without this overlay the Vizo "Display Hist Bal" tab leaves the
    early-year columns blank for long Life-of-Loan pools (e.g. Real
    Estate at 84 months needs 2019-2022 data when the report quarter
    is 2026-03).

    Codes are mapped to pools via ``cfg['pool_map']`` (case-insensitive)
    with fallback to ``cfg['default_pool']``. Rows whose mapped pool is
    empty / 'ignore' / 'exclude' are dropped. For each (year, pool) the
    annual average is computed across the available month-end snapshots
    in that year (typical 5300 coverage = 4 quarter-ends).

    Returns ``{year: {pool: avg_balance}}`` — empty dict when nothing
    is available.
    """
    cu = (config.get('credit_union') or '').strip()
    if not cu:
        return {}
    raw_map = config.get('pool_map') or {}
    pool_map_ci = {str(k).strip().lower(): str(v).strip()
                   for k, v in raw_map.items()
                   if str(k).strip() and str(v).strip()}
    default_pool = config.get('default_pool') or ''
    try:
        from cecl_credentials import get_database_url
        from sqlalchemy import create_engine, text as _sql_text
    except Exception:  # noqa: BLE001
        return {}
    try:
        eng = create_engine(get_database_url())
    except Exception as exc:  # noqa: BLE001
        print(f"    Balance history DB skipped: {type(exc).__name__}: {exc}")
        return {}
    try:
        with eng.begin() as conn:
            rows = conn.execute(
                _sql_text(
                    "SELECT as_of_date, loan_code, total_balance "
                    "FROM loan_code_history WHERE cu = :cu"
                ),
                {"cu": cu},
            ).fetchall()
    except Exception as exc:  # noqa: BLE001
        print(f"    loan_code_history read skipped: "
              f"{type(exc).__name__}: {exc}")
        return {}
    if not rows:
        return {}

    # NCUA 5300 schema quirk: a few field codes have parent/child overlap
    # where both the rollup and the lien-detail are reported with the
    # SAME total. Mapping both into the same pool double-counts the
    # balance. Drop child rows when the parent code is present for the
    # same (cu, as_of_date). loan_code -> set of child loan_codes that
    # should be skipped when this parent has a row.
    _PARENT_CHILDREN = {
        '1st mortgage real estate': {'first liens'},          # A703 vs A703A
        'other real estate': {'junior liens', 'other real estate (other)'},  # A386 vs A386A/A386B
    }
    by_date_codes: dict[str, set[str]] = {}
    for r in rows:
        d = r[0]
        if not d:
            continue
        # Only treat a code as "present" on this date if its balance is
        # non-zero. NCUA backfill writes a row per canonical field even
        # when the value is 0, so the presence of '1st Mortgage Real
        # Estate' with $0 must not suppress 'First Liens' with $15M.
        try:
            bal = float(r[2] or 0.0)
        except (TypeError, ValueError):
            bal = 0.0
        if bal <= 0:
            continue
        by_date_codes.setdefault(d.isoformat(), set()).add(
            str(r[1] or '').strip().lower()
        )
    suppressed_children: dict[str, set[str]] = {}
    for iso, codes in by_date_codes.items():
        skip: set[str] = set()
        for parent, children in _PARENT_CHILDREN.items():
            if parent in codes:
                skip.update(children & codes)
        if skip:
            suppressed_children[iso] = skip

    ncua_lookup = _build_ncua_canonical_pool_lookup(config)

    # Pool-name passthrough: rows written by the "distributed" 5300
    # backfill mode use the user's pool name directly as ``loan_code``
    # (e.g. "Real Estate"), which won't appear in pool_map (local CU
    # codes -> pools) or the NCUA canonical map. Build a case-
    # insensitive set of configured pool names so those rows route
    # straight through.
    configured_pool_names: dict[str, str] = {}
    for p in (config.get('pools') or []):
        if isinstance(p, dict):
            name = str(p.get('name') or '').strip()
            if name:
                configured_pool_names[name.lower()] = name
    for name in (config.get('pool_order') or []):
        if isinstance(name, str) and name.strip():
            configured_pool_names.setdefault(name.strip().lower(),
                                             name.strip())

    # Sum balances per (year, month-end, pool), then average per (year, pool).
    by_year_month: dict[int, dict[str, dict[str, float]]] = {}
    n_skipped = 0
    for r in rows:
        d = r[0]
        try:
            yr = int(d.year)
            mo_key = d.isoformat()
        except Exception:
            continue
        if not (2000 <= yr <= 2099):
            continue
        code_lc = str(r[1] or '').strip().lower()
        if code_lc in suppressed_children.get(mo_key, ()):
            n_skipped += 1
            continue
        # Distributed-mode rows: code IS already a configured pool name.
        if code_lc in configured_pool_names:
            pool = configured_pool_names[code_lc]
        else:
            pool = _resolve_pool_with_ncua(
                r[1], pool_map_ci, ncua_lookup, default_pool,
            )
        if not pool:
            continue
        try:
            bal = float(r[2] or 0.0)
        except (TypeError, ValueError):
            continue
        by_year_month.setdefault(yr, {}).setdefault(mo_key, {})
        by_year_month[yr][mo_key][pool] = (
            by_year_month[yr][mo_key].get(pool, 0.0) + bal
        )

    annual: dict[int, dict[str, float]] = {}
    for yr, by_mo in by_year_month.items():
        # Per pool: average across the month-ends present for that year.
        pool_sums: dict[str, float] = {}
        pool_cnts: dict[str, int] = {}
        for mo_key, by_pool in by_mo.items():
            for pool, bal in by_pool.items():
                if bal <= 0:
                    continue
                pool_sums[pool] = pool_sums.get(pool, 0.0) + bal
                pool_cnts[pool] = pool_cnts.get(pool, 0) + 1
        for pool, total in pool_sums.items():
            cnt = pool_cnts[pool]
            if cnt:
                annual.setdefault(yr, {})[pool] = total / cnt

    n_cells = sum(len(v) for v in annual.values())
    if n_cells:
        yrs = sorted(annual.keys())
        msg = (f"    Loaded balance history from DB: {len(rows)} row(s) -> "
               f"{n_cells} pool-year cell(s) ({yrs[0]}-{yrs[-1]})")
        if n_skipped:
            msg += f"; skipped {n_skipped} duplicate parent/child 5300 row(s)"
        print(msg)
    return annual


def _load_balance_history_monthly_from_db(config):
    """Return per-(pool, month-end) balances from ``loan_code_history``.

    Returns a DataFrame ``[pool, date, balance]`` (date = pd.Timestamp at
    month-end) suitable for unioning into the per-month ``monthly_balances``
    frame so the supplemental Detail_HIst Balances tab can extend its
    historical columns back into the 5300-backfill years.

    Honors the same NCUA parent/child suppression and configured-pool
    passthrough as ``_load_balance_history_from_db``.
    """
    cu = (config.get('credit_union') or '').strip()
    if not cu:
        return pd.DataFrame(columns=['pool', 'date', 'balance'])
    raw_map = config.get('pool_map') or {}
    pool_map_ci = {str(k).strip().lower(): str(v).strip()
                   for k, v in raw_map.items()
                   if str(k).strip() and str(v).strip()}
    default_pool = config.get('default_pool') or ''
    try:
        from cecl_credentials import get_database_url
        from sqlalchemy import create_engine, text as _sql_text
    except Exception:
        return pd.DataFrame(columns=['pool', 'date', 'balance'])
    try:
        eng = create_engine(get_database_url())
        with eng.begin() as conn:
            rows = conn.execute(
                _sql_text(
                    "SELECT as_of_date, loan_code, total_balance "
                    "FROM loan_code_history WHERE cu = :cu"
                ),
                {"cu": cu},
            ).fetchall()
    except Exception:
        return pd.DataFrame(columns=['pool', 'date', 'balance'])
    if not rows:
        return pd.DataFrame(columns=['pool', 'date', 'balance'])

    _PARENT_CHILDREN = {
        '1st mortgage real estate': {'first liens'},
        'other real estate': {'junior liens', 'other real estate (other)'},
    }
    by_date_codes: dict[str, set[str]] = {}
    for r in rows:
        d = r[0]
        if not d:
            continue
        try:
            bal = float(r[2] or 0.0)
        except (TypeError, ValueError):
            bal = 0.0
        if bal <= 0:
            continue
        by_date_codes.setdefault(d.isoformat(), set()).add(
            str(r[1] or '').strip().lower()
        )
    suppressed_children: dict[str, set[str]] = {}
    for iso, codes in by_date_codes.items():
        skip: set[str] = set()
        for parent, children in _PARENT_CHILDREN.items():
            if parent in codes:
                skip.update(children & codes)
        if skip:
            suppressed_children[iso] = skip

    ncua_lookup = _build_ncua_canonical_pool_lookup(config)
    configured_pool_names: dict[str, str] = {}
    for p in (config.get('pools') or []):
        if isinstance(p, dict):
            name = str(p.get('name') or '').strip()
            if name:
                configured_pool_names[name.lower()] = name
    for name in (config.get('pool_order') or []):
        if isinstance(name, str) and name.strip():
            configured_pool_names.setdefault(name.strip().lower(),
                                             name.strip())

    by_pool_date: dict[tuple[str, str], float] = {}
    for r in rows:
        d = r[0]
        if not d:
            continue
        mo_key = d.isoformat()
        code_lc = str(r[1] or '').strip().lower()
        if code_lc in suppressed_children.get(mo_key, ()):
            continue
        if code_lc in configured_pool_names:
            pool = configured_pool_names[code_lc]
        else:
            pool = _resolve_pool_with_ncua(
                r[1], pool_map_ci, ncua_lookup, default_pool,
            )
        if not pool:
            continue
        try:
            bal = float(r[2] or 0.0)
        except (TypeError, ValueError):
            continue
        if bal <= 0:
            continue
        key = (pool, mo_key)
        by_pool_date[key] = by_pool_date.get(key, 0.0) + bal

    if not by_pool_date:
        return pd.DataFrame(columns=['pool', 'date', 'balance'])
    records = [
        {'pool': pool,
         'date': pd.Timestamp(iso) + pd.offsets.MonthEnd(0),
         'balance': bal}
        for (pool, iso), bal in by_pool_date.items()
    ]
    return pd.DataFrame(records)


def _load_dq_history_from_db(config):
    """Aggregate ``loan_code_delinquency_history`` into per-year/per-pool DQ%.


    Returns ``{year: {pool: dq_pct}}`` suitable for overlaying onto
    ``hist['impaired']['warm_dq_pct']``. Rows are bucketed to the most
    recent date in their calendar year (DQ is point-in-time, so the
    year-end / latest available quarter for that year is the right
    proxy for the WARM-style annual DQ%).

    Resolution rules per (year, pool):
      * Sum ``dq_amount`` and ``total_balance`` across all loan_codes
        mapping to that pool for the chosen as_of_date.
      * If ``total_balance`` totals to > 0, return
        ``dq_amount / total_balance``.
      * Otherwise, if ANY row carries a non-null ``dq_pct``, return the
        balance-weighted average (or simple average when all balances
        are null).

    Returns an empty dict on any error (missing table, no rows, etc.).
    """
    cu = (config.get('credit_union') or '').strip()
    if not cu:
        return {}
    pool_map = config.get('pool_map') or {}
    default_pool = config.get('default_pool') or ''
    pool_map_ci = {
        str(k).strip().lower(): v
        for k, v in pool_map.items()
        if str(k).strip()
    }
    ncua_lookup = _build_ncua_canonical_pool_lookup(config)
    try:
        from cecl_credentials import get_database_url
        from sqlalchemy import create_engine, text as _sql_text
    except Exception:  # noqa: BLE001
        return {}
    try:
        eng = create_engine(get_database_url())
        with eng.begin() as conn:
            rows = conn.execute(
                _sql_text(
                    "SELECT as_of_date, loan_code, dq_amount, "
                    "       total_balance, dq_pct "
                    "FROM loan_code_delinquency_history "
                    "WHERE cu = :cu"
                ),
                {"cu": cu},
            ).fetchall()
    except Exception as exc:  # noqa: BLE001
        print(f"    DQ history DB read skipped: {type(exc).__name__}: {exc}")
        return {}
    if not rows:
        return {}

    # Pull total_balance per (as_of_date, loan_code) from
    # loan_code_history as the denominator for DQ%. The 5300 DQ
    # backfill writes dq_amount but leaves total_balance NULL on
    # delinquency rows; without this lookup every cell collapses to
    # None and the overlay returns empty.
    bal_lookup: dict[tuple[str, str], float] = {}
    try:
        eng2 = create_engine(get_database_url())
        with eng2.begin() as conn:
            bal_rows = conn.execute(
                _sql_text(
                    "SELECT as_of_date, loan_code, total_balance "
                    "FROM loan_code_history "
                    "WHERE cu = :cu AND total_balance IS NOT NULL"
                ),
                {"cu": cu},
            ).fetchall()
        for br in bal_rows:
            d = br[0].isoformat() if hasattr(br[0], 'isoformat') else str(br[0])
            code = str(br[1]).strip()
            try:
                bal_lookup[(d, code)] = float(br[2])
            except (TypeError, ValueError):
                pass
    except Exception as exc:  # noqa: BLE001
        print(f"    DQ balance lookup skipped: {type(exc).__name__}: {exc}")

    # year -> latest as_of_date in that year
    latest_in_year: dict[int, str] = {}
    for r in rows:
        d = r[0].isoformat() if hasattr(r[0], 'isoformat') else str(r[0])
        try:
            yr = int(d[:4])
        except (TypeError, ValueError):
            continue
        if yr not in latest_in_year or d > latest_in_year[yr]:
            latest_in_year[yr] = d

    # Collect contributions per (year, pool).
    by_yp: dict[tuple[int, str], dict[str, float]] = {}
    # Track which (date, pool) total_balance rows we have already added
    # so multiple NCUA codes resolving to the same pool don't multiply
    # the pool's denominator -- BUT only when the balance came from the
    # POOL-LEVEL fallback (bal_lookup, keyed by user pool name from the
    # 5300-distributed backfill). Per-NCUA-code balances (Phase 9.30,
    # stored on the DQ row itself) are independent quantities and must
    # be summed without dedupe. seen_pool_totals tracks only the fallback
    # contributions.
    seen_pool_totals: set[tuple[str, str]] = set()
    for r in rows:
        d = r[0].isoformat() if hasattr(r[0], 'isoformat') else str(r[0])
        try:
            yr = int(d[:4])
        except (TypeError, ValueError):
            continue
        if latest_in_year.get(yr) != d:
            continue  # only use the latest as_of_date per calendar year
        code = str(r[1]).strip()
        pool = _resolve_pool_with_ncua(
            code, pool_map_ci, ncua_lookup, default_pool,
        )
        if not pool:
            continue
        amt = float(r[2] or 0.0)
        tot = float(r[3]) if r[3] is not None else None
        pct = float(r[4]) if r[4] is not None else None
        # Phase 9.30: prefer the per-NCUA-code total_balance written by
        # solr_5300_delq_backfill (column r[3]). Each DQ row carries the
        # balance for its own NCUA code, so summing them per resolved
        # pool produces a denominator that stays in lockstep with the
        # numerator (dq_amount) -- both aggregated LIVE under the user's
        # current pool_map, eliminating the stale-frozen-balance class
        # of bug. Fallback to the pool-level bal_lookup only when r[3]
        # is NULL (older rows from before Phase 9.30 or rows written by
        # non-5300 sources).
        used_per_code_balance = tot is not None
        if tot is None:
            tot = bal_lookup.get((d, pool))
        agg = by_yp.setdefault((yr, pool), {
            'amount': 0.0, 'total': 0.0, 'pct_sum': 0.0,
            'pct_weight': 0.0, 'pct_count': 0,
        })
        agg['amount'] += amt
        if tot is not None:
            if used_per_code_balance:
                # Per-code balance: sum unconditionally (each NCUA code
                # contributes its own distinct balance).
                agg['total'] += tot
            else:
                # Pool-level fallback: dedupe so codes resolving to the
                # same pool don't multiply the denominator.
                key = (d, pool)
                if key not in seen_pool_totals:
                    agg['total'] += tot
                    seen_pool_totals.add(key)
        if pct is not None:
            w = tot if tot is not None else 1.0
            agg['pct_sum'] += pct * w
            agg['pct_weight'] += w
            agg['pct_count'] += 1

    out: dict[int, dict[str, float]] = {}
    for (yr, pool), agg in by_yp.items():
        pct = None
        if agg['total'] > 0:
            pct = agg['amount'] / agg['total']
        elif agg['pct_count'] > 0:
            pct = (agg['pct_sum'] / agg['pct_weight']
                   if agg['pct_weight'] else None)
        if pct is None:
            continue
        out.setdefault(yr, {})[pool] = round(pct, 6)
    return out


def _select_validated_impaired_candidate(paths, snap_prefix):
    """From impaired-NAMED but date-less candidate files, return the best one
    that validates against the impaired workbook FORMAT, or ``None``.

    "Matches the format" = ``impaired_parser.parse_file`` returns ok AND the
    workbook exposes the Impairment-Type header structure (a data header row
    and/or recognised impairment types). When a candidate carries an internal
    Period Ending that resolves to a DIFFERENT ``YYYY-MM`` than the snapshot it
    is rejected (it belongs to another period). Ties break by: internal period
    matches the snapshot, then has data rows, then newest mtime.
    """
    uniq = list(dict.fromkeys(paths or []))
    if not uniq:
        return None
    try:
        from cecl_ui.services import impaired_parser
    except Exception:  # noqa: BLE001
        return None

    def _period_ym(parsed):
        pe = parsed.get('period_ending')
        if not pe:
            return None
        s = str(pe)
        m = re.search(r'(20\d{2})[-/](\d{1,2})', s)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}"
        try:
            d = pd.to_datetime(s, errors='coerce')
            if pd.notna(d):
                return f"{d.year:04d}-{d.month:02d}"
        except Exception:  # noqa: BLE001
            pass
        return None

    scored = []
    for p in uniq:
        try:
            parsed = impaired_parser.parse_file(p)
        except Exception:  # noqa: BLE001
            continue
        if not parsed.get('ok'):
            continue
        if not (parsed.get('data_header_row') or parsed.get('impairment_types')):
            continue  # doesn't match the impaired layout
        ym = _period_ym(parsed)
        if ym and snap_prefix and ym != snap_prefix:
            continue  # internal period says a different month
        try:
            mtime = os.path.getmtime(p)
        except OSError:
            mtime = 0.0
        period_match = 1 if (ym and ym == snap_prefix) else 0
        has_rows = 1 if (parsed.get('data_rows') or []) else 0
        scored.append((period_match, has_rows, mtime, p))
    if not scored:
        return None
    scored.sort(reverse=True)
    return scored[0][3]


def find_standalone_impaired_file(config, snap):
    """Locate the standalone Impaired Loans workbook for ``snap``.

    Shared helper used by :func:`load_standalone_impaired` at report
    time AND by the run-new-quarter impaired verification intercept in
    ``cecl_ui/services/impaired_check_service.verify_for_run`` — so both
    paths agree on which file is authoritative for a given period.

    Returns the absolute path (str) on success, or ``None`` when no
    matching file is found in ``config['data_directory']`` or the
    credit-pull fallback folder.
    """
    data_dir = config.get('data_directory', '')
    if not data_dir:
        return None
    if not os.path.isabs(data_dir):
        data_dir = os.path.join(BASE, data_dir)

    snap_prefix = snap[:7] if snap else ''  # e.g. "2026-03"

    # Search for the standalone impaired loans file. Filenames vary:
    #   "2026-03 Impaired Loans - Franklin Trust FCU.xlsx"
    #   "2025-06_CECL-Migration-Impaired_Loans_-_Honolulu_FD_FCU.xlsx"
    #   "Impaired Loans - 6-2026.xlsx"          (TCP CU 2026-06)
    #   "Impaired Loans - June 2026.xlsx"       (month-name variant)
    pattern = re.compile(
        rf'^{re.escape(snap_prefix[:4])}\s*[-_]?\s*{re.escape(snap_prefix[5:7])}'
        rf'.*Impaired[\s_-]+Loans.*\.xlsx$',
        re.IGNORECASE)

    # User-supplied override pattern(s) from config['impaired_loans']['file_pattern'].
    # Accepts either a single regex string or a list. When provided, files
    # matching any user pattern are returned immediately — bypasses the strict
    # date-prefix + loose-date matching below. This is the escape hatch for
    # CU-specific naming conventions (e.g. "impaired loans july 2026.xlsx"
    # for a June 2026 period, or Vizo IDLR-renamed drops that no longer carry
    # a date token in the filename).
    _user_impaired_rxs: list[re.Pattern[str]] = []
    _imp_cfg = config.get('impaired_loans') or {}
    if isinstance(_imp_cfg, dict):
        _user_pat = _imp_cfg.get('file_pattern')
        if _user_pat:
            _raw_pats = [_user_pat] if isinstance(_user_pat, str) else list(_user_pat)
            for _p in _raw_pats:
                if isinstance(_p, str) and _p.strip():
                    try:
                        _user_impaired_rxs.append(re.compile(_p, re.IGNORECASE))
                    except re.error:
                        continue

    _loose_impaired_rx = re.compile(r"Impaired[\s_\-]+Loans", re.IGNORECASE)
    _loose_date_rx = None
    if snap_prefix and len(snap_prefix) >= 7:
        try:
            _year = snap_prefix[:4]
            _month_num = int(snap_prefix[5:7])
            _month_names = {
                1: r"jan(?:uary)?", 2: r"feb(?:ruary)?", 3: r"mar(?:ch)?",
                4: r"apr(?:il)?", 5: r"may", 6: r"jun(?:e)?",
                7: r"jul(?:y)?", 8: r"aug(?:ust)?",
                9: r"sep(?:t(?:ember)?)?", 10: r"oct(?:ober)?",
                11: r"nov(?:ember)?", 12: r"dec(?:ember)?",
            }
            _mname = _month_names.get(_month_num, "")
            _alts = [
                rf"{_year}[-_.\s]*0?{_month_num}(?!\d)",
                rf"(?<!\d)0?{_month_num}[-_.\s]*{_year}",
            ]
            if _mname:
                _alts.append(rf"{_mname}[-_.\s]*{_year}")
                _alts.append(rf"{_year}[-_.\s]*{_mname}")
            _loose_date_rx = re.compile("|".join(_alts), re.IGNORECASE)
        except (ValueError, IndexError):
            _loose_date_rx = None

    search_dirs = [data_dir]
    fb_folder = config.get('credit_pull', {}).get('fallback_report_folder', '')
    if fb_folder and fb_folder != data_dir:
        if not os.path.isabs(fb_folder):
            fb_folder = os.path.join(BASE, fb_folder)
        search_dirs.append(fb_folder)
    # Also scan the CU's client source folder (``loan_source_folder``). Many
    # CUs' impaired-loans export never lands in ``data_directory``; it is
    # dropped straight into the period folder of the client drive (e.g. Vizo
    # IDLR: ".../CECL Migration IDLR/2026/June 2026/...Impaired Loans...xlsx").
    # ``data_directory`` (and any credit-pull fallback) is searched first, so a
    # dated/period-matching file there still wins; the source folder is only
    # walked when nothing there matched — which is exactly the case that
    # previously left the impaired section empty for these CUs.
    src_folder = config.get('loan_source_folder', '')
    if src_folder:
        if not os.path.isabs(src_folder):
            src_folder = os.path.join(BASE, src_folder)
        if src_folder not in search_dirs:
            search_dirs.append(src_folder)

    # Robust filename-date extractor (handles YYYY-MM, MMDDYYYY, YYYYMMDD,
    # MM-DD-YYYY, month names, ...). Used to match an impaired file to the
    # snapshot period once we've decided it IS an impaired file — so a CU
    # can rename its export freely as long as the words "Impaired Loans"
    # and a resolvable date remain in the filename.
    try:
        from import_data import _try_common_date_layouts as _file_iso
    except Exception:  # noqa: BLE001
        _file_iso = None

    # Extract a YYYY-MM period from a folder-path fragment (or None). Handles
    # "2026-06" / "2026_06" / "2026 06" and month-name variants in either
    # order ("June 2026" / "2026 June"). Used to bind a date-stripped impaired
    # export to the period folder it was dropped into, so a filename with no
    # date still resolves to the correct snapshot (and is skipped for other
    # periods on historical re-runs).
    _months_abbr = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5,
                    'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10,
                    'nov': 11, 'dec': 12}

    def _path_period(text):
        if not text:
            return None
        m = re.search(r'(20\d{2})[-_/.\s]{1,3}(0?[1-9]|1[0-2])(?!\d)', text)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}"
        m = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*'
                      r'[-_/.\s]*((?:19|20)\d{2})', text, re.IGNORECASE)
        if m:
            return f"{int(m.group(2)):04d}-{_months_abbr[m.group(1)[:3].lower()]:02d}"
        m = re.search(r'((?:19|20)\d{2})[-_/.\s]*'
                      r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*',
                      text, re.IGNORECASE)
        if m:
            return f"{int(m.group(1)):04d}-{_months_abbr[m.group(2)[:3].lower()]:02d}"
        return None

    # Impaired-NAMED files with NO resolvable date anywhere in the filename
    # are held here and resolved via a format-validated fallback after every
    # date-based attempt fails (so dated files always win for their period).
    _dateless_impaired_candidates: list[str] = []

    for sdir in search_dirs:
        if not os.path.isdir(sdir):
            continue
        for root, dirs, files in os.walk(sdir):
            for f in files:
                if f.startswith('~$') or f.upper().startswith('DNU'):
                    continue
                if not f.lower().endswith('.xlsx'):
                    continue
                # User-supplied file_pattern(s) take priority. When any pattern
                # matches, return immediately — the CU-level config is the
                # authoritative "which file is this period's impaired list".
                for _urx in _user_impaired_rxs:
                    if _urx.search(f):
                        return os.path.join(root, f)
                is_impaired_named = bool(_loose_impaired_rx.search(f))
                # Skip the full CECL-WARM template workbook — it's the
                # historical-data source, not the standalone impaired
                # list. BUT keep files that explicitly say "Impaired
                # Loans" even when they also carry "WARM": some CUs name
                # their impaired export "CECL-WARM with Credit Migration
                # Impaired Loans NEW <date>.xlsx".
                if 'WARM' in f.upper() and not is_impaired_named:
                    continue
                # Strict date-prefix pattern (fast path).
                if pattern.match(f):
                    return os.path.join(root, f)
                if not is_impaired_named:
                    continue
                # It's an impaired-loans file — now determine its period
                # from the filename. Try the loose period matcher first,
                # then fall back to the robust date extractor and compare
                # the YYYY-MM to the snapshot. This "identify the report,
                # then find the date" path survives CU naming-convention
                # changes (e.g. a trailing MMDDYYYY token like 06302026).
                if _loose_date_rx is not None and _loose_date_rx.search(f):
                    return os.path.join(root, f)
                _iso = (_file_iso(f)
                        if (_file_iso is not None and snap_prefix) else None)
                if _iso and _iso[:7] == snap_prefix:
                    return os.path.join(root, f)
                if not _iso:
                    # No date in the FILENAME. Some CUs drop a date-stripped
                    # export into a period-named folder (e.g. Vizo IDLR:
                    # ".../2026/June 2026/CECL-WARM ... Impaired Loans NEW.xlsx").
                    # Bind the file to the period encoded in its containing
                    # folder path when present: an exact snapshot-period folder
                    # wins immediately; a DIFFERENT-period folder is skipped so
                    # historical re-runs never grab the wrong month's file.
                    try:
                        _rel_dir = os.path.relpath(root, sdir)
                    except ValueError:
                        _rel_dir = root
                    _folder_ym = _path_period(_rel_dir)
                    if _folder_ym and snap_prefix:
                        if _folder_ym == snap_prefix:
                            return os.path.join(root, f)
                        continue  # folder path says a different period
                    # No date in filename OR folder path — hold as a candidate
                    # for the format-validated fallback below (handles CUs
                    # that dropped the date token, e.g. NOVA's
                    # "TCT CECL-WARM ... Impaired Loans NEW.xlsx").
                    _dateless_impaired_candidates.append(
                        os.path.join(root, f))

    # FINAL fallback: nothing matched the snapshot period by filename date.
    # Validate any date-less impaired-NAMED candidate against the impaired
    # workbook FORMAT and return the best match, so a CU can rename/date-strip
    # its export as long as the words "Impaired Loans" remain and the layout
    # is intact.
    return _select_validated_impaired_candidate(
        _dateless_impaired_candidates, snap_prefix)


def load_standalone_impaired(config, snap, df=None):
    """Load impaired-loan data from the standalone Impaired Loans file.

    Searches for files matching patterns like:
      "2026-03 Impaired Loans - Franklin Trust FCU.xlsx"
      "2026- 03 Impaired Loans - Franklin Trust FCU.xlsx"

    Returns dict with:
      'acl_impaired': {category: provision_amount, ...}
      'spec_id_by_pool': {pool: {grade: balance_removed, ...}, ...}
      'total_spec_id': float
    or empty dict if file/tab not found.
    """
    data_dir = config.get('data_directory', '')
    if not data_dir:
        return {}
    if not os.path.isabs(data_dir):
        data_dir = os.path.join(BASE, data_dir)

    cu = config['credit_union']
    snap_prefix = snap[:7] if snap else ''  # e.g. "2026-03"
    pool_map = config.get('pool_map', {})
    default_pool = config.get('default_pool', 'Other/Uncategorized')

    # Search for the standalone impaired loans file. Filenames vary:
    #   "2026-03 Impaired Loans - Franklin Trust FCU.xlsx"
    #   "2025-06_CECL-Migration-Impaired_Loans_-_Honolulu_FD_FCU.xlsx"
    #   "Impaired Loans - 6-2026.xlsx"          (TCP CU 2026-06)
    #   "Impaired Loans - June 2026.xlsx"       (month-name variant)
    # Delegated to :func:`find_standalone_impaired_file` so the wizard's
    # run-time verification page and this report-time loader agree on
    # which file is authoritative.
    found = find_standalone_impaired_file(config, snap)

    if not found:
        return {}

    print(f"    Loading standalone impaired loans from: {os.path.basename(found)}")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(found, data_only=True)
        # Tab name varies — may have leading/trailing whitespace
        # (e.g. ' Impaired Loans' in CECL-Migration workbooks). Some
        # CUs also rename the tab entirely (e.g. TCP 2026-06 uses
        # 'Sheet2' with the "Impaired Loans" title in cell A1). Match
        # loosely: prefer a tab named "impaired loans", then any tab
        # whose name mentions "impaired", then any tab whose A1/A2
        # value contains "impaired loans", then the largest tab.
        ws = None
        for sn in wb.sheetnames:
            if sn.strip().lower() == 'impaired loans':
                ws = wb[sn]
                break
        if ws is None:
            for sn in wb.sheetnames:
                if 'impaired' in sn.strip().lower():
                    ws = wb[sn]
                    print(f"    (auto-matched tab {sn!r} via loose name)")
                    break
        if ws is None:
            for sn in wb.sheetnames:
                _ws = wb[sn]
                for _r in range(1, 4):
                    _v = _ws.cell(row=_r, column=1).value
                    if _v and 'impaired' in str(_v).lower():
                        ws = _ws
                        print(f"    (auto-matched tab {sn!r} via header cell A{_r})")
                        break
                if ws is not None:
                    break
        if ws is None:
            # Last resort: pick the largest data-bearing tab (skip
            # obvious instructions/help tabs).
            _candidates = [
                (wb[sn].max_row * wb[sn].max_column, sn)
                for sn in wb.sheetnames
                if 'instruction' not in sn.strip().lower()
                and 'help' not in sn.strip().lower()
                and 'readme' not in sn.strip().lower()
            ]
            _candidates.sort(reverse=True)
            if _candidates:
                _sz, _sn = _candidates[0]
                ws = wb[_sn]
                print(f"    (auto-matched tab {_sn!r} as largest data tab)")
        if ws is None:
            raise KeyError("No 'Impaired Loans' sheet (got: "
                           f"{wb.sheetnames!r})")
    except (KeyError, Exception) as exc:
        print(f"    Error reading impaired loans file: {exc}")
        return {}

    # ── Parse summary section ──
    # Layout varies: locate the header row by scanning the first ~22
    # rows for a cell whose value contains "Impairment Type" in the
    # right-hand columns (the left side may have a separate
    # "Impairment Type/Provision Percentage" definitions block).
    # Then find the columns labelled "Impairment Type" and "Sum of
    # Provision Amount" (or just "Provision Amount").
    acl_impaired = {}
    hdr_row = None
    type_col = prov_col = None
    for r in range(1, 23):
        for c in range(1, min(ws.max_column, 30) + 1):
            v = ws.cell(row=r, column=c).value
            if not v:
                continue
            s = str(v).strip().lower()
            if s == 'impaired type' or s == 'impairment type':
                # Prefer the right-hand "Impairment Type" header (the
                # one that has Provision Amount in the same row).
                # Verify by scanning that row for a Provision header.
                row_vals = [ws.cell(row=r, column=cc).value
                            for cc in range(1, ws.max_column + 1)]
                for cc, vv in enumerate(row_vals, start=1):
                    if vv and 'provision amount' in str(vv).lower():
                        hdr_row = r
                        type_col = c
                        prov_col = cc
                        break
                if hdr_row is not None:
                    break
        if hdr_row is not None:
            break

    if hdr_row is None:
        # Fall back to the original Franklin-Trust layout
        hdr_row = 4
        type_col = 12
        prov_col = 16

    for row in range(hdr_row + 1, hdr_row + 17):
        cat = ws.cell(row=row, column=type_col).value
        prov = ws.cell(row=row, column=prov_col).value
        if not cat or str(cat).strip().upper() in ('', 'HIDE', 'TOTAL',
                                                    'CALCULATION', 'IMPAIRMENT TYPE'):
            continue
        cat_str = str(cat).strip()
        try:
            prov_val = float(prov) if prov else 0.0
        except (ValueError, TypeError):
            prov_val = 0.0
        acl_impaired[cat_str] = prov_val

    # ── Build member→grade lookup from df ──
    # The df's member_number column is the concatenated (member+suffix)
    # form stored by import_data.derive_member_account — the exact string
    # width depends on the CU's raw file (e.g. Nucor stores an 11-char
    # zero-padded string "00001055701" = 9-char member + 2-char suffix).
    # The impaired workbook, on the other hand, ships bare integer
    # member/suffix values. To bridge the two formats we store the DB
    # key AND several normalized variants (leading-zero-stripped, and
    # int(member)+int(suffix) with the DB-detected total width).
    grade_lookup = {}  # {member_suffix_str: grade}
    pool_lookup = {}   # {member_suffix_str: loan_pool} — AIRES-matched pool
    _db_key_len = 0
    _has_pool_col = df is not None and 'loan_pool' in df.columns
    if df is not None and 'member_number' in df.columns and 'current_grade' in df.columns:
        for _, r in df.iterrows():
            mem = str(r['member_number']).strip()
            grade = r['current_grade']
            if not mem:
                continue
            grade_lookup[mem] = grade
            # Pull the pool the loan actually sits in from the extract so
            # impaired rows are removed from the SAME pool the loan is
            # reserved in (rather than re-deriving a pool from the impaired
            # file's own loan-type code, which can use a different taxonomy).
            _pool_val = ''
            if _has_pool_col:
                _pv = r['loan_pool']
                if pd.notna(_pv):
                    _pool_val = str(_pv).strip()
            if _pool_val:
                pool_lookup[mem] = _pool_val
            # Some CUs store the member key with a non-numeric separator
            # (e.g. SCI's "1718L18" = member 1718 + suffix 18). Index a
            # digits-only variant so the impaired parser's numeric
            # member+suffix keys ("171818") still find a hit.
            _digits = re.sub(r'\D+', '', mem)
            if _digits and _digits != mem:
                grade_lookup.setdefault(_digits, grade)
                if _pool_val:
                    pool_lookup.setdefault(_digits, _pool_val)
            # Also store the leading-zero-stripped variant so bare
            # int(member)+int(suffix) forms from the impaired workbook
            # can find a hit even when the DB pads the concatenated key.
            try:
                stripped = str(int(mem))
                if stripped != mem:
                    grade_lookup.setdefault(stripped, grade)
                    if _pool_val:
                        pool_lookup.setdefault(stripped, _pool_val)
            except (TypeError, ValueError):
                pass
            if not _db_key_len:
                _db_key_len = len(mem)

    # ── Parse detail rows (row 24+) ──
    spec_id_by_pool = {}  # {pool: {grade: balance_removed}}
    total_removed = 0.0
    suffix_len = config.get('account_suffix_length', 3)
    for row in range(24, ws.max_row + 1):
        imp_type = ws.cell(row=row, column=1).value
        if not imp_type or str(imp_type).strip() in ('', 'HIDE'):
            continue
        balance = ws.cell(row=row, column=5).value
        if not balance:
            continue

        member = ws.cell(row=row, column=2).value
        suffix = ws.cell(row=row, column=3).value
        loan_type = ws.cell(row=row, column=4).value
        removed = ws.cell(row=row, column=17).value  # Balance Removed

        try:
            removed_val = float(removed) if removed else 0.0
        except (ValueError, TypeError):
            removed_val = 0.0
        if removed_val <= 0:
            # "Bal Removed from Pools" is typically a formula (=E{row}, i.e.
            # the Current Balance). openpyxl reads data_only formulas as None
            # when the workbook was never recalculated in Excel, so fall back
            # to Current Balance: an individually-evaluated impaired loan has
            # its whole balance carved out of the pool.
            try:
                removed_val = float(balance) if balance else 0.0
            except (ValueError, TypeError):
                removed_val = 0.0
        if removed_val <= 0:
            continue

        # Resolve the pool by MATCHING the impaired loan against the loan
        # extract (member+suffix), so the balance is removed from the pool
        # the loan is actually reserved in. The impaired file's own
        # loan-type code is only a FALLBACK for rows that don't match any
        # extract record (its codes may use a different taxonomy than the
        # top-level pool_map — e.g. an AIRES code that resolves to a
        # different pool). Grade comes from the same matched extract row.
        lt = str(loan_type).strip() if loan_type else ''
        _code_pool = pool_map.get(lt, default_pool)

        grade = ''
        matched_pool = None
        if member is not None and suffix is not None:
            try:
                mem_int = int(float(member))
                suf_int = int(float(suffix))
            except (TypeError, ValueError):
                # Spreadsheet placeholders like 'xxxx' or blanks — skip
                # extract lookup for this row rather than crashing the
                # whole report. The balance still aggregates into the
                # fallback pool below.
                mem_int = suf_int = None
            if mem_int is not None:
                # Try multiple key variants because YAML's suffix_length
                # (e.g. Nucor's top-level account_suffix_length: 0 while
                # the raw source actually has a 2-digit padded suffix)
                # frequently disagrees with the DB's concatenated form.
                # We test the configured width AND every plausible
                # suffix-padding width; the first key present in the
                # extract lookup (which stores both the padded DB key and
                # a leading-zero-stripped variant) wins.
                _candidate_widths = [suffix_len, 0, 1, 2, 3, 4, 5]
                _tried: set[str] = set()
                _hit_key = None
                for _w in _candidate_widths:
                    if _w < 0:
                        continue
                    if _w == 0:
                        mem_key = f"{mem_int}{suf_int}"
                    else:
                        mem_key = f"{mem_int}{suf_int:0{_w}d}"
                    if mem_key in _tried:
                        continue
                    _tried.add(mem_key)
                    if mem_key in grade_lookup:
                        _hit_key = mem_key
                        break
                # As a final safety net, if the DB uses a fixed total
                # key length (e.g. 11-char zero-padded), try matching
                # against that width by zero-padding the member side
                # for each candidate suffix width.
                if _hit_key is None and _db_key_len:
                    for _sw in (2, 3, 1, 4):
                        _mw = _db_key_len - _sw
                        if _mw <= 0:
                            continue
                        mem_key = f"{mem_int:0{_mw}d}{suf_int:0{_sw}d}"
                        if mem_key in _tried:
                            continue
                        _tried.add(mem_key)
                        if mem_key in grade_lookup:
                            _hit_key = mem_key
                            break
                if _hit_key is not None:
                    grade = grade_lookup.get(_hit_key, '') or ''
                    matched_pool = pool_lookup.get(_hit_key)

        # Matched extract pool wins; otherwise fall back to the impaired
        # file's loan-type code mapping.
        pool = matched_pool if matched_pool else _code_pool

        spec_id_by_pool.setdefault(pool, {})
        spec_id_by_pool[pool][grade] = (
            spec_id_by_pool[pool].get(grade, 0) + removed_val
        )
        total_removed += removed_val

    wb.close()

    # ── Distribute "unknown grade" bucket proportionally ──
    # When the impaired file has sanitized member#/suffix values
    # (e.g. 'xxxx' placeholders) the per-row grade lookup returns ''
    # and the balance lands under an empty-grade key. Without this
    # redistribution the ACL Env by Pool tab only renders the pool
    # Total row, leaving every per-grade column at $0. Fan the
    # unknown-grade bucket across the actual grade rows using each
    # pool's current_balance distribution in df. If df has no
    # balance for the pool (e.g. excluded pools), the unknown bucket
    # is left in place so it still contributes to pool totals.
    if (df is not None and not df.empty
            and 'loan_pool' in df.columns
            and 'current_grade' in df.columns
            and 'current_balance' in df.columns):
        redistributed_total = 0.0
        redistributed_pools = 0
        for pool, grade_map in list(spec_id_by_pool.items()):
            unknown_amt = grade_map.get('', 0)
            if not unknown_amt:
                continue
            pool_df = df[df['loan_pool'] == pool]
            if pool_df.empty:
                continue
            grade_bals = pool_df.groupby('current_grade')['current_balance'].sum()
            grade_bals = grade_bals[grade_bals > 0]
            total_bal = float(grade_bals.sum())
            if total_bal <= 0:
                continue
            # Remove the unknown bucket and redistribute
            del grade_map['']
            for g, bal in grade_bals.items():
                share = unknown_amt * (float(bal) / total_bal)
                grade_map[g] = grade_map.get(g, 0) + share
            redistributed_total += unknown_amt
            redistributed_pools += 1
        if redistributed_pools:
            print(f"    Distributed ${redistributed_total:,.2f} of unknown-grade "
                  f"Spec ID across {redistributed_pools} pool(s) using current "
                  f"pool grade-balance mix")

    result = {
        'acl_impaired': acl_impaired,
        'spec_id_by_pool': spec_id_by_pool,
        'total_spec_id': sum(acl_impaired.values()),
    }

    imp_count = sum(1 for _ in acl_impaired.values() if _ > 0)
    n_pools = len(spec_id_by_pool)
    print(f"    Impaired categories: {len(acl_impaired)} "
          f"({imp_count} with provision), "
          f"Spec ID: {n_pools} pools, "
          f"Total removed: ${total_removed:,.2f}")

    return result


# ── Styling Helpers ────────────────────────────────────────────────
def hdr_row(ws, row, ncol, fill=None):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = fill or HDR_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN

def style_rows(ws, r1, r2, ncol, mcols=(), pcols=()):
    for r in range(r1, r2 + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = NORM
            cell.border = THIN
            if (r - r1) % 2 == 1:
                cell.fill = ALT_FILL
            if c in mcols:
                cell.number_format = MONEY
            elif c in pcols:
                cell.number_format = PCT

def auto_w(ws, ncol, mn=10, mx=25):
    for c in range(1, ncol + 1):
        lt = get_column_letter(c)
        best = mn
        for row in ws.iter_rows(min_col=c, max_col=c):
            for cell in row:
                if cell.value:
                    best = max(best, min(len(str(cell.value)) + 2, mx))
        ws.column_dimensions[lt].width = best

def write_title(ws, row, text_val, col=1):
    ws.cell(row=row, column=col, value=text_val).font = TITLE_FONT

def write_sub(ws, row, text_val, col=1):
    ws.cell(row=row, column=col, value=text_val).font = SUB_FONT


# ── Calculation Helpers ────────────────────────────────────────────
def calc_net_credit_change(df):
    """Return improved%, deteriorated%, net_change% for a DataFrame."""
    total = df['current_balance'].sum()
    if total == 0:
        return 0, 0, 0
    imp = df[df['migration_status'] == 'Improved']['current_balance'].sum() / total
    det = df[df['migration_status'] == 'Deteriorated']['current_balance'].sum() / total
    return imp, det, imp - det

def calc_economic_stress(config):
    """Calculate Economic Stress Index from config economic data."""
    ed = config.get('economic_data', {})
    unemp = ed.get('unemployment_rate', 0) * 100
    pop = ed.get('population', 1)
    bk_pct = (ed.get('bankruptcies', 0) / pop) * 100 if pop else 0
    fc_pct = (ed.get('foreclosures', 0) / pop) * 100 if pop else 0
    return unemp + bk_pct + fc_pct

def calc_env_factor_pool(ncc_pct, dq_variance, econ_stress):
    """Calculate environmental factor for a pool."""
    ncc_score = score_from_ranges(ncc_pct * 100, NCC_RANGES)
    dq_score = score_from_ranges(dq_variance * 100, DQ_RANGES)
    es_score = score_from_ranges(econ_stress, ES_RANGES)
    return ncc_score + dq_score + es_score

def get_acl_base_rates(grades, config):
    """Get ACL base loss rates per grade. Uses config reserve_rates as fallback."""
    return {g['label']: g['reserve_rate'] for g in grades}

def get_dist_factor(grade_idx, num_grades):
    """Get distribution factor for a grade position."""
    if grade_idx < len(DIST_FACTORS):
        return DIST_FACTORS[grade_idx] / 100.0
    return DIST_FACTORS[-1] / 100.0


# ── Admin-default + per-pool management-adjustment resolver ──────────
def _load_admin_default_mgmt_adj():
    """Read firm-wide default from admin_defaults.yaml. 0.0 on error."""
    try:
        import yaml as _yaml
        from pathlib import Path as _Path
        p = _Path(__file__).resolve().parent / 'admin_defaults.yaml'
        if not p.exists():
            return 0.0
        data = _yaml.safe_load(p.read_text(encoding='utf-8')) or {}
        return float(data.get('default_mgmt_adj', 0.0) or 0.0)
    except Exception:
        return 0.0


def _build_pool_use_default_map(config):
    """Return ``{pool_name: bool}`` from ``config['pools']``."""
    out = {}
    for p in (config.get('pools') or []):
        name = (p.get('name') or '').strip()
        if name:
            out[name] = bool(p.get('use_default_mgmt_adj'))
    return out


def _resolve_mgmt_adj_grade(pool, grade_label, grade_idx, num_grades,
                             pool_use_default, mgmt_adj_by_pool,
                             admin_default, prior_mgmt_adj_map,
                             base_rate=None):
    """Mirror of report_tct's resolver for the legacy generate_report
    path. Precedence: manual×dist > admin×dist (only when use_default
    AND no manual AND base_rate==0) > prior (carry-forward fallback,
    only when no current-period adjustment is established) > 0.
    """
    dist = get_dist_factor(grade_idx, num_grades)
    manual = mgmt_adj_by_pool.get(pool, 0) or 0
    if manual:
        return float(manual) * dist
    if (pool_use_default.get(pool, False)
            and admin_default
            and (base_rate is None or float(base_rate or 0) == 0)):
        return float(admin_default) * dist
    # Carry-forward fallback: prior period's per-grade value applies only
    # when the current period has no established adjustment above.
    pm = prior_mgmt_adj_map.get(pool, {}) if prior_mgmt_adj_map else {}
    if grade_label in pm:
        return pm[grade_label]
    return 0.0


# ══════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ══════════════════════════════════════════════════════════════════

def sheet_cover_tct(wb, cu, snap):
    """Cover sheet - TCT/Franklin Trust style."""
    ws = wb.active
    ws.title = "Cover"
    ws['B4'] = "RISK BASED PRICING"
    ws['B4'].font = Font(name='Calibri', bold=True, size=22)
    ws['B6'] = "ACL/Credit Migration Report"
    ws['B6'].font = Font(name='Calibri', bold=True, size=16)
    items = ["CECL Compliant", "Risk Change by Type",
             "Improved/Deteriorated Loan Analysis", "Environmental Factor",
             "Allowance for Credit Loss (ACL)", "Summary of Deteriorated Loans"]
    for i, item in enumerate(items):
        ws.cell(row=8 + i, column=2, value=item).font = Font(name='Calibri', size=12)
    ws['B16'] = "Prepared For:"
    ws['B16'].font = Font(name='Calibri', size=12, italic=True)
    ws['B18'] = cu
    ws['B18'].font = Font(name='Calibri', bold=True, size=20)
    ws['B21'] = f"For Period Ending"
    ws['B22'] = snap
    ws['B22'].font = Font(name='Calibri', bold=True, size=14)
    ws['B25'] = "Presented by:"
    ws['B27'] = "TCT Risk Solutions"
    ws['B27'].font = Font(name='Calibri', bold=True, size=16, color='C0392B')
    ws['B28'] = "Take Charge Today"
    ws['B28'].font = Font(name='Calibri', italic=True, size=12, color='2E86C1')
    dt = datetime.now().strftime('%B %d, %Y')
    ws['B30'] = f"Report Generated: {dt}"
    ws['B30'].font = Font(name='Calibri', size=10, color='888888')


def sheet_cover_vizo(wb, cu, snap, supplemental=False):
    """Cover sheet - Vizo/Credit Union B style."""
    ws = wb.active
    ws.title = "Cover"
    ws['B6'] = "CECL Credit Migration Report"
    ws['B6'].font = Font(name='Calibri', bold=True, size=20)
    if supplemental:
        ws['B7'] = "Supplemental Reports"
        ws['B7'].font = Font(name='Calibri', bold=True, size=16)
    ws['B10'] = cu
    ws['B10'].font = Font(name='Calibri', bold=True, size=18)
    ws['B12'] = snap
    ws['B12'].font = Font(name='Calibri', bold=True, size=14)
    ws['B16'] = "TCT Risk Solutions"
    ws['B16'].font = Font(name='Calibri', bold=True, size=14, color='C0392B')
    ws['B20'] = "All reports are confidential."
    ws['B20'].font = Font(name='Calibri', size=10, italic=True, color='888888')


def sheet_report_overview(wb, report_type="main"):
    """Report Overview / Index page (Vizo style)."""
    ws = wb.create_sheet("Report Overview")
    ws['A2'] = "Report Overview"
    ws['A2'].font = Font(name='Calibri', bold=True, size=16)
    ws['A4'] = ("The CECL Credit Migration Reports from TCT, Inc. presents a comprehensive picture "
                "of the changing nature of risk in the credit union's loan portfolio.")
    ws['A4'].font = NORM
    ws['A4'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('A4:H6')

    ws['A8'] = "Report Index:"
    ws['A8'].font = Font(name='Calibri', bold=True, size=14, color='1B4F72')

    if report_type == "main":
        sections = [
            ("Executive Summary", ["CECL Adjustment & Improved/Deteriorated",
                                    "Improved & Deteriorated Loans Risk Change By Credit Score"]),
            ("Detailed Reporting", ["Allowance & Provision for Credit Loss Reserve Analysis",
                                    "Risk Change by Credit Score - Total Loans",
                                    "Risk Change by Credit Score - Loan Pools",
                                    "Environmental Factor Provision for Loan Loss",
                                    "Loss Factor Calculation", "Delinquency Calculation"]),
        ]
    else:
        sections = [
            ("Supplemental Reporting Package", [
                "Historical Loan Balances by Credit Score",
                "Loss Factor Historical Detail",
                "Charge off and Recoveries Historical Detail",
                "Balance Adjustment Detail"]),
        ]

    r = 10
    for title, items in sections:
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=12, color='1B4F72')
        for item in items:
            r += 1
            ws.cell(row=r, column=2, value=item).font = NORM
        r += 2


def sheet_exec_summary(wb, cu, snap, df, grades, config):
    """Executive Summary – works for both TCT and Vizo formats."""
    ws = wb.create_sheet("Executive Summary")
    ws['A1'] = cu
    ws['A1'].font = TITLE_FONT
    ws['A2'] = "Executive Summary"
    ws['A2'].font = Font(bold=True, size=14)
    ws['A3'] = f"For Period Ending {snap}"
    ws['A3'].font = Font(size=12, color='555555')

    no_score = config.get('no_score_label', 'Not Reported')
    total = df['current_balance'].sum()
    imp_bal = df[df['migration_status'] == 'Improved']['current_balance'].sum()
    det_bal = df[df['migration_status'] == 'Deteriorated']['current_balance'].sum()
    unc_bal = df[df['migration_status'] == 'Unchanged']['current_balance'].sum()
    ncc = (imp_bal - det_bal) / total * 100 if total else 0
    total_reserve = df['expected_loss_amount'].sum()

    # CECL Adjustment box
    r = 5
    ws.cell(row=r, column=1, value="CECL Adjustment").font = SUB_FONT
    for label, val in [("Total Specifically Identified Allowance", 0),
                       ("Total Allowance Needed", total_reserve),
                       ("Allowance for Credit Loss Balance", 0),
                       ("Adjustment (Overfunded)", total_reserve)]:
        r += 1
        ws.cell(row=r, column=1, value=label).font = NORM
        ws.cell(row=r, column=3, value=val).number_format = MONEY

    # Improved/Deteriorated summary by grade
    r += 2
    ws.cell(row=r, column=1, value="Improved Loans Summary").font = SUB_FONT
    r += 1
    ws.cell(row=r, column=1, value="Grade"); ws.cell(row=r, column=2, value="Balance")
    hdr_row(ws, r, 2)
    grade_labels = [g['label'] for g in grades]
    imp_df = df[df['migration_status'] == 'Improved']
    for gl in grade_labels:
        r += 1
        g_bal = imp_df[imp_df['current_grade'] == gl]['current_balance'].sum()
        ws.cell(row=r, column=1, value=gl)
        ws.cell(row=r, column=2, value=g_bal).number_format = MONEY
    r += 1
    ws.cell(row=r, column=1, value="Total Improved").font = Font(bold=True)
    ws.cell(row=r, column=2, value=imp_bal).number_format = MONEY

    r += 2
    ws.cell(row=r, column=1, value="Deteriorated Loans Summary").font = SUB_FONT
    r += 1
    ws.cell(row=r, column=1, value="Grade"); ws.cell(row=r, column=2, value="Balance")
    hdr_row(ws, r, 2)
    det_df = df[df['migration_status'] == 'Deteriorated']
    for gl in grade_labels:
        r += 1
        g_bal = det_df[det_df['current_grade'] == gl]['current_balance'].sum()
        ws.cell(row=r, column=1, value=gl)
        ws.cell(row=r, column=2, value=g_bal).number_format = MONEY
    r += 1
    ws.cell(row=r, column=1, value="Total Impaired").font = Font(bold=True)
    ws.cell(row=r, column=2, value=det_bal).number_format = MONEY

    # Net Credit Change per pool
    r += 2
    ws.cell(row=r, column=1, value="Improved/Deteriorated by Pool").font = SUB_FONT
    r += 1
    for hdr_i, h in enumerate(["Pool", "Improved %", "Deteriorated %", "Net Change %"]):
        ws.cell(row=r, column=1 + hdr_i, value=h)
    hdr_row(ws, r, 4)
    pools = sorted(df['loan_pool'].unique())
    for pool in pools:
        r += 1
        pdf = df[df['loan_pool'] == pool]
        imp_p, det_p, net_p = calc_net_credit_change(pdf)
        ws.cell(row=r, column=1, value=pool)
        ws.cell(row=r, column=2, value=imp_p).number_format = PCT
        ws.cell(row=r, column=3, value=det_p).number_format = PCT
        ws.cell(row=r, column=4, value=net_p).number_format = PCT
    r += 1
    ws.cell(row=r, column=1, value="Grand Total").font = Font(bold=True)
    imp_t, det_t, net_t = calc_net_credit_change(df)
    ws.cell(row=r, column=2, value=imp_t).number_format = PCT
    ws.cell(row=r, column=3, value=det_t).number_format = PCT
    ws.cell(row=r, column=4, value=net_t).number_format = PCT

    auto_w(ws, 4)


def sheet_risk_change_all(wb, cu, snap, df, grades, config, hist=None):
    """Risk Change By Credit Score - Grand Total (dollar + percent matrices)."""
    ws = wb.create_sheet("Risk Change-All Loans")
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]
    matrix = risk_change_matrix(df, grades, no_score)

    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Risk Change By Credit Score"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A3'] = f"For Period Ending {snap}"

    # Score range map (supports optional display overrides per client)
    range_overrides = config.get('risk_change_range_labels', {})
    rng = {g['label']: range_overrides.get(g['label'], f"{g['min_score']}-{g['max_score']}") for g in grades}
    rng[no_score] = ""

    gray_fill = PatternFill('solid', fgColor='D9D9D9')
    gray_hdr_font = Font(name='Calibri', bold=True, size=10, color='000000')

    def apply_gray_header(row_num):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font = gray_hdr_font
            cell.fill = gray_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = THIN

    def apply_plain_rows(r1, r2):
        for rr in range(r1, r2 + 1):
            for cc in range(1, ncol + 1):
                cell = ws.cell(row=rr, column=cc)
                if not cell.font or not cell.font.bold:
                    cell.font = NORM
                cell.border = THIN

    # ─── Dollar Matrix ───
    r = 5
    ws.cell(row=r, column=1, value="Dollar")
    ws.cell(row=r, column=4, value="Original Credit Grade")
    r += 1
    ws.cell(row=r, column=1, value="Current Credit Grade")
    ws.cell(row=r, column=2, value="")
    for j, g in enumerate(gl):
        ws.cell(row=r, column=3 + j, value=g)
    ws.cell(row=r, column=ncol, value="Grand Total")
    ncol = 3 + len(gl)
    apply_gray_header(r)

    start = r + 1
    for i, g in enumerate(gl):
        r += 1
        ws.cell(row=r, column=1, value=g)
        ws.cell(row=r, column=2, value=rng.get(g, ''))
        rtotal = 0
        for j, og in enumerate(gl):
            v = matrix.loc[g, og] if g in matrix.index and og in matrix.columns else 0
            ws.cell(row=r, column=3 + j, value=v).number_format = MONEY
            rtotal += v
            if j < i:
                ws.cell(row=r, column=3 + j).fill = DET_FILL
            elif j > i:
                ws.cell(row=r, column=3 + j).fill = IMP_FILL
        ws.cell(row=r, column=ncol, value=rtotal).number_format = MONEY
    # Total row
    r += 1
    ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
    for j, og in enumerate(gl):
        col_total = sum(matrix.loc[g2, og] for g2 in gl if g2 in matrix.index and og in matrix.columns)
        ws.cell(row=r, column=3 + j, value=col_total).number_format = MONEY
    ws.cell(row=r, column=ncol, value=df['current_balance'].sum()).number_format = MONEY
    apply_plain_rows(start, r)

    # Balance adjustments row
    r += 1
    ws.cell(row=r, column=1, value="Loans Not Risk Rated and Adjustments")
    _imp = hist.get('impaired', {}) if hist else {}
    _tba = _imp.get('total_balance_adjustment', 0.0)
    ws.cell(row=r, column=ncol, value=_tba).number_format = MONEY
    r += 1
    ws.cell(row=r, column=1, value="Total in Portfolio").font = Font(bold=True)
    _tip = _imp.get('total_in_portfolio', df['current_balance'].sum() + _tba)
    if ncol > 3:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncol - 1)
        ws.cell(row=r, column=2, value="Section 3 Allowance for Credit Loss Calculation")
        ws.cell(row=r, column=2).font = Font(bold=True)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=ncol, value=_tip).number_format = MONEY
    apply_plain_rows(r - 1, r)

    # ─── Percent Matrix ───
    r += 2
    ws.cell(row=r, column=1, value="Percent")
    ws.cell(row=r, column=4, value="Original Credit Grade")
    r += 1
    ws.cell(row=r, column=1, value="Current Credit Grade")
    for j, g in enumerate(gl):
        ws.cell(row=r, column=3 + j, value=g)
    ws.cell(row=r, column=ncol, value="Grand Total")
    hdr_row(ws, r, ncol)

    total = df['current_balance'].sum()
    start2 = r + 1
    for i, g in enumerate(gl):
        r += 1
        ws.cell(row=r, column=1, value=g)
        ws.cell(row=r, column=2, value=rng.get(g, ''))
        rtotal = 0
        for j, og in enumerate(gl):
            v = matrix.loc[g, og] if g in matrix.index and og in matrix.columns else 0
            col_total = sum(matrix.loc[g2, og] for g2 in gl if g2 in matrix.index and og in matrix.columns)
            pct = v / col_total if col_total else 0
            ws.cell(row=r, column=3 + j, value=pct).number_format = PCT
            rtotal += v
        ws.cell(row=r, column=ncol, value=rtotal / total if total else 0).number_format = PCT
    r += 1
    ws.cell(row=r, column=1, value="Grand Total").font = Font(bold=True)
    for j in range(len(gl)):
        ws.cell(row=r, column=3 + j, value=1.0).number_format = PCT
    ws.cell(row=r, column=ncol, value=1.0).number_format = PCT
    style_rows(ws, start2, r, ncol, pcols=set(range(3, ncol + 1)))

    # ─── Net Credit Change box ───
    r += 2
    imp_bal = df[df['migration_status'] == 'Improved']['current_balance'].sum()
    det_bal = df[df['migration_status'] == 'Deteriorated']['current_balance'].sum()
    unc_bal = df[df['migration_status'] == 'Unchanged']['current_balance'].sum()
    for lbl, vd, vp in [
        ("Total-Improved", imp_bal, imp_bal / total if total else 0),
        ("Total-Deteriorated", det_bal, det_bal / total if total else 0),
        ("Total Unchanged", unc_bal, unc_bal / total if total else 0),
        ("Total In Portfolio", total, 1.0),
    ]:
        r += 1
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=vd).number_format = MONEY
        ws.cell(row=r, column=3, value=vp).number_format = PCT
        r += 1
    ws.cell(row=r, column=1, value="Net Change").font = Font(bold=True, size=12)
    ws.cell(row=r, column=2, value=imp_bal - det_bal).number_format = MONEY
    ws.cell(row=r, column=3, value=(imp_bal - det_bal) / total if total else 0).number_format = PCT

    auto_w(ws, ncol)


def sheet_impdet_summary(wb, cu, snap, df):
    """Analysis of Improved/Deteriorated Summary - all pools."""
    ws = wb.create_sheet("Improved-Deteriorated Summary")
    ws['A1'] = cu
    ws['A1'].font = TITLE_FONT
    ws['A2'] = "Analysis of Improved/Deteriorated Summary"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(df['loan_pool'].unique())
    r = 5
    for pool in pools:
        pdf = df[df['loan_pool'] == pool]
        ptotal = pdf['current_balance'].sum()
        ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=11)
        r += 1
        for h_i, h in enumerate(["", "$", "%"]):
            ws.cell(row=r, column=1 + h_i, value=h)
        hdr_row(ws, r, 3)
        r += 1
        for status in ['Improved', 'Deteriorated', 'Unchanged']:
            bal = pdf[pdf['migration_status'] == status]['current_balance'].sum()
            pct = bal / ptotal if ptotal else 0
            lbl = f"Total-{status}" if status != 'Unchanged' else "Total Unchanged"
            ws.cell(row=r, column=1, value=lbl)
            ws.cell(row=r, column=2, value=bal).number_format = MONEY
            ws.cell(row=r, column=3, value=pct).number_format = PCT
            if status == 'Improved':
                ws.cell(row=r, column=1).fill = IMP_FILL
            elif status == 'Deteriorated':
                ws.cell(row=r, column=1).fill = DET_FILL
            r += 1
        ws.cell(row=r, column=1, value="Total In Pool").font = Font(bold=True)
        ws.cell(row=r, column=2, value=ptotal).number_format = MONEY
        r += 1
        net = pdf[pdf['migration_status'] == 'Improved']['current_balance'].sum() - \
              pdf[pdf['migration_status'] == 'Deteriorated']['current_balance'].sum()
        ws.cell(row=r, column=1, value="Net Change").font = Font(bold=True, size=12)
        ws.cell(row=r, column=2, value=net).number_format = MONEY
        ws.cell(row=r, column=3, value=net / ptotal if ptotal else 0).number_format = PCT
        r += 2

    # Grand Total
    total = df['current_balance'].sum()
    ws.cell(row=r, column=1, value="Grand Total").font = Font(bold=True, size=12)
    r += 1
    for status in ['Improved', 'Deteriorated', 'Unchanged']:
        bal = df[df['migration_status'] == status]['current_balance'].sum()
        lbl = f"Total-{status}" if status != 'Unchanged' else "Total Unchanged"
        ws.cell(row=r, column=1, value=lbl)
        ws.cell(row=r, column=2, value=bal).number_format = MONEY
        ws.cell(row=r, column=3, value=bal / total if total else 0).number_format = PCT
        r += 1
    ws.cell(row=r, column=1, value="Total In Portfolio").font = Font(bold=True)
    ws.cell(row=r, column=2, value=total).number_format = MONEY
    r += 1
    net = df[df['migration_status'] == 'Improved']['current_balance'].sum() - \
          df[df['migration_status'] == 'Deteriorated']['current_balance'].sum()
    ws.cell(row=r, column=1, value="Net Change").font = Font(bold=True, size=14)
    ws.cell(row=r, column=2, value=net).number_format = MONEY
    ws.cell(row=r, column=3, value=net / total if total else 0).number_format = PCT
    auto_w(ws, 3)


def _pool_risk_sheet(wb, cu, snap, pool_df, pool_name, grades, config):
    """Risk Change per pool with matrix, net credit change, delinquency/chargeoff stubs."""
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]
    matrix = risk_change_matrix(pool_df, grades, no_score)
    safe = re.sub(r'[^\w\s-]', '', pool_name)[:25]
    ws = wb.create_sheet(safe)

    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=12)
    ws['A2'] = "Risk Change By Credit Score"
    ws['A3'] = f"For Period Ending {snap}"
    ws['A4'] = pool_name
    ws['A4'].font = Font(bold=True, size=14)

    range_overrides = config.get('risk_change_range_labels', {})
    rng = {g['label']: range_overrides.get(g['label'], f"{g['min_score']}-{g['max_score']}") for g in grades}
    rng[no_score] = ""
    ncol = 3 + len(gl)

    gray_fill = PatternFill('solid', fgColor='D9D9D9')
    gray_hdr_font = Font(name='Calibri', bold=True, size=10, color='000000')

    def apply_gray_header(row_num):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font = gray_hdr_font
            cell.fill = gray_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = THIN

    def apply_plain_rows(r1, r2):
        for rr in range(r1, r2 + 1):
            for cc in range(1, ncol + 1):
                cell = ws.cell(row=rr, column=cc)
                if not cell.font or not cell.font.bold:
                    cell.font = NORM
                cell.border = THIN

    # Dollar matrix
    r = 6
    ws.cell(row=r, column=1, value="Dollar")
    ws.cell(row=r, column=4, value="Original Credit Grade")
    r += 1
    ws.cell(row=r, column=1, value="Current Credit Grade")
    ws.cell(row=r, column=2, value="")
    for j, g in enumerate(gl):
        ws.cell(row=r, column=3 + j, value=g)
    ws.cell(row=r, column=ncol, value="Grand Total")
    apply_gray_header(r)
    start = r + 1
    for i, g in enumerate(gl):
        r += 1
        ws.cell(row=r, column=1, value=g)
        ws.cell(row=r, column=2, value=rng.get(g, ''))
        rt = 0
        for j, og in enumerate(gl):
            v = matrix.loc[g, og] if g in matrix.index and og in matrix.columns else 0
            ws.cell(row=r, column=3 + j, value=v).number_format = MONEY
            rt += v
            if j < i: ws.cell(row=r, column=3+j).fill = DET_FILL
            elif j > i: ws.cell(row=r, column=3+j).fill = IMP_FILL
        ws.cell(row=r, column=ncol, value=rt).number_format = MONEY
    r += 1
    ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
    pool_total = pool_df['current_balance'].sum()
    for j, og in enumerate(gl):
        ct = sum(matrix.loc[g2, og] for g2 in gl if g2 in matrix.index and og in matrix.columns)
        ws.cell(row=r, column=3+j, value=ct).number_format = MONEY
    ws.cell(row=r, column=ncol, value=pool_total).number_format = MONEY
    apply_plain_rows(start, r)
    r += 1
    ws.cell(row=r, column=1, value="Loans Not Risk Rated and Adjustments")
    ws.cell(row=r, column=ncol, value=0).number_format = MONEY
    r += 1
    ws.cell(row=r, column=1, value="Total in Pool").font = Font(bold=True)
    if ncol > 3:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncol - 1)
        ws.cell(row=r, column=2, value="Section 3 Allowance for Credit Loss Calculation")
        ws.cell(row=r, column=2).font = Font(bold=True)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=ncol, value=pool_total).number_format = MONEY
    apply_plain_rows(r - 1, r)

    # Percent matrix
    r += 2
    ws.cell(row=r, column=1, value="Percent")
    ws.cell(row=r, column=4, value="Original Credit Grade")
    r += 1
    ws.cell(row=r, column=1, value="Current Credit Grade")
    for j, g in enumerate(gl):
        ws.cell(row=r, column=3+j, value=g)
    ws.cell(row=r, column=ncol, value="Grand Total")
    apply_gray_header(r)
    start2 = r + 1
    for i, g in enumerate(gl):
        r += 1
        ws.cell(row=r, column=1, value=g)
        ws.cell(row=r, column=2, value=rng.get(g, ''))
        rt = 0
        for j, og in enumerate(gl):
            v = matrix.loc[g, og] if g in matrix.index and og in matrix.columns else 0
            ct = sum(matrix.loc[g2, og] for g2 in gl if g2 in matrix.index and og in matrix.columns)
            ws.cell(row=r, column=3+j, value=v/ct if ct else 0).number_format = PCT
            rt += v
        ws.cell(row=r, column=ncol, value=rt/pool_total if pool_total else 0).number_format = PCT
    r += 1
    ws.cell(row=r, column=1, value="Grand Total").font = Font(bold=True)
    for j in range(len(gl)):
        ws.cell(row=r, column=3+j, value=1.0).number_format = PCT
    ws.cell(row=r, column=ncol, value=1.0).number_format = PCT
    style_rows(ws, start2, r, ncol, pcols=set(range(3, ncol+1)))

    # Net Credit Change box
    r += 2
    ws.cell(row=r, column=1, value="Net Credit Change").font = SUB_FONT
    imp = pool_df[pool_df['migration_status']=='Improved']['current_balance'].sum()
    det = pool_df[pool_df['migration_status']=='Deteriorated']['current_balance'].sum()
    unc = pool_df[pool_df['migration_status']=='Unchanged']['current_balance'].sum()
    for lbl, vd, vp in [("Total-Improved", imp, imp/pool_total if pool_total else 0),
                        ("Total-Deteriorated", det, det/pool_total if pool_total else 0),
                        ("Total Unchanged", unc, unc/pool_total if pool_total else 0),
                        ("Total In Portfolio", pool_total, 1.0)]:
        r += 1
        ws.cell(row=r, column=1, value=lbl)
        ws.cell(row=r, column=2, value=vd).number_format = MONEY
        ws.cell(row=r, column=3, value=vp).number_format = PCT
        r += 1
    ws.cell(row=r, column=1, value="Net Change").font = Font(bold=True, size=12)
    ws.cell(row=r, column=2, value=imp - det).number_format = MONEY
    ws.cell(row=r, column=3, value=(imp-det)/pool_total if pool_total else 0).number_format = PCT

    auto_w(ws, ncol)


def sheet_pool_risk_changes(wb, cu, snap, df, grades, config):
    """Create one Risk Change sheet per pool."""
    pools = sorted(df['loan_pool'].unique())
    for pool in pools:
        _pool_risk_sheet(wb, cu, snap, df, pool, grades, config)


def sheet_acl_reserve(wb, cu, snap, df, grades, config, hist=None):
    """Allowance & Provision for Credit Loss Reserve Analysis."""
    ws = wb.create_sheet("ACL Reserve Analysis")
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]
    econ_stress = calc_economic_stress(config)

    # Compute life loss rates from historical data
    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    avg_bals = hist.get('avg_balances', {}) if hist else {}
    hist_years = hist.get('years', []) if hist else []
    dq_pct = hist.get('dq_pct', {}) if hist else {}

    pool_life_loss = {}
    for pool in sorted(df['loan_pool'].unique()):
        rates = []
        for y in hist_years:
            net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
            avg = avg_bals.get(y, {}).get(pool, 0)
            if avg > 0:
                rates.append(net / avg)
        pool_life_loss[pool] = sum(rates) / len(rates) if rates else 0

    # Compute delinquency variance per pool
    pool_dq_var = {}
    for pool in sorted(df['loan_pool'].unique()):
        dq_rates = [dq_pct.get(y, {}).get(pool, 0) for y in sorted(dq_pct.keys())]
        if len(dq_rates) >= 2:
            avg_dq = sum(dq_rates) / len(dq_rates)
            pool_dq_var[pool] = dq_rates[-1] - avg_dq
        else:
            pool_dq_var[pool] = 0

    mgmt_adj_by_pool = config.get('mgmt_adj_by_pool', {})
    pool_use_default = _build_pool_use_default_map(config)
    admin_default_mgmt_adj = _load_admin_default_mgmt_adj()
    _imp = hist.get('impaired', {}) if hist else {}
    prior_mgmt_adj = _imp.get('prior_mgmt_adj', {})
    prior_env_factor = _imp.get('prior_env_factor', {})
    spec_id_by_pool = _imp.get('spec_id_by_pool', {})
    acl_impaired = _imp.get('acl_impaired', {})
    acl_summary = _imp.get('acl_summary', {})
    spec_id_by_pool = _imp.get('spec_id_by_pool', {})
    acl_impaired = _imp.get('acl_impaired', {})
    acl_summary = _imp.get('acl_summary', {})

    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Allowance & Provision for Credit Loss Reserve Analysis"
    ws['A3'] = f"For Period Ending {snap}"

    headers = ["Current Grade", "Balance", "Specific\nIdentification",
               "Loan Loss Calc.\nBalance", "ACL Base\nLoss Rate",
               "Management\nAdjustment", "Allowance\nFactor",
               "Allowance before\nEnvironmental", "Environmental\nFactor",
               "Environmental\n Allowance", "Total Allowance"]
    pools = sorted(df['loan_pool'].unique())
    r = 5
    grand_allowance = 0

    for pool in pools:
        pdf = df[df['loan_pool'] == pool]
        pool_total = pdf['current_balance'].sum()
        # Compute env factor for this pool
        imp_p, det_p, ncc = calc_net_credit_change(pdf)
        dq_var = pool_dq_var.get(pool, 0)
        env_factor_calc = calc_env_factor_pool(ncc, dq_var, econ_stress) / 100.0  # as decimal
        env_factor = prior_env_factor.get(pool, env_factor_calc)

        ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=12)
        r += 1
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers))
        start = r + 1
        pool_allowance_before = 0

        for gi, g in enumerate(gl):
            r += 1
            g_df = pdf[pdf['current_grade'] == g]
            balance = g_df['current_balance'].sum()
            specific_id = spec_id_by_pool.get(pool, {}).get(g, 0)
            calc_bal = balance - specific_id
            # ACL base rate: life loss rate × distribution factor
            life_loss = pool_life_loss.get(pool, 0)
            dist = get_dist_factor(gi, len(gl))
            base_rate = max(0, life_loss * dist)
            if base_rate == 0 and life_loss == 0:
                # Fallback to config reserve rate only when no data exists
                base_rate = next((gr['reserve_rate'] for gr in grades if gr['label'] == g), 0.005)
                if g == no_score:
                    base_rate = np.median([gr['reserve_rate'] for gr in grades])
            mgmt_adj = _resolve_mgmt_adj_grade(
                pool, g, gi, len(gl),
                pool_use_default, mgmt_adj_by_pool,
                admin_default_mgmt_adj, prior_mgmt_adj,
                base_rate=base_rate,
            )
            factor = base_rate + mgmt_adj
            allowance_before = calc_bal * factor
            pool_allowance_before += allowance_before

            ws.cell(row=r, column=1, value=g)
            ws.cell(row=r, column=2, value=balance).number_format = MONEY
            ws.cell(row=r, column=3, value=specific_id).number_format = MONEY
            ws.cell(row=r, column=4, value=calc_bal).number_format = MONEY
            ws.cell(row=r, column=5, value=base_rate).number_format = PCT4
            ws.cell(row=r, column=6, value=mgmt_adj).number_format = PCT4
            ws.cell(row=r, column=7, value=factor).number_format = PCT4
            ws.cell(row=r, column=8, value=allowance_before).number_format = MONEY
            ws.cell(row=r, column=9, value="")
            ws.cell(row=r, column=10, value="")
            ws.cell(row=r, column=11, value="")

        # Pool total row
        r += 1
        env_allowance = pool_allowance_before * env_factor
        total_allowance = pool_allowance_before + env_allowance
        grand_allowance += total_allowance
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=r, column=2, value=pool_total).number_format = MONEY
        ws.cell(row=r, column=8, value=pool_allowance_before).number_format = MONEY
        ws.cell(row=r, column=9, value=env_factor).number_format = PCT
        ws.cell(row=r, column=10, value=env_allowance).number_format = MONEY
        ws.cell(row=r, column=11, value=total_allowance).number_format = MONEY
        style_rows(ws, start, r, len(headers), mcols={2,3,4,8,10,11}, pcols={5,6,7,9})
    pooled_balance = df['current_balance'].sum()
    pooled_spec_id = sum(sum(g.values()) for g in spec_id_by_pool.values()) if spec_id_by_pool else 0
    pooled_calc_bal = pooled_balance - pooled_spec_id
    ws.cell(row=r, column=1, value="Pooled Totals").font = Font(bold=True, size=12)
    ws.cell(row=r, column=2, value=pooled_balance).number_format = MONEY
    ws.cell(row=r, column=3, value=pooled_spec_id).number_format = MONEY
    ws.cell(row=r, column=4, value=pooled_calc_bal).number_format = MONEY
    ws.cell(row=r, column=11, value=grand_allowance).number_format = MONEY

    # Impaired loans section
    r += 2
    ws.cell(row=r, column=1, value="Impaired Loans").font = Font(bold=True)
    ws.cell(row=r, column=10, value="Allowance").font = Font(bold=True)
    total_spec_allow = 0
    for lbl in ["Delinquent Loans", "Known Losses", "Repossessions",
                "Foreclosed Real Estate", "Deceased", "Bankruptcy"]:
        imp_val = acl_impaired.get(lbl, 0)
        r += 1
        ws.cell(row=r, column=1, value=lbl)
        ws.cell(row=r, column=11, value=imp_val).number_format = MONEY
        total_spec_allow += imp_val
    r += 1
    ws.cell(row=r, column=1, value="Total Specifically Identified Allowance").font = Font(bold=True)
    ws.cell(row=r, column=11, value=total_spec_allow).number_format = MONEY
    total_allow_needed = grand_allowance + total_spec_allow
    r += 1
    ws.cell(row=r, column=1, value="Total Allowance Needed").font = Font(bold=True)
    ws.cell(row=r, column=11, value=total_allow_needed).number_format = MONEY
    acl_bal = acl_summary.get('acl_balance', config.get('acl_balance', 0))
    r += 1
    ws.cell(row=r, column=1, value=f"Allowance for Credit Loss Balance as of {snap}")
    ws.cell(row=r, column=11, value=acl_bal).number_format = MONEY
    adjustment = total_allow_needed - acl_bal
    r += 1
    adj_label = "Adjustment (Underfunded)" if adjustment >= 0 else "Adjustment (Overfunded)"
    ws.cell(row=r, column=1, value=adj_label).font = Font(bold=True)
    ws.cell(row=r, column=11, value=adjustment).number_format = MONEY

    auto_w(ws, len(headers))


def sheet_env_factor(wb, cu, snap, df, grades, config, hist=None):
    """Environmental Factor for PLL."""
    ws = wb.create_sheet("Environmental Factor")
    ed = config.get('economic_data', {})
    econ_stress = calc_economic_stress(config)
    no_score = config.get('no_score_label', 'Not Reported')
    dq_pct = hist.get('dq_pct', {}) if hist else {}

    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Environmental Factor for PLL"
    ws['A3'] = f"For Period Ending {snap}"

    # Economic Stress Index
    r = 5
    ws.cell(row=r, column=1, value="Economic Stress Index Calculation").font = SUB_FONT
    r += 1
    ws.cell(row=r, column=1, value="State")
    ws.cell(row=r, column=2, value="Unemployment Rate")
    ws.cell(row=r, column=3, value="Foreclosures")
    ws.cell(row=r, column=4, value="Bankruptcies")
    ws.cell(row=r, column=5, value="Population")
    hdr_row(ws, r, 5)
    r += 1
    ws.cell(row=r, column=1, value=ed.get('state', ''))
    ws.cell(row=r, column=2, value=ed.get('unemployment_rate', 0)).number_format = PCT
    ws.cell(row=r, column=3, value=ed.get('foreclosures', 0))
    ws.cell(row=r, column=4, value=ed.get('bankruptcies', 0))
    ws.cell(row=r, column=5, value=ed.get('population', 0)).number_format = '#,##0'
    r += 1
    ws.cell(row=r, column=1, value="County")
    ws.cell(row=r, column=2, value="Unemployment Rate")
    ws.cell(row=r, column=3, value="Bankruptcy %")
    ws.cell(row=r, column=4, value="Foreclosure %")
    ws.cell(row=r, column=5, value="Economic Stress Index")
    hdr_row(ws, r, 5)
    r += 1
    pop = ed.get('population', 1)
    ws.cell(row=r, column=1, value=ed.get('county', ''))
    ws.cell(row=r, column=2, value=ed.get('unemployment_rate', 0)).number_format = PCT
    ws.cell(row=r, column=3, value=ed.get('bankruptcies', 0) / pop if pop else 0).number_format = PCT
    ws.cell(row=r, column=4, value=ed.get('foreclosures', 0) / pop if pop else 0).number_format = PCT
    ws.cell(row=r, column=5, value=econ_stress / 100).number_format = PCT

    # Per-pool environmental factors with real delinquency variance
    r += 2
    headers = ["Portfolio Segment", "Net Credit\nChange", "Net Credit\nScore",
               "Delinquency\nVariance", "Delinquency\nScore",
               "Economic Stress\nActual", "Economic Stress\nScore",
               "Environmental\nFactor"]
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))

    pools = sorted(df['loan_pool'].unique())
    start = r + 1
    for pool in pools:
        r += 1
        pdf = df[df['loan_pool'] == pool]
        _, _, ncc = calc_net_credit_change(pdf)
        ncc_score = score_from_ranges(ncc * 100, NCC_RANGES)

        # Compute delinquency variance from historical data
        dq_rates = [dq_pct.get(y, {}).get(pool, 0) for y in sorted(dq_pct.keys())]
        if len(dq_rates) >= 2:
            avg_dq = sum(dq_rates) / len(dq_rates)
            dq_var = dq_rates[-1] - avg_dq
        else:
            dq_var = 0

        dq_score = score_from_ranges(dq_var * 100, DQ_RANGES)
        es_score = score_from_ranges(econ_stress, ES_RANGES)
        env_f = ncc_score + dq_score + es_score

        ws.cell(row=r, column=1, value=pool)
        ws.cell(row=r, column=2, value=ncc).number_format = PCT
        ws.cell(row=r, column=3, value=ncc_score / 100).number_format = PCT
        ws.cell(row=r, column=4, value=dq_var).number_format = PCT
        ws.cell(row=r, column=5, value=dq_score / 100).number_format = PCT
        ws.cell(row=r, column=6, value=econ_stress / 100).number_format = PCT
        ws.cell(row=r, column=7, value=es_score / 100).number_format = PCT
        ws.cell(row=r, column=8, value=env_f / 100).number_format = PCT
    style_rows(ws, start, r, len(headers), pcols=set(range(2, 9)))
    auto_w(ws, len(headers))


def sheet_loss_factor(wb, cu, snap, df, grades, config, hist=None):
    """Loss Factor Calculation summary."""
    ws = wb.create_sheet("Loss Factor Calculation")
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Loss Factor Calculation"
    ws['A3'] = f"For Period Ending {snap}"

    headers = ["Current Grade", "Average Balance", "Life Loss Rate",
               "Distribution Factor", "ACL Base Loss Rate", "% of Loans"]
    pools = sorted(df['loan_pool'].unique())

    # Compute life loss rate per pool from historical data
    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    avg_bals = hist.get('avg_balances', {}) if hist else {}
    years = hist.get('years', []) if hist else []
    pool_life_loss = {}
    for pool in pools:
        rates = []
        for y in years:
            net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
            avg = avg_bals.get(y, {}).get(pool, 0)
            rate = net / avg if avg > 0 else 0
            ws.cell(row=r, column=2 + yi, value=rate).number_format = PCT4
            rates.append(rate)
        avg_rate = sum(rates) / len(rates) if rates else 0
        ws.cell(row=r, column=ncol, value=avg_rate).number_format = PCT4
    style_rows(ws, start, r, ncol, pcols=set(range(2, ncol + 1)))

    # Average balances section
    r += 3
    ws.cell(row=r, column=1, value="Average Balances by Pool").font = SUB_FONT
    r += 1
    headers2 = ["Pool"] + year_strs
    ncol2 = len(headers2)
    for hi, h in enumerate(headers2):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, ncol2)
    start2 = r + 1
    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool)
        for yi, y in enumerate(years):
            avg = avg_bals.get(y, {}).get(pool, 0)
            ws.cell(row=r, column=2 + yi, value=avg).number_format = MONEY
        style_rows(ws, start2, r, ncol2, mcols=set(range(2, ncol2 + 1)))
        auto_w(ws, max(ncol, ncol2))
    else:
        ws['A5'] = "No historical data available."
        ws['A5'].font = Font(italic=True, color='888888')


def sheet_chargeoff_recovery(wb, cu, snap, config, hist=None):
    """Charge off and Recoveries summary."""
    ws = wb.create_sheet("Charge offs & Recoveries")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Charge off and Recoveries"
    ws['A3'] = f"For Period Ending {snap}"

    pools = config.get('pool_map', {})
    pool_names = sorted(set(pools.values()))

    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    years = hist.get('years', []) if hist else []
    if not years:
        years = list(range(2019, int(snap[:4]) + 1))
    year_strs = [str(y) for y in years]

    # ─── Charge offs ───
    r = 5
    headers = ["Charge offs"] + year_strs + ["ACL Charge offs"]
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    start = r + 1
    grand_co = {y: 0 for y in years}
    for pool in pool_names:
        r += 1
        ws.cell(row=r, column=1, value=pool)
        row_total = 0
        for yi, y in enumerate(years):
            val = co_data.get(y, {}).get(pool, 0)
            ws.cell(row=r, column=2 + yi, value=val).number_format = MONEY
            row_total += val
        ws.cell(row=r, column=ncol, value=row_total).number_format = MONEY
    r += 1
    ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
    acl_co_total = 0
    for yi, y in enumerate(years):
        ws.cell(row=r, column=2 + yi, value=grand_co[y]).number_format = MONEY
        acl_co_total += grand_co[y]
    ws.cell(row=r, column=ncol, value=acl_co_total).number_format = MONEY
    style_rows(ws, start, r, len(headers), mcols=set(range(2, len(headers) + 1)))

    # ─── Recoveries ───
    r += 2
    headers2 = ["Recoveries"] + year_strs + ["ACL Recoveries"]
    for hi, h in enumerate(headers2):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers2))
    start2 = r + 1
    grand_rc = {y: 0 for y in years}
    for pool in pool_names:
        r += 1
        ws.cell(row=r, column=1, value=pool)
        row_total = 0
        for yi, y in enumerate(years):
            val = rc_data.get(y, {}).get(pool, 0)
            ws.cell(row=r, column=2 + yi, value=val).number_format = MONEY
            row_total += val
        ws.cell(row=r, column=ncol, value=row_total).number_format = MONEY
    r += 1
    ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
    acl_rc_total = 0
    for yi, y in enumerate(years):
        ws.cell(row=r, column=2 + yi, value=grand_rc[y]).number_format = MONEY
        acl_rc_total += grand_rc[y]
    ws.cell(row=r, column=ncol, value=acl_rc_total).number_format = MONEY
    style_rows(ws, start2, r, len(headers2), mcols=set(range(2, len(headers2) + 1)))

    # ─── Net Charge offs ───
    r += 2
    headers3 = ["Net Charge offs"] + year_strs + ["Net Charge offs"]
    for hi, h in enumerate(headers3):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers3))
    start3 = r + 1
    grand_net = {y: 0 for y in years}
    for pool in pool_names:
        r += 1
        ws.cell(row=r, column=1, value=pool)
        row_total = 0
        for yi, y in enumerate(years):
            net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
            ws.cell(row=r, column=2 + yi, value=net).number_format = MONEY
            grand_net[y] += net
            row_total += net
        ws.cell(row=r, column=ncol, value=row_total).number_format = MONEY
    r += 1
    ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
    net_total = 0
    for yi, y in enumerate(years):
        ws.cell(row=r, column=2 + yi, value=grand_net[y]).number_format = MONEY
        net_total += grand_net[y]
    ws.cell(row=r, column=ncol, value=net_total).number_format = MONEY
    style_rows(ws, start3, r, len(headers3), mcols=set(range(2, len(headers3) + 1)))

    # ─── Life of Loan Loss Rate ───
    avg_bals = hist.get('avg_balances', {}) if hist else {}
    r += 2
    headers4 = ["Life Loss Rate"] + year_strs + ["Average"]
    for hi, h in enumerate(headers4):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers4))
    start4 = r + 1
    for pool in pool_names:
        r += 1
        ws.cell(row=r, column=1, value=pool)
        rates = []
        for yi, y in enumerate(years):
            net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
            avg = avg_bals.get(y, {}).get(pool, 0)
            rate = net / avg if avg > 0 else 0
            ws.cell(row=r, column=2 + yi, value=rate).number_format = PCT4
            rates.append(rate)
        avg_rate = sum(rates) / len(rates) if rates else 0
        ws.cell(row=r, column=len(headers4), value=avg_rate).number_format = PCT4
    style_rows(ws, start4, r, len(headers4), pcols=set(range(2, len(headers4) + 1)))

    auto_w(ws, len(headers))


def sheet_delinquency(wb, cu, snap, config, hist=None):
    """Delinquency Calculation with historical data."""
    ws = wb.create_sheet("Delinquency Calculation")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Delinquency Calculation"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(set(config.get('pool_map', {}).values()))
    dq_pct = hist.get('dq_pct', {}) if hist else {}
    years = sorted(dq_pct.keys()) if dq_pct else list(range(2019, int(snap[:4]) + 1))
    year_strs = [str(y) for y in years]

    r = 5
    headers = ["DQ %"] + year_strs + ["Average", "Variance from Avg"]
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    start = r + 1
    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool)
        rates = []
        for yi, y in enumerate(years):
            val = dq_pct.get(y, {}).get(pool, 0)
            ws.cell(row=r, column=2 + yi, value=val).number_format = PCT
            rates.append(val)
        avg = sum(rates) / len(rates) if rates else 0
        ws.cell(row=r, column=len(headers) - 1, value=avg).number_format = PCT
        # Variance = most recent - average
        current = rates[-1] if rates else 0
        ws.cell(row=r, column=len(headers), value=current - avg).number_format = PCT
    style_rows(ws, start, r, len(headers), pcols=set(range(2, len(headers) + 1)))
    auto_w(ws, len(headers))


def sheet_balance_adj(wb, cu, snap, df, grades, config):
    """FAS 114 / Balance Adjustment sheet."""
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    ws = wb.create_sheet("Balance Adjustment")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Balance Adjustment"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(df['loan_pool'].unique())
    r = 5
    headers = ["Current Grade", "Loan Report Balance", "Bal Adjustment", "Balance Sheet Total"]

    for pool in pools:
        pdf = df[df['loan_pool'] == pool]
        ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=12)
        r += 1
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers))
        start = r + 1
        pool_total = 0
        for g in gl:
            r += 1
            bal = pdf[pdf['current_grade'] == g]['current_balance'].sum()
            pool_total += bal
            ws.cell(row=r, column=1, value=g)
            ws.cell(row=r, column=2, value=bal).number_format = MONEY
            ws.cell(row=r, column=3, value=0).number_format = MONEY
            ws.cell(row=r, column=4, value=bal).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=r, column=2, value=pool_total).number_format = MONEY
        ws.cell(row=r, column=3, value=0).number_format = MONEY
        ws.cell(row=r, column=4, value=pool_total).number_format = MONEY
        style_rows(ws, start, r, len(headers), mcols={2,3,4})
        r += 2

    # Grand total
    ws.cell(row=r, column=1, value="Grand Total").font = Font(bold=True, size=12)
    total = df['current_balance'].sum()
    ws.cell(row=r, column=2, value=total).number_format = MONEY
    ws.cell(row=r, column=3, value=0).number_format = MONEY
    ws.cell(row=r, column=4, value=total).number_format = MONEY
    auto_w(ws, len(headers))


def sheet_env_ranges(wb):
    """Environmental Factor Ranges reference table."""
    ws = wb.create_sheet("Env Factor Ranges")
    ws['A1'] = "Environmental Factor Ranges"
    ws['A1'].font = Font(bold=True, size=14)

    # Net Credit Change
    r = 3
    ws.cell(row=r, column=1, value="Net Credit Change").font = SUB_FONT
    ws.cell(row=r, column=1+0, value="Range"); ws.cell(row=r, column=2, value="Score")
    ws.cell(row=r, column=4, value="Delinquency").font = SUB_FONT
    ws.cell(row=r, column=4, value="Range"); ws.cell(row=r, column=5, value="Score")
    ws.cell(row=r, column=7, value="Economic Stress Score").font = SUB_FONT
    ws.cell(row=r, column=7, value="Range"); ws.cell(row=r, column=8, value="Score")
    hdr_row(ws, r, 8)

    ncc_rows = [
        ("<-18.00%", "7.00%"), ("-17.99% to -16.00%", "6.00%"),
        ("-15.99% to -14.00%", "5.00%"), ("-13.99% to -11.00%", "4.00%"),
        ("-10.99% to -8.00%", "3.00%"), ("-7.99% to -6.00%", "2.00%"),
        ("-5.99% to -4.00%", "1.00%"), ("-3.99% to 3.99%", "0.00%"),
        ("4.00% to 5.99%", "-1.00%"), ("6.00% to 7.99%", "-2.00%"),
        ("8.00% to 8.99%", "-3.00%"), ("9.00% to 10.99%", "-4.00%"),
        ("11.00% to 12.99%", "-5.00%"), ("13.00% to 14.99%", "-6.00%"),
        (">15.00%", "-7.00%"),
    ]
    dq_rows = [
        (">5.00%", "20.00%"), ("4.00% to 4.99%", "17.00%"),
        ("3.00% to 3.99%", "12.00%"), ("2.50% to 2.99%", "8.00%"),
        ("2.00% to 2.49%", "4.00%"), ("1.50% to 1.99%", "2.50%"),
        ("1.00% to 1.49%", "1.50%"), (".50% to .99%", "0.75%"),
        ("-.49% to .49%", "0.00%"), ("-.99% to -.50%", "-0.75%"),
        ("-1.49% to -1.00%", "-1.50%"), ("-1.99% to -1.50%", "-2.50%"),
        ("-2.49% to -2.00%", "-4.00%"), ("-2.99% to -2.50%", "-8.00%"),
        ("-3.99% to -3.00%", "-12.00%"), ("-4.99% to -4.00%", "-17.00%"),
        ("<-5.00%", "-20.00%"),
    ]
    es_rows = [
        (">25.00%", "10.00%"), ("24.00% to 24.99%", "8.00%"),
        ("22.00% to 23.99%", "7.00%"), ("20.00% to 21.99%", "6.00%"),
        ("18.00% to 19.99%", "5.00%"), ("16.00% to 17.99%", "4.00%"),
        ("14.00% to 15.99%", "3.50%"), ("12.00% to 13.99%", "3.00%"),
        ("10.00% to 11.99%", "2.00%"), ("8.00% to 9.99%", "1.00%"),
        ("6.00% to 7.99%", "0.00%"), ("4.00% to 5.99%", "0.00%"),
        ("2.00% to 3.99%", "-1.00%"), (".00% to 1.99%", "-2.00%"),
    ]

    for i, (rng, sc) in enumerate(ncc_rows):
        ws.cell(row=r + 1 + i, column=1, value=rng)
        ws.cell(row=r + 1 + i, column=2, value=sc)
    for i, (rng, sc) in enumerate(dq_rows):
        ws.cell(row=r + 1 + i, column=4, value=rng)
        ws.cell(row=r + 1 + i, column=5, value=sc)
    for i, (rng, sc) in enumerate(es_rows):
        ws.cell(row=r + 1 + i, column=7, value=rng)
        ws.cell(row=r + 1 + i, column=8, value=sc)
    auto_w(ws, 8)


def sheet_grade_config(wb, grades, config):
    """Grade ranges & loan code reference."""
    ws = wb.create_sheet("Grade Ranges & Loan Codes")
    ws['A1'] = "Credit Grade Configuration"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["Grade", "Score Range", "Reserve Rate"]
    r = 3
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    for g in grades:
        r += 1
        ws.cell(row=r, column=1, value=g['label'])
        ws.cell(row=r, column=2, value=f"{g['min_score']}-{g['max_score']}")
        ws.cell(row=r, column=3, value=g['reserve_rate']).number_format = PCT

    r += 3
    ws.cell(row=r, column=1, value="Loan Type Codes").font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1, value="Code"); ws.cell(row=r, column=2, value="Loan Pool")
    hdr_row(ws, r, 2)
    for code, pool in sorted(config.get('pool_map', {}).items(), key=lambda x: x[1]):
        r += 1
        ws.cell(row=r, column=1, value=str(code))
        ws.cell(row=r, column=2, value=pool)
    auto_w(ws, 3)


def sheet_all_loans(wb, cu, snap, df, grades, config):
    """All Loans detail listing."""
    ws = wb.create_sheet("All Loans")
    no_score = config.get('no_score_label', 'Not Reported')
    ws['A1'] = "Credit Grade Analysis - All Loans"
    ws['A1'].font = Font(bold=True, size=14)
    ws['F1'] = snap

    headers = ["Member #", "Loan Pool", "Current Balance",
               "Original Score", "Original Grade",
               "Current Score", "Current Grade",
               "Migration Status", "Reserve Rate", "Expected Loss"]
    r = 2
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))

    start = r + 1
    for _, loan in df.iterrows():
        r += 1
        ws.cell(row=r, column=1, value=str(loan.get('member_number', '')))
        ws.cell(row=r, column=2, value=loan.get('loan_pool', ''))
        ws.cell(row=r, column=3, value=loan.get('current_balance', 0))
        ws.cell(row=r, column=4, value=loan.get('original_fico_score', 0))
        ws.cell(row=r, column=5, value=loan.get('original_grade', no_score))
        ws.cell(row=r, column=6, value=loan.get('current_fico_score', 0))
        ws.cell(row=r, column=7, value=loan.get('current_grade', no_score))
        ws.cell(row=r, column=8, value=loan.get('migration_status', 'Unchanged'))
        ws.cell(row=r, column=9, value=loan.get('reserve_rate', 0))
        ws.cell(row=r, column=10, value=loan.get('expected_loss_amount', 0))
        status = loan.get('migration_status', 'Unchanged')
        if status == 'Improved':
            ws.cell(row=r, column=8).fill = IMP_FILL
        elif status == 'Deteriorated':
            ws.cell(row=r, column=8).fill = DET_FILL
    style_rows(ws, start, r, len(headers), mcols={3, 10}, pcols={9})
    auto_w(ws, len(headers), mx=18)


def sheet_hist_balances(wb, cu, snap, df, grades, config, hist=None):
    """Historical Loan Balances by pool with monthly balance data."""
    ws = wb.create_sheet("Historical Balances")
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Historical Loan Balances by Pool"
    ws['A3'] = f"For Period Ending {snap}"

    monthly = hist.get('monthly_balances', pd.DataFrame()) if hist else pd.DataFrame()

    if not monthly.empty:
        # Use monthly balance data - show quarterly snapshots per pool
        # Get quarter-end dates (month-end for Mar, Jun, Sep, Dec)
        monthly['quarter'] = monthly['date'].dt.to_period('Q')
        # Get last date per quarter per pool
        qtr_data = monthly.groupby(['pool', 'quarter']).last().reset_index()
        quarters = sorted(qtr_data['quarter'].unique())
        # Limit to last 20 quarters to keep sheet manageable
        if len(quarters) > 20:
            quarters = quarters[-20:]
        qtr_strs = [str(q) for q in quarters]

        pool_names = sorted(qtr_data['pool'].unique())
        r = 5
        headers = ["Pool"] + qtr_strs
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pool_names:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            for qi, q in enumerate(quarters):
                val = qtr_data[(qtr_data['pool'] == pool) & (qtr_data['quarter'] == q)]
                bal = val['balance'].values[0] if len(val) > 0 else 0
                ws.cell(row=r, column=2 + qi, value=bal).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        for qi, q in enumerate(quarters):
            total = qtr_data[qtr_data['quarter'] == q]['balance'].sum()
            ws.cell(row=r, column=2 + qi, value=total).number_format = MONEY
        style_rows(ws, start, r, ncol, mcols=set(range(2, ncol + 1)))
        auto_w(ws, ncol)
    else:
        # Fallback: just show current data by grade
        pools = sorted(df['loan_pool'].unique())
        r = 6
        for pool in pools:
            ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=12)
            r += 1
            ws.cell(row=r, column=1, value="Current Grade")
            ws.cell(row=r, column=2, value=snap)
            hdr_row(ws, r, 2)
            pdf = df[df['loan_pool'] == pool]
            for g in gl:
                r += 1
                ws.cell(row=r, column=1, value=g)
                ws.cell(row=r, column=2, value=pdf[pdf['current_grade'] == g]['current_balance'].sum()).number_format = MONEY
            r += 1
            ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
            ws.cell(row=r, column=2, value=pdf['current_balance'].sum()).number_format = MONEY
            r += 2
        auto_w(ws, 2)


def sheet_loss_factor_hist(wb, cu, snap, df, grades, config, hist=None):
    """Loss Factor Historical Detail with charge-off/recovery data."""
    ws = wb.create_sheet("Loss Factor Historical")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Loss Factor Historical Detail"
    ws['A3'] = f"For Period Ending {snap}"

    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    avg_bals = hist.get('avg_balances', {}) if hist else {}
    years = hist.get('years', []) if hist else []

    pools = sorted(set(config.get('pool_map', {}).values()))

    if years:
        year_strs = [str(y) for y in years]
        r = 5
        # Net charge-off rates per pool per year
        headers = ["Pool"] + year_strs + ["Average Life\nLoss Rate"]
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            rates = []
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                avg = avg_bals.get(y, {}).get(pool, 0)
                rate = net / avg if avg > 0 else 0
                ws.cell(row=r, column=2 + yi, value=rate).number_format = PCT4
                rates.append(rate)
            avg_rate = sum(rates) / len(rates) if rates else 0
            ws.cell(row=r, column=ncol, value=avg_rate).number_format = PCT4

        style_rows(ws, start, r, ncol, pcols=set(range(2, ncol + 1)))

        # Average balances section
        r += 3
        ws.cell(row=r, column=1, value="Average Balances by Pool").font = SUB_FONT
        r += 1
        headers2 = ["Pool"] + year_strs
        ncol2 = len(headers2)
        for hi, h in enumerate(headers2):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol2)
        start2 = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            for yi, y in enumerate(years):
                avg = avg_bals.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=avg).number_format = MONEY
        style_rows(ws, start2, r, ncol2, mcols=set(range(2, ncol2 + 1)))
        auto_w(ws, max(ncol, ncol2))
    else:
        ws['A5'] = "No historical data available."
        ws['A5'].font = Font(italic=True, color='888888')


def sheet_chargeoff_hist(wb, cu, snap, config, hist=None):
    """Charge off / Recoveries Historical Detail."""
    ws = wb.create_sheet("Chargeoff Historical")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Charge off and Recoveries Historical Detail"
    ws['A3'] = f"For Period Ending {snap}"

    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    years = hist.get('years', []) if hist else []
    pools = sorted(set(config.get('pool_map', {}).values()))

    if years:
        year_strs = [str(y) for y in years]

        # Charge-offs by pool by year
        r = 5
        ws.cell(row=r, column=1, value="Charge offs by Year").font = SUB_FONT
        r += 1
        headers = ["Pool"] + year_strs + ["Total"]
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_tot = 0
            for yi, y in enumerate(years):
                val = co_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=val).number_format = MONEY
                row_tot += val
            ws.cell(row=r, column=ncol, value=row_tot).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        grand = 0
        for yi, y in enumerate(years):
            yt = sum(co_data.get(y, {}).get(p, 0) for p in pools)
            ws.cell(row=r, column=2 + yi, value=yt).number_format = MONEY
            grand += yt
        ws.cell(row=r, column=ncol, value=grand).number_format = MONEY
        style_rows(ws, start, r, ncol, mcols=set(range(2, ncol + 1)))

        # Recoveries by pool by year
        r += 3
        ws.cell(row=r, column=1, value="Recoveries by Year").font = SUB_FONT
        r += 1
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start2 = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_tot = 0
            for yi, y in enumerate(years):
                val = rc_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=val).number_format = MONEY
                row_tot += val
            ws.cell(row=r, column=ncol, value=row_tot).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        grand = 0
        for yi, y in enumerate(years):
            yt = sum(rc_data.get(y, {}).get(p, 0) for p in pools)
            ws.cell(row=r, column=2 + yi, value=yt).number_format = MONEY
            grand += yt
        ws.cell(row=r, column=ncol, value=grand).number_format = MONEY
        style_rows(ws, start2, r, ncol, mcols=set(range(2, ncol + 1)))

        # Net Charge offs
        r += 2
        headers3 = ["Net Charge offs"] + year_strs + ["Net Charge offs"]
        for hi, h in enumerate(headers3):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers3))
        start3 = r + 1
        grand_net = {y: 0 for y in years}
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_total = 0
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=net).number_format = MONEY
                grand_net[y] += net
                row_total += net
            ws.cell(row=r, column=len(headers3), value=row_total).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        net_total = 0
        for yi, y in enumerate(years):
            ws.cell(row=r, column=2 + yi, value=grand_net[y]).number_format = MONEY
            net_total += grand_net[y]
        ws.cell(row=r, column=len(headers3), value=net_total).number_format = MONEY
        style_rows(ws, start3, r, len(headers3), mcols=set(range(2, len(headers3) + 1)))

        # Life of Loan Loss Rate
        r += 2
        headers4 = ["Life Loss Rate"] + year_strs + ["Average"]
        for hi, h in enumerate(headers4):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers4))
        start4 = r + 1
        for pool in pool_names:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            rates = []
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                avg = avg_bals.get(y, {}).get(pool, 0)
                rate = net / avg if avg > 0 else 0
                ws.cell(row=r, column=2 + yi, value=rate).number_format = PCT4
                rates.append(rate)
            avg_rate = sum(rates) / len(rates) if rates else 0
            ws.cell(row=r, column=len(headers4), value=avg_rate).number_format = PCT4
        style_rows(ws, start4, r, len(headers4), pcols=set(range(2, len(headers4) + 1)))

        auto_w(ws, len(headers))


def sheet_delinquency(wb, cu, snap, config, hist=None):
    """Delinquency Calculation with historical data."""
    ws = wb.create_sheet("Delinquency Calculation")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Delinquency Calculation"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(set(config.get('pool_map', {}).values()))
    dq_pct = hist.get('dq_pct', {}) if hist else {}
    years = sorted(dq_pct.keys()) if dq_pct else list(range(2019, int(snap[:4]) + 1))
    year_strs = [str(y) for y in years]

    r = 5
    headers = ["DQ %"] + year_strs + ["Average", "Variance from Avg"]
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    start = r + 1
    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool)
        rates = []
        for yi, y in enumerate(years):
            val = dq_pct.get(y, {}).get(pool, 0)
            ws.cell(row=r, column=2 + yi, value=val).number_format = PCT
            rates.append(val)
        avg = sum(rates) / len(rates) if rates else 0
        ws.cell(row=r, column=len(headers) - 1, value=avg).number_format = PCT
        # Variance = most recent - average
        current = rates[-1] if rates else 0
        ws.cell(row=r, column=len(headers), value=current - avg).number_format = PCT
    style_rows(ws, start, r, len(headers), pcols=set(range(2, len(headers) + 1)))
    auto_w(ws, len(headers))


def sheet_balance_adj(wb, cu, snap, df, grades, config):
    """FAS 114 / Balance Adjustment sheet."""
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    ws = wb.create_sheet("Balance Adjustment")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Balance Adjustment"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(df['loan_pool'].unique())
    r = 5
    headers = ["Current Grade", "Loan Report Balance", "Bal Adjustment", "Balance Sheet Total"]

    for pool in pools:
        pdf = df[df['loan_pool'] == pool]
        ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=12)
        r += 1
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers))
        start = r + 1
        pool_total = 0
        for g in gl:
            r += 1
            bal = pdf[pdf['current_grade'] == g]['current_balance'].sum()
            pool_total += bal
            ws.cell(row=r, column=1, value=g)
            ws.cell(row=r, column=2, value=bal).number_format = MONEY
            ws.cell(row=r, column=3, value=0).number_format = MONEY
            ws.cell(row=r, column=4, value=bal).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=r, column=2, value=pool_total).number_format = MONEY
        ws.cell(row=r, column=3, value=0).number_format = MONEY
        ws.cell(row=r, column=4, value=pool_total).number_format = MONEY
        style_rows(ws, start, r, len(headers), mcols={2,3,4})
        r += 2

    # Grand total
    ws.cell(row=r, column=1, value="Grand Total").font = Font(bold=True, size=12)
    total = df['current_balance'].sum()
    ws.cell(row=r, column=2, value=total).number_format = MONEY
    ws.cell(row=r, column=3, value=0).number_format = MONEY
    ws.cell(row=r, column=4, value=total).number_format = MONEY
    auto_w(ws, len(headers))


def sheet_env_ranges(wb):
    """Environmental Factor Ranges reference table."""
    ws = wb.create_sheet("Env Factor Ranges")
    ws['A1'] = "Environmental Factor Ranges"
    ws['A1'].font = Font(bold=True, size=14)

    # Net Credit Change
    r = 3
    ws.cell(row=r, column=1, value="Net Credit Change").font = SUB_FONT
    ws.cell(row=r, column=1+0, value="Range"); ws.cell(row=r, column=2, value="Score")
    ws.cell(row=r, column=4, value="Delinquency").font = SUB_FONT
    ws.cell(row=r, column=4, value="Range"); ws.cell(row=r, column=5, value="Score")
    ws.cell(row=r, column=7, value="Economic Stress Score").font = SUB_FONT
    ws.cell(row=r, column=7, value="Range"); ws.cell(row=r, column=8, value="Score")
    hdr_row(ws, r, 8)

    ncc_rows = [
        ("<-18.00%", "7.00%"), ("-17.99% to -16.00%", "6.00%"),
        ("-15.99% to -14.00%", "5.00%"), ("-13.99% to -11.00%", "4.00%"),
        ("-10.99% to -8.00%", "3.00%"), ("-7.99% to -6.00%", "2.00%"),
        ("-5.99% to -4.00%", "1.00%"), ("-3.99% to 3.99%", "0.00%"),
        ("4.00% to 5.99%", "-1.00%"), ("6.00% to 7.99%", "-2.00%"),
        ("8.00% to 8.99%", "-3.00%"), ("9.00% to 10.99%", "-4.00%"),
        ("11.00% to 12.99%", "-5.00%"), ("13.00% to 14.99%", "-6.00%"),
        (">15.00%", "-7.00%"),
    ]
    dq_rows = [
        (">5.00%", "20.00%"), ("4.00% to 4.99%", "17.00%"),
        ("3.00% to 3.99%", "12.00%"), ("2.50% to 2.99%", "8.00%"),
        ("2.00% to 2.49%", "4.00%"), ("1.50% to 1.99%", "2.50%"),
        ("1.00% to 1.49%", "1.50%"), (".50% to .99%", "0.75%"),
        ("-.49% to .49%", "0.00%"), ("-.99% to -.50%", "-0.75%"),
        ("-1.49% to -1.00%", "-1.50%"), ("-1.99% to -1.50%", "-2.50%"),
        ("-2.49% to -2.00%", "-4.00%"), ("-2.99% to -2.50%", "-8.00%"),
        ("-3.99% to -3.00%", "-12.00%"), ("-4.99% to -4.00%", "-17.00%"),
        ("<-5.00%", "-20.00%"),
    ]
    es_rows = [
        (">25.00%", "10.00%"), ("24.00% to 24.99%", "8.00%"),
        ("22.00% to 23.99%", "7.00%"), ("20.00% to 21.99%", "6.00%"),
        ("18.00% to 19.99%", "5.00%"), ("16.00% to 17.99%", "4.00%"),
        ("14.00% to 15.99%", "3.50%"), ("12.00% to 13.99%", "3.00%"),
        ("10.00% to 11.99%", "2.00%"), ("8.00% to 9.99%", "1.00%"),
        ("6.00% to 7.99%", "0.00%"), ("4.00% to 5.99%", "0.00%"),
        ("2.00% to 3.99%", "-1.00%"), (".00% to 1.99%", "-2.00%"),
    ]

    for i, (rng, sc) in enumerate(ncc_rows):
        ws.cell(row=r + 1 + i, column=1, value=rng)
        ws.cell(row=r + 1 + i, column=2, value=sc)
    for i, (rng, sc) in enumerate(dq_rows):
        ws.cell(row=r + 1 + i, column=4, value=rng)
        ws.cell(row=r + 1 + i, column=5, value=sc)
    for i, (rng, sc) in enumerate(es_rows):
        ws.cell(row=r + 1 + i, column=7, value=rng)
        ws.cell(row=r + 1 + i, column=8, value=sc)
    auto_w(ws, 8)


def sheet_grade_config(wb, grades, config):
    """Grade ranges & loan code reference."""
    ws = wb.create_sheet("Grade Ranges & Loan Codes")
    ws['A1'] = "Credit Grade Configuration"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["Grade", "Score Range", "Reserve Rate"]
    r = 3
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    for g in grades:
        r += 1
        ws.cell(row=r, column=1, value=g['label'])
        ws.cell(row=r, column=2, value=f"{g['min_score']}-{g['max_score']}")
        ws.cell(row=r, column=3, value=g['reserve_rate']).number_format = PCT

    r += 3
    ws.cell(row=r, column=1, value="Loan Type Codes").font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1, value="Code"); ws.cell(row=r, column=2, value="Loan Pool")
    hdr_row(ws, r, 2)
    for code, pool in sorted(config.get('pool_map', {}).items(), key=lambda x: x[1]):
        r += 1
        ws.cell(row=r, column=1, value=str(code))
        ws.cell(row=r, column=2, value=pool)
    auto_w(ws, 3)


def sheet_all_loans(wb, cu, snap, df, grades, config):
    """All Loans detail listing."""
    ws = wb.create_sheet("All Loans")
    no_score = config.get('no_score_label', 'Not Reported')
    ws['A1'] = "Credit Grade Analysis - All Loans"
    ws['A1'].font = Font(bold=True, size=14)
    ws['F1'] = snap

    headers = ["Member #", "Loan Pool", "Current Balance",
               "Original Score", "Original Grade",
               "Current Score", "Current Grade",
               "Migration Status", "Reserve Rate", "Expected Loss"]
    r = 2
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))

    start = r + 1
    for _, loan in df.iterrows():
        r += 1
        ws.cell(row=r, column=1, value=str(loan.get('member_number', '')))
        ws.cell(row=r, column=2, value=loan.get('loan_pool', ''))
        ws.cell(row=r, column=3, value=loan.get('current_balance', 0))
        ws.cell(row=r, column=4, value=loan.get('original_fico_score', 0))
        ws.cell(row=r, column=5, value=loan.get('original_grade', no_score))
        ws.cell(row=r, column=6, value=loan.get('current_fico_score', 0))
        ws.cell(row=r, column=7, value=loan.get('current_grade', no_score))
        ws.cell(row=r, column=8, value=loan.get('migration_status', 'Unchanged'))
        ws.cell(row=r, column=9, value=loan.get('reserve_rate', 0))
        ws.cell(row=r, column=10, value=loan.get('expected_loss_amount', 0))
        status = loan.get('migration_status', 'Unchanged')
        if status == 'Improved':
            ws.cell(row=r, column=8).fill = IMP_FILL
        elif status == 'Deteriorated':
            ws.cell(row=r, column=8).fill = DET_FILL
    style_rows(ws, start, r, len(headers), mcols={3, 10}, pcols={9})
    auto_w(ws, len(headers), mx=18)


def sheet_hist_balances(wb, cu, snap, df, grades, config, hist=None):
    """Historical Loan Balances by pool with monthly balance data."""
    ws = wb.create_sheet("Historical Balances")
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Historical Loan Balances by Pool"
    ws['A3'] = f"For Period Ending {snap}"

    monthly = hist.get('monthly_balances', pd.DataFrame()) if hist else pd.DataFrame()

    if not monthly.empty:
        # Use monthly balance data - show quarterly snapshots per pool
        # Get quarter-end dates (month-end for Mar, Jun, Sep, Dec)
        monthly['quarter'] = monthly['date'].dt.to_period('Q')
        # Get last date per quarter per pool
        qtr_data = monthly.groupby(['pool', 'quarter']).last().reset_index()
        quarters = sorted(qtr_data['quarter'].unique())
        # Limit to last 20 quarters to keep sheet manageable
        if len(quarters) > 20:
            quarters = quarters[-20:]
        qtr_strs = [str(q) for q in quarters]

        pool_names = sorted(qtr_data['pool'].unique())
        r = 5
        headers = ["Pool"] + qtr_strs
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pool_names:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            for qi, q in enumerate(quarters):
                val = qtr_data[(qtr_data['pool'] == pool) & (qtr_data['quarter'] == q)]
                bal = val['balance'].values[0] if len(val) > 0 else 0
                ws.cell(row=r, column=2 + qi, value=bal).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        for qi, q in enumerate(quarters):
            total = qtr_data[qtr_data['quarter'] == q]['balance'].sum()
            ws.cell(row=r, column=2 + qi, value=total).number_format = MONEY
        style_rows(ws, start, r, ncol, mcols=set(range(2, ncol + 1)))
        auto_w(ws, ncol)
    else:
        # Fallback: just show current data by grade
        pools = sorted(df['loan_pool'].unique())
        r = 6
        for pool in pools:
            ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=12)
            r += 1
            ws.cell(row=r, column=1, value="Current Grade")
            ws.cell(row=r, column=2, value=snap)
            hdr_row(ws, r, 2)
            pdf = df[df['loan_pool'] == pool]
            for g in gl:
                r += 1
                ws.cell(row=r, column=1, value=g)
                ws.cell(row=r, column=2, value=pdf[pdf['current_grade'] == g]['current_balance'].sum()).number_format = MONEY
            r += 1
            ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
            ws.cell(row=r, column=2, value=pdf['current_balance'].sum()).number_format = MONEY
            r += 2
        auto_w(ws, 2)


def sheet_loss_factor_hist(wb, cu, snap, df, grades, config, hist=None):
    """Loss Factor Historical Detail with charge-off/recovery data."""
    ws = wb.create_sheet("Loss Factor Historical")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Loss Factor Historical Detail"
    ws['A3'] = f"For Period Ending {snap}"

    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    avg_bals = hist.get('avg_balances', {}) if hist else {}
    years = hist.get('years', []) if hist else []

    pools = sorted(set(config.get('pool_map', {}).values()))

    if years:
        year_strs = [str(y) for y in years]
        r = 5
        # Net charge-off rates per pool per year
        headers = ["Pool"] + year_strs + ["Average Life\nLoss Rate"]
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            rates = []
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                avg = avg_bals.get(y, {}).get(pool, 0)
                rate = net / avg if avg > 0 else 0
                ws.cell(row=r, column=2 + yi, value=rate).number_format = PCT4
                rates.append(rate)
            avg_rate = sum(rates) / len(rates) if rates else 0
            ws.cell(row=r, column=ncol, value=avg_rate).number_format = PCT4

        style_rows(ws, start, r, ncol, pcols=set(range(2, ncol + 1)))

        # Average balances section
        r += 3
        ws.cell(row=r, column=1, value="Average Balances by Pool").font = SUB_FONT
        r += 1
        headers2 = ["Pool"] + year_strs
        ncol2 = len(headers2)
        for hi, h in enumerate(headers2):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol2)
        start2 = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            for yi, y in enumerate(years):
                avg = avg_bals.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=avg).number_format = MONEY
        style_rows(ws, start2, r, ncol2, mcols=set(range(2, ncol2 + 1)))
        auto_w(ws, max(ncol, ncol2))
    else:
        ws['A5'] = "No historical data available."
        ws['A5'].font = Font(italic=True, color='888888')


def sheet_chargeoff_hist(wb, cu, snap, config, hist=None):
    """Charge off / Recoveries Historical Detail."""
    ws = wb.create_sheet("Chargeoff Historical")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Charge off and Recoveries Historical Detail"
    ws['A3'] = f"For Period Ending {snap}"

    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    years = hist.get('years', []) if hist else []
    pools = sorted(set(config.get('pool_map', {}).values()))

    if years:
        year_strs = [str(y) for y in years]

        # Charge-offs by pool by year
        r = 5
        ws.cell(row=r, column=1, value="Charge offs by Year").font = SUB_FONT
        r += 1
        headers = ["Pool"] + year_strs + ["Total"]
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_tot = 0
            for yi, y in enumerate(years):
                val = co_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=val).number_format = MONEY
                row_tot += val
            ws.cell(row=r, column=ncol, value=row_tot).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        grand = 0
        for yi, y in enumerate(years):
            yt = sum(co_data.get(y, {}).get(p, 0) for p in pools)
            ws.cell(row=r, column=2 + yi, value=yt).number_format = MONEY
            grand += yt
        ws.cell(row=r, column=ncol, value=grand).number_format = MONEY
        style_rows(ws, start, r, ncol, mcols=set(range(2, ncol + 1)))

        # Recoveries by pool by year
        r += 3
        ws.cell(row=r, column=1, value="Recoveries by Year").font = SUB_FONT
        r += 1
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start2 = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_tot = 0
            for yi, y in enumerate(years):
                val = rc_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=val).number_format = MONEY
                row_tot += val
            ws.cell(row=r, column=ncol, value=row_tot).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        grand = 0
        for yi, y in enumerate(years):
            yt = sum(rc_data.get(y, {}).get(p, 0) for p in pools)
            ws.cell(row=r, column=2 + yi, value=yt).number_format = MONEY
            grand += yt
        ws.cell(row=r, column=ncol, value=grand).number_format = MONEY
        style_rows(ws, start2, r, ncol, mcols=set(range(2, ncol + 1)))

        # Net Charge offs
        r += 2
        headers3 = ["Net Charge offs"] + year_strs + ["Net Charge offs"]
        for hi, h in enumerate(headers3):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers3))
        start3 = r + 1
        grand_net = {y: 0 for y in years}
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_total = 0
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=net).number_format = MONEY
                grand_net[y] += net
                row_total += net
            ws.cell(row=r, column=len(headers3), value=row_total).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        net_total = 0
        for yi, y in enumerate(years):
            ws.cell(row=r, column=2 + yi, value=grand_net[y]).number_format = MONEY
            net_total += grand_net[y]
        ws.cell(row=r, column=len(headers3), value=net_total).number_format = MONEY
        style_rows(ws, start3, r, len(headers3), mcols=set(range(2, len(headers3) + 1)))

        # Life of Loan Loss Rate
        r += 2
        headers4 = ["Life Loss Rate"] + year_strs + ["Average"]
        for hi, h in enumerate(headers4):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers4))
        start4 = r + 1
        for pool in pool_names:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            rates = []
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                avg = avg_bals.get(y, {}).get(pool, 0)
                rate = net / avg if avg > 0 else 0
                ws.cell(row=r, column=2 + yi, value=rate).number_format = PCT4
                rates.append(rate)
            avg_rate = sum(rates) / len(rates) if rates else 0
            ws.cell(row=r, column=len(headers4), value=avg_rate).number_format = PCT4
        style_rows(ws, start4, r, len(headers4), pcols=set(range(2, len(headers4) + 1)))

        auto_w(ws, len(headers))


def sheet_delinquency(wb, cu, snap, config, hist=None):
    """Delinquency Calculation with historical data."""
    ws = wb.create_sheet("Delinquency Calculation")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Delinquency Calculation"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(set(config.get('pool_map', {}).values()))
    dq_pct = hist.get('dq_pct', {}) if hist else {}
    years = sorted(dq_pct.keys()) if dq_pct else list(range(2019, int(snap[:4]) + 1))
    year_strs = [str(y) for y in years]

    r = 5
    headers = ["DQ %"] + year_strs + ["Average", "Variance from Avg"]
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    start = r + 1
    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool)
        rates = []
        for yi, y in enumerate(years):
            val = dq_pct.get(y, {}).get(pool, 0)
            ws.cell(row=r, column=2 + yi, value=val).number_format = PCT
            rates.append(val)
        avg = sum(rates) / len(rates) if rates else 0
        ws.cell(row=r, column=len(headers) - 1, value=avg).number_format = PCT
        # Variance = most recent - average
        current = rates[-1] if rates else 0
        ws.cell(row=r, column=len(headers), value=current - avg).number_format = PCT
    style_rows(ws, start, r, len(headers), pcols=set(range(2, len(headers) + 1)))
    auto_w(ws, len(headers))


def sheet_balance_adj(wb, cu, snap, df, grades, config):
    """FAS 114 / Balance Adjustment sheet."""
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    ws = wb.create_sheet("Balance Adjustment")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Balance Adjustment"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(df['loan_pool'].unique())
    r = 5
    headers = ["Current Grade", "Loan Report Balance", "Bal Adjustment", "Balance Sheet Total"]

    for pool in pools:
        pdf = df[df['loan_pool'] == pool]
        ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=12)
        r += 1
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers))
        start = r + 1
        pool_total = 0
        for g in gl:
            r += 1
            bal = pdf[pdf['current_grade'] == g]['current_balance'].sum()
            pool_total += bal
            ws.cell(row=r, column=1, value=g)
            ws.cell(row=r, column=2, value=bal).number_format = MONEY
            ws.cell(row=r, column=3, value=0).number_format = MONEY
            ws.cell(row=r, column=4, value=bal).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=r, column=2, value=pool_total).number_format = MONEY
        ws.cell(row=r, column=3, value=0).number_format = MONEY
        ws.cell(row=r, column=4, value=pool_total).number_format = MONEY
        style_rows(ws, start, r, len(headers), mcols={2,3,4})
        r += 2

    # Grand total
    ws.cell(row=r, column=1, value="Grand Total").font = Font(bold=True, size=12)
    total = df['current_balance'].sum()
    ws.cell(row=r, column=2, value=total).number_format = MONEY
    ws.cell(row=r, column=3, value=0).number_format = MONEY
    ws.cell(row=r, column=4, value=total).number_format = MONEY
    auto_w(ws, len(headers))


def sheet_env_ranges(wb):
    """Environmental Factor Ranges reference table."""
    ws = wb.create_sheet("Env Factor Ranges")
    ws['A1'] = "Environmental Factor Ranges"
    ws['A1'].font = Font(bold=True, size=14)

    # Net Credit Change
    r = 3
    ws.cell(row=r, column=1, value="Net Credit Change").font = SUB_FONT
    ws.cell(row=r, column=1+0, value="Range"); ws.cell(row=r, column=2, value="Score")
    ws.cell(row=r, column=4, value="Delinquency").font = SUB_FONT
    ws.cell(row=r, column=4, value="Range"); ws.cell(row=r, column=5, value="Score")
    ws.cell(row=r, column=7, value="Economic Stress Score").font = SUB_FONT
    ws.cell(row=r, column=7, value="Range"); ws.cell(row=r, column=8, value="Score")
    hdr_row(ws, r, 8)

    ncc_rows = [
        ("<-18.00%", "7.00%"), ("-17.99% to -16.00%", "6.00%"),
        ("-15.99% to -14.00%", "5.00%"), ("-13.99% to -11.00%", "4.00%"),
        ("-10.99% to -8.00%", "3.00%"), ("-7.99% to -6.00%", "2.00%"),
        ("-5.99% to -4.00%", "1.00%"), ("-3.99% to 3.99%", "0.00%"),
        ("4.00% to 5.99%", "-1.00%"), ("6.00% to 7.99%", "-2.00%"),
        ("8.00% to 8.99%", "-3.00%"), ("9.00% to 10.99%", "-4.00%"),
        ("11.00% to 12.99%", "-5.00%"), ("13.00% to 14.99%", "-6.00%"),
        (">15.00%", "-7.00%"),
    ]
    dq_rows = [
        (">5.00%", "20.00%"), ("4.00% to 4.99%", "17.00%"),
        ("3.00% to 3.99%", "12.00%"), ("2.50% to 2.99%", "8.00%"),
        ("2.00% to 2.49%", "4.00%"), ("1.50% to 1.99%", "2.50%"),
        ("1.00% to 1.49%", "1.50%"), (".50% to .99%", "0.75%"),
        ("-.49% to .49%", "0.00%"), ("-.99% to -.50%", "-0.75%"),
        ("-1.49% to -1.00%", "-1.50%"), ("-1.99% to -1.50%", "-2.50%"),
        ("-2.49% to -2.00%", "-4.00%"), ("-2.99% to -2.50%", "-8.00%"),
        ("-3.99% to -3.00%", "-12.00%"), ("-4.99% to -4.00%", "-17.00%"),
        ("<-5.00%", "-20.00%"),
    ]
    es_rows = [
        (">25.00%", "10.00%"), ("24.00% to 24.99%", "8.00%"),
        ("22.00% to 23.99%", "7.00%"), ("20.00% to 21.99%", "6.00%"),
        ("18.00% to 19.99%", "5.00%"), ("16.00% to 17.99%", "4.00%"),
        ("14.00% to 15.99%", "3.50%"), ("12.00% to 13.99%", "3.00%"),
        ("10.00% to 11.99%", "2.00%"), ("8.00% to 9.99%", "1.00%"),
        ("6.00% to 7.99%", "0.00%"), ("4.00% to 5.99%", "0.00%"),
        ("2.00% to 3.99%", "-1.00%"), (".00% to 1.99%", "-2.00%"),
    ]

    for i, (rng, sc) in enumerate(ncc_rows):
        ws.cell(row=r + 1 + i, column=1, value=rng)
        ws.cell(row=r + 1 + i, column=2, value=sc)
    for i, (rng, sc) in enumerate(dq_rows):
        ws.cell(row=r + 1 + i, column=4, value=rng)
        ws.cell(row=r + 1 + i, column=5, value=sc)
    for i, (rng, sc) in enumerate(es_rows):
        ws.cell(row=r + 1 + i, column=7, value=rng)
        ws.cell(row=r + 1 + i, column=8, value=sc)
    auto_w(ws, 8)


def sheet_grade_config(wb, grades, config):
    """Grade ranges & loan code reference."""
    ws = wb.create_sheet("Grade Ranges & Loan Codes")
    ws['A1'] = "Credit Grade Configuration"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["Grade", "Score Range", "Reserve Rate"]
    r = 3
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    for g in grades:
        r += 1
        ws.cell(row=r, column=1, value=g['label'])
        ws.cell(row=r, column=2, value=f"{g['min_score']}-{g['max_score']}")
        ws.cell(row=r, column=3, value=g['reserve_rate']).number_format = PCT

    r += 3
    ws.cell(row=r, column=1, value="Loan Type Codes").font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1, value="Code"); ws.cell(row=r, column=2, value="Loan Pool")
    hdr_row(ws, r, 2)
    for code, pool in sorted(config.get('pool_map', {}).items(), key=lambda x: x[1]):
        r += 1
        ws.cell(row=r, column=1, value=str(code))
        ws.cell(row=r, column=2, value=pool)
    auto_w(ws, 3)


def sheet_all_loans(wb, cu, snap, df, grades, config):
    """All Loans detail listing."""
    ws = wb.create_sheet("All Loans")
    no_score = config.get('no_score_label', 'Not Reported')
    ws['A1'] = "Credit Grade Analysis - All Loans"
    ws['A1'].font = Font(bold=True, size=14)
    ws['F1'] = snap

    headers = ["Member #", "Loan Pool", "Current Balance",
               "Original Score", "Original Grade",
               "Current Score", "Current Grade",
               "Migration Status", "Reserve Rate", "Expected Loss"]
    r = 2
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))

    start = r + 1
    for _, loan in df.iterrows():
        r += 1
        ws.cell(row=r, column=1, value=str(loan.get('member_number', '')))
        ws.cell(row=r, column=2, value=loan.get('loan_pool', ''))
        ws.cell(row=r, column=3, value=loan.get('current_balance', 0))
        ws.cell(row=r, column=4, value=loan.get('original_fico_score', 0))
        ws.cell(row=r, column=5, value=loan.get('original_grade', no_score))
        ws.cell(row=r, column=6, value=loan.get('current_fico_score', 0))
        ws.cell(row=r, column=7, value=loan.get('current_grade', no_score))
        ws.cell(row=r, column=8, value=loan.get('migration_status', 'Unchanged'))
        ws.cell(row=r, column=9, value=loan.get('reserve_rate', 0))
        ws.cell(row=r, column=10, value=loan.get('expected_loss_amount', 0))
        status = loan.get('migration_status', 'Unchanged')
        if status == 'Improved':
            ws.cell(row=r, column=8).fill = IMP_FILL
        elif status == 'Deteriorated':
            ws.cell(row=r, column=8).fill = DET_FILL
    style_rows(ws, start, r, len(headers), mcols={3, 10}, pcols={9})
    auto_w(ws, len(headers), mx=18)


def sheet_hist_balances(wb, cu, snap, df, grades, config, hist=None):
    """Historical Loan Balances by pool with monthly balance data."""
    ws = wb.create_sheet("Historical Balances")
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Historical Loan Balances by Pool"
    ws['A3'] = f"For Period Ending {snap}"

    monthly = hist.get('monthly_balances', pd.DataFrame()) if hist else pd.DataFrame()

    if not monthly.empty:
        # Use monthly balance data - show quarterly snapshots per pool
        # Get quarter-end dates (month-end for Mar, Jun, Sep, Dec)
        monthly['quarter'] = monthly['date'].dt.to_period('Q')
        # Get last date per quarter per pool
        qtr_data = monthly.groupby(['pool', 'quarter']).last().reset_index()
        quarters = sorted(qtr_data['quarter'].unique())
        # Limit to last 20 quarters to keep sheet manageable
        if len(quarters) > 20:
            quarters = quarters[-20:]
        qtr_strs = [str(q) for q in quarters]

        pool_names = sorted(qtr_data['pool'].unique())
        r = 5
        headers = ["Pool"] + qtr_strs
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pool_names:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            for qi, q in enumerate(quarters):
                val = qtr_data[(qtr_data['pool'] == pool) & (qtr_data['quarter'] == q)]
                bal = val['balance'].values[0] if len(val) > 0 else 0
                ws.cell(row=r, column=2 + qi, value=bal).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        for qi, q in enumerate(quarters):
            total = qtr_data[qtr_data['quarter'] == q]['balance'].sum()
            ws.cell(row=r, column=2 + qi, value=total).number_format = MONEY
        style_rows(ws, start, r, ncol, mcols=set(range(2, ncol + 1)))
        auto_w(ws, ncol)
    else:
        # Fallback: just show current data by grade
        pools = sorted(df['loan_pool'].unique())
        r = 6
        for pool in pools:
            ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=12)
            r += 1
            ws.cell(row=r, column=1, value="Current Grade")
            ws.cell(row=r, column=2, value=snap)
            hdr_row(ws, r, 2)
            pdf = df[df['loan_pool'] == pool]
            for g in gl:
                r += 1
                ws.cell(row=r, column=1, value=g)
                ws.cell(row=r, column=2, value=pdf[pdf['current_grade'] == g]['current_balance'].sum()).number_format = MONEY
            r += 1
            ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
            ws.cell(row=r, column=2, value=pdf['current_balance'].sum()).number_format = MONEY
            r += 2
        auto_w(ws, 2)


def sheet_loss_factor_hist(wb, cu, snap, df, grades, config, hist=None):
    """Loss Factor Historical Detail with charge-off/recovery data."""
    ws = wb.create_sheet("Loss Factor Historical")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Loss Factor Historical Detail"
    ws['A3'] = f"For Period Ending {snap}"

    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    avg_bals = hist.get('avg_balances', {}) if hist else {}
    years = hist.get('years', []) if hist else []

    pools = sorted(set(config.get('pool_map', {}).values()))

    if years:
        year_strs = [str(y) for y in years]
        r = 5
        # Net charge-off rates per pool per year
        headers = ["Pool"] + year_strs + ["Average Life\nLoss Rate"]
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            rates = []
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                avg = avg_bals.get(y, {}).get(pool, 0)
                rate = net / avg if avg > 0 else 0
                ws.cell(row=r, column=2 + yi, value=rate).number_format = PCT4
                rates.append(rate)
            avg_rate = sum(rates) / len(rates) if rates else 0
            ws.cell(row=r, column=ncol, value=avg_rate).number_format = PCT4

        style_rows(ws, start, r, ncol, pcols=set(range(2, ncol + 1)))

        # Average balances section
        r += 3
        ws.cell(row=r, column=1, value="Average Balances by Pool").font = SUB_FONT
        r += 1
        headers2 = ["Pool"] + year_strs
        ncol2 = len(headers2)
        for hi, h in enumerate(headers2):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol2)
        start2 = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            for yi, y in enumerate(years):
                avg = avg_bals.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=avg).number_format = MONEY
        style_rows(ws, start2, r, ncol2, mcols=set(range(2, ncol2 + 1)))
        auto_w(ws, max(ncol, ncol2))
    else:
        ws['A5'] = "No historical data available."
        ws['A5'].font = Font(italic=True, color='888888')


def sheet_chargeoff_hist(wb, cu, snap, config, hist=None):
    """Charge off / Recoveries Historical Detail."""
    ws = wb.create_sheet("Chargeoff Historical")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Charge off and Recoveries Historical Detail"
    ws['A3'] = f"For Period Ending {snap}"

    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    years = hist.get('years', []) if hist else []
    pools = sorted(set(config.get('pool_map', {}).values()))

    if years:
        year_strs = [str(y) for y in years]

        # Charge-offs by pool by year
        r = 5
        ws.cell(row=r, column=1, value="Charge offs by Year").font = SUB_FONT
        r += 1
        headers = ["Pool"] + year_strs + ["Total"]
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_tot = 0
            for yi, y in enumerate(years):
                val = co_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=val).number_format = MONEY
                row_tot += val
            ws.cell(row=r, column=ncol, value=row_tot).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        grand = 0
        for yi, y in enumerate(years):
            yt = sum(co_data.get(y, {}).get(p, 0) for p in pools)
            ws.cell(row=r, column=2 + yi, value=yt).number_format = MONEY
            grand += yt
        ws.cell(row=r, column=ncol, value=grand).number_format = MONEY
        style_rows(ws, start, r, ncol, mcols=set(range(2, ncol + 1)))

        # Recoveries by pool by year
        r += 3
        ws.cell(row=r, column=1, value="Recoveries by Year").font = SUB_FONT
        r += 1
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start2 = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_tot = 0
            for yi, y in enumerate(years):
                val = rc_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=val).number_format = MONEY
                row_tot += val
            ws.cell(row=r, column=ncol, value=row_tot).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        grand = 0
        for yi, y in enumerate(years):
            yt = sum(rc_data.get(y, {}).get(p, 0) for p in pools)
            ws.cell(row=r, column=2 + yi, value=yt).number_format = MONEY
            grand += yt
        ws.cell(row=r, column=ncol, value=grand).number_format = MONEY
        style_rows(ws, start2, r, ncol, mcols=set(range(2, ncol + 1)))

        # Net Charge offs
        r += 2
        headers3 = ["Net Charge offs"] + year_strs + ["Net Charge offs"]
        for hi, h in enumerate(headers3):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers3))
        start3 = r + 1
        grand_net = {y: 0 for y in years}
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_total = 0
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=net).number_format = MONEY
                grand_net[y] += net
                row_total += net
            ws.cell(row=r, column=len(headers3), value=row_total).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        net_total = 0
        for yi, y in enumerate(years):
            ws.cell(row=r, column=2 + yi, value=grand_net[y]).number_format = MONEY
            net_total += grand_net[y]
        ws.cell(row=r, column=len(headers3), value=net_total).number_format = MONEY
        style_rows(ws, start3, r, len(headers3), mcols=set(range(2, len(headers3) + 1)))

        # Life of Loan Loss Rate
        r += 2
        headers4 = ["Life Loss Rate"] + year_strs + ["Average"]
        for hi, h in enumerate(headers4):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers4))
        start4 = r + 1
        for pool in pool_names:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            rates = []
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                avg = avg_bals.get(y, {}).get(pool, 0)
                rate = net / avg if avg > 0 else 0
                ws.cell(row=r, column=2 + yi, value=rate).number_format = PCT4
                rates.append(rate)
            avg_rate = sum(rates) / len(rates) if rates else 0
            ws.cell(row=r, column=len(headers4), value=avg_rate).number_format = PCT4
        style_rows(ws, start4, r, len(headers4), pcols=set(range(2, len(headers4) + 1)))

        auto_w(ws, len(headers))


def sheet_delinquency(wb, cu, snap, config, hist=None):
    """Delinquency Calculation with historical data."""
    ws = wb.create_sheet("Delinquency Calculation")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Delinquency Calculation"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(set(config.get('pool_map', {}).values()))
    dq_pct = hist.get('dq_pct', {}) if hist else {}
    years = sorted(dq_pct.keys()) if dq_pct else list(range(2019, int(snap[:4]) + 1))
    year_strs = [str(y) for y in years]

    r = 5
    headers = ["DQ %"] + year_strs + ["Average", "Variance from Avg"]
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    start = r + 1
    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool)
        rates = []
        for yi, y in enumerate(years):
            val = dq_pct.get(y, {}).get(pool, 0)
            ws.cell(row=r, column=2 + yi, value=val).number_format = PCT
            rates.append(val)
        avg = sum(rates) / len(rates) if rates else 0
        ws.cell(row=r, column=len(headers) - 1, value=avg).number_format = PCT
        # Variance = most recent - average
        current = rates[-1] if rates else 0
        ws.cell(row=r, column=len(headers), value=current - avg).number_format = PCT
    style_rows(ws, start, r, len(headers), pcols=set(range(2, len(headers) + 1)))
    auto_w(ws, len(headers))


def sheet_balance_adj(wb, cu, snap, df, grades, config):
    """FAS 114 / Balance Adjustment sheet."""
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    ws = wb.create_sheet("Balance Adjustment")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Balance Adjustment"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(df['loan_pool'].unique())
    r = 5
    headers = ["Current Grade", "Loan Report Balance", "Bal Adjustment", "Balance Sheet Total"]

    for pool in pools:
        pdf = df[df['loan_pool'] == pool]
        ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=12)
        r += 1
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers))
        start = r + 1
        pool_total = 0
        for g in gl:
            r += 1
            bal = pdf[pdf['current_grade'] == g]['current_balance'].sum()
            pool_total += bal
            ws.cell(row=r, column=1, value=g)
            ws.cell(row=r, column=2, value=bal).number_format = MONEY
            ws.cell(row=r, column=3, value=0).number_format = MONEY
            ws.cell(row=r, column=4, value=bal).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=r, column=2, value=pool_total).number_format = MONEY
        ws.cell(row=r, column=3, value=0).number_format = MONEY
        ws.cell(row=r, column=4, value=pool_total).number_format = MONEY
        style_rows(ws, start, r, len(headers), mcols={2,3,4})
        r += 2

    # Grand total
    ws.cell(row=r, column=1, value="Grand Total").font = Font(bold=True, size=12)
    total = df['current_balance'].sum()
    ws.cell(row=r, column=2, value=total).number_format = MONEY
    ws.cell(row=r, column=3, value=0).number_format = MONEY
    ws.cell(row=r, column=4, value=total).number_format = MONEY
    auto_w(ws, len(headers))


def sheet_env_ranges(wb):
    """Environmental Factor Ranges reference table."""
    ws = wb.create_sheet("Env Factor Ranges")
    ws['A1'] = "Environmental Factor Ranges"
    ws['A1'].font = Font(bold=True, size=14)

    # Net Credit Change
    r = 3
    ws.cell(row=r, column=1, value="Net Credit Change").font = SUB_FONT
    ws.cell(row=r, column=1+0, value="Range"); ws.cell(row=r, column=2, value="Score")
    ws.cell(row=r, column=4, value="Delinquency").font = SUB_FONT
    ws.cell(row=r, column=4, value="Range"); ws.cell(row=r, column=5, value="Score")
    ws.cell(row=r, column=7, value="Economic Stress Score").font = SUB_FONT
    ws.cell(row=r, column=7, value="Range"); ws.cell(row=r, column=8, value="Score")
    hdr_row(ws, r, 8)

    ncc_rows = [
        ("<-18.00%", "7.00%"), ("-17.99% to -16.00%", "6.00%"),
        ("-15.99% to -14.00%", "5.00%"), ("-13.99% to -11.00%", "4.00%"),
        ("-10.99% to -8.00%", "3.00%"), ("-7.99% to -6.00%", "2.00%"),
        ("-5.99% to -4.00%", "1.00%"), ("-3.99% to 3.99%", "0.00%"),
        ("4.00% to 5.99%", "-1.00%"), ("6.00% to 7.99%", "-2.00%"),
        ("8.00% to 8.99%", "-3.00%"), ("9.00% to 10.99%", "-4.00%"),
        ("11.00% to 12.99%", "-5.00%"), ("13.00% to 14.99%", "-6.00%"),
        (">15.00%", "-7.00%"),
    ]
    dq_rows = [
        (">5.00%", "20.00%"), ("4.00% to 4.99%", "17.00%"),
        ("3.00% to 3.99%", "12.00%"), ("2.50% to 2.99%", "8.00%"),
        ("2.00% to 2.49%", "4.00%"), ("1.50% to 1.99%", "2.50%"),
        ("1.00% to 1.49%", "1.50%"), (".50% to .99%", "0.75%"),
        ("-.49% to .49%", "0.00%"), ("-.99% to -.50%", "-0.75%"),
        ("-1.49% to -1.00%", "-1.50%"), ("-1.99% to -1.50%", "-2.50%"),
        ("-2.49% to -2.00%", "-4.00%"), ("-2.99% to -2.50%", "-8.00%"),
        ("-3.99% to -3.00%", "-12.00%"), ("-4.99% to -4.00%", "-17.00%"),
        ("<-5.00%", "-20.00%"),
    ]
    es_rows = [
        (">25.00%", "10.00%"), ("24.00% to 24.99%", "8.00%"),
        ("22.00% to 23.99%", "7.00%"), ("20.00% to 21.99%", "6.00%"),
        ("18.00% to 19.99%", "5.00%"), ("16.00% to 17.99%", "4.00%"),
        ("14.00% to 15.99%", "3.50%"), ("12.00% to 13.99%", "3.00%"),
        ("10.00% to 11.99%", "2.00%"), ("8.00% to 9.99%", "1.00%"),
        ("6.00% to 7.99%", "0.00%"), ("4.00% to 5.99%", "0.00%"),
        ("2.00% to 3.99%", "-1.00%"), (".00% to 1.99%", "-2.00%"),
    ]

    for i, (rng, sc) in enumerate(ncc_rows):
        ws.cell(row=r + 1 + i, column=1, value=rng)
        ws.cell(row=r + 1 + i, column=2, value=sc)
    for i, (rng, sc) in enumerate(dq_rows):
        ws.cell(row=r + 1 + i, column=4, value=rng)
        ws.cell(row=r + 1 + i, column=5, value=sc)
    for i, (rng, sc) in enumerate(es_rows):
        ws.cell(row=r + 1 + i, column=7, value=rng)
        ws.cell(row=r + 1 + i, column=8, value=sc)
    auto_w(ws, 8)


def sheet_grade_config(wb, grades, config):
    """Grade ranges & loan code reference."""
    ws = wb.create_sheet("Grade Ranges & Loan Codes")
    ws['A1'] = "Credit Grade Configuration"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["Grade", "Score Range", "Reserve Rate"]
    r = 3
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    for g in grades:
        r += 1
        ws.cell(row=r, column=1, value=g['label'])
        ws.cell(row=r, column=2, value=f"{g['min_score']}-{g['max_score']}")
        ws.cell(row=r, column=3, value=g['reserve_rate']).number_format = PCT

    r += 3
    ws.cell(row=r, column=1, value="Loan Type Codes").font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1, value="Code"); ws.cell(row=r, column=2, value="Loan Pool")
    hdr_row(ws, r, 2)
    for code, pool in sorted(config.get('pool_map', {}).items(), key=lambda x: x[1]):
        r += 1
        ws.cell(row=r, column=1, value=str(code))
        ws.cell(row=r, column=2, value=pool)
    auto_w(ws, 3)


def sheet_all_loans(wb, cu, snap, df, grades, config):
    """All Loans detail listing."""
    ws = wb.create_sheet("All Loans")
    no_score = config.get('no_score_label', 'Not Reported')
    ws['A1'] = "Credit Grade Analysis - All Loans"
    ws['A1'].font = Font(bold=True, size=14)
    ws['F1'] = snap

    headers = ["Member #", "Loan Pool", "Current Balance",
               "Original Score", "Original Grade",
               "Current Score", "Current Grade",
               "Migration Status", "Reserve Rate", "Expected Loss"]
    r = 2
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))

    start = r + 1
    for _, loan in df.iterrows():
        r += 1
        ws.cell(row=r, column=1, value=str(loan.get('member_number', '')))
        ws.cell(row=r, column=2, value=loan.get('loan_pool', ''))
        ws.cell(row=r, column=3, value=loan.get('current_balance', 0))
        ws.cell(row=r, column=4, value=loan.get('original_fico_score', 0))
        ws.cell(row=r, column=5, value=loan.get('original_grade', no_score))
        ws.cell(row=r, column=6, value=loan.get('current_fico_score', 0))
        ws.cell(row=r, column=7, value=loan.get('current_grade', no_score))
        ws.cell(row=r, column=8, value=loan.get('migration_status', 'Unchanged'))
        ws.cell(row=r, column=9, value=loan.get('reserve_rate', 0))
        ws.cell(row=r, column=10, value=loan.get('expected_loss_amount', 0))
        status = loan.get('migration_status', 'Unchanged')
        if status == 'Improved':
            ws.cell(row=r, column=8).fill = IMP_FILL
        elif status == 'Deteriorated':
            ws.cell(row=r, column=8).fill = DET_FILL
    style_rows(ws, start, r, len(headers), mcols={3, 10}, pcols={9})
    auto_w(ws, len(headers), mx=18)


def sheet_hist_balances(wb, cu, snap, df, grades, config, hist=None):
    """Historical Loan Balances by pool with monthly balance data."""
    ws = wb.create_sheet("Historical Balances")
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Historical Loan Balances by Pool"
    ws['A3'] = f"For Period Ending {snap}"

    monthly = hist.get('monthly_balances', pd.DataFrame()) if hist else pd.DataFrame()

    if not monthly.empty:
        # Use monthly balance data - show quarterly snapshots per pool
        # Get quarter-end dates (month-end for Mar, Jun, Sep, Dec)
        monthly['quarter'] = monthly['date'].dt.to_period('Q')
        # Get last date per quarter per pool
        qtr_data = monthly.groupby(['pool', 'quarter']).last().reset_index()
        quarters = sorted(qtr_data['quarter'].unique())
        # Limit to last 20 quarters to keep sheet manageable
        if len(quarters) > 20:
            quarters = quarters[-20:]
        qtr_strs = [str(q) for q in quarters]

        pool_names = sorted(qtr_data['pool'].unique())
        r = 5
        headers = ["Pool"] + qtr_strs
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pool_names:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            for qi, q in enumerate(quarters):
                val = qtr_data[(qtr_data['pool'] == pool) & (qtr_data['quarter'] == q)]
                bal = val['balance'].values[0] if len(val) > 0 else 0
                ws.cell(row=r, column=2 + qi, value=bal).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        for qi, q in enumerate(quarters):
            total = qtr_data[qtr_data['quarter'] == q]['balance'].sum()
            ws.cell(row=r, column=2 + qi, value=total).number_format = MONEY
        style_rows(ws, start, r, ncol, mcols=set(range(2, ncol + 1)))
        auto_w(ws, ncol)
    else:
        # Fallback: just show current data by grade
        pools = sorted(df['loan_pool'].unique())
        r = 6
        for pool in pools:
            ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=12)
            r += 1
            ws.cell(row=r, column=1, value="Current Grade")
            ws.cell(row=r, column=2, value=snap)
            hdr_row(ws, r, 2)
            pdf = df[df['loan_pool'] == pool]
            for g in gl:
                r += 1
                ws.cell(row=r, column=1, value=g)
                ws.cell(row=r, column=2, value=pdf[pdf['current_grade'] == g]['current_balance'].sum()).number_format = MONEY
            r += 1
            ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
            ws.cell(row=r, column=2, value=pdf['current_balance'].sum()).number_format = MONEY
            r += 2
        auto_w(ws, 2)


def sheet_loss_factor_hist(wb, cu, snap, df, grades, config, hist=None):
    """Loss Factor Historical Detail with charge-off/recovery data."""
    ws = wb.create_sheet("Loss Factor Historical")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Loss Factor Historical Detail"
    ws['A3'] = f"For Period Ending {snap}"

    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    avg_bals = hist.get('avg_balances', {}) if hist else {}
    years = hist.get('years', []) if hist else []

    pools = sorted(set(config.get('pool_map', {}).values()))

    if years:
        year_strs = [str(y) for y in years]
        r = 5
        # Net charge-off rates per pool per year
        headers = ["Pool"] + year_strs + ["Average Life\nLoss Rate"]
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            rates = []
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                avg = avg_bals.get(y, {}).get(pool, 0)
                rate = net / avg if avg > 0 else 0
                ws.cell(row=r, column=2 + yi, value=rate).number_format = PCT4
                rates.append(rate)
            avg_rate = sum(rates) / len(rates) if rates else 0
            ws.cell(row=r, column=ncol, value=avg_rate).number_format = PCT4

        style_rows(ws, start, r, ncol, pcols=set(range(2, ncol + 1)))

        # Average balances section
        r += 3
        ws.cell(row=r, column=1, value="Average Balances by Pool").font = SUB_FONT
        r += 1
        headers2 = ["Pool"] + year_strs
        ncol2 = len(headers2)
        for hi, h in enumerate(headers2):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol2)
        start2 = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            for yi, y in enumerate(years):
                avg = avg_bals.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=avg).number_format = MONEY
        style_rows(ws, start2, r, ncol2, mcols=set(range(2, ncol2 + 1)))
        auto_w(ws, max(ncol, ncol2))
    else:
        ws['A5'] = "No historical data available."
        ws['A5'].font = Font(italic=True, color='888888')


def sheet_chargeoff_hist(wb, cu, snap, config, hist=None):
    """Charge off / Recoveries Historical Detail."""
    ws = wb.create_sheet("Chargeoff Historical")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Charge off and Recoveries Historical Detail"
    ws['A3'] = f"For Period Ending {snap}"

    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    years = hist.get('years', []) if hist else []
    pools = sorted(set(config.get('pool_map', {}).values()))

    if years:
        year_strs = [str(y) for y in years]

        # Charge-offs by pool by year
        r = 5
        ws.cell(row=r, column=1, value="Charge offs by Year").font = SUB_FONT
        r += 1
        headers = ["Pool"] + year_strs + ["Total"]
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_tot = 0
            for yi, y in enumerate(years):
                val = co_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=val).number_format = MONEY
                row_tot += val
            ws.cell(row=r, column=ncol, value=row_tot).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        grand = 0
        for yi, y in enumerate(years):
            yt = sum(co_data.get(y, {}).get(p, 0) for p in pools)
            ws.cell(row=r, column=2 + yi, value=yt).number_format = MONEY
            grand += yt
        ws.cell(row=r, column=ncol, value=grand).number_format = MONEY
        style_rows(ws, start, r, ncol, mcols=set(range(2, ncol + 1)))

        # Recoveries by pool by year
        r += 3
        ws.cell(row=r, column=1, value="Recoveries by Year").font = SUB_FONT
        r += 1
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start2 = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_tot = 0
            for yi, y in enumerate(years):
                val = rc_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=val).number_format = MONEY
                row_tot += val
            ws.cell(row=r, column=ncol, value=row_tot).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        grand = 0
        for yi, y in enumerate(years):
            yt = sum(rc_data.get(y, {}).get(p, 0) for p in pools)
            ws.cell(row=r, column=2 + yi, value=yt).number_format = MONEY
            grand += yt
        ws.cell(row=r, column=ncol, value=grand).number_format = MONEY
        style_rows(ws, start2, r, ncol, mcols=set(range(2, ncol + 1)))

        # Net Charge offs
        r += 2
        headers3 = ["Net Charge offs"] + year_strs + ["Net Charge offs"]
        for hi, h in enumerate(headers3):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers3))
        start3 = r + 1
        grand_net = {y: 0 for y in years}
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_total = 0
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=net).number_format = MONEY
                grand_net[y] += net
                row_total += net
            ws.cell(row=r, column=len(headers3), value=row_total).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        net_total = 0
        for yi, y in enumerate(years):
            ws.cell(row=r, column=2 + yi, value=grand_net[y]).number_format = MONEY
            net_total += grand_net[y]
        ws.cell(row=r, column=len(headers3), value=net_total).number_format = MONEY
        style_rows(ws, start3, r, len(headers3), mcols=set(range(2, len(headers3) + 1)))

        # Life of Loan Loss Rate
        r += 2
        headers4 = ["Life Loss Rate"] + year_strs + ["Average"]
        for hi, h in enumerate(headers4):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers4))
        start4 = r + 1
        for pool in pool_names:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            rates = []
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                avg = avg_bals.get(y, {}).get(pool, 0)
                rate = net / avg if avg > 0 else 0
                ws.cell(row=r, column=2 + yi, value=rate).number_format = PCT4
                rates.append(rate)
            avg_rate = sum(rates) / len(rates) if rates else 0
            ws.cell(row=r, column=len(headers4), value=avg_rate).number_format = PCT4
        style_rows(ws, start4, r, len(headers4), pcols=set(range(2, len(headers4) + 1)))

        auto_w(ws, len(headers))


def sheet_delinquency(wb, cu, snap, config, hist=None):
    """Delinquency Calculation with historical data."""
    ws = wb.create_sheet("Delinquency Calculation")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Delinquency Calculation"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(set(config.get('pool_map', {}).values()))
    dq_pct = hist.get('dq_pct', {}) if hist else {}
    years = sorted(dq_pct.keys()) if dq_pct else list(range(2019, int(snap[:4]) + 1))
    year_strs = [str(y) for y in years]

    r = 5
    headers = ["DQ %"] + year_strs + ["Average", "Variance from Avg"]
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    start = r + 1
    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool)
        rates = []
        for yi, y in enumerate(years):
            val = dq_pct.get(y, {}).get(pool, 0)
            ws.cell(row=r, column=2 + yi, value=val).number_format = PCT
            rates.append(val)
        avg = sum(rates) / len(rates) if rates else 0
        ws.cell(row=r, column=len(headers) - 1, value=avg).number_format = PCT
        # Variance = most recent - average
        current = rates[-1] if rates else 0
        ws.cell(row=r, column=len(headers), value=current - avg).number_format = PCT
    style_rows(ws, start, r, len(headers), pcols=set(range(2, len(headers) + 1)))
    auto_w(ws, len(headers))


def sheet_balance_adj(wb, cu, snap, df, grades, config):
    """FAS 114 / Balance Adjustment sheet."""
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    ws = wb.create_sheet("Balance Adjustment")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Balance Adjustment"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(df['loan_pool'].unique())
    r = 5
    headers = ["Current Grade", "Loan Report Balance", "Bal Adjustment", "Balance Sheet Total"]

    for pool in pools:
        pdf = df[df['loan_pool'] == pool]
        ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=12)
        r += 1
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers))
        start = r + 1
        pool_total = 0
        for g in gl:
            r += 1
            bal = pdf[pdf['current_grade'] == g]['current_balance'].sum()
            pool_total += bal
            ws.cell(row=r, column=1, value=g)
            ws.cell(row=r, column=2, value=bal).number_format = MONEY
            ws.cell(row=r, column=3, value=0).number_format = MONEY
            ws.cell(row=r, column=4, value=bal).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=r, column=2, value=pool_total).number_format = MONEY
        ws.cell(row=r, column=3, value=0).number_format = MONEY
        ws.cell(row=r, column=4, value=pool_total).number_format = MONEY
        style_rows(ws, start, r, len(headers), mcols={2,3,4})
        r += 2

    # Grand total
    ws.cell(row=r, column=1, value="Grand Total").font = Font(bold=True, size=12)
    total = df['current_balance'].sum()
    ws.cell(row=r, column=2, value=total).number_format = MONEY
    ws.cell(row=r, column=3, value=0).number_format = MONEY
    ws.cell(row=r, column=4, value=total).number_format = MONEY
    auto_w(ws, len(headers))


def sheet_env_ranges(wb):
    """Environmental Factor Ranges reference table."""
    ws = wb.create_sheet("Env Factor Ranges")
    ws['A1'] = "Environmental Factor Ranges"
    ws['A1'].font = Font(bold=True, size=14)

    # Net Credit Change
    r = 3
    ws.cell(row=r, column=1, value="Net Credit Change").font = SUB_FONT
    ws.cell(row=r, column=1+0, value="Range"); ws.cell(row=r, column=2, value="Score")
    ws.cell(row=r, column=4, value="Delinquency").font = SUB_FONT
    ws.cell(row=r, column=4, value="Range"); ws.cell(row=r, column=5, value="Score")
    ws.cell(row=r, column=7, value="Economic Stress Score").font = SUB_FONT
    ws.cell(row=r, column=7, value="Range"); ws.cell(row=r, column=8, value="Score")
    hdr_row(ws, r, 8)

    ncc_rows = [
        ("<-18.00%", "7.00%"), ("-17.99% to -16.00%", "6.00%"),
        ("-15.99% to -14.00%", "5.00%"), ("-13.99% to -11.00%", "4.00%"),
        ("-10.99% to -8.00%", "3.00%"), ("-7.99% to -6.00%", "2.00%"),
        ("-5.99% to -4.00%", "1.00%"), ("-3.99% to 3.99%", "0.00%"),
        ("4.00% to 5.99%", "-1.00%"), ("6.00% to 7.99%", "-2.00%"),
        ("8.00% to 8.99%", "-3.00%"), ("9.00% to 10.99%", "-4.00%"),
        ("11.00% to 12.99%", "-5.00%"), ("13.00% to 14.99%", "-6.00%"),
        (">15.00%", "-7.00%"),
    ]
    dq_rows = [
        (">5.00%", "20.00%"), ("4.00% to 4.99%", "17.00%"),
        ("3.00% to 3.99%", "12.00%"), ("2.50% to 2.99%", "8.00%"),
        ("2.00% to 2.49%", "4.00%"), ("1.50% to 1.99%", "2.50%"),
        ("1.00% to 1.49%", "1.50%"), (".50% to .99%", "0.75%"),
        ("-.49% to .49%", "0.00%"), ("-.99% to -.50%", "-0.75%"),
        ("-1.49% to -1.00%", "-1.50%"), ("-1.99% to -1.50%", "-2.50%"),
        ("-2.49% to -2.00%", "-4.00%"), ("-2.99% to -2.50%", "-8.00%"),
        ("-3.99% to -3.00%", "-12.00%"), ("-4.99% to -4.00%", "-17.00%"),
        ("<-5.00%", "-20.00%"),
    ]
    es_rows = [
        (">25.00%", "10.00%"), ("24.00% to 24.99%", "8.00%"),
        ("22.00% to 23.99%", "7.00%"), ("20.00% to 21.99%", "6.00%"),
        ("18.00% to 19.99%", "5.00%"), ("16.00% to 17.99%", "4.00%"),
        ("14.00% to 15.99%", "3.50%"), ("12.00% to 13.99%", "3.00%"),
        ("10.00% to 11.99%", "2.00%"), ("8.00% to 9.99%", "1.00%"),
        ("6.00% to 7.99%", "0.00%"), ("4.00% to 5.99%", "0.00%"),
        ("2.00% to 3.99%", "-1.00%"), (".00% to 1.99%", "-2.00%"),
    ]

    for i, (rng, sc) in enumerate(ncc_rows):
        ws.cell(row=r + 1 + i, column=1, value=rng)
        ws.cell(row=r + 1 + i, column=2, value=sc)
    for i, (rng, sc) in enumerate(dq_rows):
        ws.cell(row=r + 1 + i, column=4, value=rng)
        ws.cell(row=r + 1 + i, column=5, value=sc)
    for i, (rng, sc) in enumerate(es_rows):
        ws.cell(row=r + 1 + i, column=7, value=rng)
        ws.cell(row=r + 1 + i, column=8, value=sc)
    auto_w(ws, 8)


def sheet_grade_config(wb, grades, config):
    """Grade ranges & loan code reference."""
    ws = wb.create_sheet("Grade Ranges & Loan Codes")
    ws['A1'] = "Credit Grade Configuration"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["Grade", "Score Range", "Reserve Rate"]
    r = 3
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    for g in grades:
        r += 1
        ws.cell(row=r, column=1, value=g['label'])
        ws.cell(row=r, column=2, value=f"{g['min_score']}-{g['max_score']}")
        ws.cell(row=r, column=3, value=g['reserve_rate']).number_format = PCT

    r += 3
    ws.cell(row=r, column=1, value="Loan Type Codes").font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1, value="Code"); ws.cell(row=r, column=2, value="Loan Pool")
    hdr_row(ws, r, 2)
    for code, pool in sorted(config.get('pool_map', {}).items(), key=lambda x: x[1]):
        r += 1
        ws.cell(row=r, column=1, value=str(code))
        ws.cell(row=r, column=2, value=pool)
    auto_w(ws, 3)


def sheet_all_loans(wb, cu, snap, df, grades, config):
    """All Loans detail listing."""
    ws = wb.create_sheet("All Loans")
    no_score = config.get('no_score_label', 'Not Reported')
    ws['A1'] = "Credit Grade Analysis - All Loans"
    ws['A1'].font = Font(bold=True, size=14)
    ws['F1'] = snap

    headers = ["Member #", "Loan Pool", "Current Balance",
               "Original Score", "Original Grade",
               "Current Score", "Current Grade",
               "Migration Status", "Reserve Rate", "Expected Loss"]
    r = 2
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))

    start = r + 1
    for _, loan in df.iterrows():
        r += 1
        ws.cell(row=r, column=1, value=str(loan.get('member_number', '')))
        ws.cell(row=r, column=2, value=loan.get('loan_pool', ''))
        ws.cell(row=r, column=3, value=loan.get('current_balance', 0))
        ws.cell(row=r, column=4, value=loan.get('original_fico_score', 0))
        ws.cell(row=r, column=5, value=loan.get('original_grade', no_score))
        ws.cell(row=r, column=6, value=loan.get('current_fico_score', 0))
        ws.cell(row=r, column=7, value=loan.get('current_grade', no_score))
        ws.cell(row=r, column=8, value=loan.get('migration_status', 'Unchanged'))
        ws.cell(row=r, column=9, value=loan.get('reserve_rate', 0))
        ws.cell(row=r, column=10, value=loan.get('expected_loss_amount', 0))
        status = loan.get('migration_status', 'Unchanged')
        if status == 'Improved':
            ws.cell(row=r, column=8).fill = IMP_FILL
        elif status == 'Deteriorated':
            ws.cell(row=r, column=8).fill = DET_FILL
    style_rows(ws, start, r, len(headers), mcols={3, 10}, pcols={9})
    auto_w(ws, len(headers), mx=18)


def sheet_hist_balances(wb, cu, snap, df, grades, config, hist=None):
    """Historical Loan Balances by pool with monthly balance data."""
    ws = wb.create_sheet("Historical Balances")
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Historical Loan Balances by Pool"
    ws['A3'] = f"For Period Ending {snap}"

    monthly = hist.get('monthly_balances', pd.DataFrame()) if hist else pd.DataFrame()

    if not monthly.empty:
        # Use monthly balance data - show quarterly snapshots per pool
        # Get quarter-end dates (month-end for Mar, Jun, Sep, Dec)
        monthly['quarter'] = monthly['date'].dt.to_period('Q')
        # Get last date per quarter per pool
        qtr_data = monthly.groupby(['pool', 'quarter']).last().reset_index()
        quarters = sorted(qtr_data['quarter'].unique())
        # Limit to last 20 quarters to keep sheet manageable
        if len(quarters) > 20:
            quarters = quarters[-20:]
        qtr_strs = [str(q) for q in quarters]

        pool_names = sorted(qtr_data['pool'].unique())
        r = 5
        headers = ["Pool"] + qtr_strs
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pool_names:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            for qi, q in enumerate(quarters):
                val = qtr_data[(qtr_data['pool'] == pool) & (qtr_data['quarter'] == q)]
                bal = val['balance'].values[0] if len(val) > 0 else 0
                ws.cell(row=r, column=2 + qi, value=bal).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        for qi, q in enumerate(quarters):
            total = qtr_data[qtr_data['quarter'] == q]['balance'].sum()
            ws.cell(row=r, column=2 + qi, value=total).number_format = MONEY
        style_rows(ws, start, r, ncol, mcols=set(range(2, ncol + 1)))
        auto_w(ws, ncol)
    else:
        # Fallback: just show current data by grade
        pools = sorted(df['loan_pool'].unique())
        r = 6
        for pool in pools:
            ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=12)
            r += 1
            ws.cell(row=r, column=1, value="Current Grade")
            ws.cell(row=r, column=2, value=snap)
            hdr_row(ws, r, 2)
            pdf = df[df['loan_pool'] == pool]
            for g in gl:
                r += 1
                ws.cell(row=r, column=1, value=g)
                ws.cell(row=r, column=2, value=pdf[pdf['current_grade'] == g]['current_balance'].sum()).number_format = MONEY
            r += 1
            ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
            ws.cell(row=r, column=2, value=pdf['current_balance'].sum()).number_format = MONEY
            r += 2
        auto_w(ws, 2)


def sheet_loss_factor_hist(wb, cu, snap, df, grades, config, hist=None):
    """Loss Factor Historical Detail with charge-off/recovery data."""
    ws = wb.create_sheet("Loss Factor Historical")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Loss Factor Historical Detail"
    ws['A3'] = f"For Period Ending {snap}"

    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    avg_bals = hist.get('avg_balances', {}) if hist else {}
    years = hist.get('years', []) if hist else []

    pools = sorted(set(config.get('pool_map', {}).values()))

    if years:
        year_strs = [str(y) for y in years]
        r = 5
        # Net charge-off rates per pool per year
        headers = ["Pool"] + year_strs + ["Average Life\nLoss Rate"]
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            rates = []
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                avg = avg_bals.get(y, {}).get(pool, 0)
                rate = net / avg if avg > 0 else 0
                ws.cell(row=r, column=2 + yi, value=rate).number_format = PCT4
                rates.append(rate)
            avg_rate = sum(rates) / len(rates) if rates else 0
            ws.cell(row=r, column=ncol, value=avg_rate).number_format = PCT4

        style_rows(ws, start, r, ncol, pcols=set(range(2, ncol + 1)))

        # Average balances section
        r += 3
        ws.cell(row=r, column=1, value="Average Balances by Pool").font = SUB_FONT
        r += 1
        headers2 = ["Pool"] + year_strs
        ncol2 = len(headers2)
        for hi, h in enumerate(headers2):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol2)
        start2 = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            for yi, y in enumerate(years):
                avg = avg_bals.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=avg).number_format = MONEY
        style_rows(ws, start2, r, ncol2, mcols=set(range(2, ncol2 + 1)))
        auto_w(ws, max(ncol, ncol2))
    else:
        ws['A5'] = "No historical data available."
        ws['A5'].font = Font(italic=True, color='888888')


def sheet_chargeoff_hist(wb, cu, snap, config, hist=None):
    """Charge off / Recoveries Historical Detail."""
    ws = wb.create_sheet("Chargeoff Historical")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Charge off and Recoveries Historical Detail"
    ws['A3'] = f"For Period Ending {snap}"

    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    years = hist.get('years', []) if hist else []
    pools = sorted(set(config.get('pool_map', {}).values()))

    if years:
        year_strs = [str(y) for y in years]

        # Charge-offs by pool by year
        r = 5
        ws.cell(row=r, column=1, value="Charge offs by Year").font = SUB_FONT
        r += 1
        headers = ["Pool"] + year_strs + ["Total"]
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_tot = 0
            for yi, y in enumerate(years):
                val = co_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=val).number_format = MONEY
                row_tot += val
            ws.cell(row=r, column=ncol, value=row_tot).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        grand = 0
        for yi, y in enumerate(years):
            yt = sum(co_data.get(y, {}).get(p, 0) for p in pools)
            ws.cell(row=r, column=2 + yi, value=yt).number_format = MONEY
            grand += yt
        ws.cell(row=r, column=ncol, value=grand).number_format = MONEY
        style_rows(ws, start, r, ncol, mcols=set(range(2, ncol + 1)))

        # Recoveries by pool by year
        r += 3
        ws.cell(row=r, column=1, value="Recoveries by Year").font = SUB_FONT
        r += 1
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start2 = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_tot = 0
            for yi, y in enumerate(years):
                val = rc_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=val).number_format = MONEY
                row_tot += val
            ws.cell(row=r, column=ncol, value=row_tot).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        grand = 0
        for yi, y in enumerate(years):
            yt = sum(rc_data.get(y, {}).get(p, 0) for p in pools)
            ws.cell(row=r, column=2 + yi, value=yt).number_format = MONEY
            grand += yt
        ws.cell(row=r, column=ncol, value=grand).number_format = MONEY
        style_rows(ws, start2, r, ncol, mcols=set(range(2, ncol + 1)))

        # Net Charge offs
        r += 2
        headers3 = ["Net Charge offs"] + year_strs + ["Net Charge offs"]
        for hi, h in enumerate(headers3):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers3))
        start3 = r + 1
        grand_net = {y: 0 for y in years}
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_total = 0
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=net).number_format = MONEY
                grand_net[y] += net
                row_total += net
            ws.cell(row=r, column=len(headers3), value=row_total).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        net_total = 0
        for yi, y in enumerate(years):
            ws.cell(row=r, column=2 + yi, value=grand_net[y]).number_format = MONEY
            net_total += grand_net[y]
        ws.cell(row=r, column=len(headers3), value=net_total).number_format = MONEY
        style_rows(ws, start3, r, len(headers3), mcols=set(range(2, len(headers3) + 1)))

        # Life of Loan Loss Rate
        r += 2
        headers4 = ["Life Loss Rate"] + year_strs + ["Average"]
        for hi, h in enumerate(headers4):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers4))
        start4 = r + 1
        for pool in pool_names:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            rates = []
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                avg = avg_bals.get(y, {}).get(pool, 0)
                rate = net / avg if avg > 0 else 0
                ws.cell(row=r, column=2 + yi, value=rate).number_format = PCT4
                rates.append(rate)
            avg_rate = sum(rates) / len(rates) if rates else 0
            ws.cell(row=r, column=len(headers4), value=avg_rate).number_format = PCT4
        style_rows(ws, start4, r, len(headers4), pcols=set(range(2, len(headers4) + 1)))

        auto_w(ws, len(headers))


def sheet_delinquency(wb, cu, snap, config, hist=None):
    """Delinquency Calculation with historical data."""
    ws = wb.create_sheet("Delinquency Calculation")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Delinquency Calculation"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(set(config.get('pool_map', {}).values()))
    dq_pct = hist.get('dq_pct', {}) if hist else {}
    years = sorted(dq_pct.keys()) if dq_pct else list(range(2019, int(snap[:4]) + 1))
    year_strs = [str(y) for y in years]

    r = 5
    headers = ["DQ %"] + year_strs + ["Average", "Variance from Avg"]
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    start = r + 1
    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool)
        rates = []
        for yi, y in enumerate(years):
            val = dq_pct.get(y, {}).get(pool, 0)
            ws.cell(row=r, column=2 + yi, value=val).number_format = PCT
            rates.append(val)
        avg = sum(rates) / len(rates) if rates else 0
        ws.cell(row=r, column=len(headers) - 1, value=avg).number_format = PCT
        # Variance = most recent - average
        current = rates[-1] if rates else 0
        ws.cell(row=r, column=len(headers), value=current - avg).number_format = PCT
    style_rows(ws, start, r, len(headers), pcols=set(range(2, len(headers) + 1)))
    auto_w(ws, len(headers))


def sheet_balance_adj(wb, cu, snap, df, grades, config):
    """FAS 114 / Balance Adjustment sheet."""
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    ws = wb.create_sheet("Balance Adjustment")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Balance Adjustment"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(df['loan_pool'].unique())
    r = 5
    headers = ["Current Grade", "Loan Report Balance", "Bal Adjustment", "Balance Sheet Total"]

    for pool in pools:
        pdf = df[df['loan_pool'] == pool]
        ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=12)
        r += 1
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers))
        start = r + 1
        pool_total = 0
        for g in gl:
            r += 1
            bal = pdf[pdf['current_grade'] == g]['current_balance'].sum()
            pool_total += bal
            ws.cell(row=r, column=1, value=g)
            ws.cell(row=r, column=2, value=bal).number_format = MONEY
            ws.cell(row=r, column=3, value=0).number_format = MONEY
            ws.cell(row=r, column=4, value=bal).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=r, column=2, value=pool_total).number_format = MONEY
        ws.cell(row=r, column=3, value=0).number_format = MONEY
        ws.cell(row=r, column=4, value=pool_total).number_format = MONEY
        style_rows(ws, start, r, len(headers), mcols={2,3,4})
        r += 2

    # Grand total
    ws.cell(row=r, column=1, value="Grand Total").font = Font(bold=True, size=12)
    total = df['current_balance'].sum()
    ws.cell(row=r, column=2, value=total).number_format = MONEY
    ws.cell(row=r, column=3, value=0).number_format = MONEY
    ws.cell(row=r, column=4, value=total).number_format = MONEY
    auto_w(ws, len(headers))


def sheet_env_ranges(wb):
    """Environmental Factor Ranges reference table."""
    ws = wb.create_sheet("Env Factor Ranges")
    ws['A1'] = "Environmental Factor Ranges"
    ws['A1'].font = Font(bold=True, size=14)

    # Net Credit Change
    r = 3
    ws.cell(row=r, column=1, value="Net Credit Change").font = SUB_FONT
    ws.cell(row=r, column=1+0, value="Range"); ws.cell(row=r, column=2, value="Score")
    ws.cell(row=r, column=4, value="Delinquency").font = SUB_FONT
    ws.cell(row=r, column=4, value="Range"); ws.cell(row=r, column=5, value="Score")
    ws.cell(row=r, column=7, value="Economic Stress Score").font = SUB_FONT
    ws.cell(row=r, column=7, value="Range"); ws.cell(row=r, column=8, value="Score")
    hdr_row(ws, r, 8)

    ncc_rows = [
        ("<-18.00%", "7.00%"), ("-17.99% to -16.00%", "6.00%"),
        ("-15.99% to -14.00%", "5.00%"), ("-13.99% to -11.00%", "4.00%"),
        ("-10.99% to -8.00%", "3.00%"), ("-7.99% to -6.00%", "2.00%"),
        ("-5.99% to -4.00%", "1.00%"), ("-3.99% to 3.99%", "0.00%"),
        ("4.00% to 5.99%", "-1.00%"), ("6.00% to 7.99%", "-2.00%"),
        ("8.00% to 8.99%", "-3.00%"), ("9.00% to 10.99%", "-4.00%"),
        ("11.00% to 12.99%", "-5.00%"), ("13.00% to 14.99%", "-6.00%"),
        (">15.00%", "-7.00%"),
    ]
    dq_rows = [
        (">5.00%", "20.00%"), ("4.00% to 4.99%", "17.00%"),
        ("3.00% to 3.99%", "12.00%"), ("2.50% to 2.99%", "8.00%"),
        ("2.00% to 2.49%", "4.00%"), ("1.50% to 1.99%", "2.50%"),
        ("1.00% to 1.49%", "1.50%"), (".50% to .99%", "0.75%"),
        ("-.49% to .49%", "0.00%"), ("-.99% to -.50%", "-0.75%"),
        ("-1.49% to -1.00%", "-1.50%"), ("-1.99% to -1.50%", "-2.50%"),
        ("-2.49% to -2.00%", "-4.00%"), ("-2.99% to -2.50%", "-8.00%"),
        ("-3.99% to -3.00%", "-12.00%"), ("-4.99% to -4.00%", "-17.00%"),
        ("<-5.00%", "-20.00%"),
    ]
    es_rows = [
        (">25.00%", "10.00%"), ("24.00% to 24.99%", "8.00%"),
        ("22.00% to 23.99%", "7.00%"), ("20.00% to 21.99%", "6.00%"),
        ("18.00% to 19.99%", "5.00%"), ("16.00% to 17.99%", "4.00%"),
        ("14.00% to 15.99%", "3.50%"), ("12.00% to 13.99%", "3.00%"),
        ("10.00% to 11.99%", "2.00%"), ("8.00% to 9.99%", "1.00%"),
        ("6.00% to 7.99%", "0.00%"), ("4.00% to 5.99%", "0.00%"),
        ("2.00% to 3.99%", "-1.00%"), (".00% to 1.99%", "-2.00%"),
    ]

    for i, (rng, sc) in enumerate(ncc_rows):
        ws.cell(row=r + 1 + i, column=1, value=rng)
        ws.cell(row=r + 1 + i, column=2, value=sc)
    for i, (rng, sc) in enumerate(dq_rows):
        ws.cell(row=r + 1 + i, column=4, value=rng)
        ws.cell(row=r + 1 + i, column=5, value=sc)
    for i, (rng, sc) in enumerate(es_rows):
        ws.cell(row=r + 1 + i, column=7, value=rng)
        ws.cell(row=r + 1 + i, column=8, value=sc)
    auto_w(ws, 8)


def sheet_grade_config(wb, grades, config):
    """Grade ranges & loan code reference."""
    ws = wb.create_sheet("Grade Ranges & Loan Codes")
    ws['A1'] = "Credit Grade Configuration"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["Grade", "Score Range", "Reserve Rate"]
    r = 3
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    for g in grades:
        r += 1
        ws.cell(row=r, column=1, value=g['label'])
        ws.cell(row=r, column=2, value=f"{g['min_score']}-{g['max_score']}")
        ws.cell(row=r, column=3, value=g['reserve_rate']).number_format = PCT

    r += 3
    ws.cell(row=r, column=1, value="Loan Type Codes").font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1, value="Code"); ws.cell(row=r, column=2, value="Loan Pool")
    hdr_row(ws, r, 2)
    for code, pool in sorted(config.get('pool_map', {}).items(), key=lambda x: x[1]):
        r += 1
        ws.cell(row=r, column=1, value=str(code))
        ws.cell(row=r, column=2, value=pool)
    auto_w(ws, 3)


def sheet_all_loans(wb, cu, snap, df, grades, config):
    """All Loans detail listing."""
    ws = wb.create_sheet("All Loans")
    no_score = config.get('no_score_label', 'Not Reported')
    ws['A1'] = "Credit Grade Analysis - All Loans"
    ws['A1'].font = Font(bold=True, size=14)
    ws['F1'] = snap

    headers = ["Member #", "Loan Pool", "Current Balance",
               "Original Score", "Original Grade",
               "Current Score", "Current Grade",
               "Migration Status", "Reserve Rate", "Expected Loss"]
    r = 2
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))

    start = r + 1
    for _, loan in df.iterrows():
        r += 1
        ws.cell(row=r, column=1, value=str(loan.get('member_number', '')))
        ws.cell(row=r, column=2, value=loan.get('loan_pool', ''))
        ws.cell(row=r, column=3, value=loan.get('current_balance', 0))
        ws.cell(row=r, column=4, value=loan.get('original_fico_score', 0))
        ws.cell(row=r, column=5, value=loan.get('original_grade', no_score))
        ws.cell(row=r, column=6, value=loan.get('current_fico_score', 0))
        ws.cell(row=r, column=7, value=loan.get('current_grade', no_score))
        ws.cell(row=r, column=8, value=loan.get('migration_status', 'Unchanged'))
        ws.cell(row=r, column=9, value=loan.get('reserve_rate', 0))
        ws.cell(row=r, column=10, value=loan.get('expected_loss_amount', 0))
        status = loan.get('migration_status', 'Unchanged')
        if status == 'Improved':
            ws.cell(row=r, column=8).fill = IMP_FILL
        elif status == 'Deteriorated':
            ws.cell(row=r, column=8).fill = DET_FILL
    style_rows(ws, start, r, len(headers), mcols={3, 10}, pcols={9})
    auto_w(ws, len(headers), mx=18)


def sheet_hist_balances(wb, cu, snap, df, grades, config, hist=None):
    """Historical Loan Balances by pool with monthly balance data."""
    ws = wb.create_sheet("Historical Balances")
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Historical Loan Balances by Pool"
    ws['A3'] = f"For Period Ending {snap}"

    monthly = hist.get('monthly_balances', pd.DataFrame()) if hist else pd.DataFrame()

    if not monthly.empty:
        # Use monthly balance data - show quarterly snapshots per pool
        # Get quarter-end dates (month-end for Mar, Jun, Sep, Dec)
        monthly['quarter'] = monthly['date'].dt.to_period('Q')
        # Get last date per quarter per pool
        qtr_data = monthly.groupby(['pool', 'quarter']).last().reset_index()
        quarters = sorted(qtr_data['quarter'].unique())
        # Limit to last 20 quarters to keep sheet manageable
        if len(quarters) > 20:
            quarters = quarters[-20:]
        qtr_strs = [str(q) for q in quarters]

        pool_names = sorted(qtr_data['pool'].unique())
        r = 5
        headers = ["Pool"] + qtr_strs
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pool_names:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            for qi, q in enumerate(quarters):
                val = qtr_data[(qtr_data['pool'] == pool) & (qtr_data['quarter'] == q)]
                bal = val['balance'].values[0] if len(val) > 0 else 0
                ws.cell(row=r, column=2 + qi, value=bal).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        for qi, q in enumerate(quarters):
            total = qtr_data[qtr_data['quarter'] == q]['balance'].sum()
            ws.cell(row=r, column=2 + qi, value=total).number_format = MONEY
        style_rows(ws, start, r, ncol, mcols=set(range(2, ncol + 1)))
        auto_w(ws, ncol)
    else:
        # Fallback: just show current data by grade
        pools = sorted(df['loan_pool'].unique())
        r = 6
        for pool in pools:
            ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=12)
            r += 1
            ws.cell(row=r, column=1, value="Current Grade")
            ws.cell(row=r, column=2, value=snap)
            hdr_row(ws, r, 2)
            pdf = df[df['loan_pool'] == pool]
            for g in gl:
                r += 1
                ws.cell(row=r, column=1, value=g)
                ws.cell(row=r, column=2, value=pdf[pdf['current_grade'] == g]['current_balance'].sum()).number_format = MONEY
            r += 1
            ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
            ws.cell(row=r, column=2, value=pdf['current_balance'].sum()).number_format = MONEY
            r += 2
        auto_w(ws, 2)


def sheet_loss_factor_hist(wb, cu, snap, df, grades, config, hist=None):
    """Loss Factor Historical Detail with charge-off/recovery data."""
    ws = wb.create_sheet("Loss Factor Historical")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Loss Factor Historical Detail"
    ws['A3'] = f"For Period Ending {snap}"

    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    avg_bals = hist.get('avg_balances', {}) if hist else {}
    years = hist.get('years', []) if hist else []

    pools = sorted(set(config.get('pool_map', {}).values()))

    if years:
        year_strs = [str(y) for y in years]
        r = 5
        # Net charge-off rates per pool per year
        headers = ["Pool"] + year_strs + ["Average Life\nLoss Rate"]
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            rates = []
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                avg = avg_bals.get(y, {}).get(pool, 0)
                rate = net / avg if avg > 0 else 0
                ws.cell(row=r, column=2 + yi, value=rate).number_format = PCT4
                rates.append(rate)
            avg_rate = sum(rates) / len(rates) if rates else 0
            ws.cell(row=r, column=ncol, value=avg_rate).number_format = PCT4

        style_rows(ws, start, r, ncol, pcols=set(range(2, ncol + 1)))

        # Average balances section
        r += 3
        ws.cell(row=r, column=1, value="Average Balances by Pool").font = SUB_FONT
        r += 1
        headers2 = ["Pool"] + year_strs
        ncol2 = len(headers2)
        for hi, h in enumerate(headers2):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol2)
        start2 = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            for yi, y in enumerate(years):
                avg = avg_bals.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=avg).number_format = MONEY
        style_rows(ws, start2, r, ncol2, mcols=set(range(2, ncol2 + 1)))
        auto_w(ws, max(ncol, ncol2))
    else:
        ws['A5'] = "No historical data available."
        ws['A5'].font = Font(italic=True, color='888888')


def sheet_chargeoff_hist(wb, cu, snap, config, hist=None):
    """Charge off / Recoveries Historical Detail."""
    ws = wb.create_sheet("Chargeoff Historical")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Charge off and Recoveries Historical Detail"
    ws['A3'] = f"For Period Ending {snap}"

    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    years = hist.get('years', []) if hist else []
    pools = sorted(set(config.get('pool_map', {}).values()))

    if years:
        year_strs = [str(y) for y in years]

        # Charge-offs by pool by year
        r = 5
        ws.cell(row=r, column=1, value="Charge offs by Year").font = SUB_FONT
        r += 1
        headers = ["Pool"] + year_strs + ["Total"]
        ncol = len(headers)
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_tot = 0
            for yi, y in enumerate(years):
                val = co_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=val).number_format = MONEY
                row_tot += val
            ws.cell(row=r, column=ncol, value=row_tot).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        grand = 0
        for yi, y in enumerate(years):
            yt = sum(co_data.get(y, {}).get(p, 0) for p in pools)
            ws.cell(row=r, column=2 + yi, value=yt).number_format = MONEY
            grand += yt
        ws.cell(row=r, column=ncol, value=grand).number_format = MONEY
        style_rows(ws, start, r, ncol, mcols=set(range(2, ncol + 1)))

        # Recoveries by pool by year
        r += 3
        ws.cell(row=r, column=1, value="Recoveries by Year").font = SUB_FONT
        r += 1
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, ncol)
        start2 = r + 1
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_tot = 0
            for yi, y in enumerate(years):
                val = rc_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=val).number_format = MONEY
                row_tot += val
            ws.cell(row=r, column=ncol, value=row_tot).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        grand = 0
        for yi, y in enumerate(years):
            yt = sum(rc_data.get(y, {}).get(p, 0) for p in pools)
            ws.cell(row=r, column=2 + yi, value=yt).number_format = MONEY
            grand += yt
        ws.cell(row=r, column=ncol, value=grand).number_format = MONEY
        style_rows(ws, start2, r, ncol, mcols=set(range(2, ncol + 1)))

        # Net Charge offs
        r += 2
        headers3 = ["Net Charge offs"] + year_strs + ["Net Charge offs"]
        for hi, h in enumerate(headers3):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers3))
        start3 = r + 1
        grand_net = {y: 0 for y in years}
        for pool in pools:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            row_total = 0
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                ws.cell(row=r, column=2 + yi, value=net).number_format = MONEY
                grand_net[y] += net
                row_total += net
            ws.cell(row=r, column=len(headers3), value=row_total).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        net_total = 0
        for yi, y in enumerate(years):
            ws.cell(row=r, column=2 + yi, value=grand_net[y]).number_format = MONEY
            net_total += grand_net[y]
        ws.cell(row=r, column=len(headers3), value=net_total).number_format = MONEY
        style_rows(ws, start3, r, len(headers3), mcols=set(range(2, len(headers3) + 1)))

        # Life of Loan Loss Rate
        r += 2
        headers4 = ["Life Loss Rate"] + year_strs + ["Average"]
        for hi, h in enumerate(headers4):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers4))
        start4 = r + 1
        for pool in pool_names:
            r += 1
            ws.cell(row=r, column=1, value=pool)
            rates = []
            for yi, y in enumerate(years):
                net = co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
                avg = avg_bals.get(y, {}).get(pool, 0)
                rate = net / avg if avg > 0 else 0
                ws.cell(row=r, column=2 + yi, value=rate).number_format = PCT4
                rates.append(rate)
            avg_rate = sum(rates) / len(rates) if rates else 0
            ws.cell(row=r, column=len(headers4), value=avg_rate).number_format = PCT4
        style_rows(ws, start4, r, len(headers4), pcols=set(range(2, len(headers4) + 1)))

        auto_w(ws, len(headers))


def sheet_delinquency(wb, cu, snap, config, hist=None):
    """Delinquency Calculation with historical data."""
    ws = wb.create_sheet("Delinquency Calculation")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Delinquency Calculation"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(set(config.get('pool_map', {}).values()))
    dq_pct = hist.get('dq_pct', {}) if hist else {}
    years = sorted(dq_pct.keys()) if dq_pct else list(range(2019, int(snap[:4]) + 1))
    year_strs = [str(y) for y in years]

    r = 5
    headers = ["DQ %"] + year_strs + ["Average", "Variance from Avg"]
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    start = r + 1
    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool)
        rates = []
        for yi, y in enumerate(years):
            val = dq_pct.get(y, {}).get(pool, 0)
            ws.cell(row=r, column=2 + yi, value=val).number_format = PCT
            rates.append(val)
        avg = sum(rates) / len(rates) if rates else 0
        ws.cell(row=r, column=len(headers) - 1, value=avg).number_format = PCT
        # Variance = most recent - average
        current = rates[-1] if rates else 0
        ws.cell(row=r, column=len(headers), value=current - avg).number_format = PCT
    style_rows(ws, start, r, len(headers), pcols=set(range(2, len(headers) + 1)))
    auto_w(ws, len(headers))


def sheet_balance_adj(wb, cu, snap, df, grades, config):
    """FAS 114 / Balance Adjustment sheet."""
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    ws = wb.create_sheet("Balance Adjustment")
    ws['A1'] = cu
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "Balance Adjustment"
    ws['A3'] = f"For Period Ending {snap}"

    pools = sorted(df['loan_pool'].unique())
    r = 5
    headers = ["Current Grade", "Loan Report Balance", "Bal Adjustment", "Balance Sheet Total"]

    for pool in pools:
        pdf = df[df['loan_pool'] == pool]
        ws.cell(row=r, column=1, value=pool).font = Font(bold=True, size=12)
        r += 1
        for hi, h in enumerate(headers):
            ws.cell(row=r, column=1 + hi, value=h)
        hdr_row(ws, r, len(headers))
        start = r + 1
        pool_total = 0
        for g in gl:
            r += 1
            bal = pdf[pdf['current_grade'] == g]['current_balance'].sum()
            pool_total += bal
            ws.cell(row=r, column=1, value=g)
            ws.cell(row=r, column=2, value=bal).number_format = MONEY
            ws.cell(row=r, column=3, value=0).number_format = MONEY
            ws.cell(row=r, column=4, value=bal).number_format = MONEY
        r += 1
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=r, column=2, value=pool_total).number_format = MONEY
        ws.cell(row=r, column=3, value=0).number_format = MONEY
        ws.cell(row=r, column=4, value=pool_total).number_format = MONEY
        style_rows(ws, start, r, len(headers), mcols={2,3,4})
        r += 2

    # Grand total
    ws.cell(row=r, column=1, value="Grand Total").font = Font(bold=True, size=12)
    total = df['current_balance'].sum()
    ws.cell(row=r, column=2, value=total).number_format = MONEY
    ws.cell(row=r, column=3, value=0).number_format = MONEY
    ws.cell(row=r, column=4, value=total).number_format = MONEY
    auto_w(ws, len(headers))


def sheet_env_ranges(wb):
    """Environmental Factor Ranges reference table."""
    ws = wb.create_sheet("Env Factor Ranges")
    ws['A1'] = "Environmental Factor Ranges"
    ws['A1'].font = Font(bold=True, size=14)

    # Net Credit Change
    r = 3
    ws.cell(row=r, column=1, value="Net Credit Change").font = SUB_FONT
    ws.cell(row=r, column=1+0, value="Range"); ws.cell(row=r, column=2, value="Score")
    ws.cell(row=r, column=4, value="Delinquency").font = SUB_FONT
    ws.cell(row=r, column=4, value="Range"); ws.cell(row=r, column=5, value="Score")
    ws.cell(row=r, column=7, value="Economic Stress Score").font = SUB_FONT
    ws.cell(row=r, column=7, value="Range"); ws.cell(row=r, column=8, value="Score")
    hdr_row(ws, r, 8)

    ncc_rows = [
        ("<-18.00%", "7.00%"), ("-17.99% to -16.00%", "6.00%"),
        ("-15.99% to -14.00%", "5.00%"), ("-13.99% to -11.00%", "4.00%"),
        ("-10.99% to -8.00%", "3.00%"), ("-7.99% to -6.00%", "2.00%"),
        ("-5.99% to -4.00%", "1.00%"), ("-3.99% to 3.99%", "0.00%"),
        ("4.00% to 5.99%", "-1.00%"), ("6.00% to 7.99%", "-2.00%"),
        ("8.00% to 8.99%", "-3.00%"), ("9.00% to 10.99%", "-4.00%"),
        ("11.00% to 12.99%", "-5.00%"), ("13.00% to 14.99%", "-6.00%"),
        (">15.00%", "-7.00%"),
    ]
    dq_rows = [
        (">5.00%", "20.00%"), ("4.00% to 4.99%", "17.00%"),
        ("3.00% to 3.99%", "12.00%"), ("2.50% to 2.99%", "8.00%"),
        ("2.00% to 2.49%", "4.00%"), ("1.50% to 1.99%", "2.50%"),
        ("1.00% to 1.49%", "1.50%"), (".50% to .99%", "0.75%"),
        ("-.49% to .49%", "0.00%"), ("-.99% to -.50%", "-0.75%"),
        ("-1.49% to -1.00%", "-1.50%"), ("-1.99% to -1.50%", "-2.50%"),
        ("-2.49% to -2.00%", "-4.00%"), ("-2.99% to -2.50%", "-8.00%"),
        ("-3.99% to -3.00%", "-12.00%"), ("-4.99% to -4.00%", "-17.00%"),
        ("<-5.00%", "-20.00%"),
    ]
    es_rows = [
        (">25.00%", "10.00%"), ("24.00% to 24.99%", "8.00%"),
        ("22.00% to 23.99%", "7.00%"), ("20.00% to 21.99%", "6.00%"),
        ("18.00% to 19.99%", "5.00%"), ("16.00% to 17.99%", "4.00%"),
        ("14.00% to 15.99%", "3.50%"), ("12.00% to 13.99%", "3.00%"),
        ("10.00% to 11.99%", "2.00%"), ("8.00% to 9.99%", "1.00%"),
        ("6.00% to 7.99%", "0.00%"), ("4.00% to 5.99%", "0.00%"),
        ("2.00% to 3.99%", "-1.00%"), (".00% to 1.99%", "-2.00%"),
    ]

    for i, (rng, sc) in enumerate(ncc_rows):
        ws.cell(row=r + 1 + i, column=1, value=rng)
        ws.cell(row=r + 1 + i, column=2, value=sc)
    for i, (rng, sc) in enumerate(dq_rows):
        ws.cell(row=r + 1 + i, column=4, value=rng)
        ws.cell(row=r + 1 + i, column=5, value=sc)
    for i, (rng, sc) in enumerate(es_rows):
        ws.cell(row=r + 1 + i, column=7, value=rng)
        ws.cell(row=r + 1 + i, column=8, value=sc)
    auto_w(ws, 8)


def sheet_grade_config(wb, grades, config):
    """Grade ranges & loan code reference."""
    ws = wb.create_sheet("Grade Ranges & Loan Codes")
    ws['A1'] = "Credit Grade Configuration"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["Grade", "Score Range", "Reserve Rate"]
    r = 3
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))
    for g in grades:
        r += 1
        ws.cell(row=r, column=1, value=g['label'])
        ws.cell(row=r, column=2, value=f"{g['min_score']}-{g['max_score']}")
        ws.cell(row=r, column=3, value=g['reserve_rate']).number_format = PCT

    r += 3
    ws.cell(row=r, column=1, value="Loan Type Codes").font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1, value="Code"); ws.cell(row=r, column=2, value="Loan Pool")
    hdr_row(ws, r, 2)
    for code, pool in sorted(config.get('pool_map', {}).items(), key=lambda x: x[1]):
        r += 1
        ws.cell(row=r, column=1, value=str(code))
        ws.cell(row=r, column=2, value=pool)
    auto_w(ws, 3)


def sheet_all_loans(wb, cu, snap, df, grades, config):
    """All Loans detail listing."""
    ws = wb.create_sheet("All Loans")
    no_score = config.get('no_score_label', 'Not Reported')
    ws['A1'] = "Credit Grade Analysis - All Loans"
    ws['A1'].font = Font(bold=True, size=14)
    ws['F1'] = snap

    headers = ["Member #", "Loan Pool", "Current Balance",
               "Original Score", "Original Grade",
               "Current Score", "Current Grade",
               "Migration Status", "Reserve Rate", "Expected Loss"]
    r = 2
    for hi, h in enumerate(headers):
        ws.cell(row=r, column=1 + hi, value=h)
    hdr_row(ws, r, len(headers))

    start = r + 1
    for _, loan in df.iterrows():
        r += 1
        ws.cell(row=r, column=1, value=str(loan.get('member_number', '')))
        ws.cell(row=r, column=2, value=loan.get('loan_pool', ''))
        ws.cell(row=r, column=3, value=loan.get('current_balance', 0))
        ws.cell(row=r, column=4, value=loan.get('original_fico_score', 0))
        ws.cell(row=r, column=5, value=loan.get('original_grade', no_score))
        ws.cell(row=r, column=6, value=loan.get('current_fico_score', 0))
        ws.cell(row=r, column=7, value=loan.get('current_grade', no_score))
        ws.cell(row=r, column=8, value=loan.get('migration_status', 'Unchanged'))
        ws.cell(row=r, column=9, value=loan.get('reserve_rate', 0))
        ws.cell(row=r, column=10, value=loan.get('expected_loss_amount', 0))


# ── Main Entry Point ──────────────────────────────────────────────

def _coerce_money(value):
    """Best-effort float from an extract cell (handles ``$``/commas/blanks)."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        s = str(value).strip().replace('$', '').replace(',', '')
        if s in ('', '-'):
            return 0.0
        neg = s.startswith('(') and s.endswith(')')
        s = s.strip('()')
        try:
            v = float(s)
        except ValueError:
            return 0.0
        return -v if neg else v


def _neg_share_period_from_text(text):
    """Return ``(year, month)`` from a spelled-out or numeric month token in
    ``text`` (e.g. ``"Jun 2026"``, ``"July 2025"``, ``"6-30-26"``), else None."""
    stem = str(text)
    m = re.search(r'(?<![A-Za-z])([A-Za-z]{3,9})[\s\-_]+(20\d{2})(?!\d)', stem)
    if m:
        mo = _SUPP_MONTH_NAMES.get(m.group(1).strip().lower())
        if mo:
            return int(m.group(2)), mo
    m = re.search(r'(?<!\d)(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})(?!\d)', stem)
    if m:
        mo = int(m.group(1))
        yy = int(m.group(3))
        if yy < 100:
            yy += 2000
        if 1 <= mo <= 12 and 2000 <= yy <= 2099:
            return yy, mo
    return None


def _neg_share_period_from_path(path):
    """Period ``(year, month)`` for a negative-share file — prefer the file's
    own name, then walk up the folder names (e.g. ``.../Jun 2026/...``)."""
    p = _neg_share_period_from_text(os.path.basename(path))
    if p:
        return p
    for part in reversed(os.path.dirname(path).split(os.sep)):
        p = _neg_share_period_from_text(part)
        if p:
            return p
    return None


def _resolve_negative_share_oac(o, config, snapshot_date):
    """Resolve one ``source: negative_share`` OAC entry in-place.

    Computes a life-of-loan loss rate for the credit union's negative shares
    using the same WARM logic as the loan pools: trailing
    ``life_of_loan_months`` (default 12) of net charge-offs (charge-offs minus
    recoveries) divided by the current negative-share balance. The balance is
    the summed exposure column of the current month's negative-share extract;
    the charge-off / recovery history is gathered from the client's monthly
    summary files plus any multi-month quarterly charge-off file. Sets
    ``balance``, ``percentage`` and ``amount`` for the current snapshot so the
    Vizo and TCT builders render an up-to-date line every quarter.
    """
    try:
        snap = pd.to_datetime(snapshot_date)
        snap_year, snap_month = int(snap.year), int(snap.month)
    except Exception:
        print("  OAC negative_share: bad snapshot_date; skipped.")
        return
    try:
        life_months = int(o.get('life_of_loan_months') or 12)
    except (TypeError, ValueError):
        life_months = 12
    folder = o.get('source_folder') or config.get('data_directory') or ''
    folder = os.path.expanduser(os.path.expandvars(str(folder)))
    if not folder or not os.path.isdir(folder):
        print(f"  OAC '{o.get('title')}': source folder not found "
              f"({folder!r}); balance left as configured.")
        return
    bal_pat = re.compile(o.get('balance_pattern') or r'(?i)Negative Share File')
    sum_pat = re.compile(o.get('co_summary_pattern')
                         or r'(?i)Negative Shares Charge Off and Recovery')
    qtr_pat = re.compile(o.get('co_quarterly_pattern')
                         or r'(?i)Share COs?\s*-\s*Recoveries')
    bal_col_pref = o.get('balance_column') or 'Current Balance'

    bal_files, sum_files, qtr_files = [], [], []
    for dp, _dn, fn in os.walk(folder):
        for f in fn:
            if not f.lower().endswith(('.xlsx', '.xls', '.csv')):
                continue
            full = os.path.join(dp, f)
            if bal_pat.search(f):
                bal_files.append(full)
            elif sum_pat.search(f):
                sum_files.append(full)
            elif qtr_pat.search(f):
                qtr_files.append(full)

    # Current-month negative-share balance (member exposure to the CU).
    target = (snap_year, snap_month)
    bal_total = 0.0
    bal_pick = next(
        (f for f in bal_files if _neg_share_period_from_path(f) == target), None)
    if bal_pick:
        try:
            bdf = pd.read_excel(bal_pick, header=0)
            col = bal_col_pref if bal_col_pref in bdf.columns else (
                'Current Balance' if 'Current Balance' in bdf.columns
                else ('Balance' if 'Balance' in bdf.columns else None))
            if col is not None:
                bal_total = float(
                    pd.to_numeric(bdf[col], errors='coerce').abs().sum())
        except Exception as e:  # noqa: BLE001
            print(f"  OAC '{o.get('title')}': could not read balance file ({e}).")
    else:
        print(f"  OAC '{o.get('title')}': no negative-share balance file for "
              f"{snap_year}-{snap_month:02d}; balance left as configured.")

    # Monthly net charge-off history keyed by (year, month).
    co_by_ym = {}
    for f in sum_files:
        per = _neg_share_period_from_path(f)
        if not per:
            continue
        try:
            d = pd.read_excel(f)
        except Exception:  # noqa: BLE001
            continue
        cols = {str(c).strip().lower(): c for c in d.columns}
        co_c = next((cols[k] for k in cols if 'charge' in k and 'amount' in k), None) \
            or next((cols[k] for k in cols if 'charge' in k and 'date' not in k), None)
        rc_c = next((cols[k] for k in cols if 'recover' in k and 'amount' in k), None) \
            or next((cols[k] for k in cols if 'recover' in k and 'date' not in k), None)
        if not co_c:
            continue
        co = float(pd.to_numeric(d[co_c], errors='coerce').sum())
        rc = float(pd.to_numeric(d[rc_c], errors='coerce').sum()) if rc_c else 0.0
        co_by_ym[per] = (co, rc)
    # Multi-month quarterly files take precedence (freshest delivery, and
    # they carry month-end dates so their period is unambiguous).
    for f in qtr_files:
        try:
            raw = pd.read_excel(f, header=None)
        except Exception:  # noqa: BLE001
            continue
        hdr = None
        for i in range(min(15, len(raw))):
            vals = [str(x).strip().lower() for x in raw.iloc[i].tolist()]
            if (any(v == 'co' or 'charge' in v for v in vals)
                    and any('rec' in v for v in vals)):
                hdr = i
                break
        if hdr is None:
            continue
        try:
            d = pd.read_excel(f, header=hdr)
        except Exception:  # noqa: BLE001
            continue
        date_c = d.columns[0]
        cols = {str(c).strip().lower(): c for c in d.columns}
        co_c = next((cols[k] for k in cols if 'charge' in k and 'amount' in k), None) \
            or next((cols[k] for k in cols if k == 'co' or 'charge' in k and 'date' not in k), None)
        rc_c = next((cols[k] for k in cols if 'rec' in k and 'amount' in k), None) \
            or next((cols[k] for k in cols if 'rec' in k and 'date' not in k), None)
        for _, row in d.iterrows():
            dt = pd.to_datetime(row[date_c], errors='coerce')
            if pd.isna(dt):
                continue
            co = _coerce_money(row[co_c]) if co_c else 0.0
            rc = _coerce_money(row[rc_c]) if rc_c else 0.0
            co_by_ym[(int(dt.year), int(dt.month))] = (co, rc)

    end = snap_year * 12 + snap_month
    start = end - life_months + 1
    window = {ym: v for ym, v in co_by_ym.items()
              if start <= (ym[0] * 12 + ym[1]) <= end}
    net_total = sum((co - rc) for co, rc in window.values())

    calc_pct = (net_total / bal_total * 100.0) if bal_total > 0 else 0.0
    # Manual override from the Run Reports "Other Allowance Considerations"
    # page. When set (non-blank), it replaces the calculated life-of-loan
    # rate; the balance is still pulled from the current extract.
    override = o.get('override_percentage')
    use_override = override is not None and str(override).strip() != ''
    try:
        pct = float(override) if use_override else calc_pct
    except (TypeError, ValueError):
        pct = calc_pct
        use_override = False
    o['balance'] = round(bal_total, 2)
    o['percentage'] = round(pct, 6)
    o['amount'] = round(bal_total * pct / 100.0, 2)
    config.setdefault('_oac_calc', {})[o.get('title') or 'Negative Share Provision'] = {
        'kind': 'negative_share',
        'balance': round(bal_total, 2),
        'calculated_percentage': round(calc_pct, 6),
        'override_percentage': (round(float(override), 6) if use_override else None),
        'amount': o['amount'],
    }
    _tag = f" -> OVERRIDE {pct:.4f}%" if use_override else ""
    print(f"  OAC '{o.get('title')}': neg-share balance ${bal_total:,.2f}, "
          f"trailing {len(window)}/{life_months}mo net CO ${net_total:,.2f} "
          f"-> life-of-loan loss rate {calc_pct:.4f}%{_tag} "
          f"-> allowance ${o['amount']:,.2f}")


def _unfunded_undrawn_by_pool(config, snapshot_date, codes):
    """Return ``{pool_name: undrawn_dollars}`` for the unfunded-commitment
    codes.

    Reads the current-snapshot loan extract only (files whose filename/folder
    period matches ``snapshot_date``; undated files are allowed) and, for
    every loan whose raw ``loan_pool_code`` is one of ``codes``, adds the
    undrawn credit (``total_available_credit`` - ``current_balance``, floored
    at 0) to the pool that code maps to via ``config['pool_map']``. Codes are
    compared with leading zeros stripped so ``'0080'`` and ``'80'`` match.
    Only the snapshot file is used so an earlier month's extract sitting in
    the archive can't contaminate the drawn balances.
    """
    def _norm(c):
        s = str(c).strip()
        return s.lstrip('0') or '0'

    want = {_norm(c) for c in (codes or []) if str(c).strip()}
    if not want:
        return {}
    try:
        # generate_impdet_report builds a module-level engine from
        # os.getenv('DATABASE_URL'); make sure it's populated for standalone
        # report runs that only resolve the URL via cecl_credentials.
        if not os.getenv('DATABASE_URL'):
            os.environ['DATABASE_URL'] = get_database_url()
        from generate_impdet_report import _resolve_extract_path
    except Exception as e:  # pragma: no cover - defensive
        print(f"  OAC unfunded: extract resolver unavailable ({e}).")
        return {}

    try:
        _snap = pd.to_datetime(snapshot_date)
        target = (int(_snap.year), int(_snap.month))
    except Exception:
        target = None

    # Build the same search directories _load_extract_enrichment uses.
    data_dir = config.get('data_directory', '')
    if data_dir and not os.path.isabs(data_dir):
        data_dir = os.path.join(BASE, data_dir)
    search_dirs = []
    if data_dir and os.path.isdir(data_dir):
        search_dirs.append(data_dir)
    archive_cfg = config.get('archive_directory')
    if archive_cfg:
        archive_dir = archive_cfg if os.path.isabs(archive_cfg) \
            else os.path.join(BASE, archive_cfg)
    else:
        client_short = os.path.basename(os.path.normpath(data_dir)) if data_dir else ''
        archive_dir = os.path.join(BASE, 'Archive', client_short) if client_short else ''
    if archive_dir and os.path.isdir(archive_dir) and archive_dir not in search_dirs:
        search_dirs.append(archive_dir)
    loan_folder = config.get('loan_file_folder')
    if loan_folder:
        loan_dir = loan_folder if os.path.isabs(loan_folder) \
            else os.path.join(BASE, loan_folder)
        if os.path.isdir(loan_dir) and loan_dir not in search_dirs:
            search_dirs.append(loan_dir)

    extracts = list(config.get('loan_data_extracts') or [])
    if not extracts:
        extracts = [{
            'file_pattern': config.get('file_pattern'),
            'column_mappings': config.get('column_mappings') or {},
            'has_header': config.get('has_header', True),
            'header_row': config.get('header_row'),
        }]

    pool_map = config.get('pool_map') or {}
    norm_pool_map = {}
    for k, v in pool_map.items():
        norm_pool_map.setdefault(_norm(k), v)
    default_pool = config.get('default_pool', 'Ignore')

    def _norm_hdr(x):
        return re.sub(r'\s+', ' ', str(x)).strip()

    # Resolve each extract to a file, tagging whether its period is
    # confirmed as the snapshot month, unknown (undated path), or wrong.
    confirmed, unknown = [], []
    seen_paths = set()
    for ex in extracts:
        col_map = dict(ex.get('column_mappings') or {})
        if col_map.get('loan_pool_code') is None \
                or col_map.get('current_balance') is None \
                or col_map.get('total_available_credit') is None:
            continue
        path = _resolve_extract_path(ex.get('file_pattern'), search_dirs,
                                     snapshot_date, config)
        if not path or path in seen_paths:
            continue
        seen_paths.add(path)
        per = _neg_share_period_from_path(path) if target is not None else None
        entry = (ex, col_map, path)
        if target is not None and per == target:
            confirmed.append(entry)
        elif per is None or target is None:
            unknown.append(entry)
        # period != target -> drop (an earlier month sitting in the archive)
    # Prefer files confirmed to be the snapshot month; only fall back to
    # undated files when nothing confirms the snapshot (so a stale month in
    # the archive can't contaminate the current-quarter undrawn totals).
    chosen = confirmed if confirmed else unknown

    by_pool = {}
    counts = {}
    for ex, col_map, path in chosen:
        code_key = col_map.get('loan_pool_code')
        bal_key = col_map.get('current_balance')
        lim_key = col_map.get('total_available_credit')
        has_header = ex.get('has_header', True)
        try:
            hr = int(ex.get('header_row') or 0)
        except (TypeError, ValueError):
            hr = 0
        pd_header = (hr - 1) if hr > 1 else 0
        ext_lc = os.path.splitext(path)[1].lower()
        try:
            if ext_lc == '.csv':
                fdf = _read_csv_any(path, header=(pd_header if has_header else None))
            else:
                fdf = pd.read_excel(path, header=(pd_header if has_header else None))
        except Exception as e:  # noqa: BLE001
            print(f"  OAC unfunded: could not read extract '{path}' ({e}).")
            continue

        if has_header:
            norm_cols = {_norm_hdr(c): c for c in fdf.columns}

            def _series(key):
                actual = norm_cols.get(_norm_hdr(key))
                if actual is None and key in fdf.columns:
                    actual = key
                return fdf[actual] if actual is not None else None
        else:
            def _series(key):
                try:
                    return fdf.iloc[:, int(key)]
                except (KeyError, IndexError, ValueError):
                    return None

        codes_ser = _series(code_key)
        bal_ser = _series(bal_key)
        lim_ser = _series(lim_key)
        if codes_ser is None or bal_ser is None or lim_ser is None:
            print(f"  OAC unfunded: extract '{os.path.basename(path)}' missing "
                  f"code/balance/limit column; skipped.")
            continue
        print(f"  OAC unfunded: reading {os.path.basename(path)}")
        for i in range(len(fdf)):
            lt = _norm(codes_ser.iloc[i])
            if lt not in want:
                continue
            avail = _coerce_money(lim_ser.iloc[i]) - _coerce_money(bal_ser.iloc[i])
            if avail <= 0:
                continue
            pool = norm_pool_map.get(lt, default_pool)
            by_pool[pool] = by_pool.get(pool, 0.0) + avail
            counts[pool] = counts.get(pool, 0) + 1

    for pool in sorted(by_pool):
        print(f"  OAC unfunded: pool '{pool}' undrawn ${by_pool[pool]:,.2f} "
              f"from {counts.get(pool, 0)} loan(s)")
    return by_pool


def _expand_unfunded_commitment_oac(config, pool_eff_rate, snapshot_date):
    """Expand ``source: unfunded_commitment`` OAC templates into one resolved
    row per pool.

    Each template lists the ``loan_type_codes`` the CU flagged as *not*
    unconditionally cancelable. This computes the undrawn credit for those
    codes grouped by pool, then applies that pool's effective ACL loss rate
    (``pool_eff_rate[pool]`` — the same rate the homogeneous-pool ACL
    calculation applies to the pool) to produce the allowance. Codes that map
    to different pools each get their own ``Unfunded Commitments - <Pool>``
    row. Called from the ACL Env sheet builder (which owns the pool rates);
    idempotent because the template entry is removed once expanded.
    """
    oac = config.get('other_allowance_considerations')
    if not oac:
        return
    if not any(str((o or {}).get('source') or '').strip() == 'unfunded_commitment'
               for o in oac):
        return
    pool_eff_rate = pool_eff_rate or {}
    new_list = []
    for o in oac:
        if str((o or {}).get('source') or '').strip() != 'unfunded_commitment':
            new_list.append(o)
            continue
        codes = o.get('loan_type_codes')
        if codes is None:
            codes = o.get('loan_type_code')
        if isinstance(codes, (str, int)):
            codes = [codes]
        base_title = (str(o.get('title') or 'Unfunded Commitments').strip()
                      or 'Unfunded Commitments')
        # Per-pool manual overrides from the Run Reports "Other Allowance
        # Considerations" page, keyed by pool name (percentage points).
        overrides = o.get('override_percentage_by_pool') or {}
        by_pool = _unfunded_undrawn_by_pool(config, snapshot_date, codes or [])
        for pool in sorted(by_pool):
            bal = round(by_pool[pool], 2)
            calc_rate = float(pool_eff_rate.get(pool, 0) or 0)  # decimal
            ov = overrides.get(pool)
            use_override = ov is not None and str(ov).strip() != ''
            try:
                rate = (float(ov) / 100.0) if use_override else calc_rate
            except (TypeError, ValueError):
                rate = calc_rate
                use_override = False
            amt = round(bal * rate, 2)
            new_list.append({
                'title': f"{base_title} - {pool}",
                'balance': bal,
                'percentage': round(rate * 100.0, 6),
                'amount': amt,
            })
            config.setdefault('_oac_calc', {})[f"{base_title} - {pool}"] = {
                'kind': 'unfunded_commitment',
                'pool': pool,
                'base_title': base_title,
                'balance': bal,
                'calculated_percentage': round(calc_rate * 100.0, 6),
                'override_percentage': (round(float(ov), 6) if use_override else None),
                'amount': amt,
            }
            _tag = f" -> OVERRIDE {rate * 100:.4f}%" if use_override else ""
            print(f"  OAC '{base_title} - {pool}': undrawn ${bal:,.2f} @ pool "
                  f"ACL rate {calc_rate * 100:.4f}%{_tag} -> allowance ${amt:,.2f}")
    config['other_allowance_considerations'] = new_list


def _persist_oac_calc(client, calc, snapshot_date):
    """Write the latest calculated OAC values back to the client YAML under
    ``oac_last_calculated`` so the Run Reports override page can display them.

    Re-reads the YAML from disk (rather than dumping the in-memory config,
    which has had its OAC list expanded and carries transient keys) and only
    touches the ``oac_last_calculated`` block.
    """
    try:
        path = os.path.join(CFG_DIR, f'{client}.yaml')
        with open(path, 'r', encoding='utf-8') as f:
            raw = yaml.safe_load(f) or {}
        raw['oac_last_calculated'] = {
            'asof': str(snapshot_date),
            'rows': calc,
        }
        with open(path, 'w', encoding='utf-8') as f:
            yaml.safe_dump(raw, f, sort_keys=False, allow_unicode=True)
        print(f"  OAC: cached {len(calc)} calculated value(s) to {client}.yaml")
    except Exception as e:  # noqa: BLE001
        print(f"  OAC: could not persist calculated snapshot ({e}).")


def _apply_chargeoff_exclusions(hist, config):
    """Remove specific charge-off / recovery records from the assembled
    history per ``config['chargeoff_exclusions']``.

    Each exclusion identifies a record the CU asked to drop from the analysis.
    Matching is by year+month (from ``date``), pool (given explicitly or
    derived from ``code`` via ``pool_map``) and ``amount``. The amount is
    subtracted from the annual ``chargeoffs`` total, the ``co_monthly`` cell,
    and — when present — ``impaired.warm_net_co`` so the life-of-loan loss
    rate, the Display CO-Recov-DQ tab, and the charge-off history sheet all
    reflect the removal. Recovery exclusions (``recovery_amount``) are handled
    symmetrically. This honors a removal request without editing source files
    or re-importing.
    """
    exclusions = config.get('chargeoff_exclusions') or []
    if not exclusions or not isinstance(hist, dict):
        return

    def _norm_code(c):
        s = str(c).strip()
        return s.lstrip('0') or '0'

    def _f(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    pm = config.get('pool_map') or {}
    npm = {}
    for k, v in pm.items():
        npm.setdefault(_norm_code(k), v)
    default_pool = config.get('default_pool', 'Ignore')
    imp = hist.get('impaired') if isinstance(hist.get('impaired'), dict) else {}
    warm_net_co = imp.get('warm_net_co') \
        if isinstance(imp.get('warm_net_co'), dict) else None

    for ex in exclusions:
        if not isinstance(ex, dict):
            continue
        yr = ex.get('year')
        mo = ex.get('month')
        if (yr is None or mo is None) and ex.get('date'):
            try:
                d = pd.to_datetime(ex.get('date'))
                yr, mo = int(d.year), int(d.month)
            except Exception:  # noqa: BLE001
                pass
        try:
            yr = int(yr) if yr is not None else None
            mo = int(mo) if mo is not None else None
        except (TypeError, ValueError):
            yr = mo = None
        pool = str(ex.get('pool') or '').strip()
        if not pool and ex.get('code') is not None:
            pool = npm.get(_norm_code(ex.get('code')), default_pool)
        co_amt = _f(ex.get('amount') if ex.get('amount') is not None
                    else ex.get('chargeoff_amount'))
        rc_amt = _f(ex.get('recovery_amount'))
        label = str(ex.get('reason') or ex.get('account') or '').strip()
        if yr is None or not pool:
            print(f"  CO exclusion skipped (need year + pool/code): {ex}")
            continue

        def _subtract(annual, monthly, amt, kind):
            if amt <= 0:
                return
            cur = (annual.get(yr, {}) or {}).get(pool, 0.0)
            if cur < amt - 0.01:
                print(f"  WARNING: {kind} exclusion ${amt:,.2f} exceeds available "
                      f"${cur:,.2f} for {pool} {yr}; subtracting available only.")
                amt = cur
            if amt <= 0:
                return
            if yr in annual and pool in annual[yr]:
                annual[yr][pool] = round(annual[yr][pool] - amt, 2)
            if mo is not None and (yr, mo) in monthly and pool in monthly[(yr, mo)]:
                monthly[(yr, mo)][pool] = round(monthly[(yr, mo)][pool] - amt, 2)

        if co_amt > 0:
            _subtract(hist.setdefault('chargeoffs', {}),
                      hist.setdefault('co_monthly', {}), co_amt, 'charge-off')
            if warm_net_co is not None and pool in warm_net_co:
                warm_net_co[pool] = round(warm_net_co[pool] - co_amt, 2)
        if rc_amt > 0:
            _subtract(hist.setdefault('recoveries', {}),
                      hist.setdefault('rc_monthly', {}), rc_amt, 'recovery')
            if warm_net_co is not None and pool in warm_net_co:
                warm_net_co[pool] = round(warm_net_co[pool] + rc_amt, 2)
        print(f"  CO exclusion applied: {pool} {yr}-{(mo or 0):02d} "
              f"-CO ${co_amt:,.2f}"
              + (f" -Rec ${rc_amt:,.2f}" if rc_amt else "")
              + (f"  ({label})" if label else ""))


def _resolve_dynamic_oac(config, snapshot_date):
    """Resolve any data-derived Other Allowance Considerations in-place.

    An OAC entry with ``source: unfunded_available`` has its ``balance``
    computed from the loan extracts as the sum of undrawn credit
    (``total_available_credit`` - ``current_balance``, floored at 0) for
    loans whose raw ``loan_pool_code`` matches ``loan_type_code``. An entry
    with ``source: negative_share`` gets a life-of-loan loss rate computed
    from the CU's negative-share balance extract and charge-off / recovery
    history (see ``_resolve_negative_share_oac``). The ``amount`` is refreshed
    as ``balance * percentage / 100``. Both the Vizo and TCT report builders
    read these resolved values, so computing them here keeps the number
    current every quarter without manual entry.
    """
    oac = config.get('other_allowance_considerations') or []
    for o in oac:
        if str((o or {}).get('source') or '').strip() == 'negative_share':
            _resolve_negative_share_oac(o, config, snapshot_date)

    dynamic = [o for o in oac
               if str((o or {}).get('source') or '').strip() == 'unfunded_available']
    if not dynamic:
        return
    try:
        from generate_impdet_report import _load_extract_enrichment
    except Exception as e:  # pragma: no cover - defensive
        print(f"  OAC: enrichment loader unavailable ({e}); balances left as configured.")
        return
    enrich = _load_extract_enrichment(config, BASE, snap=snapshot_date)
    if not enrich:
        print("  OAC: no extract data found; unfunded balances left as configured.")
        return
    for o in dynamic:
        raw_codes = o.get('loan_type_code')
        if isinstance(raw_codes, (list, tuple, set)):
            targets = {str(c).strip() for c in raw_codes if str(c).strip()}
        else:
            targets = {str(raw_codes).strip()} if str(raw_codes or '').strip() else set()
        if not targets:
            continue
        total_unfunded = 0.0
        n = 0
        for rec in enrich.values():
            if str(rec.get('loan_type', '')).strip() not in targets:
                continue
            avail = _coerce_money(rec.get('total_available_credit')) \
                - _coerce_money(rec.get('current_balance'))
            if avail > 0:
                total_unfunded += avail
                n += 1
        pct = 0.0
        try:
            pct = float(o.get('percentage') or 0)
        except (TypeError, ValueError):
            pct = 0.0
        o['balance'] = round(total_unfunded, 2)
        o['amount'] = round(total_unfunded * pct / 100.0, 2)
        print(f"  OAC '{o.get('title')}': code(s) {sorted(targets)} unfunded "
              f"${total_unfunded:,.2f} from {n} loan(s) @ {pct}% "
              f"-> allowance ${o['amount']:,.2f}")


def _derive_snapshot_dq_from_extracts(config, snapshot_date):
    """Derive the snapshot quarter's delinquency from the loan-data extract(s)
    and write it to ``loan_code_delinquency_history``.

    NCUA 5300 filings lag the reported quarter, so a freshly reported period
    has no 5300 DQ history and the Display CO-Recov-DQ tab's current column
    renders blank. This mirrors the wizard's ``dq_extract_parser`` derivation
    (sum of current_balance for loans whose days_delinquent >= threshold, per
    loan code) but runs automatically at report time for the snapshot month.
    Runs when the CU's extracts carry a ``days_delinquent`` mapping and the
    period has no manually-entered / 5300 DQ rows yet; opt out via
    ``config['delinquency']['derive_from_extracts']: false``. Header names are
    whitespace-normalised so a trailing-space column (e.g. ``'DAYS DQ '``)
    still matches the mapping.
    """
    dq_cfg = config.get('delinquency') or {}
    if not dq_cfg.get('derive_from_extracts', True):
        return 0
    cu = (config.get('credit_union') or '').strip()
    if not cu:
        return 0
    try:
        from cecl_ui.services import delinquency_hist_processor as _dqp
        from cecl_ui.services import dq_extract_parser as _dqe
        from cecl_ui.services import extract_hist_processor as _ehp
        if not os.getenv('DATABASE_URL'):
            try:
                from cecl_credentials import get_database_url as _gdu
                os.environ['DATABASE_URL'] = _gdu()
            except Exception:
                pass
        from generate_impdet_report import _resolve_extract_path
        from sqlalchemy import text as _sql_text
    except Exception as _e:
        print(f"    Snapshot DQ derivation skipped (imports): {_e}")
        return 0

    try:
        snap_ts = pd.Timestamp(snapshot_date)
        snap_iso = snap_ts.date().isoformat()
        target = (int(snap_ts.year), int(snap_ts.month))
    except Exception:
        return 0

    try:
        _eng = _dqp._engine_lazy()
        with _eng.begin() as _c:
            _srcs = [str(r[0] or '') for r in _c.execute(
                _sql_text('SELECT DISTINCT source FROM loan_code_delinquency_history WHERE cu = :cu AND as_of_date = :d'),
                {'cu': cu, 'd': snap_iso}).fetchall()]
        if _srcs and not all(s.startswith('loan_extract') for s in _srcs):
            return 0
    except Exception:
        try:
            if snap_iso in _dqp.existing_dates(cu):
                return 0
        except Exception:
            pass

    threshold = int(dq_cfg.get('dq_threshold') or _dqe.DEFAULT_DQ_THRESHOLD)

    data_dir = config.get('data_directory', '')
    if data_dir and not os.path.isabs(data_dir):
        data_dir = os.path.join(BASE, data_dir)
    search_dirs = []
    if data_dir and os.path.isdir(data_dir):
        search_dirs.append(data_dir)
    archive_cfg = config.get('archive_directory')
    if archive_cfg:
        archive_dir = archive_cfg if os.path.isabs(archive_cfg) \
            else os.path.join(BASE, archive_cfg)
    else:
        client_short = os.path.basename(os.path.normpath(data_dir)) if data_dir else ''
        archive_dir = os.path.join(BASE, 'Archive', client_short) if client_short else ''
    if archive_dir and os.path.isdir(archive_dir) and archive_dir not in search_dirs:
        search_dirs.append(archive_dir)
    if not search_dirs:
        return 0

    extracts = list(config.get('loan_data_extracts') or [])
    if not extracts:
        extracts = [{
            'file_pattern': config.get('file_pattern'),
            'column_mappings': config.get('column_mappings') or {},
            'has_header': config.get('has_header', True),
            'header_row': config.get('header_row'),
        }]

    def _norm_hdr(x):
        return re.sub(r'\s+', ' ', str(x)).strip()

    by_code = {}
    seen_paths = set()
    files_used = []
    for ex in extracts:
        col_map = dict(ex.get('column_mappings') or {})
        code_key = col_map.get('loan_pool_code')
        bal_key = col_map.get('current_balance')
        dq_key = col_map.get('days_delinquent')
        if not (code_key and bal_key and dq_key):
            continue
        path = _resolve_extract_path(ex.get('file_pattern'), search_dirs, snapshot_date, config)
        if not path or path in seen_paths:
            continue
        seen_paths.add(path)
        per = _neg_share_period_from_path(path)
        if per is not None and per != target:
            continue
        try:
            hr = int(ex.get('header_row') or 0)
        except (TypeError, ValueError):
            hr = 0
        pd_header = hr - 1 if hr > 1 else 0
        has_header = ex.get('has_header', True)
        ext_lc = os.path.splitext(path)[1].lower()
        try:
            if ext_lc == '.csv':
                fdf = _read_csv_any(path, header=pd_header if has_header else None)
            else:
                fdf = pd.read_excel(path, header=pd_header if has_header else None)
        except Exception as _e:
            print(f"    Snapshot DQ: could not read '{os.path.basename(path)}' ({_e}).")
            continue
        if has_header:
            norm_cols = {_norm_hdr(c): c for c in fdf.columns}

            def _series(key):
                actual = norm_cols.get(_norm_hdr(key))
                if actual is None and key in fdf.columns:
                    actual = key
                return fdf[actual] if actual is not None else None
        else:
            def _series(key):
                try:
                    return fdf.iloc[:, int(key)]
                except (KeyError, IndexError, ValueError):
                    return None
        codes_ser = _series(code_key)
        bal_ser = _series(bal_key)
        dq_ser = _series(dq_key)
        if codes_ser is None or bal_ser is None or dq_ser is None:
            continue
        files_used.append(os.path.basename(path))
        for i in range(len(fdf)):
            code = str(codes_ser.iloc[i]).strip()
            if not code or code.lower() == 'nan':
                continue
            bal = _ehp._clean_balance(bal_ser.iloc[i])
            days = _dqe._coerce_days(dq_ser.iloc[i])
            rec = by_code.setdefault(code, {'total_balance': 0.0, 'dq_amount': 0.0, 'n_dq': 0})
            rec['total_balance'] += bal
            if days is not None and days >= threshold:
                rec['dq_amount'] += bal
                rec['n_dq'] += 1

    if not files_used or not by_code:
        return 0

    rows = []
    tot_dq = 0.0
    n_dq = 0
    for code, rec in by_code.items():
        tb = round(rec['total_balance'], 2)
        dq = round(rec['dq_amount'], 2)
        rows.append({
            'loan_code': code,
            'total_balance': tb,
            'dq_amount': dq,
            'dq_pct': round(dq / tb, 8) if tb else None,
        })
        tot_dq += dq
        n_dq += int(rec['n_dq'])

    try:
        written = _dqp.upsert_month(cu, snap_iso, rows, source='loan_extract_report')
    except Exception as _e:
        print(f"    Snapshot DQ derivation skipped (upsert): {_e}")
        return 0
    print(f"    Snapshot DQ derived from extract(s) {files_used}: {n_dq} loan(s) "
          f">= {threshold} days, ${tot_dq:,.2f} delinquent across {len(rows)} "
          f"loan code(s) -> wrote {written} row(s) for {snap_iso}.")
    return written


def generate_report(client_name, snapshot_date=None, reports=None):
    """Generate CECL reports for a client.

    Args:
        client_name: Config name (e.g. 'franklin', 'ontario', 'maple')
        snapshot_date: Optional YYYY-MM-DD string; defaults to latest in DB
        reports: Optional list of report types to generate (e.g. ['tct', 'vizo', 'vizo_supp']).
                 If None, uses the client config 'reports' section.

    Returns:
        List of output file paths that were saved.
    """
    config = load_config(client_name)
    cu = config['credit_union']
    grades = config['credit_grades']
    no_score = config.get('no_score_label', 'Not Reported')

    # Drop any grade in credit_grades that duplicates the no_score_label so
    # downstream report tabs don't render two "Not Reported" rows/columns.
    if grades:
        grades = [g for g in grades if (g.get('label') or '').strip() != no_score]
        config = dict(config)
        config['credit_grades'] = grades

    # Determine snapshot date
    if not snapshot_date:
        snapshot_date = latest_date(cu)
        if not snapshot_date:
            print(f"  No data found for {cu}")
            return []

    print(f"\n{'='*60}")
    print(f"  Generating reports for {cu} - {snapshot_date}")
    print(f"{'='*60}")

    _cp = config.get('credit_pull') or {}
    _cp_asof = _cp.get('pull_as_of_date')
    if not _cp_asof:
        _m = re.search(r'(20\d{2})[-_ ](\d{1,2})', str(_cp.get('uploaded_filename') or ''))
        if _m:
            _cp_asof = f"{_m.group(1)}-{int(_m.group(2)):02d}-01"
    if _cp_asof:
        try:
            _pt = pd.Timestamp(_cp_asof)
            _st = pd.Timestamp(snapshot_date)
            _mo = (_st.year - _pt.year) * 12 + (_st.month - _pt.month)
            if _mo > 3:
                print(f"  WARNING: credit pull is {_mo} months older than the "
                      f"report period ({_pt.date()} vs snapshot {_st.date()}). "
                      f"'Current' credit scores/grades and the credit-migration "
                      f"classification are based on this STALE pull. Upload a "
                      f"current credit pull and re-import to refresh them.")
        except Exception:
            pass

    audit = get_audit_logger()
    audit.info("BEGIN report generation | client=%s | cu=%s | date=%s | types=%s",
               client_name, cu, snapshot_date, reports or "config-default")

    # Load loan data
    df = load_loans(cu, snapshot_date, config)
    if df.empty:
        print(f"  No loan data for {snapshot_date}")
        return []

    # Business Risk Rating overrides (per-pool). When the CU's config
    # marks any pool ``brr: true`` and provides ``business_risk_ratings``
    # rules, loans in those pools are bucketed by their analyst-assigned
    # rating instead of a FICO score. Empty / missing config falls back
    # to pure-FICO behavior for every loan.
    brr_pools = {
        (p or {}).get('name') for p in (config.get('pools') or [])
        if (p or {}).get('brr') and (p or {}).get('name')
    }
    brr_rules = config.get('business_risk_ratings') or []
    if brr_pools and brr_rules:
        print(
            f"  BRR pools active: {sorted(brr_pools)} "
            f"({len(brr_rules)} rule(s))"
        )
    # Prior-snapshot BRR lookup for quarter-over-quarter rating
    # migration. The engine's BRR override sets current_grade from the
    # active snapshot's business_risk_rating; without a prior snapshot
    # to compare against, it would also write that same value into
    # original_grade (== Unchanged for every BRR loan). Looking up the
    # same loan's BRR at the most recent strictly-earlier snapshot in
    # ``monthly_loan_data`` lets us populate original_grade with the
    # *prior* rating so the Risk Change tab shows real migration.
    # First-report baselines (no prior snapshot) get an empty lookup and
    # fall back to current==original (Unchanged), which matches the
    # established baseline-then-track pattern requested by analysts.
    prior_brr_lookup = {}
    if brr_pools and brr_rules:
        prior_brr_lookup = _load_prior_brr_lookup(cu, snapshot_date, brr_pools)
        if prior_brr_lookup:
            print(
                f"  Prior-period BRR lookup: {len(prior_brr_lookup):,} "
                f"loan(s) from prior snapshot"
            )
        else:
            print("  Prior-period BRR lookup: none (baseline run — no prior snapshot)")
    df = calculate_cecl(df, grades, no_score,
                        brr_rules=brr_rules, brr_pools=brr_pools,
                        prior_brr_lookup=prior_brr_lookup)

    # Load historical data
    hist = load_historical_data(config)

    # Honor CU-requested charge-off / recovery removals
    # (config['chargeoff_exclusions']) without editing source files or
    # re-importing — subtract them from the assembled history in place.
    _apply_chargeoff_exclusions(hist, config)

    # Resolve data-derived Other Allowance Considerations (e.g. unfunded
    # commitments computed from HELOC available credit) so both the Vizo
    # and TCT builders read an up-to-date balance for this snapshot.
    _resolve_dynamic_oac(config, snapshot_date)

    # Clamp historical data to the report period. The 5300 backfill
    # (and DB-overlay tables) commonly contain quarter-end snapshots
    # for periods AFTER the report's snapshot_date when a CU's wizard
    # has been run multiple times across quarters. Without this clamp
    # a 12/31/2025 report ends up with a 2026 column on Display Hist
    # Bal (last year is auto-labeled "YTD"), and 2026 charge-offs /
    # recoveries / DQ bleed into life-of-loan rate denominators.
    try:
        snap_year = int(str(snapshot_date)[:4])
        snap_month = int(str(snapshot_date)[5:7])
    except (TypeError, ValueError):
        snap_year = None
        snap_month = None
    if snap_year is not None:
        def _drop_future(by_year):
            if not isinstance(by_year, dict):
                return by_year
            for y in [k for k in by_year if isinstance(k, int) and k > snap_year]:
                by_year.pop(y, None)
            return by_year

        for k in ('chargeoffs', 'recoveries', 'avg_balances', 'dq_pct'):
            _drop_future(hist.get(k))

        def _drop_future_month(by_ym):
            if not isinstance(by_ym, dict):
                return by_ym
            drop = []
            for k in by_ym:
                try:
                    yy, mm = int(k[0]), int(k[1])
                except (TypeError, ValueError, IndexError):
                    continue
                if yy > snap_year or (yy == snap_year and mm > snap_month):
                    drop.append(k)
            for k in drop:
                by_ym.pop(k, None)
            return by_ym

        for k in ('co_monthly', 'rc_monthly'):
            _drop_future_month(hist.get(k))

        if isinstance(hist.get('years'), list):
            hist['years'] = [y for y in hist['years'] if y <= snap_year]

        # Trim monthly_balances DataFrame to dates <= snapshot_date
        mb = hist.get('monthly_balances')
        try:
            if mb is not None and not mb.empty and 'date' in mb.columns:
                snap_ts = pd.Timestamp(snapshot_date)
                hist['monthly_balances'] = mb[mb['date'] <= snap_ts].reset_index(drop=True)
        except Exception:
            pass

    # Load impaired data from WARM working file
    impaired = load_impaired_data(config, snapshot_date)
    if impaired:
        hist['impaired'] = impaired
        # WARM workbooks ship with multi-year CO / Recoveries / DQ% +
        # per-pool per-grade hist balance time series. Fold those into
        # the top-level hist[] keys so the Display HIst Bal and Display
        # CO-Recov-DQ tabs render the full WARM history even when the
        # CU has no DB backfill / file history populated.
        _overlay_warm_history_into_hist(hist, snapshot_date)
    else:
        # No WARM file — try loading hist_bal_data from the prior TCT report
        prior = load_prior_tct_hist_bal(config, snapshot_date)
        if prior:
            # Extend with intermediate months from the monthly balances file
            hbd = prior.get('hist_bal_data', {})
            if hbd:
                # Snapshot WARM-reported pool totals for the snapshot month
                # BEFORE extension overwrites them with loan-extract sums.
                # These represent the authoritative monthly balance per pool
                # (used for Risk Change "Total in Pool" / adjustments).
                snap_ts = pd.Timestamp(snapshot_date)
                snap_ym = snap_ts.to_period('M')
                warm_snap = {}
                for pool, pdata in hbd.items():
                    dates = pdata.get('dates') or []
                    tots = pdata.get('total') or []
                    best_idx = None
                    for i, d in enumerate(dates):
                        try:
                            d_ym = pd.Timestamp(d).to_period('M')
                        except Exception:
                            continue
                        if d_ym <= snap_ym:
                            best_idx = i
                        if d_ym == snap_ym:
                            break
                    if best_idx is not None and best_idx < len(tots):
                        try:
                            warm_snap[str(pool).strip()] = float(tots[best_idx])
                        except (TypeError, ValueError):
                            continue
                prior['warm_snapshot_balances'] = warm_snap

                extend_hist_bal_with_monthly(hbd, hist.get('monthly_balances'))
                extend_hist_bal_with_db(hbd, df, snapshot_date, grades, config)
                n_dates = max(len(d.get('dates', [])) for d in hbd.values())
                print(f"    Extended hist bal: {len(hbd)} pools, {n_dates} months")

            # Merge current-year CO/recovery from file-parsed data into
            # the prior report's warm_co/warm_rc so we get the authoritative
            # prior-year values plus fresh current-year data.
            snap_year = int(snapshot_date[:4])
            if prior.get('warm_co') is not None:
                cur_co = hist.get('chargeoffs', {}).get(snap_year, {})
                cur_rc = hist.get('recoveries', {}).get(snap_year, {})
                if cur_co:
                    prior['warm_co'][snap_year] = cur_co
                if cur_rc:
                    prior['warm_rc'][snap_year] = cur_rc
                # Compute net for current year
                if cur_co or cur_rc:
                    net_yr = {}
                    all_pools = set(list(cur_co.keys()) + list(cur_rc.keys()))
                    for p in all_pools:
                        net_yr[p] = cur_co.get(p, 0) - cur_rc.get(p, 0)
                    prior['warm_net'][snap_year] = net_yr

                # Merge current-year monthly CO/RC data before recalculating
                # ACL totals so windowing can use monthly granularity
                co_monthly_file = hist.get('co_monthly', {})
                rc_monthly_file = hist.get('rc_monthly', {})
                wco_m = prior.get('warm_co_monthly', {})
                wrc_m = prior.get('warm_rc_monthly', {})
                if co_monthly_file or rc_monthly_file:
                    for ym, pools_d in co_monthly_file.items():
                        if ym[0] >= snap_year:
                            wco_m[ym] = pools_d
                    for ym, pools_d in rc_monthly_file.items():
                        if ym[0] >= snap_year:
                            wrc_m[ym] = pools_d
                    if wco_m:
                        prior['warm_co_monthly'] = wco_m
                    if wrc_m:
                        prior['warm_rc_monthly'] = wrc_m

                # Recalculate ACL totals using acl_months window with
                # monthly-level precision for the earliest (partial) year
                acl_map = prior.get('acl_months', {})
                snap_month = int(snapshot_date[5:7])
                for pool in set(p for yr in prior['warm_co'].values()
                                for p in yr):
                    pool_acl = acl_map.get(pool, 36)
                    earliest_abs = (snap_year * 12 + snap_month) - pool_acl + 1
                    earliest_yr = (earliest_abs - 1) // 12
                    earliest_mo = earliest_abs - earliest_yr * 12

                    co_tot = 0
                    rc_tot = 0
                    for y in prior['warm_co']:
                        if y < earliest_yr:
                            continue
                        if y == earliest_yr:
                            # Partial year: sum only months in the window
                            partial = 0
                            has_monthly = False
                            for m in range(earliest_mo, 13):
                                v = wco_m.get((y, m), {}).get(pool, 0)
                                if v:
                                    has_monthly = True
                                partial += v
                            if has_monthly:
                                co_tot += partial
                            else:
                                full = prior['warm_co'].get(y, {}).get(pool, 0)
                                months_in = 12 - earliest_mo + 1
                                co_tot += full * months_in / 12 if full else 0
                        else:
                            co_tot += prior['warm_co'].get(y, {}).get(pool, 0)

                    for y in prior['warm_rc']:
                        if y < earliest_yr:
                            continue
                        if y == earliest_yr:
                            partial = 0
                            has_monthly = False
                            for m in range(earliest_mo, 13):
                                v = wrc_m.get((y, m), {}).get(pool, 0)
                                if v:
                                    has_monthly = True
                                partial += v
                            if has_monthly:
                                # Monthly RC may be stored negative; align sign
                                full_year = prior['warm_rc'].get(y, {}).get(pool, 0)
                                if full_year and (full_year > 0) != (partial > 0):
                                    partial = -partial
                                rc_tot += partial
                            else:
                                full = prior['warm_rc'].get(y, {}).get(pool, 0)
                                months_in = 12 - earliest_mo + 1
                                rc_tot += full * months_in / 12 if full else 0
                        else:
                            rc_tot += prior['warm_rc'].get(y, {}).get(pool, 0)

                    prior['warm_co_totals'][pool] = co_tot
                    prior['warm_rc_totals'][pool] = rc_tot
                    prior['warm_net_co'][pool] = co_tot - rc_tot

                n_co_yr = len(prior['warm_co'])
                print(f"    Merged CO/RC: {n_co_yr} years "
                      f"(added {snap_year} from file-parsed data)")

                # ── Overlay prior-report CO/Rc onto top-level hist keys ──
                # Going forward, historical (prior-year) data should come from
                # the prior TCT report, not from re-parsing source files. Only
                # the current snapshot year is taken from raw-file parsing.
                # This makes prior-year totals stable across runs and aligned
                # with the previously-validated report.
                #
                # User can force a full re-parse of historicals by deleting
                # the prior report from the Reports/ folder.
                hist['chargeoffs'] = {
                    y: dict(pools) for y, pools in prior['warm_co'].items()
                }
                hist['recoveries'] = {
                    y: dict(pools) for y, pools in prior['warm_rc'].items()
                }
                hist['co_monthly'] = {
                    ym: dict(pools)
                    for ym, pools in prior.get('warm_co_monthly', {}).items()
                }
                hist['rc_monthly'] = {
                    ym: dict(pools)
                    for ym, pools in prior.get('warm_rc_monthly', {}).items()
                }
                hist['years'] = sorted(set(hist['chargeoffs'])
                                       | set(hist['recoveries'])
                                       | set(hist.get('avg_balances') or {}))
                tot_co = sum(sum(p.values())
                             for p in hist['chargeoffs'].values())
                tot_rc = sum(sum(p.values())
                             for p in hist['recoveries'].values())
                print(f"    Historical CO/Rc sourced from prior TCT report: "
                      f"${tot_co:,.2f} CO / ${tot_rc:,.2f} Rc "
                      f"(years {hist['years'][0]}-{hist['years'][-1]})")

            # Merge current-year DQ% from file-parsed data
            file_dq = hist.get('dq_pct', {})
            if file_dq and prior.get('warm_dq_pct') is not None:
                cur_dq = file_dq.get(snap_year, {})
                if cur_dq:
                    prior['warm_dq_pct'][snap_year] = cur_dq

            hist['impaired'] = prior
        else:
            # No WARM file and no prior TCT report — build a fresh hist_bal_data
            # from the monthly balances workbook + current snapshot grades so the
            # Historical Trends Balance tab still has data to display.
            fresh = build_hist_bal_from_monthly(
                hist.get('monthly_balances'), df, snapshot_date, grades, config,
            )
            if fresh:
                hbd = fresh.get('hist_bal_data', {})
                n_dates = max((len(d.get('dates', [])) for d in hbd.values()), default=0)
                print(f"    Built fresh hist bal from monthly file: "
                      f"{len(hbd)} pools, {n_dates} months")
                hist['impaired'] = fresh

    # ── Load standalone Impaired Loans file (if available) ──
    standalone_imp = load_standalone_impaired(config, snapshot_date, df)
    if standalone_imp:
        imp = hist.get('impaired', {})
        # Overlay fresh impaired data onto whatever was loaded
        imp['acl_impaired'] = standalone_imp['acl_impaired']
        imp['spec_id_by_pool'] = standalone_imp['spec_id_by_pool']
        imp['total_spec_id'] = standalone_imp['total_spec_id']
        hist['impaired'] = imp

    # ── Load wizard-entered impaired loans (highest precedence) ──
    # When the user has entered/uploaded impaired loans in the setup
    # wizard, those rows are the most current source of truth and
    # override anything loaded from WARM / standalone file / prior TCT
    # baseline. The wizard's impaired_parser has already resolved
    # loan_pool + credit_grade via lookup against the loan-data extract.
    wizard_imp = load_wizard_impaired(config)
    if wizard_imp:
        imp = hist.get('impaired', {})
        imp['acl_impaired'] = wizard_imp['acl_impaired']
        imp['spec_id_by_pool'] = wizard_imp['spec_id_by_pool']
        imp['total_spec_id'] = wizard_imp['total_spec_id']
        hist['impaired'] = imp

    # ── 5300 DQ fallback ──
    # When the credit union has no historical delinquency rows in the
    # DB but has a charter number on file, pull DQ% history directly
    # from NCUA Form 5300 and write it to
    # ``loan_code_delinquency_history``. Subsequent runs find the rows
    # already in place and skip the fetch. Opt out per CU via
    # ``cfg['delinquency']['use_5300_fallback']: false``. Requires
    # ``cfg['charter_number']`` to know which CU to query.
    dq_cfg = config.get('delinquency') or {}
    if (dq_cfg.get('use_5300_fallback', True)
            and config.get('charter_number')):
        try:
            from cecl_ui.services import delinquency_hist_processor as _dqp
            try:
                _hv = _dqp.history_matrix(config['credit_union'])
                _existing_dq = set((_hv or {}).get('months') or [])
            except Exception:  # noqa: BLE001
                _hv = {}
                _existing_dq = set()
            # Detect stale-zero rows: a prior backfill run wrote DQ
            # rows for every expected quarter but with dq_amount=0
            # across the board (e.g. the canonical CSV was incomplete
            # when the backfill ran, or the wrong Solr core was used).
            # Without this check the auto-trigger silently no-ops on
            # subsequent runs because rows exist -- exactly Central
            # Keystone FCU / Census FCU's symptom (2026-06-19).
            _dq_total = 0.0
            for _m, _cells in (_hv or {}).get('cells', {}).items():
                for _cd, _cell in (_cells or {}).items():
                    _dq_total += float((_cell or {}).get('amount') or 0.0)
                    if _dq_total > 0:
                        break
                if _dq_total > 0:
                    break
            _stale_zero = bool(_existing_dq) and _dq_total == 0.0
            _needs_refill = (not _existing_dq) or _stale_zero
            if _needs_refill:
                from cecl_ui.services import (
                    solr_5300_delq_backfill as _solr_dq,
                )
                _snap_iso = pd.Timestamp(snapshot_date).date().isoformat()
                _months_back = int(dq_cfg.get('history_months') or 84)
                _solr_url = (
                    dq_cfg.get('solr_url')
                    or 'http://searchserver1.tctrisk.com:8983/solr'
                )
                _solr_core = dq_cfg.get('solr_core') or 'ncua'
                if _stale_zero:
                    print(
                        f"    5300 DQ backfill: existing rows total "
                        f"$0 across {len(_existing_dq)} quarter(s); "
                        f"re-running with overwrite=True to recover."
                    )
                _res = _solr_dq.backfill_missing_delinquency_quarters(
                    config['credit_union'],
                    int(config['charter_number']),
                    _solr_url, _solr_core,
                    _snap_iso, _months_back,
                    overwrite=_stale_zero,
                )
                if _res.get('ok'):
                    _filled = len(_res.get('months_filled') or [])
                    print(
                        f"    5300 DQ backfill: filled {_filled} "
                        f"quarter(s), wrote "
                        f"{_res.get('rows_written', 0)} row(s) for "
                        f"charter {config['charter_number']}"
                    )
                else:
                    print(
                        f"    5300 DQ fallback skipped: "
                        f"{_res.get('error')}"
                    )
        except Exception as _e:  # noqa: BLE001
            print(f"    5300 DQ fallback skipped: {_e}")

    try:
        _derive_snapshot_dq_from_extracts(config, snapshot_date)
    except Exception as _e:
        print(f"    Snapshot DQ derivation skipped: {_e}")

    # ── Overlay DQ% from loan_code_delinquency_history table ──
    # The wizard's "Historical DQ" step writes rows here from three
    # sources (loan-extract derivation, 5300 backfill, manual entry).
    # DB rows take precedence over the WARM-derived warm_dq_pct: any
    # (year, pool) cell present in the DB overwrites the WARM value;
    # other cells are left alone.
    db_dq = _load_dq_history_from_db(config)
    if db_dq:
        imp = hist.get('impaired') or {}
        existing = imp.get('warm_dq_pct') or {}
        for yr, by_pool in db_dq.items():
            existing.setdefault(yr, {}).update(by_pool)
        imp['warm_dq_pct'] = existing
        hist['impaired'] = imp
        n_cells = sum(len(v) for v in db_dq.values())
        print(f"    Overlaid DQ% from loan_code_delinquency_history: "
              f"{len(db_dq)} year(s), {n_cells} pool-year cell(s).")
    else:
        # Soft-miss diagnostic. If the CU has a charter number on file
        # and the DQ overlay returned nothing, the Display CO-Recov-DQ
        # tab's DQ% column will render blank/zero for every pool/year.
        # Surface a clear warning so the user can tell the difference
        # between "genuinely no DQ history" and "the data pipeline is
        # silently failing". Typical causes: stale-zero rows blocked
        # the auto-trigger (now fixed by overwrite=_stale_zero above);
        # Solr was unreachable; the canonical CSV is missing fields
        # for this CU's 5300 form variant.
        if config.get('charter_number'):
            print(
                f"    WARNING: DQ% overlay produced no cells for "
                f"{config.get('credit_union')!r}. Display CO-Recov-DQ "
                f"tab will show blank DQ% across all pools/years. "
                f"Inspect loan_code_delinquency_history for this CU "
                f"and verify the 5300 Solr fetch is returning non-zero "
                f"values for the canonical DQ field codes."
            )

    # ── Fallback: load impaired data from prior TCT baseline ──
    # When the source WARM has no 'Impaired Loans' tab and there's no
    # standalone Impaired Loans file, pull from the previously-generated
    # TCT model (which carries 'Impaired Loans' + 'Impaired Loans Pivot' tabs).
    # Even when impaired data is already present from a more current source
    # (wizard / standalone file), still call the baseline loader so that its
    # historical warm_dq_pct / warm_co / warm_rc data from the
    # 'Display CO-Recov -DQ' tab gets merged in (impaired keys are skipped
    # to preserve precedence).
    imp_now = hist.get('impaired', {})
    has_imp = bool(imp_now.get('acl_impaired')) or bool(imp_now.get('spec_id_by_pool'))
    baseline_imp = load_impaired_from_tct_baseline(config, snapshot_date)
    if baseline_imp:
        if not has_imp:
            imp_now.update(baseline_imp)
        else:
            # Only merge historical CO/RC/DQ keys; preserve current impaired data.
            for k, v in baseline_imp.items():
                if k.startswith('warm_') and k not in imp_now:
                    imp_now[k] = v
        hist['impaired'] = imp_now

    # ── Carry-forward WARM Months from prior reports (per-pool ACL months) ──
    # Long-term plan: TCT reports are replacing the legacy CECL-Migration-WARM
    # xlsx files. Going forward, each new quarter inherits its per-pool WARM
    # Months from the most recent prior TCT report; the legacy WARM xlsx is
    # only consulted as a fallback during phase-out (or for the very first TCT
    # generation when no prior TCT exists yet).
    #
    # Override priority (highest wins):
    #   1. Current quarter's CECL-Migration-WARM xlsx (if user maintains one)
    #   2. Most recent prior TCT report's "> Detail_HIst Balances" WARM column
    #   3. Most recent prior CECL-Migration-WARM xlsx (legacy fallback)
    imp_now = hist.get('impaired', {})
    cur_acl = dict(imp_now.get('acl_months', {}) or {})
    pool_order = imp_now.get('pool_order', []) or list(
        imp_now.get('hist_bal_data', {}).keys()
    )
    pre_count = len(cur_acl)

    needs_fill = [p for p in pool_order if p not in cur_acl]
    if needs_fill or not cur_acl:
        # 2. Prior TCT report
        prior_tct = _find_prior_tct_report(config, snapshot_date)
        added_tct = 0
        if prior_tct:
            tct_acl = _load_acl_months_from_tct(prior_tct)
            for pool, months in tct_acl.items():
                if pool not in cur_acl:
                    cur_acl[pool] = months
                    added_tct += 1
            if added_tct:
                print(f"    Carried forward WARM Months from prior TCT "
                      f"({os.path.basename(prior_tct)}): {added_tct} pools")

        # 3. Legacy WARM xlsx (phase-out fallback)
        still_missing = [p for p in pool_order if p not in cur_acl]
        if still_missing or not cur_acl:
            prior_warm = _find_prior_warm_xlsx(config, snapshot_date)
            if prior_warm:
                prior_acl = _load_acl_months_from_warm_xlsx(prior_warm)
                added_warm = 0
                for pool, months in prior_acl.items():
                    if pool not in cur_acl:
                        cur_acl[pool] = months
                        added_warm += 1
                if added_warm:
                    print(f"    Carried forward WARM Months from legacy WARM xlsx "
                          f"({os.path.basename(prior_warm)}): {added_warm} pools")

        if len(cur_acl) > pre_count:
            imp_now['acl_months'] = cur_acl
            hist['impaired'] = imp_now

    # ── YAML overrides for per-pool ACL months and risk_rated flag ──
    # Lets users edit pool settings post-setup without touching WARM files.
    cfg_acl_overrides = config.get('acl_months_by_pool') or {}
    if cfg_acl_overrides:
        imp_now = hist.get('impaired', {})
        cur_acl = dict(imp_now.get('acl_months', {}) or {})
        applied = 0
        for pool, months in cfg_acl_overrides.items():
            try:
                m = int(months)
            except (TypeError, ValueError):
                continue
            if m > 0:
                cur_acl[pool] = m
                applied += 1
        if applied:
            imp_now['acl_months'] = cur_acl
            hist['impaired'] = imp_now
            print(f"    YAML acl_months_by_pool override: {applied} pool(s)")

    cfg_nrr = set(config.get('not_risk_rated', []) or [])
    if cfg_nrr:
        imp_now = hist.get('impaired', {})
        rr_map = dict(imp_now.get('risk_rated', {}) or {})
        for pool in cfg_nrr:
            rr_map[pool] = False
        imp_now['risk_rated'] = rr_map
        hist['impaired'] = imp_now

    # ── 5300 ACL fallback ──
    # When the credit union's monthly-balance / ACL source does not
    # include an ACL row for the snapshot quarter (or stores $0), the
    # report engine can pull the value directly from NCUA Form 5300.
    # Opt-in per CU via ``cfg['acl']['use_5300_fallback']``; requires
    # ``cfg['charter_number']`` to know which CU to query. The fetcher
    # probes AAS0048 (current 5300 / 5300SF "Allowance for Credit
    # Losses on Loans") by default — per user direction this is the
    # canonical source. Power users can override the probe order via
    # ``cfg['acl']['solr_fields']`` (a list, e.g. ``["AAS0048",
    # "A007"]`` to enable legacy backstops for historical quarters).
    # Only missing or zero quarter-end values are filled — explicit
    # user values from the file / YAML history always win.
    acl_cfg = config.get('acl') or {}
    if acl_cfg.get('use_5300_fallback') and config.get('charter_number'):
        try:
            from cecl_ui.services import solr_5300_acl as _solr_acl
            # Per-CU / per-run override: cfg['acl']['solr_fields'] lets
            # the user (via Run New Quarter wizard) pick a specific
            # 5300 field or supply an explicit probe order. Empty /
            # missing list falls back to the default probe order baked
            # into ``_solr_acl._ACL_FIELD_ORDER`` (currently AAS0048).
            _solr_field_order = [
                str(f).strip().upper()
                for f in (acl_cfg.get('solr_fields') or [])
                if str(f).strip()
            ]
            if not _solr_field_order:
                _solr_field_order = list(_solr_acl._ACL_FIELD_ORDER)
            # NOTE: legacy field backstops (A007/A718A3/A718A5/A719)
            # are intentionally NOT auto-appended. Per user direction
            # AAS0048 is the only canonical source; appending A007 etc.
            # produced incorrect values (A007 is a related-but-different
            # aggregate, not ACL-on-Loans). Users who want backstops
            # for historical quarters list them explicitly in
            # ``acl.solr_fields``.
            alll_by_date = hist.get('alll_by_date') or {}
            snap_dt = pd.Timestamp(snapshot_date)

            # Decide which quarter-ends to probe. Always probe the
            # snapshot quarter. Also probe any quarter-end already in
            # alll_by_date that currently holds 0 (those came from the
            # file with a blank/zero cell).
            from datetime import date as _date

            def _qe_for(ts):
                ts = pd.Timestamp(ts)
                m = ts.month
                qm = 3 * ((m - 1) // 3 + 1)
                yr = ts.year
                last = {3: 31, 6: 30, 9: 30, 12: 31}[qm]
                return _date(yr, qm, last).isoformat()

            wanted: set[str] = {_qe_for(snap_dt)}
            for dt, val in list(alll_by_date.items()):
                try:
                    if float(val) == 0.0:
                        wanted.add(_qe_for(dt))
                except (TypeError, ValueError):
                    continue

            results = _solr_acl.fetch_acl_history(
                config['charter_number'],
                sorted(wanted),
                fields=tuple(_solr_field_order),
            ) or {}

            filled = 0
            field_used = None
            for qe_iso, info in results.items():
                ts = pd.Timestamp(qe_iso)
                cur = alll_by_date.get(ts)
                if cur is None or float(cur or 0) == 0.0:
                    alll_by_date[ts] = float(info['value'])
                    field_used = info.get('field') or field_used
                    filled += 1
            if filled:
                hist['alll_by_date'] = alll_by_date
                print(f"    5300 ACL fallback: filled {filled} quarter(s) "
                      f"for charter {config['charter_number']} "
                      f"(field {field_used or 'mixed'}; "
                      f"probe order: {', '.join(_solr_field_order)})")
            else:
                print(f"    5300 ACL fallback: no non-zero ACL value "
                      f"found in Solr for charter {config['charter_number']} "
                      f"across {len(wanted)} quarter(s) "
                      f"(probed {', '.join(_solr_field_order)})")
        except Exception as _e:  # noqa: BLE001
            print(f"    5300 ACL fallback skipped: {_e}")

    # ── Set ACL Balance from Monthly loan balances file (ALLL Balance row) ──
    alll_by_date = hist.get('alll_by_date', {})
    if alll_by_date:
        snap_dt = pd.Timestamp(snapshot_date)
        snap_ym = snap_dt.to_period('M')
        acl_bal = None
        for dt, val in alll_by_date.items():
            if pd.Timestamp(dt).to_period('M') == snap_ym:
                acl_bal = val
                break
        if acl_bal is not None:
            imp = hist.get('impaired', {})
            acl_sum = imp.get('acl_summary', {})
            acl_sum['acl_balance'] = acl_bal
            imp['acl_summary'] = acl_sum
            imp['acl_balance'] = acl_bal  # also at top level for report_vizo
            hist['impaired'] = imp
            print(f"    ACL Balance from monthly file: ${acl_bal:,.2f}")

    # ── Compute balance adjustments from monthly file vs loan file ──
    _compute_balance_adjustments(df, hist, config, snapshot_date)

    # ── Delinquency by credit-grade migration (fallback) ──
    # ``hist['impaired']['dq_by_status'|'dq_by_pool']`` feeds the
    # "Delinquency by Credit Grade Migration" pie on every Risk Change tab
    # and has historically had a single producer: the ``DQ Data Entry`` tab
    # of a legacy CECL-Migration-WARM workbook. Wizard-onboarded CUs never
    # get one, so the pie plotted literal zeros (docs/pdf_migration/
    # 04_blank_charts.md). Derive the same split from the loan-level frame
    # plus ``days_delinquent`` off the loan extract when -- and only when --
    # the WARM did not supply it, so WARM-fed CUs are untouched.
    try:
        from dq_migration_split import fill_missing_dq_migration
        fill_missing_dq_migration(hist, config, snapshot_date, df, grades,
                                  no_score=no_score, workspace_root=BASE)
    except Exception as _dqe_exc:  # noqa: BLE001 - never block a report
        print(f"    DQ migration split skipped: {_dqe_exc}")

    # ── Charge-off by credit-grade migration (fallback) ──
    # Mirror of the DQ block above for ``co_by_status`` / ``co_by_pool``.
    # Recovers each charged-off loan's origination and at-charge-off scores
    # from the charge-off feed and the loan-snapshot history; refuses rather
    # than guesses when coverage is too thin. See
    # docs/pdf_migration/09_chargeoff_score_history.md.
    try:
        from co_migration_split import fill_missing_co_migration
        fill_missing_co_migration(hist, config, snapshot_date, grades,
                                  no_score=no_score)
    except Exception as _coe_exc:  # noqa: BLE001 - never block a report
        print(f"    CO migration split skipped: {_coe_exc}")

    # Determine which reports to generate
    if reports is None:
        rpt_cfg = config.get('reports', {})
        reports = [k for k, v in rpt_cfg.items() if v]
    if not reports:
        reports = ['tct']  # default fallback
    # The Vizo PDF is a rendering of the Vizo workbook -- requesting it implies
    # the Vizo report, so selecting it alone still produces output.
    if 'vizo_pdf' in reports and 'vizo' not in reports:
        reports = list(reports) + ['vizo']

    os.makedirs(RPT_DIR, exist_ok=True)
    saved = []
    failed_integrity = []

    # Optional data-driven PDF alongside the Vizo workbook. Snapshot pristine
    # inputs BEFORE the loop mutates config/hist (OAC expansion, Impr-Deter
    # stashes) so the from-data recompute starts clean. Enabled by the config
    # flag OR by 'vizo_pdf' appearing in the requested reports list.
    _want_vizo_pdf = ('vizo' in reports
                      and (bool(config.get('reports', {}).get('vizo_pdf'))
                           or 'vizo_pdf' in reports))
    # The ACL sidecar is written on EVERY vizo run so future quarters can diff
    # against it without opening this workbook.
    _want_vizo_sidecar = 'vizo' in reports
    if _want_vizo_pdf or _want_vizo_sidecar:
        import copy as _copy
        _pdf_config = _copy.deepcopy(config)
        _pdf_hist = _copy.deepcopy(hist)

    for rpt_type in reports:
        if rpt_type == 'vizo_pdf':
            continue  # a modifier on the 'vizo' report, not a tab of its own
        try:
            if rpt_type == 'tct':
                wb, fname = compose_tct_new(client_name, snapshot_date, df, config, grades, hist)
            elif rpt_type == 'vizo':
                wb, fname = compose_vizo_main_new(client_name, snapshot_date, df, config, grades, hist)
            elif rpt_type == 'vizo_supp':
                wb, fname = compose_vizo_supp_new(client_name, snapshot_date, df, config, grades, hist)
            elif rpt_type == 'mgmt_adj_napkin':
                from report_mgmt_adj_napkin import compose_mgmt_adj_napkin
                wb, fname = compose_mgmt_adj_napkin(client_name, snapshot_date, df, config, grades, hist)
            elif rpt_type == 'acl_funding':
                from report_acl_funding import compose_acl_funding
                wb, fname = compose_acl_funding(client_name, snapshot_date, df, config, grades, hist)
            else:
                print(f"  Unknown report type: {rpt_type}")
                continue

            output_path = os.path.join(RPT_DIR, fname)

            # "All Loans" and "Risk Change-All Loans" tabs are intentionally
            # left unlocked so users can sort/filter. (Previously protected
            # with a password; removed per user request.)

            wb.save(output_path)
            print(f"  Saved {rpt_type}: {output_path}")
            saved.append(output_path)
            log_report_generation(client_name, cu, snapshot_date, rpt_type, output_path, success=True)

            # Post-processing: patch charts
            if rpt_type == 'vizo':
                try:
                    patch_dq_pie_zero_labels(output_path)
                    patch_impdet_charts(output_path)
                    patch_drawing_onecell_to_twocell(output_path)
                    patch_remove_chart_borders_and_axis_lines(output_path)
                    print(f"  Patched charts in {fname}")
                except Exception as e:
                    print(f"  Warning: Chart patching failed: {e}")

            if rpt_type == 'vizo' and _want_vizo_pdf:
                try:
                    from cecl_report_web.assembly import (
                        render_report_pdf_from_data)
                    pdf_bytes = render_report_pdf_from_data(
                        client_name, snapshot_date, _pdf_config,
                        grades=grades, hist=_pdf_hist, df=df)
                    pdf_path = os.path.join(
                        RPT_DIR, fname.rsplit('.xlsx', 1)[0] + '.pdf')
                    with open(pdf_path, 'wb') as _pf:
                        _pf.write(pdf_bytes)
                    print(f"  Saved vizo PDF: {pdf_path}")
                    saved.append(pdf_path)
                except Exception as _pdf_exc:  # noqa: BLE001
                    print(f"  Warning: vizo PDF generation failed: {_pdf_exc}")

            if rpt_type == 'vizo' and _want_vizo_sidecar:
                try:
                    from cecl_report_web import acl_store
                    from cecl_report_web import from_data as _fd
                    from report_vizo import compute_acl_environmental
                    _env = compute_acl_environmental(
                        df, grades, _pdf_config, _pdf_hist, snapshot_date)
                    if _env.get('acl_pools'):
                        _shape = _fd._acl_current_shape(
                            _env['acl_pools'], _env['acl_summary'],
                            _env['acl_impaired'])
                        acl_store.write_acl_snapshot(
                            RPT_DIR, cu, snapshot_date, _shape)
                        print(f"  Wrote ACL sidecar for {snapshot_date}")
                except Exception as _sc_exc:  # noqa: BLE001
                    print(f"  ACL sidecar skipped: {_sc_exc}")

            # Integrity gate. On 2026-08-31 a namespace bug in
            # patch_impdet_charts produced workbooks Excel refused to open,
            # and nothing caught it -- openpyxl loaded them fine and the zip
            # was intact. A report a client cannot open must never ship
            # silently, so validate every saved workbook and record failures
            # loudly. See report_integrity for what is checked.
            try:
                from report_integrity import check_and_report
                if not check_and_report(output_path, fname):
                    failed_integrity.append(fname)
                    log_report_generation(client_name, cu, snapshot_date,
                                          rpt_type, output_path, success=False)
            except Exception as e:  # noqa: BLE001
                print(f"  Warning: integrity check could not run: {e}")

        except Exception as e:
            print(f"  ERROR generating {rpt_type}: {e}")
            log_report_generation(client_name, cu, snapshot_date, rpt_type, None, success=False)
            import traceback
            traceback.print_exc()

    if saved:
        print(f"\n  {len(saved)} report(s) saved to {RPT_DIR}")
    else:
        print(f"\n  No reports were generated.")

    if failed_integrity:
        print("")
        print(f"  *** {len(failed_integrity)} report(s) FAILED the"
              f" integrity check and must not be delivered: ***")
        for _f in failed_integrity:
            print(f"      {_f}")
        print("      Diagnose with: python report_integrity.py <path> --excel")

    # Persist the resolved Other Allowance Consideration values so the Run
    # Reports "Other Allowance Considerations" override page can show the
    # latest calculated rate/balance/amount alongside any manual override.
    _oac_calc = config.get('_oac_calc')
    if _oac_calc:
        _persist_oac_calc(client_name, _oac_calc, snapshot_date)

    return saved


if __name__ == '__main__':
    import sys
    parser = argparse.ArgumentParser(
        description="Generate CECL reports (TCT, Vizo, Supplemental)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python generate_report.py --client franklin --date 2025-12-31
  python generate_report.py --client franklin --reports tct vizo vizo_supp
  python generate_report.py --all --date 2025-12-31
  python generate_report.py --list
        """,
    )
    parser.add_argument('--client', help='Client config name (e.g., "franklin")')
    parser.add_argument('--date', help='Snapshot date (YYYY-MM-DD), defaults to latest')
    parser.add_argument('--reports', nargs='+', choices=['tct', 'vizo', 'vizo_supp'],
                        help='Report types to generate (overrides config)')
    parser.add_argument('--all', action='store_true', help='Generate for all clients')
    parser.add_argument('--list', action='store_true', help='List available clients')
    args = parser.parse_args()

    if args.list:
        print("Available Clients:")
        print(f"  {'Config':20s}  {'Credit Union':40s}  {'Reports'}")
        print(f"  {'-'*20}  {'-'*40}  {'-'*20}")
        for c in list_clients():
            cfg = load_config(c)
            rpts = [k for k, v in cfg.get('reports', {}).items() if v]
            print(f"  {c:20s}  {cfg['credit_union']:40s}  {', '.join(rpts)}")
        sys.exit(0)

    if args.all:
        log_session_start('generate_report.py', f'--all --date={args.date}')
        clients = list_clients()
        print(f"Processing {len(clients)} client(s): {', '.join(clients)}")
        for client_name in clients:
            generate_report(client_name, args.date, args.reports)
        log_session_end('generate_report.py')
    elif args.client:
        log_session_start('generate_report.py', f'--client {args.client} --date={args.date} --reports={args.reports}')
        generate_report(args.client, args.date, args.reports)
        log_session_end('generate_report.py')
    else:
        parser.print_help()
