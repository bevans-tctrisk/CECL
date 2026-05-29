"""
generate_impdet_report.py  — Generate Improved / Deteriorated Loans Report

Creates a standalone Excel workbook with four tabs:
  1. Key-Deteriorated Type  — Description and methodology overview
  2. Improved Loans Summary  — Summary by pool of improved loans
  3. Deteriorated Loans Summary  — Summary by pool of deteriorated loans
  4. All Loans  — Loan-level detail with NCC status

Usage:
    python generate_impdet_report.py --client sample
    python generate_impdet_report.py --client sample --date 2025-12-31
"""

import os, re, argparse
from datetime import datetime
import numpy as np
import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers,
)
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text
from dotenv import load_dotenv
from cecl_engine import assign_credit_grade, build_grade_order
from import_data import derive_member_account, _normalize_col_map_for_no_header

load_dotenv()
# Honour CECL_WORKSPACE_ROOT so the data root can be decoupled from the
# code location; falls back to historical layout when the env var is unset.
BASE = os.environ.get('CECL_WORKSPACE_ROOT') or os.path.dirname(os.path.abspath(__file__))
CFG_DIR = os.path.join(BASE, 'client_configs')
RPT_DIR = os.path.join(BASE, 'Reports')
engine = create_engine(os.getenv('DATABASE_URL'))

# ── Styles ─────────────────────────────────────────────────────────
FNT_TITLE = Font(name='Calibri', bold=True, size=14, color='1B4F72')
FNT_SUB   = Font(name='Calibri', bold=True, size=12, color='1B4F72')
FNT_HDR   = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
FNT_BOLD  = Font(name='Calibri', bold=True, size=10)
FNT_NORM  = Font(name='Calibri', size=10)
FNT_DESC  = Font(name='Calibri', size=10, italic=True)
FNT_KEY   = Font(name='Calibri', bold=True, size=11)

FILL_HDR  = PatternFill('solid', fgColor='1B4F72')
FILL_IMP  = PatternFill('solid', fgColor='D5F5E3')
FILL_DET  = PatternFill('solid', fgColor='FADBD8')
FILL_TOT  = PatternFill('solid', fgColor='D6EAF8')
FILL_ALT  = PatternFill('solid', fgColor='EBF5FB')

THIN = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin'),
)
MONEY  = '#,##0.00'
NUM    = '#,##0'
PCT    = '0.00%'
CENTER = Alignment(horizontal='center', vertical='center')
LEFT   = Alignment(horizontal='left', vertical='center')
WRAP   = Alignment(horizontal='left', vertical='top', wrap_text=True)


# ── Helpers ────────────────────────────────────────────────────────
def load_config(client):
    path = os.path.join(CFG_DIR, f'{client}.yaml')
    with open(path, 'r', encoding='utf-8') as f:
        cfg = yaml.safe_load(f)
    excl = set((cfg.get('excluded_pools') or []))
    if excl:
        pm = cfg.get('pool_map') or {}
        cfg['pool_map'] = {
            k: ('Exclude' if v in excl else v) for k, v in pm.items()
        }
        if cfg.get('default_pool') in excl:
            cfg['default_pool'] = 'Exclude'
    return cfg


def load_loans(cu, snap, config=None):
    df = pd.read_sql(
        text("SELECT * FROM monthly_loan_data WHERE credit_union=:c AND snapshot_date=:s"),
        engine, params={"c": cu, "s": snap},
    )
    if df is None or df.empty or 'loan_pool' not in df.columns:
        return df
    excl = set((config.get('excluded_pools') or [])) if config else set()
    excl.add('Exclude')
    mask = df['loan_pool'].isin(excl)
    if mask.any():
        df = df.loc[~mask].copy()
    return df


def latest_date(cu):
    with engine.connect() as c:
        r = c.execute(
            text("SELECT MAX(snapshot_date) FROM monthly_loan_data WHERE credit_union=:c"),
            {"c": cu},
        ).fetchone()
    return str(r[0]) if r and r[0] else None


def snap_display(snap):
    """Format '2025-12-31' → '12/31/2025'."""
    d = datetime.strptime(snap, '%Y-%m-%d')
    return d.strftime('%m/%d/%Y')


def loan_ncc_status(orig_grade, cur_grade, grade_labels, n_top=3, no_score='Not Reported'):
    """Per-loan NCC status using the top_grades_double_drop logic."""
    if orig_grade not in grade_labels or cur_grade not in grade_labels:
        return "Unchanged"
    # Loans with no original score have no baseline — always Unchanged
    if orig_grade == no_score:
        return "Unchanged"
    j = grade_labels.index(orig_grade)   # original grade index (lower = better)
    i = grade_labels.index(cur_grade)    # current grade index
    if i > j:                            # current is worse
        if j < n_top and (i - j) < 2:
            return "Unchanged"           # small drop within top grades
        return "Deteriorated"
    elif i < j:                          # current is better
        return "Improved"
    return "Unchanged"


# ── WARM file data loading ─────────────────────────────────────────
# Column name mapping from WARM "All Loans" tab to internal DataFrame fields
_WARM_COL_MAP = {
    'Member #':                'member_number_raw',
    'Loan Suffix':             'loan_suffix',
    'Member #-Suffix':         'member_suffix',
    'Loan Type':               'loan_type',
    'Loan Pool':               'loan_pool',
    'Current Balance':         'current_balance',
    'Original Credit Score':   'original_fico_score',
    'Original Credit Grade':   'original_grade',
    'Current Credit Score':    'current_fico_score',
    'Current Credit Grade':    'current_grade',
    'Open Date':               'open_date',
    'Interest Rate':           'interest_rate',
    'Days Delinquent':         'days_delinquent',
    'Original Loan Amount':    'original_loan_amount',
    'Net Credit Change Status': 'warm_ncc_status',
}
# Columns 16-38 — keep as-is with shorter internal names
_WARM_EXTRA_MAP = {
    "Total Available Credit (Lines of Credit, Credit Cards and HELOC's)": 'total_available_credit',
    "Balance at Other Lender (2nd Mortgage and HELOC's)":                 'balance_other_lender',
    "Colateral Value (for unsecured loans, leave blank)":                 'collateral_value',
    'Total Loans':                      'total_loans',
    'LTV':                              'ltv',
    'Years loan has been on books':     'years_on_books',
    '$ of Pricipal Paid':               'principal_paid',
    '% of Pricipal Paid':               'pct_principal_paid',
    '1 year of payments Balance':       'yr1_balance',
    '2 years of payments Balance':      'yr2_balance',
    '3+ years of payments Balance':     'yr3_balance',
    'Total of 1+ years of payments made': 'total_1yr_payments',
    'Collateral Equity Count':          'collateral_eq_count',
    'Collateral Equity':                'collateral_equity',
    'Count of Unused Credit':           'unused_credit_count',
    '$ Unused Credit':                  'unused_credit',
    '% Unused Credit':                  'pct_unused_credit',
    'Amount at Risk':                   'amount_at_risk',
    'Contingency Risk':                 'contingency_risk',
    'Loans at Risk Count':              'loans_at_risk_count',
    'Contingency Risk Count':           'contingency_risk_count',
    'Total Risk Count':                 'total_risk_count',
    'Net Credit Change/Pool':           'ncc_pool',
}
_WARM_COL_MAP.update(_WARM_EXTRA_MAP)


def find_warm_file(config, snap):
    """Locate the CECL-Migration-WARM workbook for the given client/snapshot."""
    cu = config['credit_union']
    snap_prefix = snap[:7]  # "2025-12"

    # Directories to search
    data_dir = config.get('data_directory', '')
    if data_dir and not os.path.isabs(data_dir):
        data_dir = os.path.join(BASE, data_dir)
    fb_folder = config.get('credit_pull', {}).get('fallback_report_folder', '')
    if fb_folder and not os.path.isabs(fb_folder):
        fb_folder = os.path.join(BASE, fb_folder)

    search_dirs = []
    if data_dir and os.path.isdir(data_dir):
        search_dirs.append(data_dir)
    if fb_folder and fb_folder != data_dir and os.path.isdir(fb_folder):
        search_dirs.append(fb_folder)

    # Exact filename match first
    target_name = f"{snap_prefix} CECL-Migration-WARM - {cu}.xlsx"
    for sdir in search_dirs:
        for root, dirs, files in os.walk(sdir):
            for f in files:
                if f.startswith('~$') or f.upper().startswith('DNU'):
                    continue
                if f == target_name:
                    return os.path.join(root, f)

    # Fallback: pattern match
    pattern = re.compile(rf'^{re.escape(snap_prefix)}.*CECL-Migration-WARM.*\.xlsx$', re.IGNORECASE)
    for sdir in search_dirs:
        for root, dirs, files in os.walk(sdir):
            for f in files:
                if f.startswith('~$') or f.upper().startswith('DNU'):
                    continue
                if pattern.match(f):
                    return os.path.join(root, f)
    return None


def load_warm_all_loans(warm_path, config):
    """Read the 'All Loans' tab from the WARM workbook.

    Returns a DataFrame with all 38 columns mapped to internal field names,
    plus a 'member_number' column constructed from Member# + Suffix for DB matching.
    Returns None if the sheet doesn't exist or is empty.
    """
    try:
        # Use pandas read_excel — reads cached formula values, much faster than openpyxl
        # Header is on row 2 (0-indexed: row 1), skip row 1 (title), use cols A:AL (38)
        raw = pd.read_excel(
            warm_path,
            sheet_name='All Loans',
            header=1,        # 0-indexed: row 2 in Excel is header
            usecols=range(38),  # columns A through AL (0-37)
            engine='openpyxl',
        )
    except ValueError:
        # Sheet doesn't exist
        return None
    except Exception as e:
        print(f"    WARNING: Could not read WARM All Loans: {e}")
        return None

    if raw.empty:
        return None

    # Strip whitespace from column names
    raw.columns = [str(c).strip() for c in raw.columns]

    # Rename columns using our mapping
    rename = {k: v for k, v in _WARM_COL_MAP.items() if k in raw.columns}
    df = raw.rename(columns=rename)

    # Drop rows where both member and balance are missing (empty rows at end)
    id_col = 'member_number_raw' if 'member_number_raw' in df.columns else df.columns[0]
    df = df.dropna(subset=[id_col, 'current_balance'], how='all')

    # Construct DB-compatible member_number: str(Member#) + str(Suffix).zfill(suffix_length)
    suffix_len = config.get('account_suffix_length', 3)
    if 'member_number_raw' in df.columns and 'loan_suffix' in df.columns:
        df['member_number'] = df.apply(
            lambda row: str(int(row['member_number_raw'])) + str(int(row['loan_suffix'])).zfill(suffix_len)
            if pd.notna(row['member_number_raw']) and pd.notna(row['loan_suffix'])
            else None,
            axis=1,
        )
    else:
        df['member_number'] = df.get('member_number_raw')

    # Clean up loan_pool trailing spaces
    if 'loan_pool' in df.columns:
        df['loan_pool'] = df['loan_pool'].astype(str).str.strip()

    # Ensure numeric columns are numeric
    for col in ['current_balance', 'original_fico_score', 'current_fico_score',
                'original_loan_amount', 'interest_rate', 'days_delinquent',
                'total_available_credit', 'balance_other_lender', 'collateral_value',
                'total_loans', 'years_on_books', 'principal_paid', 'pct_principal_paid',
                'yr1_balance', 'yr2_balance', 'yr3_balance', 'total_1yr_payments',
                'collateral_eq_count', 'collateral_equity', 'unused_credit_count',
                'unused_credit', 'pct_unused_credit', 'amount_at_risk',
                'contingency_risk', 'loans_at_risk_count', 'contingency_risk_count',
                'total_risk_count']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    print(f"    Loaded {len(df)} loans from WARM 'All Loans' tab")
    return df


def _style_row(ws, row_num, cols, font=FNT_NORM, fill=None, fmt=None, align=None):
    """Apply styling to a range of cells in a row."""
    for c in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font = font
        cell.border = THIN
        if fill:
            cell.fill = fill
        if fmt:
            cell.number_format = fmt
        if align:
            cell.alignment = align


# ═══════════════════════════════════════════════════════════════════
# SHEET 1 — Key-Deteriorated Type
# ═══════════════════════════════════════════════════════════════════
def _sheet_key(wb, cu, snap):
    ws = wb.active
    ws.title = "Key-Deteriorated Type"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.column_dimensions['A'].width = 100

    r = 1
    ws.cell(row=r, column=1, value=cu).font = FNT_TITLE
    r += 1
    ws.cell(row=r, column=1,
            value="Risk Based Pricing Credit Grade Improved and Deteriorated Loans"
            ).font = FNT_SUB
    r += 1
    ws.cell(row=r, column=1, value="For Period Ending").font = FNT_BOLD
    r += 1
    d = datetime.strptime(snap, '%Y-%m-%d')
    ws.cell(row=r, column=1, value=d).font = FNT_BOLD
    ws.cell(row=r, column=1).number_format = 'MM/DD/YYYY'
    r += 1
    ws.cell(row=r, column=1, value="Report Description and Key").font = FNT_KEY
    r += 2

    # Explanation text
    paragraphs = [
        ("Explanation and Overview", True),
        ("In the presence of risk-based pricing, it is prudent for Credit Unions to examine "
         "the changing nature of their loans and borrowers.  By analyzing the factors that "
         "determine loan pricing and then identifying significant changes in those factors "
         "we can better estimate actual risk in the loan portfolio and take actions to manage "
         "that risk.", False),
        ("", False),
        ("Risk Factors Employed to Update Risk Status", True),
        ("Four specific risk factors are included in this updating risk assessment.  They are;", False),
        ("  Borrowers Original Credit Score", False),
        ("  Borrowers Current Credit score", False),
        ("  Loan Balance and/or Credit Limit", False),
        ("  Current Loan Balance to Collateral Value", False),
        ("", False),
        ("Methodology", True),
        ("A logic based method is employed to sift all loans in the portfolio using these "
         "factors with the outcome of identifying those loans with a risk that has significantly "
         "increased and require special attention by the Credit Union;", False),
        ("The sifting is performed using the four factors listed above.", False),
        ("An initial sorting process identifies those loans that require an analysis of risk "
         "as identified by LTV for secured loans, and borrower capacity for unsecured loans. ", False),
        ("Current risk is then examined by measuring the current LTV and borrower performance "
         "on secured loans and borrower capacity and performance on unsecured loans.", False),
        ("A final step on credit cards, lines of credit and HELOCs is to determine if changes "
         "to the limit and availability of credit are indicated.", False),
    ]
    for txt, is_header in paragraphs:
        cell = ws.cell(row=r, column=1, value=txt)
        cell.font = FNT_KEY if is_header else FNT_NORM
        cell.alignment = WRAP
        r += 1


# ═══════════════════════════════════════════════════════════════════
# SHEET 2 — Improved Loans Summary
# ═══════════════════════════════════════════════════════════════════
def _sheet_improved(wb, cu, snap, df, pool_order, hide_prefix="HIDE-"):
    ws = wb.create_sheet("Improved Loans Summary")
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = 'landscape'

    # Column widths: A=pool name, B=hidden tag, C-D=count+bal, E-F=equity, G-H=unused
    for col, w in {'A': 22, 'B': 24, 'C': 10, 'D': 16, 'E': 10, 'F': 18, 'G': 10, 'H': 16}.items():
        ws.column_dimensions[col].width = w

    ncols = 8
    imp = df[df['ncc_status'] == 'Improved']

    # Header rows
    r = 1
    ws.cell(row=r, column=1, value=cu).font = FNT_TITLE
    r += 1
    ws.cell(row=r, column=1, value="Summary of Credit Grade Improved Loans").font = FNT_SUB
    r += 1
    ws.cell(row=r, column=1, value="Hide Row").font = FNT_NORM
    ws.cell(row=r, column=2, value="Improved").font = FNT_NORM
    r += 1
    ws.cell(row=r, column=1, value=f"For Period Ending {snap_display(snap)}").font = FNT_BOLD
    r += 1

    # Sub-headers
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(row=r, column=3, value="Total Credit Grade Improved Loans").font = FNT_BOLD
    ws.cell(row=r, column=3).alignment = CENTER
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    ws.cell(row=r, column=5, value="Current Collateral Equity").font = FNT_BOLD
    ws.cell(row=r, column=5).alignment = CENTER
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
    ws.cell(row=r, column=7, value="Unused Credit").font = FNT_BOLD
    ws.cell(row=r, column=7).alignment = CENTER
    r += 1

    # Column labels
    labels = ['Loan Pools', '', 'Number', 'Balance', 'Number', 'Equity in Collateral',
              'Number', '$ of Unused Credit']
    for c, lbl in enumerate(labels, 1):
        cell = ws.cell(row=r, column=c, value=lbl)
        cell.font = FNT_HDR
        cell.fill = FILL_HDR
        cell.alignment = CENTER
        cell.border = THIN
    r += 1

    # Data rows by pool
    grand_num = 0
    grand_bal = 0.0
    grand_eq_num = 0
    grand_eq_val = 0.0
    grand_uc_num = 0
    grand_uc_val = 0.0

    for pool in pool_order:
        pool_imp = imp[imp['loan_pool'] == pool]
        num = len(pool_imp)
        bal = pool_imp['current_balance'].sum()
        # Collateral equity and unused credit from WARM data
        eq_num = int(pool_imp['collateral_eq_count'].sum()) if 'collateral_eq_count' in pool_imp.columns else 0
        eq_val = pool_imp['collateral_equity'].sum() if 'collateral_equity' in pool_imp.columns else 0.0
        uc_num = int(pool_imp['unused_credit_count'].sum()) if 'unused_credit_count' in pool_imp.columns else 0
        uc_val = pool_imp['unused_credit'].sum() if 'unused_credit' in pool_imp.columns else 0.0

        tag = f"{pool}-Improved"
        ws.cell(row=r, column=1, value=pool).font = FNT_NORM
        ws.cell(row=r, column=2, value=tag).font = FNT_NORM
        ws.cell(row=r, column=3, value=num).font = FNT_NORM
        ws.cell(row=r, column=3).number_format = NUM
        ws.cell(row=r, column=4, value=bal).font = FNT_NORM
        ws.cell(row=r, column=4).number_format = MONEY
        ws.cell(row=r, column=5, value=eq_num).font = FNT_NORM
        ws.cell(row=r, column=5).number_format = NUM
        ws.cell(row=r, column=6, value=eq_val).font = FNT_NORM
        ws.cell(row=r, column=6).number_format = MONEY
        ws.cell(row=r, column=7, value=uc_num).font = FNT_NORM
        ws.cell(row=r, column=7).number_format = NUM
        ws.cell(row=r, column=8, value=uc_val).font = FNT_NORM
        ws.cell(row=r, column=8).number_format = MONEY
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = THIN
            ws.cell(row=r, column=c).alignment = CENTER if c > 2 else LEFT

        # Alternate row fill
        if (pool_order.index(pool) % 2) == 1:
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = FILL_ALT

        grand_num += num
        grand_bal += bal
        grand_eq_num += eq_num
        grand_eq_val += eq_val
        grand_uc_num += uc_num
        grand_uc_val += uc_val
        r += 1

    # Grand Total
    ws.cell(row=r, column=1, value="Grand Total ").font = FNT_BOLD
    ws.cell(row=r, column=3, value=grand_num).font = FNT_BOLD
    ws.cell(row=r, column=3).number_format = NUM
    ws.cell(row=r, column=4, value=grand_bal).font = FNT_BOLD
    ws.cell(row=r, column=4).number_format = MONEY
    ws.cell(row=r, column=5, value=grand_eq_num).font = FNT_BOLD
    ws.cell(row=r, column=5).number_format = NUM
    ws.cell(row=r, column=6, value=grand_eq_val).font = FNT_BOLD
    ws.cell(row=r, column=6).number_format = MONEY
    ws.cell(row=r, column=7, value=grand_uc_num).font = FNT_BOLD
    ws.cell(row=r, column=7).number_format = NUM
    ws.cell(row=r, column=8, value=grand_uc_val).font = FNT_BOLD
    ws.cell(row=r, column=8).number_format = MONEY
    for c in range(1, ncols + 1):
        ws.cell(row=r, column=c).fill = FILL_TOT
        ws.cell(row=r, column=c).border = THIN
        ws.cell(row=r, column=c).alignment = CENTER if c > 2 else LEFT


# ═══════════════════════════════════════════════════════════════════
# SHEET 3 — Deteriorated Loans Summary
# ═══════════════════════════════════════════════════════════════════
def _sheet_deteriorated(wb, cu, snap, df, pool_order):
    ws = wb.create_sheet("Deteriorated Loans Summary")
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = 'landscape'

    # Columns: A=pool, B=tag, C-D=total, E-F=amount at risk, G-H=contingency, I-J=total at risk
    ncols = 10
    for col, w in {'A': 22, 'B': 26, 'C': 10, 'D': 16,
                   'E': 10, 'F': 16, 'G': 10, 'H': 16,
                   'I': 10, 'J': 16}.items():
        ws.column_dimensions[col].width = w

    det = df[df['ncc_status'] == 'Deteriorated']

    r = 1
    ws.cell(row=r, column=1, value=cu).font = FNT_TITLE
    r += 1
    ws.cell(row=r, column=1, value="Summary of Credit Grade Deteriorated Loans").font = FNT_SUB
    r += 1
    ws.cell(row=r, column=1, value="Hide Row").font = FNT_NORM
    ws.cell(row=r, column=2, value="Deteriorated").font = FNT_NORM
    r += 1
    ws.cell(row=r, column=1, value=f"For Period Ending {snap_display(snap)}").font = FNT_BOLD
    r += 1

    # Sub-headers
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(row=r, column=3, value="Total Credit Grade Impaired Loans").font = FNT_BOLD
    ws.cell(row=r, column=3).alignment = CENTER
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    ws.cell(row=r, column=5, value="Current Amount at Risk").font = FNT_BOLD
    ws.cell(row=r, column=5).alignment = CENTER
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
    ws.cell(row=r, column=7, value="Contingency Risk").font = FNT_BOLD
    ws.cell(row=r, column=7).alignment = CENTER
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=10)
    ws.cell(row=r, column=9, value="Total at Risk").font = FNT_BOLD
    ws.cell(row=r, column=9).alignment = CENTER
    r += 1

    labels = ['Loan Pools', '', 'Number', 'Balance', 'Number', 'Amount at Risk',
              'Number', 'Amount at Risk', 'Number', 'Amount at Risk']
    for c, lbl in enumerate(labels, 1):
        cell = ws.cell(row=r, column=c, value=lbl)
        cell.font = FNT_HDR
        cell.fill = FILL_HDR
        cell.alignment = CENTER
        cell.border = THIN
    r += 1

    # Accumulators for grand total
    g = [0] * 8  # num, bal, risk_num, risk_amt, cont_num, cont_amt, tot_num, tot_amt

    for pool in pool_order:
        pool_det = det[det['loan_pool'] == pool]
        num = len(pool_det)
        bal = pool_det['current_balance'].sum()
        # Amount at Risk from WARM data
        risk_num = int(pool_det['loans_at_risk_count'].sum()) if 'loans_at_risk_count' in pool_det.columns else num
        risk_amt = pool_det['amount_at_risk'].sum() if 'amount_at_risk' in pool_det.columns else bal
        # Contingency Risk from WARM data
        cont_num = int(pool_det['contingency_risk_count'].sum()) if 'contingency_risk_count' in pool_det.columns else 0
        cont_amt = pool_det['contingency_risk'].sum() if 'contingency_risk' in pool_det.columns else 0.0
        tot_num = risk_num + cont_num
        tot_amt = risk_amt + cont_amt

        tag = f"{pool}-Deteriorated"
        vals = [pool, tag, num, bal, risk_num, risk_amt, cont_num, cont_amt, tot_num, tot_amt]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = FNT_NORM
            cell.border = THIN
            cell.alignment = CENTER if c > 2 else LEFT
            if c in (4, 6, 8, 10):
                cell.number_format = MONEY
            elif c in (3, 5, 7, 9):
                cell.number_format = NUM

        if (pool_order.index(pool) % 2) == 1:
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = FILL_ALT

        g[0] += num; g[1] += bal
        g[2] += risk_num; g[3] += risk_amt
        g[4] += cont_num; g[5] += cont_amt
        g[6] += tot_num; g[7] += tot_amt
        r += 1

    # Grand Total
    gt_vals = ['Grand Total ', None, g[0], g[1], g[2], g[3], g[4], g[5], g[6], g[7]]
    for c, v in enumerate(gt_vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = FNT_BOLD
        cell.fill = FILL_TOT
        cell.border = THIN
        cell.alignment = CENTER if c > 2 else LEFT
        if c in (4, 6, 8, 10):
            cell.number_format = MONEY
        elif c in (3, 5, 7, 9):
            cell.number_format = NUM

    r += 1
    ws.cell(row=r, column=1, value="Comments:").font = FNT_BOLD


# ═══════════════════════════════════════════════════════════════════
# SHEET 4 — All Loans  (38 columns matching Franklin Trust layout)
# ═══════════════════════════════════════════════════════════════════
# Columns A–R (1–18): written as data values
# (header_label, df_field, number_format)
ALL_LOANS_DATA_COLS = [
    ('Member #',                                                      'member_number',         None),
    ('Loan Suffix',                                                   'loan_suffix',           None),
    ('Member #-Suffix',                                               'member_suffix',         None),
    ('Loan Type',                                                     'loan_type',             None),
    ('Loan Pool',                                                     'loan_pool',             None),
    ('Current Balance',                                               'current_balance',       MONEY),
    ('Original Credit Score',                                         'original_fico_score',   NUM),
    ('Original Credit Grade',                                         'original_grade',        None),
    ('Current Credit Score',                                          'current_fico_score',    NUM),
    ('Current Credit Grade',                                          'current_grade',         None),
    ('Open Date',                                                     'open_date',             'MM/DD/YYYY'),
    ('Interest Rate',                                                 'interest_rate',         '0.0000%'),
    ('Days Delinquent',                                               'days_delinquent',       NUM),
    ('Original Loan Amount',                                          'original_loan_amount',  MONEY),
    ('Net Credit Change Status',                                      'ncc_status',            None),
    ("Total Available Credit (Lines of Credit, Credit Cards and HELOC's)",
                                                                      'total_available_credit', MONEY),
    ("Balance at Other Lender (2nd Mortgage and HELOC's)",            'balance_other_lender',  MONEY),
    ('Colateral Value (for unsecured loans, leave blank)',            'collateral_value',      MONEY),
]

# Columns S–AL (19–38): written as Excel formulas
# (header_label, formula_template with {r} placeholder for row number, number_format)
ALL_LOANS_FORMULA_COLS = [
    ('Total Loans',                             '=F{r}+Q{r}',                                                                                              MONEY),
    ('LTV',                                     '=IF(R{r}=0,"No Value",S{r}/R{r})',                                                                        None),
    ('Years loan has been on books',            '=($F$1-K{r})/365',                                                                                        '0.00'),
    ('$ of Pricipal Paid',                      '=IF(N{r}-F{r}<0,0,N{r}-F{r})',                                                                            MONEY),
    ('% of Pricipal Paid',                      '=IFERROR(V{r}/N{r},0)',                                                                                   PCT),
    ('1 year of payments Balance',              '=IF(U{r}>2,0,IF(U{r}<1,0,F{r}))',                                                                         MONEY),
    ('2 years of payments Balance',             '=IF(U{r}>3,0,IF(U{r}<2,0,F{r}))',                                                                         MONEY),
    ('3+ years of payments Balance',            '=IF(U{r}<3,0,F{r})',                                                                                      MONEY),
    ('Total of 1+ years of payments made',      '=SUM(X{r}:Z{r})',                                                                                         MONEY),
    ('Collateral Equity Count',                 '=IF(AC{r}>0,1,0)',                                                                                         NUM),
    ('Collateral Equity',                       '=IF(S{r}>R{r},0,R{r}-S{r})',                                                                              MONEY),
    ('Count of Unused Credit',                  '=IF(AE{r}>0,1,0)',                                                                                        NUM),
    ('$ Unused Credit',                         '=IF(P{r}-S{r}<=0,0,P{r}-S{r})',                                                                           MONEY),
    ('% Unused Credit',                         '=IFERROR(AE{r}/P{r},0)',                                                                                  PCT),
    ('Amount at Risk',                          '=IF(T{r}<=1,0,IF(S{r}-R{r}>=F{r},F{r},S{r}-R{r}))',                                                      MONEY),
    ('Contingency Risk',                        '=IF(F{r}>P{r},0,IF(P{r}=0,0,IF(Q{r}+P{r}<R{r},0,IF(Q{r}+P{r}-R{r}-AG{r}>P{r},P{r}-AG{r},Q{r}+P{r}-R{r}-AG{r}))))', MONEY),
    ('Loans at Risk Count',                     '=IF(AG{r}=0,0,1)',                                                                                        NUM),
    ('Contingency Risk Count',                  '=IF(AH{r}=0,0,1)',                                                                                        NUM),
    ('Total Risk Count',                        '=IF(AI{r}=1,1,IF(AJ{r}=1,1,0))',                                                                         NUM),
    ('Net Credit Change/Pool',                  '=E{r}&"-"&O{r}',                                                                                          None),
]

_NCOLS = len(ALL_LOANS_DATA_COLS) + len(ALL_LOANS_FORMULA_COLS)  # 38

# Column widths for the 38 All Loans columns (A–AL)
_ALL_LOANS_WIDTHS = [
    12, 10, 14, 10, 18, 14, 16, 14, 14, 14,   # A-J
    12, 11, 13, 16, 20, 22, 22, 22, 13, 8,     # K-T
    14, 14, 14, 16, 16, 16, 18, 16, 16, 16,    # U-AD
    14, 14, 14, 14, 14, 16, 14, 26,            # AE-AL
]


def _sheet_all_loans(wb, cu, snap, df):
    ws = wb.create_sheet("All Loans")
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = 'landscape'

    ncols = _NCOLS  # 38

    # Column widths
    for i, w in enumerate(_ALL_LOANS_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Title row  — F1 holds the report date (referenced by formula =($F$1-K{r})/365)
    r = 1
    ws.cell(row=r, column=1, value="Credit Grade Improved Loans").font = FNT_TITLE
    d = datetime.strptime(snap, '%Y-%m-%d')
    ws.cell(row=r, column=6, value=d).font = FNT_BOLD
    ws.cell(row=r, column=6).number_format = 'MM/DD/YYYY'
    r += 1

    # Header row  — combine data + formula column headers
    all_headers = [(lbl, fmt) for lbl, _, fmt in ALL_LOANS_DATA_COLS] + \
                  [(lbl, fmt) for lbl, _, fmt in ALL_LOANS_FORMULA_COLS]
    for c, (label, _) in enumerate(all_headers, 1):
        cell = ws.cell(row=r, column=c, value=label)
        cell.font = FNT_HDR
        cell.fill = FILL_HDR
        cell.alignment = CENTER
        cell.border = THIN
    r += 1

    # Sort: Improved first, then Deteriorated, then Unchanged, then by pool, balance desc
    status_order = {'Improved': 0, 'Deteriorated': 1, 'Unchanged': 2}
    df_sorted = df.copy()
    df_sorted['_sort'] = df_sorted['ncc_status'].map(status_order)
    df_sorted = df_sorted.sort_values(
        ['_sort', 'loan_pool', 'current_balance'],
        ascending=[True, True, False],
    )

    n_data = len(ALL_LOANS_DATA_COLS)    # 18 (columns A–R)
    n_form = len(ALL_LOANS_FORMULA_COLS) # 20 (columns S–AL)

    for _, loan in df_sorted.iterrows():
        # ── Columns A–R: data values ──
        for c, (_, field, fmt) in enumerate(ALL_LOANS_DATA_COLS, 1):
            val = loan.get(field, '')
            if val == '' or (isinstance(val, float) and pd.isna(val)):
                val = None
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = FNT_NORM
            cell.border = THIN
            if fmt and val is not None:
                cell.number_format = fmt
            cell.alignment = CENTER if c > 1 else LEFT

        # ── Columns S–AL: Excel formulas ──
        for idx, (_, tmpl, fmt) in enumerate(ALL_LOANS_FORMULA_COLS):
            c = n_data + 1 + idx   # column number (19–38)
            formula = tmpl.format(r=r)
            cell = ws.cell(row=r, column=c, value=formula)
            cell.font = FNT_NORM
            cell.border = THIN
            if fmt:
                cell.number_format = fmt
            cell.alignment = CENTER

        # Color-code based on NCC status
        status = loan.get('ncc_status', '')
        if status == 'Improved':
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = FILL_IMP
        elif status == 'Deteriorated':
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = FILL_DET
        r += 1

    # Auto-filter
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{r - 1}"


# ═══════════════════════════════════════════════════════════════════
# Raw-extract enrichment (used when no WARM 'All Loans' tab exists)
# ═══════════════════════════════════════════════════════════════════
# Fields we want to pull from the configured loan_data_extracts (or the
# top-level column_mappings) when only the slim monthly_loan_data row is
# available. Each value is the column_mappings key whose mapped column
# holds the raw value for that field.
_EXTRACT_FIELDS = [
    'loan_type',                # from loan_pool_code
    'open_date',
    'interest_rate',
    'days_delinquent',
    'original_loan_amount',
    'total_available_credit',
]


def _resolve_extract_path(file_pattern, search_dirs):
    """Find the first file in any search_dir whose name matches file_pattern."""
    if not file_pattern:
        return None
    try:
        rx = re.compile(file_pattern)
    except re.error:
        return None
    for sdir in search_dirs:
        if not sdir or not os.path.isdir(sdir):
            continue
        for root, _dirs, files in os.walk(sdir):
            for f in files:
                if f.startswith('~$') or f.upper().startswith('DNU'):
                    continue
                if rx.search(f):
                    return os.path.join(root, f)
    return None


def _split_suffix_for_row(member_only, full_account, ma_cfg, raw_suffix):
    """Return the user-facing loan-suffix string for one row.

    For 'split' mode, raw_suffix already holds the suffix column value.
    For 'delimiter' mode, the suffix is whatever follows the delimiter,
    which here we reconstruct from full_account = member_only + suffix.
    For 'fixed_suffix' mode, the trailing N chars of full_account.
    """
    if raw_suffix is not None and str(raw_suffix).strip() not in ('', 'nan', 'None'):
        return str(raw_suffix).strip()
    mode = (ma_cfg or {}).get('mode') or 'fixed_suffix'
    full = str(full_account or '').strip()
    mem = str(member_only or '').strip()
    if not full or not mem:
        return ''
    if mode == 'delimiter' and full.startswith(mem):
        return full[len(mem):]
    if mode == 'fixed_suffix':
        try:
            n = int((ma_cfg or {}).get('suffix_length') or 0)
        except (TypeError, ValueError):
            n = 0
        if n > 0 and len(full) > n:
            return full[-n:]
    return ''


def _load_extract_enrichment(config, workspace_root):
    """Read configured loan extract files to populate the All Loans tab
    fields that don't live in monthly_loan_data (loan_type, open_date,
    interest_rate, days_delinquent, original_loan_amount,
    total_available_credit) plus the per-loan Member# / Suffix split.

    Returns a dict keyed by full_account_str (str) -> dict of fields.
    Falls back to {} when no extracts can be located.
    """
    enrich = {}
    data_dir = config.get('data_directory', '')
    if data_dir and not os.path.isabs(data_dir):
        data_dir = os.path.join(workspace_root, data_dir)
    search_dirs = [data_dir] if data_dir and os.path.isdir(data_dir) else []

    extracts = list(config.get('loan_data_extracts') or [])
    if not extracts:
        # Synthesize one entry from top-level config so this single code
        # path works for both styles.
        extracts = [{
            'label': 'top-level',
            'file_pattern': config.get('file_pattern'),
            'column_mappings': config.get('column_mappings') or {},
            'member_account': config.get('member_account'),
            'has_header': config.get('has_header', True),
            'header_row': config.get('header_row'),
        }]

    for ex in extracts:
        col_map = dict(ex.get('column_mappings') or {})
        if not col_map:
            continue
        path = _resolve_extract_path(ex.get('file_pattern'), search_dirs)
        if not path:
            print(f"    Extract '{ex.get('label')}' not found in {data_dir}; skipping enrichment.")
            continue
        has_header = ex.get('has_header', True)
        try:
            hr_cfg = int(ex.get('header_row') or 0)
        except (TypeError, ValueError):
            hr_cfg = 0
        pd_header = (hr_cfg - 1) if hr_cfg > 1 else 0
        ext_lc = os.path.splitext(path)[1].lower()
        try:
            if has_header:
                if ext_lc == '.csv':
                    df = pd.read_csv(path, header=pd_header)
                else:
                    df = pd.read_excel(path, header=pd_header)
                df.columns = [str(c).strip() for c in df.columns]
            else:
                if ext_lc == '.csv':
                    df = pd.read_csv(path, header=None)
                else:
                    df = pd.read_excel(path, header=None)
                col_map = _normalize_col_map_for_no_header(col_map)
        except Exception as e:
            print(f"    WARNING: could not read extract '{path}': {e}")
            continue

        # Build a per-row config snapshot derive_member_account expects.
        per_cfg = dict(config)
        per_cfg['column_mappings'] = col_map
        per_cfg['member_account'] = ex.get('member_account') or config.get('member_account')
        try:
            member_only, full_account = derive_member_account(df, per_cfg, has_header)
        except Exception as e:
            print(f"    WARNING: derive_member_account failed for '{path}': {e}")
            continue

        ma_cfg = per_cfg.get('member_account') or {}
        # Suffix column when in split mode
        raw_suffix_series = None
        if (ma_cfg.get('mode') == 'split') and col_map.get('loan_suffix') is not None:
            try:
                raw_suffix_series = (
                    df[col_map['loan_suffix']] if has_header
                    else df.iloc[:, col_map['loan_suffix']]
                )
            except Exception:
                raw_suffix_series = None

        # Map our impdet field names to the column_mappings keys.
        FIELD_TO_MAPKEY = {
            'loan_type':              'loan_pool_code',
            'open_date':              'open_date',
            'interest_rate':          'interest_rate',
            'days_delinquent':        'days_delinquent',
            'original_loan_amount':   'original_loan_amount',
            'total_available_credit': 'total_available_credit',
        }
        # Pre-resolve column series once per field
        field_series = {}
        for our_field, map_key in FIELD_TO_MAPKEY.items():
            ref = col_map.get(map_key)
            if ref is None or (isinstance(ref, str) and not ref):
                continue
            try:
                field_series[our_field] = (
                    df[ref] if has_header else df.iloc[:, ref]
                )
            except (KeyError, IndexError):
                continue

        n_added = 0
        for i in range(len(df)):
            full_str = str(full_account.iloc[i] or '').strip()
            if not full_str:
                continue
            mem_str = str(member_only.iloc[i] or '').strip()
            raw_suf = raw_suffix_series.iloc[i] if raw_suffix_series is not None else None
            suffix = _split_suffix_for_row(mem_str, full_str, ma_cfg, raw_suf)
            row_out = {
                'member_number_raw': mem_str,
                'loan_suffix': suffix,
                'member_suffix': full_str,
            }
            for our_field, ser in field_series.items():
                val = ser.iloc[i]
                if pd.isna(val):
                    continue
                row_out[our_field] = val
            # Prefer the richest record; don't let an extract that lacks
            # most columns overwrite a fuller one.
            prev = enrich.get(full_str)
            if prev is None or len([k for k in row_out if row_out.get(k) not in (None, '')]) > \
                              len([k for k in prev if prev.get(k) not in (None, '')]):
                enrich[full_str] = row_out
            n_added += 1
        print(f"    Loaded {n_added} row(s) from extract '{ex.get('label')}'")

    print(f"  Enrichment dictionary: {len(enrich)} unique loan(s) from extracts")
    return enrich


# ═══════════════════════════════════════════════════════════════════
# Main orchestration
# ═══════════════════════════════════════════════════════════════════
def generate_report(client, snap=None):
    config = load_config(client)
    cu = config['credit_union']
    grades = config['credit_grades']
    no_score = config.get('no_score_label', 'Not Reported')
    n_top = config.get('top_grades_double_drop', 3)

    # Resolve snapshot date
    if not snap:
        snap = latest_date(cu)
    if not snap:
        print(f"ERROR: No loan data found for '{cu}'")
        return

    print(f"Generating Improved/Deteriorated report for {cu}, period ending {snap}")

    # ── Try to load rich data from WARM "All Loans" tab ──────────
    warm_path = find_warm_file(config, snap)
    warm_df = None
    if warm_path:
        print(f"  Found WARM file: {os.path.basename(warm_path)}")
        warm_df = load_warm_all_loans(warm_path, config)

    if warm_df is not None and not warm_df.empty:
        # ── Use WARM data as the primary data source ─────────────
        df = warm_df.copy()

        # Apply our own credit grading from the WARM scores
        grade_labels = [g['label'] for g in grades] + [no_score]
        df['original_grade'] = df['original_fico_score'].apply(
            lambda s: assign_credit_grade(int(s) if pd.notna(s) else 0, grades, no_score)
        )
        df['current_grade'] = df['current_fico_score'].apply(
            lambda s: assign_credit_grade(int(s) if pd.notna(s) else 0, grades, no_score)
        )

        # Per-loan NCC status using our top_grades_double_drop logic
        not_risk_rated = set(config.get('not_risk_rated', []))
        df['ncc_status'] = df.apply(
            lambda row: (
                "Unchanged" if row['loan_pool'] in not_risk_rated
                else loan_ncc_status(
                    row['original_grade'], row['current_grade'], grade_labels, n_top, no_score
                )
            ), axis=1,
        )

        # Update NCC-dependent derived columns with our classification
        df['ncc_pool'] = df['loan_pool'] + '-' + df['ncc_status']

        # Use Member # (raw) for the All Loans "Member #" column
        if 'member_number_raw' in df.columns:
            df['member_number'] = df['member_number_raw']

        print(f"  Using WARM data — all 38 columns populated")

    else:
        # ── Fallback: DB-only approach (limited columns) ─────────
        print("  No WARM 'All Loans' data available — using DB only")
        df = load_loans(cu, snap, config)
        if df.empty:
            print(f"ERROR: No loans found for {cu} on {snap}")
            return
        print(f"  Loaded {len(df)} loans from DB")

        # Apply credit grades
        grade_labels = [g['label'] for g in grades] + [no_score]
        df['original_grade'] = df['original_fico_score'].apply(
            lambda s: assign_credit_grade(s, grades, no_score)
        )
        df['current_grade'] = df['current_fico_score'].apply(
            lambda s: assign_credit_grade(s, grades, no_score)
        )

        # Per-loan NCC status
        not_risk_rated = set(config.get('not_risk_rated', []))
        df['ncc_status'] = df.apply(
            lambda row: (
                "Unchanged" if row['loan_pool'] in not_risk_rated
                else loan_ncc_status(
                    row['original_grade'], row['current_grade'], grade_labels, n_top, no_score
                )
            ), axis=1,
        )

        # Derived columns — blanks for columns not in DB
        df['ncc_pool'] = df['loan_pool'] + '-' + df['ncc_status']
        df['amount_at_risk'] = df.apply(
            lambda row: row['current_balance'] if row['ncc_status'] == 'Deteriorated' else 0.0,
            axis=1,
        )
        df['contingency_risk'] = 0.0

        # Try to enrich from the configured loan_data_extracts so the
        # All Loans tab has Member#/Suffix split + Loan Type / Open Date /
        # Interest Rate / Days Delinquent / Original Loan Amount / Credit
        # Limit even without a WARM workbook on disk.
        enrich = _load_extract_enrichment(config, BASE)

        def _enr_get(acct, field, default=None):
            row = enrich.get(str(acct).strip()) if enrich else None
            if not row:
                return default
            val = row.get(field, default)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return default
            return val

        # Member#/Suffix split — fall back to DB member_number when no
        # extract row matches (preserves existing behavior for CUs
        # without loan_data_extracts).
        df['member_number_raw'] = df['member_number'].apply(
            lambda a: _enr_get(a, 'member_number_raw', str(a))
        )
        df['loan_suffix'] = df['member_number'].apply(
            lambda a: _enr_get(a, 'loan_suffix', '')
        )
        df['member_suffix'] = df['member_number'].apply(
            lambda a: _enr_get(a, 'member_suffix', str(a))
        )
        # Re-point the "Member #" column written to the All Loans tab so
        # it shows the member-only string (matches WARM-mode output).
        df['member_number'] = df['member_number_raw']

        df['loan_type'] = df['member_suffix'].apply(lambda a: _enr_get(a, 'loan_type'))
        df['open_date'] = df['member_suffix'].apply(lambda a: _enr_get(a, 'open_date'))
        df['interest_rate'] = df['member_suffix'].apply(lambda a: _enr_get(a, 'interest_rate'))
        df['days_delinquent'] = df['member_suffix'].apply(lambda a: _enr_get(a, 'days_delinquent'))
        df['original_loan_amount'] = df['member_suffix'].apply(
            lambda a: _enr_get(a, 'original_loan_amount')
        )
        df['total_available_credit'] = df['member_suffix'].apply(
            lambda a: _enr_get(a, 'total_available_credit')
        )

        # Interest rate convention: WARM stores as a percent (e.g. 6.625)
        # but the All Loans tab formats column L as "0.0000%" which expects
        # a decimal fraction. Convert >1 values down to fraction.
        df['interest_rate'] = df['interest_rate'].apply(
            lambda v: (v / 100.0) if isinstance(v, (int, float)) and v is not None
                      and not (isinstance(v, float) and pd.isna(v)) and v > 1
                      else v
        )

        df['balance_other_lender'] = None
        df['collateral_value'] = None
        df['total_loans'] = df['current_balance']
        df['ltv'] = None
        df['years_on_books'] = None
        df['principal_paid'] = None
        df['pct_principal_paid'] = None
        df['yr1_balance'] = None
        df['yr2_balance'] = None
        df['yr3_balance'] = None
        df['total_1yr_payments'] = None
        df['collateral_eq_count'] = 0
        df['collateral_equity'] = 0.0
        df['unused_credit_count'] = 0
        df['unused_credit'] = 0.0
        df['pct_unused_credit'] = None
        df['loans_at_risk_count'] = 1
        df['contingency_risk_count'] = 0
        df['total_risk_count'] = 1

    # Pool order from config
    pool_order = config.get('pool_order', sorted(df['loan_pool'].unique()))

    # Status summary
    for status in ['Improved', 'Deteriorated', 'Unchanged']:
        sub = df[df['ncc_status'] == status]
        print(f"  {status:14s}: {len(sub):6,d} loans, ${sub['current_balance'].sum():>14,.2f}")

    # Build workbook
    wb = Workbook()
    _sheet_key(wb, cu, snap)
    _sheet_improved(wb, cu, snap, df, pool_order)
    _sheet_deteriorated(wb, cu, snap, df, pool_order)
    _sheet_all_loans(wb, cu, snap, df)

    # "All Loans" tab is intentionally left unlocked so users can sort/filter.
    # (Previously protected with a password; removed per user request.)

    # Save
    os.makedirs(RPT_DIR, exist_ok=True)
    snap_prefix = snap[:7]   # "2025-12"
    safe_cu = cu.replace(' ', '_')
    fname = f"{snap_prefix} Improved Deteriorated Loans - {safe_cu}.xlsx"
    out_path = os.path.join(RPT_DIR, fname)
    wb.save(out_path)
    print(f"  Saved: {out_path}")
    return out_path


# ── CLI ────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Generate Improved/Deteriorated Loans Report")
    parser.add_argument('--client', required=True, help="Client config name (e.g. sample, franklin)")
    parser.add_argument('--date', default=None, help="Snapshot date YYYY-MM-DD (default: latest)")
    args = parser.parse_args()
    generate_report(args.client, args.date)


if __name__ == '__main__':
    main()
