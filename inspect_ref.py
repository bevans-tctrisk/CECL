"""Inspect the reference Excel report in detail."""
import os, warnings, sys, json
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

REF = os.path.join(os.environ['TEMP'], 'ref_report.xlsx')
import time
print(f"Loading {REF}...", flush=True)
t0 = time.time()
wb = load_workbook(REF, data_only=True)
print(f"Loaded in {time.time()-t0:.0f}s", flush=True)

# Report sheets are 1-48 (Cover through Grade Ranges & Loan Codes)
REPORT_SHEETS = wb.sheetnames[:48]

def dump_sheet(name, max_rows=None):
    ws = wb[name]
    print(f"\n{'='*80}")
    print(f"SHEET: {name}")
    print(f"  Dims: {ws.max_row} rows x {ws.max_column} cols")
    print(f"  Orientation: {ws.page_setup.orientation}")
    print(f"  Charts: {len(ws._charts)}")
    print(f"  Merged cells: {len(ws.merged_cells.ranges)}")

    # Column widths
    cw = {}
    for c in range(1, min(ws.max_column + 1, 30)):
        lt = get_column_letter(c)
        w = ws.column_dimensions[lt].width
        if w:
            cw[lt] = round(w, 1)
    print(f"  Col widths: {cw}")

    # Row heights (non-default)
    rh = {}
    mr = max_rows or ws.max_row
    for r in range(1, min(mr + 1, 100)):
        h = ws.row_dimensions[r].height
        if h and h != 15:
            rh[r] = h
    if rh:
        print(f"  Row heights: {rh}")

    # Merged ranges
    if ws.merged_cells.ranges:
        merges = sorted([str(m) for m in ws.merged_cells.ranges])
        print(f"  Merges: {merges[:20]}")
        if len(merges) > 20:
            print(f"    ... and {len(merges)-20} more")

    # Cell content with formatting (first max_rows rows)
    limit = max_rows or 90
    print(f"  --- Content (first {limit} rows) ---")
    for r in range(1, min(ws.max_row + 1, limit + 1)):
        for c in range(1, min(ws.max_column + 1, 25)):
            cell = ws.cell(r, c)
            if cell.value is not None:
                font = cell.font
                fill = cell.fill
                align = cell.alignment
                border = cell.border

                parts = [f"R{r}C{c}"]
                val = repr(cell.value)
                if len(val) > 60:
                    val = val[:60] + "..."
                parts.append(val)

                if font:
                    fb = "B" if font.bold else ""
                    fi = "I" if font.italic else ""
                    fc = ""
                    if font.color and font.color.rgb and font.color.rgb != '00000000':
                        fc = font.color.rgb
                    parts.append(f"f={font.name}/{font.size}{fb}{fi}")
                    if fc:
                        parts.append(f"fc={fc}")

                if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb not in ('00000000', '0'):
                    parts.append(f"bg={fill.fgColor.rgb}")

                if cell.number_format and cell.number_format != 'General':
                    parts.append(f"fmt={cell.number_format}")

                if align:
                    ha = align.horizontal or ""
                    va = align.vertical or ""
                    wrap = "W" if align.wrap_text else ""
                    if ha or va or wrap:
                        parts.append(f"align={ha}/{va}/{wrap}")

                print("    " + "  |  ".join(parts))

    # Chart details
    for ci, chart in enumerate(ws._charts):
        print(f"  Chart {ci}: type={type(chart).__name__} title={chart.title} "
              f"w={chart.width} h={chart.height}")


if len(sys.argv) > 1:
    sheets = sys.argv[1:]
else:
    # Dump key report sheets
    sheets = [
        'Cover',
        'Introduction',
        'Executive Summary',
        'Executive Summary (2)',
        'Executive Summary (3)',
        'Risk Change by Credit Score',
    ]

for s in sheets:
    if s in wb.sheetnames:
        dump_sheet(s)
    else:
        print(f"Sheet '{s}' not found")
