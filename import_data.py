"""
Config-driven CECL data import pipeline.
Reads client YAML configs and imports loan data from any CU's file format.

Usage:
    python import_data.py --client ontario
    python import_data.py --client ontario --file "AIRESLOANS 2025-12.xlsx"
    python import_data.py --all
    python import_data.py --list
"""
import os
import re
import sys
import shutil
import argparse
import calendar
from datetime import date

import pandas as pd
import yaml
from sqlalchemy import create_engine, text
from dotenv import load_dotenv
from cecl_audit_log import get_audit_logger, log_data_import, log_session_start, log_session_end
from cecl_credentials import get_database_url

load_dotenv()

# Honour CECL_WORKSPACE_ROOT so the data root can be decoupled from the
# code location; falls back to historical layout when the env var is unset.
BASE_FOLDER = os.environ.get('CECL_WORKSPACE_ROOT') or os.path.dirname(os.path.abspath(__file__))
CONFIG_FOLDER = os.path.join(BASE_FOLDER, 'client_configs')
UPLOAD_FOLDER = os.path.join(BASE_FOLDER, 'Raw_Uploads')
ARCHIVE_FOLDER = os.path.join(BASE_FOLDER, 'Archive')

db_url = get_database_url()
engine = create_engine(db_url)


def resolve_path(path_value, base=BASE_FOLDER):
    """Resolve configured paths: keep absolute paths, join relative paths to base."""
    if not path_value:
        return ''
    return path_value if os.path.isabs(path_value) else os.path.join(base, path_value)


def load_client_config(client_name):
    """Load a client YAML config file."""
    config_path = os.path.join(CONFIG_FOLDER, f'{client_name}.yaml')
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Config not found: {config_path}")
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def list_clients():
    """List available client configs (excluding template)."""
    clients = []
    for f in os.listdir(CONFIG_FOLDER):
        if f.endswith('.yaml') and not f.startswith('_'):
            clients.append(os.path.splitext(f)[0])
    return sorted(clients)


def extract_snapshot_date(source_text, config):
    """Extract snapshot date using the client's date regex pattern."""
    date_pattern = config['date_pattern']
    date_fmt = config.get('date_format', 'YYYY-MM')
    text = str(source_text)
    match = re.search(date_pattern, text)

    def _from_month_name(year_token, mon_token):
        month = _MONTH_MAP.get(str(mon_token)[:3].lower())
        if not month:
            return None
        try:
            year = int(year_token)
        except (TypeError, ValueError):
            return None
        # Normalize 2-digit years (e.g. "25" -> 2025) so month-name
        # filenames like "Aires December 25.v2.xlsx" resolve correctly.
        # Without this a token of "25" yields year 25 AD and a bogus date.
        if year < 100:
            year += 2000
        last_day = calendar.monthrange(year, month)[1]
        return date(year, month, last_day).isoformat()

    if not match:
        # Fallback: handle month-name formats like "Mar_2026", "March-2026",
        # or "2026_Mar" that the configured numeric regex won't catch.
        mon_first = re.search(
            r"(?i)(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*"
            r"[-_ \.]+(20\d{2})",
            text,
        )
        if mon_first:
            return _from_month_name(mon_first.group(2), mon_first.group(1))
        year_first = re.search(
            r"(?i)(20\d{2})[-_ \.]+"
            r"(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*",
            text,
        )
        if year_first:
            return _from_month_name(year_first.group(1), year_first.group(2))
        # Month-name + 2-digit year (e.g. "December 25", "Mar 26" from
        # "Aires December 25.v2.xlsx"). Only attempted when NO 4-digit
        # year appears anywhere in the name — that guard avoids reading
        # the day out of a "Month DD YYYY" filename as a 2-digit year.
        # ``_from_month_name`` normalizes the 2-digit year to 20xx.
        if not re.search(r"20\d{2}", text):
            mon_yy = re.search(
                r"(?i)(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)"
                r"[a-z]*[-_ \.]+(\d{2})(?!\d)",
                text,
            )
            if mon_yy:
                return _from_month_name(mon_yy.group(2), mon_yy.group(1))
        # Fallback: "DDMonYY" / "DDMonYYYY" smushed-together filenames like
        # "30JUN25" or "01Jan2026" (no separators between day/month/year).
        # The day is captured but ignored — we still snap to month-end below.
        ddmonyy = re.search(
            r"(?i)(?<!\d)(\d{1,2})(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*"
            r"[-_ \.]?(\d{2}|\d{4})(?!\d)",
            text,
        )
        if ddmonyy:
            yr_token = ddmonyy.group(3)
            if len(yr_token) == 2:
                yr_token = "20" + yr_token
            return _from_month_name(yr_token, ddmonyy.group(2))
        # Final fallback: try common numeric layouts (YYYYMMDD, MMDDYYYY,
        # YYYY-MM, MM-YYYY...) so we still find a date when the configured
        # date_pattern doesn't fit this particular filename.
        return _try_common_date_layouts(text)

    # If either captured group is a month name, route through the name parser
    # rather than treating it as a number.
    g1 = match.group(1) if match.lastindex and match.lastindex >= 1 else None
    g2 = match.group(2) if match.lastindex and match.lastindex >= 2 else None
    if g1 and g1[:3].lower() in _MONTH_MAP:
        return _from_month_name(g2, g1)
    if g2 and g2[:3].lower() in _MONTH_MAP:
        return _from_month_name(g1, g2)

    try:
        if date_fmt == 'MMDDYY':
            month, day, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
            year += 2000 if year < 100 else 0
            return date(year, month, day).isoformat()
        elif date_fmt == 'MMDDYYYY':
            month, day, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
            return date(year, month, day).isoformat()
        elif date_fmt == 'MMYY':
            month, year = int(match.group(1)), int(match.group(2))
            year += 2000 if year < 100 else 0
            last_day = calendar.monthrange(year, month)[1]
            return date(year, month, last_day).isoformat()
        elif date_fmt == 'MMYYYY':
            month, year = int(match.group(1)), int(match.group(2))
            last_day = calendar.monthrange(year, month)[1]
            return date(year, month, last_day).isoformat()
        elif date_fmt == 'YYYYMMDD':
            year, month, day = int(match.group(1)), int(match.group(2)), int(match.group(3))
            return date(year, month, day).isoformat()
        elif date_fmt == 'YYYYQ':
            year, quarter = int(match.group(1)), int(match.group(2))
            month = quarter * 3
            last_day = calendar.monthrange(year, month)[1]
            return date(year, month, last_day).isoformat()
        elif date_fmt == 'QYYYY':
            quarter, year = int(match.group(1)), int(match.group(2))
            month = quarter * 3
            last_day = calendar.monthrange(year, month)[1]
            return date(year, month, last_day).isoformat()
        else:
            year, month = int(match.group(1)), int(match.group(2))
            last_day = calendar.monthrange(year, month)[1]
            return date(year, month, last_day).isoformat()
    except (ValueError, calendar.IllegalMonthError) as exc:
        # The captured groups don't form a real calendar date (most often the
        # date_pattern is wrong for this filename — e.g. ``(\d{4})(\d{2})``
        # against ``03312026`` gives month=20). Try a small set of common
        # filename date layouts on the same text before giving up, so a
        # mis-configured client config still produces a date when one is
        # plainly present in the filename.
        fallback = _try_common_date_layouts(text)
        if fallback is not None:
            print(f"  WARN: configured date_pattern={date_pattern!r} "
                  f"format={date_fmt} failed on '{source_text}' ({exc}); "
                  f"recovered via fallback -> {fallback}")
            return fallback
        print(f"  WARN: could not parse date from '{source_text}' "
              f"(format={date_fmt}): {exc}")
        return None


# Common filename date layouts, tried in order, used both as a regex *and*
# to interpret the captured groups. Year groups are anchored to ``20\d{2}``
# so the heuristic doesn't grab e.g. "0331" out of "03312026" as a year.
_FALLBACK_DATE_LAYOUTS: list[tuple[str, str]] = [
    # (regex, kind) — kind tells the loop how to read the groups.
    (r"(20\d{2})[-_./ ](\d{1,2})[-_./ ](\d{1,2})(?!\d)", "YMD"),
    (r"(20\d{2})(\d{2})(\d{2})",               "YMD"),
    (r"(?<!\d)(\d{1,2})[-_./ ](\d{1,2})[-_./ ](20\d{2})", "MDY"),
    (r"(\d{2})(\d{2})(20\d{2})",               "MDY"),
    # Month-Day-2digit-year (e.g. "6-30-26", "12-31-25"). Placed after the
    # 4-digit-year forms so a full year always wins. YY -> 2000+YY, bounded
    # to a plausible window so it doesn't grab unrelated number triples.
    # Separators exclude spaces so a trailing code digit + spaced date
    # (e.g. "CECLCC1 6-30-26") isn't misread as "1 6-30".
    (r"(?<!\d)(\d{1,2})[-_./](\d{1,2})[-_./](\d{2})(?!\d)", "MDYY"),
    (r"(20\d{2})[-_./ ](\d{1,2})(?!\d)",       "YM"),
    (r"(20\d{2})(\d{2})(?!\d)",                "YM"),
    (r"(?<!\d)(\d{1,2})[-_./ ](20\d{2})",      "MY"),
    # Quarter-end forms: "2025Q4", "2026-Q1", "2026_Q1".
    # Q1->Mar 31, Q2->Jun 30, Q3->Sep 30, Q4->Dec 31.
    (r"(20\d{2})[-_ ]?[Qq]([1-4])(?!\d)",       "YQ"),
    # "Q4 2025" / "Q1-2026" / "Q1_2026".
    (r"(?<![A-Za-z])[Qq]([1-4])[-_ ]?(20\d{2})(?!\d)", "QY"),
    # Two-digit-year fallback: anchored to 19-30 to keep ambiguity low.
    # Used for AIRES-style filenames like "25-12 AIRES LOANS v2.xlsx"
    # where the wizard's auto-derived date_pattern (which expects a
    # 4-digit year) doesn't match. YY → 2000+YY.
    (r"(?<!\d)(\d{2})[-_./ ](\d{2})(?!\d)",    "YYM"),
]


def _try_common_date_layouts(text: str) -> str | None:
    """Best-effort filename date recovery; returns ISO date or ``None``."""
    for rx, kind in _FALLBACK_DATE_LAYOUTS:
        m = re.search(rx, text)
        if not m:
            continue
        try:
            if kind == "YMD":
                y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
                return date(y, mo, d).isoformat()
            if kind == "MDY":
                mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
                return date(y, mo, d).isoformat()
            if kind == "MDYY":
                mo, d, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
                # Bound the 2-digit year to a plausible window and validate
                # month/day so a random "a-b-cc" triple isn't misread.
                if not (1 <= mo <= 12) or not (1 <= d <= 31) \
                        or yy < 19 or yy > 40:
                    continue
                return date(2000 + yy, mo, d).isoformat()
            if kind == "YM":
                y, mo = int(m.group(1)), int(m.group(2))
                last = calendar.monthrange(y, mo)[1]
                return date(y, mo, last).isoformat()
            if kind == "MY":
                mo, y = int(m.group(1)), int(m.group(2))
                last = calendar.monthrange(y, mo)[1]
                return date(y, mo, last).isoformat()
            if kind == "YYM":
                yy, mo = int(m.group(1)), int(m.group(2))
                # Restrict YY to a plausible recent-CU window to limit
                # false matches (would also reject e.g. "12-25" tail of a
                # phone number). Months still validated below.
                if yy < 19 or yy > 30 or mo < 1 or mo > 12:
                    continue
                y = 2000 + yy
                last = calendar.monthrange(y, mo)[1]
                return date(y, mo, last).isoformat()
            if kind == "YQ":
                y, q = int(m.group(1)), int(m.group(2))
                mo = q * 3
                last = calendar.monthrange(y, mo)[1]
                return date(y, mo, last).isoformat()
            if kind == "QY":
                q, y = int(m.group(1)), int(m.group(2))
                mo = q * 3
                last = calendar.monthrange(y, mo)[1]
                return date(y, mo, last).isoformat()
        except (ValueError, calendar.IllegalMonthError):
            continue
    return None


def _compile_file_patterns(value, *, label=""):
    """Compile a YAML ``file_pattern`` value into a list of regex objects.

    Accepts either a single regex string (legacy single-pattern form) or a
    list of regex strings (new multi-pattern form, first-match-wins). Empty
    or ``None`` returns an empty list. Invalid regexes are warned and
    skipped so a typo in one pattern doesn't blow up the whole import.

    The multi-pattern form lets a single ``loan_data_extracts`` entry match
    files whose names changed across years (e.g. anonymized old-style
    ``Aires Loan <date> Credit Union B.xlsx`` alongside modern
    ``<acct> - Nova Aires Loan <date>.xlsx``). Same shape semantics as a
    single regex — order in the list is preserved for diagnostics.
    """
    if value is None or value == "":
        return []
    if isinstance(value, str):
        items = [value]
    elif isinstance(value, (list, tuple)):
        items = [v for v in value if isinstance(v, str) and v]
    else:
        return []
    out = []
    for s in items:
        try:
            out.append(re.compile(s, re.IGNORECASE))
        except re.error as exc:
            tag = f" ({label})" if label else ""
            print(f"  WARNING: invalid file_pattern {s!r}{tag}: {exc}; skipping.")
    return out


def _patterns_match_any(patterns, *texts):
    """Return True if any compiled pattern matches any of the texts."""
    for p in patterns:
        for t in texts:
            if t and p.search(t):
                return True
    return False


def clean_balance(series, balance_format):
    """Clean a balance column based on config rules."""
    s = series.astype(str)
    for char in balance_format.get('remove_chars', []):
        s = s.str.replace(char, '', regex=False)
    if balance_format.get('accounting_negatives', False):
        s = s.str.replace('(', '-', regex=False).str.replace(')', '', regex=False)
    return pd.to_numeric(s, errors='coerce')


# pandas' default ``na_values`` treats the bare string "NA" as NaN. Several
# credit unions use "NA" as a real loan-type code (New Auto), so reading with
# the defaults silently turns every New Auto loan into a blank code. This list
# is pandas' default NaN-token set with "NA" removed, so ONLY the literal
# "NA" changes behaviour (it is preserved as a string); every other token
# ('', 'N/A', 'NaN', 'NULL', ...) keeps its normal NaN handling.
_NA_VALUES_KEEP_LITERAL = [
    '', '#N/A', '#N/A N/A', '#NA', '-1.#IND', '-1.#QNAN', '-NaN', '-nan',
    '1.#IND', '1.#QNAN', '<NA>', 'N/A', 'NULL', 'NaN', 'None', 'n/a', 'nan',
    'null',
]


def map_pool_codes(series, config):
    """Map raw loan pool codes to pool names using config."""
    split_char = config.get('pool_code_split')
    if split_char:
        raw = series.astype(str).str.split(split_char).str[0].str.strip()
    else:
        raw = series.astype(str).str.strip()
    # Normalize float strings like "85.0" to "85" for numeric codes. Guard with
    # isinstance(str): a string-dtype column can still carry float NaN entries
    # (blank cells), and float NaN has no ``.replace`` — leave those untouched
    # so they fall through to ``default_pool`` below instead of crashing.
    raw = raw.apply(
        lambda x: str(int(float(x)))
        if isinstance(x, str) and x.replace('.', '', 1).isdigit()
        else x
    )
    pool_map = {str(k): v for k, v in config['pool_map'].items()}
    default = config.get('default_pool', 'Other/Uncategorized')
    # Normalize the legacy 'Ignore' sentinel to the canonical 'Exclude' so
    # downstream report filters (generate_report._apply_excluded_pools) drop
    # these rows uniformly. Future re-imports will write 'Exclude' to the DB
    # and old 'Ignore'-tagged rows are also filtered at report time.
    pool_map = {k: ('Exclude' if v == 'Ignore' else v) for k, v in pool_map.items()}
    if default == 'Ignore':
        default = 'Exclude'
    return raw.map(pool_map).fillna(default)


# Month name -> number for sorting credit pull sheet names
_MONTH_MAP = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
}


def _sort_credit_pull_sheets(sheet_names):
    """Sort credit pull sheet names most-recent-first.

    Handles names like 'Dec-25 Credit Pull', 'Dec-24 Credit Pull', etc.
    """
    def _sort_key(name):
        m = re.search(r'([A-Za-z]{3})-(\d{2})', name)
        if m:
            month = _MONTH_MAP.get(m.group(1).lower(), 0)
            year = int(m.group(2))
            return (-year, -month)
        return (0, 0)
    return sorted(sheet_names, key=_sort_key)


def load_credit_pull_scores(config):
    """
    Load current credit scores from a credit pull source.
    Priority:
      1. Standalone credit pull file (latest bureau scores)
      2. Credit pull tabs in existing CECL/WARM report (older bureau scores)
      3. WARM "All Loans" current score column (per-loan carry-forward scores)
    Returns a tuple of:
      - member_scores: dict {member_number (int): fico_score (int)} — member-level
      - acct_scores: dict {account_number (str): fico_score (int)} — per-loan level
      - pull_as_of: pd.Timestamp or None — best-effort date of the latest credit
        pull source (config override > standalone file mtime > fallback report mtime)
    """
    cp_config = config.get('credit_pull', {})
    if not cp_config:
        return {}, {}, None

    scores = {}
    warm_acct_scores = {}  # per-loan scores from WARM All Loans
    pull_as_of = None      # tracked across the source files we read

    # Optional explicit override from the wizard.
    pull_override = (cp_config.get('pull_as_of_date') or '').strip()
    if pull_override:
        try:
            pull_as_of = pd.Timestamp(pull_override)
        except (ValueError, TypeError):
            print(f"    WARNING: Could not parse pull_as_of_date '{pull_override}'")
            pull_as_of = None

    # 1. Try standalone credit pull file
    file_pattern = cp_config.get('file_pattern')
    if file_pattern:
        source_folder = resolve_path(cp_config.get('source_folder', ''))
        if os.path.isdir(source_folder):
            pattern_re = re.compile(file_pattern, re.IGNORECASE)
            matching_files = []
            for root, _dirs, files in os.walk(source_folder):
                for fname in files:
                    if fname.startswith('~$'):
                        continue
                    if pattern_re.search(fname):
                        matching_files.append(os.path.join(root, fname))

            for fpath in sorted(matching_files, key=os.path.getmtime, reverse=True):
                    fname = os.path.basename(fpath)
                    member_col = cp_config.get('member_column', 'Member Number')
                    score_col = cp_config.get('score_column', 'FICO')
                    df = pd.read_excel(fpath)
                    # Normalise headers to match the wizard-saved column
                    # name shape (collapse wrap-text CR/LF inside header
                    # cells to a single space; rewrite blank/'nan'/
                    # 'Unnamed: N' to col_<LETTER>). Same shape that
                    # ``cecl_ui/services/sample_parser._clean_header``
                    # produces, so a credit-pull file with a wrap-text
                    # ``Credit\nScore`` header still resolves against the
                    # YAML's ``score_column: Credit Score`` mapping.
                    _ucp_rx = re.compile(r"^unnamed:\s*\d+(?:_level_\d+)*$",
                                          re.IGNORECASE)
                    _ncols = []
                    for _i, _c in enumerate(df.columns):
                        _s = re.sub(r"\s+", " ", str(_c)).strip() if _c is not None else ""
                        _low = _s.lower()
                        if (not _s) or _low == "nan" or _ucp_rx.match(_s):
                            _ncols.append(f"col_{_excel_idx_to_letter(_i)}")
                        else:
                            _ncols.append(_s)
                    df.columns = _ncols
                    if member_col in df.columns and score_col in df.columns:
                        # Strip the same suffix the Aires file uses so the
                        # member-number key matches what import_file builds.
                        suffix_length = config.get('account_suffix_length', 0) or 0
                        for _, row in df.iterrows():
                            m = row[member_col]
                            s = row[score_col]
                            if pd.notna(m) and pd.notna(s):
                                try:
                                    score_val = int(float(s))
                                    if score_val <= 0:
                                        continue  # skip zero = no data
                                    m_str = str(m).strip()
                                    if m_str.endswith('.0'):
                                        m_str = m_str[:-2]
                                    # Always store the raw value as an int
                                    # key (covers credit-pull files that
                                    # carry just the bare member number).
                                    # Additionally, when the value looks
                                    # like a full account (member + fixed
                                    # suffix), also store the stripped
                                    # member-only key so loans whose join
                                    # key is the bare member still match.
                                    try:
                                        raw_key = int(float(m_str))
                                        scores[raw_key] = score_val
                                    except (ValueError, TypeError):
                                        raw_key = None
                                    if suffix_length and len(m_str) > suffix_length:
                                        try:
                                            stripped_key = int(m_str[:-suffix_length])
                                            scores.setdefault(stripped_key, score_val)
                                        except (ValueError, TypeError):
                                            pass
                                except (ValueError, TypeError):
                                    pass
                        print(f"    Credit pull file: {fname} ({len(scores)} scores loaded)")
                        if pull_as_of is None:
                            try:
                                pull_as_of = pd.Timestamp(os.path.getmtime(fpath), unit='s')
                            except (OSError, ValueError):
                                pass
                        break  # headers matched + scores loaded; stop here
                    else:
                        # File matched the regex but its headers don't carry the
                        # configured member/score columns. Common when a shared
                        # temp/sample folder holds files for multiple CUs with
                        # divergent header conventions (Account # vs Account
                        # Number, FICO vs Credit Score, etc.). Skip this file
                        # and let the loop try the next match.
                        print(f"    Credit pull file {fname}: headers do not contain "
                              f"'{member_col}' and/or '{score_col}' — skipping.")
                        continue

    # 2. Also check credit pull tabs in existing CECL report to fill gaps
    # (older credit pulls may cover members not in the latest standalone file)
    warm_file_path = None  # track the WARM file for step 3
    report_pattern = cp_config.get('fallback_report_pattern')
    report_folder = cp_config.get('fallback_report_folder')
    if report_pattern and report_folder:
        folder_path = resolve_path(report_folder)
        if os.path.isdir(folder_path):
            pattern_re = re.compile(report_pattern, re.IGNORECASE)
            sheet_pattern = cp_config.get('fallback_sheet_pattern', 'Credit Pull')
            member_idx = cp_config.get('fallback_member_col', 0)
            score_idx = cp_config.get('fallback_score_col', 1)
            matching_reports = []
            for root, _dirs, files in os.walk(folder_path):
                for fname in files:
                    if fname.startswith('~$'):
                        continue
                    if pattern_re.search(fname):
                        matching_reports.append(os.path.join(root, fname))

            for fpath in sorted(matching_reports, key=os.path.getmtime, reverse=True):
                    warm_file_path = fpath
                    if pull_as_of is None:
                        try:
                            pull_as_of = pd.Timestamp(os.path.getmtime(fpath), unit='s')
                        except (OSError, ValueError):
                            pass
                    fname = os.path.basename(fpath)
                    xl = pd.ExcelFile(fpath)
                    cp_sheets = [s for s in xl.sheet_names
                                 if sheet_pattern.lower() in s.lower()]
                    if not cp_sheets:
                        continue
                    # Sort credit pull sheets: most recent first (by date in name)
                    # so we fill newest scores first and older sheets fill gaps
                    cp_sheets = _sort_credit_pull_sheets(cp_sheets)
                    for sheet_name in cp_sheets:
                        sheet_count = 0
                        df = pd.read_excel(xl, sheet_name, header=None, skiprows=1)
                        for _, row in df.iterrows():
                            m = row.iloc[member_idx]
                            s = row.iloc[score_idx]
                            if pd.notna(m) and pd.notna(s):
                                try:
                                    score_val = int(float(s))
                                    mem_id = int(float(m))
                                    if score_val > 0 and mem_id not in scores:
                                        scores[mem_id] = score_val
                                        sheet_count += 1
                                except (ValueError, TypeError):
                                    pass
                        print(f"    Credit pull tab: '{sheet_name}' from {fname} (+{sheet_count} scores, total {len(scores)})")
                    break  # use first matching report

    # 3. Fill remaining gaps from WARM "All Loans" current score column.
    # The WARM workbook carries forward scores from prior periods per loan.
    # This is useful on first import when there's no previous snapshot in the DB.
    # Scores are keyed per-loan (member-suffix) since different loans for the
    # same member may carry different scores.
    warm_scores_cfg = cp_config.get('warm_scores', {})
    if warm_scores_cfg:
        warm_sheet = warm_scores_cfg.get('sheet', 'All Loans')
        warm_member_col = warm_scores_cfg.get('member_col', 0)
        warm_suffix_col = warm_scores_cfg.get('suffix_col', 1)
        warm_score_col = warm_scores_cfg.get('score_col', 8)
        warm_skip_rows = warm_scores_cfg.get('skip_rows', 2)  # header + column labels
        suffix_length = config.get('account_suffix_length', 0)

        # Use the same WARM file found in step 2, or search for it
        warm_path = warm_file_path
        if not warm_path:
            warm_pattern = warm_scores_cfg.get('file_pattern', report_pattern)
            warm_folder = warm_scores_cfg.get('folder', report_folder)
            if warm_pattern and warm_folder:
                folder_path = resolve_path(warm_folder)
                if os.path.isdir(folder_path):
                    pattern_re = re.compile(warm_pattern, re.IGNORECASE)
                    for root, _dirs, files in os.walk(folder_path):
                        for fname in sorted(files, reverse=True):
                            if fname.startswith('~$'):
                                continue
                            if pattern_re.search(fname):
                                warm_path = os.path.join(root, fname)
                                break
                        if warm_path:
                            break

        if warm_path:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(warm_path, data_only=True, read_only=True)
                if warm_sheet in wb.sheetnames:
                    ws = wb[warm_sheet]
                    warm_count = 0
                    warm_member_count = 0
                    for row_idx, row in enumerate(ws.iter_rows(
                            min_row=warm_skip_rows + 1, values_only=True)):
                        if len(row) <= max(warm_member_col, warm_suffix_col, warm_score_col):
                            continue
                        m = row[warm_member_col]
                        suffix = row[warm_suffix_col]
                        s = row[warm_score_col]
                        if m is not None and s is not None:
                            try:
                                mem_id = int(float(m))
                                score_val = int(float(s))
                                suffix_str = str(int(float(suffix))).zfill(suffix_length) if suffix is not None and suffix_length > 0 else str(int(float(suffix))) if suffix is not None else ''
                                # Build account-level key (member+suffix) matching loan file formatformat
                                acct_key = f"{mem_id}{suffix_str}"
                                if score_val > 0:
                                    if acct_key not in warm_acct_scores:
                                        warm_acct_scores[acct_key] = score_val
                                        warm_count += 1
                                    # Also add member-level for credit pull gap-fill
                                    if mem_id not in scores:
                                        scores[mem_id] = score_val
                                        warm_member_count += 1
                            except (ValueError, TypeError):
                                pass
                    if warm_count > 0:
                        print(f"    WARM '{warm_sheet}' scores: +{warm_member_count} members, +{warm_count} accounts (total {len(scores)} members)")
                wb.close()
            except Exception as e:
                print(f"    WARNING: Could not read WARM scores: {e}")

    return scores, warm_acct_scores, pull_as_of


def extract_member_number(account_series, suffix_length):
    """Strip the loan suffix from account numbers to get the member number."""
    acct_str = account_series.astype(str).str.strip()
    # Drop trailing ".0" produced when pandas reads numeric account columns as float
    acct_str = acct_str.str.replace(r'\.0+$', '', regex=True)
    if suffix_length and suffix_length > 0:
        return acct_str.str[:-suffix_length].astype(int)
    return pd.to_numeric(acct_str, errors='coerce').fillna(0).astype(int)


def _clean_id_series(series):
    """Trim whitespace and strip trailing '.0' that pandas leaves on
    numeric-looking ID columns. Returns a string series."""
    s = series.astype(str).str.strip()
    return s.str.replace(r'\.0+$', '', regex=True)


def derive_member_account(df, config, has_header):
    """Return (member_only_str, full_account_str) honoring the three input
    modes captured by the wizard:

      * fixed_suffix : single column, last N chars are the account/suffix.
      * delimiter    : single column, member & account split by a delimiter.
      * split        : two columns; member col + account/suffix col.

    Falls back to the legacy `account_suffix_length` behavior when the
    `member_account` block is absent (older configs).
    """
    col_map = config['column_mappings']
    ma = config.get('member_account') or {}
    mode = ma.get('mode') or 'fixed_suffix'

    def _col(field):
        ref = col_map[field]
        return df[ref] if has_header else df.iloc[:, ref]

    member_raw = _clean_id_series(_col('member_number'))

    if mode == 'split' and col_map.get('loan_suffix'):
        suffix_raw = _clean_id_series(_col('loan_suffix'))
        # Pad suffix to 3 chars by default (match historical convention).
        # Honour explicit suffix_length=0 (caller wants the suffix used
        # verbatim with no zero-pad). Plain ``or 3`` would silently flip
        # 0 back to 3 because 0 is falsy in Python.
        _sl_raw = ma.get('suffix_length')
        try:
            pad_len = int(_sl_raw) if _sl_raw is not None else 3
        except (TypeError, ValueError):
            pad_len = 3
        suffix_padded = suffix_raw.str.zfill(pad_len) if pad_len > 0 else suffix_raw
        full = member_raw + suffix_padded
        return member_raw, full

    if mode == 'delimiter':
        delim = ma.get('delimiter') or '-'
        # Split once: left=member, right=account
        parts = member_raw.str.split(delim, n=1, expand=True)
        member_only = parts[0].fillna(member_raw)
        # Reconstruct the "full account" identifier without the delimiter so
        # the DB key matches the credit-pull / WARM convention of one
        # contiguous string.
        if parts.shape[1] > 1:
            account_part = parts[1].fillna('')
            full = member_only + account_part
        else:
            full = member_only
        return member_only, full

    # mode == 'fixed_suffix' (or unknown): use legacy suffix-length logic.
    suffix_length = int(
        ma.get('suffix_length')
        if ma.get('suffix_length') is not None
        else config.get('account_suffix_length', 0) or 0
    )
    if suffix_length > 0:
        member_only = member_raw.str[:-suffix_length]
    else:
        member_only = member_raw
    return member_only, member_raw


def _load_previous_fico(cu_name, current_snapshot):
    """Load current_fico_score from the most recent previous snapshot.

    Returns a dict mapping raw account number (str) -> previous current_fico_score,
    keyed at the loan level (not member level) so that each loan can carry forward
    its own score from the prior period.
    """
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT MAX(snapshot_date) FROM monthly_loan_data "
                 "WHERE credit_union = :cu AND snapshot_date < :snap"),
            {"cu": cu_name, "snap": current_snapshot}
        ).fetchone()
    prev_date = row[0] if row and row[0] else None
    if not prev_date:
        return {}

    prev_df = pd.read_sql(
        text("SELECT member_number, current_fico_score FROM monthly_loan_data "
             "WHERE credit_union = :cu AND snapshot_date = :snap"),
        engine, params={"cu": cu_name, "snap": str(prev_date)}
    )
    if prev_df.empty:
        return {}

    # Build lookup keyed by raw account number (includes suffix)
    scores = {}
    for _, r in prev_df.iterrows():
        acct = str(r['member_number']).strip()
        fico = int(r['current_fico_score'])
        if fico > 0:
            scores[acct] = fico
    return scores


def _excel_letter_to_index(letter: str) -> int:
    """'A' -> 0, 'Z' -> 25, 'AA' -> 26, etc. Raises ValueError on bad input."""
    s = str(letter).strip().upper()
    if not s or not s.isalpha():
        raise ValueError(f"Not a column letter: {letter!r}")
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n - 1


def _excel_idx_to_letter(idx: int) -> str:
    """0 -> 'A', 25 -> 'Z', 26 -> 'AA', etc. Inverse of _excel_letter_to_index."""
    if idx < 0:
        return ""
    s = ""
    n = idx
    while True:
        s = chr(ord("A") + (n % 26)) + s
        n = n // 26 - 1
        if n < 0:
            break
    return s


def _normalize_col_map_for_no_header(col_map):
    """When the loan extracts have no header row, the wizard stores mapping
    values like ``"col_A"`` (or already-integer positions on legacy configs).
    Return a copy with every value coerced to a 0-based integer position."""
    out = {}
    for field, val in (col_map or {}).items():
        if val is None or val == "":
            continue
        if isinstance(val, int):
            out[field] = val
            continue
        s = str(val).strip()
        if s.lower().startswith("col_"):
            try:
                out[field] = _excel_letter_to_index(s[4:])
                continue
            except ValueError:
                pass
        try:
            out[field] = int(s)
        except ValueError:
            try:
                out[field] = _excel_letter_to_index(s)
            except ValueError:
                # Leave as-is; will fail loudly downstream with a clear msg.
                out[field] = val
    return out


def import_file(file_path, config, snapshot_date, credit_pull_scores=None,
                warm_acct_scores=None, pull_as_of=None,
                wiped_snapshots=None):
    """Import a single file using the client config. Returns count of imported rows.

    ``wiped_snapshots`` (optional ``set[str]``) controls the pre-insert
    DELETE strategy. When non-None (full-folder imports from
    ``process_client``), the FIRST file for any given ``snapshot_date``
    does ONE wholesale ``DELETE WHERE (credit_union, snapshot_date)`` and
    records the snapshot in the set; subsequent files for the same
    snapshot just APPEND. This is the correct behavior for multi-extract
    CUs where two or more files contribute to the same snapshot AND may
    produce overlapping ``loan_pool`` values (e.g. Destinations CU's
    ``ceclce`` + ``cecloe`` both producing ``All Other Unsecured Loans/LOC``
    rows for codes 10/11 vs 34). When None (specific-file CLI re-import
    via ``--file=X.xlsx``), falls back to the legacy per-pool delete so
    re-running a single file does not nuke other extracts' rows for the
    same snapshot.
    """
    cu_name = config['credit_union']
    col_map = config['column_mappings']
    has_header = config.get('has_header', True)
    suffix_length = config.get('account_suffix_length', 0)

    # When the wizard set a "fixed loan pool code for every row" on this
    # file, the loan_pool_code column is optional — we synthesize a
    # constant series for it below.
    static_pool_code = (col_map.get('loan_pool_code_static') or '').strip()

    ext = os.path.splitext(file_path)[1].lower()
    if has_header:
        # header_row is 1-indexed; pandas wants 0-indexed. 0/1/missing
        # → default header=0. Required for AIRES-style extracts whose
        # real headers live on row 2 (row 1 = column position numbers).
        try:
            hr_cfg = int(config.get('header_row') or 0)
        except (TypeError, ValueError):
            hr_cfg = 0
        pd_header = hr_cfg - 1 if hr_cfg > 1 else 0
        if ext == '.csv':
            df = pd.read_csv(file_path, header=pd_header,
                             keep_default_na=False,
                             na_values=_NA_VALUES_KEEP_LITERAL)
        else:
            df = pd.read_excel(file_path, header=pd_header,
                               keep_default_na=False,
                               na_values=_NA_VALUES_KEEP_LITERAL)
        # Normalise header cells to match what the wizard's sample_parser
        # emits: collapse internal whitespace (wrap-text Excel cells often
        # contain CR/LF inside a single header cell), and replace any
        # blank / 'nan' / pandas 'Unnamed: N' placeholder with the
        # ``col_<LETTER>`` form so the user's saved column_mappings —
        # which may reference e.g. ``col_H`` for an unlabelled column —
        # resolves to the right pandas column.
        _unnamed_rx_runtime = re.compile(r"^unnamed:\s*\d+(?:_level_\d+)*$",
                                         re.IGNORECASE)
        _normed_cols = []
        for _i, _c in enumerate(df.columns):
            _s = re.sub(r"\s+", " ", str(_c)).strip() if _c is not None else ""
            _low = _s.lower()
            if (not _s) or _low == "nan" or _unnamed_rx_runtime.match(_s):
                _normed_cols.append(f"col_{_excel_idx_to_letter(_i)}")
            else:
                _normed_cols.append(_s)
        df.columns = _normed_cols

        required = ['member_number', 'current_balance']
        if not static_pool_code:
            required.append('loan_pool_code')
        missing = [
            f for f in required
            if not col_map.get(f) or col_map.get(f) not in df.columns
        ]
        if missing:
            # Header-row auto-detection. The configured header_row didn't
            # surface the required columns — common when a CU changes an
            # export's layout between periods (e.g. AIRES cardholder files
            # that moved their header from row 1 to row 4). Scan the first
            # rows for the one that actually carries the mapped column
            # names and re-read with it, so a stale header_row no longer
            # silently drops a whole file (which had been dropping every
            # credit-card loan for Central Keystone's June extract).
            wanted = {
                re.sub(r"\s+", " ", str(v)).strip().lower()
                for k, v in col_map.items()
                if k in ('member_number', 'current_balance', 'loan_pool_code')
                and isinstance(v, str) and v.strip()
            }
            if wanted:
                try:
                    probe = (
                        pd.read_csv(file_path, header=None, nrows=15, dtype=str)
                        if ext == '.csv'
                        else pd.read_excel(file_path, header=None, nrows=15)
                    )
                except Exception:  # noqa: BLE001
                    probe = None
                best_i, best_hits = None, 0
                if probe is not None:
                    for _i in range(len(probe)):
                        cells = {
                            re.sub(r"\s+", " ", str(v)).strip().lower()
                            for v in probe.iloc[_i].tolist()
                            if v is not None and str(v).strip()
                        }
                        hits = len(wanted & cells)
                        if hits > best_hits:
                            best_hits, best_i = hits, _i
                if (best_i is not None and best_i != pd_header
                        and best_hits >= min(2, len(wanted))):
                    if ext == '.csv':
                        df = pd.read_csv(file_path, header=best_i)
                    else:
                        df = pd.read_excel(file_path, header=best_i)
                    _normed_cols = []
                    for _i2, _c in enumerate(df.columns):
                        _s = (re.sub(r"\s+", " ", str(_c)).strip()
                              if _c is not None else "")
                        _low = _s.lower()
                        if (not _s) or _low == "nan" \
                                or _unnamed_rx_runtime.match(_s):
                            _normed_cols.append(
                                f"col_{_excel_idx_to_letter(_i2)}")
                        else:
                            _normed_cols.append(_s)
                    df.columns = _normed_cols
                    print(f"    Auto-detected header on row {best_i + 1} "
                          f"(configured header_row={hr_cfg} did not match "
                          f"columns {missing}).")
        for field in required:
            src_col = col_map.get(field)
            if not src_col or src_col not in df.columns:
                raise ValueError(f"Required column '{field}' mapped to '{src_col}' not found. "
                                 f"Available: {list(df.columns)[:20]}")
    else:
        if ext == '.csv':
            df = pd.read_csv(file_path, header=None)
        else:
            df = pd.read_excel(file_path, header=None)

        # Wizard stores mapping values as "col_A"/"col_B" strings; the
        # importer needs 0-based integer positions when reading without
        # headers. Translate once up-front and rebind the local for the rest
        # of the function (and for derive_member_account, which reads
        # col_map via config).
        col_map = _normalize_col_map_for_no_header(col_map)
        config = dict(config)
        config['column_mappings'] = col_map

        required = ['member_number', 'current_balance']
        if not static_pool_code:
            required.append('loan_pool_code')
        for field in required:
            pos = col_map.get(field)
            if not isinstance(pos, int) or pos < 0 or pos >= len(df.columns):
                raise ValueError(f"Required column '{field}' mapped to position {pos!r} "
                                 f"but file has only {len(df.columns)} columns")

    # Build the output DataFrame — access by name or position index
    def col(field):
        if field == 'loan_pool_code' and static_pool_code:
            return pd.Series([static_pool_code] * len(df), index=df.index)
        return df[col_map[field]] if has_header else df.iloc[:, col_map[field]]

    # AIRES file provides the original FICO score (optional)
    if col_map.get('original_fico_score') or 'original_fico_score' in col_map:
        try:
            original_fico = pd.to_numeric(col('original_fico_score'), errors='coerce').fillna(0).astype(int)
        except KeyError:
            original_fico = pd.Series([0] * len(df), index=df.index, dtype=int)
    else:
        original_fico = pd.Series([0] * len(df), index=df.index, dtype=int)

    # Optional: current FICO read directly from the loan-data extract.
    # When the wizard's column-mapping step assigns ``current_fico_score``
    # to a column in the extract, those values are the authoritative
    # current scores and take priority over WARM/credit-pull lookups.
    # EXCEPTION: when ``current_fico_score`` and ``original_fico_score``
    # both map to the SAME source column, that column holds only the
    # origination score (the wizard auto-suggest landed on a single
    # ``Credit Score`` column for both fields). Treating it as the
    # current score would force original==current and silently disable
    # the credit-pull join. In that case, leave current FICO unmapped
    # from the extract and let the priority chain fall through to the
    # credit-pull / WARM / previous-snapshot sources.
    extract_current_fico: pd.Series | None = None
    cur_col = col_map.get('current_fico_score') or ''
    orig_col = col_map.get('original_fico_score') or ''
    same_col = bool(cur_col) and bool(orig_col) and \
        str(cur_col).strip().lower() == str(orig_col).strip().lower()
    if (col_map.get('current_fico_score') or 'current_fico_score' in col_map) \
            and not same_col:
        try:
            extract_current_fico = (
                pd.to_numeric(col('current_fico_score'), errors='coerce')
                .fillna(0)
                .astype(int)
            )
        except KeyError:
            extract_current_fico = None
    elif same_col:
        print("    Note: current_fico_score and original_fico_score map to "
              "the same column — falling through to credit-pull/WARM for "
              "current scores.")

    # Extract member-only & full-account identifiers honoring the wizard's
    # member/account format selection (fixed-suffix / delimiter / split).
    member_only_str, full_account_str = derive_member_account(df, config, has_header)
    raw_account = full_account_str           # full account string (DB key)
    raw_acct_str = full_account_str          # alias used downstream
    # member_numbers is used as a join key against credit_pull_scores (int keys).
    member_numbers = pd.to_numeric(member_only_str, errors='coerce').fillna(0).astype(int)

    # ----------------------------------------------------------------------
    # Original-FICO baseline fallback (wizard's "Original Score Baseline"
    # step). Fills in original_fico for loans whose extract doesn't carry
    # the original score (e.g. VISA / credit-card files). Two lookup keys
    # are tried, in order of specificity:
    #   1. full account string (member + suffix) — exact loan match.
    #   2. member-only string  — member-level fallback when no suffix was
    #      mapped on the baseline file.
    # Only loans with original_fico == 0 are touched; existing scores from
    # the extract are preserved.
    # ----------------------------------------------------------------------
    baseline_cfg = config.get('original_fico_baseline') or {}
    baseline_rows = baseline_cfg.get('rows') or []
    if baseline_rows:
        full_lookup: dict[str, int] = {}
        member_lookup: dict[str, int] = {}
        for r in baseline_rows:
            mem = str(r.get('member') or '').strip()
            if not mem:
                continue
            try:
                score = int(r.get('score') or 0)
            except (TypeError, ValueError):
                continue
            if score <= 0:
                continue
            suf = str(r.get('suffix') or '').strip()
            if suf:
                # Match the importer's convention of zero-padding suffixes.
                ma_b = config.get('member_account') or {}
                pad_len = int(ma_b.get('suffix_length') or 0) or len(suf)
                if pad_len > 0 and suf.isdigit():
                    suf_padded = suf.zfill(pad_len)
                else:
                    suf_padded = suf
                full_lookup[mem + suf_padded] = score
                full_lookup[mem + suf] = score  # also key as-given
            else:
                member_lookup[mem] = score
        if full_lookup or member_lookup:
            missing_mask = (original_fico == 0)
            # Optional pool scope: only fill within the listed loan_pool
            # value(s). Empty/missing list = apply globally.
            scoped_pools = [
                str(p).strip() for p in (baseline_cfg.get('pools') or [])
                if str(p).strip()
            ]
            if scoped_pools:
                # Derive the mapped pool name the same way clean_data does
                # later on (map raw loan_pool_code -> pool_map names). Use
                # the static value when the file carries no code column
                # (e.g. credit-card files with loan_pool_code_static set).
                try:
                    if static_pool_code:
                        raw_code_series = pd.Series(
                            [static_pool_code] * len(df), index=df.index
                        )
                    else:
                        raw_code_series = col('loan_pool_code')
                    pool_series = map_pool_codes(
                        raw_code_series, config
                    ).astype(str).str.strip()
                except KeyError:
                    pool_series = pd.Series([''] * len(df), index=df.index)
                pool_mask = pool_series.isin(scoped_pools)
                missing_mask = missing_mask & pool_mask
            if missing_mask.any():
                # Try full-account match first.
                filled = pd.Series(
                    [0] * len(df), index=df.index, dtype='Int64'
                )
                if full_lookup:
                    filled = raw_acct_str.map(full_lookup).astype('Int64')
                # Member-only fallback for still-unmatched rows.
                if member_lookup:
                    member_filled = member_only_str.map(member_lookup).astype('Int64')
                    filled = filled.fillna(member_filled)
                apply_mask = missing_mask & filled.notna()
                if apply_mask.any():
                    original_fico = original_fico.where(
                        ~apply_mask, filled
                    ).fillna(0).astype(int)
                    scope_msg = (
                        f" (pool scope: {', '.join(scoped_pools)})"
                        if scoped_pools else ""
                    )
                    print(
                        f"    Original-FICO baseline: filled {int(apply_mask.sum())} "
                        f"loan(s) from {len(baseline_rows)} baseline row(s)"
                        f"{scope_msg}"
                    )

    # Current FICO priority:
    #   0. Extract column mapped to ``current_fico_score`` (when present)
    #   1. WARM "All Loans" per-loan scores (authoritative source matching WARM's final scores)
    #   2. Credit pull (member-level) for loans not in WARM
    #   3. Previous snapshot per-loan carry-forward
    #   4. Original FICO
    # raw_acct_str is set above (alias of full_account_str)

    if extract_current_fico is not None:
        # Extract column wins. Treat 0/blank as missing and fall back to
        # WARM/credit-pull/previous-snapshot/original in that order so a
        # CU whose extract has current FICO on most-but-not-all loans
        # still benefits from the other sources for the gaps.
        current_fico = extract_current_fico.where(extract_current_fico > 0)
        ext_matched = int(current_fico.notna().sum())
        unmatched = int(current_fico.isna().sum())

        if unmatched > 0 and warm_acct_scores:
            warm_mapped = raw_acct_str.map(warm_acct_scores)
            current_fico = current_fico.fillna(warm_mapped)
            unmatched = int(current_fico.isna().sum())

        cp_filled = 0
        if unmatched > 0 and credit_pull_scores:
            cp_mapped = member_numbers.map(credit_pull_scores)
            before = int(current_fico.notna().sum())
            current_fico = current_fico.fillna(cp_mapped)
            cp_filled = int(current_fico.notna().sum()) - before
            unmatched = int(current_fico.isna().sum())

        parts = [f"Extract current-FICO matched: {ext_matched}"]
        if cp_filled > 0:
            parts.append(f"credit pull: {cp_filled}")
        if unmatched > 0:
            parts.append(f"fallback to original: {unmatched}")
        print(f"    {', '.join(parts)}")

        # Final fallback to original score
        current_fico = current_fico.fillna(original_fico).astype(int)

    elif warm_acct_scores:
        # Primary: WARM per-loan scores (exactly matches WARM's computed current scores)
        current_fico = raw_acct_str.map(warm_acct_scores)
        warm_matched = current_fico.notna().sum()
        unmatched = current_fico.isna().sum()

        # Secondary: credit pull (member-level) for loans not in WARM
        cp_filled = 0
        if unmatched > 0 and credit_pull_scores:
            cp_mapped = member_numbers.map(credit_pull_scores)
            current_fico = current_fico.fillna(cp_mapped)
            cp_filled = current_fico.notna().sum() - warm_matched
            unmatched = current_fico.isna().sum()

        parts = [f"WARM per-loan matched: {warm_matched}"]
        if cp_filled > 0:
            parts.append(f"credit pull: {cp_filled}")
        if unmatched > 0:
            parts.append(f"fallback to original: {unmatched}")
        print(f"    {', '.join(parts)}")

        # Final fallback to original score
        current_fico = current_fico.fillna(original_fico).astype(int)

    elif credit_pull_scores:
        current_fico = member_numbers.map(credit_pull_scores)
        matched = current_fico.notna().sum()
        unmatched = current_fico.isna().sum()

        if unmatched > 0:
            prev_scores = _load_previous_fico(cu_name, snapshot_date)
            if prev_scores:
                prev_mapped = raw_acct_str.map(prev_scores)
                prev_filled_count = prev_mapped.notna().sum()
                current_fico = current_fico.fillna(prev_mapped)
                unmatched = current_fico.isna().sum()
            else:
                prev_filled_count = 0

            parts = [f"Credit pull matched: {matched}"]
            if prev_filled_count > 0:
                parts.append(f"previous snapshot: {prev_filled_count}")
            if unmatched > 0:
                parts.append(f"fallback to original: {unmatched}")
            print(f"    {', '.join(parts)}")
        else:
            print(f"    Credit pull matched: {matched}")

        # Final fallback to original score
        current_fico = current_fico.fillna(original_fico).astype(int)
    else:
        current_fico = original_fico.copy()
        print(f"    WARNING: No credit pull data — current score = original score")

    # ----------------------------------------------------------------------
    # Original-score fallback for loans newer than the credit pull.
    # When the wizard's `prefer_original_for_new_loans` flag is set, any loan
    # whose open_date is after the credit-pull as-of date gets its current
    # score replaced with the loan-file's original score, on the theory that
    # the original was pulled at origination — i.e. more recent than the
    # bureau-wide pull. Without an open_date column or a known pull date this
    # block is a no-op (the bureau scores already loaded above stand).
    # ----------------------------------------------------------------------
    cp_cfg = config.get('credit_pull') or {}
    if cp_cfg.get('prefer_original_for_new_loans'):
        open_date_col = (config.get('column_mappings') or {}).get('open_date')
        if open_date_col is not None and (
            (has_header and open_date_col in df.columns)
            or (not has_header and isinstance(open_date_col, int)
                and open_date_col < len(df.columns))
        ):
            open_date_series = pd.to_datetime(
                df[open_date_col] if has_header else df.iloc[:, open_date_col],
                errors='coerce',
            )
            # Per-member "most recent loans" method: for each member, take the
            # credit score from their most-recently-opened loan (non-zero) and
            # apply it as the current score to every loan that member has, so
            # credit migration can be shown without a separate bureau pull.
            #
            # Scope depends on whether a real credit pull is configured:
            #   * pull configured  -> only override loans opened AFTER the pull
            #     (their origination score is fresher than the bureau-wide pull).
            #   * NO pull configured -> derive the current score for EVERY loan
            #     from the member's most-recent origination score. This is the
            #     analyst's manual "build a credit pull from the AIRES file"
            #     process, now automatic.
            if pull_as_of is not None:
                scope_mask = open_date_series.notna() & (
                    open_date_series > pd.Timestamp(pull_as_of))
                scope_label = f"newer than {pd.Timestamp(pull_as_of).date()}"
            else:
                scope_mask = pd.Series(True, index=df.index)
                scope_label = "all loans (no credit pull)"
            # Per-member: pick the original score from the most-recently-opened
            # loan and apply it to every loan that member has.
            tmp = pd.DataFrame({
                'member': member_numbers,
                'open_date': open_date_series,
                'orig_fico': original_fico,
            })
            # Only consider loans with a positive original score and a real
            # open date; zeros/blanks represent missing data and shouldn't
            # override real scores or drive the "most recent" comparison.
            tmp_valid = tmp[(tmp['orig_fico'] > 0) & tmp['open_date'].notna()]
            if not tmp_valid.empty:
                idx_max = tmp_valid.groupby('member')['open_date'].idxmax()
                member_to_latest = (
                    tmp_valid.loc[idx_max].set_index('member')['orig_fico']
                )
                replacement = member_numbers.map(member_to_latest)
                apply_mask = scope_mask & replacement.notna()
                if apply_mask.any():
                    current_fico = current_fico.where(
                        ~apply_mask, replacement.astype('Int64')
                    )
                    # Coerce back to int after the masked update.
                    current_fico = current_fico.fillna(0).astype(int)
                    print(
                        f"    Original-score fallback: {int(apply_mask.sum())} "
                        f"loan(s) [{scope_label}] now use member's "
                        f"most-recent original score"
                    )
        elif cp_cfg.get('prefer_original_for_new_loans'):
            print("    Original-score fallback: open_date column not mapped — "
                  "cannot derive most-recent-loan scores.")

    clean_data = pd.DataFrame({
        'credit_union': cu_name,
        'snapshot_date': snapshot_date,
        'member_number': raw_account,
        'current_balance': clean_balance(
            col('current_balance'), config.get('balance_format', {})
        ),
        'current_fico_score': current_fico,
        'original_fico_score': original_fico,
        'loan_pool': map_pool_codes(col('loan_pool_code'), config),
    })

    # ── Charge-off exclusion ──
    # Some cores leave charged-off loans in the extract under their original
    # (active) loan-type code, carrying the written-off amount in a dedicated
    # "charge off amount" column even though the GL / balance sheet has already
    # removed them. When ``chargeoff_exclude_column`` is set (a header name, or
    # a 0-based column index for headerless files), any row with a non-zero
    # value there is routed to the ``Exclude`` pool so it drops out of the
    # reserve population and the loan totals reconcile to the balance sheet.
    _co_ref = config.get('chargeoff_exclude_column')
    if _co_ref is not None and _co_ref != '':
        try:
            _co_series = df[_co_ref] if has_header else df.iloc[:, int(_co_ref)]
            _co_mask = pd.to_numeric(_co_series, errors='coerce').fillna(0) != 0
            _n_co = int(_co_mask.sum())
            if _n_co:
                clean_data.loc[_co_mask, 'loan_pool'] = 'Exclude'
                print(f"    Charge-off exclusion: {_n_co} loan(s) with a non-zero "
                      f"'{_co_ref}' routed to Exclude (already charged off).")
        except (KeyError, IndexError, ValueError) as _co_exc:
            print(f"    WARNING: chargeoff_exclude_column {_co_ref!r} not usable: "
                  f"{_co_exc}")

    # report engine routes this through ``cecl_engine.assign_business_risk_grade``
    # when the loan's pool is flagged ``brr: true`` in the CU's YAML.
    # Stored as text so analyst-defined rating labels (e.g. "Pass",
    # "Special Mention") survive intact alongside numeric ratings.
    if col_map.get('business_risk_rating') or 'business_risk_rating' in col_map:
        try:
            brr_series = col('business_risk_rating').astype(str).str.strip()
            brr_series = brr_series.where(
                ~brr_series.isin(('', 'nan', 'NaN', 'None')), None
            )
            clean_data['business_risk_rating'] = brr_series
        except KeyError:
            clean_data['business_risk_rating'] = None
    else:
        clean_data['business_risk_rating'] = None

    # When original FICO is 0 but current is known, treat as unchanged (WARM convention)
    mask = (clean_data['original_fico_score'] == 0) & (clean_data['current_fico_score'] > 0)
    if mask.any():
        clean_data.loc[mask, 'original_fico_score'] = clean_data.loc[mask, 'current_fico_score']
        print(f"    Original FICO gap-fill: {mask.sum()} loans set original = current")

    clean_data = clean_data.dropna(subset=['current_balance'])
    clean_data = clean_data[clean_data['current_balance'] > 0]

    if len(clean_data) == 0:
        return 0

    with engine.begin() as conn:
        # Make sure the optional ``business_risk_rating`` column exists on
        # the table. The schema was originally created implicitly by
        # ``to_sql(if_exists='append')`` on the first import; adding the
        # column here as an idempotent ALTER lets pre-existing CECL
        # databases pick up BRR support without a manual migration step.
        # Postgres ``ADD COLUMN IF NOT EXISTS`` is a no-op when present.
        try:
            conn.execute(text(
                "ALTER TABLE monthly_loan_data "
                "ADD COLUMN IF NOT EXISTS business_risk_rating TEXT"
            ))
        except Exception as alter_err:  # pragma: no cover - safety net
            # If the dialect doesn't speak IF NOT EXISTS (rare; e.g.
            # SQLite < 3.35) we'd rather log + drop the BRR column from
            # the DataFrame than block the whole import.
            print(f"    [warn] could not ensure business_risk_rating column: {alter_err}")
            if 'business_risk_rating' in clean_data.columns:
                clean_data = clean_data.drop(columns=['business_risk_rating'])

        # Two delete strategies, picked by the ``wiped_snapshots``
        # parameter (controlled by ``process_client``):
        #
        #   (a) Full-folder import (``wiped_snapshots`` is a set): the
        #       FIRST file for a snapshot does ONE wholesale
        #       ``DELETE WHERE (cu, snap)`` and records the snapshot;
        #       subsequent files for the same snapshot just APPEND. This
        #       is the correct behavior for multi-extract CUs where two
        #       files for the same snapshot may produce overlapping
        #       ``loan_pool`` values (e.g. Destinations ``ceclce`` and
        #       ``cecloe`` both producing ``All Other Unsecured Loans/LOC``
        #       rows). The previous per-pool delete had the second file
        #       silently wipe the first file's rows for overlapping pools.
        #
        #   (b) Specific-file re-import (``wiped_snapshots`` is None,
        #       e.g. ``--file=X.xlsx``): fall back to the per-pool delete
        #       so re-running a single file does not nuke other extracts'
        #       rows for the same snapshot.
        if wiped_snapshots is not None:
            snap_key = str(snapshot_date)
            if snap_key not in wiped_snapshots:
                conn.execute(
                    text(
                        "DELETE FROM monthly_loan_data "
                        "WHERE credit_union = :cu "
                        "AND snapshot_date = :sd"
                    ),
                    {"cu": cu_name, "sd": snapshot_date},
                )
                wiped_snapshots.add(snap_key)
        else:
            pools_in_file = sorted({str(p) for p in clean_data['loan_pool'].dropna().unique()})
            if pools_in_file:
                conn.execute(
                    text(
                        "DELETE FROM monthly_loan_data "
                        "WHERE credit_union = :cu "
                        "AND snapshot_date = :sd "
                        "AND loan_pool = ANY(:pools)"
                    ),
                    {"cu": cu_name, "sd": snapshot_date, "pools": pools_in_file},
                )
        clean_data.to_sql('monthly_loan_data', conn, if_exists='append', index=False)

    return len(clean_data)


def process_client(client_name, specific_file=None, scan_folder_override=None):
    """Process all matching files for a client."""
    config = load_client_config(client_name)
    cu_name = config['credit_union']
    # ``file_pattern`` accepts either a single regex string (legacy) or a
    # list of regex strings (multi-pattern, first-match-wins). Empty list
    # means no top-level catch-all is enforced — only per-extract patterns
    # decide which files to import.
    top_pattern_res = _compile_file_patterns(
        config.get('file_pattern'), label='top-level',
    )

    # Per-file loan-data extracts (wizard "Column Mappings" step). When
    # present, each entry overrides ``column_mappings`` / ``member_account``
    # / ``has_header`` / ``account_suffix_length`` for files whose name
    # matches the extract's own ``file_pattern``. Falls back to the
    # top-level mapping when no per-file pattern matches.
    extracts_raw = config.get('loan_data_extracts') or []
    extracts: list[tuple[list[re.Pattern], dict]] = []
    for e in extracts_raw:
        label = (e or {}).get('label', '?')
        pats = _compile_file_patterns(
            (e or {}).get('file_pattern'), label=label,
        )
        if pats:
            extracts.append((pats, e))

    # Explicit scan-folder override (e.g. the "re-import a single period"
    # action stages just that period's files into a temp folder and imports
    # ONLY them). Takes precedence over config and is never archived.
    if scan_folder_override:
        scan_folder = str(scan_folder_override)
        if not os.path.isdir(scan_folder):
            raise FileNotFoundError(
                f"Scan folder override not found: {scan_folder}")
        recursive_scan = True
    else:
        # Optional custom loan source folder (absolute or relative), useful for external client folders.
        configured_loan_folder = config.get('loan_file_folder')
        if configured_loan_folder:
            scan_folder = resolve_path(configured_loan_folder)
            if not os.path.isdir(scan_folder):
                raise FileNotFoundError(f"Loan file folder not found: {scan_folder}")
        else:
            # Look in per-client subfolder first, then fallback to main Raw_Uploads
            client_upload = os.path.join(UPLOAD_FOLDER, client_name)
            if os.path.isdir(client_upload):
                scan_folder = client_upload
            else:
                scan_folder = UPLOAD_FOLDER

        recursive_scan = bool(config.get('loan_file_recursive', False))
    archive_imported = bool(config.get('archive_imported_files', True))
    if scan_folder_override:
        # Never sweep an ad-hoc override scan into the Archive.
        archive_imported = False

    client_archive = None
    if archive_imported:
        archive_dir = config.get('archive_directory')
        client_archive = resolve_path(archive_dir) if archive_dir else os.path.join(ARCHIVE_FOLDER, client_name)
        os.makedirs(client_archive, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"Processing: {cu_name}")
    print(f"Scanning: {scan_folder}")
    print(f"Recursive scan: {'Yes' if recursive_scan else 'No'}")
    print(f"Archive imported files: {'Yes' if archive_imported else 'No'}")
    print(f"{'='*60}")

    # Load credit pull scores (current FICO) before processing loan files
    credit_pull_scores, warm_acct_scores, pull_as_of = load_credit_pull_scores(config)
    if pull_as_of is not None:
        print(f"  Credit-pull as-of date: {pd.Timestamp(pull_as_of).date()}")

    files_to_process = []
    if recursive_scan:
        for root, _, files in os.walk(scan_folder):
            for filename in files:
                files_to_process.append((root, filename))
    else:
        for filename in os.listdir(scan_folder):
            files_to_process.append((scan_folder, filename))

    files_to_process.sort(key=lambda x: os.path.relpath(os.path.join(x[0], x[1]), scan_folder).lower())

    files_processed = 0
    # Per-snapshot wholesale-wipe tracker (see ``import_file`` for full
    # rationale). On full-folder imports this is an empty set we mutate
    # as each new snapshot is encountered; the first file for a given
    # snapshot wipes everything for (cu, snap) then APPENDs, and
    # subsequent files for the same snapshot just APPEND so multiple
    # extracts contributing overlapping ``loan_pool`` values all survive.
    # On specific-file imports (``--file=X.xlsx``) we pass ``None`` to
    # preserve the legacy per-pool delete semantics so re-running a
    # single file does not nuke other extracts' rows for the same snap.
    wiped_snapshots: set[str] | None = None if specific_file else set()
    for root, filename in files_to_process:
        file_path = os.path.join(root, filename)
        if not os.path.isfile(file_path):
            continue
        if filename.startswith("~$"):
            continue

        relative_file = os.path.relpath(file_path, scan_folder)
        if specific_file and filename != specific_file and relative_file != specific_file:
            continue

        # Route this file to a per-file extract (column_mappings,
        # member_account, has_header overrides) when one matches. Falls
        # back to the top-level mapping for back-compat.
        per_file_cfg = config
        matched_extract = None
        for pats, extract in extracts:
            if _patterns_match_any(pats, filename, relative_file):
                matched_extract = extract
                break
        if matched_extract is not None:
            per_file_cfg = dict(config)
            per_file_cfg['column_mappings'] = dict(
                matched_extract.get('column_mappings') or {}
            )
            ma_over = matched_extract.get('member_account')
            if ma_over:
                per_file_cfg['member_account'] = dict(ma_over)
                if (ma_over.get('mode') == 'fixed_suffix'
                        and ma_over.get('suffix_length') is not None):
                    per_file_cfg['account_suffix_length'] = int(
                        ma_over.get('suffix_length') or 0
                    )
            if 'has_header' in matched_extract:
                per_file_cfg['has_header'] = bool(
                    matched_extract.get('has_header')
                )
            # Per-file header_row override (1-indexed). Lets AIRES-style
            # extracts (row 1 = position numbers, row 2 = real headers)
            # coexist with conventional extracts in the same client.
            if 'header_row' in matched_extract:
                try:
                    per_file_cfg['header_row'] = int(
                        matched_extract.get('header_row') or 0
                    )
                except (TypeError, ValueError):
                    per_file_cfg['header_row'] = 0
            # Phase 9.22: per-extract ``pool_code_split`` override. When
            # present on the matched extract (including ``""`` meaning
            # "no split"), it wins over the CU-level value. CUMA-style
            # mortgage files ship loan codes like ``15/15 ARM`` where
            # ``/`` is part of the code; without an explicit per-file
            # ``""`` here, ``map_pool_codes`` would truncate to ``15`` and
            # route to the default pool (typically ``Ignore``), silently
            # excluding those loans from the report.
            if 'pool_code_split' in matched_extract:
                per_file_cfg['pool_code_split'] = (
                    matched_extract.get('pool_code_split') or ''
                )
            # Per-extract ``pool_map`` override. Some CUs ship a second loan
            # extract whose loan-type codes use a *different* numbering scheme
            # that collides with the CU-level codes (e.g. code ``10`` means
            # one pool in the legacy AIRES file and another in a newer "v2"
            # export). When present, the extract's map fully replaces the
            # CU-level ``pool_map`` for that file so the schemes stay isolated.
            if matched_extract.get('pool_map'):
                per_file_cfg['pool_map'] = dict(matched_extract.get('pool_map'))
            # Per-extract charge-off exclusion column (name or index). Routes
            # rows with a non-zero charge-off amount to the Exclude pool.
            if matched_extract.get('chargeoff_exclude_column') not in (None, ''):
                per_file_cfg['chargeoff_exclude_column'] = (
                    matched_extract.get('chargeoff_exclude_column')
                )
            label_txt = matched_extract.get('label') or '(unlabeled)'
            print(f"    Using extract mapping: {label_txt}")
        elif extracts:
            # We have per-file extracts configured but none match this
            # file. Top-level pattern is the global catch-all; if the
            # file matched that we'd still be here. Skip with a warning
            # rather than risk mis-mapping with the first extract's
            # columns.
            if not _patterns_match_any(top_pattern_res, filename, relative_file):
                continue
            print(f"    WARNING: {filename} matched top-level file_pattern but "
                  "no loan_data_extracts entry. Using top-level mapping as "
                  "fallback.")

        if (not _patterns_match_any(top_pattern_res, filename, relative_file)
                and matched_extract is None):
            continue

        print(f"\n  File: {relative_file}")

        date_source = relative_file if str(config.get('date_source', 'filename')).lower() == 'path' else filename
        snapshot_date = extract_snapshot_date(date_source, config)
        if not snapshot_date and date_source != relative_file:
            # Fallback: allow a filename regex to match a dated folder segment in recursive paths.
            snapshot_date = extract_snapshot_date(relative_file, config)
        if not snapshot_date:
            # Final fallback: filename has no parseable date but the config
            # has an explicit ``report_period``. Two cases handled:
            #   (a) filename mentions the report_period's own month — use
            #       that month-end (e.g. ``December Loan File - Upload``
            #       against ``report_period: 2025-12``).
            #   (b) filename mentions ANY month name (Jan-Dec) — use that
            #       month with the report_period's YEAR. Covers monthly
            #       drops within a quarterly run (e.g. ``February Loan
            #       File Upload`` and ``March Loan File Upload`` staged
            #       alongside the Jan file when running ``2026-03``).
            # The report_period month is preferred if present, falling
            # through to the first other month name otherwise.
            rp = str(config.get('report_period') or '').strip()
            rp_match = re.match(r'^(20\d{2})-(0[1-9]|1[0-2])$', rp)
            if rp_match:
                rp_year = int(rp_match.group(1))
                rp_month = int(rp_match.group(2))
                src_lower = str(date_source).lower()
                candidate_month: int | None = None
                rp_names = (
                    calendar.month_name[rp_month].lower(),
                    calendar.month_abbr[rp_month].lower(),
                )
                if any(name and re.search(rf'(?<![a-z]){name}(?![a-z])', src_lower)
                       for name in rp_names if name):
                    candidate_month = rp_month
                else:
                    for m in range(1, 13):
                        if m == rp_month:
                            continue
                        names = (
                            calendar.month_name[m].lower(),
                            calendar.month_abbr[m].lower(),
                        )
                        if any(name and re.search(rf'(?<![a-z]){name}(?![a-z])', src_lower)
                               for name in names if name):
                            candidate_month = m
                            break
                if candidate_month is not None:
                    last_day = calendar.monthrange(rp_year, candidate_month)[1]
                    snapshot_date = date(rp_year, candidate_month, last_day).isoformat()
                    matched_name = calendar.month_name[candidate_month].lower()
                    print(f"    Fallback: matched month name '{matched_name}' to "
                          f"report_period year {rp_year} -> {snapshot_date}")
                else:
                    # (c) The filename carries no date and no month name at
                    # all (e.g. "CECLOE.xls"). An undated snapshot file in
                    # the current upload belongs to the period being run, so
                    # stamp it to the report_period month-end. Gated on
                    # report_period being set, and only reached after every
                    # date-parsing attempt above has failed.
                    last_day = calendar.monthrange(rp_year, rp_month)[1]
                    snapshot_date = date(rp_year, rp_month, last_day).isoformat()
                    print(f"    Fallback: no date in filename; using "
                          f"report_period -> {snapshot_date}")
        if not snapshot_date:
            print(f"    SKIPPED: Could not extract date from filename")
            continue

        try:
            count = import_file(file_path, per_file_cfg, snapshot_date, credit_pull_scores, warm_acct_scores, pull_as_of, wiped_snapshots=wiped_snapshots)
            if count > 0:
                if archive_imported and client_archive:
                    archive_target = os.path.join(client_archive, relative_file)
                    os.makedirs(os.path.dirname(archive_target), exist_ok=True)
                    shutil.move(file_path, archive_target)
                print(f"    SUCCESS: Imported {count} loans for {snapshot_date}")
                log_data_import(client_name, cu_name, file_path, count, success=True)
                files_processed += 1
            else:
                print(f"    WARNING: No valid loan records found")
                log_data_import(client_name, cu_name, file_path, 0, success=False)
        except Exception as e:
            print(f"    ERROR: {e}")
            log_data_import(client_name, cu_name, file_path, 0, success=False)

    if files_processed == 0:
        print(f"\n  No new files found to import.")
    else:
        print(f"\n  Imported {files_processed} file(s).")

    return files_processed


def main():
    parser = argparse.ArgumentParser(description="Import loan data for CECL analysis")
    parser.add_argument('--client', help='Client config name (e.g., "ontario")')
    parser.add_argument('--file', help='Specific filename to import')
    parser.add_argument('--all', action='store_true', help='Process all configured clients')
    parser.add_argument('--list', action='store_true', help='List available client configs')
    args = parser.parse_args()

    if args.list:
        print("Available clients:")
        for c in list_clients():
            cfg = load_client_config(c)
            print(f"  {c:20s} -> {cfg['credit_union']}")
        return

    if args.all:
        log_session_start('import_data.py', '--all')
        for client_name in list_clients():
            process_client(client_name)
        log_session_end('import_data.py')
    elif args.client:
        log_session_start('import_data.py', f'--client {args.client} --file={args.file}')
        process_client(args.client, args.file)
        log_session_end('import_data.py')
    else:
        parser.print_help()
        print("\nAvailable clients:", ', '.join(list_clients()))


if __name__ == '__main__':
    main()