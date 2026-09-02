"""New-CU setup wizard (CECL Migration Model).

A multi-step form. Each step writes its data into Flask's session under the
key ``setup_state``; the final step builds and saves the YAML.

Steps:
  1. Identity        -- CU name, short name, state, county, optional data dir,
                        and whether the user has Manual WARM workbook(s)
  2. WARM Upload     -- (if yes) upload the latest CECL-Migration-WARM file;
                        seed pools / grades / history range from it
  3. Historical Data -- (if no WARM) scan a folder or upload prior WARM
                        workbooks, charge-off tracking, recovery, and
                        impaired-loan files for the WARM -> CM comparison
  4. Sample File     -- upload one quarterly file; auto-suggest mappings/pools
  5. Pools           -- pool_map (raw code -> pool name), default_pool, split char
                        (seeded from WARM Grade Ranges & Loan Codes cols S/T,
                        or from a user-uploaded Loan Code Map file)
  6. Files           -- file_pattern, date_pattern, account_suffix_length
  7. Columns         -- column_mappings (member#, balance, FICO, pool code, ...)
  8. Grades          -- credit_grades (label, min/max FICO, reserve_rate)
  9. Credit Pull     -- credit_pull file pattern + fallback CECL-report tabs
 10. Economic Data   -- state/county + unemployment/foreclosures/etc.
 11. Mgmt Adj        -- ltv_baseline, probability_factor
 12. Reports         -- which Excel reports to generate (TCT/Vizo/Vizo-Supp)
 13. Review & Save   -- show YAML preview, confirm, write file
"""
from __future__ import annotations

import re
import shutil
import tempfile
from pathlib import Path
from typing import Any

from flask import (
    Blueprint, current_app, flash, jsonify, redirect, render_template, request,
    session, url_for,
)
from werkzeug.utils import secure_filename

from cecl_ui.services import balance_check as balance_check_service, chargeoff_hist_processor, co_recov_parser, column_mapping_suggestions, config_service, cu_locator, delinquency_hist_processor, dq_extract_parser, extract_hist_processor, extract_hist_service, geo_service, hist_parser, impaired_parser, monthly_bal_parser, monthly_co_recov_aggregator, pipeline_service, recovery_hist_processor, risk_tier_parser, sample_parser, solr_5300_backfill, solr_5300_co_backfill, solr_5300_delq_backfill, solr_5300_recov_backfill, warm_parser, wizard_drafts
from cecl_ui.services import admin_defaults


setup_bp = Blueprint("setup", __name__)

STATE_KEY = "setup_state"

# Steps shown when the CU has a WARM file (history comes from WARM, skip historical step)
WIZARD_STEPS_WARM = [
    ("identity",    "1. CU Identity"),
    ("warm",        "2. Loan Pools"),
    ("pools",       "3. Loan Code Mapping"),
    ("balances",    "4. Balance Titles"),
    ("baseline",    "5. Historical Balances"),
    ("dq_hist",      "6. Historical DQ"),
    ("monthly_bal", "7. Monthly Balance File"),
    ("grades",      "8. Credit Grades and Business Risk Ratings"),
    ("credit_pull",  "9. Credit Pull"),
    ("orig_score",  "10. Original Score Baseline"),
    ("sample",      "11. Loan Data Extract(s)"),
    ("columns",     "12. Column Mappings"),
    ("balance_check","13. Balance Adjustment"),
    ("co_recov",     "14. Charge-Offs & Recoveries"),
    ("impaired",     "15. Impaired Loans"),
    ("files",        "16. File Format"),
    ("economic",     "17. Economic Data"),
    ("mgmt_adj",     "18. Mgmt Adjustments"),
    ("reports",      "19. Reports"),
    ("review",       "20. Review & Save"),
]

# Steps shown when the CU has no WARM file (historical data uploaded separately)
WIZARD_STEPS_NO_WARM = [
    ("identity",    "1. CU Identity"),
    ("loan_pools",  "2. Loan Pools"),
    ("pools",       "3. Loan Code Mapping"),
    ("historical",  "4. Historical Balances"),
    ("co_history",   "5. Historical Charge-Offs"),
    ("recov_history","6. Historical Recoveries"),
    ("dq_hist",      "7. Historical DQ"),
    ("monthly_bal", "8. Monthly Balance File"),
    ("grades",       "9. Credit Grades and Business Risk Ratings"),
    ("credit_pull",  "10. Credit Pull"),
    ("orig_score",  "11. Original Score Baseline"),
    ("sample",      "12. Loan Data Extract(s)"),
    ("columns",     "13. Column Mappings"),
    ("balance_check","14. Balance Adjustment"),
    ("co_recov",     "15. Charge-Offs & Recoveries"),
    ("impaired",     "16. Impaired Loans"),
    ("files",        "17. File Format"),
    ("economic",     "18. Economic Data"),
    ("mgmt_adj",     "19. Mgmt Adjustments"),
    ("reports",      "20. Reports"),
    ("review",       "21. Review & Save"),
]

# Steps shown when the user is in the CECL Simple (SCALE) model.
# Kept here (not in scale_setup) so _wizard_ctx can branch without an
# import cycle.
WIZARD_STEPS_SCALE = [
    ("identity",       "1. CU Identity"),
    ("scale_solr",     "2. Solr & Period"),
    ("scale_template", "3. Template & Map"),
    ("scale_qfactors", "4. Q-Factors"),
    ("scale_lol",      "5. Life of Loan Months"),
    ("scale_mgmt_adj", "6. Mgmt Adjustments"),
    ("scale_impaired", "7. Impaired Loans"),
    ("scale_review",   "8. Review"),
    ("scale_run",      "9. Run"),
]

# Default (identity not yet answered)
WIZARD_STEPS = WIZARD_STEPS_NO_WARM


def _state() -> dict[str, Any]:
    if STATE_KEY not in session:
        session[STATE_KEY] = _default_state()
    _strip_legacy_pool_map_stubs(session[STATE_KEY])
    # Self-heal: sessions/drafts saved before economic_data existed (or
    # adopted from a raw YAML config) can lack this key, which breaks the
    # Step 1 / Step 7 templates that read state.economic_data.*.
    ed = session[STATE_KEY].setdefault("economic_data", {})
    for _k, _v in _default_state()["economic_data"].items():
        ed.setdefault(_k, _v)
    return session[STATE_KEY]


# Legacy `_default_state()` (pre-2026-06-09) pre-seeded `pool_map` with
# these 7 example code->pool pairs. They were intended as
# illustrative placeholders but lingered in every CU's session/draft and
# surfaced as "Unrecognized pool names" on Step 3 even when the CU's
# real data contained no such codes. _default_state() now starts with
# an empty pool_map; this shim cleans up sessions and drafts saved
# before that change so users don't have to manually Ignore the stubs.
_LEGACY_POOL_MAP_STUBS = {
    "AUTO_NEW": "New Vehicle",
    "AUTO_USED": "Used Vehicle",
    "MORT": "Mortgage Loans",
    "CC": "Credit Cards",
    "SIG": "Signature Loans",
    "HELOC": "HELOC",
    "SECURED": "Share Secured",
}


def _strip_legacy_pool_map_stubs(state: dict[str, Any]) -> bool:
    """Remove the 7 hard-coded legacy `pool_map` example entries — but ONLY
    when the pool_map still looks like a pristine default (contains
    nothing but legacy stub keys, each still holding the legacy default
    value). Once the user has added any other loan code, or edited any
    stub's pool name, the pool_map is treated as real user data and
    left alone forever (via the ``_legacy_stubs_purged`` sentinel).

    This prevents a very sharp edge case: several of the legacy stub
    keys (notably ``CC`` -> ``Credit Cards``, ``HELOC`` -> ``HELOC``,
    ``MORT`` -> ``Mortgage Loans``) are ALSO the natural mapping any
    real credit union would enter. Without the sentinel, saving one of
    those mappings would silently strip it on the very next page render
    and every subsequent validator would flag the code as unmapped —
    even though the disk draft still shows the correct save.

    Safe to call on every request. Returns True when changes were made.
    """
    pm = state.get("pool_map")
    if not isinstance(pm, dict) or not pm:
        return False
    # Once we've decided this draft has real user data, never touch it.
    if state.get("_legacy_stubs_purged"):
        return False
    # If ANY entry is not an untouched legacy stub, treat the whole
    # pool_map as user data and mark it so we never re-check.
    for k, v in pm.items():
        if k not in _LEGACY_POOL_MAP_STUBS or _LEGACY_POOL_MAP_STUBS[k] != v:
            state["_legacy_stubs_purged"] = True
            return False
    # pool_map is exclusively unmodified legacy stubs — safe to purge.
    removed = False
    for k in list(_LEGACY_POOL_MAP_STUBS.keys()):
        if pm.pop(k, None) is not None:
            removed = True
    state["_legacy_stubs_purged"] = True
    return removed


def _default_state() -> dict[str, Any]:
    return {
        # identity
        "credit_union": "",
        "short_name": "",
        "charter_number": "",
        "data_directory": "",
        # Optional raw-data folder pasted by the user on Step 1
        # (Identity). When non-blank the wizard's auto-scan walks
        # this path with cecl_ui.services.auto_setup to populate as
        # much of the state as it can BEFORE the user touches any
        # downstream step. Persisted so the user can re-scan after
        # adding more files to the folder.
        "raw_data_folder": "",
        # Multiple raw-data folder paths (WARM history, client uploads,
        # credit pulls, ...). The wizard classifies + merges them and
        # decides how to set the CU up. ``raw_data_folder`` mirrors the
        # first path for single-folder self-heal reuse.
        "raw_data_folders": [],
        # Optional per-folder type hints, parallel to ``raw_data_folders``
        # ("" = auto-detect). Informational only — the classifier still
        # decides file types by content/name.
        "raw_data_folder_types": [],
        # Report period (YYYY-MM) the user is configuring this CU to
        # run reports for. Acts as the global "as-of" anchor: any
        # source file / column whose period is AFTER this is ignored
        # (e.g. Apr-Dec columns in a Jan-Dec balance sheet when the
        # report period is 2026-03). Blank means "use the latest
        # available data".
        "report_period": "",
        # files
        "file_pattern": r"LOANDATA.*\.(xlsx|xls|csv)$",
        "date_pattern": r"(\d{4})-(\d{2})",
        # ``date_format`` tells ``import_data.extract_snapshot_date`` how
        # to interpret the regex capture groups. Must be one of the values
        # in ``sample_parser.ACCEPTED_DATE_FORMATS`` (currently:
        # YYYY-MM, YYYYMMDD, MMDDYY, MMYY, MMYYYY, YYYYQ, QYYYY).
        # Anything else silently falls through to YYYY-MM and will yield
        # a bogus snapshot date at import time.
        "date_format": "YYYY-MM",
        "account_suffix_length": 3,
        # How the loan-data extract represents member + account numbers.
        # mode = "fixed_suffix" -> member & account share a column; the last
        #                          `suffix_length` chars are the account suffix.
        # mode = "delimiter"    -> member & account share a column, separated
        #                          by `delimiter` (e.g. "-", "L").
        # mode = "split"        -> member and account live in two separate
        #                          columns. column_mappings["member_number"]
        #                          holds the member column, column_mappings
        #                          ["loan_suffix"] holds the account/suffix
        #                          column.
        "member_account": {
            "mode": "fixed_suffix",
            "suffix_length": 3,
            "delimiter": "-",
        },
        # Whether real loan-data extracts include a header row. The Sample
        # step has its own per-sample `state.sample.has_header` toggle that
        # controls how the wizard parses the uploaded sample file; this
        # top-level flag is what gets written to the generated YAML and used
        # by the importer at run-time.
        "has_header": True,
        # columns
        "column_mappings": {
            "member_number": "MEMBER_ID",
            "loan_suffix": "",
            "current_balance": "BALANCE",
            "original_fico_score": "FICO_SCORE",
            "current_fico_score": "",
            "loan_pool_code": "LOAN_TYPE",
            "days_delinquent": "DQ_DAYS",
            "interest_rate": "INT_RATE",
            "open_date": "OPEN_DATE",
            "original_loan_amount": "ORIG_AMT",
            "total_available_credit": "",
        },
        # pools
        #
        # Empty by default. The wizard auto-populates this from the
        # uploaded sample file (one row per distinct loan_pool_code) on
        # the Sample step, and the WARM-detection branch on Step 2 seeds
        # it from the WARM workbook's pool map. Pre-seeding with example
        # codes here would pollute every new CU's Loan Code Mapping step
        # with stub rows ("AUTO_NEW", "MORT", ...) that don't exist in
        # the CU's real data.
        "pool_map": {},
        "default_pool": "Ignore",
        "pool_code_split": "/",
        # grades
        "credit_grades": [dict(g) for g in config_service.DEFAULT_CREDIT_GRADES],
        "no_score_label": "Not Reported",
        # credit pull
        "credit_pull": {
            "file_pattern": "Credit Pull.*\\.xlsx$",
            "source_folder": "",
            "member_column": "Member Number",
            "score_column": "FICO",
            "fallback_report_pattern": "CECL-Migration.*\\.xlsx$",
            "fallback_report_folder": "",
            "fallback_sheet_pattern": "Credit Pull",
            "fallback_member_col": 0,
            "fallback_score_col": 1,
            # When True, the fallback_report_folder/pattern are auto-populated
            # from the configured WARM/CECL report (data_directory + file_pattern
            # from Step 3) rather than entered manually.
            "use_configured_report": False,
            # When True, Option 1 (standalone file) is active.
            "use_standalone_file": False,
            # Last uploaded standalone credit-pull file (display only).
            "uploaded_filename": "",
            # Use original-score column for loans opened after the credit pull
            # (or when no credit pull is configured at all).
            "prefer_original_for_new_loans": True,
            # Optional override for the credit-pull as-of date (YYYY-MM-DD).
            # If blank, the engine will infer it from the file mtime.
            "pull_as_of_date": "",
        },
        # Original Credit Score baseline (one-time upload). Used by the
        # importer to fill in ``original_fico_score`` for loans whose
        # monthly extract doesn't include an original score (e.g. VISA
        # credit-card files). Shape::
        #   {
        #     "saved_path":        str,    # where the uploaded file lives
        #     "uploaded_filename": str,    # display name
        #     "member_column":     str,    # header for the member# col
        #     "suffix_column":     str,    # optional header for the suffix col
        #     "score_column":      str,    # header for the original FICO col
        #     "rows": [                    # parsed lookup rows
        #         {"member": str, "suffix": str, "score": int},
        #     ],
        #     "row_count":         int,
        #     "preview":           [[...]],# first few rows for the UI
        #     "headers":           [str],  # detected column headers
        #   }
        "orig_score_baseline": {
            "saved_path": "",
            "uploaded_filename": "",
            "member_column": "",
            "suffix_column": "",
            "score_column": "",
            "rows": [],
            "row_count": 0,
            "preview": [],
            "headers": [],
            # Names of loan pools the baseline should apply to. Empty list
            # means "all pools" (legacy behaviour). Used to scope the
            # importer's fill so a Visa baseline doesn't bleed into other
            # pools.
            "pools": [],
        },
        # economic
        "economic_data": {
            "state": "",
            "county": "",
            "unemployment_rate": 0.04,
            "foreclosures": 0,
            "bankruptcies": 0,
            "population": 0,
        },
        # mgmt adj
        "mgmt_adj": {"ltv_baseline": 0.9, "probability_factor": 0.35},
        # Per-pool overlay added to the base loss rate (decimal).
        # e.g. {"New Auto": 0.005} adds 0.50% to the base rate for New Auto.
        "mgmt_adj_by_pool": {},
        # balance fmt
        "balance_remove_chars": ["$", ","],
        "accounting_negatives": True,
        # reports
        "reports": {
            "tct": True,
            "vizo": False,
            "vizo_supp": False,
            "impdet": False,
            "mgmt_adj_napkin": True,
            "acl_funding": True,
            "vizo_pdf": False,
        },
        # sample file analysis (populated by sample step)
        "sample": None,  # type: dict | None
        # WARM workbook analysis (populated by WARM step)
        "has_warm_files": None,   # "yes" | "no" | None (unanswered)
        "warm": None,             # type: dict | None
        # Historical data scan result (populated by step 3 historical)
        "hist_scan": None,        # type: dict | None
        # Historical step — which source the CU is providing for the
        # monthly historical loan balances:
        #   "single_workbook"        — one spreadsheet with all months
        #                              (the original behavior)
        #   "monthly_loan_extracts"  — one loan-data extract per month
        #                              (e.g. AIRES file)
        #   "monthly_balance_sheets" — one balance sheet per month
        "hist_balance_source": "single_workbook",
        # Historical step — explicit, display-only provenance record for the
        # monthly historical balances, so a validator can see HOW they were
        # compiled. When set (e.g. by a setup script that derived balances
        # from the imported loan-data extracts), it is authoritative for the
        # "How these balances were compiled" panel; otherwise the panel
        # derives a best-effort description from ``hist_balance_source``.
        # Shape: {"method": str, "summary": str, "inputs": [str, ...],
        #         "coverage": {"start","end","months","pools"},
        #         "generated_by": str, "validation_hint": str,
        #         "generated_at": "YYYY-MM-DD..."}
        "hist_balance_provenance": {},
        # Historical step — state for the "monthly loan-data extracts" source.
        # Populated when hist_balance_source == "monthly_loan_extracts".
        #
        # Shape::
        #   {
        #     "target_period":  "YYYY-MM-DD",  # last day of report month
        #     "history_months": 84,             # how far back to scan
        #     "folder_path":    "",             # CU's historical extracts folder
        #     "anchor_files":   [               # anchor month uploads
        #         {"name": str, "path": str,
        #          "signature": str,            # header-row signature
        #          "detected_date": str,        # YYYY-MM-DD or ""
        #          "detected_source": str,      # filename | mtime | ""
        #          "detected_confidence": str,  # high | medium | low | none
        #          "override_date": str,        # user-entered, beats detected
        #          "profile_id":    str},       # which profile this file uses
        #     ],
        #     "profiles": [                     # one per distinct header layout
        #         {"id": str, "signature": str, "headers": [str, ...],
        #          "column_mappings": {field: header},  # field-> source header
        #          "label": str},                # user-friendly name
        #     ],
        #     "scan_results": None,             # folder-scan output (later)
        #   }
        "hist_extracts": {
            "target_period": "",
            "history_months": 84,
            "folder_path": "",
            "anchor_files": [],
            "profiles": [],
            "scan_results": None,
            # Signatures the user has explicitly ignored. Files matching
            # any of these layouts are excluded from the unmapped-layouts
            # panel and skipped during Process & Save.
            "ignored_signatures": [],
            # NCUA 5300 backfill — Option 2 of 3 for filling missing
            # quarter-end months in loan_code_history (Option 1: real
            # extracts via Process & Save above; Option 3: manual entry).
            "solr_backfill": {
                "solr_url": "http://searchserver1.tctrisk.com:8983/solr",
                "core": "ncua",
                # {loan_code: comma-separated 5300 field codes}
                "loan_code_fields": {},
                # "distributed" (default): fetch the QUARTER's TOTAL
                # loan balance and distribute it across user pools in
                # the same % ratios as the EARLIEST month of the
                # historical balance file(s) uploaded above. This is
                # the right answer for almost every CU we onboard.
                # "per_loan_code": one row per canonical 5300 bucket,
                # downstream loader maps NCUA code -> user pool via
                # pool_map / NCUA classifier. Only used in special
                # cases where the 5300 categories line up cleanly with
                # this CU's pools.
                "mode": "distributed",
                # When True (the default), the wizard automatically runs
                # the 5300 balance backfill once, when the user leaves the
                # Historical Balances step, so filling early quarters from
                # the 5300 is the out-of-the-box behavior. The analyst can
                # untick it on that step to opt out, and the manual "Run"
                # button still works either way.
                "enabled": True,
                "auto_ran": False,
                "last_test": None,
                "last_run": None,
            },
        },
        # Historical step — optional recoveries flag (non-WARM path)
        "no_hist_recoveries": False,
        # Historical step — source choice for charge-offs / recoveries.
        # One of: "single_workbook" (default — one file w/ all history),
        # "monthly_files" (one CO/recov file per month), or
        # "5300_backfill" (auto-pull from NCUA Solr).
        "hist_co_source": "single_workbook",
        "hist_recov_source": "single_workbook",
        # Historical step — pool mapping detected/edited for the uploaded
        # Historical Month-End Balances file. Shape::
        #   {
        #     "sheet": str,                   # sheet read from the workbook
        #     "labels": [str, ...],           # distinct col-A labels detected
        #     "mapping": {label: pool_name},  # user-confirmed (or pre-filled)
        #     "source": "warm" | "manual" | "",
        #   }
        "hist_pool_map": None,
        # Per-pool config table (seeded from WARM "BS CO DQ Data Enter").
        # List of dicts, in display/report order::
        #   {"name": str,
        #    "risk_rated": bool,
        #    "acl_months": int,
        #    "use_default_mgmt_adj": bool,
        #    "excluded": bool}   # excluded pools are dropped from the
        #                          analysis entirely (not in single-line, not
        #                          in grade breakouts, not in totals).
        "pool_settings": [],
        # Per-pool manual carry-forward adjustments captured on the
        # Balance Adjustment step. Keyed by canonical pool name ->
        # {"amount": float, "note": str}.
        "balance_adjustments": {},
        # ACL Amount captured (or entered manually) for the
        # "Allowance for Credit Loss" line on Step 2. Single dollar value.
        "acl_balance": 0.0,
        # Business Risk Rating (Step 6 — Credit Grades).
        # When uses_brr is True the user can list rating buckets the way
        # they list credit grades. Each entry is {"label": str, "criteria": str}.
        "uses_brr": False,
        "business_risk_ratings": [],
        # "Other Allowance Considerations" — optional ACL overlays the user
        # adds in Step 2. Each row::
        #   {"title": str, "balance": float, "percentage": float}
        # The product (balance * percentage) is computed at render/save time.
        "include_other_allowance": False,
        "other_allowance_considerations": [],
        # Data-derived "Negative Share Provision" overlay (Step 8). When
        # enabled, the report engine computes a life-of-loan loss rate for the
        # CU's negative shares (trailing N months of net charge-offs / current
        # negative-share balance) and applies it as an Other Allowance
        # Consideration. Reusable across CUs — only the file patterns/folder
        # differ. See generate_report._resolve_negative_share_oac.
        "include_negative_share": False,
        "negative_share": {
            "title": "Negative Share Provision",
            "life_of_loan_months": 12,
            "source_folder": "",
            "balance_column": "Current Balance",
            "balance_pattern": r"(?i)Negative Share File",
            "co_summary_pattern": r"(?i)Negative Shares Charge Off and Recovery",
            "co_quarterly_pattern": r"(?i)Share COs?\s*-\s*Recoveries",
        },
        # Data-derived "Unfunded Commitments" overlay (Step 8). When enabled,
        # the report engine sums undrawn credit (credit limit - current
        # balance) for the loan type codes the CU flagged as NOT
        # unconditionally cancelable, grouped by pool, and applies each pool's
        # ACL loss rate. One "Unfunded Commitments - <Pool>" row per pool.
        # See generate_report._expand_unfunded_commitment_oac.
        "include_unfunded_commitment": False,
        "unfunded_commitment": {
            "title": "Unfunded Commitments",
            "loan_type_codes": [],
        },
        # Sample files step uploads and optional flags
        "sample_uploads": {
            "loan_balance_files": [],
            "loan_data_files": [],
            "co_files": [],
            "recov_files": [],
            "impaired_files": [],
            "credit_pull_files": [],
            "no_recoveries": False,
            "no_credit_pull": False,
            "loan_data_error": "",
        },
        # Monthly Balance File step (Step 5).
        # Captures the recurring quarterly file the CU sends with month-end
        # totals by pool/type. On the WARM path it fills the establishing
        # quarter's columns; on the no-WARM path it carries the full history.
        "monthly_bal": {
            "filename": "",
            "saved_path": "",
            "sheet": "",
            "header_row": 0,           # 1-based row containing month-end dates
            "pool_name_col": "A",      # column letter holding pool/type labels
            "first_date_col": "",      # column letter where the first date sits
            "parsed_dates": [],        # list of "YYYY-MM-DD" strings
            "parsed_pool_labels": [],  # list of distinct labels found in col
            "pool_map": {},            # raw label -> wizard pool name
            "file_pattern": "",        # glob/regex the importer should expect
            "notes": "",
        },
    }


# Action names that signal "user explicitly clicked Save & Next on this
# step" — i.e. they consider the step done. Catches the bare ``next``
# (used on most templates' Save & Next button) and any ``*_and_next`` /
# ``*_next`` variant (per-month / per-year layouts, manual ACL save,
# etc.). Plain "save" actions (Save Progress) are NOT included; they
# don't advance to the next step.
_SAVE_AND_NEXT_RX = re.compile(r"^(?:next|.+_(?:and_)?next)$")


def _mark_step_user_completed(state: dict[str, Any], step_key: str) -> None:
    """Record that the user explicitly finished *step_key* via Save & Next.

    The breadcrumb stepper consults ``state["_user_completed_steps"]``
    to render a green ✓ regardless of the current data-completion
    predicate, so a user who clicked Save & Next on a step sees it
    reflected as done even if a *recommended* HIL still applies.
    """
    if not step_key:
        return
    done = state.setdefault("_user_completed_steps", [])
    if not isinstance(done, list):
        done = []
        state["_user_completed_steps"] = done
    if step_key not in done:
        done.append(step_key)


# Provenance state keys that must never be silently dropped by a save from a
# session that predates them (see _preserve_provenance_records).
_PROVENANCE_KEYS = ("hist_balance_provenance", "co_recov_provenance")


def _provenance_record_has_content(key: str, rec: Any) -> bool:
    """True when ``rec`` is a populated provenance record for ``key``."""
    if not isinstance(rec, dict) or not rec:
        return False
    if key == "co_recov_provenance":
        for side in ("chargeoff", "recovery"):
            sub = rec.get(side) or {}
            if isinstance(sub, dict) and (sub.get("method") or sub.get("no_recoveries")):
                return True
        return False
    # hist_balance_provenance and future single-record keys.
    return bool(rec.get("method"))


def _preserve_provenance_records(state: dict[str, Any]) -> None:
    """Permanent guard against silently dropping provenance records.

    A save from an in-memory session that predates a provenance record (e.g.
    one written straight to the draft/config by a setup or backfill script)
    would otherwise overwrite the disk draft — and any regenerated YAML — with
    a state that no longer carries the record. Before persisting, restore any
    provenance record the live state is missing from the still-current on-disk
    draft, falling back to the saved YAML config. Preserve-only: never clears a
    record the session already holds, and only reads disk (no recompute), so it
    is safe to call on every save.
    """
    missing = [
        k for k in _PROVENANCE_KEYS
        if not _provenance_record_has_content(k, state.get(k))
    ]
    if not missing:
        return
    try:
        ws = current_app.config["WORKSPACE_ROOT"]
    except Exception:  # noqa: BLE001
        return
    slug = (state.get("short_name") or state.get("credit_union") or "").strip()
    if not slug:
        return
    # Source 1: the current on-disk draft (about to be overwritten).
    disk: dict[str, Any] | None = None
    try:
        disk = wizard_drafts.load_draft(
            ws, slug, model=state.get("model") or "migration"
        )
    except Exception:  # noqa: BLE001
        disk = None
    # Source 2: the saved YAML config (loaded lazily, only if still needed).
    saved: dict[str, Any] | None = None
    short = (state.get("short_name") or "").strip()
    for key in missing:
        if isinstance(disk, dict) and _provenance_record_has_content(key, disk.get(key)):
            state[key] = disk[key]
            continue
        if short:
            if saved is None:
                try:
                    saved = config_service.load_client_config(ws, short) or {}
                except Exception:  # noqa: BLE001
                    saved = {}
            if _provenance_record_has_content(key, saved.get(key)):
                state[key] = saved[key]


def _save_state(state: dict[str, Any]) -> None:
    # Detect a "Save & Next"-style POST and record the user's explicit
    # done-signal for the active step. This drives the green ✓ on the
    # breadcrumb. Safe to run on every save: only POST + matching
    # action triggers the mark.
    try:
        if request.method == "POST":
            action_val = (request.form.get("action") or "").strip()
            if action_val and _SAVE_AND_NEXT_RX.match(action_val):
                _mark_step_user_completed(state, state.get("_active_step", "") or "")
    except RuntimeError:
        # No active request context (e.g. background save) — skip.
        pass

    session[STATE_KEY] = state
    session.modified = True
    # Auto-save to disk on every step so wizard work survives a session
    # reset (e.g. clicking "+ New CU" or session expiry). Requires a
    # short_name/credit_union to have been entered — Step 1 sets these,
    # so anything past Step 1 auto-persists. Never raise: an auto-save
    # failure must not block the request.
    if not (state.get("short_name") or state.get("credit_union")):
        return
    # Permanent guard: don't let a stale session drop provenance records that
    # already exist on disk (draft or saved config) before we overwrite them.
    _preserve_provenance_records(state)
    try:
        wizard_drafts.save_draft(
            current_app.config["WORKSPACE_ROOT"],
            state,
            active_step=state.get("_active_step", "") or "",
            model=state.get("model") or "migration",
        )
    except Exception:  # noqa: BLE001
        pass


def _wizard_ctx(active: str) -> dict[str, Any]:
    st = _state()
    if st.get("model") == "scale":
        steps = WIZARD_STEPS_SCALE
    else:
        has_warm = st.get("has_warm_files")
        steps = WIZARD_STEPS_WARM if has_warm == "yes" else WIZARD_STEPS_NO_WARM
    # Remember the last step the user touched so Save / Resume can return
    # them to it.
    st["_active_step"] = active
    session.modified = True

    # Snapshot any non-blank pool_map assignments into a durable history
    # dict so a subsequent extract delete-and-re-add cycle (which
    # otherwise wipes state["pool_map"] via the empty-seed branch in
    # _apply_sample_to_state) can restore user assignments automatically.
    # Runs on EVERY wizard GET so history is captured before any
    # self-heal or seed path could clear pool_map. Idempotent.
    try:
        _snapshot_pool_map_history(st)
    except Exception:  # noqa: BLE001
        pass

    # Compute per-step status badges + HIL list when the user has run
    # the Step 1 auto-scan. Without an auto-scan, every step renders
    # as "pending" (no badges shown).
    step_status: dict[str, str] = {}
    hil_needs: list[dict[str, Any]] = []
    next_hil_key: str | None = None
    # YAML-schema → wizard-schema self-heal: drafts created via
    # ``home.adopt_config_to_completed`` carry the raw YAML config as
    # the draft payload (``pools`` / ``monthly_balance`` / ``acl`` /
    # ``historical_file_formats``). The wizard renders from a different
    # schema (``pool_settings`` / ``monthly_bal`` / ``acl_balance`` /
    # ``co_columns``), so without translation Step 2 shows zero pools
    # even though the YAML has every pool defined. Runs OUTSIDE the
    # ``_auto_scan_completed`` gate because adopted drafts skip auto-scan.
    try:
        from cecl_ui.services import auto_setup as _auto_setup_adopt
        _adopt_msgs = _auto_setup_adopt.selfheal_adopt_yaml_schema_into_state(st)
        if _adopt_msgs:
            _save_state(st)
            for _m in _adopt_msgs:
                flash(f"Recovered from adopted config: {_m}", "success")
    except Exception:  # noqa: BLE001
        pass
    # SCALE drafts have their own short, self-contained step flow and do
    # not use the Migration auto-scan / human-in-the-loop gate system, so
    # the badge/remaining-items computation (which is all Migration-model
    # checks) is skipped entirely for them.
    if (st.get("model") != "scale"
            and (st.get("_auto_scan_completed")
                 or (st.get("autoderive") or {}).get("ok"))):
        try:
            from cecl_ui.services import auto_setup
            # Proactive self-heal: stale drafts saved before the
            # Phase 9.4 bulk-stage code (or where only one side of a
            # combined CO+Recov classification got staged) get their
            # hist_scan.monthly_co_files / monthly_recov_files lists
            # refilled here. Fires on EVERY wizard GET so the badges on
            # Steps 5 and 6 reflect reality regardless of which step
            # the user is currently viewing.
            try:
                _heal_msgs = auto_setup.selfheal_hist_scan_from_folder(st)
                if _heal_msgs:
                    _save_state(st)
                    for _m in _heal_msgs:
                        flash(f"Auto-scan: {_m}", "success")
            except Exception:  # noqa: BLE001
                # Self-heal is best-effort; never break the wizard.
                pass
            # Flat pool-seed self-heal: drafts saved before the
            # flat-hist-bal pool seeding shipped have populated
            # parsed_pool_labels but empty pool_settings. Read-only on
            # state (no folder I/O) so it's cheap to run on every GET.
            try:
                _flat_msgs = auto_setup.selfheal_flat_pool_seed_from_state(st)
                if _flat_msgs:
                    _save_state(st)
                    for _m in _flat_msgs:
                        flash(f"Auto-scan: {_m}", "success")
            except Exception:  # noqa: BLE001
                pass
            # Single-hist-bal layout self-heal: when monthly_bal.source is
            # 'single' with a saved_path but parsed_pool_labels is empty
            # (e.g. label column was mis-elected on a Vizo-style workbook
            # with loan codes in col A and descriptive labels in col B),
            # re-run analyse_file to recover the correct layout.
            try:
                _layout_msgs = auto_setup.selfheal_single_hist_bal_layout(st)
                if _layout_msgs:
                    _save_state(st)
                    for _m in _layout_msgs:
                        flash(f"Auto-scan: {_m}", "success")
            except Exception:  # noqa: BLE001
                pass
            # Heuristic auto-mapping of balance labels to pool_settings
            # names. Targets workbook-source CUs without a WARM
            # balance_title_map: matches each parsed_pool_label to a
            # configured pool by keyword category (credit_card,
            # real_estate, share_secured, recreational, new_vehicle,
            # used_vehicle, unsecured, etc). Only fills BLANK entries
            # so user edits are preserved across refreshes.
            try:
                _automap_msgs = auto_setup.selfheal_auto_map_balance_labels(st)
                if _automap_msgs:
                    _save_state(st)
                    for _m in _automap_msgs:
                        flash(f"Auto-scan: {_m}", "success")
            except Exception:  # noqa: BLE001
                pass
            # Exact-match precedence: correct any non-blank monthly_bal
            # pool_map entries whose label case-insensitively equals a
            # configured pool name but is mapped to a DIFFERENT pool
            # (typically because an earlier keyword-heuristic pass
            # collapsed sibling pools like "New Auto"/"Indirect New
            # Auto" onto the same target). Runs on every GET so a stale
            # session that survived a prior bad Save is repaired
            # automatically on the next page view.
            try:
                _exact_msgs = auto_setup.selfheal_enforce_exact_match_pool_map(
                    st
                )
                if _exact_msgs:
                    _save_state(st)
                    for _m in _exact_msgs:
                        flash(_m, "info")
            except Exception:  # noqa: BLE001
                pass
            hil_needs = auto_setup.compute_hil_needs(st)
            need_keys = {n["step_key"] for n in hil_needs}
            # Severity-aware HIL set: any 'required' entry blocks the
            # green check; 'recommended'-only entries do not.
            required_keys = {
                n["step_key"] for n in hil_needs
                if str(n.get("severity") or "").lower() == "required"
            }
            report = st.get("_auto_scan_report") or {}
            step_complete = auto_setup.compute_step_completion(st)
            user_done_raw = st.get("_user_completed_steps") or []
            user_done: set[str] = set(
                k for k in user_done_raw if isinstance(k, str)
            )
            for key, _label in steps:
                # Priority:
                #   1. Required HIL → ⚠ amber (data integrity always wins).
                #   2. User explicitly clicked Save & Next on this step
                #      → green ✓ (their stated intent overrides any
                #      lingering recommended HIL).
                #   3. Step has data sufficient for completion AND no
                #      required HIL → green ✓.
                #   4. Else if step has any HIL entry → amber ⚠.
                #   5. Else if auto-scan filled it → green ✓ (legacy).
                #   6. Else pending.
                if key in required_keys:
                    step_status[key] = "hil"
                elif key in user_done:
                    step_status[key] = "auto"
                elif key in step_complete:
                    step_status[key] = "auto"
                elif key in need_keys:
                    step_status[key] = "hil"
                elif key in report:
                    step_status[key] = "auto"
                else:
                    step_status[key] = "pending"
            # Filter the HIL banner: hide steps whose badge is now ✓
            # (recommended-only items on completed steps no longer block
            # progress, so don't list them as "needs your input").
            hil_needs = [
                n for n in hil_needs
                if step_status.get(n["step_key"]) != "auto"
            ]
            next_hil_key = auto_setup.first_hil_step_key(st, steps)
            # If the original first-HIL step is now complete, advance to
            # the next still-flagged step (if any).
            still_hil = {n["step_key"] for n in hil_needs}
            if next_hil_key and next_hil_key not in still_hil:
                next_hil_key = None
                for key, _label in steps:
                    if key in still_hil:
                        next_hil_key = key
                        break
        except Exception:  # noqa: BLE001
            # Defensive: stepper rendering must never 500.
            step_status = {}
            hil_needs = []
            next_hil_key = None

    return {
        "steps": steps,
        "active": active,
        "state": st,
        "step_endpoints": STEP_ENDPOINTS,
        "step_status": step_status,
        "hil_needs": hil_needs,
        "next_hil_key": next_hil_key,
    }


# Map wizard step keys to their Flask endpoint names so the stepper can
# render clickable breadcrumbs.
STEP_ENDPOINTS: dict[str, str] = {
    "identity":      "setup.step1_identity",
    # SCALE wizard steps (resolved against scale_setup blueprint)
    "scale_solr":     "scale_setup.step_solr",
    "scale_template": "scale_setup.step_template_map",
    "scale_qfactors": "scale_setup.step_qfactors",
    "scale_lol":      "scale_setup.step_lol",
    "scale_mgmt_adj": "scale_setup.step_mgmt_adj",
    "scale_impaired": "scale_setup.step_impaired",
    "scale_review":   "scale_setup.step_review",
    "scale_run":      "scale_setup.step_run",
    "warm":          "setup.step2_warm",
    "loan_pools":    "setup.step_loan_pools",
    "balances":      "setup.step3_balances",
    "baseline":      "setup.step3_baseline",
    "historical":    "setup.step3_historical",
    "co_history":    "setup.step3a_co_history",
    "recov_history": "setup.step3b_recov_history",
    "monthly_bal":   "setup.step5_monthly_bal",
    "grades":        "setup.step5_grades",
    "sample":        "setup.step2_sample",
    "columns":       "setup.step3_columns",
    "pools":         "setup.step4_pools",
    "balance_check": "setup.step_balance_check",
    "co_recov":      "setup.step_co_recov",
    "dq_hist":       "setup.step_dq_hist",
    "impaired":      "setup.step_impaired",
    "files":         "setup.step2_files",
    "credit_pull":   "setup.step6_credit_pull",
    "orig_score":    "setup.step_orig_score",
    "economic":      "setup.step7_economic",
    "mgmt_adj":      "setup.step8_mgmt_adj",
    "reports":       "setup.step9_reports",
    "review":        "setup.step10_review",
}


# ---------- helpers for parsing form rows ----------

def _parse_kv_rows(form, key_prefix: str, val_prefix: str) -> dict[str, str]:
    """Combine paired form fields like key_0/val_0, key_1/val_1, ... into a dict."""
    keys = form.getlist(key_prefix)
    vals = form.getlist(val_prefix)
    out: dict[str, str] = {}
    for k, v in zip(keys, vals):
        k = (k or "").strip()
        v = (v or "").strip()
        if k:
            out[k] = v
    return out


# ---------- save-progress / resume-draft ----------

@setup_bp.route("/save-progress", methods=["POST"])
def save_progress():
    """Persist the in-progress wizard state to disk.

    Uses ``state['short_name']`` (or credit_union name) as the file key
    so each CU has its own draft. Returns to home unless the form
    supplies ``return_to=stay``, in which case it bounces back to the
    same step.
    """
    st = _state()
    active = (request.form.get("active_step") or st.get("_active_step") or "identity")
    if not (st.get("short_name") or st.get("credit_union")):
        flash("Add the credit union name on Step 1 before saving progress.", "error")
        ep = STEP_ENDPOINTS.get(active, "setup.step1_identity")
        return redirect(url_for(ep))
    try:
        path = wizard_drafts.save_draft(
            current_app.config["WORKSPACE_ROOT"], st, active_step=active,
            model=st.get("model") or "migration",
        )
    except Exception as exc:  # noqa: BLE001
        flash(f"Could not save progress: {exc}", "error")
        ep = STEP_ENDPOINTS.get(active, "setup.step1_identity")
        return redirect(url_for(ep))

    label = st.get("credit_union") or st.get("short_name")
    flash(f"Saved progress for {label} ({path.name}).", "success")

    if request.form.get("return_to") == "stay":
        ep = STEP_ENDPOINTS.get(active, "setup.step1_identity")
        return redirect(url_for(ep))
    # Clear the session so the user can start a new CU without bleed-over.
    session.pop(STATE_KEY, None)
    return redirect(url_for("home.index"))


@setup_bp.route("/resume/<model>/<key>", methods=["GET", "POST"])
@setup_bp.route("/resume/<key>", methods=["GET", "POST"])
def resume_draft(key: str, model: str = "migration"):
    """Load a saved draft into the session and jump to its last step.

    The ``<model>`` URL segment is optional for backward compatibility
    with old bookmarks; it defaults to ``migration``.
    """
    data = wizard_drafts.load_draft(
        current_app.config["WORKSPACE_ROOT"], key, model=model,
    )
    if not data:
        flash(f"No saved progress found for '{key}' ({model}).", "error")
        return redirect(url_for("home.index"))
    # Make sure the model flag in state matches the file we loaded so
    # the next Save round-trips back to the right file.
    if model == "scale":
        data["model"] = "scale"
    _strip_legacy_pool_map_stubs(data)
    session[STATE_KEY] = data
    session.modified = True
    active = (data.get(wizard_drafts.DRAFT_META_KEY) or {}).get("active_step") \
        or data.get("_active_step") or "identity"
    ep = STEP_ENDPOINTS.get(active, "setup.step1_identity")
    flash(f"Resumed {data.get('credit_union') or key}.", "success")
    return redirect(url_for(ep))


@setup_bp.route("/delete-draft/<model>/<key>", methods=["POST"])
@setup_bp.route("/delete-draft/<key>", methods=["POST"])
def delete_draft(key: str, model: str = "migration"):
    if wizard_drafts.delete_draft(
        current_app.config["WORKSPACE_ROOT"], key, model=model,
    ):
        flash(f"Deleted saved progress for '{key}' ({model}).", "success")
    else:
        flash(f"No saved progress found for '{key}' ({model}).", "error")
    return redirect(url_for("home.index"))


@setup_bp.route("/discard", methods=["POST"])
def discard_session():
    """Clear the in-memory wizard state (no disk effect)."""
    session.pop(STATE_KEY, None)
    flash("Cleared the in-progress session.", "success")
    return redirect(url_for("home.index"))


# =================================================================
# Step 1 — Identity
# =================================================================
@setup_bp.route("/", methods=["GET"])
def root():
    return redirect(url_for("setup.step1_identity"))


@setup_bp.route("/start", methods=["GET", "POST"])
def start_warm_choice():
    """Entry point for the CECL-Migration model.

    The old binary "upload a WARM workbook vs. start from scratch" gate has
    been replaced by a single Identity step that lets the user add one or
    more raw-data folder paths (WARM history, client uploads, credit pulls,
    ...) and let the wizard classify them and decide how to set the CU up —
    including whether a WARM workbook is available for the WARM -> CM
    comparison. This endpoint is kept so existing links resolve; it simply
    forwards to the Identity step.
    """
    # Start a brand-new Migration setup from a clean slate so a prior CU's
    # data can't bleed into this one.
    session.pop(STATE_KEY, None)
    session.modified = True
    return redirect(url_for("setup.step1_identity"))


@setup_bp.route("/step/identity", methods=["GET", "POST"])
def step1_identity():
    state = _state()
    if request.method == "POST":
        cu = request.form.get("credit_union", "").strip()
        sn = request.form.get("short_name", "").strip().lower()
        if not cu:
            flash("Credit union name is required.", "error")
        else:
            if not sn:
                sn = config_service.slugify(cu)
            state["credit_union"] = cu
            state["short_name"] = sn
            # NCUA charter number — digits only, optional.
            charter_raw = request.form.get("charter_number", "").strip()
            charter_clean = "".join(ch for ch in charter_raw if ch.isdigit())
            if charter_raw and not charter_clean:
                flash("Charter number must contain digits — saved as blank.", "info")
            state["charter_number"] = charter_clean
            # Global "report period" anchor (YYYY-MM). Used by Step 14
            # and other steps to ignore data past this period. Accept
            # blank (= "use latest available"); otherwise normalise to
            # ``YYYY-MM`` and validate. Anything unparseable -> blank
            # with an info flash, so a typo never silently locks the
            # user into a wrong anchor.
            rp_raw = (request.form.get("report_period") or "").strip()
            rp_norm = ""
            if rp_raw:
                rp_try = rp_raw.replace("/", "-")
                # Accept YYYY-MM and YYYY-MM-DD; keep only YYYY-MM.
                parts = rp_try.split("-")
                try:
                    y = int(parts[0])
                    m = int(parts[1])
                    if 2000 <= y <= 2100 and 1 <= m <= 12:
                        rp_norm = f"{y:04d}-{m:02d}"
                except (ValueError, IndexError):
                    pass
                if not rp_norm:
                    flash(
                        f"Report period '{rp_raw}' isn't a valid YYYY-MM "
                        "value -- saved as blank (use latest data).",
                        "info",
                    )
            state["report_period"] = rp_norm
            # NOTE: historical data directory is collected later in the
            # wizard (Files step), not on this Identity step.
            state["economic_data"]["state"] = request.form.get("state", "").strip()
            state["economic_data"]["county"] = request.form.get("county", "").strip()
            # Raw-data folders for auto-scan (Step 1 paste fields). The user
            # can add multiple paths (e.g. a WARM-history folder, a client
            # data-upload folder, a credit-pull folder), optionally tag each
            # with a type hint, and let the wizard decide how to set the CU
            # up. Accept both the new multi-value ``raw_data_folders`` field
            # and the legacy single ``raw_data_folder`` field.
            _raw_paths = request.form.getlist("raw_data_folders")
            _raw_types = request.form.getlist("raw_data_folder_types")
            _pairs: list[tuple[str, str]] = []
            for _i, _p in enumerate(_raw_paths):
                _pn = _normalize_folder_path(_p)
                if not _pn:
                    continue
                _lbl = (_raw_types[_i] if _i < len(_raw_types) else "").strip()
                _pairs.append((_pn, _lbl))
            _single = _normalize_folder_path(
                request.form.get("raw_data_folder", "") or ""
            )
            if _single:
                _pairs.append((_single, ""))
            # De-dupe by path (case-insensitive), keeping the first label,
            # preserving order.
            _seen: set[str] = set()
            _folders: list[str] = []
            _ftypes: list[str] = []
            for _p, _l in _pairs:
                if _p.lower() in _seen:
                    continue
                _seen.add(_p.lower())
                _folders.append(_p)
                _ftypes.append(_l)
            state["raw_data_folders"] = _folders
            state["raw_data_folder_types"] = _ftypes
            # Keep the legacy single-folder key populated (first path) so
            # single-folder self-heal reuse keeps working.
            state["raw_data_folder"] = _folders[0] if _folders else ""
            _save_state(state)

            # --- Auto-scan branch ---------------------------------------
            # When the user clicks "Scan & auto-fill" we run the folder
            # classifier + populater, then redirect them to the first
            # step that still needs human input. Identity save above
            # already happened so short_name / charter / economic
            # fields are persisted before the scan runs.
            action = (request.form.get("action") or "").strip().lower()
            if action == "scan_folder":
                # SCALE draws its data from the NCUA 5300 (Solr), not a
                # raw-data folder scan, so skip the Migration auto-scan
                # entirely and go straight to the SCALE Solr step.
                if state.get("model") == "scale":
                    _save_state(state)
                    return redirect(url_for("scale_setup.step_solr"))
                folders = state.get("raw_data_folders") or []
                if not folders:
                    flash(
                        "Add at least one raw-data folder path before clicking "
                        "Scan & auto-configure.",
                        "error",
                    )
                else:
                    from cecl_ui.services import auto_setup
                    snapshot = state.get("report_period") or None
                    findings = auto_setup.scan_folders_for_setup(
                        folders, snapshot_yyyymm=snapshot,
                        folder_labels=state.get("raw_data_folder_types") or [],
                    )
                    if not findings.get("ok"):
                        flash(
                            f"Auto-scan failed: {findings.get('error')}",
                            "error",
                        )
                    else:
                        report = auto_setup.apply_findings_to_state(
                            state, findings,
                            current_app.config["WORKSPACE_ROOT"],
                        )
                        # Record every folder scanned for the report panel.
                        state["_auto_scan_folders"] = list(
                            findings.get("folders") or folders
                        )
                        # When the scan detected a WARM workbook, import its
                        # establishing-quarter balances as the historical
                        # baseline (best-effort) so the WARM -> CM comparison
                        # / balance-adjustment validation has data to compare
                        # against. Mirrors the manual WARM-upload path.
                        if state.get("has_warm_files") == "yes":
                            _warm = state.get("warm") or {}
                            _wsrc = _warm.get("saved_path") or _warm.get("source_path")
                            _cu = state.get("credit_union") or _warm.get("cu_name")
                            _as_of = _warm.get("as_of_date")
                            if _wsrc and _cu and _as_of:
                                try:
                                    _res = pipeline_service.import_warm_as_baseline(
                                        cu_name=_cu,
                                        warm_source_path=str(_wsrc),
                                        as_of_date=_as_of,
                                        overwrite=False,
                                    )
                                    if _res.get("ok"):
                                        report.setdefault("warm", []).append(
                                            "Historical WARM baseline imported "
                                            f"for {_as_of} — WARM comparison ready."
                                        )
                                except Exception as _exc:  # noqa: BLE001
                                    report.setdefault("warm", []).append(
                                        f"WARM baseline import skipped ({_exc})."
                                    )
                        # Persist the populated state immediately so a
                        # crash mid-step doesn't lose the scan results.
                        _save_state(state)
                        # Friendly summary flash.
                        n_steps = len(report)
                        n_msgs = sum(len(v) for v in report.values())
                        warm_note = (
                            " WARM workbook detected — WARM comparison path active."
                            if state.get("has_warm_files") == "yes"
                            else " No WARM workbook found — building reports from raw data."
                        )
                        flash(
                            f"Auto-scan complete across {len(folders)} folder(s): "
                            f"filled {n_msgs} item(s) across {n_steps} step(s). "
                            f"Errors: {len(findings.get('errors') or [])}.{warm_note}",
                            "success" if n_msgs else "info",
                        )
                        # Route to the first REQUIRED step only. Recommended
                        # review checkpoints (pools, grades, mgmt-adj, ...) are
                        # surfaced as stepper badges but do NOT halt the flow,
                        # so a good auto-scan drives all the way through to
                        # Review. The wizard still stops on a genuinely
                        # blocking gap that couldn't be auto-derived (e.g. a
                        # missing monthly-balance source or file pattern).
                        if state.get("model") == "scale":
                            steps = WIZARD_STEPS_SCALE
                        elif state.get("has_warm_files") == "yes":
                            steps = WIZARD_STEPS_WARM
                        else:
                            steps = WIZARD_STEPS_NO_WARM
                        next_key = auto_setup.first_hil_step_key(
                            state, steps, severity="required"
                        )
                        # "review" is always a required step (final Save); when
                        # it's the first one left, everything upstream is
                        # satisfied — route straight to Review.
                        if next_key and next_key != "review" \
                                and next_key in STEP_ENDPOINTS:
                            return redirect(url_for(STEP_ENDPOINTS[next_key]))
                        # Nothing blocking upstream -> jump to review. SCALE
                        # drafts finalize via scale_setup.step_review;
                        # only Migration drafts hit step10_review.
                        if state.get("model") == "scale":
                            return redirect(url_for("scale_setup.step_review"))
                        return redirect(url_for("setup.step10_review"))
                # Fall through to re-render Step 1 with the flashed
                # error and any partial state.
                return redirect(url_for("setup.step1_identity"))
            # ------------------------------------------------------------

            # Conditional routing:
            #  * SCALE model -> jump straight to the SCALE Solr step.
            #  * Migration with WARM file -> WARM upload.
            #  * Migration without WARM -> historical / loan pools.
            if state.get("model") == "scale":
                return redirect(url_for("scale_setup.step_solr"))
            if state.get("has_warm_files") == "yes":
                return redirect(url_for("setup.step2_warm"))
            return redirect(url_for("setup.step_loan_pools"))
    return render_template("setup/step1_identity.html", states=geo_service.states(), **_wizard_ctx("identity"))


# JSON: counties for a given state (used by Step 1 to populate the dropdown)
@setup_bp.route("/api/counties")
def api_counties():
    state = request.args.get("state", "").strip()
    return jsonify({"state": state, "counties": geo_service.counties_for_state(state)})


# JSON: auto-detect state + county for a credit union by charter number.
# Sources: state from the Solr 5300 ``ncua`` core (walks back through prior
# quarters when the latest one indexes ``custate='NA'``); county from
# NCUA's quarterly FOICU.txt (downloaded + cached on demand).
@setup_bp.route("/api/locate-cu")
def api_locate_cu():
    charter_raw = (request.args.get("charter") or "").strip()
    if not charter_raw.isdigit():
        return jsonify({
            "ok": False,
            "errors": ["charter must be a positive integer"],
        }), 400
    try:
        result = cu_locator.lookup_cu_geography(
            int(charter_raw),
            workspace_root=current_app.config["WORKSPACE_ROOT"],
        )
    except Exception as exc:  # pragma: no cover — defensive
        return jsonify({
            "ok": False,
            "charter": int(charter_raw),
            "errors": [f"lookup failed: {exc}"],
        }), 500
    return jsonify(result)


# =================================================================
# Step 2 — WARM workbook upload (optional)
# =================================================================

_WARM_DIR = Path(tempfile.gettempdir()) / "cecl_ui_warm"


def _save_warm_upload(file_storage) -> Path:
    _WARM_DIR.mkdir(parents=True, exist_ok=True)
    fn = secure_filename(file_storage.filename or "warm.xlsx")
    target = _WARM_DIR / fn
    file_storage.save(target)
    return target


def _apply_warm_to_state(state: dict[str, Any], analysis: dict[str, Any]) -> None:
    """Merge WARM-workbook findings into wizard state.

    Pre-fills credit_union, as-of date hints, pool_map, and credit_grades when
    the user is still on default values for those fields.
    """
    defaults = _default_state()

    # Identity & economic baseline from "BS CO DQ Data Enter" — prefer the
    # WARM values for blank fields. We never overwrite something the user has
    # already typed in Step 1.
    bid = analysis.get("baseline_identity") or {}
    if bid:
        if not state.get("credit_union") and bid.get("cu_name"):
            state["credit_union"] = bid["cu_name"]
        if not state.get("charter_number") and bid.get("charter_number"):
            state["charter_number"] = bid["charter_number"]
        econ = state.setdefault("economic_data", {})
        if not econ.get("state") and bid.get("state"):
            econ["state"] = bid["state"]
        if not econ.get("county") and bid.get("county"):
            econ["county"] = bid["county"]
        # Numeric fields: only seed when still at the default-state value
        # (so a user's manual entry wins).
        defaults_econ = defaults["economic_data"]
        for key in ("unemployment_rate", "foreclosures",
                    "bankruptcies", "population"):
            if econ.get(key) == defaults_econ.get(key) and bid.get(key):
                econ[key] = bid[key]
        # Short name follows CU name.
        if not state.get("short_name") and state.get("credit_union"):
            state["short_name"] = config_service.slugify(state["credit_union"])
        # As-of date hint for the wizard summary (kept on the warm dict).
        if bid.get("period_end_date"):
            analysis["as_of_date"] = analysis.get("as_of_date") \
                or bid["period_end_date"]

    # CU name — fall back to whatever the cu_name scan picked up if the
    # baseline_identity didn't have one.
    if not state.get("credit_union") and analysis.get("cu_name"):
        state["credit_union"] = analysis["cu_name"]
        if not state.get("short_name"):
            state["short_name"] = config_service.slugify(analysis["cu_name"])

    # Pool map — seed from WARM ``Grade Ranges & Loan Codes`` cols S/T
    # (real raw loan codes -> pool name). Fall back to {pool: pool} if the
    # WARM file didn't have that mapping.
    #
    # Reseed when:
    #   * the user hasn't customized yet (still equal to factory defaults), OR
    #   * the current map is exactly what we last seeded from a WARM file
    #     (so re-uploading WARM picks up new codes / corrected pool names).
    # Once the user edits the table on the Loan Pools step, we leave it
    # alone and only flash a hint.
    code_map = analysis.get("loan_code_pool_map") or {}
    prev_seed = state.get("_warm_seeded_pool_map") or {}
    user_untouched = (
        state["pool_map"] == defaults["pool_map"]
        or (prev_seed and state["pool_map"] == prev_seed)
    )
    if code_map:
        if user_untouched:
            state["pool_map"] = dict(code_map)
            state["_warm_seeded_pool_map"] = dict(code_map)
            distinct_pools = len({v for v in code_map.values() if v})
            flash(
                f"Loaded initial Loan Code Mapping from WARM "
                f"&ldquo;Grade Ranges &amp; Loan Codes&rdquo; tab "
                f"(cols S/T): {len(code_map)} code(s) mapped to "
                f"{distinct_pools} pool(s).",
                "success",
            )
        else:
            flash(
                f"WARM &ldquo;Grade Ranges &amp; Loan Codes&rdquo; tab has "
                f"{len(code_map)} code &rarr; pool entries (cols S/T), "
                "but your Loan Code Mapping has been edited &mdash; not "
                "overwriting. Clear or re-upload from the Loan Pools step "
                "if you want to start over.",
                "info",
            )
    elif state["pool_map"] == defaults["pool_map"] and analysis.get("pools"):
        state["pool_map"] = {p: p for p in analysis["pools"]}

    # Per-pool settings — seed from WARM "BS CO DQ Data Enter" tab unless
    # the user has already filled some in.
    def _is_real_pool_row(s: dict[str, Any]) -> bool:
        nm = (s.get("name") or "").strip().lower()
        if not nm:
            return False
        if nm.startswith("allowance for credit loss") or nm == "allowance":
            return False
        if nm.startswith("credit grade deteriorated"):
            return False
        if nm.startswith("grand total") or nm.startswith("total") \
                or nm.startswith("hide") or nm == "exclude":
            return False
        return True

    warm_pool_settings = [
        s for s in (analysis.get("pool_settings") or [])
        if _is_real_pool_row(s)
    ]
    existing_ps = state.get("pool_settings") or []
    if existing_ps:
        # Scrub any stale ACL/CGD rows that were saved by an earlier version.
        cleaned = [s for s in existing_ps if _is_real_pool_row(s)]
        if len(cleaned) != len(existing_ps):
            state["pool_settings"] = cleaned
    if warm_pool_settings and not state.get("pool_settings"):
        state["pool_settings"] = [dict(s) for s in warm_pool_settings]

    # ACL balance — seed from WARM if the user hasn't typed one yet.
    if not state.get("acl_balance") and analysis.get("acl_balance"):
        try:
            state["acl_balance"] = float(analysis["acl_balance"])
        except (TypeError, ValueError):
            state["acl_balance"] = 0.0

    # Credit grades — replace defaults with grades from the workbook.
    if state["credit_grades"] == defaults["credit_grades"] and analysis.get("grades"):
        state["credit_grades"] = list(analysis["grades"])

    state["warm"] = analysis


def _merge_autoderive_into_state(state: dict[str, Any], cfg: dict[str, Any]) -> list[str]:
    """Overlay the validated resolver-path (``warm_autoderive``) derivations
    onto wizard state.

    Fields the wizard already emits (column mappings, member/account key,
    loan-code map, credit grades, credit-pull SSN join) are written only when
    still at their factory defaults so a user's edits always win. The WARM
    reserve pinning + monthly book are stashed under ``state['warm_reserve']``
    for ``config_service.build_yaml_from_wizard`` to emit. Returns a list of
    human-readable field names that were populated (for the flash summary).
    """
    defaults = _default_state()
    filled: list[str] = []

    # Column mappings — traced from the WARM Data-tab formulas. This is what
    # lets WARM CUs skip the manual Column Mappings step.
    cm = {k: v for k, v in (cfg.get("column_mappings") or {}).items() if v}
    if cm and state.get("column_mappings") == defaults["column_mappings"]:
        merged = dict(state["column_mappings"])
        merged.update(cm)
        state["column_mappings"] = merged
        filled.append(f"{len(cm)} column mapping(s)")

    # Headerless positional extract: when the assembled cfg marks the loan
    # file as header-less (integer column_mappings from the WARM Data-tab
    # column trace), propagate that so build_yaml emits has_header: false.
    if cfg.get("has_header") is False:
        state["has_header"] = False
        filled.append("headerless extract (positional columns)")

    # Member / account key (mode + suffix).
    ma = cfg.get("member_account") or {}
    if ma.get("mode") and state.get("member_account") == defaults["member_account"]:
        state["member_account"] = {
            k: ma[k] for k in ("mode", "suffix_length", "delimiter") if k in ma
        }
        if ma.get("suffix_length") is not None:
            try:
                state["account_suffix_length"] = int(ma["suffix_length"])
            except (TypeError, ValueError):
                pass
        filled.append(f"member/account = {ma.get('mode')}")

    # Loan-code -> pool map (keep-first: ignores legacy/duplicate code blocks
    # like Maple's CUDL rows that the old parser mis-mapped).
    pm = cfg.get("pool_map") or {}
    prev = state.get("_warm_seeded_pool_map") or {}
    if pm and (not state.get("pool_map")
               or state.get("pool_map") == defaults["pool_map"]
               or state.get("pool_map") == prev):
        state["pool_map"] = dict(pm)
        state["_warm_seeded_pool_map"] = dict(pm)
        filled.append(f"{len(pm)} loan code(s)")

    # Credit grades.
    cg = cfg.get("credit_grades") or []
    if cg and state.get("credit_grades") == defaults["credit_grades"]:
        state["credit_grades"] = [dict(g) for g in cg]

    # Credit-pull block: WARM fallback (report-time score/ACL/CO/DQ load) +
    # SSN join key. Overlaid onto state only when the user hasn't already
    # pointed the fallback at a folder of their own.
    cp = cfg.get("credit_pull") or {}
    if cp:
        scp = dict(state.get("credit_pull") or {})
        if not scp.get("fallback_report_folder"):
            for k in ("fallback_report_folder", "fallback_report_pattern",
                      "fallback_sheet_pattern", "fallback_member_col",
                      "fallback_score_col", "member_column", "score_column",
                      "pull_as_of_date", "use_standalone_file"):
                if cp.get(k) is not None:
                    scp[k] = cp[k]
            if cp.get("fallback_report_folder"):
                filled.append("credit-pull fallback")
        if cp.get("loan_join_column"):
            scp["loan_join_column"] = cp["loan_join_column"]
            if cp.get("member_column"):
                scp["member_column"] = cp["member_column"]
            filled.append(f"SSN pull-join &rarr; {cp['loan_join_column']}")
        state["credit_pull"] = scp

    # Reserve pinning + monthly book — emitted by build_yaml_from_wizard.
    state["warm_reserve"] = {
        "base_loss_rate_by_pool_grade": cfg.get("base_loss_rate_by_pool_grade"),
        "warm_allowance_pools": cfg.get("warm_allowance_pools"),
        "not_risk_rated": cfg.get("not_risk_rated"),
        "monthly_balance": cfg.get("monthly_balance"),
    }
    if cfg.get("base_loss_rate_by_pool_grade"):
        filled.append(
            f"reserve pins for {len(cfg['base_loss_rate_by_pool_grade'])} pool(s)")

    return filled


def _refresh_stale_warm_analysis(state: dict[str, Any]) -> bool:
    """Passive re-parse of the cached WARM workbook.

    If the saved WARM file is still on disk but the cached analysis is
    missing fields the current parser version produces (most commonly
    ``balance_titles`` / ``bs_loan_type_map`` from the BS Data tab —
    added/extended in later commits), re-run ``analyse_warm_file`` and
    merge any populated-now-but-empty-before fields in. Never clobbers
    user edits. Returns ``True`` if state was updated.

    Called from wizard steps that consume WARM-derived BS-Data mappings
    (Step 2, Step 7) so existing drafts pick up the richer parser
    output without forcing a re-upload.
    """
    w = state.get("warm") or {}
    path = w.get("saved_path")
    if not path:
        return False
    if w.get("balance_titles") and w.get("bs_loan_type_map"):
        return False
    try:
        from pathlib import Path as _P
        wp = _P(path)
        if not wp.exists():
            return False
        fresh = warm_parser.analyse_warm_file(
            wp, original_filename=w.get("filename") or wp.name,
        )
        if not fresh.get("ok"):
            return False
        changed = False
        for _k in ("balance_titles", "bs_loan_type_map",
                   "loan_code_pool_map", "pool_monthly_balances",
                   "pool_settings", "pools", "grades",
                   "co_pools_with_data", "recov_pools_with_data",
                   "history_start", "history_end",
                   "history_months", "as_of_date", "cu_name",
                   "sheets_found", "sheets_missing",
                   "baseline_identity", "acl_balance"):
            _new = fresh.get(_k)
            _old = w.get(_k)
            if _new and not _old:
                w[_k] = _new
                changed = True
        if changed:
            state["warm"] = w
            _save_state(state)
        return changed
    except Exception:  # noqa: BLE001
        return False


@setup_bp.route("/step/warm", methods=["GET", "POST"])
def step2_warm():
    state = _state()
    if request.method == "GET":
        _refresh_stale_warm_analysis(state)

    if request.method == "POST":
        action = request.form.get("action", "")

        if action == "upload":
            f = request.files.get("warm_file")
            if not f or not f.filename:
                flash("Pick a WARM workbook first.", "error")
            else:
                try:
                    saved = _save_warm_upload(f)
                    analysis = warm_parser.analyse_warm_file(
                        saved, original_filename=f.filename
                    )
                    if not analysis.get("ok"):
                        flash(f"Could not parse WARM file: {analysis.get('error')}",
                              "error")
                    else:
                        # Remember where the upload landed so the user can
                        # promote it to the historical baseline (Reports/).
                        analysis["saved_path"] = str(saved)
                        _apply_warm_to_state(state, analysis)
                        # Validated resolver-path auto-derive on the same WARM:
                        # column mappings (Data-tab formula trace), reserve col-G
                        # pins, SSN credit-pull join, keep-first loan-code map,
                        # WARM pool order. Surfaced as a tiered review + merged
                        # into state so the generated config ties to the WARM.
                        ad_filled: list[str] = []
                        try:
                            from cecl_ui.services.setup import warm_autoderive
                            _ad = warm_autoderive.derive(
                                str(saved),
                                short_name=state.get("short_name") or "",
                                snapshot_date=state.get("report_period") or None,
                            )
                            state["autoderive"] = {
                                "ok": bool(_ad.get("ok")),
                                "overall_tier": _ad.get("overall_tier"),
                                "decisions": _ad.get("decisions") or [],
                                "error": _ad.get("error"),
                            }
                            if _ad.get("ok"):
                                ad_filled = _merge_autoderive_into_state(
                                    state, _ad["config"])
                                # Stash the WARM Data-tab column POSITIONS so the
                                # loan-extract step can swap the named mappings
                                # for 0-based integers if the delivered extract
                                # turns out to be headerless (Franklin/Bridgeton
                                # positional AIRES files).
                                state["_warm_column_indices"] = (
                                    (_ad.get("warm") or {}).get("column_indices")
                                    or {})
                                # Steps the WARM auto-derive fully covers ->
                                # rendered green (review-only) so the user flows
                                # to the interactive steps (loan extract, monthly
                                # balance/ACL, historical baseline, credit-pull
                                # options, balance-adjustment tie, impaired).
                                state["_warm_autoderived_steps"] = [
                                    "pools", "balances", "grades", "columns",
                                    "economic", "mgmt_adj", "files", "dq_hist",
                                    "co_recov", "orig_score", "credit_pull",
                                ]
                        except Exception as _ad_exc:  # noqa: BLE001
                            state["autoderive"] = {
                                "ok": False, "overall_tier": "red",
                                "decisions": [], "error": str(_ad_exc)}
                        _save_state(state)
                        flash(
                            f"Parsed {f.filename!s}: {len(analysis['pools'])} pools, "
                            f"{len(analysis['grades'])} grades, "
                            f"{analysis['history_months']} months of history "
                            f"({analysis['history_start']} to {analysis['history_end']}).",
                            "success",
                        )
                        if ad_filled:
                            flash(
                                "Auto-derived from WARM: " + ", ".join(ad_filled)
                                + ". Review below.",
                                "success",
                            )
                except Exception as exc:  # noqa: BLE001
                    flash(f"Upload failed: {exc}", "error")
            # fall through to re-render with the analysis visible

        elif action == "clear":
            defaults = _default_state()
            state["warm"] = None
            state["pool_map"] = dict(defaults["pool_map"])
            state["credit_grades"] = [dict(g) for g in defaults["credit_grades"]]
            _save_state(state)
            flash("WARM analysis cleared. Defaults restored.", "info")

        elif action in ("import_baseline", "import_baseline_overwrite"):
            # Baseline import moved to Step 3 (step3_baseline). Forward there
            # so any old form posts still work.
            return redirect(url_for("setup.step3_baseline"))

        elif action in ("next", "skip"):
            # Both WARM and no-WARM paths now go to the dedicated
            # Loan Code Mapping step (Step 3) before any further
            # data-collection steps.
            return redirect(url_for("setup.step4_pools"))

        elif action == "save_pool_settings":
            names = request.form.getlist("ps_name")
            # Preserve existing use_default_mgmt_adj values (now edited on
            # the Mgmt Adjustments step, not this one).
            prior_use_def = {
                (p.get("name") or "").strip(): bool(p.get("use_default_mgmt_adj"))
                for p in (state.get("pool_settings") or [])
            }
            updated: list[dict[str, Any]] = []
            for i, name in enumerate(names):
                nm = (name or "").strip()
                if not nm:
                    continue
                rr_val = request.form.get(f"ps_risk_rated_{i}", "no").strip().lower()
                excluded = (rr_val == "excluded")
                brr = (rr_val == "yes_brr")
                rr = (rr_val in ("yes", "yes_brr"))
                acl_raw = (request.form.get(f"ps_acl_months_{i}", "") or "").strip()
                try:
                    acl = int(acl_raw) if acl_raw else 0
                except ValueError:
                    acl = 0
                use_def = prior_use_def.get(nm, False)
                updated.append({
                    "name": nm,
                    "risk_rated": rr,
                    "brr": brr,
                    "acl_months": acl,
                    "use_default_mgmt_adj": use_def,
                    "excluded": excluded,
                })
            state["pool_settings"] = updated

            # Sync the canonical ordered pool list (state.warm.pools) so any
            # pool the user added/removed/reordered here also drives the
            # downstream loan-code mapping step and the final YAML.
            warm = state.get("warm") or {}
            warm["pools"] = [p["name"] for p in updated]
            state["warm"] = warm

            _save_state(state)
            flash(f"Saved settings for {len(updated)} pool(s).", "success")
            # Move on to the dedicated Loan Code Mapping step.
            return redirect(url_for("setup.step4_pools"))

    changed = _apply_score_based_risk_default(state)
    if changed:
        _save_state(state)
        flash(
            "Defaulted "
            + ", ".join(changed)
            + " to <strong>No (single line)</strong> because "
            + ("this pool has" if len(changed) == 1 else "these pools have")
            + " no credit scores. Change any back to per-grade below if needed.",
            "info",
        )
    return render_template("setup/step2_warm.html", **_wizard_ctx("warm"))


# Pools whose scored balance is below this fraction of their total balance are
# treated as having "no credit scores" and default to single-line
# (risk_rated=False) in Step 2. A small non-zero threshold absorbs a handful of
# stray scores in an otherwise unscored pool (e.g. one scored loan out of
# hundreds), matching the "all balances in Not Reported" case.
_NO_SCORE_SINGLE_LINE_FRAC = 0.01


def _apply_score_based_risk_default(state: dict[str, Any]) -> list[str]:
    """Default pools with (near-)zero credit-score coverage to single-line.

    For any pool whose imported loans carry essentially no credit scores, set
    ``risk_rated=False`` so Step 2 defaults it to a single line instead of a
    per-grade breakout. Runs at most once per draft (guarded by
    ``_score_default_applied``) so it never fights a manual choice the user
    makes afterward, and only ever flips a scored-out pool *off* — scored pools
    are left untouched. No-ops when the CU has no imported loan data yet, so it
    naturally activates on the first Step 2 visit after an import.

    Returns the list of pool names that were flipped (for a flash message).
    """
    if state.get("_score_default_applied"):
        return []
    pool_settings = state.get("pool_settings") or []
    if not pool_settings:
        return []
    cu = (state.get("credit_union") or "").strip()
    if not cu:
        return []
    try:
        coverage = pipeline_service.pool_score_coverage_for_cu(cu)
    except Exception:  # noqa: BLE001
        coverage = {}
    if not coverage:
        # No imported data to judge by — leave defaults and try again on the
        # next visit (once the loans have been imported).
        return []
    cov_ci = {str(k).strip().lower(): v for k, v in coverage.items()}
    changed: list[str] = []
    for p in pool_settings:
        nm = (p.get("name") or "").strip()
        if not nm or nm.lower() in ("ignore", "exclude"):
            continue
        info = cov_ci.get(nm.lower())
        if not info or info.get("total", 0) <= 0:
            continue
        if info["frac"] < _NO_SCORE_SINGLE_LINE_FRAC and p.get("risk_rated"):
            p["risk_rated"] = False
            p["brr"] = False
            changed.append(nm)
    # Data was available, so this decision is final — don't re-run and clobber
    # the user's subsequent edits.
    state["_score_default_applied"] = True
    return changed


# =================================================================
# Step 2 (no-WARM path) — Loan Pools
# Mirrors the per-pool settings table from step2_warm so credit unions
# without a WARM workbook can still declare their canonical pool list
# (name, risk-rated breakout, ACL months, mgmt-adj overlay, exclusion)
# before any downstream step that needs it.
# =================================================================
@setup_bp.route("/step/loan-pools", methods=["GET", "POST"])
def step_loan_pools():
    state = _state()
    if request.method == "POST":
        action = request.form.get("action", "")

        if action == "apply_pool_grouping":
            names = request.form.getlist("grp_pool_name")
            choices: dict[str, str] = {}
            for i, nm in enumerate(names):
                nm = (nm or "").strip()
                if not nm:
                    continue
                raw = (request.form.get(f"grp_choice_{i}", "keep") or "keep").strip().lower()
                choices[nm] = "split" if raw == "split" else "keep"
            try:
                from cecl_ui.services import auto_setup as _auto_setup
                result = _auto_setup.apply_pool_grouping_choice(state, choices)
            except Exception as exc:  # noqa: BLE001
                flash(f"Could not apply pool grouping: {exc}", "error")
            else:
                if result.get("ok"):
                    _save_state(state)
                    flash(
                        f"Rebuilt pool list: {result['pools_count']} pool(s) "
                        f"({result['kept_count']} kept, "
                        f"{result['split_count']} split) &mdash; "
                        f"{result['mappings_count']} sub-category mapping(s) "
                        f"propagated to Step 8 Monthly Balance map.",
                        "success",
                    )
                else:
                    flash(
                        result.get("error")
                        or "Could not rebuild from pool seed.",
                        "error",
                    )
            return redirect(url_for("setup.step_loan_pools"))

        if action == "save_pool_settings":
            names = request.form.getlist("ps_name")
            prior_use_def = {
                (p.get("name") or "").strip(): bool(p.get("use_default_mgmt_adj"))
                for p in (state.get("pool_settings") or [])
            }
            updated: list[dict[str, Any]] = []
            for i, name in enumerate(names):
                nm = (name or "").strip()
                if not nm:
                    continue
                rr_val = request.form.get(f"ps_risk_rated_{i}", "no").strip().lower()
                excluded = (rr_val == "excluded")
                brr = (rr_val == "yes_brr")
                rr = (rr_val in ("yes", "yes_brr"))
                acl_raw = (request.form.get(f"ps_acl_months_{i}", "") or "").strip()
                try:
                    acl = int(acl_raw) if acl_raw else 0
                except ValueError:
                    acl = 0
                use_def = prior_use_def.get(nm, False)
                updated.append({
                    "name": nm,
                    "risk_rated": rr,
                    "brr": brr,
                    "acl_months": acl,
                    "use_default_mgmt_adj": use_def,
                    "excluded": excluded,
                })
            state["pool_settings"] = updated

            # Keep state.warm.pools in sync so the downstream Loan Code
            # Mapping step (and final YAML) see the same ordered list
            # the user just confirmed. The no-WARM path has no parsed
            # WARM analysis, so we only set the `pools` slot.
            warm = state.get("warm") or {}
            warm["pools"] = [p["name"] for p in updated]
            state["warm"] = warm

            _save_state(state)
            flash(f"Saved settings for {len(updated)} pool(s).", "success")
            return redirect(url_for("setup.step4_pools"))

        if action in ("next", "skip"):
            return redirect(url_for("setup.step4_pools"))

    changed = _apply_score_based_risk_default(state)
    if changed:
        _save_state(state)
        flash(
            "Defaulted "
            + ", ".join(changed)
            + " to <strong>No (single line)</strong> because "
            + ("this pool has" if len(changed) == 1 else "these pools have")
            + " no credit scores. Change any back to per-grade below if needed.",
            "info",
        )
    return render_template("setup/step_loan_pools.html", **_wizard_ctx("loan_pools"))


# =================================================================
# Step 3 — Balance Titles → Loan Pool mapping (WARM path)
# =================================================================
@setup_bp.route("/step/balances", methods=["GET", "POST"])
def step3_balances():
    """Map each CU-supplied balance title (read from BS Data above the
    Pool Order anchor) to one of the established Loan Pools, or mark it
    as ignored.

    The titles are the raw labels the credit union uses in its monthly
    balance-sheet feed (e.g. "New Autos", "3rd Party First Mortgages").
    Each one belongs in exactly one Loan Pool from Step 2 — or it should
    be left out of the report entirely (e.g. spacers, sub-totals).
    """
    state = _state()
    warm = state.get("warm") or {}
    titles = warm.get("balance_titles") or []
    pools = [p["name"] for p in (state.get("pool_settings") or [])
             if (p.get("name") or "").strip()]
    if not pools:
        pools = warm.get("pools") or []

    if request.method == "POST":
        action = request.form.get("action", "")
        if action in ("save", "next"):
            keys = request.form.getlist("bt_title")
            mapping: dict[str, str] = {}
            for i, title in enumerate(keys):
                t = (title or "").strip()
                if not t:
                    continue
                pool = (request.form.get(f"bt_pool_{i}", "") or "").strip()
                # "" / "__ignore__" both mean the row is intentionally
                # excluded from the report.
                if pool and pool != "__ignore__":
                    mapping[t] = pool
                else:
                    mapping[t] = ""
            state["balance_title_map"] = mapping
            _save_state(state)
            kept = sum(1 for v in mapping.values() if v)
            ignored = len(mapping) - kept
            flash(
                f"Saved balance-title mapping: {kept} mapped, {ignored} ignored.",
                "success",
            )
            if action == "next":
                return redirect(url_for("setup.step3_baseline"))
            return redirect(url_for("setup.step3_balances"))
        if action == "back":
            return redirect(url_for("setup.step2_warm"))

    saved_map = state.get("balance_title_map") or {}
    return render_template(
        "setup/step3_balances.html",
        titles=titles,
        pools=pools,
        saved_map=saved_map,
        **_wizard_ctx("balances"),
    )


# =================================================================
# Step 3 — Historical Baseline import + ACL Balance (WARM path)
# =================================================================
@setup_bp.route("/step/baseline", methods=["GET", "POST"])
def step3_baseline():
    state = _state()
    if request.method == "POST":
        action = request.form.get("action", "")

        if action in ("import_baseline", "import_baseline_overwrite"):
            warm = state.get("warm") or {}
            src = warm.get("saved_path")
            cu_name = state.get("credit_union") or warm.get("cu_name") or ""
            as_of = warm.get("as_of_date") or ""
            if not src or not Path(src).exists():
                flash(
                    "WARM upload is no longer available on disk &mdash; "
                    "re-upload the file from Step 1 before importing.",
                    "error",
                )
            elif not cu_name:
                flash(
                    "Set the credit-union legal name on Step 1 before importing "
                    "the historical baseline.",
                    "error",
                )
            elif not as_of:
                flash(
                    "Could not determine the WARM as-of date &mdash; "
                    "re-upload or pick a different file in Step 1.",
                    "error",
                )
            else:
                result = pipeline_service.import_warm_as_baseline(
                    cu_name=cu_name,
                    warm_source_path=src,
                    as_of_date=as_of,
                    overwrite=(action == "import_baseline_overwrite"),
                )
                if not result["ok"]:
                    flash(f"Import failed: {result['error']}", "error")
                elif result["skipped"]:
                    warm["baseline_imported_to"] = result["dest_path"]
                    warm["baseline_filename"] = result["filename"]
                    warm["baseline_already_existed"] = True
                    state["warm"] = warm
                    _save_state(state)
                    flash(
                        f"A baseline file already exists at "
                        f"{result['filename']} &mdash; not overwritten.",
                        "info",
                    )
                else:
                    warm["baseline_imported_to"] = result["dest_path"]
                    warm["baseline_filename"] = result["filename"]
                    warm["baseline_already_existed"] = False
                    state["warm"] = warm
                    _save_state(state)
                    flash(
                        f"Imported as historical baseline: {result['filename']}. "
                        "Future report runs for this CU will pull historical "
                        "balances, charge-offs, and recoveries from this file.",
                        "success",
                    )
            return redirect(url_for("setup.step3_baseline"))

        if action in ("save_acl", "next"):
            acl_bal_raw = (request.form.get("acl_balance", "") or "").strip()
            acl_bal_raw = acl_bal_raw.replace(",", "").replace("$", "")
            if acl_bal_raw:
                try:
                    state["acl_balance"] = float(acl_bal_raw)
                except ValueError:
                    pass
            _save_state(state)
            if action == "next":
                return redirect(url_for("setup.step_dq_hist"))
            flash(
                f"Saved Allowance for Credit Loss: ${state.get('acl_balance') or 0:,.2f}.",
                "success",
            )
            return redirect(url_for("setup.step3_baseline"))

        if action == "back":
            return redirect(url_for("setup.step2_warm"))

    return render_template("setup/step3_baseline.html", **_wizard_ctx("baseline"))


# =================================================================
# Step 5 — Monthly Balance File (recurring quarterly upload contract)
# =================================================================
_MONTHLY_BAL_DIR = Path(tempfile.gettempdir()) / "cecl_ui_monthly_bal"
_ACL_FILE_DIR = Path(tempfile.gettempdir()) / "cecl_ui_acl_files"


def _save_monthly_bal_upload(file_storage) -> Path:
    _MONTHLY_BAL_DIR.mkdir(parents=True, exist_ok=True)
    fn = secure_filename(file_storage.filename or "monthly_balances.xlsx")
    target = _MONTHLY_BAL_DIR / fn
    file_storage.save(target)
    return target


def _ingest_annual_workbook(
    state: dict,
    mb: dict,
    target: Path,
    year_raw: str | int | None,
) -> tuple[str, str]:
    """Analyse a per-year balance workbook and register it on ``mb``.

    Shared by Step 3 (Historical) and Step 5 (Monthly Balance File) so
    the same workbook ingest behaviour applies on both pages. Returns
    ``(category, message)`` where category is one of
    ``success`` / ``warning`` / ``error`` suitable for ``flash``.
    """
    try:
        analysis = monthly_bal_parser.analyse_per_year_file(target)
    except Exception as exc:  # noqa: BLE001
        return ("error", f"Could not analyse {target.name}: {exc}")

    try:
        year_int = int(year_raw) if year_raw not in (None, "") else None
    except (TypeError, ValueError):
        year_int = None
    if year_int is None:
        year_int = analysis.get("detected_year")
    if year_int is None:
        return (
            "error",
            f"Saved {target.name}, but no calendar year was supplied "
            "and none could be detected from the filename.",
        )

    entry = {
        "filename": target.name,
        "saved_path": str(target),
        "year": int(year_int),
        "period_count": len(analysis.get("period_columns") or []),
    }
    files = mb.setdefault("year_files", [])
    files[:] = [e for e in files if e.get("filename") != entry["filename"]]
    files.append(entry)
    files.sort(key=lambda e: e.get("year") or 0)

    if not analysis.get("ok"):
        return (
            "warning",
            f"Uploaded {target.name} for {year_int}, but auto-detect "
            f"failed: {analysis.get('error', 'unknown error')}. "
            "Fill in the layout fields by hand and click Save.",
        )

    layout = mb.setdefault("per_year_layout", {})
    layout["sheet"] = analysis.get("sheet", "")
    layout["label_col"] = analysis.get("label_col", "B")
    layout["header_row"] = analysis.get("header_row", 1)
    layout["period_columns"] = analysis.get("period_columns", [])

    existing_labels = list(mb.get("parsed_pool_labels") or [])
    seen = {s.lower() for s in existing_labels}
    for lab in analysis.get("pool_labels", []):
        if lab.lower() not in seen:
            existing_labels.append(lab)
            seen.add(lab.lower())
    mb["parsed_pool_labels"] = existing_labels

    combined_map: dict[str, str] = {}
    for _k, _v in (state.get("balance_title_map") or {}).items():
        if _k:
            combined_map[_k] = (_v or "")
    _hpm = state.get("hist_pool_map") or {}
    for _k, _v in (_hpm.get("mapping") or {}).items():
        if _k and _k not in combined_map:
            combined_map[_k] = (_v or "")
    seeded, status = monthly_bal_parser.seed_pool_map(
        existing_labels, combined_map)
    existing_pm = mb.get("pool_map") or {}
    for label, pool in seeded.items():
        if label not in existing_pm or not existing_pm.get(label):
            existing_pm[label] = pool
    mb["pool_map"] = existing_pm
    mb["label_status"] = status

    # Merge auto-detected ACL history from this year file (if any) into
    # mb["acl"]["history"]. Only seeds row/label when the user has not
    # already set them so manual overrides are preserved.
    acl_hist = analysis.get("acl_history") or {}
    acl_row = analysis.get("acl_row")
    if acl_hist or acl_row:
        acl_state = mb.setdefault("acl", {})
        if acl_row and not acl_state.get("row"):
            acl_state["row"] = int(acl_row)
            acl_state["label"] = analysis.get("acl_label", "")
        if acl_hist:
            existing_hist = acl_state.get("history") or {}
            existing_hist.update(acl_hist)
            acl_state["history"] = existing_hist
    acl_msg = ""
    if acl_hist:
        acl_msg = (f" ACL history captured for {len(acl_hist)} "
                   f"month-end(s) from row {acl_row} "
                   f"({analysis.get('acl_label','')}).")

    return (
        "success",
        f"Uploaded {target.name} for {year_int}: detected sheet "
        f"{analysis.get('sheet')!r}, header row "
        f"{analysis.get('header_row')}, label column "
        f"{analysis.get('label_col')}, "
        f"{entry['period_count']} month-end column(s), "
        f"{len(analysis.get('pool_labels', []))} pool label(s)."
        + acl_msg,
    )


def _ingest_per_month_workbook(
    state: dict,
    mb: dict,
    target: Path,
    period_raw: str,
    stop_row: int | None = None,
    as_of_cell: str | None = None,
) -> tuple[str, str]:
    """Analyse one monthly balance-sheet file and register it on ``mb``.

    Mirrors :func:`_ingest_annual_workbook` for the per_month source.
    Returns ``(category, message)`` where category is one of
    ``success`` / ``warning`` / ``error`` suitable for ``flash``.
    """
    try:
        analysis = monthly_bal_parser.analyse_per_month_file(
            target, stop_row=stop_row, as_of_cell=as_of_cell,
            acl_row=((mb.get("acl") or {}).get("row") or None),
            acl_col=((mb.get("acl") or {}).get("col") or None))
    except Exception as exc:  # noqa: BLE001
        return ("error", f"Could not analyse {target.name}: {exc}")

    period = (period_raw or "").strip()
    if not period and analysis.get("detected_period"):
        period = analysis["detected_period"]
    if not period:
        return (
            "error",
            f"Saved {target.name}, but no month-end date was supplied "
            "and none could be detected in the file. Rename it to "
            "include YYYYMMDD or add an 'As of:' cell and re-upload.",
        )

    entry = {
        "filename": target.name,
        "saved_path": str(target),
        "period": period,
    }
    files = mb.setdefault("monthly_files", [])
    files[:] = [e for e in files if e.get("filename") != entry["filename"]]
    files.append(entry)
    files.sort(key=lambda e: e.get("period") or "")

    if not analysis.get("ok"):
        return (
            "warning",
            f"Uploaded {target.name} for {period}, but auto-detect "
            f"failed: {analysis.get('error', 'unknown error')}. "
            "Fill in the layout fields by hand and click Save.",
        )

    layout = mb.setdefault("per_month_layout", {})
    layout["sheet"] = analysis.get("sheet", "")
    layout["label_col"] = analysis.get("pool_name_col", "A")
    layout["balance_col"] = analysis.get("balance_col", "B")
    layout["header_row"] = analysis.get("header_row", 1)
    if as_of_cell is not None:
        layout["as_of_cell"] = (as_of_cell or "").strip().upper()
    # Persist effective + auto-detected stop rows so the runtime
    # aggregator can hard-cap on the same boundary the wizard used.
    if isinstance(stop_row, int) and stop_row > 0:
        layout["stop_row"] = int(stop_row)
    elif analysis.get("stop_row"):
        layout["stop_row"] = int(analysis["stop_row"])
    if analysis.get("auto_stop_row"):
        layout["auto_stop_row"] = int(analysis["auto_stop_row"])

    existing_labels = list(mb.get("parsed_pool_labels") or [])
    seen = {s.lower() for s in existing_labels}
    for lab in analysis.get("parsed_pool_labels", []):
        if lab.lower() not in seen:
            existing_labels.append(lab)
            seen.add(lab.lower())
    mb["parsed_pool_labels"] = existing_labels

    combined_map: dict[str, str] = {}
    for _k, _v in (state.get("balance_title_map") or {}).items():
        if _k:
            combined_map[_k] = (_v or "")
    _hpm = state.get("hist_pool_map") or {}
    for _k, _v in (_hpm.get("mapping") or {}).items():
        if _k and _k not in combined_map:
            combined_map[_k] = (_v or "")
    seeded, status = monthly_bal_parser.seed_pool_map(
        existing_labels, combined_map)
    existing_pm = mb.get("pool_map") or {}
    for label, pool in seeded.items():
        if label not in existing_pm or not existing_pm.get(label):
            existing_pm[label] = pool
    mb["pool_map"] = existing_pm
    mb["label_status"] = status

    acl_row = analysis.get("acl_row")
    acl_value = analysis.get("acl_value")
    acl_msg = ""
    if acl_row and acl_value is not None:
        acl_state = mb.setdefault("acl", {})
        if not acl_state.get("row"):
            acl_state["row"] = int(acl_row)
            acl_state["label"] = analysis.get("acl_label", "")
        hist = acl_state.get("history") or {}
        hist[period] = float(acl_value)
        acl_state["history"] = hist
        acl_msg = (f" ACL ${acl_value:,.0f} captured from row "
                   f"{acl_row} ({analysis.get('acl_label','')}).")

    return (
        "success",
        f"Uploaded {target.name} for {period}: parsed "
        f"{len(analysis.get('parsed_pool_labels', []))} pool label(s) "
        f"from sheet {analysis.get('sheet')!r} "
        f"(labels in column {analysis.get('pool_name_col')}, balances "
        f"in column {analysis.get('balance_col')})." + acl_msg,
    )


def _save_acl_file_upload(file_storage) -> Path:
    _ACL_FILE_DIR.mkdir(parents=True, exist_ok=True)
    fn = secure_filename(file_storage.filename or "acl.xlsx")
    target = _ACL_FILE_DIR / fn
    file_storage.save(target)
    return target


def _parse_pool_mapping_file(
    path: Path,
    label_col_idx: int = 0,
    pool_col_idx: int = 1,
    has_header: bool | None = None,
) -> tuple[dict[str, str], str, list[str]]:
    """Parse a workbook label -> pool name mapping file.

    Accepts ``.xlsx`` / ``.xlsm`` / ``.xls`` / ``.csv``.

    Args:
        path: file to read.
        label_col_idx: 0-based column index that holds the workbook label.
        pool_col_idx: 0-based column index that holds the target pool name.
        has_header: if None, auto-detect a header row by keyword; if True
            always skip the first row; if False never skip.

    Returns ``(mapping, error, header_preview)`` where ``mapping`` is
    ``{label: pool}``, ``error`` is "" on success, and ``header_preview``
    is the first non-blank row's cell values (used by the UI to render
    column-picker dropdowns when the user reuploads with explicit
    column choices). Empty pool values are preserved (= "ignore").
    """
    suffix = path.suffix.lower()
    rows: list[list[Any]] = []
    try:
        if suffix == ".csv":
            import csv
            with open(path, "r", encoding="utf-8-sig", newline="") as fh:
                rows = [list(r) for r in csv.reader(fh)]
        elif suffix in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                rows.append(list(r))
            wb.close()
        elif suffix == ".xls":
            import xlrd  # type: ignore
            book = xlrd.open_workbook(str(path))
            sh = book.sheet_by_index(0)
            for ri in range(sh.nrows):
                rows.append([sh.cell_value(ri, ci)
                             for ci in range(sh.ncols)])
        else:
            return ({}, f"Unsupported file type: {suffix}", [])
    except Exception as exc:  # noqa: BLE001
        return ({}, f"Could not read {path.name}: {exc}", [])

    rows = [r for r in rows
            if any((str(c).strip() if c is not None else "") for c in r)]
    if not rows:
        return ({}, "File is empty.", [])

    header_preview = [
        (str(c).strip() if c is not None else "") for c in rows[0]
    ]

    # Auto-detect header row when not explicitly told.
    drop_first = bool(has_header)
    if has_header is None:
        first = rows[0]
        first_label = (str(first[label_col_idx]).strip().lower()
                       if len(first) > label_col_idx
                       and first[label_col_idx] is not None else "")
        first_pool = (str(first[pool_col_idx]).strip().lower()
                      if len(first) > pool_col_idx
                      and first[pool_col_idx] is not None else "")
        header_keywords = {"label", "labels", "name", "title", "category",
                           "description", "workbook label",
                           "balance label", "account", "type"}
        pool_keywords = {"pool", "pools", "maps to", "target", "loan pool",
                         "mapped pool", "destination", "category"}
        if first_label in header_keywords or first_pool in pool_keywords:
            drop_first = True
    if drop_first:
        rows = rows[1:]

    mapping: dict[str, str] = {}
    for r in rows:
        if not r:
            continue
        label = (str(r[label_col_idx]).strip()
                 if len(r) > label_col_idx and r[label_col_idx] is not None
                 else "")
        if not label:
            continue
        pool = ""
        if len(r) > pool_col_idx and r[pool_col_idx] is not None:
            pool = str(r[pool_col_idx]).strip()
        if pool.lower() in ("ignore", "-- ignore --", "— ignore —", "none",
                            "n/a", "na", "skip", "exclude"):
            pool = ""
        mapping[label] = pool

    if not mapping:
        return ({}, "No mapping rows found.", header_preview)
    return (mapping, "", header_preview)


def _apply_pool_mapping(
    mb: dict, mapping: dict[str, str], source_name: str
) -> tuple[int, int, int]:
    """Merge a parsed label->pool mapping into ``mb``.

    Returns ``(applied, added_labels, matched_existing)``.
    """
    existing_pm = mb.get("pool_map") or {}
    existing_labels = list(mb.get("parsed_pool_labels") or [])
    lc_to_existing = {
        s.strip().lower(): s for s in existing_labels if s
    }
    seen_lc = set(lc_to_existing.keys())
    applied = 0
    added_labels = 0
    matched_existing = 0
    for lab, pool in mapping.items():
        key = (lab or "").strip().lower()
        if not key:
            continue
        if key in seen_lc:
            canonical = lc_to_existing[key]
            existing_pm[canonical] = pool
            matched_existing += 1
        else:
            existing_pm[lab] = pool
            existing_labels.append(lab)
            seen_lc.add(key)
            lc_to_existing[key] = lab
            added_labels += 1
        applied += 1
    mb["parsed_pool_labels"] = existing_labels
    mb["pool_map"] = existing_pm
    msg_parts = [f"Imported {applied} pool mapping(s) from {source_name}"]
    if matched_existing:
        msg_parts.append(
            f"{matched_existing} matched existing parsed label(s)")
    if added_labels:
        msg_parts.append(f"{added_labels} new label(s) added")
    flash("; ".join(msg_parts) + ".", "success")
    return (applied, added_labels, matched_existing)


def _save_acl_form(mb: dict, form) -> None:
    """Persist ACL-source form fields onto ``mb['acl']``.

    Reads ``acl_source`` plus the per-source inputs (``acl_row``,
    ``acl_sep_sheet``/``acl_sep_cell``, ``acl_manual_*``) from the POSTed
    form and writes them into the wizard's ``mb['acl']`` substate without
    overwriting auto-detected history.
    """
    acl_state = mb.setdefault("acl", {})
    src = (form.get("acl_source") or "").strip()
    if src in ("monthly_file", "separate", "manual"):
        acl_state["source"] = src
    # Row override for the monthly-file source.
    row_raw = form.get("acl_row")
    if row_raw is not None and str(row_raw).strip():
        try:
            acl_state["row"] = int(row_raw)
        except (TypeError, ValueError):
            pass
    # Column override for the monthly-file source. When the auto-
    # detected pool-balance column doesn't hold the ACL number
    # (e.g. ACL value sits in a different column than loan balances)
    # the user can pin a specific Excel column letter here.
    col_raw = form.get("acl_col")
    if col_raw is not None:
        col_clean = str(col_raw).strip().upper()
        if col_clean and col_clean.isalpha():
            acl_state["col"] = col_clean
        elif col_clean == "":
            acl_state.pop("col", None)
    # Separate-file metadata.
    sep = acl_state.setdefault("separate_file", {})
    sep["sheet"] = (form.get("acl_sep_sheet") or "").strip()
    sep["cell"] = (form.get("acl_sep_cell") or "").strip().upper()
    # Manual entries (3 month-end values).
    manual = acl_state.setdefault("manual", {})
    for key in ("month1", "month2", "month3"):
        date_field = f"acl_manual_{key}_date"
        val_field = f"acl_manual_{key}_value"
        if date_field in form:
            manual[f"{key}_date"] = (form.get(date_field) or "").strip()
        if val_field in form:
            raw = (form.get(val_field) or "").strip()
            if raw == "":
                manual[f"{key}_value"] = None
            else:
                # Strip currency symbols, thousands separators, and any
                # surrounding whitespace. Accept accounting-style
                # parentheses notation for negatives, e.g. "(896,484.48)".
                cleaned = (
                    raw.replace("$", "")
                    .replace(",", "")
                    .replace(" ", "")
                )
                negate = False
                if cleaned.startswith("(") and cleaned.endswith(")"):
                    cleaned = cleaned[1:-1]
                    negate = True
                try:
                    val = float(cleaned)
                    if negate:
                        val = -val
                    manual[f"{key}_value"] = val
                except ValueError:
                    pass
    # NCUA 5300 fallback toggle. The checkbox name is consistently
    # ``acl_use_5300_fallback``; the hidden ``acl_5300_fallback_present``
    # field tells us the form actually rendered the control so an
    # unticked box means "user disabled it" rather than "field absent".
    if "acl_5300_fallback_present" in form:
        acl_state["use_5300_fallback"] = bool(form.get("acl_use_5300_fallback"))


def _persist_per_month_layout(mb: dict, form) -> None:
    """Persist the per_month layout + pool_map + notes + ACL from ``form``.

    Used by both the ``save_per_month_layout`` action (Save button) and the
    ``save_per_month_layout_and_next`` action (Save & Next button) so we
    never silently lose user-entered fields when advancing to the next
    step.
    """
    layout = mb.setdefault("per_month_layout", {})
    layout["sheet"] = (form.get("pm_sheet", "") or "").strip()
    layout["label_col"] = (
        form.get("pm_label_col", "A") or "A"
    ).strip().upper()
    layout["balance_col"] = (
        form.get("pm_balance_col", "B") or "B"
    ).strip().upper()
    try:
        layout["header_row"] = int(form.get("pm_header_row", "1") or 1)
    except ValueError:
        layout["header_row"] = 1
    mb["pool_map"] = _parse_kv_rows(form, "map_label", "map_pool")
    mb["file_pattern"] = (form.get("file_pattern", "") or "").strip()
    mb["notes"] = (form.get("notes", "") or "").strip()
    _save_acl_form(mb, form)


def _persist_per_year_layout(mb: dict, form) -> None:
    """Persist the per_year layout + pool_map + notes + ACL from ``form``.

    Used by both the ``save_per_year_layout`` (Save) and
    ``save_per_year_layout_and_next`` (Save & Next) actions.
    """
    layout = mb.setdefault("per_year_layout", {})
    layout["label_col"] = (
        form.get("py_label_col", "B") or "B"
    ).strip().upper()
    try:
        layout["header_row"] = int(form.get("py_header_row", "1") or 1)
    except ValueError:
        layout["header_row"] = 1
    mb["pool_map"] = _parse_kv_rows(form, "map_label", "map_pool")
    mb["notes"] = (form.get("notes", "") or "").strip()
    _save_acl_form(mb, form)


def _persist_manual_grid(mb: dict, form) -> int:
    """Persist the manual pool × month grid from ``form``.

    Returns the count of pools in the saved grid (caller can use it for
    a flash message). Used by both ``save_manual`` and
    ``save_manual_and_next`` so Save & Next never drops the grid.
    """
    months = [
        (m or "").strip()
        for m in form.getlist("manual_month")
        if (m or "").strip()
    ]
    mb["manual_months"] = months
    pools = form.getlist("manual_pool")
    grid: dict[str, dict[str, float]] = {}
    for pool in pools:
        pool = (pool or "").strip()
        if not pool:
            continue
        row: dict[str, float] = {}
        for m in months:
            raw = (
                form.get(f"mv__{pool}__{m}") or ""
            ).strip().replace("$", "").replace(",", "")
            if raw == "":
                continue
            try:
                row[m] = float(raw)
            except ValueError:
                pass
        grid[pool] = row
    mb["manual_entries"] = grid
    mb["notes"] = (form.get("notes", "") or "").strip()
    _save_acl_form(mb, form)
    return len(grid)


@setup_bp.route("/step/monthly_bal", methods=["GET", "POST"])
def step5_monthly_bal():
    """Establish the recurring 'monthly balance by pool/type' file contract.

    On the WARM path this file fills the establishing-quarter columns that
    were intentionally skipped in Step 4. On the no-WARM path it provides
    the full historical balance series as well.
    """
    state = _state()
    # Backfill WARM-derived BS-Data fields for drafts that uploaded the
    # WARM workbook before the column-detection fix landed.  Safe no-op
    # when the cached analysis is already complete.
    if request.method == "GET":
        _refresh_stale_warm_analysis(state)
    has_warm = state.get("has_warm_files") == "yes"
    mb = state.setdefault("monthly_bal", {
        "filename": "", "saved_path": "", "sheet": "",
        "header_row": 0, "pool_name_col": "A", "first_date_col": "",
        "parsed_dates": [], "parsed_pool_labels": [], "pool_map": {},
        "file_pattern": "", "notes": "",
    })
    mb.setdefault("label_status", {})
    # Passive re-seed: if we have parsed labels but some are still un-mapped,
    # try to fill the blanks from the current balance_title_map +
    # hist_pool_map.mapping. Never overwrites a non-empty existing entry, so
    # user edits are preserved. Lets users who uploaded their file before
    # the hist_pool_map source was added pick up the new auto-mappings
    # without re-uploading.
    _parsed_labels = mb.get("parsed_pool_labels") or []
    _existing_map = mb.get("pool_map") or {}
    if _parsed_labels and any(not _existing_map.get(lbl) for lbl in _parsed_labels):
        _combined: dict[str, str] = {}
        for _k, _v in (state.get("balance_title_map") or {}).items():
            if _k:
                _combined[_k] = (_v or "")
        _hpm0 = state.get("hist_pool_map") or {}
        for _k, _v in (_hpm0.get("mapping") or {}).items():
            if _k and _k not in _combined:
                _combined[_k] = (_v or "")
        # Also pull the WARM-derived BS-Data loan-type→pool map directly,
        # so users who haven't yet visited Step 3 (Balance Titles) still
        # get pool_map suggestions out-of-the-box from the WARM workbook.
        _warm0 = state.get("warm") or {}
        for _k, _v in (_warm0.get("bs_loan_type_map") or {}).items():
            if _k and _k not in _combined:
                _combined[_k] = (_v or "")
        for _bt in (_warm0.get("balance_titles") or []):
            _t = (_bt.get("title") or "").strip() if isinstance(_bt, dict) else ""
            _p = (_bt.get("suggested_pool") or "").strip() if isinstance(_bt, dict) else ""
            if _t and _t not in _combined:
                _combined[_t] = _p
        if _combined:
            _seeded0, _status0 = monthly_bal_parser.seed_pool_map(
                _parsed_labels, _combined,
            )
            _changed = False
            for _lbl, _pool in _seeded0.items():
                if not _existing_map.get(_lbl) and _pool:
                    _existing_map[_lbl] = _pool
                    _changed = True
            if _changed:
                mb["pool_map"] = _existing_map
                # Refresh status only for labels we just filled in.
                _cur_status = mb.get("label_status") or {}
                for _lbl, _st in _status0.items():
                    if _existing_map.get(_lbl) and _cur_status.get(_lbl) != "matched":
                        _cur_status[_lbl] = _st
                mb["label_status"] = _cur_status
                _save_state(state)
    # How the credit union delivers month-end pool balances. One of:
    #   "single"    — one quarterly file with all months in column-band
    #                 layout (current default; uses Section 1 + 2 below).
    #   "per_month" — one balance-sheet style file per month.
    #   "manual"    — user enters pool × month grid by hand.
    mb.setdefault("source", "single")
    # Per-month-file mode: each entry is {filename, saved_path, period}.
    mb.setdefault("monthly_files", [])
    # Passive re-analyse for per_month mode: if the user uploaded files
    # before the parser knew about their balance-sheet flavour (e.g. a
    # GL-style sheet with "LOANS TO MEMBERS" instead of just "LOANS"),
    # ``parsed_pool_labels`` may still be empty even though files exist.
    # Re-run analysis on each saved file once and merge its labels in.
    # Never overwrites an existing pool_map entry the user has set.
    if (mb.get("source") or "single") == "per_month":
        _pm_files = [e for e in (mb.get("monthly_files") or [])
                     if e.get("saved_path")]
        if _pm_files and not (mb.get("parsed_pool_labels") or []):
            _seen_labels: list[str] = []
            _seen_set: set[str] = set()
            _pm_layout = mb.get("per_month_layout") or {}
            for _e in _pm_files:
                try:
                    _r = monthly_bal_parser.analyse_per_month_file(
                        _e.get("saved_path") or "")
                except Exception:  # noqa: BLE001
                    continue
                if not _r.get("ok"):
                    continue
                # Adopt layout defaults from first successful analyse if
                # the user-saved layout has blanks.
                if not _pm_layout.get("sheet") and _r.get("sheet"):
                    _pm_layout["sheet"] = _r.get("sheet") or ""
                if not _pm_layout.get("label_col") and _r.get("pool_name_col"):
                    _pm_layout["label_col"] = _r.get("pool_name_col") or ""
                if not _pm_layout.get("balance_col") and _r.get("balance_col"):
                    _pm_layout["balance_col"] = _r.get("balance_col") or ""
                if not _pm_layout.get("header_row") and _r.get("header_row"):
                    _pm_layout["header_row"] = int(_r.get("header_row") or 0)
                for _lbl in (_r.get("parsed_pool_labels") or []):
                    if _lbl and _lbl not in _seen_set:
                        _seen_set.add(_lbl)
                        _seen_labels.append(_lbl)
            if _seen_labels:
                mb["parsed_pool_labels"] = _seen_labels
                mb["per_month_layout"] = _pm_layout
                # Seed pool_map suggestions from any existing WARM mapping.
                _combined_pm: dict[str, str] = {}
                for _k, _v in (state.get("balance_title_map") or {}).items():
                    if _k:
                        _combined_pm[_k] = (_v or "")
                _hpm_pm = state.get("hist_pool_map") or {}
                for _k, _v in (_hpm_pm.get("mapping") or {}).items():
                    if _k and _k not in _combined_pm:
                        _combined_pm[_k] = _hpm_value_for_seed(_v)
                # WARM BS Data direct sources (loan_type→pool + balance_titles).
                _warm_pm = state.get("warm") or {}
                for _k, _v in (_warm_pm.get("bs_loan_type_map") or {}).items():
                    if _k and _k not in _combined_pm:
                        _combined_pm[_k] = (_v or "")
                for _bt in (_warm_pm.get("balance_titles") or []):
                    _t = (_bt.get("title") or "").strip() if isinstance(_bt, dict) else ""
                    _p = (_bt.get("suggested_pool") or "").strip() if isinstance(_bt, dict) else ""
                    if _t and _t not in _combined_pm:
                        _combined_pm[_t] = _p
                if _combined_pm:
                    _seeded_pm, _status_pm = monthly_bal_parser.seed_pool_map(
                        _seen_labels, _combined_pm,
                    )
                    _existing_pm = mb.get("pool_map") or {}
                    for _lbl, _pool in _seeded_pm.items():
                        if not _existing_pm.get(_lbl) and _pool:
                            _existing_pm[_lbl] = _pool
                    mb["pool_map"] = _existing_pm
                    _cur_status_pm = mb.get("label_status") or {}
                    for _lbl, _st in _status_pm.items():
                        if (mb.get("pool_map") or {}).get(_lbl) and \
                                _cur_status_pm.get(_lbl) != "matched":
                            _cur_status_pm[_lbl] = _st
                    mb["label_status"] = _cur_status_pm
                _save_state(state)
    # Common layout across all per-month files (so the user only specifies
    # sheet / label-col / balance-col / header-row once). Leave values
    # empty so the upload_per_month auto-detect can fill them; the
    # template renders display-defaults via ``{{ pm.label_col or 'A' }}``.
    mb.setdefault("per_month_layout", {
        "sheet": "", "label_col": "", "balance_col": "", "header_row": 0,
    })
    # Saved source folder for the "scan a folder for more months"
    # workflow — remembered so the user can re-scan next quarter with
    # a single click.
    mb.setdefault("per_month_source_folder", "")
    # Per-year mode: one workbook per calendar year, each with all 12
    # month-end balances as columns. ``year_files`` entries:
    # ``{filename, saved_path, year, period_count}``.
    mb.setdefault("year_files", [])
    mb.setdefault("per_year_layout", {
        "sheet": "", "label_col": "", "header_row": 0,
        "period_columns": [],
    })
    # Manual-entry mode: explicit list of month-end dates plus the
    # {pool_name: {YYYY-MM-DD: float}} grid the user fills in.
    mb.setdefault("manual_months", [])
    mb.setdefault("manual_entries", {})
    # ACL (Allowance for Credit Loss) source configuration. One of:
    #   "monthly_file" — read from a row of the monthly balance file.
    #   "separate"     — uploaded standalone file with the ACL value(s).
    #   "manual"       — user-entered values for the months in the quarter.
    mb.setdefault("acl", {
        "source": "monthly_file",   # default: same file as the pool balances
        "row": 0,                   # row number in monthly file (1-based)
        "label": "",                # auto-detected/last-confirmed label
        "history": {},              # {"YYYY-MM-DD": float}
        "separate_file": {
            "filename": "", "saved_path": "",
            "sheet": "", "cell": "",  # e.g. "B14"
            "value": None,
        },
        "manual": {                 # 3 month-end balances for the quarter
            "month1_date": "", "month1_value": None,
            "month2_date": "", "month2_value": None,
            "month3_date": "", "month3_value": None,
        },
        # NCUA 5300 fallback — when True and the file source produces
        # no ACL value (or $0) for the snapshot quarter, the report
        # engine pulls the ACL on Loans value directly from NCUA Form
        # 5300. Default ON for new wizard runs; users can untick if
        # the CU explicitly does not want 5300 data substituted.
        "use_5300_fallback": True,
    })
    # Backfill any sub-keys that may be missing on older drafts (the
    # outer ``setdefault`` above only seeds when ``acl`` is absent).
    _acl = mb["acl"]
    _acl.setdefault("source", "monthly_file")
    _acl.setdefault("row", 0)
    _acl.setdefault("label", "")
    _acl.setdefault("history", {})
    _acl.setdefault("separate_file", {
        "filename": "", "saved_path": "",
        "sheet": "", "cell": "", "value": None,
    })
    _man = _acl.setdefault("manual", {})
    for _mk in ("month1", "month2", "month3"):
        _man.setdefault(f"{_mk}_date", "")
        _man.setdefault(f"{_mk}_value", None)
    _acl.setdefault("use_5300_fallback", True)

    # Pre-fill manual ACL month-end dates from the WARM as-of date so the
    # user only has to type the three balances. We populate any blank date
    # cell — never overwrite a date the user has already entered.
    warm_as_of = (state.get("warm") or {}).get("as_of_date") or ""
    if warm_as_of:
        try:
            from datetime import date as _date
            from calendar import monthrange as _mr
            y, m, d = warm_as_of.split("-")
            anchor = _date(int(y), int(m), int(d))
            # Walk back to month-end of (anchor month - i) for i=2,1,0.
            ends: list[str] = []
            ay, am = anchor.year, anchor.month
            for back in (2, 1, 0):
                mm = am - back
                yy = ay
                while mm <= 0:
                    mm += 12
                    yy -= 1
                ends.append(_date(yy, mm, _mr(yy, mm)[1]).isoformat())
            man = mb["acl"].setdefault("manual", {})
            for k, iso in zip(("month1", "month2", "month3"), ends):
                if not man.get(f"{k}_date"):
                    man[f"{k}_date"] = iso
            # Seed manual-entry month list with the same three dates if the
            # user hasn't customised it.
            if not mb.get("manual_months"):
                mb["manual_months"] = list(ends)
        except (ValueError, TypeError):
            pass

    if request.method == "POST":
        action = request.form.get("action", "")
        # Enter-key submissions in a text input post the form without
        # any submit-button value, so ``action`` is empty. Pick a sane
        # default per source mode so the user's typed input is saved
        # rather than dropped (or worse, the form defaulting to the
        # first submit button which is "Back").
        if not action:
            src = (mb.get("source") or "single")
            if src == "per_month":
                action = "save_per_month_layout"
            elif src == "per_year":
                action = "save_per_year_layout"
            elif src == "manual":
                action = "save_manual"
            else:
                action = "save"

        if action == "upload":
            f = request.files.get("monthly_bal_file")
            if f and f.filename:
                try:
                    target = _save_monthly_bal_upload(f)
                    mb["filename"] = target.name
                    mb["saved_path"] = str(target)
                    # Auto-detect layout.
                    analysis = monthly_bal_parser.analyse_file(target)
                    if analysis.get("ok"):
                        mb["sheet"] = analysis.get("sheet", "") or mb["sheet"]
                        mb["header_row"] = analysis.get("header_row") or mb["header_row"]
                        mb["pool_name_col"] = analysis.get("pool_name_col") or mb["pool_name_col"]
                        mb["first_date_col"] = analysis.get("first_date_col") or mb["first_date_col"]
                        mb["parsed_dates"] = analysis.get("dates") or []
                        mb["parsed_pool_labels"] = analysis.get("parsed_pool_labels") or []
                        # Seed pool_map from any prior label->pool mapping the
                        # wizard has already collected: WARM balance_title_map
                        # (Step 3 Balance Titles) AND the historical-file
                        # hist_pool_map (Step 3 Historical). A monthly-balance
                        # file usually mirrors the historical-balance file's
                        # label layout, so seeding from both means the user
                        # rarely has to map anything by hand.
                        combined_map: dict[str, str] = {}
                        for _k, _v in (state.get("balance_title_map") or {}).items():
                            if _k:
                                combined_map[_k] = (_v or "")
                        _hpm = state.get("hist_pool_map") or {}
                        for _k, _v in (_hpm.get("mapping") or {}).items():
                            if _k and _k not in combined_map:
                                combined_map[_k] = _hpm_value_for_seed(_v)
                        seeded, status = monthly_bal_parser.seed_pool_map(
                            mb["parsed_pool_labels"],
                            combined_map,
                        )
                        # Preserve any user edits already in mb["pool_map"].
                        existing = mb.get("pool_map") or {}
                        for label, pool in seeded.items():
                            if label not in existing or not existing.get(label):
                                existing[label] = pool
                        mb["pool_map"] = existing
                        mb["label_status"] = status
                        # Seed ACL row + history if the parser found one and
                        # the user hasn't already configured an override.
                        acl_row = analysis.get("acl_row")
                        acl_hist = analysis.get("acl_history") or {}
                        acl_state = mb.setdefault("acl", {})
                        if acl_row and not acl_state.get("row"):
                            acl_state["row"] = int(acl_row)
                            acl_state["label"] = analysis.get("acl_label", "")
                        if acl_hist:
                            acl_state["history"] = acl_hist
                        flash(
                            f"Parsed {target.name}: "
                            f"{len(mb['parsed_dates'])} date column(s), "
                            f"{len(mb['parsed_pool_labels'])} pool label(s)."
                            + (f" ACL row auto-detected: row {acl_row} "
                               f"({analysis.get('acl_label','')})."
                               if acl_row else ""),
                            "success",
                        )
                    else:
                        flash(
                            f"Saved {target.name}, but auto-detect failed: "
                            f"{analysis.get('error', 'unknown error')}",
                            "warning",
                        )
                except Exception as exc:  # noqa: BLE001
                    flash(f"Upload failed: {exc}", "error")
            else:
                flash("Choose a monthly balance file to upload.", "error")
            _save_state(state)
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "remove":
            removed_name = mb.get("filename") or ""
            saved_path = mb.get("saved_path") or ""
            # Best-effort delete of the saved upload file.
            if saved_path:
                try:
                    p = Path(saved_path)
                    if p.is_file():
                        p.unlink()
                except Exception:  # noqa: BLE001
                    pass
            # Clear all state derived from the uploaded sample.
            for key in (
                "filename", "saved_path", "parsed_dates",
                "parsed_pool_labels", "pool_map", "label_status",
            ):
                if key in mb:
                    mb[key] = "" if key in ("filename", "saved_path") else (
                        [] if key in ("parsed_dates", "parsed_pool_labels")
                        else {}
                    )
            # Reset ACL auto-detected bits (preserve user-set overrides).
            acl_state = mb.get("acl") or {}
            if acl_state.get("source") in (None, "", "monthly_file"):
                acl_state.pop("row", None)
                acl_state.pop("label", None)
                acl_state.pop("history", None)
                mb["acl"] = acl_state
            flash(
                f"Removed {removed_name}." if removed_name
                else "Cleared monthly balance file state.",
                "success",
            )
            _save_state(state)
            return redirect(url_for("setup.step5_monthly_bal"))

        if action in ("save", "next"):
            # In per_month / manual modes the Save and Save & Next buttons
            # post the dedicated actions ``save_per_month_layout[_and_next]``
            # / ``save_manual[_and_next]`` which handle persistence. If we
            # land here in those modes it's a legacy bare action=next (e.g.
            # someone re-posting an older form): just advance without
            # touching the single-mode layout fields.
            if action == "next" and (mb.get("source") or "single") != "single":
                return redirect(url_for("setup.step5_grades"))
            # Guarded layout writes: only overwrite a previously-detected
            # layout value (sheet / header_row / pool_name_col / first_date_col)
            # if the submitted form actually carries a non-empty value for it.
            # This prevents partial form posts from silently clobbering the
            # auto-detected layout when a monthly_bal file is already uploaded.
            # Log a warning whenever we observe a clobber attempt so the
            # offending submitter (template / handler / JS) can be traced.
            _has_file = bool(mb.get("saved_path"))
            _form_sheet = (request.form.get("sheet", "") or "").strip()
            _form_hdr_raw = request.form.get("header_row", "")
            _form_pnc = (request.form.get("pool_name_col", "") or "").strip().upper()
            _form_fdc = (request.form.get("first_date_col", "") or "").strip().upper()

            def _guard(field: str, new_val, prior_val):
                if _has_file and prior_val and not new_val:
                    current_app.logger.warning(
                        "monthly_bal save: refusing to clobber %s=%r with empty "
                        "value (saved_path=%s, action=%s)",
                        field, prior_val, mb.get("saved_path"), action,
                    )
                    return prior_val
                return new_val if new_val else prior_val

            mb["sheet"] = _guard("sheet", _form_sheet, mb.get("sheet", ""))
            try:
                _new_hdr = int(_form_hdr_raw or 0)
            except ValueError:
                _new_hdr = 0
            mb["header_row"] = _guard(
                "header_row", _new_hdr, mb.get("header_row", 0)
            )
            mb["pool_name_col"] = _guard(
                "pool_name_col", _form_pnc, mb.get("pool_name_col", "")
            ) or "A"
            mb["first_date_col"] = _guard(
                "first_date_col", _form_fdc, mb.get("first_date_col", "")
            )
            mb["file_pattern"] = (request.form.get("file_pattern", "") or "").strip()
            mb["notes"] = (request.form.get("notes", "") or "").strip()
            # Re-extract pool/type labels from the (possibly user-adjusted)
            # layout. The auto-detection on upload only ever looks at
            # column A; if the user moved pool_name_col elsewhere (or
            # corrected the sheet / header_row), the parsed_pool_labels
            # list won't reflect the actual labels until we re-scan
            # against the new layout.
            if mb.get("saved_path") and mb.get("sheet") and mb.get("header_row") \
                    and mb.get("pool_name_col"):
                try:
                    _re = monthly_bal_parser.extract_pool_labels(
                        mb["saved_path"],
                        mb["sheet"],
                        int(mb["header_row"]),
                        mb["pool_name_col"],
                    )
                    if _re.get("ok") and _re.get("labels"):
                        mb["parsed_pool_labels"] = _re["labels"]
                        # Seed the pool_map from any prior label->pool
                        # context (WARM balance_title_map + historical
                        # hist_pool_map) for the newly-found labels.
                        _combined_map: dict[str, str] = {}
                        for _k, _v in (state.get("balance_title_map") or {}).items():
                            if _k:
                                _combined_map[_k] = (_v or "")
                        _hpm = state.get("hist_pool_map") or {}
                        for _k, _v in (_hpm.get("mapping") or {}).items():
                            if _k and _k not in _combined_map:
                                _combined_map[_k] = (_v or "")
                        _seeded, _status = monthly_bal_parser.seed_pool_map(
                            _re["labels"], _combined_map,
                        )
                        mb["label_status"] = _status
                        # The form-submitted pool_map is parsed below;
                        # we'll merge seed values for any labels the
                        # user hasn't already mapped.
                        mb["_seeded_pool_map"] = _seeded
                except Exception as exc:  # noqa: BLE001
                    current_app.logger.warning(
                        "monthly_bal save: re-extract pool labels failed: %s",
                        exc,
                    )
            # Persist any pool-label -> wizard-pool mapping rows.
            _form_map = _parse_kv_rows(
                request.form, "map_label", "map_pool"
            )
            # Merge any newly-seeded labels (from a column change) so
            # they appear in the pool-label mapping table on the next
            # render — but never overwrite a user-supplied selection.
            _seeded = mb.pop("_seeded_pool_map", {}) or {}
            for _lbl, _pool in _seeded.items():
                if _lbl not in _form_map:
                    _form_map[_lbl] = _pool
            # Defensive exact-match precedence: if a label case-
            # insensitively matches a configured pool name AND the
            # form-submitted mapping points at a DIFFERENT pool, the
            # form value is almost certainly a stale artifact of an
            # earlier keyword-heuristic auto-map that collapsed an
            # exact-match label onto a similar-category pool (e.g.
            # ``New Auto`` label → ``Indirect New Auto`` pool while a
            # ``New Auto`` pool exists on Step 2). Users can't easily
            # notice this in browser dropdowns and it silently breaks
            # Step 14's Balance Adjustment comparison. Correct the
            # exact-match collision here at save time so any stale
            # browser form can't re-pollute a good on-disk mapping.
            try:
                _ps_names = [
                    (p.get("name") or "").strip()
                    for p in (state.get("pool_settings") or [])
                    if isinstance(p, dict) and (p.get("name") or "").strip()
                ]
                _ps_by_cf = {n.casefold(): n for n in _ps_names}
                _exact_fixes: list[tuple[str, str, str]] = []
                for _lbl, _pool in list(_form_map.items()):
                    _lbl_cf = (_lbl or "").strip().casefold()
                    _exact = _ps_by_cf.get(_lbl_cf)
                    if _exact and (_pool or "").strip() != _exact:
                        _form_map[_lbl] = _exact
                        _exact_fixes.append((_lbl, _pool or "", _exact))
                if _exact_fixes:
                    _sample = ", ".join(
                        f"'{_a}' → '{_c}' (was '{_b or '—'}')"
                        for _a, _b, _c in _exact_fixes[:3]
                    )
                    _more = (
                        f" (+{len(_exact_fixes) - 3} more)"
                        if len(_exact_fixes) > 3 else ""
                    )
                    flash(
                        "Auto-corrected pool mapping where a balance "
                        "label exactly matches a configured pool: "
                        f"{_sample}{_more}.",
                        "info",
                    )
            except Exception:  # noqa: BLE001
                pass
            mb["pool_map"] = _form_map
            # Persist the ACL source selection + per-source values.
            _save_acl_form(mb, request.form)
            state["monthly_bal"] = mb
            _save_state(state)
            if action == "next":
                # Both flows: monthly_bal → grades → credit_pull → sample.
                return redirect(url_for("setup.step5_grades"))
            flash("Saved monthly balance file settings.", "success")
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "acl_set_source":
            # Lightweight handler invoked by the ACL-source radios'
            # onchange auto-submit. Persists ONLY the selected source
            # so the page re-renders with the matching sub-panel
            # (monthly_file row override, separate-file uploader, or
            # manual three-month entry table) without running the full
            # single-mode save path against potentially-empty fields.
            src = (request.form.get("acl_source") or "").strip()
            if src in ("monthly_file", "separate", "manual"):
                mb.setdefault("acl", {})["source"] = src
                state["monthly_bal"] = mb
                _save_state(state)
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "acl_scan_balance_files":
            # Re-scan all already-uploaded balance files (per-month or
            # per-year) for the ACL/ALLL line and merge into
            # mb["acl"]["history"]. Preserves any existing row override.
            _save_acl_form(mb, request.form)
            acl_state = mb.setdefault("acl", {})
            hist = acl_state.get("history") or {}
            scanned = 0
            populated = 0
            for yf in (mb.get("year_files") or []):
                sp = yf.get("saved_path")
                if not sp or not Path(sp).is_file():
                    continue
                scanned += 1
                try:
                    res = monthly_bal_parser.analyse_per_year_file(sp)
                except Exception:  # noqa: BLE001
                    continue
                ah = res.get("acl_history") or {}
                if ah:
                    if not acl_state.get("row") and res.get("acl_row"):
                        acl_state["row"] = int(res["acl_row"])
                        acl_state["label"] = res.get("acl_label", "")
                    for k, v in ah.items():
                        hist[k] = float(v)
                        populated += 1
            for mf in (mb.get("monthly_files") or []):
                sp = mf.get("saved_path")
                period = mf.get("period")
                if not sp or not Path(sp).is_file() or not period:
                    continue
                scanned += 1
                try:
                    res = monthly_bal_parser.analyse_per_month_file(
                        sp,
                        acl_row=acl_state.get("row") or None,
                        acl_col=acl_state.get("col") or None,
                    )
                except Exception:  # noqa: BLE001
                    continue
                if res.get("acl_value") is not None:
                    if not acl_state.get("row") and res.get("acl_row"):
                        acl_state["row"] = int(res["acl_row"])
                        acl_state["label"] = res.get("acl_label", "")
                    elif acl_state.get("row") and res.get("acl_label"):
                        # Refresh the displayed label from the row we read.
                        acl_state["label"] = res.get("acl_label", "")
                    hist[period] = float(res["acl_value"])
                    populated += 1
            acl_state["history"] = hist
            mb["acl"] = acl_state
            state["monthly_bal"] = mb
            _save_state(state)
            if scanned == 0:
                flash("No uploaded balance files to scan.", "warning")
            elif populated == 0:
                flash(
                    f"Scanned {scanned} balance file(s), but no ACL/ALLL "
                    "row was found. Use a Separate file or Manual entry.",
                    "warning",
                )
            else:
                flash(
                    f"Scanned {scanned} balance file(s); captured "
                    f"{populated} ACL value(s).",
                    "success",
                )
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "acl_refresh_row":
            # User changed the ACL row number; re-extract that row's history.
            _save_acl_form(mb, request.form)
            try:
                target_row = int(request.form.get("acl_row") or 0)
            except ValueError:
                target_row = 0
            mb_source = (mb.get("source") or "").strip()
            saved_path = mb.get("saved_path")
            if not target_row:
                flash("Pick a row number before refreshing.", "error")
            elif mb_source == "per_month":
                # Walk every uploaded per-month file, reading the user's
                # chosen row at the layout's balance column. This lets
                # the user override auto-detect when it picks the wrong
                # ACL line.
                acl_state = mb.setdefault("acl", {})
                acl_state["row"] = target_row
                hist: dict[str, float] = {}
                scanned = 0
                populated = 0
                last_label = ""
                for mf in (mb.get("monthly_files") or []):
                    sp = mf.get("saved_path")
                    period = mf.get("period")
                    if not sp or not Path(sp).is_file() or not period:
                        continue
                    scanned += 1
                    try:
                        res = monthly_bal_parser.analyse_per_month_file(
                            sp, acl_row=target_row,
                            acl_col=acl_state.get("col") or None,
                        )
                    except Exception:  # noqa: BLE001
                        continue
                    if res.get("acl_value") is not None:
                        hist[period] = float(res["acl_value"])
                        populated += 1
                        if res.get("acl_label"):
                            last_label = res["acl_label"]
                acl_state["history"] = hist
                if last_label:
                    acl_state["label"] = last_label
                state["monthly_bal"] = mb
                _save_state(state)
                if scanned == 0:
                    flash(
                        "No uploaded per-month files to refresh.",
                        "warning",
                    )
                elif populated == 0:
                    flash(
                        f"Read row {target_row} from {scanned} file(s) "
                        "but found no numeric balances. Double-check "
                        "the row number.",
                        "warning",
                    )
                else:
                    flash(
                        f"Refreshed ACL row {target_row}"
                        + (f" ({last_label})" if last_label else "")
                        + f": {populated} of {scanned} month(s) parsed.",
                        "success",
                    )
                return redirect(url_for("setup.step5_monthly_bal"))
            elif not saved_path:
                flash(
                    "Upload a monthly balance file and pick a row number "
                    "before refreshing.",
                    "error",
                )
            else:
                res = monthly_bal_parser.extract_row_history(
                    saved_path, mb.get("sheet", ""), mb.get("header_row", 0),
                    target_row,
                )
                if res.get("ok"):
                    acl_state = mb.setdefault("acl", {})
                    acl_state["row"] = target_row
                    acl_state["label"] = res.get("label", "")
                    acl_state["history"] = res.get("history", {})
                    flash(
                        f"Refreshed ACL row {target_row} "
                        f"({res.get('label','')}): "
                        f"{len(res.get('history') or {})} month(s) parsed.",
                        "success",
                    )
                else:
                    flash(f"ACL refresh failed: {res.get('error')}", "error")
            state["monthly_bal"] = mb
            _save_state(state)
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "acl_upload_separate":
            _save_acl_form(mb, request.form)
            f = request.files.get("acl_separate_file")
            if not f or not f.filename:
                flash("Choose a file containing the ACL value to upload.", "error")
            else:
                try:
                    target = _save_acl_file_upload(f)
                    sep = mb.setdefault("acl", {}).setdefault("separate_file", {})
                    sep["filename"] = target.name
                    sep["saved_path"] = str(target)
                    flash(
                        f"Uploaded ACL file: {target.name}. "
                        f"Enter the sheet + cell to extract from below.",
                        "success",
                    )
                except Exception as exc:  # noqa: BLE001
                    flash(f"ACL file upload failed: {exc}", "error")
            state["monthly_bal"] = mb
            _save_state(state)
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "acl_extract_separate":
            # Read a single cell from the uploaded standalone ACL file.
            _save_acl_form(mb, request.form)
            sep = mb.setdefault("acl", {}).setdefault("separate_file", {})
            saved_path = sep.get("saved_path")
            sheet = sep.get("sheet", "")
            cell = (sep.get("cell") or "").strip().upper()
            if not saved_path or not sheet or not cell:
                flash("Provide a saved file, sheet name, and cell address "
                      "(e.g. B14) before extracting.", "error")
            else:
                try:
                    from openpyxl import load_workbook as _lw
                    wb = _lw(saved_path, read_only=True, data_only=True)
                    if sheet not in wb.sheetnames:
                        flash(f"Sheet '{sheet}' not found in {sep['filename']}.",
                              "error")
                    else:
                        raw = wb[sheet][cell].value
                        v = monthly_bal_parser._coerce_number(raw)
                        if v is None:
                            flash(f"Cell {sheet}!{cell} contained no numeric "
                                  f"value (got: {raw!r}).", "warning")
                        else:
                            sep["value"] = abs(v)
                            flash(f"Extracted ACL value from {sheet}!{cell}: "
                                  f"${abs(v):,.2f}", "success")
                    wb.close()
                except Exception as exc:  # noqa: BLE001
                    flash(f"Could not read ACL value: {exc}", "error")
            state["monthly_bal"] = mb
            _save_state(state)
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "set_source":
            src = (request.form.get("monthly_bal_source") or "").strip()
            if src in ("single", "per_month", "manual"):
                mb["source"] = src
                _save_state(state)
                flash(f"Switched to '{src}' source mode.", "success")
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "upload_per_month":
            f = request.files.get("per_month_file")
            period = (request.form.get("per_month_period") or "").strip()
            if not f or not f.filename:
                flash("Choose a file to upload.", "error")
            else:
                try:
                    target = _save_monthly_bal_upload(f)
                    category, message = _ingest_per_month_workbook(
                        state, mb, target, period)
                    flash(message, category)
                except Exception as exc:  # noqa: BLE001
                    flash(f"Upload failed: {exc}", "error")
            _save_state(state)
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "remove_per_month":
            target_name = (request.form.get("filename") or "").strip()
            files = mb.get("monthly_files") or []
            removed = None
            for e in files:
                if e.get("filename") == target_name:
                    removed = e
                    break
            if removed:
                files.remove(removed)
                # Best-effort disk cleanup — but only for files we
                # actually copied into the managed upload dir. Folder-
                # scanned entries reference the user's original file
                # and must NEVER be deleted.
                if not removed.get("external"):
                    sp = removed.get("saved_path") or ""
                    if sp:
                        try:
                            p = Path(sp)
                            if p.is_file():
                                p.unlink()
                        except Exception:  # noqa: BLE001
                            pass
                flash(f"Removed {target_name}.", "success")
            else:
                flash(f"File '{target_name}' not found.", "error")
            mb["monthly_files"] = files
            _save_state(state)
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "scan_per_month_folder":
            folder_raw = (request.form.get("scan_folder") or "").strip()
            pattern_raw = (request.form.get("scan_pattern") or "").strip()
            mb["per_month_source_folder"] = folder_raw
            if pattern_raw:
                mb["file_pattern"] = pattern_raw
            if not folder_raw:
                flash("Enter a folder path to scan.", "error")
                _save_state(state)
                return redirect(url_for("setup.step5_monthly_bal"))
            try:
                folder = Path(folder_raw).expanduser()
            except Exception as exc:  # noqa: BLE001
                flash(f"Invalid folder path: {exc}", "error")
                _save_state(state)
                return redirect(url_for("setup.step5_monthly_bal"))
            if not folder.is_dir():
                flash(
                    f"Folder not found or not a directory: {folder}",
                    "error",
                )
                _save_state(state)
                return redirect(url_for("setup.step5_monthly_bal"))

            # Compile the regex once. Empty pattern = match every
            # supported balance-sheet file.
            try:
                fp_rx = re.compile(pattern_raw, re.IGNORECASE) \
                    if pattern_raw else None
            except re.error as exc:
                flash(f"Filename pattern is not a valid regex: {exc}",
                      "error")
                _save_state(state)
                return redirect(url_for("setup.step5_monthly_bal"))

            supported_exts = {".xls", ".xlsx", ".xlsm", ".csv"}
            files = mb.setdefault("monthly_files", [])
            # Build period+filename indexes for de-dupe.
            existing_by_name = {e.get("filename"): e for e in files}
            existing_periods = {e.get("period") for e in files
                                if e.get("period")}

            scanned = 0
            added: list[str] = []
            replaced: list[str] = []
            skipped_no_period: list[str] = []
            skipped_parse: list[str] = []
            for entry_path in sorted(folder.iterdir()):
                if not entry_path.is_file():
                    continue
                if entry_path.suffix.lower() not in supported_exts:
                    continue
                if fp_rx is not None and not fp_rx.search(entry_path.name):
                    continue
                scanned += 1
                analysis = monthly_bal_parser.analyse_per_month_file(
                    entry_path,
                    acl_row=((mb.get("acl") or {}).get("row") or None),
                    acl_col=((mb.get("acl") or {}).get("col") or None),
                )
                period = analysis.get("detected_period") or ""
                if not period:
                    skipped_no_period.append(entry_path.name)
                    continue
                if not analysis.get("ok") and not analysis.get(
                        "parsed_pool_labels"):
                    # Parser failed AND we found nothing useful — still
                    # add the entry so the importer can try with the
                    # saved layout, but warn.
                    skipped_parse.append(entry_path.name)
                # Auto-capture ACL value into mb["acl"]["history"].
                _acl_row = analysis.get("acl_row")
                _acl_val = analysis.get("acl_value")
                if _acl_row and _acl_val is not None:
                    _acl_state = mb.setdefault("acl", {})
                    if not _acl_state.get("row"):
                        _acl_state["row"] = int(_acl_row)
                        _acl_state["label"] = analysis.get("acl_label", "")
                    _hist = _acl_state.get("history") or {}
                    _hist[period] = float(_acl_val)
                    _acl_state["history"] = _hist
                entry = {
                    "filename": entry_path.name,
                    "saved_path": str(entry_path),
                    "period": period,
                    "external": True,
                }
                if entry_path.name in existing_by_name:
                    files[:] = [e for e in files
                                if e.get("filename") != entry_path.name]
                    files.append(entry)
                    replaced.append(entry_path.name)
                elif period in existing_periods:
                    # Skip duplicate periods so we don't double-count.
                    continue
                else:
                    files.append(entry)
                    existing_periods.add(period)
                    added.append(entry_path.name)

            files.sort(key=lambda e: e.get("period") or "")

            msgs = []
            if added:
                msgs.append(f"added {len(added)} new file(s)")
            if replaced:
                msgs.append(f"refreshed {len(replaced)}")
            if skipped_no_period:
                msgs.append(
                    f"skipped {len(skipped_no_period)} with no detectable "
                    f"month-end (rename them to include YYYYMMDD or add "
                    "an 'As of:' cell)"
                )
            if skipped_parse:
                msgs.append(
                    f"{len(skipped_parse)} file(s) couldn't be parsed; "
                    "they'll be retried at run time using the saved layout"
                )
            if scanned == 0:
                flash(
                    f"Scanned {folder}: no matching files found "
                    f"(pattern={pattern_raw!r}).",
                    "warning",
                )
            else:
                flash(
                    f"Scanned {folder}: {scanned} file(s) matched; "
                    + ", ".join(msgs) + ".",
                    "success" if added or replaced else "warning",
                )
            _save_state(state)
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "save_per_month_layout":
            _persist_per_month_layout(mb, request.form)
            _save_state(state)
            flash("Saved per-month layout.", "success")
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "save_per_month_layout_and_next":
            _persist_per_month_layout(mb, request.form)
            _save_state(state)
            flash("Saved per-month layout.", "success")
            return redirect(url_for("setup.step5_grades"))

        if action == "upload_per_year":
            f = request.files.get("per_year_file")
            year_raw = (request.form.get("per_year_year") or "").strip()
            if not f or not f.filename:
                flash("Choose an annual balance workbook to upload.", "error")
                _save_state(state)
                return redirect(url_for("setup.step5_monthly_bal"))
            try:
                target = _save_monthly_bal_upload(f)
                category, message = _ingest_annual_workbook(
                    state, mb, target, year_raw)
                flash(message, category)
            except Exception as exc:  # noqa: BLE001
                flash(f"Upload failed: {exc}", "error")
            _save_state(state)
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "remove_per_year":
            target_name = (request.form.get("filename") or "").strip()
            files = mb.get("year_files") or []
            removed = None
            for e in files:
                if e.get("filename") == target_name:
                    removed = e
                    break
            if removed:
                files.remove(removed)
                if not removed.get("external"):
                    sp = removed.get("saved_path") or ""
                    if sp:
                        try:
                            p = Path(sp)
                            if p.is_file():
                                p.unlink()
                        except Exception:  # noqa: BLE001
                            pass
                flash(f"Removed {target_name}.", "success")
            else:
                flash(f"File '{target_name}' not found.", "error")
            mb["year_files"] = files
            _save_state(state)
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "save_per_year_layout":
            _persist_per_year_layout(mb, request.form)
            _save_state(state)
            flash("Saved per-year layout.", "success")
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "save_per_year_layout_and_next":
            _persist_per_year_layout(mb, request.form)
            _save_state(state)
            flash("Saved per-year layout.", "success")
            return redirect(url_for("setup.step5_grades"))

        if action == "save_manual":
            n_pools = _persist_manual_grid(mb, request.form)
            _save_state(state)
            flash(
                f"Saved manual balances for {n_pools} pool(s) × "
                f"{len(mb.get('manual_months') or [])} month(s).",
                "success",
            )
            return redirect(url_for("setup.step5_monthly_bal"))

        if action == "save_manual_and_next":
            n_pools = _persist_manual_grid(mb, request.form)
            _save_state(state)
            flash(
                f"Saved manual balances for {n_pools} pool(s) × "
                f"{len(mb.get('manual_months') or [])} month(s).",
                "success",
            )
            return redirect(url_for("setup.step5_grades"))

        if action == "back":
            return redirect(url_for("setup.step_dq_hist"))

    # Pool choices for the mapping dropdowns: prefer the editable Step 2
    # pool_settings list (so what the user sees here matches what they
    # configured on the Pools step), then fall back to WARM-derived pool
    # names, then to whatever pools showed up in pool_map.
    pool_choices: list[str] = []
    for p in (state.get("pool_settings") or []):
        if not isinstance(p, dict):
            continue
        name = (p.get("name") or "").strip()
        if name and not p.get("excluded") and name not in pool_choices:
            pool_choices.append(name)
    if not pool_choices and has_warm:
        warm_pools = ((state.get("warm") or {}).get("pools")) or []
        for p in warm_pools:
            if isinstance(p, dict):
                name = p.get("name", "")
            else:
                name = str(p or "")
            if name and name not in pool_choices:
                pool_choices.append(name)
    if not pool_choices:
        pool_choices = sorted({v for v in (state.get("pool_map") or {}).values() if v})

    # Compute latest balance per parsed label for the Pool Label Mapping
    # table (helps users decide which pool to map each label to).  Same
    # data the runtime aggregator uses, just with no label_to_pool so we
    # can read raw_rows directly. Always wrapped in try/except — never
    # break the page render on a parser hiccup.
    label_balances: dict[str, float] = {}
    label_balance_period: str = ""
    try:
        src = (mb.get("source") or "").strip()
        if src == "per_month":
            files = mb.get("monthly_files") or []
            layout = mb.get("per_month_layout") or {}
            if files and layout:
                res = monthly_bal_parser.pool_balances_for_per_month_files(
                    files, layout, label_to_pool=None,
                )
                by_period = (res or {}).get("by_period") or {}
                if by_period:
                    latest = max(by_period.keys())
                    label_balance_period = latest
                    for row in (by_period[latest].get("raw_rows") or []):
                        lab = (row.get("label") or "").strip()
                        bal = row.get("balance")
                        if lab and isinstance(bal, (int, float)):
                            label_balances[lab.lower()] = float(bal)
        elif src == "per_year":
            files = mb.get("year_files") or []
            layout = mb.get("per_year_layout") or {}
            if files and layout:
                res = monthly_bal_parser.pool_balances_for_per_year_files(
                    files, layout, label_to_pool=None,
                )
                by_period = (res or {}).get("by_period") or {}
                if by_period:
                    latest = max(by_period.keys())
                    label_balance_period = latest
                    for row in (by_period[latest].get("raw_rows") or []):
                        lab = (row.get("label") or "").strip()
                        bal = row.get("balance")
                        if lab and isinstance(bal, (int, float)):
                            label_balances[lab.lower()] = float(bal)
    except Exception:  # noqa: BLE001
        label_balances = {}
        label_balance_period = ""

    return render_template(
        "setup/step5_monthly_bal.html",
        mb=mb,
        pool_choices=pool_choices,
        has_warm=has_warm,
        files_used=_monthly_balance_files_used(state),
        label_balances=label_balances,
        label_balance_period=label_balance_period,
        **_wizard_ctx("monthly_bal"),
    )


# =================================================================
# Step 3 — Sample file upload (auto-suggest mappings)
# =================================================================

_SAMPLE_DIR = Path(tempfile.gettempdir()) / "cecl_ui_samples"


def _save_sample_upload(file_storage) -> Path:
    _SAMPLE_DIR.mkdir(parents=True, exist_ok=True)
    fn = secure_filename(file_storage.filename or "sample")
    target = _SAMPLE_DIR / fn
    file_storage.save(target)
    return target


def _loan_data_entry_analysis(analysis: dict[str, Any]) -> dict[str, Any]:
    """Slim copy of a sample-parser analysis suitable for caching on a
    `loan_data_files` entry. Keeps headers / column suggestions / first few
    sample rows so the per-file Column Mappings step can render dropdowns
    and a preview without re-parsing the file on every page render."""
    return {
        "headers": list(analysis.get("headers") or []),
        "column_suggestions": dict(analysis.get("column_suggestions") or {}),
        "pool_code_suggestions": list(analysis.get("pool_code_suggestions") or []),
        "sample_rows": list(analysis.get("sample_rows") or [])[:5],
        "has_header": bool(analysis.get("has_header")),
        "header_row": int(analysis.get("header_row") or 0),
        # Carry the parser's auto-detected member-account layout (if any)
        # so Step 3 can show a "✓ auto-detected from sample" hint and
        # ``_seed_loan_data_entry_mapping`` can prefer it over the
        # historical hardcoded ``fixed_suffix=3`` default.
        "member_account_suggestion": dict(analysis.get("member_account_suggestion") or {})
            if analysis.get("member_account_suggestion") else None,
        # Date format the filename heuristic guessed for this sample
        # (one of ``sample_parser.ACCEPTED_DATE_FORMATS``). Surfaced to
        # Step 2 so the user sees what value will land in the YAML.
        "date_format": str(analysis.get("date_format") or "YYYY-MM"),
    }


def _norm_filename_for_match(name: str) -> str:
    """Normalize a filename for fuzzy matching: drop extension, lowercase,
    strip everything except [a-z0-9]. Lets us match
    'Credit Cardholder - AIRES 03312026 v2.xlsx' against
    'Credit_Cardholder_-_AIRES_03312026_v2.xlsx' regardless of separators."""
    import os as _os
    import re as _re
    if not name:
        return ""
    stem = _os.path.splitext(_os.path.basename(name))[0]
    return _re.sub(r"[^a-z0-9]", "", stem.lower())


def _filename_from_profile_label(label: str) -> str:
    """Extract the source filename from a profile label of the form
    'Profile N (filename.xlsx)'. Returns '' if no parens present."""
    if not label:
        return ""
    i = label.rfind("(")
    j = label.rfind(")")
    if i == -1 or j == -1 or j <= i:
        return ""
    return label[i + 1:j]


def _seed_loan_data_entry_mapping(
    state: dict[str, Any], entry: dict[str, Any], analysis: dict[str, Any]
) -> str | None:
    """Compute a header signature for ``analysis`` and seed the per-file
    ``column_mappings`` / ``member_account`` on ``entry``.

    If a Step 2 historical-extract profile matches the same signature, copy
    its saved mapping verbatim. Otherwise fall back to the parser's
    column_suggestions. Returns the matched profile label (or None if no
    match) so the caller can surface a user-visible flash.
    """
    headers = list(analysis.get("headers") or [])
    sig = ""
    try:
        sig = extract_hist_service.compute_header_signature(headers)
    except Exception:  # noqa: BLE001
        sig = ""
    entry["signature"] = sig

    # Look for a Step 2 profile with the same header signature.
    he = state.get("hist_extracts") or {}
    matched = None
    if sig:
        for prof in (he.get("profiles") or []):
            if prof.get("signature") == sig and (prof.get("column_mappings") or {}):
                matched = prof
                break

    # Fallback: same file uploaded on Step 6 with a different has_header
    # setting (or column-letter placeholders) produces a different
    # signature than the Step 2 profile. Match by (filename + column
    # count) so the user still gets their mapping auto-copied.
    if not matched:
        entry_name = str(entry.get("name") or "")
        col_count = len(headers)
        norm_entry = _norm_filename_for_match(entry_name)
        if norm_entry and col_count:
            for prof in (he.get("profiles") or []):
                if not (prof.get("column_mappings") or {}):
                    continue
                if len(prof.get("headers") or []) != col_count:
                    continue
                prof_name = (
                    prof.get("source_filename")
                    or _filename_from_profile_label(prof.get("label") or "")
                )
                if _norm_filename_for_match(prof_name) == norm_entry:
                    matched = prof
                    break

    if matched:
        entry["column_mappings"] = dict(matched.get("column_mappings") or {})
        entry["member_account"] = dict(matched.get("member_account") or {
            "mode": "fixed_suffix", "suffix_length": 3, "delimiter": "-",
        })
        return matched.get("label") or matched.get("id") or "historical profile"

    # No match — seed from sample-parser column_suggestions if entry is
    # not already mapped. Never destroys an existing mapping on the entry.
    if not entry.get("column_mappings"):
        entry["column_mappings"] = dict(analysis.get("column_suggestions") or {})
    if not entry.get("member_account"):
        # Prefer the parser's per-file inference over the historical
        # ``fixed_suffix=3`` default. The inferred dict already carries
        # mode/delimiter/suffix_length keys; we strip the metadata keys
        # (``confidence``, ``rationale``) before storing on the entry
        # so the YAML emit path stays clean.
        sug = analysis.get("member_account_suggestion") or {}
        if sug.get("mode"):
            entry["member_account"] = {
                "mode": sug.get("mode"),
                "suffix_length": int(sug.get("suffix_length") or 0),
                "delimiter": sug.get("delimiter") or "-",
            }
        else:
            entry["member_account"] = {
                "mode": "fixed_suffix", "suffix_length": 3, "delimiter": "-",
            }

    # WARM headerless overlay: for a positional (header-less) core extract the
    # WARM Data-tab column trace gives authoritative 0-based positions
    # (Franklin/Bridgeton AIRES files). Prefer them over the parser's
    # col_<LETTER> suggestions so the file imports correctly, and mark the
    # extract header-less so build_yaml emits integer column_mappings.
    if not bool(analysis.get("has_header")):
        widx = {k: int(v)
                for k, v in (state.get("_warm_column_indices") or {}).items()
                if isinstance(v, int)}
        if widx:
            ecm = dict(entry.get("column_mappings") or {})
            ecm.update(widx)
            entry["column_mappings"] = ecm
            entry["has_header"] = False
            state["has_header"] = False
            tcm = dict(state.get("column_mappings") or {})
            tcm.update(widx)
            state["column_mappings"] = tcm
    return None


def _snapshot_pool_map_history(state: dict[str, Any]) -> int:
    """Capture non-blank ``state["pool_map"]`` entries into
    ``state["_pool_map_history"]`` (a durable code -> pool memory).

    Idempotent. Only writes when a code's current pool_map value is
    non-blank AND either not present in history or differs. New history
    entries never overwrite non-blank existing history unless the current
    pool_map value differs (user reassignment wins). Returns the number
    of history entries captured or updated.

    The history exists to survive an extract delete-and-re-add cycle
    which otherwise wipes ``state["pool_map"]`` values (see
    ``_apply_sample_to_state``'s empty-seed branch that re-seeds fresh
    codes with blank values). The seed path consults history to
    restore user assignments transparently.
    """
    pm = state.get("pool_map") or {}
    if not isinstance(pm, dict) or not pm:
        return 0
    hist = state.setdefault("_pool_map_history", {})
    if not isinstance(hist, dict):
        hist = {}
        state["_pool_map_history"] = hist
    n = 0
    for code, pool in pm.items():
        if not isinstance(code, str):
            continue
        val = (pool or "").strip() if isinstance(pool, str) else ""
        if not val:
            continue
        prev = hist.get(code)
        if prev != val:
            hist[code] = val
            n += 1
    return n


def _apply_sample_to_state(state: dict[str, Any], analysis: dict[str, Any]) -> None:
    """Merge sample-file suggestions into wizard state.

    Only fills fields the user hasn't deviated from defaults; never destroys
    pre-existing user input on a field-by-field basis.
    """
    defaults = _default_state()

    # File patterns — always overwrite (they were just defaults).
    if state.get("file_pattern") in ("", defaults["file_pattern"]):
        state["file_pattern"] = analysis["file_pattern"]
    if state.get("date_pattern") in ("", defaults["date_pattern"]):
        state["date_pattern"] = analysis["date_pattern"]
    # Date format — only auto-set when the user is still on the default
    # so a fresh sample upload doesn't clobber a deliberate Step 2 pick.
    if state.get("date_format") in ("", defaults["date_format"]):
        guessed_df = analysis.get("date_format")
        if guessed_df and guessed_df in sample_parser.ACCEPTED_DATE_FORMATS:
            state["date_format"] = guessed_df

    # Column mappings — overwrite per-field where the suggestion is non-empty
    # AND the user is still on the default value for that field.
    for sys_field, suggested_header in analysis["column_suggestions"].items():
        current = state["column_mappings"].get(sys_field, "")
        default_val = defaults["column_mappings"].get(sys_field, "")
        if not current or current == default_val:
            state["column_mappings"][sys_field] = suggested_header

    # Cross-CU learned suggestions — fill any field the sample parser
    # couldn't auto-detect, using the most common header other credit
    # unions have mapped to that field. Restricted to headers that
    # actually exist in this sample. Never overrides a non-empty mapping.
    headers_now = analysis.get("headers") or []
    if headers_now:
        already_set = {
            f for f, v in state["column_mappings"].items() if v
        }
        learned = column_mapping_suggestions.suggest_for_headers(
            headers_now, skip_fields=already_set
        )
        for sys_field, learned_header in learned.items():
            if not state["column_mappings"].get(sys_field):
                state["column_mappings"][sys_field] = learned_header

    # Pool map — if the user is still on the default seed map, replace it
    # with placeholders for every distinct code we found. Blank by default,
    # but any code we've previously seen assigned (in
    # ``state["_pool_map_history"]``) is restored to its remembered pool
    # so a Step 12 extract delete-and-re-add cycle doesn't wipe Step 3
    # mappings the user configured earlier in the session.
    if state["pool_map"] == defaults["pool_map"] and analysis["pool_code_suggestions"]:
        hist = state.get("_pool_map_history") or {}
        state["pool_map"] = {
            code: hist.get(code, "") for code in analysis["pool_code_suggestions"]
        }

    state["sample"] = analysis


# =================================================================
# Step 3 — Historical Data (WARM workbooks + supplemental files)
# =================================================================

_HIST_DIR = Path(tempfile.gettempdir()) / "cecl_ui_hist"


def _save_hist_upload(file_storage, subfolder: str = "") -> Path:
    dest = _HIST_DIR / subfolder if subfolder else _HIST_DIR
    dest.mkdir(parents=True, exist_ok=True)
    fn = secure_filename(file_storage.filename or "upload.xlsx")
    target = dest / fn
    file_storage.save(target)
    return target


def _save_hist_folder_uploads(
    files,
    folder_labels: list[str],
    include_paths: set[str] | None = None,
) -> Path:
    scans_root = _HIST_DIR / "folder_scans"
    scans_root.mkdir(parents=True, exist_ok=True)
    target_root = Path(tempfile.mkdtemp(prefix="scan_", dir=str(scans_root)))

    for index, file_storage in enumerate(files):
        if not file_storage or not file_storage.filename:
            continue
        rel_original = (file_storage.filename or "upload").replace("\\", "/")
        if include_paths is not None and rel_original not in include_paths:
            continue
        rel_name = rel_original
        rel_parts = [secure_filename(part) for part in rel_name.split("/") if part and part not in (".", "..")]
        if not rel_parts:
            rel_parts = ["upload"]
        target = target_root
        for part in rel_parts[:-1]:
            target /= part
        target.mkdir(parents=True, exist_ok=True)
        file_storage.save(target / rel_parts[-1])
    return target_root


def _hist_period(filename: str) -> str | None:
    m = re.search(r"(20\d{2})[-_\s](\d{2})", filename)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    month_map = {
        "jan": "01", "feb": "02", "mar": "03", "apr": "04",
        "may": "05", "jun": "06", "jul": "07", "aug": "08",
        "sep": "09", "sept": "09", "oct": "10", "nov": "11", "dec": "12",
    }
    m2 = re.search(
        r"(jan|feb|mar|apr|may|jun|jul|aug|sep(?:t)?|oct|nov|dec)[a-z]*[-_\s]+?(20\d{2})",
        filename,
        re.IGNORECASE,
    )
    if m2:
        mon = month_map.get(m2.group(1).lower()[:3], "")
        if mon:
            return f"{m2.group(2)}-{mon}"
    return None


def _add_hist_file(state: dict[str, Any], key: str, file_storage, subfolder: str) -> str:
    saved = _save_hist_upload(file_storage, subfolder=subfolder)
    entry = {"name": saved.name, "period": _hist_period(file_storage.filename or saved.name), "path": str(saved)}
    existing = state.get("hist_scan") or {}
    rows = [e for e in (existing.get(key) or []) if e.get("name") != saved.name]
    rows.append(entry)
    existing[key] = rows
    existing.setdefault("ok", True)
    existing.setdefault("folder", "Uploaded files")
    for list_key in (
        "warm_files", "co_files", "recov_files", "impaired_files",
        "credit_pull_files", "monthly_files", "loan_data_files", "other_files",
        "monthly_co_files", "monthly_recov_files",
    ):
        existing.setdefault(list_key, [])
    state["hist_scan"] = existing
    _save_state(state)
    return saved.name


def _add_hist_file_from_path(
    state: dict[str, Any], key: str, src_path: Path, subfolder: str
) -> str:
    """Like ``_add_hist_file`` but ingests an existing server-side file.

    Copies ``src_path`` into the wizard's temp dir so subsequent reads come
    from a stable location, then registers it in ``state.hist_scan[key]``.
    """
    import shutil

    dest_dir = _HIST_DIR / subfolder if subfolder else _HIST_DIR
    dest_dir.mkdir(parents=True, exist_ok=True)
    fn = secure_filename(src_path.name) or "upload.xlsx"
    target = dest_dir / fn
    if target.resolve() != src_path.resolve():
        shutil.copy2(src_path, target)
    entry = {
        "name": target.name,
        "period": _hist_period(src_path.name),
        "path": str(target),
    }
    existing = state.get("hist_scan") or {}
    rows = [e for e in (existing.get(key) or []) if e.get("name") != target.name]
    rows.append(entry)
    existing[key] = rows
    existing.setdefault("ok", True)
    existing.setdefault("folder", "Uploaded files")
    for list_key in (
        "warm_files", "co_files", "recov_files", "impaired_files",
        "credit_pull_files", "monthly_files", "loan_data_files", "other_files",
        "monthly_co_files", "monthly_recov_files",
    ):
        existing.setdefault(list_key, [])
    state["hist_scan"] = existing
    _save_state(state)
    return target.name


def _normalize_folder_path(folder_str: str) -> str:
    """Trim whitespace + surrounding quotes a user may have pasted in."""
    s = (folder_str or "").strip()
    if len(s) >= 2 and s[0] in ('"', "'") and s[-1] == s[0]:
        s = s[1:-1].strip()
    return s


def _flash_co_recov_scan_result(res: dict[str, Any], kind: str) -> None:
    """Standard flash output for the CO/Recov folder-scan helpers."""
    kind_label = "charge-off" if kind == "co" else "recoveries"
    if not res.get("scanned") and res.get("error"):
        flash(f"{kind_label.capitalize()} folder scan failed: {res['error']}",
              "error")
        return
    parts = [f"Scanned {res.get('scanned', 0)} {kind_label} file(s)"]
    if res.get("added"):
        parts.append(f"added {len(res['added'])}")
    if res.get("ignored"):
        parts.append(f"skipped {len(res['ignored'])} via ignored layouts")
    if res.get("skipped"):
        parts.append(f"failed {len(res['skipped'])}")
    msg = "; ".join(parts) + "."
    flash(msg, "success" if res.get("added") or res.get("ignored") else "info")


_MONTHLY_CO_RECOV_GLOBS = ("*.xlsx", "*.xlsm", "*.xls", "*.csv")


def _discover_co_recov_files(folder: Path) -> list[Path]:
    """Recursive glob for ``_MONTHLY_CO_RECOV_GLOBS`` under ``folder``,
    dropping hidden/temp files and de-duping by lowercased basename.
    """
    found: list[Path] = []
    for pat in _MONTHLY_CO_RECOV_GLOBS:
        found.extend(folder.rglob(pat))
    seen: set[str] = set()
    files: list[Path] = []
    for p in sorted(found, key=lambda x: x.name.lower()):
        if p.name.startswith("~$") or p.name.startswith("."):
            continue
        try:
            rel_parts = p.relative_to(folder).parts[:-1]
        except ValueError:
            rel_parts = ()
        if any(part.startswith("~$") or part.startswith(".")
               for part in rel_parts):
            continue
        key = p.name.lower()
        if key in seen:
            continue
        seen.add(key)
        files.append(p)
    return files


def _file_layout_signature(path: Path) -> tuple[str, list[str]]:
    """Return ``(signature, headers)`` for a CO/Recov file.

    Returns ``("", [])`` on failure rather than raising so a single bad
    file doesn't abort an entire folder scan.
    """
    try:
        head = extract_hist_service.read_extract_headers(path)
        if not head.get("ok"):
            return "", []
        headers = head.get("headers") or []
        return extract_hist_service.compute_header_signature(headers), headers
    except Exception:  # noqa: BLE001
        return "", []


def _co_recov_ignored_set(state: dict[str, Any], kind: str) -> set[str]:
    key = "co_ignored_signatures" if kind == "co" else "recov_ignored_signatures"
    return set(state.get(key) or [])


def _co_recov_scan_state_key(kind: str, scope: str) -> str:
    """State key for the persisted scan-results card (per scope).

    ``scope`` is ``"monthly"`` (no-WARM Historical step) or ``"sample"``
    (step_co_recov).  Each scope keeps its own folder / counts / layouts so
    one page's scan doesn't overwrite the other's display.
    """
    return f"{kind}_scan_data_{scope}"


def _build_co_recov_layouts(
    state: dict[str, Any], kind: str, files: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Group scanned files by signature.

    ``files`` items are ``{name, path, signature, headers}``.
    """
    ignored = _co_recov_ignored_set(state, kind)
    by_sig: dict[str, dict[str, Any]] = {}
    for entry in files:
        sig = entry.get("signature") or ""
        if not sig:
            continue
        slot = by_sig.setdefault(sig, {
            "signature": sig,
            "sample_file": entry.get("name", ""),
            "sample_path": entry.get("path", ""),
            "sample_headers": entry.get("headers") or [],
            "file_count": 0,
            "ignored": sig in ignored,
        })
        slot["file_count"] += 1
    return sorted(
        by_sig.values(),
        key=lambda s: (s.get("ignored", False), -s.get("file_count", 0)),
    )


def _refresh_layouts_in_scan_data(
    state: dict[str, Any], kind: str, scope: str,
) -> None:
    """Refresh the ``ignored`` flag on stored layouts (after ignore /
    unignore) without re-reading any files.
    """
    key = _co_recov_scan_state_key(kind, scope)
    data = state.get(key) or {}
    layouts = list(data.get("layouts") or [])
    if not layouts:
        return
    ignored = _co_recov_ignored_set(state, kind)
    for layout in layouts:
        layout["ignored"] = layout.get("signature") in ignored
    data["layouts"] = sorted(
        layouts,
        key=lambda s: (s.get("ignored", False), -s.get("file_count", 0)),
    )
    state[key] = data


def _persist_co_recov_scan_data(
    state: dict[str, Any],
    kind: str,
    scope: str,
    out: dict[str, Any],
    scanned_meta: list[dict[str, Any]],
) -> None:
    """Persist a scan-results card payload to ``state`` for the template.

    ``out`` is the dict returned by ``_scan_*_co_recov_folder``;
    ``scanned_meta`` is the per-file list used to compute layouts.
    """
    layouts = _build_co_recov_layouts(state, kind, scanned_meta)
    state[_co_recov_scan_state_key(kind, scope)] = {
        "folder": out.get("folder", ""),
        "scanned": out.get("scanned", 0),
        "added_count": len(out.get("added") or []),
        "ignored_count": len(out.get("ignored") or []),
        "skipped_count": len(out.get("skipped") or []),
        "error": out.get("error"),
        "layouts": layouts,
    }


def _remove_files_with_signature(
    state: dict[str, Any], kind: str, signature: str, *, is_sample: bool,
) -> int:
    """Drop any already-ingested files whose header signature matches
    ``signature``.  Returns the count removed.
    """
    if not signature:
        return 0
    if is_sample:
        bucket_key = "co_files" if kind == "co" else "recov_files"
        uploads = state.get("sample_uploads") or {}
        rows = list(uploads.get(bucket_key) or [])
    else:
        bucket_key = "monthly_co_files" if kind == "co" else "monthly_recov_files"
        hs = state.get("hist_scan") or {}
        rows = list(hs.get(bucket_key) or [])
    if not rows:
        return 0
    removed = 0
    kept: list[dict[str, Any]] = []
    for entry in rows:
        p = entry.get("path")
        if not p:
            kept.append(entry)
            continue
        sig, _ = _file_layout_signature(Path(p))
        if sig == signature:
            removed += 1
            try:
                Path(p).unlink(missing_ok=True)
            except Exception:  # noqa: BLE001
                pass
            continue
        kept.append(entry)
    if removed:
        if is_sample:
            uploads[bucket_key] = kept
            state["sample_uploads"] = uploads
        else:
            hs[bucket_key] = kept
            state["hist_scan"] = hs
    return removed


def _scan_monthly_co_recov_folder(
    state: dict[str, Any], folder_str: str, kind: str
) -> dict[str, Any]:
    """Glob ``folder_str`` for monthly CO/Recov files and ingest each.

    ``kind`` is ``"co"`` or ``"recov"``.  Files whose header layout the
    user has marked ignored (``state[f"{kind}_ignored_signatures"]``) are
    skipped.  Returns a small summary dict for flashing in the UI; the
    grouped layouts are persisted to ``state[f"{kind}_scan_layouts"]``
    so the template can render Ignore / Unignore controls.
    """
    list_key = "monthly_co_files" if kind == "co" else "monthly_recov_files"
    subfolder = "co" if kind == "co" else "recov"
    out: dict[str, Any] = {
        "ok": False, "error": None,
        "added": [], "skipped": [], "ignored": [], "scanned": 0,
        "folder": folder_str,
    }
    folder_str = _normalize_folder_path(folder_str)
    if not folder_str:
        out["error"] = "Enter a folder path before scanning."
        return out
    folder = Path(folder_str)
    if not folder.exists():
        out["error"] = f"Folder not found: {folder}"
        return out
    if not folder.is_dir():
        out["error"] = f"Path is not a folder: {folder}"
        return out

    files = _discover_co_recov_files(folder)
    out["scanned"] = len(files)
    if not files:
        try:
            entries = sorted(p.name for p in folder.iterdir())[:8]
        except Exception:  # noqa: BLE001
            entries = []
        sample = (", ".join(entries)) if entries else "(empty)"
        out["error"] = (
            f"No .xlsx/.xlsm/.xls/.csv files found in {folder} (recursive). "
            f"First items in folder: {sample}"
        )
        return out

    ignored_sigs = _co_recov_ignored_set(state, kind)
    scanned_meta: list[dict[str, Any]] = []
    for p in files:
        sig, headers = _file_layout_signature(p)
        scanned_meta.append({
            "name": p.name, "path": str(p),
            "signature": sig, "headers": headers,
        })
        if sig and sig in ignored_sigs:
            out["ignored"].append(p.name)
            continue
        try:
            _add_hist_file_from_path(state, list_key, p, subfolder)
            out["added"].append(p.name)
        except Exception as exc:  # noqa: BLE001
            out["skipped"].append(f"{p.name}: {exc}")
    _persist_co_recov_scan_data(state, kind, "monthly", out, scanned_meta)
    out["ok"] = bool(out["added"]) or bool(out["ignored"])
    return out


def _read_co_recov_column_form(kind: str) -> dict[str, Any]:
    """Pull the Column Mapping form fields off ``request.form`` for the
    monthly CO/Recov card.  ``kind`` is ``"co"`` or ``"recov"``.

    Returns the shape persisted in ``state[f"{kind}_columns"]``:

        {
          "loan_code": header_name,
          "amount":    header_name,
          "date":      header_name,
          "member_number": header_name,
          "loan_suffix":   header_name,
          "member_account": {
            "mode": "fixed_suffix" | "delimiter" | "split",
            "suffix_length": int,
            "delimiter": str,
          },
        }
    """
    prefix = "co" if kind == "co" else "recov"

    def _val(field: str) -> str:
        return (request.form.get(f"{prefix}_{field}") or "").strip()

    mode = _val("member_account_mode") or "fixed_suffix"
    if mode not in ("fixed_suffix", "delimiter", "split"):
        mode = "fixed_suffix"
    try:
        suffix_len = int(_val("member_account_suffix_length") or "3")
    except ValueError:
        suffix_len = 3
    suffix_len = max(0, min(9, suffix_len))
    delim = _val("member_account_delimiter") or "-"

    result: dict[str, Any] = {
        "loan_code":     _val("loan_code_col"),
        "amount":        _val("amount_col"),
        "date":          _val("date_col"),
        "member_number": _val("member_number_col"),
        "loan_suffix":   _val("loan_suffix_col"),
        "member_account": {
            "mode": mode,
            "suffix_length": suffix_len,
            "delimiter": delim,
        },
    }
    # Worksheet/tab override for single-workbook CO+Recov files (both on
    # separate tabs of one workbook). Only include the key when the form
    # actually carried the field so a mapping form that omits it (e.g. the
    # monthly-files card) doesn't wipe a previously-chosen tab.
    _sheet_raw = request.form.get(f"{prefix}_sheet")
    if _sheet_raw is not None:
        result["sheet"] = _sheet_raw.strip()
    return result


def _refresh_co_recov_inspect(
    state: dict[str, Any], kind: str, *, force: bool = False
) -> None:
    """Re-run the column inspector against the first uploaded CO/Recov
    file and seed default column mappings.  Called after each upload so
    the column-mapping UI always reflects the most recently-uploaded
    file.  Looks at the per-month list first (``monthly_co_files`` /
    ``monthly_recov_files``) and falls back to the single-workbook list
    (``co_files`` / ``recov_files``), so the same inspect/mapping UI
    works for both upload modes.

    User-edited column mappings are preserved unless ``force=True``.
    """
    monthly_key = "monthly_co_files" if kind == "co" else "monthly_recov_files"
    workbook_key = "co_files" if kind == "co" else "recov_files"
    inspect_key = "co_inspect" if kind == "co" else "recov_inspect"
    columns_key = "co_columns" if kind == "co" else "recov_columns"
    suggested_field = "co_amount" if kind == "co" else "recov_amount"

    hist_scan = state.get("hist_scan") or {}
    files = hist_scan.get(monthly_key) or hist_scan.get(workbook_key) or []
    if not files:
        state.pop(inspect_key, None)
        _save_state(state)
        return
    first = files[0]
    path_str = first.get("path") or ""
    if not path_str or not Path(path_str).exists():
        state.pop(inspect_key, None)
        _save_state(state)
        return
    saved_sheet = (state.get(columns_key) or {}).get("sheet") or None
    res = monthly_co_recov_aggregator.inspect_file(
        Path(path_str), sheet=saved_sheet
    )
    state[inspect_key] = res

    # Seed mapping on first upload (or force) using the suggested headers.
    existing = state.get(columns_key) or {}
    have_user_edit = bool(
        existing.get("loan_code") or existing.get("amount")
    )
    suggested = (res.get("suggested") or {})
    if (not have_user_edit) or force:
        # Preserve any previously-saved member/account block + the user's
        # mode choice; just refresh the auto-detectable header pickers.
        ma = existing.get("member_account") or {
            "mode": "fixed_suffix", "suffix_length": 3, "delimiter": "-",
        }
        state[columns_key] = {
            "loan_code":     suggested.get("code") or "",
            "amount":        suggested.get(suggested_field) or "",
            "date":          suggested.get("date") or existing.get("date") or "",
            "member_number": (existing.get("member_number")
                              or suggested.get("member") or ""),
            "loan_suffix":   (existing.get("loan_suffix")
                              or suggested.get("account") or ""),
            "member_account": ma,
            # Preserve the user's tab choice across a re-seed.
            "sheet":         existing.get("sheet", ""),
        }
    else:
        # Backfill any individual fields that are still blank from the
        # fresh suggestions. This catches the case where a prior
        # inspect's whitespace-laden header values never round-tripped
        # through the form, leaving amount/date blank while loan_code
        # was filled.  Updating in place preserves the user's edits.
        cols = dict(existing)
        if not cols.get("loan_code"):
            cols["loan_code"] = suggested.get("code") or ""
        if not cols.get("amount"):
            cols["amount"] = suggested.get(suggested_field) or ""
        if not cols.get("date"):
            cols["date"] = suggested.get("date") or ""
        if not cols.get("member_number"):
            cols["member_number"] = suggested.get("member") or ""
        if not cols.get("loan_suffix"):
            cols["loan_suffix"] = suggested.get("account") or ""
        cols.setdefault("member_account", {
            "mode": "fixed_suffix", "suffix_length": 3, "delimiter": "-",
        })
        state[columns_key] = cols
    _save_state(state)


# Common loan pool names offered as datalist suggestions when no WARM map is
# available. Kept short on purpose — the user can type anything.
_DEFAULT_POOL_SUGGESTIONS = [
    "New Vehicle", "Used Vehicle", "New Recreational", "Used Recreational",
    "Other Secured", "Unsecured", "eBuilder", "Share Secured", "VISA",
    "Credit Card", "Mortgage Loans", "HELOC", "Signature Loans",
]


def _expand_numeric_tokens(text: str) -> set[str]:
    """Return all integer codes implied by ``text``.

    Handles single numbers, comma-separated lists, and dash ranges so that a
    label like ``"NEW VEHICLE 22-26"`` yields ``{"22","23","24","25","26"}``
    and ``"75-90, 92-94,99"`` yields ``{"75",...,"90","92","93","94","99"}``.
    Only treats 1- to 4-digit integers as loan codes to avoid swallowing
    things like phone numbers or years.
    """
    out: set[str] = set()
    # Find each digit run, then look one char left/right for dash to expand.
    # Easier: scan for "NN-MM" range patterns first, then individual numbers.
    range_rx = re.compile(r"(?<!\d)(\d{1,4})\s*-\s*(\d{1,4})(?!\d)")
    for m in range_rx.finditer(text):
        a, b = int(m.group(1)), int(m.group(2))
        if a > b or (b - a) > 200:
            continue
        for n in range(a, b + 1):
            out.add(str(n))
    # Strip the ranges so we don't double-count.
    stripped = range_rx.sub(" ", text)
    for m in re.finditer(r"(?<!\d)(\d{1,4})(?!\d)", stripped):
        out.add(m.group(1))
    return out


def _alpha_tokens(text: str) -> list[str]:
    """Return lowercase alphabetic tokens (length >= 3, ignoring filler words)."""
    raw = re.findall(r"[A-Za-z]{2,}", text.lower())
    filler = {"the", "and", "for", "loan", "loans", "type", "types", "pool",
              "pools", "hide", "total"}
    return [t for t in raw if t not in filler]


def _auto_map_one_label(
    label: str,
    *,
    bs_map_ci: dict[str, str],
    pool_map: dict[str, str],
    pool_map_keys_ci: dict[str, str],
    pool_names: list[str],
) -> tuple[str, str]:
    """Try several strategies to map ``label`` to a pool.

    Returns ``(pool_name, strategy)``. ``pool_name`` is "" if nothing matched.
    Strategies (in priority order):

      1. ``warm``       — exact case-insensitive match against the WARM
                          BS Data label->pool map.
      2. ``code-token`` — any numeric or alphanumeric token in ``label``
                          matches a raw code in ``state.pool_map`` whose
                          value (pool name) is non-empty.
      3. ``code-substr``— any raw code in ``state.pool_map`` (length >= 2)
                          appears as a whole-word substring of ``label``.
      4. ``name-substr``— any known pool name appears (case-insensitive) as
                          a substring of ``label``.
      5. ``name-tokens``— all alphabetic tokens of a known pool name appear
                          in ``label`` (e.g. "New Vehicle" matches
                          "NEW VEHICLE 22-26").
    """
    label_low = label.strip().lower()

    # 1) WARM exact match
    hit = bs_map_ci.get(label_low)
    if hit:
        return hit, "warm"

    # 2) code-token match (numbers + alphanumeric tokens)
    tokens: set[str] = set()
    tokens.update(_expand_numeric_tokens(label))
    for t in re.findall(r"[A-Za-z0-9_]+", label):
        tokens.add(t)
        tokens.add(t.lower())
        tokens.add(t.upper())
    for tok in tokens:
        pool = pool_map_keys_ci.get(tok.lower())
        if pool:
            return pool, "code-token"

    # 3) code-substr — pool_map raw code appears inside label
    label_alnum = re.sub(r"[^A-Za-z0-9]", " ", label).lower()
    label_alnum_spaced = f" {label_alnum} "
    for code, pool in pool_map.items():
        if not pool or not code or len(code) < 2:
            continue
        code_low = code.lower()
        if f" {code_low} " in label_alnum_spaced:
            return pool, "code-substr"

    # 4) name-substr — pool name appears in label
    for name in pool_names:
        if not name:
            continue
        name_low = name.lower()
        if len(name_low) >= 3 and name_low in label_low:
            return name, "name-substr"

    # 5) name-tokens — all alphabetic tokens of a pool name appear in label
    label_tokens = set(_alpha_tokens(label))
    if label_tokens:
        for name in pool_names:
            name_tokens = _alpha_tokens(name)
            if not name_tokens:
                continue
            if all(t in label_tokens for t in name_tokens):
                return name, "name-tokens"

    return "", ""


def _seed_hist_pool_map(state: dict[str, Any], analysis: dict[str, Any]) -> None:
    """Pre-fill the historical loan-type -> pool mapping from a fresh
    balance-file analysis.

    Combines several sources for auto-mapping (most authoritative first):

      * ``state.warm.bs_loan_type_map`` — column-A label -> pool from the
        uploaded WARM workbook's ``BS Data`` tab.
      * ``state.pool_map`` — raw loan code -> pool name from the loan-data
        sample step (Step 7). Lets historical labels share loan codes with
        the loan extract.
      * ``state.sample.pool_code_suggestions`` — distinct raw codes detected
        in the loan extract (used to widen the pool_map key set).
      * ``state.warm.pools`` and configured pool-map values — used for
        fuzzy name-substring matching.

    Preserves any prior user edits.
    """
    labels: list[str] = analysis.get("labels") or []
    existing = state.get("hist_pool_map") or {}
    prior_map: dict[str, str] = dict(existing.get("mapping") or {})

    warm = state.get("warm") or {}
    bs_map_raw: dict[str, str] = warm.get("bs_loan_type_map") or {}
    bs_map_ci = {k.strip().lower(): v for k, v in bs_map_raw.items() if k and v}

    pool_map: dict[str, str] = {
        (k or "").strip(): (v or "").strip()
        for k, v in (state.get("pool_map") or {}).items()
        if (k or "").strip()
    }
    pool_map_keys_ci = {k.lower(): v for k, v in pool_map.items() if v}

    # Build the universe of known pool names for fuzzy matching.
    pool_names: list[str] = []
    seen_names: set[str] = set()
    def _add_name(n: str) -> None:
        n = (n or "").strip()
        low = n.lower()
        if n and low not in seen_names:
            pool_names.append(n)
            seen_names.add(low)
    for v in bs_map_raw.values():
        _add_name(v)
    for v in pool_map.values():
        _add_name(v)
    for v in warm.get("pools") or []:
        _add_name(v)
    # Prefer longer names first so "New Vehicle" matches before "Vehicle".
    pool_names.sort(key=lambda s: (-len(s), s.lower()))

    strategies_used: set[str] = set()
    new_mapping: dict[str, str] = {}
    for label in labels:
        if prior_map.get(label):
            new_mapping[label] = prior_map[label]
            continue
        pool, strat = _auto_map_one_label(
            label,
            bs_map_ci=bs_map_ci,
            pool_map=pool_map,
            pool_map_keys_ci=pool_map_keys_ci,
            pool_names=pool_names,
        )
        new_mapping[label] = pool
        if strat:
            strategies_used.add(strat)

    if "warm" in strategies_used:
        source = "warm"
    elif strategies_used:
        source = "auto"
    elif existing:
        source = "manual"
    else:
        source = ""

    state["hist_pool_map"] = {
        "sheet": analysis.get("sheet", ""),
        "labels": labels,
        "mapping": new_mapping,
        "source": source,
        "strategies": sorted(strategies_used),
    }
    _save_state(state)


def _save_hist_pool_map_from_form(state: dict[str, Any], form) -> None:
    """Persist user edits to the loan-type -> pool mapping table.

    The "__ignore__" sentinel is preserved in the saved mapping so the
    template can distinguish "user explicitly chose to ignore this row"
    from "auto-seed left it blank and the user hasn't reviewed it yet."
    Downstream consumers that fold mapping values into other pool maps
    treat the sentinel as empty (see ``_hpm_value_for_seed`` below).
    """
    existing = state.get("hist_pool_map") or {}
    labels = list(existing.get("labels") or [])
    mapping: dict[str, str] = {}
    # Form fields are named pool_for_<index> with a parallel label_<index>.
    for idx, label in enumerate(labels):
        # Prefer the hidden label echo (in case the labels list changed mid-edit).
        echoed = (form.get(f"label_{idx}") or "").strip()
        key = echoed or label
        pool = (form.get(f"pool_for_{idx}") or "").strip()
        # Keep "__ignore__" as the stored sentinel so the completion check
        # can tell an explicit "ignore" apart from an unreviewed blank.
        mapping[key] = pool
    existing["labels"] = labels
    existing["mapping"] = mapping
    existing["source"] = "manual"
    state["hist_pool_map"] = existing
    # Propagate explicit user edits to the Monthly Balance File pool_map
    # (Step 8), which Step 14 (Balance Adjustment) and the report runtime
    # both read from. We only touch labels that ALSO appear in the
    # monthly_bal parsed labels and only when the new value is non-empty
    # and different from what is already there. The "__ignore__" sentinel
    # is normalised away via _hpm_value_for_seed so it doesn't blank out
    # an already-good Step 8 mapping.
    try:
        mb = state.get("monthly_bal") or {}
        mb_labels = {str(l) for l in (mb.get("parsed_pool_labels") or []) if l}
        if mb_labels:
            mb_map = dict(mb.get("pool_map") or {})
            changed = False
            for lbl, val in mapping.items():
                if not lbl or lbl not in mb_labels:
                    continue
                new_val = _hpm_value_for_seed(val)
                if not new_val:
                    continue
                if mb_map.get(lbl) != new_val:
                    mb_map[lbl] = new_val
                    changed = True
            if changed:
                mb["pool_map"] = mb_map
                state["monthly_bal"] = mb
    except Exception:
        # Propagation is best-effort; never block the Step 4 save.
        pass
    _save_state(state)


def _hpm_value_for_seed(value: Any) -> str:
    """Normalize a hist_pool_map.mapping value for downstream seeding.

    Returns an empty string for the "__ignore__" sentinel so it doesn't
    leak into other pool maps (Step 5 monthly balance pool_map, etc.).
    """
    v = (value or "")
    if isinstance(v, str) and v.strip() == "__ignore__":
        return ""
    return v


# ---------------------------------------------------------------------------
# Monthly loan-data extract — anchor uploads
# ---------------------------------------------------------------------------

def _ensure_hist_extracts(state: dict[str, Any]) -> dict[str, Any]:
    """Return state['hist_extracts'], initialising it if missing."""
    he = state.get("hist_extracts")
    if not isinstance(he, dict):
        he = {
            "target_period": "",
            "history_months": 84,
            "folder_path": "",
            "anchor_files": [],
            "profiles": [],
            "scan_results": None,
        }
        state["hist_extracts"] = he
    he.setdefault("target_period", "")
    he.setdefault("history_months", 84)
    he.setdefault("folder_path", "")
    he.setdefault("anchor_files", [])
    he.setdefault("profiles", [])
    he.setdefault("scan_results", None)
    he.setdefault("ignored_signatures", [])
    sb = he.get("solr_backfill")
    if not isinstance(sb, dict):
        sb = {}
        he["solr_backfill"] = sb
    sb.setdefault("solr_url", "http://searchserver1.tctrisk.com:8983/solr")
    sb.setdefault("core", "ncua")
    sb.setdefault("loan_code_fields", {})
    # "distributed" (default) fetches the quarter's TOTAL loan balance
    # and apportions across user pools using the earliest historical-
    # balance month's pool ratios -- the right answer for almost every
    # CU we onboard. "per_loan_code" maps each canonical NCUA bucket
    # through pool_map / NCUA inference and is only used in special
    # cases where the 5300 categories line up cleanly with this CU's
    # pools.
    sb.setdefault("mode", "distributed")
    sb.setdefault("last_test", None)
    sb.setdefault("last_run", None)
    # Parallel slot for the charge-off backfill (Solr URL/core are
    # shared with the loan-balance backfill above; this slot only
    # records the most recent run result for the CO panel).
    co_sb = he.get("co_solr_backfill")
    if not isinstance(co_sb, dict):
        co_sb = {}
        he["co_solr_backfill"] = co_sb
    co_sb.setdefault("last_run", None)
    # Parallel slot for the recovery backfill.
    rec_sb = he.get("recov_solr_backfill")
    if not isinstance(rec_sb, dict):
        rec_sb = {}
        he["recov_solr_backfill"] = rec_sb
    rec_sb.setdefault("last_run", None)
    return he


def _earliest_month_pool_distribution(state: dict) -> dict:
    """Compute the {pool_name: ratio} distribution from the EARLIEST
    month of the user's uploaded historical balance file(s).

    Picks the loader matching ``state.hist_balance_source``:
      * ``annual_balance_sheets`` -> per_year_files aggregator
      * ``monthly_balance_sheets`` -> per_month_files aggregator
      * ``single_workbook`` -> reads the wide-format workbook directly,
        picking the earliest date column on the header row.
      * ``monthly_loan_extracts`` -> not supported (returns ``error``).

    Returns a dict::

        {
            "ok": bool,
            "error": str,
            "period": "YYYY-MM-DD",   # earliest period found
            "total": float,           # sum across all pools that month
            "pool_distribution": {pool: ratio},  # ratios sum to 1.0
            "by_pool": {pool: balance},          # raw $ per pool
            "source": "annual" | "per_month",
        }
    """
    out: dict = {
        "ok": False,
        "error": "",
        "period": "",
        "total": 0.0,
        "pool_distribution": {},
        "by_pool": {},
        "source": "",
    }
    mb_state = state.get("monthly_bal") or {}
    label_to_pool = mb_state.get("pool_map") or {}
    source = state.get("hist_balance_source") or ""
    by_period: dict = {}
    if source == "annual_balance_sheets":
        out["source"] = "annual"
        ann = monthly_bal_parser.pool_balances_for_per_year_files(
            mb_state.get("year_files") or [],
            mb_state.get("per_year_layout") or {},
            label_to_pool,
            exclude_labels=mb_state.get("exclude_labels") or [],
        )
        if not ann.get("ok") and not (ann.get("by_period") or {}):
            out["error"] = (
                f"Could not read annual balance workbook(s): "
                f"{ann.get('error') or 'unknown error'}"
            )
            return out
        by_period = ann.get("by_period") or {}
    elif source == "monthly_balance_sheets":
        out["source"] = "per_month"
        pm = monthly_bal_parser.pool_balances_for_per_month_files(
            mb_state.get("monthly_files") or [],
            mb_state.get("per_month_layout") or {},
            label_to_pool,
            exclude_labels=mb_state.get("exclude_labels") or [],
        )
        if not pm.get("ok") and not (pm.get("by_period") or {}):
            out["error"] = (
                f"Could not read monthly balance file(s): "
                f"{pm.get('error') or 'unknown error'}"
            )
            return out
        by_period = pm.get("by_period") or {}
    elif source == "single_workbook":
        # Single wide-format workbook with one column per month-end.
        # Find the earliest date column on the header row, then ask the
        # per-period extractor for that period.
        out["source"] = "single"
        saved_path = mb_state.get("saved_path") or ""
        sheet = mb_state.get("sheet") or ""
        header_row = mb_state.get("header_row") or 0
        pool_name_col = mb_state.get("pool_name_col") or ""
        if not (saved_path and sheet and header_row and pool_name_col):
            out["error"] = (
                "Single-workbook layout is incomplete (sheet / header row / "
                "pool name column). Re-save Step 3 layout."
            )
            return out
        try:
            from openpyxl import load_workbook  # local import
            wb = load_workbook(saved_path, read_only=True, data_only=True)
            try:
                if sheet not in wb.sheetnames:
                    out["error"] = f"Sheet '{sheet}' not found in workbook."
                    return out
                ws = wb[sheet]
                date_isos: list[str] = []
                for cell in ws[int(header_row)]:
                    d = monthly_bal_parser.normalize_to_month_end(cell.value)
                    if d is not None:
                        date_isos.append(d.isoformat())
            finally:
                wb.close()
        except Exception as exc:  # noqa: BLE001
            out["error"] = f"Could not open balance workbook: {exc}"
            return out
        if not date_isos:
            out["error"] = (
                f"No date columns detected on row {header_row} of the "
                f"single workbook."
            )
            return out
        earliest_iso = min(date_isos)
        sw = monthly_bal_parser.pool_balances_for_latest_period(
            saved_path, sheet, int(header_row), pool_name_col,
            label_to_pool=label_to_pool, period=earliest_iso,
            exclude_labels=mb_state.get("exclude_labels") or [],
        )
        if not sw.get("ok"):
            out["error"] = (
                f"Could not read single workbook: "
                f"{sw.get('error') or 'unknown error'}"
            )
            return out
        by_period = {sw.get("period") or earliest_iso: {
            "by_pool": sw.get("by_pool") or {},
            "raw_rows": sw.get("raw_rows") or [],
        }}
    elif source == "monthly_loan_extracts":
        # No balance workbook to read — derive the earliest-month pool
        # distribution from rows already extracted into
        # ``loan_code_history`` (per-extract loan-code totals). The user
        # must have run the per-extract Source-D ingestion first; if no
        # rows exist we surface a clear error.
        out["source"] = "monthly_loan_extracts"
        cu = (state.get("credit_union") or "").strip()
        if not cu:
            out["error"] = (
                "Credit union name is required to look up loan-code "
                "history. Re-save Step 1."
            )
            return out
        try:
            from sqlalchemy import create_engine, text as _sql_text  # local
            from cecl_credentials import get_database_url as _gdb
            eng = create_engine(_gdb())
            with eng.begin() as cx:
                row = cx.execute(_sql_text(
                    "SELECT MIN(as_of_date) FROM loan_code_history "
                    "WHERE cu = :cu"
                ), {"cu": cu}).fetchone()
                earliest_d = row[0] if row else None
                if earliest_d is None:
                    out["error"] = (
                        "No rows in loan_code_history for this credit union. "
                        "Run the historical extracts ingest on Step 3 first, "
                        "then come back."
                    )
                    return out
                rows = cx.execute(_sql_text(
                    "SELECT loan_code, total_balance "
                    "FROM loan_code_history "
                    "WHERE cu = :cu AND as_of_date = :d"
                ), {"cu": cu, "d": earliest_d}).fetchall()
        except Exception as exc:  # noqa: BLE001
            out["error"] = (
                f"Could not query loan_code_history: {exc}"
            )
            return out
        # Map each raw loan_code to a configured pool via pool_map.
        pmap = state.get("pool_map") or {}
        pmap_ci = {str(k).strip().lower(): v for k, v in pmap.items()}
        excluded = {str(p).strip().lower()
                    for p in (state.get("excluded_pools") or [])}
        excluded.update({"ignore", "exclude"})
        by_pool: dict[str, float] = {}
        for code, bal in rows:
            try:
                bal_f = float(bal or 0)
            except (TypeError, ValueError):
                continue
            if bal_f <= 0:
                continue
            code_s = str(code or "").strip()
            if not code_s:
                continue
            pool = (
                pmap.get(code_s)
                or pmap_ci.get(code_s.lower())
                or code_s  # fall back to using the code as-is
            )
            if not pool or str(pool).strip().lower() in excluded:
                continue
            by_pool[pool] = by_pool.get(pool, 0.0) + bal_f
        by_period = {earliest_d.isoformat(): {"by_pool": by_pool}}
    else:
        out["error"] = (
            "Distributed mode requires the historical balance source to be "
            "Single workbook, Annual balance sheets, Monthly balance sheets, "
            "or Monthly loan extracts (Step 3)."
        )
        return out

    if not by_period:
        out["error"] = "No periods found in the uploaded balance file(s)."
        return out
    earliest = min(by_period.keys())
    by_pool = (by_period.get(earliest) or {}).get("by_pool") or {}
    # Drop zero/negative entries before normalizing.
    pos = {p: float(b) for p, b in by_pool.items() if float(b or 0) > 0}
    total = sum(pos.values())
    if total <= 0:
        # Source-aware guidance: workbook-based sources (single/annual/per-month)
        # need the user to map parsed labels to canonical pool names on
        # Step 8 (Monthly Balance File). The monthly_loan_extracts source
        # uses Step 3 (Loan Code Mapping).
        if source == "monthly_loan_extracts":
            _hint = "Check the Loan-Code Mapping on Step 3."
        else:
            _mb_state = state.get("monthly_bal") or {}
            _label_n = len(_mb_state.get("parsed_pool_labels") or [])
            _mapped_n = sum(
                1 for v in (_mb_state.get("pool_map") or {}).values()
                if str(v or "").strip()
            )
            if _label_n and _mapped_n == 0:
                _hint = (
                    f"{_label_n} balance label(s) detected but none are "
                    "mapped to a pool yet. Open Step 8 (Monthly Balance "
                    "File) and map each label to a configured pool."
                )
            elif _label_n and _mapped_n < _label_n:
                _hint = (
                    f"{_mapped_n} of {_label_n} balance label(s) mapped on "
                    "Step 8 (Monthly Balance File). Map the remaining "
                    f"{_label_n - _mapped_n} label(s) so their balances "
                    "flow into the right pools."
                )
            else:
                _hint = (
                    "Check the balance label-to-pool mapping on Step 8 "
                    "(Monthly Balance File)."
                )
        out["error"] = (
            f"Earliest month ({earliest}) had no positive pool balances. "
            + _hint
        )
        out["period"] = earliest
        return out
    pd_ratios = {p: (b / total) for p, b in pos.items()}
    out.update({
        "ok": True,
        "period": earliest,
        "total": round(total, 2),
        "pool_distribution": pd_ratios,
        "by_pool": {p: round(b, 2) for p, b in pos.items()},
    })
    return out


def _extracted_balance_preview(state: dict, max_recent_months: int = 6) -> dict:
    """Build a confirmation/sanity-check preview of the historical
    balances actually extracted from the uploaded balance file(s) +
    any NCUA 5300 distributed-backfill rows written to the DB.

    Returns a dict::

        {
            "ok": bool,
            "error": str,
            "source": "annual" | "per_month" | "single" | "",
            "source_files": [str, ...],   # file names that produced data
            "period_count": int,          # distinct month-ends
            "earliest_period": "YYYY-MM-DD",
            "latest_period": "YYYY-MM-DD",
            "pool_count": int,
            "recent_periods": ["YYYY-MM-DD", ...],   # most-recent N
            "top_pools": [str, ...],                 # top by latest balance
            "totals_by_period": {period_iso: total_$},
            "balances_grid": {pool: {period_iso: $, ...}},  # for top_pools
            "backfill": {                # 5300 distributed last_run summary
                "present": bool,
                "ok": bool,
                "months_filled": int,
                "rows_written": int,
                "error": str,
                "earliest_filled": "YYYY-MM-DD",
                "latest_filled": "YYYY-MM-DD",
            },
        }

    Cheap (<200ms typically); meant to be computed at template render
    time and never raise.
    """
    out: dict = {
        "ok": False, "error": "", "source": "",
        "source_files": [], "period_count": 0,
        "earliest_period": "", "latest_period": "",
        "latest_period_with_data": "",
        "pool_count": 0, "recent_periods": [], "top_pools": [],
        "totals_by_period": {}, "balances_grid": {},
        "backfill": {
            "present": False, "ok": False,
            "months_filled": 0, "rows_written": 0, "error": "",
            "earliest_filled": "", "latest_filled": "",
        },
    }

    # ----- Backfill summary (independent of source) --------------------
    try:
        sb_run = (
            ((state.get("hist_extracts") or {}).get("solr_backfill") or {})
            .get("last_run") or {}
        )
    except Exception:  # noqa: BLE001
        sb_run = {}
    if sb_run:
        out["backfill"]["present"] = True
        out["backfill"]["ok"] = bool(sb_run.get("ok"))
        months = sb_run.get("months_filled") or sb_run.get("filled_months") or []
        if isinstance(months, list):
            out["backfill"]["months_filled"] = len(months)
            # Each entry may be a plain ISO string OR a dict like
            # {"period": "2018-10-31", "rows_written": ...}. Normalise to
            # a list of strings before min/max so we never compare dicts.
            period_strs: list[str] = []
            for m in months:
                if isinstance(m, dict):
                    p = m.get("period") or m.get("as_of") or m.get("date")
                    if p:
                        period_strs.append(str(p))
                elif m:
                    period_strs.append(str(m))
            if period_strs:
                out["backfill"]["earliest_filled"] = min(period_strs)
                out["backfill"]["latest_filled"] = max(period_strs)
        try:
            out["backfill"]["rows_written"] = int(
                sb_run.get("rows_written") or 0
            )
        except (TypeError, ValueError):
            pass
        if not sb_run.get("ok") and sb_run.get("error"):
            out["backfill"]["error"] = str(sb_run.get("error"))

    # ----- Extracted balance grid -------------------------------------
    mb_state = state.get("monthly_bal") or {}
    label_to_pool = mb_state.get("pool_map") or {}
    source = state.get("hist_balance_source") or ""
    by_period: dict = {}
    source_files: list[str] = []
    try:
        if source == "annual_balance_sheets":
            out["source"] = "annual"
            year_files = mb_state.get("year_files") or []
            if not year_files:
                # User picked the source but hasn't uploaded yet -- stay quiet.
                return out
            res = monthly_bal_parser.pool_balances_for_per_year_files(
                year_files,
                mb_state.get("per_year_layout") or {},
                label_to_pool,
                exclude_labels=mb_state.get("exclude_labels") or [],
            )
            by_period = res.get("by_period") or {}
            for yf in year_files:
                if isinstance(yf, dict):
                    nm = yf.get("name") or yf.get("path") or ""
                    if nm:
                        source_files.append(str(nm).split("/")[-1].split("\\")[-1])
        elif source == "monthly_balance_sheets":
            out["source"] = "per_month"
            monthly_files = mb_state.get("monthly_files") or []
            if not monthly_files:
                return out
            res = monthly_bal_parser.pool_balances_for_per_month_files(
                monthly_files,
                mb_state.get("per_month_layout") or {},
                label_to_pool,
                exclude_labels=mb_state.get("exclude_labels") or [],
            )
            by_period = res.get("by_period") or {}
            for mf in monthly_files:
                if isinstance(mf, dict):
                    nm = mf.get("name") or mf.get("path") or ""
                    if nm:
                        source_files.append(str(nm).split("/")[-1].split("\\")[-1])
        elif source == "single_workbook":
            # Single workbook needs a per-month walk; defer to the
            # earliest-month helper which already opens it once. The
            # full grid would require a wide-table reshape we don't
            # ship today -- show the earliest month only.
            saved_path = mb_state.get("saved_path") or ""
            if not saved_path:
                # No file uploaded yet -- stay quiet.
                return out
            earliest = _earliest_month_pool_distribution(state)
            if earliest.get("ok"):
                out["source"] = "single"
                by_period = {earliest["period"]: {
                    "by_pool": earliest.get("by_pool") or {}
                }}
                source_files.append(
                    str(saved_path).split("/")[-1].split("\\")[-1]
                )
            else:
                out["error"] = earliest.get("error") or "preview unavailable"
                return out
        else:
            # No source picked yet -- stay quiet (no card).
            return out
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Could not build balance preview: {exc}"
        return out

    if not by_period:
        out["error"] = (
            "No periods extracted from the uploaded balance file(s). "
            "Check the pool mapping below."
        )
        return out

    out["source_files"] = source_files
    sorted_periods = sorted(by_period.keys())
    out["period_count"] = len(sorted_periods)
    out["earliest_period"] = sorted_periods[0]
    out["latest_period"] = sorted_periods[-1]

    # Compute totals per period + collect every pool that has any balance
    totals: dict[str, float] = {}
    for per in sorted_periods:
        by_pool = (by_period.get(per) or {}).get("by_pool") or {}
        tot = 0.0
        for pool, bal in by_pool.items():
            try:
                b = float(bal or 0)
            except (TypeError, ValueError):
                continue
            if b <= 0:
                continue
            tot += b
        totals[per] = round(tot, 2)

    # Drop trailing periods with zero total (workbooks often pre-allocate
    # future month columns that haven't been filled in yet). Surface this
    # to the template as "latest_period_with_data" so the displayed grid
    # never appears empty just because the workbook has empty future cells.
    periods_with_data = [p for p in sorted_periods if totals.get(p, 0) > 0]
    if not periods_with_data:
        out["error"] = (
            "Every extracted period summed to zero. Check that the pool "
            "mapping below assigns at least one balance-sheet label to a "
            "real pool (not 'Ignore')."
        )
        return out
    latest_with_data = periods_with_data[-1]
    out["latest_period_with_data"] = latest_with_data

    # Pick top pools by balance at the latest period that actually has data
    pool_to_latest: dict[str, float] = {}
    latest_pools = (by_period.get(latest_with_data) or {}).get("by_pool") or {}
    for pool, bal in latest_pools.items():
        try:
            b = float(bal or 0)
        except (TypeError, ValueError):
            continue
        if b > 0:
            pool_to_latest[str(pool)] = b
    # Top N pools by latest balance
    top_pools = [
        p for p, _ in sorted(
            pool_to_latest.items(), key=lambda kv: kv[1], reverse=True
        )
    ][:10]
    out["pool_count"] = len(pool_to_latest)
    out["top_pools"] = top_pools

    # Show the most-recent N months WITH DATA as columns
    recent = periods_with_data[-max_recent_months:]
    out["recent_periods"] = recent
    out["totals_by_period"] = {p: totals.get(p, 0.0) for p in recent}

    grid: dict[str, dict[str, float]] = {}
    for pool in top_pools:
        row: dict[str, float] = {}
        for per in recent:
            by_pool = (by_period.get(per) or {}).get("by_pool") or {}
            try:
                row[per] = round(float(by_pool.get(pool, 0) or 0), 2)
            except (TypeError, ValueError):
                row[per] = 0.0
        grid[pool] = row
    out["balances_grid"] = grid
    out["ok"] = True
    return out
    used = {p.get("id") for p in profiles}
    n = 1
    while f"p{n}" in used:
        n += 1
    return f"p{n}"


def _find_or_create_profile(
    he: dict[str, Any], headers: list[str], signature: str, source_filename: str
) -> dict[str, Any]:
    """Return the profile dict matching ``signature``, creating it if new."""
    for prof in he["profiles"]:
        if prof.get("signature") == signature:
            return prof
    new_id = _next_profile_id(he["profiles"])
    label = f"Profile {len(he['profiles']) + 1} ({source_filename})"
    prof = {
        "id": new_id,
        "signature": signature,
        "headers": list(headers),
        "column_mappings": {},
        "member_account": {
            "mode": "fixed_suffix",
            "suffix_length": 3,
            "delimiter": "-",
        },
        "label": label,
        "source_filename": source_filename,
    }
    he["profiles"].append(prof)
    return prof


def _add_hist_extract_anchor(
    state: dict[str, Any], file_storage
) -> tuple[str, dict[str, Any]]:
    """Save an anchor extract upload and attach it to a profile.

    Returns ``(saved_filename, anchor_entry)``.  Raises on save / read errors.
    """
    he = _ensure_hist_extracts(state)
    saved = _save_hist_upload(file_storage, subfolder="extract_anchors")
    head_info = extract_hist_service.read_extract_headers(saved)
    headers = head_info.get("headers") or []
    if not head_info.get("ok") or not headers:
        raise RuntimeError(
            head_info.get("error") or "Could not read a header row from this file."
        )
    signature = extract_hist_service.compute_header_signature(headers)
    profile = _find_or_create_profile(he, headers, signature, saved.name)
    date_info = extract_hist_service.detect_as_of_date(
        file_storage.filename or saved.name, path=saved
    )
    entry = {
        "name": saved.name,
        "path": str(saved),
        "signature": signature,
        "detected_date": date_info.get("date", ""),
        "detected_source": date_info.get("source", ""),
        "detected_confidence": date_info.get("confidence", "none"),
        "override_date": "",
        "profile_id": profile["id"],
    }
    # Replace any prior entry with the same saved filename.
    he["anchor_files"] = [
        e for e in he["anchor_files"] if e.get("name") != saved.name
    ]
    he["anchor_files"].append(entry)
    _save_state(state)
    return saved.name, entry


def _add_hist_extract_anchor_from_path(
    state: dict[str, Any], source_path: Path
) -> tuple[str, dict[str, Any]]:
    """Copy an existing on-disk extract into the anchor area and register it.

    Mirrors :func:`_add_hist_extract_anchor` but for files the user picked
    from the scan results (no upload involved).
    """
    he = _ensure_hist_extracts(state)
    dest_dir = _HIST_DIR / "extract_anchors"
    dest_dir.mkdir(parents=True, exist_ok=True)
    fn = secure_filename(source_path.name) or "anchor.xlsx"
    target = dest_dir / fn
    # Avoid clobbering a different file already saved under this name.
    if target.exists() and target.resolve() != source_path.resolve():
        stem, suffix = target.stem, target.suffix
        n = 2
        while (dest_dir / f"{stem}_{n}{suffix}").exists():
            n += 1
        target = dest_dir / f"{stem}_{n}{suffix}"
    if target.resolve() != source_path.resolve():
        import shutil
        shutil.copy2(source_path, target)
    head_info = extract_hist_service.read_extract_headers(target)
    headers = head_info.get("headers") or []
    if not head_info.get("ok") or not headers:
        raise RuntimeError(
            head_info.get("error") or "Could not read a header row from this file."
        )
    signature = extract_hist_service.compute_header_signature(headers)
    profile = _find_or_create_profile(he, headers, signature, target.name)
    date_info = extract_hist_service.detect_as_of_date(
        source_path.name, path=target
    )
    entry = {
        "name": target.name,
        "path": str(target),
        "signature": signature,
        "detected_date": date_info.get("date", ""),
        "detected_source": date_info.get("source", ""),
        "detected_confidence": date_info.get("confidence", "none"),
        "override_date": "",
        "profile_id": profile["id"],
    }
    he["anchor_files"] = [
        e for e in he["anchor_files"] if e.get("name") != target.name
    ]
    he["anchor_files"].append(entry)
    _save_state(state)
    return target.name, entry


def _remove_hist_extract_anchor(state: dict[str, Any], name: str) -> bool:
    """Remove an anchor file by saved name.  Also drops orphan profiles."""
    he = _ensure_hist_extracts(state)
    before = len(he["anchor_files"])
    he["anchor_files"] = [e for e in he["anchor_files"] if e.get("name") != name]
    removed = len(he["anchor_files"]) != before
    # Drop profiles no longer referenced by any anchor file or scan result.
    used_pids = {e.get("profile_id") for e in he["anchor_files"]}
    he["profiles"] = [p for p in he["profiles"] if p.get("id") in used_pids]
    if removed:
        _save_state(state)
    return removed


def _aggregate_history_by_pool(
    view: dict[str, Any] | None,
    pool_map: dict[str, str] | None,
) -> dict[str, Any] | None:
    """Roll a per-loan-code history matrix up to per-pool.

    ``view`` is the dict returned by ``chargeoff_hist_processor.history_matrix``
    (or its recovery / delinquency twins). ``pool_map`` is
    ``state['pool_map']``: a case-insensitive lookup from raw loan_code
    (numeric like ``'31'`` or NCUA canonical like ``'Used Vehicles'``)
    to a configured pool name. Codes whose mapped pool is empty,
    "Ignore", or missing are dropped from the rolled-up view (those
    codes still surface in the inline "unmapped codes" picker).

    Returns a new dict with the same shape but with column key = pool
    name. ``cells[m][pool]`` carries an aggregated ``amount`` across
    every contributing raw code. Returns ``view`` unchanged when no
    pool_map is supplied (so the table still renders something useful
    when the user hasn't done any mapping yet).
    """
    if not view:
        return view
    if not isinstance(view, dict):
        return view
    if "codes" not in view or "cells" not in view:
        return view
    pm_ci: dict[str, str] = {}
    for k, v in (pool_map or {}).items():
        if not isinstance(k, str):
            continue
        ks = k.strip().lower()
        if not ks:
            continue
        pm_ci[ks] = (v or "").strip() if isinstance(v, str) else ""
    if not pm_ci:
        return view

    def _pool_for(code: str) -> str:
        if not isinstance(code, str):
            return ""
        return pm_ci.get(code.strip().lower(), "")

    months = list(view.get("months") or [])
    raw_cells = view.get("cells") or {}
    raw_totals_code = view.get("totals_by_code") or {}

    new_cells: dict[str, dict[str, dict[str, float]]] = {}
    new_totals_month: dict[str, dict[str, Any]] = {}
    new_totals_code: dict[str, dict[str, Any]] = {}
    pool_first_month: dict[str, str] = {}
    pool_last_month: dict[str, str] = {}
    pools_seen: set[str] = set()
    row_count = 0

    # Months are stored newest-first; iterate oldest->newest for
    # NEW/DROP first/last detection.
    for m in reversed(months):
        per_month = raw_cells.get(m) or {}
        agg: dict[str, float] = {}
        for code, cell in per_month.items():
            pool = _pool_for(code)
            if not pool or pool.lower() == "ignore":
                continue
            amt = float((cell or {}).get("amount") or 0.0)
            agg[pool] = agg.get(pool, 0.0) + amt
        if not agg:
            continue
        new_cells[m] = {p: {"amount": a} for p, a in agg.items()}
        tm_amount = sum(agg.values())
        new_totals_month[m] = {"amount": tm_amount, "codes": len(agg)}
        for p, a in agg.items():
            pools_seen.add(p)
            tc = new_totals_code.setdefault(p, {"amount": 0.0, "months": 0})
            tc["amount"] += a
            tc["months"] += 1
            if p not in pool_first_month:
                pool_first_month[p] = m
            pool_last_month[p] = m
            row_count += 1

    pools_sorted = sorted(pools_seen)
    months_sorted = sorted(new_cells.keys(), reverse=True)
    new_in_month: dict[str, list[str]] = {m: [] for m in months_sorted}
    dropped_after_month: dict[str, list[str]] = {m: [] for m in months_sorted}
    last_month_iso = months_sorted[0] if months_sorted else ""
    for p in pools_sorted:
        fm = pool_first_month.get(p)
        if fm and fm in new_in_month:
            new_in_month[fm].append(p)
        lm = pool_last_month.get(p)
        if lm and lm != last_month_iso and lm in dropped_after_month:
            dropped_after_month[lm].append(p)

    return {
        "months": months_sorted,
        "codes": pools_sorted,
        "cells": new_cells,
        "totals_by_month": new_totals_month,
        "totals_by_code": new_totals_code,
        "new_in_month": new_in_month,
        "dropped_after_month": dropped_after_month,
        "row_count": row_count,
        "_aggregated_by_pool": True,
    }


def _aggregate_dq_history_by_pool(
    view: dict[str, Any] | None,
    pool_map: dict[str, str] | None,
) -> dict[str, Any] | None:
    """Pool-rollup variant for the DQ history matrix.

    DQ cells carry ``dq_pct`` / ``dq_amount`` / ``total_balance``; the
    template formats either ``dq_pct`` directly or ``amount/total``.
    Aggregation sums ``amount`` and ``total_balance`` per pool, then
    derives the rolled-up ``dq_pct`` from those sums (balance-weighted).
    """
    if not view or not isinstance(view, dict):
        return view
    if "codes" not in view or "cells" not in view:
        return view
    pm_ci: dict[str, str] = {}
    for k, v in (pool_map or {}).items():
        if not isinstance(k, str):
            continue
        ks = k.strip().lower()
        if not ks:
            continue
        pm_ci[ks] = (v or "").strip() if isinstance(v, str) else ""
    if not pm_ci:
        return view

    def _pool_for(code: str) -> str:
        if not isinstance(code, str):
            return ""
        return pm_ci.get(code.strip().lower(), "")

    months = list(view.get("months") or [])
    raw_cells = view.get("cells") or {}
    new_cells: dict[str, dict[str, dict[str, Any]]] = {}
    pools_seen: set[str] = set()
    totals_code: dict[str, dict[str, Any]] = {}
    row_count = 0

    for m in months:
        per_month = raw_cells.get(m) or {}
        agg: dict[str, dict[str, Any]] = {}
        for code, cell in per_month.items():
            pool = _pool_for(code)
            if not pool or pool.lower() == "ignore":
                continue
            amt = (cell or {}).get("amount")
            tot = (cell or {}).get("total_balance")
            pct = (cell or {}).get("dq_pct")
            slot = agg.setdefault(pool, {
                "amount": 0.0, "total_balance": 0.0,
                "_pct_num": 0.0, "_pct_den": 0.0,
                "_have_amount": False, "_have_total": False,
            })
            try:
                amt_v = float(amt) if amt is not None else 0.0
            except (TypeError, ValueError):
                amt_v = 0.0
            slot["amount"] += amt_v
            if amt is not None:
                slot["_have_amount"] = True
            if tot is not None:
                try:
                    tv = float(tot)
                except (TypeError, ValueError):
                    tv = 0.0
                slot["total_balance"] += tv
                slot["_have_total"] = True
            elif pct is not None:
                # No balance available; weight the % by amount so the
                # rolled-up dq_pct stays sensible across mixed sources.
                try:
                    pv = float(pct)
                except (TypeError, ValueError):
                    pv = 0.0
                slot["_pct_num"] += pv * max(amt_v, 1.0)
                slot["_pct_den"] += max(amt_v, 1.0)
        if not agg:
            continue
        out_cells: dict[str, dict[str, Any]] = {}
        for pool, slot in agg.items():
            cell: dict[str, Any] = {"amount": slot["amount"]}
            tot_val = slot["total_balance"] if slot["_have_total"] else None
            if tot_val is not None and tot_val > 0:
                cell["total_balance"] = tot_val
                cell["dq_pct"] = slot["amount"] / tot_val
            elif slot["_pct_den"] > 0:
                cell["dq_pct"] = slot["_pct_num"] / slot["_pct_den"]
                cell["total_balance"] = None
            else:
                cell["total_balance"] = None
                cell["dq_pct"] = None
            out_cells[pool] = cell
            pools_seen.add(pool)
            tc = totals_code.setdefault(
                pool, {"amount": 0.0, "months": 0}
            )
            tc["amount"] += slot["amount"]
            tc["months"] += 1
            row_count += 1
        new_cells[m] = out_cells

    pools_sorted = sorted(pools_seen)
    months_sorted = sorted(new_cells.keys(), reverse=True)
    return {
        "months": months_sorted,
        "codes": pools_sorted,
        "cells": new_cells,
        "totals_by_code": totals_code,
        "row_count": row_count,
        "_aggregated_by_pool": True,
    }


# =================================================================
# Historical-balance provenance — surface HOW the monthly historical
# balances were compiled so an analyst can validate them.
# =================================================================
# Plain-English descriptions for each user-selectable source. The
# "derived_from_loan_extracts" method is not user-selectable; it is
# recorded explicitly (e.g. by a setup script) when the balances were
# computed from the imported loan-data snapshots rather than uploaded.
_HIST_SOURCE_DESCRIPTIONS = {
    "single_workbook": (
        "Uploaded spreadsheet",
        "Read directly from a single workbook that lists each pool's balance "
        "by month.",
    ),
    "monthly_loan_extracts": (
        "One loan-data extract per month",
        "Each month's pool balance is the sum of the loan balances in that "
        "month's loan-data extract file.",
    ),
    "monthly_balance_sheets": (
        "One balance sheet per month",
        "Each month's pool balance is read from that month's balance-sheet "
        "file.",
    ),
    "annual_balance_sheets": (
        "One balance sheet per year",
        "Each year's pool balance is read from that year's balance-sheet file.",
    ),
    "derived_from_loan_extracts": (
        "Derived from imported loan extracts",
        "Not uploaded as a balance file. Each pool's monthly balance was "
        "computed as the sum of current_balance for all loans in that pool, "
        "from that month's imported loan-data extract (monthly_loan_data).",
    ),
}


def build_derived_loan_extract_provenance(cu_name: str) -> dict[str, Any]:
    """Build the provenance record for balances derived from loan extracts.

    Shared by the setup wizard and the balance-build scripts so scripted
    setups record the same honest, validator-friendly provenance. Queries the
    DB for the real coverage (date range / months / pools).
    """
    label, summary = _HIST_SOURCE_DESCRIPTIONS["derived_from_loan_extracts"]
    coverage = {}
    try:
        coverage = pipeline_service.balance_coverage_for_cu(cu_name) or {}
    except Exception:  # noqa: BLE001
        coverage = {}
    from datetime import datetime as _dt
    return {
        "method": "derived_from_loan_extracts",
        "label": label,
        "summary": summary,
        "inputs": ["Imported monthly loan-data extracts (monthly_loan_data)"],
        "coverage": coverage,
        "generated_by": "Auto-derived during setup (not an uploaded balance file)",
        "validation_hint": (
            "To validate a cell: open that month's loan-data extract, filter to "
            "the pool, and sum the current-balance column — it should equal the "
            "balance shown here."
        ),
        "generated_at": _dt.now().strftime("%Y-%m-%d %H:%M"),
    }


def _hist_balance_provenance_view(state: dict[str, Any]) -> dict[str, Any]:
    """Return a display-ready provenance record for the Historical Balances
    step. An explicit ``state['hist_balance_provenance']`` always wins (it is
    the honest record of how the data was actually produced); otherwise a
    best-effort description is derived from the configured
    ``hist_balance_source``. ``recorded`` is False when the source has not been
    confirmed, so the UI can prompt the analyst instead of masking the gap.
    """
    explicit = state.get("hist_balance_provenance")
    if isinstance(explicit, dict) and explicit.get("method"):
        view = dict(explicit)
        view.setdefault("recorded", True)
        if not view.get("coverage"):
            try:
                view["coverage"] = pipeline_service.balance_coverage_for_cu(
                    (state.get("credit_union") or "").strip()
                ) or {}
            except Exception:  # noqa: BLE001
                view["coverage"] = {}
        return view

    # Fallback: the live session may predate a provenance record that was
    # written straight to the saved config (e.g. by a setup/backfill script,
    # or a prior save the current session never reloaded). Read it back from
    # the persisted YAML so the panel still reflects the real, recorded source
    # without forcing the analyst to re-resume the draft.
    short = (state.get("short_name") or "").strip()
    if short:
        try:
            saved = config_service.load_client_config(
                current_app.config["WORKSPACE_ROOT"], short
            ) or {}
        except Exception:  # noqa: BLE001
            saved = {}
        saved_prov = saved.get("hist_balance_provenance")
        if isinstance(saved_prov, dict) and saved_prov.get("method"):
            view = dict(saved_prov)
            view["recorded"] = True
            if not view.get("coverage"):
                try:
                    view["coverage"] = pipeline_service.balance_coverage_for_cu(
                        (state.get("credit_union") or "").strip()
                    ) or {}
                except Exception:  # noqa: BLE001
                    view["coverage"] = {}
            return view

    source = (state.get("hist_balance_source") or "").strip()
    desc = _HIST_SOURCE_DESCRIPTIONS.get(source)
    # Treat the unconfigured default as "not recorded": single_workbook is the
    # initial value, so if no workbook is actually attached we can't honestly
    # claim it — flag it so the analyst confirms the real source.
    mb = state.get("monthly_bal") or {}
    has_single_wb = bool(mb.get("saved_path") or mb.get("filename"))
    if source == "single_workbook" and not has_single_wb:
        return {
            "method": "",
            "recorded": False,
            "label": "Source not recorded",
            "summary": (
                "The source of these historical balances has not been "
                "confirmed. Choose how they were compiled below so a reviewer "
                "can validate them."
            ),
            "inputs": [],
            "coverage": {},
        }
    if not desc:
        return {"method": source, "recorded": False, "label": "",
                "summary": "", "inputs": [], "coverage": {}}
    label, summary = desc
    return {
        "method": source,
        "recorded": True,
        "label": label,
        "summary": summary,
        "inputs": [],
        "coverage": {},
        "validation_hint": "",
    }


# =================================================================
# Charge-off / Recovery provenance — mirror of the balance panel so an
# analyst can see (and validate) HOW the historical charge-offs and
# recoveries were compiled: CU-provided files, NCUA 5300 backfill,
# monthly summaries, or a mix.
# =================================================================
def _co_recov_cu_file_inputs(state: dict[str, Any], section: str) -> list[dict[str, str]]:
    """Return the CU-provided file layouts feeding this section (co|recov).

    Reads the wizard-state shape first (``co_recov_formats`` /
    ``co_columns`` / ``recov_columns``); falls back to the saved config
    shape (``historical_file_formats``) so a stale session still shows the
    persisted source.
    """
    want = "co_columns" if section == "co" else "recov_columns"
    inputs: list[dict[str, str]] = []
    for f in (state.get("co_recov_formats") or []):
        cols = f.get(want) or {}
        if (cols.get("amount_col") is not None
                or cols.get("code_static")
                or cols.get("code_col") is not None):
            inputs.append({
                "name": f.get("name") or "(unnamed format)",
                "file_pattern": f.get("file_pattern") or "",
            })
    if inputs:
        return inputs
    single = state.get(want) or {}
    if single.get("amount_col") is not None:
        inputs.append({
            "name": "Charge-off file" if section == "co" else "Recovery file",
            "file_pattern": "",
        })
    if inputs:
        return inputs
    # Fallback: saved config's historical_file_formats.
    short = (state.get("short_name") or "").strip()
    if not short:
        return inputs
    try:
        cfg = config_service.load_client_config(
            current_app.config["WORKSPACE_ROOT"], short
        ) or {}
    except Exception:  # noqa: BLE001
        return inputs
    hff = cfg.get("historical_file_formats") or {}
    dst = "chargeoff" if section == "co" else "recovery"
    formats = hff.get("formats")
    if isinstance(formats, list) and formats:
        for f in formats:
            if f.get(dst):
                inputs.append({
                    "name": f.get("name") or "(unnamed format)",
                    "file_pattern": f.get("file_pattern") or "",
                })
    elif hff.get(dst):
        inputs.append({
            "name": "Charge-off file" if section == "co" else "Recovery file",
            "file_pattern": "",
        })
    return inputs


def _co_recov_provenance_view(state: dict[str, Any], section: str) -> dict[str, Any]:
    """Return a display-ready provenance record for the CO / recovery step.

    Combines the CU-provided file layouts with the DB-stored origins (NCUA
    5300 backfill, monthly summaries, uploaded workbook) so a validator can
    see where each part of the history came from and how to tie it out.
    ``recorded`` is False only when nothing can be stated.
    """
    is_recov = (section == "recov")
    noun = "recovery" if is_recov else "charge-off"

    if is_recov and state.get("no_hist_recoveries"):
        return {
            "recorded": True,
            "no_recoveries": True,
            "label": "No historical recoveries",
            "summary": ("This credit union was marked as having no historical "
                        "recoveries, so none are included in the analysis."),
            "cu_files": [],
            "db_sources": [],
            "validation_hint": "",
        }

    cu_name = (state.get("credit_union") or "").strip()
    cu_files = _co_recov_cu_file_inputs(state, section)

    cov = {}
    try:
        cov = pipeline_service.co_recov_coverage_for_cu(cu_name) or {}
    except Exception:  # noqa: BLE001
        cov = {}
    sec_cov = cov.get("recovery" if is_recov else "chargeoff") or {}
    db_groups = [
        {
            "label": g.get("group") or "",
            "years": g.get("years") or "",
            "total": float(g.get("total") or 0),
        }
        for g in (sec_cov.get("groups") or [])
        if float(g.get("total") or 0) or (g.get("rows") or 0)
    ]

    has_cu = bool(cu_files)
    has_5300 = any("5300" in g["label"] for g in db_groups)
    has_db = bool(db_groups)

    if not has_cu and not has_db:
        return {
            "recorded": False,
            "label": "Source not recorded",
            "summary": (f"The source of these historical {noun}s has not been "
                        "confirmed. Configure it below so a reviewer can "
                        "validate the numbers."),
            "cu_files": [],
            "db_sources": [],
        }

    # Method + human label.
    if has_cu and has_5300:
        method = "cu_files+5300"
        label = f"CU-provided {noun} files + NCUA 5300 backfill"
    elif has_cu and has_db:
        method = "cu_files+db"
        label = f"CU-provided {noun} files + supplemental backfill"
    elif has_cu:
        method = "cu_files"
        label = f"CU-provided {noun} files"
    elif has_5300:
        method = "5300_backfill"
        label = "NCUA 5300 Call Report backfill"
    else:
        method = "db"
        label = "Backfilled from stored history"

    parts: list[str] = []
    if has_cu:
        parts.append(
            f"The credit union's own {noun} file(s) are parsed at report time "
            f"(each row's amount summed by loan pool and year)."
        )
    if has_db:
        parts.append(
            "Years not covered by the CU files are filled from stored history "
            "(e.g. the NCUA 5300 Call Report backfill)."
        )
    if has_cu and has_db:
        parts.append(
            "Where both exist, the CU's own files take precedence per year; the "
            "backfill only fills the years the CU files don't cover."
        )
    summary = " ".join(parts)

    hint_parts: list[str] = []
    if has_cu:
        hint_parts.append(
            f"open the matching CU {noun} file for a period, filter to the loan "
            f"type/pool, and sum the {noun} amount column"
        )
    if has_5300:
        hint_parts.append(
            f"for backfilled years, tie to the CU's NCUA 5300 Call Report "
            f"({noun} schedule)"
        )
    validation_hint = ("To validate a year: " + "; ".join(hint_parts) + "."
                       if hint_parts else "")

    return {
        "recorded": True,
        "method": method,
        "label": label,
        "summary": summary,
        "cu_files": cu_files,
        "db_sources": db_groups,
        "validation_hint": validation_hint,
    }


def _files_used_for_specs(
    state: dict[str, Any], specs: list[dict[str, str]]
) -> dict[str, Any]:
    """Scan the CU's data directory for files matching each spec's pattern.

    ``specs`` is a list of ``{"name", "file_pattern"}``. Returns
    ``{data_directory, ok, error, groups: [{name, file_pattern, match_count,
    error, files: [{name, period, subdir}]}]}`` — a display-ready record of
    exactly which files the report will read, so a reviewer can validate the
    source data. Shared by the charge-off/recovery, loan-extract and
    monthly-balance "Files used" panels.
    """
    data_dir = (state.get("data_directory") or "").strip()
    date_pattern = state.get("date_pattern") or ""
    date_format = state.get("date_format") or "YYYY-MM"
    out: dict[str, Any] = {
        "data_directory": data_dir,
        "ok": bool(data_dir),
        "error": None,
        "groups": [],
    }
    if not data_dir:
        out["error"] = "No data directory is set for this credit union yet."
        return out
    for spec in specs:
        fp = (spec.get("file_pattern") or "").strip()
        if not fp:
            continue
        scan = _scan_data_directory(data_dir, fp, date_pattern, date_format)
        matched = [r for r in (scan.get("files") or []) if r.get("matched")]
        matched.sort(key=lambda r: (r.get("date") or ""), reverse=True)
        out["groups"].append({
            "name": spec.get("name") or "(unnamed)",
            "file_pattern": fp,
            "error": scan.get("error"),
            "match_count": len(matched),
            "files": [
                {"name": r.get("name"),
                 "period": r.get("date") or "",
                 "subdir": r.get("subdir") or ""}
                for r in matched
            ],
        })
    return out


def _co_recov_files_used(state: dict[str, Any]) -> dict[str, Any]:
    """Describe the actual charge-off / recovery files the report will read.

    For each configured CO/recovery format, scan the data directory for files
    matching its ``file_pattern``. Especially useful when the config was built
    by a script (no per-file sample uploads) or the CU has many monthly files.
    """
    formats = list(state.get("co_recov_formats") or [])
    # Fall back to the saved config's historical_file_formats.formats so a
    # stale / script-built session still shows the persisted patterns.
    if not formats:
        short = (state.get("short_name") or "").strip()
        if short:
            try:
                cfg = config_service.load_client_config(
                    current_app.config["WORKSPACE_ROOT"], short
                ) or {}
                hff = cfg.get("historical_file_formats") or {}
                cfg_formats = hff.get("formats")
                if isinstance(cfg_formats, list):
                    formats = cfg_formats
            except Exception:  # noqa: BLE001
                formats = []
    specs = [
        {"name": f.get("name") or "(unnamed format)",
         "file_pattern": f.get("file_pattern") or ""}
        for f in formats
    ]
    return _files_used_for_specs(state, specs)


def _loan_extract_files_used(state: dict[str, Any]) -> dict[str, Any]:
    """Describe the actual loan-data extract files the report will read.

    Builds one spec per loan extract (from the wizard's ``loan_data_files``,
    else the saved config's ``loan_data_extracts``, else the single top-level
    ``file_pattern``) and scans the data directory for matches.
    """
    specs: list[dict[str, str]] = []
    for lf in ((state.get("sample_uploads") or {}).get("loan_data_files") or []):
        fp = (lf.get("file_pattern") or "").strip()
        if fp:
            specs.append({"name": lf.get("name") or "Loan extract",
                          "file_pattern": fp})
    if not specs:
        short = (state.get("short_name") or "").strip()
        if short:
            try:
                cfg = config_service.load_client_config(
                    current_app.config["WORKSPACE_ROOT"], short
                ) or {}
                for ex in (cfg.get("loan_data_extracts") or []):
                    fp = (ex.get("file_pattern") or "").strip()
                    if fp:
                        specs.append({
                            "name": ex.get("label") or ex.get("name")
                            or "Loan extract",
                            "file_pattern": fp,
                        })
            except Exception:  # noqa: BLE001
                pass
    if not specs:
        fp = (state.get("file_pattern") or "").strip()
        if fp:
            specs.append({"name": "Loan data extract", "file_pattern": fp})
    return _files_used_for_specs(state, specs)


def _monthly_balance_files_used(state: dict[str, Any]) -> dict[str, Any]:
    """Describe the monthly-balance source the report will read.

    Handles each ``monthly_bal.source`` mode: a discovery ``monthly_file_pattern``
    is scanned against the data directory; explicit uploads (single workbook /
    per-month / per-year) are listed directly; and when balances are DERIVED
    from the loan extracts (no separate file), a note explains the origin using
    the recorded historical-balance provenance so a reviewer can still tie out.
    """
    mb = state.get("monthly_bal") or {}
    source = (mb.get("source") or "single").strip().lower()
    pattern = (mb.get("monthly_file_pattern") or "").strip()

    if pattern:
        out = _files_used_for_specs(
            state, [{"name": "Monthly balance snapshots", "file_pattern": pattern}]
        )
        out["note"] = None
        return out

    data_dir = (state.get("data_directory") or "").strip()
    out = {"data_directory": data_dir, "ok": True, "error": None,
           "groups": [], "note": None, "inputs": []}

    def _explicit(entries: list, label: str) -> None:
        files = []
        for e in entries or []:
            nm = (e.get("filename") or e.get("name")
                  or (Path(e.get("saved_path", "")).name
                      if e.get("saved_path") else ""))
            if nm:
                files.append({"name": nm, "period": e.get("period") or "",
                              "subdir": ""})
        if files:
            out["groups"].append({
                "name": label, "file_pattern": "", "error": None,
                "match_count": len(files), "files": files})

    if source == "single":
        nm = (mb.get("filename")
              or (Path(mb.get("saved_path", "")).name
                  if mb.get("saved_path") else ""))
        if nm:
            out["groups"].append({
                "name": "Monthly balance workbook", "file_pattern": "",
                "error": None, "match_count": 1,
                "files": [{"name": nm, "period": "", "subdir": ""}]})
    elif source == "per_month":
        _explicit(mb.get("monthly_files"), "Per-month balance files")
    elif source == "per_year":
        _explicit(mb.get("year_files"), "Per-year balance workbooks")

    if not out["groups"]:
        if source == "manual":
            out["note"] = ("Monthly balances were entered manually on this "
                           "step, so there is no separate source file.")
        else:
            prov = None
            try:
                prov = _hist_balance_provenance_view(state)
            except Exception:  # noqa: BLE001
                prov = None
            if prov and prov.get("recorded"):
                out["note"] = (
                    "No standalone monthly-balance file is configured; "
                    "balances are derived. " + (prov.get("summary") or "")
                ).strip()
                out["inputs"] = list(prov.get("inputs") or [])
            else:
                out["note"] = (
                    "No standalone monthly-balance file is configured. "
                    "Balances may be derived from the imported loan extracts.")
    return out


def _uploaded_balance_periods(state: dict[str, Any]) -> set[str]:
    """Month-end periods (ISO) the CU already has real balance data for, read
    from its configured historical-balance source. Used so the auto 5300
    backfill only fills EARLIER quarters (annual / monthly balance sheets;
    single_workbook returns empty and relies on the report's workbook-wins
    merge)."""
    mb = state.get("monthly_bal") or {}
    source = state.get("hist_balance_source") or ""
    label_to_pool = mb.get("pool_map") or {}
    try:
        if source == "annual_balance_sheets":
            ann = monthly_bal_parser.pool_balances_for_per_year_files(
                mb.get("year_files") or [], mb.get("per_year_layout") or {},
                label_to_pool, exclude_labels=mb.get("exclude_labels") or [])
            return set((ann.get("by_period") or {}).keys())
        if source == "monthly_balance_sheets":
            pm = monthly_bal_parser.pool_balances_for_per_month_files(
                mb.get("monthly_files") or [], mb.get("per_month_layout") or {},
                label_to_pool, exclude_labels=mb.get("exclude_labels") or [])
            return set((pm.get("by_period") or {}).keys())
    except Exception:  # noqa: BLE001
        return set()
    return set()


def _auto_run_5300_balance_backfill(state: dict[str, Any]) -> None:
    """Default behavior: when the analyst leaves the Historical Balances step,
    automatically fill the earlier quarters' balances from the NCUA 5300
    (distributed mode) — so 5300 backfill is the out-of-the-box default rather
    than a button the analyst has to remember. Opt out via
    ``solr_backfill['enabled'] = False``. Runs at most once per draft, only in
    distributed mode, only when the prerequisites (charter, period, a
    computable pool distribution) are present, and never blocks navigation.
    The manual "Run" button remains available for precise control.
    """
    he = state.get("hist_extracts") or {}
    sb = he.get("solr_backfill") or {}
    if not sb.get("enabled", True) or sb.get("auto_ran"):
        return
    if (sb.get("mode") or "distributed") != "distributed":
        return  # per-loan-code is a manual special case
    lr = sb.get("last_run") or {}
    if isinstance(lr, dict) and lr.get("ok"):
        return  # already backfilled (e.g. manual run)
    cu = (state.get("credit_union") or "").strip()
    try:
        charter = int(re.sub(r"\D", "", state.get("charter_number") or ""))
    except (TypeError, ValueError):
        charter = 0
    period = (he.get("target_period") or "").strip()
    months = int(he.get("history_months") or 84)
    if not (cu and charter and period):
        return  # not enough info yet — analyst can run it manually later
    dist = _earliest_month_pool_distribution(state)
    if not dist.get("ok") or not dist.get("pool_distribution"):
        return  # no distribution (e.g. no uploaded balance file) — skip quietly
    existing: set[str] = set()
    try:
        hv = extract_hist_processor.history_matrix(cu)
        existing = set((hv or {}).get("months") or [])
    except Exception:  # noqa: BLE001
        existing = set()
    existing |= _uploaded_balance_periods(state)
    sb["auto_ran"] = True
    try:
        res = solr_5300_backfill.backfill_missing_quarters_distributed(
            cu, charter, sb.get("solr_url"), sb.get("core"),
            period, months,
            pool_distribution=dist["pool_distribution"],
            existing_dates=existing,
            source_period_iso=dist.get("period") or "",
        )
    except Exception as exc:  # noqa: BLE001
        flash(
            "Auto 5300 balance backfill was skipped (" f"{exc}"
            "). You can run it manually from the Historical Balances step.",
            "warning",
        )
        _save_state(state)
        return
    sb["last_run"] = res
    _save_state(state)
    if res.get("ok"):
        filled = len(res.get("months_filled") or [])
        if filled:
            flash(
                f"Automatically backfilled {filled} earlier month(s) of pool "
                f"balances from the NCUA 5300 (distributed by the "
                f"{dist.get('period')} pool mix). Untick 'Auto-fill from 5300' "
                "on the Historical Balances step to disable this.",
                "success",
            )
    else:
        flash(f"Auto 5300 balance backfill: {res.get('error')}", "warning")


@setup_bp.route("/step/historical", methods=["GET", "POST"])
@setup_bp.route("/step/co-history", methods=["GET", "POST"],
                endpoint="step3a_co_history")
@setup_bp.route("/step/recov-history", methods=["GET", "POST"],
                endpoint="step3b_recov_history")
def step3_historical():
    """Historical Data step — only shown for CUs without a WARM file.

    Split across three wizard steps (one URL each) so the page stays
    navigable; the same view function handles all three by branching on
    ``request.endpoint``:

      - ``/step/historical``       — Historical loan balances by pool/type
      - ``/step/co-history``       — Historical charge-offs
      - ``/step/recov-history``    — Historical recoveries (optional)
    """
    state = _state()

    # Which slice of the legacy combined page is this URL rendering?
    _ep = request.endpoint or "setup.step3_historical"
    if _ep.endswith("step3a_co_history"):
        section = "co"
        active_key = "co_history"
        next_endpoint = "setup.step3b_recov_history"
    elif _ep.endswith("step3b_recov_history"):
        section = "recov"
        active_key = "recov_history"
        next_endpoint = "setup.step_dq_hist"
    else:
        section = "balances"
        active_key = "historical"
        next_endpoint = "setup.step3a_co_history"
    back_endpoint = _ep if _ep.startswith("setup.") else f"setup.{_ep}"

    # Auto-scan self-heal: when Step 1's auto-scan classified CO/Recov
    # files in the folder but they never made it into hist_scan (e.g. the
    # user ran the scan before this bulk-stage code was wired up), refill
    # hist_scan.monthly_co_files / monthly_recov_files now so Step 5 and
    # Step 6 render the per-file tables on the next page load.
    if request.method == "GET" and state.get("_auto_scan_completed"):
        folder = state.get("_auto_scan_folder") or ""
        if folder:
            hs = state.get("hist_scan") or {}
            need_co = section in ("co",) and not (hs.get("monthly_co_files") or hs.get("co_files"))
            need_recov = section in ("recov",) and not (hs.get("monthly_recov_files") or hs.get("recov_files"))
            if need_co or need_recov:
                try:
                    from cecl_ui.services import auto_setup as _auto_setup
                    cls = _auto_setup.classify_folder(folder)
                    msgs: list[str] = []
                    if need_co:
                        msgs.extend(_auto_setup._stage_hist_scan_files(
                            state, cls.get("co_files") or [],
                            "monthly_co_files", "co",
                        ))
                    if need_recov:
                        msgs.extend(_auto_setup._stage_hist_scan_files(
                            state, cls.get("recov_files") or [],
                            "monthly_recov_files", "recov",
                        ))
                    if msgs:
                        _save_state(state)
                        for m in msgs:
                            flash(f"Auto-scan: {m}", "success")
                except Exception:  # noqa: BLE001
                    # Self-heal is best-effort; never block the page render.
                    pass

    if request.method == "POST":
        action = request.form.get("action", "next")
        # Persist the recoveries checkbox only on forms that include the
        # marker (the recoveries toggle form). Otherwise other source-
        # picker forms on the page would clobber the user's selection.
        if request.form.get("recov_toggle_marker") == "1" or action == "toggle_no_recov":
            state["no_hist_recoveries"] = (request.form.get("no_hist_recoveries") == "on")

        if action == "set_hist_balance_source":
            choice = (request.form.get("hist_balance_source") or "").strip()
            if choice in (
                "single_workbook",
                "monthly_loan_extracts",
                "monthly_balance_sheets",
                "annual_balance_sheets",
            ):
                state["hist_balance_source"] = choice
                # Annual balance sheets feed the SAME pipeline as Step 5's
                # per_year monthly-balance mode. Auto-flip the Step 5 source
                # so the user only configures it in one place.
                if choice == "annual_balance_sheets":
                    mb = state.setdefault("monthly_bal", {})
                    mb["source"] = "per_year"
                # Monthly balance sheets feed Step 5's per_month mode.
                elif choice == "monthly_balance_sheets":
                    mb = state.setdefault("monthly_bal", {})
                    mb["source"] = "per_month"
                _save_state(state)
                return redirect(url_for(back_endpoint))
            flash("Please choose one of the balance-source options.", "error")

        elif action == "upload_annual_year":
            mb = state.setdefault("monthly_bal", {})
            mb["source"] = "per_year"
            f = request.files.get("annual_year_file")
            year_raw = (request.form.get("annual_year") or "").strip()
            if not f or not f.filename:
                flash("Choose an annual balance workbook to upload.", "error")
            else:
                try:
                    target = _save_monthly_bal_upload(f)
                    category, message = _ingest_annual_workbook(
                        state, mb, target, year_raw)
                    flash(message, category)
                except Exception as exc:  # noqa: BLE001
                    flash(f"Upload failed: {exc}", "error")
            _save_state(state)
            return redirect(url_for(back_endpoint))

        elif action == "scan_annual_folder":
            mb = state.setdefault("monthly_bal", {})
            mb["source"] = "per_year"
            folder_raw = (request.form.get("annual_folder") or "").strip()
            folder = _normalize_folder_path(folder_raw)
            mb["annual_folder"] = folder
            if not folder:
                flash("Enter a folder path to scan.", "error")
            else:
                try:
                    p = Path(folder)
                    if not p.is_dir():
                        flash(
                            f"Folder not found or not accessible: {folder}",
                            "error",
                        )
                    else:
                        candidates = [
                            f for f in sorted(p.iterdir())
                            if f.is_file()
                            and f.suffix.lower() in (".xlsx", ".xlsm", ".xls")
                            and not f.name.startswith(("~$", "."))
                        ]
                        if not candidates:
                            flash(
                                f"No .xlsx/.xls workbooks found in {folder}.",
                                "warning",
                            )
                        else:
                            import shutil
                            ok_n, warn_n, err_n = 0, 0, 0
                            for src in candidates:
                                _MONTHLY_BAL_DIR.mkdir(
                                    parents=True, exist_ok=True)
                                fn = secure_filename(src.name)
                                dest = _MONTHLY_BAL_DIR / fn
                                try:
                                    if dest.resolve() != src.resolve():
                                        shutil.copy2(src, dest)
                                except Exception:  # noqa: BLE001
                                    dest = src
                                category, _msg = _ingest_annual_workbook(
                                    state, mb, dest, year_raw="")
                                if category == "success":
                                    ok_n += 1
                                elif category == "warning":
                                    warn_n += 1
                                else:
                                    err_n += 1
                            parts = [f"Scanned {folder}:",
                                     f"{ok_n} ingested"]
                            if warn_n:
                                parts.append(f"{warn_n} layout warning(s)")
                            if err_n:
                                parts.append(f"{err_n} skipped")
                            flash(
                                ", ".join(parts) + ".",
                                "success" if ok_n else "warning",
                            )
                except Exception as exc:  # noqa: BLE001
                    flash(f"Folder scan failed: {exc}", "error")
            _save_state(state)
            return redirect(url_for(back_endpoint))

        elif action == "remove_annual_year":
            mb = state.setdefault("monthly_bal", {})
            target_name = (request.form.get("filename") or "").strip()
            files = mb.get("year_files") or []
            removed = None
            for e in files:
                if e.get("filename") == target_name:
                    removed = e
                    break
            if removed:
                files.remove(removed)
                sp = removed.get("saved_path") or ""
                if sp and not removed.get("external"):
                    try:
                        p = Path(sp)
                        if p.is_file():
                            p.unlink()
                    except Exception:  # noqa: BLE001
                        pass
                flash(f"Removed {target_name}.", "success")
            else:
                flash(f"File '{target_name}' not found.", "error")
            mb["year_files"] = files
            _save_state(state)
            return redirect(url_for(back_endpoint))

        elif action == "save_annual_pool_map":
            mb = state.setdefault("monthly_bal", {})
            pool_map = mb.setdefault("pool_map", {})
            for key in list(request.form.keys()):
                if key.startswith("pm_label_"):
                    idx = key[len("pm_label_"):]
                    label = (request.form.get(key) or "").strip()
                    if not label:
                        continue
                    pool = (
                        request.form.get(f"pm_pool_{idx}") or "").strip()
                    # "__ignore__" sentinel from the dropdown means
                    # intentionally excluded -- store as empty string.
                    if pool == "__ignore__":
                        pool = ""
                    pool_map[label] = pool
            mb["pool_map"] = pool_map
            _save_state(state)
            flash("Saved pool mapping for annual balance workbooks.",
                  "success")
            return redirect(url_for(back_endpoint))

        elif action == "upload_per_month_step3":
            mb = state.setdefault("monthly_bal", {})
            mb["source"] = "per_month"
            f = request.files.get("per_month_file")
            period_raw = (request.form.get("per_month_period") or "").strip()
            stop_row_raw = (request.form.get("per_month_stop_row") or "").strip()
            as_of_cell_raw = (request.form.get("per_month_as_of_cell") or "").strip().upper()
            try:
                stop_row_val: int | None = int(stop_row_raw) if stop_row_raw else None
            except (TypeError, ValueError):
                stop_row_val = None
            if not f or not f.filename:
                # Allow re-analysing the most recent uploaded sample with
                # a new stop_row — don't force the user to re-pick the file.
                files = mb.get("monthly_files") or []
                last = files[-1] if files else None
                last_path = (last or {}).get("saved_path") or ""
                if last_path and Path(last_path).is_file():
                    try:
                        target = Path(last_path)
                        category, message = _ingest_per_month_workbook(
                            state, mb, target, last.get("period") or "",
                            stop_row=stop_row_val,
                            as_of_cell=as_of_cell_raw)
                        flash(message, category)
                    except Exception as exc:  # noqa: BLE001
                        flash(f"Re-analyse failed: {exc}", "error")
                else:
                    flash(
                        "Choose a monthly balance sheet to upload.",
                        "error",
                    )
            else:
                try:
                    target = _save_monthly_bal_upload(f)
                    category, message = _ingest_per_month_workbook(
                        state, mb, target, period_raw,
                        stop_row=stop_row_val,
                        as_of_cell=as_of_cell_raw)
                    flash(message, category)
                except Exception as exc:  # noqa: BLE001
                    flash(f"Upload failed: {exc}", "error")
            _save_state(state)
            return redirect(url_for(back_endpoint))

        elif action == "reset_step3_balance_sheets":
            mb = state.setdefault("monthly_bal", {})
            # Delete temp-stashed sample workbooks + the pool-mapping
            # upload (best-effort; never blocks the reset).
            for entry in (mb.get("monthly_files") or []):
                sp = entry.get("saved_path") or ""
                if sp and not entry.get("external"):
                    try:
                        pp = Path(sp)
                        if pp.is_file():
                            pp.unlink()
                    except Exception:  # noqa: BLE001
                        pass
            pm_up = mb.get("pool_mapping_upload") or {}
            sp = pm_up.get("saved_path") or ""
            if sp:
                try:
                    pp = Path(sp)
                    if pp.is_file():
                        pp.unlink()
                except Exception:  # noqa: BLE001
                    pass
            # Wipe everything Step 3 touches under monthly_bal but keep
            # the chosen source (per_month) so the radio stays sticky.
            kept_source = mb.get("source") or ""
            for k in (
                "monthly_files", "per_month_layout", "per_month_source_folder",
                "pool_map", "parsed_pool_labels", "label_status",
                "acl", "pool_mapping_upload", "year_files",
                "per_year_layout", "annual_folder",
            ):
                mb.pop(k, None)
            if kept_source:
                mb["source"] = kept_source
            flash(
                "Cleared all Step 3 monthly-balance uploads, mappings, "
                "and detected layout. Start fresh by uploading a sample.",
                "success",
            )
            _save_state(state)
            return redirect(url_for(back_endpoint))

        elif action == "scan_per_month_folder_step3":
            mb = state.setdefault("monthly_bal", {})
            mb["source"] = "per_month"
            folder_raw = (request.form.get("per_month_folder") or "").strip()
            folder = _normalize_folder_path(folder_raw)
            mb["per_month_source_folder"] = folder
            if not folder:
                flash("Enter a folder path to scan.", "error")
            else:
                try:
                    p = Path(folder)
                    if not p.is_dir():
                        flash(
                            f"Folder not found or not accessible: {folder}",
                            "error",
                        )
                    else:
                        candidates = [
                            f for f in sorted(p.iterdir())
                            if f.is_file()
                            and f.suffix.lower() in (
                                ".xlsx", ".xlsm", ".xls", ".csv")
                            and not f.name.startswith(("~$", "."))
                        ]
                        if not candidates:
                            flash(
                                f"No .xlsx/.xls/.csv workbooks found in "
                                f"{folder}.",
                                "warning",
                            )
                        else:
                            import shutil
                            ok_n, warn_n, err_n = 0, 0, 0
                            for src in candidates:
                                _MONTHLY_BAL_DIR.mkdir(
                                    parents=True, exist_ok=True)
                                fn = secure_filename(src.name)
                                dest = _MONTHLY_BAL_DIR / fn
                                try:
                                    if dest.resolve() != src.resolve():
                                        shutil.copy2(src, dest)
                                except Exception:  # noqa: BLE001
                                    dest = src
                                category, _msg = _ingest_per_month_workbook(
                                    state, mb, dest, period_raw="")
                                if category == "success":
                                    ok_n += 1
                                elif category == "warning":
                                    warn_n += 1
                                else:
                                    err_n += 1
                            parts = [f"Scanned {folder}:",
                                     f"{ok_n} ingested"]
                            if warn_n:
                                parts.append(f"{warn_n} layout warning(s)")
                            if err_n:
                                parts.append(
                                    f"{err_n} skipped (no detectable "
                                    "month-end date)")
                            flash(
                                ", ".join(parts) + ".",
                                "success" if ok_n else "warning",
                            )
                except Exception as exc:  # noqa: BLE001
                    flash(f"Folder scan failed: {exc}", "error")
            _save_state(state)
            return redirect(url_for(back_endpoint))

        elif action == "remove_per_month_step3":
            mb = state.setdefault("monthly_bal", {})
            target_name = (request.form.get("filename") or "").strip()
            files = mb.get("monthly_files") or []
            removed = None
            for e in files:
                if e.get("filename") == target_name:
                    removed = e
                    break
            if removed:
                files.remove(removed)
                if not removed.get("external"):
                    sp = removed.get("saved_path") or ""
                    if sp:
                        try:
                            pp = Path(sp)
                            if pp.is_file():
                                pp.unlink()
                        except Exception:  # noqa: BLE001
                            pass
                flash(f"Removed {target_name}.", "success")
            else:
                flash(f"File '{target_name}' not found.", "error")
            mb["monthly_files"] = files
            _save_state(state)
            return redirect(url_for(back_endpoint))

        elif action == "remove_all_per_month_step3":
            mb = state.setdefault("monthly_bal", {})
            files = mb.get("monthly_files") or []
            removed_count = 0
            for e in files:
                if not e.get("external"):
                    sp = e.get("saved_path") or ""
                    if sp:
                        try:
                            pp = Path(sp)
                            if pp.is_file():
                                pp.unlink()
                        except Exception:  # noqa: BLE001
                            pass
                removed_count += 1
            mb["monthly_files"] = []
            if removed_count:
                flash(
                    f"Removed all {removed_count} monthly workbook(s).",
                    "success",
                )
            else:
                flash("No monthly workbooks to remove.", "info")
            _save_state(state)
            return redirect(url_for(back_endpoint))

        elif action == "remove_all_annual_step3":
            mb = state.setdefault("monthly_bal", {})
            files = mb.get("year_files") or []
            removed_count = 0
            for e in files:
                if not e.get("external"):
                    sp = e.get("saved_path") or ""
                    if sp:
                        try:
                            pp = Path(sp)
                            if pp.is_file():
                                pp.unlink()
                        except Exception:  # noqa: BLE001
                            pass
                removed_count += 1
            mb["year_files"] = []
            if removed_count:
                flash(
                    f"Removed all {removed_count} annual workbook(s).",
                    "success",
                )
            else:
                flash("No annual workbooks to remove.", "info")
            _save_state(state)
            return redirect(url_for(back_endpoint))

        elif action == "remove_pool_mapping_upload":
            mb = state.setdefault("monthly_bal", {})
            pm_up = mb.get("pool_mapping_upload") or {}
            fn = pm_up.get("filename") or ""
            sp = pm_up.get("saved_path") or ""
            if sp:
                try:
                    pp = Path(sp)
                    # Only delete if the file lives in the managed
                    # temp upload dir — never touch user-provided
                    # paths outside of it.
                    if pp.is_file() and "cecl_ui_monthly_bal" in str(pp).lower():
                        pp.unlink()
                except Exception:  # noqa: BLE001
                    pass
            mb["pool_mapping_upload"] = {}
            if fn:
                flash(
                    f"Removed pool-mapping file: {fn}. "
                    "Label-to-pool entries it added remain in the table "
                    "above; edit or clear them as needed.",
                    "success",
                )
            else:
                flash("No pool-mapping file to remove.", "info")
            _save_state(state)
            return redirect(url_for(back_endpoint))

        elif action == "save_per_month_pool_map":
            mb = state.setdefault("monthly_bal", {})
            pool_map = mb.setdefault("pool_map", {})
            for key in list(request.form.keys()):
                if key.startswith("pm_label_"):
                    idx = key[len("pm_label_"):]
                    label = (request.form.get(key) or "").strip()
                    if not label:
                        continue
                    pool = (
                        request.form.get(f"pm_pool_{idx}") or "").strip()
                    if pool == "__ignore__":
                        pool = ""
                    pool_map[label] = pool
            mb["pool_map"] = pool_map
            _save_state(state)
            flash("Saved pool mapping for monthly balance sheets.",
                  "success")
            return redirect(url_for(back_endpoint))

        elif action == "upload_pool_mapping":
            mb = state.setdefault("monthly_bal", {})
            f = request.files.get("pool_mapping_file")

            def _col_to_zero_idx(value: str, default_letter: str) -> int:
                """Accept a column letter (A,B,...,AA) OR a 1-based number."""
                raw = (value or "").strip()
                if not raw:
                    raw = default_letter
                if raw.isdigit():
                    n = int(raw)
                    return max(0, n - 1)
                # Letter form.
                idx = monthly_bal_parser._col_letter_to_idx(raw)
                return max(0, idx - 1)

            label_col_idx = _col_to_zero_idx(
                request.form.get("pm_file_label_col") or "", "A")
            pool_col_idx = _col_to_zero_idx(
                request.form.get("pm_file_pool_col") or "", "B")
            header_choice = (request.form.get("pm_file_header") or "auto"
                             ).strip().lower()
            has_header: bool | None
            if header_choice == "yes":
                has_header = True
            elif header_choice == "no":
                has_header = False
            else:
                has_header = None

            def _idx_to_letter(idx: int) -> str:
                """0-based index -> letter (0->'A')."""
                n = idx + 1
                s = ""
                while n > 0:
                    n, r = divmod(n - 1, 26)
                    s = chr(ord("A") + r) + s
                return s or "A"

            label_letter = _idx_to_letter(label_col_idx)
            pool_letter = _idx_to_letter(pool_col_idx)
            if not f or not f.filename:
                # Allow re-parsing the previously uploaded file with new
                # column choices — don't force the user to re-pick it.
                prior = mb.get("pool_mapping_upload") or {}
                prior_path = prior.get("saved_path") or ""
                if prior_path and Path(prior_path).is_file():
                    try:
                        target = Path(prior_path)
                        fn = target.name
                        mapping, err, header_preview = (
                            _parse_pool_mapping_file(
                                target,
                                label_col_idx=label_col_idx,
                                pool_col_idx=pool_col_idx,
                                has_header=has_header,
                            )
                        )
                        sample_pairs = list(
                            (mapping or {}).items())[:8]
                        mb["pool_mapping_upload"] = {
                            "filename": fn,
                            "saved_path": str(target),
                            "header_preview": header_preview,
                            "label_col": label_letter,
                            "pool_col": pool_letter,
                            "header_choice": header_choice,
                            "sample_pairs": sample_pairs,
                            "row_count": len(mapping or {}),
                        }
                        if err:
                            flash(f"Could not parse {fn}: {err}", "error")
                        else:
                            _apply_pool_mapping(mb, mapping, fn)
                    except Exception as exc:  # noqa: BLE001
                        flash(f"Re-parse failed: {exc}", "error")
                else:
                    flash(
                        "Choose a pool-mapping file (.csv/.xlsx) to upload.",
                        "error",
                    )
            else:
                try:
                    _MONTHLY_BAL_DIR.mkdir(parents=True, exist_ok=True)
                    fn = secure_filename(f.filename or "pool_mapping.csv")
                    target = _MONTHLY_BAL_DIR / fn
                    f.save(target)
                    mapping, err, header_preview = _parse_pool_mapping_file(
                        target,
                        label_col_idx=label_col_idx,
                        pool_col_idx=pool_col_idx,
                        has_header=has_header,
                    )
                    sample_pairs = list((mapping or {}).items())[:8]
                    # Stash the file path + preview so the UI can render
                    # column-picker dropdowns for a re-parse.
                    mb["pool_mapping_upload"] = {
                        "filename": fn,
                        "saved_path": str(target),
                        "header_preview": header_preview,
                        "label_col": label_letter,
                        "pool_col": pool_letter,
                        "header_choice": header_choice,
                        "sample_pairs": sample_pairs,
                        "row_count": len(mapping or {}),
                    }
                    if err:
                        flash(f"Could not parse {fn}: {err}", "error")
                    else:
                        _apply_pool_mapping(mb, mapping, fn)
                except Exception as exc:  # noqa: BLE001
                    flash(f"Upload failed: {exc}", "error")
            _save_state(state)
            return redirect(url_for(back_endpoint))

        elif action == "set_hist_extract_target":
            he = _ensure_hist_extracts(state)
            tp = (request.form.get("target_period") or "").strip()
            try:
                months = int(request.form.get("history_months") or "84")
            except (TypeError, ValueError):
                months = 84
            months = max(1, min(months, 240))
            he["target_period"] = tp
            he["history_months"] = months
            _save_state(state)
            flash("Saved target reporting period.", "success")

        elif action == "upload_hist_extract":
            f = request.files.get("hist_extract_file")
            if f and f.filename:
                try:
                    saved_name, entry = _add_hist_extract_anchor(state, f)
                    msg = f"Added extract: {saved_name}"
                    if entry["detected_date"]:
                        msg += (
                            f" \u2014 detected period {entry['detected_date']} "
                            f"({entry['detected_confidence']})"
                        )
                    else:
                        msg += " \u2014 could not detect period; set it below."
                    flash(msg, "success")
                except Exception as exc:  # noqa: BLE001
                    flash(f"Upload failed: {exc}", "error")
            else:
                flash("Choose a loan-data extract file to upload.", "error")

        elif action == "remove_hist_extract":
            name = (request.form.get("anchor_name") or "").strip()
            if name and _remove_hist_extract_anchor(state, name):
                flash(f"Removed extract: {name}", "info")

        elif action == "remove_extract_and_ignore":
            # Remove the anchor AND add its signature to ignored_signatures
            # so every file sharing that layout is excluded going forward.
            name = (request.form.get("anchor_name") or "").strip()
            he = _ensure_hist_extracts(state)
            sig = ""
            for e in he.get("anchor_files") or []:
                if e.get("name") == name:
                    sig = (e.get("signature") or "").strip()
                    break
            removed = name and _remove_hist_extract_anchor(state, name)
            if removed:
                if sig and sig not in he.get("ignored_signatures", []):
                    he.setdefault("ignored_signatures", []).append(sig)
                    _save_state(state)
                flash(
                    f"Removed extract {name} and will ignore matching files.",
                    "info",
                )

        elif action == "ignore_layout":
            sig = (request.form.get("signature") or "").strip()
            he = _ensure_hist_extracts(state)
            if sig and sig not in he.get("ignored_signatures", []):
                he.setdefault("ignored_signatures", []).append(sig)
                _save_state(state)
                flash("Layout ignored \u2014 files with this header signature "
                      "will be skipped.", "info")

        elif action == "unignore_layout":
            sig = (request.form.get("signature") or "").strip()
            he = _ensure_hist_extracts(state)
            ignored = he.get("ignored_signatures") or []
            if sig and sig in ignored:
                he["ignored_signatures"] = [s for s in ignored if s != sig]
                _save_state(state)
                flash("Layout restored \u2014 files will appear as unmapped "
                      "again.", "info")

        elif action == "remove_profile":
            # Drop a column profile (and any anchor files attached to it)
            # so the matching files reappear as an unmapped layout in the
            # scan results, where they can be ignored or re-anchored.
            pid = (request.form.get("profile_id") or "").strip()
            he = _ensure_hist_extracts(state)
            profile = next(
                (p for p in he.get("profiles") or [] if p.get("id") == pid),
                None,
            )
            if not profile:
                flash("Unknown column profile.", "error")
            else:
                label = profile.get("label") or pid
                he["anchor_files"] = [
                    e for e in he.get("anchor_files") or []
                    if e.get("profile_id") != pid
                ]
                he["profiles"] = [
                    p for p in he.get("profiles") or [] if p.get("id") != pid
                ]
                _save_state(state)
                flash(
                    f"Removed profile {label}. Matching files are back in "
                    "the scan results as an unmapped layout \u2014 click "
                    "Ignore layout there if you want to skip them.",
                    "info",
                )

        elif action == "use_scan_file_as_anchor":
            # Pick a file from the scan results and register it as an anchor.
            # Validate the path is actually one the wizard scanned, to avoid
            # arbitrary file access via crafted form values.
            sample_path = (request.form.get("sample_path") or "").strip()
            he = _ensure_hist_extracts(state)
            scan = he.get("scan_results") or {}
            scanned_paths = {
                (f.get("path") or "") for f in (scan.get("files") or [])
            }
            if not sample_path or sample_path not in scanned_paths:
                flash("Could not locate that file in the scan results.", "error")
            else:
                src = Path(sample_path)
                if not src.exists() or not src.is_file():
                    flash(f"File no longer exists: {sample_path}", "error")
                else:
                    try:
                        saved_name, entry = _add_hist_extract_anchor_from_path(
                            state, src
                        )
                        flash(
                            f"Added {saved_name} as an anchor \u2014 map its "
                            "columns above to enable processing.",
                            "success",
                        )
                    except Exception as exc:  # noqa: BLE001
                        flash(f"Could not register anchor: {exc}", "error")

        elif action == "set_extract_date":
            name = (request.form.get("anchor_name") or "").strip()
            new_date = (request.form.get("override_date") or "").strip()
            he = _ensure_hist_extracts(state)
            for entry in he["anchor_files"]:
                if entry.get("name") == name:
                    entry["override_date"] = new_date
                    break
            _save_state(state)

        elif action == "set_hist_extract_folder":
            he = _ensure_hist_extracts(state)
            he["folder_path"] = (request.form.get("folder_path") or "").strip()
            _save_state(state)
            flash("Saved historical extracts folder path.", "success")

        elif action == "scan_hist_extract_folder":
            he = _ensure_hist_extracts(state)
            folder_str = (request.form.get("folder_path") or
                          he.get("folder_path") or "").strip()
            he["folder_path"] = folder_str
            if not folder_str:
                flash("Enter a folder path before scanning.", "error")
            elif not he.get("target_period"):
                flash(
                    "Set the target reporting period before scanning so "
                    "the wizard knows which months to look for.",
                    "error",
                )
            else:
                try:
                    folder = Path(folder_str)
                    result = extract_hist_service.scan_folder(
                        folder,
                        he.get("target_period", ""),
                        int(he.get("history_months") or 84),
                        known_profiles=he.get("profiles") or [],
                    )
                except Exception as exc:  # noqa: BLE001
                    result = {"ok": False, "error": str(exc)}
                he["scan_results"] = result
                if not result.get("ok"):
                    flash(
                        f"Scan failed: {result.get('error') or 'unknown error'}",
                        "error",
                    )
                else:
                    s = result.get("summary") or {}
                    msg = (
                        f"Scanned {s.get('scanned', 0)} file(s) &mdash; "
                        f"{s.get('covered_months', 0)} of "
                        f"{len(result.get('expected_months') or [])} months "
                        f"covered"
                    )
                    if s.get("missing_months"):
                        msg += f", {s['missing_months']} missing"
                    if s.get("multiple_months"):
                        msg += f", {s['multiple_months']} months have >1 file"
                    if s.get("new_signatures"):
                        msg += (
                            f", {s['new_signatures']} new column layout(s) "
                            "detected"
                        )
                    if s.get("unreadable"):
                        msg += f", {s['unreadable']} unreadable"
                    flash(msg, "success")
            _save_state(state)

        elif action == "clear_hist_extract_scan":
            he = _ensure_hist_extracts(state)
            he["scan_results"] = None
            _save_state(state)

        elif action == "process_hist_extracts":
            he = _ensure_hist_extracts(state)
            scan = he.get("scan_results") or {}
            if not scan.get("ok"):
                flash(
                    "Run a folder scan before processing.",
                    "error",
                )
            else:
                cu = (state.get("credit_union") or "").strip()
                try:
                    result = extract_hist_processor.process_scan(
                        cu,
                        scan,
                        he.get("profiles") or [],
                        anchor_files=he.get("anchor_files") or [],
                        ignored_signatures=he.get("ignored_signatures") or [],
                    )
                except Exception as exc:  # noqa: BLE001
                    result = {
                        "ok": False,
                        "error": str(exc),
                        "months_processed": [],
                        "months_skipped": [],
                    }
                scan["processed"] = result
                he["scan_results"] = scan
                _save_state(state)
                if not result.get("ok"):
                    flash(
                        f"Processing failed: "
                        f"{result.get('error') or 'unknown error'}",
                        "error",
                    )
                else:
                    done = len(result.get("months_processed") or [])
                    skipped = len(result.get("months_skipped") or [])
                    flash(
                        f"Processed {done} month(s), wrote "
                        f"{result.get('rows_written', 0)} row(s) "
                        f"covering {result.get('loan_codes_total', 0)} "
                        f"distinct loan code(s); {skipped} skipped.",
                        "success" if not skipped else "warning",
                    )

        elif action == "save_solr_settings":
            he = _ensure_hist_extracts(state)
            sb = he["solr_backfill"]
            sb["solr_url"] = (request.form.get("solr_url") or "").strip() or sb["solr_url"]
            sb["core"] = (request.form.get("solr_core") or "").strip() or sb["core"]
            mode_raw = (request.form.get("solr_backfill_mode") or "").strip().lower()
            if mode_raw in ("per_loan_code", "distributed"):
                sb["mode"] = mode_raw
            _save_state(state)
            flash("Saved 5300 Solr settings.", "success")

        elif action == "set_solr_backfill_mode":
            he = _ensure_hist_extracts(state)
            sb = he["solr_backfill"]
            mode_raw = (request.form.get("solr_backfill_mode") or "").strip().lower()
            if mode_raw in ("per_loan_code", "distributed"):
                sb["mode"] = mode_raw
                _save_state(state)
                label = ("Per-loan-code" if mode_raw == "per_loan_code"
                         else "Distributed (earliest-month ratios)")
                flash(f"5300 backfill mode set to: {label}.", "success")
            else:
                flash("Invalid backfill mode.", "error")

        elif action == "set_solr_backfill_enabled":
            he = _ensure_hist_extracts(state)
            sb = he["solr_backfill"]
            enabled = request.form.get("solr_backfill_enabled") == "on"
            sb["enabled"] = enabled
            # Re-arm the one-shot auto-run when re-enabled so toggling it back
            # on actually triggers a backfill on the next "next".
            if enabled:
                sb["auto_ran"] = False
            _save_state(state)
            flash(
                "Auto-fill balances from NCUA 5300: "
                + ("ON — will run when you continue past this step."
                   if enabled else "OFF."),
                "success",
            )

        elif action == "test_solr_fetch":
            he = _ensure_hist_extracts(state)
            sb = he["solr_backfill"]
            charter_raw = (state.get("charter_number") or "").strip()
            period = (he.get("target_period") or "").strip()
            if not charter_raw:
                flash("Set the CU's charter number on the Identity step first.",
                      "error")
            elif not period:
                flash("Set the target reporting period above first.", "error")
            else:
                # Test against the most-recent quarter-end <= target_period.
                try:
                    qends = solr_5300_backfill.expected_quarter_ends(period, 3)
                except Exception:  # noqa: BLE001
                    qends = []
                test_period = qends[0] if qends else period
                try:
                    charter_int = int(re.sub(r"\D", "", charter_raw))
                except ValueError:
                    charter_int = 0
                sb["last_test"] = solr_5300_backfill.test_connection(
                    sb["solr_url"], sb["core"], charter_int, test_period,
                )
                _save_state(state)
                if sb["last_test"].get("ok"):
                    flash(
                        f"Solr OK \u2014 fetched {sb['last_test']['field_count']}"
                        f" fields for charter {charter_int} at {test_period}.",
                        "success",
                    )
                else:
                    flash(
                        f"Solr test failed: {sb['last_test'].get('error')}",
                        "error",
                    )

        elif action == "run_solr_backfill":
            he = _ensure_hist_extracts(state)
            sb = he["solr_backfill"]
            cu = (state.get("credit_union") or "").strip()
            charter_raw = (state.get("charter_number") or "").strip()
            period = (he.get("target_period") or "").strip()
            months = int(he.get("history_months") or 84)
            try:
                charter_int = int(re.sub(r"\D", "", charter_raw))
            except ValueError:
                charter_int = 0
            if not cu:
                flash("Set the credit union on the Identity step first.", "error")
            elif not charter_int:
                flash("Set the CU's charter number on the Identity step first.",
                      "error")
            elif not period:
                flash("Set the target reporting period above first.", "error")
            else:
                # Skip months we already have data for (any source).
                existing: set[str] = set()
                try:
                    hv = extract_hist_processor.history_matrix(cu)
                    existing = set((hv or {}).get("months") or [])
                except Exception:  # noqa: BLE001
                    existing = set()
                db_months_before = set(existing)
                # Also skip months already provided by uploaded balance
                # sheet workbooks (annual or per-month). Without this,
                # the 5300 backfill would happily re-fill quarter-ends
                # that the user already covered via Excel uploads —
                # 87 months would be filled even when the annual files
                # already supply most of them.
                upload_months: set[str] = set()
                try:
                    mb_state = state.get("monthly_bal") or {}
                    source = state.get("hist_balance_source") or ""
                    label_to_pool = mb_state.get("pool_map") or {}
                    if source == "annual_balance_sheets":
                        ann = monthly_bal_parser.pool_balances_for_per_year_files(
                            mb_state.get("year_files") or [],
                            mb_state.get("per_year_layout") or {},
                            label_to_pool,
                            exclude_labels=mb_state.get("exclude_labels") or [],
                        )
                        upload_months = set(
                            (ann.get("by_period") or {}).keys())
                    elif source == "monthly_balance_sheets":
                        pm = monthly_bal_parser.pool_balances_for_per_month_files(
                            mb_state.get("monthly_files") or [],
                            mb_state.get("per_month_layout") or {},
                            label_to_pool,
                            exclude_labels=mb_state.get("exclude_labels") or [],
                        )
                        upload_months = set(
                            (pm.get("by_period") or {}).keys())
                    existing |= upload_months
                except Exception:  # noqa: BLE001
                    # Non-fatal — fall back to history_matrix only.
                    pass

                # Clean up: prior backfills may have written 5300-source
                # rows for months that the user has now covered via
                # uploaded balance-sheet workbooks. Those rows are
                # harmless at report time (uploads win in
                # _load_balance_history_from_db) but they pollute the
                # DB and confuse the "X months already filled" math.
                # Drop any 5300-source rows whose as_of_date matches an
                # upload month.
                cleanup_removed = 0
                try:
                    overlap = sorted(db_months_before & upload_months)
                    if overlap:
                        from sqlalchemy import text as _sql_text
                        eng = extract_hist_processor._engine_lazy()
                        with eng.begin() as conn:
                            res = conn.execute(
                                _sql_text(
                                    "DELETE FROM loan_code_history "
                                    "WHERE cu = :cu "
                                    "AND source LIKE '5300:%' "
                                    "AND as_of_date = ANY(:dates)"
                                ),
                                {"cu": cu, "dates": overlap},
                            )
                            cleanup_removed = int(res.rowcount or 0)
                except Exception:  # noqa: BLE001
                    cleanup_removed = 0

                mode = sb.get("mode") or "per_loan_code"
                added_to_pool_map: list[str] = []
                dist_info: dict = {}
                if mode == "distributed":
                    dist_info = _earliest_month_pool_distribution(state)
                    if not dist_info.get("ok"):
                        flash(
                            "5300 backfill (distributed): "
                            + (dist_info.get("error")
                               or "could not compute pool distribution"),
                            "error",
                        )
                        sb["last_run"] = {
                            "ok": False,
                            "mode": "distributed",
                            "error": dist_info.get("error") or "no distribution",
                        }
                        _save_state(state)
                        return redirect(url_for("setup.step3_historical"))
                    sb["last_run"] = (
                        solr_5300_backfill
                        .backfill_missing_quarters_distributed(
                            cu, charter_int, sb["solr_url"], sb["core"],
                            period, months,
                            pool_distribution=dist_info["pool_distribution"],
                            # Distributed mode wipes prior 5300:% rows on
                            # entry, so the only months we should treat
                            # as already-covered are the ones supplied
                            # by the user's uploaded balance file(s).
                            # Passing the full `existing` set (which is
                            # built from history_matrix BEFORE cleanup)
                            # would cause every quarter to be skipped
                            # as "already covered" after the cleanup
                            # deletes those very rows.
                            existing_dates=upload_months,
                            source_period_iso=dist_info.get("period") or "",
                        )
                    )
                else:
                    sb["last_run"] = (
                        solr_5300_backfill.backfill_missing_quarters(
                            cu, charter_int, sb["solr_url"], sb["core"],
                            period, months,
                            existing_dates=existing,
                        )
                    )
                    # Seed any new canonical loan codes into pool_map so
                    # they appear in the Loan Code Mapping step for
                    # pool assignment. (Distributed mode writes pool
                    # names directly, no seeding needed.)
                    pool_map = state.setdefault("pool_map", {})
                    for code in (sb["last_run"].get("new_loan_codes")
                                 or []):
                        if code not in pool_map:
                            pool_map[code] = ""
                            added_to_pool_map.append(code)
                lr = sb["last_run"]
                _save_state(state)
                if lr.get("ok"):
                    filled = len(lr.get("months_filled") or [])
                    quarters = len({m.get("quarter_end")
                                    for m in (lr.get("months_filled") or [])})
                    skipped = len(lr.get("months_skipped") or [])
                    none_yet = len(lr.get("months_no_data") or [])
                    stale = int(lr.get("stale_rows_removed") or 0)
                    mode_label = ("distributed (earliest-month ratios)"
                                  if mode == "distributed"
                                  else "per loan code")
                    msg = (
                        f"5300 backfill [{mode_label}]: filled {filled} "
                        f"month(s) across {quarters} quarter(s) "
                        f"({lr.get('rows_written', 0)} row(s)); "
                        f"{skipped} skipped, "
                        f"{none_yet} quarter(s) had no Solr doc."
                    )
                    if mode == "distributed" and dist_info:
                        pools_n = len(dist_info.get("pool_distribution") or {})
                        msg += (
                            f" Distribution from earliest month "
                            f"{dist_info.get('period')} "
                            f"(${dist_info.get('total', 0):,.0f} across "
                            f"{pools_n} pool(s))."
                        )
                    if filled == 0:
                        msg += (
                            f" (Already covered: {len(upload_months)} month(s) "
                            f"from uploaded workbooks, "
                            f"{len(db_months_before - upload_months)} month(s) "
                            f"from prior backfill/extracts.)"
                        )
                    if cleanup_removed:
                        msg += (
                            f" Removed {cleanup_removed} redundant 5300 row(s) "
                            f"for month(s) now covered by uploaded workbooks."
                        )
                    if stale:
                        msg += (
                            f" Removed {stale} stale row(s)."
                        )
                    if added_to_pool_map:
                        msg += (
                            f" Added {len(added_to_pool_map)} new loan code(s) "
                            f"to Loan Code Mapping — assign them to pools "
                            f"on the Pools step."
                        )
                    flash(msg, "success" if filled else "warning")
                else:
                    flash(
                        f"5300 backfill failed: {lr.get('error')}",
                        "error",
                    )

        elif action == "run_solr_co_backfill":
            he = _ensure_hist_extracts(state)
            sb = he["solr_backfill"]
            co_sb = he["co_solr_backfill"]
            cu = (state.get("credit_union") or "").strip()
            charter_raw = (state.get("charter_number") or "").strip()
            period = (he.get("target_period") or "").strip()
            months = int(he.get("history_months") or 84)
            try:
                charter_int = int(re.sub(r"\D", "", charter_raw))
            except ValueError:
                charter_int = 0
            if not cu:
                flash("Set the credit union on the Identity step first.", "error")
            elif not charter_int:
                flash("Set the CU's charter number on the Identity step first.",
                      "error")
            elif not period:
                flash("Set the target reporting period above first.", "error")
            else:
                # Skip months already populated in the CO history table.
                existing: set[str] = set()
                try:
                    hv = chargeoff_hist_processor.history_matrix(cu)
                    existing = set((hv or {}).get("months") or [])
                except Exception:  # noqa: BLE001
                    existing = set()
                co_sb["last_run"] = (
                    solr_5300_co_backfill
                    .backfill_missing_chargeoff_quarters(
                        cu, charter_int, sb["solr_url"], sb["core"],
                        period, months,
                        existing_dates=existing,
                    )
                )
                _save_state(state)
                lr = co_sb["last_run"]
                if lr.get("ok"):
                    filled = len(lr.get("months_filled") or [])
                    quarters = len({m.get("quarter_end")
                                    for m in (lr.get("months_filled") or [])})
                    skipped = len(lr.get("months_skipped") or [])
                    none_yet = len(lr.get("months_no_data") or [])
                    stale = int(lr.get("stale_rows_removed") or 0)
                    msg = (
                        f"5300 charge-off backfill: filled {filled} month(s) "
                        f"across {quarters} quarter(s) "
                        f"({lr.get('rows_written', 0)} row(s)); "
                        f"{skipped} skipped, {none_yet} quarter(s) had no "
                        f"Solr doc."
                    )
                    if stale:
                        msg += (
                            f" Removed {stale} stale row(s) for loan codes "
                            f"no longer in the canonical map."
                        )
                    flash(msg, "success" if filled else "warning")
                else:
                    flash(
                        f"5300 charge-off backfill failed: {lr.get('error')}",
                        "error",
                    )

        elif action == "run_solr_recov_backfill":
            he = _ensure_hist_extracts(state)
            sb = he["solr_backfill"]
            rec_sb = he["recov_solr_backfill"]
            cu = (state.get("credit_union") or "").strip()
            charter_raw = (state.get("charter_number") or "").strip()
            period = (he.get("target_period") or "").strip()
            months = int(he.get("history_months") or 84)
            try:
                charter_int = int(re.sub(r"\D", "", charter_raw))
            except ValueError:
                charter_int = 0
            if not cu:
                flash("Set the credit union on the Identity step first.", "error")
            elif not charter_int:
                flash("Set the CU's charter number on the Identity step first.",
                      "error")
            elif not period:
                flash("Set the target reporting period above first.", "error")
            else:
                existing: set[str] = set()
                try:
                    hv = recovery_hist_processor.history_matrix(cu)
                    existing = set((hv or {}).get("months") or [])
                except Exception:  # noqa: BLE001
                    existing = set()
                rec_sb["last_run"] = (
                    solr_5300_recov_backfill
                    .backfill_missing_recovery_quarters(
                        cu, charter_int, sb["solr_url"], sb["core"],
                        period, months,
                        existing_dates=existing,
                    )
                )
                _save_state(state)
                lr = rec_sb["last_run"]
                if lr.get("ok"):
                    filled = len(lr.get("months_filled") or [])
                    quarters = len({m.get("quarter_end")
                                    for m in (lr.get("months_filled") or [])})
                    skipped = len(lr.get("months_skipped") or [])
                    none_yet = len(lr.get("months_no_data") or [])
                    stale = int(lr.get("stale_rows_removed") or 0)
                    msg = (
                        f"5300 recovery backfill: filled {filled} month(s) "
                        f"across {quarters} quarter(s) "
                        f"({lr.get('rows_written', 0)} row(s)); "
                        f"{skipped} skipped, {none_yet} quarter(s) had no "
                        f"Solr doc."
                    )
                    if stale:
                        msg += (
                            f" Removed {stale} stale row(s) for loan codes "
                            f"no longer in the canonical map."
                        )
                    flash(msg, "success" if filled else "warning")
                else:
                    flash(
                        f"5300 recovery backfill failed: {lr.get('error')}",
                        "error",
                    )

        elif action == "save_extract_profile_mapping":
            he = _ensure_hist_extracts(state)
            pid = (request.form.get("profile_id") or "").strip()
            profile = next(
                (p for p in he["profiles"] if p.get("id") == pid), None
            )
            try:
                import tempfile, os
                _dbg = os.path.join(tempfile.gettempdir(), "cecl_save_mapping.log")
                with open(_dbg, "a", encoding="utf-8") as _f:
                    _f.write(f"--- pid={pid!r} profile_found={bool(profile)} "
                             f"form_keys={list(request.form.keys())}\n"
                             f"    form_values={ {k: request.form.get(k) for k in request.form.keys()} }\n")
            except Exception:
                pass
            if not profile:
                flash("Unknown column profile.", "error")
            else:
                import re as _re_hdr
                def _norm_hdr(s: str) -> str:
                    return _re_hdr.sub(r"\s+", " ", str(s or "")).strip()
                # Normalise stored headers in-place so older drafts captured
                # before _clean_header was hardened (e.g. AIRES exports with
                # multi-line header cells like "Current \nLoan Bal") still
                # match the values browsers send back from the form.
                raw_headers = profile.get("headers") or []
                headers = [_norm_hdr(h) for h in raw_headers]
                if headers != raw_headers:
                    profile["headers"] = headers
                hdr_set = set(headers)

                # Member / Account number format (same options as the
                # later "Column Mappings" step).
                ma_mode = (
                    request.form.get("member_account_mode")
                    or "fixed_suffix"
                ).strip()
                if ma_mode not in ("fixed_suffix", "delimiter", "split"):
                    ma_mode = "fixed_suffix"
                try:
                    ma_suffix_len = int(
                        request.form.get("member_account_suffix_length", "3")
                    )
                except (TypeError, ValueError):
                    ma_suffix_len = 3
                ma_suffix_len = max(0, min(ma_suffix_len, 9))
                ma_delim = (
                    request.form.get("member_account_delimiter") or "-"
                ).strip() or "-"
                member_account = {
                    "mode": ma_mode,
                    "suffix_length": ma_suffix_len,
                    "delimiter": ma_delim,
                }

                # Fields we collect for the historical-extract roll-up.
                required = [
                    "member_number",
                    "current_balance",
                    "loan_pool_code",
                ]
                if ma_mode == "split":
                    # Separate member & suffix columns -> suffix required.
                    required.append("loan_suffix")
                optional = ("loan_suffix", "original_fico_score")
                new_map: dict[str, str] = {}
                for fld in tuple(required) + optional:
                    val = _norm_hdr(request.form.get(fld) or "")
                    if val:
                        new_map[fld] = val
                missing = [f for f in required if not new_map.get(f)]
                unknown = [
                    f"{f} -> {v}" for f, v in new_map.items()
                    if v not in hdr_set
                ]
                try:
                    import tempfile, os
                    _dbg = os.path.join(tempfile.gettempdir(), "cecl_save_mapping.log")
                    with open(_dbg, "a", encoding="utf-8") as _f:
                        _f.write(f"    new_map={new_map}\n"
                                 f"    missing={missing} unknown={unknown}\n"
                                 f"    headers_sample={list(hdr_set)[:5]}\n")
                except Exception:
                    pass
                if missing:
                    flash(
                        f"Profile {profile.get('label') or pid}: missing "
                        "required column(s): " + ", ".join(missing),
                        "error",
                    )
                elif unknown:
                    flash(
                        f"Profile {profile.get('label') or pid}: these "
                        "mappings don't match a header in the file: "
                        + "; ".join(unknown),
                        "error",
                    )
                else:
                    profile["column_mappings"] = new_map
                    profile["member_account"] = member_account
                    # Seed the wizard-wide column_mappings from the first
                    # profile to get mapped, so the later "Column Mappings"
                    # step pre-populates instead of starting blank.
                    state.setdefault("column_mappings", {})
                    if not state["column_mappings"]:
                        state["column_mappings"] = dict(new_map)
                        state["member_account"] = dict(member_account)
                        state["account_suffix_length"] = (
                            ma_suffix_len
                            if ma_mode == "fixed_suffix"
                            else 0
                        )
                        flash(
                            f"Saved mapping for {profile.get('label') or pid} "
                            "and seeded the wizard's Column Mappings step.",
                            "success",
                        )
                    else:
                        flash(
                            f"Saved mapping for "
                            f"{profile.get('label') or pid}.",
                            "success",
                        )
                    _save_state(state)

        elif action == "upload_balance_history":
            f = request.files.get("balance_history_file")
            if f and f.filename:
                try:
                    saved_name = _add_hist_file(state, "monthly_files", f, "balance_history")
                    flash(f"Saved historical balances file: {saved_name}", "success")
                    # Parse col-A labels and seed the loan-type -> pool mapping.
                    try:
                        saved_path = Path(
                            (state.get("hist_scan") or {}).get("monthly_files", [])[-1]["path"]
                        )
                        analysis = hist_parser.extract_balance_labels(saved_path)
                        if analysis.get("ok") and analysis.get("labels"):
                            _seed_hist_pool_map(state, analysis)
                            flash(
                                f"Detected {len(analysis['labels'])} loan-type "
                                f"labels in column A \u2014 review the pool "
                                f"mapping below.",
                                "info",
                            )
                        elif analysis.get("error"):
                            flash(
                                f"Could not read loan-type labels: "
                                f"{analysis['error']}",
                                "error",
                            )
                    except Exception as exc:  # noqa: BLE001
                        flash(f"Could not analyse balances file: {exc}", "error")
                except Exception as exc:  # noqa: BLE001
                    flash(f"Upload failed: {exc}", "error")
            else:
                flash("Choose a historical balances file to upload.", "error")

        elif action == "remove_balance_history":
            target_name = (request.form.get("filename") or "").strip()
            hs = state.get("hist_scan") or {}
            files = list(hs.get("monthly_files") or [])
            removed = None
            for e in files:
                if e.get("name") == target_name:
                    removed = e
                    break
            if removed:
                files.remove(removed)
                sp = removed.get("path") or ""
                if sp:
                    try:
                        pp = Path(sp)
                        # Only delete files we manage in the wizard's
                        # temp dir — never touch user-provided source
                        # paths.
                        if pp.is_file() and "cecl_ui_hist" in str(pp).lower():
                            pp.unlink()
                    except Exception:  # noqa: BLE001
                        pass
                hs["monthly_files"] = files
                state["hist_scan"] = hs
                # If a remaining file is still present, re-seed the
                # loan-type -> pool mapping from the most recent one
                # so the section 1b dropdowns stay in sync. Otherwise
                # clear the cached mapping entirely.
                if files:
                    try:
                        last_path = Path(files[-1].get("path") or "")
                        if last_path.is_file():
                            analysis = hist_parser.extract_balance_labels(
                                last_path
                            )
                            if analysis.get("ok") and analysis.get("labels"):
                                _seed_hist_pool_map(state, analysis)
                    except Exception:  # noqa: BLE001
                        pass
                else:
                    state["hist_pool_map"] = None
                flash(
                    f"Removed {target_name} and refreshed the loan-type "
                    f"\u2192 pool mapping.",
                    "success",
                )
                _save_state(state)
            else:
                flash(f"File '{target_name}' not found.", "error")
            return redirect(url_for(back_endpoint))

        elif action == "remove_all_balance_history":
            hs = state.get("hist_scan") or {}
            files = list(hs.get("monthly_files") or [])
            removed_count = 0
            for e in files:
                sp = e.get("path") or ""
                if sp:
                    try:
                        pp = Path(sp)
                        if pp.is_file() and "cecl_ui_hist" in str(pp).lower():
                            pp.unlink()
                    except Exception:  # noqa: BLE001
                        pass
                removed_count += 1
            hs["monthly_files"] = []
            state["hist_scan"] = hs
            state["hist_pool_map"] = None
            if removed_count:
                flash(
                    f"Removed all {removed_count} historical balance "
                    f"workbook(s) and cleared the loan-type \u2192 pool "
                    f"mapping.",
                    "success",
                )
            else:
                flash(
                    "No historical balance workbooks to remove.", "info"
                )
            _save_state(state)
            return redirect(url_for(back_endpoint))

        elif action == "save_hist_pool_map":
            _save_hist_pool_map_from_form(state, request.form)
            flash("Saved loan-type \u2192 pool mapping.", "success")

        elif action == "upload_co":
            f = request.files.get("co_file")
            if f and f.filename:
                try:
                    saved_name = _add_hist_file(state, "co_files", f, "co")
                    flash(f"Saved historical charge-off file: {saved_name}", "success")
                    # Inspect the workbook so the column-mapping UI can
                    # appear if the auto-detector misses any required
                    # column. Non-fatal if inspection fails — the user
                    # can still try Aggregate, which surfaces its own
                    # per-file error.
                    try:
                        _refresh_co_recov_inspect(state, "co")
                    except Exception:  # noqa: BLE001
                        pass
                except Exception as exc:  # noqa: BLE001
                    flash(f"Upload failed: {exc}", "error")
            else:
                flash("Choose a historical charge-off file to upload.", "error")

        elif action == "upload_recov":
            f = request.files.get("recov_file")
            if f and f.filename:
                try:
                    saved_name = _add_hist_file(state, "recov_files", f, "recov")
                    flash(f"Saved historical recoveries file: {saved_name}", "success")
                    try:
                        _refresh_co_recov_inspect(state, "recov")
                    except Exception:  # noqa: BLE001
                        pass
                except Exception as exc:  # noqa: BLE001
                    flash(f"Upload failed: {exc}", "error")
            else:
                flash("Choose a historical recoveries file to upload.", "error")

        elif action == "set_hist_co_source":
            choice = (request.form.get("hist_co_source") or "").strip()
            if choice in ("single_workbook", "monthly_files", "5300_backfill"):
                state["hist_co_source"] = choice

        elif action == "set_hist_recov_source":
            choice = (request.form.get("hist_recov_source") or "").strip()
            if choice in ("single_workbook", "monthly_files", "5300_backfill"):
                state["hist_recov_source"] = choice

        elif action == "upload_co_monthly":
            files = request.files.getlist("co_monthly_files")
            saved_n, errs = 0, []
            for f in files:
                if not (f and f.filename):
                    continue
                try:
                    _add_hist_file(state, "monthly_co_files", f, "co")
                    saved_n += 1
                except Exception as exc:  # noqa: BLE001
                    errs.append(f"{f.filename}: {exc}")
            if saved_n:
                flash(f"Saved {saved_n} monthly charge-off file(s).", "success")
                _refresh_co_recov_inspect(state, "co")
            if errs:
                flash("Some uploads failed: " + "; ".join(errs), "error")
            if not saved_n and not errs:
                flash("Choose one or more monthly charge-off files to upload.", "error")

        elif action == "upload_recov_monthly":
            files = request.files.getlist("recov_monthly_files")
            saved_n, errs = 0, []
            for f in files:
                if not (f and f.filename):
                    continue
                try:
                    _add_hist_file(state, "monthly_recov_files", f, "recov")
                    saved_n += 1
                except Exception as exc:  # noqa: BLE001
                    errs.append(f"{f.filename}: {exc}")
            if saved_n:
                flash(f"Saved {saved_n} monthly recoveries file(s).", "success")
                _refresh_co_recov_inspect(state, "recov")
            if errs:
                flash("Some uploads failed: " + "; ".join(errs), "error")
            if not saved_n and not errs:
                flash("Choose one or more monthly recoveries files to upload.", "error")

        elif action == "remove_co_monthly":
            name = (request.form.get("filename") or "").strip()
            hs = state.get("hist_scan") or {}
            hs["monthly_co_files"] = [
                e for e in (hs.get("monthly_co_files") or []) if e.get("name") != name
            ]
            state["hist_scan"] = hs
            if name:
                flash(f"Removed {name}.", "success")

        elif action == "remove_recov_monthly":
            name = (request.form.get("filename") or "").strip()
            hs = state.get("hist_scan") or {}
            hs["monthly_recov_files"] = [
                e for e in (hs.get("monthly_recov_files") or []) if e.get("name") != name
            ]
            state["hist_scan"] = hs
            if name:
                flash(f"Removed {name}.", "success")

        elif action in ("clear_all_co_monthly", "clear_all_recov_monthly"):
            kind = "co" if action == "clear_all_co_monthly" else "recov"
            list_key = (
                "monthly_co_files" if kind == "co" else "monthly_recov_files"
            )
            kind_label = "charge-off" if kind == "co" else "recoveries"
            hs = state.get("hist_scan") or {}
            prev_n = len(hs.get(list_key) or [])
            hs[list_key] = []
            state["hist_scan"] = hs
            # Also drop any cached scan-results table so the page redraws clean.
            state.pop(_co_recov_scan_state_key(kind, "monthly"), None)
            if prev_n:
                flash(
                    f"Cleared {prev_n} monthly {kind_label} file(s).",
                    "success",
                )
            else:
                flash(
                    f"No monthly {kind_label} files to clear.",
                    "info",
                )

        elif action == "process_co_monthly":
            res = monthly_co_recov_aggregator.aggregate_all(state, "co")
            state["monthly_co_aggregate"] = res
            if res.get("ok"):
                msg = (
                    f"Charge-off aggregation: wrote "
                    f"{res['total_rows_written']} row(s) across "
                    f"{len(res['months_written'])} month(s) "
                    f"(${res['total_amount']:,.2f} total)."
                )
                flash(msg, "success")
            else:
                flash(
                    "Charge-off aggregation failed: "
                    + (res.get("error") or "see file errors below."),
                    "error",
                )

        elif action == "process_co_workbook":
            res = monthly_co_recov_aggregator.aggregate_workbook(state, "co")
            state["monthly_co_aggregate"] = res
            if res.get("ok"):
                msg = (
                    f"Charge-off workbook aggregation: wrote "
                    f"{res['total_rows_written']} row(s) across "
                    f"{len(res['months_written'])} month(s) "
                    f"(${res['total_amount']:,.2f} total)."
                )
                flash(msg, "success")
            else:
                flash(
                    "Charge-off workbook aggregation failed: "
                    + (res.get("error") or "see file errors below."),
                    "error",
                )

        elif action == "process_recov_monthly":
            res = monthly_co_recov_aggregator.aggregate_all(state, "recov")
            state["monthly_recov_aggregate"] = res
            if res.get("ok"):
                msg = (
                    f"Recoveries aggregation: wrote "
                    f"{res['total_rows_written']} row(s) across "
                    f"{len(res['months_written'])} month(s) "
                    f"(${res['total_amount']:,.2f} total)."
                )
                flash(msg, "success")
            else:
                flash(
                    "Recoveries aggregation failed: "
                    + (res.get("error") or "see file errors below."),
                    "error",
                )

        elif action == "process_recov_workbook":
            res = monthly_co_recov_aggregator.aggregate_workbook(state, "recov")
            state["monthly_recov_aggregate"] = res
            if res.get("ok"):
                msg = (
                    f"Recoveries workbook aggregation: wrote "
                    f"{res['total_rows_written']} row(s) across "
                    f"{len(res['months_written'])} month(s) "
                    f"(${res['total_amount']:,.2f} total)."
                )
                flash(msg, "success")
            else:
                flash(
                    "Recoveries workbook aggregation failed: "
                    + (res.get("error") or "see file errors below."),
                    "error",
                )

        elif action == "save_co_columns":
            cols = _read_co_recov_column_form("co")
            if "sheet" not in cols and (state.get("co_columns") or {}).get("sheet"):
                cols["sheet"] = state["co_columns"]["sheet"]
            state["co_columns"] = cols
            sig = (state.get("co_active_signature") or "").strip()
            if sig:
                layout_map = dict(state.get("co_layout_columns") or {})
                layout_map[sig] = cols
                state["co_layout_columns"] = layout_map
                flash(
                    f"Saved charge-off column mapping for layout {sig[:8]}\u2026.",
                    "success",
                )
            else:
                flash("Saved charge-off column mapping.", "success")

        elif action == "save_recov_columns":
            cols = _read_co_recov_column_form("recov")
            if "sheet" not in cols and (state.get("recov_columns") or {}).get("sheet"):
                cols["sheet"] = state["recov_columns"]["sheet"]
            state["recov_columns"] = cols
            sig = (state.get("recov_active_signature") or "").strip()
            if sig:
                layout_map = dict(state.get("recov_layout_columns") or {})
                layout_map[sig] = cols
                state["recov_layout_columns"] = layout_map
                flash(
                    f"Saved recoveries column mapping for layout {sig[:8]}\u2026.",
                    "success",
                )
            else:
                flash("Saved recoveries column mapping.", "success")

        elif action in (
            "copy_co_mapping_to_recov", "copy_recov_mapping_from_co",
        ):
            # Convenience action for CUs whose Charge-Off and Recovery
            # data live in the SAME file (e.g. Destinations CU's
            # "Recoveries-Chg-offs" workbooks have both a Charge-off
            # Amount column and a Recovery Amount column in every row).
            # Copies the saved charge-off column mapping into the
            # recoveries slot, swapping the Amount column to the
            # recoveries-side suggestion and (when available) swapping
            # the Date column to a recoveries-specific date header.
            src = state.get("co_columns") or {}
            if not (src.get("loan_code") or src.get("amount")):
                flash(
                    "Save the charge-off column mapping first, then click "
                    "this button to copy it to recoveries.",
                    "error",
                )
            else:
                recov_ins = state.get("recov_inspect") or {}
                suggested = recov_ins.get("suggested") or {}
                recov_headers = recov_ins.get("headers") or []
                existing_recov = state.get("recov_columns") or {}

                # New amount: prefer the recov-side suggestion from the
                # inspector (so the SAME file's "Recovery Amount" header
                # wins over the copied "Charge-off Amount"); fall back
                # to whatever recov already had; finally fall back to
                # the CO amount (single-amount-column files).
                new_amount = (
                    suggested.get("recov_amount")
                    or existing_recov.get("amount")
                    or src.get("amount")
                    or ""
                )

                # New date: scan recov headers for a recovery-flavoured
                # date column (e.g. "Recovery Date"); fall back to
                # whatever the CO mapping used (single-date files).
                new_date = src.get("date") or ""
                for h in recov_headers:
                    hs = str(h).lower()
                    if "recov" in hs and "date" in hs:
                        new_date = h
                        break

                # Build the recov mapping as a deep copy of CO with the
                # two overrides. member_account block is copied as-is
                # (same file => same account-key shape).
                ma = dict(src.get("member_account") or {})
                new_cols = {
                    "loan_code":     src.get("loan_code") or "",
                    "amount":        new_amount,
                    "date":          new_date,
                    "member_number": src.get("member_number") or "",
                    "loan_suffix":   src.get("loan_suffix") or "",
                    "member_account": ma or {
                        "mode": "fixed_suffix",
                        "suffix_length": 3,
                        "delimiter": "-",
                    },
                }
                state["recov_columns"] = new_cols

                # Mirror into the per-layout map so the staleness
                # round-trip preserves the copied mapping.  Prefer the
                # recov side's active signature; if it's not set,
                # adopt the CO side's signature so the copied mapping
                # tracks with the same file layout.
                rsig = (state.get("recov_active_signature") or "").strip()
                if not rsig:
                    rsig = (state.get("co_active_signature") or "").strip()
                    if rsig:
                        state["recov_active_signature"] = rsig
                if rsig:
                    layout_map = dict(state.get("recov_layout_columns") or {})
                    layout_map[rsig] = new_cols
                    state["recov_layout_columns"] = layout_map

                _save_state(state)
                flash(
                    "Copied charge-off column mapping into recoveries"
                    + (f" (amount column swapped to '{new_amount}'"
                       + (f", date column swapped to '{new_date}'"
                          if new_date != src.get("date") else "")
                       + ")."
                       if new_amount != src.get("amount") else "."),
                    "success",
                )

        elif action in (
            "map_unmapped_co_codes", "map_unmapped_recov_codes",
        ):
            # Inline mapping of unmapped charge-off / recovery loan codes
            # straight into the global state['pool_map'] (the same store
            # Step 3 Loan Code Mapping edits). Form posts a list of
            # ``unmapped_code`` values plus indexed ``unmapped_pool_<i>``
            # selects; we zip them and merge non-empty selections.
            form = request.form
            codes = form.getlist("unmapped_code")
            pool_map = dict(state.get("pool_map") or {})
            saved = 0
            for i, raw in enumerate(codes):
                raw = (raw or "").strip()
                if not raw:
                    continue
                pool = (form.get(f"unmapped_pool_{i}") or "").strip()
                if not pool:
                    continue
                pool_map[raw] = pool
                saved += 1
            state["pool_map"] = pool_map
            _save_state(state)
            kind_label = (
                "charge-off" if action == "map_unmapped_co_codes"
                else "recoveries"
            )
            if saved:
                flash(
                    f"Mapped {saved} {kind_label} loan code(s) into the "
                    "global Step 3 Loan Code Mapping.",
                    "success",
                )
            else:
                flash(
                    "No mappings selected. Pick a pool for at least one "
                    "code before clicking Save.",
                    "info",
                )

        elif action in (
            "save_co_loan_code_map", "save_recov_loan_code_map",
        ):
            kind = "co" if action == "save_co_loan_code_map" else "recov"
            sig_key = (
                "co_active_signature" if kind == "co"
                else "recov_active_signature"
            )
            map_key = (
                "co_loan_code_map" if kind == "co"
                else "recov_loan_code_map"
            )
            sig = (state.get(sig_key) or "").strip()
            if not sig:
                flash(
                    "Click 'Column Mapping' on a layout row first so "
                    "the wizard knows which layout to save the loan-code "
                    "mapping under.",
                    "error",
                )
            else:
                form = request.form
                # Form posts pairs of fields named
                # ``<kind>_lcm_<i>_raw`` and ``<kind>_lcm_<i>_pool``;
                # collect them by index.
                prefix = "co_lcm_" if kind == "co" else "recov_lcm_"
                indices: set[str] = set()
                for fname in form.keys():
                    if fname.startswith(prefix) and (
                        fname.endswith("_raw") or fname.endswith("_pool")
                    ):
                        idx = fname[len(prefix):].rsplit("_", 1)[0]
                        indices.add(idx)
                pairs: dict[str, str] = {}
                for idx in indices:
                    raw = (form.get(f"{prefix}{idx}_raw") or "").strip()
                    pool = (form.get(f"{prefix}{idx}_pool") or "").strip()
                    if raw and pool:
                        pairs[raw] = pool
                full_map = dict(state.get(map_key) or {})
                full_map[sig] = pairs
                state[map_key] = full_map
                kind_label = (
                    "charge-off" if kind == "co" else "recoveries"
                )
                flash(
                    f"Saved {len(pairs)} loan-code mapping(s) for the "
                    f"{kind_label} layout {sig[:8]}\u2026. They will be "
                    "applied the next time you aggregate.",
                    "success",
                )

        elif action == "reinspect_co_monthly":
            _refresh_co_recov_inspect(state, "co", force=True)
            flash("Re-inspected the first charge-off file.", "success")

        elif action == "reinspect_recov_monthly":
            _refresh_co_recov_inspect(state, "recov", force=True)
            flash("Re-inspected the first recoveries file.", "success")

        elif action in ("set_co_sheet", "set_recov_sheet"):
            _kind = "co" if action == "set_co_sheet" else "recov"
            _cols_key = "co_columns" if _kind == "co" else "recov_columns"
            chosen = (request.form.get(f"{_kind}_sheet") or "").strip()
            cols = dict(state.get(_cols_key) or {})
            cols["sheet"] = chosen
            state[_cols_key] = cols
            # Re-inspect against the chosen tab so the header pickers and
            # suggestions reflect that tab (force re-seeds the columns since
            # a different tab usually has a different layout).
            _refresh_co_recov_inspect(state, _kind, force=True)
            label = "charge-off" if _kind == "co" else "recoveries"
            if chosen:
                flash(
                    f"Reading the {label} workbook from the "
                    f"\u201c{chosen}\u201d tab.",
                    "success",
                )
            else:
                flash(
                    f"Auto-detecting the {label} tab (densest sheet).",
                    "success",
                )

        elif action == "scan_co_monthly_folder":
            folder_str = (request.form.get("co_monthly_folder") or "").strip()
            hs = state.get("hist_scan") or {}
            hs["monthly_co_folder"] = folder_str
            state["hist_scan"] = hs
            res = _scan_monthly_co_recov_folder(state, folder_str, "co")
            _flash_co_recov_scan_result(res, "co")
            if res.get("added"):
                _refresh_co_recov_inspect(state, "co")

        elif action == "scan_recov_monthly_folder":
            folder_str = (request.form.get("recov_monthly_folder") or "").strip()
            hs = state.get("hist_scan") or {}
            hs["monthly_recov_folder"] = folder_str
            state["hist_scan"] = hs
            res = _scan_monthly_co_recov_folder(state, folder_str, "recov")
            _flash_co_recov_scan_result(res, "recov")
            if res.get("added"):
                _refresh_co_recov_inspect(state, "recov")

        elif action in ("ignore_co_layout", "ignore_recov_layout"):
            kind = "co" if action == "ignore_co_layout" else "recov"
            # Accept one OR many signatures (the wizard's bulk-ignore
            # form posts multiple "signature" inputs).
            raw_sigs = request.form.getlist("signature")
            new_sigs = [s.strip() for s in raw_sigs if s and s.strip()]
            if not new_sigs:
                one = (request.form.get("signature") or "").strip()
                if one:
                    new_sigs = [one]
            sig_key = f"{kind}_ignored_signatures"
            sigs = list(state.get(sig_key) or [])
            added = 0
            removed_total = 0
            for sig in new_sigs:
                if sig not in sigs:
                    sigs.append(sig)
                    added += 1
                removed_total += _remove_files_with_signature(
                    state, kind, sig, is_sample=False,
                )
            state[sig_key] = sigs
            _refresh_layouts_in_scan_data(state, kind, "monthly")
            _refresh_layouts_in_scan_data(state, kind, "sample")
            kind_label = "charge-off" if kind == "co" else "recoveries"
            if not new_sigs:
                flash(
                    f"Pick at least one {kind_label} layout to ignore.",
                    "error",
                )
            else:
                count = len(new_sigs)
                noun = "layout" if count == 1 else "layouts"
                msg = (
                    f"Ignored {count} {kind_label} {noun} \u2014 future "
                    f"scans will skip matching files."
                )
                if removed_total:
                    msg += f" Removed {removed_total} already-added file(s)."
                flash(msg, "success")

        elif action in ("unignore_co_layout", "unignore_recov_layout"):
            kind = "co" if action == "unignore_co_layout" else "recov"
            sig = (request.form.get("signature") or "").strip()
            sig_key = f"{kind}_ignored_signatures"
            sigs = [s for s in (state.get(sig_key) or []) if s != sig]
            state[sig_key] = sigs
            _refresh_layouts_in_scan_data(state, kind, "monthly")
            _refresh_layouts_in_scan_data(state, kind, "sample")
            flash("Layout un-ignored. Re-run the folder scan to pick up matching files.", "success")

        elif action in ("clear_co_scan_monthly", "clear_recov_scan_monthly"):
            kind = "co" if action == "clear_co_scan_monthly" else "recov"
            state.pop(_co_recov_scan_state_key(kind, "monthly"), None)
            flash("Scan results cleared.", "success")

        elif action in (
            "use_co_scan_file_as_columns",
            "use_recov_scan_file_as_columns",
        ):
            kind = "co" if action == "use_co_scan_file_as_columns" else "recov"
            list_key = (
                "monthly_co_files" if kind == "co" else "monthly_recov_files"
            )
            kind_label = "charge-off" if kind == "co" else "recoveries"
            sample_path = (request.form.get("sample_path") or "").strip()
            hs = state.get("hist_scan") or {}
            files = list(hs.get(list_key) or [])
            # Find the chosen file in the already-added list and promote
            # it to index 0 so ``_refresh_co_recov_inspect`` (which seeds
            # off ``files[0]``) reads its columns.
            target_idx = None
            for i, entry in enumerate(files):
                if (entry.get("path") or "") == sample_path:
                    target_idx = i
                    break
            if not sample_path or target_idx is None:
                flash(
                    f"Could not locate that {kind_label} file in the scan "
                    "results.",
                    "error",
                )
            else:
                files.insert(0, files.pop(target_idx))
                hs[list_key] = files
                state["hist_scan"] = hs
                # Stash the file's signature so save_co_columns /
                # save_recov_columns / save_*_loan_code_map know which
                # layout slot to write to.  Restore any previously-saved
                # per-layout column mapping so the form shows that
                # layout's last-saved values rather than auto-detected.
                sig_key = (
                    "co_active_signature" if kind == "co"
                    else "recov_active_signature"
                )
                cols_key = "co_columns" if kind == "co" else "recov_columns"
                layout_cols_key = (
                    "co_layout_columns" if kind == "co"
                    else "recov_layout_columns"
                )
                sig, _ = _file_layout_signature(Path(sample_path))
                state[sig_key] = sig or ""
                saved_layout = (
                    (state.get(layout_cols_key) or {}).get(sig)
                    if sig else None
                )
                force_seed = True
                if saved_layout:
                    state[cols_key] = dict(saved_layout)
                    force_seed = False
                _refresh_co_recov_inspect(state, kind, force=force_seed)
                flash(
                    f"Loaded {Path(sample_path).name} into the "
                    f"{kind_label} Column Mapping card below.",
                    "success",
                )

        elif action == "remove_co_workbook":
            name = (request.form.get("filename") or "").strip()
            hs = state.get("hist_scan") or {}
            hs["co_files"] = [
                e for e in (hs.get("co_files") or []) if e.get("name") != name
            ]
            state["hist_scan"] = hs
            if name:
                flash(f"Removed {name}.", "success")

        elif action == "remove_recov_workbook":
            name = (request.form.get("filename") or "").strip()
            hs = state.get("hist_scan") or {}
            hs["recov_files"] = [
                e for e in (hs.get("recov_files") or []) if e.get("name") != name
            ]
            state["hist_scan"] = hs
            if name:
                flash(f"Removed {name}.", "success")

        elif action in ("next", "skip"):
            # Leaving the Historical Balances step: auto-fill earlier quarters
            # from the NCUA 5300 by default (opt-out on that step). No-op for
            # the CO / recovery sub-steps and when already run / disabled.
            if section == "balances":
                _auto_run_5300_balance_backfill(state)
            _save_state(state)
            return redirect(url_for(next_endpoint))

        _save_state(state)

    # Re-attach profile_id / profile_label on any stored scan results so
    # that anchors uploaded *after* the last scan are reflected without
    # forcing the user to click Scan Folder again.  Signatures don't
    # change, so this is just a lookup against the current profiles list.
    # Also rebuild ``new_signatures`` (unmapped layouts) every render with
    # accurate per-layout counts so the UI can show the user exactly which
    # column shapes still need an anchor + mapping.
    he_view = state.get("hist_extracts") or {}
    scan_view = he_view.get("scan_results") or {}
    ignored_sigs_set = set(he_view.get("ignored_signatures") or [])
    if scan_view.get("ok") and scan_view.get("files"):
        sig_to_prof = {
            (p.get("signature") or ""): p
            for p in (he_view.get("profiles") or [])
        }
        changed = False
        new_sigs: dict[str, dict[str, Any]] = {}
        ignored_view: dict[str, dict[str, Any]] = {}
        for entry in scan_view["files"]:
            sig = entry.get("signature") or ""
            prof = sig_to_prof.get(sig) if sig else None
            if prof:
                if entry.get("profile_id") != prof.get("id"):
                    entry["profile_id"] = prof.get("id")
                    entry["profile_label"] = prof.get("label")
                    changed = True
                continue
            # Unmapped (or sig=='' = unreadable headers — skip those here).
            if entry.get("profile_id"):
                entry["profile_id"] = None
                entry["profile_label"] = None
                changed = True
            if not sig:
                continue
            if sig in ignored_sigs_set:
                slot = ignored_view.setdefault(sig, {
                    "signature": sig,
                    "sample_file": entry.get("name", ""),
                    "file_count": 0,
                    "in_window_count": 0,
                })
                slot["file_count"] += 1
                if entry.get("in_window"):
                    slot["in_window_count"] += 1
                continue
            slot = new_sigs.setdefault(sig, {
                "signature": sig,
                "headers": [],
                "sample_file": entry.get("name", ""),
                "sample_path": entry.get("path", ""),
                "file_count": 0,
                "in_window_count": 0,
            })
            slot["file_count"] += 1
            if entry.get("in_window"):
                slot["in_window_count"] += 1
        # Sort biggest-first so the most valuable layout to map is on top.
        sorted_sigs = sorted(
            new_sigs.values(),
            key=lambda s: (s.get("in_window_count", 0), s.get("file_count", 0)),
            reverse=True,
        )
        sorted_ignored = sorted(
            ignored_view.values(),
            key=lambda s: (s.get("in_window_count", 0), s.get("file_count", 0)),
            reverse=True,
        )
        prior = scan_view.get("new_signatures") or []
        prior_ign = scan_view.get("ignored_layouts") or []
        if (changed
                or len(prior) != len(sorted_sigs)
                or len(prior_ign) != len(sorted_ignored)):
            scan_view["new_signatures"] = sorted_sigs
            scan_view["ignored_layouts"] = sorted_ignored
            summary = scan_view.setdefault("summary", {})
            summary["new_signatures"] = len(sorted_sigs)
            summary["ignored_layouts"] = len(sorted_ignored)
            _save_state(state)
        else:
            # Always refresh in-memory copy (counts may differ even when
            # profile_id attachments didn't change).
            scan_view["new_signatures"] = sorted_sigs
            scan_view["ignored_layouts"] = sorted_ignored
            scan_view.setdefault("summary", {})["new_signatures"] = len(sorted_sigs)
            scan_view.setdefault("summary", {})["ignored_layouts"] = len(sorted_ignored)

    # Pull the saved history matrix for this CU so the user can see
    # what's actually in the loan_code_history table.  Only fetch when
    # the user is on the monthly-extracts path (the other paths don't
    # populate this table).
    history_view = None
    cu = (state.get("credit_union") or "").strip()
    if state.get("hist_balance_source") == "monthly_loan_extracts":
        if cu:
            try:
                history_view = extract_hist_processor.history_matrix(cu)
            except Exception as exc:  # noqa: BLE001
                history_view = {"error": str(exc), "row_count": 0,
                                "months": [], "codes": [], "cells": {}}
    # Always pull CO/Recov history matrices for the "what's loaded"
    # tables in sections 2 and 3 (regardless of source choice).
    co_history_view = None
    recov_history_view = None
    if cu:
        try:
            co_history_view = chargeoff_hist_processor.history_matrix(cu)
        except Exception as exc:  # noqa: BLE001
            co_history_view = {"error": str(exc), "row_count": 0,
                               "months": [], "codes": [], "cells": {}}
        try:
            recov_history_view = recovery_hist_processor.history_matrix(cu)
        except Exception as exc:  # noqa: BLE001
            recov_history_view = {"error": str(exc), "row_count": 0,
                                  "months": [], "codes": [], "cells": {}}

    # Auto-refresh CO/Recov column-inspect snapshots if they're stale
    # (i.e. the cached `<kind>_inspect.filename` no longer matches the
    # first uploaded file).  This catches the case where the user
    # uploaded files A, B, C, then later replaced or scanned in a
    # different set — without this guard the dropdowns keep auto-
    # populating from the OLD inspection until the user clicks
    # "Re-inspect" manually.
    for _kind in ("co", "recov"):
        _monthly_key = "monthly_co_files" if _kind == "co" else "monthly_recov_files"
        _workbook_key = "co_files" if _kind == "co" else "recov_files"
        _ins_key = "co_inspect" if _kind == "co" else "recov_inspect"
        _hs = state.get("hist_scan") or {}
        _files = _hs.get(_monthly_key) or _hs.get(_workbook_key) or []
        _ins = state.get(_ins_key) or {}
        if not _files:
            if _ins:
                state.pop(_ins_key, None)
                _save_state(state)
            continue
        _first_name = (_files[0].get("name") or "").strip()
        _ins_name = (_ins.get("filename") or "").strip()
        # Detect cached inspects from before the whitespace-normalisation
        # fix: any header still containing ``\n`` / ``\t`` / runs of
        # spaces means the saved column mapping (which the browser
        # collapsed) can never round-trip, so force a fresh inspect.
        _stale_ws = any(
            isinstance(h, str) and ("\n" in h or "\t" in h
                                    or "  " in h)
            for h in (_ins.get("headers") or [])
        )
        # Detect cached inspects that errored out (e.g. the legacy
        # ".xls openpyxl-not-supported" error from before xlrd support
        # was added). If the cache has an error AND empty headers but
        # the file still exists on disk, force a fresh inspect so
        # post-fix code paths take over without requiring the user to
        # click "Re-inspect" manually.
        _stale_error = bool(
            _ins.get("error") and not (_ins.get("headers") or [])
        )
        if _first_name and (_first_name != _ins_name or _stale_ws
                            or _stale_error):
            # Preserve any saved per-layout mapping for the file now at
            # files[0]: only force a fresh seed when no saved mapping
            # exists for that layout.
            _sig_key = (
                "co_active_signature" if _kind == "co"
                else "recov_active_signature"
            )
            _layout_cols_key = (
                "co_layout_columns" if _kind == "co"
                else "recov_layout_columns"
            )
            _cols_key = "co_columns" if _kind == "co" else "recov_columns"
            try:
                _path = Path(_files[0].get("path") or "")
                _sig, _ = _file_layout_signature(_path) if _path else ("", [])
                _saved = (state.get(_layout_cols_key) or {}).get(_sig)
                _existing = state.get(_cols_key) or {}
                if _sig:
                    state[_sig_key] = _sig
                _force = True
                if _saved:
                    state[_cols_key] = dict(_saved)
                    _force = False
                elif _existing.get("loan_code") or _existing.get("amount") \
                        or _existing.get("date"):
                    # User saved a column mapping but no per-layout
                    # signature was active at save time (typical when
                    # there's only one uploaded file and the user never
                    # clicked the explicit layout picker). Treat the
                    # top-level mapping as authoritative for the current
                    # file too, so the staleness re-inspect doesn't
                    # clobber their picks with auto-suggestions on the
                    # very next page render.
                    _force = False
                _refresh_co_recov_inspect(state, _kind, force=_force)
                # Mirror any backfilled column values into the per-
                # layout map so the staleness round-trip persists.
                if _sig:
                    _layout_map = dict(state.get(_layout_cols_key) or {})
                    _layout_map[_sig] = dict(state.get(_cols_key) or {})
                    state[_layout_cols_key] = _layout_map
                    _save_state(state)
            except Exception:  # noqa: BLE001
                # Non-fatal — fall back to whatever inspect is cached.
                pass
    view_mode = (request.args.get("matrix_view") or "balance").strip()
    if view_mode not in ("balance", "count"):
        view_mode = "balance"

    # Pools the user established on Step 2 — used to constrain the
    # loan-type -> pool dropdown (Section 1b) so the user can only pick
    # from real pools or « ignore » (same UX as the WARM Balance
    # Titles step).
    step2_pools = [
        (p.get("name") or "").strip()
        for p in (state.get("pool_settings") or [])
        if (p.get("name") or "").strip()
        and not p.get("excluded")
    ]

    # Live preview of the earliest-month pool distribution that the
    # "distributed" 5300 backfill mode would use. Cheap (<100ms) so we
    # always compute it; template only renders the panel when mode is
    # 'distributed'.
    try:
        solr_dist_preview = _earliest_month_pool_distribution(state)
    except Exception:  # noqa: BLE001
        solr_dist_preview = {"ok": False, "error": "preview failed"}

    # Full historical-balance extraction preview -- the user-visible
    # confirmation that running the auto-scan (or manually picking the
    # source on Step 4) actually pulled month-end balances out of the
    # uploaded workbook(s). Always computed so the panel always shows
    # something (an error message in the worst case).
    try:
        extracted_bal_preview = _extracted_balance_preview(state)
    except Exception as exc:  # noqa: BLE001
        extracted_bal_preview = {
            "ok": False,
            "error": f"preview unavailable: {exc}",
            "source": "", "source_files": [], "period_count": 0,
            "earliest_period": "", "latest_period": "",
            "latest_period_with_data": "",
            "pool_count": 0, "recent_periods": [], "top_pools": [],
            "totals_by_period": {}, "balances_grid": {},
            "backfill": {"present": False, "ok": False,
                         "months_filled": 0, "rows_written": 0,
                         "error": "", "earliest_filled": "",
                         "latest_filled": ""},
        }

    # Distinct loan codes from the active CO/Recov sample file so the
    # template can render a per-layout loan-code -> pool mapping table.
    # ``co_distinct_codes`` / ``recov_distinct_codes`` are lists of raw
    # strings; the saved mapping is a dict in
    # ``state.{kind}_loan_code_map[active_signature]``.
    def _distinct_for(_kind: str) -> list[str]:
        ins_key = "co_inspect" if _kind == "co" else "recov_inspect"
        cols_key = "co_columns" if _kind == "co" else "recov_columns"
        ins = state.get(ins_key) or {}
        fname = (ins.get("filename") or "").strip()
        if not fname:
            return []
        # Locate the inspected file's path on disk.
        list_key = (
            "monthly_co_files" if _kind == "co" else "monthly_recov_files"
        )
        candidates: list[str] = []
        for entry in (state.get("hist_scan") or {}).get(list_key) or []:
            if (entry.get("name") or "") == fname:
                p = (entry.get("path") or "").strip()
                if p:
                    candidates.append(p)
        single_list_key = (
            "co_files" if _kind == "co" else "recov_files"
        )
        for entry in (state.get("hist_scan") or {}).get(single_list_key) or []:
            if (entry.get("name") or "") == fname:
                p = (entry.get("path") or "").strip()
                if p:
                    candidates.append(p)
        for path_str in candidates:
            p = Path(path_str)
            if not p.exists():
                continue
            code_header = (
                (state.get(cols_key) or {}).get("loan_code") or ""
            ).strip() or None
            saved_sheet = (state.get(cols_key) or {}).get("sheet") or None
            try:
                return monthly_co_recov_aggregator.extract_distinct_loan_codes(
                    p, code_header, sheet=saved_sheet
                )
            except Exception:  # noqa: BLE001
                return []
        return []

    co_distinct_codes = _distinct_for("co")
    recov_distinct_codes = _distinct_for("recov")

    # Also surface every distinct loan_code already present in the DB
    # history tables (typically NCUA canonical names like "Used Vehicles"
    # written by the 5300 backfills, plus any user-uploaded codes).
    # Merge into the per-section distinct list so the inline "map
    # unmapped codes" UI lets the user assign 5300 codes to pools too,
    # even when no raw CO/Recov file is loaded.
    def _merge_db_codes(existing: list[str], view: dict[str, Any] | None) -> list[str]:
        if not view:
            return existing
        codes = list(view.get("codes") or [])
        if not codes:
            return existing
        seen = {c.strip().lower() for c in existing if isinstance(c, str)}
        out = list(existing)
        for c in codes:
            if not isinstance(c, str):
                continue
            key = c.strip().lower()
            if not key or key in seen:
                continue
            seen.add(key)
            out.append(c)
        return out

    co_distinct_codes = _merge_db_codes(co_distinct_codes, co_history_view)
    recov_distinct_codes = _merge_db_codes(
        recov_distinct_codes, recov_history_view
    )

    # Group every uploaded CO/Recov file by header signature so the
    # template can render a layout picker.  The picker lets the user
    # switch the Column Mapping card between layouts even when the
    # files were uploaded individually (no folder scan).
    def _layouts_for(_kind: str) -> list[dict[str, Any]]:
        list_key = (
            "monthly_co_files" if _kind == "co" else "monthly_recov_files"
        )
        sig_key = (
            "co_active_signature" if _kind == "co"
            else "recov_active_signature"
        )
        layout_cols_key = (
            "co_layout_columns" if _kind == "co"
            else "recov_layout_columns"
        )
        loan_map_key = (
            "co_loan_code_map" if _kind == "co"
            else "recov_loan_code_map"
        )
        rows = (state.get("hist_scan") or {}).get(list_key) or []
        active_sig = (state.get(sig_key) or "").strip()
        saved_cols = state.get(layout_cols_key) or {}
        saved_lcm = state.get(loan_map_key) or {}
        groups: dict[str, dict[str, Any]] = {}
        for entry in rows:
            p_str = (entry.get("path") or "").strip()
            if not p_str:
                continue
            p = Path(p_str)
            if not p.exists():
                continue
            sig, _h = _file_layout_signature(p)
            if not sig:
                continue
            g = groups.setdefault(sig, {
                "signature": sig,
                "sample_path": p_str,
                "sample_file": p.name,
                "file_count": 0,
                "is_active": (sig == active_sig),
                "has_column_mapping": sig in saved_cols,
                "has_loan_code_map": bool(saved_lcm.get(sig)),
                "loan_code_map_count": len(saved_lcm.get(sig) or {}),
            })
            g["file_count"] += 1
        return sorted(
            groups.values(), key=lambda g: g["sample_file"].lower()
        )

    co_layouts = _layouts_for("co")
    recov_layouts = _layouts_for("recov")

    # Roll the CO/Recov "what's loaded" tables up by mapped pool name so
    # they reflect the user's Step 3 Loan Code Mapping (5300 NCUA names
    # collapsed into pools). The raw matrix is still used above for the
    # unmapped-codes picker.
    pool_map_for_agg = state.get("pool_map") or {}
    co_history_view_pool = _aggregate_history_by_pool(
        co_history_view, pool_map_for_agg
    )
    recov_history_view_pool = _aggregate_history_by_pool(
        recov_history_view, pool_map_for_agg
    )

    # Worksheet/tab names for the single-workbook CO & Recov uploads, so the
    # column-mapping card can offer a tab picker when charge-offs and
    # recoveries live on separate tabs of one workbook.
    def _sheet_names_for(list_key: str) -> list[str]:
        files = (state.get("hist_scan") or {}).get(list_key) or []
        if not files:
            return []
        p = Path((files[0].get("path") or "").strip())
        if not p.exists():
            return []
        return monthly_co_recov_aggregator.list_sheet_names(p)

    co_sheet_names = _sheet_names_for("co_files")
    recov_sheet_names = _sheet_names_for("recov_files")

    return render_template(
        "setup/step3_historical.html",
        pool_suggestions=_DEFAULT_POOL_SUGGESTIONS,
        step2_pools=step2_pools,
        history_view=history_view,
        co_history_view=co_history_view_pool or co_history_view,
        recov_history_view=recov_history_view_pool or recov_history_view,
        matrix_view=view_mode,
        co_sheet_names=co_sheet_names,
        recov_sheet_names=recov_sheet_names,
        solr_canonical_map=solr_5300_backfill.load_canonical_map(),
        co_canonical_map=solr_5300_co_backfill.load_canonical_map(),
        recov_canonical_map=solr_5300_recov_backfill.load_canonical_map(),
        solr_dist_preview=solr_dist_preview,
        extracted_bal_preview=extracted_bal_preview,
        co_distinct_codes=co_distinct_codes,
        recov_distinct_codes=recov_distinct_codes,
        co_layouts=co_layouts,
        recov_layouts=recov_layouts,
        section=section,
        hist_balance_provenance=_hist_balance_provenance_view(state),
        co_provenance=_co_recov_provenance_view(state, "co"),
        recov_provenance=_co_recov_provenance_view(state, "recov"),
        **_wizard_ctx(active_key),
    )


def _add_sample_upload(state: dict[str, Any], key: str, file_storage, subfolder: str) -> str:
    """Save a sample-step file and record it in state['sample_uploads']."""
    dest = _SAMPLE_DIR / subfolder
    dest.mkdir(parents=True, exist_ok=True)
    fn = secure_filename(file_storage.filename or "upload.xlsx")
    target = dest / fn
    file_storage.save(target)
    entry = {"name": fn, "path": str(target)}
    uploads = state.setdefault("sample_uploads", {
        "loan_balance_files": [], "co_files": [], "recov_files": [],
        "impaired_files": [], "credit_pull_files": [],
        "no_recoveries": False, "no_credit_pull": False,
    })
    rows = [e for e in (uploads.get(key) or []) if e.get("name") != fn]
    rows.append(entry)
    uploads[key] = rows
    _save_state(state)
    return fn


def _add_sample_upload_from_path(
    state: dict[str, Any], key: str, src_path: Path, subfolder: str
) -> str:
    """Like ``_add_sample_upload`` but ingests an existing server-side file
    (e.g. one found by a folder scan).  Copies into ``_SAMPLE_DIR/subfolder``
    and registers in ``state.sample_uploads[key]``.
    """
    import shutil

    dest = _SAMPLE_DIR / subfolder
    dest.mkdir(parents=True, exist_ok=True)
    fn = secure_filename(src_path.name) or "upload.xlsx"
    target = dest / fn
    if target.resolve() != src_path.resolve():
        shutil.copy2(src_path, target)
    entry = {"name": fn, "path": str(target)}
    uploads = state.setdefault("sample_uploads", {
        "loan_balance_files": [], "co_files": [], "recov_files": [],
        "impaired_files": [], "credit_pull_files": [],
        "no_recoveries": False, "no_credit_pull": False,
    })
    rows = [e for e in (uploads.get(key) or []) if e.get("name") != fn]
    rows.append(entry)
    uploads[key] = rows
    _save_state(state)
    return fn


def _scan_sample_co_recov_folder(
    state: dict[str, Any], folder_str: str, kind: str
) -> dict[str, Any]:
    """Glob ``folder_str`` for sample CO/Recov files and ingest each.

    ``kind`` is ``"co"`` or ``"recov"``.  Returns a small summary dict
    suitable for flashing in the UI.
    """
    list_key = "co_files" if kind == "co" else "recov_files"
    subfolder = "sample_co" if kind == "co" else "sample_recov"
    out: dict[str, Any] = {
        "ok": False, "error": None,
        "added": [], "skipped": [], "ignored": [], "scanned": 0,
        "folder": folder_str,
    }
    folder_str = _normalize_folder_path(folder_str)
    if not folder_str:
        out["error"] = "Enter a folder path before scanning."
        return out
    folder = Path(folder_str)
    if not folder.exists():
        out["error"] = f"Folder not found: {folder}"
        return out
    if not folder.is_dir():
        out["error"] = f"Path is not a folder: {folder}"
        return out

    files = _discover_co_recov_files(folder)
    out["scanned"] = len(files)
    if not files:
        try:
            entries = sorted(p.name for p in folder.iterdir())[:8]
        except Exception:  # noqa: BLE001
            entries = []
        sample = (", ".join(entries)) if entries else "(empty)"
        out["error"] = (
            f"No .xlsx/.xlsm/.xls/.csv files found in {folder} (recursive). "
            f"First items in folder: {sample}"
        )
        return out

    ignored_sigs = _co_recov_ignored_set(state, kind)
    scanned_meta: list[dict[str, Any]] = []
    for p in files:
        sig, headers = _file_layout_signature(p)
        scanned_meta.append({
            "name": p.name, "path": str(p),
            "signature": sig, "headers": headers,
        })
        if sig and sig in ignored_sigs:
            out["ignored"].append(p.name)
            continue
        try:
            _add_sample_upload_from_path(state, list_key, p, subfolder)
            out["added"].append(p.name)
        except Exception as exc:  # noqa: BLE001
            out["skipped"].append(f"{p.name}: {exc}")
    _persist_co_recov_scan_data(state, kind, "sample", out, scanned_meta)
    out["ok"] = bool(out["added"]) or bool(out["ignored"])
    return out


@setup_bp.route("/step/sample", methods=["GET", "POST"])
def step2_sample():
    """Loan Data Extract(s) step.

    The user uploads one of the CU's quarterly loan-data extracts (e.g. an
    Aries file). We parse it to seed column mappings and pool codes for the
    File Format / Column Mapping / Loan Code Mapping steps. Charge-offs,
    recoveries, impaired loans, and credit-pull files now live in their own
    dedicated steps.
    """
    state = _state()
    if request.method == "POST":
        action = request.form.get("action", "")

        if action == "upload":
            # Loan data file — also runs sample analysis for column detection
            f = request.files.get("sample_file")
            uploads = state.setdefault("sample_uploads", {
                "loan_balance_files": [], "loan_data_files": [],
                "co_files": [], "recov_files": [], "impaired_files": [],
                "credit_pull_files": [], "no_recoveries": False, "no_credit_pull": False,
                "loan_data_error": "",
            })
            if not f or not f.filename:
                uploads["loan_data_error"] = "Pick a loan data file first."
                _save_state(state)
                flash("Pick a loan data file first.", "error")
                return redirect(url_for("setup.step2_sample"))
            else:
                try:
                    saved = _save_sample_upload(f)
                    analysis = sample_parser.analyse_sample_file(
                        saved, original_filename=f.filename
                    )
                    if not analysis.get("ok"):
                        msg = f"Could not parse: {analysis.get('error')}"
                        uploads["loan_data_error"] = msg
                        _save_state(state)
                        flash(msg, "error")
                        return redirect(url_for("setup.step2_sample"))
                    else:
                        _apply_sample_to_state(state, analysis)
                        # Record in sample_uploads as well — append so a CU
                        # with multiple loan-data extracts can stack them.
                        # Each entry carries its own has_header / header_row
                        # so the user can override per-file in the UI.
                        uploads["loan_data_error"] = ""
                        existing = [
                            e for e in (uploads.get("loan_data_files") or [])
                            if e.get("name") != analysis["filename"]
                        ]
                        entry = {
                            "name": analysis["filename"],
                            "path": str(saved),
                            "has_header": bool(analysis.get("has_header")),
                            "header_row": int(analysis.get("header_row") or 1),
                            "analysis": _loan_data_entry_analysis(analysis),
                        }
                        matched_label = _seed_loan_data_entry_mapping(
                            state, entry, analysis
                        )
                        existing.append(entry)
                        uploads["loan_data_files"] = existing
                        _save_state(state)
                        if matched_label:
                            flash(
                                f"Parsed {f.filename!s}: matched Step 2 historical "
                                f"profile {matched_label!r} — column mapping copied "
                                "automatically. Review on the Column Mappings step.",
                                "success",
                            )
                        else:
                            flash(
                                f"Parsed {f.filename!s}: detected "
                                f"{len(analysis['headers'])} columns, "
                                f"{len(analysis['pool_code_suggestions'])} distinct "
                                "pool codes. Suggestions applied to later steps.",
                                "success",
                            )
                        return redirect(url_for("setup.step2_sample"))
                except Exception as exc:  # noqa: BLE001
                    uploads["loan_data_error"] = f"Upload failed: {exc}"
                    _save_state(state)
                    flash(f"Upload failed: {exc}", "error")
                    return redirect(url_for("setup.step2_sample"))

        elif action == "remove_loan_data":
            target = (request.form.get("filename") or "").strip()
            uploads = state.setdefault("sample_uploads", {})
            uploads["loan_data_files"] = [
                e for e in (uploads.get("loan_data_files") or [])
                if e.get("name") != target
            ]
            _save_state(state)
            flash(f"Removed loan data file: {target}", "info")
            return redirect(url_for("setup.step2_sample"))

        elif action == "set_file_header":
            # Per-file header override: re-parse the chosen loan-data file
            # with the user's has_header / header_row settings. Updates the
            # entry in loan_data_files and refreshes state.sample so the
            # column-mapping suggestions reflect the new layout.
            target = (request.form.get("filename") or "").strip()
            uploads = state.setdefault("sample_uploads", {})
            files = uploads.get("loan_data_files") or []
            entry = next((e for e in files if e.get("name") == target), None)
            if not entry:
                flash(f"File not found in upload list: {target}", "error")
                return redirect(url_for("setup.step2_sample"))
            saved = entry.get("path")
            if not saved or not Path(saved).exists():
                flash(
                    f"Original file is no longer on disk \u2014 please "
                    f"re-upload {target} before changing header settings.",
                    "error",
                )
                return redirect(url_for("setup.step2_sample"))
            if request.form.get("has_header") == "on":
                try:
                    hr = int(request.form.get("header_row", "1"))
                except ValueError:
                    hr = 1
                hr = max(1, hr)
            else:
                hr = 0  # explicit "no header"
            try:
                analysis = sample_parser.analyse_sample_file(
                    saved, original_filename=target, header_row=hr,
                )
                if not analysis.get("ok"):
                    flash(f"Could not re-parse {target}: {analysis.get('error')}",
                          "error")
                else:
                    entry["has_header"] = bool(analysis.get("has_header"))
                    entry["header_row"] = int(analysis.get("header_row") or 0)
                    entry["analysis"] = _loan_data_entry_analysis(analysis)
                    # Refresh signature; if the entry has no per-file
                    # column mapping yet (e.g. set_file_header is the
                    # user's first action after upload), seed it now.
                    had_mapping = bool(entry.get("column_mappings"))
                    matched_label = _seed_loan_data_entry_mapping(
                        state, entry, analysis
                    )
                    _apply_sample_to_state(state, analysis)
                    _save_state(state)
                    if hr == 0:
                        flash(f"Re-parsed {target} with no header row.", "success")
                    else:
                        flash(
                            f"Re-parsed {target} using row {hr} as the header.",
                            "success",
                        )
                    if matched_label and not had_mapping:
                        flash(
                            f"Matched Step 2 historical profile "
                            f"{matched_label!r} — column mapping copied "
                            f"automatically for {target}.",
                            "success",
                        )
            except Exception as exc:  # noqa: BLE001
                flash(f"Re-parse failed for {target}: {exc}", "error")
            return redirect(url_for("setup.step2_sample"))

        elif action == "clear":
            defaults = _default_state()
            state["sample"] = None
            state["file_pattern"] = defaults["file_pattern"]
            state["date_pattern"] = defaults["date_pattern"]
            state["column_mappings"] = dict(defaults["column_mappings"])
            state["pool_map"] = dict(defaults["pool_map"])
            _save_state(state)
            flash("Sample analysis cleared. Defaults restored.", "info")

        elif action == "set_header":
            # Re-parse the previously uploaded sample with a user-chosen
            # header row (or "no header"). Requires that we still have the
            # saved file on disk from the original upload.
            sample = state.get("sample") or {}
            saved = sample.get("saved_path")
            orig_name = sample.get("filename", "sample")
            if not saved or not Path(saved).exists():
                flash(
                    "Original sample file is no longer on disk \u2014 please "
                    "re-upload it before setting the header row.",
                    "error",
                )
            else:
                if request.form.get("has_header") == "on":
                    try:
                        hr = int(request.form.get("header_row", "1"))
                    except ValueError:
                        hr = 1
                    hr = max(1, hr)
                else:
                    hr = 0  # explicit "no header"
                try:
                    analysis = sample_parser.analyse_sample_file(
                        saved, original_filename=orig_name, header_row=hr,
                    )
                    if not analysis.get("ok"):
                        flash(f"Could not re-parse: {analysis.get('error')}",
                              "error")
                    else:
                        _apply_sample_to_state(state, analysis)
                        _save_state(state)
                        if hr == 0:
                            flash("Re-parsed with no header row.", "success")
                        else:
                            flash(
                                f"Re-parsed using row {hr} as the header.",
                                "success",
                            )
                except Exception as exc:  # noqa: BLE001
                    flash(f"Re-parse failed: {exc}", "error")

        elif action in ("next", "skip"):
            _save_state(state)
            return redirect(url_for("setup.step3_columns"))

    return render_template("setup/step2_sample.html", **_wizard_ctx("sample"))


# =================================================================
# Charge-Offs & Recoveries
# =================================================================
@setup_bp.route("/step/co-recov", methods=["GET", "POST"])
def step_co_recov():
    """Combined Charge-Offs and Recoveries upload + column-mapping step."""
    state = _state()

    def _read_int(form_key: str, default: int | None = None) -> int | None:
        raw = (request.form.get(form_key) or "").strip()
        if raw == "":
            return default
        try:
            return int(raw)
        except (TypeError, ValueError):
            return default

    if request.method == "POST":
        action = request.form.get("action", "")

        if action == "upload_sample_co":
            f = request.files.get("sample_co_file")
            if f and f.filename:
                try:
                    saved_name = _add_sample_upload(state, "co_files", f, "sample_co")
                    flash(f"Saved charge-off tracking file: {saved_name}", "success")
                    # Seed column-mapping suggestions from the first upload
                    if not (state.get("co_columns") or {}).get("code_col") in (None, ""):
                        pass
                    else:
                        files = (state.get("sample_uploads") or {}).get("co_files") or []
                        if files:
                            sug = co_recov_parser._suggest_columns(files[-1]["path"])
                            existing = state.get("co_columns") or {}
                            existing.update({k: v for k, v in sug.items()
                                             if k not in existing or existing.get(k) in (None, "")})
                            state["co_columns"] = existing
                            _save_state(state)
                    return redirect(url_for("setup.step_co_recov"))
                except Exception as exc:  # noqa: BLE001
                    flash(f"Upload failed: {exc}", "error")
            else:
                flash("Choose a charge-off tracking file to upload.", "error")

        elif action == "upload_sample_recov":
            f = request.files.get("sample_recov_file")
            if f and f.filename:
                try:
                    saved_name = _add_sample_upload(state, "recov_files", f, "sample_recov")
                    flash(f"Saved recoveries file: {saved_name}", "success")
                    if not (state.get("recov_columns") or {}).get("code_col") in (None, ""):
                        pass
                    else:
                        files = (state.get("sample_uploads") or {}).get("recov_files") or []
                        if files:
                            sug = co_recov_parser._suggest_columns(files[-1]["path"])
                            existing = state.get("recov_columns") or {}
                            existing.update({k: v for k, v in sug.items()
                                             if k not in existing or existing.get(k) in (None, "")})
                            state["recov_columns"] = existing
                            _save_state(state)
                    return redirect(url_for("setup.step_co_recov"))
                except Exception as exc:  # noqa: BLE001
                    flash(f"Upload failed: {exc}", "error")
            else:
                flash("Choose a recoveries file to upload.", "error")

        elif action == "use_co_files_for_recov":
            # Many CUs publish a single workbook that holds both charge-off
            # rows AND recovery rows (e.g. Destinations CU's
            # "Recoveries-Chg-offs.xls" files). Mirror the Charge-Off file
            # list into the Recoveries list so the user doesn't have to
            # re-upload or re-scan.
            import shutil
            from pathlib import Path as _P
            uploads = state.setdefault("sample_uploads", {})
            co_entries = list(uploads.get("co_files") or [])
            if not co_entries:
                flash(
                    "No Charge-Off files to copy. Upload or scan some first.",
                    "error",
                )
            else:
                existing_recov = list(uploads.get("recov_files") or [])
                existing_names = {e.get("name") for e in existing_recov}
                dest_dir = _SAMPLE_DIR / "sample_recov"
                dest_dir.mkdir(parents=True, exist_ok=True)
                copied = 0
                skipped = 0
                for entry in co_entries:
                    name = entry.get("name") or ""
                    src = entry.get("path") or ""
                    if not name:
                        continue
                    if name in existing_names:
                        skipped += 1
                        continue
                    target = dest_dir / name
                    try:
                        src_p = _P(src)
                        if src_p.exists() and target.resolve() != src_p.resolve():
                            shutil.copy2(src_p, target)
                    except Exception:  # noqa: BLE001
                        # If the staged CO copy is missing, fall back to
                        # registering the same path so the entry still has
                        # something to point at. The reader will surface a
                        # clear "missing file" error if it's truly gone.
                        target = _P(src) if src else (dest_dir / name)
                    existing_recov.append({"name": name, "path": str(target)})
                    existing_names.add(name)
                    copied += 1
                uploads["recov_files"] = existing_recov
                # Clear the "no recoveries" flag if the user previously
                # checked it — they now have files to process.
                if copied:
                    uploads["no_recoveries"] = False
                # Seed Recoveries column-mapping suggestions if still empty.
                rc = state.get("recov_columns") or {}
                if not (rc.get("code_col") or rc.get("amount_col") or rc.get("date_col")):
                    try:
                        sug = co_recov_parser._suggest_columns(
                            existing_recov[-1]["path"]
                        )
                        existing = state.get("recov_columns") or {}
                        existing.update({
                            k: v for k, v in sug.items()
                            if k not in existing or existing.get(k) in (None, "")
                        })
                        state["recov_columns"] = existing
                    except Exception:  # noqa: BLE001
                        pass
                if copied and skipped:
                    flash(
                        f"Copied {copied} file(s) from Charge-Offs "
                        f"({skipped} already in the Recoveries list).",
                        "success",
                    )
                elif copied:
                    flash(
                        f"Copied {copied} file(s) from Charge-Offs to Recoveries.",
                        "success",
                    )
                else:
                    flash(
                        "All Charge-Off files were already in the Recoveries list.",
                        "info",
                    )
                _save_state(state)
            return redirect(url_for("setup.step_co_recov"))

        elif action in ("set_sample_co_sheet", "set_sample_recov_sheet"):
            kind = "co" if action == "set_sample_co_sheet" else "recov"
            cfg_key = "co_columns" if kind == "co" else "recov_columns"
            list_key = "co_files" if kind == "co" else "recov_files"
            chosen = (request.form.get(f"{kind}_sheet") or "").strip()
            cfg = dict(state.get(cfg_key) or {})
            cfg["sheet"] = chosen
            # A different tab usually has a different column layout, so drop
            # the old index picks and re-seed suggestions from the new tab.
            for _k in ("member_col", "account_col", "code_col",
                       "amount_col", "date_col", "orig_score_col",
                       "curr_score_col"):
                cfg.pop(_k, None)
            files = (state.get("sample_uploads") or {}).get(list_key) or []
            if files:
                try:
                    sug = co_recov_parser._suggest_columns(
                        files[-1]["path"], sheet=(chosen or None)
                    )
                    for _k, _v in sug.items():
                        cfg[_k] = _v
                except Exception:  # noqa: BLE001
                    pass
            state[cfg_key] = cfg
            _save_state(state)
            label = "charge-off" if kind == "co" else "recoveries"
            if chosen:
                flash(
                    f"Reading the {label} file from the \u201c{chosen}\u201d tab.",
                    "success",
                )
            else:
                flash(
                    f"Reading the {label} file from all tabs (auto).",
                    "success",
                )
            return redirect(url_for("setup.step_co_recov"))

        elif action in ("save_co_columns", "save_recov_columns"):
            kind = "co" if action == "save_co_columns" else "recov"
            cfg_key = "co_columns" if kind == "co" else "recov_columns"

            ma_mode = (request.form.get(f"{kind}_member_account_mode") or "split").strip()
            if ma_mode not in ("fixed_suffix", "delimiter", "split"):
                ma_mode = "split"
            try:
                ma_suffix = int((request.form.get(f"{kind}_member_account_suffix_length") or "3").strip())
            except ValueError:
                ma_suffix = 3
            ma_suffix = max(0, min(9, ma_suffix))
            ma_delim = (request.form.get(f"{kind}_member_account_delimiter") or "-").strip() or "-"

            # Map the mode-specific column pickers back to the canonical
            # member_col / account_col slots so downstream consumers keep
            # working without changes. In fixed_suffix / delimiter modes
            # the single combined-cell dropdown drives member_col and
            # account_col is left None.
            if ma_mode == "split":
                member_col = _read_int(f"{kind}_member_col", None)
                account_col = _read_int(f"{kind}_account_col", None)
            else:
                member_col = _read_int(f"{kind}_combined_col", None)
                account_col = None

            cfg = {
                "has_header": request.form.get(f"{kind}_has_header") == "on",
                "skip_rows": _read_int(f"{kind}_skip_rows", 0) or 0,
                "member_col": member_col,
                "account_col": account_col,
                "code_col": _read_int(f"{kind}_code_col", None),
                "amount_col": _read_int(f"{kind}_amount_col", None),
                "date_col": _read_int(f"{kind}_date_col", None),
                "member_account": {
                    "mode": ma_mode,
                    "suffix_length": ma_suffix,
                    "delimiter": ma_delim,
                },
            }
            # Optional credit-score columns (charge-off only) feed the CO
            # credit-grade migration split; left unset when the file has none.
            if kind == "co":
                cfg["orig_score_col"] = _read_int(f"{kind}_orig_score_col", None)
                cfg["curr_score_col"] = _read_int(f"{kind}_curr_score_col", None)
            # Preserve the chosen worksheet/tab across a mapping save.
            _prev_sheet = (state.get(cfg_key) or {}).get("sheet")
            if _prev_sheet:
                cfg["sheet"] = _prev_sheet
            if cfg["code_col"] is None or cfg["amount_col"] is None:
                flash("Code and Amount columns are required.", "error")
            else:
                state[cfg_key] = cfg
                _save_state(state)
                flash(
                    ("Charge-off" if kind == "co" else "Recoveries")
                    + " column mapping saved.",
                    "success",
                )
            return redirect(url_for("setup.step_co_recov"))

        elif action == "assign_codes_bulk":
            codes = request.form.getlist("map_code")
            pools = request.form.getlist("map_pool")
            if len(codes) != len(pools):
                flash(
                    "Bulk save: mismatched code/pool counts; nothing saved.",
                    "error",
                )
            else:
                pm = state.setdefault("pool_map", {})
                mapped_n = 0
                unmapped_n = 0
                for raw_code, raw_pool in zip(codes, pools):
                    code = (raw_code or "").strip()
                    new_pool = (raw_pool or "").strip()
                    if not code:
                        continue
                    pm[code] = new_pool
                    if new_pool:
                        mapped_n += 1
                    else:
                        unmapped_n += 1
                if mapped_n or unmapped_n:
                    _save_state(state)
                parts = []
                if mapped_n:
                    parts.append(f"{mapped_n} code(s) assigned to a pool")
                if unmapped_n:
                    parts.append(
                        f"{unmapped_n} code(s) left unmapped (will use default)"
                    )
                if parts:
                    flash("Saved: " + "; ".join(parts) + ".", "success")
                else:
                    flash("Nothing to save.", "info")
            return redirect(url_for("setup.step_co_recov"))

        elif action == "assign_code":
            code = (request.form.get("code") or "").strip()
            new_pool = (request.form.get("new_pool") or "").strip()
            if not code:
                flash("Missing loan code.", "error")
            else:
                pm = state.setdefault("pool_map", {})
                pm[code] = new_pool
                _save_state(state)
                if new_pool:
                    flash(f"Loan code '{code}' -> {new_pool}", "success")
                else:
                    flash(f"Loan code '{code}' left unmapped (will use default).",
                          "info")
            return redirect(url_for("setup.step_co_recov"))

        elif action in ("scan_sample_co_folder", "scan_sample_recov_folder"):
            kind = "co" if action == "scan_sample_co_folder" else "recov"
            folder_key = (
                "sample_co_folder" if kind == "co" else "sample_recov_folder"
            )
            folder_str = (request.form.get(folder_key) or "").strip()
            state[folder_key] = folder_str
            res = _scan_sample_co_recov_folder(state, folder_str, kind)
            _flash_co_recov_scan_result(res, kind)
            if res.get("added"):
                # Seed column-mapping suggestions from the latest added file
                cfg_key = "co_columns" if kind == "co" else "recov_columns"
                files = (
                    (state.get("sample_uploads") or {}).get(
                        "co_files" if kind == "co" else "recov_files"
                    )
                    or []
                )
                if files and (state.get(cfg_key) or {}).get("code_col") in (
                    None, "",
                ):
                    try:
                        sug = co_recov_parser._suggest_columns(files[-1]["path"])
                        existing = state.get(cfg_key) or {}
                        existing.update({
                            k: v for k, v in sug.items()
                            if k not in existing or existing.get(k) in (None, "")
                        })
                        state[cfg_key] = existing
                    except Exception:  # noqa: BLE001
                        pass
            _save_state(state)
            return redirect(url_for("setup.step_co_recov"))

        elif action in (
            "ignore_sample_co_layout", "ignore_sample_recov_layout",
            "unignore_sample_co_layout", "unignore_sample_recov_layout",
        ):
            kind = "co" if "_co_" in action else "recov"
            do_ignore = action.startswith("ignore_")
            sig = (request.form.get("signature") or "").strip()
            sig_key = f"{kind}_ignored_signatures"
            sigs = list(state.get(sig_key) or [])
            if do_ignore:
                if sig and sig not in sigs:
                    sigs.append(sig)
                removed = _remove_files_with_signature(
                    state, kind, sig, is_sample=True,
                )
                kind_label = "charge-off" if kind == "co" else "recoveries"
                msg = f"Ignored layout — future scans will skip matching {kind_label} files."
                if removed:
                    msg += f" Removed {removed} already-added file(s)."
                flash(msg, "success")
            else:
                sigs = [s for s in sigs if s != sig]
                flash("Layout un-ignored. Re-run the folder scan to pick up matching files.", "success")
            state[sig_key] = sigs
            _refresh_layouts_in_scan_data(state, kind, "monthly")
            _refresh_layouts_in_scan_data(state, kind, "sample")
            _save_state(state)
            return redirect(url_for("setup.step_co_recov"))

        elif action in ("clear_sample_co_scan", "clear_sample_recov_scan"):
            kind = "co" if action == "clear_sample_co_scan" else "recov"
            state.pop(_co_recov_scan_state_key(kind, "sample"), None)
            flash("Scan results cleared.", "success")
            _save_state(state)
            return redirect(url_for("setup.step_co_recov"))

        elif action in ("remove_sample_co_file", "remove_sample_recov_file"):
            list_key = (
                "co_files" if action == "remove_sample_co_file"
                else "recov_files"
            )
            target_name = (request.form.get("filename") or "").strip()
            uploads = state.setdefault("sample_uploads", {})
            current = uploads.get(list_key) or []
            kept: list[dict[str, Any]] = []
            removed_path: str | None = None
            for entry in current:
                if entry.get("name") == target_name:
                    removed_path = entry.get("path") or removed_path
                    continue
                kept.append(entry)
            if len(kept) == len(current):
                flash(f"{target_name}: not found in upload list.", "info")
            else:
                uploads[list_key] = kept
                # Best-effort: delete the staged copy from %TEMP% so the
                # working area stays tidy. Never fail the request on a
                # filesystem hiccup.
                if removed_path:
                    try:
                        from pathlib import Path as _P
                        _p = _P(removed_path)
                        if _p.exists() and _SAMPLE_DIR in _p.parents:
                            _p.unlink()
                    except Exception:  # noqa: BLE001
                        pass
                flash(f"Removed {target_name}.", "success")
            _save_state(state)
            return redirect(url_for("setup.step_co_recov"))

        elif action in ("clear_all_sample_co", "clear_all_sample_recov"):
            kind = "co" if action == "clear_all_sample_co" else "recov"
            list_key = "co_files" if kind == "co" else "recov_files"
            kind_label = "charge-off" if kind == "co" else "recoveries"
            uploads = state.setdefault("sample_uploads", {})
            current = uploads.get(list_key) or []
            prev_n = len(current)
            # Best-effort: delete each staged copy from %TEMP%.
            for entry in current:
                p = entry.get("path") or ""
                if not p:
                    continue
                try:
                    from pathlib import Path as _P
                    _p = _P(p)
                    if _p.exists() and _SAMPLE_DIR in _p.parents:
                        _p.unlink()
                except Exception:  # noqa: BLE001
                    pass
            uploads[list_key] = []
            # Also drop the cached scan-results table so the page redraws clean.
            state.pop(_co_recov_scan_state_key(kind, "sample"), None)
            if prev_n:
                flash(
                    f"Cleared {prev_n} {kind_label} file(s).",
                    "success",
                )
            else:
                flash(
                    f"No {kind_label} files to clear.",
                    "info",
                )
            _save_state(state)
            return redirect(url_for("setup.step_co_recov"))

        elif action in ("next", "skip"):
            uploads = state.setdefault("sample_uploads", {})
            uploads["no_recoveries"] = (request.form.get("no_sample_recoveries") == "on")
            _save_state(state)
            return redirect(url_for("setup.step_impaired"))

        _save_state(state)

    # ---- GET (or fall-through) -----------------------------------------
    co_cfg = state.get("co_columns") or {}
    recov_cfg = state.get("recov_columns") or {}
    su = state.get("sample_uploads") or {}

    co_files = su.get("co_files") or []
    recov_files = su.get("recov_files") or []

    # Worksheet/tab names of the loaded workbook(s) so the file section can
    # offer a tab picker (e.g. reuse the charge-off workbook for recoveries
    # and point this side at its “Recoveries” tab).
    co_sheet_names = (
        monthly_co_recov_aggregator.list_sheet_names(Path(co_files[-1]["path"]))
        if co_files else []
    )
    recov_sheet_names = (
        monthly_co_recov_aggregator.list_sheet_names(
            Path(recov_files[-1]["path"]))
        if recov_files else []
    )

    co_preview = (co_recov_parser.inspect_file(
                      co_files[-1]["path"], sheet=(co_cfg.get("sheet") or None))
                  if co_files else None)
    recov_preview = (co_recov_parser.inspect_file(
                        recov_files[-1]["path"],
                        sheet=(recov_cfg.get("sheet") or None))
                     if recov_files else None)

    co_validation = (co_recov_parser.validate_codes(state, "co")
                     if co_files and co_cfg.get("code_col") not in (None, "")
                     else None)
    recov_validation = (co_recov_parser.validate_codes(state, "recov")
                        if recov_files and recov_cfg.get("code_col") not in (None, "")
                        else None)

    pool_choices = balance_check_service.canonical_pool_order(state)

    return render_template(
        "setup/step_co_recov.html",
        co_cfg=co_cfg, recov_cfg=recov_cfg,
        co_preview=co_preview, recov_preview=recov_preview,
        co_validation=co_validation, recov_validation=recov_validation,
        pool_choices=pool_choices,
        co_sheet_names=co_sheet_names,
        recov_sheet_names=recov_sheet_names,
        files_used=_co_recov_files_used(state),
        **_wizard_ctx("co_recov"),
    )


# =================================================================
# Historical Delinquency (DQ %)
# =================================================================
def _ensure_dq_hist(state: dict[str, Any]) -> dict[str, Any]:
    """Return ``state['dq_hist']``, initialising defaults if missing."""
    dh = state.get("dq_hist")
    if not isinstance(dh, dict):
        dh = {}
        state["dq_hist"] = dh
    dh.setdefault("dq_threshold_days", 60)
    dh.setdefault("manual_rows", [])
    dh.setdefault("last_extract_run", None)
    dh.setdefault("last_solr_run", None)
    return dh


@setup_bp.route("/step/dq-history", methods=["GET", "POST"])
def step_dq_hist():
    """Historical Delinquency step.

    Three sources, all writing to ``loan_code_delinquency_history``:
      A) Derive from uploaded historical loan-data extracts.
      B) 5300 call-report backfill via Solr.
      C) Manual entry of dq_pct per loan_code per quarter.
    """
    state = _state()
    dh = _ensure_dq_hist(state)
    he = _ensure_hist_extracts(state)

    if request.method == "POST":
        action = request.form.get("action", "")
        cu = (state.get("credit_union") or "").strip()

        # ---- Inline NCUA / loan-code -> pool mapping ----------------
        if action == "map_unmapped_dq_codes":
            form = request.form
            codes = form.getlist("unmapped_code")
            pool_map = dict(state.get("pool_map") or {})
            saved = 0
            for i, raw in enumerate(codes):
                raw = (raw or "").strip()
                if not raw:
                    continue
                pool = (form.get(f"unmapped_pool_{i}") or "").strip()
                if not pool:
                    continue
                pool_map[raw] = pool
                saved += 1
            state["pool_map"] = pool_map
            _save_state(state)
            if saved:
                flash(
                    f"Mapped {saved} DQ loan code(s) into the global "
                    "Step 3 Loan Code Mapping.",
                    "success",
                )
            else:
                flash(
                    "No mappings selected. Pick a pool for at least one "
                    "code before clicking Save.",
                    "info",
                )
            return redirect(url_for("setup.step_dq_hist"))

        # ---- Settings -----------------------------------------------
        if action == "save_settings":
            try:
                thr = int(request.form.get("dq_threshold_days") or 60)
            except ValueError:
                thr = 60
            dh["dq_threshold_days"] = max(1, min(360, thr))
            _save_state(state)
            flash("DQ settings saved.", "success")
            return redirect(url_for("setup.step_dq_hist"))

        # ---- Source A: extract derivation ---------------------------
        if action == "upload_dq_extract":
            f = request.files.get("dq_extract_file")
            if f and f.filename:
                try:
                    saved_name = _add_sample_upload(
                        state, "dq_extract_files", f, "dq_extracts",
                    )
                    flash(
                        f"Saved loan-data extract: {saved_name}",
                        "success",
                    )
                except Exception as exc:  # noqa: BLE001
                    flash(f"Upload failed: {exc}", "error")
            else:
                flash("Choose a loan-data extract file to upload.", "error")
            return redirect(url_for("setup.step_dq_hist"))

        if action == "remove_dq_extract":
            name = (request.form.get("name") or "").strip()
            uploads = state.setdefault("sample_uploads", {})
            rows = [
                e for e in (uploads.get("dq_extract_files") or [])
                if e.get("name") != name
            ]
            uploads["dq_extract_files"] = rows
            _save_state(state)
            flash(f"Removed {name}.", "success")
            return redirect(url_for("setup.step_dq_hist"))

        if action == "process_dq_extracts":
            files = (state.get("sample_uploads") or {}).get(
                "dq_extract_files"
            ) or []
            mapping = state.get("column_mappings") or {}
            res = dq_extract_parser.process_files(
                cu,
                files,
                mapping,
                dq_threshold=int(dh.get("dq_threshold_days") or 60),
            )
            dh["last_extract_run"] = res
            _save_state(state)
            if res.get("ok"):
                flash(
                    f"Derived DQ history from "
                    f"{sum(1 for fi in res['files'] if fi['ok'])}/"
                    f"{len(res['files'])} file(s); "
                    f"wrote {res['rows_written']} row(s).",
                    "success",
                )
            else:
                flash(
                    f"DQ extract processing failed: "
                    f"{res.get('error') or 'see per-file errors'}",
                    "error",
                )
            return redirect(url_for("setup.step_dq_hist"))

        # ---- Source B: 5300 Solr backfill ---------------------------
        if action == "run_solr_dq_backfill":
            sb = he["solr_backfill"]
            charter_raw = (state.get("charter_number") or "").strip()
            period = (he.get("target_period") or "").strip()
            # WARM CUs don't visit the Historical step (which is where
            # ``target_period`` is normally captured).  Fall back to the
            # global Report Period anchor set on Step 1 (Identity).
            if not period:
                period = balance_check_service._report_period_cutoff(state)
                if period:
                    he["target_period"] = period
            months = int(he.get("history_months") or 84)
            try:
                charter_int = int(re.sub(r"\D", "", charter_raw))
            except ValueError:
                charter_int = 0
            if not cu:
                flash(
                    "Set the credit union on the Identity step first.",
                    "error",
                )
            elif not charter_int:
                flash(
                    "Set the CU's charter number on the Identity step "
                    "first.",
                    "error",
                )
            elif not period:
                flash(
                    "Set the target reporting period on the Historical "
                    "step (or Identity for WARM CUs) first.",
                    "error",
                )
            else:
                existing: set[str] = set()
                try:
                    hv = delinquency_hist_processor.history_matrix(cu)
                    existing = set((hv or {}).get("months") or [])
                except Exception:  # noqa: BLE001
                    existing = set()
                dh["last_solr_run"] = (
                    solr_5300_delq_backfill
                    .backfill_missing_delinquency_quarters(
                        cu, charter_int, sb["solr_url"], sb["core"],
                        period, months,
                        existing_dates=existing,
                    )
                )
                _save_state(state)
                lr = dh["last_solr_run"]
                if lr.get("ok"):
                    filled = len(lr.get("months_filled") or [])
                    skipped = len(lr.get("months_skipped") or [])
                    none_yet = len(lr.get("months_no_data") or [])
                    stale = int(lr.get("stale_rows_removed") or 0)
                    msg = (
                        f"5300 DQ backfill: filled {filled} quarter(s) "
                        f"({lr.get('rows_written', 0)} row(s)); "
                        f"{skipped} skipped, {none_yet} had no Solr doc."
                    )
                    if stale:
                        msg += f" Removed {stale} stale row(s)."
                    flash(msg, "success" if filled else "warning")
                else:
                    flash(
                        f"5300 DQ backfill failed: {lr.get('error')}",
                        "error",
                    )
            return redirect(url_for("setup.step_dq_hist"))

        # ---- Source D: WARM workbook DQ % history -------------------
        if action == "load_warm_dq_history":
            warm_meta = state.get("warm") or {}
            saved_path = warm_meta.get("saved_path") or ""
            if not cu:
                flash(
                    "Set the credit union on the Identity step first.",
                    "error",
                )
                return redirect(url_for("setup.step_dq_hist"))
            if not saved_path:
                flash(
                    "No WARM workbook is attached to this CU. Re-upload "
                    "it on the Start screen, then try again.",
                    "error",
                )
                return redirect(url_for("setup.step_dq_hist"))
            res = warm_parser.parse_dq_history_from_warm(saved_path)
            dh["last_warm_run"] = res
            if not res.get("ok"):
                flash(
                    f"Could not load DQ history from WARM: "
                    f"{res.get('error')}",
                    "error",
                )
                _save_state(state)
                return redirect(url_for("setup.step_dq_hist"))
            try:
                delinquency_hist_processor.ensure_table()
            except Exception as exc:  # noqa: BLE001
                flash(f"DB error: {exc}", "error")
                return redirect(url_for("setup.step_dq_hist"))
            # Replace WARM rows only — leave 5300 / extract / manual alone.
            try:
                delinquency_hist_processor.delete_rows_by_source_prefix(
                    cu, "WARM:",
                )
            except Exception as exc:  # noqa: BLE001
                flash(
                    f"Could not clear prior WARM DQ rows: {exc}", "error",
                )
                return redirect(url_for("setup.step_dq_hist"))
            warm_fn = warm_meta.get("filename") or Path(saved_path).name
            source_tag = f"WARM:{warm_fn}"
            written = 0
            try:
                eng = delinquency_hist_processor._engine_lazy()
                with eng.begin() as conn:
                    for row in res.get("rows") or []:
                        pct = row.get("dq_pct")
                        if pct is None:
                            continue
                        conn.execute(
                            delinquency_hist_processor._UPSERT,
                            {
                                "cu": cu,
                                "as_of_date": row["as_of_date"],
                                "loan_code": row["loan_code"],
                                "dq_amount": 0.0,
                                "total_balance": None,
                                "dq_pct": float(pct),
                                "source": source_tag,
                            },
                        )
                        written += 1
            except Exception as exc:  # noqa: BLE001
                flash(f"DB write failed: {exc}", "error")
                _save_state(state)
                return redirect(url_for("setup.step_dq_hist"))
            res["rows_written"] = written
            res["source_tag"] = source_tag
            dh["last_warm_run"] = res
            _save_state(state)
            flash(
                f"Loaded {written} DQ row(s) from WARM "
                f"&ldquo;{res.get('sheet')}&rdquo; tab "
                f"({res.get('date_columns')} month(s) &times; "
                f"{res.get('pool_count')} pool(s)).",
                "success",
            )
            return redirect(url_for("setup.step_dq_hist"))

        # ---- Source C: manual entry ---------------------------------
        if action == "save_manual_dq":
            dates = request.form.getlist("manual_date")
            codes = request.form.getlist("manual_code")
            pcts = request.form.getlist("manual_pct")
            grouped: dict[str, list[dict[str, Any]]] = {}
            for d, c, p in zip(dates, codes, pcts, strict=False):
                d = (d or "").strip()
                c = (c or "").strip()
                p = (p or "").strip()
                if not d or not c:
                    continue
                try:
                    pct_val = float(p) if p else None
                except ValueError:
                    pct_val = None
                if pct_val is not None and pct_val > 1.5:
                    # User typed "3" meaning 3% — convert to fraction.
                    pct_val = pct_val / 100.0
                grouped.setdefault(d, []).append({
                    "loan_code": c,
                    "dq_amount": 0.0,
                    "total_balance": None,
                    "dq_pct": pct_val,
                })
            try:
                delinquency_hist_processor.ensure_table()
            except Exception as exc:  # noqa: BLE001
                flash(f"DB error: {exc}", "error")
                return redirect(url_for("setup.step_dq_hist"))
            # Replace MANUAL rows only — leave 5300 / extract rows alone.
            try:
                delinquency_hist_processor.delete_rows_by_source_prefix(
                    cu, "manual:",
                )
            except Exception as exc:  # noqa: BLE001
                flash(f"Could not clear prior manual rows: {exc}", "error")
                return redirect(url_for("setup.step_dq_hist"))
            written = 0
            for d, rows in grouped.items():
                # Manual rows are merged on top of any existing rows for
                # the same (cu, as_of_date) from other sources by going
                # one loan_code at a time via a per-code upsert that
                # preserves rows we don't touch. Easiest impl: load
                # existing rows for that date, overlay manual rows by
                # loan_code, then upsert_month (which DELETEs by date).
                eng = delinquency_hist_processor._engine_lazy()
                from sqlalchemy import text as _sql_text
                with eng.begin() as conn:
                    existing_rows = conn.execute(
                        _sql_text(
                            "SELECT loan_code, dq_amount, total_balance, "
                            "       dq_pct, source "
                            "FROM loan_code_delinquency_history "
                            "WHERE cu = :cu AND as_of_date = :d"
                        ),
                        {"cu": cu, "d": d},
                    ).fetchall()
                merged: dict[str, dict[str, Any]] = {}
                for er in existing_rows:
                    merged[er[0]] = {
                        "loan_code": er[0],
                        "dq_amount": float(er[1] or 0.0),
                        "total_balance": (
                            float(er[2]) if er[2] is not None else None
                        ),
                        "dq_pct": (
                            float(er[3]) if er[3] is not None else None
                        ),
                        "_source": er[4] or "",
                    }
                for mr in rows:
                    merged[mr["loan_code"]] = {**mr, "_source": "manual:"}
                # upsert_month deletes the date wholesale then writes
                # everything in ``rows`` under one source tag. To keep
                # mixed sources accurate, write twice with different
                # source tags using direct SQL via the _UPSERT path.
                conn = eng.connect()
                trans = conn.begin()
                try:
                    conn.execute(
                        _sql_text(
                            "DELETE FROM loan_code_delinquency_history "
                            "WHERE cu = :cu AND as_of_date = :d"
                        ),
                        {"cu": cu, "d": d},
                    )
                    for code, rec in merged.items():
                        conn.execute(
                            delinquency_hist_processor._UPSERT,
                            {
                                "cu": cu,
                                "as_of_date": d,
                                "loan_code": code,
                                "dq_amount": float(rec.get("dq_amount") or 0.0),
                                "total_balance": rec.get("total_balance"),
                                "dq_pct": rec.get("dq_pct"),
                                "source": rec.get("_source") or "manual:",
                            },
                        )
                        written += 1
                    trans.commit()
                except Exception as exc:  # noqa: BLE001
                    trans.rollback()
                    flash(f"DB write failed for {d}: {exc}", "error")
                finally:
                    conn.close()
            _save_state(state)
            flash(
                f"Saved {written} manual DQ row(s) across "
                f"{len(grouped)} date(s).",
                "success",
            )
            return redirect(url_for("setup.step_dq_hist"))

        # ---- Navigation ---------------------------------------------
        if action == "back":
            _save_state(state)
            has_warm = (state.get("identity") or {}).get("has_warm") == "yes"
            if has_warm:
                return redirect(url_for("setup.step3_baseline"))
            return redirect(url_for("setup.step3_historical"))

        if action in ("next", "skip"):
            _save_state(state)
            return redirect(url_for("setup.step5_monthly_bal"))

        _save_state(state)

    # ---- GET ---------------------------------------------------------
    cu = (state.get("credit_union") or "").strip()
    matrix = (
        delinquency_hist_processor.history_matrix(cu) if cu
        else {"months": [], "codes": [], "cells": {}, "row_count": 0}
    )
    # Roll the stored-history table up by mapped pool name so the
    # display reflects the user's Step 3 Loan Code Mapping.  Raw
    # ``matrix`` is still used by the inline "unmapped codes" picker.
    matrix_pool = _aggregate_dq_history_by_pool(
        matrix, state.get("pool_map") or {}
    )
    # Build NCUA-canonical-name -> configured-pool suggestions for the
    # inline unmapped-codes picker. Mirrors the auto-classification
    # report-engine uses (``_build_ncua_canonical_pool_lookup``) so the
    # wizard pre-selects the same pool the report would route a 5300
    # canonical name to (e.g. "Used Vehicles" -> "Indirect Auto" when
    # the CU has Auto-style pools). Users can still pick a different
    # pool from the dropdown to override.
    dq_suggestions: dict[str, str] = {}
    try:
        import importlib
        _gr = importlib.import_module("generate_report")
        _cfg_for_lookup = {
            "pools": [
                {"name": (p.get("name") or "").strip()}
                for p in (state.get("pool_settings") or [])
                if (p.get("name") or "").strip() and not p.get("excluded")
            ],
            "pool_order": [
                (p.get("name") or "").strip()
                for p in (state.get("pool_settings") or [])
                if (p.get("name") or "").strip() and not p.get("excluded")
            ],
            "excluded_pools": [
                (p.get("name") or "").strip()
                for p in (state.get("pool_settings") or [])
                if (p.get("name") or "").strip() and p.get("excluded")
            ],
        }
        _ncua_lookup = _gr._build_ncua_canonical_pool_lookup(_cfg_for_lookup)
        _pool_map_lc = {
            str(k).strip().lower(): (v or "").strip()
            for k, v in (state.get("pool_map") or {}).items()
        }
        for c in (matrix.get("codes") or []):
            key = str(c).strip().lower()
            if not key:
                continue
            if (_pool_map_lc.get(key) or "").strip():
                continue  # already mapped via Step 3
            sug = _ncua_lookup.get(key) or _ncua_lookup.get(
                " ".join(key.split())
            )
            if sug:
                dq_suggestions[c] = sug
    except Exception:  # noqa: BLE001
        # Suggestions are a UX nicety; never break the page render.
        dq_suggestions = {}
    map_status = solr_5300_delq_backfill.map_status()
    extract_files = (
        (state.get("sample_uploads") or {}).get("dq_extract_files") or []
    )
    return render_template(
        "setup/step_dq_hist.html",
        dh=dh,
        matrix=matrix,
        matrix_pool=matrix_pool,
        map_status=map_status,
        extract_files=extract_files,
        solr_backfill=he.get("solr_backfill") or {},
        target_period=(
            he.get("target_period")
            or balance_check_service._report_period_cutoff(state)
            or ""
        ),
        history_months=he.get("history_months") or 84,
        dq_suggestions=dq_suggestions,
        **_wizard_ctx("dq_hist"),
    )


# =================================================================
# Impaired Loans
# =================================================================
@setup_bp.route("/step/impaired", methods=["GET", "POST"])
def step_impaired():
    """Impaired Loans upload + impairment-type editor + calculations step."""
    state = _state()
    if request.method == "POST":
        action = request.form.get("action", "")

        if action == "upload_sample_impaired":
            f = request.files.get("sample_impaired_file")
            if f and f.filename:
                try:
                    saved_name = _add_sample_upload(state, "impaired_files", f, "sample_impaired")
                    flash(f"Saved impaired loans file: {saved_name}", "success")
                    files = (state.get("sample_uploads") or {}).get("impaired_files") or []
                    if files:
                        parsed = impaired_parser.parse_file(files[-1]["path"])
                        if parsed.get("ok"):
                            existing = state.get("impaired") or {}
                            # Only overwrite the editable types list when
                            # it's empty so a re-upload doesn't blow away
                            # user edits to the impairment types.
                            if not (existing.get("types") or []):
                                existing["types"] = parsed.get("impairment_types") or []
                            existing["dq_ranges"] = (existing.get("dq_ranges")
                                                     or parsed.get("dq_ranges") or [])
                            existing["period_ending"] = parsed.get("period_ending") or existing.get("period_ending")
                            existing["cu_name_in_file"] = parsed.get("cu_name") or ""
                            existing["data_rows"] = parsed.get("data_rows") or []
                            existing["last_parse_error"] = None
                            lookup = impaired_parser.recompute_all(existing, state)
                            existing["lookup_status"] = lookup
                            state["impaired"] = existing
                            _save_state(state)
                            flash(
                                f"Parsed {len(parsed.get('impairment_types') or [])} "
                                f"impairment types and "
                                f"{len(existing.get('data_rows') or [])} loan rows. "
                                f"Loan-data lookup matched "
                                f"{lookup.get('matched', 0)} / "
                                f"{lookup.get('matched', 0) + lookup.get('unmatched', 0)} rows.",
                                "success",
                            )
                        else:
                            existing = state.get("impaired") or {}
                            existing["last_parse_error"] = parsed.get("error")
                            state["impaired"] = existing
                            _save_state(state)
                            flash(f"Could not parse impaired loans file: {parsed.get('error')}",
                                  "error")
                    return redirect(url_for("setup.step_impaired"))
                except Exception as exc:  # noqa: BLE001
                    flash(f"Upload failed: {exc}", "error")
            else:
                flash("Choose an impaired loans file to upload.", "error")

        elif action == "remove_sample_impaired":
            target = (request.form.get("filename") or "").strip()
            uploads = state.setdefault("sample_uploads", {})
            before = uploads.get("impaired_files") or []
            uploads["impaired_files"] = [
                e for e in before if e.get("name") != target
            ]
            # If no impaired files remain, clear parsed/derived data so
            # the UI doesn't show stale rows from the removed file.
            if not uploads["impaired_files"]:
                imp = state.get("impaired") or {}
                imp["data_rows"] = []
                imp["lookup_status"] = None
                imp["last_parse_error"] = None
                imp["period_ending"] = None
                imp["cu_name_in_file"] = ""
                state["impaired"] = imp
            _save_state(state)
            flash(f"Removed impaired loans file: {target}", "info")
            return redirect(url_for("setup.step_impaired"))

        elif action == "save_impaired_types":
            names = request.form.getlist("type_name")
            pcts = request.form.getlist("type_pct")
            new_types: list[dict[str, Any]] = []
            for nm, pc in zip(names, pcts):
                nm_s = (nm or "").strip()
                if not nm_s:
                    continue
                pc_s = (pc or "").strip()
                if pc_s == "" or pc_s.lower() == "variable":
                    pct_val: float | None = None
                else:
                    try:
                        if pc_s.endswith("%"):
                            pct_val = float(pc_s[:-1].strip()) / 100.0
                        else:
                            pct_val = float(pc_s)
                    except ValueError:
                        pct_val = None
                new_types.append({"name": nm_s, "provision_pct": pct_val})
            existing = state.setdefault("impaired", {})
            existing["types"] = new_types
            # Recompute calculations with the new percentages.
            lookup = impaired_parser.recompute_all(existing, state)
            existing["lookup_status"] = lookup
            _save_state(state)
            flash(
                f"Saved {len(new_types)} impairment type(s). "
                f"Recalculated provisions for "
                f"{len(existing.get('data_rows') or [])} row(s).",
                "success",
            )
            return redirect(url_for("setup.step_impaired"))

        elif action == "reparse_impaired":
            files = (state.get("sample_uploads") or {}).get("impaired_files") or []
            if not files:
                flash("No impaired loans file uploaded yet.", "error")
            else:
                parsed = impaired_parser.parse_file(files[-1]["path"])
                if parsed.get("ok"):
                    existing = state.setdefault("impaired", {})
                    existing["types"] = parsed.get("impairment_types") or []
                    existing["dq_ranges"] = parsed.get("dq_ranges") or []
                    existing["period_ending"] = parsed.get("period_ending")
                    existing["cu_name_in_file"] = parsed.get("cu_name") or ""
                    existing["data_rows"] = parsed.get("data_rows") or []
                    existing["last_parse_error"] = None
                    lookup = impaired_parser.recompute_all(existing, state)
                    existing["lookup_status"] = lookup
                    _save_state(state)
                    flash("Re-parsed and recalculated from the uploaded file.",
                          "success")
                else:
                    flash(f"Could not parse: {parsed.get('error')}", "error")
            return redirect(url_for("setup.step_impaired"))

        elif action == "refresh_lookup":
            existing = state.setdefault("impaired", {})
            lookup = impaired_parser.recompute_all(existing, state)
            existing["lookup_status"] = lookup
            _save_state(state)
            flash(
                f"Re-ran loan-data lookup: matched "
                f"{lookup.get('matched', 0)} / "
                f"{lookup.get('matched', 0) + lookup.get('unmatched', 0)} row(s).",
                "success" if lookup.get("ok") else "error",
            )
            return redirect(url_for("setup.step_impaired"))

        elif action == "save_impaired_rows":
            # Editable / add / remove of the Data Entry section.
            fields = impaired_parser.INPUT_FIELDS
            arrays = {fld: request.form.getlist(f"row_{fld}") for fld in fields}
            n = max((len(a) for a in arrays.values()), default=0)
            numeric_money = {
                "current_balance", "other_lender_balance",
                "collateral_value", "allowance_provided",
            }
            new_rows: list[dict[str, Any]] = []
            for i in range(n):
                raw = {fld: (arrays[fld][i] if i < len(arrays[fld]) else "") for fld in fields}
                if not any((str(v) or "").strip() for v in raw.values()):
                    continue  # skip fully-blank row
                row: dict[str, Any] = {}
                for fld in fields:
                    val = (raw[fld] or "").strip()
                    if fld in numeric_money:
                        row[fld] = impaired_parser._to_float(val) if val else None
                    elif fld == "days_dq":
                        try:
                            row[fld] = int(float(val)) if val else None
                        except (TypeError, ValueError):
                            row[fld] = None
                    else:
                        row[fld] = val
                new_rows.append(row)
            existing = state.setdefault("impaired", {})
            existing["data_rows"] = new_rows
            lookup = impaired_parser.recompute_all(existing, state)
            existing["lookup_status"] = lookup
            _save_state(state)
            flash(
                f"Saved {len(new_rows)} loan row(s). Loan-data lookup matched "
                f"{lookup.get('matched', 0)} / "
                f"{lookup.get('matched', 0) + lookup.get('unmatched', 0)}.",
                "success",
            )
            return redirect(url_for("setup.step_impaired"))

        elif action == "assign_impaired_codes_bulk":
            codes = request.form.getlist("map_code")
            pools = request.form.getlist("map_pool")
            if len(codes) != len(pools):
                flash(
                    "Bulk save: mismatched code/pool counts; nothing saved.",
                    "error",
                )
            else:
                pm = state.setdefault("pool_map", {})
                mapped_n = 0
                unmapped_n = 0
                for raw_code, raw_pool in zip(codes, pools):
                    code = (raw_code or "").strip()
                    new_pool = (raw_pool or "").strip()
                    if not code:
                        continue
                    pm[code] = new_pool
                    if new_pool:
                        mapped_n += 1
                    else:
                        unmapped_n += 1
                if mapped_n or unmapped_n:
                    # Recompute lookup once after all updates so the impaired
                    # table reflects the new mapping in a single pass.
                    existing = state.setdefault("impaired", {})
                    lookup = impaired_parser.recompute_all(existing, state)
                    existing["lookup_status"] = lookup
                    _save_state(state)
                parts = []
                if mapped_n:
                    parts.append(f"{mapped_n} code(s) assigned to a pool")
                if unmapped_n:
                    parts.append(
                        f"{unmapped_n} code(s) left unmapped (will use default)"
                    )
                if parts:
                    flash("Saved: " + "; ".join(parts) + ".", "success")
                else:
                    flash("Nothing to save.", "info")
            return redirect(url_for("setup.step_impaired"))

        elif action == "assign_impaired_code":
            code = (request.form.get("code") or "").strip()
            new_pool = (request.form.get("new_pool") or "").strip()
            if not code:
                flash("Missing loan code.", "error")
            else:
                pm = state.setdefault("pool_map", {})
                pm[code] = new_pool
                # Recompute lookup so the impaired table reflects the new mapping
                existing = state.setdefault("impaired", {})
                lookup = impaired_parser.recompute_all(existing, state)
                existing["lookup_status"] = lookup
                _save_state(state)
                if new_pool:
                    flash(f"Loan code '{code}' -> {new_pool}", "success")
                else:
                    flash(f"Loan code '{code}' left unmapped (will use default).",
                          "info")
            return redirect(url_for("setup.step_impaired"))

        elif action in ("next", "skip"):
            _save_state(state)
            return redirect(url_for("setup.step2_files"))

        _save_state(state)

    impaired = state.get("impaired") or {}
    # Auto-migrate stale rows from older schema (missing new keys).
    rows = impaired.get("data_rows") or []
    if rows and not all(isinstance(r, dict) and "current_balance" in r for r in rows):
        impaired["data_rows"] = []
        impaired["lookup_status"] = None
        state["impaired"] = impaired
        _save_state(state)
        flash(
            "Cleared impaired-loan rows from a previous version. "
            "Please re-upload the WARM file to repopulate.",
            "warning",
        )

    # Build loan-code validation summary for impaired rows whose resolved
    # Loan Pool is empty or "Ignore" — regardless of whether the row was
    # matched in the loan-data extract or fell back to the data-entry code.
    # This surfaces BOTH (a) codes absent from pool_map (fall through to
    # default_pool, which is typically "Ignore") AND (b) codes explicitly
    # mapped to "Ignore" but which the user now wants to re-pool.
    pool_map = state.get("pool_map") or {}
    pool_split = state.get("pool_code_split") or None
    pool_choices = sorted({
        v for v in pool_map.values()
        if v and str(v).strip().lower() not in ("ignore", "exclude")
    })
    # Also add named pools from pool_settings so the user can pick a pool
    # that isn't yet present anywhere in pool_map.
    for p in state.get("pool_settings") or []:
        name = (p.get("name") or "").strip() if isinstance(p, dict) else ""
        if name and name.lower() not in ("ignore", "exclude"):
            if name not in pool_choices:
                pool_choices = sorted(set(pool_choices) | {name})

    def _is_ignore_pool(v: Any) -> bool:
        s = (str(v) if v is not None else "").strip().lower()
        return s == "" or s == "ignore"

    code_counts: dict[str, int] = {}
    for r in impaired.get("data_rows") or []:
        # Surface every row whose resolved Loan Pool is blank or "Ignore",
        # not just those that fell back to the data-entry loan code.
        if not _is_ignore_pool(r.get("loan_pool")):
            continue
        # For MATCHED rows (unmatched_in_loan_data=False) the pool came
        # from the loan extract's pool_code column lookup — so the code
        # the user needs to remap is the EXTRACT'S raw pool code, not the
        # impaired row's loan_type. For UNMATCHED rows the impaired row's
        # loan_type IS the key under pool_map. Prefer the extract code
        # when present (Phase 9.24b).
        raw = r.get("loan_pool_code_resolved")
        if raw is None or str(raw).strip() == "":
            raw = r.get("loan_type")
        if raw is None or str(raw).strip() == "":
            continue
        code = str(raw).strip()
        if pool_split and pool_split in code:
            code = code.split(pool_split, 1)[0].strip()
        code_counts[code] = code_counts.get(code, 0) + 1
    impaired_validation = {
        "unmapped_codes": sorted(code_counts.keys(), key=str.lower),
        "counts": code_counts,
        "pool_choices": pool_choices,
    }

    return render_template(
        "setup/step_impaired.html",
        impaired=impaired,
        impaired_validation=impaired_validation,
        **_wizard_ctx("impaired"),
    )


# =================================================================
# Step 3 — Loan file format
# =================================================================

# File extensions we'll show in the data-directory scan preview.  Keep this
# tight to avoid drowning the table in unrelated files.
_FILE_SCAN_EXTS = {".xlsx", ".xls", ".xlsm", ".csv", ".txt"}
_FILE_SCAN_LIMIT = 200  # cap so a noisy folder doesn't blow up the page


def _suggest_file_patterns(filename: str) -> dict[str, str] | None:
    """Heuristically derive ``file_pattern`` + ``date_pattern`` + ``date_format``
    from a sample loan-data filename.

    Delegates to :func:`sample_parser.guess_filename_patterns` so that the
    returned ``date_format`` is one ``import_data`` actually understands
    (see ``sample_parser.ACCEPTED_DATE_FORMATS``). Returns ``None`` when
    no recognizable date is found.
    """
    if not filename:
        return None
    try:
        guess = sample_parser.guess_filename_patterns(filename)
    except Exception:  # noqa: BLE001
        return None
    # ``guess_filename_patterns`` always returns a dict (defaults to
    # YYYY-MM); only treat it as a real suggestion when a date pattern
    # actually matched the filename. The default is a benign fallback,
    # not an inferred match.
    name = Path(filename).name
    stem = Path(name).stem
    if not any(
        re.search(rx, stem) for _desc, rx in sample_parser._DATE_RX_CANDIDATES
    ):
        return None
    return {
        "file_pattern": guess.get("file_pattern") or "",
        "date_pattern": guess.get("date_pattern") or "",
        "date_format": guess.get("date_format") or "YYYY-MM",
    }


def _scan_data_directory(
    data_directory: str,
    file_pattern: str,
    date_pattern: str,
    date_format: str = "YYYY-MM",
) -> dict[str, Any]:
    """Walk ``data_directory`` and report which files match ``file_pattern``.

    When ``date_format`` is supplied, the per-file date preview is
    resolved through ``import_data.extract_snapshot_date`` so it matches
    exactly what the importer will produce at run-time. This catches
    silent format / pattern mismatches (e.g. ``date_format='MDY'`` —
    not a recognized value, falls through to YYYY-MM, yields garbage
    months) before the user moves off Step 2.
    """
    import re as _re
    from datetime import datetime as _dt

    out: dict[str, Any] = {
        "ok": False, "error": None, "directory": data_directory,
        "total_files": 0, "matched": 0, "truncated": False, "files": [],
    }
    if not data_directory:
        out["error"] = "No data directory set yet (Step 1)."
        return out
    root = Path(data_directory)
    if not root.exists():
        out["error"] = f"Directory does not exist: {data_directory}"
        return out
    if not root.is_dir():
        out["error"] = f"Not a directory: {data_directory}"
        return out

    try:
        file_re = _re.compile(file_pattern, _re.IGNORECASE) if file_pattern else None
    except _re.error as exc:
        out["error"] = f"Invalid file_pattern regex: {exc}"
        return out
    try:
        date_re = _re.compile(date_pattern) if date_pattern else None
    except _re.error as exc:
        out["error"] = f"Invalid date_pattern regex: {exc}"
        return out

    # Defer the heavy ``import_data`` import to the first call that
    # actually has both a date_pattern and date_format. Matches what
    # the importer does at run-time, so silent format mismatches
    # surface in the scan preview.
    extract_date = None
    if date_pattern and date_format:
        try:
            import import_data as _id  # type: ignore[import-not-found]
            extract_date = _id.extract_snapshot_date
        except Exception:  # noqa: BLE001
            extract_date = None

    rows: list[dict[str, Any]] = []
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.lower() not in _FILE_SCAN_EXTS:
            continue
        if path.name.startswith("~$"):
            continue
        out["total_files"] += 1
        name = path.name
        try:
            sub = str(path.parent.relative_to(root))
            if sub == ".":
                sub = ""
        except ValueError:
            sub = ""
        matched = bool(file_re.search(name)) if file_re else False
        date_str: str | None = None
        date_err: str | None = None
        if matched and date_re:
            # Preferred path: route through extract_snapshot_date so the
            # preview reflects what the importer will actually do. Falls
            # back to the simpler 2-group regex parse when import_data
            # could not be loaded.
            if extract_date is not None:
                try:
                    iso = extract_date(
                        name,
                        {"date_pattern": date_pattern,
                         "date_format": date_format},
                    )
                    if iso:
                        # Render YYYY-MM-DD when import_data found a day,
                        # else just YYYY-MM. The date_format dictates which.
                        date_str = iso
                    else:
                        date_err = (
                            f"format {date_format} could not resolve "
                            f"a date from this filename"
                        )
                except Exception as exc:  # noqa: BLE001
                    date_err = str(exc)
            else:
                m = date_re.search(name)
                if m and m.lastindex and m.lastindex >= 2:
                    try:
                        yr, mo = int(m.group(1)), int(m.group(2))
                        if 1900 <= yr <= 2100 and 1 <= mo <= 12:
                            date_str = f"{yr:04d}-{mo:02d}"
                        else:
                            date_err = f"out-of-range yr/mo: {yr}/{mo}"
                    except (ValueError, IndexError) as exc:
                        date_err = str(exc)
                elif m:
                    date_err = "regex matched but lacks 2 capture groups"
                else:
                    date_err = "no date match"
        if matched:
            out["matched"] += 1
        try:
            mtime = path.stat().st_mtime
        except OSError:
            mtime = 0
        rows.append({
            "name": name, "subdir": sub, "matched": matched,
            "date": date_str, "date_error": date_err, "_mtime": mtime,
        })

    # Newest first; truncate.
    rows.sort(key=lambda r: r["_mtime"], reverse=True)
    if len(rows) > _FILE_SCAN_LIMIT:
        out["truncated"] = True
        rows = rows[:_FILE_SCAN_LIMIT]
    for r in rows:
        r.pop("_mtime", None)
    out["files"] = rows
    out["ok"] = True
    return out


@setup_bp.route("/step/files", methods=["GET", "POST"])
def step2_files():
    state = _state()
    # Form values to render with: usually pulled from state, but on a "test"
    # action we re-render with the just-typed (unsaved) values so the preview
    # reflects them.
    # Phase 9.35d: defensive `.get(..., default)` for top-level scalars.
    # Adopted drafts (created via "Edit setup" from an existing YAML in
    # `cecl_ui/routes/home.py adopt_config_to_completed`) hydrate the wizard
    # state directly from the YAML schema and do NOT pass through
    # `_default_state()`. Top-level scalars like `pool_code_split` that exist
    # in `_default_state()` but not in every YAML would otherwise raise
    # KeyError when the user navigates to this step. The companion structural
    # fix is in `selfheal_adopt_yaml_schema_into_state` which proactively
    # seeds these defaults on every wizard GET — these defensive `.get()`
    # calls are belt-and-suspenders so the page still renders even if a
    # future code path bypasses the self-heal.
    form_values: dict[str, Any] = {
        "file_pattern": state.get("file_pattern") or r"LOANDATA.*\.(xlsx|xls|csv)$",
        "date_pattern": state.get("date_pattern") or r"(\d{4})-(\d{2})",
        "date_format": state.get("date_format") or "YYYY-MM",
        "pool_code_split": state.get("pool_code_split", "/"),
        "balance_remove_chars": state.get("balance_remove_chars", ["$", ","]),
        "accounting_negatives": state.get("accounting_negatives", True),
    }
    scan: dict[str, Any] | None = None

    # If the user uploaded a sample loan-data file earlier in the wizard,
    # offer it as the basis for auto-suggested patterns.
    sample_loan_files = (
        (state.get("sample_uploads") or {}).get("loan_data_files") or []
    )
    # Guard against entries that carry no ``path`` (e.g. script-built configs
    # whose loan-data files were never uploaded through the wizard) — a hard
    # ``["path"]`` subscript here 500s the whole step.
    _last_loan_path = (
        sample_loan_files[-1].get("path") if sample_loan_files else None
    )
    sample_loan_name = Path(_last_loan_path).name if _last_loan_path else ""
    suggestion: dict[str, str] | None = None
    if sample_loan_name:
        suggestion = _suggest_file_patterns(sample_loan_name)

    # On the very first GET, if patterns are still the defaults and we have
    # a sample loan filename, prefill the form with the suggestion so the
    # user only has to confirm.
    _DEFAULTS = (r"LOANDATA.*\.(xlsx|xls|csv)$", r"(\d{4})-(\d{2})")
    is_default = (
        form_values["file_pattern"] == _DEFAULTS[0]
        and form_values["date_pattern"] == _DEFAULTS[1]
    )
    if request.method == "GET" and suggestion and is_default:
        form_values["file_pattern"] = suggestion["file_pattern"]
        form_values["date_pattern"] = suggestion["date_pattern"]
        # Only override the saved date_format when the user is still on
        # the wizard default; if they've already locked in a value we
        # don't want a sample re-detection to clobber it on revisit.
        if (state.get("date_format") or "YYYY-MM") == "YYYY-MM":
            form_values["date_format"] = suggestion.get(
                "date_format", "YYYY-MM"
            )

    # Auto-remediation (Phase: month-name date rescue). Independent of the
    # first-load prefill above: when the file_pattern was already
    # customized (so ``is_default`` is False) but the saved
    # (date_pattern, date_format) pair CANNOT resolve a date from the
    # uploaded sample, adopt the sample-derived suggestion's date fields
    # so the red "no date" / ✗ preview flips green. This is the common
    # case for month-name filenames (e.g. "Aires December 25.v2.xlsx")
    # whose file_pattern matched but whose date_pattern was left at the
    # numeric YYYY-MM default. Values land in form_values only — the user
    # still confirms by clicking Next, which persists them.
    if (
        request.method == "GET"
        and suggestion
        and sample_loan_name
        and not is_default
    ):
        try:
            _cur = sample_parser.validate_date_format(
                form_values["date_pattern"],
                form_values["date_format"],
                sample_loan_name,
            )
        except Exception:  # noqa: BLE001
            _cur = {"ok": False}
        if not _cur.get("ok"):
            try:
                _sug = sample_parser.validate_date_format(
                    suggestion["date_pattern"],
                    suggestion.get("date_format", "YYYY-MM"),
                    sample_loan_name,
                )
            except Exception:  # noqa: BLE001
                _sug = {"ok": False}
            if _sug.get("ok"):
                form_values["date_pattern"] = suggestion["date_pattern"]
                form_values["date_format"] = suggestion.get(
                    "date_format", "YYYY-MM"
                )
                flash(
                    "Adjusted the date pattern to match this CU's filename "
                    f"style — '{sample_loan_name}' now reads as "
                    f"{_sug.get('preview')}. Click Next to save it.",
                    "success",
                )

    if request.method == "POST":
        action = request.form.get("action", "save")
        # Read everything off the form first.
        fp = request.form.get("file_pattern", "").strip()
        dp = request.form.get("date_pattern", "").strip()
        df = (request.form.get("date_format") or "").strip() or "YYYY-MM"
        if df not in sample_parser.ACCEPTED_DATE_FORMATS:
            # Reject silently-invalid choices BEFORE they reach state /
            # YAML / import_data. The select on the form is restricted to
            # accepted values so this only fires on direct form tampering.
            flash(
                f"Date format {df!r} is not recognized; "
                f"falling back to YYYY-MM.",
                "error",
            )
            df = "YYYY-MM"
        pcs = request.form.get("pool_code_split", "").strip() or "/"
        an = request.form.get("accounting_negatives") == "on"
        rcs_str = request.form.get("balance_remove_chars", "$,").strip()
        rcs = [c for c in rcs_str.split(",") if c]

        form_values.update({
            "file_pattern": fp, "date_pattern": dp, "date_format": df,
            "pool_code_split": pcs,
            "balance_remove_chars": rcs, "accounting_negatives": an,
        })

        if action == "remove_file":
            # Delete a file from the configured data directory. The filename
            # is supplied as a relative path (subdir/name) so we resolve it
            # against data_directory and reject any escape attempts.
            target_rel = (request.form.get("filename") or "").strip()
            data_dir = state.get("data_directory") or ""
            if not target_rel or not data_dir:
                flash("Nothing to remove.", "error")
            else:
                try:
                    base = Path(data_dir).resolve()
                    candidate = (base / target_rel).resolve()
                    # Refuse to delete anything outside the data directory.
                    if base != candidate and base not in candidate.parents:
                        flash(
                            f"Refusing to delete file outside data "
                            f"directory: {target_rel}",
                            "error",
                        )
                    elif not candidate.exists() or not candidate.is_file():
                        flash(f"File not found: {target_rel}", "error")
                    else:
                        candidate.unlink()
                        flash(f"Deleted {candidate.name}.", "success")
                except OSError as exc:
                    flash(f"Could not delete {target_rel}: {exc}", "error")
            # Re-scan so the user immediately sees the file is gone.
            scan = _scan_data_directory(data_dir, fp, dp, df)
        elif action == "suggest":
            # User clicked "Suggest from sample". Overwrite file/date
            # patterns from the sample filename heuristic and re-render
            # without saving so they can review/test.
            if suggestion:
                form_values["file_pattern"] = suggestion["file_pattern"]
                form_values["date_pattern"] = suggestion["date_pattern"]
                form_values["date_format"] = suggestion.get(
                    "date_format", "YYYY-MM"
                )
                flash(
                    f"Suggested patterns from sample: "
                    f"{sample_loan_name} (format "
                    f"{form_values['date_format']}).",
                    "success",
                )
            else:
                flash(
                    "Could not detect a date in the sample filename "
                    + (f"({sample_loan_name})." if sample_loan_name
                       else "(no sample uploaded)."),
                    "info",
                )
        elif action == "test":
            # Test the patterns against the configured data directory but DO
            # NOT save state — let the user iterate.
            scan = _scan_data_directory(
                state.get("data_directory", ""), fp, dp, df,
            )
            if scan.get("error"):
                flash(scan["error"], "error")
            else:
                flash(
                    f"Tested patterns: {scan['matched']} of {scan['total_files']} "
                    f"files matched.",
                    "success" if scan["matched"] > 0 else "info",
                )
        else:  # save / next
            state["file_pattern"] = fp
            state["date_pattern"] = dp
            state["date_format"] = df
            state["pool_code_split"] = pcs
            state["accounting_negatives"] = an
            state["balance_remove_chars"] = rcs
            _save_state(state)
            # File Format is now late in the wizard — next is Economic Data.
            return redirect(url_for("setup.step7_economic"))
    else:
        # GET: auto-run a scan if we have a directory + patterns, so the user
        # sees feedback immediately on first load.
        if state.get("data_directory") and form_values.get("file_pattern"):
            scan = _scan_data_directory(
                state.get("data_directory", ""),
                form_values["file_pattern"],
                form_values["date_pattern"],
                form_values["date_format"],
            )

    # Live validation preview: shows whether the current
    # (date_pattern, date_format) pair can resolve a date out of the
    # uploaded sample filename. Surfaces silent-format-fallthrough
    # bugs (e.g. ``date_format='MDY'`` lands at YYYY-MM and produces
    # garbage at run-time) right on this page.
    date_format_preview: dict[str, Any] | None = None
    if form_values.get("date_pattern"):
        try:
            date_format_preview = sample_parser.validate_date_format(
                form_values["date_pattern"],
                form_values["date_format"],
                sample_loan_name,
            )
        except Exception:  # noqa: BLE001
            date_format_preview = None

    return render_template(
        "setup/step2_files.html",
        form_values=form_values,
        scan=scan,
        sample_loan_name=sample_loan_name,
        suggestion=suggestion,
        date_format_choices=sample_parser.ACCEPTED_DATE_FORMATS,
        date_format_labels=sample_parser.DATE_FORMAT_LABELS,
        date_format_preview=date_format_preview,
        **_wizard_ctx("files"),
    )


# =================================================================
# Step 3 — Column mappings
# =================================================================
@setup_bp.route("/step/columns", methods=["GET", "POST"])
def step3_columns():
    state = _state()

    def _files_list() -> list[dict[str, Any]]:
        uploads = state.get("sample_uploads") or {}
        return list(uploads.get("loan_data_files") or [])

    def _hydrate_legacy_entries() -> bool:
        """Backfill ``analysis`` / ``signature`` / per-file mapping on any
        loan_data_files entry that pre-dates the per-file mapping refactor,
        OR whose cached ``analysis.headers`` is empty (a stale entry from
        before the user toggled ``has_header`` / ``header_row`` on Step 9).
        Returns True if state was mutated."""
        uploads = state.get("sample_uploads") or {}
        files = uploads.get("loan_data_files") or []
        mutated = False
        for entry in files:
            cached = entry.get("analysis") or {}
            cached_headers = list(cached.get("headers") or [])
            import re as _re_hdr
            _SYN = _re_hdr.compile(r"^col_[A-Z]+$")
            all_synthetic = bool(cached_headers) and all(
                _SYN.match(str(h)) for h in cached_headers
            )
            any_real = bool(cached_headers) and any(
                not _SYN.match(str(h)) for h in cached_headers
            )
            # Stale-mapping check: the entry's saved column_mappings
            # reference header names that no longer exist in the cached
            # headers list — this happens when the file was originally
            # parsed with header_row=1 (real names captured + mapped),
            # then re-parsed with header_row=0 (synthetic col_<LETTER>
            # names took over) without clearing the mappings. The user
            # is stuck with ``Balance (not in current sample)`` ghost
            # options that always win over fresh picks. Recover by
            # re-parsing with header_row=1.
            cm = entry.get("column_mappings") or {}
            cm_real_values = [
                v for k, v in cm.items()
                if k != "loan_pool_code_static"
                and v
                and not _SYN.match(str(v))
            ]
            stale_mappings_against_synthetic = bool(
                cm_real_values
                and all_synthetic
            )
            needs_reparse = (
                not entry.get("analysis")
                or entry.get("signature") is None
                # Empty headers cache while the user has indicated the
                # file has a header row — re-parse so Step 10 can build
                # real dropdowns.
                or (not cached_headers and bool(entry.get("has_header")))
                # has_header=True but every header is a synthetic
                # ``col_<LETTER>`` — leftover from a stale parse where
                # auto-detect missed the header row but the user later
                # toggled the checkbox; force a fresh parse so the
                # dropdowns can show the real header names.
                or (bool(entry.get("has_header")) and all_synthetic)
                # has_header=False but headers contain real names —
                # the user toggled the checkbox off after a real-header
                # parse; re-parse with header_row=0 so col_<LETTER>
                # placeholders take over.
                or (not entry.get("has_header") and any_real)
                # Saved mappings reference real header names but the
                # cache is synthetic — entry is inconsistent.
                or stale_mappings_against_synthetic
            )
            if not needs_reparse:
                continue
            saved = entry.get("path")
            if not saved or not Path(saved).exists():
                # Can't re-parse — leave it; the template will show
                # an empty mapping form for this file and the user can
                # re-upload from Step 9.
                continue
            try:
                hr_int = int(entry.get("header_row") or 0)
            except (TypeError, ValueError):
                hr_int = 0
            # Respect an explicit "no header" entry (header_row==0 AND
            # has_header==False). Otherwise pass the user's chosen
            # 1-based row, falling back to auto-detect when unset.
            if stale_mappings_against_synthetic:
                # The entry is inconsistent: saved mappings reference
                # real header names but the cached headers are
                # synthetic. Force a real-header re-parse so the
                # mappings can re-attach to a fresh ``analysis``.
                header_row_arg: int | None = 1
            elif not entry.get("has_header") and hr_int == 0:
                header_row_arg = 0
            elif hr_int >= 1:
                header_row_arg = hr_int
            else:
                header_row_arg = None
            try:
                analysis = sample_parser.analyse_sample_file(
                    saved,
                    original_filename=entry.get("name") or Path(saved).name,
                    header_row=header_row_arg,
                )
            except Exception:  # noqa: BLE001
                continue
            if not analysis.get("ok"):
                continue
            entry["analysis"] = _loan_data_entry_analysis(analysis)
            entry["has_header"] = bool(analysis.get("has_header"))
            entry["header_row"] = int(analysis.get("header_row") or 0)
            # Drop any column_mappings whose value no longer exists in
            # the freshly-parsed headers list. Otherwise the dropdown
            # ends up showing a ``(not in current sample)`` ghost option
            # that always wins over a real header pick.
            new_headers = list(entry["analysis"].get("headers") or [])
            old_map = entry.get("column_mappings") or {}
            entry["column_mappings"] = {
                k: v
                for k, v in old_map.items()
                if k == "loan_pool_code_static" or v in new_headers
            }
            # Seed mapping from prior top-level state (if any) so the
            # user doesn't lose their previous work; otherwise fall
            # through to suggestion-based seeding.
            top_map = state.get("column_mappings") or {}
            top_ma = state.get("member_account") or {}
            if top_map and not entry.get("column_mappings"):
                entry["column_mappings"] = {
                    k: v for k, v in top_map.items() if v in new_headers
                }
            if top_ma and not entry.get("member_account"):
                entry["member_account"] = dict(top_ma)
            _seed_loan_data_entry_mapping(state, entry, analysis)
            mutated = True
        if mutated:
            _save_state(state)
        return mutated

    def _learned_for_headers(headers: list[str]) -> dict[str, str]:
        if not headers:
            return {}
        try:
            return column_mapping_suggestions.suggest_for_headers(headers)
        except Exception:  # noqa: BLE001
            return {}

    def _default_file_pattern(sample_name: str) -> str:
        """Best-effort regex seed derived from an uploaded sample filename.

        Strips the extension and any trailing date-ish tokens (digits,
        underscores, dashes, dots) so that e.g. ``Mortgages_2024_12.xlsx``
        seeds ``(?i)mortgages`` and matches future months like
        ``Mortgages_2025_01.xlsx``.
        """
        stem = Path(sample_name or "").stem
        # Drop trailing date-like suffix (digits + separators) repeatedly.
        stem = re.sub(r"[\s_\-.]*\d[\d_\-./]*$", "", stem).strip(" _-.")
        if not stem:
            return ""
        # Make the pattern tolerant of whitespace/underscore swaps that
        # happen when uploaded filenames are sanitized to disk (e.g.
        # "DM - Credit-Migr_CC 30JUN25.xlsx" gets saved as
        # "DM_-_Credit-Migr_CC_30JUN25.xlsx"). Any run of space or
        # underscore in the source becomes "[\s_]+" in the regex.
        parts = re.split(r"[\s_]+", stem)
        escaped = r"[\s_]+".join(re.escape(p) for p in parts if p)
        return "(?i)" + escaped

    def _file_view(entry: dict[str, Any], idx: int) -> dict[str, Any]:
        """Build the per-file template context."""
        analysis = entry.get("analysis") or {}
        headers = list(analysis.get("headers") or [])
        suggestions = dict(analysis.get("column_suggestions") or {})
        mapping = dict(entry.get("column_mappings") or {})
        # Fall back to suggestions for any unset field so the dropdown
        # starts on a reasonable value.
        for fld, sug in suggestions.items():
            if not mapping.get(fld):
                mapping[fld] = sug
        return {
            "idx": idx,
            "name": entry.get("name") or f"File {idx + 1}",
            "headers": headers,
            "suggestions": suggestions,
            "mapping": mapping,
            "member_account": dict(entry.get("member_account") or {
                "mode": "fixed_suffix", "suffix_length": 3, "delimiter": "-",
            }),
            # Surfaced for the "auto-detected from sample" hint on the
            # member-account radios. ``None`` when the parser couldn't
            # infer anything (template renders no hint in that case).
            "member_account_suggestion": (
                dict(analysis.get("member_account_suggestion") or {})
                if analysis.get("member_account_suggestion") else None
            ),
            # Top-level state's member_account dict, so the template can
            # render a "Custom override (differs from top-level)" pill
            # whenever the per-file dict diverges. Lets the user reason
            # about which extracts inherit vs. override.
            "top_level_member_account": dict(
                state.get("member_account") or {}
            ),
            "has_header": bool(entry.get("has_header")),
            "header_row": int(entry.get("header_row") or 0),
            "sample_rows": list(analysis.get("sample_rows") or [])[:5],
            "learned": _learned_for_headers(headers),
            "signature": entry.get("signature") or "",
            "file_pattern": (
                entry.get("file_pattern")
                or _default_file_pattern(entry.get("name") or "")
            ),
            "sample_filename": entry.get("name") or "",
            # Phase 9.22: per-extract pool_code_split override. ``None`` =
            # inherit top-level; ``""`` (empty string explicitly stored on
            # the entry) = no split; any other string = split on that
            # character. The template renders a 4-option dropdown
            # (inherit, none, "/", "-") based on this value.
            "pool_code_split": (
                entry["pool_code_split"]
                if "pool_code_split" in entry
                else None
            ),
        }

    if request.method == "POST":
        _hydrate_legacy_entries()
        files = _files_list()
        if not files:
            flash(
                "No loan-data files uploaded yet. Upload at least one "
                "extract on the previous step before mapping columns.",
                "error",
            )
            return redirect(url_for("setup.step2_sample"))

        required = ["member_number", "current_balance"]
        # ``days_delinquent`` is optional: overdraft / mortgage-summary
        # extracts legitimately have no delinquency column, and the
        # importer defaults an unmapped value to 0. Requiring it blocked
        # multi-extract CUs (e.g. ODP / CUMA files) from advancing past
        # this step.
        optional = ("loan_suffix", "loan_pool_code", "original_fico_score",
                    "current_fico_score", "days_delinquent",
                    "interest_rate", "open_date",
                    "original_loan_amount", "total_available_credit",
                    "business_risk_rating")

        all_errors: list[str] = []
        new_entries: list[tuple[dict[str, Any], dict[str, Any]]] = []
        # tuples of (original_entry, new_values_dict_to_apply)

        # First pass: detect any per-file ``has_header`` toggles.
        # When the checkbox state on the form differs from what was
        # saved on the entry, the cached ``analysis.headers`` is stale
        # — re-parse the file with the new header setting BEFORE we
        # process column mappings, otherwise dropdowns are populated
        # with column-letter placeholders (col_A/col_B/...) and any
        # mapping values the user picks against those placeholders
        # won't match the file's real headers.
        reparsed_labels: list[str] = []
        import re as _re_hdr
        _SYNTHETIC_HDR = _re_hdr.compile(r"^col_[A-Z]+$")
        for i, entry in enumerate(files):
            prefix = f"f{i}__"
            has_header_flag = request.form.get(prefix + "has_header") == "on"
            cached_has_header = bool(entry.get("has_header"))
            cached_headers = list(
                (entry.get("analysis") or {}).get("headers") or []
            )
            all_synthetic = bool(cached_headers) and all(
                _SYNTHETIC_HDR.match(str(h)) for h in cached_headers
            )
            any_real = bool(cached_headers) and any(
                not _SYNTHETIC_HDR.match(str(h)) for h in cached_headers
            )
            # Re-parse when (a) the checkbox state changed since the
            # entry was last analysed, or (b) the cached headers don't
            # match the requested header mode (synthetic col_* names
            # while header mode is on, or real names while off — both
            # are leftovers from before the form is in sync with the
            # parsed file).
            stale_for_header_mode = (
                (has_header_flag and all_synthetic)
                or (not has_header_flag and any_real)
            )
            if has_header_flag == cached_has_header and not stale_for_header_mode:
                continue
            saved = entry.get("path")
            if not saved or not Path(saved).exists():
                # Can't re-parse without the original upload — just
                # update the flag so the user can manually re-upload.
                entry["has_header"] = has_header_flag
                continue
            new_header_row = 1 if has_header_flag else 0
            try:
                analysis = sample_parser.analyse_sample_file(
                    saved,
                    original_filename=entry.get("name") or Path(saved).name,
                    header_row=new_header_row,
                )
            except Exception:  # noqa: BLE001
                analysis = None
            if not analysis or not analysis.get("ok"):
                entry["has_header"] = has_header_flag
                continue
            entry["analysis"] = _loan_data_entry_analysis(analysis)
            entry["has_header"] = bool(analysis.get("has_header"))
            entry["header_row"] = int(analysis.get("header_row") or 0)
            # Drop any column_mappings whose value no longer exists in
            # the freshly-parsed headers list. Keep entries that still
            # match (e.g. user previously had real header names mapped
            # and the re-parse confirmed them) and the static-pool-code
            # synthetic key, which isn't a real column.
            new_headers = list(entry["analysis"].get("headers") or [])
            old_map = entry.get("column_mappings") or {}
            entry["column_mappings"] = {
                k: v
                for k, v in old_map.items()
                if k == "loan_pool_code_static" or v in new_headers
            }
            # Re-seed any unset suggestion-driven defaults from the
            # fresh analysis so the user starts with a reasonable
            # baseline against the new headers.
            for fld, sug in (entry["analysis"].get("column_suggestions") or {}).items():
                if not entry["column_mappings"].get(fld):
                    entry["column_mappings"][fld] = sug
            reparsed_labels.append(entry.get("name") or f"File {i + 1}")

        if reparsed_labels:
            _save_state(state)
            flash(
                "Re-parsed loan-data extract(s) after the "
                "<em>This file includes a header row</em> toggle changed: "
                + ", ".join(f"<code>{n}</code>" for n in reparsed_labels)
                + ". Please re-confirm column mappings against the "
                "newly-detected headers, then click Save again.",
                "info",
            )
            return render_template(
                "setup/step3_columns.html",
                files_view=[
                    _file_view(e, i) for i, e in enumerate(_files_list())
                ],
                **_wizard_ctx("columns"),
            )

        for i, entry in enumerate(files):
            prefix = f"f{i}__"
            has_header_flag = request.form.get(prefix + "has_header") == "on"

            ma_mode = (request.form.get(prefix + "member_account_mode")
                       or "fixed_suffix").strip()
            if ma_mode not in ("fixed_suffix", "delimiter", "split"):
                ma_mode = "fixed_suffix"

            try:
                ma_suffix_len = int(
                    request.form.get(prefix + "member_account_suffix_length", "3")
                )
            except (TypeError, ValueError):
                ma_suffix_len = 3
            ma_suffix_len = max(0, min(ma_suffix_len, 9))

            ma_delim = (
                request.form.get(prefix + "member_account_delimiter") or "-"
            ).strip() or "-"

            member_account = {
                "mode": ma_mode,
                "suffix_length": ma_suffix_len,
                "delimiter": ma_delim,
            }

            file_required = list(required)
            if ma_mode == "split":
                file_required.append("loan_suffix")

            new_map: dict[str, str] = {}
            for fld in tuple(file_required) + optional:
                val = (request.form.get(prefix + fld) or "").strip()
                if val:
                    new_map[fld] = val

            # Optional "fixed loan pool code for every row" — when set, the
            # importer ignores the loan_pool_code column and synthesizes a
            # constant series from this value.
            static_code = (
                request.form.get(prefix + "loan_pool_code_static") or ""
            ).strip()
            if static_code:
                new_map["loan_pool_code_static"] = static_code

            label = entry.get("name") or f"File {i + 1}"
            missing = [f for f in file_required if not new_map.get(f)]
            # Require EITHER a loan_pool_code column OR a static code.
            if not new_map.get("loan_pool_code") and not static_code:
                missing.append("loan_pool_code (or a fixed code)")
            if missing:
                all_errors.append(
                    f"{label}: missing required column(s) "
                    + ", ".join(missing)
                )
                # Still record the partial values so the form re-renders
                # with what the user entered.
                new_entries.append((entry, {
                    "column_mappings": new_map,
                    "member_account": member_account,
                    "has_header": has_header_flag,
                }))
                continue

            headers = list(((entry.get("analysis") or {}).get("headers")) or [])
            if headers:
                unknown = [
                    f"{f} -> {v}" for f, v in new_map.items()
                    if f != "loan_pool_code_static" and v not in headers
                ]
                if unknown:
                    all_errors.append(
                        f"{label}: mappings don't match any column in the "
                        "file: " + "; ".join(unknown)
                    )

            # Warn on duplicates within this file (non-fatal).
            seen: dict[str, str] = {}
            dupes: list[str] = []
            for f, v in new_map.items():
                if f == "loan_pool_code_static":
                    continue
                if v in seen:
                    dupes.append(f"{seen[v]} & {f} -> {v}")
                else:
                    seen[v] = f
            if dupes:
                flash(
                    f"Heads-up &mdash; in {label}, these system fields point "
                    "at the same column: " + "; ".join(dupes)
                    + ". Allowed but unusual.",
                    "info",
                )

            file_pattern = (
                request.form.get(prefix + "file_pattern") or ""
            ).strip()
            # Validate that it compiles, but don't block save — surface as
            # an error so the user fixes it before running reports.
            if file_pattern:
                try:
                    re.compile(file_pattern)
                except re.error as exc:
                    all_errors.append(
                        f"{label}: filename pattern is not a valid regex "
                        f"({exc}). Example: (?i)mortgages"
                    )

            # Phase 9.22: per-extract pool_code_split override.
            # Form value semantics:
            #   "__inherit__" → no per-entry override (inherit top-level)
            #   "__none__"    → explicitly disable splitting on this file
            #   any other str → split on that exact character
            # CUMA-style mortgage extracts ship loan codes like "15/15 ARM"
            # where '/' is part of the code itself; splitting on '/' would
            # truncate them to '15' and route to the default pool (Ignore).
            pcs_raw = (
                request.form.get(prefix + "pool_code_split")
                or "__inherit__"
            ).strip()
            if pcs_raw == "__inherit__":
                entry_pool_split: str | None = None
                entry_pool_split_set = True  # signal: clear any prior value
            elif pcs_raw == "__none__":
                entry_pool_split = ""
                entry_pool_split_set = True
            else:
                entry_pool_split = pcs_raw
                entry_pool_split_set = True

            new_entries.append((entry, {
                "column_mappings": new_map,
                "member_account": member_account,
                "has_header": has_header_flag,
                "file_pattern": file_pattern,
                "pool_code_split": entry_pool_split,
                "pool_code_split_set": entry_pool_split_set,
            }))

        # Apply (always, so the form re-renders with the user's edits even
        # if some files had errors).
        for orig_entry, updates in new_entries:
            orig_entry["column_mappings"] = updates["column_mappings"]
            orig_entry["member_account"] = updates["member_account"]
            orig_entry["has_header"] = updates["has_header"]
            orig_entry["file_pattern"] = updates["file_pattern"]
            if updates.get("pool_code_split_set"):
                _val = updates.get("pool_code_split")
                if _val is None:
                    orig_entry.pop("pool_code_split", None)
                else:
                    orig_entry["pool_code_split"] = _val

        if all_errors:
            for msg in all_errors:
                flash(msg, "error")
            _save_state(state)
            return render_template(
                "setup/step3_columns.html",
                files_view=[_file_view(e, i) for i, e in enumerate(files)],
                **_wizard_ctx("columns"),
            )

        # Mirror the first file's mapping into top-level state for
        # back-compat with downstream services / YAML writers.
        first = new_entries[0][1]
        state["column_mappings"] = dict(first["column_mappings"])
        state["member_account"] = dict(first["member_account"])
        state["has_header"] = bool(first["has_header"])
        ma_mode = first["member_account"].get("mode", "fixed_suffix")
        ma_suffix_len = int(first["member_account"].get("suffix_length", 3))
        state["account_suffix_length"] = (
            ma_suffix_len if ma_mode == "fixed_suffix" else 0
        )

        # Cross-CU learning: record each file's mapping so future credit
        # unions with similar headers get smarter defaults.
        for _orig, upd in new_entries:
            try:
                column_mapping_suggestions.record_mapping(upd["column_mappings"])
            except Exception:  # noqa: BLE001 - non-fatal
                pass

        # Pool-code re-derivation: rebuild state["pool_map"] as the union
        # of distinct pool codes across ALL loan-data extracts using each
        # file's CURRENT loan_pool_code column. Preserves existing pool
        # assignments for codes that survive the rebuild, drops codes
        # that no longer appear in any extract, and seeds new codes with
        # an empty pool name so they surface in the Loan Code Mapping
        # step (Step 3) AND the Balance Adjustment step (Step 14) for
        # the user to map.
        #
        # Each entry is re-scanned only when its loan_pool_code column
        # changed since the entry's cached column_suggestions snapshot;
        # unchanged entries reuse their cached pool_code_suggestions so
        # the per-file IO cost stays bounded.
        old_pool_map = dict(state.get("pool_map") or {})
        codes_by_file: list[tuple[str, list[str], bool]] = []
        rescan_failures: list[str] = []
        for orig_entry, updates in new_entries:
            cm = updates.get("column_mappings") or {}
            new_pool_col = (cm.get("loan_pool_code") or "").strip()
            saved = orig_entry.get("path")
            label = orig_entry.get("name") or "(unnamed file)"
            analysis = orig_entry.get("analysis") or {}
            cached_suggestions = list(
                analysis.get("pool_code_suggestions") or []
            )
            cached_col = (
                (analysis.get("column_suggestions") or {}).get(
                    "loan_pool_code"
                )
                or ""
            ).strip()

            if not new_pool_col:
                # Static-code-only entries contribute their static value
                # via the static-code loop below; nothing to scan here.
                codes_by_file.append((label, [], False))
                continue

            col_changed = new_pool_col != cached_col
            need_rescan = col_changed or not cached_suggestions
            if not need_rescan:
                codes_by_file.append((label, cached_suggestions, False))
                continue

            if not saved or not Path(saved).exists():
                # Staged file is gone from %TEMP% (e.g. server restart
                # since upload). Best we can do is keep the cached
                # suggestions so we don't silently lose codes.
                codes_by_file.append((label, cached_suggestions, False))
                continue

            try:
                fresh = sample_parser.extract_pool_codes(
                    saved,
                    column_name=new_pool_col,
                    header_row=orig_entry.get("header_row"),
                    split_char=(
                        orig_entry["pool_code_split"]
                        if "pool_code_split" in orig_entry
                        else state.get("pool_code_split", "/")
                    ),
                )
            except Exception:  # noqa: BLE001
                fresh = []
                rescan_failures.append(label)

            if fresh:
                analysis["pool_code_suggestions"] = fresh
            analysis.setdefault("column_suggestions", {})[
                "loan_pool_code"
            ] = new_pool_col
            orig_entry["analysis"] = analysis
            codes_by_file.append((label, fresh, True))

        # Union all codes across files. Preserves first-seen order for
        # determinism in flash messages (dict insertion order matches).
        seen_codes: list[str] = []
        seen_set: set[str] = set()
        for _label, codes, _ in codes_by_file:
            for c in codes:
                if c and c not in seen_set:
                    seen_set.add(c)
                    seen_codes.append(c)

        any_rescanned = any(rescanned for _l, _c, rescanned in codes_by_file)

        if seen_codes:
            # Rebuild pool_map preserving prior pool assignments for
            # codes that survive. Codes new to this rebuild get an
            # empty pool name so they show up on Steps 3/14.
            new_pool_map = {
                code: old_pool_map.get(code, "") for code in seen_codes
            }
            state["pool_map"] = new_pool_map

            # Mirror first file's fresh data to top-level state["sample"]
            # (downstream YAML writers / legacy services still read it).
            first_entry = new_entries[0][0]
            first_analysis = first_entry.get("analysis") or {}
            first_fresh = list(first_analysis.get("pool_code_suggestions") or [])
            first_pool_col = (
                (first_analysis.get("column_suggestions") or {}).get(
                    "loan_pool_code"
                )
                or ""
            ).strip()
            if first_fresh:
                sample = state.get("sample") or {}
                sample["pool_code_suggestions"] = first_fresh
                if first_pool_col:
                    sample.setdefault("column_suggestions", {})[
                        "loan_pool_code"
                    ] = first_pool_col
                state["sample"] = sample

            # Summarise the rebuild for the user. Only flash when
            # something actually changed (added or dropped codes), or
            # when at least one file was re-scanned this POST.
            added = sum(1 for c in seen_codes if c not in old_pool_map)
            dropped_set = {c for c in old_pool_map if c not in seen_set}
            dropped = len(dropped_set)
            if any_rescanned and (added or dropped):
                parts = []
                if added:
                    parts.append(f"added {added} new code(s)")
                if dropped:
                    parts.append(f"dropped {dropped} stale code(s)")
                file_count = sum(
                    1 for _l, _c, rescanned in codes_by_file if rescanned
                )
                flash(
                    f"Pool-code mapping rebuilt from {file_count} re-scanned "
                    "extract(s): " + ", ".join(parts) + f" (now {len(seen_codes)} "
                    "code(s) in scope across all extracts). Visit the "
                    "<strong>Loan Code Mapping</strong> step to assign "
                    "pools to any new codes.",
                    "info",
                )

        if rescan_failures:
            flash(
                "Couldn't re-scan pool codes for: "
                + ", ".join(f"<code>{n}</code>" for n in rescan_failures)
                + ". The staged file may have been moved or is locked. "
                "Re-upload it on the Sample step if needed.",
                "warning",
            )

        # Warn (non-fatal) if multiple files share the same pattern — they
        # will collide at import time. Two files with no pattern is fine
        # only when the top-level file_pattern (Step 8) routes them; we
        # can't tell here, so just info.
        pat_counts: dict[str, int] = {}
        for _orig, upd in new_entries:
            p = upd.get("file_pattern") or ""
            if p:
                pat_counts[p] = pat_counts.get(p, 0) + 1
        dupe_pats = [p for p, n in pat_counts.items() if n > 1]
        if dupe_pats:
            flash(
                "Heads-up &mdash; two or more loan-data extracts share the "
                "same monthly filename pattern: "
                + ", ".join(f"<code>{p}</code>" for p in dupe_pats)
                + ". Each file in the import folder will be routed to the "
                "FIRST matching mapping. Make the patterns more specific "
                "if you want them routed to different mappings.",
                "warning",
            )

        # Seed pool_map with any per-file "fixed loan pool code" values so
        # the Loan Code Mapping step (Step 4) lists them alongside any
        # column-derived codes.
        static_codes = []
        for _orig, upd in new_entries:
            sc = (upd["column_mappings"] or {}).get("loan_pool_code_static")
            if sc and sc not in static_codes:
                static_codes.append(sc)
        if static_codes:
            pmap = state.get("pool_map") or {}
            added = []
            for code in static_codes:
                if code not in pmap:
                    pmap[code] = ""
                    added.append(code)
            state["pool_map"] = pmap
            if added:
                flash(
                    "Added fixed loan code(s) to the Loan Code Mapping list: "
                    + ", ".join(added)
                    + ". Map them to a pool on the next step.",
                    "info",
                )

        _save_state(state)
        return redirect(url_for("setup.step_balance_check"))

    # GET
    _hydrate_legacy_entries()
    files = _files_list()
    # Self-heal: drafts saved before sample_parser._clean_header normalised
    # internal whitespace may carry headers like "Current \nLoan Bal" or
    # "Credit\nScore" inside `analysis.headers` AND in column_mappings.
    # Browsers normalise whitespace inside <option value="..."> on form
    # submit, so saved values won't match the rendered options after
    # POST → dropdown appears to reset on every save. Rewrite both lists
    # in-place to the normalised shape.
    #
    # Additionally: drafts predating the ``col_<LETTER>`` placeholder for
    # blank header cells stored the literal string ``nan`` (or pandas'
    # ``Unnamed: N``) for every unlabelled column. All blanks collapsing
    # to the same option made the dropdown unable to distinguish between
    # them — every pick snapped to the leftmost blank. Rewrite blanks to
    # ``col_A``/``col_B``/... using the column index and clear any saved
    # mapping value that is ambiguous (literal ``nan`` / ``Unnamed: N``).
    _ws_re = re.compile(r"\s+")
    _unnamed_rx = re.compile(r"^unnamed:\s*\d+(?:_level_\d+)*$", re.IGNORECASE)

    def _norm(s: object) -> str:
        return _ws_re.sub(" ", str(s)).strip()

    def _is_blank_header_value(s: str) -> bool:
        if not s:
            return True
        low = s.lower()
        return low == "nan" or bool(_unnamed_rx.match(s))

    def _col_letter(i: int) -> str:
        if i < 0:
            return ""
        out = ""
        n = i
        while True:
            out = chr(ord("A") + (n % 26)) + out
            n = n // 26 - 1
            if n < 0:
                break
        return out

    healed = False
    for entry in files:
        analysis = entry.get("analysis") or {}
        old_headers = list(analysis.get("headers") or [])
        new_headers: list[str] = []
        for i, h in enumerate(old_headers):
            n = _norm(h) if h else ""
            if _is_blank_header_value(n):
                new_headers.append(f"col_{_col_letter(i)}")
            else:
                new_headers.append(n)
        if new_headers != old_headers:
            analysis["headers"] = new_headers
            entry["analysis"] = analysis
            healed = True
        # Normalise per-file column_mappings values. Clear any saved value
        # that is ambiguous (literal nan / Unnamed: N) — the user will
        # re-pick from the now-disambiguated dropdown.
        cm = entry.get("column_mappings") or {}
        for fld, val in list(cm.items()):
            if not val or fld == "loan_pool_code_static":
                continue
            n = _norm(val)
            if _is_blank_header_value(n):
                cm[fld] = ""
                healed = True
            elif n != val:
                cm[fld] = n
                healed = True
        # Also normalise column_suggestions values so seeding stays in sync.
        sug = analysis.get("column_suggestions") or {}
        for fld, val in list(sug.items()):
            if not val:
                continue
            n = _norm(val)
            if _is_blank_header_value(n):
                sug[fld] = ""
                healed = True
            elif n != val:
                sug[fld] = n
                healed = True
    # Mirror the heal into top-level state.column_mappings (used by
    # downstream services / YAML emitters / Step-4 lookups).
    top_map = state.get("column_mappings") or {}
    for fld, val in list(top_map.items()):
        if not val or fld == "loan_pool_code_static":
            continue
        n = _norm(val)
        if _is_blank_header_value(n):
            top_map[fld] = ""
            healed = True
        elif n != val:
            top_map[fld] = n
            healed = True
    # Also heal state.sample.headers (Step 2 upload snapshot) so any
    # later code that reads it sees the disambiguated shape.
    samp = state.get("sample") or {}
    samp_headers = list(samp.get("headers") or [])
    if samp_headers:
        new_samp: list[str] = []
        for i, h in enumerate(samp_headers):
            n = _norm(h) if h else ""
            if _is_blank_header_value(n):
                new_samp.append(f"col_{_col_letter(i)}")
            else:
                new_samp.append(n)
        if new_samp != samp_headers:
            samp["headers"] = new_samp
            state["sample"] = samp
            healed = True
    if healed:
        _save_state(state)
    return render_template(
        "setup/step3_columns.html",
        files_view=[_file_view(e, i) for i, e in enumerate(files)],
        files_used=_loan_extract_files_used(state),
        **_wizard_ctx("columns"),
    )


# =================================================================
# Step 4 — Pool map
# =================================================================
@setup_bp.route("/step/pools", methods=["GET", "POST"])
def step4_pools():
    state = _state()

    # Build a case-insensitive lookup of the canonical pool names the user
    # configured on Step 2 (pool_settings) / WARM. Used so that an uploaded
    # Loan Code Map whose pool column contains e.g. "AUTO - NEW" is matched
    # against the canonical "Auto - New" instead of being treated as a new
    # custom pool. Includes existing pool_map values as a final fallback.
    def _build_canon_pool_map() -> dict[str, str]:
        canon: dict[str, str] = {}
        for ps in (state.get("pool_settings") or []):
            n = ((ps or {}).get("name") or "").strip()
            if n and n.lower() not in canon:
                canon[n.lower()] = n
        for n in ((state.get("warm") or {}).get("pools") or []):
            s = (n or "").strip() if isinstance(n, str) else ""
            if s and s.lower() not in canon:
                canon[s.lower()] = s
        for v in (state.get("pool_map") or {}).values():
            s = (v or "").strip() if isinstance(v, str) else ""
            if s and s.lower() not in canon and s not in ("Ignore", "Exclude"):
                canon[s.lower()] = s
        return canon

    def _canon_pool_name(name: str, canon: dict[str, str]) -> str:
        """Snap ``name`` to its canonical-cased form if a case-insensitive
        match exists; otherwise return the original string unchanged."""
        if not name or not isinstance(name, str):
            return name
        return canon.get(name.strip().lower(), name)

    if request.method == "POST":
        action = request.form.get("action", "save")

        if action == "add_codes":
            # Pull raw codes from the sample analysis that aren't already in
            # the pool_map and add them as empty rows the user can name.
            sample = state.get("sample") or {}
            sample_codes = sample.get("pool_code_suggestions") or []
            existing = set(state.get("pool_map", {}).keys())
            added = []
            for code in sample_codes:
                if code and code not in existing:
                    state["pool_map"][code] = ""
                    added.append(code)
            if added:
                _save_state(state)
                flash(
                    f"Added {len(added)} unmapped raw code(s) from the sample: "
                    + ", ".join(added),
                    "success",
                )
            else:
                flash("No new raw codes to add &mdash; all sample codes are "
                      "already in the map.", "info")
            return redirect(url_for("setup.step4_pools"))

        # ---- Clear pool mappings (undo a polluted upload) ----------
        # Wipes all values in state["pool_map"] back to empty so the
        # user can re-map from scratch. Drops any staged upload review
        # AND deletes saved poolmap_*.xlsx files in the samples dir
        # so a stale file from a prior CU can't leak again.
        if action == "clear_pool_map":
            scope = (request.form.get("clear_scope") or "values").strip().lower()
            pool_map = state.get("pool_map") or {}
            cleared_count = 0
            if scope == "all":
                # Drop every code -> pool entry entirely
                cleared_count = len(pool_map)
                state["pool_map"] = {}
            else:
                # Keep the codes (so the user doesn't lose the list);
                # just blank the pool-name values.
                for code in list(pool_map.keys()):
                    if pool_map.get(code):
                        cleared_count += 1
                    pool_map[code] = ""
                state["pool_map"] = pool_map
            # Drop staged upload review + saved poolmap files in TEMP
            staged = state.pop("pool_map_upload", None)
            removed_files: list[str] = []
            try:
                if _SAMPLE_DIR.exists():
                    for p in _SAMPLE_DIR.glob("poolmap_*.xlsx"):
                        try:
                            p.unlink()
                            removed_files.append(p.name)
                        except OSError:
                            pass
                    for p in _SAMPLE_DIR.glob("poolmap_*.xls"):
                        try:
                            p.unlink()
                            removed_files.append(p.name)
                        except OSError:
                            pass
                    for p in _SAMPLE_DIR.glob("poolmap_*.csv"):
                        try:
                            p.unlink()
                            removed_files.append(p.name)
                        except OSError:
                            pass
            except Exception:  # noqa: BLE001
                pass
            _save_state(state)
            parts = []
            if cleared_count:
                if scope == "all":
                    parts.append(
                        f"removed {cleared_count} code &rarr; pool entries"
                    )
                else:
                    parts.append(
                        f"cleared pool name on {cleared_count} code(s)"
                    )
            if staged:
                parts.append("discarded staged upload review")
            if removed_files:
                parts.append(
                    f"deleted {len(removed_files)} cached upload file(s)"
                )
            if parts:
                flash(
                    "Pool mappings reset: " + ", ".join(parts) + ".",
                    "success",
                )
            else:
                flash("Nothing to clear &mdash; pool map was already empty.",
                      "info")
            return redirect(url_for("setup.step4_pools"))

        # ---- Upload a code -> pool-name map file -------------------
        if action == "upload_map":
            f = request.files.get("pool_map_file")
            if not f or not f.filename:
                flash("Please choose a file to upload.", "error")
                return redirect(url_for("setup.step4_pools"))
            try:
                _SAMPLE_DIR.mkdir(parents=True, exist_ok=True)
                fn = secure_filename(f.filename)
                target = _SAMPLE_DIR / f"poolmap_{fn}"
                f.save(target)
                parsed = sample_parser.parse_pool_map_file(target)
            except Exception as exc:  # noqa: BLE001
                flash(f"Could not parse pool-map file: {exc}", "error")
                return redirect(url_for("setup.step4_pools"))

            if not parsed["rows"]:
                flash("Pool-map file had no usable rows after the header.",
                      "error")
                return redirect(url_for("setup.step4_pools"))

            # Stash for the review panel; cleared on apply/dismiss.
            state["pool_map_upload"] = {
                "filename": fn,
                "saved_path": str(target),
                "code_column": parsed["code_column"],
                "name_column": parsed["name_column"],
                "header_row": parsed.get("header_row", 1),
                "headers": parsed["headers"],
                "rows": parsed["rows"],
            }
            _save_state(state)
            flash(
                f"Loaded {parsed['row_count']} mappings from "
                f"<code>{fn}</code> &mdash; review and apply below.",
                "success",
            )
            return redirect(url_for("setup.step4_pools"))

        if action == "dismiss_upload":
            state.pop("pool_map_upload", None)
            _save_state(state)
            return redirect(url_for("setup.step4_pools"))

        # ---- Re-parse the staged upload with user-picked columns ----
        # ``apply_header_row``  - reread file at chosen row, auto-detect
        #                        code / name columns from the new
        #                        header row (clears stale picks).
        # ``rechoose_columns``  - reread file at chosen row, honor the
        #                        user-picked code_col / name_col.
        if action in ("rechoose_columns", "apply_header_row"):
            upload = state.get("pool_map_upload") or {}
            saved_path = upload.get("saved_path")
            if not saved_path or not Path(saved_path).exists():
                flash("No staged upload to re-parse \u2014 please upload again.",
                      "error")
                state.pop("pool_map_upload", None)
                _save_state(state)
                return redirect(url_for("setup.step4_pools"))
            hdr_raw = (request.form.get("header_row") or "").strip()
            try:
                header_row = int(hdr_raw) if hdr_raw else None
            except ValueError:
                header_row = None
            if header_row is not None and header_row < 1:
                header_row = 1
            if action == "apply_header_row":
                # Force auto-detection so the dropdowns repopulate from
                # the new header row.
                code_col = None
                name_col = None
            else:
                code_col = (request.form.get("code_col") or "").strip() or None
                name_col = (request.form.get("name_col") or "").strip() or None
                if code_col and name_col and code_col == name_col:
                    flash(
                        "Code column and pool-name column must be different.",
                        "error",
                    )
                    return redirect(url_for("setup.step4_pools"))
            try:
                parsed = sample_parser.parse_pool_map_file(
                    Path(saved_path),
                    code_col=code_col,
                    name_col=name_col,
                    header_row=header_row,
                )
            except ValueError as exc:
                # The chosen row didn't yield 2+ usable columns via
                # pandas. Salvage the raw cells of that row so the
                # user can still adjust selections.
                if header_row:
                    raw_cells = sample_parser._read_row_cells(
                        Path(saved_path), header_row
                    )
                    raw_headers = [c for c in raw_cells if c]
                else:
                    raw_headers = []
                if raw_headers:
                    state["pool_map_upload"] = {
                        "filename": upload.get("filename"),
                        "saved_path": saved_path,
                        "code_column": (
                            raw_headers[0] if raw_headers else ""
                        ),
                        "name_column": (
                            raw_headers[1] if len(raw_headers) > 1 else ""
                        ),
                        "header_row": header_row or 1,
                        "headers": raw_headers,
                        "rows": [],
                    }
                    _save_state(state)
                flash(
                    f"Header row {header_row or '(auto)'} doesn't look "
                    f"like a usable header: {exc}",
                    "error",
                )
                return redirect(url_for("setup.step4_pools"))
            except Exception as exc:  # noqa: BLE001
                flash(f"Could not re-parse pool-map file: {exc}", "error")
                return redirect(url_for("setup.step4_pools"))
            if not parsed["rows"]:
                # Still keep the new headers so the dropdowns refresh
                # even if the row had no usable data below it.
                row_for_cells = parsed.get("header_row", 1)
                raw_cells = sample_parser._read_row_cells(
                    Path(saved_path), row_for_cells
                )
                merged_headers = list(parsed["headers"])
                for cell in raw_cells:
                    if cell and cell not in merged_headers:
                        merged_headers.append(cell)
                state["pool_map_upload"] = {
                    "filename": upload.get("filename"),
                    "saved_path": saved_path,
                    "code_column": parsed["code_column"],
                    "name_column": parsed["name_column"],
                    "header_row": parsed.get("header_row", 1),
                    "headers": merged_headers,
                    "rows": [],
                }
                _save_state(state)
                flash(
                    f"Header row {parsed.get('header_row', 1)} loaded "
                    f"({len(parsed['headers'])} column(s)) but no data "
                    "rows below it &mdash; pick a different header row "
                    "or the right code / name columns.",
                    "warning",
                )
                return redirect(url_for("setup.step4_pools"))
            # Always pull the raw cell values of the chosen header
            # row directly from the file so the dropdowns have every
            # column visible (pandas can collapse blank trailing cols
            # or rename duplicates in ways that hide picks).
            row_for_cells = parsed.get("header_row", 1)
            raw_cells = sample_parser._read_row_cells(
                Path(saved_path), row_for_cells
            )
            merged_headers = list(parsed["headers"])
            for cell in raw_cells:
                if cell and cell not in merged_headers:
                    merged_headers.append(cell)
            state["pool_map_upload"] = {
                "filename": upload.get("filename"),
                "saved_path": saved_path,
                "code_column": parsed["code_column"],
                "name_column": parsed["name_column"],
                "header_row": parsed.get("header_row", 1),
                "headers": merged_headers,
                "rows": parsed["rows"],
            }
            _save_state(state)
            flash(
                f"Re-parsed (header row {parsed.get('header_row', 1)}): "
                f"code = <code>{parsed['code_column']}</code>, "
                f"name = <code>{parsed['name_column']}</code> &mdash; "
                f"{parsed['row_count']} mapping(s) ready to review.",
                "success",
            )
            return redirect(url_for("setup.step4_pools"))

        # ---- Resolve "unrecognized" pool names ---------------------
        # Each row in the extras panel posts an ``extra_name`` hidden
        # field, an ``extra_choice_<i>`` radio (ignore|create|merge),
        # and an ``extra_merge_<i>`` <select> giving the target Step-2
        # pool when the choice is "merge".
        #   - Ignore:  rewrite every pool_map entry mapping to that
        #              name to the "Ignore" sentinel.
        #   - Create:  append a new entry to ``pool_settings`` and
        #              redirect to Step 2 so the user can set ACL Months.
        #   - Merge:   rewrite every pool_map entry mapping to that
        #              name to the chosen existing Step-2 pool.
        if action == "resolve_extras":
            names = request.form.getlist("extra_name")
            pool_map = dict(state.get("pool_map") or {})
            pool_settings = list(state.get("pool_settings") or [])
            existing_names_lower = {
                ((p or {}).get("name") or "").strip().lower(): (
                    ((p or {}).get("name") or "").strip()
                )
                for p in pool_settings
                if ((p or {}).get("name") or "").strip()
            }
            created: list[str] = []
            ignored: list[str] = []
            merged: list[str] = []
            for i, raw in enumerate(names):
                nm = (raw or "").strip()
                if not nm:
                    continue
                choice = (request.form.get(
                    f"extra_choice_{i}", "ignore"
                ) or "ignore").strip().lower()
                if choice == "create":
                    if nm.lower() not in existing_names_lower:
                        pool_settings.append({
                            "name": nm,
                            "risk_rated": False,
                            "brr": False,
                            "acl_months": 0,
                            "use_default_mgmt_adj": False,
                            "excluded": False,
                        })
                        existing_names_lower[nm.lower()] = nm
                        created.append(nm)
                elif choice == "merge":
                    target_raw = (request.form.get(
                        f"extra_merge_{i}", ""
                    ) or "").strip()
                    target = existing_names_lower.get(target_raw.lower())
                    if not target:
                        flash(
                            f"Cannot merge <strong>{nm}</strong>: please "
                            "pick an existing Step 2 pool from the "
                            "dropdown.",
                            "error",
                        )
                        continue
                    changed = False
                    for c, n in list(pool_map.items()):
                        if (
                            isinstance(n, str)
                            and n.strip().lower() == nm.lower()
                        ):
                            pool_map[c] = target
                            changed = True
                    if changed:
                        merged.append(f"{nm} &rarr; {target}")
                else:  # ignore (default)
                    changed = False
                    for c, n in list(pool_map.items()):
                        if (
                            isinstance(n, str)
                            and n.strip().lower() == nm.lower()
                        ):
                            pool_map[c] = "Ignore"
                            changed = True
                    if changed:
                        ignored.append(nm)
            state["pool_map"] = pool_map
            state["pool_settings"] = pool_settings
            _save_state(state)
            if ignored:
                flash(
                    f"Set {len(ignored)} pool name(s) to <strong>Ignore</strong>: "
                    + ", ".join(ignored)
                    + ".",
                    "success",
                )
            if merged:
                flash(
                    f"Merged {len(merged)} pool name(s) into existing "
                    "Step 2 pools: " + ", ".join(merged) + ".",
                    "success",
                )
            if created:
                flash(
                    f"Added {len(created)} new pool(s) to Step 2: "
                    + ", ".join(created)
                    + ". Please set the <strong>ACL Months</strong> for "
                    "each (and any other settings) before continuing.",
                    "info",
                )
                if state.get("has_warm_files") == "yes":
                    return redirect(url_for("setup.step2_warm"))
                return redirect(url_for("setup.step_loan_pools"))
            if not ignored and not merged:
                flash(
                    "No unrecognized pool names were selected for action.",
                    "info",
                )
            return redirect(url_for("setup.step4_pools"))

        # ---- Apply Off-loan-file groups (from upload review) -------
        # Codes that the uploaded Loan Code Map references but that
        # don't appear in the Loan Data Extract are grouped by their
        # pool name. Each group posts:
        #   off_name        - hidden, pool name from the file
        #   off_codes_<i>   - hidden, comma-joined list of codes
        #   off_choice_<i>  - radio: ignore | merge | create
        #   off_merge_<i>   - <select>, target Step-2 pool (merge only)
        if action == "apply_offfile_groups":
            upload = state.get("pool_map_upload") or {}
            names = request.form.getlist("off_name")
            pool_map = dict(state.get("pool_map") or {})
            pool_settings = list(state.get("pool_settings") or [])
            existing_names_lower = {
                ((p or {}).get("name") or "").strip().lower(): (
                    ((p or {}).get("name") or "").strip()
                )
                for p in pool_settings
                if ((p or {}).get("name") or "").strip()
            }
            ignored: list[str] = []
            merged: list[str] = []
            created: list[str] = []
            processed_codes: set[str] = set()
            for i, raw in enumerate(names):
                nm = (raw or "").strip()
                if not nm:
                    continue
                codes_raw = request.form.get(f"off_codes_{i}", "") or ""
                codes = [c.strip() for c in codes_raw.split(",") if c.strip()]
                if not codes:
                    continue
                choice = (request.form.get(
                    f"off_choice_{i}", "ignore"
                ) or "ignore").strip().lower()
                target_name: str | None = None
                if choice == "create":
                    if nm.lower() not in existing_names_lower:
                        pool_settings.append({
                            "name": nm,
                            "risk_rated": False,
                            "brr": False,
                            "acl_months": 0,
                            "use_default_mgmt_adj": False,
                            "excluded": False,
                        })
                        existing_names_lower[nm.lower()] = nm
                        created.append(nm)
                    target_name = existing_names_lower[nm.lower()]
                elif choice == "merge":
                    tgt_raw = (request.form.get(
                        f"off_merge_{i}", ""
                    ) or "").strip()
                    target_name = existing_names_lower.get(tgt_raw.lower())
                    if not target_name:
                        flash(
                            f"Cannot merge <strong>{nm}</strong>: please "
                            "pick an existing Step 2 pool.",
                            "error",
                        )
                        continue
                    merged.append(f"{nm} &rarr; {target_name} ({len(codes)})")
                else:  # ignore (default)
                    target_name = "Ignore"
                    ignored.append(f"{nm} ({len(codes)})")
                for c in codes:
                    pool_map[c] = target_name
                    processed_codes.add(c)

            # Remove processed rows from the staged upload so they
            # don't reappear on the next page load.
            if processed_codes and upload.get("rows"):
                kept = [
                    r for r in upload["rows"]
                    if (r.get("code") or "").strip() not in processed_codes
                ]
                if kept:
                    upload["rows"] = kept
                    state["pool_map_upload"] = upload
                else:
                    state.pop("pool_map_upload", None)

            state["pool_map"] = pool_map
            state["pool_settings"] = pool_settings
            _save_state(state)

            if ignored:
                flash(
                    f"Ignored {len(ignored)} off-file pool group(s): "
                    + ", ".join(ignored) + ".",
                    "success",
                )
            if merged:
                flash(
                    f"Merged {len(merged)} off-file pool group(s) into "
                    "existing Step 2 pools: " + ", ".join(merged) + ".",
                    "success",
                )
            if created:
                flash(
                    f"Added {len(created)} new pool(s) to Step 2 from "
                    "off-file groups: " + ", ".join(created)
                    + ". Please set <strong>ACL Months</strong> and "
                    "other settings before continuing.",
                    "info",
                )
                if state.get("has_warm_files") == "yes":
                    return redirect(url_for("setup.step2_warm"))
                return redirect(url_for("setup.step_loan_pools"))
            if not ignored and not merged:
                flash("No off-file groups were selected.", "info")
            return redirect(url_for("setup.step4_pools"))

        if action == "apply_upload":
            upload = state.get("pool_map_upload") or {}
            rows = upload.get("rows") or []
            picks = set(request.form.getlist("apply_code"))
            applied_new = 0
            applied_update = 0
            applied_added_unmapped = 0
            sample_codes = set(
                ((state.get("sample") or {}).get("pool_code_suggestions"))
                or []
            )
            pool_map = state.get("pool_map") or {}
            canon = _build_canon_pool_map()
            for r in rows:
                code = r.get("code", "")
                name = _canon_pool_name(r.get("name", ""), canon)
                if not code or code not in picks:
                    continue
                existing_name = pool_map.get(code)
                if code in pool_map:
                    if existing_name != name:
                        pool_map[code] = name
                        applied_update += 1
                else:
                    pool_map[code] = name
                    if code in sample_codes:
                        applied_new += 1
                    else:
                        applied_added_unmapped += 1
            state["pool_map"] = pool_map
            state.pop("pool_map_upload", None)
            _save_state(state)
            parts = []
            if applied_update:
                parts.append(f"{applied_update} updated")
            if applied_new:
                parts.append(f"{applied_new} new (in loan file)")
            if applied_added_unmapped:
                parts.append(
                    f"{applied_added_unmapped} new (not in loan file &mdash; "
                    "may show up in charge-offs/recoveries/impaired)"
                )
            if parts:
                flash("Applied: " + ", ".join(parts) + ".", "success")
            else:
                flash("No mappings selected.", "info")
            return redirect(url_for("setup.step4_pools"))

        # Default: save and advance.  Also handles the "Save progress" /
        # "Save progress & exit" buttons in the stepper, which submit this
        # same form (via HTML5 form="step-form") so any in-DOM edits the
        # user made are persisted instead of being silently dropped.
        new_map = _parse_kv_rows(request.form, "pool_code", "pool_name")
        state["pool_map"] = new_map
        state["default_pool"] = request.form.get(
            "default_pool", "Ignore"
        ).strip() or "Ignore"
        _save_state(state)

        if action in ("save_progress_stay", "save_progress_exit"):
            # _save_state above already auto-persisted the draft; just
            # mirror the /save-progress UX for these two action values.
            label = state.get("credit_union") or state.get("short_name") or "draft"
            flash(f"Saved progress for {label}.", "success")
            if action == "save_progress_exit":
                session.pop(STATE_KEY, None)
                return redirect(url_for("home.index"))
            return redirect(url_for("setup.step4_pools"))

        # Soft-validate names against the WARM pool list (warning only).
        warm_pools = ((state.get("warm") or {}).get("pools")) or []
        if warm_pools:
            warm_set = set(warm_pools)
            unmapped = [code for code, name in new_map.items() if not name]
            mismatched = [
                f"{code} -> {name}"
                for code, name in new_map.items()
                if name and name not in warm_set
                and name != state["default_pool"]
            ]
            if unmapped:
                flash(
                    f"{len(unmapped)} pool code(s) have no name yet: "
                    + ", ".join(unmapped[:10])
                    + (" ..." if len(unmapped) > 10 else "")
                    + ". They'll fall through to the default pool.",
                    "info",
                )
            if mismatched:
                flash(
                    "These pool names don't match any name in the WARM workbook "
                    "&mdash; historical data won't line up unless that's intentional: "
                    + "; ".join(mismatched[:6])
                    + (" ..." if len(mismatched) > 6 else ""),
                    "info",
                )

        # Loan Code Mapping is now wizard step 3.  WARM users continue
        # to Balance Titles (step3_balances); no-WARM users continue
        # to Historical Balances (step3_historical).
        if state.get("has_warm_files") == "yes":
            return redirect(url_for("setup.step3_balances"))
        return redirect(url_for("setup.step3_historical"))
    sample_codes = ((state.get("sample") or {}).get("pool_code_suggestions")) or []
    known = set((state.get("pool_map") or {}).keys())
    unmapped_sample = [c for c in sample_codes if c and c not in known]

    # If a pool-map upload is staged for review, classify each row so the
    # template can group them and pick sensible default checkboxes.
    upload_review = None
    upload = state.get("pool_map_upload")
    if upload:
        sample_set = set(c for c in sample_codes if c)
        pool_map = state.get("pool_map") or {}
        canon = _build_canon_pool_map()
        items = []
        for r in upload.get("rows") or []:
            code = r.get("code", "")
            # Snap the file's pool name to its canonical case (e.g.
            # "AUTO - NEW" -> "Auto - New") so case-only differences
            # against existing pool_map entries register as matches
            # instead of conflicts / new pools.
            name = _canon_pool_name(r.get("name", ""), canon)
            in_map = code in pool_map
            existing_name = pool_map.get(code, "")
            in_loan_file = code in sample_set
            if in_map and existing_name == name:
                status = "match"
            elif in_map and existing_name and existing_name != name:
                status = "conflict"
            elif in_map and not existing_name:
                status = "fill"
            elif in_loan_file:
                status = "new_in_loans"
            else:
                status = "new_off_loans"
            # Default-on for everything except fully off-loan-file (let user
            # opt in) and exact matches (no-op).
            default_checked = status not in ("match", "new_off_loans")
            items.append({
                "code": code,
                "name": name,
                "existing_name": existing_name,
                "status": status,
                "default_checked": default_checked,
            })
        upload_review = {
            "filename": upload.get("filename"),
            "code_column": upload.get("code_column"),
            "name_column": upload.get("name_column"),
            "entries": items,
            "counts": {
                "match": sum(1 for i in items if i["status"] == "match"),
                "conflict": sum(1 for i in items if i["status"] == "conflict"),
                "fill": sum(1 for i in items if i["status"] == "fill"),
                "new_in_loans": sum(1 for i in items if i["status"] == "new_in_loans"),
                "new_off_loans": sum(1 for i in items if i["status"] == "new_off_loans"),
            },
        }

        # Off-loan-file groups: bucket every upload row whose code is
        # NOT in the loan-data extract by its (canonicalized) pool name
        # from the file. The user gets one row per group with the same
        # Ignore / Merge / Create choices used by the extras panel.
        _off_groups: dict[str, list[str]] = {}
        for it in items:
            if it["status"] != "new_off_loans":
                continue
            nm = (it.get("name") or "").strip()
            cd = (it.get("code") or "").strip()
            if not nm or not cd:
                continue
            bucket = _off_groups.setdefault(nm, [])
            if cd not in bucket:
                bucket.append(cd)
        upload_review["off_file_groups"] = [
            {"name": k, "codes": sorted(v), "count": len(v)}
            for k, v in sorted(_off_groups.items(), key=lambda kv: kv[0].lower())
        ]

    # Established pool names: STRICTLY the names configured on Step 2
    # (``pool_settings``). "Ignore" is always offered separately by the
    # template. Anything in ``pool_map`` whose value isn't here is
    # surfaced to the user via the ``extras`` panel below for explicit
    # resolution (Ignore vs. Create new pool on Step 2).
    pool_choices: list[str] = []
    _seen: set[str] = set()
    for ps in (state.get("pool_settings") or []):
        s = ((ps or {}).get("name") or "").strip()
        if not s or s.lower() in _seen or s in ("Ignore", "Exclude"):
            continue
        _seen.add(s.lower())
        pool_choices.append(s)

    # Pool names referenced in pool_map (or the default_pool) that are
    # NOT configured on Step 2 and aren't the "Ignore" sentinel. The
    # template renders one row per extra with radio buttons so the user
    # can either drop those codes (Ignore) or promote the name to a new
    # Step-2 pool (Create new loan pool).
    _pool_lower = {p.lower() for p in pool_choices}
    _extras_map: dict[str, list[str]] = {}
    for _code, _name in (state.get("pool_map") or {}).items():
        s = (_name or "").strip() if isinstance(_name, str) else ""
        if not s or s in ("Ignore", "Exclude"):
            continue
        if s.lower() in _pool_lower:
            continue
        _extras_map.setdefault(s, []).append(_code or "")
    extras = [
        {"name": k, "codes": sorted(v), "count": len(v)}
        for k, v in sorted(_extras_map.items(), key=lambda kv: kv[0].lower())
    ]

    return render_template(
        "setup/step4_pools.html",
        unmapped_sample=unmapped_sample,
        upload_review=upload_review,
        pool_choices=pool_choices,
        extras=extras,
        **_wizard_ctx("pools"),
    )


# =================================================================
# Step — Balance Adjustment (reconciliation)
# =================================================================
@setup_bp.route("/step/balance-check", methods=["GET", "POST"])
def step_balance_check():
    """Side-by-side reconciliation of per-pool balances.

    Compares the latest period in the Monthly Balance file against the sum
    of ``current_balance`` from the Loan Data Extract(s), grouped by the
    same canonical pool name. Lets the user record an optional manual
    carry-forward adjustment per pool that downstream steps / the report
    engine can consume.
    """
    state = _state()
    has_warm = state.get("has_warm_files") == "yes"

    if request.method == "POST":
        action = request.form.get("action", "")

        if action == "remap_code":
            code = (request.form.get("code", "") or "").strip()
            new_pool = (request.form.get("new_pool", "") or "").strip()
            if not code:
                flash("Missing loan code.", "error")
            else:
                pmap = state.setdefault("pool_map", {})
                old = pmap.get(code, "")
                pmap[code] = new_pool
                _save_state(state)
                if new_pool:
                    flash(
                        f"Loan code {code} reassigned: "
                        f"{old or '(unmapped)'} \u2192 {new_pool}.",
                        "success",
                    )
                else:
                    flash(
                        f"Loan code {code} cleared (was {old or 'unmapped'}).",
                        "success",
                    )
            return redirect(url_for("setup.step_balance_check"))

        if action == "refresh":
            return redirect(url_for("setup.step_balance_check"))

        if action in ("save", "next"):
            if action == "next":
                return redirect(url_for("setup.step_co_recov"))
            return redirect(url_for("setup.step_balance_check"))

        if action == "back":
            return redirect(url_for("setup.step4_pools"))

    comparison = balance_check_service.compare(state)
    pool_choices = balance_check_service.canonical_pool_order(state)
    return render_template(
        "setup/step_balance_check.html",
        comparison=comparison,
        pool_choices=pool_choices,
        has_warm=has_warm,
        report_period=state.get("report_period") or "",
        **_wizard_ctx("balance_check"),
    )


# =================================================================
# Step 5 — Credit grades
# =================================================================
@setup_bp.route("/step/grades", methods=["GET", "POST"])
def step5_grades():
    state = _state()
    # Global bounds (admin-configurable). Top grade's max auto-fills to
    # ``score_ceiling``; bottom grade's min auto-fills to
    # ``score_floor``. All other maxes derive as
    # ``next_higher_grade.min - 1``.
    _admin_cfg = admin_defaults.load() or {}
    try:
        score_floor = int(_admin_cfg.get("credit_score_min", 350))
    except (TypeError, ValueError):
        score_floor = 350
    try:
        score_ceiling = int(_admin_cfg.get("credit_score_max", 900))
    except (TypeError, ValueError):
        score_ceiling = 900
    if score_floor >= score_ceiling:
        score_floor, score_ceiling = 350, 900
    if request.method == "POST":
        action = request.form.get("action", "save")

        # Reset actions: replace state grades and redirect back to grades.
        if action == "reset_defaults":
            state["credit_grades"] = [
                dict(g) for g in config_service.DEFAULT_CREDIT_GRADES
            ]
            _save_state(state)
            flash("Reset to default 6-tier credit grades.", "success")
            return redirect(url_for("setup.step5_grades"))

        if action == "reset_warm":
            warm_grades = ((state.get("warm") or {}).get("grades")) or []
            if not warm_grades:
                flash("No WARM grades available to load.", "error")
            else:
                state["credit_grades"] = [dict(g) for g in warm_grades]
                _save_state(state)
                flash(
                    f"Loaded {len(warm_grades)} grade band(s) from WARM workbook.",
                    "success",
                )
            return redirect(url_for("setup.step5_grades"))

        # Import risk tiers / FICO ranges from an uploaded CU document
        # (.docx or .xlsx). Filenames vary widely between CUs, so the analyst
        # uploads the file here; we parse it tolerantly and pre-fill the grade
        # bands for confirmation. Reserve rates are never in these documents,
        # so they're preserved (by label) or left at 0 for the user to enter.
        if action == "parse_risk_tiers":
            f = request.files.get("risk_tier_file")
            if not f or not f.filename:
                flash("Choose a risk-tier document (.docx or .xlsx) to upload.", "error")
                return redirect(url_for("setup.step5_grades"))
            dest = _SAMPLE_DIR / "risk_tiers"
            dest.mkdir(parents=True, exist_ok=True)
            fn = secure_filename(f.filename) or "risk_tiers.docx"
            target = dest / fn
            try:
                f.save(target)
                result = risk_tier_parser.parse_risk_tiers(target)
            except Exception as exc:  # noqa: BLE001
                flash(f"Could not read the uploaded document: {exc}", "error")
                return redirect(url_for("setup.step5_grades"))
            if not result.get("ok"):
                flash(result.get("error") or "Could not parse the document.", "error")
                return redirect(url_for("setup.step5_grades"))
            existing_rates = {
                (g.get("label") or "").strip().lower(): float(g.get("reserve_rate") or 0.0)
                for g in (state.get("credit_grades") or [])
            }
            tiers = sorted(
                result["tiers"],
                key=lambda t: (t["min_score"] if t["min_score"] is not None else -1),
                reverse=True,
            )
            grades: list[dict[str, Any]] = []
            for i, t in enumerate(tiers):
                mn = t.get("min_score")
                mx = t.get("max_score")
                if mx is None:
                    if i == 0:
                        mx = score_ceiling
                    else:
                        prev_min = tiers[i - 1].get("min_score")
                        mx = (prev_min - 1) if prev_min is not None else score_ceiling
                if i == len(tiers) - 1 or mn is None:
                    mn = score_floor
                grades.append({
                    "label": t["label"],
                    "min_score": mn,
                    "max_score": mx,
                    "reserve_rate": existing_rates.get(t["label"].strip().lower(), 0.0),
                })
            state["credit_grades"] = grades
            _save_state(state)
            flash(
                f"Imported {len(grades)} tier(s) from {fn}. Review the ranges "
                "and enter a reserve rate for each before saving.",
                "success",
            )
            for w in (result.get("warnings") or []):
                flash(w, "info")
            return redirect(url_for("setup.step5_grades"))

        # Save-progress hijacks: the stepper submits THIS form (via
        # ``form="step-form"``) so any in-DOM edits to the BRR table or
        # the no-score label persist. We DON'T re-validate / overwrite
        # the credit-grade bands here \u2014 partial edits to a row would
        # raise spurious errors and refuse to save the BRR changes.
        if action in ("save_progress_stay", "save_progress_exit"):
            state["no_score_label"] = request.form.get(
                "no_score_label",
                state.get("no_score_label") or "Not Reported",
            ).strip() or "Not Reported"
            state["uses_brr"] = request.form.get("uses_brr") == "on"
            brr_rows: list[dict[str, Any]] = []
            if state["uses_brr"]:
                brr_labels = request.form.getlist("brr_label")
                brr_criteria = request.form.getlist("brr_criteria")
                for lbl, crit in zip(brr_labels, brr_criteria):
                    lbl = (lbl or "").strip()
                    crit = (crit or "").strip()
                    if not lbl and not crit:
                        continue
                    brr_rows.append({"label": lbl, "criteria": crit})
            state["business_risk_ratings"] = brr_rows
            _save_state(state)
            label = (
                state.get("credit_union")
                or state.get("short_name")
                or "draft"
            )
            flash(f"Saved progress for {label}.", "success")
            if action == "save_progress_exit":
                session.pop(STATE_KEY, None)
                return redirect(url_for("home.index"))
            return redirect(url_for("setup.step5_grades"))

        # Default action = save: parse rows, validate, store.
        # Capture BRR + no_score_label up-front so an unrelated grade
        # validation error doesn't silently drop those edits.
        state["no_score_label"] = request.form.get(
            "no_score_label",
            state.get("no_score_label") or "Not Reported",
        ).strip() or "Not Reported"
        state["uses_brr"] = request.form.get("uses_brr") == "on"
        _brr_rows: list[dict[str, Any]] = []
        if state["uses_brr"]:
            for _lbl, _crit in zip(
                request.form.getlist("brr_label"),
                request.form.getlist("brr_criteria"),
            ):
                _lbl = (_lbl or "").strip()
                _crit = (_crit or "").strip()
                if not _lbl and not _crit:
                    continue
                _brr_rows.append({"label": _lbl, "criteria": _crit})
        state["business_risk_ratings"] = _brr_rows
        _save_state(state)

        labels = request.form.getlist("grade_label")
        mins = request.form.getlist("grade_min")
        # Max is derived from admin bounds + adjacent mins; the form no
        # longer collects it. Any legacy hidden ``grade_max`` values are
        # ignored on purpose.
        # Preserve any existing reserve_rate values keyed by label so we
        # don't lose data the user set in a previous (now-removed) field.
        existing_rates = {
            (g.get("label") or "").strip(): float(g.get("reserve_rate") or 0.0)
            for g in (state.get("credit_grades") or [])
        }
        # Collect (label, min) pairs, dropping wholly-empty rows.
        raw_rows: list[tuple[str, int | None]] = []
        errors: list[str] = []
        for idx, (lbl, mn) in enumerate(zip(labels, mins), start=1):
            lbl = (lbl or "").strip()
            mn_raw = (mn or "").strip()
            if not lbl and not mn_raw:
                continue  # skip totally-empty row
            if not lbl:
                errors.append(f"Row {idx}: label is required.")
                continue
            mn_val: int | None
            if mn_raw == "":
                mn_val = None
            else:
                try:
                    mn_val = int(mn_raw)
                except ValueError:
                    errors.append(
                        f"Row {idx} ({lbl}): min Credit Score must be an integer."
                    )
                    continue
            raw_rows.append((lbl, mn_val))

        # The bottom grade's min is auto-set to the admin floor; if the
        # user left the bottom row's min blank that's expected.
        if raw_rows:
            last_lbl, last_mn = raw_rows[-1]
            if last_mn is None:
                raw_rows[-1] = (last_lbl, score_floor)

        # Every other row needs a min within bounds.
        for idx, (lbl, mn_val) in enumerate(raw_rows, start=1):
            if mn_val is None:
                errors.append(
                    f"Row {idx} ({lbl}): enter a Min Credit Score "
                    "(only the bottom grade may leave it blank)."
                )
                continue
            if mn_val < score_floor or mn_val > score_ceiling:
                errors.append(
                    f"Row {idx} ({lbl}): Min Credit Score {mn_val} is "
                    f"outside the admin bounds [{score_floor}, "
                    f"{score_ceiling}]."
                )

        # Sort highest-min first (the table is rendered top-down by
        # range), then derive each row's max.
        if not errors and raw_rows:
            raw_rows.sort(key=lambda t: (t[1] if t[1] is not None else -1),
                          reverse=True)
            grades: list[dict[str, Any]] = []
            for i, (lbl, mn_val) in enumerate(raw_rows):
                if i == 0:
                    mx_val = score_ceiling
                else:
                    prev_min = raw_rows[i - 1][1]
                    mx_val = (prev_min - 1) if prev_min is not None else score_ceiling
                if i == len(raw_rows) - 1:
                    mn_val = score_floor  # enforce floor on bottom
                if mn_val is not None and mn_val > mx_val:
                    errors.append(
                        f"Row {i + 1} ({lbl}): Min {mn_val} must be at "
                        f"most {mx_val} (one less than the next-higher "
                        "grade's Min)."
                    )
                grades.append({
                    "label": lbl,
                    "min_score": mn_val if mn_val is not None else score_floor,
                    "max_score": mx_val,
                    "reserve_rate": existing_rates.get(lbl, 0.0),
                })
        else:
            grades = []

        if errors:
            for e in errors:
                flash(e, "error")
            return render_template(
                "setup/step5_grades.html",
                draft_grades=[
                    {"label": (l or "").strip(),
                     "min_score": (int(m) if (m or "").strip().lstrip("-").isdigit() else None),
                     "max_score": None,
                     "reserve_rate": existing_rates.get((l or "").strip(), 0.0)}
                    for l, m in zip(labels, mins)
                ],
                admin_min=score_floor,
                admin_max=score_ceiling,
                **_wizard_ctx("grades"),
            )

        if not grades:
            flash("Define at least one credit-grade band.", "error")
            return render_template(
                "setup/step5_grades.html",
                admin_min=score_floor,
                admin_max=score_ceiling,
                **_wizard_ctx("grades"),
            )

        state["credit_grades"] = grades
        # ``no_score_label`` / ``uses_brr`` / ``business_risk_ratings``
        # were captured at the top of this branch so partial grade-row
        # edits can't silently drop them. Re-save now that grades pass.

        _save_state(state)
        # Both flows: grades → credit_pull → sample.
        return redirect(url_for("setup.step6_credit_pull"))
    return render_template(
        "setup/step5_grades.html",
        admin_min=score_floor,
        admin_max=score_ceiling,
        **_wizard_ctx("grades"),
    )


# =================================================================
# Step 6 — Credit pull source
# =================================================================
def _read_credit_pull_headers(path: Path) -> list[str]:
    """Return the column-header strings from the first row of a credit-pull
    file. Empty list on failure or empty file. Embedded whitespace (newlines,
    tabs, multiple spaces) inside a header cell is normalised to a single
    space — browsers do the same when round-tripping `<option value="...">`
    on form submit, so we must store the same shape to keep the dropdown
    selection sticky across saves.
    """
    headers: list[str] = []

    def _norm(cell: object) -> str:
        return re.sub(r"\s+", " ", str(cell)).strip()

    try:
        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xlsm", ".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                ws = wb[wb.sheetnames[0]]
                first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                headers = [_norm(c) for c in first if c not in (None, "")]
                headers = [h for h in headers if h]
            finally:
                wb.close()
        elif suffix == ".csv":
            import csv
            with open(path, "r", encoding="utf-8-sig", newline="") as fh:
                reader = csv.reader(fh)
                row = next(reader, [])
                headers = [_norm(h) for h in row if h and _norm(h)]
    except Exception:  # noqa: BLE001
        return []
    return headers


def _detect_credit_pull_columns(path: Path) -> tuple[str | None, str | None]:
    """Read the first sheet/row of a credit-pull file and pick best-guess
    member-# and score column names. Returns (member_col, score_col); either
    may be None if no plausible header was found.
    """
    headers = _read_credit_pull_headers(path)
    if not headers:
        return None, None

    def _pick(patterns: list[str]) -> str | None:
        # Exact (case-insensitive) match first.
        lowered = {h.lower(): h for h in headers}
        for p in patterns:
            if p.lower() in lowered:
                return lowered[p.lower()]
        # Substring fallback.
        for p in patterns:
            for h in headers:
                if p.lower() in h.lower():
                    return h
        return None

    member = _pick([
        "Member Number", "Member #", "Member No", "Member ID",
        "Account Number", "Acct Number", "Acct #", "Account #",
        "Member", "Account",
    ])
    score = _pick([
        "FICO", "Current FICO", "Current Score", "Credit Score",
        "Score", "Bureau Score",
    ])
    return member, score


def _auto_seed_credit_pull_from_sample(state: dict[str, Any]) -> bool:
    """If a credit-pull file was uploaded in Step 3 but Step 8 hasn't been
    configured yet, populate ``state['credit_pull']`` from that file:
    folder + exact-name pattern + best-guess member/score columns.
    Returns True if any field was set.
    """
    cp = state.setdefault("credit_pull", {})
    # Already configured (user uploaded directly on Step 8, or has typed
    # values manually) — don't clobber.
    if cp.get("uploaded_filename") or cp.get("source_folder"):
        return False
    files = ((state.get("sample_uploads") or {}).get("credit_pull_files")) or []
    if not files:
        return False
    # Most recently uploaded entry sits at the end of the list.
    entry = files[-1]
    fpath = Path(entry.get("path") or "")
    if not fpath.exists():
        return False
    cp["source_folder"] = str(fpath.parent)
    # Generalized pattern wildcards month / year tokens so next quarter's
    # drop of the same file matches without re-running the wizard.
    cp["file_pattern"] = sample_parser.guess_filename_patterns(
        fpath.name
    )["file_pattern"]
    cp["uploaded_filename"] = fpath.name
    member, score = _detect_credit_pull_columns(fpath)
    if member:
        cp["member_column"] = member
    if score:
        cp["score_column"] = score
    return True


@setup_bp.route("/step/credit-pull", methods=["GET", "POST"])
def step6_credit_pull():
    state = _state()
    if request.method == "GET":
        if _auto_seed_credit_pull_from_sample(state):
            _save_state(state)
            cp = state["credit_pull"]
            bits = [f"file <code>{cp['uploaded_filename']}</code>"]
            if cp.get("member_column"):
                bits.append(f"member column <code>{cp['member_column']}</code>")
            if cp.get("score_column"):
                bits.append(f"score column <code>{cp['score_column']}</code>")
            flash(
                "Auto-loaded credit-pull file from Step 3: "
                + ", ".join(bits)
                + ". Review and adjust below if needed.",
                "info",
            )
        # Normalize mutually-exclusive options on GET as well, so any state
        # left over from earlier defaults (e.g. prefer_original_for_new_loans
        # = True alongside an Option-1 upload) doesn't show two boxes ticked.
        cp = state.get("credit_pull") or {}
        if cp.get("use_standalone_file"):
            if cp.get("use_configured_report") or cp.get("prefer_original_for_new_loans"):
                cp["use_configured_report"] = False
                cp["prefer_original_for_new_loans"] = False
                _save_state(state)
        elif cp.get("use_configured_report"):
            if cp.get("prefer_original_for_new_loans"):
                cp["prefer_original_for_new_loans"] = False
                _save_state(state)
    if request.method == "POST":
        cp = state["credit_pull"]
        action = request.form.get("action", "")

        # --- Option 1: file picker upload ---------------------------------
        if action == "upload_credit_pull":
            up = request.files.get("credit_pull_file")
            if up and up.filename:
                saved = _save_sample_upload(up)
                cp["source_folder"] = str(saved.parent)
                cp["file_pattern"] = sample_parser.guess_filename_patterns(
                    saved.name
                )["file_pattern"]
                cp["uploaded_filename"] = saved.name
                cp["use_standalone_file"] = True
                # Option 1 wins — turn off Option 2 / Option 3.
                cp["use_configured_report"] = False
                cp["prefer_original_for_new_loans"] = False
                # Auto-detect the member/score columns from the upload so
                # the column-mapping table is pre-filled.
                m_col, s_col = _detect_credit_pull_columns(saved)
                if m_col:
                    cp["member_column"] = m_col
                if s_col:
                    cp["score_column"] = s_col
                flash(f"Credit-pull file saved: {saved.name}", "success")
                # A fresh credit-pull replacement is the canonical trigger
                # for "I want my reports re-run with new scores."  If the
                # CU is already set up (YAML + Archive folder exist) the
                # previously-imported loan extracts are sitting in
                # Archive/<short>/ — restore them so the next run_import
                # has something to chew on.  No-op for first-time setup.
                sn = (state.get("short_name") or "").strip()
                if sn:
                    try:
                        res = pipeline_service.restore_archived_files(sn)
                        n_restored = len(res.get("restored") or [])
                        if n_restored:
                            preview = ", ".join((res.get("restored") or [])[:3])
                            if n_restored > 3:
                                preview += f", … (+{n_restored - 3} more)"
                            flash(
                                f"Restored {n_restored} archived loan file(s) "
                                f"into Raw_Uploads/{sn}/ so the next import "
                                f"will re-score them: {preview}.",
                                "info",
                            )
                        errs = res.get("errors") or []
                        if errs:
                            flash(
                                f"Archive restore had {len(errs)} error(s); "
                                f"first: {errs[0]}",
                                "warning",
                            )
                    except Exception as exc:  # noqa: BLE001
                        flash(
                            f"Archive restore skipped (non-fatal): {exc}",
                            "warning",
                        )
            else:
                flash("Please choose a credit-pull file to upload.", "warning")
            _save_state(state)
            return redirect(url_for("setup.step6_credit_pull"))
        if action == "clear_credit_pull":
            cp["source_folder"] = None
            cp["file_pattern"] = None
            cp.pop("uploaded_filename", None)
            cp["use_standalone_file"] = False
            _save_state(state)
            flash("Cleared standalone credit-pull file.", "info")
            return redirect(url_for("setup.step6_credit_pull"))

        # --- Option 2: replace the WARM workbook from this step ----------
        if action == "replace_warm_file":
            up = request.files.get("warm_file")
            if not up or not up.filename:
                flash("Pick a WARM workbook to upload first.", "warning")
            else:
                try:
                    saved = _save_warm_upload(up)
                    analysis = warm_parser.analyse_warm_file(
                        saved, original_filename=up.filename
                    )
                    if not analysis.get("ok"):
                        flash(
                            f"Could not parse WARM file: {analysis.get('error')}",
                            "error",
                        )
                    else:
                        analysis["saved_path"] = str(saved)
                        _apply_warm_to_state(state, analysis)
                        # If Option 2 is active, repoint its folder/pattern
                        # at the new file.
                        if cp.get("use_configured_report"):
                            p = Path(saved)
                            cp["fallback_report_folder"] = str(p.parent)
                            cp["fallback_report_pattern"] = (
                                "^" + re.escape(p.name) + "$"
                            )
                        _save_state(state)
                        flash(
                            f"Replaced WARM workbook: {up.filename}",
                            "success",
                        )
                except Exception as exc:  # noqa: BLE001
                    flash(f"WARM upload failed: {exc}", "error")
            return redirect(url_for("setup.step6_credit_pull"))

        # --- Option 1 enable/disable + remaining fields (full form save) -
        cp["use_standalone_file"] = bool(
            request.form.get("use_standalone_file")
        )

        for fld in ("file_pattern", "source_folder", "member_column",
                    "score_column", "fallback_report_pattern",
                    "fallback_report_folder", "fallback_sheet_pattern"):
            cp[fld] = request.form.get(fld, "").strip() or None

        # --- Option 2: use the configured WARM/CECL report ---------------
        cp["use_configured_report"] = bool(
            request.form.get("use_configured_report")
        )
        if cp["use_configured_report"]:
            warm = state.get("warm") or {}
            warm_path = warm.get("saved_path")
            if warm_path and Path(warm_path).exists():
                p = Path(warm_path)
                cp["fallback_report_folder"] = str(p.parent)
                cp["fallback_report_pattern"] = "^" + re.escape(p.name) + "$"
            else:
                flash(
                    "Can't use the configured WARM workbook — no WARM file "
                    "has been uploaded in Step 2.",
                    "warning",
                )
                cp["use_configured_report"] = False

        try:
            cp["fallback_member_col"] = int(
                request.form.get("fallback_member_col", "0")
            )
            cp["fallback_score_col"] = int(
                request.form.get("fallback_score_col", "1")
            )
        except ValueError:
            pass

        # New: original-score fallback options
        cp["prefer_original_for_new_loans"] = bool(
            request.form.get("prefer_original_for_new_loans")
        )

        # Enforce mutually-exclusive options 1 / 2 / 3.
        # Priority: Option 1 (standalone) > Option 2 (WARM tabs) > Option 3.
        if cp.get("use_standalone_file"):
            cp["use_configured_report"] = False
            cp["prefer_original_for_new_loans"] = False
        elif cp.get("use_configured_report"):
            cp["prefer_original_for_new_loans"] = False
        pull_date = request.form.get("pull_as_of_date", "").strip()
        if pull_date:
            try:
                # validate yyyy-mm-dd
                from datetime import date as _date
                _date.fromisoformat(pull_date)
                cp["pull_as_of_date"] = pull_date
            except ValueError:
                flash(
                    f"Pull-as-of date '{pull_date}' isn't a valid YYYY-MM-DD; "
                    "left blank — engine will infer from file timestamp.",
                    "info",
                )
                cp["pull_as_of_date"] = ""
        else:
            cp["pull_as_of_date"] = ""

        # Soft warning: original-score-only mode requires open_date column for
        # the per-loan freshness comparison to mean anything.
        no_pull = not (cp.get("file_pattern") or cp.get("fallback_report_pattern"))
        if cp["prefer_original_for_new_loans"]:
            mapped_open = (state.get("column_mappings") or {}).get("open_date")
            if not mapped_open:
                flash(
                    "Original-score fallback is enabled, but no 'open_date' "
                    "column is mapped in step 5 — every loan will use the "
                    "credit-pull score (or original score if no pull is found).",
                    "info",
                )
        if no_pull and not cp["prefer_original_for_new_loans"]:
            flash(
                "No credit-pull source configured and original-score fallback "
                "is OFF — current score will equal original score for every loan.",
                "info",
            )

        _save_state(state)
        # Save-progress buttons in the stepper hijack the cp-main form via
        # the HTML5 form= attribute and submit with action=save_progress_*.
        # Honour those by redirecting to home / back-to-self instead of
        # advancing to the next wizard step — without this branch the user's
        # column-mapping selections still save, but they never get to see
        # the persisted page.
        if action == "save_progress_exit":
            session.pop(STATE_KEY, None)
            return redirect(url_for("home.index"))
        if action == "save_progress_stay":
            return redirect(url_for("setup.step6_credit_pull"))
        return redirect(url_for("setup.step_orig_score"))
    # GET — gather column headers from the uploaded credit-pull file (if any)
    # so the Member#/Score selects can be populated.
    cp = state.get("credit_pull") or {}
    cp_headers: list[str] = []
    cp_path = None
    if cp.get("source_folder") and cp.get("uploaded_filename"):
        candidate = Path(cp["source_folder"]) / cp["uploaded_filename"]
        if candidate.exists():
            cp_path = candidate
            cp_headers = _read_credit_pull_headers(candidate)
    # Self-heal: existing drafts may have stored a header containing embedded
    # whitespace ('Credit\r\nScore'). Browsers normalise that on form submit,
    # so the dropdown's selected value would never match — making the picker
    # appear to reset. Re-map the saved column to the normalised header.
    if cp_headers:
        norm_lookup = {re.sub(r"\s+", " ", h).strip().lower(): h for h in cp_headers}
        changed = False
        for fld in ("member_column", "score_column"):
            saved = cp.get(fld) or ""
            if saved and saved not in cp_headers:
                key = re.sub(r"\s+", " ", str(saved)).strip().lower()
                fixed = norm_lookup.get(key)
                if fixed and fixed != saved:
                    cp[fld] = fixed
                    changed = True
        if changed:
            _save_state(state)
    return render_template(
        "setup/step6_credit_pull.html",
        cp_headers=cp_headers,
        **_wizard_ctx("credit_pull"),
    )


# =================================================================
# Step — Original Credit Score Baseline (one-time upload)
# =================================================================
_ORIG_SCORE_DIR = Path(tempfile.gettempdir()) / "cecl_ui_orig_score"


def _save_orig_score_upload(file_storage) -> Path:
    """Save the uploaded baseline file under a deterministic temp folder."""
    _ORIG_SCORE_DIR.mkdir(parents=True, exist_ok=True)
    fn = secure_filename(file_storage.filename) or "baseline.csv"
    target = _ORIG_SCORE_DIR / fn
    file_storage.save(str(target))
    return target


def _detect_orig_score_columns(headers: list[str]) -> tuple[str, str, str]:
    """Best-guess (member, suffix, score) column names from a header list."""
    if not headers:
        return "", "", ""

    def _pick(patterns: list[str]) -> str:
        lowered = {h.lower(): h for h in headers}
        for p in patterns:
            if p.lower() in lowered:
                return lowered[p.lower()]
        for p in patterns:
            for h in headers:
                if p.lower() in h.lower():
                    return h
        return ""

    member = _pick([
        "Member Number", "Member #", "Member No", "Member ID",
        "Account Number", "Acct Number", "Acct #", "Account #",
        "Member", "Account",
    ])
    suffix = _pick([
        "Loan Suffix", "Suffix", "Loan #", "Loan Number", "Note", "Sub Acct",
    ])
    score = _pick([
        "Original FICO", "Original Score", "Original Credit Score",
        "Orig FICO", "Orig Score", "Baseline FICO", "Baseline Score",
        "FICO", "Credit Score", "Score",
    ])
    return member, suffix, score


def _read_orig_score_table(path: Path) -> tuple[list[str], list[list[str]]]:
    """Return (headers, rows) for a CSV/XLSX baseline file. ``rows`` is the
    full table as list-of-list-of-strings (None -> ""), preserving the
    user's column order. Empty list on failure.
    """
    headers: list[str] = []
    rows: list[list[str]] = []
    try:
        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xlsm", ".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                ws = wb[wb.sheetnames[0]]
                it = ws.iter_rows(values_only=True)
                first = next(it, None)
                if first is None:
                    return [], []
                headers = [str(c).strip() if c is not None else "" for c in first]
                for r in it:
                    if r is None:
                        continue
                    if all(c is None or str(c).strip() == "" for c in r):
                        continue
                    rows.append([
                        "" if c is None else str(c).strip()
                        for c in r
                    ])
            finally:
                wb.close()
        elif suffix == ".csv":
            import csv
            with open(path, "r", encoding="utf-8-sig", newline="") as fh:
                reader = csv.reader(fh)
                first = next(reader, None)
                if first is None:
                    return [], []
                headers = [str(c).strip() for c in first]
                for r in reader:
                    if not r or all((c or "").strip() == "" for c in r):
                        continue
                    rows.append([(c or "").strip() for c in r])
    except Exception:  # noqa: BLE001
        return [], []
    return headers, rows


def _build_orig_score_rows(
    headers: list[str], rows: list[list[str]],
    member_col: str, suffix_col: str, score_col: str,
) -> tuple[list[dict[str, Any]], int]:
    """Slice the parsed table down to the {member, suffix, score} lookup
    rows. Skips rows with no member# or no numeric score. Returns
    (rows, skipped_count).
    """
    if not headers or not rows or not member_col or not score_col:
        return [], 0
    idx_member = headers.index(member_col) if member_col in headers else -1
    idx_score = headers.index(score_col) if score_col in headers else -1
    idx_suffix = headers.index(suffix_col) if (
        suffix_col and suffix_col in headers
    ) else -1
    if idx_member < 0 or idx_score < 0:
        return [], 0
    out: list[dict[str, Any]] = []
    skipped = 0
    for r in rows:
        mem = (r[idx_member] if idx_member < len(r) else "").strip()
        if not mem:
            skipped += 1
            continue
        # Strip any leading zeros artifact like "12345.0" from Excel.
        mem = re.sub(r"\.0$", "", mem)
        score_raw = (r[idx_score] if idx_score < len(r) else "").strip()
        try:
            score = int(float(score_raw))
        except (TypeError, ValueError):
            skipped += 1
            continue
        if score <= 0:
            skipped += 1
            continue
        suf = ""
        if idx_suffix >= 0 and idx_suffix < len(r):
            suf = (r[idx_suffix] or "").strip()
            suf = re.sub(r"\.0$", "", suf)
        out.append({"member": mem, "suffix": suf, "score": score})
    return out, skipped


@setup_bp.route("/step/orig-score", methods=["GET", "POST"])
def step_orig_score():
    state = _state()
    osb = state.setdefault("orig_score_baseline", {
        "saved_path": "", "uploaded_filename": "",
        "member_column": "", "suffix_column": "", "score_column": "",
        "rows": [], "row_count": 0, "preview": [], "headers": [],
        "pools": [],
    })
    osb.setdefault("pools", [])

    if request.method == "POST":
        action = request.form.get("action", "next")

        if action == "upload_baseline":
            up = request.files.get("baseline_file")
            if not up or not up.filename:
                flash("Pick a baseline file to upload first.", "warning")
                return redirect(url_for("setup.step_orig_score"))
            try:
                saved = _save_orig_score_upload(up)
            except Exception as exc:  # noqa: BLE001
                flash(f"Upload failed: {exc}", "error")
                return redirect(url_for("setup.step_orig_score"))
            headers, full_rows = _read_orig_score_table(saved)
            if not headers:
                flash(
                    "Couldn't read that file. Use a .csv / .xlsx with a "
                    "header row (member#, optional suffix, original FICO).",
                    "error",
                )
                return redirect(url_for("setup.step_orig_score"))
            m_col, s_col, sc_col = _detect_orig_score_columns(headers)
            osb["saved_path"] = str(saved)
            osb["uploaded_filename"] = saved.name
            osb["headers"] = headers
            osb["preview"] = full_rows[:8]
            osb["_full_rows"] = full_rows  # stashed for re-parse on column save
            osb["member_column"] = m_col
            osb["suffix_column"] = s_col
            osb["score_column"] = sc_col
            built, skipped = _build_orig_score_rows(
                headers, full_rows, m_col, s_col, sc_col,
            )
            osb["rows"] = built
            osb["row_count"] = len(built)
            _save_state(state)
            msg = f"Loaded {len(built)} baseline score(s) from {saved.name}."
            if skipped:
                msg += f" Skipped {skipped} row(s) with no member# or invalid score."
            if not (m_col and sc_col):
                msg += " Pick the Member# and Score columns below, then click Save."
            flash(msg, "success")
            return redirect(url_for("setup.step_orig_score"))

        if action == "clear_baseline":
            osb.update({
                "saved_path": "", "uploaded_filename": "",
                "member_column": "", "suffix_column": "", "score_column": "",
                "rows": [], "row_count": 0, "preview": [], "headers": [],
                "pools": [],
            })
            osb.pop("_full_rows", None)
            _save_state(state)
            flash("Cleared baseline upload.", "info")
            return redirect(url_for("setup.step_orig_score"))

        if action == "save_columns":
            m_col = (request.form.get("member_column") or "").strip()
            s_col = (request.form.get("suffix_column") or "").strip()
            sc_col = (request.form.get("score_column") or "").strip()
            osb["member_column"] = m_col
            osb["suffix_column"] = s_col
            osb["score_column"] = sc_col
            # Pool scope (checkbox list). Empty selection = apply to all pools.
            selected_pools = [
                (p or "").strip()
                for p in request.form.getlist("apply_pools")
                if (p or "").strip()
            ]
            osb["pools"] = selected_pools
            # Reparse from the stashed full-rows snapshot, or re-read the file
            # if the snapshot was lost (e.g. session reload).
            full_rows = osb.get("_full_rows")
            headers = osb.get("headers") or []
            if not full_rows and osb.get("saved_path"):
                p = Path(osb["saved_path"])
                if p.exists():
                    headers, full_rows = _read_orig_score_table(p)
                    osb["headers"] = headers
                    osb["preview"] = (full_rows or [])[:8]
                    osb["_full_rows"] = full_rows
            built, skipped = _build_orig_score_rows(
                headers, full_rows or [], m_col, s_col, sc_col,
            )
            osb["rows"] = built
            osb["row_count"] = len(built)
            _save_state(state)
            flash(
                f"Saved column mapping. {len(built)} baseline score(s) ready"
                + (f"; {skipped} skipped." if skipped else "."),
                "success",
            )
            return redirect(url_for("setup.step_orig_score"))

        # action == "next" (skip or continue)
        _save_state(state)
        return redirect(url_for("setup.step2_sample"))

    return render_template(
        "setup/step_orig_score.html", **_wizard_ctx("orig_score"),
    )


# =================================================================
# Step 7 — Economic data
# =================================================================
@setup_bp.route("/step/economic", methods=["GET", "POST"])
def step7_economic():
    state = _state()
    fetched: dict[str, Any] | None = None
    if request.method == "POST":
        action = request.form.get("action", "next")
        ed = state["economic_data"]
        ed["state"] = request.form.get("state", "").strip()
        ed["county"] = request.form.get("county", "").strip()

        def _f(field: str, cast=float):
            try:
                return cast(request.form.get(field, "").strip() or 0)
            except ValueError:
                return 0

        ed["unemployment_rate"] = _f("unemployment_rate", float)
        ed["foreclosures"] = _f("foreclosures", int)
        ed["bankruptcies"] = _f("bankruptcies", int)
        ed["population"] = _f("population", int)

        if action == "fetch":
            if not ed["state"]:
                flash("Pick a state before auto-fetching.", "error")
            else:
                fetched = pipeline_service.fetch_economic_data(
                    ed["state"], ed["county"]
                )
                if "error" in fetched:
                    flash(f"Could not auto-fetch: {fetched['error']}", "error")
                else:
                    updated = []
                    for k in ("unemployment_rate", "foreclosures",
                              "bankruptcies", "population"):
                        if k in fetched and fetched[k] is not None:
                            ed[k] = fetched[k]
                            updated.append(k)
                    if updated:
                        flash(
                            "Auto-fetch updated: "
                            + ", ".join(updated).replace("_", " ")
                            + ". Review and click Next.",
                            "success",
                        )
                    else:
                        flash(
                            "Auto-fetch returned no usable values; "
                            "enter figures manually below.",
                            "info",
                        )
                    if "foreclosures" not in fetched or fetched.get("foreclosures") is None:
                        flash(
                            "Foreclosures aren't published by the federal "
                            "APIs — enter a county-level estimate manually "
                            "if you have one.",
                            "info",
                        )
            _save_state(state)
        else:
            _save_state(state)
            return redirect(url_for("setup.step8_mgmt_adj"))
    return render_template(
        "setup/step7_economic.html",
        fetched=fetched,
        states=geo_service.states(),
        **_wizard_ctx("economic"),
    )


# =================================================================
# Step 8 — Management adjustments
# =================================================================
@setup_bp.route("/step/mgmt-adj", methods=["GET", "POST"])
def step8_mgmt_adj():
    state = _state()

    # Build the list of pool names the user might want to adjust.
    # Prefer the order established by the WARM workbook so this step mirrors
    # the pool layout the WARM analysis defined. We use, in order of
    # preference: state["pool_settings"] (per-pool WARM settings, canonical
    # WARM order), then state["warm"]["pools"], then any pool_map values
    # the user has configured. default_pool is appended last if missing.
    pool_names: list[str] = []
    seen: set[str] = set()

    def _add(name: object) -> None:
        s = (name or "").strip() if isinstance(name, str) else ""
        if s and s.lower() not in seen:
            seen.add(s.lower())
            pool_names.append(s)

    for ps in (state.get("pool_settings") or []):
        _add((ps or {}).get("name"))
    for name in ((state.get("warm") or {}).get("pools") or []):
        _add(name)
    for v in (state.get("pool_map") or {}).values():
        _add(v)
    _add(state.get("default_pool"))
    # Note: do NOT sort — WARM ordering is intentional.

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()
        try:
            state["mgmt_adj"]["ltv_baseline"] = float(
                request.form.get("ltv_baseline", "0.9")
            )
            state["mgmt_adj"]["probability_factor"] = float(
                request.form.get("probability_factor", "0.35")
            )
        except ValueError:
            flash("Numeric values required for LTV / probability factor.", "error")

        # Per-pool overlays: parse a percentage and store as decimal.
        # Empty / 0 entries are dropped so the YAML stays clean.
        new_overlay: dict[str, float] = {}
        bad_pools: list[str] = []
        for pname in pool_names:
            raw = (request.form.get(f"pool_pct__{pname}", "") or "").strip()
            if not raw:
                continue
            try:
                pct = float(raw)
            except ValueError:
                bad_pools.append(pname)
                continue
            # Stored as a decimal (e.g. 0.5% -> 0.005).
            decimal = pct / 100.0
            if abs(decimal) > 1e-9:
                new_overlay[pname] = round(decimal, 6)
        if bad_pools:
            flash(
                "Ignored non-numeric overlays for: " + ", ".join(bad_pools),
                "error",
            )
        state["mgmt_adj_by_pool"] = new_overlay

        # Per-pool "use industry default management adjustment" flag.
        # Persist into state.pool_settings so the YAML / report engine
        # can pick it up (alongside risk_rated, acl_months, brr).
        existing_ps = state.get("pool_settings") or []
        # Case-insensitive lookup so ``Visa`` and ``visa`` don't create
        # duplicate entries.
        ps_by_name = {
            (p.get("name") or "").strip().lower(): p for p in existing_ps
        }
        # Pools whose risk-rated setting is configured elsewhere
        # (Step 2 WARM / Step 4 Pools). When step8 is the first step
        # touching pool_settings (e.g. user skipped earlier configuration
        # for the default_pool), default risk_rated to True so the
        # report renders per-grade detail — matches report_tct.py's
        # ``risk_rated_map.get(pool, True)`` fallback. The user can flip
        # it on Step 4 if a pool is genuinely non-risk-rated.
        default_pool_lc = (state.get("default_pool") or "").strip().lower()
        for pname in pool_names:
            checked = request.form.get(f"pool_use_default__{pname}") == "on"
            entry = ps_by_name.get(pname.lower())
            if entry is None:
                # default_pool ("Other/Uncategorized" or similar) is
                # almost always an Ignore bucket → leave NRR. Real pools
                # default to risk_rated=True to match report-engine
                # default and avoid silently collapsing per-grade detail.
                is_default_bucket = pname.strip().lower() == default_pool_lc
                entry = {
                    "name": pname,
                    "risk_rated": not is_default_bucket,
                    "brr": False,
                    "acl_months": 0,
                    "use_default_mgmt_adj": checked,
                    "excluded": False,
                }
                existing_ps.append(entry)
                ps_by_name[pname.lower()] = entry
            else:
                entry["use_default_mgmt_adj"] = checked
        state["pool_settings"] = existing_ps

        # Other Allowance Considerations (optional) — moved here from the
        # Credit Grades step. These are pool-agnostic ACL overlays the user
        # adds on top of mgmt adjustments.
        include = request.form.get("include_other_allowance") == "on"
        state["include_other_allowance"] = include
        oac: list[dict[str, Any]] = []
        if include:
            titles = request.form.getlist("oac_title")
            bals = request.form.getlist("oac_balance")
            pcts = request.form.getlist("oac_percentage")
            for j in range(len(titles)):
                title = (titles[j] or "").strip()
                bal_raw = (bals[j] if j < len(bals) else "" or "").strip().replace(",", "").replace("$", "")
                pct_raw = (pcts[j] if j < len(pcts) else "" or "").strip().replace("%", "")
                if not title and not bal_raw and not pct_raw:
                    continue
                try:
                    bal = float(bal_raw) if bal_raw else 0.0
                except ValueError:
                    bal = 0.0
                try:
                    pct = float(pct_raw) if pct_raw else 0.0
                except ValueError:
                    pct = 0.0
                if pct < 0:
                    pct = 0.0
                if pct > 100:
                    pct = 100.0
                oac.append({
                    "title": title or "(untitled)",
                    "balance": bal,
                    "percentage": pct,
                })
        state["other_allowance_considerations"] = oac

        # Negative Share Provision (data-derived Other Allowance
        # Consideration). Reusable across CUs: the engine recomputes a
        # life-of-loan loss rate each quarter from the negative-share balance
        # extract and charge-off / recovery history, so we only persist the
        # file patterns / folder here.
        ns_on = request.form.get("include_negative_share") == "on"
        state["include_negative_share"] = ns_on
        if ns_on:
            ns = dict(state.get("negative_share") or {})
            ns["title"] = (
                (request.form.get("ns_title") or "").strip()
                or "Negative Share Provision"
            )
            try:
                ns["life_of_loan_months"] = int(
                    request.form.get("ns_life_months", "12") or "12"
                )
            except ValueError:
                ns["life_of_loan_months"] = 12
            ns["source_folder"] = (request.form.get("ns_source_folder") or "").strip()
            ns["balance_column"] = (
                (request.form.get("ns_balance_column") or "").strip()
                or "Current Balance"
            )
            ns["balance_pattern"] = (
                (request.form.get("ns_balance_pattern") or "").strip()
                or r"(?i)Negative Share File"
            )
            ns["co_summary_pattern"] = (
                (request.form.get("ns_co_summary_pattern") or "").strip()
                or r"(?i)Negative Shares Charge Off and Recovery"
            )
            ns["co_quarterly_pattern"] = (
                (request.form.get("ns_co_quarterly_pattern") or "").strip()
                or r"(?i)Share COs?\s*-\s*Recoveries"
            )
            state["negative_share"] = ns

        # Unfunded Commitments (data-derived Other Allowance Consideration).
        # The user supplies the loan type codes flagged as NOT unconditionally
        # cancelable; balances and rates are resolved at report time per pool.
        uc_on = request.form.get("include_unfunded_commitment") == "on"
        state["include_unfunded_commitment"] = uc_on
        if uc_on:
            uc = dict(state.get("unfunded_commitment") or {})
            uc["title"] = (
                (request.form.get("uc_title") or "").strip()
                or "Unfunded Commitments"
            )
            raw_codes = request.form.get("uc_loan_type_codes") or ""
            codes = [c.strip() for c in re.split(r"[\s,;]+", raw_codes) if c.strip()]
            # De-dupe preserving order.
            seen_c: set[str] = set()
            uc["loan_type_codes"] = [
                c for c in codes if not (c in seen_c or seen_c.add(c))
            ]
            state["unfunded_commitment"] = uc

        _save_state(state)

        # Stepper "Save progress" / "Save progress & exit" buttons submit
        # this same form (via HTML5 form="step-form") so the user's typed
        # overlays land in state above before we redirect. Mirror the
        # /save-progress UX rather than advancing to the next step.
        if action in ("save_progress_stay", "save_progress_exit"):
            label = state.get("credit_union") or state.get("short_name") or "draft"
            flash(f"Saved progress for {label}.", "success")
            if action == "save_progress_exit":
                session.pop(STATE_KEY, None)
                return redirect(url_for("home.index"))
            return redirect(url_for("setup.step8_mgmt_adj"))

        return redirect(url_for("setup.step9_reports"))

    # Pre-format saved overlays (decimals) back to percentages for the form.
    saved_overlay = state.get("mgmt_adj_by_pool") or {}
    pool_overlay_pct: dict[str, str] = {}
    for pname in pool_names:
        d = saved_overlay.get(pname)
        if d is None:
            pool_overlay_pct[pname] = ""
        else:
            # Show as plain percent without trailing zeros.
            v = d * 100.0
            pool_overlay_pct[pname] = ("%g" % v)

    # Per-pool "use industry default mgmt adj" checkbox state, sourced
    # from state.pool_settings (was previously edited on Step 2 WARM).
    ps_by_name_lookup = {
        (p.get("name") or "").strip(): bool(p.get("use_default_mgmt_adj"))
        for p in (state.get("pool_settings") or [])
    }
    pool_use_default = {p: ps_by_name_lookup.get(p, False) for p in pool_names}

    return render_template(
        "setup/step8_mgmt_adj.html",
        pool_names=pool_names,
        pool_overlay_pct=pool_overlay_pct,
        pool_use_default=pool_use_default,
        **_wizard_ctx("mgmt_adj"),
    )


# =================================================================
# Step 9 — Reports to generate
# =================================================================
@setup_bp.route("/step/reports", methods=["GET", "POST"])
def step9_reports():
    state = _state()
    if request.method == "POST":
        sel = {
            "tct":       request.form.get("tct") == "on",
            "vizo":      request.form.get("vizo") == "on",
            "vizo_supp": request.form.get("vizo_supp") == "on",
            "impdet":    request.form.get("impdet") == "on",
            "mgmt_adj_napkin": request.form.get("mgmt_adj_napkin") == "on",
            "acl_funding": request.form.get("acl_funding") == "on",
            "vizo_pdf": request.form.get("vizo_pdf") == "on",
        }
        if not any(sel.values()):
            flash(
                "Pick at least one report \u2014 you can always change this "
                "per-run later.",
                "error",
            )
            return render_template(
                "setup/step9_reports.html", **_wizard_ctx("reports")
            )
        if sel["vizo_supp"] and not sel["vizo"]:
            flash(
                "Vizo Supplemental is the historical companion to the main "
                "Vizo report \u2014 you usually want both turned on together.",
                "info",
            )
        state["reports"].update(sel)
        _save_state(state)
        return redirect(url_for("setup.step10_review"))
    return render_template("setup/step9_reports.html", **_wizard_ctx("reports"))


# =================================================================
# Step 10 — Review & Save
# =================================================================
def _build_review_summary(state: dict[str, Any]) -> dict[str, Any]:
    """Build a small dict of human-friendly highlights for the review screen."""
    cm = state.get("column_mappings", {}) or {}
    mapped = {k: v for k, v in cm.items() if v}

    pool_map = state.get("pool_map", {}) or {}
    named_pools = sorted({(v or "").strip() for v in pool_map.values()
                          if (v or "").strip()})
    unnamed_codes = [k for k, v in pool_map.items() if not (v or "").strip()]

    grades = state.get("credit_grades", []) or []

    cp = state.get("credit_pull", {}) or {}
    cp_sources: list[str] = []
    if (cp.get("source_folder") or cp.get("file_pattern")):
        cp_sources.append("Standalone file")
    if cp.get("fallback_report_pattern"):
        cp_sources.append("Fallback CECL report sheet")
    if cp.get("prefer_original_for_new_loans"):
        cp_sources.append("Original-score fallback for newer loans")

    ed = state.get("economic_data", {}) or {}

    overlay = state.get("mgmt_adj_by_pool", {}) or {}

    bt_map = state.get("balance_title_map") or {}
    bt_titles = state.get("warm", {}).get("balance_titles") or [] if state.get("warm") else []
    bt_total = len(bt_titles)
    bt_mapped = sum(1 for v in bt_map.values() if v)
    bt_ignored = bt_total - bt_mapped if bt_total else 0

    reports = state.get("reports", {}) or {}
    selected_reports = [name for name, on in (
        ("TCT", reports.get("tct")),
        ("Vizo", reports.get("vizo")),
        ("Vizo Supplemental", reports.get("vizo_supp")),
        ("Improved/Deteriorated", reports.get("impdet")),
        ("Management Adj Worksheet", reports.get("mgmt_adj_napkin")),
        ("ACL Funding Worksheet", reports.get("acl_funding")),
        ("Vizo Migration PDF", reports.get("vizo_pdf")),
    ) if on]

    return {
        "mapped_columns": mapped,
        "unmapped_columns": [k for k in cm if not cm[k]],
        "named_pools": named_pools,
        "unnamed_pool_codes": unnamed_codes,
        "default_pool": state.get("default_pool", ""),
        "grade_count": len(grades),
        "credit_pull_sources": cp_sources,
        "economic_state": ed.get("state", ""),
        "economic_county": ed.get("county", ""),
        "unemployment_rate": ed.get("unemployment_rate", 0),
        "pool_overlay": overlay,
        "selected_reports": selected_reports,
        "balance_titles_total": bt_total,
        "balance_titles_mapped": bt_mapped,
        "balance_titles_ignored": bt_ignored,
    }


def _review_warnings(state: dict[str, Any], summary: dict[str, Any]) -> list[str]:
    warns: list[str] = []
    if not state.get("credit_union"):
        warns.append("Credit union name is blank — go back to step 1.")
    if not state.get("short_name"):
        warns.append("Short name is blank — go back to step 1.")
    if not summary["mapped_columns"]:
        warns.append("No column mappings were configured (step 13 — Column Mappings).")
    if summary["unnamed_pool_codes"]:
        codes = summary["unnamed_pool_codes"]
        # List the actual codes so the user can find them quickly. Cap at
        # 10 to keep the warning readable when something has gone really
        # wrong (e.g. all 62 codes blank).
        shown = ", ".join(repr(c) for c in codes[:10])
        more = f" (+{len(codes) - 10} more)" if len(codes) > 10 else ""
        warns.append(
            f"{len(codes)} pool code(s) have no name on step 3 (Loan Code "
            f"Mapping) — they'll fall through to '{summary['default_pool']}'. "
            f"Codes: {shown}{more}."
        )
    if not summary["selected_reports"]:
        warns.append("No reports selected (step 20 — Reports).")
    if not summary["economic_state"]:
        warns.append("Economic state is blank (step 18 — Economic Data) — Q-factor fetch won't work.")
    return warns


def _copy_sample_uploads_to_raw(
    state: dict[str, Any], workspace_root: Path | str, short_name: str
) -> tuple[int, list[str]]:
    """Copy any wizard-uploaded sample files (loan / CO / recov / impaired)
    into ``Raw_Uploads/<short_name>/`` so the report engine can find them.

    Returns (files_copied, list_of_filenames_copied). Existing files in the
    destination are left in place UNLESS the wizard-uploaded source is
    newer (preserves real quarterly drops from the CU while still
    honoring intentional re-uploads from the wizard, e.g. when the
    user replaces a wrong impaired-loans file mid-setup).
    """
    dest_dir = config_service.raw_uploads_dir(workspace_root) / short_name
    dest_dir.mkdir(parents=True, exist_ok=True)

    su = state.get("sample_uploads") or {}
    keys = (
        "loan_data_files",
        "loan_balance_files",
        "co_files",
        "recov_files",
        "impaired_files",
    )
    copied: list[str] = []
    for key in keys:
        for entry in su.get(key) or []:
            src = Path(entry.get("path") or "")
            if not src.exists() or not src.is_file():
                continue
            target = dest_dir / src.name
            if target.exists():
                try:
                    if src.stat().st_mtime <= target.stat().st_mtime:
                        continue
                except OSError:
                    continue
            try:
                shutil.copy2(src, target)
            except OSError:
                continue
            copied.append(src.name)
    return len(copied), copied


@setup_bp.route("/step/review", methods=["GET", "POST"])
def step10_review():
    state = _state()
    # SCALE guard: SCALE-model drafts must finalize via the SCALE
    # wizard's Review step (which routes to Run and then SCALE Runs
    # dashboard). This handler is Migration-only -- if a SCALE draft
    # reaches it (e.g. via stale bookmark, or the Step 1 auto-scan
    # fallback prior to the Phase SCALE.route-guard fix), redirect
    # before any migration completion state is written.
    if state.get("model") == "scale":
        flash(
            "SCALE drafts finalize on the SCALE Review step, not the "
            "Migration review. Redirected you to the SCALE Review.",
            "info",
        )
        return redirect(url_for("scale_setup.step_review"))
    # Permanent guard: restore any provenance record the live session is
    # missing (from the on-disk draft or saved config) before we build and
    # persist the YAML, so a stale session can't drop it from the final config.
    _preserve_provenance_records(state)
    yaml_dict = config_service.build_yaml_from_wizard(state)
    import yaml as _yaml
    yaml_text = _yaml.safe_dump(
        yaml_dict, sort_keys=False, default_flow_style=False, allow_unicode=True
    )

    ws = current_app.config["WORKSPACE_ROOT"]
    sn = state.get("short_name") or config_service.slugify(
        state.get("credit_union") or ""
    )
    config_exists = (
        bool(sn)
        and (config_service.configs_dir(ws) / f"{sn}.yaml").exists()
    )

    summary = _build_review_summary(state)
    warnings = _review_warnings(state, summary)

    if request.method == "POST":
        overwrite = request.form.get("overwrite") == "on"
        if not sn:
            flash(
                "Cannot save — short name is blank. Go back to step 1.",
                "error",
            )
        else:
            # Make sure the engine can find the user's uploaded sample files.
            # If the wizard never set a data_directory, default it to the
            # per-CU Raw_Uploads folder so CO / recov / impaired files copied
            # below are picked up by load_chargeoff_recovery_history etc.
            ws_path = Path(ws)
            raw_dir = config_service.raw_uploads_dir(ws_path) / sn
            if not yaml_dict.get("data_directory"):
                yaml_dict["data_directory"] = str(raw_dir)
                state["data_directory"] = str(raw_dir)
            try:
                target = config_service.save_client_config(
                    ws, sn, yaml_dict, overwrite=overwrite
                )
            except FileExistsError:
                flash(
                    f"A config named '{sn}.yaml' already exists. "
                    "Tick 'Overwrite' to replace it, or change the short "
                    "name in step 1.",
                    "error",
                )
            else:
                # Copy sample files captured during the wizard into the
                # per-CU Raw_Uploads folder so the very first
                # Generate-Reports run has data to import.
                n_copied, copied_names = _copy_sample_uploads_to_raw(
                    state, ws, sn
                )
                # Stamp the Migration draft as completed so it appears
                # under "Completed setup" on the home dashboard. The
                # draft file is retained so the user can still Edit
                # setup or Delete from there. If the user reached
                # review without ever clicking "Save Progress", write
                # the draft first so mark_completed has something to
                # stamp.
                try:
                    wizard_drafts.save_draft(
                        ws, state, active_step="review",
                        model="migration",
                    )
                    wizard_drafts.mark_completed(
                        ws, sn, model="migration",
                    )
                except Exception:  # noqa: BLE001
                    # Non-fatal: completion stamp is best-effort and
                    # must not block the redirect to the run flow.
                    pass
                # Auto-run the importer so the user doesn't have to make
                # an extra click on the run page after finishing setup.
                # Best-effort: any failure flashes an error but still
                # delivers the user to the run dashboard so they can
                # diagnose / re-run from there.
                n_imported = 0
                import_error: str | None = None
                try:
                    n_imported = pipeline_service.run_import(sn)
                except Exception as exc:  # noqa: BLE001
                    import_error = str(exc) or exc.__class__.__name__
                # Done — wipe wizard state, send user to the run flow for this CU.
                session.pop(STATE_KEY, None)
                msg = f"Saved {target.name}. Raw_Uploads/{sn}/ folder created."
                if n_copied:
                    preview = ", ".join(copied_names[:4])
                    if n_copied > 4:
                        preview += f", … (+{n_copied - 4} more)"
                    msg += f" Copied {n_copied} sample file(s): {preview}."
                flash(msg, "success")
                if import_error:
                    flash(
                        f"Setup saved, but the auto-import failed: "
                        f"{import_error}. You can re-run the importer "
                        f"from the run page.",
                        "error",
                    )
                elif n_imported:
                    flash(
                        f"Imported {n_imported} loan(s) from "
                        f"Raw_Uploads/{sn}/.",
                        "success",
                    )
                else:
                    flash(
                        "No loan files were imported automatically. "
                        "Use the run page to upload or re-scan once "
                        "files are available.",
                        "info",
                    )
                return redirect(
                    url_for("run.client_dashboard", short_name=sn)
                )
        return render_template(
            "setup/step10_review.html",
            yaml_text=yaml_text, summary=summary, warnings=warnings,
            config_exists=config_exists, short_name=sn,
            **_wizard_ctx("review"),
        )

    return render_template(
        "setup/step10_review.html",
        yaml_text=yaml_text, summary=summary, warnings=warnings,
        config_exists=config_exists, short_name=sn,
        **_wizard_ctx("review"),
    )


@setup_bp.route("/cancel", methods=["POST"])
def cancel():
    session.pop(STATE_KEY, None)
    flash("Setup cancelled.", "info")
    return redirect(url_for("home.index"))
