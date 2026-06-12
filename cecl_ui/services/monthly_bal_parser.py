"""Parser for the recurring 'Monthly Balance by Pool/Type' workbook a CU
sends each quarter.

Public API
==========
- ``analyse_file(path)`` -> dict with detected layout + parsed labels/dates.
- ``normalize_to_month_end(value)`` -> ``datetime.date | None``. Snaps any
  date-ish value to the last day of its calendar month.
"""
from __future__ import annotations

import calendar
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string


# ---------------------------------------------------------------------------
# Date normalization
# ---------------------------------------------------------------------------

_MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


def _last_day(year: int, month: int) -> date:
    return date(year, month, calendar.monthrange(year, month)[1])


def _parse_string_date(s: str) -> tuple[int, int, int] | None:
    """Return (year, month, day) or None. Day defaults to 1 if absent."""
    s = s.strip()
    if not s:
        return None

    # ISO YYYY-MM-DD or YYYY-MM
    m = re.match(r"^(\d{4})[-/](\d{1,2})(?:[-/](\d{1,2}))?$", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3) or 1)
        if 1 <= mo <= 12:
            return y, mo, d

    # M/D/YYYY or M-D-YYYY (US-style)
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", s)
    if m:
        mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000 if y < 70 else 1900
        if 1 <= mo <= 12:
            return y, mo, d

    # MM-YYYY or MM/YYYY (no day)
    m = re.match(r"^(\d{1,2})[/-](\d{4})$", s)
    if m:
        mo, y = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12:
            return y, mo, 1

    # "Mon YYYY" / "Month YYYY" / "Mon-YY"
    m = re.match(r"^([A-Za-z]+)[\s\-/]+(\d{2,4})$", s)
    if m:
        mo = _MONTHS.get(m.group(1).lower())
        y = int(m.group(2))
        if y < 100:
            y += 2000 if y < 70 else 1900
        if mo:
            return y, mo, 1

    # "YYYY Mon" reverse
    m = re.match(r"^(\d{4})[\s\-/]+([A-Za-z]+)$", s)
    if m:
        y = int(m.group(1))
        mo = _MONTHS.get(m.group(2).lower())
        if mo:
            return y, mo, 1

    return None


def normalize_to_month_end(value: Any) -> date | None:
    """Snap any date-ish value to the last day of its calendar month.

    Returns ``None`` for unparseable input or for dates that fall before
    the year 1900 -- typos like ``"06/30/203"`` (meant to be 2023) get
    rejected here so they never poison downstream "earliest period"
    pickers.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        if value.year < 1900:
            return None
        return _last_day(value.year, value.month)
    if isinstance(value, date):
        if value.year < 1900:
            return None
        return _last_day(value.year, value.month)
    if isinstance(value, (int, float)):
        # Excel serial date (1900-based, ignoring the 1900 leap-year bug).
        # Restrict to a plausible CECL window (year 2000 through ~2100) so
        # balance values like 615.10, 1000, or 2,349,876.41 do not get
        # misread as serial dates and poison header-row detection.
        # Serial 36526 = 2000-01-01, serial 73415 = 2100-12-31.
        try:
            f = float(value)
        except (TypeError, ValueError):
            return None
        if not (36526 <= f <= 73415):
            return None
        try:
            from openpyxl.utils.datetime import from_excel
            d = from_excel(f)
            if isinstance(d, datetime):
                if d.year < 1900:
                    return None
                return _last_day(d.year, d.month)
            if isinstance(d, date):
                if d.year < 1900:
                    return None
                return _last_day(d.year, d.month)
        except Exception:  # noqa: BLE001
            return None
    if isinstance(value, str):
        parts = _parse_string_date(value)
        if parts:
            y, mo, _d = parts
            if y < 1900:
                return None
            return _last_day(y, mo)
    return None


# ---------------------------------------------------------------------------
# File analysis
# ---------------------------------------------------------------------------

_LIKELY_LABEL_HEADERS = {
    "loan type", "loan pool", "pool", "type", "category", "description",
    "loan category",
}

_SKIP_LABELS = {
    "total", "total loans", "subtotal", "sub-total", "sub total",
    "grand total",
}

# Phrases that, when present in a label, identify the row holding the
# Allowance for Credit Loss balance per month.
_ACL_LABEL_PATTERNS = [
    "alll balance",
    "acl balance",
    "allowance for credit loss",
    "allowance for loan loss",
    "allowance for loan and lease loss",
    "allowance balance",
]


def _is_acl_label(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    s = value.strip().lower()
    if not s:
        return False
    return any(p in s for p in _ACL_LABEL_PATTERNS)


def _find_acl_row_in_grid(
    rows: list[list[Any]],
    preferred_col: int | None = None,
) -> tuple[int | None, str, int | None]:
    """Scan a 2D grid for the row holding the ACL/ALLL line item.

    Returns ``(row_idx_0based, label, col_idx_0based)`` or
    ``(None, "", None)``. The ACL label often sits OUTSIDE the LOANS
    section (e.g. in 'Other Liabilities' or as a contra-asset), so we
    search the entire grid. ``preferred_col`` is checked first when
    given, then we fall back to any column.
    """
    cols_to_try: list[int | None]
    if preferred_col is not None:
        cols_to_try = [preferred_col, None]
    else:
        cols_to_try = [None]
    for col_pref in cols_to_try:
        for r, row in enumerate(rows):
            if col_pref is not None:
                if col_pref < len(row) and _is_acl_label(row[col_pref]):
                    return r, str(row[col_pref]).strip(), col_pref
            else:
                for c, v in enumerate(row):
                    if _is_acl_label(v):
                        return r, str(v).strip(), c
        if col_pref is not None:
            # If we found a hit in preferred column we'd have returned;
            # otherwise loop continues with col_pref=None.
            continue
    return None, "", None


def _is_label_row(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    s = value.strip()
    if not s:
        return False
    if s.lower() in _SKIP_LABELS:
        return False
    return True


def _scan_sheet(ws) -> dict[str, Any]:
    """Find header row + first date column on a single worksheet."""
    best: dict[str, Any] | None = None

    # Look at the first ~30 rows for a row containing 2+ date-like cells.
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=30, values_only=False), start=1
    ):
        date_cols: list[tuple[int, date, Any]] = []
        for cell in row:
            if cell.value is None:
                continue
            norm = normalize_to_month_end(cell.value)
            if norm is not None:
                date_cols.append((cell.column, norm, cell.value))
        if len(date_cols) >= 2:
            score = len(date_cols)
            if best is None or score > best["score"]:
                best = {
                    "score": score,
                    "header_row": row_idx,
                    "first_date_col_idx": date_cols[0][0],
                    "dates": [
                        {
                            "col": get_column_letter(c),
                            "raw": str(raw),
                            "normalized": d.isoformat(),
                        }
                        for (c, d, raw) in date_cols
                    ],
                }

    if not best:
        return {"ok": False, "error": "No date-like header row found in first 30 rows."}

    # Pool/type label column: scan EVERY column to the left of the first
    # date column for the one that yields the most string labels. Many
    # Vizo-style workbooks have col A = loan-type-code (integer) and
    # col B = description (the human-readable pool label). Hard-coding
    # col A would yield zero parsed_pool_labels in that case.
    first_date_col = best["first_date_col_idx"]
    candidate_cols = list(range(1, max(2, first_date_col)))  # at least col A
    label_col_idx = 1
    labels: list[str] = []
    seen: set[str] = set()
    acl_row: int | None = None
    acl_label: str = ""
    best_label_count = -1
    for try_col in candidate_cols:
        try_labels: list[str] = []
        try_seen: set[str] = set()
        try_acl_row: int | None = None
        try_acl_label: str = ""
        for row_idx in range(best["header_row"] + 1, best["header_row"] + 60):
            cell = ws.cell(row=row_idx, column=try_col)
            if cell.value is None:
                continue
            # ACL row detection (independent of pool-label collection).
            if try_acl_row is None and _is_acl_label(cell.value):
                try_acl_row = row_idx
                try_acl_label = str(cell.value).strip()
            if not _is_label_row(cell.value):
                continue
            s = cell.value.strip()
            key = s.lower()
            if key in try_seen:
                continue
            try_seen.add(key)
            try_labels.append(s)
        # Pick the column with the most labels. Tie-break favours the
        # earlier (leftmost) column to match historical behaviour.
        if len(try_labels) > best_label_count:
            best_label_count = len(try_labels)
            label_col_idx = try_col
            labels = try_labels
            seen = try_seen
            acl_row = try_acl_row
            acl_label = try_acl_label

    # Extract per-date ACL history if a row was identified.
    acl_history: dict[str, float] = {}
    if acl_row is not None:
        for entry in best["dates"]:
            col_letter = entry["col"]
            col_idx = _col_letter_to_idx(col_letter)
            if not col_idx:
                continue
            v = _coerce_number(ws.cell(row=acl_row, column=col_idx).value)
            if v is not None:
                acl_history[entry["normalized"]] = abs(v)

    return {
        "ok": True,
        "sheet": ws.title,
        "header_row": best["header_row"],
        "pool_name_col": get_column_letter(label_col_idx),
        "first_date_col": get_column_letter(best["first_date_col_idx"]),
        "dates": best["dates"],
        "parsed_pool_labels": labels,
        "acl_row": acl_row,
        "acl_label": acl_label,
        "acl_history": acl_history,
    }


def analyse_file(path: str | Path) -> dict[str, Any]:
    """Open the workbook, pick the most data-rich sheet, and return layout."""
    p = Path(path)
    if not p.exists():
        return {"ok": False, "error": f"File not found: {path}"}

    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"Could not open workbook: {exc}"}

    best: dict[str, Any] | None = None
    errors: list[str] = []
    try:
        for ws in wb.worksheets:
            try:
                result = _scan_sheet(ws)
            except Exception as exc:  # noqa: BLE001
                errors.append(f"{ws.title}: {exc}")
                continue
            if not result.get("ok"):
                continue
            score = len(result.get("dates", [])) + len(result.get("parsed_pool_labels", []))
            if best is None or score > best["_score"]:
                result["_score"] = score
                best = result
    finally:
        wb.close()

    if not best:
        msg = "Could not detect a header row with month-end dates."
        if errors:
            msg += " (" + "; ".join(errors[:3]) + ")"
        return {"ok": False, "error": msg}

    best.pop("_score", None)
    return best


def extract_row_history(
    saved_path: str | Path,
    sheet: str,
    header_row: int,
    target_row: int,
) -> dict[str, Any]:
    """Re-read a single row's per-date values from the monthly balance file.

    Used when the user overrides the auto-detected ACL row number.

    Returns ``{"ok": bool, "error": str|None, "label": str,
    "history": {"YYYY-MM-DD": float, ...}}``.
    """
    p = Path(saved_path)
    if not p.exists():
        return {"ok": False, "error": f"File not found: {saved_path}",
                "label": "", "history": {}}
    if not sheet or not header_row or not target_row:
        return {"ok": False, "error": "Missing sheet / header_row / target_row",
                "label": "", "history": {}}
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"Could not open workbook: {exc}",
                "label": "", "history": {}}
    try:
        if sheet not in wb.sheetnames:
            return {"ok": False, "error": f"Sheet '{sheet}' not found",
                    "label": "", "history": {}}
        ws = wb[sheet]
        # Walk the header row, collect (col_idx, normalized_date) pairs.
        date_cols: list[tuple[int, date]] = []
        for cell in ws[header_row]:
            d = normalize_to_month_end(cell.value)
            if d is not None:
                date_cols.append((cell.column, d))
        history: dict[str, float] = {}
        for col_idx, d in date_cols:
            v = _coerce_number(ws.cell(row=target_row, column=col_idx).value)
            if v is not None:
                history[d.isoformat()] = abs(v)
        label_cell = ws.cell(row=target_row, column=1)
        label = str(label_cell.value).strip() if label_cell.value else ""
        return {"ok": True, "error": None, "label": label, "history": history}
    finally:
        wb.close()


def extract_pool_labels(
    saved_path: str | Path,
    sheet: str,
    header_row: int,
    pool_name_col: str,
) -> dict[str, Any]:
    """Re-read pool/type labels from a specific column on a specific sheet.

    Used after the user manually adjusts the layout fields (sheet,
    header_row, pool_name_col) — auto-detection in :func:`_scan_sheet`
    only ever looks at column A, so when labels live elsewhere we need
    to re-scan against the user-supplied column.

    Returns ``{"ok": bool, "error": str|None, "labels": [...]}``.
    Labels are de-duplicated case-insensitively, preserving first-seen
    order, and rows that look like dates / numbers / total / subtotal
    rollups are skipped.
    """
    p = Path(saved_path)
    if not p.exists():
        return {"ok": False, "error": f"File not found: {saved_path}", "labels": []}
    if not sheet or not header_row or not pool_name_col:
        return {"ok": False,
                "error": "Missing sheet / header_row / pool_name_col",
                "labels": []}
    col_idx = _col_letter_to_idx(pool_name_col)
    if not col_idx:
        return {"ok": False,
                "error": f"Invalid column letter: {pool_name_col!r}",
                "labels": []}
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"Could not open workbook: {exc}",
                "labels": []}
    try:
        if sheet not in wb.sheetnames:
            return {"ok": False, "error": f"Sheet '{sheet}' not found",
                    "labels": []}
        ws = wb[sheet]
        labels: list[str] = []
        seen: set[str] = set()
        # Scan a generous window below the header; stop only on hard EOF.
        max_scan = max(header_row + 200, ws.max_row or (header_row + 200))
        for row_idx in range(header_row + 1, max_scan + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue
            if not _is_label_row(cell.value):
                continue
            s = str(cell.value).strip()
            if not s:
                continue
            key = s.lower()
            if key in seen:
                continue
            seen.add(key)
            labels.append(s)
        return {"ok": True, "error": None, "labels": labels}
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Pool-map auto-seeding
# ---------------------------------------------------------------------------

def seed_pool_map(
    parsed_labels: list[str],
    balance_title_map: dict[str, str] | None,
) -> tuple[dict[str, str], dict[str, str]]:
    """Return (pool_map, status_per_label).

    ``status_per_label`` is one of: ``matched``, ``ignored`` (mapped to
    empty in WARM), or ``new`` (label not present in WARM map).
    """
    btm = {k.strip().lower(): v for k, v in (balance_title_map or {}).items()}
    pool_map: dict[str, str] = {}
    status: dict[str, str] = {}
    for label in parsed_labels:
        key = label.strip().lower()
        if key in btm:
            value = (btm[key] or "").strip()
            if value:
                pool_map[label] = value
                status[label] = "matched"
            else:
                pool_map[label] = ""
                status[label] = "ignored"
        else:
            pool_map[label] = ""
            status[label] = "new"
    return pool_map, status


# ---------------------------------------------------------------------------
# Per-pool balance extraction (latest period)
# ---------------------------------------------------------------------------

def _coerce_number(v: Any) -> float | None:
    """Best-effort numeric coercion for a balance cell."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None
    s = str(v).strip()
    if not s:
        return None
    neg = False
    # Accountants commonly wrap negatives in parentheses; some workbooks
    # use angle brackets (e.g. "<561,006.75>") instead. Treat both as
    # negative-marker wrappers.
    if (s.startswith("(") and s.endswith(")")) or \
            (s.startswith("<") and s.endswith(">")):
        neg = True
        s = s[1:-1]
    s = s.replace("$", "").replace(",", "").replace(" ", "")
    if s.startswith("-"):
        neg = True
        s = s[1:]
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n


def _col_letter_to_idx(letter: str) -> int:
    """1-based column index from a letter (A->1)."""
    s = str(letter or "").strip().upper()
    if not s.isalpha():
        return 0
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def pool_balances_for_latest_period(
    saved_path: str | Path,
    sheet: str,
    header_row: int,
    pool_name_col: str,
    label_to_pool: dict[str, str] | None = None,
    period: str | None = None,
    exclude_labels: set[str] | list[str] | None = None,
) -> dict[str, Any]:
    """Read the monthly balance file and return per-pool balances for the
    latest detected period (or for ``period`` if given as ISO ``YYYY-MM-DD``).

    Returns::

        {
          "ok": bool,
          "error": str | None,
          "period": "YYYY-MM-DD",
          "by_pool": {pool_name: balance_float},
          "raw_rows": [{label, balance, mapped_pool}, ...],
        }

    ``label_to_pool`` is the wizard's ``monthly_bal.pool_map`` (raw label ->
    canonical pool name). Labels that map to an empty/None pool are
    treated as "ignored" and dropped from ``by_pool`` totals (but still
    appear in ``raw_rows``).

    ``exclude_labels`` is a case-insensitive set of raw labels that
    should NEVER contribute to ``by_pool`` totals - typically the
    parent/summary rows in a hierarchical Vizo-style 'Balance Sheets
    <CU>' workbook where each parent row is the sum of its sub-category
    rows beneath it. Excluded rows still appear in ``raw_rows`` (with
    ``mapped_pool`` cleared) so the UI can show them as ``Skipped``.
    """
    p = Path(saved_path)
    if not p.exists():
        return {"ok": False, "error": f"File not found: {saved_path}",
                "period": "", "by_pool": {}, "raw_rows": []}
    if not sheet or not header_row or not pool_name_col:
        return {"ok": False, "error": "Missing sheet / header_row / pool_name_col",
                "period": "", "by_pool": {}, "raw_rows": []}

    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"Could not open workbook: {exc}",
                "period": "", "by_pool": {}, "raw_rows": []}

    try:
        if sheet not in wb.sheetnames:
            return {"ok": False, "error": f"Sheet '{sheet}' not found",
                    "period": "", "by_pool": {}, "raw_rows": []}
        ws = wb[sheet]

        label_col_idx = _col_letter_to_idx(pool_name_col) or 1

        # Walk the header row, collect (col_idx, normalized_date) pairs.
        date_cols: list[tuple[int, date]] = []
        for cell in ws[header_row]:
            d = normalize_to_month_end(cell.value)
            if d is not None:
                date_cols.append((cell.column, d))
        if not date_cols:
            return {"ok": False,
                    "error": f"No date columns found on row {header_row}",
                    "period": "", "by_pool": {}, "raw_rows": []}

        # Pick the target period: explicit, else latest.
        # When ``period`` is provided we prefer an exact match; if no
        # column hits that month-end we fall back to the latest column
        # whose date is on or BEFORE the anchor (so a March anchor on a
        # Jan-Dec workbook picks the March column, not April or
        # December). Only when every column is AFTER the anchor do we
        # fall through to "latest" so a freshly-uploaded file with no
        # historical columns still produces something.
        target: tuple[int, date] | None = None
        if period:
            try:
                want = date.fromisoformat(period)
                for c, d in date_cols:
                    if d == want:
                        target = (c, d)
                        break
                if target is None:
                    on_or_before = [(c, d) for c, d in date_cols if d <= want]
                    if on_or_before:
                        target = max(on_or_before, key=lambda t: t[1])
            except ValueError:
                target = None
        if target is None:
            target = max(date_cols, key=lambda t: t[1])
        target_col, target_date = target

        ltp = {
            (k or "").strip().lower(): (v or "").strip()
            for k, v in (label_to_pool or {}).items()
        }
        excl = {
            (s or "").strip().lower()
            for s in (exclude_labels or [])
            if (s or "").strip()
        }

        by_pool: dict[str, float] = {}
        raw_rows: list[dict[str, Any]] = []
        seen_labels: set[str] = set()

        # Walk data rows below the header; stop after a stretch of blanks.
        max_row = ws.max_row or (header_row + 200)
        blanks = 0
        for r in range(header_row + 1, max_row + 1):
            label_cell = ws.cell(row=r, column=label_col_idx).value
            if not _is_label_row(label_cell):
                blanks += 1
                if blanks >= 15:
                    break
                continue
            blanks = 0
            label = str(label_cell).strip()
            key = label.lower()
            if key in seen_labels:
                continue
            seen_labels.add(key)

            bal = _coerce_number(
                ws.cell(row=r, column=target_col).value
            )
            is_excluded = key in excl
            mapped = "" if is_excluded else ltp.get(key, "")
            raw_rows.append({
                "label": label,
                "balance": bal,
                "mapped_pool": mapped,
                "excluded": is_excluded,
            })
            if mapped and bal is not None:
                by_pool[mapped] = by_pool.get(mapped, 0.0) + bal
    finally:
        wb.close()

    return {
        "ok": True,
        "error": None,
        "period": target_date.isoformat(),
        "by_pool": by_pool,
        "raw_rows": raw_rows,
    }


# ---------------------------------------------------------------------------
# Per-month (individual balance sheet) file analysis
# ---------------------------------------------------------------------------
#
# In ``per_month`` mode the CU sends one file per month-end (a typical
# "Detailed Balance Sheet" produced by their core processor). The file is a
# vertical layout: one column holds account labels, another holds the
# month-end balance for a single period. The wide-format auto-detector
# (``analyse_file`` above) does not apply.
#
# ``analyse_per_month_file`` opens the workbook (xls / xlsx / csv) and
# tries to locate the "LOANS" section, then identifies the label column
# and the balance column by scanning the rows beneath it. It returns the
# same surface shape ``analyse_file`` does (sheet / header_row /
# pool_name_col / parsed_pool_labels) plus a couple of per-month extras
# (``balance_col`` and ``detected_period``).

_LOAN_SECTION_PATTERNS = (
    "loans",
    "loan portfolio",
    "loan balances",
    "loans to members",
    "loans to member",
    "member loans",
    "member loan accounts",
    "loans receivable",
    "loan receivable",
    "loan accounts",
    "total loans to members",
    "loans and leases",
    "loans & leases",
)
# Substring fallback when no exact match was found. We only trip on
# strings that START WITH one of these phrases so detail rows like
# "Auto Loans" / "Used Auto Loans" don't accidentally anchor the
# section.
_LOAN_SECTION_START_PREFIXES = (
    "loans to ",
    "loans receivable",
    "loan accounts",
    "member loans",
    "loans and leases",
    "loans & leases",
)
# Exact-match section-end phrases (lower-cased, stripped).
_LOAN_SECTION_END = (
    "total loans", "net loans", "total loan", "accounts receivable",
    "total accounts receivable", "cash", "investments", "fixed assets",
    "other assets", "total assets", "liabilities", "equity",
)
# Substring patterns; if any of these appears anywhere in a row's text
# we treat the row as the end of the LOANS section. "total net loans"
# is the canonical hard-stop marker per CU convention; we also catch
# variations like "TOTAL GROSS LOANS" / "NET LOANS".
_LOAN_SECTION_END_SUBSTRINGS = (
    "total net loans", "total gross loans", "total loans", "net loans",
    "total receivables", "total cash", "total investments",
    "total fixed assets", "total other assets", "total assets",
    "total liabilities", "total equity",
)


def _is_loan_section_end(value: Any) -> bool:
    """Return True when ``value`` looks like a totals/section-end label.
    Substring match (case-insensitive) so "TOTAL NET LOANS" trips even
    though the exact-match list only has "total loans".
    """
    if not isinstance(value, str):
        return False
    s = value.strip().lower()
    if not s:
        return False
    if s in _LOAN_SECTION_END:
        return True
    return any(p in s for p in _LOAN_SECTION_END_SUBSTRINGS)
# Cells/labels we should never treat as a pool/account row.
_PER_MONTH_SKIP_PHRASES = (
    "loans", "loan portfolio", "balance", "as of", "produced",
    "in usd", "detailed balance sheet", "balance sheet",
)


def _iter_xls_rows(path: Path) -> tuple[str, list[list[Any]]] | None:
    """Read a legacy .xls file via xlrd. Returns (sheet_name, rows)."""
    try:
        import xlrd  # type: ignore[import-untyped]
    except ImportError:
        return None
    try:
        wb = xlrd.open_workbook(str(path))
    except Exception:  # noqa: BLE001
        return None
    best: tuple[str, list[list[Any]]] | None = None
    best_score = -1
    for sname in wb.sheet_names():
        s = wb.sheet_by_name(sname)
        rows: list[list[Any]] = []
        for r in range(s.nrows):
            row = [s.cell_value(r, c) for c in range(s.ncols)]
            rows.append(row)
        # Score by total non-empty cells.
        score = sum(1 for row in rows for v in row
                    if v not in (None, "", 0))
        if score > best_score:
            best_score = score
            best = (sname, rows)
    return best


def _iter_xlsx_rows(path: Path) -> tuple[str, list[list[Any]]] | None:
    """Read .xlsx via openpyxl. Returns (sheet_name, rows) for best sheet."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return None
    best: tuple[str, list[list[Any]]] | None = None
    best_score = -1
    try:
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            score = sum(1 for row in rows for v in row
                        if v not in (None, ""))
            if score > best_score:
                best_score = score
                best = (ws.title, rows)
    finally:
        wb.close()
    return best


def _iter_csv_rows(path: Path) -> tuple[str, list[list[Any]]] | None:
    """Read CSV. Returns (sheet_name, rows)."""
    import csv
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            rows = [list(r) for r in csv.reader(fh)]
    except Exception:  # noqa: BLE001
        return None
    return (path.stem, rows)


def _load_grid(path: Path) -> tuple[str, list[list[Any]]] | None:
    """Dispatch to the right reader based on extension. Returns
    ``(sheet_name, rows)`` or ``None``.
    """
    ext = path.suffix.lower()
    if ext == ".xls":
        return _iter_xls_rows(path)
    if ext in (".xlsx", ".xlsm"):
        return _iter_xlsx_rows(path)
    if ext == ".csv":
        return _iter_csv_rows(path)
    # Unknown / pdf: not supported here yet.
    return None


def _looks_like_money(v: Any) -> bool:
    """Loose check: cell value plausibly a balance ($1k+, finite)."""
    n = _coerce_number(v)
    if n is None:
        return False
    return abs(n) >= 1000


def _find_loan_section(rows: list[list[Any]]) -> int | None:
    """Return 0-based row index where a 'LOANS' section header sits, or
    ``None``. Match cell strings exactly equal to a known phrase (after
    lower/strip) to avoid grabbing 'Auto Loans' detail rows. Falls back
    to a starts-with match against ``_LOAN_SECTION_START_PREFIXES`` when
    no exact match is found, which catches phrasings like ``LOANS TO
    MEMBERS`` (common on credit-union GL-style balance sheets).
    """
    fallback: int | None = None
    for r, row in enumerate(rows):
        for v in row:
            if not isinstance(v, str):
                continue
            s = v.strip().lower()
            if not s:
                continue
            if s in _LOAN_SECTION_PATTERNS:
                return r
            if fallback is None and any(
                s.startswith(p) for p in _LOAN_SECTION_START_PREFIXES
            ):
                fallback = r
    return fallback


def _detect_period_from_rows(rows: list[list[Any]]) -> date | None:
    """Search the first ~20 rows for an 'As of: <date>' style cell."""
    rx = re.compile(r"as of[: ]+(.+)$", re.IGNORECASE)
    for row in rows[:20]:
        for v in row:
            if not isinstance(v, str):
                continue
            m = rx.search(v.strip())
            if m:
                d = normalize_to_month_end(m.group(1).strip())
                if d:
                    return d
    return None


def _detect_period_from_name(name: str) -> date | None:
    """Pull a date from a filename in any common layout.

    Supports (in priority order):
    - YYYYMMDD / YYYY-MM-DD / YYYY_MM_DD
    - MMDDYYYY / MM-DD-YYYY
    - MMDDYY    (e.g. ``123125`` -> 2025-12-31)
    - YYYYMM / YYYY-MM
    - MMYYYY
    Two-digit years are window-mapped to 2000-2069 / 1970-1999.
    Each pattern is tried against EVERY match in the stem so a leading
    member-number like ``878339`` doesn't shadow a trailing date.
    """
    stem = Path(name).stem

    def _yy_to_yyyy(yy: int) -> int:
        return 2000 + yy if yy < 70 else 1900 + yy

    # YYYYMMDD
    for m in re.finditer(
            r"(?<!\d)(20\d{2})[-_]?(\d{2})[-_]?(\d{2})(?!\d)", stem):
        try:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if 1 <= mo <= 12 and 1 <= d <= 31:
                return _last_day(y, mo)
        except ValueError:
            continue
    # MMDDYYYY
    for m in re.finditer(
            r"(?<!\d)(\d{2})[-_]?(\d{2})[-_]?(20\d{2})(?!\d)", stem):
        try:
            mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if 1 <= mo <= 12 and 1 <= d <= 31:
                return _last_day(y, mo)
        except ValueError:
            continue
    # MMDDYY (no separators) — e.g. "123125" -> 12/31/2025
    for m in re.finditer(r"(?<!\d)(\d{2})(\d{2})(\d{2})(?!\d)", stem):
        try:
            mo, d, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if 1 <= mo <= 12 and 1 <= d <= 31:
                return _last_day(_yy_to_yyyy(yy), mo)
        except ValueError:
            continue
    # YYYYMM / YYYY-MM
    for m in re.finditer(r"(?<!\d)(20\d{2})[-_]?(\d{2})(?!\d)", stem):
        try:
            y, mo = int(m.group(1)), int(m.group(2))
            if 1 <= mo <= 12:
                return _last_day(y, mo)
        except ValueError:
            continue
    # MMYYYY
    for m in re.finditer(r"(?<!\d)(\d{2})[-_]?(20\d{2})(?!\d)", stem):
        try:
            mo, y = int(m.group(1)), int(m.group(2))
            if 1 <= mo <= 12:
                return _last_day(y, mo)
        except ValueError:
            continue
    return None


def _extract_period_from_cell(
    rows: list[list[Any]], cell_ref: str,
) -> date | None:
    """Read an explicit Excel-style cell reference (e.g. ``B3``) from the
    pre-loaded grid and snap it to a month-end. Strips a leading
    ``"As of:"`` if present.
    """
    ref = (cell_ref or "").strip()
    if not ref:
        return None
    try:
        col_letter, row_num = coordinate_from_string(ref)
        col_idx = column_index_from_string(col_letter) - 1
        r = int(row_num) - 1
    except Exception:
        return None
    if r < 0 or r >= len(rows):
        return None
    row = rows[r]
    if col_idx < 0 or col_idx >= len(row):
        return None
    val = row[col_idx]
    if val is None:
        return None
    # Normalise via the standard helper; strip any leading "As of:" label.
    if isinstance(val, str):
        cleaned = re.sub(r"^\s*as of[:\s]*", "", val,
                         flags=re.IGNORECASE).strip()
        return normalize_to_month_end(cleaned) if cleaned else None
    return normalize_to_month_end(val)


def analyse_per_month_file(
    path: str | Path,
    stop_row: int | None = None,
    as_of_cell: str | None = None,
    acl_row: int | None = None,
    acl_col: str | None = None,
) -> dict[str, Any]:
    """Detect layout + pool labels in a single-month balance-sheet file.

    ``stop_row`` (1-based) is an optional hard cap: any row at or below
    this row index is ignored. When omitted the parser auto-stops at
    the first "TOTAL NET LOANS" / "TOTAL LOANS" / "NET LOANS" marker.

    ``as_of_cell`` is an optional Excel-style cell reference (e.g.
    ``"B3"``). When supplied, the period is read from that cell first
    and only falls back to the auto-detection (in-file ``As of:`` row,
    then filename) if the cell is empty / unparseable.

    Returns::

        {
          "ok": bool,
          "error": str | None,
          "sheet": str,                # sheet name (or filename stem for csv)
          "header_row": int,           # 1-based row above the first data row
          "pool_name_col": str,        # column letter of the label column
          "balance_col": str,          # column letter of the balance column
          "parsed_pool_labels": [str], # labels found in the LOANS section
          "rows": [{"label": str, "balance": float|None}],
          "detected_period": str,      # ISO YYYY-MM-DD or ""
          "stop_row": int | None,      # 1-based row that ended scanning
          "auto_stop_row": int | None, # 1-based auto-detected end row
        }
    """
    p = Path(path)
    if not p.exists():
        return {"ok": False, "error": f"File not found: {path}",
                "sheet": "", "header_row": 0, "pool_name_col": "",
                "balance_col": "", "parsed_pool_labels": [], "rows": [],
                "detected_period": ""}

    loaded = _load_grid(p)
    if loaded is None:
        return {"ok": False,
                "error": f"Unsupported or unreadable file type: {p.suffix}",
                "sheet": "", "header_row": 0, "pool_name_col": "",
                "balance_col": "", "parsed_pool_labels": [], "rows": [],
                "detected_period": ""}
    sheet_name, rows = loaded
    if not rows:
        return {"ok": False, "error": "Empty workbook",
                "sheet": sheet_name, "header_row": 0, "pool_name_col": "",
                "balance_col": "", "parsed_pool_labels": [], "rows": [],
                "detected_period": ""}

    # Period: explicit cell wins, then in-file "As of:" marker, then
    # the filename, else "".
    period = (
        _extract_period_from_cell(rows, as_of_cell or "")
        or _detect_period_from_rows(rows)
        or _detect_period_from_name(p.name)
    )
    period_iso = period.isoformat() if period else ""

    # Find the LOANS section header.
    loans_idx = _find_loan_section(rows)
    if loans_idx is None:
        return {"ok": False,
                "error": "Could not find a 'LOANS' section in the file.",
                "sheet": sheet_name, "header_row": 0, "pool_name_col": "",
                "balance_col": "", "parsed_pool_labels": [], "rows": [],
                "detected_period": period_iso}

    # Pick the label column = the column the "LOANS" header text sits in
    # (or the closest non-numeric text column to its left if it's at col 0).
    loans_row = rows[loans_idx]
    label_col_idx: int | None = None
    for c, v in enumerate(loans_row):
        if isinstance(v, str) and v.strip().lower() in _LOAN_SECTION_PATTERNS:
            label_col_idx = c
            break
    if label_col_idx is None:
        return {"ok": False, "error": "Could not pin the label column",
                "sheet": sheet_name, "header_row": 0, "pool_name_col": "",
                "balance_col": "", "parsed_pool_labels": [], "rows": [],
                "detected_period": period_iso}

    # Walk rows below the LOANS header; pool/account labels usually sit
    # one column to the LEFT of the section header (so that section
    # headers like "LOANS" / "TOTAL LOANS" stand out). Detect: scan the
    # first ~15 non-empty rows below to find which column holds short
    # text labels with adjacent numeric balances. The detail-label column
    # is the one with the most rows whose text != one of the section
    # phrases AND whose row also has a money cell elsewhere.
    detail_col_counts: dict[int, int] = {}
    balance_col_counts: dict[int, int] = {}
    end_idx = len(rows)
    auto_end_idx = len(rows)
    for r in range(loans_idx + 1, min(loans_idx + 80, len(rows))):
        row = rows[r]
        # Stop at section end (TOTAL NET LOANS / TOTAL LOANS / etc.) —
        # substring match, so "TOTAL NET LOANS" trips even though
        # the exact-match list has only "total loans".
        is_end = False
        for v in row:
            if _is_loan_section_end(v):
                is_end = True
                auto_end_idx = min(auto_end_idx, r)
                end_idx = min(end_idx, r)
                break
        if is_end:
            break
        # Identify text + money columns in this row.
        text_cols = [c for c, v in enumerate(row)
                     if isinstance(v, str) and v.strip()
                     and v.strip().lower() not in _PER_MONTH_SKIP_PHRASES]
        money_cols = [c for c, v in enumerate(row) if _looks_like_money(v)]
        if text_cols and money_cols:
            for c in text_cols:
                detail_col_counts[c] = detail_col_counts.get(c, 0) + 1
            for c in money_cols:
                balance_col_counts[c] = balance_col_counts.get(c, 0) + 1

    if not detail_col_counts or not balance_col_counts:
        return {"ok": False,
                "error": "Found a LOANS section but no labelled balance "
                         "rows beneath it.",
                "sheet": sheet_name, "header_row": loans_idx + 1,
                "pool_name_col": "", "balance_col": "",
                "parsed_pool_labels": [], "rows": [],
                "detected_period": period_iso}

    detail_col = max(detail_col_counts.items(), key=lambda kv: kv[1])[0]
    # Balance sheets always show balances to the RIGHT of the label
    # column. When a column to the LEFT of the label has more numeric
    # cells (e.g. a GL-account-number column where every row carries a
    # 6-digit code) we'd otherwise pick it as the balance column and
    # produce huge bogus totals. Restrict the balance-column candidates
    # to those at or right of the label column when at least one such
    # candidate exists; otherwise fall back to the global max.
    right_candidates = {c: n for c, n in balance_col_counts.items()
                        if c > detail_col}
    if right_candidates:
        balance_col = max(right_candidates.items(),
                          key=lambda kv: kv[1])[0]
    else:
        balance_col = max(balance_col_counts.items(),
                          key=lambda kv: kv[1])[0]

    # Honour an explicit user-supplied stop_row (1-based). When the
    # user sets stop_row=N we treat row N as the first EXCLUDED row
    # (i.e. "stop scanning at row N"). Always wins over auto-detect
    # if it sits above the auto stop.
    effective_end_idx = end_idx
    if isinstance(stop_row, int) and stop_row > 0:
        manual_end = max(loans_idx + 1, stop_row - 1)
        effective_end_idx = min(effective_end_idx, manual_end)

    # Now extract labels + balances from the loans section.
    parsed_labels: list[str] = []
    extracted: list[dict[str, Any]] = []
    seen: set[str] = set()
    for r in range(loans_idx + 1, effective_end_idx):
        row = rows[r]
        if detail_col >= len(row):
            continue
        cell = row[detail_col]
        if not isinstance(cell, str):
            continue
        label = cell.strip()
        if not label:
            continue
        if label.lower() in _PER_MONTH_SKIP_PHRASES:
            continue
        if _is_loan_section_end(label):
            break
        bal = _coerce_number(row[balance_col]) \
            if balance_col < len(row) else None
        if bal is None:
            # Skip section labels that have no balance on their row.
            continue
        key = label.lower()
        if key in seen:
            continue
        seen.add(key)
        parsed_labels.append(label)
        extracted.append({"label": label, "balance": bal})

    if not parsed_labels:
        return {"ok": False,
                "error": "Loans section found but no pool/account labels "
                         "with balances were extractable.",
                "sheet": sheet_name, "header_row": loans_idx + 1,
                "pool_name_col": get_column_letter(detail_col + 1),
                "balance_col": get_column_letter(balance_col + 1),
                "parsed_pool_labels": [], "rows": [],
                "detected_period": period_iso}

    # Detect ACL/ALLL line. The row may sit far below the LOANS section
    # (typically under 'Other Liabilities' or as a contra-asset). Scan
    # the whole grid; prefer hits in the same label column we picked.
    # When the caller passed an explicit ``acl_row`` (1-based), honour
    # it instead of auto-detecting -- the user has already chosen.
    if isinstance(acl_row, int) and acl_row > 0 and acl_row - 1 < len(rows):
        ridx = acl_row - 1
        row = rows[ridx]
        # Try the same label column as the loans section, then any
        # string cell on the row.
        lbl_val: Any = ""
        if detail_col < len(row):
            lbl_val = row[detail_col]
        if not isinstance(lbl_val, str) or not lbl_val.strip():
            for v in row:
                if isinstance(v, str) and v.strip():
                    lbl_val = v
                    break
        acl_row_idx = ridx
        acl_label = (str(lbl_val).strip() if isinstance(lbl_val, str) else "")
    else:
        acl_row_idx, acl_label, _acl_col = _find_acl_row_in_grid(
            rows, preferred_col=detail_col)
    # Pick the column to read the ACL value from. User override wins;
    # otherwise fall back to the auto-detected balance column for the
    # loans section.
    acl_col_idx = balance_col
    if isinstance(acl_col, str) and acl_col.strip():
        ovr = _col_letter_to_idx(acl_col) - 1
        if ovr >= 0:
            acl_col_idx = ovr
    acl_value: float | None = None
    if acl_row_idx is not None and acl_col_idx < len(rows[acl_row_idx]):
        acl_value = _coerce_number(rows[acl_row_idx][acl_col_idx])
        if acl_value is not None:
            acl_value = abs(acl_value)

    return {
        "ok": True,
        "error": None,
        "sheet": sheet_name,
        "header_row": loans_idx + 1,           # 1-based row of "LOANS" cell
        "pool_name_col": get_column_letter(detail_col + 1),
        "balance_col": get_column_letter(balance_col + 1),
        "parsed_pool_labels": parsed_labels,
        "rows": extracted,
        "detected_period": period_iso,
        "acl_row": (acl_row_idx + 1) if acl_row_idx is not None else None,
        "acl_label": acl_label,
        "acl_value": acl_value,
        "acl_col": get_column_letter(acl_col_idx + 1)
            if acl_col_idx is not None and acl_col_idx >= 0 else "",
        "auto_acl_col": get_column_letter(balance_col + 1),
        "auto_stop_row": (auto_end_idx + 1) if auto_end_idx < len(rows) else None,
        "stop_row": (effective_end_idx + 1)
            if effective_end_idx < len(rows) else None,
    }


def pool_balances_for_per_month_files(
    monthly_files: list[dict[str, Any]],
    layout: dict[str, Any],
    label_to_pool: dict[str, str] | None = None,
    exclude_labels: set[str] | list[str] | None = None,
) -> dict[str, Any]:
    """Aggregate per-pool balances across a list of single-month files.

    ``monthly_files`` is the wizard's ``monthly_bal.monthly_files`` list:
    each entry has ``{filename, saved_path, period}``. ``layout`` is
    ``monthly_bal.per_month_layout`` with ``sheet`` / ``label_col`` /
    ``balance_col`` / ``header_row`` keys.

    Returns ``{ok, error, by_period: {period: {by_pool: {pool: amt},
    raw_rows: [...]}}}``.
    """
    sheet = (layout or {}).get("sheet", "")
    label_col = (layout or {}).get("label_col", "A") or "A"
    balance_col = (layout or {}).get("balance_col", "B") or "B"
    try:
        header_row = int((layout or {}).get("header_row") or 1)
    except (TypeError, ValueError):
        header_row = 1
    try:
        stop_row_raw = (layout or {}).get("stop_row")
        stop_row = int(stop_row_raw) if stop_row_raw else 0
    except (TypeError, ValueError):
        stop_row = 0
    label_col_idx = _col_letter_to_idx(label_col) or 1
    balance_col_idx = _col_letter_to_idx(balance_col) or 2

    ltp = {
        (k or "").strip().lower(): (v or "").strip()
        for k, v in (label_to_pool or {}).items()
    }
    excl = {
        (s or "").strip().lower()
        for s in (exclude_labels or [])
        if (s or "").strip()
    }

    by_period: dict[str, dict[str, Any]] = {}
    errors: list[str] = []
    for entry in (monthly_files or []):
        period = (entry.get("period") or "").strip()
        saved_path = entry.get("saved_path") or ""
        if not period or not saved_path:
            continue
        p = Path(saved_path)
        if not p.exists():
            errors.append(f"{period}: file missing ({p.name})")
            continue
        loaded = _load_grid(p)
        if loaded is None:
            errors.append(f"{period}: unreadable file ({p.name})")
            continue
        _sn, rows = loaded
        # If the user gave a sheet name and we have an xlsx, prefer that
        # sheet specifically. (Our _load_grid currently picks the densest
        # sheet; that's usually correct for these one-tab balance sheets.)
        by_pool: dict[str, float] = {}
        raw_rows: list[dict[str, Any]] = []
        seen: set[str] = set()
        start = max(header_row, 1)
        # When the user supplied a manual stop_row we cap the scan there;
        # otherwise we scan the whole grid and rely on substring totals
        # detection to bail out.
        end_exclusive = len(rows)
        if stop_row and stop_row > start:
            end_exclusive = min(end_exclusive, stop_row - 1)
        hit_end = False
        for r in range(start, end_exclusive):
            row = rows[r]
            # Stop at the first totals/end marker anywhere in the row.
            for v in row:
                if _is_loan_section_end(v):
                    hit_end = True
                    break
            if hit_end:
                break
            if label_col_idx - 1 >= len(row):
                continue
            cell = row[label_col_idx - 1]
            if not isinstance(cell, str):
                continue
            label = cell.strip()
            if not label:
                continue
            if label.lower() in _PER_MONTH_SKIP_PHRASES:
                continue
            if _is_loan_section_end(label):
                # Stop at the first totals/end marker.
                break
            key = label.lower()
            if key in seen:
                continue
            seen.add(key)
            bal = (_coerce_number(row[balance_col_idx - 1])
                   if balance_col_idx - 1 < len(row) else None)
            is_excluded = key in excl
            mapped = "" if is_excluded else ltp.get(key, "")
            raw_rows.append({"label": label, "balance": bal,
                             "mapped_pool": mapped,
                             "excluded": is_excluded})
            if mapped and bal is not None:
                by_pool[mapped] = by_pool.get(mapped, 0.0) + bal
        by_period[period] = {"by_pool": by_pool, "raw_rows": raw_rows}

    return {"ok": True,
            "error": "; ".join(errors) if errors else None,
            "by_period": by_period}


# ---------------------------------------------------------------------------
# Per-YEAR balance-sheet files
# ---------------------------------------------------------------------------
#
# Some CUs (e.g. Census FCU) keep a single workbook per calendar year with
# all 12 month-end balances arranged as columns. Layout (Census 2024):
#
#   Row 2:  | | | DEC 2024 | NOV 2024 | OCT 2024 | ... | JAN 2024 | DEC 2023 | NET CHANGE
#   Row 5:  | LOANS
#   Row 6+: <#> | <Pool Name> | <bal> | <bal> | ...
#   Row N:  | TOTAL LOANS    (stop here)
#
# ``analyse_per_year_file`` finds the header row + label/period columns; the
# runtime aggregator ``pool_balances_for_per_year_files`` walks the LOANS
# section once per year file and emits per-period rows.


def _parse_period_header_cell(v: Any) -> date | None:
    """Parse a single header cell into a month-end date, or None.

    Accepts datetime objects and the common ``"MMM YYYY"`` / ``"MMM YY"``
    spelling used in the Census workbook header row (``"DEC 2024"``,
    ``"SEPT 2024"``, ``"JULY 2024"`` …). Falls back to the general
    ``normalize_to_month_end`` for ISO and slash dates.
    """
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return _last_day(v.year, v.month)
    if isinstance(v, date):
        return _last_day(v.year, v.month)
    if not isinstance(v, str):
        return None
    s = v.strip()
    if not s:
        return None
    # "MONTH YEAR" (case-insensitive). Try the explicit per-year header
    # spelling first so we don't accidentally hit fallback paths.
    m = re.match(r"^([A-Za-z]+)\s+(\d{2,4})$", s)
    if m:
        mo = _MONTHS.get(m.group(1).lower())
        y = int(m.group(2))
        if y < 100:
            y += 2000 if y < 70 else 1900
        if mo:
            return _last_day(y, mo)
    return normalize_to_month_end(s)


def _detect_year_from_name(name: str) -> int | None:
    m = re.search(r"(20\d{2})", name)
    if m:
        return int(m.group(1))
    return None


def analyse_per_year_file(path: str | Path) -> dict[str, Any]:
    """Detect layout + pool labels in a single-year balance-sheet file.

    Returns::

        {
          "ok": bool,
          "error": str | None,
          "sheet": str,
          "header_row": int,           # 1-based row carrying the period labels
          "label_col": str,            # column letter of the pool-name column
          "period_columns": [          # one entry per detected month-end column
              {"col": "C", "period": "2024-12-31"},
              ...
          ],
          "pool_labels": [str],
          "detected_year": int | None,
        }
    """
    p = Path(path)
    empty = {
        "ok": False, "error": "", "sheet": "", "header_row": 0,
        "label_col": "", "period_columns": [], "pool_labels": [],
        "detected_year": _detect_year_from_name(p.name),
    }
    if not p.exists():
        empty["error"] = f"File not found: {path}"
        return empty

    loaded = _load_grid(p)
    if loaded is None:
        empty["error"] = f"Unsupported or unreadable file type: {p.suffix}"
        return empty
    sheet_name, rows = loaded
    if not rows:
        empty["sheet"] = sheet_name
        empty["error"] = "Sheet is empty."
        return empty

    # Find the header row: pick the row in the first 15 with the most
    # parseable month-year cells (>=3).
    best_hdr_idx: int | None = None
    best_hdr_cols: list[tuple[int, date]] = []
    for r in range(min(15, len(rows))):
        parsed_here: list[tuple[int, date]] = []
        for c, v in enumerate(rows[r]):
            d = _parse_period_header_cell(v)
            if d:
                parsed_here.append((c, d))
        if len(parsed_here) >= 3 and len(parsed_here) > len(best_hdr_cols):
            best_hdr_idx = r
            best_hdr_cols = parsed_here
    if best_hdr_idx is None:
        empty["sheet"] = sheet_name
        empty["error"] = ("No header row with month-year labels was found "
                          "in the first 15 rows.")
        return empty

    # Restrict to a single calendar year when possible — use the filename
    # year hint to filter out the trailing 'DEC 2023' carryover.
    yr_hint = _detect_year_from_name(p.name)
    if yr_hint:
        filtered = [(c, d) for c, d in best_hdr_cols if d.year == yr_hint]
        if filtered:
            best_hdr_cols = filtered

    # Pick the label column: the first column to the LEFT of the leftmost
    # period column whose value at any LOANS-section row is a non-empty
    # text label.
    leftmost_period_col = min(c for c, _ in best_hdr_cols)
    loans_idx = _find_loan_section(rows[best_hdr_idx + 1:])
    if loans_idx is None:
        empty["sheet"] = sheet_name
        empty["header_row"] = best_hdr_idx + 1
        empty["error"] = ("Header row detected but no LOANS section found "
                          "below it.")
        return empty
    loans_idx += best_hdr_idx + 1  # back to absolute row index

    label_col: int | None = None
    for c in range(leftmost_period_col - 1, -1, -1):
        text_hits = 0
        for r in range(loans_idx + 1,
                       min(loans_idx + 30, len(rows))):
            if c >= len(rows[r]):
                continue
            v = rows[r][c]
            if isinstance(v, str) and v.strip():
                text_hits += 1
        if text_hits >= 3:
            label_col = c
            break
    if label_col is None:
        empty["sheet"] = sheet_name
        empty["header_row"] = best_hdr_idx + 1
        empty["error"] = "Could not locate the pool-name column."
        return empty

    # Walk from LOANS row downward, collecting pool labels until we hit a
    # TOTAL/section-end marker.
    pool_labels: list[str] = []
    seen: set[str] = set()
    for r in range(loans_idx + 1, len(rows)):
        row = rows[r]
        if label_col >= len(row):
            continue
        cell = row[label_col]
        if not isinstance(cell, str):
            continue
        label = cell.strip()
        if not label:
            continue
        lc = label.lower()
        if lc in _LOAN_SECTION_END:
            break
        if lc in _PER_MONTH_SKIP_PHRASES:
            continue
        # Require at least one numeric value in any of the period columns.
        has_num = False
        for c, _d in best_hdr_cols:
            if c < len(row) and _coerce_number(row[c]) is not None:
                has_num = True
                break
        if not has_num:
            continue
        if lc in seen:
            continue
        seen.add(lc)
        pool_labels.append(label)

    if not pool_labels:
        return {
            "ok": False,
            "error": "No pool rows with numeric balances were found.",
            "sheet": sheet_name, "header_row": best_hdr_idx + 1,
            "label_col": get_column_letter(label_col + 1),
            "period_columns": [], "pool_labels": [],
            "detected_year": yr_hint,
        }

    period_columns = [
        {"col": get_column_letter(c + 1), "period": d.isoformat()}
        for c, d in sorted(best_hdr_cols, key=lambda t: t[1])
    ]

    # Detect ACL/ALLL line. Scan the whole grid (it usually lives below
    # the LOANS section in 'Other Liabilities' or as a contra-asset);
    # prefer the same label column we chose for pools.
    acl_row_idx, acl_label, _acl_col = _find_acl_row_in_grid(
        rows, preferred_col=label_col)
    acl_history: dict[str, float] = {}
    if acl_row_idx is not None:
        acl_row_data = rows[acl_row_idx]
        for c, d in best_hdr_cols:
            if c < len(acl_row_data):
                v = _coerce_number(acl_row_data[c])
                if v is not None:
                    acl_history[d.isoformat()] = abs(v)

    return {
        "ok": True, "error": None,
        "sheet": sheet_name, "header_row": best_hdr_idx + 1,
        "label_col": get_column_letter(label_col + 1),
        "period_columns": period_columns,
        "pool_labels": pool_labels,
        "detected_year": yr_hint,
        "acl_row": (acl_row_idx + 1) if acl_row_idx is not None else None,
        "acl_label": acl_label,
        "acl_history": acl_history,
    }


def pool_balances_for_per_year_files(
    year_files: list[dict[str, Any]],
    layout: dict[str, Any],
    label_to_pool: dict[str, str] | None = None,
    exclude_labels: set[str] | list[str] | None = None,
) -> dict[str, Any]:
    """Aggregate per-pool balances across a list of single-year files.

    ``year_files`` is the wizard's ``monthly_bal.year_files`` list:
    each entry has ``{filename, saved_path, year}``. ``layout`` is
    ``monthly_bal.per_year_layout`` with ``sheet`` / ``header_row`` /
    ``label_col`` keys; per-file period columns are re-detected by
    re-parsing the header row using ``_parse_period_header_cell`` so
    different years (whose column letters typically match) all line
    up. If the wizard has stored an explicit ``period_columns`` list
    it is used as a fallback when the per-file header can't be read.

    Returns ``{ok, error, by_period: {period_iso: {by_pool: {pool: amt},
    raw_rows: [...]}}}``.
    """
    layout = layout or {}
    label_col = (layout.get("label_col") or "A").upper()
    try:
        header_row = int(layout.get("header_row") or 1)
    except (TypeError, ValueError):
        header_row = 1
    fallback_period_cols = layout.get("period_columns") or []
    label_col_idx = _col_letter_to_idx(label_col) or 1

    ltp = {
        (k or "").strip().lower(): (v or "").strip()
        for k, v in (label_to_pool or {}).items()
    }
    excl = {
        (s or "").strip().lower()
        for s in (exclude_labels or [])
        if (s or "").strip()
    }

    by_period: dict[str, dict[str, Any]] = {}
    errors: list[str] = []
    for entry in (year_files or []):
        saved_path = entry.get("saved_path") or ""
        year_hint = entry.get("year")
        if not saved_path:
            continue
        p = Path(saved_path)
        if not p.exists():
            errors.append(f"{year_hint or p.name}: file missing")
            continue
        loaded = _load_grid(p)
        if loaded is None:
            errors.append(f"{p.name}: unreadable file")
            continue
        _sn, rows = loaded

        # Re-detect period columns on this file's header row so the
        # importer doesn't depend on every file sharing the same
        # column layout. Fall back to layout.period_columns if header
        # parse misses.
        period_cols: list[tuple[int, date]] = []
        if 1 <= header_row <= len(rows):
            for c, v in enumerate(rows[header_row - 1]):
                d = _parse_period_header_cell(v)
                if d:
                    period_cols.append((c, d))
        # Filter to the year hint if we have one (drops trailing prior-Dec).
        if year_hint:
            try:
                yi = int(year_hint)
                filt = [(c, d) for c, d in period_cols if d.year == yi]
                if filt:
                    period_cols = filt
            except (TypeError, ValueError):
                pass
        if not period_cols and fallback_period_cols:
            for spec in fallback_period_cols:
                ci = _col_letter_to_idx(spec.get("col") or "") or 0
                try:
                    d = date.fromisoformat(spec.get("period") or "")
                except (TypeError, ValueError):
                    continue
                if ci:
                    period_cols.append((ci - 1, _last_day(d.year, d.month)))
        if not period_cols:
            errors.append(f"{p.name}: no period columns detected")
            continue

        for c, d in period_cols:
            period_iso = d.isoformat()
            slot = by_period.setdefault(
                period_iso, {"by_pool": {}, "raw_rows": []})
            seen: set[str] = set()
            for r in range(header_row, len(rows)):
                row = rows[r]
                if label_col_idx - 1 >= len(row):
                    continue
                cell = row[label_col_idx - 1]
                if not isinstance(cell, str):
                    continue
                label = cell.strip()
                if not label:
                    continue
                lc = label.lower()
                if lc in _LOAN_SECTION_END:
                    break
                if lc in _PER_MONTH_SKIP_PHRASES:
                    continue
                key = lc
                if key in seen:
                    continue
                seen.add(key)
                bal = (_coerce_number(row[c]) if c < len(row) else None)
                is_excluded = key in excl
                mapped = "" if is_excluded else ltp.get(key, "")
                slot["raw_rows"].append({
                    "label": label, "balance": bal,
                    "mapped_pool": mapped, "period": period_iso,
                    "excluded": is_excluded,
                })
                if mapped and bal is not None:
                    slot["by_pool"][mapped] = (
                        slot["by_pool"].get(mapped, 0.0) + bal)

    return {"ok": True,
            "error": "; ".join(errors) if errors else None,
            "by_period": by_period}

