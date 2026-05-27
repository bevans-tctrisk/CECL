"""Helpers for the "monthly loan-data extracts" historical source.

Used by the no-WARM Historical Data step when the credit union is going to
provide one loan-data extract per month and have the wizard build the
historical balance database from them.

This module is intentionally narrow:

* read the header row of an extract file (xlsx / xlsm / xls / csv),
* compute a deterministic header signature so we can recognise files that
  share the same column layout,
* detect the as-of date for an extract from its filename (and file mtime
  as a last-resort fallback).

Actual roll-up / write to the database lives in a separate (later) slice.
"""
from __future__ import annotations

import hashlib
import re
from calendar import monthrange
from datetime import date, datetime
from pathlib import Path
from typing import Any


# ---------------------------------------------------------------------------
# Header reading + signature
# ---------------------------------------------------------------------------

def read_extract_headers(path: Path, sheet: str | None = None) -> dict[str, Any]:
    """Return the header row of an extract file.

    Result::

        {"ok": bool, "error": str | None, "sheet": str, "headers": [str, ...]}

    For Excel workbooks we read row 1 of the first visible sheet (or the
    sheet name supplied by the caller).  For CSV we read the first row.
    """
    out: dict[str, Any] = {"ok": False, "error": None, "sheet": "", "headers": []}
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            import csv
            with path.open("r", encoding="utf-8-sig", newline="") as fh:
                reader = csv.reader(fh)
                row = next(reader, [])
            out["headers"] = [_clean_header(c) for c in row]
            out["sheet"] = ""
            out["ok"] = True
            return out

        # Excel — prefer openpyxl for .xlsx/.xlsm, fall back to pandas for .xls.
        if suffix in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            target = sheet
            if not target:
                for name in wb.sheetnames:
                    ws = wb[name]
                    if getattr(ws, "sheet_state", "visible") == "visible":
                        target = name
                        break
                if not target:
                    target = wb.sheetnames[0]
            ws = wb[target]
            header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            out["sheet"] = target
            out["headers"] = [_clean_header(c) for c in header_cells]
            out["ok"] = True
            wb.close()
            return out

        # .xls or anything pandas can read.
        import pandas as pd
        df = pd.read_excel(path, sheet_name=sheet or 0, nrows=0)
        out["headers"] = [_clean_header(c) for c in df.columns]
        out["sheet"] = sheet or ""
        out["ok"] = True
        return out
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Could not read header row: {exc}"
        return out


def _clean_header(cell: Any) -> str:
    if cell is None:
        return ""
    return str(cell).strip()


def compute_header_signature(headers: list[str]) -> str:
    """Stable signature for a list of column headers.

    Whitespace and case are normalised so that "Member ID" and "member id "
    map to the same signature.  Empty trailing columns are trimmed.
    """
    norm = [re.sub(r"\s+", " ", h).strip().lower() for h in headers]
    while norm and not norm[-1]:
        norm.pop()
    joined = "\u241f".join(norm)  # unit-separator-ish char, unlikely in a header
    return hashlib.sha1(joined.encode("utf-8")).hexdigest()[:16]


# ---------------------------------------------------------------------------
# As-of date detection
# ---------------------------------------------------------------------------

_MONTH_NAMES = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


def detect_as_of_date(filename: str, path: Path | None = None) -> dict[str, Any]:
    """Infer an extract's as-of date from its filename.

    Returns::

        {"date": "YYYY-MM-DD" | "",
         "source": "filename" | "mtime" | "",
         "confidence": "high" | "medium" | "low" | "none"}

    "high"   – filename includes a full Y-M-D match.
    "medium" – filename includes Y-M; we assume end-of-month.
    "low"    – we fell back to file mtime.
    "none"   – nothing matched; caller should ask the user.
    """
    out: dict[str, Any] = {"date": "", "source": "", "confidence": "none"}
    base = (filename or "").replace("\\", "/").rsplit("/", 1)[-1]

    # YYYY-MM-DD / YYYY_MM_DD / YYYYMMDD
    m = re.search(r"(?<!\d)(20\d{2})[-_]?(\d{2})[-_]?(\d{2})(?!\d)", base)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        iso = _safe_date(y, mo, d)
        if iso:
            return {"date": iso, "source": "filename", "confidence": "high"}

    # MM-DD-YYYY / MM_DD_YYYY / MMDDYYYY  (US style)
    m = re.search(r"(?<!\d)(\d{2})[-_]?(\d{2})[-_]?(20\d{2})(?!\d)", base)
    if m:
        mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        iso = _safe_date(y, mo, d)
        if iso:
            return {"date": iso, "source": "filename", "confidence": "high"}

    # YYYY-MM or YYYY_MM  (no day → end of month)
    m = re.search(r"(?<!\d)(20\d{2})[-_](\d{2})(?!\d)", base)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        iso = _end_of_month(y, mo)
        if iso:
            return {"date": iso, "source": "filename", "confidence": "medium"}

    # YYYYMM (six digits) — but only if it isn't really a YYYYMMDD captured above.
    m = re.search(r"(?<!\d)(20\d{2})(\d{2})(?!\d)", base)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12:
            iso = _end_of_month(y, mo)
            if iso:
                return {"date": iso, "source": "filename", "confidence": "medium"}

    # "Mar 2026", "March_2026", "Mar-2026"
    m = re.search(
        r"(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|"
        r"jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|"
        r"oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)[\s\-_]*?(20\d{2})",
        base,
        re.IGNORECASE,
    )
    if m:
        mo = _MONTH_NAMES.get(m.group(1).lower())
        y = int(m.group(2))
        if mo:
            iso = _end_of_month(y, mo)
            if iso:
                return {"date": iso, "source": "filename", "confidence": "medium"}

    # Last resort: file modified time (rounded to end-of-month).
    if path is not None:
        try:
            mtime = datetime.fromtimestamp(path.stat().st_mtime)
            iso = _end_of_month(mtime.year, mtime.month)
            if iso:
                return {"date": iso, "source": "mtime", "confidence": "low"}
        except Exception:  # noqa: BLE001
            pass

    return out


def _safe_date(y: int, m: int, d: int) -> str:
    try:
        return date(y, m, d).isoformat()
    except ValueError:
        return ""


def _end_of_month(y: int, m: int) -> str:
    try:
        last = monthrange(y, m)[1]
        return date(y, m, last).isoformat()
    except ValueError:
        return ""


# ---------------------------------------------------------------------------
# Folder scan
# ---------------------------------------------------------------------------

_EXTRACT_EXTS = {".xlsx", ".xlsm", ".xls", ".csv"}


def _period_key(iso_date: str) -> str:
    """``"2026-03-31"`` -> ``"2026-03"``.  Returns ``""`` on bad input."""
    if not iso_date or len(iso_date) < 7:
        return ""
    return iso_date[:7]


def _months_back(target_iso: str, count: int) -> list[str]:
    """Return ``["YYYY-MM", ...]`` for the target month going back ``count``
    months inclusive of the target month.  Newest first.
    """
    if not target_iso:
        return []
    try:
        y, m, _d = (int(x) for x in target_iso.split("-")[:3])
    except (ValueError, AttributeError):
        return []
    months: list[str] = []
    for _ in range(max(1, count)):
        months.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return months


def scan_folder(
    folder: Path,
    target_period: str,
    history_months: int,
    known_profiles: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Walk a folder of monthly loan-data extracts and report what's there.

    The scanner reads only the header row of each candidate file, so it
    stays cheap even on large folders.  Pool / loan-code roll-up happens
    in a later slice.

    Result shape::

        {
          "ok": bool,
          "error": str | None,
          "folder": str,
          "target_period": str,
          "history_months": int,
          "expected_months": [str, ...],            # newest first
          "files": [
            {"name", "path", "detected_date",
             "detected_source", "detected_confidence",
             "period": "YYYY-MM" | "",
             "signature", "profile_id" | None,
             "profile_label" | None,
             "in_window": bool,
             "error": str | None},
            ...
          ],
          "months": {
            "YYYY-MM": {
              "files": [file name, ...],
              "status": "ok" | "missing" | "multiple",
            },
            ...
          },
          "new_signatures": [
            {"signature", "headers", "sample_file"},
            ...
          ],
          "summary": {
            "scanned": int, "in_window": int, "out_of_window": int,
            "covered_months": int, "missing_months": int,
            "multiple_months": int, "unreadable": int,
            "new_signatures": int,
          },
        }
    """
    out: dict[str, Any] = {
        "ok": False,
        "error": None,
        "folder": str(folder),
        "target_period": target_period,
        "history_months": history_months,
        "expected_months": [],
        "files": [],
        "months": {},
        "new_signatures": [],
        "summary": {
            "scanned": 0, "in_window": 0, "out_of_window": 0,
            "covered_months": 0, "missing_months": 0,
            "multiple_months": 0, "unreadable": 0,
            "new_signatures": 0,
        },
    }

    if not folder or not Path(folder).exists() or not Path(folder).is_dir():
        out["error"] = f"Folder not found: {folder}"
        return out

    expected = _months_back(target_period, history_months or 1)
    out["expected_months"] = expected
    window_set = set(expected)

    sig_to_profile: dict[str, dict[str, Any]] = {
        (p.get("signature") or ""): p for p in (known_profiles or [])
    }

    # Stable file order so the UI is deterministic.  ``rglob`` walks the
    # folder recursively so monthly files tucked into per-year (or other)
    # subfolders are picked up too.
    root = Path(folder)
    candidates = sorted(
        (p for p in root.rglob("*")
         if p.is_file()
         and p.suffix.lower() in _EXTRACT_EXTS
         and not p.name.startswith("~$")),
        key=lambda p: str(p).lower(),
    )

    new_sigs_seen: dict[str, dict[str, Any]] = {}

    for path in candidates:
        try:
            rel_name = str(path.relative_to(root))
        except ValueError:
            rel_name = path.name
        entry: dict[str, Any] = {
            "name": rel_name,
            "path": str(path),
            "detected_date": "",
            "detected_source": "",
            "detected_confidence": "none",
            "period": "",
            "signature": "",
            "profile_id": None,
            "profile_label": None,
            "in_window": False,
            "error": None,
        }
        date_info = detect_as_of_date(path.name, path=path)
        entry["detected_date"] = date_info.get("date", "")
        entry["detected_source"] = date_info.get("source", "")
        entry["detected_confidence"] = date_info.get("confidence", "none")
        entry["period"] = _period_key(entry["detected_date"])

        head_info = read_extract_headers(path)
        headers = head_info.get("headers") or []
        if not head_info.get("ok") or not headers:
            entry["error"] = head_info.get("error") or "Could not read headers."
            out["summary"]["unreadable"] += 1
        else:
            sig = compute_header_signature(headers)
            entry["signature"] = sig
            prof = sig_to_profile.get(sig)
            if prof:
                entry["profile_id"] = prof.get("id")
                entry["profile_label"] = prof.get("label")
            elif sig not in new_sigs_seen:
                new_sigs_seen[sig] = {
                    "signature": sig,
                    "headers": list(headers),
                    "sample_file": path.name,
                }

        if entry["period"] and entry["period"] in window_set:
            entry["in_window"] = True
            out["summary"]["in_window"] += 1
        else:
            out["summary"]["out_of_window"] += 1

        out["files"].append(entry)
        out["summary"]["scanned"] += 1

    # Per-month bucket (only for the expected window).
    for ym in expected:
        files_in = [f["name"] for f in out["files"] if f["period"] == ym]
        if not files_in:
            status = "missing"
            out["summary"]["missing_months"] += 1
        elif len(files_in) > 1:
            status = "multiple"
            out["summary"]["multiple_months"] += 1
            out["summary"]["covered_months"] += 1
        else:
            status = "ok"
            out["summary"]["covered_months"] += 1
        out["months"][ym] = {"files": files_in, "status": status}

    out["new_signatures"] = list(new_sigs_seen.values())
    out["summary"]["new_signatures"] = len(new_sigs_seen)
    out["ok"] = True
    return out
