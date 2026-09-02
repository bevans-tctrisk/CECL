"""Aggregate per-month Charge-Off / Recovery files into the historical DB.

The wizard's Historical step lets a user upload many monthly CO / Recovery
spreadsheets at once (one file per month).  Each file typically has a header
row + a few transaction rows like::

    Account #   Loan Type Code   Recovery Amount   Charge off Amount   Date
    28883L102   43               197.15            (blank)             ...
    49198L102   42                                 100                 ...

Some credit unions ship combined CO+Recovery files (both amount columns in
one workbook), others ship separate files.  This module reads each file,
sums per ``loan_code`` for the requested kind (``"co"`` or ``"recov"``), and
writes the result via ``chargeoff_hist_processor`` /
``recovery_hist_processor`` ``upsert_month``.

Files dated to the same ``as_of_date`` are merged before upsert so the
processor's "replace prior rows" behaviour doesn't drop earlier files of
the same month.
"""
from __future__ import annotations

import calendar
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from cecl_ui.services import (
    chargeoff_hist_processor,
    extract_hist_service,
    recovery_hist_processor,
)


# Header-name keyword sets used to auto-locate columns when the file has a
# header row.  Lowered + stripped before matching.
_LOAN_CODE_HEADERS = (
    "loan type code", "loan_type_code", "loan code", "loan_code",
    "type code", "type_code", "pool code", "pool_code",
    "security code", "security_code",
    "loan type", "loan_type", "loan class", "loan_class",
    "product code", "product_code", "product type", "product_type",
    "collateral code", "collateral_code", "collateral type",
    "collateral_type",
)
_MEMBER_HEADERS = (
    "member number", "member_number", "member#", "member #", "member id",
    "member", "acct member", "member no",
)
_ACCOUNT_HEADERS = (
    "account number", "account_number", "account#", "account #",
    "loan number", "loan_number", "loan#", "loan #",
    "suffix", "account suffix", "loan suffix", "share suffix",
    "id",  # last-resort fallback
)
_CO_AMOUNT_HEADERS = (
    "charge off amount", "charge-off amount", "chargeoff amount",
    "charge off", "chargeoff", "chg off amount", "chg-off amount",
    "co amount", "co_amount",
)
_RECOV_AMOUNT_HEADERS = (
    "recovery amount", "recovery_amount", "recoveries amount",
    "recovery", "recoveries", "recov amount", "recov_amount",
)
_DATE_HEADERS = (
    "charge off date", "chargeoff date", "charge-off date",
    "recovery date", "recoveries date",
    "effective date", "transaction date", "posting date", "date",
)


def file_layout_signature(path: Path) -> str:
    """Return the header signature for a CO/Recov file, or '' on failure.

    Used by the per-layout column-mapping plumbing so the aggregator can
    look up which mapping applies to each file.  Mirrors
    ``setup._file_layout_signature`` but lives here to avoid importing
    from a Flask routes module.
    """
    try:
        head = extract_hist_service.read_extract_headers(path)
        if not head.get("ok"):
            return ""
        headers = head.get("headers") or []
        return extract_hist_service.compute_header_signature(headers)
    except Exception:  # noqa: BLE001
        return ""


def _resolve_overrides_for_file(
    state: dict[str, Any], kind: str, path: Path
) -> tuple[str | None, str | None, str | None, dict[str, str], str | None]:
    """Return ``(code_header, amount_header, date_header, code_map, sheet)``
    for one file, preferring the per-layout mapping (keyed by the file's
    header signature) over the shared default.  ``code_map`` translates
    raw loan codes into pool names; empty when the user hasn't set one.
    ``sheet`` is the worksheet/tab the user chose (None = auto-detect).
    """
    cols_key = "co_columns" if kind == "co" else "recov_columns"
    layout_cols_key = (
        "co_layout_columns" if kind == "co" else "recov_layout_columns"
    )
    lcm_key = (
        "co_loan_code_map" if kind == "co" else "recov_loan_code_map"
    )
    shared = state.get(cols_key) or {}
    sig = file_layout_signature(path)
    layout_map = (state.get(layout_cols_key) or {}).get(sig) if sig else None
    overrides = layout_map if layout_map else shared
    code_header = (overrides.get("loan_code") or "").strip() or None
    amount_header = (overrides.get("amount") or "").strip() or None
    date_header = (overrides.get("date") or "").strip() or None
    # A per-layout mapping can pin its own tab; fall back to the shared
    # ``{kind}_columns`` sheet so single-workbook uploads still route.
    sheet = (
        (overrides.get("sheet") or "").strip()
        or (shared.get("sheet") or "").strip()
        or None
    )
    # The Loan Code Mapping step (wizard step 3) writes a single
    # ``state["pool_map"]`` dict shared by every consumer
    # (generate_report, step4_pools, etc.). Per-layout
    # ``co_loan_code_map[sig]`` / ``recov_loan_code_map[sig]``
    # overrides from older drafts win when set.
    global_map = state.get("pool_map") or {}
    layout_lcm = (
        (state.get(lcm_key) or {}).get(sig) if sig else None
    ) or {}
    code_map: dict[str, str] = {}
    for k, v in global_map.items():
        ks = str(k).strip()
        vs = str(v).strip()
        if ks and vs:
            code_map[ks] = vs
    for k, v in layout_lcm.items():
        ks = str(k).strip()
        vs = str(v).strip()
        if ks and vs:
            code_map[ks] = vs
    return code_header, amount_header, date_header, code_map, sheet


def parse_loan_code_map_file(path: Path) -> dict[str, Any]:
    """Parse a 2-column user-supplied loan-code mapping file (CSV/XLSX).

    Expected layout: a header row whose first column is something like
    "Loan Code" / "Code" / "Loan Type" and second column is something
    like "Pool" / "Pool Name" / "Pool Code".  If no header row is
    detected the first non-empty row is treated as data.

    Returns ``{"ok": bool, "rows": [{code, pool}, ...], "error": str|None}``.
    Codes/pools are stripped strings.  Empty rows are skipped.
    """
    try:
        raw = _read_rows(path)
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "rows": [], "error": str(exc)}
    if not raw:
        return {"ok": False, "rows": [], "error": "File is empty."}
    code_kw = (
        "loan code", "loan_code", "code", "loan type code",
        "loan_type_code", "loan type", "loan_type", "type code",
        "type", "raw code", "raw_code",
    )
    pool_kw = (
        "pool", "pool name", "pool_name", "target pool",
        "target", "mapped pool", "mapped_pool", "category",
    )
    code_col = 0
    pool_col = 1
    start_row = 0
    for idx in range(min(5, len(raw))):
        cells = [
            str(c).strip().lower() if c is not None else ""
            for c in raw[idx]
        ]
        if not any(cells):
            continue
        c_idx = next(
            (i for i, c in enumerate(cells) if c in code_kw or any(k in c for k in code_kw)),
            None,
        )
        p_idx = next(
            (i for i, c in enumerate(cells) if c in pool_kw or any(k in c for k in pool_kw)),
            None,
        )
        if c_idx is not None and p_idx is not None and c_idx != p_idx:
            code_col = c_idx
            pool_col = p_idx
            start_row = idx + 1
            break
    rows: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in raw[start_row:]:
        if not row:
            continue
        code_v = (
            _loan_code_str(row[code_col])
            if code_col < len(row) else None
        )
        pool_v = (
            str(row[pool_col]).strip()
            if pool_col < len(row) and row[pool_col] not in (None, "")
            else ""
        )
        if not code_v or not pool_v:
            continue
        if code_v in seen:
            continue
        seen.add(code_v)
        rows.append({"code": code_v, "pool": pool_v})
    if not rows:
        return {
            "ok": False,
            "rows": [],
            "error": (
                "No loan-code/pool pairs found. The file should have two "
                "columns: loan code, pool name."
            ),
        }
    return {"ok": True, "rows": rows, "error": None}


def extract_distinct_loan_codes(
    path: Path, code_header: str | None, *, limit: int = 200,
    sheet: str | None = None,
) -> list[str]:
    """Return up to ``limit`` distinct loan codes found in ``path`` for the
    Loan Code column ``code_header`` (header-name match, case- and
    whitespace-insensitive).  Falls back to the auto-detected code column
    when ``code_header`` is empty.  ``sheet`` pins which tab to read.
    Empty list on parse failure.
    """
    try:
        raw = _read_rows(path, sheet=sheet)
    except Exception:  # noqa: BLE001
        return []
    if not raw:
        return []
    hdr_idx = _find_header_row(raw)
    if hdr_idx is None:
        return []
    headers = [str(c) if c is not None else "" for c in raw[hdr_idx]]
    if code_header:
        col = _pick_column_by_name(headers, code_header)
    else:
        col = _pick_column(headers, _LOAN_CODE_HEADERS)
    if col is None:
        return []
    seen: dict[str, None] = {}
    for row in raw[hdr_idx + 1:]:
        if col >= len(row):
            continue
        code = _loan_code_str(row[col])
        if code is None:
            continue
        if code not in seen:
            seen[code] = None
            if len(seen) >= limit:
                break
    return sorted(seen.keys())


def list_sheet_names(path: Path) -> list[str]:
    """Return the worksheet/tab names of an Excel workbook.

    Empty list for CSV files or when the workbook can't be opened. Used by
    the wizard's CO/Recovery step to let the user pick which tab to read
    when a single workbook holds charge-offs and recoveries on separate
    tabs (auto-detection reads only the densest tab).
    """
    ext = path.suffix.lower()
    if ext == ".csv":
        return []
    try:
        if ext == ".xls":
            import xlrd  # type: ignore

            wb = xlrd.open_workbook(str(path), on_demand=True)
            return [str(n) for n in wb.sheet_names()]
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        return [str(n) for n in wb.sheetnames]
    except Exception:  # noqa: BLE001
        return []


def _read_rows(path: Path, sheet: str | None = None) -> list[list[Any]]:
    """Return all sheet rows as a list of lists. CSV or XLSX/XLSM/XLS.

    When ``sheet`` names a worksheet in the workbook (case-insensitive), only
    that tab is read — this is how a CU whose charge-offs and recoveries live
    on separate tabs of one workbook tells us which tab feeds each side.
    When ``sheet`` is empty or not found, falls back to picking the worksheet
    with the most non-blank cells (prevents an empty leading sheet from
    masking the real data tab).
    """
    ext = path.suffix.lower()
    if ext == ".csv":
        import csv

        out: list[list[Any]] = []
        with open(path, newline="", encoding="utf-8-sig") as fh:
            for row in csv.reader(fh):
                out.append(list(row))
        return out

    want = (sheet or "").strip().lower()

    if ext == ".xls":
        # Legacy BIFF format — openpyxl can't read these. Use xlrd.
        import xlrd  # type: ignore

        wb_xls = xlrd.open_workbook(str(path))
        if want:
            for sh in wb_xls.sheets():
                if str(sh.name).strip().lower() == want:
                    return [
                        list(sh.row_values(r)) if sh.row_values(r) else []
                        for r in range(sh.nrows)
                    ]
        best_rows_xls: list[list[Any]] = []
        best_score_xls = -1
        for sh in wb_xls.sheets():
            sheet_rows: list[list[Any]] = []
            for r in range(sh.nrows):
                row_vals = sh.row_values(r)
                sheet_rows.append(list(row_vals) if row_vals else [])
            score = sum(
                1 for row in sheet_rows
                for c in row if c not in (None, "")
            )
            if score > best_score_xls:
                best_score_xls = score
                best_rows_xls = sheet_rows
        return best_rows_xls

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if want:
        for ws in wb.worksheets:
            if str(ws.title).strip().lower() == want:
                return [
                    list(r) if r else []
                    for r in ws.iter_rows(values_only=True)
                ]
    best_rows: list[list[Any]] = []
    best_score = -1
    for ws in wb.worksheets:
        sheet_rows: list[list[Any]] = []
        for r in ws.iter_rows(values_only=True):
            sheet_rows.append(list(r) if r else [])
        score = sum(
            1 for row in sheet_rows
            for c in row if c not in (None, "")
        )
        if score > best_score:
            best_score = score
            best_rows = sheet_rows
    return best_rows


def _norm(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip()).lower()


def _find_header_row(rows: list[list[Any]]) -> int | None:
    """Find the first row whose cells look like headers (contain a loan-code
    column AND at least one amount column)."""
    for i, row in enumerate(rows[:25]):  # don't scan forever
        cells = [_norm(c) for c in row]
        has_code = any(
            any(k in c for k in _LOAN_CODE_HEADERS) for c in cells
        )
        has_amount = any(
            any(k in c for k in (*_CO_AMOUNT_HEADERS, *_RECOV_AMOUNT_HEADERS))
            for c in cells
        )
        if has_code and has_amount:
            return i
    return None


def _pick_column(
    headers: list[str], keywords: tuple[str, ...]
) -> int | None:
    """Return the index of the first header that exactly equals or contains
    one of ``keywords``.  Prefers exact equality over substring match."""
    norm = [_norm(h) for h in headers]
    for kw in keywords:
        if kw in norm:
            return norm.index(kw)
    for kw in keywords:
        for i, h in enumerate(norm):
            if kw in h:
                return i
    return None


def _to_float(v: Any) -> float | None:
    """Coerce a cell to a float.  Returns None for blank/non-numeric."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        if f != f:  # NaN
            return None
        return f
    s = str(v).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace("$", "").replace(",", "").strip()
    if not s:
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return -f if neg else f


def _loan_code_str(v: Any) -> str | None:
    """Normalise a loan-code cell to a short uppercase string.

    Strips trailing ``"/anything"`` suffixes (e.g. ``"VA/01"`` -> ``"VA"``)
    and converts numeric codes (43, 42) to ``"43"`` / ``"42"``.
    """
    if v is None or v == "":
        return None
    if isinstance(v, float) and v != v:  # NaN
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    if not s:
        return None
    if "/" in s:
        s = s.split("/", 1)[0].strip()
    return s.upper()


def _pick_column_by_name(headers: list[str], name: str) -> int | None:
    if not name:
        return None
    norm = [_norm(h) for h in headers]
    target = _norm(name)
    if target in norm:
        return norm.index(target)
    for i, h in enumerate(norm):
        if target and target in h:
            return i
    return None


def inspect_file(path: Path, sheet: str | None = None) -> dict[str, Any]:
    """Peek at a CO/Recov file and return its header row + suggested column
    matches for the column-mapping UI.  ``sheet`` pins which tab to read
    (None = densest tab).

    Returns::

        {
          "ok": bool, "error": str | None,
          "filename": str,
          "headers": [str, ...],
          "suggested": {"code": str, "co_amount": str,
                        "recov_amount": str, "date": str},
          "preview_rows": [[...], ...],  # up to 5 rows after the header
        }
    """
    out: dict[str, Any] = {
        "ok": False, "error": None,
        "filename": path.name,
        "headers": [], "suggested": {}, "preview_rows": [],
    }
    try:
        raw = _read_rows(path, sheet=sheet)
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Could not read file: {exc}"
        return out
    if not raw:
        out["error"] = "File is empty."
        return out

    hdr_idx = _find_header_row(raw)
    if hdr_idx is None:
        # Fallback: the keyword heuristic missed (file uses non-standard
        # header text). Pick the row with the most non-blank string
        # cells inside the first 25 rows so the manual dropdowns at
        # least get useful options.
        best_i, best_score = -1, 0
        for i, row in enumerate(raw[:25]):
            score = sum(
                1 for c in row
                if isinstance(c, str) and c.strip()
            )
            if score > best_score:
                best_score, best_i = score, i
        if best_score >= 2:
            hdr_idx = best_i
            out["error"] = (
                "Couldn't auto-detect the header row from the usual "
                "keywords \u2014 used the row with the most text cells "
                f"(row {best_i + 1}) as a best guess. Confirm and pick "
                "the columns manually below."
            )
        else:
            out["error"] = (
                "Could not find a header row. Pick the columns "
                "manually below."
            )
            # Still expose the first row as a best-effort header list.
            # Normalise embedded whitespace (newlines / tabs / runs of
            # spaces) so the option ``value`` posted by the browser
            # round-trips cleanly back into the saved column mapping.
            out["headers"] = [
                re.sub(r"\s+", " ", str(c)).strip() if c is not None else ""
                for c in (raw[0] if raw else [])
            ]
            out["preview_rows"] = [list(r) for r in raw[1:6]]
            return out

    # Normalise embedded whitespace in header cells (Excel often wraps
    # headers across lines, leaving \n / \t inside the value). Browsers
    # collapse whitespace inside ``<option value="...">`` attributes
    # when submitting forms, so without this the saved value ("Chargeoff
    # Amount") would never match the option text ("Chargeoff\nAmount")
    # on subsequent renders and the dropdown would reset to "choose".
    headers = [
        re.sub(r"\s+", " ", str(c)).strip() if c is not None else ""
        for c in raw[hdr_idx]
    ]

    def _name_at(idx: int | None) -> str:
        if idx is None or idx >= len(headers):
            return ""
        return headers[idx]

    out["headers"] = headers
    out["preview_rows"] = [
        list(r) for r in raw[hdr_idx + 1: hdr_idx + 6]
    ]
    out["suggested"] = {
        "code": _name_at(_pick_column(headers, _LOAN_CODE_HEADERS)),
        "co_amount": _name_at(_pick_column(headers, _CO_AMOUNT_HEADERS)),
        "recov_amount": _name_at(
            _pick_column(headers, _RECOV_AMOUNT_HEADERS)
        ),
        "date": _name_at(_pick_column(headers, _DATE_HEADERS)),
        "member": _name_at(_pick_column(headers, _MEMBER_HEADERS)),
        "account": _name_at(_pick_column(headers, _ACCOUNT_HEADERS)),
    }
    out["ok"] = True
    return out


def parse_file(
    path: Path,
    kind: str,
    *,
    code_header: str | None = None,
    amount_header: str | None = None,
    date_header: str | None = None,
    sheet: str | None = None,
) -> dict[str, Any]:
    """Parse one monthly CO/Recov file.

    ``kind`` is ``"co"`` or ``"recov"``.  Optional ``code_header`` /
    ``amount_header`` force a specific header-name match (case/whitespace
    insensitive); when not supplied we fall back to keyword auto-detection.
    ``sheet`` pins which worksheet/tab to read (None = densest tab).

    Returns::

        {
          "ok": bool, "error": str | None,
          "filename": str,
          "as_of_date": "YYYY-MM-DD" | "",
          "date_source": "filename" | "mtime" | "header" | "",
          "date_confidence": "high"|"medium"|"low"|"none",
          "rows": [{"loan_code": str, "amount": float}, ...],  # aggregated
          "total_rows": int,        # raw rows used
          "total_amount": float,    # sum across rows
          "empty_ok": bool,         # True when file parsed cleanly but had
                                    # zero non-blank amount rows
        }
    """
    out: dict[str, Any] = {
        "ok": False, "error": None,
        "filename": path.name,
        "as_of_date": "", "date_source": "", "date_confidence": "none",
        "rows": [], "total_rows": 0, "total_amount": 0.0,
        "empty_ok": False,
    }
    try:
        raw = _read_rows(path, sheet=sheet)
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Could not read file: {exc}"
        return out
    if not raw:
        out["error"] = "File is empty."
        return out

    hdr_idx = _find_header_row(raw)
    if hdr_idx is None:
        # Fallback to densest-text-row heuristic so files with
        # non-standard header text still parse when the user has
        # supplied an explicit column mapping.
        best_i, best_score = -1, 0
        for i, row in enumerate(raw[:25]):
            score = sum(
                1 for c in row
                if isinstance(c, str) and c.strip()
            )
            if score > best_score:
                best_score, best_i = score, i
        if best_score >= 2 and (code_header or amount_header):
            hdr_idx = best_i
        else:
            out["error"] = (
                "Could not find a header row with a Loan Code column and an "
                "amount column."
            )
            return out

    headers = [str(c) if c is not None else "" for c in raw[hdr_idx]]
    if code_header:
        code_col = _pick_column_by_name(headers, code_header)
    else:
        code_col = _pick_column(headers, _LOAN_CODE_HEADERS)
    if amount_header:
        amount_col = _pick_column_by_name(headers, amount_header)
    else:
        amount_keywords = (
            _CO_AMOUNT_HEADERS if kind == "co" else _RECOV_AMOUNT_HEADERS
        )
        amount_col = _pick_column(headers, amount_keywords)
    if date_header:
        date_col = _pick_column_by_name(headers, date_header)
    else:
        date_col = _pick_column(headers, _DATE_HEADERS)

    if code_col is None:
        if code_header:
            out["error"] = (
                f"Configured Loan Code column '{code_header}' was not found "
                f"in the header row."
            )
        else:
            out["error"] = (
                "Could not find a Loan Code column in the header row."
            )
        return out
    if amount_col is None:
        label = "Charge-Off" if kind == "co" else "Recovery"
        if amount_header:
            out["error"] = (
                f"Configured {label} Amount column '{amount_header}' was not "
                f"found in the header row."
            )
        else:
            out["error"] = (
                f"Could not find a {label} Amount column in the header row."
            )
        return out

    # Aggregate by loan code.
    agg: dict[str, float] = defaultdict(float)
    used = 0
    header_dates: list[Any] = []
    for row in raw[hdr_idx + 1:]:
        if not row:
            continue
        code = _loan_code_str(
            row[code_col] if code_col < len(row) else None
        )
        amt = _to_float(
            row[amount_col] if amount_col < len(row) else None
        )
        if code is None or amt is None or amt == 0:
            continue
        agg[code] += amt
        used += 1
        if date_col is not None and date_col < len(row):
            dv = row[date_col]
            if dv not in (None, ""):
                header_dates.append(dv)

    # Date detection.
    det = extract_hist_service.detect_as_of_date(path.name, path)
    out["as_of_date"] = det.get("date") or ""
    out["date_source"] = det.get("source") or ""
    out["date_confidence"] = det.get("confidence") or "none"

    # If filename detection was low/none, try the latest in-row date.
    if (out["date_confidence"] in ("low", "none")) and header_dates:
        latest = _latest_in_row_date(header_dates)
        if latest:
            out["as_of_date"] = latest
            out["date_source"] = "header"
            out["date_confidence"] = "medium"

    out["rows"] = [
        {"loan_code": k, "amount": round(v, 2)}
        for k, v in sorted(agg.items())
    ]
    out["total_rows"] = used
    out["total_amount"] = round(sum(agg.values()), 2)
    out["empty_ok"] = used == 0
    out["ok"] = True
    return out


def _latest_in_row_date(values: list[Any]) -> str | None:
    """Best-effort: scan a list of date-ish values and return the latest as
    YYYY-MM-DD."""
    from datetime import date, datetime

    latest: date | None = None
    for v in values:
        d: date | None = None
        if isinstance(v, datetime):
            d = v.date()
        elif isinstance(v, date):
            d = v
        else:
            s = str(v).strip()
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%Y/%m/%d"):
                try:
                    d = datetime.strptime(s, fmt).date()
                    break
                except ValueError:
                    continue
        if d is None:
            continue
        if latest is None or d > latest:
            latest = d
    return latest.isoformat() if latest else None


def aggregate_all(state: dict[str, Any], kind: str) -> dict[str, Any]:
    """Parse every monthly file in ``state.hist_scan`` for ``kind`` and
    upsert into the historical CO/Recov DB.

    ``kind`` is ``"co"`` (uses ``monthly_co_files``) or ``"recov"`` (uses
    ``monthly_recov_files``).
    """
    out: dict[str, Any] = {
        "ok": False, "error": None,
        "kind": kind,
        "files": [],          # per-file summary list
        "months_written": [], # distinct as_of_date values written
        "total_files": 0,
        "total_parsed": 0,
        "total_rows_written": 0,
        "total_amount": 0.0,
    }

    cu = (state.get("credit_union") or "").strip()
    if not cu:
        out["error"] = "Set the credit union on the Identity step first."
        return out

    hs = state.get("hist_scan") or {}
    key = "monthly_co_files" if kind == "co" else "monthly_recov_files"
    files = hs.get(key) or []
    out["total_files"] = len(files)
    if not files:
        out["error"] = "No monthly files have been uploaded yet."
        return out

    # First pass: parse each file. ``parsed_per_date`` is a normal dict so
    # an empty-but-validated month registers a key with an empty inner dict,
    # which downstream upserts as a zero-row baseline for that month.  Each
    # file is resolved against its own per-layout column mapping (keyed by
    # the file's header signature) with fallback to the shared default.
    parsed_per_date: dict[str, dict[str, float]] = {}
    per_date_sources: dict[str, list[str]] = defaultdict(list)
    for entry in files:
        name = entry.get("name") or ""
        path_str = entry.get("path") or ""
        item: dict[str, Any] = {
            "name": name, "ok": False, "error": None,
            "as_of_date": "", "rows": 0, "amount": 0.0,
            "empty": False,
        }
        if not path_str or not Path(path_str).exists():
            item["error"] = "Saved file is missing on disk."
            out["files"].append(item)
            continue
        code_header, amount_header, date_header, code_map, sheet = (
            _resolve_overrides_for_file(state, kind, Path(path_str))
        )
        res = parse_file(
            Path(path_str), kind,
            code_header=code_header, amount_header=amount_header,
            date_header=date_header, sheet=sheet,
        )
        item["as_of_date"] = res.get("as_of_date") or ""
        if not res.get("ok"):
            item["error"] = res.get("error") or "Parse failed."
            out["files"].append(item)
            continue
        if not item["as_of_date"]:
            item["error"] = (
                "Could not determine an as-of date for this file. Rename "
                "the file to include a date (e.g. MMDDYYYY)."
            )
            out["files"].append(item)
            continue
        item["ok"] = True
        item["rows"] = res["total_rows"]
        item["amount"] = res["total_amount"]
        item["empty"] = bool(res.get("empty_ok"))
        out["total_parsed"] += 1
        out["files"].append(item)
        per_date_sources[item["as_of_date"]].append(name)
        # Register the date even if no rows so an empty file still serves
        # as a baseline marker for the month.
        bucket = parsed_per_date.setdefault(item["as_of_date"], {})
        for r in res["rows"]:
            raw_code = r["loan_code"]
            tgt = code_map.get(raw_code, raw_code) if code_map else raw_code
            if not tgt or str(tgt).strip().casefold() in ("__ignore__", "ignore", "exclude"):
                continue
            bucket[tgt] = bucket.get(tgt, 0.0) + r["amount"]

    if not parsed_per_date:
        out["error"] = "Nothing to aggregate \u2014 see per-file errors."
        return out

    # Second pass: upsert each (cu, as_of_date) bundle.
    field = "chargeoff_amount" if kind == "co" else "recovery_amount"
    processor = (
        chargeoff_hist_processor if kind == "co"
        else recovery_hist_processor
    )
    write_errors: list[str] = []
    try:
        processor.ensure_table()
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Could not prepare DB table: {exc}"
        return out

    for as_of, code_to_amt in sorted(parsed_per_date.items()):
        rows = [
            {"loan_code": code, field: round(amt, 2)}
            for code, amt in sorted(code_to_amt.items())
        ]
        src = "monthly:" + ",".join(per_date_sources[as_of])
        try:
            n = processor.upsert_month(cu, as_of, rows, src)
        except Exception as exc:  # noqa: BLE001
            write_errors.append(f"{as_of}: {exc}")
            continue
        out["months_written"].append(as_of)
        out["total_rows_written"] += n
        out["total_amount"] += sum(code_to_amt.values())

    if write_errors and not out["months_written"]:
        out["error"] = "Database write failed: " + "; ".join(write_errors)
        return out
    if write_errors:
        out["error"] = "Some months failed: " + "; ".join(write_errors)
    out["ok"] = bool(out["months_written"])
    out["total_amount"] = round(out["total_amount"], 2)
    return out


# ---------------------------------------------------------------------------
# Workbook-mode aggregation (single_workbook source)
# ---------------------------------------------------------------------------
#
# Some CUs ship a single (or a couple of) transaction-level workbook(s) that
# contain MANY months of charge-off / recovery rows side by side, with each
# row carrying its own date.  Unlike monthly files (one as-of date per file),
# these need PER-ROW date binning: each row's date is snapped to its
# month-end, then per-(month_end, loan_code) sums are upserted.
#
# This is the back-end for the "single_workbook" CO/Recov source on the
# wizard's Historical Charge-Offs / Recoveries step.  The same files often
# carry both a Charge-Off Amount and a Recovery Amount column, so the user
# typically uploads them once per kind.

def _to_month_end(value: Any) -> str | None:
    """Snap a date-ish cell to ``YYYY-MM-DD`` month-end.  Returns None for
    blanks/unparseable values."""
    d: date | None = None
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        d = value.date()
    elif isinstance(value, date):
        d = value
    else:
        s = str(value).strip()
        if not s:
            return None
        # Excel may serialise pure-date cells as floats (serial day count).
        # Try a few common text formats.
        for fmt in (
            "%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m-%d-%Y",
            "%m/%d/%y", "%m-%d-%y", "%d/%m/%Y",
        ):
            try:
                d = datetime.strptime(s, fmt).date()
                break
            except ValueError:
                continue
        if d is None:
            try:
                d = datetime.fromisoformat(s).date()
            except ValueError:
                return None
    last = calendar.monthrange(d.year, d.month)[1]
    return date(d.year, d.month, last).isoformat()


def aggregate_workbook(state: dict[str, Any], kind: str) -> dict[str, Any]:
    """Parse every file in ``state.hist_scan.{co_files,recov_files}`` (the
    ``single_workbook`` upload bucket), bin each row by month-end of its
    date column, sum the requested amount column per ``(month_end,
    loan_code)``, and upsert into the historical CO / Recov DB.

    ``kind`` is ``"co"`` or ``"recov"``.  Designed for transaction-level
    workbooks that span many months — one row per loan/charge-off/recovery
    event.  Mirrors :func:`aggregate_all` shape so the wizard renders the
    same status panel.
    """
    out: dict[str, Any] = {
        "ok": False, "error": None,
        "kind": kind,
        "files": [],
        "months_written": [],
        "total_files": 0,
        "total_parsed": 0,
        "total_rows_written": 0,
        "total_amount": 0.0,
    }

    cu = (state.get("credit_union") or "").strip()
    if not cu:
        out["error"] = "Set the credit union on the Identity step first."
        return out

    hs = state.get("hist_scan") or {}
    key = "co_files" if kind == "co" else "recov_files"
    files = hs.get(key) or []
    out["total_files"] = len(files)
    if not files:
        out["error"] = "No workbook has been uploaded yet."
        return out

    col_key = "co_columns" if kind == "co" else "recov_columns"
    # Shared overrides serve as the fallback when a file's signature has
    # no per-layout mapping in ``state[f"{kind}_layout_columns"]``.
    _ = state.get(col_key) or {}

    # Accumulate across all files so duplicate (month, code) rows from
    # different files merge correctly before the single per-month upsert.
    by_month: dict[str, dict[str, float]] = defaultdict(
        lambda: defaultdict(float)
    )
    per_month_sources: dict[str, set[str]] = defaultdict(set)

    for entry in files:
        name = entry.get("name") or ""
        path_str = entry.get("path") or ""
        item: dict[str, Any] = {
            "name": name, "ok": False, "error": None,
            "as_of_date": "",  # filled with min..max range for display
            "rows": 0, "amount": 0.0, "empty": False,
            "months": 0,
        }
        if not path_str or not Path(path_str).exists():
            item["error"] = "Saved file is missing on disk."
            out["files"].append(item)
            continue
        code_header, amount_header, date_header, code_map, sheet = (
            _resolve_overrides_for_file(state, kind, Path(path_str))
        )
        try:
            raw = _read_rows(Path(path_str), sheet=sheet)
        except Exception as exc:  # noqa: BLE001
            item["error"] = f"Could not read file: {exc}"
            out["files"].append(item)
            continue
        if not raw:
            item["error"] = "File is empty."
            out["files"].append(item)
            continue

        hdr_idx = _find_header_row(raw)
        if hdr_idx is None:
            best_i, best_score = -1, 0
            for i, row in enumerate(raw[:25]):
                score = sum(
                    1 for c in row
                    if isinstance(c, str) and c.strip()
                )
                if score > best_score:
                    best_score, best_i = score, i
            if best_score >= 2 and (code_header or amount_header):
                hdr_idx = best_i
            else:
                item["error"] = (
                    "Could not find a header row with a Loan Code column "
                    "and an amount column."
                )
                out["files"].append(item)
                continue

        headers = [str(c) if c is not None else "" for c in raw[hdr_idx]]
        if code_header:
            code_col = _pick_column_by_name(headers, code_header)
        else:
            code_col = _pick_column(headers, _LOAN_CODE_HEADERS)
        if amount_header:
            amount_col = _pick_column_by_name(headers, amount_header)
        else:
            amount_keywords = (
                _CO_AMOUNT_HEADERS if kind == "co" else _RECOV_AMOUNT_HEADERS
            )
            amount_col = _pick_column(headers, amount_keywords)
        if date_header:
            date_col = _pick_column_by_name(headers, date_header)
        else:
            # Prefer the kind-specific date header (e.g. "Charge Off Date")
            # before generic "date".
            kind_date_kw = (
                ("charge off date", "chargeoff date", "charge-off date")
                if kind == "co"
                else ("recovery date", "recoveries date")
            )
            date_col = _pick_column(headers, kind_date_kw)
            if date_col is None:
                date_col = _pick_column(headers, _DATE_HEADERS)

        if code_col is None:
            item["error"] = (
                "Could not find a Loan Code column"
                + (f" matching '{code_header}'" if code_header else "")
                + " in the header row."
            )
            out["files"].append(item)
            continue
        if amount_col is None:
            label = "Charge-Off" if kind == "co" else "Recovery"
            item["error"] = (
                f"Could not find a {label} Amount column"
                + (f" matching '{amount_header}'" if amount_header else "")
                + " in the header row."
            )
            out["files"].append(item)
            continue
        if date_col is None:
            item["error"] = (
                "Could not find a Date column"
                + (f" matching '{date_header}'" if date_header else "")
                + " in the header row. Workbook mode bins each row to its "
                "own month-end, so a date column is required."
            )
            out["files"].append(item)
            continue

        # Filename-derived fallback month for rows whose date cell is
        # blank (e.g. recovery rows in a "Dec_2025_..." file that only
        # has a Charge-Off Date populated for new charge-offs).
        det = extract_hist_service.detect_as_of_date(Path(path_str).name, Path(path_str))
        fallback_month = (det.get("date") or "") if det else ""
        if fallback_month:
            try:
                fy, fm, fd = (int(x) for x in fallback_month.split("-"))
                last = calendar.monthrange(fy, fm)[1]
                fallback_month = date(fy, fm, last).isoformat()
            except Exception:  # noqa: BLE001
                fallback_month = ""

        used = 0
        amt_sum = 0.0
        file_months: set[str] = set()
        for row in raw[hdr_idx + 1:]:
            if not row:
                continue
            code = _loan_code_str(
                row[code_col] if code_col < len(row) else None
            )
            amt = _to_float(
                row[amount_col] if amount_col < len(row) else None
            )
            if code is None or amt is None or amt == 0:
                continue
            dv = row[date_col] if date_col < len(row) else None
            month_end = _to_month_end(dv)
            if month_end is None:
                month_end = fallback_month or None
            if month_end is None:
                continue
            tgt_code = code_map.get(code, code) if code_map else code
            if not tgt_code or str(tgt_code).strip().casefold() in ("__ignore__", "ignore", "exclude"):
                continue
            by_month[month_end][tgt_code] += amt
            per_month_sources[month_end].add(name)
            used += 1
            amt_sum += amt
            file_months.add(month_end)

        item["ok"] = True
        item["rows"] = used
        item["amount"] = round(amt_sum, 2)
        item["months"] = len(file_months)
        item["empty"] = used == 0
        if file_months:
            ms = sorted(file_months)
            item["as_of_date"] = (
                ms[0] if len(ms) == 1 else f"{ms[0]} \u2192 {ms[-1]}"
            )
        out["total_parsed"] += 1
        out["files"].append(item)

    if not by_month:
        out["error"] = (
            "Nothing to aggregate \u2014 no rows had a non-zero amount and "
            "a parseable date.  See per-file errors below."
        )
        return out

    processor = (
        chargeoff_hist_processor if kind == "co"
        else recovery_hist_processor
    )
    field = "chargeoff_amount" if kind == "co" else "recovery_amount"
    try:
        processor.ensure_table()
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Could not prepare DB table: {exc}"
        return out

    write_errors: list[str] = []
    for as_of, code_to_amt in sorted(by_month.items()):
        rows = [
            {"loan_code": code, field: round(amt, 2)}
            for code, amt in sorted(code_to_amt.items())
        ]
        src = "workbook:" + ",".join(sorted(per_month_sources[as_of]))
        try:
            n = processor.upsert_month(cu, as_of, rows, src)
        except Exception as exc:  # noqa: BLE001
            write_errors.append(f"{as_of}: {exc}")
            continue
        out["months_written"].append(as_of)
        out["total_rows_written"] += n
        out["total_amount"] += sum(code_to_amt.values())

    if write_errors and not out["months_written"]:
        out["error"] = "Database write failed: " + "; ".join(write_errors)
        return out
    if write_errors:
        out["error"] = "Some months failed: " + "; ".join(write_errors)
    out["ok"] = bool(out["months_written"])
    out["total_amount"] = round(out["total_amount"], 2)
    return out

