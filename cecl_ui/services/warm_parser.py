"""Parse a Manual WARM workbook and extract historical CECL data.

Used by the new-CU wizard's optional "WARM Upload" step.  Given the most
recent quarterly ``CECL-Migration-WARM - <CU>.xlsx`` file, this returns a
summary of:

* the CU name and as-of date detected in the workbook
* the loan pool list (from ``BS Data`` / ``CO Hist Data``)
* the credit grade definitions (from ``Grade Ranges & Loan Codes``)
* historical date range and month count (from ``HIst Bal Data``)
* per-pool charge-off and recovery counts that have at least one non-zero
  monthly observation

The module never writes to session state directly -- the route does that.

Both sample WARM workbooks inspected (Tongass FCU and Siskiyou CU) share the
same template: same sheet names, same row layout for the data tabs.  The
parser tolerates missing optional sheets (e.g. Siskiyou is missing the
``Recoveries`` historical tab).

NOTE: pandas / openpyxl imports are deferred inside ``analyse_warm_file`` to
keep app startup fast on network drives.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any


# Sheets we look for.  Any not present is silently skipped.
SHEET_HIST_BAL = "HIst Bal Data"          # historical pool balances by grade
SHEET_CO_HIST = "CO Hist Data"            # charge-off history by pool
SHEET_BS_DATA = "BS Data"                 # balance sheet history by pool
SHEET_BS_CO_DQ = "BS CO DQ Data Enter"   # per-pool risk/ACL/mgmt-adj settings
SHEET_RECOVERIES = "Recoveries"           # historical recoveries (optional)
SHEET_GRADES = "Grade Ranges & Loan Codes"
SHEET_DISPLAY_HIST = "Display HIst Bal"   # has a "Year #" / "Month #" header

_FN_DATE_RX = re.compile(r"(20\d{2})[-_](\d{2})")  # filename like "2026-03 ..."

# Pool labels that aren't real loan pools (totals, balancing rows, summary lines,
# and impairment-rollup rows which belong on the Impaired Loans tab — not in the
# loan-pool list).
_POOL_BLOCKLIST_PREFIXES = (
    "total", "sub total", "subtotal", "allowance", "balancer",
    "pool order", "loan pools", "grand total", "net ", "recover",
    "hide", "exclude",
    "impaired", "consumer problem", "business problem",
    "credit grade deteriorated",
)

# Grade labels that aren't real credit grades (hidden, totals, helper rows).
_GRADE_BLOCKLIST_PREFIXES = (
    "hide", "total", "minimum", "max for", "maximum",
)


def analyse_warm_file(path: Path, original_filename: str) -> dict[str, Any]:
    """Parse ``path`` and return a summary dict.

    Returns dict with at minimum::

        {
          "ok": bool,
          "error": str | None,
          "filename": str,
          "cu_name": str | None,
          "as_of_date": str | None,         # YYYY-MM-DD
          "pools": list[str],
          "grades": list[{label, min_score, max_score, reserve_rate}],
          "history_start": str | None,      # YYYY-MM-DD
          "history_end":   str | None,
          "history_months": int,
          "co_pools_with_data": int,
          "recov_pools_with_data": int,
          "sheets_found": list[str],
          "sheets_missing": list[str],
        }
    """
    out: dict[str, Any] = {
        "ok": False, "error": None, "filename": original_filename,
        "cu_name": None, "as_of_date": None,
        "pools": [], "grades": [],
        "history_start": None, "history_end": None, "history_months": 0,
        "co_pools_with_data": 0, "recov_pools_with_data": 0,
        "sheets_found": [], "sheets_missing": [],
        # {loan_type_label (col A): pool_name (col B)} read from BS Data
        "bs_loan_type_map": {},
        # {raw loan code (col S): pool name (col T)} read from Grade
        # Ranges & Loan Codes — used to seed the wizard's pool_map.
        "loan_code_pool_map": {},
        # Per-pool settings from BS CO DQ Data Enter:
        # [{name, risk_rated (bool), acl_months (int), use_default_mgmt_adj (bool)}, ...]
        "pool_settings": [],
        # ACL Amount captured from the row labelled "Allowance for Credit Loss"
        # on BS CO DQ Data Enter (single dollar value, not a loan pool).
        "acl_balance": 0.0,
        # Identity & economic baseline pulled from "BS CO DQ Data Enter":
        #   M1 charter, M2 period, M3 cu_name, L7 state, M7 county,
        #   N7 unemp_rate (decimal), O7 foreclosures, P7 bankruptcies, Q7 population.
        "baseline_identity": {},
        # Per-pool monthly total balances pulled from BS Data, the block to
        # the right of the Pool Order / Loan Pools anchor. Format::
        #
        #   {
        #     "as_of_date": "2025-12-31",
        #     "current_quarter": ["2025-10-31", "2025-11-30", "2025-12-31"],
        #     "pools": [
        #       {"name": "Collateralized Loans",
        #        "history":         [{"date": "...", "balance": 12345.67}, ...],
        #        "current_quarter": [{"date": "...", "balance": ...}, ...]},
        #       ...
        #     ],
        #   }
        #
        # The 3 months in the establishing quarter are split off because
        # those will be re-derived from the user's quarterly upload and
        # should not be trusted as historical truth.
        "pool_monthly_balances": {},
    }

    try:
        # Deferred imports — both are slow on the network drive.
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"openpyxl import failed: {exc}"
        return out

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Could not open workbook: {exc}"
        return out

    try:
        names = set(wb.sheetnames)
        wanted = [SHEET_HIST_BAL, SHEET_CO_HIST, SHEET_BS_DATA,
                  SHEET_BS_CO_DQ,
                  SHEET_RECOVERIES, SHEET_GRADES, SHEET_DISPLAY_HIST]
        out["sheets_found"]   = [s for s in wanted if s in names]
        out["sheets_missing"] = [s for s in wanted if s not in names]

        # --- as-of date: first try filename, then Display HIst Bal headers
        as_of = _date_from_filename(original_filename)
        if not as_of and SHEET_DISPLAY_HIST in names:
            as_of = _date_from_display_hist(wb[SHEET_DISPLAY_HIST])
        out["as_of_date"] = as_of

        # --- CU name: top-left of CO Hist Data or HIst Bal Data row 2
        for sheet in (SHEET_CO_HIST, SHEET_HIST_BAL, SHEET_BS_DATA):
            if sheet in names:
                cu = _first_string_in_row(wb[sheet], row=1) \
                     or _first_string_in_row(wb[sheet], row=2)
                if cu and "FCU" in cu.upper() or (cu and "CU" in cu.upper()):
                    out["cu_name"] = cu.strip()
                    break

        # --- Pools: distinct, ordered names from BS Data col A (rows 2+)
        if SHEET_BS_DATA in names:
            out["pools"] = _pool_names_from_bs(wb[SHEET_BS_DATA])
            out["bs_loan_type_map"] = _bs_loan_type_to_pool_map(wb[SHEET_BS_DATA])
            out["balance_titles"] = _balance_titles_above_pool_order(
                wb[SHEET_BS_DATA]
            )
            out["pool_monthly_balances"] = _pool_monthly_balances_from_bs(
                wb[SHEET_BS_DATA], out.get("as_of_date")
            )
        elif SHEET_CO_HIST in names:
            out["pools"] = _pool_names_from_co(wb[SHEET_CO_HIST])

        # --- Per-pool settings (risk-rated / ACL months / default mgmt adj)
        if SHEET_BS_CO_DQ in names:
            ps_result = _pool_settings_from_bs_co_dq(wb[SHEET_BS_CO_DQ])
            out["pool_settings"] = ps_result.get("settings", [])
            out["acl_balance"]   = ps_result.get("acl_balance", 0.0)
            out["baseline_identity"] = _identity_from_bs_co_dq(
                wb[SHEET_BS_CO_DQ]
            )

        # --- Grade definitions + raw loan-code -> pool map (cols S/T)
        if SHEET_GRADES in names:
            out["grades"] = _grades_from_sheet(wb[SHEET_GRADES])
            out["loan_code_pool_map"] = _loan_code_to_pool_map_from_grades(
                wb[SHEET_GRADES]
            )

        # --- History range / month count from HIst Bal Data row 5
        if SHEET_HIST_BAL in names:
            start, end, months = _history_range(wb[SHEET_HIST_BAL])
            out["history_start"]  = start
            out["history_end"]    = end
            out["history_months"] = months

        # --- Pools with non-zero CO history
        if SHEET_CO_HIST in names:
            out["co_pools_with_data"] = _count_pools_with_data(
                wb[SHEET_CO_HIST], start_row=5, start_col=3
            )

        # --- Pools with non-zero recovery history (optional sheet)
        if SHEET_RECOVERIES in names:
            out["recov_pools_with_data"] = _count_pools_with_data(
                wb[SHEET_RECOVERIES], start_row=3, start_col=3
            )

        out["ok"] = True
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Parse error: {exc}"
    finally:
        wb.close()

    return out


# ---------- helpers ----------

def _date_from_filename(filename: str) -> str | None:
    m = _FN_DATE_RX.search(filename)
    if not m:
        return None
    year, month = m.group(1), m.group(2)
    # last day of month is good enough for reporting dates
    last_day = {"01":"31","02":"28","03":"31","04":"30","05":"31","06":"30",
                "07":"31","08":"31","09":"30","10":"31","11":"30","12":"31"}.get(month, "01")
    return f"{year}-{month}-{last_day}"


def _first_string_in_row(ws, row: int) -> str | None:
    for r in ws.iter_rows(min_row=row, max_row=row, values_only=True):
        for v in r:
            if isinstance(v, str) and v.strip():
                return v.strip()
    return None


def _date_from_display_hist(ws) -> str | None:
    """``Display HIst Bal`` rows 2/3/4 hold Month/Year/Day numbers, with the
    rightmost non-empty column being the current as-of period."""
    rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
    if len(rows) < 4:
        return None
    months = rows[1]; years = rows[2]; days = rows[3]
    last_year = last_month = last_day = None
    for i in range(1, min(len(months), len(years), len(days))):
        if years[i] and months[i]:
            try:
                last_year  = int(years[i])
                last_month = int(months[i])
                last_day   = int(days[i]) if days[i] else None
            except (TypeError, ValueError):
                continue
    if not (last_year and last_month):
        return None
    if not last_day:
        last_day = {1:31,2:28,3:31,4:30,5:31,6:30,7:31,8:31,9:30,10:31,
                    11:30,12:31}[last_month]
    return f"{last_year:04d}-{last_month:02d}-{last_day:02d}"


def _is_real_pool_label(name: str) -> bool:
    low = name.strip().lower()
    if not low:
        return False
    return not any(low.startswith(p) for p in _POOL_BLOCKLIST_PREFIXES)


def _pool_names_from_bs(ws) -> list[str]:
    """Read the canonical loan-pool list from ``BS Data``.

    The sheet has a labelled block::

        Pool Order | Loan Pools         | <date> | <date> | ...
                 1 | Collateralized Loans
                 2 | RV Loans
                 ...
                21 | Exclude
                   | Grand Total

    We scan the first ~80 rows / ~10 cols for a cell whose value is exactly
    ``Pool Order``, then read pool names from the column where ``Loan Pools``
    appears in the same row (typically the next non-empty column). Names
    starting with ``HIDE`` / ``Exclude`` / ``Grand Total`` are dropped, and
    pool collection stops at the first ``Grand Total`` or after a stretch
    of blanks.

    Falls back to the legacy "col A from row 2 onward" scan if the labelled
    block isn't found, so older WARM templates keep working.
    """
    anchor = _find_pool_order_anchor(ws)
    if anchor is not None:
        header_row, label_col, name_col = anchor
        pools: list[str] = []
        seen: set[str] = set()
        blanks = 0
        for r in range(header_row + 1, header_row + 200):
            v = ws.cell(row=r, column=name_col).value
            if v is None or (isinstance(v, str) and not v.strip()):
                blanks += 1
                if blanks >= 5:
                    break
                continue
            blanks = 0
            if not isinstance(v, str):
                continue
            name = v.strip()
            low = name.lower()
            if low.startswith("grand total"):
                break
            if not _is_real_pool_label(name):
                continue
            if low in seen:
                continue
            seen.add(low)
            pools.append(name)
            if len(pools) >= 60:
                break
        if pools:
            return pools

    # ── Legacy fallback: col A row 2+ ──
    pools: list[str] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[0]
        if isinstance(v, str):
            name = v.strip()
            if name and name.lower() not in seen and _is_real_pool_label(name):
                pools.append(name)
                seen.add(name.lower())
        if len(pools) >= 60:
            break
    return pools


def _find_pool_order_anchor(ws) -> tuple[int, int, int] | None:
    """Locate the ``Pool Order`` / ``Loan Pools`` header on ``BS Data``.

    Returns ``(header_row, label_col, name_col)`` where ``label_col`` is the
    column containing the literal string "Pool Order" and ``name_col`` is
    the column containing "Loan Pools" (where pool names are listed below).
    Returns ``None`` if the block can't be found.
    """
    for r in range(1, 80):
        for c in range(1, 12):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().lower() == "pool order":
                # Walk right looking for "Loan Pools"
                for c2 in range(c + 1, c + 8):
                    v2 = ws.cell(row=r, column=c2).value
                    if isinstance(v2, str) and v2.strip().lower() == "loan pools":
                        return r, c, c2
                # No explicit "Loan Pools" header — treat the next column as
                # the names column.
                return r, c, c + 1
    return None


def _pool_monthly_balances_from_bs(
    ws, as_of_date: str | None
) -> dict[str, Any]:
    """Read per-pool monthly total balances from ``BS Data``.

    The block lives below the ``Pool Order`` / ``Loan Pools`` anchor.  Layout::

        Row N:    | Pool Order | Loan Pools           | <date1> | <date2> | ...
        Row N+1:  |          1 | Collateralized Loans | 0       | 0       | ...
        Row N+2:  |          2 | RV Loans             | 0       | 0       | ...
        ...

    Returns a dict::

        {
          "as_of_date": "2025-12-31" | None,
          "dates": ["2016-01-31", ..., "2025-12-31"],
          "current_quarter": ["2025-10-31", "2025-11-30", "2025-12-31"],
          "pools": [
            {"name": "Collateralized Loans",
             "history":         [{"date": "...", "balance": 0.0}, ...],
             "current_quarter": [{"date": "...", "balance": 0.0}, ...]},
            ...
          ],
        }

    The ``current_quarter`` split is keyed off ``as_of_date`` (any month in
    the same calendar quarter as as_of, including as_of itself, is treated
    as "to be re-derived from the upcoming quarterly upload" and excluded
    from the history series).  HIDE / Exclude / Grand Total rows are
    dropped.  When the anchor can't be found, returns an empty dict.
    """
    anchor = _find_pool_order_anchor(ws)
    if anchor is None:
        return {}
    header_row, _label_col, name_col = anchor

    # Build the date column list by walking right across the header row.
    dates: list[tuple[int, str]] = []  # (col_index, iso_date)
    for c in range(name_col + 1, name_col + 400):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            # tolerate one empty header cell, but stop after several
            empty_run = 0
            for cc in range(c, c + 4):
                if ws.cell(row=header_row, column=cc).value is None:
                    empty_run += 1
            if empty_run >= 3:
                break
            continue
        iso = _coerce_iso_date(v)
        if iso:
            dates.append((c, iso))
    if not dates:
        return {}

    # Determine which dates fall inside the establishing quarter.
    cq_set: set[str] = set()
    if as_of_date:
        try:
            ay, am, _ad = as_of_date.split("-")
            ay_i, am_i = int(ay), int(am)
            q_start_month = ((am_i - 1) // 3) * 3 + 1
            cq_months = {q_start_month, q_start_month + 1, q_start_month + 2}
            cq_set = {
                iso for (_c, iso) in dates
                if int(iso.split("-")[0]) == ay_i
                and int(iso.split("-")[1]) in cq_months
                and iso <= as_of_date
            }
        except (ValueError, IndexError):
            cq_set = set()

    # Walk pool rows below the anchor, mirroring _pool_names_from_bs logic.
    pools_out: list[dict[str, Any]] = []
    seen: set[str] = set()
    blanks = 0
    for r in range(header_row + 1, header_row + 200):
        v = ws.cell(row=r, column=name_col).value
        if v is None or (isinstance(v, str) and not v.strip()):
            blanks += 1
            if blanks >= 5:
                break
            continue
        blanks = 0
        if not isinstance(v, str):
            continue
        name = v.strip()
        if name.lower().startswith("grand total"):
            break
        if not _is_real_pool_label(name):
            continue
        if name.lower() in seen:
            continue
        seen.add(name.lower())

        history: list[dict[str, Any]] = []
        cq: list[dict[str, Any]] = []
        for col_idx, iso in dates:
            cell = ws.cell(row=r, column=col_idx).value
            if isinstance(cell, (int, float)):
                bal = float(cell)
            else:
                bal = 0.0
            entry = {"date": iso, "balance": bal}
            if iso in cq_set:
                cq.append(entry)
            else:
                history.append(entry)
        pools_out.append({
            "name": name,
            "history": history,
            "current_quarter": cq,
        })
        if len(pools_out) >= 60:
            break

    return {
        "as_of_date": as_of_date,
        "dates": [iso for (_c, iso) in dates],
        "current_quarter": sorted(cq_set),
        "pools": pools_out,
    }


def _coerce_iso_date(v: Any) -> str | None:
    """Best-effort YYYY-MM-DD coercion of a header-row cell value."""
    if v is None:
        return None
    # openpyxl returns datetime for date-typed cells
    try:
        from datetime import date, datetime
        if isinstance(v, datetime):
            return v.date().isoformat()
        if isinstance(v, date):
            return v.isoformat()
    except Exception:  # noqa: BLE001
        pass
    if isinstance(v, str):
        s = v.strip()
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
        if m:
            return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


def _balance_titles_above_pool_order(ws) -> list[dict[str, Any]]:
    """Read the credit-union-supplied balance titles from ``BS Data``,
    above the ``Pool Order`` / ``Loan Pools`` anchor.

    Layout::

        Row 1: Pools | Types | Pool | <dates>...
        Row 2: New Autos          | 1,3,7,15 | Collateralized Loans | <bal>
        Row 3: Used Autos         | 2,4,...  | Collateralized Loans | <bal>
        ...
        Row 17: Sub-Total Loans   |          | Blank                  (skipped)
        Row 19: 3 Dealers         |          | Dealer Loans
        Row 21: Total Loans       |          | Blank                  (skipped)
        ...
        Row N: Pool Order         | Loan Pools | <dates>              (anchor)

    Each accepted row becomes ``{title, note, suggested_pool}`` where
    ``title`` is col A (the CU-supplied raw label that appears in the
    monthly data), ``note`` is col B (loan-code list, free-form), and
    ``suggested_pool`` is col C (the workbook's hint at which loan pool
    the row belongs in -- empty / 'Blank' rows are dropped).
    """
    anchor = _find_pool_order_anchor(ws)
    end_row = (anchor[0] - 1) if anchor else 60
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=2, max_row=end_row, values_only=True):
        if not row:
            continue
        a = row[0]
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        title = a.strip() if isinstance(a, str) else ""
        if not title:
            continue
        if not _is_real_pool_label(title):
            continue
        suggested = c.strip() if isinstance(c, str) else ""
        # Skip the ALLOWANCE row and any row whose suggested-pool cell is
        # literally "Blank" (totals/spacers).
        if suggested.lower() == "blank":
            continue
        low = title.lower()
        if low in seen:
            continue
        seen.add(low)
        if isinstance(b, float) and b.is_integer():
            note = str(int(b))
        else:
            note = str(b).strip() if b is not None else ""
        out.append({
            "title": title,
            "note": note,
            "suggested_pool": suggested,
        })
    return out


def _bs_loan_type_to_pool_map(ws) -> dict[str, str]:
    """From ``BS Data``, build a mapping of *loan-type label* (col A, row 2+)
    to *pool name* (col B).

    The BS Data tab has the layout::

        Row 1: Loan Type | Loan Pool | <date> | <date> | ...
        Row 2: NEW VEHICLE 22-26 | New Vehicle | ...
        Row 3: USED VEHICLE 27-31 | Used Vehicle | ...
        ...
        Row N: Total Loans | (blank)
        Row N+1: (blank rows)
        Row M: Pool Order | Loan Pools | <date> | ...   (canonical pool list)

    We stop reading at the first row whose col A label is a blocklist prefix
    (e.g. "Total Loans", "Pool Order") or after a stretch of blanks.
    """
    mapping: dict[str, str] = {}
    blanks = 0
    for row in ws.iter_rows(min_row=2, max_row=80, values_only=True):
        if not row:
            continue
        a = row[0]
        b = row[1] if len(row) > 1 else None
        a_s = a.strip() if isinstance(a, str) else ""
        b_s = b.strip() if isinstance(b, str) else ""
        if not a_s and not b_s:
            blanks += 1
            if blanks >= 3:
                break
            continue
        blanks = 0
        # Stop once we hit totals/summary/canonical-pool-list rows.
        if a_s and not _is_real_pool_label(a_s):
            break
        if a_s and b_s and _is_real_pool_label(b_s):
            mapping[a_s] = b_s
    return mapping


def _loan_code_to_pool_map_from_grades(ws) -> dict[str, str]:
    """From ``Grade Ranges & Loan Codes``, build a mapping of *raw loan
    code* (col S = index 18) to *pool name* (col T = index 19).

    Scans rows 2..200 (header is row 1). Stops counting after a stretch
    of blank rows. Codes are normalised to a stripped string; integer-valued
    codes are emitted without a trailing ``.0``.
    """
    mapping: dict[str, str] = {}
    blanks = 0
    for row in ws.iter_rows(min_row=2, max_row=200, values_only=True):
        # row may be shorter than 20 cells on early rows
        if row is None or len(row) < 20:
            blanks += 1
            if blanks >= 10:
                break
            continue
        code_v = row[18]
        pool_v = row[19]
        if code_v in (None, "") and pool_v in (None, ""):
            blanks += 1
            if blanks >= 10:
                break
            continue
        blanks = 0
        # Normalise code: ints come back as floats from some workbooks.
        if isinstance(code_v, float) and code_v.is_integer():
            code_s = str(int(code_v))
        else:
            code_s = str(code_v).strip() if code_v is not None else ""
        pool_s = str(pool_v).strip() if pool_v is not None else ""
        if not code_s or not pool_s:
            continue
        # Skip rows whose pool cell looks like a header/total marker.
        if not _is_real_pool_label(pool_s):
            continue
        # Skip header rows like ("Collateral Code", "Loan Pool").
        if pool_s.lower() in {"loan pool", "pool", "pool name", "loan pools"}:
            continue
        if code_s.lower().endswith(" code") or code_s.lower() in {"code", "loan code", "collateral code"}:
            continue
        # First write wins (preserves order of appearance on the sheet).
        mapping.setdefault(code_s, pool_s)
    return mapping


def _pool_settings_from_bs_co_dq(ws) -> dict[str, Any]:
    """Read per-pool settings from the ``BS CO DQ Data Enter`` tab.

    Layout::

        Row 4: Loan Pools | Risk Rated Yes/No | Balance Sheet Total |
               Delinquent Balances | DQ % | CG or RR | ACL Months |
               ACL Quarters | Pool Order | Default Loss Rate Yes/No
        Row 5..N: per-pool data, then a row labelled "Allowance for Credit
        Loss" (col A) followed by ancillary lines ("Credit Grade
        Deteriorated...", etc.) which are NOT loan pools.

    Returns::

        {"settings":    [ {name, risk_rated, acl_months,
                           use_default_mgmt_adj}, ... ],   # true loan pools
         "acl_balance": float}                              # ACL row col C

    Pool collection stops as soon as the ``Allowance for Credit Loss`` row
    is encountered; everything below it is ignored on purpose (the user
    will configure "Other Allowance Considerations" by hand in the wizard).
    """
    settings: list[dict[str, Any]] = []
    acl_balance: float = 0.0
    blanks = 0
    for row in ws.iter_rows(min_row=5, max_row=80, values_only=True):
        if row is None or len(row) < 10:
            blanks += 1
            if blanks >= 3:
                break
            continue
        name_v = row[0]
        name = name_v.strip() if isinstance(name_v, str) else ""
        if not name:
            blanks += 1
            if blanks >= 3:
                break
            continue
        blanks = 0
        low = name.lower()
        # Stop entirely once we hit the ACL row — capture its balance and
        # ignore everything beneath (Credit Grade Deteriorated, etc.).
        if low.startswith("allowance for credit loss") or low == "allowance":
            try:
                bv = row[2]
                if bv not in (None, ""):
                    acl_balance = float(bv)
            except (TypeError, ValueError):
                acl_balance = 0.0
            break
        # Skip totals/hidden/excluded rows but keep scanning.
        if low.startswith("grand total") or low.startswith("hide") \
                or low == "exclude" or low.startswith("total"):
            continue
        # Skip impairment-rollup / non-pool labels (Impaired Loans, Consumer
        # Problem Loans, Business Problem Loans, etc.). These show up above
        # the ACL row in some workbooks but are NOT loan pools.
        if not _is_real_pool_label(name):
            continue
        rr_v = row[1]
        rr = isinstance(rr_v, str) and rr_v.strip().lower() == "yes"
        acl_v = row[6]
        try:
            acl_m = int(acl_v) if acl_v not in (None, "") else 0
        except (TypeError, ValueError):
            acl_m = 0
        dl_v = row[9]
        use_default = isinstance(dl_v, str) and dl_v.strip().lower() == "yes"
        settings.append({
            "name": name,
            "risk_rated": rr,
            "acl_months": acl_m,
            "use_default_mgmt_adj": use_default,
        })
    return {"settings": settings, "acl_balance": acl_balance}


def _identity_from_bs_co_dq(ws) -> dict[str, Any]:
    """Read identity + economic baseline cells from ``BS CO DQ Data Enter``.

    Layout (label cell -> value cell)::

        L1 "Charter Number"     -> M1
        L2 "For Period Ending"  -> M2  (datetime)
        L3 "Credit Union Name"  -> M3
        Row 6 headers (L..Q):  State | County | Unemp % | FC/Person | BK | Population
        Row 7 values  (L..Q)

    Returns::

        {"charter_number": str,
         "period_end_date": "YYYY-MM-DD" | "",
         "cu_name": str,
         "state": str,
         "county": str,
         "unemployment_rate": float,   # stored as decimal (0.04 = 4%)
         "foreclosures": int,
         "bankruptcies": int,
         "population": int}
    """
    import datetime as _dt
    out = {
        "charter_number": "",
        "period_end_date": "",
        "cu_name": "",
        "state": "",
        "county": "",
        "unemployment_rate": 0.0,
        "foreclosures": 0,
        "bankruptcies": 0,
        "population": 0,
    }
    rows = list(ws.iter_rows(min_row=1, max_row=8, max_col=18, values_only=True))
    def cell(r1: int, col_letter: str):
        ci = ord(col_letter.upper()) - ord('A')
        if r1 - 1 >= len(rows):
            return None
        row = rows[r1 - 1]
        if row is None or ci >= len(row):
            return None
        return row[ci]

    # M1 charter
    v = cell(1, "M")
    if v not in (None, ""):
        if isinstance(v, float) and v.is_integer():
            out["charter_number"] = str(int(v))
        else:
            out["charter_number"] = "".join(ch for ch in str(v).strip() if ch.isdigit())

    # M2 period ending
    v = cell(2, "M")
    if isinstance(v, _dt.datetime):
        out["period_end_date"] = v.date().isoformat()
    elif isinstance(v, _dt.date):
        out["period_end_date"] = v.isoformat()
    elif isinstance(v, str) and v.strip():
        out["period_end_date"] = v.strip()

    # M3 cu name
    v = cell(3, "M")
    if isinstance(v, str):
        out["cu_name"] = v.strip()

    # L7..Q7
    v = cell(7, "L")
    if isinstance(v, str):
        out["state"] = v.strip()
    v = cell(7, "M")
    if isinstance(v, str):
        out["county"] = v.strip()

    def _to_int(val):
        if val in (None, ""):
            return 0
        try:
            return int(round(float(val)))
        except (TypeError, ValueError):
            return 0

    def _to_float(val):
        if val in (None, ""):
            return 0.0
        try:
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    out["unemployment_rate"] = _to_float(cell(7, "N"))
    out["foreclosures"]      = _to_int(cell(7, "O"))
    out["bankruptcies"]      = _to_int(cell(7, "P"))
    out["population"]        = _to_int(cell(7, "Q"))
    return out


def _pool_names_from_co(ws) -> list[str]:
    """``CO Hist Data`` col A starting row 5 lists each pool once per section."""
    pools: list[str] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=5, max_row=80, values_only=True):
        v = row[0]
        if isinstance(v, str):
            name = v.strip()
            if not _is_real_pool_label(name):
                continue
            if name not in seen:
                pools.append(name)
                seen.add(name)
    return pools


def _grades_from_sheet(ws) -> list[dict[str, Any]]:
    """Rows 3+ of ``Grade Ranges & Loan Codes`` are: order, label, range_str,
    range_min, distribution_factor, default_loss_rate, default_mgmt_adj."""
    grades: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=3, max_row=20, values_only=True):
        if len(row) < 6:
            continue
        order, label, range_str, range_min, _dist, loss_rate = row[:6]
        if label is None or str(label).strip() == "":
            continue
        label_s = str(label).strip()
        # Skip helper / hidden / total rows
        low = label_s.lower()
        if any(low.startswith(p) for p in _GRADE_BLOCKLIST_PREFIXES):
            continue
        try:
            mn = int(range_min) if range_min not in (None, "") else 0
        except (TypeError, ValueError):
            mn = 0
        mx = _max_from_range_str(str(range_str) if range_str else "", mn)
        try:
            rate = float(loss_rate) if loss_rate not in (None, "") else 0.0
        except (TypeError, ValueError):
            rate = 0.0
        grades.append({
            "label": label_s,
            "min_score": mn,
            "max_score": mx,
            "reserve_rate": rate,
        })
    return grades


def _max_from_range_str(rng: str, fallback_min: int) -> int:
    """Parse strings like '730+', '729-680', '680-729', '<549', '549 or less'."""
    s = rng.strip()
    if not s:
        return fallback_min
    if s.endswith("+"):
        return 850
    if s.startswith("<") or "less" in s.lower():
        nums = re.findall(r"\d+", s)
        return int(nums[0]) if nums else fallback_min
    nums = [int(n) for n in re.findall(r"\d+", s)]
    if not nums:
        return fallback_min
    return max(nums)


def _history_range(ws) -> tuple[str | None, str | None, int]:
    """Row 5 of ``HIst Bal Data`` lists every monthly date as a header.
    Return (start, end, count)."""
    rows = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))
    if not rows:
        return None, None, 0
    dates: list[str] = []
    for v in rows[0]:
        if v is None:
            continue
        s = str(v)
        m = re.search(r"(\d{4}-\d{2}-\d{2})", s)
        if m:
            dates.append(m.group(1))
    if not dates:
        return None, None, 0
    return dates[0], dates[-1], len(dates)


def _count_pools_with_data(ws, start_row: int, start_col: int) -> int:
    """Count rows whose label (col A) is text and whose data columns from
    ``start_col`` onward contain at least one non-zero numeric value."""
    count = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        label = row[0]
        if not isinstance(label, str) or not label.strip():
            continue
        # only first 60 pool rows; defensively cap
        if count >= 60:
            break
        # check data
        for v in row[start_col - 1: start_col - 1 + 200]:
            if isinstance(v, (int, float)) and v not in (0, 0.0):
                count += 1
                break
    return count


# ─────────────────────────────────────────────────────────────────
# Public helper: scan a folder for historical data files
# ─────────────────────────────────────────────────────────────────

# Patterns used to classify files found in a historical data folder.
_WARM_FILE_RX = re.compile(
    r"^(\d{4}-\d{2})(?:-\d{2})?\s+CECL-Migration-WARM.*\.xlsx$", re.IGNORECASE
)
_CO_FILE_RX = re.compile(
    r"(charge[\s_\-]*off|charge_off_track|co[\s_\-]*hist)", re.IGNORECASE
)
_RECOV_FILE_RX = re.compile(
    r"(recov(?:er(?:y|ies))?|historical[\s_\-]*recov)", re.IGNORECASE
)
_IMPAIRED_FILE_RX = re.compile(
    r"(impaired[\s_\-]*loan|cecl[\s_\-]*migration[\s_\-]*impaired)", re.IGNORECASE
)
_CREDIT_PULL_RX = re.compile(
    r"(credit[\s_\-]*pull|fico[\s_\-]*pull)", re.IGNORECASE
)
_MONTHLY_BAL_RX = re.compile(
    r"(monthly[\s_\-]*bal|loan[\s_\-]*bal[\s_\-]*hist)", re.IGNORECASE
)
_LOAN_DATA_RX = re.compile(
    r"(aires|loan[\s_\-]*data|loan[\s_\-]*file|lndn)", re.IGNORECASE
)

# Month-name → zero-padded month number map (for filenames like "Mar_2026").
_MONTH_NAME_MAP = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "may": "05", "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "sept": "09", "oct": "10", "nov": "11", "dec": "12",
}
_MONTH_NAME_RX = re.compile(
    r"(jan|feb|mar|apr|may|jun|jul|aug|sep(?:t)?|oct|nov|dec)[a-z]*"
    r"[-_\s]+?(20\d{2})",
    re.IGNORECASE,
)


def _period_from_filename(name: str) -> str | None:
    """Return 'YYYY-MM' from a filename, handling both numeric and month-name dates."""
    # Numeric first: 2026-03, 2026_03, etc.
    m = re.search(r"(20\d{2})[-_\s](\d{2})", name)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    # Month-name: Mar_2026, March 2026, etc.
    m2 = _MONTH_NAME_RX.search(name)
    if m2:
        mon = _MONTH_NAME_MAP.get(m2.group(1).lower()[:3], "")
        if mon:
            return f"{m2.group(2)}-{mon}"
    return None


def scan_historical_folder(folder: str | Path) -> dict[str, Any]:
    """Scan *folder* and classify relevant historical data files into:

    * ``warm_files``     – CECL-Migration-WARM quarterly workbooks
    * ``co_files``       – charge-off tracking files
    * ``recov_files``    – recovery history files
    * ``impaired_files`` – impaired-loan summary files
    * ``credit_pull_files`` – standalone credit-pull files
    * ``monthly_files``  – monthly balance / ALLL files
    * ``loan_data_files`` – loan data extracts (e.g. Aries loan files)
    * ``other_files``    – everything else (supported extensions only)

    Returns::

        {
            "ok": bool,
            "error": str | None,
            "folder": str,
            "warm_files": [{"name": str, "period": str | None, "path": str}, ...],
            "co_files":       [...],
            "recov_files":    [...],
            "impaired_files": [...],
            "credit_pull_files": [...],
            "monthly_files":  [...],
            "loan_data_files": [...],
            "other_files":    [...],
            "history_start":  str | None,   # earliest WARM period detected
            "history_end":    str | None,   # latest WARM period detected
            "warm_months":    int,           # total months from WARM file headers
        }
    """
    folder = Path(folder)
    result: dict[str, Any] = {
        "ok": False,
        "error": None,
        "folder": str(folder),
        "warm_files": [],
        "co_files": [],
        "recov_files": [],
        "impaired_files": [],
        "credit_pull_files": [],
        "monthly_files": [],
        "loan_data_files": [],
        "other_files": [],
        "history_start": None,
        "history_end": None,
        "warm_months": 0,
    }

    if not folder.exists():
        result["error"] = f"Folder does not exist: {folder}"
        return result
    if not folder.is_dir():
        result["error"] = f"Not a directory: {folder}"
        return result

    result["ok"] = True

    allowed_exts = {".xlsx", ".xls", ".xlsm", ".csv", ".txt"}
    for path in sorted(folder.rglob("*")):
        if not path.is_file():
            continue
        name = path.name
        if path.suffix.lower() not in allowed_exts:
            continue
        if name.startswith("~$"):
            continue
        if name.upper().startswith("DNU"):
            continue

        entry = {"name": name, "period": None, "path": str(path)}

        m = _WARM_FILE_RX.match(name)
        if m:
            entry["period"] = m.group(1)
            result["warm_files"].append(entry)
            continue

        lname = name.lower()
        if _CO_FILE_RX.search(lname):
            entry["period"] = _period_from_filename(name)
            result["co_files"].append(entry)
        elif _IMPAIRED_FILE_RX.search(lname):
            entry["period"] = _period_from_filename(name)
            result["impaired_files"].append(entry)
        elif _CREDIT_PULL_RX.search(lname):
            entry["period"] = _period_from_filename(name)
            result["credit_pull_files"].append(entry)
        elif _RECOV_FILE_RX.search(lname):
            entry["period"] = _period_from_filename(name)
            result["recov_files"].append(entry)
        elif _MONTHLY_BAL_RX.search(lname):
            entry["period"] = _period_from_filename(name)
            result["monthly_files"].append(entry)
        elif _LOAN_DATA_RX.search(lname):
            entry["period"] = _period_from_filename(name)
            result["loan_data_files"].append(entry)
        else:
            result["other_files"].append(entry)

    # Sort WARM files by period descending so newest is first.
    result["warm_files"].sort(key=lambda e: e["period"] or "", reverse=True)

    # Derive history span from WARM file list.
    periods = sorted(
        e["period"] for e in result["warm_files"] if e["period"]
    )
    if periods:
        result["history_start"] = periods[0]
        result["history_end"] = periods[-1]
        # Rough month count: difference between first and last period + 1
        try:
            y0, m0 = int(periods[0][:4]), int(periods[0][5:7])
            y1, m1 = int(periods[-1][:4]), int(periods[-1][5:7])
            result["warm_months"] = (y1 - y0) * 12 + (m1 - m0) + 1
        except (ValueError, IndexError):
            pass

    return result
