"""Lightweight parser for the Historical Month-End Balances workbook.

Used by the new-CU wizard's Historical Data step (non-WARM path) to extract
the distinct loan-type labels from column A of the uploaded file so the user
can map each label to a pool.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any


# Labels in col A that aren't real loan-type rows.
_LABEL_BLOCKLIST_PREFIXES = (
    "total", "sub total", "subtotal", "allowance", "balancer",
    "pool order", "net ", "recover", "grand total",
)


def _is_real_label(name: str) -> bool:
    low = name.strip().lower()
    if not low:
        return False
    return not any(low.startswith(p) for p in _LABEL_BLOCKLIST_PREFIXES)


def extract_balance_labels(path: Path, sheet: str | None = None) -> dict[str, Any]:
    """Open ``path`` and return the distinct loan-type labels in column A.

    Returns dict::

        {
          "ok": bool,
          "error": str | None,
          "sheet": str,             # the sheet we actually read
          "labels": list[str],      # distinct col-A labels, in order
        }

    If ``sheet`` is None we pick the first sheet whose row 1, col A header
    contains "loan" (case-insensitive) or "type", falling back to the first
    visible sheet.
    """
    out: dict[str, Any] = {
        "ok": False, "error": None, "sheet": "", "labels": [],
    }
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"openpyxl import failed: {exc}"
        return out

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Could not open workbook: {exc}"
        return out

    try:
        # Pick sheet.
        target = sheet
        if not target:
            for nm in wb.sheetnames:
                ws = wb[nm]
                first_a = None
                for r in ws.iter_rows(min_row=1, max_row=1, max_col=2,
                                      values_only=True):
                    if r:
                        first_a = r[0]
                    break
                if isinstance(first_a, str) and (
                    "loan" in first_a.lower() or "type" in first_a.lower()
                ):
                    target = nm
                    break
            if not target:
                target = wb.sheetnames[0]

        out["sheet"] = target
        ws = wb[target]

        labels: list[str] = []
        seen: set[str] = set()
        blanks = 0
        for row in ws.iter_rows(min_row=2, max_row=200, max_col=1,
                                values_only=True):
            v = row[0] if row else None
            if not isinstance(v, str) or not v.strip():
                blanks += 1
                if blanks >= 5:
                    break
                continue
            blanks = 0
            name = v.strip()
            key = name.lower()
            if key in seen:
                continue
            if not _is_real_label(name):
                # Stop at the first totals/summary row — anything past it is
                # usually a second block (e.g. "Pool Order") we don't want.
                break
            labels.append(name)
            seen.add(key)
            if len(labels) >= 60:
                break

        out["labels"] = labels
        out["ok"] = True
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Parse error: {exc}"
    finally:
        wb.close()

    return out
