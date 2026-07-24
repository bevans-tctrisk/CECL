"""Auto-setup orchestrator for the Migration wizard.

Given a folder full of credit-union raw data, this module:

  1. Classifies the files (loan extracts, balance sheets, impaired loans,
     credit pulls, WARM workbooks, charge-off & recovery files, etc.)
     by leaning on ``warm_parser.scan_historical_folder``.
  2. Runs the appropriate per-file parser on the most-representative
     file of each type (sample_parser for loan extracts, monthly_bal_parser
     for balance sheet workbooks, etc.).
  3. Mutates a wizard ``state`` dict to populate as much as it can
     auto-derive (file paths, file patterns, column mappings, pool codes,
     mgmt-adj baselines, report defaults, etc.).
  4. Returns a per-step "what was filled" report and a list of HIL steps
     that still need user attention.

The wizard's existing per-step handlers continue to own all editing UX;
this module is only an upfront populater.

Patterns learned from the Census FCU greenfield script
(``scripts/setup_census_scratch.py``) are baked in:

  * mgmt_adj baselines (ltv_baseline=0.9, probability_factor=0.35).
  * reports default: Vizo-only when ``"vizo financial"`` is in the folder
    path (the Vizo Financial corporate-CU client convention); otherwise TCT.
  * impaired-loans block emitted only when an impaired-loans file is
    discovered for the target snapshot.
  * credit-pull block emitted only when a matching credit-pull file
    is discovered for the target snapshot.
  * pool_code_split default ``"/"``.
  * credit_grades default 5-grade industry standard with reserve_rate on
    every entry (so ``cecl_engine.get_reserve_rate`` doesn't crash).
"""
from __future__ import annotations

import re
import shutil
import tempfile
from pathlib import Path
from typing import Any

from cecl_ui.services import (
    config_service,
    column_mapping_suggestions,
    monthly_bal_parser,
    sample_parser,
    warm_parser,
)


# --------------------------------------------------------------------------
# File classification
# --------------------------------------------------------------------------

# Extra patterns layered on top of ``warm_parser.scan_historical_folder``.
# That function already buckets WARM / CO / recov / impaired / credit-pull
# / monthly-bal / loan-data, but we want to flag annual balance-sheet
# workbooks separately (file names like ``2025 Balance Sheets.xlsx``).
_ANNUAL_BAL_RX = re.compile(
    r"(?:^|[^a-z])(?P<yr>20\d{2})\s*balance\s*sheets?",
    re.IGNORECASE,
)

# A "monthly detailed balance sheet" file (per-month workbook).
# Pattern: name contains "balance" + "sheet" + an 8-digit date.
_MONTHLY_DETAIL_RX = re.compile(
    r"detailed?\s*balance\s*sheet.*?(?P<dt>\d{8})",
    re.IGNORECASE,
)

# A consolidated SINGLE historical balance workbook (one file, wide
# layout with one column per month-end). Filename hints:
#   * "Historical Loan Balances"
#   * "Balance History"
#   * "Balance_History_-_<CU>"
#   * "Loan Balance History"
#   * "Historical Balance(s)"
# When found, we set hist_balance_source = "single_workbook" and
# analyse_file() detects sheet / header_row / pool_name_col automatically.
_SINGLE_HIST_BAL_RX = re.compile(
    r"(?:historical[\s_\-]*(?:loan[\s_\-]*)?balances?"
    r"|balance[\s_\-]*history"
    r"|loan[\s_\-]*balance[\s_\-]*history)",
    re.IGNORECASE,
)

# Recognise a 5300 / call-report file (informational only).
_5300_RX = re.compile(r"(?:^|[^a-z])(?:5300|call[\s_\-]*report)", re.IGNORECASE)

# Broader loan-extract patterns. warm_parser's _LOAN_DATA_RX only matches
# "aires" / "loan data" / "loan file" / "lndn" — too narrow for CUs like
# Census FCU whose monthly extracts are named "Dec 2025 Loans V2.xlsx" or
# "Dec 2025 Cuma Loans.xlsx". We re-classify ``other_files`` against this
# richer pattern, but ONLY when the file ALSO has a month/year fingerprint
# in its name (so we don't accidentally grab "Loan Type Codes.xlsx" etc).
_LOAN_EXTRA_RX = re.compile(
    r"(?:^|[\s_\-])(?:loans?(?:[\s_\-]*v?\d+)?|"
    r"cuma[\s_\-]*loans?|"
    # CUMA MTG Servicing reports are the mortgage-portfolio loan
    # extract for CUs that ship them (Vizo client convention, e.g.
    # ``CUMA MTG Servicing Report 12-31-2025 V2.xls``). Treat as a
    # loan-data extract.
    r"cuma[\s_\-]*mtg(?:[\s_\-]*servicing)?|"
    r"cuma[\s_\-]*servicing|"
    r"active[\s_\-]*loans?|"
    r"loan[\s_\-]*portfolio|"
    r"loan[\s_\-]*detail|"
    r"loan[\s_\-]*list|"
    # Symitar/Episys CECL bucket extracts: ceclce (Closed-End),
    # cecloe (Open-End), ceclcc / ceclcc1 (Credit Card). Some CUs
    # space-separate ("cecl ce"), some run together ("ceclce").
    r"cecl[\s_\-]*(?:ce|oe|cc\d*))"
    r"(?:[\s_\-]|\.|$)",
    re.IGNORECASE,
)

# A name "looks dated" when it has a YYYY-MM, YYYYMMDD, MM-YYYY, or
# month-name + year combo.
_DATE_HINT_RX = re.compile(
    r"(20\d{2}[-_]?\d{2})"
    r"|(\d{2}[-_]?20\d{2})"
    r"|(\d{8})"
    r"|((?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*[-_\s]+20\d{2})",
    re.IGNORECASE,
)

# Filenames that explicitly are NOT loan extracts even when they contain
# the word "loans" (mapping / reference workbooks, not data).
_LOAN_FALSE_POSITIVE_RX = re.compile(
    r"(loan[\s_\-]*type[\s_\-]*code|loan[\s_\-]*code[\s_\-]*map|"
    r"loan[\s_\-]*pool[\s_\-]*map|collateral[\s_\-]*code|"
    r"impaired[\s_\-]*loan|standalone[\s_\-]*impaired|"
    r"collateral[\s_\-]*dependent|"
    r"improved[\s_\-]*deteriorated|"
    r"impr[\s_\-]*deter|"
    # Loan-participation registers and "individually identified" impaired
    # registers both contain the word "loan(s)" but are NOT the monthly
    # loan-data extract this tool consumes.
    r"loan[\s_\-]*participation|participation[\s_\-]*information|"
    r"individually[\s_\-]*identified|"
    r"watchlist|risk[\s_\-]*rating|"
    r"aires[\s_\-]*shares|shares[\s_\-]*aires|"
    r"loan[\s_\-]*code|loan[\s_\-]*pool[\s_\-]*name)",
    re.IGNORECASE,
)

# Aires/AIRES files come in TWO varieties: "Aires Loans" (the loan
# portfolio extract this tool consumes) and "Aires Shares" (the member
# share/deposit export, NOT loan data). Any filename matching this
# pattern is excluded from ``loan_data_files`` even if warm_parser's
# ``_LOAN_DATA_RX`` already matched on the bare "aires" token.
_AIRES_SHARES_RX = re.compile(
    r"aires[\s_\-]*shares|shares[\s_\-]*aires",
    re.IGNORECASE,
)

# Phase 9.26: filename heuristic for loan-data extracts whose loan codes
# legitimately contain ``/`` (so the global ``pool_code_split = "/"`` would
# corrupt them). CUMA / mortgage servicing exports ship codes like
# ``15/15 ARM``, ``30/30 FIXED`` where the ``/`` is part of the label, not a
# prefix:description separator. When a loan-extract candidate's filename
# matches this regex the auto-scan calls ``analyse_sample_file`` with
# ``split_char=""`` AND records ``pool_code_split=""`` on the staged extract
# entry so the importer's per-file override (Phase 9.22) takes effect at
# runtime. Filename keywords: ``CUMA``, ``MTG``, ``Mortgage``, ``Servicing``.
_NO_SPLIT_FILENAME_RX = re.compile(
    r"(?:^|[\s_\-])(?:cuma|mtg|mortgage|servicing)(?:[\s_\-]|$|\.)",
    re.IGNORECASE,
)

# Phase 9.26: value-shape detector for composite mortgage codes. Matches
# raw values like ``15/15 ARM``, ``30/30 FIXED``, ``5/1 ARM``, ``15 YR FIXED``
# — anything where a ``/`` separator is followed by alphabetic content
# rather than a pure prefix:description split. Used as a fallback when the
# filename heuristic above misses (some CUs ship CUMA-style data in a
# generically-named workbook).
_COMPOSITE_CODE_VALUE_RX = re.compile(
    r"^\s*\d{1,3}\s*/\s*\d{1,3}\s+[A-Za-z]",
)


def _looks_like_no_split_extract(
    name: str,
    sample_values: list[str] | None = None,
) -> bool:
    """Return True if this loan-extract candidate carries composite codes
    where ``/`` is part of the label (CUMA mortgage convention).

    Two independent triggers — either flips it to no-split:

    * Filename matches CUMA/MTG/Mortgage/Servicing regex, OR
    * ANY of the supplied raw pool-code values matches the
      ``\\d+/\\d+\\s+[A-Z]`` composite-code shape.
    """
    if name and _NO_SPLIT_FILENAME_RX.search(str(name)):
        return True
    if sample_values:
        for val in sample_values:
            if val and _COMPOSITE_CODE_VALUE_RX.match(str(val)):
                return True
    return False

# A consolidated "Balance Sheets <CU>" pool-grouping workbook (Vizo
# Financial analyst convention). Filename starts with "Balance Sheets"
# (or "Balance Sheet") followed by a NON-year token (the CU's short
# name). Structure: col A holds NCUA loan codes ("NNN-NN") on sub-cat
# rows and an integer on pool-header rows; col B holds labels with the
# pool-header rows bolded. See ``extract_pools_from_balance_workbook``.
_CONSOLIDATED_POOL_BAL_RX = re.compile(
    r"^balance\s*sheets?\s+(?!20\d{2}\b)\S",
    re.IGNORECASE,
)

# Sub-category code in col A (e.g. "708-01", "702-03"). Used by
# ``extract_pools_from_balance_workbook``.
_SUBCAT_CODE_RX = re.compile(r"^\d{3}-\d{2}$")

# Grand-total row in col B (terminates the pool walk).
_GRAND_TOTAL_RX = re.compile(r"^\s*total\s+loans\s*$", re.IGNORECASE)


def _looks_like_loan_extract(name: str) -> bool:
    """True when a filename looks like a monthly loan-data extract."""
    if _LOAN_FALSE_POSITIVE_RX.search(name):
        return False
    if not _LOAN_EXTRA_RX.search(name):
        return False
    # Require either a date fingerprint OR the file is named generically
    # with "AIRES"/"LNDN" (those are already pre-classified by
    # warm_parser, so a hit here means the file has SOME month/year hint).
    return bool(_DATE_HINT_RX.search(name))


# Fallback period extractor: warm_parser._period_from_filename only
# handles YYYY-MM and month-name+year; CUMA MTG Servicing reports use
# MM-DD-YYYY (e.g. "CUMA MTG Servicing Report 12-31-2025 V2.xls"). This
# helper layers on the missing shapes so ``_pick_latest_for_period``
# can still respect snapshot for those files.
_MD_YYYY_RX = re.compile(r"(?<!\d)(\d{1,2})[-_/](\d{1,2})[-_/](20\d{2})(?!\d)")
_MM_YYYY_RX = re.compile(r"(?<!\d)(\d{1,2})[-_/](20\d{2})(?!\d)")
# Compact ISO date fingerprint (e.g. ``20230228_Cottonwood CU_Loans.xlsx``).
# Many core exports lead the filename with a bare YYYYMMDD stamp.
_YYYYMMDD_RX = re.compile(r"(?<!\d)(20\d{2})(\d{2})(\d{2})(?!\d)")


def _fallback_period_from_filename(name: str) -> str | None:
    """Return ``YYYY-MM`` from YYYYMMDD / MM-DD-YYYY / MM-YYYY shapes."""
    m = _YYYYMMDD_RX.search(name)
    if m:
        mo = int(m.group(2))
        if 1 <= mo <= 12:
            return f"{m.group(1)}-{mo:02d}"
    m = _MD_YYYY_RX.search(name)
    if m:
        mo = int(m.group(1))
        if 1 <= mo <= 12:
            return f"{m.group(3)}-{mo:02d}"
    m = _MM_YYYY_RX.search(name)
    if m:
        mo = int(m.group(1))
        if 1 <= mo <= 12:
            return f"{m.group(2)}-{mo:02d}"
    return None


def _bucket_extra(entry: dict[str, Any], folder_path: str) -> str | None:
    """Return an extra bucket name for an ``other_files`` entry, or None."""
    name = entry.get("name", "")
    if _ANNUAL_BAL_RX.search(name):
        m = _ANNUAL_BAL_RX.search(name)
        try:
            entry["year"] = int(m.group("yr"))
        except (ValueError, AttributeError, TypeError):
            entry["year"] = None
        return "annual_balance_files"
    if _MONTHLY_DETAIL_RX.search(name):
        return "monthly_detail_balance_files"
    if _SINGLE_HIST_BAL_RX.search(name):
        return "single_hist_bal_files"
    if _CONSOLIDATED_POOL_BAL_RX.search(name):
        return "consolidated_pool_balance_files"
    if _5300_RX.search(name):
        return "five_thirtythousand_files"
    if _looks_like_loan_extract(name):
        # Promote to loan_data_files. period_from_filename was already
        # set by warm_parser when it had a chance; re-derive here for
        # safety.
        if not entry.get("period"):
            entry["period"] = warm_parser._period_from_filename(name)  # noqa: SLF001
        # warm_parser._period_from_filename only handles YYYY-MM and
        # month-name+year shapes. CUMA MTG Servicing reports are named
        # like "CUMA MTG Servicing Report 12-31-2025 V2.xls" (MM-DD-YYYY)
        # and a few other CU exports use MM/YYYY or M-YYYY. Layer those
        # on as a fallback so multi-extract pickup respects snapshot.
        if not entry.get("period"):
            entry["period"] = _fallback_period_from_filename(name)
        return "loan_data_files"
    return None


def classify_folder(folder: str | Path) -> dict[str, Any]:
    """Walk *folder* and bucket every supported file by purpose.

    Wraps ``warm_parser.scan_historical_folder`` (which already does WARM /
    CO / recov / impaired / credit-pull / monthly-bal / loan-data) and
    layers on annual / monthly balance-sheet workbook detection plus 5300
    detection. Returns the same shape as ``scan_historical_folder`` plus
    additional bucket keys.
    """
    base = warm_parser.scan_historical_folder(folder)
    if not base.get("ok"):
        return base

    base.setdefault("annual_balance_files", [])
    base.setdefault("monthly_detail_balance_files", [])
    base.setdefault("single_hist_bal_files", [])
    base.setdefault("five_thirtythousand_files", [])
    base.setdefault("consolidated_pool_balance_files", [])

    # warm_parser._LOAN_DATA_RX matches the bare "aires" token, so it
    # promotes BOTH "Aires Loans <date>.xlsx" (the loan extract this
    # tool consumes) AND "Aires Shares <date>.xlsx" (the member
    # share/deposit export — NOT loan data). It also matches other
    # "loan(s)"-bearing report names that are NOT the monthly extract
    # (loan-participation registers, "individually identified" impaired
    # registers, improved/deteriorated reports, code maps, ...). Demote
    # any such entry from ``loan_data_files`` to ``other_files`` so the
    # sample picker and Step 13 only see real loan extracts. Files routed
    # through ``other_files`` never come back because ``_bucket_extra`` ->
    # ``_looks_like_loan_extract`` re-checks ``_LOAN_FALSE_POSITIVE_RX``.
    _ldf = base.get("loan_data_files") or []
    if _ldf:
        _kept: list[dict[str, Any]] = []
        _demoted: list[dict[str, Any]] = []
        for _e in _ldf:
            _nm = _e.get("name") or ""
            if _AIRES_SHARES_RX.search(_nm) or _LOAN_FALSE_POSITIVE_RX.search(_nm):
                _demoted.append(_e)
            else:
                _kept.append(_e)
        base["loan_data_files"] = _kept
        if _demoted:
            base.setdefault("other_files", []).extend(_demoted)

    # Re-route files that fell into ``other_files`` into our extra buckets.
    leftover: list[dict[str, Any]] = []
    folder_str = str(folder)
    for entry in base.get("other_files") or []:
        extra = _bucket_extra(entry, folder_str)
        if extra:
            base[extra].append(entry)
        else:
            leftover.append(entry)
    base["other_files"] = leftover

    # warm_parser._CO_FILE_RX is strict ("charge off" / "co hist"). Many
    # CUs (Destinations, Shuford-style) name their combined CO+Recovery
    # files "Recoveries - Charge-offs", "Recoveries and Charged off",
    # "Recoveries-Chg-offs" — none of which start with the CO keyword and
    # so end up exclusively in recov_files (or vice-versa). Cross-list any
    # entry whose filename matches the BROADER pattern so Step 5 and
    # Step 6 both see combined files.
    _broad_co_rx = re.compile(
        r"(charge[ds]?[\s_\-]*off|chg[\s_\-]*off|co[\s_\-]*hist|charge_off_track)",
        re.IGNORECASE,
    )
    _broad_recov_rx = re.compile(
        r"(recov(?:er(?:y|ies))?|historical[\s_\-]*recov)",
        re.IGNORECASE,
    )

    def _names_in(bucket: list[dict[str, Any]]) -> set[str]:
        return {(e.get("name") or "").lower() for e in bucket}

    co_list = base.setdefault("co_files", [])
    rc_list = base.setdefault("recov_files", [])
    co_names = _names_in(co_list)
    rc_names = _names_in(rc_list)

    for entry in list(rc_list):
        nm = (entry.get("name") or "")
        if _broad_co_rx.search(nm) and nm.lower() not in co_names:
            co_list.append(entry)
            co_names.add(nm.lower())
    for entry in list(co_list):
        nm = (entry.get("name") or "")
        if _broad_recov_rx.search(nm) and nm.lower() not in rc_names:
            rc_list.append(entry)
            rc_names.add(nm.lower())

    return base


# --------------------------------------------------------------------------
# Consolidated "Balance Sheets <CU>" pool extractor
# --------------------------------------------------------------------------


def _cell_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _is_bold(cell: Any) -> bool:
    try:
        return bool(cell.font and cell.font.bold)
    except Exception:
        return False


def extract_pools_from_balance_workbook(
    xlsx_path: str | Path,
) -> dict[str, Any]:
    """Walk a consolidated 'Balance Sheets <CU>' workbook.

    Convention learned from Destinations CU 66333 (Vizo Financial client):

      * Col A holds either an INTEGER (pool number, 1/4/5/...), a
        ``"NNN-NN"`` NCUA loan code (sub-category), or is empty.
      * Col B holds the label (pool name OR sub-category name).
      * **BOLD col B with INTEGER col A** => POOL row.
      * **BOLD col B with EMPTY col A** and leading whitespace in label
        => CONTINUATION of the previous pool's name (some pool names
        span two rows, e.g. ``"Total Loans/Lines of Credit Secured" +
        "   by 1st Lien 1-4 Family Residential Properties"``).
      * Sub-category rows (``NNN-NN`` col A) flow into the NEXT pool row.
      * ``"Total Loans"`` row (col B only, no children) is the grand
        total: terminates the walk.

    Returns::

        {
            "ok": True,
            "sheet": str,
            "pools": [
                {"name": str,
                 "subcategories": [{"code": "NNN-NN", "label": str}, ...]},
                ...
            ],
            "subcategory_to_pool": {label: pool_name, ...},
            "code_to_pool":        {code:  pool_name, ...},
        }
    """
    try:
        import openpyxl  # local import — heavy
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"openpyxl not available: {exc}"}

    path = Path(xlsx_path)
    if not path.exists():
        return {"ok": False, "error": f"file not found: {path}"}
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"openpyxl load failed: {exc}"}

    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    pools: list[dict[str, Any]] = []
    subcat_buf: list[dict[str, str]] = []
    pending_pool_name: str | None = None

    max_row = ws.max_row or 0
    for r in range(1, max_row + 1):
        a_cell = ws.cell(row=r, column=1)
        b_cell = ws.cell(row=r, column=2)
        a_raw = a_cell.value
        b_raw = b_cell.value

        a_txt = _cell_text(a_raw)
        b_txt_raw = "" if b_raw is None else str(b_raw)
        b_txt = b_txt_raw.strip()

        if not a_txt and not b_txt:
            continue

        # Grand-total row -> stop processing.
        if not a_txt and _GRAND_TOTAL_RX.match(b_txt):
            break

        # Sub-category row (NCUA loan code in col A).
        if _SUBCAT_CODE_RX.match(a_txt):
            if b_txt:
                subcat_buf.append({"code": a_txt, "label": b_txt})
            continue

        # Pool row: BOLD col B, with integer in col A OR (empty col A
        # AND leading whitespace in label = continuation of prior pool).
        if _is_bold(b_cell) and b_txt:
            is_continuation = (
                not a_txt
                and (b_txt_raw != b_txt_raw.lstrip())
            )
            is_integer_a = False
            if a_txt:
                try:
                    int(float(a_txt))
                    is_integer_a = True
                except (ValueError, TypeError):
                    is_integer_a = False

            if is_continuation and pending_pool_name and pools:
                pending_pool_name = f"{pending_pool_name} {b_txt}".strip()
                pools[-1]["name"] = pending_pool_name
                continue

            if is_integer_a:
                pool = {
                    "name": b_txt,
                    "subcategories": list(subcat_buf),
                }
                pools.append(pool)
                pending_pool_name = b_txt
                subcat_buf = []
                continue
        # Anything else: ignore (header rows, blank labels, formulas).

    # Build flat lookup dicts.
    subcat_to_pool: dict[str, str] = {}
    code_to_pool: dict[str, str] = {}
    for p in pools:
        for sc in p.get("subcategories") or []:
            subcat_to_pool[sc["label"]] = p["name"]
            code_to_pool[sc["code"]] = p["name"]

    return {
        "ok": True,
        "sheet": sheet_name,
        "pools": pools,
        "subcategory_to_pool": subcat_to_pool,
        "code_to_pool": code_to_pool,
    }


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------

_AUTOSCAN_TEMP = Path(tempfile.gettempdir()) / "cecl_ui_autoscan"


def _copy_into_tmp(src: Path, subfolder: str) -> Path:
    """Copy *src* into the wizard's working temp area and return the new path.

    The wizard's per-step upload handlers (Step 2 Sample, Step 15 Impaired,
    Step 10 Credit Pull, Step 5 Monthly Balance, etc.) all store uploaded
    files in dedicated subfolders of ``%TEMP%`` so the parsers can read
    them on subsequent requests without holding open handles. We mirror
    that pattern here so post-scan navigation between steps remains
    consistent.
    """
    dest_dir = _AUTOSCAN_TEMP / subfolder
    dest_dir.mkdir(parents=True, exist_ok=True)
    target = dest_dir / src.name
    try:
        shutil.copy2(src, target)
    except OSError:
        pass
    return target


def _safe_period(entry: dict[str, Any]) -> str | None:
    """Return ``YYYY-MM`` period string from a classifier entry, or None."""
    period = entry.get("period")
    if not period:
        return None
    period = str(period)
    if len(period) >= 7 and period[4] == "-":
        return period[:7]
    return None


def _pick_latest_for_period(
    entries: list[dict[str, Any]], snapshot_yyyymm: str | None
) -> dict[str, Any] | None:
    """Pick the entry whose period matches snapshot, else the latest period.

    When multiple entries share the target period, prefer those whose
    filename does NOT contain a subsystem keyword (cuma/mortgage/premier
    etc.) — those are usually the broader / primary extract for a CU
    that runs more than one loan system.

    Files with no period sort to the end. Returns None when *entries* is empty.
    """
    if not entries:
        return None

    def _is_subsystem(name: str) -> bool:
        # These keywords typically denote a partial / subsystem extract.
        return bool(
            re.search(
                r"(cuma|mortgage|premier|redicash|line[\s_\-]*of[\s_\-]*credit|"
                r"share[\s_\-]*secured)",
                name or "",
                re.IGNORECASE,
            )
        )

    if snapshot_yyyymm:
        matches = [e for e in entries if _safe_period(e) == snapshot_yyyymm]
        if matches:
            # Prefer non-subsystem files first.
            primary = [e for e in matches if not _is_subsystem(e.get("name", ""))]
            if primary:
                return primary[0]
            return matches[0]
    # Fall back to most-recent period (entries are NOT guaranteed sorted).
    sortable = sorted(
        entries,
        key=lambda e: (
            _safe_period(e) or "0000-00",
            0 if _is_subsystem(e.get("name", "")) else 1,  # non-subsystem wins ties
        ),
        reverse=True,
    )
    return sortable[0] if sortable else None


# --------------------------------------------------------------------------
# Delivered code -> pool map workbook (Loan Type / Collateral Code -> Pool)
# --------------------------------------------------------------------------

# Filename hints for a delivered "loan code -> pool" reference workbook
# (e.g. "<CU> Loan Type Codes and Collateral Codes with Pools.xlsx"). These
# are NOT loan extracts; they carry the analyst's intended code->pool mapping.
_CODE_MAP_RX = re.compile(
    r"(?:loan[\s_\-]*type[\s_\-]*codes?"
    r"|collateral[\s_\-]*codes?"
    r"|codes?[\s_\-]*(?:and|&|with|to)[\s_\-]*(?:collateral|pools?)"
    r"|code[\s_\-]*map(?:ping)?)",
    re.IGNORECASE,
)


def _looks_like_code_map(name: str) -> bool:
    n = name or ""
    if not n.lower().endswith((".xlsx", ".xls")):
        return False
    nl = n.lower()
    # Require a pool/code-map signal so we don't grab transaction files that
    # merely mention "Loan Type Code" (e.g. "CO & Recs by Loan Type Code
    # 02282026.xlsx" — a charge-off/recovery file, NOT a code->pool map).
    if not (re.search(r"pools?\b", nl)
            or re.search(r"code[\s_\-]*map(?:ping)?", nl)):
        return False
    return bool(_CODE_MAP_RX.search(n))


def _norm_hdr(h: Any) -> str:
    return re.sub(r"\s+", " ", str(h).strip()).lower() if h is not None else ""


def _parse_code_map_file(
    path: str, known_codes: "list[str] | set[str] | None" = None
) -> dict[str, str]:
    """Parse a delivered code->pool workbook -> ``{code(str): pool(str)}``.

    Robust to header naming: the POOL column is the one whose header contains
    ``pool`` (preferring an exact ``pool`` over a coarser ``group``); the CODE
    column is chosen by best overlap of its values with *known_codes* (the
    codes already seeded on the loan sample), falling back to a ``code``-named
    header. Across sheets/columns the parse that maps the most known codes
    (tiebreak: most distinct pools = the finer, reviewed granularity) wins.
    Best-effort: returns ``{}`` on any failure.
    """
    try:
        import openpyxl
    except Exception:  # noqa: BLE001
        return {}
    known = {str(k).strip() for k in (known_codes or []) if str(k).strip()}

    def _score(m: dict[str, str]) -> tuple[int, int]:
        mk = len(set(m) & known) if known else len(m)
        return (mk, len({v for v in m.values()}))

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:  # noqa: BLE001
        return {}
    best: dict[str, str] = {}
    try:
        for sh in wb.sheetnames:
            rows = [
                r for r in wb[sh].iter_rows(values_only=True)
                if any(c is not None for c in r)
            ]
            if len(rows) < 2:
                continue
            hdr = rows[0]
            pool_idx = None
            for i, h in enumerate(hdr):
                if _norm_hdr(h) in ("pool", "pool name", "cecl pool"):
                    pool_idx = i
                    break
            if pool_idx is None:
                for i, h in enumerate(hdr):
                    if "pool" in _norm_hdr(h):
                        pool_idx = i
                        break
            if pool_idx is None:
                continue
            code_idx = None
            if known:
                best_ov = 0
                for i in range(len(hdr)):
                    if i == pool_idx:
                        continue
                    vals = {
                        str(r[i]).strip()
                        for r in rows[1:] if i < len(r) and r[i] is not None
                    }
                    ov = len(vals & known)
                    if ov > best_ov:
                        best_ov = ov
                        code_idx = i
            if code_idx is None:
                for i, h in enumerate(hdr):
                    hn = _norm_hdr(h)
                    if ("code" in hn and "lookup" not in hn
                            and "key" not in hn and i != pool_idx):
                        code_idx = i
                        break
            if code_idx is None:
                continue
            cmap: dict[str, str] = {}
            for r in rows[1:]:
                if code_idx >= len(r) or pool_idx >= len(r):
                    continue
                c, p = r[code_idx], r[pool_idx]
                if c is None or p is None:
                    continue
                c, p = str(c).strip(), str(p).strip()
                if c and p and c.lower() != "nan" and p.lower() != "nan":
                    cmap.setdefault(c, p)
            if _score(cmap) > _score(best):
                best = cmap
    finally:
        wb.close()
    return best


def _apply_code_map_to_pool_map(
    state: dict[str, Any], code_map_files: list[dict[str, Any]]
) -> list[str]:
    """Fill ``pool_map`` VALUES from a delivered code->pool workbook.

    Only proposes: fills codes currently blank / ``Ignore`` and registers the
    proposed pool names into ``pool_settings``. The delivered map is a strong
    PROPOSAL (not ground truth — it can carry errors), so the caller keeps the
    pools step flagged for human review.
    """
    pm = state.get("pool_map") or {}
    if not pm or not code_map_files:
        return []
    known = list(pm.keys())
    best_map: dict[str, str] = {}
    best_src = ""

    def _score(m: dict[str, str]) -> tuple[int, int]:
        return (len(set(m) & set(known)), len({v for v in m.values()}))

    for entry in code_map_files:
        path = entry.get("path") or entry.get("saved_path")
        if not path:
            continue
        m = _parse_code_map_file(path, known)
        if _score(m) > _score(best_map):
            best_map = m
            best_src = entry.get("name") or str(path)
    if not best_map:
        return []

    filled = 0
    proposed_pools: set[str] = set()
    for code in list(pm.keys()):
        cur = pm.get(code)
        if (not cur or str(cur).strip().lower() == "ignore") and code in best_map:
            pool = best_map[code]
            pm[code] = pool
            proposed_pools.add(pool)
            filled += 1
    if not filled:
        return []

    existing = {
        (p.get("name") or "").strip().lower()
        for p in (state.get("pool_settings") or [])
    }
    ps = state.setdefault("pool_settings", [])
    added = 0
    for pool in sorted(proposed_pools):
        key = pool.strip().lower()
        if key and key != "ignore" and key not in existing:
            ps.append({
                "name": pool,
                "risk_rated": False,
                "brr": False,
                "acl_months": _default_acl_months_for_pool(pool),
                "use_default_mgmt_adj": False,
                "excluded": False,
            })
            existing.add(key)
            added += 1
    state["_code_map_proposed_count"] = filled
    return [
        f"pool_map: pre-filled {filled} code(s) from delivered code map "
        f"'{best_src}' ({len(proposed_pools)} distinct pool(s); {added} new "
        "pool(s) added to Loan Pools) — REVIEW these proposals before saving"
    ]


# --------------------------------------------------------------------------
# Findings runner
# --------------------------------------------------------------------------


def scan_folder_for_setup(
    folder: str | Path,
    snapshot_yyyymm: str | None = None,
    _prebuilt_classification: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Scan *folder* and produce a ``findings`` dict for state population.

    Calls the existing per-file parsers (sample_parser, monthly_bal_parser)
    on the most-representative file of each type. Does not mutate any
    state; use ``apply_findings_to_state`` for that.

    Returns::

        {
            "ok": bool,
            "error": str | None,
            "folder": str,
            "classification": <classify_folder result>,
            "sample": <sample_parser.analyse_sample_file result> | None,
            "annual_bal": <monthly_bal_parser.analyse_per_year_file result> | None,
            "monthly_bal": <monthly_bal_parser.analyse_per_month_file result> | None,
            "warm": dict | None,
            "impaired_file": dict | None,   # entry that will be copied to Raw_Uploads
            "credit_pull_file": dict | None,
            "co_file": dict | None,
            "recov_file": dict | None,
            "messages": list[str],            # human-readable narration
            "errors":   list[str],
        }
    """
    folder = Path(folder)
    out: dict[str, Any] = {
        "ok": False,
        "error": None,
        "folder": str(folder),
        "classification": None,
        "sample": None,
        "sample_extracts": [],
        "annual_bal": None,
        "monthly_bal": None,
        "single_hist_bal": None,
        "warm": None,
        "impaired_file": None,
        "credit_pull_file": None,
        "co_file": None,
        "recov_file": None,
        "pool_seed": None,
        "code_map_files": [],
        "messages": [],
        "errors": [],
    }

    cls = _prebuilt_classification or classify_folder(folder)
    out["classification"] = cls
    if not cls.get("ok"):
        out["error"] = cls.get("error") or "folder scan failed"
        return out

    out["ok"] = True
    out["messages"].append(
        f"Classified {len(cls.get('loan_data_files') or [])} loan extract(s), "
        f"{len(cls.get('annual_balance_files') or [])} annual balance sheet(s), "
        f"{len(cls.get('monthly_files') or [])} monthly balance file(s), "
        f"{len(cls.get('impaired_files') or [])} impaired-loan file(s), "
        f"{len(cls.get('credit_pull_files') or [])} credit-pull file(s), "
        f"{len(cls.get('co_files') or [])} CO file(s), "
        f"{len(cls.get('recov_files') or [])} recovery file(s), "
        f"{len(cls.get('warm_files') or [])} WARM file(s)."
    )

    # ----- Loan extract sample (drives column_mappings + pool_codes) -----
    loan_files = cls.get("loan_data_files") or []
    sample_entry = _pick_latest_for_period(loan_files, snapshot_yyyymm)
    sample_candidates: list[dict[str, Any]] = []
    if sample_entry:
        sample_candidates.append(sample_entry)
        # If multiple files share the snapshot period, queue them so
        # multi-extract CUs (e.g. AIRES + CUMA) stage all their
        # snapshot extracts on Step 13 and the best-scoring one
        # still drives the top-level state.
        if snapshot_yyyymm:
            for e in loan_files:
                if e is sample_entry:
                    continue
                if _safe_period(e) == snapshot_yyyymm:
                    sample_candidates.append(e)
        else:
            # No snapshot pinned -- group remaining files by file_pattern
            # shape and add the latest from each shape distinct from the
            # primary so multi-extract CUs without an anchor still stage
            # all their unique extracts.
            try:
                primary_pat = sample_parser.guess_filename_patterns(
                    sample_entry["name"]
                ).get("file_pattern", "") or ""
            except Exception:  # noqa: BLE001
                primary_pat = ""
            seen_patterns: set[str] = {primary_pat} if primary_pat else set()
            remaining = sorted(
                (e for e in loan_files if e is not sample_entry),
                key=lambda e: _safe_period(e) or "0000-00",
                reverse=True,
            )
            for e in remaining:
                try:
                    pat = sample_parser.guess_filename_patterns(
                        e["name"]
                    ).get("file_pattern", "") or ""
                except Exception:  # noqa: BLE001
                    pat = ""
                if pat and pat not in seen_patterns:
                    seen_patterns.add(pat)
                    sample_candidates.append(e)

    best_analysis: dict[str, Any] | None = None
    best_entry: dict[str, Any] | None = None
    best_score = -1
    extracts: list[dict[str, Any]] = []
    for cand in sample_candidates:
        src = Path(cand["path"])
        try:
            tmp = _copy_into_tmp(src, "sample")
        except Exception as exc:  # noqa: BLE001
            out["errors"].append(f"copy failed for {cand['name']}: {exc}")
            continue
        # Phase 9.26: detect "no-split" CUMA/mortgage extracts BEFORE
        # analyse_sample_file runs, so the seeded pool_code_suggestions
        # carry the full composite labels (e.g. ``15/15 ARM``) instead of
        # being silently truncated to ``"15"`` by the default ``"/"`` split.
        # First-pass uses filename-only; if that misses we re-check the
        # actual sample values after analysis and re-run if needed.
        no_split = _looks_like_no_split_extract(cand["name"])
        try:
            analysis = sample_parser.analyse_sample_file(
                str(tmp), cand["name"],
                split_char=("" if no_split else "/"),
            )
        except Exception as exc:  # noqa: BLE001
            out["errors"].append(
                f"sample_parser failed on {cand['name']}: {exc}"
            )
            continue
        if not analysis.get("ok"):
            out["errors"].append(
                f"sample analysis returned not-ok for "
                f"{cand['name']}: {analysis.get('error')}"
            )
            continue
        # Second-pass detection on raw sample-rows values (catches CUs
        # that ship CUMA-style codes in generically-named workbooks).
        if not no_split:
            pool_col = (analysis.get("column_suggestions") or {}).get(
                "loan_pool_code"
            )
            sample_rows = analysis.get("sample_rows") or []
            if pool_col and sample_rows:
                raw_vals = []
                for row in sample_rows:
                    if isinstance(row, dict):
                        v = row.get(pool_col)
                        if v is not None:
                            raw_vals.append(str(v))
                if _looks_like_no_split_extract(cand["name"], raw_vals):
                    no_split = True
                    try:
                        analysis = sample_parser.analyse_sample_file(
                            str(tmp), cand["name"], split_char="",
                        )
                    except Exception:  # noqa: BLE001
                        pass
        analysis["saved_path"] = str(tmp)
        ex_entry = {
            "name": cand["name"],
            "saved_path": str(tmp),
            "analysis": analysis,
        }
        if no_split:
            ex_entry["pool_code_split"] = ""
        extracts.append(ex_entry)
        # Score = #column mappings + #pool codes. Higher = better
        # foothold for the wizard.
        score = (
            len(analysis.get("column_suggestions") or {})
            + len(analysis.get("pool_code_suggestions") or [])
        )
        if score > best_score:
            best_score = score
            best_analysis = analysis
            best_entry = cand

    if best_analysis and best_entry:
        out["sample"] = best_analysis
        out["messages"].append(
            f"Analysed sample loan extract: {best_entry['name']} "
            f"({len(best_analysis.get('headers') or [])} columns, "
            f"{len(best_analysis.get('pool_code_suggestions') or [])} pool codes)"
        )

    out["sample_extracts"] = extracts
    if len(extracts) > 1:
        names = ", ".join(e["name"] for e in extracts)
        out["messages"].append(
            f"Staged {len(extracts)} loan-extract sample(s) for Step 13: {names}"
        )

    # ----- Annual balance-sheet workbooks (drives monthly_bal per_year) ---
    annual_entries = sorted(
        cls.get("annual_balance_files") or [],
        key=lambda e: e.get("year") or 0,
        reverse=True,
    )
    if annual_entries:
        # Analyse the most-recent one to capture the layout (sheet, cols, etc).
        recent = annual_entries[0]
        try:
            an = monthly_bal_parser.analyse_per_year_file(recent["path"])
        except Exception as exc:  # noqa: BLE001
            out["errors"].append(
                f"monthly_bal_parser.analyse_per_year_file failed on "
                f"{recent['name']}: {exc}"
            )
        else:
            if an.get("ok"):
                out["annual_bal"] = {
                    "layout": {
                        "sheet": an.get("sheet"),
                        "header_row": an.get("header_row"),
                        "label_col": an.get("label_col"),
                        "period_columns": an.get("period_columns") or [],
                    },
                    "pool_labels": an.get("pool_labels") or [],
                    "files": annual_entries,
                }
                out["messages"].append(
                    f"Analysed annual balance-sheet layout from "
                    f"{recent['name']}: sheet={an.get('sheet')}, "
                    f"{len(an.get('pool_labels') or [])} pool labels."
                )

    # ----- Per-month detailed balance sheet (drives monthly_bal per_month) -
    if not out["annual_bal"] and cls.get("monthly_detail_balance_files"):
        # Pick the one matching snapshot if possible.
        m_entries = cls.get("monthly_detail_balance_files") or []
        recent = _pick_latest_for_period(m_entries, snapshot_yyyymm)
        if recent:
            try:
                an = monthly_bal_parser.analyse_per_month_file(recent["path"])
            except Exception as exc:  # noqa: BLE001
                out["errors"].append(
                    f"monthly_bal_parser.analyse_per_month_file failed on "
                    f"{recent['name']}: {exc}"
                )
            else:
                if an.get("ok"):
                    out["monthly_bal"] = {
                        "layout": {
                            "sheet": an.get("sheet"),
                            "header_row": an.get("header_row"),
                            "label_col": an.get("label_col"),
                            "balance_col": an.get("balance_col"),
                            "as_of_cell": an.get("as_of_cell"),
                            "stop_row": an.get("stop_row"),
                        },
                        "pool_labels": an.get("pool_labels") or [],
                        "files": m_entries,
                        "representative": recent,
                    }
                    out["messages"].append(
                        f"Analysed per-month balance layout from "
                        f"{recent['name']}."
                    )

    # ----- Consolidated pool-grouping balance workbook ------------------
    # (Vizo Financial convention: ``Balance Sheets <CU>.xlsx`` carries the
    # authoritative pool list AND the sub-category -> pool mapping.)
    cpb_entries = cls.get("consolidated_pool_balance_files") or []
    if cpb_entries:
        recent = cpb_entries[0]
        try:
            seed = extract_pools_from_balance_workbook(recent["path"])
        except Exception as exc:  # noqa: BLE001
            out["errors"].append(
                f"extract_pools_from_balance_workbook failed on "
                f"{recent['name']}: {exc}"
            )
        else:
            if seed.get("ok") and seed.get("pools"):
                seed["source_file"] = recent.get("name")
                seed["source_path"] = recent.get("path")
                out["pool_seed"] = seed
                out["messages"].append(
                    f"Extracted {len(seed['pools'])} pool(s) and "
                    f"{len(seed.get('subcategory_to_pool') or {})} "
                    f"sub-category mapping(s) from {recent['name']}"
                )
            elif not seed.get("ok"):
                out["errors"].append(
                    f"pool extraction returned not-ok for "
                    f"{recent['name']}: {seed.get('error')}"
                )

    # ----- Single consolidated historical balance workbook --------------
    # (Used when neither annual nor monthly-detail workbooks were found.
    # File names like "Historical Loan Balances.xlsx" or
    # "Balance_History_-_<CU>.xlsx" -- a single wide-format workbook with
    # one column per month-end.)
    if (
        not out["annual_bal"]
        and not out["monthly_bal"]
        and cls.get("single_hist_bal_files")
    ):
        shb_entries = cls.get("single_hist_bal_files") or []
        # Prefer the largest/newest file if multiple exist.
        recent = shb_entries[0]
        try:
            an = monthly_bal_parser.analyse_file(recent["path"])
        except Exception as exc:  # noqa: BLE001
            out["errors"].append(
                f"monthly_bal_parser.analyse_file failed on "
                f"{recent['name']}: {exc}"
            )
        else:
            if an.get("ok"):
                out["single_hist_bal"] = {
                    "entry": recent,
                    "sheet": an.get("sheet"),
                    "header_row": an.get("header_row"),
                    "pool_name_col": an.get("pool_name_col"),
                    "pool_labels": an.get("parsed_pool_labels") or [],
                    "dates": an.get("dates") or [],
                }
                out["messages"].append(
                    f"Analysed single historical balance workbook "
                    f"{recent['name']}: sheet={an.get('sheet')}, "
                    f"{len(an.get('parsed_pool_labels') or [])} pool labels, "
                    f"{len(an.get('dates') or [])} date columns."
                )
            else:
                out["errors"].append(
                    f"analyse_file failed on {recent['name']}: "
                    f"{an.get('error')}"
                )

    # ----- Fallback: consolidated pool-grouping workbook as hist source -
    # Many CUs (especially Vizo Financial clients) ship a single
    # ``Balance Sheets <CU>.xlsx`` workbook that is BOTH the pool-seed
    # AND the historical balance time-series (loan codes in col A,
    # category names in col B, one column per month-end in cols C+).
    # When no annual / per-month / single-hist workbook was found but
    # we DID find a consolidated pool workbook, reuse it as the single
    # historical balance source.
    if (
        not out["annual_bal"]
        and not out["monthly_bal"]
        and not out["single_hist_bal"]
        and cls.get("consolidated_pool_balance_files")
    ):
        cpb = (cls.get("consolidated_pool_balance_files") or [])[0]
        try:
            an = monthly_bal_parser.analyse_file(cpb["path"])
        except Exception as exc:  # noqa: BLE001
            out["errors"].append(
                f"monthly_bal_parser.analyse_file (consolidated fallback) "
                f"failed on {cpb['name']}: {exc}"
            )
        else:
            if an.get("ok") and (an.get("dates") or []):
                # ``analyse_file`` always picks col A as the label column,
                # but the Vizo consolidated workbooks put loan codes in
                # col A and descriptive labels in col B. Pick whichever
                # column overlaps best with the pool-seed sub-category
                # mapping we extracted from the same workbook.
                pool_name_col = an.get("pool_name_col") or "A"
                seed = out.get("pool_seed") or {}
                sub_to_pool = (seed.get("subcategory_to_pool") or {})
                if sub_to_pool:
                    sub_keys = {
                        str(k).strip().casefold()
                        for k in sub_to_pool
                        if str(k).strip()
                    }
                    try:
                        from openpyxl import load_workbook as _lw
                        _wb = _lw(cpb["path"], read_only=True, data_only=True)
                        try:
                            _ws = _wb[an.get("sheet")]
                            best_col = pool_name_col
                            best_hits = -1
                            for col_letter in ("A", "B", "C"):
                                col_idx = ord(col_letter) - ord("A") + 1
                                hits = 0
                                for row in _ws.iter_rows(
                                    min_row=int(an.get("header_row") or 1) + 1,
                                    max_row=int(an.get("header_row") or 1) + 80,
                                    min_col=col_idx,
                                    max_col=col_idx,
                                    values_only=True,
                                ):
                                    val = row[0]
                                    if val is None:
                                        continue
                                    if (
                                        str(val).strip().casefold()
                                        in sub_keys
                                    ):
                                        hits += 1
                                if hits > best_hits:
                                    best_hits = hits
                                    best_col = col_letter
                            if best_hits > 0:
                                pool_name_col = best_col
                        finally:
                            _wb.close()
                    except Exception:  # noqa: BLE001
                        pass
                # Re-read parsed pool labels from the elected column so
                # the wizard's Step 4 displays the descriptive names that
                # match the user's pool map (rather than loan codes).
                pool_labels = an.get("parsed_pool_labels") or []
                if pool_name_col != (an.get("pool_name_col") or "A"):
                    try:
                        from openpyxl import load_workbook as _lw
                        _wb = _lw(cpb["path"], read_only=True, data_only=True)
                        try:
                            _ws = _wb[an.get("sheet")]
                            col_idx = (
                                ord(pool_name_col.upper()) - ord("A") + 1
                            )
                            new_labels = []
                            for row in _ws.iter_rows(
                                min_row=int(an.get("header_row") or 1) + 1,
                                max_row=int(an.get("header_row") or 1) + 200,
                                min_col=col_idx,
                                max_col=col_idx,
                                values_only=True,
                            ):
                                val = row[0]
                                if val is None:
                                    continue
                                txt = str(val).strip()
                                if txt and txt not in new_labels:
                                    new_labels.append(txt)
                            if new_labels:
                                pool_labels = new_labels
                        finally:
                            _wb.close()
                    except Exception:  # noqa: BLE001
                        pass
                out["single_hist_bal"] = {
                    "entry": cpb,
                    "sheet": an.get("sheet"),
                    "header_row": an.get("header_row"),
                    "pool_name_col": pool_name_col,
                    "pool_labels": pool_labels,
                    "dates": an.get("dates") or [],
                }
                out["messages"].append(
                    f"Using consolidated balance workbook {cpb['name']!r} "
                    f"as single historical balance source "
                    f"(sheet={an.get('sheet')}, pool col={pool_name_col}, "
                    f"{len(pool_labels)} labels, "
                    f"{len(an.get('dates') or [])} date columns)."
                )

    # ----- WARM (if any) -------------------------------------------------
    warm_entry = _pick_latest_for_period(
        cls.get("warm_files") or [], snapshot_yyyymm
    )
    if warm_entry:
        src = Path(warm_entry["path"])
        tmp = _copy_into_tmp(src, "warm")
        try:
            analysis = warm_parser.analyse_warm_file(tmp, warm_entry["name"])
        except Exception as exc:  # noqa: BLE001
            out["errors"].append(
                f"warm_parser failed on {warm_entry['name']}: {exc}"
            )
        else:
            out["warm"] = {
                "entry": warm_entry,
                "saved_path": str(tmp),
                "analysis": analysis,
            }
            out["messages"].append(
                f"Found WARM workbook for {warm_entry.get('period')}: "
                f"{warm_entry['name']}"
            )

    # ----- Companion files: impaired, credit pull, CO, recov -------------
    out["impaired_file"] = _pick_latest_for_period(
        cls.get("impaired_files") or [], snapshot_yyyymm
    )
    out["credit_pull_file"] = _pick_latest_for_period(
        cls.get("credit_pull_files") or [], snapshot_yyyymm
    )
    out["co_file"] = _pick_latest_for_period(
        cls.get("co_files") or [], snapshot_yyyymm
    )
    out["recov_file"] = _pick_latest_for_period(
        cls.get("recov_files") or [], snapshot_yyyymm
    )

    # ----- Delivered code -> pool map workbook(s) ------------------------
    # Reference workbooks (not loan extracts) carrying the analyst's intended
    # loan-code -> pool mapping; land in ``other_files`` from classification.
    code_map_files = [
        e for e in (cls.get("other_files") or [])
        if _looks_like_code_map(e.get("name") or "")
    ]
    out["code_map_files"] = code_map_files
    if code_map_files:
        out["messages"].append(
            f"Detected {len(code_map_files)} loan-code map workbook(s) "
            "(code -> pool proposals for the Loan Code Mapping step)"
        )

    return out


# Bucket keys produced by ``classify_folder`` that hold lists of file entries.
# Used to merge classifications across multiple scanned folders.
_CLASSIFICATION_BUCKETS: tuple[str, ...] = (
    "loan_data_files",
    "warm_files",
    "co_files",
    "recov_files",
    "impaired_files",
    "credit_pull_files",
    "monthly_files",
    "annual_balance_files",
    "monthly_detail_balance_files",
    "single_hist_bal_files",
    "consolidated_pool_balance_files",
    "five_thirtythousand_files",
    "other_files",
)

# Display names for the optional per-folder type hints the user can set on
# the Identity step. Purely cosmetic (used in scan narration); the
# classifier still decides file types by content/name.
_FOLDER_LABEL_NAMES: dict[str, str] = {
    "warm": "WARM history",
    "data": "Client data / loan extracts",
    "credit_pull": "Credit pull",
    "co_recov": "Charge-offs / Recoveries",
    "impaired": "Impaired loans",
    "balances": "Balance sheets",
}


def _merge_classifications(
    parts: list[tuple[str, dict[str, Any]]]
) -> dict[str, Any]:
    """Union the file buckets from several ``classify_folder`` results.

    ``parts`` is a list of ``(folder, classification)`` tuples. Entries are
    de-duplicated by absolute path (falling back to name) so a file that
    appears under two overlapping folders is only classified once. Each
    surviving entry is tagged with ``_source_folder`` for later narration.
    Returns a single classification dict with ``ok=True`` when at least one
    part classified successfully.
    """
    merged: dict[str, Any] = {"ok": False, "error": None}
    for key in _CLASSIFICATION_BUCKETS:
        merged[key] = []
    seen: set[str] = set()
    errors: list[str] = []

    for folder, cls in parts:
        if not cls.get("ok"):
            errors.append(f"{folder}: {cls.get('error') or 'scan failed'}")
            continue
        merged["ok"] = True
        for key in _CLASSIFICATION_BUCKETS:
            for entry in cls.get(key) or []:
                ident = str(
                    entry.get("path") or entry.get("name") or id(entry)
                ).lower()
                if ident in seen:
                    continue
                seen.add(ident)
                entry.setdefault("_source_folder", folder)
                merged[key].append(entry)

    if not merged["ok"] and errors:
        merged["error"] = "; ".join(errors)
    return merged


def scan_folders_for_setup(
    folders: list[str | Path],
    snapshot_yyyymm: str | None = None,
    folder_labels: list[str] | None = None,
) -> dict[str, Any]:
    """Scan several folders and produce a single merged ``findings`` dict.

    This is the multi-folder entry point behind the wizard's "add folder
    paths and let the tool decide" flow. The user can point at, e.g., a
    WARM-history folder, a client-data-upload folder, and a credit-pull
    folder; every folder is classified, the buckets are unioned, and the
    same downstream logic as :func:`scan_folder_for_setup` runs against the
    combined set — so the wizard auto-derives column mappings, pools, file
    patterns, balance layout, companion files, and (crucially) whether a
    WARM workbook is available for the WARM -> CM comparison.

    ``folder_labels`` is an OPTIONAL list, parallel to ``folders``, of
    user-supplied type hints ("warm", "data", "credit_pull", ...). They are
    purely informational — the classifier still decides file types by
    content/name — but they are echoed in the per-folder narration so the
    scan report reflects the user's intent.

    Returns the same shape as :func:`scan_folder_for_setup` plus a
    ``"folders"`` list. ``"folder"`` is set to the most representative folder
    (the one contributing loan extracts, else the first) so single-folder
    self-heal reuse keeps working.
    """
    norm = [str(f).strip() for f in (folders or []) if str(f).strip()]
    if not norm:
        return {
            "ok": False,
            "error": "no folder paths provided",
            "folder": "",
            "folders": [],
            "classification": None,
            "messages": [],
            "errors": [],
        }

    # Align optional labels to the (blank-filtered) folder list.
    labels_in = list(folder_labels or [])
    label_by_folder: dict[str, str] = {}
    _li = 0
    for f in (folders or []):
        fs = str(f).strip()
        lbl = labels_in[_li] if _li < len(labels_in) else ""
        _li += 1
        if fs:
            label_by_folder[fs] = (lbl or "").strip()

    parts: list[tuple[str, dict[str, Any]]] = []
    for f in norm:
        try:
            parts.append((f, classify_folder(f)))
        except Exception as exc:  # noqa: BLE001 - one bad folder shouldn't abort
            parts.append((f, {"ok": False, "error": f"{type(exc).__name__}: {exc}"}))

    merged_cls = _merge_classifications(parts)

    # Pick the most representative folder for self-heal reuse: prefer the one
    # that contributed loan extracts, else the first provided folder.
    primary = norm[0]
    for entry in merged_cls.get("loan_data_files") or []:
        sf = entry.get("_source_folder")
        if sf:
            primary = sf
            break

    findings = scan_folder_for_setup(
        primary, snapshot_yyyymm, _prebuilt_classification=merged_cls
    )
    findings["folders"] = norm
    findings["folder_labels"] = {
        f: label_by_folder.get(f, "") for f in norm
    }
    # Prepend a per-folder narration line so the scan report shows what each
    # path contributed (with the user's optional type hint, if any).
    if findings.get("ok"):
        per_folder = []
        for f, cls in parts:
            hint = label_by_folder.get(f, "")
            tag = f" [{_FOLDER_LABEL_NAMES.get(hint, hint)}]" if hint else ""
            if cls.get("ok"):
                per_folder.append(
                    f"{f}{tag}: {len(cls.get('loan_data_files') or [])} extract(s), "
                    f"{len(cls.get('warm_files') or [])} WARM, "
                    f"{len(cls.get('credit_pull_files') or [])} credit-pull, "
                    f"{len(cls.get('impaired_files') or [])} impaired."
                )
            else:
                per_folder.append(
                    f"{f}{tag}: {cls.get('error') or 'scan failed'}"
                )
        findings.setdefault("messages", [])
        findings["messages"] = per_folder + findings["messages"]
    return findings


# --------------------------------------------------------------------------
# State population
# --------------------------------------------------------------------------


# Heuristic: when the folder path contains "vizo financial" we know the CU
# is hosted under Vizo Financial Corporate CU's analyst tree. That client
# convention turns OFF TCT reports and turns ON Vizo + Vizo Supplemental
# + ImpDet. Pure-TCT setup is the safer fallback for any other CU.
_VIZO_FOLDER_HINT = re.compile(r"vizo[\s_\-]*financial", re.IGNORECASE)


def _set_reports_default_from_folder(state: dict[str, Any], folder: str) -> None:
    if _VIZO_FOLDER_HINT.search(folder or ""):
        state["reports"] = {
            "tct": False,
            "vizo": True,
            "vizo_supp": True,
            "impdet": True,
        }
    # else leave whatever default was already in state (TCT-only by default).


def _apply_mgmt_adj_defaults(state: dict[str, Any]) -> None:
    """Apply the Census-learned analyst-standard mgmt_adj baselines.

    Only writes when the current values are still the in-code default.
    """
    cur = state.get("mgmt_adj") or {}
    # In-code default: ltv_baseline=0.9, probability_factor=0.35. If the
    # state already matches those, that's because nothing's been edited
    # yet — re-writing is a no-op. We still write them explicitly so the
    # auto-fill report can show "mgmt_adj baselines applied".
    cur.setdefault("ltv_baseline", 0.9)
    cur.setdefault("probability_factor", 0.35)
    state["mgmt_adj"] = cur


def _apply_sample_to_state(
    state: dict[str, Any],
    sample: dict[str, Any],
    representative_name: str,
) -> list[str]:
    """Mutate state with the sample-file analysis. Returns step-level messages."""
    msgs: list[str] = []

    # state["sample"] is what Step 2 Sample / Step 12 Sample (no-WARM)
    # reads in subsequent renders.
    state["sample"] = {
        "filename": representative_name,
        "saved_path": sample.get("saved_path"),
        "has_header": bool(sample.get("has_header")),
        "header_row": sample.get("header_row"),
        "headers": sample.get("headers") or [],
        "sample_rows": sample.get("sample_rows") or [],
        "column_suggestions": sample.get("column_suggestions") or {},
        "pool_code_suggestions": sample.get("pool_code_suggestions") or [],
        "file_pattern": sample.get("file_pattern"),
        "date_pattern": sample.get("date_pattern"),
        "date_format": sample.get("date_format"),
    }
    msgs.append(
        f"sample analysis loaded ({len(state['sample']['headers'])} columns, "
        f"{len(state['sample']['pool_code_suggestions'])} pool codes)"
    )

    # Column mappings: apply suggestions over any blank-or-default cells only.
    cm = state.setdefault("column_mappings", {})
    suggestions = sample.get("column_suggestions") or {}
    filled = []
    for field, header in suggestions.items():
        if not header:
            continue
        cur = cm.get(field, "")
        if not cur or cur == "MEMBER_ID" or cur == "BALANCE" or cur == "FICO_SCORE" \
                or cur == "LOAN_TYPE" or cur == "DQ_DAYS" or cur == "INT_RATE" \
                or cur == "OPEN_DATE" or cur == "ORIG_AMT":
            # blank or in-code default placeholder -> overwrite
            cm[field] = header
            filled.append(field)
    if filled:
        msgs.append(f"column_mappings filled: {', '.join(sorted(filled))}")

    # Cross-CU learned suggestions — fill any field the sample parser
    # couldn't auto-detect, using the header most other credit unions have
    # mapped to that field (restricted to headers present in THIS sample).
    # This is the SAME learned store the interactive Columns step consults
    # (routes.setup._apply_sample_to_state); applying it during the
    # auto-scan means cryptic core headers keyword-matching can't infer
    # (e.g. 'col_cd' -> loan_pool_code, 'User_Defined_Field_7' ->
    # original_fico_score) are filled up front, shrinking the manual step.
    # Never overrides a real (non-placeholder) mapping.
    _placeholder_cols = {
        "MEMBER_ID", "BALANCE", "FICO_SCORE", "LOAN_TYPE",
        "DQ_DAYS", "INT_RATE", "OPEN_DATE", "ORIG_AMT",
    }
    headers_now = sample.get("headers") or []
    if headers_now:
        already_real = {
            f for f, v in cm.items()
            if v and v not in _placeholder_cols
        }
        try:
            learned = column_mapping_suggestions.suggest_for_headers(
                headers_now, skip_fields=already_real
            )
        except Exception:  # noqa: BLE001
            learned = {}
        learned_filled = []
        for field, header in learned.items():
            cur = cm.get(field, "")
            if not cur or cur in _placeholder_cols:
                cm[field] = header
                learned_filled.append(field)
        if learned_filled:
            msgs.append(
                "column_mappings filled from cross-CU history: "
                + ", ".join(sorted(learned_filled))
            )

    # File pattern + date pattern (drives importer file discovery).
    if sample.get("file_pattern"):
        state["file_pattern"] = sample["file_pattern"]
        msgs.append(f"file_pattern set to {sample['file_pattern']}")
    if sample.get("date_pattern"):
        state["date_pattern"] = sample["date_pattern"]
        msgs.append(f"date_pattern set to {sample['date_pattern']}")
    # ``date_format`` tells import_data.extract_snapshot_date how to read the
    # date_pattern capture groups. It MUST travel with date_pattern — a
    # month-first name like "03-2026" yields date_pattern (\d{2})-(20\d{2})
    # whose format is MMYYYY, not the YYYY-MM default. Dropping it here (the
    # original bug) left every MM-YYYY / MMDDYY / YYYYMMDD CU silently
    # misparsing dates until a human corrected the Files step.
    if sample.get("date_format"):
        state["date_format"] = sample["date_format"]
        msgs.append(f"date_format set to {sample['date_format']}")
    # has_header
    state["has_header"] = bool(sample.get("has_header"))

    # Pool map seed: every detected code -> "Ignore" so user only has to
    # rename pools, not type the codes. We keep existing entries untouched
    # so re-running the scan doesn't clobber the user's renames. If we
    # have a durable history entry for a code (from a prior session on
    # this draft), restore that instead of the "Ignore" placeholder so a
    # Step 12 extract delete-and-re-add cycle doesn't force the user to
    # re-map codes they already assigned.
    pm = state.setdefault("pool_map", {})
    hist = state.get("_pool_map_history") or {}
    pool_codes = sample.get("pool_code_suggestions") or []
    seeded = 0
    for code in pool_codes:
        if not code:
            continue
        if code not in pm:
            pm[code] = hist.get(code) or "Ignore"
            seeded += 1
    if seeded:
        msgs.append(
            f"pool_map seeded with {seeded} code(s) (all set to 'Ignore' — "
            "rename on the Loan Code Mapping step)"
        )

    # Fallback pool-code seeding: when the sample parser detected NO pool
    # column (pool_code_suggestions empty -> pm still empty) but the learned
    # store just mapped loan_pool_code to a REAL header, read the distinct
    # codes straight from that column so the user only renames pools instead
    # of typing every code. Bounded + fully guarded; a distinct-count cap
    # avoids mis-seeding if the mapped column is not actually a code column.
    lpc_header = cm.get("loan_pool_code")
    if (not pm) and lpc_header and lpc_header not in _placeholder_cols \
            and sample.get("saved_path"):
        try:
            import pandas as pd
            # sample_parser reports header_row 1-based; pandas wants 0-based.
            # Try the most-likely offset first, then fall back, selecting
            # whichever read makes the mapped header appear as a column.
            _hr = sample.get("header_row")
            _cands = ([_hr - 1, _hr, 0] if isinstance(_hr, int) else [0])
            _df = None
            for _h in _cands:
                if _h is None or _h < 0:
                    continue
                try:
                    _tmp = pd.read_excel(sample["saved_path"], header=_h, dtype=str)
                except Exception:  # noqa: BLE001
                    continue
                if lpc_header in _tmp.columns:
                    _df = _tmp
                    break
            if _df is not None:
                _split = state.get("pool_code_split") or "/"
                _seen: list[str] = []
                _seenset: set[str] = set()
                for _v in _df[lpc_header].dropna().tolist():
                    _c = str(_v).strip()
                    if _split and _split in _c:
                        _c = _c.split(_split)[0].strip()
                    if _c and _c.lower() != "nan" and _c not in _seenset:
                        _seenset.add(_c)
                        _seen.append(_c)
                # Only seed when the distinct-code count is pool-code-plausible
                # (a real loan-type/collateral-code column has few distinct
                # values; hundreds means we mapped the wrong column).
                if _seen and len(_seen) <= 100:
                    for _code in _seen:
                        if _code not in pm:
                            pm[_code] = hist.get(_code) or "Ignore"
                    msgs.append(
                        f"pool_map seeded with {len(_seen)} code(s) read from "
                        f"the '{lpc_header}' column (all 'Ignore' — rename on "
                        "the Loan Code Mapping step)"
                    )
        except Exception:  # noqa: BLE001
            pass

    # Sensible split-char default (learned from Census FCU greenfield).
    if not state.get("pool_code_split"):
        state["pool_code_split"] = "/"

    return msgs


def _apply_annual_bal_to_state(
    state: dict[str, Any], annual: dict[str, Any]
) -> list[str]:
    """Mutate state.monthly_bal with annual-balance-sheet findings."""
    msgs: list[str] = []
    mb = state.setdefault("monthly_bal", {})
    mb["source"] = "per_year"
    mb["per_year_layout"] = annual.get("layout") or {}
    files = []
    for entry in annual.get("files") or []:
        files.append({
            "year": entry.get("year"),
            "filename": entry.get("name"),
            "saved_path": entry.get("path"),
        })
    mb["year_files"] = files
    mb["parsed_pool_labels"] = annual.get("pool_labels") or []
    # User still needs to map labels -> pool names on Step 5.
    msgs.append(
        f"monthly_bal source=per_year with {len(files)} workbook(s), "
        f"{len(mb['parsed_pool_labels'])} pool labels detected"
    )
    return msgs


def _apply_monthly_bal_to_state(
    state: dict[str, Any], mbal: dict[str, Any]
) -> list[str]:
    """Mutate state.monthly_bal with per-month-detail findings."""
    msgs: list[str] = []
    mb = state.setdefault("monthly_bal", {})
    mb["source"] = "per_month"
    mb["per_month_layout"] = mbal.get("layout") or {}
    files = []
    for entry in mbal.get("files") or []:
        files.append({
            "filename": entry.get("name"),
            "saved_path": entry.get("path"),
            "period": entry.get("period"),
        })
    mb["monthly_files"] = files
    mb["parsed_pool_labels"] = mbal.get("pool_labels") or []
    msgs.append(
        f"monthly_bal source=per_month with {len(files)} file(s), "
        f"{len(mb['parsed_pool_labels'])} pool labels detected"
    )
    return msgs


def _apply_single_hist_bal_to_state(
    state: dict[str, Any], shb: dict[str, Any]
) -> list[str]:
    """Mutate state.monthly_bal with single-historical-workbook findings.

    Sets the ``single_workbook`` layout fields (saved_path / sheet /
    header_row / pool_name_col) so the runtime balance-preview helper +
    Step 4's green confidence card can render immediately without the
    user re-uploading the file or re-saving the layout.
    """
    msgs: list[str] = []
    mb = state.setdefault("monthly_bal", {})
    mb["source"] = "single"
    entry = shb.get("entry") or {}
    saved_path = entry.get("path") or ""
    mb["saved_path"] = saved_path
    mb["filename"] = entry.get("name") or ""
    mb["sheet"] = shb.get("sheet") or ""
    mb["header_row"] = shb.get("header_row") or 0
    mb["pool_name_col"] = shb.get("pool_name_col") or ""
    mb["parsed_pool_labels"] = shb.get("pool_labels") or []
    msgs.append(
        f"monthly_bal source=single with workbook {entry.get('name')!r} "
        f"(sheet={shb.get('sheet')!r}, header row {shb.get('header_row')}, "
        f"pool label col {shb.get('pool_name_col')!r}, "
        f"{len(mb['parsed_pool_labels'])} pool labels, "
        f"{len(shb.get('dates') or [])} date columns)"
    )
    return msgs


# Pool-name keywords that imply a long life-of-loan (mortgages, HE, real
# estate). Used to pick a sensible ACL-months default when seeding pools
# from a 'Balance Sheets <CU>' workbook. Everything else defaults to 36.
_LONG_LOL_HINTS = re.compile(
    r"mortgage|first\s*lien|junior\s*lien|home\s*equity|heloc|"
    r"real\s*estate|residential|mobile\s*home|land\s*loan|"
    r"business\s*loan|commercial",
    re.IGNORECASE,
)

# Aggregate / total / grand-total row labels that should be excluded
# when seeding pool_settings from a flat hist-balance workbook (col A
# typically ends with a "Total" row that sums the pools above it).
_AGG_LABEL_RX = re.compile(
    r"^\s*(?:grand[\s_\-]*)?total(?:\s+loans?)?\s*$|"
    r"^\s*sub[\s_\-]*total\s*$|"
    r"^\s*all[\s_\-]*loans?\s*$",
    re.IGNORECASE,
)


def _default_acl_months_for_pool(name: str) -> int:
    """Sensible ACL-months default based on pool-name keywords."""
    if not name:
        return 36
    return 84 if _LONG_LOL_HINTS.search(name) else 36


def _purge_stale_pool_map_values(state: dict[str, Any]) -> int:
    """Rewrite any ``state["pool_map"]`` value that doesn't match a
    current ``pool_settings`` pool name (or the sentinels ``Ignore`` /
    ``Exclude``) to ``"Ignore"``.

    Defends against cross-CU pool-map pollution: pool_map entries from a
    previous CU's pool-map upload survive in the active Flask session
    after the user starts a new CU (no code path clears them on CU
    change). When auto-scan replaces ``pool_settings`` with the new
    CU's pools, leftover values become "unrecognized pool names" on
    Step 3. Run this AFTER ``pool_settings`` has been populated.

    Returns the number of entries rewritten.
    """
    pm = state.get("pool_map")
    if not isinstance(pm, dict) or not pm:
        return 0
    valid_names: set[str] = {
        (p.get("name") or "").strip()
        for p in (state.get("pool_settings") or [])
        if isinstance(p, dict) and p.get("name")
    }
    if not valid_names:
        # No pool_settings yet -> can't validate; do nothing.
        return 0
    valid_names |= {"Ignore", "Exclude"}
    purged = 0
    for code, value in list(pm.items()):
        if not value:
            continue
        if str(value).strip() in valid_names:
            continue
        pm[code] = "Ignore"
        purged += 1
    return purged


def _apply_flat_pool_labels_to_state(
    state: dict[str, Any],
    labels: list[str],
    src: str,
) -> dict[str, list[str]]:
    """Seed Step 2 (pool_settings) + Step 8 (monthly_bal.pool_map) from a
    FLAT historical-balance workbook (one pool per row in col A; no
    sub-category hierarchy).

    Mirrors ``_apply_pool_seed_to_state`` but for the flat layout: each
    label IS a canonical pool name, so ``pool_map`` is seeded with each
    label mapping to itself. Aggregate rows ("Total", "Grand Total",
    "Sub-total", "All Loans") are filtered out.

    Never overwrites a user's existing entries: ``pool_settings`` is
    only populated when empty; ``monthly_bal.pool_map`` entries are
    only added for labels not yet keyed in the map.
    """
    out: dict[str, list[str]] = {}
    cleaned: list[str] = []
    seen_norm: set[str] = set()
    for raw in labels or []:
        nm = (raw or "").strip()
        if not nm:
            continue
        if _AGG_LABEL_RX.match(nm):
            continue
        key = nm.lower()
        if key in seen_norm:
            continue
        seen_norm.add(key)
        cleaned.append(nm)
    if not cleaned:
        return out

    # ---- state["pool_settings"] (Step 2 "Loan Pools" table) ------------
    if not state.get("pool_settings"):
        ps_rows: list[dict[str, Any]] = []
        for nm in cleaned:
            ps_rows.append({
                "name": nm,
                "risk_rated": False,
                "brr": False,
                "acl_months": _default_acl_months_for_pool(nm),
                "use_default_mgmt_adj": False,
                "excluded": False,
            })
        state["pool_settings"] = ps_rows
        out.setdefault("loan_pools", []).append(
            f"pool_settings seeded with {len(ps_rows)} pool(s) from {src} "
            f"(review the risk-rated / ACL-months / excluded toggles)"
        )

    # Mirror to state.warm.pools so downstream steps and the final YAML
    # see the same ordered list.
    warm = state.get("warm") or {}
    if not warm.get("pools"):
        warm["pools"] = [p["name"] for p in (state.get("pool_settings") or [])]
        state["warm"] = warm

    # ---- state["monthly_bal"]["pool_map"] (label -> itself) -----------
    mb = state.setdefault("monthly_bal", {})
    pm = mb.setdefault("pool_map", {})
    added = 0
    for nm in cleaned:
        if nm not in pm or not pm.get(nm):
            pm[nm] = nm
            added += 1
    if added:
        out.setdefault("monthly_bal", []).append(
            f"pool_map seeded with {added} label -> pool mapping(s) from "
            f"{src} (each label adopted as its own pool name)"
        )

    # Defensive: clear any pool_map values left over from a prior CU's
    # pool-map upload that don't match the freshly-seeded pool_settings.
    purged = _purge_stale_pool_map_values(state)
    if purged:
        out.setdefault("pools", []).append(
            f"Cleared {purged} stale pool_map value(s) (carried over from "
            f"a prior CU) — re-map on Step 3."
        )

    return out


def _apply_pool_seed_to_state(
    state: dict[str, Any], seed: dict[str, Any]
) -> dict[str, list[str]]:
    """Mutate state.pool_settings / state.warm.pools / state.monthly_bal.pool_map
    from a consolidated 'Balance Sheets <CU>' workbook extraction.

    Never overwrites a user's existing entries: ``pool_settings`` is only
    populated when empty; ``monthly_bal.pool_map`` entries are only
    added for sub-category labels that are not yet keyed in the map.

    Returns a ``{step_key: [messages]}`` dict ready to merge into the
    main report.
    """
    out: dict[str, list[str]] = {}
    pools = seed.get("pools") or []
    if not pools:
        return out
    src = seed.get("source_file") or "balance-sheet workbook"

    # ---- state["pool_settings"] (Step 2 "Loan Pools" table) ------------
    if not state.get("pool_settings"):
        ps_rows: list[dict[str, Any]] = []
        for p in pools:
            nm = (p.get("name") or "").strip()
            if not nm:
                continue
            ps_rows.append({
                "name": nm,
                "risk_rated": False,
                "brr": False,
                "acl_months": _default_acl_months_for_pool(nm),
                "use_default_mgmt_adj": False,
                "excluded": False,
            })
        state["pool_settings"] = ps_rows
        out.setdefault("loan_pools", []).append(
            f"pool_settings seeded with {len(ps_rows)} pool(s) from {src} "
            f"(review the risk-rated / ACL-months / excluded toggles)"
        )

    # Mirror to state.warm.pools so downstream steps and the final YAML
    # see the same ordered list (the no-WARM step2 handler does the same
    # sync on save — we do it eagerly here).
    warm = state.get("warm") or {}
    if not warm.get("pools"):
        warm["pools"] = [p["name"] for p in (state.get("pool_settings") or [])]
        state["warm"] = warm

    # ---- state["monthly_bal"]["pool_map"] (Step 8 sub-cat -> pool) ----
    mb = state.setdefault("monthly_bal", {})
    pm = mb.setdefault("pool_map", {})
    subcat_to_pool = seed.get("subcategory_to_pool") or {}
    added = 0
    for label, pool_name in subcat_to_pool.items():
        if not label:
            continue
        if label not in pm or not pm.get(label):
            pm[label] = pool_name
            added += 1
    if added:
        out.setdefault("monthly_bal", []).append(
            f"pool_map seeded with {added} sub-category mapping(s) from {src} "
            f"(parent pools applied automatically)"
        )
    # Mirror sub-cat labels into parsed_pool_labels so the Step 8 mapping
    # table renders the right row set when the auto-scan ran before the
    # user uploaded a balance file.
    parsed = mb.get("parsed_pool_labels") or []
    for label in subcat_to_pool:
        if label and label not in parsed:
            parsed.append(label)
    mb["parsed_pool_labels"] = parsed

    # Hierarchical 'Balance Sheets <CU>' workbooks have BOTH parent pool
    # rows (showing the sum) AND sub-category rows underneath them. The
    # monthly_bal_parser walks every label row, so without an exclusion
    # list the parent rows would double-count their own sub-categories
    # whenever the parent label is also keyed in the pool_map. Stash the
    # parent pool names here so balance_check / preview / 5300-backfill
    # can pass them as exclude_labels into the parser.
    excl_existing = mb.get("exclude_labels") or []
    excl_existing_norm = {
        (s or "").strip().lower() for s in excl_existing
        if (s or "").strip()
    }
    for p in pools:
        nm = (p.get("name") or "").strip()
        if not nm:
            continue
        if nm.lower() in excl_existing_norm:
            continue
        excl_existing.append(nm)
        excl_existing_norm.add(nm.lower())
    mb["exclude_labels"] = excl_existing

    # ---- Stash full seed on state for UI inspection -------------------
    state["_pool_seed"] = {
        "source_file": src,
        "pools": pools,
        "subcategory_to_pool": subcat_to_pool,
        "code_to_pool": seed.get("code_to_pool") or {},
    }

    # Defensive: clear any pool_map values left over from a prior CU's
    # pool-map upload that don't match the freshly-seeded pool_settings.
    purged = _purge_stale_pool_map_values(state)
    if purged:
        out.setdefault("pools", []).append(
            f"Cleared {purged} stale pool_map value(s) (carried over from "
            f"a prior CU) — re-map on Step 3."
        )

    return out


def apply_pool_grouping_choice(
    state: dict[str, Any],
    choices: dict[str, str],
) -> dict[str, Any]:
    """Rebuild ``state["pool_settings"]`` + ``state["warm"]["pools"]`` +
    ``state["monthly_bal"]["pool_map"]`` from ``state["_pool_seed"]``
    using the user's per-pool keep/split choices.

    Args:
        state:    Wizard state dict (mutated in place).
        choices:  ``{pool_name: "keep" | "split"}``. Pools absent from
                  this map default to ``"keep"``. Pools with no
                  sub-categories are always kept (split would be a
                  no-op).

    Semantics:
        * **keep**  — Parent pool stays a single ``pool_settings`` row;
          all of its sub-category labels point to it in
          ``monthly_bal.pool_map``.
        * **split** — Parent pool is REMOVED from ``pool_settings``;
          each sub-category becomes its OWN ``pool_settings`` row, and
          each sub-cat label points to itself in
          ``monthly_bal.pool_map``.

    User-edited per-pool settings (``risk_rated`` / ``brr`` /
    ``acl_months`` / ``excluded`` / ``use_default_mgmt_adj``) are
    preserved by NAME when the rebuilt list contains a pool with the
    same name as an existing row.

    Returns ``{"ok": bool, "pools_count": int, "mappings_count": int,
    "split_count": int, "kept_count": int}`` or
    ``{"ok": False, "error": str}``.
    """
    seed = state.get("_pool_seed") or {}
    pools_seed = seed.get("pools") or []
    if not pools_seed:
        return {"ok": False, "error": "No pool seed available on state."}

    # Preserve user-edited per-row settings by name.
    existing = {
        (p.get("name") or "").strip(): p
        for p in (state.get("pool_settings") or [])
        if (p.get("name") or "").strip()
    }

    def _row_for(name: str) -> dict[str, Any]:
        if name in existing:
            return dict(existing[name])
        return {
            "name": name,
            "risk_rated": False,
            "brr": False,
            "acl_months": _default_acl_months_for_pool(name),
            "use_default_mgmt_adj": False,
            "excluded": False,
        }

    new_settings: list[dict[str, Any]] = []
    managed_map: dict[str, str] = {}  # sub-cat label -> pool name
    split_count = 0
    kept_count = 0

    for pool in pools_seed:
        pname = (pool.get("name") or "").strip()
        if not pname:
            continue
        subcats = pool.get("subcategories") or []
        choice = (choices.get(pname) or "keep").strip().lower()
        # Split-with-no-subcats is meaningless — fall back to keep.
        if choice == "split" and subcats:
            for sc in subcats:
                label = (sc.get("label") or "").strip()
                if not label:
                    continue
                new_settings.append(_row_for(label))
                managed_map[label] = label
            split_count += 1
        else:
            new_settings.append(_row_for(pname))
            for sc in subcats:
                label = (sc.get("label") or "").strip()
                if label:
                    managed_map[label] = pname
            kept_count += 1

    state["pool_settings"] = new_settings
    warm = state.get("warm") or {}
    warm["pools"] = [p["name"] for p in new_settings]
    state["warm"] = warm

    mb = state.setdefault("monthly_bal", {})
    pm = mb.setdefault("pool_map", {})
    # Only overwrite entries that came from our seed — labels outside
    # the seed (e.g. user-added rows on Step 8) are left alone.
    for label, pool_name in managed_map.items():
        pm[label] = pool_name
    parsed = mb.get("parsed_pool_labels") or []
    for label in managed_map:
        if label not in parsed:
            parsed.append(label)
    mb["parsed_pool_labels"] = parsed

    state["_pool_grouping_choices"] = {
        (pool.get("name") or "").strip(): (
            "split"
            if (
                (choices.get(pool.get("name") or "") or "keep").lower()
                == "split"
                and (pool.get("subcategories") or [])
            )
            else "keep"
        )
        for pool in pools_seed
        if (pool.get("name") or "").strip()
    }

    return {
        "ok": True,
        "pools_count": len(new_settings),
        "mappings_count": len(managed_map),
        "split_count": split_count,
        "kept_count": kept_count,
    }


def _apply_companion_file(
    state: dict[str, Any],
    kind: str,
    entry: dict[str, Any] | None,
    state_list_key: str,
    tmp_subfolder: str,
) -> list[str]:
    """Stage a companion file (impaired / credit-pull / CO / recov / etc.)
    into ``state['sample_uploads'][state_list_key]`` so the Review-step
    file copier picks it up.
    """
    if not entry:
        return []
    src = Path(entry["path"])
    if not src.exists():
        return []
    tmp = _copy_into_tmp(src, tmp_subfolder)
    su = state.setdefault("sample_uploads", {})
    bucket = su.setdefault(state_list_key, [])
    # Don't add duplicate names.
    if any(e.get("name") == src.name for e in bucket):
        return [f"{kind} file already staged: {src.name}"]
    bucket.append({"name": src.name, "path": str(tmp)})
    return [f"{kind} file staged: {src.name}"]


def _stage_hist_scan_files(
    state: dict[str, Any],
    entries: list[dict[str, Any]] | None,
    hist_scan_key: str,
    tmp_subfolder: str,
) -> list[str]:
    """Bulk-stage a list of classifier entries into ``state['hist_scan'][key]``.

    Each entry is shaped ``{name, period, path}`` (the warm_parser/classifier
    output). Files are copied into the auto-scan temp area so subsequent
    requests can read them, and registered under ``hist_scan[<hist_scan_key>]``
    in the same shape that ``routes/setup._add_hist_file_from_path`` writes.
    Existing ``hist_scan`` entries with the same name are NOT duplicated;
    user-uploaded files are preserved.

    Used so Step 5 (Historical Charge-Offs) and Step 6 (Historical Recoveries)
    pick up every monthly CO/Recov file from the folder scan, not just the
    single representative file kept in ``sample_uploads``.
    """
    if not entries:
        return []
    hs = state.get("hist_scan")
    if not isinstance(hs, dict):
        hs = {}
        state["hist_scan"] = hs
    # Match the keyset that `_add_hist_file*` writes so downstream code
    # never trips over a missing list.
    for k in (
        "warm_files", "co_files", "recov_files", "impaired_files",
        "credit_pull_files", "monthly_files", "loan_data_files", "other_files",
        "monthly_co_files", "monthly_recov_files",
    ):
        hs.setdefault(k, [])
    hs.setdefault("ok", True)
    hs.setdefault("folder", state.get("_auto_scan_folder") or "Auto-scanned folder")
    rows = list(hs.get(hist_scan_key) or [])
    existing_names = {e.get("name") for e in rows}
    staged = 0
    for entry in entries:
        try:
            src = Path(entry.get("path") or "")
        except (TypeError, ValueError):
            continue
        if not src.exists():
            continue
        if src.name in existing_names:
            continue
        tmp = _copy_into_tmp(src, tmp_subfolder)
        rows.append({
            "name": src.name,
            "period": entry.get("period"),
            "path": str(tmp),
        })
        existing_names.add(src.name)
        staged += 1
    hs[hist_scan_key] = rows
    if staged:
        return [f"{staged} {hist_scan_key.replace('_', ' ')} staged from folder scan"]
    return []


def _stage_loan_data_extracts(
    state: dict[str, Any],
    extracts: list[dict[str, Any]] | None,
) -> list[str]:
    """Bulk-stage analysed loan-data extracts into
    ``state['sample_uploads']['loan_data_files']`` so the Column Mappings
    step (``step3_columns``) renders dropdowns for each detected extract
    without the user having to re-upload them on Step 12.

    Each ``extracts`` entry is shaped ``{name, saved_path, analysis}``
    where ``analysis`` is the full ``sample_parser.analyse_sample_file``
    return dict. Builds Step-12-shaped entries (mirrors the upload
    handler in ``routes/setup.step2_sample``) and seeds per-file
    ``column_mappings`` from the file's own ``column_suggestions`` plus
    a ``member_account`` block from the parser's per-file inference.
    Skips entries whose name is already present in ``loan_data_files``
    so user-uploaded files are preserved on a re-scan.
    """
    if not extracts:
        return []
    su = state.setdefault("sample_uploads", {})
    bucket = su.setdefault("loan_data_files", [])
    existing_names = {e.get("name") for e in bucket}
    msgs: list[str] = []
    for ex in extracts:
        name = ex.get("name") or ""
        path = ex.get("saved_path") or ""
        analysis = ex.get("analysis") or {}
        if not name or not path or not analysis:
            continue
        if name in existing_names:
            msgs.append(f"loan extract already staged: {name}")
            continue
        headers = list(analysis.get("headers") or [])
        suggestions = dict(analysis.get("column_suggestions") or {})
        # Slim ``analysis`` matches the shape produced by
        # ``routes/setup._loan_data_entry_analysis`` so Step 13 renders
        # dropdowns immediately without lazy hydrate on first GET.
        slim = {
            "headers": headers,
            "column_suggestions": suggestions,
            "pool_code_suggestions": list(
                analysis.get("pool_code_suggestions") or []
            ),
            "sample_rows": list(analysis.get("sample_rows") or [])[:5],
            "has_header": bool(analysis.get("has_header")),
            "header_row": int(analysis.get("header_row") or 0),
            "member_account_suggestion": (
                dict(analysis.get("member_account_suggestion") or {})
                if analysis.get("member_account_suggestion") else None
            ),
            "date_format": str(analysis.get("date_format") or "YYYY-MM"),
        }
        # Per-file member_account: prefer parser inference, fall back
        # to the historical fixed_suffix=3 default. Strip metadata keys.
        ma_sug = analysis.get("member_account_suggestion") or {}
        if ma_sug.get("mode"):
            member_account = {
                "mode": ma_sug.get("mode"),
                "suffix_length": int(ma_sug.get("suffix_length") or 0),
                "delimiter": ma_sug.get("delimiter") or "-",
            }
        else:
            member_account = {
                "mode": "fixed_suffix", "suffix_length": 3, "delimiter": "-",
            }
        # Per-file column_mappings seeded from this file's own header
        # suggestions. Step 13 renders these as dropdown defaults; user
        # can override per-file. Only include suggestions whose value
        # is actually one of the file's headers (avoids ghost options).
        col_map = {
            field: header
            for field, header in suggestions.items()
            if header and header in headers
        }
        entry = {
            "name": name,
            "path": path,
            "has_header": bool(analysis.get("has_header")),
            "header_row": int(analysis.get("header_row") or 0),
            "analysis": slim,
            "column_mappings": col_map,
            "member_account": member_account,
            "file_pattern": analysis.get("file_pattern") or "",
        }
        # Phase 9.26: per-extract pool_code_split override (CUMA/mortgage
        # files where ``/`` is part of the label, not a prefix:desc split).
        if "pool_code_split" in ex:
            entry["pool_code_split"] = ex.get("pool_code_split") or ""
        bucket.append(entry)
        existing_names.add(name)
        msgs.append(
            f"loan extract staged: {name} "
            f"({len(headers)} cols, {len(col_map)} suggested mappings)"
        )
    return msgs


def selfheal_hist_scan_from_folder(state: dict[str, Any]) -> list[str]:
    """Proactively bulk-stage CO/Recov files into ``hist_scan`` for any draft
    whose Step-1 auto-scan completed but whose ``monthly_co_files`` /
    ``monthly_recov_files`` lists are still empty.

    Safe to call from any wizard GET hook. Returns the staging messages so
    callers can decide whether to flash them.

    Two reasons this self-heal is needed:

    1. Drafts saved BEFORE the Phase 9.4 bulk-stage code was wired into
       ``apply_findings_to_state`` only have the single representative
       file in ``sample_uploads``; the per-month list is empty.
    2. Combined CO+Recov files (Destinations-style "Recoveries -
       Charge-offs") are routed into BOTH ``classify_folder['co_files']``
       and ``classify_folder['recov_files']`` via the broadened cross-list
       overlay, but ``apply_findings_to_state`` ran on a draft before that
       overlay existed -- so only one bucket got bulk-staged.

    The old step3_historical lazy self-heal only fired when the user
    landed on Step 5 or Step 6, so the OTHER section's badge stayed
    stale until the user navigated there. Running this on every wizard
    GET (from ``_wizard_ctx``) makes the breadcrumb truthful across
    every step.
    """
    if not state.get("_auto_scan_completed"):
        return []
    folder = state.get("_auto_scan_folder") or ""
    hs = state.get("hist_scan") or {}
    need_co = not (hs.get("monthly_co_files") or hs.get("co_files"))
    need_recov = not (hs.get("monthly_recov_files") or hs.get("recov_files"))
    msgs: list[str] = []
    # If files are already staged but the source picker is still on its
    # bare default, just flip the picker -- no need to reclassify.
    if not (need_co or need_recov):
        msgs.extend(_auto_flip_hist_source(state))
        return msgs
    if not folder or not Path(folder).exists():
        # Folder unreachable: still try to flip the source picker for
        # the side that already has files staged.
        msgs.extend(_auto_flip_hist_source(state))
        return msgs
    try:
        cls = classify_folder(folder)
    except Exception:  # noqa: BLE001 -- never let self-heal block the page
        msgs.extend(_auto_flip_hist_source(state))
        return msgs
    if need_co:
        msgs.extend(_stage_hist_scan_files(
            state, cls.get("co_files") or [], "monthly_co_files", "co",
        ))
    if need_recov:
        msgs.extend(_stage_hist_scan_files(
            state, cls.get("recov_files") or [], "monthly_recov_files", "recov",
        ))
    # When per-month files are staged, flip the section's source picker
    # from the default "single_workbook" to "monthly_files" so the user
    # actually SEES those files on Step 5/6 instead of an empty single-
    # workbook upload field. Respects user choice: only flips when the
    # current value is still the bare default and no single workbook
    # has been uploaded.
    msgs.extend(_auto_flip_hist_source(state))
    return msgs


def selfheal_single_hist_bal_layout(state: dict[str, Any]) -> list[str]:
    """Re-run ``monthly_bal_parser.analyse_file`` on
    ``monthly_bal.saved_path`` when a draft has
    ``monthly_bal.source == 'single'`` + a saved file path but
    ``parsed_pool_labels`` is empty (or ``pool_name_col`` points at a
    column that yields no labels). Targets two stale-draft cases:

      * Older drafts seeded by an earlier ``analyse_file`` whose
        column-election heuristic picked column A on a workbook with
        loan codes in A and descriptive labels in B (e.g. Emergency
        Responders CU's Vizo-style ``Historical Balances`` workbook).
      * Drafts saved before a parser fix landed.

    Always runs (no auto-scan gate) so user-uploaded single workbooks
    self-repair too. Read-only on the workbook (load_workbook
    read_only=True). Never overwrites user-entered ``pool_map``
    entries — only fills in blanks for newly-discovered labels.
    """
    mb = state.get("monthly_bal") or {}
    if (mb.get("source") or "") != "single":
        return []
    saved_path = mb.get("saved_path") or ""
    if not saved_path:
        return []
    # Only heal when there's actually nothing useful in state already.
    labels = mb.get("parsed_pool_labels") or []
    if labels:
        return []
    try:
        from pathlib import Path as _Path
        if not _Path(saved_path).exists():
            return []
    except Exception:  # noqa: BLE001
        return []
    try:
        from cecl_ui.services import monthly_bal_parser as _mbp
        an = _mbp.analyse_file(saved_path)
    except Exception:  # noqa: BLE001
        return []
    if not an.get("ok"):
        return []
    new_labels = an.get("parsed_pool_labels") or []
    if not new_labels:
        return []
    msgs: list[str] = []
    prior_pnc = mb.get("pool_name_col") or ""
    new_pnc = an.get("pool_name_col") or ""
    new_sheet = an.get("sheet") or ""
    new_hdr = int(an.get("header_row") or 0)
    new_fdc = an.get("first_date_col") or ""
    if new_sheet:
        mb["sheet"] = new_sheet
    if new_hdr:
        mb["header_row"] = new_hdr
    if new_pnc:
        mb["pool_name_col"] = new_pnc
    if new_fdc:
        mb["first_date_col"] = new_fdc
    mb["parsed_pool_labels"] = new_labels
    if an.get("dates"):
        mb["parsed_dates"] = an.get("dates") or []
    # Seed pool_map for the newly-parsed labels without overwriting
    # any user selection. Use seed_pool_map to fuzzy-match against
    # WARM balance_title_map + historical hist_pool_map so common
    # labels get auto-mapped where possible.
    existing = mb.get("pool_map") or {}
    try:
        from cecl_ui.services import monthly_bal_parser as _mbp2
        _combined: dict[str, str] = {}
        for _k, _v in (state.get("balance_title_map") or {}).items():
            if _k:
                _combined[_k] = (_v or "")
        _hpm = state.get("hist_pool_map") or {}
        for _k, _v in (_hpm.get("mapping") or {}).items():
            if _k and _k not in _combined:
                _combined[_k] = (_v or "")
        _seeded, _status = _mbp2.seed_pool_map(new_labels, _combined)
        mb["label_status"] = _status
    except Exception:  # noqa: BLE001
        _seeded = {lab: "" for lab in new_labels}
    seed_changed = False
    for lab, val in _seeded.items():
        if lab and lab not in existing:
            existing[lab] = val or ""
            seed_changed = True
    if seed_changed:
        mb["pool_map"] = existing
    state["monthly_bal"] = mb
    if prior_pnc and new_pnc and prior_pnc.upper() != new_pnc.upper():
        msgs.append(
            f"Historical balance layout: re-detected label column "
            f"{new_pnc} (was {prior_pnc}); recovered {len(new_labels)} "
            f"pool label(s) from "
            f"{mb.get('filename') or _Path(saved_path).name}."
        )
    else:
        msgs.append(
            f"Historical balance layout: recovered "
            f"{len(new_labels)} pool label(s) from "
            f"{mb.get('filename') or _Path(saved_path).name} "
            f"(label column {new_pnc or 'A'})."
        )
    return msgs


# ---------------------------------------------------------------------------
# Heuristic auto-mapping of balance labels to pool_settings names
# ---------------------------------------------------------------------------

# Category keyword patterns. Order in ``_CATEGORY_ORDER`` matters:
# more specific categories are evaluated first so that, e.g.,
# "new recreational vehicle" classifies as ``recreational``, not
# ``new_vehicle``. Vehicle handling is special-cased after the
# recreational check (see ``_classify_label_category``) so labels
# with truncated suffixes (e.g. ``AUTO SPECIAL - USED A``) still
# resolve correctly via ``new``/``used`` qualifier detection.
_LABEL_KEYWORDS: dict[str, tuple[str, ...]] = {
    "credit_card": (
        r"\b(credit\s*card|ccard|visa\s*card|mastercard)\b",
    ),
    "real_estate": (
        r"\b(mortgage|first\s*lien|1st\s*lien|junior\s*lien|"
        r"2nd\s*lien|home\s*equity|heloc|real\s*estate|residential|"
        r"1-4\s*family|unimproved|construction\s*loan|land\s*loan)\b",
    ),
    "share_secured": (
        r"\b(share\s*secured|share\s*draft|share\s*pledge|"
        r"cd\s*secured|certificate\s*secured|shr\s*sec)\b",
    ),
    "recreational": (
        r"\b(motorcycle|motor\s*cycle|atv|snowmobile|boat|marine|"
        r"recreational|equipment|trailer|jet\s*ski|camper|rv)\b",
    ),
    "unsecured": (
        r"\b(signature|unsecured|personal\s*loan|christmas\s*loan|"
        r"holiday\s*loan|short[\s\-]*term|rapid|cosigner|"
        r"co[\s\-]?signer)\b",
        r"\bline\s*of\s*credit\b",
        r"\bloc\b",
    ),
    "commercial": (
        r"\b(business\s*loan|commercial|sba|business\s*line)\b",
    ),
    "aggregate": (
        r"\b(totals?|subtotals?|grand[\s_\-]*totals?|"
        r"all[\s_\-]*loans?|loan[\s_\-]*account[\s_\-]*totals?)\b",
    ),
    "other_consumer": (
        r"\b(other\s+consumer|miscellaneous|misc\s+loan|"
        r"consumer\s+loan)\b",
    ),
}

_VEHICLE_KEYWORDS_RX = re.compile(
    r"\b(vehicles?|vehciles?|auto|autos|car|cars|truck|trucks|van|vans)\b",
    re.IGNORECASE,
)
_NEW_QUALIFIER_RX = re.compile(r"\bnew\b", re.IGNORECASE)
_USED_QUALIFIER_RX = re.compile(r"\bused\b", re.IGNORECASE)

# Categories evaluated BEFORE vehicle-context detection so e.g.
# "new recreational vehicle" picks recreational over new_vehicle.
_CATEGORY_ORDER_HIGH: tuple[str, ...] = (
    "credit_card",
    "real_estate",
    "share_secured",
    "recreational",
)
# Categories evaluated AFTER vehicle-context detection.
_CATEGORY_ORDER_LOW: tuple[str, ...] = (
    "unsecured",
    "commercial",
    "aggregate",
    "other_consumer",
)


def _classify_label_category(text: str) -> str:
    """Return the first matching category for ``text``, or ``""``.

    Vehicle classification is special-cased: any label with a vehicle
    keyword (``auto``/``vehicle``/``car``/``truck``/etc, including the
    common typo ``vehciles``) plus a ``new``/``used`` qualifier
    anywhere in the string resolves to ``new_vehicle`` /
    ``used_vehicle`` even when the qualifier is truncated (e.g.
    ``AUTO SPECIAL - USED A`` from a 20-char-truncated Symitar export).
    """
    if not text:
        return ""
    s = str(text).strip().lower()
    if not s:
        return ""
    # Pass 1: high-specificity categories (recreational beats vehicle).
    for cat in _CATEGORY_ORDER_HIGH:
        for pat in _LABEL_KEYWORDS.get(cat, ()):
            try:
                if re.search(pat, s):
                    return cat
            except re.error:  # pragma: no cover
                continue
    # Pass 2: vehicle context with new/used qualifier.
    if _VEHICLE_KEYWORDS_RX.search(s):
        has_new = bool(_NEW_QUALIFIER_RX.search(s))
        has_used = bool(_USED_QUALIFIER_RX.search(s))
        if has_used and not has_new:
            return "used_vehicle"
        if has_new and not has_used:
            return "new_vehicle"
        if has_new and has_used:
            # Ambiguous; prefer used (more conservative provisioning).
            return "used_vehicle"
        return "vehicle_unspecified"
    # Pass 3: low-specificity categories.
    for cat in _CATEGORY_ORDER_LOW:
        for pat in _LABEL_KEYWORDS.get(cat, ()):
            try:
                if re.search(pat, s):
                    return cat
            except re.error:  # pragma: no cover
                continue
    return ""


def _build_pool_category_index(
    pool_settings: list[dict[str, Any]] | None,
) -> dict[str, str]:
    """Return ``{category: pool_name}`` for the configured pools.

    Pools without a category match are skipped. If multiple pools share
    a category the first one wins (preserves user ordering on Step 2).
    """
    out: dict[str, str] = {}
    for ps in (pool_settings or []):
        nm = (ps.get("name") if isinstance(ps, dict) else "") or ""
        nm = nm.strip()
        if not nm:
            continue
        cat = _classify_label_category(nm)
        if cat and cat not in out:
            out[cat] = nm
    return out


def _suggest_pool_for_label(
    label: str,
    pool_index: dict[str, str],
    pool_names: list[str] | None = None,
) -> str:
    """Best-effort pool-name suggestion for ``label``.

    Precedence: (1) exact case-insensitive match of ``label`` against
    any pool name in ``pool_names`` wins outright — this handles CUs
    where multiple pools fall in the same keyword category
    (``Indirect New Auto`` and ``New Auto`` both classify as
    ``new_vehicle``, so the keyword heuristic would collapse them onto
    the same pool). (2) Falls back to the keyword-category index.
    Returns ``""`` when no confident match. Vehicle labels missing
    new/used qualifiers are routed to whichever vehicle pool exists,
    preferring ``other_consumer`` as a last-resort fallback. Aggregate
    rows (``Total``/``Subtotal``/``Loan Account Totals``) are
    intentionally left blank.
    """
    lab_norm = (label or "").strip().casefold()
    if lab_norm and pool_names:
        for nm in pool_names:
            if (nm or "").strip().casefold() == lab_norm:
                return nm
    cat = _classify_label_category(label)
    if not cat or cat == "aggregate":
        return ""
    if cat in pool_index:
        return pool_index[cat]
    # Soft fallbacks: route niche categories to a general pool when the
    # exact match isn't configured on Step 2.
    if cat == "vehicle_unspecified":
        for alt in ("new_vehicle", "used_vehicle", "other_consumer"):
            if alt in pool_index:
                return pool_index[alt]
    if cat == "recreational":
        if "other_consumer" in pool_index:
            return pool_index["other_consumer"]
    if cat == "commercial":
        if "other_consumer" in pool_index:
            return pool_index["other_consumer"]
    return ""


def selfheal_auto_map_balance_labels(state: dict[str, Any]) -> list[str]:
    """Auto-map blank ``monthly_bal.pool_map`` entries to configured
    ``pool_settings`` using keyword heuristics.

    Targets workbook-source CUs without a WARM ``balance_title_map``
    (the typical state for any CU that auto-detected a single
    historical balance workbook). Only fills BLANK entries so user
    edits are preserved. Idempotent: subsequent calls re-evaluate
    blanks and either fill or leave them — never overwrites a
    non-empty mapping.
    """
    mb = state.get("monthly_bal") or {}
    if not isinstance(mb, dict):
        return []
    labels = mb.get("parsed_pool_labels") or []
    if not labels:
        return []
    pool_settings = state.get("pool_settings") or []
    pool_index = _build_pool_category_index(pool_settings)
    if not pool_index:
        return []
    pool_names = [
        (ps.get("name") or "").strip()
        for ps in pool_settings
        if isinstance(ps, dict) and (ps.get("name") or "").strip()
    ]
    pool_map = mb.get("pool_map") or {}
    if not isinstance(pool_map, dict):
        pool_map = {}
    filled = 0
    for lab in labels:
        if not lab:
            continue
        existing = (pool_map.get(lab) or "").strip()
        if existing:
            continue
        suggestion = _suggest_pool_for_label(lab, pool_index, pool_names)
        if suggestion:
            pool_map[lab] = suggestion
            filled += 1
    if not filled:
        return []
    mb["pool_map"] = pool_map
    state["monthly_bal"] = mb
    return [
        f"Auto-mapped {filled} balance label(s) to configured pool(s) "
        f"based on keyword match. Review on Step 8 (Monthly Balance "
        f"File) and adjust any mismatches."
    ]


def selfheal_enforce_exact_match_pool_map(
    state: dict[str, Any],
) -> list[str]:
    """Correct non-blank ``monthly_bal.pool_map`` entries whose label
    case-insensitively equals a configured pool name but is currently
    mapped to a DIFFERENT pool.

    Complements ``selfheal_auto_map_balance_labels`` (which only fills
    blanks). When the keyword heuristic ran on an earlier setup pass
    and collapsed exact-name labels onto a sibling pool (e.g.
    ``New Auto`` → ``Indirect New Auto``), this heal snaps them back to
    the exact pool.

    Safe on every GET. Idempotent — subsequent calls find no
    corrections to make. Never touches labels whose current mapping
    doesn't have a case-insensitive exact-name candidate in
    ``pool_settings``.
    """
    mb = state.get("monthly_bal") or {}
    if not isinstance(mb, dict):
        return []
    pool_map = mb.get("pool_map") or {}
    if not isinstance(pool_map, dict) or not pool_map:
        return []
    pool_settings = state.get("pool_settings") or []
    pool_names = [
        (ps.get("name") or "").strip()
        for ps in pool_settings
        if isinstance(ps, dict) and (ps.get("name") or "").strip()
    ]
    if not pool_names:
        return []
    pool_lookup = {nm.casefold(): nm for nm in pool_names}
    corrections: list[tuple[str, str, str]] = []
    for label, mapped in list(pool_map.items()):
        lab_key = (label or "").strip().casefold()
        if not lab_key:
            continue
        exact = pool_lookup.get(lab_key)
        if not exact:
            continue
        current = (mapped or "").strip()
        if current == exact:
            continue
        pool_map[label] = exact
        corrections.append((label, current or "(blank)", exact))
    if not corrections:
        return []
    mb["pool_map"] = pool_map
    state["monthly_bal"] = mb
    sample = "; ".join(
        f"{lab!r}: {old} \u2192 {new}" for lab, old, new in corrections[:3]
    )
    tail = "" if len(corrections) <= 3 else f" (+{len(corrections) - 3} more)"
    return [
        f"Corrected {len(corrections)} exact-match label mapping(s): "
        f"{sample}{tail}."
    ]


def selfheal_adopt_yaml_schema_into_state(state: dict[str, Any]) -> list[str]:
    """Translate YAML-config schema keys into wizard-state schema keys
    in-place when a draft was created via ``adopt_config_to_completed``
    (or any path that writes the raw YAML dict as the wizard state).

    The adoption path in ``cecl_ui/routes/home.py`` writes the YAML
    config verbatim as the draft payload. That works for run-time
    report generation (which reads YAML keys) but the wizard renders
    every step from a DIFFERENT schema (``pool_settings``, ``monthly_bal``,
    ...). Without translation, the wizard's Loan Pools table is empty
    even though the YAML's ``pools`` block has every pool definition.

    Idempotent. Safe to call on every GET. Only fills empty wizard-side
    fields — never overwrites user edits.
    """
    msgs: list[str] = []

    # ---- Phase 9.35d: seed missing top-level scalar defaults --------
    #
    # Adopted drafts hydrate from YAML and may be missing top-level
    # scalars that `_default_state()` would otherwise provide. Each
    # entry is a (key, default) pair mirroring the canonical default in
    # `cecl_ui/routes/setup.py _default_state()`. Uses `setdefault` so
    # existing values (including legitimate user-zero/empty values like
    # `accounting_negatives=False`) are preserved.
    _scalar_defaults: list[tuple[str, Any]] = [
        ("file_pattern", r"LOANDATA.*\.(xlsx|xls|csv)$"),
        ("date_pattern", r"(\d{4})-(\d{2})"),
        ("date_format", "YYYY-MM"),
        ("pool_code_split", "/"),
        ("default_pool", "Ignore"),
        ("balance_remove_chars", ["$", ","]),
        ("accounting_negatives", True),
        ("has_header", True),
        ("account_suffix_length", 3),
        ("no_score_label", "Not Reported"),
        ("raw_data_folder", ""),
        ("report_period", ""),
    ]
    _seeded: list[str] = []
    for _k, _default in _scalar_defaults:
        if _k not in state:
            state[_k] = _default
            _seeded.append(_k)
    if _seeded:
        msgs.append(
            f"Seeded {len(_seeded)} missing top-level default(s) from the "
            f"adopted YAML config: {', '.join(_seeded)}."
        )

    # ---- pools (YAML) -> pool_settings (wizard) ----------------------
    if not state.get("pool_settings"):
        yaml_pools = state.get("pools")
        if isinstance(yaml_pools, list) and yaml_pools:
            ps_rows: list[dict[str, Any]] = []
            for p in yaml_pools:
                if not isinstance(p, dict):
                    continue
                nm = str(p.get("name") or "").strip()
                if not nm:
                    continue
                ps_rows.append({
                    "name": nm,
                    "risk_rated": bool(p.get("risk_rated", True)),
                    "brr": bool(p.get("brr", False)),
                    "acl_months": p.get("acl_months"),
                    "use_default_mgmt_adj": bool(
                        p.get("use_default_mgmt_adj", False)
                    ),
                    "excluded": bool(p.get("excluded", False)),
                })
            if ps_rows:
                state["pool_settings"] = ps_rows
                # Mirror to warm.pools so downstream steps see the same list.
                warm = state.get("warm") or {}
                if not warm.get("pools"):
                    warm["pools"] = [r["name"] for r in ps_rows]
                    state["warm"] = warm
                msgs.append(
                    f"Recovered {len(ps_rows)} loan pool(s) from the adopted "
                    f"YAML config — review on Step 2 (Loan Pools)."
                )

    # ---- monthly_balance (YAML) -> monthly_bal (wizard) -------------
    mb = state.setdefault("monthly_bal", {})
    yaml_mbal = state.get("monthly_balance")
    if (
        isinstance(yaml_mbal, dict)
        and yaml_mbal
        and not (mb.get("saved_path") or mb.get("parsed_pool_labels"))
    ):
        src_raw = str(yaml_mbal.get("source") or "").strip().lower()
        # Map YAML source name -> wizard source name.
        src_map = {
            "single_workbook": "single",
            "single": "single",
            "per_year": "per_year",
            "annual": "per_year",
            "per_month": "per_month",
            "monthly": "per_month",
            "monthly_loan_extracts": "monthly_loan_extracts",
        }
        wizard_src = src_map.get(src_raw, src_raw or "single")
        mb["source"] = wizard_src
        if yaml_mbal.get("saved_path"):
            mb["saved_path"] = str(yaml_mbal["saved_path"])
        if yaml_mbal.get("filename"):
            mb["filename"] = str(yaml_mbal["filename"])
        layout = yaml_mbal.get("layout") or {}
        if isinstance(layout, dict):
            if layout.get("sheet"):
                mb["sheet"] = str(layout["sheet"])
            if layout.get("label_col"):
                mb["pool_name_col"] = str(layout["label_col"])
            if layout.get("first_date_col"):
                mb["first_date_col"] = str(layout["first_date_col"])
            if layout.get("header_row") is not None:
                try:
                    mb["header_row"] = int(layout["header_row"])
                except (TypeError, ValueError):
                    pass
        ypm = yaml_mbal.get("pool_map")
        if isinstance(ypm, dict) and ypm and not mb.get("pool_map"):
            mb["pool_map"] = dict(ypm)
            if not mb.get("parsed_pool_labels"):
                mb["parsed_pool_labels"] = list(ypm.keys())
        state["monthly_bal"] = mb
        if msgs or src_raw:
            msgs.append(
                "Recovered Monthly Balance File layout from the adopted YAML "
                "config — review on Step 8 (Monthly Balance File)."
            )

    return msgs


def selfheal_flat_pool_seed_from_state(state: dict[str, Any]) -> list[str]:
    """Proactively seed Step 2 ``pool_settings`` + Step 8
    ``monthly_bal.pool_map`` from a previously-detected flat
    ``Historical Balance Sheets`` workbook when those state keys are
    still empty.

    Targets drafts saved BEFORE the flat-pool-labels seeding wired into
    ``apply_findings_to_state`` for ``single_hist_bal`` findings. Reads
    only state (``monthly_bal.parsed_pool_labels`` /
    ``monthly_bal.filename``) — no folder I/O — so it's safe to call
    from any wizard GET hook.

    Skips silently when:
      * Auto-scan never completed
      * ``pool_settings`` already populated (preserves user edits)
      * The hist-balance source is not the flat single workbook shape
        (consolidated pool-seed workbooks have their own seeding path)
      * No pool labels were parsed
    """
    if not state.get("_auto_scan_completed"):
        return []
    if state.get("pool_settings"):
        return []
    mb = state.get("monthly_bal") or {}
    if (mb.get("source") or "") != "single":
        return []
    labels = mb.get("parsed_pool_labels") or []
    if not labels:
        return []
    src = mb.get("filename") or "single historical balance workbook"
    flat_report = _apply_flat_pool_labels_to_state(state, labels, src)
    msgs: list[str] = []
    for key in ("loan_pools", "monthly_bal"):
        for m in flat_report.get(key, []):
            msgs.append(m)
    return msgs


def _auto_flip_hist_source(state: dict[str, Any]) -> list[str]:
    """If monthly CO/Recov files have been staged but the user-facing
    source picker on Step 5/6 still says ``single_workbook`` (the bare
    default), flip it to ``monthly_files`` so the per-file table
    renders. Only fires when the corresponding ``co_files`` /
    ``recov_files`` (legacy single-workbook uploads) list is empty,
    so a user who deliberately uploaded a single workbook is preserved.
    Returns human-readable messages for the auto-scan report.
    """
    msgs: list[str] = []
    hs = state.get("hist_scan") or {}
    # CO side
    if (
        state.get("hist_co_source") in (None, "", "single_workbook")
        and (hs.get("monthly_co_files") or [])
        and not (hs.get("co_files") or [])
    ):
        if state.get("hist_co_source") != "monthly_files":
            state["hist_co_source"] = "monthly_files"
            msgs.append(
                "Charge-off source set to 'Monthly charge-off files' "
                "(per-month files detected by folder scan)."
            )
    # Recov side
    if (
        state.get("hist_recov_source") in (None, "", "single_workbook")
        and (hs.get("monthly_recov_files") or [])
        and not (hs.get("recov_files") or [])
    ):
        if state.get("hist_recov_source") != "monthly_files":
            state["hist_recov_source"] = "monthly_files"
            msgs.append(
                "Recovery source set to 'Monthly recovery files' "
                "(per-month files detected by folder scan)."
            )
    return msgs


def _apply_credit_pull_to_state(
    state: dict[str, Any], entry: dict[str, Any] | None
) -> list[str]:
    """Stage credit-pull file (similar to companion_file but writes to
    state.credit_pull as well so Step 10 can render it correctly)."""
    if not entry:
        return []
    src = Path(entry["path"])
    if not src.exists():
        return []
    tmp = _copy_into_tmp(src, "credit_pull")
    cp = state.setdefault("credit_pull", {})
    cp["uploaded_filename"] = src.name
    cp["source_folder"] = str(src.parent)
    cp["use_standalone_file"] = True
    # Also stage into sample_uploads so Review-step copies into Raw_Uploads.
    su = state.setdefault("sample_uploads", {})
    bucket = su.setdefault("credit_pull_files", [])
    if not any(e.get("name") == src.name for e in bucket):
        bucket.append({"name": src.name, "path": str(tmp)})
    return [f"credit_pull file staged: {src.name}"]


def _apply_data_directory(
    state: dict[str, Any],
    workspace_root: str | Path,
) -> list[str]:
    """Point state.data_directory at the per-CU Raw_Uploads folder so the
    importer + report engines can find any companion files we'll copy
    there on Review submit.
    """
    sn = state.get("short_name") or config_service.slugify(
        state.get("credit_union") or ""
    )
    if not sn:
        return []
    raw = config_service.raw_uploads_dir(workspace_root) / sn
    state["data_directory"] = str(raw)
    return [f"data_directory set to {raw}"]


# --------------------------------------------------------------------------
# 5300 distributed backfill (auto-run at end of scan)
# --------------------------------------------------------------------------


def _compute_earliest_month_pool_distribution(
    state: dict[str, Any],
) -> dict[str, Any]:
    """Service-side mirror of ``routes/setup._earliest_month_pool_distribution``.

    Only handles the two ``hist_balance_source`` values that the
    auto-scan can ever produce -- ``annual_balance_sheets`` and
    ``monthly_balance_sheets``. Single-workbook + monthly-loan-extract
    flavours are picked by the user manually on Step 4 and the
    auto-runner just skips them.

    Returns a dict with keys ``ok``, ``error``, ``period``, ``total``,
    ``pool_distribution`` (ratios that sum to 1.0), ``by_pool`` (raw $
    per pool at the earliest month), and ``source``.
    """
    out: dict[str, Any] = {
        "ok": False, "error": "", "period": "", "total": 0.0,
        "pool_distribution": {}, "by_pool": {}, "source": "",
    }
    mb_state = state.get("monthly_bal") or {}
    label_to_pool = mb_state.get("pool_map") or {}
    source = state.get("hist_balance_source") or ""
    by_period: dict[str, Any] = {}
    if source == "annual_balance_sheets":
        out["source"] = "annual"
        ann = monthly_bal_parser.pool_balances_for_per_year_files(
            mb_state.get("year_files") or [],
            mb_state.get("per_year_layout") or {},
            label_to_pool,
        )
        if not ann.get("ok") and not (ann.get("by_period") or {}):
            out["error"] = (
                f"Could not read annual balance workbook(s): "
                f"{ann.get('error') or 'unknown error'}"
            )
            return out
        by_period = ann.get("by_period") or {}
    elif source == "monthly_balance_sheets":
        out["source"] = "per_month"
        pm = monthly_bal_parser.pool_balances_for_per_month_files(
            mb_state.get("monthly_files") or [],
            mb_state.get("per_month_layout") or {},
            label_to_pool,
        )
        if not pm.get("ok") and not (pm.get("by_period") or {}):
            out["error"] = (
                f"Could not read monthly balance file(s): "
                f"{pm.get('error') or 'unknown error'}"
            )
            return out
        by_period = pm.get("by_period") or {}
    else:
        out["error"] = (
            f"Auto-run only supports annual / monthly balance sheets "
            f"(hist_balance_source={source!r})"
        )
        return out
    if not by_period:
        out["error"] = "No periods extracted from balance file(s)."
        return out
    earliest = sorted(by_period.keys())[0]
    by_pool = by_period[earliest] or {}
    total = sum(float(v or 0) for v in by_pool.values())
    if total <= 0:
        out["error"] = (
            f"Earliest month ({earliest}) has zero total balance "
            f"-- check pool_map labels."
        )
        return out
    out["ok"] = True
    out["period"] = earliest
    out["total"] = total
    out["by_pool"] = dict(by_pool)
    out["pool_distribution"] = {
        k: float(v or 0) / total
        for k, v in by_pool.items()
        if (v or 0) > 0
    }
    return out


def _auto_run_distributed_backfill(
    state: dict[str, Any],
) -> list[str]:
    """If all preconditions are met, run the distributed 5300 backfill
    and return summary messages. Stores result on
    ``state['hist_extracts']['solr_backfill']['last_run']``. Never
    raises -- any failure is captured in the returned messages and on
    ``last_run`` so a scan can never 500 because of this auto-run.

    Preconditions:
      * ``credit_union`` and a numeric ``charter_number`` on Identity
      * ``report_period`` (used as ``target_period``)
      * ``hist_balance_source`` is annual_balance_sheets or
        monthly_balance_sheets
      * ``monthly_bal.pool_map`` has at least one non-Ignore mapping
        so the earliest-month distribution is non-empty
    """
    msgs: list[str] = []
    # Lazy import keeps the rest of auto_setup decoupled from the DB +
    # HTTP transport that the backfill drags in.
    try:
        from cecl_ui.services import solr_5300_backfill  # noqa: F401
    except Exception as exc:  # noqa: BLE001
        return [f"5300 backfill skipped: import failed ({exc})"]

    cu = (state.get("credit_union") or "").strip()
    charter_raw = (state.get("charter_number") or "").strip()
    try:
        charter_int = (
            int(re.sub(r"\D", "", charter_raw)) if charter_raw else 0
        )
    except (TypeError, ValueError):
        charter_int = 0
    report_period = (state.get("report_period") or "").strip()
    source = (state.get("hist_balance_source") or "").strip()

    # Initialise hist_extracts so the result has a place to live and
    # the user lands on Step 4 with the mode already set to distributed.
    he = state.get("hist_extracts")
    if not isinstance(he, dict):
        he = {}
        state["hist_extracts"] = he
    he.setdefault("target_period", report_period)
    he.setdefault("history_months", 84)
    sb = he.get("solr_backfill")
    if not isinstance(sb, dict):
        sb = {}
        he["solr_backfill"] = sb
    sb.setdefault("solr_url", "http://searchserver1.tctrisk.com:8983/solr")
    sb.setdefault("core", "ncua")
    sb["mode"] = "distributed"
    # If a prior wizard pass left target_period blank, fill from the
    # Identity step's report_period.
    if not (he.get("target_period") or "").strip():
        he["target_period"] = report_period
    target_period = (he.get("target_period") or "").strip()
    months = int(he.get("history_months") or 84)

    missing: list[str] = []
    if not cu:
        missing.append("credit_union")
    if not charter_int:
        missing.append("charter_number")
    if not target_period:
        missing.append("report_period")
    if source not in ("annual_balance_sheets", "monthly_balance_sheets"):
        missing.append(f"hist_balance_source ({source!r})")
    if missing:
        return [
            "5300 backfill (distributed) skipped at auto-scan time -- "
            "still needs: " + ", ".join(missing)
        ]

    try:
        dist = _compute_earliest_month_pool_distribution(state)
    except Exception as exc:  # noqa: BLE001
        return [f"5300 backfill skipped: distribution error ({exc})"]
    if not dist.get("ok"):
        return [
            "5300 backfill (distributed) skipped: "
            + (dist.get("error") or "no pool distribution available")
        ]

    # Quarter-ends already covered by the user's uploaded balance file(s)
    # are passed as ``existing_dates`` so the backfill doesn't overwrite
    # them.
    upload_months: set[str] = set()
    try:
        mb_state = state.get("monthly_bal") or {}
        label_to_pool = mb_state.get("pool_map") or {}
        if source == "annual_balance_sheets":
            ann = monthly_bal_parser.pool_balances_for_per_year_files(
                mb_state.get("year_files") or [],
                mb_state.get("per_year_layout") or {},
                label_to_pool,
            )
            upload_months = set((ann.get("by_period") or {}).keys())
        else:  # monthly_balance_sheets
            pm = monthly_bal_parser.pool_balances_for_per_month_files(
                mb_state.get("monthly_files") or [],
                mb_state.get("per_month_layout") or {},
                label_to_pool,
            )
            upload_months = set((pm.get("by_period") or {}).keys())
    except Exception:  # noqa: BLE001
        upload_months = set()

    try:
        from cecl_ui.services import solr_5300_backfill
        result = solr_5300_backfill.backfill_missing_quarters_distributed(
            cu, charter_int, sb["solr_url"], sb["core"],
            target_period, months,
            pool_distribution=dist["pool_distribution"],
            existing_dates=upload_months,
            source_period_iso=dist.get("period") or "",
        )
    except Exception as exc:  # noqa: BLE001
        sb["last_run"] = {
            "ok": False, "mode": "distributed", "error": str(exc),
        }
        return [f"5300 backfill failed: {type(exc).__name__}: {exc}"]

    sb["last_run"] = result
    if result.get("ok"):
        filled = len(result.get("months_filled") or [])
        rows = int(result.get("rows_written") or 0)
        msgs.append(
            f"5300 backfill (distributed): filled {filled} quarter-end(s), "
            f"wrote {rows} row(s) using earliest-month ratios "
            f"from {dist.get('period')}"
        )
    else:
        msgs.append(
            "5300 backfill (distributed) error: "
            + (result.get("error") or "unknown")
        )
    return msgs


def _is_real_warm_pool_row(s: dict[str, Any]) -> bool:
    """True when a WARM ``pool_settings`` row is a genuine loan pool.

    Filters out the WARM workbook's ACL / Credit-Grade-Deteriorated /
    Grand Total / Hide / Exclude sentinel rows so only real pools seed the
    wizard's Loan Pools step. Mirrors the route-level ``_is_real_pool_row``
    inside ``setup._apply_warm_to_state``.
    """
    nm = (s.get("name") or "").strip().lower()
    if not nm:
        return False
    if nm.startswith("allowance for credit loss") or nm == "allowance":
        return False
    if nm.startswith("credit grade deteriorated"):
        return False
    if (nm.startswith("grand total") or nm.startswith("total")
            or nm.startswith("hide") or nm == "exclude"):
        return False
    return True


def _apply_warm_findings_to_state(
    state: dict[str, Any], warm: dict[str, Any]
) -> list[str]:
    """Populate wizard state from an auto-scanned WARM workbook.

    This is the auto-scan counterpart to ``setup._apply_warm_to_state``
    (which only runs on the manual single-file WARM upload path). Without
    it, a folder scan that *detects* a WARM workbook would set
    ``has_warm_files='yes'`` but leave pools, grades, and identity empty —
    stranding the user on an empty Loan Pools step. Here the WARM workbook
    is treated as the authoritative source for pools / grades on a WARM CU.

    Never raises; returns human-readable messages for the scan report.
    ``warm`` is the ``findings['warm']`` dict
    ``{entry, saved_path, analysis}``.
    """
    msgs: list[str] = []
    analysis = (warm or {}).get("analysis") or {}
    if not analysis.get("ok"):
        return msgs

    # Carry the file locations on the analysis so downstream WARM steps and
    # the WARM -> CM comparison can find the workbook.
    if warm.get("saved_path"):
        analysis.setdefault("saved_path", warm["saved_path"])
    entry = warm.get("entry") or {}
    if entry.get("path"):
        analysis.setdefault("source_path", entry["path"])

    # ----- Identity / economic baseline (fill blanks only) --------------
    bid = analysis.get("baseline_identity") or {}
    if bid:
        if not state.get("credit_union") and bid.get("cu_name"):
            state["credit_union"] = bid["cu_name"]
        if not state.get("charter_number") and bid.get("charter_number"):
            state["charter_number"] = bid["charter_number"]
        econ = state.setdefault("economic_data", {})
        for key in ("state", "county", "unemployment_rate",
                    "foreclosures", "bankruptcies", "population"):
            if not econ.get(key) and bid.get(key):
                econ[key] = bid[key]
        if not state.get("short_name") and state.get("credit_union"):
            state["short_name"] = config_service.slugify(state["credit_union"])
        if bid.get("period_end_date"):
            analysis["as_of_date"] = (
                analysis.get("as_of_date") or bid["period_end_date"]
            )
    if not state.get("credit_union") and analysis.get("cu_name"):
        state["credit_union"] = analysis["cu_name"]
        if not state.get("short_name"):
            state["short_name"] = config_service.slugify(analysis["cu_name"])
    if state.get("credit_union"):
        msgs.append(f"Identity seeded from WARM for {state['credit_union']!r}")

    # ----- Loan-code -> pool map ---------------------------------------
    # Priority: WARM's own code map (cols S/T) > a real code map already
    # derived from a loan-extract sample > a {pool: pool} placeholder from
    # the WARM pool names. Never clobber real raw codes with placeholders.
    code_map = analysis.get("loan_code_pool_map") or {}
    if code_map:
        state["pool_map"] = dict(code_map)
        state["_warm_seeded_pool_map"] = dict(code_map)
        distinct = len({v for v in code_map.values() if v})
        msgs.append(
            f"Loan Code Mapping seeded from WARM 'Grade Ranges & Loan Codes' "
            f"(cols S/T): {len(code_map)} code(s) -> {distinct} pool(s)"
        )
    elif not (state.get("pool_map") or {}) and analysis.get("pools"):
        state["pool_map"] = {p: p for p in analysis["pools"] if p}
        msgs.append(
            f"Loan Code Mapping seeded with {len(state['pool_map'])} WARM "
            "pool name(s) (no raw code map in WARM — refine on the mapping step)"
        )

    # ----- Per-pool settings (the Loan Pools step) — WARM authoritative --
    warm_ps = [
        s for s in (analysis.get("pool_settings") or [])
        if _is_real_warm_pool_row(s)
    ]
    if warm_ps:
        state["pool_settings"] = [dict(s) for s in warm_ps]
        msgs.append(
            f"{len(warm_ps)} loan pool(s) seeded from WARM "
            "'BS CO DQ Data Enter'"
        )

    # ----- ACL balance --------------------------------------------------
    if not state.get("acl_balance") and analysis.get("acl_balance"):
        try:
            state["acl_balance"] = float(analysis["acl_balance"])
        except (TypeError, ValueError):
            pass

    # ----- Credit grades — WARM authoritative ---------------------------
    if analysis.get("grades"):
        state["credit_grades"] = list(analysis["grades"])
        msgs.append(
            f"{len(analysis['grades'])} credit grade(s) seeded from WARM"
        )

    # ----- Monthly balances from WARM 'BS Data' -------------------------
    # A WARM CU's per-pool monthly balance history lives inside the WARM
    # workbook. When the folder scan found no separate monthly-balance file,
    # use that series as a 'manual' monthly_balance source so the report's
    # balance-adjustment logic has data and the wizard's Monthly Balance step
    # is satisfied from WARM alone. Never clobber a real monthly-balance file
    # the scan already wired up (per_year / per_month / single).
    mb_existing = (state.get("monthly_bal") or {}).get("source") or ""
    if mb_existing.strip() not in (
        "per_year", "per_month", "single", "single_workbook", "manual"
    ):
        pmb = analysis.get("pool_monthly_balances") or {}
        pool_rows = pmb.get("pools") or []
        entries: dict[str, dict[str, float]] = {}
        all_dates: set[str] = set()
        for prow in pool_rows:
            name = (prow.get("name") or "").strip()
            if not name or not _is_real_warm_pool_row({"name": name}):
                continue
            row_map: dict[str, float] = {}
            for series in ("history", "current_quarter"):
                for pt in (prow.get(series) or []):
                    d = (pt.get("date") or "").strip()
                    if not d:
                        continue
                    try:
                        row_map[d] = float(pt.get("balance") or 0.0)
                    except (TypeError, ValueError):
                        continue
                    all_dates.add(d)
            if row_map:
                entries[name] = row_map
        if entries:
            mbst = state.setdefault("monthly_bal", {})
            mbst["source"] = "manual"
            mbst["manual_entries"] = entries
            mbst["manual_months"] = sorted(all_dates)
            mbst["parsed_pool_labels"] = list(entries.keys())
            mbst["_warm_derived"] = True
            msgs.append(
                f"Monthly balances seeded from WARM 'BS Data': {len(entries)} "
                f"pool(s) x {len(all_dates)} month(s) (manual source)"
            )

    state["warm"] = analysis
    return msgs


# Process-level guard so the learned-store warm-start runs at most once per
# worker (idempotent anyway, but this avoids re-globbing client_configs on
# every scan).
_LEARNED_STORE_WARMED = False


def _ensure_learned_store_warm(workspace_root: str | Path) -> None:
    """Backfill the cross-CU learned column store from validated client
    configs, once per process. Fully best-effort — never raises.
    """
    global _LEARNED_STORE_WARMED
    if _LEARNED_STORE_WARMED:
        return
    _LEARNED_STORE_WARMED = True
    try:
        cfg_dir = Path(workspace_root) / "client_configs"
        if cfg_dir.is_dir():
            column_mapping_suggestions.backfill_from_config_dir(str(cfg_dir))
    except Exception:  # noqa: BLE001
        pass


def apply_findings_to_state(
    state: dict[str, Any],
    findings: dict[str, Any],
    workspace_root: str | Path,
) -> dict[str, list[str]]:
    """Mutate *state* with everything *findings* can supply.

    Returns a dict ``{step_key: [messages]}`` describing what was filled
    on each wizard step. Step keys mirror the keys in
    ``WIZARD_STEPS_NO_WARM`` / ``WIZARD_STEPS_WARM`` so the stepper UI
    can render per-step status badges off the same dict.
    """
    report: dict[str, list[str]] = {}

    # Warm-start the cross-CU learned column store from the already-validated
    # client configs (idempotent, once per process) so a fresh CU's cryptic
    # core headers are recognised from prior onboardings. Best-effort.
    _ensure_learned_store_warm(workspace_root)

    if not findings.get("ok"):
        report.setdefault("identity", []).append(
            f"Scan failed: {findings.get('error') or 'unknown error'}"
        )
        return report

    folder = findings.get("folder") or ""

    # ----- Sample / columns / pools / files steps ----------------------
    sample = findings.get("sample")
    if sample:
        rep = findings.get("classification", {}).get("loan_data_files") or []
        name = (rep and rep[0].get("name")) or ""
        msgs = _apply_sample_to_state(state, sample, name)
        report.setdefault("sample", []).extend(msgs)
        # column_mappings and file_pattern/date_pattern affect multiple
        # steps; mirror to those keys for stepper UX.
        if any("column_mappings" in m for m in msgs):
            report.setdefault("columns", []).append("column_mappings auto-filled from sample headers")
        if any("file_pattern" in m or "date_pattern" in m for m in msgs):
            report.setdefault("files", []).append("file_pattern and date_pattern auto-filled")

    # Bulk-stage every analysed loan-data extract into
    # ``sample_uploads.loan_data_files`` so Step 13 (Column Mappings)
    # renders dropdowns immediately without the user having to re-upload
    # each file on Step 12. Mirrors ``_stage_hist_scan_files`` for CO/Recov.
    extract_msgs = _stage_loan_data_extracts(
        state, findings.get("sample_extracts") or []
    )
    if extract_msgs:
        report.setdefault("columns", []).extend(extract_msgs)
        report.setdefault("sample", []).extend(extract_msgs)
        if any("pool_map seeded" in m for m in msgs):
            report.setdefault("pools", []).append(
                "pool_map seeded with detected codes (all 'Ignore' — rename to your real pool names)"
            )

    # ----- Delivered code -> pool map: pre-fill pool_map VALUES ----------
    # Runs AFTER the sample seeded pool_map with the raw codes, so the code
    # column can be matched by overlap. Proposals only (kept under review).
    code_map_msgs = _apply_code_map_to_pool_map(
        state, findings.get("code_map_files") or []
    )
    if code_map_msgs:
        report.setdefault("pools", []).extend(code_map_msgs)
        report.setdefault("loan_pools", []).extend(code_map_msgs)

    # ----- monthly_bal step (annual or per-month) ----------------------
    if findings.get("annual_bal"):
        msgs = _apply_annual_bal_to_state(state, findings["annual_bal"])
        report.setdefault("monthly_bal", []).extend(msgs)
        # When we found annual workbooks, surface that on Step 3 historical
        # too because that step has a related "hist_balance_source" toggle.
        state["hist_balance_source"] = "annual_balance_sheets"
        report.setdefault("historical", []).append(
            "hist_balance_source set to annual_balance_sheets"
        )
    elif findings.get("monthly_bal"):
        msgs = _apply_monthly_bal_to_state(state, findings["monthly_bal"])
        report.setdefault("monthly_bal", []).extend(msgs)
        state["hist_balance_source"] = "monthly_balance_sheets"
        report.setdefault("historical", []).append(
            "hist_balance_source set to monthly_balance_sheets"
        )
    elif findings.get("single_hist_bal"):
        shb = findings["single_hist_bal"]
        msgs = _apply_single_hist_bal_to_state(state, shb)
        report.setdefault("monthly_bal", []).extend(msgs)
        state["hist_balance_source"] = "single_workbook"
        report.setdefault("historical", []).append(
            "hist_balance_source set to single_workbook"
        )
        # Seed Step 2 pool_settings + Step 8 pool_map from the workbook's
        # pool labels when no consolidated 'Balance Sheets <CU>'
        # pool_seed is also present (flat hist-bal workbooks ARE the
        # authoritative pool list for CUs that ship them, e.g. TCP CU's
        # ``Historical Balance Sheets-TCP CU 65384.xlsx``).
        if not findings.get("pool_seed"):
            shb_entry = shb.get("entry") or {}
            shb_src = shb_entry.get("name") or "single historical balance workbook"
            flat_report = _apply_flat_pool_labels_to_state(
                state, shb.get("pool_labels") or [], shb_src
            )
            for key, fmsgs in flat_report.items():
                report.setdefault(key, []).extend(fmsgs)
            # Mirror loan_pools msgs to "pools" stepper bucket too.
            if "loan_pools" in flat_report:
                report.setdefault("pools", []).extend(
                    flat_report["loan_pools"]
                )

    # ----- Consolidated pool-grouping balance workbook --------------------
    # (Vizo Financial 'Balance Sheets <CU>.xlsx' carries the authoritative
    # pool list AND the sub-category -> pool mapping; runs INDEPENDENTLY of
    # annual / per-month balance findings so a CU can have both.)
    if findings.get("pool_seed"):
        ps_report = _apply_pool_seed_to_state(state, findings["pool_seed"])
        for key, msgs in ps_report.items():
            report.setdefault(key, []).extend(msgs)
        # Mirror to "pools" stepper bucket as well so the badge shows on
        # Step 2 (no-WARM path) regardless of how the step list keys it.
        if "loan_pools" in ps_report:
            report.setdefault("pools", []).extend(ps_report["loan_pools"])

    # ----- WARM presence ------------------------------------------------
    if findings.get("warm"):
        state["has_warm_files"] = "yes"
        report.setdefault("warm", []).append(
            "WARM workbook detected — WARM-path wizard steps will be used."
        )
        # Apply the WARM workbook's pools, grades, and identity so the
        # wizard doesn't strand the user on an empty Loan Pools step. WARM
        # is authoritative for pools/grades on a WARM CU.
        warm_msgs = _apply_warm_findings_to_state(state, findings["warm"])
        if warm_msgs:
            report.setdefault("warm", []).extend(warm_msgs)
            # Mirror pool/grade seeding onto the steps that render them so
            # the stepper badges light up on both wizard paths.
            report.setdefault("loan_pools", []).extend(
                m for m in warm_msgs if "pool" in m.lower()
            )
            report.setdefault("pools", []).extend(
                m for m in warm_msgs if "Loan Code Mapping" in m
            )
            report.setdefault("grades", []).extend(
                m for m in warm_msgs if "grade" in m.lower()
            )
            report.setdefault("monthly_bal", []).extend(
                m for m in warm_msgs if "Monthly balances" in m
            )
    else:
        # Don't override an explicit prior answer; only set when unanswered.
        if state.get("has_warm_files") is None:
            state["has_warm_files"] = "no"
            report.setdefault("loan_pools", []).append(
                "No WARM workbook found — non-WARM wizard path active."
            )

    # ----- Companion files: impaired, credit pull, CO, recov ----------
    cls = findings.get("classification") or {}
    if findings.get("impaired_file"):
        msgs = _apply_companion_file(
            state, "impaired", findings["impaired_file"],
            "impaired_files", "impaired"
        )
        report.setdefault("impaired", []).extend(msgs)
    if findings.get("credit_pull_file"):
        msgs = _apply_credit_pull_to_state(state, findings["credit_pull_file"])
        report.setdefault("credit_pull", []).extend(msgs)
    if findings.get("co_file"):
        msgs = _apply_companion_file(
            state, "charge-off", findings["co_file"],
            "co_files", "co"
        )
        report.setdefault("co_recov", []).extend(msgs)
        report.setdefault("co_history", []).extend(msgs)
    # Stage ALL monthly CO files into hist_scan.monthly_co_files so Step 5
    # (Historical Charge-Offs) renders the per-file table and the
    # `monthly_co_recov_aggregator` can aggregate them on the next visit.
    bulk_co_msgs = _stage_hist_scan_files(
        state, cls.get("co_files") or [], "monthly_co_files", "co"
    )
    if bulk_co_msgs:
        report.setdefault("co_history", []).extend(bulk_co_msgs)
    if findings.get("recov_file"):
        msgs = _apply_companion_file(
            state, "recovery", findings["recov_file"],
            "recov_files", "recov"
        )
        report.setdefault("co_recov", []).extend(msgs)
        report.setdefault("recov_history", []).extend(msgs)
    bulk_recov_msgs = _stage_hist_scan_files(
        state, cls.get("recov_files") or [], "monthly_recov_files", "recov"
    )
    if bulk_recov_msgs:
        report.setdefault("recov_history", []).extend(bulk_recov_msgs)

    # Flip Step 5/6 source picker to "monthly_files" when per-month
    # files were staged so the user immediately sees them on those
    # steps instead of the empty single-workbook upload field.
    flip_msgs = _auto_flip_hist_source(state)
    for m in flip_msgs:
        if "Charge-off" in m:
            report.setdefault("co_history", []).append(m)
        elif "Recovery" in m:
            report.setdefault("recov_history", []).append(m)

    # ----- Folder-derived defaults -------------------------------------
    _set_reports_default_from_folder(state, folder)
    report.setdefault("reports", []).append(
        "reports defaults applied (Vizo-client convention if folder is under "
        "Vizo Financial; else TCT)"
    )

    _apply_mgmt_adj_defaults(state)
    report.setdefault("mgmt_adj", []).append(
        "mgmt_adj baselines applied (ltv_baseline=0.9, probability_factor=0.35)"
    )

    # ----- Raw_Uploads pointer -----------------------------------------
    msgs = _apply_data_directory(state, workspace_root)
    if msgs:
        report.setdefault("identity", []).extend(msgs)

    # ----- Distributed 5300 backfill (auto, non-fatal) -----------------
    # Runs after everything else has populated hist_balance_source +
    # monthly_bal.pool_map so the earliest-month distribution can be
    # computed. Skipped gracefully if any precondition is missing.
    try:
        bf_msgs = _auto_run_distributed_backfill(state)
    except Exception as exc:  # noqa: BLE001
        bf_msgs = [f"5300 backfill skipped: {type(exc).__name__}: {exc}"]
    if bf_msgs:
        report.setdefault("historical", []).extend(bf_msgs)

    # ----- Persist the report on state so UI can read it back ----------
    state["_auto_scan_completed"] = True
    state["_auto_scan_folder"] = folder
    state["_auto_scan_report"] = report
    state["_auto_scan_messages"] = list(findings.get("messages") or [])
    state["_auto_scan_errors"] = list(findings.get("errors") or [])

    return report


# --------------------------------------------------------------------------
# HIL detection
# --------------------------------------------------------------------------


def compute_hil_needs(state: dict[str, Any]) -> list[dict[str, Any]]:
    """Walk *state* and return ordered list of steps still needing user input.

    Each entry: ``{step_key, severity, reasons: list[str]}`` where severity
    is ``"required"`` (the wizard cannot proceed without user input) or
    ``"recommended"`` (we filled a default but the user should review).

    The order matches the step list ordering so callers can pick the
    first entry as "next step to visit".
    """
    needs: list[dict[str, Any]] = []

    def add(key: str, severity: str, *reasons: str) -> None:
        needs.append({
            "step_key": key,
            "severity": severity,
            "reasons": list(reasons),
        })

    # identity — required CU name + report period
    if not state.get("credit_union"):
        add("identity", "required", "Credit Union name is blank")
    if not state.get("report_period"):
        add("identity", "recommended", "report_period (target YYYY-MM) is blank")

    # economic — state + county
    econ = state.get("economic_data") or {}
    if not econ.get("state"):
        add("economic", "required", "Economic state is blank")
    elif not econ.get("county"):
        add("economic", "recommended", "Economic county is blank")

    # loan_pools / warm (Step 2) — ensure the wizard lands here after
    # Step 1 auto-scan so the user can confirm or expand the pool list
    # BEFORE mapping loan codes (Step 3) into those pools. Suppressed
    # once the user explicitly clicks Save & Next on Step 2 (recorded
    # in _user_completed_steps via the Phase 9.9 mechanism). Both
    # keys are emitted; first_hil_step_key filters to whichever step
    # belongs to the active step list (loan_pools = NO_WARM path,
    # warm = WARM path).
    user_done = set(state.get("_user_completed_steps") or [])
    pool_settings = state.get("pool_settings") or []
    has_named_pool = any((p.get("name") or "").strip() for p in pool_settings)
    auto_scan_done = bool(state.get("_auto_scan_completed"))
    if "loan_pools" not in user_done and "warm" not in user_done:
        if not has_named_pool:
            add("loan_pools", "required",
                "No loan pools defined yet — add at least one pool before "
                "mapping loan codes")
            add("warm", "required",
                "No loan pools defined yet — upload a WARM file or add "
                "pools manually before mapping loan codes")
        elif auto_scan_done:
            n = len(pool_settings)
            add("loan_pools", "recommended",
                f"{n} loan pool(s) auto-detected — review names, ACL months, "
                "and risk-rated flags before continuing")
            add("warm", "recommended",
                f"{n} loan pool(s) auto-detected — review names, ACL months, "
                "and risk-rated flags before continuing")

    # pools — ALL seeded codes default to 'Ignore' which means user has not
    # renamed any. That's the most common 'must review' step after a scan.
    pm = state.get("pool_map") or {}
    if pm:
        ignore_count = sum(1 for v in pm.values() if str(v).lower() == "ignore")
        named_count = len(pm) - ignore_count
        if ignore_count == len(pm):
            add("pools", "required",
                f"All {len(pm)} detected loan codes are set to 'Ignore' — "
                "rename each to your real pool names")
        elif ignore_count:
            add("pools", "recommended",
                f"{ignore_count} of {len(pm)} loan codes still set to 'Ignore'")
        elif state.get("_code_map_proposed_count") and "pools" not in user_done:
            add("pools", "recommended",
                f"{int(state['_code_map_proposed_count'])} loan code(s) were "
                "pre-mapped to pools from the delivered code-map workbook — "
                "verify the assignments (a delivered map can contain errors)")
    else:
        add("pools", "recommended", "No loan codes detected — upload a sample on the Loan Data Extracts step")

    # monthly_bal — at least one source must be picked AND mapped
    mb = state.get("monthly_bal") or {}
    mb_source = (mb.get("source") or "").strip()
    if mb_source in ("per_year", "per_month", "single", "single_workbook"):
        labels = mb.get("parsed_pool_labels") or []
        mapping = mb.get("pool_map") or {}
        if labels and not mapping:
            add("monthly_bal", "required",
                f"{len(labels)} balance-sheet labels detected but none mapped to pools")
        elif labels:
            unmapped = [lab for lab in labels if not mapping.get(lab)]
            if unmapped:
                add("monthly_bal", "recommended",
                    f"{len(unmapped)} balance-sheet label(s) still unmapped")
    elif mb_source == "manual":
        # Manual ACL entry mode — no labels to map; treated as
        # complete once the user has picked the mode.
        pass
    else:
        # No source picked yet.
        add("monthly_bal", "required", "Monthly Balance source not yet chosen")

    # grades — only HIL when user hasn't customised reserve_rate yet.
    grades = state.get("credit_grades") or []
    if not grades:
        add("grades", "required", "Credit grades are empty")
    else:
        defaults = config_service.DEFAULT_CREDIT_GRADES
        # Defaults dict shape: list of {label, min_fico, max_fico, reserve_rate}
        is_default = (len(grades) == len(defaults)) and all(
            grades[i].get("reserve_rate") == defaults[i].get("reserve_rate")
            for i in range(len(grades))
        )
        if is_default:
            add("grades", "recommended",
                "Credit grades using factory default reserve rates — review "
                "against your analyst's CECL methodology")

    # credit_pull — recommended unless the user explicitly opted out.
    cp = state.get("credit_pull") or {}
    if not cp.get("uploaded_filename") and not cp.get("use_configured_report"):
        # Was a credit-pull file found in the auto-scan?
        cp_file_present = bool(
            (state.get("sample_uploads") or {}).get("credit_pull_files")
        )
        if not cp_file_present:
            add("credit_pull", "recommended",
                "No credit-pull file found for the target period — current "
                "FICO will equal original FICO (no grade movement)")

    # impaired — recommended review when a file was staged.
    if (state.get("sample_uploads") or {}).get("impaired_files"):
        add("impaired", "recommended",
            "Impaired-loans file staged — review parsed rows before final save")

    # sample / loan extract — required when the auto-scan completed but
    # found NO loan-data extract. Without a loan file the column_mappings,
    # file_pattern, and pool_map remain PLACEHOLDER DEFAULTS that pass the
    # "blank" checks below (member_number='MEMBER_ID', file_pattern=
    # 'LOANDATA.*') yet won't match the CU's real data — so those steps
    # would otherwise show no HIL and mislead the user into thinking the
    # values were derived. Scoped to the non-WARM path (WARM CUs derive
    # column mappings from the WARM Data tab, not a loan sample).
    if state.get("_auto_scan_completed") and state.get("has_warm_files") != "yes":
        sample_staged = bool(
            (state.get("sample_uploads") or {}).get("loan_data_files")
        )
        if not sample_staged:
            add("sample", "required",
                "No loan-data extract was found in the scanned folder — "
                "upload the monthly loan file so column mappings, file "
                "pattern, and loan-code pools can be derived (they are "
                "currently placeholder defaults, not values from your data)")

    # files — recommended when we auto-filled from sample
    if not state.get("file_pattern"):
        add("files", "required", "File pattern is blank")

    # columns — required when member_number unmapped, plus detection of
    # factory PLACEHOLDER sentinels that survived an auto-scan (the sample's
    # column-suggester could not match them to a real header). Placeholders
    # are non-blank so they pass a naive "unmapped" test, yet they are NOT
    # real columns. loan_pool_code is the most damaging: a placeholder there
    # yields an EMPTY pool_map (no loan codes extracted from the sample).
    cm = state.get("column_mappings") or {}
    _placeholder_cols = {
        "MEMBER_ID", "BALANCE", "FICO_SCORE", "LOAN_TYPE",
        "DQ_DAYS", "INT_RATE", "OPEN_DATE", "ORIG_AMT",
    }
    _scanned = bool(state.get("_auto_scan_completed"))
    _mnum = cm.get("member_number")
    if not _mnum or (_scanned and _mnum in _placeholder_cols):
        add("columns", "required", "member_number column is unmapped")
    if _scanned:
        _lpc = cm.get("loan_pool_code")
        if isinstance(_lpc, str) and _lpc in _placeholder_cols:
            add("columns", "required",
                "loan_pool_code is still the placeholder 'LOAN_TYPE' — the "
                "loan file's pool/collateral-code column was not auto-detected; "
                "map it on the Column Mappings step (until then pool_map is empty)")
        _other = sorted(
            k for k, v in cm.items()
            if k not in ("member_number", "loan_pool_code")
            and isinstance(v, str) and v in _placeholder_cols
        )
        if _other:
            add("columns", "recommended",
                "Columns still at placeholder defaults (not matched to a real "
                "header): " + ", ".join(_other))

    # mgmt_adj — recommended review (we applied baseline defaults)
    if state.get("_auto_scan_completed"):
        add("mgmt_adj", "recommended",
            "ltv_baseline=0.9 and probability_factor=0.35 applied — review "
            "against your analyst's CECL methodology")

    # review — always shown last; the user must click Save to write YAML.
    add("review", "required", "Confirm and save the final YAML")

    return needs


def first_hil_step_key(
    state: dict[str, Any],
    step_list: list[tuple[str, str]],
    severity: str | None = None,
) -> str | None:
    """Pick the first HIL step that exists in *step_list* (in display order).

    When ``severity`` is given (e.g. ``"required"``), only HIL entries of that
    severity are considered. This lets the post-auto-scan router drive the
    wizard straight through *recommended* review checkpoints (pools, grades,
    mgmt-adj, ...) and stop only on genuinely blocking gaps that cannot be
    auto-derived — landing on the Review step when nothing required remains.
    """
    needs = compute_hil_needs(state)
    if severity:
        need_keys = {n["step_key"] for n in needs if n["severity"] == severity}
    else:
        need_keys = {n["step_key"] for n in needs}
    for key, _label in step_list:
        if key in need_keys:
            return key
    return None


def compute_step_completion(state: dict[str, Any]) -> set[str]:
    """Return wizard step keys whose saved data passes a basic completeness
    check.

    Used by the stepper to display a green ✓ on steps the user has
    filled in -- whether by Step 1 auto-scan, by manual edit on the
    page, or by a mix of both. A step that is "complete" here gets the
    ``auto`` badge in the UI, even if it still has *recommended* HIL
    entries (the HIL banner can flag those as optional review items).

    Steps with *required* HIL entries (e.g. CU name blank) will fail
    their predicate here and continue to render as ``hil``.

    The set covers every step key from any of the three step lists
    (WARM, NO_WARM, SCALE). The caller filters to its active list.
    """
    done: set[str] = set()

    # ----- WARM auto-derive -----
    # The setup wizard's WARM upload runs the validated resolver path
    # (warm_autoderive) which fully derives these steps from the WARM
    # workbook. Mark them complete so the stepper shows them green
    # (review-only) and the HIL banner steers the user to the interactive
    # steps. User edits still win (they can open any step to change it).
    done |= set(state.get("_warm_autoderived_steps") or [])

    # ----- identity -----
    if (state.get("credit_union") or "").strip() and (state.get("report_period") or "").strip():
        done.add("identity")

    # ----- loan_pools / warm (Step 2 on either path) -----
    pool_settings = state.get("pool_settings") or []
    has_named_pool = any((p.get("name") or "").strip() for p in pool_settings)
    if has_named_pool:
        done.add("loan_pools")
        # WARM path's "warm" step is the upload checkpoint -- it's
        # complete once the WARM file is parsed and pools were
        # extracted. Treat a non-empty pool_settings list as proof.
        warm = state.get("warm") or {}
        if warm.get("saved_path") or warm.get("balance_titles") or has_named_pool:
            done.add("warm")

    # ----- pools (Step 3 -- Loan Code Mapping) -----
    pm = state.get("pool_map") or {}
    if pm:
        # At least one code mapped to a real pool name (anything other
        # than blank or 'Ignore').
        if any(
            (str(v).strip() and str(v).strip().lower() != "ignore")
            for v in pm.values()
        ):
            done.add("pools")
    else:
        # No codes detected -- nothing to do, treat as complete so the
        # user isn't stuck with a perpetual ⚠ on an empty step.
        done.add("pools")

    # ----- balances (WARM path, Step 4 -- Balance Titles) -----
    btm = state.get("balance_title_map") or {}
    if btm:
        done.add("balances")

    # ----- baseline / historical -----
    warm = state.get("warm") or {}
    if warm.get("imported_baseline") or state.get("hist_balance_source"):
        done.add("baseline")
        done.add("historical")
    # Also count an explicit Solr-backfill mode as completion of the
    # historical step on the no-WARM path.
    he = state.get("hist_extracts") or {}
    if he.get("solr_backfill") and isinstance(he["solr_backfill"], dict) and he["solr_backfill"].get("mode"):
        done.add("historical")

    # ----- co_history / recov_history / dq_hist (NO_WARM Steps 5/6/7) -----
    # CRITICAL: Step 5 (Historical Charge-Offs) and Step 6 (Recoveries)
    # render their per-file tables from ``hist_scan.monthly_co_files`` /
    # ``hist_scan.monthly_recov_files`` -- NOT from ``sample_uploads.*``.
    # ``sample_uploads.co_files`` is only the single representative file
    # the auto-scan stages for sample analysis. If the bulk-stage for
    # that section never ran (old draft from before Phase 9.4, or only
    # one side of a combined CO+Recov classification got cross-listed),
    # the step page renders empty even though the breadcrumb would
    # otherwise show green. When auto-scan completed, REQUIRE the
    # corresponding hist_scan list (the real source for the page); when
    # auto-scan never ran (pure manual draft), accept the sample-uploads
    # / state-level path as before.
    sample_uploads = state.get("sample_uploads") or {}
    hist_scan = state.get("hist_scan") or {}
    auto_scan_done = bool(state.get("_auto_scan_completed"))
    has_monthly_co = bool(hist_scan.get("monthly_co_files") or hist_scan.get("co_files"))
    has_monthly_recov = bool(hist_scan.get("monthly_recov_files") or hist_scan.get("recov_files"))
    if auto_scan_done:
        if has_monthly_co:
            done.add("co_history")
        if has_monthly_recov:
            done.add("recov_history")
    else:
        if sample_uploads.get("co_files") or state.get("co_files") or has_monthly_co:
            done.add("co_history")
        if sample_uploads.get("recov_files") or state.get("recov_files") or has_monthly_recov:
            done.add("recov_history")
    dq = state.get("dq_hist") or {}
    if state.get("dq_files") or dq.get("uploaded") or dq.get("source"):
        done.add("dq_hist")

    # ----- monthly_bal -----
    mb = state.get("monthly_bal") or {}
    mb_source = (mb.get("source") or "").strip()
    if mb_source:
        labels = mb.get("parsed_pool_labels") or []
        mapping = mb.get("pool_map") or {}
        if labels and all(mapping.get(lab) for lab in labels):
            done.add("monthly_bal")
        elif not labels and mb_source in ("manual", "single", "single_workbook"):
            # No labels were parsed (e.g. manual ACL entry mode) --
            # treat as complete once the user has picked a source.
            done.add("monthly_bal")

    # ----- grades -----
    if state.get("credit_grades"):
        done.add("grades")

    # ----- credit_pull -----
    cp = state.get("credit_pull") or {}
    if cp.get("uploaded_filename") or cp.get("skipped") or cp.get("use_configured_report"):
        done.add("credit_pull")

    # ----- orig_score -- informational, always considered complete -----
    done.add("orig_score")

    # ----- sample (Loan Data Extract(s)) -----
    extracts = state.get("loan_data_extracts") or []
    loan_files = (
        state.get("loan_data_files")
        or (state.get("sample_uploads") or {}).get("loan_data_files")
        or []
    )
    if (extracts and any(e.get("file_pattern") for e in extracts)) or loan_files:
        done.add("sample")

    # ----- columns (Column Mappings) -----
    cm = state.get("column_mappings") or {}
    if cm.get("member_number") and cm.get("current_balance"):
        done.add("columns")

    # ----- balance_check / co_recov -- informational sanity steps -----
    done.add("balance_check")
    done.add("co_recov")

    # ----- impaired -----
    if sample_uploads.get("impaired_files") or (state.get("impaired") or {}).get("skipped"):
        done.add("impaired")

    # ----- files (File Format) -----
    if (state.get("file_pattern") or "").strip() and (state.get("date_format") or "").strip():
        done.add("files")

    # ----- economic -----
    econ = state.get("economic_data") or {}
    if (econ.get("state") or "").strip() and (econ.get("county") or "").strip():
        done.add("economic")

    # ----- mgmt_adj -----
    ma = state.get("mgmt_adj") or {}
    if ma.get("ltv_baseline") is not None and ma.get("probability_factor") is not None:
        done.add("mgmt_adj")

    # ----- reports -----
    rp = state.get("reports") or {}
    if any(bool(v) for v in rp.values()):
        done.add("reports")

    # ----- review -- never auto-complete; user must click Save -----

    return done
