"""Parser & calculator for the standard "Impaired Loans" workbook.

The CECL Impaired Loans tab in a CU's WARM workbook follows the same
shape across every credit union (Bridgeton, Franklin, Tongass, ...):

* Cells **A1**, **A2**, **A3/B3**: CU name, "Impaired Loans" title,
  "Report for Period Ending" label + date.
* Rows **5..19** (or **7..25**, depending on CU): impairment-type
  names in column A and provision percentages in column B. Rows whose
  name is ``HIDE`` are placeholders.
* Rows **22-23**: section banner ("Data Entry" / "Calculation" / "TCT")
  and the per-loan column headers.
* Rows **35..343**: the data block. The *first* 10 columns
  (``A:J`` -- Impairment Type, Member #, Loan Suffix, Loan Type,
  Current Balance, Days Delinquent, Balance at Other Lender,
  Collateral Value, Allowance Provided, Notes) are CU-entered data.
  Columns ``K:Q`` are spreadsheet calculations and ``R:S`` are
  XLOOKUP-from-loan-data values.

This module:

1. ``parse_file(path)`` -- pulls A1/A2/A3, the impairment types
   table, and the **input** rows from A:J only. The calculation /
   lookup columns are deliberately ignored -- we recompute them.
2. ``compute_calculations(rows, types, dq_ranges)`` -- recreates the
   K:Q formulas in Python.
3. ``lookup_from_loan_data(rows, state)`` -- recreates the R:S/T/U
   XLOOKUPs by reading the latest sample loan-data file and applying
   the wizard's pool_map and credit-grade mappings.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _norm(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _digits(v: Any) -> str:
    """Digits-only form of a member / suffix / account value.

    Strips a trailing Excel ``.0`` (integer-valued floats) and any
    non-digit characters so ``19238``, ``19238.0`` and ``'19238-'`` all
    collapse to the comparable digit string ``'19238'``.
    """
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return "".join(ch for ch in s if ch.isdigit())


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    if s == "":
        return None
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


def _to_pct(v: Any) -> float | None:
    """Coerce a Provision Percentage cell to a float (0.5 == 50%).

    Accepts numbers, percent strings (``50%``), and the sentinel
    string ``Variable`` (returns None -- meaning the row falls back
    to the DQ-range table).
    """
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip()
    if not s or s.lower() == "variable":
        return None
    if s.endswith("%"):
        try:
            return float(s[:-1].strip()) / 100.0
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def _coerce(v: Any) -> Any:
    """JSON/Flask-Session-friendly value."""
    if v is None or isinstance(v, (int, float, str, bool)):
        return v
    try:
        import datetime as _dt
        if isinstance(v, (_dt.datetime, _dt.date)):
            return v.isoformat()
    except Exception:  # noqa: BLE001
        pass
    return str(v)


# Canonical key names for the 10 input columns (A:J).
INPUT_FIELDS: tuple[str, ...] = (
    "impairment_type",       # A
    "member",                # B
    "suffix",                # C
    "loan_type",             # D
    "current_balance",       # E
    "days_dq",               # F
    "other_lender_balance",  # G
    "collateral_value",      # H
    "allowance_provided",    # I
    "notes",                 # J
)


# ---------------------------------------------------------------------------
# parse_file
# ---------------------------------------------------------------------------


def parse_file(filepath: str | Path) -> dict[str, Any]:
    """Parse an Impaired Loans workbook.

    Returns::
        {
            "ok": bool, "error": str|None,
            "filename": str,
            "cu_name": str,
            "period_ending": str|None,
            "impairment_types": [{"name": str, "provision_pct": float|None}, ...],
            "data_rows": [{INPUT_FIELDS keys -> raw values}, ...],
            "summary_row": int|None,
            "data_header_row": int|None,
        }
    """
    out: dict[str, Any] = {
        "ok": False, "error": None,
        "filename": Path(str(filepath)).name,
        "cu_name": "", "period_ending": None,
        "impairment_types": [],
        "data_rows": [],
        "summary_row": None, "data_header_row": None,
    }
    p = Path(str(filepath))
    if not p.exists():
        out["error"] = f"File not found: {filepath}"
        return out

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        out["error"] = f"openpyxl not available: {exc}"
        return out

    try:
        wb = load_workbook(p, data_only=True)
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Could not open workbook: {exc}"
        return out

    ws = None
    for sn in wb.sheetnames:
        if sn.strip().lower() == "impaired loans":
            ws = wb[sn]
            break
    if ws is None:
        for sn in wb.sheetnames:
            if "impair" in sn.lower() and "pivot" not in sn.lower():
                ws = wb[sn]
                break
    # Phase 9.36b — mirror the tab-name ladder in
    # ``generate_report.load_standalone_impaired``: some CUs (e.g. TCP
    # 2026-06) ship the workbook with the impaired data on a tab named
    # ``Sheet2`` and the title "Impaired Loans" only in cell A1/A2/A3.
    # Fall through to header-cell + largest-tab detection so the wizard
    # run-time verification page can still parse those workbooks.
    if ws is None:
        for sn in wb.sheetnames:
            _ws = wb[sn]
            for _r in range(1, 4):
                _v = _ws.cell(row=_r, column=1).value
                if _v and "impaired" in str(_v).lower():
                    ws = _ws
                    break
            if ws is not None:
                break
    if ws is None:
        _candidates = [
            (wb[sn].max_row * wb[sn].max_column, sn)
            for sn in wb.sheetnames
            if "instruction" not in sn.strip().lower()
            and "help" not in sn.strip().lower()
            and "readme" not in sn.strip().lower()
        ]
        _candidates.sort(reverse=True)
        if _candidates:
            ws = wb[_candidates[0][1]]
    if ws is None:
        wb.close()
        out["error"] = "No 'Impaired Loans' worksheet found."
        return out

    max_row = ws.max_row or 0

    # --- A1 / A3 metadata ---
    out["cu_name"] = _norm(ws.cell(row=1, column=1).value)
    a3 = _norm(ws.cell(row=3, column=1).value)
    b3 = ws.cell(row=3, column=2).value
    if "period" in a3.lower() and b3 is not None:
        try:
            import datetime as _dt
            if isinstance(b3, (_dt.datetime, _dt.date)):
                out["period_ending"] = b3.isoformat()
            else:
                out["period_ending"] = str(b3).strip()
        except Exception:  # noqa: BLE001
            out["period_ending"] = str(b3).strip()

    # --- Locate the two "Impairment Type" header rows ---
    summary_row = None
    data_header_row = None
    for r in range(1, min(max_row, 60) + 1):
        a = _norm(ws.cell(row=r, column=1).value).lower()
        b = _norm(ws.cell(row=r, column=2).value).lower()
        if a == "impairment type":
            if b.startswith("member"):
                data_header_row = data_header_row or r
            elif b.startswith("provision"):
                summary_row = summary_row or r
            elif summary_row is None:
                summary_row = r
    out["summary_row"] = summary_row
    out["data_header_row"] = data_header_row

    # --- Impairment types (rows below summary header in cols A & B) ---
    if summary_row:
        r = summary_row + 1
        blanks = 0
        while r <= max_row:
            if data_header_row and r >= data_header_row - 2:
                break
            name = _norm(ws.cell(row=r, column=1).value)
            pct_raw = ws.cell(row=r, column=2).value
            if name.lower() == "data entry":
                break
            if not name:
                blanks += 1
                if blanks >= 2:
                    break
                r += 1
                continue
            blanks = 0
            if name.upper() == "HIDE":
                r += 1
                continue
            out["impairment_types"].append({
                "name": name,
                "provision_pct": _to_pct(pct_raw),
            })
            r += 1

    # --- Data rows: cells A:J only ---
    if data_header_row:
        # The Bridgeton template has 11 placeholder rows (24..34) above
        # the actual data. They have empty A:J but populated R/S formulas.
        # We just skip any row whose first 4 cells are blank.
        for r in range(data_header_row + 1, max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 11)]
            if not any(_norm(v) for v in row_vals[:4]):
                continue
            row: dict[str, Any] = {}
            for key, raw in zip(INPUT_FIELDS, row_vals):
                if key in ("current_balance", "other_lender_balance",
                          "collateral_value", "allowance_provided"):
                    row[key] = _to_float(raw)
                elif key == "days_dq":
                    f = _to_float(raw)
                    row[key] = int(f) if f is not None else None
                elif key in ("member", "suffix"):
                    row[key] = _coerce(raw)
                else:
                    row[key] = _norm(raw) if raw is not None else ""
            out["data_rows"].append(row)

    wb.close()
    out["ok"] = True
    return out


# ---------------------------------------------------------------------------
# compute_calculations -- K:Q
# ---------------------------------------------------------------------------


def _member_suffix_key(member: Any, suffix: Any) -> str:
    """Build the ``B&"-"&C`` key used by the spreadsheet's XLOOKUPs."""
    m = "" if member is None else str(member).strip()
    s = "" if suffix is None else str(suffix).strip()
    if m == "" and s == "":
        return ""
    # Strip a trailing ".0" on integer-valued floats (Excel idiosyncrasy).
    for var in ("m", "s"):
        v = locals()[var]
        if v.endswith(".0"):
            try:
                if float(v) == int(float(v)):
                    locals()[var] = str(int(float(v)))
            except ValueError:
                pass
    # The locals() trick above doesn't actually rebind in CPython, redo:
    if m.endswith(".0"):
        try:
            if float(m) == int(float(m)):
                m = str(int(float(m)))
        except ValueError:
            pass
    if s.endswith(".0"):
        try:
            if float(s) == int(float(s)):
                s = str(int(float(s)))
        except ValueError:
            pass
    return f"{m}-{s}"


def _provision_for_type(name: str, types: list[dict[str, Any]],
                        dq_ranges: list[dict[str, Any]] | None,
                        days_dq: int | None) -> float | None:
    """Return the provision percentage for an impairment type.

    Mirrors the spreadsheet's nested ``IF(A=$A$5, $B$5, IF(A=$A$6, ...))``.
    If the matching row has a ``Variable`` provision (None) and DQ
    ranges are configured, falls through to the largest DQ range whose
    minimum days is <= ``days_dq``.
    """
    name_l = (name or "").strip().lower()
    matched_pct: float | None = None
    matched = False
    for t in types or []:
        if (t.get("name") or "").strip().lower() == name_l:
            matched = True
            matched_pct = t.get("provision_pct")
            break
    if not matched:
        return None
    if matched_pct is not None:
        return matched_pct
    # Variable -> DQ-range fallback
    if dq_ranges and days_dq is not None:
        # Pick range with the largest min_days that <= days_dq.
        best = None
        for r in dq_ranges:
            mn = r.get("min_days")
            if mn is None:
                continue
            try:
                mn_i = int(mn)
            except (TypeError, ValueError):
                continue
            if days_dq >= mn_i and (best is None or mn_i > best[0]):
                pct = r.get("provision_pct")
                if pct is not None:
                    best = (mn_i, float(pct))
        if best is not None:
            return best[1]
    return None


def compute_calculations(rows: list[dict[str, Any]],
                         types: list[dict[str, Any]],
                         dq_ranges: list[dict[str, Any]] | None = None
                         ) -> list[dict[str, Any]]:
    """Recreate the K:Q spreadsheet columns for each row.

    Adds these keys to each row in-place and returns the list::
        member_suffix       (K)  =B&"-"&C
        total_loans         (L)  =E + G
        ltv                 (M)  =L / H   (None if H==0)
        lgd                 (N)  =I if I!=""
                                  else max(0, L - H)
        pct_at_risk         (O)  =100% if I!=""
                                  else provision_pct(impairment_type)
        provision_amount    (P)  =N * O
        balance_removed     (Q)  =E
    """
    for row in rows:
        e = row.get("current_balance") or 0.0
        g = row.get("other_lender_balance") or 0.0
        h = row.get("collateral_value") or 0.0
        i_allow = row.get("allowance_provided")
        days = row.get("days_dq")
        ipt = row.get("impairment_type") or ""

        row["member_suffix"] = _member_suffix_key(
            row.get("member"), row.get("suffix"))

        total = float(e) + float(g)
        row["total_loans"] = total

        if not h:
            row["ltv"] = None  # "No Value"
        else:
            row["ltv"] = total / float(h)

        if i_allow is not None:
            row["lgd"] = float(i_allow)
            row["pct_at_risk"] = 1.0
        else:
            row["lgd"] = max(0.0, total - float(h))
            row["pct_at_risk"] = _provision_for_type(
                ipt, types, dq_ranges, days)

        if row["pct_at_risk"] is None:
            row["provision_amount"] = None
        else:
            row["provision_amount"] = row["lgd"] * row["pct_at_risk"]

        row["balance_removed"] = float(e)
    return rows


# ---------------------------------------------------------------------------
# lookup_from_loan_data -- R:S (+ T/U for diff)
# ---------------------------------------------------------------------------


def _resolve_pool_code(code: Any,
                       pool_map: dict[str, Any],
                       default_pool: str,
                       pool_split: str | None = None) -> str:
    """Map a raw loan code to a wizard pool name.

    Mirrors the lookup used when building the loan-data index so that the
    impaired step can fall back to the data-entry loan code when a row
    is not found in the loan extract.
    """
    if code is None:
        return default_pool or ""
    s = str(code).strip()
    if not s:
        return default_pool or ""
    if pool_split and pool_split in s:
        s = s.split(pool_split, 1)[0].strip()
    if s in (pool_map or {}) and pool_map[s]:
        return str(pool_map[s])
    s2 = s.lstrip("0") or s
    if s2 in (pool_map or {}) and pool_map[s2]:
        return str(pool_map[s2])
    return default_pool or ""


def _coerce_member_part(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        try:
            if float(s) == int(float(s)):
                return str(int(float(s)))
        except ValueError:
            pass
    return s


def _build_loan_index(loan_path: str | Path,
                      state: dict[str, Any],
                      header_row: int | None = None,
                      entry: dict[str, Any] | None = None) -> dict[str, dict[str, Any]]:
    """Build {member-suffix: {pool, grade, balance}} from a loan extract.

    ``header_row`` is the per-file 1-indexed header row stored on the
    loan_data_files entry (e.g. ``2`` for AIRES extracts whose row 1 is
    column position numbers and row 2 is the real header). When omitted
    or <= 1 we fall back to pandas' default header=0.

    ``entry`` is the loan_data_files entry dict. When provided, its
    ``column_mappings`` and ``member_account`` override the top-level
    ``state`` values for THIS file only — required for CUs that ship
    multiple extracts with different column shapes (e.g. Symitar
    ceclXX files with ACCTBS/ACTTYP/CURBAL/RISKSC + a CUMA Mortgage
    workbook with Account/Type/NEW PRINC BALANCE).
    """
    index: dict[str, dict[str, Any]] = {}
    p = Path(str(loan_path))
    if not p.exists():
        return index
    try:
        import pandas as pd
    except ImportError:
        return index

    sample = state.get("sample") or {}
    has_header = bool(sample.get("has_header", True))

    # Translate (has_header, header_row) into pandas' header= arg.
    if has_header:
        try:
            hr_i = int(header_row) if header_row is not None else 0
        except (TypeError, ValueError):
            hr_i = 0
        pd_header: int | None = hr_i - 1 if hr_i > 1 else 0
    else:
        pd_header = None

    suffix = (p.suffix or "").lower()
    try:
        if suffix in (".xlsx", ".xls", ".xlsm"):
            df = pd.read_excel(p, header=pd_header, dtype=object)
        elif suffix == ".csv":
            df = pd.read_csv(p, header=pd_header,
                             dtype=object, keep_default_na=False)
        else:
            return index
    except Exception:  # noqa: BLE001
        return index

    if df is None or df.empty:
        return index

    # Phase 9.24c — prefer per-entry column_mappings / member_account over
    # the top-level state values so multi-extract CUs (Symitar + CUMA
    # alongside, etc.) honour each file's own configured shape. Falls back
    # to the top-level dict for any key not overridden by the entry.
    top_cm = state.get("column_mappings") or {}
    top_ma = state.get("member_account") or {}
    if entry is not None:
        ecm = entry.get("column_mappings") or {}
        cm = {**top_cm, **{k: v for k, v in ecm.items() if v}}
        ema = entry.get("member_account") or {}
        ma = {**top_ma, **{k: v for k, v in ema.items() if v not in (None, "")}}
    else:
        cm = top_cm
        ma = top_ma
    pool_map = state.get("pool_map") or {}
    default_pool = state.get("default_pool") or ""
    grades = state.get("credit_grades") or []
    no_score = state.get("no_score_label") or "Not Reported"
    pool_split = state.get("pool_code_split") or None

    # Build a normalised header lookup once so column-name matches tolerate
    # leading/trailing whitespace and case differences (Symitar extracts
    # ship headers like ' CURBAL ' or '   CURBAL   ' that exact-match
    # against the saved 'CURBAL' would miss).
    def _norm_header(s: Any) -> str:
        return " ".join(str(s).split()).strip().lower()

    norm_cols = {_norm_header(c): c for c in df.columns}

    def _resolve_col(field_name: str):
        target = cm.get(field_name) or ""
        if target == "" or target is None:
            return None
        # If the mapping is a header name, use it directly when has_header.
        if has_header and target in df.columns:
            return target
        # Whitespace/case-insensitive header lookup (handles ' CURBAL ').
        nt = _norm_header(target)
        if nt and nt in norm_cols:
            return norm_cols[nt]
        # Otherwise treat as 0-based index.
        try:
            idx = int(target)
            if 0 <= idx < len(df.columns):
                return df.columns[idx]
        except (TypeError, ValueError):
            pass
        # Fall back to header lookup even when has_header is false.
        if target in df.columns:
            return target
        return None

    member_col = _resolve_col("member_number")
    suffix_col = _resolve_col("loan_suffix")
    pool_col = _resolve_col("loan_pool_code")
    bal_col = _resolve_col("current_balance")
    fico_col = _resolve_col("original_fico_score")
    if member_col is None:
        return index

    mode = (ma.get("mode") or "fixed_suffix").lower()
    # Honour an explicit suffix_length=0 (combined-fixed mode where the
    # member-number column already holds the full account, no suffix).
    # Plain ``int(... or 3)`` would silently override 0 -> 3 because 0 is
    # falsy in Python; that produced phantom mismatches against impaired
    # files for CUs configured with combined-fixed/no-suffix accounts.
    _sl_raw = ma.get("suffix_length")
    try:
        suffix_len = int(_sl_raw) if _sl_raw is not None else 3
    except (TypeError, ValueError):
        suffix_len = 3
    delim = ma.get("delimiter") or "-"

    def _grade_for(score: Any) -> str:
        f = _to_float(score)
        if f is None:
            return no_score
        for g in grades:
            try:
                # Accept either {min,max} or {min_score,max_score}.
                lo = g.get("min", g.get("min_score"))
                hi = g.get("max", g.get("max_score"))
                lo_n = _to_float(lo)
                hi_n = _to_float(hi)
                # Skip the catch-all "no score" row that uses 0/0.
                if lo_n == 0 and hi_n == 0:
                    continue
                if (lo_n is None or f >= lo_n) and (hi_n is None or f <= hi_n):
                    return str(g.get("label") or g.get("name") or "")
            except Exception:  # noqa: BLE001
                continue
        return no_score

    def _pool_for(code: Any) -> str:
        return _resolve_pool_code(code, pool_map, default_pool, pool_split)

    for _, row in df.iterrows():
        raw_member = row.get(member_col)
        if raw_member is None or _norm(raw_member) == "":
            continue
        if mode == "split" and suffix_col is not None:
            m = _coerce_member_part(raw_member)
            s = _coerce_member_part(row.get(suffix_col))
            key = f"{m}-{s}"
            acct_raw = f"{m}{s}"
        elif mode == "delimiter":
            s_raw = _coerce_member_part(raw_member)
            if delim and delim in s_raw:
                m, s = s_raw.split(delim, 1)
                key = f"{m.strip()}-{s.strip()}"
                acct_raw = f"{m.strip()}{s.strip()}"
            else:
                key = f"{s_raw}-"
                acct_raw = s_raw
        else:  # fixed_suffix
            s_raw = _coerce_member_part(raw_member)
            if suffix_len > 0 and len(s_raw) > suffix_len:
                m = s_raw[:-suffix_len]
                s = s_raw[-suffix_len:].lstrip("0") or "0"
                key = f"{m}-{s}"
            else:
                key = f"{s_raw}-"
            acct_raw = s_raw

        # Phase 9.24b — also carry the raw extract loan_pool_code value
        # so the impaired-step validation card can surface codes from the
        # loan extract whose mapping resolved to Ignore (matched-row case).
        raw_pool_code = ""
        if pool_col is not None:
            _rpc = row.get(pool_col)
            if _rpc is not None:
                raw_pool_code = str(_rpc).strip()

        index[key] = {
            "loan_pool": _pool_for(row.get(pool_col)) if pool_col else default_pool,
            "loan_pool_code_raw": raw_pool_code,
            "credit_grade": _grade_for(row.get(fico_col)) if fico_col else no_score,
            "balance_from_loan_report": _to_float(row.get(bal_col)) if bal_col else None,
            "_acct": _digits(acct_raw),
        }
    return index


def lookup_from_loan_data(rows: list[dict[str, Any]],
                          state: dict[str, Any]) -> dict[str, Any]:
    """Enrich each row with R/S/T/U values pulled from the loan extract.

    Returns a status dict ``{ok, error, matched, unmatched, source}``.
    """
    status: dict[str, Any] = {
        "ok": False, "error": None, "matched": 0,
        "unmatched": 0, "source": None,
    }
    pool_map = state.get("pool_map") or {}
    default_pool = state.get("default_pool") or ""
    pool_split = state.get("pool_code_split") or None
    no_score = state.get("no_score_label") or "Not Reported"

    def _fallback_from_data_entry(row: dict[str, Any]) -> None:
        """Fill loan_pool/credit_grade/balance from the row's own data-entry."""
        cb = row.get("current_balance")
        row["loan_pool"] = _resolve_pool_code(
            row.get("loan_type"), pool_map, default_pool, pool_split)
        # No extract row matched — clear any stale resolved code.
        row["loan_pool_code_resolved"] = ""
        row["credit_grade"] = no_score
        row["balance_from_loan_report"] = (
            float(cb) if cb is not None else None)
        row["balance_difference"] = 0.0 if cb is not None else None
        row["unmatched_in_loan_data"] = True

    sample_uploads = state.get("sample_uploads") or {}
    # The wizard's Sample step records loan-data extracts under
    # ``loan_data_files``; ``loan_files`` was the legacy key. Try both.
    su = (sample_uploads.get("loan_data_files")
          or sample_uploads.get("loan_files")
          or [])
    if not su:
        status["error"] = "No sample loan-data file uploaded yet (Step 5 — Sample)."
        for row in rows:
            _fallback_from_data_entry(row)
        return status

    # Build a combined index across ALL uploaded loan-data extracts so
    # impaired loans can match against any of them (e.g. a separate CC
    # extract alongside the main loan file). Later files overwrite
    # earlier ones on key collision.
    index: dict[str, dict[str, Any]] = {}
    sources: list[str] = []
    for entry in su:
        if not isinstance(entry, dict):
            continue
        p = entry.get("path")
        if not p:
            continue
        # Per-file header_row (1-indexed). AIRES-style extracts use
        # header_row=2; conventional extracts use 1 (or absent).
        try:
            hr_entry = int(entry.get("header_row") or 0) or None
        except (TypeError, ValueError):
            hr_entry = None
        sub = _build_loan_index(p, state, header_row=hr_entry, entry=entry)
        if sub:
            index.update(sub)
            sources.append(entry.get("name") or Path(p).name)
    if not sources:
        status["error"] = "No loan-data file on disk could be read."
        for row in rows:
            _fallback_from_data_entry(row)
        return status
    status["source"] = ", ".join(sources)

    # Phase 9.37b — build a leading-zero-normalised secondary index so
    # extract keys like ``000001002-4`` still match impaired rows keyed as
    # ``1002-4`` (or vice versa). Some CU loan extracts (Symitar AIRES)
    # zero-pad Account Number to 9 digits as a text string, while the
    # impaired workbook ships raw ints.
    def _strip_lead(k: str) -> str:
        if "-" in k:
            a, b = k.split("-", 1)
            return f"{a.lstrip('0') or '0'}-{b.lstrip('0') or '0'}"
        return k.lstrip("0") or "0"

    norm_index: dict[str, dict[str, Any]] = {}
    for _k, _v in index.items():
        _nk = _strip_lead(_k)
        # First-in wins for duplicates; the primary index (padded form)
        # is authoritative when both exist.
        norm_index.setdefault(_nk, _v)

    # Phase 9.41 — account-reconstruction index. Keyed by the digits-only
    # combined account (member digits + suffix digits) captured when the
    # loan index was built. Lets an impaired row whose member/suffix are
    # split differently than the extract still resolve: e.g. the AIRES
    # extract packs a 4-digit zero-padded loan suffix ("192380004") while
    # the CU is configured with suffix_length=3 (so the extract mis-splits
    # to "192380-4"), and the impaired file lists member 19238 / suffix 4.
    # We reconstruct the extract's full account from the impaired member +
    # a zero-padded suffix and require an EXACT account-digit match, so it
    # cannot introduce spurious hits. Driven from the impaired side (few
    # rows), it is a last-resort fallback after the direct / leading-zero
    # paths miss.
    acct_index: dict[str, dict[str, Any]] = {}
    for _v in index.values():
        _a = _v.get("_acct")
        if _a:
            acct_index.setdefault(_a, _v)
            acct_index.setdefault(_a.lstrip("0") or "0", _v)

    def _match_by_account(member: Any, suffix: Any):
        md = _digits(member)
        if not md:
            return None
        sd = _digits(suffix)
        sd_core = sd.lstrip("0")
        # Loan suffixes are short sequence numbers; only reconstruct when
        # the significant suffix is small so we don't concatenate a long
        # suffix into a *different* member's account by coincidence.
        if len(sd_core) > 4:
            return None
        md_variants = [md]
        md_stripped = md.lstrip("0")
        if md_stripped and md_stripped != md:
            md_variants.append(md_stripped)
        # Reconstruct the extract account as member + the suffix zero-padded
        # to widths from its own length up to +4 leading zeros (covers the
        # common 2-, 3- and 4-digit zero-padded loan-suffix conventions).
        # Longest (most-padded) width first so the structurally correct
        # reconstruction wins before shorter, more ambiguous forms.
        base_w = max(len(sd_core), 1)
        for _md in md_variants:
            cands: list[str] = []
            for w in range(base_w + 4, base_w - 1, -1):
                cands.append(_md + sd_core.zfill(w))
            cands.append(_md + sd)        # suffix exactly as entered
            if not sd_core:
                cands.append(_md)         # no / zero suffix
            seen: set[str] = set()
            for c in cands:
                if not c or c in seen:
                    continue
                seen.add(c)
                hit = acct_index.get(c) or acct_index.get(c.lstrip("0") or "0")
                if hit is not None:
                    return hit
        return None

    for row in rows:
        key = row.get("member_suffix") or _member_suffix_key(
            row.get("member"), row.get("suffix"))
        match = index.get(key)
        if match is None:
            # Try alternate normalisations of suffix (drop leading zeros)
            m, _, s = key.partition("-")
            alt = f"{m}-{s.lstrip('0') or '0'}"
            match = index.get(alt)
        if match is None:
            # Phase 9.37b — try the leading-zero-normalised index
            match = norm_index.get(_strip_lead(key))
        if match is None:
            # Phase 9.41 — reconstruct the extract account from the
            # impaired member + zero-padded suffix (handles suffix-length /
            # zero-pad disagreements between the two files).
            match = _match_by_account(row.get("member"), row.get("suffix"))
        if match is None:
            status["unmatched"] += 1
            _fallback_from_data_entry(row)
        else:
            status["matched"] += 1
            # Phase 9.24b — if the extract row's pool_col was blank or
            # resolved to Ignore, fall back to the impaired row's own
            # loan_type resolution. Lets the user remap pool_map[loan_type]
            # via the impaired validation card to fix Ignore-resolving rows
            # whose loan extract had no useful pool info.
            extract_pool = match.get("loan_pool") or ""
            _ep_l = extract_pool.strip().lower()
            if _ep_l and _ep_l != "ignore":
                row["loan_pool"] = extract_pool
            else:
                row["loan_pool"] = _resolve_pool_code(
                    row.get("loan_type"), pool_map, default_pool, pool_split)
            # Propagate the extract's raw loan_pool_code so the wizard's
            # impaired validation card can offer to remap it when present.
            row["loan_pool_code_resolved"] = match.get("loan_pool_code_raw") or ""
            row["credit_grade"] = match.get("credit_grade") or ""
            row["balance_from_loan_report"] = match.get("balance_from_loan_report")
            cb = row.get("current_balance") or 0.0
            blr = match.get("balance_from_loan_report")
            row["balance_difference"] = (
                None if blr is None else float(cb) - float(blr)
            )
            row["unmatched_in_loan_data"] = False
    status["ok"] = True
    return status


# ---------------------------------------------------------------------------
# Aggregator helper used by the wizard route
# ---------------------------------------------------------------------------


def recompute_all(impaired_state: dict[str, Any],
                  wizard_state: dict[str, Any]) -> dict[str, Any]:
    """Recompute K:Q and R:S/T/U on the rows in ``impaired_state``.

    Returns the lookup-status dict from ``lookup_from_loan_data``.
    """
    rows = impaired_state.get("data_rows") or []
    # NOTE: parse_file stores the impairment-type table under
    # "impairment_types" (not "types"); reading the wrong key left the
    # provision-percentage lookup empty, so every row relying on a
    # type-table percentage (e.g. "Delinquent Loans" @ 35%) resolved to
    # None and the verification page showed $0 provision.
    types = (impaired_state.get("impairment_types")
             or impaired_state.get("types") or [])
    dq = impaired_state.get("dq_ranges") or []
    compute_calculations(rows, types, dq)
    return lookup_from_loan_data(rows, wizard_state)
