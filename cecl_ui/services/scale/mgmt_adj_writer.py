"""SCALE Management Adjustment writer.

The SCALE template's ``Management Adjustment`` tab has the structure
(per ``cecl_ui/data/scale/templates/*.xlsx``)::

    D1  "Default Mgmt Adj"
    D2  <default decimal, e.g. 0.0011>      E2  "Yes"  (sentinel)
    A3..F3 headers (Loan Pool / Final / Hard Code / Default / Include Y/N / CECL Lifetime Rate)
    A4..A16   13 pool labels (pulled via formula from 'Scale Calculation'!C9:C21)
    C4..C16   per-pool *Hard Code* mgmt adj (decimal)
    E4..E16   "Yes"/"No"  -- toggles the Default lookup
    A18..F18  whole-portfolio overlay row (same shape)

The workbook formulas already implement the rule the user described:
``Final = IF(HardCode + BaseRate == 0, Default, HardCode)`` and
``Default = IF(IncludeYN == "Yes", $D$2, 0)``. The wizard's job is
only to write the inputs: ``D2`` (admin default), and per-row
``C`` (hard code) and ``E`` ("Yes"/"No" checkbox).
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


SHEET_NAME = "Management Adjustment"
DEFAULT_CELL = "D2"
DEFAULT_SENTINEL_CELL = "E2"   # must equal each row's E to enable default
POOL_ROWS = list(range(4, 17))         # rows 4..16 (13 pools)
PORTFOLIO_ROW = 18

POOL_REF_SHEET = "Scale Calculation"
POOL_REF_COL = 3              # column C
POOL_REF_ROW_START = 9         # C9 corresponds to Management Adjustment A4
POOL_REF_ROW_END = 21          # C21 corresponds to Management Adjustment A16


def list_pool_names(template_path: str | Path) -> list[str]:
    """Return the 13 pool labels in the order they appear on the template.

    Reads the literal strings from ``Scale Calculation``!C9:C21 because
    the ``Management Adjustment`` A column holds formulas, not values.
    Empty cells (e.g. C22) are skipped; the ``Total`` row is also skipped.
    """
    try:
        wb = openpyxl.load_workbook(template_path, data_only=False)
    except Exception:  # noqa: BLE001
        return []
    try:
        if POOL_REF_SHEET not in wb.sheetnames:
            return []
        ws = wb[POOL_REF_SHEET]
        names: list[str] = []
        for r in range(POOL_REF_ROW_START, POOL_REF_ROW_END + 1):
            v = ws.cell(row=r, column=POOL_REF_COL).value
            if v is None:
                continue
            s = str(v).strip()
            if not s or s.lower() == "total":
                continue
            names.append(s)
        return names
    finally:
        wb.close()


def _coerce_decimal(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def apply_mgmt_adj(
    workbook_path: str | Path, mgmt_state: dict | None,
) -> dict:
    """Write Management Adjustment inputs into ``workbook_path``.

    ``mgmt_state`` shape (see :func:`scale_setup._default_scale_block`)::

        {
          "default_pct":          float,   # decimal, e.g. 0.0011
          "pool_rows": [                    # one per pool, in template order
              {"name": str,
               "hard_code_pct": float,    # decimal
               "use_default":   bool},
              ...
          ],
          "portfolio":  {"hard_code_pct": float, "use_default": bool},
        }

    Returns ``{ok, sheet_missing, pools_written, portfolio_written,
    default_written, error}``.
    """
    result: dict[str, Any] = {
        "ok": False,
        "sheet_missing": False,
        "default_written": False,
        "pools_written": 0,
        "portfolio_written": False,
        "error": "",
    }
    if not mgmt_state:
        result["ok"] = True
        return result
    try:
        wb = openpyxl.load_workbook(workbook_path)
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"Could not open workbook: {exc}"
        return result
    try:
        if SHEET_NAME not in wb.sheetnames:
            result["sheet_missing"] = True
            result["error"] = f"Sheet {SHEET_NAME!r} not found."
            return result
        ws = wb[SHEET_NAME]

        # 1. Default rate cell.
        default_pct = _coerce_decimal(mgmt_state.get("default_pct"))
        ws[DEFAULT_CELL].value = default_pct
        result["default_written"] = True

        # 2. Per-pool rows. Walk the rows in template order; pad/truncate
        # ``pool_rows`` to match POOL_ROWS length.
        pool_rows = list(mgmt_state.get("pool_rows") or [])
        for idx, row_no in enumerate(POOL_ROWS):
            pr = pool_rows[idx] if idx < len(pool_rows) else {}
            ws.cell(row=row_no, column=3).value = _coerce_decimal(
                pr.get("hard_code_pct")
            )
            ws.cell(row=row_no, column=5).value = (
                "Yes" if pr.get("use_default") else "No"
            )
            result["pools_written"] += 1

        # 3. Portfolio overlay row.
        port = mgmt_state.get("portfolio") or {}
        ws.cell(row=PORTFOLIO_ROW, column=3).value = _coerce_decimal(
            port.get("hard_code_pct")
        )
        ws.cell(row=PORTFOLIO_ROW, column=5).value = (
            "Yes" if port.get("use_default") else "No"
        )
        result["portfolio_written"] = True

        wb.save(workbook_path)
        result["ok"] = True
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"Write failed: {exc}"
    finally:
        wb.close()
    return result
