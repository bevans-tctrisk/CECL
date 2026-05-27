"""Write firm-wide Environmental Factor Ranges into a SCALE workbook.

The TCT tab ``Environmental Factor Ranges`` holds two source-of-truth
tables (delinquency and economic stress); the Vizo tab
``Envir Factor Ranges-Vizo`` only references the TCT tab via formulas
so we only write to the TCT tab.

Cell layout (1-based row numbers, see template inspection 2026-05-27):

* Delinquency  -- ``J9:K25`` (17 rows). Col J = minimum % (decimal),
  Col K = score (decimal). Display cells C9:D25 are formula-driven
  off J/K and update automatically.
* Economic stress (primary)  -- ``L9:M22`` (14 rows). Col L = minimum
  % (decimal), Col M = score (decimal). Display cells E9:F22 are
  formula-driven off L/M and update automatically.
* Economic stress (secondary helper)  -- ``O10:R22`` (rows 10..22).
  This is a hardcoded range-display table that does NOT reference
  L/M so we rewrite it here to stay in sync. Row 9 is special
  (``P9`` is the ">25%" text, ``R9`` carries the top score) and is
  also rewritten.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from cecl_ui.services import admin_defaults


TCT_TAB_NAME = "Environmental Factor Ranges"

_DELQ_FIRST_ROW = 9
_DELQ_MIN_COL = "J"
_DELQ_SCORE_COL = "K"

_ECON_FIRST_ROW = 9
_ECON_MIN_COL = "L"
_ECON_SCORE_COL = "M"

# Secondary econ-stress helper table (range-display) -- cols O/P/R.
_ECON2_MIN_COL = "O"
_ECON2_MAX_COL = "P"
_ECON2_SCORE_COL = "R"


def _coerce_pair(row: Any) -> tuple[float, float] | None:
    if not isinstance(row, (list, tuple)) or len(row) < 2:
        return None
    try:
        return float(row[0]), float(row[1])
    except (TypeError, ValueError):
        return None


def apply_env_factor_ranges(
    workbook_path: str | Path,
    ranges: dict[str, Any] | None = None,
) -> dict:
    """Write the env-factor ranges from ``ranges`` (or admin defaults).

    Returns ``{ok, applied_delq, applied_econ, skipped, error}``.
    Missing tab returns ``{ok: False, error: ...}`` and writes nothing.
    """
    out: dict[str, Any] = {
        "ok": False,
        "applied_delq": 0,
        "applied_econ": 0,
        "skipped": [],
        "error": "",
    }
    if ranges is None:
        ranges = (admin_defaults.load().get("env_factor_ranges") or {})

    delq_rows = ranges.get("delinquency") or []
    econ_rows = ranges.get("econ_stress") or []

    delq_expected = admin_defaults.DELINQUENCY_ROW_COUNT
    econ_expected = admin_defaults.ECON_STRESS_ROW_COUNT

    try:
        wb = openpyxl.load_workbook(workbook_path)
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Could not open workbook: {exc}"
        return out
    if TCT_TAB_NAME not in wb.sheetnames:
        out["error"] = f"Tab {TCT_TAB_NAME!r} not found in workbook."
        return out
    ws = wb[TCT_TAB_NAME]

    # Delinquency table (J9:K25)
    for i in range(delq_expected):
        pair = _coerce_pair(delq_rows[i]) if i < len(delq_rows) else None
        if pair is None:
            out["skipped"].append(f"delinquency row {i + 1}")
            continue
        row = _DELQ_FIRST_ROW + i
        ws[f"{_DELQ_MIN_COL}{row}"].value = pair[0]
        ws[f"{_DELQ_SCORE_COL}{row}"].value = pair[1]
        out["applied_delq"] += 1

    # Economic stress primary table (L9:M22)
    econ_pairs: list[tuple[float, float] | None] = [None] * econ_expected
    for i in range(econ_expected):
        pair = _coerce_pair(econ_rows[i]) if i < len(econ_rows) else None
        econ_pairs[i] = pair
        if pair is None:
            out["skipped"].append(f"econ_stress row {i + 1}")
            continue
        row = _ECON_FIRST_ROW + i
        ws[f"{_ECON_MIN_COL}{row}"].value = pair[0]
        ws[f"{_ECON_SCORE_COL}{row}"].value = pair[1]
        out["applied_econ"] += 1

    # Secondary helper table O/P/R. Row 9 P-cell is ">NN%" text; R9
    # mirrors M9. Rows 10..22: O = L_row, P = L_(row-1) - 0.0001,
    # R = M_row.
    if econ_pairs[0] is not None:
        top_min, top_score = econ_pairs[0]
        ws[f"{_ECON2_MAX_COL}{_ECON_FIRST_ROW}"].value = (
            f">{int(round(top_min * 100))}%"
        )
        ws[f"{_ECON2_SCORE_COL}{_ECON_FIRST_ROW}"].value = top_score

    for i in range(1, econ_expected):
        prev = econ_pairs[i - 1]
        cur = econ_pairs[i]
        if cur is None or prev is None:
            continue
        row = _ECON_FIRST_ROW + i
        ws[f"{_ECON2_MIN_COL}{row}"].value = cur[0]
        ws[f"{_ECON2_MAX_COL}{row}"].value = round(prev[0] - 0.0001, 6)
        ws[f"{_ECON2_SCORE_COL}{row}"].value = cur[1]

    wb.save(workbook_path)
    out["ok"] = True
    return out
