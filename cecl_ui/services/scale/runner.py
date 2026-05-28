"""End-to-end SCALE run.

``run_single_quarter(state, workspace_root)`` resolves the template and
mapping for the chosen period, fetches the Solr doc, writes the cells,
and saves an output workbook under
``Generated_Reports/<short>/<YYYY-MM>/CECL_SCALE_<short>.xlsx``.

``run_multi_quarter(state, workspace_root, quarters)`` repeats the fill
for ``quarters`` consecutive quarter-ends (newest -> oldest), writing
each quarter's data into the same workbook. After the first iteration
the output file is used as the template so writes accumulate (the
period-specific mapping CSVs point to different cells per quarter so
prior writes are preserved). Quarters with no mapping CSV are skipped
and reported.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
import re
import shutil
import zipfile

import openpyxl

from . import (
    env_factor_writer, impaired_loader, mapping_loader, mgmt_adj_writer,
    qfactor_loader, runs_service, solr_fetcher, template_loader,
)


_THEME_ZIP_PATH = "xl/theme/theme1.xml"
_DEFAULT_OFFICE_THEME_CACHE: bytes | None = None


def _default_office_theme_bytes() -> bytes:
    """Return ``xl/theme/theme1.xml`` from a fresh openpyxl workbook.

    Fresh openpyxl workbooks ship with the standard Office theme (the
    one VS Excel uses for blank workbooks). Cached after first call.
    """
    global _DEFAULT_OFFICE_THEME_CACHE
    if _DEFAULT_OFFICE_THEME_CACHE is not None:
        return _DEFAULT_OFFICE_THEME_CACHE
    buf = BytesIO()
    wb = openpyxl.Workbook()
    wb.save(buf)
    buf.seek(0)
    with zipfile.ZipFile(buf, "r") as zf:
        _DEFAULT_OFFICE_THEME_CACHE = zf.read(_THEME_ZIP_PATH)
    return _DEFAULT_OFFICE_THEME_CACHE


def _read_workbook_theme(path: str | Path) -> bytes | None:
    """Return ``xl/theme/theme1.xml`` bytes from an xlsx file, or None."""
    try:
        with zipfile.ZipFile(path, "r") as zf:
            return zf.read(_THEME_ZIP_PATH)
    except (KeyError, zipfile.BadZipFile, FileNotFoundError):
        return None


def _replace_workbook_theme(path: str | Path, theme_bytes: bytes) -> None:
    """Rewrite ``xl/theme/theme1.xml`` inside an xlsx zip in-place."""
    src = Path(path)
    tmp = src.with_suffix(src.suffix + ".themetmp")
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(
        tmp, "w", zipfile.ZIP_DEFLATED
    ) as zout:
        for item in zin.infolist():
            data = (
                theme_bytes if item.filename == _THEME_ZIP_PATH
                else zin.read(item.filename)
            )
            zout.writestr(item, data)
    tmp.replace(src)


VALID_Q_MONTHS = {3, 6, 9, 12}


def prev_quarter(ym: str) -> str:
    """``2025-12`` -> ``2025-09`` (matches multi_quarter.py CLI)."""
    y_str, m_str = ym.split("-")
    y, m = int(y_str), int(m_str)
    if m not in VALID_Q_MONTHS:
        raise ValueError("period month must be one of 03, 06, 09, 12")
    if m == 3:
        return f"{y - 1}-12"
    if m == 6:
        return f"{y}-03"
    if m == 9:
        return f"{y}-06"
    return f"{y}-09"


def make_periods(start_period: str, quarters: int) -> list[str]:
    if quarters < 1:
        raise ValueError("quarters must be >= 1")
    periods = [start_period]
    while len(periods) < quarters:
        periods.append(prev_quarter(periods[-1]))
    return periods


def _output_path(workspace_root: str, short: str, period: str) -> Path:
    safe_short = short or "cu"
    out_dir = Path(workspace_root) / "Generated_Reports" / safe_short / period
    out_dir.mkdir(parents=True, exist_ok=True)
    # Prefix with the YYYY-MM period so the workbook self-identifies once it
    # leaves the period folder (email attachments, audit copies, etc.).
    # apply_report_variant() appends _TCT/_Vizo suffixes to this stem.
    return out_dir / f"{period}_CECL_SCALE_{safe_short}.xlsx"


def apply_qfactors(
    workbook_path: str | Path, entries: list[dict],
) -> dict:
    """Write Q-factor values (bps -> decimal) into ``workbook_path``.

    ``entries`` come from :func:`qfactor_loader.merge_with_overrides`
    (each has ``sheet``, ``cell``, ``effective_bps``).
    """
    result = {
        "applied": 0,
        "total": len(entries),
        "missing_sheets": [],
        "issues": [],
    }
    if not entries:
        return result
    wb = openpyxl.load_workbook(workbook_path)
    for row in entries:
        sheet = row["sheet"]
        cell = row["cell"]
        if sheet not in wb.sheetnames:
            result["missing_sheets"].append(f"{sheet} (for {cell})")
            continue
        try:
            wb[sheet][cell].value = float(row.get("effective_bps", 0.0)) / 10000.0
            result["applied"] += 1
        except Exception as exc:  # noqa: BLE001
            result["issues"].append(f"{sheet}!{cell}: {exc}")
    wb.save(workbook_path)
    return result


def _qfactor_entries_from_state(state: dict) -> list[dict]:
    scale = state.get("scale") or {}
    defaults = qfactor_loader.load_defaults()
    return qfactor_loader.merge_with_overrides(
        defaults, scale.get("qfactor_overrides") or {}
    )


def _impaired_rows_from_state(state: dict) -> list[dict]:
    scale = state.get("scale") or {}
    imp = scale.get("impaired_file") or {}
    parsed = imp.get("parsed") or {}
    if not parsed.get("ok"):
        return []
    return list(parsed.get("rows") or [])


def _normalize_mgmt_adj_state(
    mgmt_state: dict, template_path: str | Path,
) -> dict:
    """Re-align the wizard's name-keyed mgmt_adj rows to template order.

    Wizard state stores per-pool rows in
    ``mgmt_state["pool_rows"]`` as ``{pool_name: {hard_code_pct,
    use_default}}``. The writer needs an ordered list matching the
    template's pool rows. Pool names missing from saved state get a
    zero/off default. Returns the shape ``mgmt_adj_writer.apply_mgmt_adj``
    expects.
    """
    if not isinstance(mgmt_state, dict):
        return {}
    try:
        names = mgmt_adj_writer.list_pool_names(template_path)
    except Exception:  # noqa: BLE001
        names = []
    saved = mgmt_state.get("pool_rows") or {}
    if not isinstance(saved, dict):
        saved = {}
    ordered: list[dict] = []
    for name in names:
        row = saved.get(name) or {}
        ordered.append({
            "name": name,
            "hard_code_pct": float(row.get("hard_code_pct") or 0.0),
            "use_default": bool(row.get("use_default")),
        })
    return {
        "default_pct": float(mgmt_state.get("default_pct") or 0.0),
        "pool_rows": ordered,
        "portfolio": mgmt_state.get("portfolio") or {},
    }


# Tab sets per report variant. Shared tabs (Management Adjustment,
# Historical Data, Industry Data) are NOT listed below and are kept
# visible for every variant. ``DNU `` sheets are always hidden.
_TCT_ONLY_SHEETS = {
    "Cover", "Introduction", "Scale Calculation", "Env Factor by Pool",
    "Environmental Factor Ranges", " Impaired Loans ASC 310-10", "Calc tab",
}
_VIZO_ONLY_SHEETS = {
    "Cover-Vizo", "Executive Summary-Vizo", "Scale Calculation-Vizo",
    "Env Factor by Pool-Vizo", " Impaired Loans-Vizo", "Introduction-Vizo",
    "Explanation of ACL Calc-Vizo", "Envir Factor Ranges-Vizo",
    "New Report Calc-Vizo",
}


def _hide_other_variant(
    workbook_path: str | Path, keep_variant: str,
) -> tuple[list[str], list[str]]:
    """Hide the sheets that don't belong to ``keep_variant`` ('tct' or 'vizo').

    ``DNU `` sheets are always hidden. Returns (hidden, kept) name lists.
    Shared tabs (Management Adjustment, Historical Data, Industry Data)
    are kept visible for both variants.
    """
    hidden: list[str] = []
    kept: list[str] = []
    wb = openpyxl.load_workbook(workbook_path)
    for name in list(wb.sheetnames):
        ws = wb[name]
        if name.startswith("DNU "):
            hide = True
        elif keep_variant == "tct" and name in _VIZO_ONLY_SHEETS:
            hide = True
        elif keep_variant == "vizo" and name in _TCT_ONLY_SHEETS:
            hide = True
        else:
            hide = False
        if hide:
            ws.sheet_state = "hidden"
            hidden.append(name)
        else:
            ws.sheet_state = "visible"
            kept.append(name)
    visible = [s for s in wb.sheetnames if wb[s].sheet_state == "visible"]
    if visible:
        wb.active = wb.sheetnames.index(visible[0])
    wb.save(workbook_path)
    return hidden, kept


def apply_report_variant(
    workbook_path: str | Path, variant: str,
) -> dict:
    """Produce one themed output file per requested variant.

    A workbook can only carry a single ``xl/theme/theme1.xml``. The
    master template ships with the Vizo theme; the TCT variant uses the
    default Office theme. To honour both, we write a per-variant copy of
    ``workbook_path`` (suffixed ``_TCT.xlsx`` / ``_Vizo.xlsx``), hide the
    other variant's sheets in the copy, and inject the correct theme.

    Hidden (not deleted) so cross-sheet formulas (e.g.
    ``='Scale Calculation'!C9``) keep evaluating.

    ``variant`` in {``tct``, ``vizo``, ``both``}. When at least one
    themed file is written the master ``workbook_path`` is removed so
    the user is not confused by an extra single-theme file.

    Returns ``{variant, outputs:[{path,label,hidden,kept,theme}],
    hidden, kept}`` where ``hidden``/``kept`` mirror the last output for
    back-compat with existing call sites.
    """
    v = (variant or "both").strip().lower()
    if v not in ("tct", "vizo", "both"):
        v = "both"
    master = Path(workbook_path)
    targets: list[str] = []
    if v in ("tct", "both"):
        targets.append("tct")
    if v in ("vizo", "both"):
        targets.append("vizo")

    vizo_theme = _read_workbook_theme(master)
    tct_theme = _default_office_theme_bytes()

    outputs: list[dict] = []
    last_hidden: list[str] = []
    last_kept: list[str] = []
    for which in targets:
        suffix = "_TCT" if which == "tct" else "_Vizo"
        out_path = master.with_name(master.stem + suffix + master.suffix)
        shutil.copy2(master, out_path)
        hidden, kept = _hide_other_variant(out_path, which)
        theme_bytes = tct_theme if which == "tct" else vizo_theme
        theme_label = "Office default" if which == "tct" else "Vizo"
        if theme_bytes is not None:
            try:
                _replace_workbook_theme(out_path, theme_bytes)
            except Exception:  # noqa: BLE001
                # Theme swap is best-effort; the file is still usable.
                theme_label += " (theme swap failed)"
        outputs.append({
            "path": str(out_path),
            "label": "TCT" if which == "tct" else "Vizo",
            "hidden": hidden,
            "kept": kept,
            "theme": theme_label,
        })
        last_hidden, last_kept = hidden, kept

    # Remove the un-themed master once at least one themed copy exists.
    if outputs and master.exists():
        try:
            master.unlink()
        except OSError:
            pass

    return {
        "variant": v,
        "outputs": outputs,
        "hidden": last_hidden,
        "kept": last_kept,
    }


def _report_variant_from_state(state: dict) -> str:
    scale = state.get("scale") or {}
    return (scale.get("report_variant") or "both").strip().lower()


def fill_template(
    template_path: str | Path,
    out_path: str | Path,
    mapping_rows: list[dict],
    fields: dict[str, Any],
) -> dict:
    wb = openpyxl.load_workbook(template_path)
    applied = 0
    issues: list[str] = []
    missing_fields: list[str] = []
    missing_sheets: list[str] = []
    for row in mapping_rows:
        code = row["field_code"]
        sheet = row["sheet"]
        cell = row["cell"]
        if code not in fields:
            missing_fields.append(code)
            continue
        if sheet not in wb.sheetnames:
            missing_sheets.append(f"{sheet} (for {code})")
            continue
        try:
            wb[sheet][cell].value = solr_fetcher.coerce_numeric(fields[code])
            applied += 1
        except Exception as exc:  # noqa: BLE001
            issues.append(f"Failed to write {code} to {sheet}!{cell}: {exc}")
    wb.save(out_path)
    return {
        "applied": applied,
        "issues": issues,
        "missing_fields": missing_fields,
        "missing_sheets": missing_sheets,
    }


# Matches A1-style column letter (1-3 letters) followed by row digits,
# e.g. ``AY145`` in ``'Historical Data'!AY145``. Used by the column
# retargeter to scope replacements to actual cell refs and ignore
# things like text labels or sheet names that happen to contain the
# same letters.
_COL_LETTERS_RE = re.compile(r"^[A-Z]{1,3}$")


def _column_from_cell_coord(cell: str) -> str:
    """``AZ20`` -> ``AZ``; ``20`` or ``""`` -> ``""``."""
    s = (cell or "").strip().upper()
    out = []
    for ch in s:
        if ch.isalpha():
            out.append(ch)
        else:
            break
    return "".join(out)


def _target_column_from_mapping_rows(rows: list[dict]) -> str:
    """Derive the Historical Data column letter the mapping writes to.

    Mapping CSV rows that target Historical Data all use the same
    column letter (e.g. every ``AZ20``/``AZ145``/... cell for the
    2026-03 mapping). Returns the most common letter found, or ``""``
    if no Historical Data rows exist.
    """
    counts: dict[str, int] = {}
    for r in rows:
        if (r.get("sheet") or "").strip() != "Historical Data":
            continue
        col = _column_from_cell_coord(r.get("cell") or "")
        if _COL_LETTERS_RE.match(col):
            counts[col] = counts.get(col, 0) + 1
    if not counts:
        return ""
    return max(counts, key=counts.get)


def _detect_prior_column(workbook_path: str | Path,
                         anchor_cells: list[tuple[str, str]]) -> str:
    """Detect what Historical Data column the carried workbook's
    formulas currently reference.

    Looks at ``anchor_cells`` (list of ``(sheet, coord)``) in order,
    extracts the column letter from the first formula matching
    ``'Historical Data'!<col><row>``. Returns ``""`` if nothing found.
    """
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=False)
    except Exception:  # noqa: BLE001
        return ""
    pat = re.compile(r"Historical Data'?!([A-Z]{1,3})\d+")
    for sheet, coord in anchor_cells:
        if sheet not in wb.sheetnames:
            continue
        try:
            v = wb[sheet][coord].value
        except Exception:  # noqa: BLE001
            continue
        if isinstance(v, str) and v.startswith("="):
            m = pat.search(v)
            if m:
                return m.group(1).upper()
    return ""


# Stable cells in the SCALE template that reference Historical Data
# with the quarter-specific column. Used to detect what column the
# carried workbook currently points at. M9 lives in the historical
# CECL ACL Lifetime Loss Rate column on Scale Calculation; B3 is the
# Env Factor by Pool "For Period Ending" header.
_PRIOR_COLUMN_ANCHORS: list[tuple[str, str]] = [
    ("Scale Calculation", "M9"),
    ("Env Factor by Pool", "C7"),
    ("Scale Calculation-Vizo", "Q25"),
    ("Cover", "B35"),
]


def shift_historical_data_column_refs(
    workbook_path: str | Path,
    prior_col: str,
    target_col: str,
) -> dict:
    """Retarget every formula that references ``Historical Data!<prior_col>...``
    to ``Historical Data!<target_col>...``.

    When a quarter is generated by copying the prior quarter's
    workbook (carry-history mode) the new period's values are written
    into a new column on Historical Data, but the downstream tabs'
    formulas still point at the prior column. This walks every cell on
    every sheet and rewrites column-letter references inside formulas
    that mention ``Historical Data``. Cell refs that don't sit behind
    ``Historical Data!`` are not touched.

    Returns ``{ok, cells_updated, formulas_updated, prior_col,
    target_col, error}``.
    """
    result = {
        "ok": False,
        "cells_updated": 0,
        "formulas_updated": 0,
        "prior_col": prior_col,
        "target_col": target_col,
        "error": "",
    }
    if not prior_col or not target_col:
        result["error"] = "prior_col and target_col are required"
        return result
    if prior_col == target_col:
        result["ok"] = True
        return result
    if not _COL_LETTERS_RE.match(prior_col) or not _COL_LETTERS_RE.match(target_col):
        result["error"] = (
            f"prior_col/target_col must be A1 column letters; "
            f"got prior={prior_col!r} target={target_col!r}"
        )
        return result

    # Only rewrite a column letter that is (a) not preceded by another
    # letter (so ``BAY`` doesn't match ``AY``) and (b) directly followed
    # by a row digit (so ``AYZ`` literal text doesn't match). The
    # ``Historical Data`` substring test on the whole formula is the
    # primary scoping mechanism; the lookbehind/lookahead just keeps us
    # from corrupting unrelated cell refs inside the same formula.
    sub_pat = re.compile(
        r"(?<![A-Z])" + re.escape(prior_col) + r"(?=\d)"
    )
    try:
        wb = openpyxl.load_workbook(workbook_path)
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"Could not open workbook: {exc}"
        return result

    cells_updated = 0
    formulas_updated = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                if "Historical Data" not in v:
                    continue
                new_v, n = sub_pat.subn(target_col, v)
                if n and new_v != v:
                    cell.value = new_v
                    cells_updated += 1
                    formulas_updated += n
    if cells_updated:
        wb.save(workbook_path)
    result["ok"] = True
    result["cells_updated"] = cells_updated
    result["formulas_updated"] = formulas_updated
    return result


def _period_end_date(period: str) -> datetime | None:
    """``"YYYY-MM"`` / ``"YYYY_MM"`` -> ``datetime`` at month-end.

    Returns ``None`` if the period string can't be parsed.
    """
    import calendar
    m = re.match(r"^(\d{4})[-_](\d{1,2})$", (period or "").strip())
    if not m:
        return None
    y, mo = int(m.group(1)), int(m.group(2))
    last = calendar.monthrange(y, mo)[1]
    return datetime(y, mo, last)


def seed_new_historical_data_column(
    workbook_path: str | Path,
    target_col: str,
    period: str,
) -> dict:
    """Seed the row-1/2/3/4/5/8 cells in the new Historical Data column.

    When ``run_quarter_carry_history`` copies the prior quarter's
    workbook, the new period's data goes into a new column on the
    Historical Data tab but the column's header/snapshot cells are
    empty (the mapping CSV only covers the 5300 field rows, not the
    metadata rows). That leaves ``{col}8`` (quarter-end date) and the
    ``{col}2..{col}5`` Scale Calculation snapshot formulas blank, so
    downstream tabs that reference them error out.

    Mirrors what the manual workbook does when carried forward:
      - ``{col}1``  = "Copy and paste values before adding new column"
      - ``{col}2``  = ``='Scale Calculation'!U29``
      - ``{col}3``  = ``='Scale Calculation'!U30``
      - ``{col}4``  = ``='Scale Calculation'!U31``
      - ``{col}5``  = ``='Scale Calculation'!U33``
      - ``{col}8``  = the period-end date (e.g. ``3/31/2026``)

    Only writes cells that are currently empty so the function is safe
    to re-run.
    """
    result: dict[str, Any] = {
        "ok": False, "target_col": target_col, "period": period,
        "cells_written": [], "error": "",
    }
    if not target_col or not _COL_LETTERS_RE.match(target_col):
        result["error"] = f"invalid target_col {target_col!r}"
        return result
    end_date = _period_end_date(period)
    if end_date is None:
        result["error"] = f"could not parse period {period!r}"
        return result
    try:
        wb = openpyxl.load_workbook(workbook_path)
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"open failed: {exc}"
        return result
    if "Historical Data" not in wb.sheetnames:
        result["error"] = "Historical Data sheet missing"
        return result
    ws = wb["Historical Data"]
    desired = [
        ("1", "Copy and paste values before adding new column"),
        ("2", "='Scale Calculation'!U29"),
        ("3", "='Scale Calculation'!U30"),
        ("4", "='Scale Calculation'!U31"),
        ("5", "='Scale Calculation'!U33"),
        ("8", end_date),
    ]
    written = []
    for row, val in desired:
        coord = f"{target_col}{row}"
        cur = ws[coord].value
        if cur not in (None, ""):
            continue
        ws[coord] = val
        if row == "8":
            # Use a date number-format so Excel renders it as a date
            # rather than a serial number.
            ws[coord].number_format = "m/d/yyyy"
        written.append(coord)
    if written:
        wb.save(workbook_path)
    result["ok"] = True
    result["cells_written"] = written
    return result


def propagate_column_formulas(
    workbook_path: str | Path,
    prev_col: str,
    target_col: str,
    sheet: str = "Historical Data",
    start_row: int = 2,
) -> dict:
    """Drag-fill formulas from ``prev_col`` into ``target_col``.

    For every row >= ``start_row`` in ``sheet`` where the ``prev_col``
    cell holds a formula AND the ``target_col`` cell is empty, copies
    ``prev_col``'s formula into ``target_col`` and translates the
    relative cell references (e.g. ``AY`` → ``AZ``, ``$C$9`` stays).
    Mirrors Excel's "drag the fill handle one column right" behaviour.

    Returns ``{ok, cells_written: [coords], error}``. Only rewrites
    empty cells so it is safe to re-run.
    """
    from openpyxl.formula.translate import Translator
    result: dict[str, Any] = {
        "ok": False, "prev_col": prev_col, "target_col": target_col,
        "cells_written": [], "error": "",
    }
    if (not prev_col or not target_col
            or not _COL_LETTERS_RE.match(prev_col)
            or not _COL_LETTERS_RE.match(target_col)):
        result["error"] = (
            f"invalid prev_col/target_col {prev_col!r}/{target_col!r}")
        return result
    if prev_col == target_col:
        result["error"] = "prev_col == target_col, nothing to do"
        return result
    try:
        wb = openpyxl.load_workbook(workbook_path)
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"open failed: {exc}"
        return result
    if sheet not in wb.sheetnames:
        result["error"] = f"sheet {sheet!r} missing"
        return result
    ws = wb[sheet]
    written = []
    skipped_errors = []
    for row in range(start_row, ws.max_row + 1):
        src = ws[f"{prev_col}{row}"].value
        dst = ws[f"{target_col}{row}"].value
        if dst not in (None, ""):
            continue
        if not (isinstance(src, str) and src.startswith("=")):
            continue
        try:
            new_formula = Translator(
                src, origin=f"{prev_col}{row}",
            ).translate_formula(f"{target_col}{row}")
        except Exception as exc:  # noqa: BLE001
            skipped_errors.append((row, str(exc)))
            continue
        ws[f"{target_col}{row}"] = new_formula
        written.append(f"{target_col}{row}")
    if written:
        wb.save(workbook_path)
    result["ok"] = True
    result["cells_written"] = written
    if skipped_errors:
        result["error"] = (
            f"{len(skipped_errors)} formula(s) failed to translate; "
            f"first: r{skipped_errors[0][0]} {skipped_errors[0][1]}"
        )
    return result


# Historical Data rows 139-142 are the four environmental factors
# (unemployment, foreclosures, bankruptcies, population) that mirror
# the CECL Migration Model's Step 7 (Economic Data) fetch. Carry-
# history just clones the prior quarter's column, so the target column
# either holds stale numeric values from the prior quarter or, when
# the prior column was a "source-label" position, stray strings like
# 'BLS' / 'Sofi' / 'EAFCR'. Always overwrite with fresh fetched
# values; fall back to the prior column's numeric value when a fetch
# fails or no API is available (foreclosures has no free federal
# source).
_ENV_FACTOR_ROW_MAP: list[tuple[int, str]] = [
    (139, "unemployment_rate"),
    (140, "foreclosures"),
    (141, "bankruptcies"),
    (142, "population"),
]


def apply_env_factors_to_historical_data(
    workbook_path: str | Path,
    target_col: str,
    prior_col: str,
    state_name: str,
    county_name: str = "",
    sheet: str = "Historical Data",
) -> dict:
    """Write fresh environmental-factor values into the new quarter
    column on ``Historical Data`` rows 139-142.

    Calls :func:`fetch_econ_data.fetch_economic_data` (same source as
    the Migration Model) and writes results to ``{target_col}139``
    (unemployment), ``{target_col}141`` (bankruptcies), and
    ``{target_col}142`` (population). Foreclosures (``{target_col}140``)
    has no federal API and always falls back to the prior column.
    When a fetch fails for any other key, that row also falls back to
    the prior column. Target cells are always overwritten (the
    carry-clone may leave stale source-label strings there).
    """
    result: dict[str, Any] = {
        "ok": False,
        "target_col": target_col,
        "prior_col": prior_col,
        "state_name": state_name,
        "county_name": county_name,
        "fetched": {},
        "fallback_rows": [],
        "cells_written": [],
        "error": "",
    }
    if not target_col or not _COL_LETTERS_RE.match(target_col):
        result["error"] = f"invalid target_col {target_col!r}"
        return result
    if not prior_col or not _COL_LETTERS_RE.match(prior_col):
        result["error"] = f"invalid prior_col {prior_col!r}"
        return result
    fetched: dict[str, Any] = {}
    if state_name:
        try:
            import importlib
            fed = importlib.import_module("fetch_econ_data")
            fetched = fed.fetch_economic_data(state_name, county_name) or {}
        except Exception as exc:  # noqa: BLE001
            result["error"] = f"fetch failed: {exc}"
            fetched = {}
    else:
        result["error"] = "state_name empty; using prior column for all rows"
    try:
        wb = openpyxl.load_workbook(workbook_path)
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"open failed: {exc}"
        return result
    if sheet not in wb.sheetnames:
        result["error"] = f"sheet {sheet!r} missing"
        return result
    ws = wb[sheet]
    cells_written: list[str] = []
    for row, key in _ENV_FACTOR_ROW_MAP:
        value = fetched.get(key)
        # Coerce empty / missing / non-numeric to None so we fall back.
        if isinstance(value, str):
            try:
                value = float(value)
            except (TypeError, ValueError):
                value = None
        if value is None or value == 0:
            prior_val = ws[f"{prior_col}{row}"].value
            value = prior_val
            result["fallback_rows"].append(row)
        else:
            result["fetched"][key] = value
        coord = f"{target_col}{row}"
        ws[coord] = value
        cells_written.append(coord)
    wb.save(workbook_path)
    result["cells_written"] = cells_written
    result["ok"] = True
    return result


def run_single_quarter(state: dict, workspace_root: str) -> dict:
    """Drive a one-quarter SCALE fill from wizard state.

    Returns a result dict suitable for stashing back into
    ``state["scale"]["last_run"]`` and rendering on the review page.
    """
    scale = state.get("scale") or {}
    period = (scale.get("period") or "").strip()
    solr_url = (scale.get("solr_url") or "").strip()
    solr_core = (scale.get("solr_core") or "").strip()
    charter_raw = state.get("charter_number") or ""
    short = state.get("short_name") or ""

    errors: list[str] = []
    if not period:
        errors.append("Target period (YYYY-MM) is required.")
    if not solr_url:
        errors.append("Solr URL is required.")
    if not solr_core:
        errors.append("Solr core is required.")
    if not charter_raw:
        errors.append("Charter number is missing on the Identity step.")
    if errors:
        return {"ok": False, "errors": errors, "ran_at": "", "output_path": ""}
    try:
        charter = int(charter_raw)
    except (TypeError, ValueError):
        return {
            "ok": False,
            "errors": [f"Charter number {charter_raw!r} is not numeric."],
            "ran_at": "",
            "output_path": "",
        }

    tmpl = template_loader.resolve_template(
        period, scale.get("template_override_path") or None
    )
    if not tmpl["ok"]:
        return {"ok": False, "errors": [tmpl["message"]], "ran_at": "",
                "output_path": ""}
    mp = template_loader.resolve_map(
        period, scale.get("map_override_path") or None
    )
    if not mp["ok"]:
        return {"ok": False, "errors": [mp["message"]], "ran_at": "",
                "output_path": ""}

    try:
        doc = solr_fetcher.fetch_doc(
            solr_url=solr_url,
            core=solr_core,
            charter=charter,
            period=period,
            username=scale.get("solr_user") or None,
            password=scale.get("solr_pass") or None,
        )
    except LookupError as exc:
        return {"ok": False, "errors": [str(exc)], "ran_at": "",
                "output_path": ""}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "errors": [f"Solr fetch failed: {exc}"],
                "ran_at": "", "output_path": ""}

    try:
        rows = mapping_loader.load_rows(mp["path"])
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "errors": [f"Mapping load failed: {exc}"],
                "ran_at": "", "output_path": ""}

    out_path = _output_path(workspace_root, short, period)
    fill = fill_template(tmpl["path"], out_path, rows, doc)

    qf_entries = _qfactor_entries_from_state(state)
    qf_result = apply_qfactors(out_path, qf_entries)

    mgmt_state = (state.get("scale") or {}).get("mgmt_adj") or {}
    mgmt_state_norm = _normalize_mgmt_adj_state(mgmt_state, tmpl["path"])
    mgmt_result = mgmt_adj_writer.apply_mgmt_adj(out_path, mgmt_state_norm)

    imp_rows = _impaired_rows_from_state(state)
    imp_result = impaired_loader.apply_impaired_rows(out_path, imp_rows)

    env_result = env_factor_writer.apply_env_factor_ranges(out_path)

    variant = _report_variant_from_state(state)
    variant_result = apply_report_variant(out_path, variant)
    primary_output = (
        variant_result["outputs"][0]["path"]
        if variant_result["outputs"] else str(out_path)
    )

    prior_acl = _inject_prior_acl(
        workspace_root, short, period, variant_result["outputs"]
    )

    return {
        "ok": True,
        "ran_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "output_path": primary_output,
        "output_files": variant_result["outputs"],
        "period": period,
        "charter": charter,
        "prior_acl": prior_acl,
        "template_path": tmpl["path"],
        "template_source": tmpl["source"],
        "template_message": tmpl.get("message", ""),
        "map_path": mp["path"],
        "map_source": mp["source"],
        "applied": fill["applied"],
        "total_rows": len(rows),
        "issues": fill["issues"],
        "missing_fields": fill["missing_fields"],
        "missing_sheets": fill["missing_sheets"],
        "qfactor_applied": qf_result["applied"],
        "qfactor_total": qf_result["total"],
        "qfactor_missing_sheets": qf_result["missing_sheets"],
        "qfactor_issues": qf_result["issues"],
        "mgmt_adj_ok": mgmt_result["ok"],
        "mgmt_adj_pools_written": mgmt_result["pools_written"],
        "mgmt_adj_default_written": mgmt_result["default_written"],
        "mgmt_adj_portfolio_written": mgmt_result["portfolio_written"],
        "mgmt_adj_error": mgmt_result["error"],
        "impaired_applied": imp_result["applied"],
        "impaired_cleared": imp_result["cleared"],
        "impaired_error": imp_result["error"],
        "env_factor_ok": env_result["ok"],
        "env_factor_applied_delq": env_result["applied_delq"],
        "env_factor_applied_econ": env_result["applied_econ"],
        "env_factor_error": env_result["error"],
        "report_variant": variant_result["variant"],
        "report_variant_hidden": variant_result["hidden"],
        "errors": [],
    }


def run_multi_quarter(
    state: dict,
    workspace_root: str,
    quarters: int = 32,
) -> dict:
    """Drive a multi-quarter SCALE fill from wizard state.

    Iterates newest -> oldest. Resolves the starting template once
    (period-exact if available, else newest <= start). After the first
    successful iteration, the output file becomes the template for
    subsequent quarters so writes accumulate. Each period MUST have an
    exact-match mapping CSV in ``cecl_ui/data/scale/maps/`` or that
    quarter is skipped (recorded in ``skipped``).

    Returns ``{ok, ran_at, output_path, periods, iterations, skipped,
    errors}`` where ``iterations`` is one entry per attempted quarter.
    """
    scale = state.get("scale") or {}
    start_period = (scale.get("period") or "").strip()
    solr_url = (scale.get("solr_url") or "").strip()
    solr_core = (scale.get("solr_core") or "").strip()
    charter_raw = state.get("charter_number") or ""
    short = state.get("short_name") or ""

    errors: list[str] = []
    if not start_period:
        errors.append("Target period (YYYY-MM) is required.")
    if not solr_url:
        errors.append("Solr URL is required.")
    if not solr_core:
        errors.append("Solr core is required.")
    if not charter_raw:
        errors.append("Charter number is missing on the Identity step.")
    if errors:
        return {
            "ok": False, "errors": errors, "ran_at": "",
            "output_path": "", "periods": [], "iterations": [], "skipped": [],
        }
    try:
        charter = int(charter_raw)
    except (TypeError, ValueError):
        return {
            "ok": False,
            "errors": [f"Charter number {charter_raw!r} is not numeric."],
            "ran_at": "", "output_path": "", "periods": [],
            "iterations": [], "skipped": [],
        }

    try:
        periods = make_periods(start_period, quarters)
    except Exception as exc:  # noqa: BLE001
        return {
            "ok": False, "errors": [f"Period list failed: {exc}"],
            "ran_at": "", "output_path": "", "periods": [],
            "iterations": [], "skipped": [],
        }

    # Resolve the starting template once. Per-CU override always wins.
    tmpl = template_loader.resolve_template(
        start_period, scale.get("template_override_path") or None
    )
    if not tmpl["ok"]:
        return {
            "ok": False, "errors": [tmpl["message"]], "ran_at": "",
            "output_path": "", "periods": periods, "iterations": [],
            "skipped": [],
        }

    out_path = _output_path(workspace_root, short, start_period)
    iterations: list[dict] = []
    skipped: list[dict] = []
    successes = 0
    current_template: str = tmpl["path"]

    for idx, p in enumerate(periods, start=1):
        # Per-quarter mapping override only applies to the starting
        # period (the user uploaded it for that period); older quarters
        # always use canonical maps.
        override_map = (
            scale.get("map_override_path")
            if p == start_period else None
        )
        mp = template_loader.resolve_map(p, override_map or None)
        if not mp["ok"]:
            skipped.append({"period": p, "reason": mp["message"]})
            continue

        try:
            doc = solr_fetcher.fetch_doc(
                solr_url=solr_url,
                core=solr_core,
                charter=charter,
                period=p,
                username=scale.get("solr_user") or None,
                password=scale.get("solr_pass") or None,
            )
        except LookupError as exc:
            skipped.append({"period": p, "reason": f"Solr: {exc}"})
            continue
        except Exception as exc:  # noqa: BLE001
            skipped.append({"period": p, "reason": f"Solr fetch failed: {exc}"})
            continue

        try:
            rows = mapping_loader.load_rows(mp["path"])
        except Exception as exc:  # noqa: BLE001
            skipped.append({"period": p, "reason": f"Map load failed: {exc}"})
            continue

        fill = fill_template(current_template, out_path, rows, doc)
        successes += 1
        iterations.append({
            "period": p,
            "index": idx,
            "map_path": mp["path"],
            "map_source": mp["source"],
            "applied": fill["applied"],
            "total_rows": len(rows),
            "missing_fields_count": len(fill["missing_fields"]),
            "missing_sheets_count": len(fill["missing_sheets"]),
            "issues_count": len(fill["issues"]),
        })
        # After first successful write, accumulate into the output file.
        current_template = str(out_path)

    # Apply Q-factor overlays once at the end so they sit on top of the
    # most recent 5300 writes (won't be overwritten by accumulation).
    qf_result = {"applied": 0, "total": 0, "missing_sheets": [], "issues": []}
    imp_result = {"applied": 0, "cleared": 0, "error": ""}
    mgmt_result = {"ok": False, "pools_written": 0, "default_written": False,
                   "portfolio_written": False, "error": ""}
    variant_result: dict = {
        "variant": "both", "hidden": [], "kept": [], "outputs": [],
    }
    env_result = {"ok": False, "applied_delq": 0, "applied_econ": 0,
                  "skipped": [], "error": ""}
    if successes > 0:
        qf_entries = _qfactor_entries_from_state(state)
        qf_result = apply_qfactors(out_path, qf_entries)
        mgmt_state = (state.get("scale") or {}).get("mgmt_adj") or {}
        mgmt_state_norm = _normalize_mgmt_adj_state(mgmt_state, tmpl["path"])
        mgmt_result = mgmt_adj_writer.apply_mgmt_adj(out_path, mgmt_state_norm)
        imp_rows = _impaired_rows_from_state(state)
        imp_result = impaired_loader.apply_impaired_rows(out_path, imp_rows)
        env_result = env_factor_writer.apply_env_factor_ranges(out_path)
        variant_result = apply_report_variant(
            out_path, _report_variant_from_state(state)
        )

    primary_output = (
        variant_result["outputs"][0]["path"]
        if variant_result.get("outputs") else str(out_path)
    )

    prior_acl = _inject_prior_acl(
        workspace_root, short, start_period, variant_result.get("outputs") or []
    ) if successes > 0 else {"prior_period": "", "prior_path": "", "applied": [], "error": ""}
    return {
        "ok": successes > 0,
        "ran_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "output_path": primary_output,
        "output_files": variant_result.get("outputs", []),
        "period": start_period,
        "charter": charter,
        "quarters_requested": quarters,
        "quarters_written": successes,
        "periods": periods,
        "iterations": iterations,
        "skipped": skipped,
        "template_path": tmpl["path"],
        "template_source": tmpl["source"],
        "template_message": tmpl.get("message", ""),
        "qfactor_applied": qf_result["applied"],
        "qfactor_total": qf_result["total"],
        "qfactor_missing_sheets": qf_result["missing_sheets"],
        "qfactor_issues": qf_result["issues"],
        "mgmt_adj_ok": mgmt_result["ok"],
        "mgmt_adj_pools_written": mgmt_result["pools_written"],
        "mgmt_adj_default_written": mgmt_result["default_written"],
        "mgmt_adj_portfolio_written": mgmt_result["portfolio_written"],
        "mgmt_adj_error": mgmt_result["error"],
        "impaired_applied": imp_result["applied"],
        "impaired_cleared": imp_result["cleared"],
        "impaired_error": imp_result["error"],
        "env_factor_ok": env_result["ok"],
        "env_factor_applied_delq": env_result["applied_delq"],
        "env_factor_applied_econ": env_result["applied_econ"],
        "env_factor_error": env_result["error"],
        "report_variant": variant_result["variant"],
        "report_variant_hidden": variant_result["hidden"],
        "prior_acl": prior_acl,
        "errors": [] if successes > 0 else ["No quarters were successfully written."],
    }


def _inject_prior_acl(workspace_root: str, short: str, period: str,
                      output_files: list[dict]) -> dict:
    """Hard-code Prior ACL values from the previous quarter's report
    into Historical Data!AY2..AY5 of every output file.

    Looks up `Generated_Reports/<short>/<prev_quarter>/` for the
    newest SCALE workbook (prefers Vizo variant) and copies its
    Total Expected Losses / ACL Balance / Adjustment / ACL Ratio
    cells. Silently no-ops when the prior report is missing -- e.g.
    first ever run for this CU.
    """
    result: dict[str, Any] = {
        "prior_period": "", "prior_path": "",
        "applied": [], "error": "",
    }
    try:
        prev = prev_quarter(period)
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"Could not compute prior quarter: {exc}"
        return result
    result["prior_period"] = prev
    prior_path = runs_service.find_prior_report(workspace_root, short, prev)
    if prior_path is None:
        # Fallback: newest report strictly older than target period.
        fallback = runs_service.find_newest_prior_report(
            workspace_root, short, period
        )
        if fallback is None:
            return result
        prev, prior_path = fallback
        result["prior_period"] = prev
    result["prior_path"] = str(prior_path)
    read = runs_service.read_prior_acl_values(prior_path)
    if not read["ok"]:
        result["error"] = read["error"]
        return result
    values = read["values"]
    for entry in output_files or []:
        path = entry.get("path") or ""
        if not path:
            continue
        write = runs_service.write_prior_acl_values(path, values)
        result["applied"].append({
            "path": path,
            "ok": write["ok"],
            "written": write["written"],
            "error": write["error"],
        })
    return result


def run_quarter_carry_history(state: dict, workspace_root: str) -> dict:
    """Run only the target quarter, carrying historical data from a
    prior quarter's report (the "normal process" mode).

    Copies the most recent prior SCALE workbook (`*_Vizo.xlsx`
    preferred) to the new output path, unhides any tabs the variant
    apply step previously hid, then overlays the target quarter's
    Solr mapping, Q-factors, mgmt-adj, impaired, and env factors --
    leaving every earlier quarter's data and the Historical Data tab
    intact.

    Returns a dict shaped like `run_single_quarter` plus extra
    `carry_*` provenance keys.
    """
    scale = state.get("scale") or {}
    period = (scale.get("period") or "").strip()
    solr_url = (scale.get("solr_url") or "").strip()
    solr_core = (scale.get("solr_core") or "").strip()
    charter_raw = state.get("charter_number") or ""
    short = state.get("short_name") or ""

    errors: list[str] = []
    if not period:
        errors.append("Target period (YYYY-MM) is required.")
    if not solr_url:
        errors.append("Solr URL is required.")
    if not solr_core:
        errors.append("Solr core is required.")
    if not charter_raw:
        errors.append("Charter number is missing on the Identity step.")
    if not short:
        errors.append("Short name is missing on the Identity step.")
    if errors:
        return {"ok": False, "errors": errors, "ran_at": "",
                "output_path": ""}
    try:
        charter = int(charter_raw)
    except (TypeError, ValueError):
        return {
            "ok": False,
            "errors": [f"Charter number {charter_raw!r} is not numeric."],
            "ran_at": "", "output_path": "",
        }

    try:
        prev_period = prev_quarter(period)
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "errors": [str(exc)], "ran_at": "",
                "output_path": ""}

    prior_path = runs_service.find_prior_report(workspace_root, short, prev_period)
    if prior_path is None:
        # Search further back -- the user may have skipped a quarter.
        fallback = runs_service.find_newest_prior_report(
            workspace_root, short, period
        )
        if fallback is None:
            return {
                "ok": False,
                "errors": [
                    f"No prior SCALE report found for {short}. Run the "
                    "'Re-pull all from 5300' mode for the first run, "
                    "or generate a baseline through the wizard.",
                ],
                "ran_at": "", "output_path": "",
            }
        prev_period, prior_path = fallback

    mp = template_loader.resolve_map(
        period, scale.get("map_override_path") or None
    )
    if not mp["ok"]:
        return {"ok": False, "errors": [mp["message"]], "ran_at": "",
                "output_path": ""}

    try:
        doc = solr_fetcher.fetch_doc(
            solr_url=solr_url,
            core=solr_core,
            charter=charter,
            period=period,
            username=scale.get("solr_user") or None,
            password=scale.get("solr_pass") or None,
        )
    except LookupError as exc:
        return {"ok": False, "errors": [str(exc)], "ran_at": "",
                "output_path": ""}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "errors": [f"Solr fetch failed: {exc}"],
                "ran_at": "", "output_path": ""}

    try:
        rows = mapping_loader.load_rows(mp["path"])
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "errors": [f"Mapping load failed: {exc}"],
                "ran_at": "", "output_path": ""}

    out_path = _output_path(workspace_root, short, period)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(prior_path, out_path)
    runs_service.unhide_all_sheets(out_path)

    # Now overlay the target quarter's data on top of the carried workbook.
    fill = fill_template(out_path, out_path, rows, doc)

    # The carried workbook's downstream formulas (Scale Calculation,
    # Env Factor by Pool, Executive Summary-Vizo, Cover, etc.) still
    # reference the prior quarter's Historical Data column. Shift them
    # to the target quarter's column so the new 5300 data actually
    # feeds the calcs (mirrors the manual "drag the formula one column
    # right" step in the analyst workflow).
    target_col = _target_column_from_mapping_rows(rows)
    prior_col = _detect_prior_column(out_path, _PRIOR_COLUMN_ANCHORS)
    if not prior_col:
        # Fallback: derive from the prior period's canonical mapping.
        try:
            prior_mp = template_loader.resolve_map(prev_period, None)
            if prior_mp.get("ok"):
                prior_rows = mapping_loader.load_rows(prior_mp["path"])
                prior_col = _target_column_from_mapping_rows(prior_rows)
        except Exception:  # noqa: BLE001
            prior_col = ""
    shift_result = shift_historical_data_column_refs(
        out_path, prior_col, target_col,
    )

    # Seed the new Historical Data column's header/snapshot cells
    # ({col}1 label, {col}2..{col}5 Scale Calc snapshot formulas,
    # {col}8 quarter-end date). The mapping CSV only covers the 5300
    # data rows so these would otherwise be blank, causing #VALUE!/
    # #REF! errors in downstream tabs that reference them (Calc tab,
    # Executive Summary-Vizo, Cover, etc.).
    seed_result = seed_new_historical_data_column(
        out_path, target_col, period,
    )

    # Drag-fill prior column's formulas into the new column for rows
    # below the 5300-mapped data band (Management Adjustments,
    # Combined Balances, the per-pool SUMIFs in rows 162-245, etc.).
    # The mapping CSV doesn't cover those rows; without this step they
    # stay blank and Scale Calculation / Env Factor by Pool tabs that
    # reference them error out.
    propagate_result = propagate_column_formulas(
        out_path, prior_col, target_col,
        sheet="Historical Data", start_row=2,
    )

    # Refresh the four environmental factors in the new column
    # (Historical Data rows 139-142). Mirrors the Migration Model's
    # Step 7 fetch. Falls back to the prior column's values for any
    # missing fetched key (foreclosures has no federal API).
    econ_state = (state.get("economic_data") or {})
    env_state_name = str(econ_state.get("state") or "").strip()
    env_county_name = str(econ_state.get("county") or "").strip()
    env_factors_result = apply_env_factors_to_historical_data(
        out_path, target_col, prior_col,
        env_state_name, env_county_name,
    )

    qf_entries = _qfactor_entries_from_state(state)
    qf_result = apply_qfactors(out_path, qf_entries)

    mgmt_state = (state.get("scale") or {}).get("mgmt_adj") or {}
    mgmt_state_norm = _normalize_mgmt_adj_state(mgmt_state, str(out_path))
    mgmt_result = mgmt_adj_writer.apply_mgmt_adj(out_path, mgmt_state_norm)

    imp_rows = _impaired_rows_from_state(state)
    imp_result = impaired_loader.apply_impaired_rows(out_path, imp_rows)

    env_result = env_factor_writer.apply_env_factor_ranges(out_path)

    variant = _report_variant_from_state(state)
    variant_result = apply_report_variant(out_path, variant)
    primary_output = (
        variant_result["outputs"][0]["path"]
        if variant_result["outputs"] else str(out_path)
    )

    prior_acl = _inject_prior_acl(
        workspace_root, short, period, variant_result["outputs"]
    )

    return {
        "ok": True,
        "ran_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "output_path": primary_output,
        "output_files": variant_result["outputs"],
        "period": period,
        "charter": charter,
        "carry_mode": True,
        "carry_from_period": prev_period,
        "carry_from_path": str(prior_path),
        "map_path": mp["path"],
        "map_source": mp["source"],
        "applied": fill["applied"],
        "total_rows": len(rows),
        "issues": fill["issues"],
        "missing_fields": fill["missing_fields"],
        "missing_sheets": fill["missing_sheets"],
        "col_shift_ok": shift_result["ok"],
        "col_shift_prior": shift_result["prior_col"],
        "col_shift_target": shift_result["target_col"],
        "col_shift_cells_updated": shift_result["cells_updated"],
        "col_shift_formulas_updated": shift_result["formulas_updated"],
        "col_shift_error": shift_result["error"],
        "col_seed_ok": seed_result["ok"],
        "col_seed_cells_written": seed_result["cells_written"],
        "col_seed_error": seed_result["error"],
        "col_propagate_ok": propagate_result["ok"],
        "col_propagate_cells_written": propagate_result["cells_written"],
        "col_propagate_error": propagate_result["error"],
        "col_env_ok": env_factors_result["ok"],
        "col_env_fetched": env_factors_result["fetched"],
        "col_env_fallback_rows": env_factors_result["fallback_rows"],
        "col_env_cells_written": env_factors_result["cells_written"],
        "col_env_state": env_factors_result["state_name"],
        "col_env_county": env_factors_result["county_name"],
        "col_env_error": env_factors_result["error"],
        "qfactor_applied": qf_result["applied"],
        "qfactor_total": qf_result["total"],
        "qfactor_missing_sheets": qf_result["missing_sheets"],
        "qfactor_issues": qf_result["issues"],
        "mgmt_adj_ok": mgmt_result["ok"],
        "mgmt_adj_pools_written": mgmt_result["pools_written"],
        "mgmt_adj_default_written": mgmt_result["default_written"],
        "mgmt_adj_portfolio_written": mgmt_result["portfolio_written"],
        "mgmt_adj_error": mgmt_result["error"],
        "impaired_applied": imp_result["applied"],
        "impaired_cleared": imp_result["cleared"],
        "impaired_error": imp_result["error"],
        "env_factor_ok": env_result["ok"],
        "env_factor_applied_delq": env_result["applied_delq"],
        "env_factor_applied_econ": env_result["applied_econ"],
        "env_factor_error": env_result["error"],
        "report_variant": variant_result["variant"],
        "report_variant_hidden": variant_result["hidden"],
        "prior_acl": prior_acl,
        "errors": [],
    }
