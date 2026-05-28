"""Enforce wrap-text and row-height on the Vizo Explanation-of-ACL-Calc tab.

Templates are patched on disk, but `run_quarter_carry_history` seeds new
output workbooks from the *previous quarter's report* rather than the
master template, so old reports that pre-date the formatting fix carry
the wrong row heights / unwrapped text forward forever. This module
re-applies the fix to any workbook handed to it so every newly-written
output is correct regardless of how it was seeded.

Keep in sync with `scripts/fix_vizo_explanation_wrap.py` (the one-shot
template patcher); the two share the same intent and constants.
"""
from __future__ import annotations

import math
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

SHEET = "Explanation of ACL Calc-Vizo"
TARGET_ROWS = (19, 23, 25)
COLUMN = "A"

CHARS_PER_LINE = 105
LINE_HEIGHT = 15.75
PADDING_LINES = 0.3


def _compute_height(text: str) -> float:
    lines = max(1, math.ceil(len(text) / CHARS_PER_LINE))
    return round((lines + PADDING_LINES) * LINE_HEIGHT, 2)


def apply_vizo_explanation_wrap(workbook_path: str | Path) -> dict:
    """Set wrap_text + grow row height on rows 19/23/25 of the Vizo
    Explanation tab so long paragraphs render fully. No-op if the sheet
    is missing (TCT-only variants). Never raises."""
    path = Path(workbook_path)
    result: dict = {"ok": False, "applied": [], "error": ""}
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"load failed: {exc}"
        return result

    try:
        if SHEET not in wb.sheetnames:
            result["ok"] = True
            return result
        ws = wb[SHEET]
        changed = False
        for row in TARGET_ROWS:
            cell = ws[f"{COLUMN}{row}"]
            text = "" if cell.value is None else str(cell.value)
            if not text.strip():
                continue

            existing = cell.alignment or Alignment()
            cell.alignment = Alignment(
                horizontal=existing.horizontal,
                vertical="top",
                wrap_text=True,
                indent=existing.indent,
                shrink_to_fit=False,
                text_rotation=existing.text_rotation,
            )

            needed = _compute_height(text)
            rd = ws.row_dimensions[row]
            current = rd.height or 0
            if needed > current + 0.1:
                rd.height = needed
                result["applied"].append(
                    f"A{row}: height {current:.2f}->{needed:.2f}"
                )
            else:
                result["applied"].append(
                    f"A{row}: wrap set (height ok at {current:.2f})"
                )
            changed = True

        if changed:
            wb.save(path)
        result["ok"] = True
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"apply failed: {exc}"
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass
    return result
