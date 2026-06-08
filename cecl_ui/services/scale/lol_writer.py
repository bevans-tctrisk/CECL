"""SCALE Life-of-Loan months writer.

The SCALE template's ``Industry Data`` tab holds the per-pool
"ACL Lifetime Months" (Weighted Average Maturity) values that drive
the Life-of-Loan loss-rate calculation on the ``Calc tab`` (rows
``AF7:AF19`` reference ``Industry Data!B3:B15``).

The 13 ``Industry Data`` pool rows correspond 1:1 with
``Scale Calculation!C9:C21`` (mirrored via formulas ``='Scale
Calculation'!C9`` in column A). Each call to :func:`read_lol_months`
joins the two views by row index so callers receive a list of
``{name, months}`` entries in template order.

Wizard state stores **overrides only** as
``state["scale"]["life_of_loan_overrides"]`` keyed by pool name, so
the per-CU edit survives a template-pool reorder. :func:`apply_lol_overrides`
writes the overrides to a generated workbook before the run; pools
without an override keep the template's baseline value.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


SHEET_NAME = "Industry Data"
MONTHS_COL = 2          # column B
ROW_START = 3           # B3 corresponds to Scale Calculation!C9
ROW_END = 15            # B15 corresponds to Scale Calculation!C21

# Reuse the same pool-name source as the Mgmt Adj writer so both
# wizard steps stay in lock-step with the template.
POOL_REF_SHEET = "Scale Calculation"
POOL_REF_COL = 3
POOL_REF_ROW_START = 9
POOL_REF_ROW_END = 21


def _coerce_months(value: Any) -> int | None:
    """Return an int >= 0, or None when the cell is blank/non-numeric."""
    if value is None or value == "":
        return None
    try:
        n = int(round(float(value)))
    except (TypeError, ValueError):
        return None
    if n < 0:
        return None
    return n


def read_lol_months(template_path: str | Path) -> list[dict[str, Any]]:
    """Return the per-pool LoL months from the template, in template order.

    Output shape::

        [{"name": str, "months": int | None, "row": int}, ...]

    ``row`` is the 1-based ``Industry Data`` row (3..15) so the writer
    can target it directly. Empty pool names are skipped (e.g. a 13th
    row that's not in use for that template).
    """
    try:
        wb = openpyxl.load_workbook(template_path, data_only=False)
    except Exception:  # noqa: BLE001
        return []
    try:
        if POOL_REF_SHEET not in wb.sheetnames:
            return []
        names_ws = wb[POOL_REF_SHEET]
        names: list[str] = []
        for r in range(POOL_REF_ROW_START, POOL_REF_ROW_END + 1):
            v = names_ws.cell(row=r, column=POOL_REF_COL).value
            if v is None:
                names.append("")
                continue
            s = str(v).strip()
            if s.lower() == "total":
                names.append("")
                continue
            names.append(s)

        months: list[int | None] = [None] * len(names)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
            for i in range(len(names)):
                row_no = ROW_START + i
                if row_no > ROW_END:
                    break
                months[i] = _coerce_months(ws.cell(row=row_no, column=MONTHS_COL).value)

        out: list[dict[str, Any]] = []
        for i, name in enumerate(names):
            if not name:
                continue
            out.append({
                "name": name,
                "months": months[i],
                "row": ROW_START + i,
            })
        return out
    finally:
        wb.close()


def apply_lol_overrides(
    workbook_path: str | Path, overrides: dict[str, Any] | None,
) -> dict:
    """Write per-pool Life-of-Loan month overrides to ``workbook_path``.

    ``overrides`` is the wizard state dict ``{pool_name: months_int}``.
    Pool names not present in the dict (or with non-positive values)
    are left untouched so the template's defaults flow through.

    Returns ``{ok, sheet_missing, pools_written, skipped, error}``.
    """
    result: dict[str, Any] = {
        "ok": False,
        "sheet_missing": False,
        "pools_written": 0,
        "skipped": [],
        "error": "",
    }
    cleaned: dict[str, int] = {}
    if isinstance(overrides, dict):
        for k, v in overrides.items():
            if not k:
                continue
            n = _coerce_months(v)
            if n is None or n <= 0:
                continue
            cleaned[str(k).strip()] = n
    if not cleaned:
        result["ok"] = True
        return result

    try:
        wb = openpyxl.load_workbook(workbook_path)
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"Could not open workbook: {exc}"
        return result
    try:
        if SHEET_NAME not in wb.sheetnames or POOL_REF_SHEET not in wb.sheetnames:
            result["sheet_missing"] = True
            result["error"] = (
                f"Sheet {SHEET_NAME!r} or {POOL_REF_SHEET!r} not found."
            )
            return result
        names_ws = wb[POOL_REF_SHEET]
        ws = wb[SHEET_NAME]

        # Build name -> Industry Data row map.
        pool_to_row: dict[str, int] = {}
        for i in range(POOL_REF_ROW_END - POOL_REF_ROW_START + 1):
            v = names_ws.cell(
                row=POOL_REF_ROW_START + i, column=POOL_REF_COL,
            ).value
            if v is None:
                continue
            s = str(v).strip()
            if not s or s.lower() == "total":
                continue
            pool_to_row[s] = ROW_START + i

        for name, months in cleaned.items():
            row_no = pool_to_row.get(name)
            if row_no is None:
                result["skipped"].append(name)
                continue
            ws.cell(row=row_no, column=MONTHS_COL).value = months
            result["pools_written"] += 1

        wb.save(workbook_path)
        result["ok"] = True
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"Write failed: {exc}"
    finally:
        wb.close()
    return result
