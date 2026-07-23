"""Discovery + helpers for running new SCALE quarters against an
already-configured credit union.

The "set up" status of a CU is inferred from two sources:

* ``wizard_drafts/<short>.json`` — the persisted wizard state. Required
  to rehydrate solr, q-factor overrides, mgmt-adj, impaired settings,
  etc. without forcing the user back through the wizard.
* ``Generated_Reports/<short>/<period>/CECL_SCALE_<short>*.xlsx`` —
  the actual quarterly outputs. We use the newest of these as the
  source for "carry historical from prior report" and for the
  Executive Summary-Vizo Prior ACL section.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

from cecl_ui.services import wizard_drafts


# Glob accepts both the historical (CECL_SCALE_<short>...xlsx) and the
# YYYY-MM-prefixed (2026-03_CECL_SCALE_<short>...xlsx) layouts so older
# runs stay discoverable after the naming change.
_REPORT_GLOB = "*CECL_SCALE_*.xlsx"

# Source cells in the prior workbook that drive the Prior ACL block on
# Executive Summary-Vizo. In a fresh template these are the formula
# targets of Historical Data!AZ2..AZ5. We read them from the prior xlsx
# (cached values written by Excel; openpyxl raw values for the U30 row
# which is itself a formula) and write them as hard-coded numbers into
# Historical Data!AY2..AY5 of the new workbook so the Vizo Prior ACL
# OFFSET formulas (=OFFSET('Historical Data'!AZ2,0,-1)) resolve.
_PRIOR_ACL_SOURCES = [
    ("Scale Calculation", "U29", "AY2"),  # Total Expected Losses on Loans
    ("Scale Calculation", "U30", "AY3"),  # Current ACL Balance (formula → Historical Data!AZ177)
    ("Scale Calculation", "U31", "AY4"),  # Adjustment
    ("Scale Calculation", "U33", "AY5"),  # ACL/Total Loans
]


def _reports_root(workspace_root: str | Path) -> Path:
    return Path(workspace_root) / "Generated_Reports"


def _period_sort_key(period: str) -> tuple[int, int]:
    try:
        y, m = period.split("-")
        return (int(y), int(m))
    except (ValueError, AttributeError):
        return (0, 0)


def list_scale_cus(workspace_root: str | Path) -> list[dict[str, Any]]:
    """Return one entry per CU that has a SCALE workbook on disk OR a
    completed SCALE wizard draft (so freshly-configured CUs surface on
    the SCALE Runs index before their first quarter has been run).

    Each entry::

        {
          "short_name": "connections_cu",
          "credit_union": "Connections Credit Union",  # from draft if available
          "latest_period": "2025-12",  # empty string if no workbook yet
          "latest_path": "Z:/.../CECL_SCALE_connections_cu_Vizo.xlsx",
          "periods": ["2025-12", "2025-09", ...],
          "draft_present": True,
        }
    """
    root = _reports_root(workspace_root)

    # Index drafts by slug. We accept either a SCALE draft (preferred,
    # has the Solr/period config) or fall back to a Migration draft
    # purely so the dashboard can show the CU's display name.
    drafts_by_slug: dict[str, dict[str, Any]] = {}
    for d in (wizard_drafts.list_drafts(workspace_root) or []):
        slug = d.get("key") or ""
        if not slug:
            continue
        # Prefer SCALE entry; only let migration win if no SCALE exists.
        existing = drafts_by_slug.get(slug)
        if existing is None or (
            existing.get("model") != "scale" and d.get("model") == "scale"
        ):
            drafts_by_slug[slug] = d

    out: list[dict[str, Any]] = []
    seen_shorts: set[str] = set()
    if root.exists():
        for cu_dir in sorted(root.iterdir()):
            if not cu_dir.is_dir():
                continue
            short = cu_dir.name
            periods: list[str] = []
            for period_dir in cu_dir.iterdir():
                if not period_dir.is_dir():
                    continue
                if not any(period_dir.glob(_REPORT_GLOB)):
                    continue
                periods.append(period_dir.name)
            if not periods:
                continue
            periods.sort(key=_period_sort_key, reverse=True)
            latest = periods[0]
            latest_path = _pick_report_in_period(cu_dir / latest)
            draft = drafts_by_slug.get(short)
            out.append({
                "short_name": short,
                "credit_union": (draft or {}).get("credit_union") or short,
                "latest_period": latest,
                "latest_path": str(latest_path) if latest_path else "",
                "periods": periods,
                "draft_present": (
                    draft is not None and draft.get("model") == "scale"
                ),
            })
            seen_shorts.add(short)

    # Also surface CUs that have a completed SCALE wizard draft but no
    # workbook on disk yet (fresh setup, first-quarter run pending).
    # Without this, the SCALE Runs index would hide any CU the user
    # just finished configuring in the SCALE wizard.
    for slug, draft in drafts_by_slug.items():
        if slug in seen_shorts:
            continue
        if draft.get("model") != "scale":
            continue
        out.append({
            "short_name": slug,
            "credit_union": draft.get("credit_union") or slug,
            "latest_period": "",
            "latest_path": "",
            "periods": [],
            "draft_present": True,
        })

    out.sort(key=lambda r: r["credit_union"].lower())
    return out


def _pick_report_in_period(period_dir: Path) -> Path | None:
    """Pick the best SCALE workbook from a period directory.

    Prefers ``*_Vizo.xlsx`` (preserves the original theme) over
    ``*_TCT.xlsx`` over an un-suffixed master. All variants share the
    same Historical Data / Scale Calculation tabs so either works for
    history carry-over.
    """
    candidates = list(period_dir.glob(_REPORT_GLOB))
    if not candidates:
        return None

    def rank(p: Path) -> int:
        name = p.name.lower()
        if name.endswith("_vizo.xlsx"):
            return 0
        if name.endswith("_tct.xlsx"):
            return 1
        return 2

    candidates.sort(key=lambda p: (rank(p), p.name))
    return candidates[0]


def find_prior_report(
    workspace_root: str | Path,
    short_name: str,
    prior_period: str,
) -> Path | None:
    """Locate a workbook for the given prior period, or None.

    ``prior_period`` is the YYYY-MM the *previous* quarter — e.g. when
    running 2026-03 the caller passes "2025-12".
    """
    root = _reports_root(workspace_root) / short_name / prior_period
    if not root.exists():
        return None
    return _pick_report_in_period(root)


def find_newest_prior_report(
    workspace_root: str | Path,
    short_name: str,
    before_period: str,
) -> tuple[str, Path] | None:
    """Return ``(period, path)`` of newest SCALE workbook strictly
    older than ``before_period``, or None when there is none.

    Useful when the user is running a quarter that's more than one
    step ahead of the most recent report on disk.
    """
    cu_dir = _reports_root(workspace_root) / short_name
    if not cu_dir.exists():
        return None
    before_key = _period_sort_key(before_period)
    candidates: list[tuple[tuple[int, int], str, Path]] = []
    for period_dir in cu_dir.iterdir():
        if not period_dir.is_dir():
            continue
        k = _period_sort_key(period_dir.name)
        if k >= before_key:
            continue
        pick = _pick_report_in_period(period_dir)
        if pick is not None:
            candidates.append((k, period_dir.name, pick))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return (candidates[0][1], candidates[0][2])


def read_prior_acl_values(prior_path: str | Path) -> dict[str, Any]:
    """Read the four Prior-ACL values from a previously-generated SCALE
    workbook.

    Returns ``{"ok": bool, "values": {"AY2": ..., "AY3": ..., ...},
    "error": str}``. Tries the file's cached values first (set by Excel
    when the user last opened/saved); falls back to reading raw cell
    values where the source itself is a literal number.
    """
    result: dict[str, Any] = {"ok": False, "values": {}, "error": ""}
    p = Path(prior_path)
    if not p.exists():
        result["error"] = f"Prior workbook not found: {p}"
        return result

    # Pass 1 — cached evaluated values (preferred).
    cached: dict[str, Any] = {}
    try:
        wb_cached = openpyxl.load_workbook(p, data_only=True, read_only=True)
        for sheet, src_coord, dest_coord in _PRIOR_ACL_SOURCES:
            if sheet not in wb_cached.sheetnames:
                continue
            v = wb_cached[sheet][src_coord].value
            cached[dest_coord] = v
        wb_cached.close()
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"Could not read cached values from {p.name}: {exc}"

    # Pass 2 — raw values, used only for source cells whose cached
    # value is None (file was never opened/saved by Excel).
    raw: dict[str, Any] = {}
    needs_raw = any(cached.get(d) is None for _, _, d in _PRIOR_ACL_SOURCES)
    if needs_raw:
        try:
            wb_raw = openpyxl.load_workbook(p, data_only=False, read_only=True)
            for sheet, src_coord, dest_coord in _PRIOR_ACL_SOURCES:
                if sheet not in wb_raw.sheetnames:
                    continue
                v = wb_raw[sheet][src_coord].value
                if isinstance(v, (int, float)):
                    raw[dest_coord] = v
            wb_raw.close()
        except Exception as exc:  # noqa: BLE001
            if not result["error"]:
                result["error"] = (
                    f"Could not read raw values from {p.name}: {exc}"
                )

    merged: dict[str, Any] = {}
    for _, _, dest in _PRIOR_ACL_SOURCES:
        v = cached.get(dest)
        if v is None:
            v = raw.get(dest)
        if isinstance(v, (int, float)):
            merged[dest] = float(v)
    result["values"] = merged
    result["ok"] = bool(merged)
    if not merged and not result["error"]:
        result["error"] = (
            "Prior workbook had no numeric values for the four Prior ACL "
            "cells. Open it in Excel once and save so values cache, or "
            "rerun the prior quarter."
        )
    return result


def write_prior_acl_values(
    new_workbook_path: str | Path,
    values: dict[str, Any],
) -> dict[str, Any]:
    """Hard-code Prior ACL values into the Historical Data prior column.

    The fresh template ships those cells as formulas pointing into
    Scale Calculation!T29..T33 (the manual "copy previous column"
    workflow). We bypass that by writing literal numbers that the
    Executive Summary-Vizo "Prior ACL" OFFSET formulas read.

    The target column is detected dynamically: each carried quarter adds
    a new current column to Historical Data, so the column the
    ``=OFFSET('Historical Data'!<anchor>N,0,-1)`` prior formulas resolve
    to drifts right over time (AY -> AZ -> BA ...). We read that anchor
    from Executive Summary-Vizo and write to ``anchor - 1`` so the values
    land in the exact cells the report reads — overwriting any leftover
    ``='Scale Calculation'!U..`` formula that would otherwise leak the
    CURRENT column's numbers into the "prior" block. Falls back to the
    incoming coords (legacy AY2..AY5) when the anchor can't be detected.
    """
    result: dict[str, Any] = {"ok": False, "written": 0, "error": ""}
    if not values:
        return result
    p = Path(new_workbook_path)
    if not p.exists():
        result["error"] = f"Workbook not found: {p}"
        return result
    try:
        wb = openpyxl.load_workbook(p)
        if "Historical Data" not in wb.sheetnames:
            result["error"] = "Historical Data tab not found in target workbook."
            return result
        ws = wb["Historical Data"]
        target_col = _detect_prior_acl_target_column(wb)
        result["target_col"] = target_col
        for coord, val in values.items():
            if target_col:
                row = re.sub(r"^[A-Za-z]+", "", str(coord)) or ""
                if row:
                    coord = f"{target_col}{row}"
            ws[coord].value = val
            result["written"] += 1
        wb.save(p)
        result["ok"] = True
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"Failed writing Prior ACL: {exc}"
    return result


def _detect_prior_acl_target_column(wb) -> str:
    """Return the Historical Data column the Vizo 'Prior ACL' block reads.

    The prior rows on Executive Summary-Vizo are
    ``=OFFSET('Historical Data'!<anchor>N,0,-1)``; the value they display
    lives one column LEFT of ``<anchor>``. Returns that column letter
    (e.g. ``AZ`` when the anchor is ``BA``), or ``""`` if not found.
    """
    from openpyxl.utils import column_index_from_string, get_column_letter
    _rx = re.compile(
        r"OFFSET\('?Historical Data'?!\$?([A-Z]{1,3})\$?\d+\s*,\s*0\s*,\s*-\s*1\s*\)",
        re.IGNORECASE,
    )
    for sheet in ("Executive Summary-Vizo", "Executive Summary-TCT",
                  "Executive Summary"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in range(10, 26):
            v = ws.cell(row=row, column=3).value
            if isinstance(v, str) and v.startswith("="):
                m = _rx.search(v)
                if m:
                    anchor = column_index_from_string(m.group(1).upper())
                    if anchor > 1:
                        return get_column_letter(anchor - 1)
    return ""



def unhide_all_sheets(workbook_path: str | Path) -> dict[str, Any]:
    """Reset every sheet's ``sheet_state`` to "visible".

    Used when the prior workbook is a themed variant (``*_Vizo.xlsx``
    or ``*_TCT.xlsx``) where ``apply_report_variant`` hid one half of
    the tabs. Carrying that hidden state into the new run would
    suppress the wrong variant.
    """
    result: dict[str, Any] = {"ok": False, "unhidden": 0, "error": ""}
    p = Path(workbook_path)
    if not p.exists():
        result["error"] = f"Workbook not found: {p}"
        return result
    try:
        wb = openpyxl.load_workbook(p)
        for name in wb.sheetnames:
            if wb[name].sheet_state != "visible":
                wb[name].sheet_state = "visible"
                result["unhidden"] += 1
        wb.save(p)
        result["ok"] = True
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"Failed unhiding sheets: {exc}"
    return result


def load_state_for_cu(
    workspace_root: str | Path,
    short_name: str,
) -> dict[str, Any] | None:
    """Rehydrate the wizard state JSON for the given CU, or None.

    Prefers the SCALE draft (``<slug>__scale.json``); falls back to a
    Migration draft so reruns work for CUs that were originally
    configured only via the TCT/Vizo wizard. Callers can hand the
    returned dict straight to ``runner.run_single_quarter`` and
    friends; ``scale_runs.run`` separately seeds any missing
    ``state['scale']`` fields from admin defaults.
    """
    data = wizard_drafts.load_draft(workspace_root, short_name, model="scale")
    if data is not None:
        return data
    return wizard_drafts.load_draft(
        workspace_root, short_name, model="migration",
    )
