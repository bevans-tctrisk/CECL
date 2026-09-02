"""SCALE Impaired Loans loader.

Reads the standardized ``CECL-SCALE Impaired Loans`` workbook a credit
union uploads and copies the data-entry rows (cols A:J) into the SCALE
output workbook's `` Impaired Loans ASC 310-10`` sheet, preserving the
template's calc columns K:Q (formulas) and summary K6:Q26.

Pools and impairment types are canonical across all credit unions for
the SCALE methodology, so we do NOT remap pool names or look up credit
grades — we just stream the rows through.

Source layout (per template ``CECL-SCALE Impaired Loans.xlsx``):
    Sheet:       " Impaired Loans ASC 310-10" (note leading space)
    Header row:  27
    Data rows:   28..412 (leading blank/placeholder rows are skipped;
                 read stops at the first blank Impairment Type after data)
    Columns:     A=Impairment Type, B=Member #, C=Loan Suffix,
                 D=Loan Pool, E=Current Balance, F=Days Delinquent,
                 G=Balance at Other Lender, H=Collateral Value,
                 I=Allowance Provided, J=Notes
"""
from __future__ import annotations

import warnings
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font


SHEET_CANDIDATES = (
    " Impaired Loans ASC 310-10",
    "Impaired Loans ASC 310-10",
)

_VIZO_SHEET_CANDIDATES = (
    " Impaired Loans-Vizo",
    "Impaired Loans-Vizo",
)

# 1-based column letters mirrored as 0-based indexes for clarity.
_COLS = ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J")
_FIELDS = (
    "impairment_type", "member", "suffix", "loan_pool",
    "current_balance", "days_delinquent", "other_lender_balance",
    "collateral_value", "allowance_provided", "notes",
)
_DATA_START_ROW = 28  # OUTPUT first data-entry row (below header row 27)
_DATA_END_ROW = 414  # last calc-formula row (clear/hide/repair scope)
# Summary SUMIFs aggregate the band $A$28:$A$412, so data writes are
# capped here — a loan past row 412 would be excluded from the totals.
_SUMMARY_LAST_ROW = 412
# Source uploads vary in layout (the header "Impairment Type" sits on
# row 27 or 29 depending on where the file came from), so the parser
# locates the header row and reads the row after it. This is only the
# fallback when the header label can't be found.
_FALLBACK_DATA_START_ROW = 30
_HEADER_LABEL = "impairment type"
# Vizo Impaired tab N/O (Total Loans, LTV) formula band. Row 28 is the
# first calc row (27 is the header).
_VIZO_FORMULA_START_ROW = 28

# Optional cleanup bands on the TCT impaired tab that are data-entry
# placeholders in the template but are usually left empty in generated
# reports.
_OPTIONAL_HIDE_BANDS = (
    (10, 23),
)

# "No impaired loans" note. When the CU didn't send an impaired-loans
# file, or sent one with no impaired loans, every data row (28..414) is
# hidden and the tab prints as an empty grid. Row 25 is the blank gap
# between the summary total (24) and the detail-table banner (26) on both
# impaired tabs, and sits inside both tabs' print areas.
# Both tabs print their data-entry block and their calculation block on
# separate pages (the two blocks are far too wide to share one), so the
# note has to be repeated in each block or page 2 still reads as unfinished.
# Data entry is A:J on both; calculation is K:Q on the TCT tab and L:S on
# the Vizo one (its K is a hairline spacer column).
_NOTE_ROW = 25
_NOTE_SPANS_TCT = ("A25:J25", "K25:Q25")
_NOTE_SPANS_VIZO = ("A25:J25", "L25:S25")
_NOTE_TEXT = (
    "The credit union reported no impaired loans for this quarter."
)
_NOTE_ROW_HEIGHT = 21.0


def _find_sheet(wb) -> str | None:
    names = {s.strip().lower(): s for s in wb.sheetnames}
    for cand in SHEET_CANDIDATES:
        hit = names.get(cand.strip().lower())
        if hit:
            return hit
    # Fallback: any sheet containing "impaired" + "310" in its name.
    for s in wb.sheetnames:
        low = s.lower()
        if "impaired" in low and "310" in low:
            return s
    return None


def _find_vizo_sheet(wb) -> str | None:
    names = {s.strip().lower(): s for s in wb.sheetnames}
    for cand in _VIZO_SHEET_CANDIDATES:
        hit = names.get(cand.strip().lower())
        if hit:
            return hit
    return None


def _repair_tct_prov_formula(wb, tct_sheet: str) -> int:
    """Restore the TCT Impaired tab's provision-% column (O) formula.

    Older templates / carried-forward prior-quarter reports hold a variant
    that skips ``$A$8`` (Foreclosed Real Estate) and instead references the
    empty ``$A$10``, so Foreclosed Real Estate loans resolve to a 0%
    provision instead of 100%. Rewrite every data row to the correct
    5-category lookup. Idempotent / harmless on already-correct workbooks.
    Returns the number of cells rewritten.
    """
    ws = wb[tct_sheet]
    fixed = 0
    for r in range(_DATA_START_ROW, _DATA_END_ROW + 1):
        ws[f"O{r}"] = (
            f'=IF(I{r}<>"",100%,IF(A{r}=$A$5,$B$5,'
            f'IF(A{r}=$A$6,$B$6,IF(A{r}=$A$7,$B$7,'
            f'IF(A{r}=$A$8,$B$8,IF(A{r}=$A$9,$B$9,0))))))'
        )
        fixed += 1
    return fixed


def _repair_impaired_summary_totals(wb, tct_sheet: str) -> int:
    """Total the impaired summary (row 24) over the actual per-loan
    columns instead of the per-impairment-type SUMIF rows.

    The template's ``N24/P24/Q24`` = ``SUM(N5:N10)`` etc., where rows
    5-10 are ``SUMIF`` by impairment type against the fixed category
    table (A5:A9). A CU whose loans use a type NOT in that table (e.g.
    "Other Impaired Loans") sums to 0, so the individual-basis provision
    (Scale Calculation!U27 = P24) came out blank even though the loans
    imported. Sum the data band directly so the totals capture every
    loan regardless of its impairment-type label.
    """
    ws = wb[tct_sheet]
    band = f"{_DATA_START_ROW}:{_SUMMARY_LAST_ROW}"  # 28:412
    fixed = 0
    for col in ("N", "P", "Q"):
        lo, hi = band.split(":")
        ws[f"{col}24"] = f"=SUM({col}{lo}:{col}{hi})"
        fixed += 1
    return fixed


def _repair_vizo_impaired_formulas(wb, tct_sheet: str) -> int:
    """Restore the Vizo Impaired Loans tab's Total Loans (N) and LTV (O)
    columns to reference the computed TCT columns.

    Some workbooks (carried forward from a prior quarter) hold a broken
    variant of these formulas -- ``=IFERROR(E+G,"")`` for N and
    ``=IFERROR(IF(H=0,"No Value",N/H),"")`` for O -- that recomputes off
    the Vizo tab's own cells. Those cells are string-guarded
    (``=IF(TCT!x="","",TCT!x)``), so a blank other-lender/collateral cell
    makes the arithmetic ``number + ""`` -> #VALUE!, which ``IFERROR``
    then blanks. The template's correct form reads the already-computed
    TCT ``L`` (Total Loans) and ``M`` (LTV) values, so we rewrite to that.
    Returns the number of cells rewritten.
    """
    vizo = _find_vizo_sheet(wb)
    if not vizo:
        return 0
    ws = wb[vizo]
    q = f"'{tct_sheet}'"
    fixed = 0
    for r in range(_VIZO_FORMULA_START_ROW, _DATA_END_ROW + 1):
        ws[f"N{r}"] = f'=IF({q}!A{r}="","",{q}!L{r})'
        ws[f"O{r}"] = f'=IF({q}!A{r}="","",{q}!M{r})'
        fixed += 2
    return fixed



def _impaired_sheet_candidates(wb) -> list[str]:
    """All worksheets that look like an impaired-loans data sheet.

    CUs frequently keep the canonical blank `` Impaired Loans ASC
    310-10`` tab AND a period-prefixed working copy (e.g. `` 26-6
    Impaired Loans ASC 310-10``) that holds the quarter's actual rows.
    Some source files name the tab simply `` Impaired Loans`` (no
    ``ASC 310-10``), so match any ``impaired`` sheet — preferring the
    ``310`` ones — while excluding the Vizo mirror tab.
    """
    strict: list[str] = []
    loose: list[str] = []
    for s in wb.sheetnames:
        low = s.lower()
        if "impaired" not in low or "vizo" in low:
            continue
        (strict if "310" in low else loose).append(s)
    return strict + loose


def _find_data_start_row(ws, max_scan: int = 45) -> int:
    """Row where source data begins = the row after the "Impairment
    Type" header. Source files differ (header on row 27 or 29), so
    detect it instead of assuming a fixed offset."""
    for r in range(1, max_scan + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip().lower() == _HEADER_LABEL:
            return r + 1
    return _FALLBACK_DATA_START_ROW


def _row_is_real(balance: Any, allowance: Any) -> bool:
    """A real impaired loan carries a balance (or an explicit
    allowance). Rows with only an impairment type / filled-down member
    label are template scaffold or CU padding — treat as empty."""
    return bool(_coerce_num(balance) or _coerce_num(allowance))


def _count_data_rows(ws) -> int:
    """Count real data rows (balance/allowance) below the header.

    Leading blank/placeholder rows are skipped; reading stops at the
    first blank Impairment Type after the data begins.
    """
    start = _find_data_start_row(ws)
    count = 0
    started = False
    for r in range(start, _DATA_END_ROW + 1):
        a = ws.cell(row=r, column=1).value
        real = _row_is_real(ws.cell(row=r, column=5).value,
                            ws.cell(row=r, column=9).value)
        if not started:
            if a in (None, "") or not real:
                continue
            started = True
        else:
            if a in (None, ""):
                break
            if not real:
                continue
        count += 1
    return count


def _select_data_sheet(wb) -> str | None:
    """Pick the impaired sheet that actually holds the quarter's rows.

    Prefer the matching sheet with the most data-entry rows so a
    period-prefixed working tab wins over a leftover blank template
    tab. Falls back to ``_find_sheet`` when no candidate has data.
    """
    best: str | None = None
    best_rows = 0
    for name in _impaired_sheet_candidates(wb):
        n = _count_data_rows(wb[name])
        if n > best_rows:
            best_rows = n
            best = name
    if best is not None:
        return best
    return _find_sheet(wb)


def _coerce_num(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", "").replace("$", ""))
    except (TypeError, ValueError):
        return 0.0


def _str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_has_entry_data(ws, row_num: int) -> bool:
    """True when any entry column A:J has a meaningful value.

    Treat blank strings and zero-like placeholders as empty so the
    template's unused scaffold rows can be hidden.
    """
    for c in range(1, 11):
        v = ws.cell(row=row_num, column=c).value
        if v in (None, ""):
            continue
        if isinstance(v, str) and not v.strip():
            continue
        if isinstance(v, (int, float)) and float(v) == 0.0:
            continue
        if isinstance(v, str) and v.strip() in {"$", "$ -", "-", "0", "0%", "0.0%"}:
            continue
        return True
    return False


def _hide_unused_rows(ws, applied_rows: int) -> int:
    """Hide unused data-entry rows on the impaired worksheet.

    - Keep populated rows visible.
    - Hide placeholder bands when empty (10:23, 28:29).
    - Hide the trailing data-entry rows after the last populated row.
    Returns the count of rows currently hidden by this pass.
    """
    hidden = 0

    for start, end in _OPTIONAL_HIDE_BANDS:
        for r in range(start, end + 1):
            # These are template scaffold rows (summary filler / spacer),
            # not user data-entry rows. On Vizo they may contain mirror
            # formulas, so hide unconditionally.
            ws.row_dimensions[r].hidden = True
            hidden += 1

    # Show the active filled range (if any), hide everything below it.
    used_end = _DATA_START_ROW + max(applied_rows, 0) - 1
    for r in range(_DATA_START_ROW, _DATA_END_ROW + 1):
        hide = r > used_end
        ws.row_dimensions[r].hidden = hide
        hidden += 1 if hide else 0

    return hidden


def _apply_empty_note(ws, applied_rows: int,
                      spans: tuple[str, ...]) -> bool:
    """Write (or clear) the "no impaired loans" note on an impaired tab.

    One note per printed block (see ``_NOTE_SPANS_*``). Idempotent in both
    directions: a carried-forward workbook holding last quarter's note has
    it removed as soon as real rows land, and re-running an empty quarter
    doesn't stack duplicate merges. Returns True when the note is present
    after this pass.

    The note deliberately sits above the detail table rather than in a
    data row — rows 28+ hold the K:Q / L:S calc formulas, which would
    render as ``$0`` / ``No Value`` junk alongside the text.
    """
    existing = {str(m) for m in ws.merged_cells.ranges}
    # Match whatever face the tab's detail rows use (Arial on the TCT tab,
    # Calibri on the Vizo one) so the note doesn't read as pasted in.
    base_name = ws.cell(row=_DATA_START_ROW, column=1).font.name or "Calibri"

    for span in spans:
        col = ws[span.split(":")[0]].column
        cell = ws.cell(row=_NOTE_ROW, column=col)
        if applied_rows > 0:
            if span in existing:
                ws.unmerge_cells(span)
            if isinstance(cell.value, str) and cell.value.strip() == _NOTE_TEXT:
                cell.value = None
            continue
        if span not in existing:
            ws.merge_cells(span)
        cell.value = _NOTE_TEXT
        cell.font = Font(name=base_name, size=11, italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if applied_rows > 0:
        return False
    ws.row_dimensions[_NOTE_ROW].hidden = False
    ws.row_dimensions[_NOTE_ROW].height = _NOTE_ROW_HEIGHT
    return True


def parse_file(path: str | Path) -> dict[str, Any]:
    """Parse a CU's SCALE Impaired Loans workbook.

    Returns ``{ok, error, sheet_used, cu_name, period, rows, row_count,
    total_balance}``. ``rows`` is a list of dicts using ``_FIELDS``
    above plus a derived ``member_suffix`` key.
    """
    out: dict[str, Any] = {
        "ok": False,
        "error": "",
        "sheet_used": "",
        "cu_name": "",
        "period": "",
        "rows": [],
        "row_count": 0,
        "total_balance": 0.0,
    }
    p = Path(path)
    if not p.exists():
        out["error"] = f"File not found: {p}"
        return out
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(p, data_only=True, read_only=False)
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Failed to open workbook: {exc}"
        return out

    sheet = _select_data_sheet(wb)
    if not sheet:
        out["error"] = (
            f"Could not find ' Impaired Loans ASC 310-10' sheet. "
            f"Sheets present: {wb.sheetnames}"
        )
        return out
    ws = wb[sheet]
    out["sheet_used"] = sheet
    # Header metadata (best-effort)
    try:
        out["cu_name"] = _str(ws["A2"].value)
    except Exception:  # noqa: BLE001
        pass
    try:
        out["period"] = _str(ws["B5"].value)
    except Exception:  # noqa: BLE001
        pass

    rows: list[dict[str, Any]] = []
    total = 0.0
    started = False
    for r in range(_find_data_start_row(ws), _DATA_END_ROW + 1):
        a = ws.cell(row=r, column=1).value
        record = {}
        for idx, field in enumerate(_FIELDS):
            v = ws.cell(row=r, column=idx + 1).value
            if field in (
                "current_balance", "days_delinquent",
                "other_lender_balance", "collateral_value",
                "allowance_provided",
            ):
                record[field] = _coerce_num(v)
            else:
                record[field] = _str(v)
        record["member_suffix"] = (
            f"{record['member']}-{record['suffix']}"
            if record["member"] or record["suffix"]
            else ""
        )
        # A real impaired loan carries a balance (or explicit
        # allowance). Rows with only an impairment type / filled-down
        # member label are scaffold or CU padding — skip them so the
        # output starts at the first actual loan.
        placeholder = not _row_is_real(
            record["current_balance"], record["allowance_provided"]
        )
        if not started:
            if a in (None, "") or placeholder:
                continue
            started = True
        else:
            if a in (None, ""):
                # First fully blank Impairment Type after data = end.
                break
            if placeholder:
                continue
        rows.append(record)
        total += record["current_balance"]

    out["ok"] = True
    out["rows"] = rows
    out["row_count"] = len(rows)
    out["total_balance"] = round(total, 2)
    return out


def apply_impaired_rows(
    workbook_path: str | Path, rows: list[dict[str, Any]],
) -> dict[str, Any]:
    """Write ``rows`` into the output workbook's impaired sheet.

    Clears A28:J<end> first so a re-run doesn't leave stale rows.
    Calc columns K:Q are left untouched — the template formulas pick up
    the new data automatically.
    """
    result = {
        "ok": False,
        "applied": 0,
        "cleared": 0,
        "sheet_used": "",
        "error": "",
    }
    if not rows:
        # Still clear any prior writes so re-runs with an emptied list
        # remove old data.
        rows = []
    # Drop placeholder rows (only an impairment type / filled-down
    # member label, no balance or allowance) so the output starts at the
    # first real loan. These come from CU source padding and can also be
    # carried in stored wizard state parsed before the parser skipped
    # them.
    rows = [
        row for row in rows
        if _row_is_real(row.get("current_balance"),
                        row.get("allowance_provided"))
    ]
    p = Path(workbook_path)
    if not p.exists():
        result["error"] = f"Output workbook not found: {p}"
        return result
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(p)
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"Failed to open output workbook: {exc}"
        return result
    sheet = _find_sheet(wb)
    if not sheet:
        result["error"] = (
            "Output template is missing ' Impaired Loans ASC 310-10' sheet."
        )
        return result
    ws = wb[sheet]
    result["sheet_used"] = sheet
    # Clear A:J for the full data range.
    cleared = 0
    for r in range(_DATA_START_ROW, _DATA_END_ROW + 1):
        for c in range(1, 11):  # cols A..J
            cell = ws.cell(row=r, column=c)
            if cell.value not in (None, ""):
                cell.value = None
                cleared += 1
    # Write rows (cap at the summary SUMIF band so no loan lands past
    # row 412 where it would be excluded from the totals).
    applied = 0
    max_rows = _SUMMARY_LAST_ROW - _DATA_START_ROW + 1
    for i, row in enumerate(rows[:max_rows]):
        r = _DATA_START_ROW + i
        ws.cell(row=r, column=1).value = row.get("impairment_type") or None
        ws.cell(row=r, column=2).value = row.get("member") or None
        ws.cell(row=r, column=3).value = row.get("suffix") or None
        ws.cell(row=r, column=4).value = row.get("loan_pool") or None
        ws.cell(row=r, column=5).value = row.get("current_balance") or None
        ws.cell(row=r, column=6).value = row.get("days_delinquent") or None
        ws.cell(row=r, column=7).value = row.get("other_lender_balance") or None
        ws.cell(row=r, column=8).value = row.get("collateral_value") or None
        ws.cell(row=r, column=9).value = row.get("allowance_provided") or None
        ws.cell(row=r, column=10).value = row.get("notes") or None
        applied += 1
    # Repair the Vizo Impaired tab's Total Loans/LTV columns (N/O) so
    # carried-forward workbooks with the broken IFERROR variant don't
    # blank them out. Harmless on already-correct workbooks.
    vizo_fixed = _repair_vizo_impaired_formulas(wb, sheet)
    # Repair the TCT Impaired tab's provision-% column (O) so carried-
    # forward / older-template workbooks that skip $A$8 (Foreclosed Real
    # Estate) are corrected. Harmless on already-correct workbooks.
    tct_prov_fixed = _repair_tct_prov_formula(wb, sheet)
    # Total the summary (row 24) over the actual per-loan columns so the
    # individual-basis provision (Scale Calculation!U27 = P24) captures
    # loans whose impairment type isn't in the template category table.
    tct_totals_fixed = _repair_impaired_summary_totals(wb, sheet)
    hidden_rows = _hide_unused_rows(ws, applied)
    # No rows means the CU either sent no impaired-loans file or sent one
    # with no impaired loans; note it on the tab so the page doesn't print
    # as a bare, seemingly-unfinished grid.
    empty_note = _apply_empty_note(ws, applied, _NOTE_SPANS_TCT)
    hidden_rows_vizo = 0
    vizo_sheet = _find_vizo_sheet(wb)
    if vizo_sheet:
        hidden_rows_vizo = _hide_unused_rows(wb[vizo_sheet], applied)
        _apply_empty_note(wb[vizo_sheet], applied, _NOTE_SPANS_VIZO)
    try:
        wb.save(p)
    except PermissionError as exc:
        result["error"] = (
            f"Permission denied saving {p.name}. The workbook is likely "
            f"open in Excel or locked by a sync client (Egnyte / OneDrive). "
            f"Close it everywhere and re-run. (details: {exc})"
        )
        return result
    except OSError as exc:
        result["error"] = f"Failed to save workbook {p.name}: {exc}"
        return result
    result["ok"] = True
    result["applied"] = applied
    result["cleared"] = cleared
    result["vizo_formulas_fixed"] = vizo_fixed
    result["tct_prov_formulas_fixed"] = tct_prov_fixed
    result["tct_summary_totals_fixed"] = tct_totals_fixed
    result["rows_hidden"] = hidden_rows
    result["rows_hidden_vizo"] = hidden_rows_vizo
    result["empty_note"] = empty_note
    if len(rows) > max_rows:
        result["error"] = (
            f"Truncated: source has {len(rows)} rows but template only "
            f"supports {max_rows} (rows {_DATA_START_ROW}..{_SUMMARY_LAST_ROW})."
        )
    return result
