"""Normalize any SCALE Vizo output to the approved redesign layout.

The three run modes (fresh single-quarter, multi-quarter, and the normal
"carry history" process) all funnel through ``apply_report_variant``. Only
fresh template runs inherit the redesigned template structure; carry-history
runs are seeded from the *prior quarter's report* and would otherwise keep
the old tab order / names forever. This module re-applies the structural
redesign to any Vizo workbook so every generated report is consistent,
regardless of how it was seeded. It is idempotent: a report already in the
new layout is left unchanged.

The Env Factor by Pool-Vizo (with the merged Environmental Factor Ranges
block) and the redesigned Cover-Vizo are copied from the canonical template
so the merge/rewire and cover art stay in one source of truth.
"""
from __future__ import annotations

from copy import copy, deepcopy
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.properties import PageSetupProperties

RENAMES = {
    "Introduction-Vizo": "Appendix-Vizo",
    "Explanation of ACL Calc-Vizo": "Appendix 2-Vizo",
}

# Sheets copied wholesale from the template (formula/art driven, period
# agnostic) so the environ-merge, cover redesign, Executive Summary
# formatting, and the Historical Summary (New Report Calc AE:AI mirror)
# live in one place.
TEMPLATE_SOURCED = (
    "Env Factor by Pool-Vizo", "Cover-Vizo", "Executive Summary-Vizo",
    "Historical Summary",
)

# The cover redesign shifted the CU-name/date cells up (A15->A10, A16->A11).
# Downstream 'Prepared For'/'Quarter Ended' headers must follow.
COVER_REF_FIXES = {
    "Cover-Vizo'!A15": "Cover-Vizo'!A10",
    "Cover-Vizo'!A16": "Cover-Vizo'!A11",
    "Cover-Vizo'!$A$15": "Cover-Vizo'!$A$10",
    "Cover-Vizo'!$A$16": "Cover-Vizo'!$A$11",
}

REDUNDANT_HIDDEN = "Envir Factor Ranges-Vizo"

CHANGE_ANALYSIS = "Change Analysis"

# Blue sheet-tab color flags the tabs that get converted to PDF
# (Cover-Vizo through Appendix 2-Vizo).
BLUE_TAB = "FF0070C0"
BLUE_TABS = (
    "Cover-Vizo", "Executive Summary-Vizo", "Scale Calculation-Vizo",
    " Impaired Loans-Vizo", "Historical Summary", "Env Factor by Pool-Vizo",
    "Envir Factor Ranges-Vizo", "Change Analysis", "New Report Calc-Vizo",
    "Appendix-Vizo", "Appendix 2-Vizo",
)

PAGE_NUM_TABS = (
    "Executive Summary-Vizo", "Scale Calculation-Vizo", " Impaired Loans-Vizo",
    "Historical Summary", "Change Analysis", "Env Factor by Pool-Vizo",
    "New Report Calc-Vizo", "Appendix-Vizo", "Appendix 2-Vizo",
)

# Force these tabs to print on exactly one page so different PDF converters
# can't introduce their own page breaks. (Change Analysis sets its own
# landscape fit-to-page in the change-analysis builder.)
FIT_ONE_PAGE_TABS = (
    "Cover-Vizo", "Scale Calculation-Vizo", "Env Factor by Pool-Vizo",
    "Historical Summary",
)

# Narrow tabs (~5 columns) that must print portrait, not landscape.
PORTRAIT_TABS = (
    "Historical Summary",
)

# Full redesign order; only sheets actually present are reordered, unknown
# sheets keep their relative order at the end.
TARGET_ORDER = (
    "Cover", "Introduction", "Scale Calculation", "Env Factor by Pool",
    "Environmental Factor Ranges", " Impaired Loans ASC 310-10", "Calc tab",
    "Cover-Vizo", "Executive Summary-Vizo", "Scale Calculation-Vizo",
    " Impaired Loans-Vizo", "Historical Summary", "Env Factor by Pool-Vizo",
    "Envir Factor Ranges-Vizo", "Change Analysis", "New Report Calc-Vizo",
    "Appendix-Vizo", "Appendix 2-Vizo",
    "Management Adjustment", "Historical Data", "Industry Data",
    "DNU Background", "DNU Tab 4 - Adj. to Loss Rate",
    "DNU Tab 2 - Individually Assess", "DNU Tab 3 - Qualitative Adj",
    "DNU Scale Org",
)


def _deep_copy_sheet(src, dst) -> None:
    for row in src.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            nc = dst.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                nc.font = copy(cell.font)
                nc.border = copy(cell.border)
                nc.fill = copy(cell.fill)
                nc.number_format = cell.number_format
                nc.protection = copy(cell.protection)
                nc.alignment = copy(cell.alignment)
    for mc in src.merged_cells.ranges:
        dst.merge_cells(str(mc))
    for col, dim in src.column_dimensions.items():
        d = dst.column_dimensions[col]
        d.width = dim.width
        d.hidden = dim.hidden
    for ridx, dim in src.row_dimensions.items():
        d = dst.row_dimensions[ridx]
        d.height = dim.height
        d.hidden = dim.hidden
    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    dst.freeze_panes = src.freeze_panes
    ps, sps = src.page_setup, dst.page_setup
    sps.orientation = ps.orientation
    sps.paperSize = ps.paperSize
    sps.scale = ps.scale
    sps.fitToWidth = ps.fitToWidth
    sps.fitToHeight = ps.fitToHeight
    dst.sheet_properties.pageSetUpPr = copy(src.sheet_properties.pageSetUpPr)
    for a in ("left", "right", "top", "bottom", "header", "footer"):
        setattr(dst.page_margins, a, getattr(src.page_margins, a))
    dst.print_options.horizontalCentered = src.print_options.horizontalCentered
    dst.print_options.verticalCentered = src.print_options.verticalCentered
    for side in ("left", "center", "right"):
        getattr(dst.oddFooter, side).text = getattr(src.oddFooter, side).text
        getattr(dst.oddHeader, side).text = getattr(src.oddHeader, side).text
    if src.print_area:
        dst.print_area = src.print_area
    for im in getattr(src, "_images", []):
        try:
            data = bytes(im._data())
        except Exception:  # noqa: BLE001
            continue
        nim = XLImage(BytesIO(data))
        anchor = getattr(im, "anchor", None)
        if anchor is not None:
            try:
                nim.anchor = deepcopy(anchor)
            except Exception:  # noqa: BLE001
                pass
        nim.width = im.width
        nim.height = im.height
        dst.add_image(nim)


def _replace_from_template(wb, tmpl_wb, sheet_name: str) -> None:
    if sheet_name not in tmpl_wb.sheetnames:
        return
    if sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(sheet_name)
        del wb[sheet_name]
        new_ws = wb.create_sheet(sheet_name, idx)
    else:
        # Not in the carried report yet (e.g. new Historical Summary tab);
        # add it — final position is set by the reorder step.
        new_ws = wb.create_sheet(sheet_name)
    _deep_copy_sheet(tmpl_wb[sheet_name], new_ws)


def _fix_cover_refs(wb) -> None:
    """Retarget stale 'Cover-Vizo' cell refs after the cover redesign moved
    the CU-name/date cells up (idempotent)."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "Cover-Vizo'!" in v:
                    for old, new in COVER_REF_FIXES.items():
                        if old in v:
                            v = v.replace(old, new)
                    cell.value = v


def normalize_vizo_layout(output_path: str | Path,
                          template_path: str | Path | None = None) -> dict:
    """Reshape a SCALE Vizo workbook to the approved redesign layout.

    Idempotent. ``template_path`` supplies the merged Env Factor and
    redesigned Cover; when absent those two sheets are left as-is.
    """
    result = {"ok": False, "renamed": [], "error": ""}
    path = Path(output_path)
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"load failed: {exc}"
        return result

    try:
        # 1. Renames (skip if the new name already exists).
        for old, new in RENAMES.items():
            if old in wb.sheetnames and new not in wb.sheetnames:
                wb[old].title = new
                result["renamed"].append(f"{old}->{new}")

        # 2. Change Analysis placeholder.
        if CHANGE_ANALYSIS not in wb.sheetnames:
            wb.create_sheet(CHANGE_ANALYSIS)

        # 3. Template-sourced sheets (merged environ + redesigned cover).
        if template_path:
            try:
                tmpl_wb = openpyxl.load_workbook(str(template_path))
                for name in TEMPLATE_SOURCED:
                    _replace_from_template(wb, tmpl_wb, name)
            except Exception as exc:  # noqa: BLE001
                result["error"] = f"template copy skipped: {exc}"

        # 4. Hide the now-redundant ranges tab.
        if REDUNDANT_HIDDEN in wb.sheetnames:
            wb[REDUNDANT_HIDDEN].sheet_state = "hidden"

        # 5. Retarget stale cover cell references (post cover redesign).
        _fix_cover_refs(wb)

        # 6. Page-number footers.
        for tab in PAGE_NUM_TABS:
            if tab in wb.sheetnames:
                wb[tab].oddFooter.right.text = "Page &P of &N"

        # 6b. Force fit-to-one-page (consistent across PDF converters).
        for tab in FIT_ONE_PAGE_TABS:
            if tab in wb.sheetnames:
                ws = wb[tab]
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 1
                ws.page_setup.scale = None
                pr = ws.sheet_properties.pageSetUpPr
                if pr is None:
                    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
                else:
                    pr.fitToPage = True

        # 6b-2. Force portrait on the narrow tabs (carried workbooks can
        # still hold the old landscape setting).
        for tab in PORTRAIT_TABS:
            if tab in wb.sheetnames:
                wb[tab].page_setup.orientation = "portrait"

        # 6c. Blue tab colors flag the tabs that get converted to PDF.
        for tab in BLUE_TABS:
            if tab in wb.sheetnames:
                wb[tab].sheet_properties.tabColor = BLUE_TAB

        # 7. Reorder: known sheets in target order, unknown appended.
        known = [s for s in TARGET_ORDER if s in wb.sheetnames]
        extra = [s for s in wb.sheetnames if s not in TARGET_ORDER]
        wb._sheets = [wb[name] for name in known + extra]

        if "Cover-Vizo" in wb.sheetnames:
            wb.active = wb.sheetnames.index("Cover-Vizo")

        wb.save(path)
        result["ok"] = True
    except Exception as exc:  # noqa: BLE001
        result["error"] = f"normalize failed: {exc}"
    return result
