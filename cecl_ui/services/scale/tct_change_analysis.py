"""SCALE TCT period-over-period Change Analysis tab.

Adds a "Change Analysis" worksheet to generated SCALE TCT workbooks,
positioned immediately after "Calc tab". The tab compares the current
quarter to the most recent prior SCALE TCT report for the same CU.
"""
from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties

SCALE_CALC_SHEET = "Scale Calculation"
IMPAIRED_SHEET = " Impaired Loans ASC 310-10"
SHEET_NAME = "Change Analysis"

POOL_ROW_START = 9
POOL_ROW_END = 21

_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
F_TITLE = Font(name="Arial", size=14, bold=True)
F_SUB = Font(name="Arial", size=10, italic=True, color="595959")
F_HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_CELL = Font(name="Arial", size=10)
F_BOLD = Font(name="Arial", size=10, bold=True)
F_NOTE = Font(name="Arial", size=10)
FILL_HDR = PatternFill("solid", fgColor="305496")
FILL_TOT = PatternFill("solid", fgColor="D9E1F2")
FILL_FLAG = PatternFill("solid", fgColor="FCE4D6")
# Vizo variant (Change Analysis): match the accent-1 theme blues used on
# the other Vizo tabs (e.g. ' Impaired Loans-Vizo'), resolved via the Vizo
# theme that is swapped into the workbook after this tab is written.
FILL_HDR_VIZO = PatternFill("solid", fgColor=Color(theme=4))
FILL_TOT_VIZO = PatternFill("solid", fgColor=Color(theme=4, tint=0.5999))
ACCT = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
PCT = "0.0%"
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center")


def _num(v: Any) -> float:
    if v in (None, ""):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _s(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _is_significant(delta: float, prior: float, current: float) -> bool:
    d = abs(delta)
    if prior == 0 and current != 0:
        return abs(current) >= 1000
    if current == 0 and prior != 0:
        return abs(prior) >= 1000
    if d >= 25000:
        return True
    if prior and abs(delta / prior) >= 0.15 and d >= 5000:
        return True
    return False


def _extract_period_shortname(
    path: Path,
) -> tuple[str | None, str | None, str | None]:
    m = re.match(r"^(\d{4}-\d{2})_CECL_SCALE_(.+)_(TCT|Vizo)$", path.stem)
    if not m:
        return None, None, None
    return m.group(1), m.group(2), m.group(3)


def _find_prior_report(
    cur_path: Path,
    period: str,
    short_name: str,
    suffix: str,
) -> tuple[Path | None, str | None]:
    base = cur_path.parents[2] if len(cur_path.parents) >= 3 else None
    if base is None:
        return None, None
    cu_root = base / short_name
    if not cu_root.exists():
        return None, None

    best_path: Path | None = None
    best_period: str | None = None
    for p in cu_root.glob(f"*/*_CECL_SCALE_*_{suffix}.xlsx"):
        pp, sn, sf = _extract_period_shortname(p)
        if not pp or sn != short_name or sf != suffix:
            continue
        if pp >= period:
            continue
        if best_period is None or pp > best_period:
            best_period = pp
            best_path = p
    return best_path, best_period


def _set_table_header(ws, row: int, labels: list[str],
                      fill: PatternFill = FILL_HDR) -> None:
    for c, lbl in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=c, value=lbl)
        cell.font = F_HDR
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = CENTER


def _set_active_cover(wb) -> None:
    """Select the variant's cover tab (Vizo output hides the TCT 'Cover')."""
    for name in ("Cover-Vizo", "Cover"):
        if name in wb.sheetnames:
            wb.active = wb.sheetnames.index(name)
            return


def _find_sheet_name(wb, wanted: str) -> str | None:
    key = wanted.strip().lower()
    for s in wb.sheetnames:
        if s.strip().lower() == key:
            return s
    return None


def _parse_scale_snapshot(wb) -> dict[str, Any]:
    calc_name = _find_sheet_name(wb, SCALE_CALC_SHEET)
    if not calc_name:
        raise KeyError(f"Missing sheet: {SCALE_CALC_SHEET}")
    ws = wb[calc_name]

    pools: dict[str, dict[str, float]] = {}
    order: list[str] = []
    for r in range(POOL_ROW_START, POOL_ROW_END + 1):
        pool = _s(ws.cell(r, 3).value)
        if not pool or pool.lower() == "total":
            continue
        pools[pool] = {
            "balance": _num(ws.cell(r, 5).value),
            "spec_id": _num(ws.cell(r, 7).value),
            "total_allow": _num(ws.cell(r, 21).value),  # col U
        }
        order.append(pool)

    total_allow = _num(ws.cell(23, 21).value)   # U23
    pooled_loans = _num(ws.cell(23, 9).value)   # I23

    impaired_name = _find_sheet_name(wb, IMPAIRED_SHEET)
    impaired: dict[str, float] = {}
    total_spec_allow = 0.0
    if impaired_name:
        iw = wb[impaired_name]
        for r in range(5, 10):
            cat = _s(iw.cell(r, 12).value)      # L
            if cat:
                impaired[cat] = _num(iw.cell(r, 16).value)  # P
        total_spec_allow = _num(iw.cell(24, 16).value)      # P24

    return {
        "pools": pools,
        "order": order,
        "impaired": impaired,
        "totals": {
            "total_allow_needed": total_allow,
            "total_spec_allow": total_spec_allow,
            "pooled_total_allow": total_allow - total_spec_allow,
            "pooled_loans": pooled_loans,
        },
    }


def append_change_analysis(
    workbook_path: str | Path, sheet_name: str = SHEET_NAME,
    is_vizo: bool = False,
) -> dict[str, Any]:
    out: dict[str, Any] = {
        "ok": False,
        "sheet": sheet_name,
        "prior_path": "",
        "prior_period": "",
        "error": "",
    }

    p = Path(workbook_path)
    period, short_name, suffix = _extract_period_shortname(p)
    if not period or not short_name:
        out["error"] = "Filename does not match SCALE naming pattern."
        return out

    # Vizo variant uses the accent-1 theme blues + landscape; the TCT tab
    # keeps its original RGB palette / portrait.
    vizo = is_vizo
    fill_hdr = FILL_HDR_VIZO if vizo else FILL_HDR
    fill_tot = FILL_TOT_VIZO if vizo else FILL_TOT

    try:
        wb = load_workbook(p)
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Could not open current workbook: {exc}"
        return out

    calc_name = _find_sheet_name(wb, SCALE_CALC_SHEET)
    if not calc_name:
        out["error"] = f"Missing sheet: {SCALE_CALC_SHEET}"
        return out
    calc_ws = wb[calc_name]

    impaired_name = _find_sheet_name(wb, IMPAIRED_SHEET)
    impaired_ws = wb[impaired_name] if impaired_name else None

    # Best-effort cached snapshot for commentary/flagging only.
    # Current-column display uses live formulas written below.
    cur_cached: dict[str, Any] = {"pools": {}, "totals": {}, "impaired": {}}
    try:
        cur_cached_wb = load_workbook(p, data_only=True, read_only=True)
        cur_cached = _parse_scale_snapshot(cur_cached_wb)
        cur_cached_wb.close()
    except Exception:  # noqa: BLE001
        cur_cached = {"pools": {}, "totals": {}, "impaired": {}}

    prior_path, prior_period = _find_prior_report(p, period, short_name, suffix)

    existing_idx = None
    if sheet_name in wb.sheetnames:
        existing_idx = wb.sheetnames.index(sheet_name)
        del wb[sheet_name]
    if is_vizo and existing_idx is not None:
        # Fill the template's placeholder in its existing position.
        insert_at = existing_idx
    elif "Calc tab" in wb.sheetnames:
        insert_at = wb.sheetnames.index("Calc tab") + 1
    else:
        insert_at = len(wb.sheetnames)
    ws = wb.create_sheet(sheet_name, insert_at)

    ws.sheet_view.showGridLines = False
    # Vizo 'Change Analysis' is landscape, so column A is wide (and wraps
    # below) to show the full pool/label names that clip in the narrower TCT
    # portrait tab.
    col_widths = (("A", 65), ("B", 14), ("C", 14), ("D", 13), ("E", 10)) if vizo \
        else (("A", 42), ("B", 18), ("C", 18), ("D", 16), ("E", 12))
    for col, width in col_widths:
        ws.column_dimensions[col].width = width

    # Vizo 'Change Analysis' prints landscape, fit to one page, to match
    # the other Vizo tabs.
    if vizo:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.scale = None
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.sheet_properties.tabColor = "FF0070C0"  # PDF-convertible tab flag

    cu_label = short_name.replace("_", " ").title()
    ws.cell(row=1, column=1, value=cu_label).font = F_TITLE
    ws.cell(row=2, column=1, value="Change Analysis - Period over Period").font = F_BOLD

    if not prior_path or not prior_period:
        ws.cell(row=3, column=1, value=f"Current period: {period}").font = F_SUB
        ws.cell(
            row=5,
            column=1,
            value=(
                "No prior SCALE report is available for comparison. "
                "Run at least one earlier quarter for this CU to enable "
                "change analysis."
            ),
        ).font = F_NOTE
        _set_active_cover(wb)
        try:
            wb.save(p)
            out["ok"] = True
            return out
        except Exception as exc:  # noqa: BLE001
            out["error"] = f"Failed to save workbook: {exc}"
            return out

    try:
        prior_wb = load_workbook(prior_path, data_only=True, read_only=True)
        prior = _parse_scale_snapshot(prior_wb)
        prior_wb.close()
    except Exception as exc:  # noqa: BLE001
        ws.cell(row=3, column=1, value=f"Current: {period} | Prior: {prior_period}").font = F_SUB
        ws.cell(row=5, column=1, value=f"Prior report could not be read: {exc}").font = F_NOTE
        _set_active_cover(wb)
        try:
            wb.save(p)
            out["ok"] = True
            out["prior_path"] = str(prior_path)
            out["prior_period"] = prior_period
            return out
        except Exception as save_exc:  # noqa: BLE001
            out["error"] = f"Failed to save workbook: {save_exc}"
            return out

    ws.cell(row=3, column=1, value=f"Current: {period} | Prior: {prior_period}").font = F_SUB

    def _q(sheet_name: str) -> str:
        return "'" + sheet_name.replace("'", "''") + "'"

    cur_pools: list[tuple[str, int]] = []
    for rr in range(POOL_ROW_START, POOL_ROW_END + 1):
        pool = _s(calc_ws.cell(rr, 3).value)
        if pool and pool.lower() != "total":
            cur_pools.append((pool, rr))

    all_pools = [p for p, _ in cur_pools] + [x for x in prior["order"] if x not in {p for p, _ in cur_pools}]
    row_by_pool = {p: r0 for p, r0 in cur_pools}
    cache_has_signal = (
        abs(cur_cached.get("totals", {}).get("total_allow_needed", 0.0)) > 0.01
        or abs(cur_cached.get("totals", {}).get("total_spec_allow", 0.0)) > 0.01
        or any(
            abs(v.get("total_allow", 0.0)) > 0.01
            for v in (cur_cached.get("pools", {}) or {}).values()
        )
    )
    pool_rows: list[dict[str, Any]] = []
    for pool in all_pools:
        pr = prior["pools"].get(pool, {})
        cur_allow = cur_cached.get("pools", {}).get(pool, {}).get("total_allow", 0.0)
        prior_allow = pr.get("total_allow", 0.0)
        delta = cur_allow - prior_allow
        pool_rows.append({
            "pool": pool,
            "calc_row": row_by_pool.get(pool),
            "current": cur_allow,
            "prior": prior_allow,
            "delta": delta,
            "pct": (delta / prior_allow) if prior_allow else None,
        })

    r = 5
    ws.cell(row=r, column=1, value="Pool ACL Allowance Variance").font = F_BOLD
    r += 1
    _set_table_header(ws, r, ["Pool", "Current", "Prior", "$ Change", "% Change"], fill_hdr)
    r += 1
    for row in pool_rows:
        flag = cache_has_signal and _is_significant(row["delta"], row["prior"], row["current"])
        ws.cell(row=r, column=1, value=row["pool"]).font = F_CELL
        if row.get("calc_row") is not None:
            ws.cell(
                row=r,
                column=2,
                value=f"={_q(calc_name)}!U{int(row['calc_row'])}",
            ).number_format = ACCT
        else:
            ws.cell(row=r, column=2, value=0).number_format = ACCT
        ws.cell(row=r, column=3, value=row["prior"]).number_format = ACCT
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = ACCT
        pct_cell = ws.cell(row=r, column=5, value=f"=IF(C{r}=0,\"\",D{r}/C{r})")
        pct_cell.number_format = PCT
        for cidx in range(1, 6):
            cell = ws.cell(row=r, column=cidx)
            cell.border = BORDER
            if cidx > 1:
                cell.font = F_CELL
            if flag:
                cell.fill = FILL_FLAG
        r += 1

    cur_pool_formula = (
        f"={_q(calc_name)}!U23-{_q(impaired_name)}!P24"
        if impaired_name else f"={_q(calc_name)}!U23"
    )
    cur_pool = cur_cached.get("totals", {}).get("pooled_total_allow", 0.0)
    prior_pool = prior["totals"].get("pooled_total_allow", 0.0)
    pct = ((cur_pool - prior_pool) / prior_pool) if prior_pool else None
    for col, val, fmt in (
        (1, "Pooled Total Allowance", None),
        (2, cur_pool_formula, ACCT),
        (3, prior_pool, ACCT),
        (4, None, ACCT),
        (5, None, PCT),
    ):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = F_BOLD
        cell.fill = fill_tot
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt
    ws.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = ACCT
    ws.cell(row=r, column=5, value=f"=IF(C{r}=0,\"\",D{r}/C{r})").number_format = PCT
    r += 2

    ws.cell(row=r, column=1, value="Impaired Loan Reserve Variance").font = F_BOLD
    r += 1
    _set_table_header(ws, r, ["Impairment Type", "Current", "Prior", "$ Change", ""], fill_hdr)
    r += 1
    imp_rows: dict[str, int] = {}
    if impaired_ws is not None:
        for rr in range(5, 10):
            label = _s(impaired_ws.cell(rr, 1).value)
            if label:
                imp_rows[label] = rr
    cats = list(imp_rows.keys()) + [x for x in prior["impaired"].keys() if x not in imp_rows]
    for cat in cats:
        cv = cur_cached.get("impaired", {}).get(cat, 0.0)
        pv = prior["impaired"].get(cat, 0.0)
        ws.cell(row=r, column=1, value=cat).font = F_CELL
        if cat in imp_rows and impaired_name:
            ws.cell(
                row=r,
                column=2,
                value=f"={_q(impaired_name)}!P{int(imp_rows[cat])}",
            ).number_format = ACCT
        else:
            ws.cell(row=r, column=2, value=0).number_format = ACCT
        ws.cell(row=r, column=3, value=pv).number_format = ACCT
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = ACCT
        for cidx in range(1, 5):
            cell = ws.cell(row=r, column=cidx)
            cell.border = BORDER
            if cidx > 1:
                cell.font = F_CELL
        r += 1
    cimp_formula = f"={_q(impaired_name)}!P24" if impaired_name else "=0"
    cimp = cur_cached.get("totals", {}).get("total_spec_allow", 0.0)
    pimp = prior["totals"].get("total_spec_allow", 0.0)
    for col, val, fmt in (
        (1, "Total Specifically Identified", None),
        (2, cimp_formula, ACCT),
        (3, pimp, ACCT),
        (4, None, ACCT),
    ):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = F_BOLD
        cell.fill = fill_tot
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt
    ws.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = ACCT
    r += 2

    ws.cell(row=r, column=1, value="Summary").font = F_BOLD
    r += 1
    _set_table_header(ws, r, ["Metric", "Current", "Prior", "$ Change", ""], fill_hdr)
    r += 1
    for label, key in (
        ("Total Final CECL ACL", "total_allow_needed"),
        ("Pooled CECL ACL", "pooled_total_allow"),
        ("Specific/Impaired CECL ACL", "total_spec_allow"),
        ("Pooled Basis Loans", "pooled_loans"),
    ):
        cv = cur_cached.get("totals", {}).get(key, 0.0)
        pv = prior["totals"].get(key, 0.0)
        ws.cell(row=r, column=1, value=label).font = F_CELL
        if key == "total_allow_needed":
            ws.cell(row=r, column=2, value=f"={_q(calc_name)}!U23").number_format = ACCT
        elif key == "pooled_total_allow":
            ws.cell(row=r, column=2, value=cur_pool_formula).number_format = ACCT
        elif key == "total_spec_allow":
            ws.cell(row=r, column=2, value=cimp_formula).number_format = ACCT
        else:  # pooled_loans
            ws.cell(row=r, column=2, value=f"={_q(calc_name)}!I23").number_format = ACCT
        ws.cell(row=r, column=3, value=pv).number_format = ACCT
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = ACCT
        for cidx in range(1, 5):
            cell = ws.cell(row=r, column=cidx)
            cell.border = BORDER
            if cidx > 1:
                cell.font = F_CELL
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Analysis of Significant Changes").font = F_BOLD
    r += 1
    notes: list[str] = []
    if cache_has_signal:
        total_delta = cur_cached.get("totals", {}).get("total_allow_needed", 0.0) - prior["totals"].get("total_allow_needed", 0.0)
        if abs(total_delta) >= 1000:
            notes.append(
                f"Total Final CECL ACL {'increased' if total_delta > 0 else 'decreased'} "
                f"${abs(total_delta):,.0f} versus the prior quarter to "
                f"${cur_cached.get('totals', {}).get('total_allow_needed', 0.0):,.0f}."
            )
        sig = [x for x in sorted(pool_rows, key=lambda z: -abs(z['delta'])) if _is_significant(x['delta'], x['prior'], x['current'])][:4]
        for row in sig:
            pct_txt = f" ({row['pct']:+.1%})" if row["pct"] is not None else ""
            notes.append(f"{row['pool']} changed ${row['delta']:+,.0f}{pct_txt} quarter over quarter.")
        imp_delta = cimp - pimp
        if abs(imp_delta) >= 1000:
            notes.append(
                f"Impaired-specific reserves {'rose' if imp_delta > 0 else 'fell'} "
                f"${abs(imp_delta):,.0f} to ${cimp:,.0f}."
            )
    if not notes:
        notes.append(
            "Current-period values are linked to this workbook's live "
            "calculation tabs and may update after Excel recalculates."
        )
    for n in notes:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        cell = ws.cell(row=r, column=1, value="- " + n)
        cell.font = F_NOTE
        cell.alignment = LEFT_WRAP
        ws.row_dimensions[r].height = 14 * max(1, math.ceil(len(n) / 95)) + 4
        r += 1

    if vizo:
        # Wrap long pool/label names in column A so nothing is clipped.
        for row_cells in ws.iter_rows(min_col=1, max_col=1):
            a_cell = row_cells[0]
            if isinstance(a_cell.value, str):
                al = a_cell.alignment
                a_cell.alignment = Alignment(
                    horizontal=al.horizontal,
                    vertical=al.vertical or "center",
                    wrap_text=True,
                )

    _set_active_cover(wb)

    try:
        wb.save(p)
    except Exception as exc:  # noqa: BLE001
        out["error"] = f"Failed to save workbook: {exc}"
        return out

    out["ok"] = True
    out["prior_path"] = str(prior_path)
    out["prior_period"] = prior_period
    return out


def append_tct_change_analysis(workbook_path: str | Path) -> dict[str, Any]:
    """TCT variant: 'Change Analysis' tab after 'Calc tab'."""
    return append_change_analysis(workbook_path, "Change Analysis", is_vizo=False)


def append_vizo_change_analysis(workbook_path: str | Path) -> dict[str, Any]:
    """Vizo variant: fills the 'Change Analysis' placeholder tab."""
    return append_change_analysis(workbook_path, "Change Analysis", is_vizo=True)
