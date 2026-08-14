"""Parse a credit union's risk-tier / FICO score-range document into
credit-grade bands the setup wizard can pre-fill.

Filenames for these documents vary widely between CUs, so discovery is left to
the analyst (upload on the Credit Grades step); this module only has to be
tolerant of *content* layout. It handles:

  * ``.docx`` — parsed straight from the zip's ``word/document.xml`` (no
    third-party dependency), reading both real tables and free-form
    paragraphs (many CUs type the tiers as plain text).
  * ``.xlsx`` — first worksheet, via openpyxl (already a project dependency).

Range notations understood: ``700+``, ``650-699``, ``>=700``, ``below 500``,
``549 or less``, and a two-column ``min | max`` layout in spreadsheets.

The result is always *proposed* — the caller shows it for confirmation; nothing
is applied silently, because free-form documents are error-prone and reserve
rates are never present in these files.
"""
from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

_W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"

# A tier label: a letter optionally followed by +/-/digits/letters (A, A+, B2,
# "Prime"), kept short so prose sentences aren't mistaken for labels.
_TIER = r"[A-Za-z][A-Za-z0-9+\-]{0,19}"

# Score-range forms, most specific first.
_RANGE = re.compile(
    r"""
      (?P<lo>\d{3})\s*[-\u2013\u2014to]+\s*(?P<hi>\d{3})   # 650-699 / 650 to 699
    | (?P<plus>\d{3})\s*\+                                 # 700+
    | (?:>=?\s*)(?P<ge>\d{3})                              # >=700 / >700
    | (?:below|under|less\s+than|<)\s*(?P<lt>\d{3})        # below 500
    | (?P<le>\d{3})\s*(?:or\s*less|and\s*below|or\s*below) # 549 or less
    """,
    re.I | re.X,
)

_SKIP_LABELS = {
    "tier", "tiers", "score", "scores", "grade", "grades", "range", "ranges",
    "here", "fico", "credit", "rating", "ratings", "risk", "band", "bands",
    "our", "the", "these", "reserve", "rate", "rates",
}


def _para_text(p) -> str:
    return "".join(t.text or "" for t in p.iter(_W + "t"))


def _docx_lines_tables(path: Path) -> tuple[list[str], list[list[list[str]]]]:
    with zipfile.ZipFile(path) as z:
        root = ET.fromstring(z.read("word/document.xml"))
    body = root.find(_W + "body")
    lines: list[str] = []
    tables: list[list[list[str]]] = []
    if body is None:
        return lines, tables
    for el in body:
        tag = el.tag.replace(_W, "")
        if tag == "p":
            txt = _para_text(el).strip()
            if txt:
                lines.append(txt)
        elif tag == "tbl":
            rows: list[list[str]] = []
            for tr in el.findall(_W + "tr"):
                rows.append([
                    " ".join(_para_text(p).strip() for p in tc.findall(_W + "p")).strip()
                    for tc in tr.findall(_W + "tc")
                ])
            if rows:
                tables.append(rows)
    return lines, tables


def _xlsx_lines_tables(path: Path) -> tuple[list[str], list[list[list[str]]]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c).strip() for c in row]
            if any(cells):
                rows.append(cells)
    finally:
        wb.close()
    lines = [" ".join(c for c in r if c) for r in rows]
    return lines, ([rows] if rows else [])


def _range_from_match(m: re.Match) -> tuple[int | None, int | None]:
    d = m.groupdict()
    if d.get("lo"):
        return int(d["lo"]), int(d["hi"])
    if d.get("plus"):
        return int(d["plus"]), None
    if d.get("ge"):
        return int(d["ge"]), None
    if d.get("lt"):
        return None, int(d["lt"]) - 1
    if d.get("le"):
        return None, int(d["le"])
    return None, None


def _tier_from_row(cells: list[str]) -> tuple[str, int | None, int | None] | None:
    """Best-effort (label, min, max) from a spreadsheet/table row."""
    if not cells:
        return None
    label = (cells[0] or "").strip()
    if not label or label.lower() in _SKIP_LABELS or not re.fullmatch(_TIER, label):
        return None
    joined = " ".join(cells[1:]).strip()
    m = _RANGE.search(joined)
    if m:
        lo, hi = _range_from_match(m)
        return label, lo, hi
    # Two-column numeric min | max (e.g. "A", "700", "850").
    nums = [int(c) for c in cells[1:] if re.fullmatch(r"\d{3}", c.strip())]
    if len(nums) >= 2:
        return label, min(nums[0], nums[1]), max(nums[0], nums[1])
    if len(nums) == 1:
        # Single number + a trailing "+" cell → open-topped.
        if any("+" in c or re.search(r"(?i)\+|above|over", c) for c in cells[1:]):
            return label, nums[0], None
        return label, nums[0], None
    return None


def _extract(lines: list[str], tables: list[list[list[str]]]) -> list[dict[str, Any]]:
    out: list[tuple[str, int | None, int | None]] = []
    seen: set[str] = set()

    # 1) Tables first (most reliable).
    for rows in tables:
        for r in rows:
            got = _tier_from_row(r)
            if got and got[0].lower() not in seen:
                out.append(got)
                seen.add(got[0].lower())

    # 2) Free-form paragraph text. Insert a break before a tier letter that
    #    immediately follows a digit so merged lines split (e.g. the common
    #    "600-649D 550-599" where a soft line-break was lost on save).
    if not out:
        blob = "\n".join(lines)
        blob = re.sub(r"(\d)\s*([A-Za-z])\s+(\d{3})", r"\1\n\2 \3", blob)
        for ln in blob.splitlines():
            ln = ln.strip()
            mlabel = re.match(r"^(" + _TIER + r")\b\s*(.*)$", ln)
            if not mlabel:
                continue
            label, rest = mlabel.group(1), mlabel.group(2)
            if label.lower() in _SKIP_LABELS:
                continue
            rm = _RANGE.search(rest)
            if not rm:
                continue
            lo, hi = _range_from_match(rm)
            if label.lower() not in seen:
                out.append((label, lo, hi))
                seen.add(label.lower())

    return [{"label": lbl, "min_score": lo, "max_score": hi} for lbl, lo, hi in out]


def _warnings(tiers: list[dict[str, Any]]) -> list[str]:
    warns: list[str] = []
    if not tiers:
        return warns
    if len(tiers) < 2:
        warns.append("Only one tier was found — double-check the document layout.")
    # Sort by min (open/None last) to check ordering / gaps.
    ranked = sorted(
        tiers,
        key=lambda t: (t.get("min_score") if t.get("min_score") is not None else -1),
        reverse=True,
    )
    for i in range(len(ranked) - 1):
        cur_min = ranked[i].get("min_score")
        nxt_max = ranked[i + 1].get("max_score")
        if cur_min is not None and nxt_max is not None:
            if nxt_max >= cur_min:
                warns.append(
                    f"Ranges for {ranked[i]['label']} and {ranked[i+1]['label']} "
                    "overlap — please review."
                )
            elif nxt_max < cur_min - 1:
                warns.append(
                    f"Gap between {ranked[i+1]['label']} (\u2264{nxt_max}) and "
                    f"{ranked[i]['label']} (\u2265{cur_min}) — please review."
                )
    warns.append(
        "Reserve rates aren't defined in tier documents — enter them for each "
        "tier before saving."
    )
    return warns


def parse_risk_tiers(path: str | Path) -> dict[str, Any]:
    """Parse a risk-tier document into proposed credit-grade bands.

    Returns::

        {"ok": bool,
         "tiers": [{"label", "min_score", "max_score"}, ...],
         "warnings": [str, ...],
         "raw_lines": [str, ...],   # for a preview / troubleshooting
         "error": str}             # only when ok is False
    """
    p = Path(path)
    suffix = p.suffix.lower()
    try:
        if suffix == ".docx":
            lines, tables = _docx_lines_tables(p)
        elif suffix in (".xlsx", ".xlsm"):
            lines, tables = _xlsx_lines_tables(p)
        elif suffix == ".doc":
            return {"ok": False, "tiers": [], "warnings": [], "raw_lines": [],
                    "error": ("Legacy .doc isn't supported — save it as .docx "
                              "(or .xlsx) and re-upload.")}
        else:
            return {"ok": False, "tiers": [], "warnings": [], "raw_lines": [],
                    "error": f"Unsupported file type '{suffix}'. Upload a .docx or .xlsx."}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "tiers": [], "warnings": [], "raw_lines": [],
                "error": f"Could not read the file: {exc}"}

    tiers = _extract(lines, tables)
    if not tiers:
        return {
            "ok": False, "tiers": [], "warnings": [],
            "raw_lines": lines[:40],
            "error": ("No risk tiers / score ranges were found. Expected rows "
                      "like 'A  700+' or 'B  650-699'."),
        }
    return {
        "ok": True,
        "tiers": tiers,
        "warnings": _warnings(tiers),
        "raw_lines": lines[:40],
    }
