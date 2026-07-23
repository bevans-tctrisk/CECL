"""New-client pool_map bootstrap (Setup Router).

A brand-new client has no ``pool_map``, so the tie-to-book detector has nothing
to map. This module proposes an initial ``code -> pool`` map WITHOUT a prior
config, then validates it by tie-to-book — the same ground-truth check.

Phase-0 learning: numeric codes are NOT standardized across CUs, so we key off
**descriptions**. Few extracts carry a description *for the pool code itself*,
but most carry categorical descriptors (purpose / collateral) that correlate
with it. The bootstrap:
  1. reads the canonical pool names + book balances from the monthly file;
  2. aggregates, per pool code, its dominant purpose/collateral descriptor tokens;
  3. scores those tokens against each book pool's keyword lexicon;
  4. assembles a candidate ``pool_map`` and validates it by tie-to-book;
  5. returns per-pool tie status + a list of AMBIGUOUS codes (low confidence or
     pool doesn't tie) for human / AI resolution — the mechanical ~80% is
     proposed, the judgment ~20% is surfaced (design principle).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field

import pandas as pd

# Pool inference is two ORTHOGONAL dimensions, because book pools are
# combinations of them (e.g. AG REAL ESTATE = ag-segment x real-estate-collateral).
# Keying on a single keyword set conflates AG LOAN with AG REAL ESTATE.
_SEGMENT_KW: dict[str, list[str]] = {
    "ag": ["agricultural", "operating", "livestock", "farm", "crop",
           "ag ", "bare land", "land purchase"],
    "business": ["commercial", "business"],
    "consumer": ["consumer", "signature", "personal", "debt consolidation",
                 "vehicle", "auto", "home", "residential", "recreational",
                 "refinance", "new construction", "heloc"],
}
_COLLATERAL_KW: dict[str, list[str]] = {
    "real_estate": ["mortgage", "real estate", "bare land", "land",
                    "dwelling", "deed of trust", "heloc", "home equity",
                    "mobile home", "1-4 family"],
    "auto": ["auto", "vehicle", "car", "truck", "recreational"],
    "shares": ["shares", "share", "savings", "certificate", "pledged"],
    "equipment": ["equipment"],
    "signature": ["signature", "unsecured"],
}

# (segment, collateral) -> canonical book-pool label.
_COMBINE: dict[tuple[str, str], str] = {
    ("consumer", "auto"): "AUTO",
    ("consumer", "real_estate"): "RESIDENTIAL REAL ESTATE",
    ("consumer", "signature"): "UNSECURED",
    ("consumer", "equipment"): "UNSECURED",
    ("business", "real_estate"): "BUSINESS REAL ESTATE",
    ("business", "equipment"): "BUSINESS LOAN",
    ("business", "signature"): "BUSINESS LOAN",
    ("business", "auto"): "BUSINESS LOAN",
    ("ag", "real_estate"): "AG REAL ESTATE",
    ("ag", "equipment"): "AG LOAN",
    ("ag", "signature"): "AG LOAN",
    ("ag", "auto"): "AG LOAN",
}
# Collateral 'shares' always -> SHARE SECURED regardless of segment.
# Segment-only fallback when collateral is unknown.
_SEG_FALLBACK = {"consumer": "UNSECURED", "business": "BUSINESS LOAN",
                 "ag": "AG LOAN"}


def _top(kw_map: dict[str, list[str]], text: str) -> tuple[str | None, int, int]:
    """Return (best_key, best_hits, runner_up_hits) for a keyword map."""
    scores = {k: sum(1 for w in ws if w in text) for k, ws in kw_map.items()}
    ranked = sorted(scores.values(), reverse=True)
    best = max(scores, key=scores.get) if any(scores.values()) else None
    top = ranked[0] if ranked else 0
    second = ranked[1] if len(ranked) > 1 else 0
    return best, top, second


def _match_book_pool(label: str | None, book_pools: dict) -> str | None:
    if not label:
        return None
    ll = label.lower()
    for p in book_pools:
        if p.lower() == ll:
            return p
    for p in book_pools:
        if ll in p.lower() or p.lower() in ll:
            return p
    return None


def _infer_pool(tokens: list[str], book_pools: dict
                ) -> tuple[str | None, str]:
    """Infer (book_pool, confidence) from descriptor tokens via segment x
    collateral. confidence: high (both dims clear) | low (one dim / defaulted)
    | none (no signal)."""
    text = " ".join(tokens).lower()
    seg, seg_top, _ = _top(_SEGMENT_KW, text)
    col, col_top, _ = _top(_COLLATERAL_KW, text)

    if col == "shares":
        return _match_book_pool("SHARE SECURED", book_pools), "high"

    label = None
    conf = "none"
    if seg and col and (seg, col) in _COMBINE:
        label, conf = _COMBINE[(seg, col)], "high"
    elif col == "real_estate":
        # Real-estate collateral, segment unclear -> default residential (the
        # common case); low confidence so it surfaces for review.
        label = "RESIDENTIAL REAL ESTATE" if not seg else \
            _COMBINE.get((seg, "real_estate"), "RESIDENTIAL REAL ESTATE")
        conf = "high" if seg else "low"
    elif col == "auto":
        label, conf = ("AUTO" if seg in (None, "consumer") else
                       _COMBINE.get((seg, "auto"), "AUTO")), \
            ("high" if seg in (None, "consumer") else "low")
    elif seg:
        label, conf = _SEG_FALLBACK.get(seg), "low"

    # Equipment collateral is irreducibly ambiguous between BUSINESS LOAN and
    # AG LOAN (the same "EQUIPMENT PURCHASE" descriptor appears in both; the CU
    # encodes the split in the code NUMBER, not the description). Never claim
    # high confidence — surface for tie-guided / AI / CU-reference resolution.
    if col == "equipment" and conf == "high":
        conf = "low"
    return _match_book_pool(label, book_pools), conf


@dataclass
class CodeProposal:
    code: str
    pool: str | None
    balance: float
    confidence: str          # high | low | none
    tokens: list[str] = field(default_factory=list)
    scores: dict = field(default_factory=dict)


@dataclass
class BootstrapResult:
    pool_map: dict[str, str]
    code_proposals: list[CodeProposal]
    book_pools: dict[str, float]
    pools_tied: int
    pools_total: int
    total_abs_diff: float
    ambiguous: list[str]
    evidence: str = ""


def book_pools_from_monthly(path: str, snapshot: str) -> dict[str, float]:
    """Read ``{pool_name: balance}`` for ``snapshot`` from a Monthly Balances
    by Pool file (pools down col A, dates across row 1)."""
    import openpyxl
    from openpyxl.utils import column_index_from_string as ci  # noqa: F401
    import datetime as _dt

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    y, m = int(snapshot[:4]), int(snapshot[5:7])
    col_idx = None
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, _dt.datetime) and v.year == y and v.month == m:
            col_idx = c
            break
    out: dict[str, float] = {}
    if col_idx is None:
        return out
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name is None:
            continue
        name = str(name).strip()
        low = name.lower()
        if not name or any(w in low for w in (
                "total", "alll", "provision", "unfunded", "percent",
                "matches", "risky", "%")):
            continue
        val = ws.cell(row=r, column=col_idx).value
        try:
            out[name] = float(val) if val is not None else 0.0
        except (TypeError, ValueError):
            continue
    return out


def _descriptor_columns(df: pd.DataFrame) -> list[str]:
    """Categorical descriptor columns (low-cardinality text) useful for pool
    inference — e.g. PURPOSE_CODE_DESCRIPTION, COLLATERAL_CODE_DESCRIPTION.
    Requires a 'descrip' name, excludes geo/identity columns and free-text."""
    geo = re.compile(r"address|city|county|region|state|zip|postal|name|date"
                     r"|\bid\b|number|phone|vin", re.I)
    out = []
    for c in df.columns:
        if not re.search(r"descrip", str(c), re.I) or geo.search(str(c)):
            continue
        s = df[c].dropna().astype(str).str.strip()
        if s.empty:
            continue
        txt = s[~s.str.replace(".", "", 1).str.isdigit()]
        if txt.empty:
            continue
        if 1 <= txt.nunique() <= 40:      # categorical, not free-text
            out.append(c)
    return out


def _score_pool(tokens: list[str], book_pools: dict[str, float]) -> dict[str, int]:
    """Score each book pool by keyword hits in the descriptor tokens."""
    text = " ".join(tokens).lower()
    canon_score: dict[str, int] = {}
    for canon, kws in _CANON_KEYWORDS.items():
        canon_score[canon] = sum(1 for kw in kws if kw in text)

    # Real-estate disambiguation: if 'mortgage'/'real estate' present, route to
    # residential/business/ag RE by the accompanying purpose tokens.
    if "mortgage" in text or "real estate" in text:
        if any(t in text for t in _RE_AG):
            canon_score["ag real estate"] += 3
        elif any(t in text for t in _RE_BUS):
            canon_score["business real estate"] += 3
        else:
            canon_score["residential real estate"] += 3

    # Map canonical categories onto the actual book pool names (fuzzy contains).
    pool_score: dict[str, int] = {}
    for pool in book_pools:
        pl = pool.lower()
        best = 0
        for canon, sc in canon_score.items():
            if sc <= 0:
                continue
            # canonical label shares words with the book pool name?
            canon_words = set(canon.split())
            pool_words = set(pl.split())
            if canon_words & pool_words or canon in pl or pl in canon:
                best = max(best, sc)
        if best:
            pool_score[pool] = best
    return pool_score


def bootstrap(df: pd.DataFrame, code_col: str, balance_col: str,
              book_pools: dict[str, float], *, tol: float = 1.0,
              exclude_pools: set[str] | None = None) -> BootstrapResult:
    exclude = {p.lower() for p in (exclude_pools or set())}
    targets = {p: b for p, b in book_pools.items() if p.lower() not in exclude}

    desc_cols = _descriptor_columns(df)
    bal = pd.to_numeric(df[balance_col], errors="coerce").fillna(0.0)
    codes = df[code_col].astype(str).str.strip()

    proposals: list[CodeProposal] = []
    pool_map: dict[str, str] = {}
    for code, grp in df.groupby(codes):
        code_bal = float(bal[grp.index].sum())
        tokens: list[str] = []
        for dc in desc_cols:
            vc = grp[dc].dropna().astype(str).str.strip()
            vc = vc[~vc.str.replace(".", "", 1).str.isdigit()]
            if not vc.empty:
                tokens.append(vc.mode().iloc[0])   # dominant descriptor
        best_pool, conf = _infer_pool(tokens, targets)
        if best_pool:
            pool_map[code] = best_pool
        proposals.append(CodeProposal(code=code, pool=best_pool,
                                      balance=round(code_bal, 2),
                                      confidence=conf, tokens=tokens,
                                      scores={}))

    # Validate: group extract by the candidate map and tie to the book.
    mapped = codes.map(pool_map)
    by_pool = bal.groupby(mapped).sum()
    tied = 0
    total_abs = 0.0
    for pool, book in targets.items():
        got = float(by_pool.get(pool, 0.0))
        total_abs += abs(got - book)
        if abs(got - book) <= tol:
            tied += 1
    ambiguous = [p.code for p in proposals if p.confidence != "high"]

    return BootstrapResult(
        pool_map=pool_map, code_proposals=proposals, book_pools=targets,
        pools_tied=tied, pools_total=len(targets),
        total_abs_diff=round(total_abs, 2), ambiguous=ambiguous,
        evidence=(f"bootstrapped {len(pool_map)} codes from descriptors "
                  f"{desc_cols}; {tied}/{len(targets)} pools tie to book "
                  f"(diff ${total_abs:,.2f}); {len(ambiguous)} ambiguous code(s)"))
