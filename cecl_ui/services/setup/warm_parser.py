"""WARM workbook parser (Setup Router — WARM ingestion path).

For CUs built from a delivered CECL-Migration-WARM workbook, the WARM IS a
filled-in configuration spec. Rather than reverse-engineering config from raw
extracts (the from-scratch path), read the authoritative mappings directly
from the WARM's well-known tabs.

Tab map (per analyst):
  * "Grade Ranges & Loan Codes" — loan-code->pool map, credit grade ranges,
    default loss/mgmt-adj rates, business-risk-rating legend, grade-movement.
  * "BS Data"                    — monthly balances incl ACL/ALLL.
  * "BS CO DQ Data Enter"        — pools (grade break-out y/n), ACL months.
  * "Mmm-yy Data" (e.g. Jun-26)  — member#/suffix, code->pool quirks, orig/curr score calc.
  * "Mmm-YY Credit Pull"         — credit pull history.
  * "CO Hist Data"/"CO Data"/"Recov Data" — CO/recovery/DQ history + detail.
  * "Risk Change Data Entry"     — mgmt adjustment + credit score movement.

This module currently implements the highest-value core (Grade Ranges & Loan
Codes). Other tabs are added incrementally.
"""

from __future__ import annotations

import glob
import os
import re
from dataclasses import dataclass, field

import openpyxl
from openpyxl.utils import column_index_from_string as _ci


def find_latest_warm(folder: str) -> str | None:
    """Return the most recent real WARM workbook in ``folder`` (root only),
    excluding DNU/backup/temp copies."""
    best = None
    best_key = ""
    for f in glob.glob(os.path.join(folder, "*.xls*")):
        b = os.path.basename(f)
        bl = b.lower()
        if "warm" not in bl or bl.startswith("~$") or "dnu" in bl \
                or "combined" in bl or "comined" in bl:
            continue
        m = re.match(r"(\d{4})-(\d{2})", b)
        key = m.group(0) if m else "0000-00"
        if key > best_key:
            best_key, best = key, f
    return best


def _txt(v) -> str:
    return "" if v is None else str(v).strip()


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


@dataclass
class WarmGradeLoanCodes:
    pool_map: dict[str, str] = field(default_factory=dict)
    credit_grades: list[dict] = field(default_factory=list)
    default_mgmt_adj_by_grade: dict[str, float] = field(default_factory=dict)
    brr_labels: dict[str, str] = field(default_factory=dict)
    not_reported_min: float | None = None
    ltv_baseline: float | None = None
    probability_factor: float | None = None


def parse_grade_loan_codes(ws) -> WarmGradeLoanCodes:
    out = WarmGradeLoanCodes()

    # ---- credit grades (cols A..G) ----
    # Header row has 'Grade Order' in col A; data until 'Total'.
    grade_rows = []
    for r in range(1, min(ws.max_row, 30) + 1):
        a = _txt(ws.cell(row=r, column=1).value)
        g = _txt(ws.cell(row=r, column=2).value)
        if g and g.lower() not in ("grade", "total") and not g.lower().startswith("hide") \
                and a.isdigit():
            rng_min = _num(ws.cell(row=r, column=4).value)
            loss = _num(ws.cell(row=r, column=6).value)
            madj = _num(ws.cell(row=r, column=7).value)
            grade_rows.append((int(a), g.strip(), rng_min, loss, madj))
        low = g.lower()
        if low == "minimum for not reported":
            out.not_reported_min = _num(ws.cell(row=r, column=4).value)
    # min for 'Not Reported' also lands in grade_rows if it had a numeric order;
    # convert to grades with min/max derived from the next-higher grade's min.
    graded = [t for t in grade_rows if t[2] is not None]
    graded.sort(key=lambda t: -(t[2] or 0))
    nr_min = out.not_reported_min or 900
    for i, (order, label, gmin, loss, madj) in enumerate(graded):
        gmax = int(nr_min) - 1 if i == 0 else int(graded[i - 1][2]) - 1
        out.credit_grades.append({
            "label": label, "min_score": int(gmin), "max_score": gmax,
            "reserve_rate": loss if loss is not None else 0.0011,
        })
        if madj is not None:
            out.default_mgmt_adj_by_grade[label] = madj
    for label, val in (("Not Reported", None),):
        pass

    # ---- mgmt-adj calc params (col F/G labelled block) ----
    for r in range(1, min(ws.max_row, 30) + 1):
        lbl = _txt(ws.cell(row=r, column=6).value).lower()
        if lbl == "ltv baseline":
            out.ltv_baseline = _num(ws.cell(row=r, column=7).value)
        elif lbl == "probability factor":
            out.probability_factor = _num(ws.cell(row=r, column=7).value)

    # ---- business risk rating legend (cols I..L) ----
    for r in range(3, min(ws.max_row, 40) + 1):
        code = _txt(ws.cell(row=r, column=_ci("I")).value)
        label = _txt(ws.cell(row=r, column=_ci("J")).value)
        if code and label and code.lower() not in ("rr order",):
            out.brr_labels[code] = label

    # ---- loan code -> pool (cols S..T) ----
    cS, cT = _ci("S"), _ci("T")
    for r in range(1, ws.max_row + 1):
        code = _txt(ws.cell(row=r, column=cS).value)
        pool = _txt(ws.cell(row=r, column=cT).value)
        if not code or not pool:
            continue
        if code.lower() in ("collateral code", "loan types") or pool.lower() == "loan pool":
            continue
        # Normalize the intentional sentinels.
        if pool == "**No pool**":
            pool = "Ignore"
        # Keep the FIRST mapping for a code. Some WARMs carry a legacy/detail
        # block further down the S:T columns that re-maps a code to a non-
        # reporting sub-pool (e.g. Maple code 27 -> 'Used Vehicle' in the main
        # block, then '-> USED VEHICLE LOANS-CUDL' later). The first (main)
        # block holds the reporting-pool mapping; later duplicates would orphan
        # those loans out of the reporting pool.
        if code not in out.pool_map:
            out.pool_map[code] = pool
    return out


def _to_month_end(v) -> str | None:
    """Normalize a BS Data date header (datetime or 'YYYY-MM') to month-end ISO."""
    import calendar
    import datetime as _dt
    if isinstance(v, _dt.datetime):
        y, m = v.year, v.month
    else:
        m2 = re.match(r"^(\d{4})[-/](\d{1,2})", _txt(v))
        if not m2:
            return None
        y, m = int(m2.group(1)), int(m2.group(2))
    if not (1 <= m <= 12):
        return None
    return f"{y:04d}-{m:02d}-{calendar.monthrange(y, m)[1]:02d}"


@dataclass
class WarmBsData:
    entries: dict[str, dict[str, float]] = field(default_factory=dict)  # pool -> {date: bal}
    alll: dict[str, float] = field(default_factory=dict)                # date -> ACL (abs)
    months: list[str] = field(default_factory=list)


def parse_bs_data(ws) -> WarmBsData:
    """Parse the 'BS Data' tab: monthly balances by pool (dates across cols,
    pools down rows) + the ACL/ALLL row. Returns absolute ACL values."""
    out = WarmBsData()
    # Date columns from the header row (row 1), from col C onward.
    date_cols: dict[int, str] = {}
    for c in range(3, ws.max_column + 1):
        iso = _to_month_end(ws.cell(row=1, column=c).value)
        if iso:
            date_cols[c] = iso
    out.months = [date_cols[c] for c in sorted(date_cols)]

    for r in range(2, ws.max_row + 1):
        label_a = _txt(ws.cell(row=r, column=1).value)
        pool = _txt(ws.cell(row=r, column=2).value) or label_a
        low = (label_a or pool).lower()
        if not (label_a or pool):
            continue
        if any(low.startswith(w) for w in
               ("total", "grand total", "balancer", "hide")):
            continue
        is_acl = "allowance" in low
        for c, iso in date_cols.items():
            val = _num(ws.cell(row=r, column=c).value)
            if val is None:
                continue
            if is_acl:
                out.alll[iso] = abs(val)
            else:
                out.entries.setdefault(pool, {})[iso] = val
    return out


# ---------------------------------------------------------------------------
# "Mmm-yy Data" tab — derive column_mappings from the WARM's own formulas.
#
# The Data tab pastes the raw core extract (e.g. Symitar LND* cols) on the left
# and builds the final CECL fields via formulas. Reading those formulas
# (data_only=False) traces each config field back to the raw source column, so
# we can assemble column_mappings + member_account + score logic authoritatively.
# ---------------------------------------------------------------------------

_MON = {m: i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], 1)}


def _data_tab_key(s: str):
    m = re.match(r"^([A-Z][a-z]{2})-(\d{2})", s)
    if not m:
        return (0, 0)
    return (int(m.group(2)), _MON.get(m.group(1), 0))


def _cross_sheet_names(formula: str) -> list[str]:
    """Sheet names referenced by a formula (quoted or bare, before '!')."""
    names = [m.group(1) for m in re.finditer(r"'([^']+)'!", formula)]
    names += [m.group(1) for m in
              re.finditer(r"(?<![A-Za-z0-9_'])([A-Za-z_][A-Za-z0-9_.]*)!", formula)]
    return names


def _col_refs(formula: str) -> list[tuple[str, int]]:
    """Same-sheet cell references (col_letter, row) — cross-sheet refs removed."""
    f = re.sub(r"'[^']*'![^\s,()]*", " ", formula)          # 'Sheet Name'!A1
    f = re.sub(r"(?<![A-Za-z0-9_'])[A-Za-z_][A-Za-z0-9_.]*![^\s,()]*",
               " ", f)                                        # Sheet!A1
    refs = []
    for m in re.finditer(r"(?<![A-Za-z0-9_$.])\$?([A-Z]{1,3})\$?(\d+)", f):
        refs.append((m.group(1), int(m.group(2))))
    return refs


def _trace(hdr: dict, datarow: dict, col_idx: int, depth: int = 0, seen=None):
    """Follow a cell's formula chain to the raw source column(s). Returns
    (raw_header_names_in_order, cross_sheet_names). All refs stay on the data
    row, so only ``hdr`` (row 1 names) and ``datarow`` (the data row values) are
    needed. A cell with no formula is a raw pasted value — its header is the
    source."""
    seen = set() if seen is None else seen
    if depth > 6 or col_idx in seen:
        return [], set()
    seen.add(col_idx)
    v = datarow.get(col_idx)
    if not (isinstance(v, str) and v.startswith("=")):
        h = _txt(hdr.get(col_idx))
        return ([h] if h else []), set()
    raws: list[str] = []
    cross = set(_cross_sheet_names(v))
    for cl, _rr in _col_refs(v):
        try:
            ci2 = _ci(cl)
        except ValueError:
            continue
        r2, c2 = _trace(hdr, datarow, ci2, depth + 1, seen)
        for x in r2:
            if x not in raws:
                raws.append(x)
        cross |= c2
    return raws, cross


def _detect_member_account(hdr: dict, datarow: dict, mem_c, suf_c) -> dict:
    """Infer member_account mode from the Member/Suffix helper formulas.

    Patterns seen: fixed_suffix  Member=ROUNDDOWN(<acct>*0.001,0), Suffix=RIGHT(<acct>,n)
                   split         Suffix=RIGHT(<sep col>,n) (distinct raw suffix column)
                   delimiter     Member uses LEFT/FIND/MID on a delimited id
    """
    res: dict = {}
    mem_f = datarow.get(mem_c) if mem_c else None
    suf_f = datarow.get(suf_c) if suf_c else None
    mem_f = mem_f if isinstance(mem_f, str) else ""
    suf_f = suf_f if isinstance(suf_f, str) else ""
    if mem_c:
        raws, _ = _trace(hdr, datarow, mem_c)
        if raws:
            res["member_source"] = raws[0]
    sm = re.search(r"RIGHT\(\s*\$?([A-Z]{1,3})\$?\d+\s*,\s*(\d+)\)", suf_f, re.I)
    if sm:
        suf_col, n = sm.group(1), int(sm.group(2))
        mm = re.search(r"\$?([A-Z]{1,3})\$?\d+", mem_f)
        mem_col = mm.group(1) if mm else None
        if mem_col and suf_col.upper() == mem_col.upper():
            res["mode"] = "fixed_suffix"
            res["suffix_length"] = n
        else:
            res["mode"] = "split"
            res["suffix_length"] = n
            src = _txt(hdr.get(_ci(suf_col)))
            if src:
                res["suffix_source"] = src
    elif re.search(r"LEFT\(|FIND\(|MID\(", mem_f, re.I):
        res["mode"] = "delimiter"
    return res


@dataclass
class WarmDataTab:
    tab: str = ""
    data_row: int | None = None
    column_mappings: dict = field(default_factory=dict)  # config field -> raw header
    member_account: dict = field(default_factory=dict)
    original_score_source: str = ""
    current_score_source: str = ""   # 'credit_pull' or a raw header
    ssn_column: str = ""             # raw loan SSN header (credit-pull join key)


def parse_data_tab(ws) -> WarmDataTab:
    """Derive column_mappings/member_account/score sources from a Data tab's
    formulas. ``ws`` must expose formulas (workbook loaded data_only=False);
    read_only is fine — this reads only the header row + first data row."""
    out = WarmDataTab(tab=ws.title)
    rows = ws.iter_rows()
    first = next(rows, None)
    if first is None:
        return out
    hdr = {i: c.value for i, c in enumerate(first, 1) if _txt(c.value)}
    name2idx = {_txt(v).lower(): k for k, v in hdr.items()}

    key_idx = None
    for key in ("current balance", "loan type", "interest rate"):
        if key in name2idx:
            key_idx = name2idx[key]
            break
    if not key_idx:
        return out

    datarow = None
    for row in rows:
        v = row[key_idx - 1].value if key_idx - 1 < len(row) else None
        if isinstance(v, str) and v.startswith("="):
            datarow = {i: c.value for i, c in enumerate(row, 1)}
            out.data_row = row[0].row
            break
    if datarow is None:
        return out

    def first_raw(header_key):
        c = name2idx.get(header_key)
        if not c:
            return None, set()
        raws, cross = _trace(hdr, datarow, c)
        return (raws[0] if raws else None), cross

    for field_name, hkey in (
            ("current_balance", "current balance"),
            ("loan_pool_code", "loan type"),
            ("interest_rate", "interest rate"),
            ("original_loan_amount", "original loan amount")):
        src, _ = first_raw(hkey)
        if src:
            out.column_mappings[field_name] = src

    org_src, _ = first_raw("org score")
    if not org_src:
        org_src, _ = first_raw("original credit score")
    if org_src:
        out.original_score_source = org_src
        out.column_mappings["original_fico_score"] = org_src

    cur_c = name2idx.get("curr score from pull") or name2idx.get("current credit score")
    if cur_c:
        raws, cross = _trace(hdr, datarow, cur_c)
        if any("credit pull" in s.lower() for s in cross):
            out.current_score_source = "credit_pull"
        elif raws:
            out.current_score_source = raws[0]
            out.column_mappings["current_fico_score"] = raws[0]

    out.member_account = _detect_member_account(
        hdr, datarow, name2idx.get("member"), name2idx.get("suffix"))
    if out.member_account.get("member_source"):
        out.column_mappings["member_number"] = out.member_account["member_source"]

    # open_date (loan origination date): the WARM's 'Open Date' final column is
    # a snapshot placeholder (=EOMONTH(snapshot)), so formula-tracing won't find
    # it. Grab the real origination date straight from the raw extract header
    # (Symitar LNDOPEN, or a generic *open*date* column).
    for _idx, _nm in hdr.items():
        _u = _txt(_nm).upper()
        if _u == "LNDOPEN" or (("OPEN" in _u) and ("DATE" in _u)) or _u.endswith("OPEN"):
            out.column_mappings["open_date"] = _txt(_nm)
            break

    # SSN column carried from the raw extract (credit-pull join key for CUs
    # whose pull is keyed by SSN rather than member number, e.g. Symitar/DEXA
    # 'MBR SSN STRING'). Detected here so the loan side can join the pull on it.
    for _idx, _nm in hdr.items():
        _l = _txt(_nm).lower()
        if "ssn" in _l or "social security" in _l:
            out.ssn_column = _txt(_nm)
            break
    return out


def _mmm_yy_to_month_end(s: str) -> str | None:
    import calendar
    m = re.match(r"^([A-Z][a-z]{2})-(\d{2})", s)
    if not m:
        return None
    mon = _MON.get(m.group(1))
    yy = 2000 + int(m.group(2))
    if not mon:
        return None
    return f"{yy:04d}-{mon:02d}-{calendar.monthrange(yy, mon)[1]:02d}"


@dataclass
class WarmCreditPull:
    tab: str = ""
    period: str | None = None            # month-end ISO derived from tab name
    scores: dict = field(default_factory=dict)   # member -> FICO (int)
    member_header: str = ""              # header of the pull's key column


def parse_credit_pull(ws) -> WarmCreditPull:
    """Parse a 'Mmm-YY Credit Pull' tab: 2 cols (Member Number, FICO)."""
    out = WarmCreditPull(tab=ws.title, period=_mmm_yy_to_month_end(ws.title))
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header:
        out.member_header = _txt(header[0])
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        mem = _txt(row[0])
        if not mem or mem.lower().startswith("member"):
            continue
        fico = _num(row[1]) if len(row) > 1 else None
        if fico is None:
            continue
        out.scores[mem] = int(fico)
    return out


@dataclass
class WarmPoolSettings:
    pools: list = field(default_factory=list)   # per-pool dicts
    snapshot_date: str | None = None
    charter_number: str | None = None
    cu_name: str | None = None
    core_processor: str | None = None
    economic_stress: dict = field(default_factory=dict)  # state/county/unemp


def parse_bs_co_dq(ws) -> WarmPoolSettings:
    """Parse 'BS CO DQ Data Enter': per-pool risk-rated / grade-breakout / ACL
    horizon / pool order + CU metadata (charter #, period, core, econ stress)."""
    out = WarmPoolSettings()
    rows = [tuple(r) for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80),
                                           values_only=True)]

    # metadata is in a label/value block (cols L/M/N ~ idx 11/12/13)
    def _meta(label_sub):
        for r in rows:
            if len(r) > 12 and _txt(r[11]).lower().startswith(label_sub):
                return r[12]
        return None
    ch = _meta("charter")
    out.charter_number = _txt(ch) or None
    per = _meta("for period")
    out.snapshot_date = _to_month_end(per) if per is not None else None
    out.cu_name = _txt(_meta("credit union")) or None
    out.core_processor = _txt(_meta("economic stress")) or None  # placeholder; refined below
    # economic stress row: label row "State | County | Unemp %" then a value row
    for i, r in enumerate(rows):
        if len(r) > 13 and _txt(r[11]).lower() == "state" and _txt(r[12]).lower() == "county":
            if i + 1 < len(rows):
                vr = rows[i + 1]
                out.economic_stress = {
                    "state": _txt(vr[11]) or None,
                    "county": _txt(vr[12]) or None,
                    "unemployment": _num(vr[13]) if len(vr) > 13 else None,
                }
            break
    # core processor label
    cp = _meta("core processor")
    if cp is not None:
        out.core_processor = _txt(cp) or None

    # locate the pool table header row (col A == 'Loan Pools')
    hdr_i = None
    for i, r in enumerate(rows):
        if _txt(r[0]).lower() == "loan pools":
            hdr_i = i
            break
    if hdr_i is None:
        return out
    for r in rows[hdr_i + 1:]:
        pool = _txt(r[0])
        low = pool.lower()
        if not pool or low.startswith(("total", "grand")):
            if low.startswith(("total", "grand")):
                break
            continue
        out.pools.append({
            "pool": pool,
            "risk_rated": _txt(r[1]).lower().startswith("y") if len(r) > 1 else None,
            "balance": _num(r[2]) if len(r) > 2 else None,
            "dq_balance": _num(r[3]) if len(r) > 3 else None,
            "dq_pct": _num(r[4]) if len(r) > 4 else None,
            "grade_basis": _txt(r[5]) or None if len(r) > 5 else None,   # CG or RR
            "acl_months": _num(r[6]) if len(r) > 6 else None,
            "acl_quarters": _num(r[7]) if len(r) > 7 else None,
            "pool_order": _num(r[8]) if len(r) > 8 else None,
            "grade_broken_out": _txt(r[9]).lower().startswith("y") if len(r) > 9 else None,
        })
    return out


def parse_acl_reserve_rates(ws) -> dict:
    """Parse 'ACL Env by Pool Mgmt Adj': per-pool per-grade ACL Base Loss Rate
    (col E) + the pool Total blended rate + environmental factor (col I).

    Returns {pool: {'grades': {grade: rate}, 'blended': rate, 'env_factor': f}}.
    These are the analyst-computed loss rates; feeding them back as
    ``base_loss_rate_by_pool_grade`` lets the report reproduce the WARM exactly
    (the firm-wide model can't re-derive them without the same CO history).
    """
    out: dict = {}
    cur = None
    for r in ws.iter_rows(values_only=True):
        if not r:
            continue
        a = _txt(r[0])
        if not a:
            continue
        low = a.lower()
        # pool header row: col A set, col B empty, not a header/total/banner
        if (a not in ("Current Grade", "Total")
                and (len(r) < 2 or r[1] is None)
                and not low.startswith(("tongass", "allowance", "for period",
                                        "pooled"))):
            cur = a
            out[cur] = {"grades": {}, "blended": None, "env_factor": None,
                        "balance_only": False, "_grade_bal": 0.0}
            continue
        if cur is None or a == "Current Grade":
            continue
        if a == "Total":
            out[cur]["blended"] = _num(r[6]) if len(r) > 6 else None  # col G = E+F
            out[cur]["env_factor"] = _num(r[8]) if len(r) > 8 else None
            out[cur]["allow_before"] = _num(r[7]) if len(r) > 7 else None
            total_bal = _num(r[1]) if len(r) > 1 else None
            # balance-only pool: carries a book balance but no per-grade
            # loan-level detail (overdraft / participations). These render as
            # a single Total row (NRR) in the report.
            out[cur]["balance_only"] = bool(
                total_bal and total_bal > 0.005
                and out[cur]["_grade_bal"] <= 0.005)
            out[cur].pop("_grade_bal", None)
            cur = None
            continue
        if a.lower().startswith("hide"):
            continue
        gbal = _num(r[1]) if len(r) > 1 else None
        out[cur]["_grade_bal"] += abs(gbal or 0)
        base = _num(r[4]) if len(r) > 4 else None   # col E = ACL base loss rate
        factor = _num(r[6]) if len(r) > 6 else None  # col G = allowance factor (E+F)
        # The override pins the FULL allowance factor (base + mgmt adj) so the
        # report reproduces the WARM even when the mgmt-adj carry-forward is
        # unavailable. ``max_base`` tracks the true base rate (col E) so the
        # analyst-mgmt-adj-only pools are still detected for warm_allowance_pools.
        if factor is not None:
            out[cur]["grades"][a] = factor
        if base is not None:
            out[cur]["max_base"] = max(out[cur].get("max_base", 0.0), abs(base))
    return out


def parse_warm(path: str) -> dict:
    """Parse a WARM workbook into partial config pieces. Returns a dict with
    keys populated per tab available."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    result: dict = {"source_warm": os.path.basename(path),
                    "sheets": list(wb.sheetnames)}
    if "Grade Ranges & Loan Codes" in wb.sheetnames:
        glc = parse_grade_loan_codes(wb["Grade Ranges & Loan Codes"])
        result["pool_map"] = glc.pool_map
        result["credit_grades"] = glc.credit_grades
        result["default_mgmt_adj_by_grade"] = glc.default_mgmt_adj_by_grade
        result["brr_labels"] = glc.brr_labels
        result["mgmt_adj_params"] = {
            "ltv_baseline": glc.ltv_baseline,
            "probability_factor": glc.probability_factor,
            "not_reported_min": glc.not_reported_min,
        }
    if "BS Data" in wb.sheetnames:
        bs = parse_bs_data(wb["BS Data"])
        result["monthly_balance"] = {"source": "manual",
                                     "months": bs.months,
                                     "entries": bs.entries}
        result["alll"] = bs.alll
    if "BS CO DQ Data Enter" in wb.sheetnames:
        ps = parse_bs_co_dq(wb["BS CO DQ Data Enter"])
        result["pool_settings"] = ps.pools
        result["snapshot_date"] = ps.snapshot_date
        result["cu_meta"] = {"charter_number": ps.charter_number,
                             "cu_name": ps.cu_name,
                             "core_processor": ps.core_processor,
                             "economic_stress": ps.economic_stress}
    if "ACL Env by Pool Mgmt Adj" in wb.sheetnames:
        result["acl_reserve_rates"] = parse_acl_reserve_rates(
            wb["ACL Env by Pool Mgmt Adj"])
    # data + credit-pull tab inventory
    result["data_tabs"] = [s for s in wb.sheetnames
                           if re.match(r"^[A-Z][a-z]{2}-\d{2} Data$", s)]
    result["credit_pull_tabs"] = [s for s in wb.sheetnames
                                  if re.search(r"Credit Pull", s)]
    if result["credit_pull_tabs"]:
        latest_cp = sorted(result["credit_pull_tabs"], key=_data_tab_key)[-1]
        cp = parse_credit_pull(wb[latest_cp])
        result["credit_pull"] = {"tab": cp.tab, "period": cp.period,
                                 "count": len(cp.scores), "scores": cp.scores,
                                 "member_header": cp.member_header}
    wb.close()

    # column_mappings need the FORMULAS (data_only=False). read_only is fine —
    # parse_data_tab reads only the header row + first data row.
    if result["data_tabs"]:
        latest = sorted(result["data_tabs"], key=_data_tab_key)[-1]
        wbf = openpyxl.load_workbook(path, data_only=False, read_only=True)
        try:
            dt = parse_data_tab(wbf[latest])
            result["data_tab_used"] = latest
            result["column_mappings"] = dt.column_mappings
            result["member_account"] = dt.member_account
            result["score_sources"] = {"original": dt.original_score_source,
                                        "current": dt.current_score_source}
            # SSN-keyed credit pull: when the pull's key column is an SSN
            # (header says so, or the keys are 9-digit) and the loan extract
            # carries an SSN column, the loans must join the pull on that
            # column rather than the member number. Surface it so the
            # assembler/router can emit credit_pull.loan_join_column.
            _cpd = result.get("credit_pull") or {}
            if _cpd and dt.ssn_column:
                _mh = (_cpd.get("member_header") or "").lower()
                _keys = list((_cpd.get("scores") or {}).keys())[:200]
                _hdr_ssn = "ssn" in _mh or "social" in _mh
                _val_ssn = bool(_keys) and sum(
                    1 for k in _keys
                    if re.fullmatch(r"\d{9}", str(k).strip())) >= max(1, len(_keys) // 2)
                if _hdr_ssn or _val_ssn:
                    _cpd["ssn_keyed"] = True
                    _cpd["loan_join_column"] = dt.ssn_column
                    _cpd["member_column"] = _cpd.get("member_header") or "SSN"
                    result["credit_pull"] = _cpd
        finally:
            wbf.close()
    return result
