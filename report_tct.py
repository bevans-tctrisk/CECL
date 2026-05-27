"""
TCT Report Builder – Produces a multi-sheet Excel workbook matching
the production TCT Risk Solutions CECL-Migration-WARM format.

Sheet map (matches the reference workbook):
  Cover
  Introduction
  Executive Summary  /  Executive Summary (2)  /  Executive Summary (3)
  Risk Change by Credit Score
  Improved Deteriorated Summary
  Historical Trends Balance
  Risk ChangeType 01..N  /  Risk ChangeType Total Loans
  Env Factor by Pool
  ACL Env by Pool Mgmt Adj
  Pool_Balance Adjust
  Envir Fact Ranges
  Grade Ranges & Loan Codes
"""
import os, re, math
from datetime import datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers, NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    PieChart, BarChart, LineChart, DoughnutChart, Reference,
)
from openpyxl.chart.series import DataPoint
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.legend import Legend as ChartLegend
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.text import (
    Paragraph as DrawParagraph, ParagraphProperties, CharacterProperties,
    RegularTextRun, Font as DrawingFont, RichTextProperties,
)
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break

from cecl_engine import risk_change_matrix

# ══════════════════════════════════════════════════════════════════
# CONSTANTS & STYLING
# ══════════════════════════════════════════════════════════════════

HIDDEN_GRADES = ['Hide-E', 'Hide-F', 'Hide-G', 'Hide-H', 'Hide-I']

# ── Arial fonts (headers, labels, matrices) ──────────────────────
FNT_A35B  = Font(name='Arial', bold=True, size=35)
FNT_A28B  = Font(name='Arial', bold=True, size=28)
FNT_A26B  = Font(name='Arial', bold=True, size=26)
FNT_A18   = Font(name='Arial', size=18)
FNT_A18B  = Font(name='Arial', bold=True, size=18)
FNT_A16B  = Font(name='Arial', bold=True, size=16)
FNT_A14B  = Font(name='Arial', bold=True, size=14)
FNT_A12B  = Font(name='Arial', bold=True, size=12)
FNT_A12   = Font(name='Arial', size=12)
FNT_A11   = Font(name='Arial', size=11)
FNT_A10B  = Font(name='Arial', bold=True, size=10)
FNT_A10   = Font(name='Arial', size=10)
FNT_A9    = Font(name='Arial', size=9)

# ── Red fonts for hidden grades ──────────────────────────────────
FNT_A12R  = Font(name='Arial', size=12, color='FF0000')
FNT_A12BR = Font(name='Arial', bold=True, size=12, color='FF0000')
FNT_A10BR = Font(name='Arial', bold=True, size=10, color='FF0000')
FNT_A10R  = Font(name='Arial', size=10, color='FF0000')

# ── Times New Roman fonts (body paragraphs, ACL sheet) ───────────
FNT_T14B  = Font(name='Times New Roman', bold=True, size=14)
FNT_T12B  = Font(name='Times New Roman', bold=True, size=12)
FNT_T12   = Font(name='Times New Roman', size=12)
FNT_T11   = Font(name='Times New Roman', size=11)
FNT_T10B  = Font(name='Times New Roman', bold=True, size=10)
FNT_T10   = Font(name='Times New Roman', size=10)
FNT_T8    = Font(name='Times New Roman', size=8)

# ── Calibri exception (Economic Stress table in Env Factor) ──────
FNT_C11   = Font(name='Calibri', size=11)

# ── Fills ─────────────────────────────────────────────────────────
FILL_IMP  = PatternFill('solid', fgColor='2E7D32')   # medium dark green – improved
FILL_DET  = PatternFill('solid', fgColor='C62828')   # medium dark red – deteriorated

# ── Risk Change Modern Table Styling ─────────────────────────────
RC_HDR_FILL = PatternFill('solid', fgColor='2C3E50')   # dark slate header
RC_TOT_FILL = PatternFill('solid', fgColor='34495E')   # total row slate
RC_HDR_FNT  = Font(name='Arial', bold=True, size=11, color='FFFFFF')
RC_HDR_FNT2 = Font(name='Arial', bold=True, size=10, color='FFFFFF')
RC_TOT_FNT  = Font(name='Arial', bold=True, size=11, color='FFFFFF')
FNT_A12W    = Font(name='Arial', size=12, color='FFFFFF')
FNT_A12BW   = Font(name='Arial', bold=True, size=12, color='FFFFFF')
RC_BDR = Border(
    left=Side('thin', color='BDC3C7'), right=Side('thin', color='BDC3C7'),
    top=Side('thin', color='BDC3C7'), bottom=Side('thin', color='BDC3C7'),
)

# ── Border ────────────────────────────────────────────────────────
THIN = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin'),
)

# ── Number Formats ────────────────────────────────────────────────
ACCT_FMT     = '_("$"* #,##0_);_("$"* \\(#,##0\\);_("$"* "-"??_);_(@_)'
DOLLAR_FMT   = '"$"#,##0'
COMMA_PARENS = '#,##0_);(#,##0)'
NET_CHG_FMT  = '"$"#,##0_);\\("$"#,##0\\)'
PCT          = '0.00%'
PCT4         = '0.0000%'
PCT0         = '0%'

# ── Environmental Factor Score Tables ─────────────────────────────
NCC_RANGES = [
    (-999, -18.00, 7), (-18.00, -16.00, 6), (-16.00, -14.00, 5),
    (-14.00, -11.00, 4), (-11.00, -8.00, 3), (-8.00, -6.00, 2),
    (-6.00, -4.00, 1), (-4.00, 4.00, 0), (4.00, 6.00, -1),
    (6.00, 8.00, -2), (8.00, 9.00, -3), (9.00, 11.00, -4),
    (11.00, 13.00, -5), (13.00, 15.00, -6), (15.00, 999, -7),
]
DQ_RANGES = [
    (5.00, 999, 20), (4.00, 5.00, 17), (3.00, 4.00, 12),
    (2.50, 3.00, 8), (2.00, 2.50, 4), (1.50, 2.00, 2.5),
    (1.00, 1.50, 1.5), (0.50, 1.00, 0.75), (-0.50, 0.50, 0),
    (-1.00, -0.50, -0.75), (-1.50, -1.00, -1.5), (-2.00, -1.50, -2.5),
    (-2.50, -2.00, -4), (-3.00, -2.50, -8), (-4.00, -3.00, -12),
    (-5.00, -4.00, -17), (-999, -5.00, -20),
]
ES_RANGES = [
    (25.00, 999, 10), (24.00, 25.00, 8), (22.00, 24.00, 7),
    (20.00, 22.00, 6), (18.00, 20.00, 5), (16.00, 18.00, 4),
    (14.00, 16.00, 3.5), (12.00, 14.00, 3), (10.00, 12.00, 2),
    (8.00, 10.00, 1), (6.00, 8.00, 0), (4.00, 6.00, 0),
    (2.00, 4.00, -1), (0.00, 2.00, -2),
]
DIST_FACTORS = [10.52, 22.93, 45.15, 116.10, 141.17, 152.04, 160.21]

CHART_COLORS = [
    '4472C4', 'ED7D31', 'A5A5A5', 'FFC000', '5B9BD5', '70AD47',
    '264478', '9B57A0', '636363', '255E91', 'BF8F00',
]


# ══════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════

def _all_grades(grades, no_score):
    """Return complete grade list: visible + hidden + no_score (11 grades)."""
    visible = {g['label'] for g in grades}
    hidden = [h for h in HIDDEN_GRADES if h.replace('Hide-', '') not in visible]
    n_hidden = max(0, len(HIDDEN_GRADES) - (len(grades) - 5))
    return [g['label'] for g in grades] + hidden[:n_hidden] + [no_score]


def _is_hidden(grade_label):
    return grade_label in HIDDEN_GRADES


def _grade_ranges(grades, no_score):
    """Return dict of grade_label -> score range string."""
    ranges = {}
    for g in grades:
        if g['max_score'] >= 900:
            ranges[g['label']] = f"{g['min_score']}+"
        elif g['min_score'] <= 1:
            ranges[g['label']] = f"{g['max_score']} or less"
        else:
            ranges[g['label']] = f"{g['min_score']}-{g['max_score']}"
    for h in HIDDEN_GRADES:
        ranges[h] = 'N/A'
    ranges[no_score] = ''
    return ranges


def _matrix_val(matrix, current, original):
    """Get matrix value, returning 0 for hidden grades or missing entries."""
    if _is_hidden(current) or _is_hidden(original):
        return 0
    if current in matrix.index and original in matrix.columns:
        return matrix.loc[current, original]
    return 0


def _score(value, ranges):
    v = value * 100 if abs(value) < 1 else value
    for lo, hi, s in ranges:
        if lo <= v < hi:
            return s
    return 0


def _dist_factor(idx):
    return DIST_FACTORS[min(idx, len(DIST_FACTORS) - 1)] / 100.0


def _pool_life_loss(pools, hist):
    """Compute life loss rate per pool from historical data."""
    co = hist.get('chargeoffs', {}) if hist else {}
    rc = hist.get('recoveries', {}) if hist else {}
    ab = hist.get('avg_balances', {}) if hist else {}
    years = hist.get('years', []) if hist else []
    result = {}
    for pool in pools:
        rates = []
        for y in years:
            net = co.get(y, {}).get(pool, 0) - rc.get(y, {}).get(pool, 0)
            avg = ab.get(y, {}).get(pool, 0)
            if avg > 0:
                rates.append(net / avg)
        result[pool] = sum(rates) / len(rates) if rates else 0
    return result


def _sort_pools(pools, config):
    """Sort pools by config pool_order if available, else alphabetically.
    Non-risk-rated pools are always placed at the end."""
    nrr = set(config.get('not_risk_rated', []))
    order_list = config.get('pool_order', [])
    if order_list:
        order_map = {name: i for i, name in enumerate(order_list)}
        fallback = len(order_list)
        rr = sorted((p for p in pools if p not in nrr), key=lambda p: order_map.get(p, fallback))
        nr = sorted((p for p in pools if p in nrr), key=lambda p: order_map.get(p, fallback))
    else:
        rr = sorted(p for p in pools if p not in nrr)
        nr = sorted(p for p in pools if p in nrr)
    return rr + nr


def _merge_pool_orders(pools, warm_order, extra=None):
    """Order ``pools`` by ``warm_order`` first, then append any remaining
    pools (e.g. non-risk-rated pools that have no entry in the WARM
    template) in their original order. If ``extra`` iterables are passed,
    pools from those are also included so that pools known to the WARM
    file but missing from the loan DataFrame are not silently dropped.
    Never drops a pool from any input.
    """
    pools = list(pools)
    extras = []
    if extra:
        for it in extra:
            extras.extend(list(it or []))
    if not warm_order and not extras:
        return pools
    def _ok(p):
        if not p or p in seen:
            return False
        s = str(p).strip()
        if not s or s == 'Exclude':
            return False
        if s.upper().startswith('HIDE'):
            return False
        if s.lower() in ('grand total', 'total', 'excluded'):
            return False
        return True

    seen = set()
    out = []
    for p in (warm_order or []):
        if _ok(p):
            out.append(p)
            seen.add(p)
    for p in pools:
        if _ok(p):
            out.append(p)
            seen.add(p)
    for p in extras:
        if _ok(p):
            out.append(p)
            seen.add(p)
    return out


# ── Admin-default + per-pool management-adjustment resolver ──────────
def _load_admin_default_mgmt_adj():
    """Read the firm-wide default management adjustment from
    ``admin_defaults.yaml`` at the workspace root. Returns ``0.0`` on
    any error (missing file / parse failure / missing key).
    """
    try:
        import yaml as _yaml
        from pathlib import Path as _Path
        p = _Path(__file__).resolve().parent / 'admin_defaults.yaml'
        if not p.exists():
            return 0.0
        data = _yaml.safe_load(p.read_text(encoding='utf-8')) or {}
        return float(data.get('default_mgmt_adj', 0.0) or 0.0)
    except Exception:
        return 0.0


def _build_pool_use_default_map(config):
    """Return ``{pool_name: bool}`` from ``config['pools']`` so the
    engine knows which pools opted in to the admin default mgmt-adj."""
    out = {}
    for p in (config.get('pools') or []):
        name = (p.get('name') or '').strip()
        if name:
            out[name] = bool(p.get('use_default_mgmt_adj'))
    return out


def _resolve_mgmt_adj_grade(pool, grade_label, grade_idx, no_score_label,
                             pool_use_default, mgmt_adj_by_pool,
                             admin_default, prior_mgmt_adj_map,
                             base_rate=None):
    """Resolve the per-(pool, grade) management adjustment.

    Precedence (highest first):
      1. ``prior_mgmt_adj_map[pool][grade_label]`` — preserves the
         value from the prior period's report so historical reports
         remain stable.
      2. Manual overlay typed on wizard Step 16
         (``mgmt_adj_by_pool[pool]``) multiplied by the per-grade
         distribution factor.
      3. Admin firm-wide default multiplied by the distribution
         factor — applied **only** when ALL of the following hold,
         matching the Migration model's behaviour:
           a. ``pool_use_default[pool]`` is True (user opted in on
              wizard Step 16),
           b. no manual overlay is set for the pool,
           c. the pool's ACL base loss rate calculation is zero
              (``base_rate == 0`` — i.e. there is no historical loss
              data to drive the rate, so the firm-wide default fills
              the gap).
      4. ``0.0`` otherwise.

    A pool-level value (manual or default) gets distributed across
    grades using the same ``DIST_FACTORS`` table that the manual
    overlay path has always used, so the default and manual paths
    behave identically per the user's spec.
    """
    pm = prior_mgmt_adj_map.get(pool, {}) if prior_mgmt_adj_map else {}
    if grade_label in pm:
        return pm[grade_label]
    dist = (_dist_factor(len(DIST_FACTORS) - 1)
            if grade_label == no_score_label
            else _dist_factor(grade_idx))
    manual = mgmt_adj_by_pool.get(pool, 0) or 0
    if manual:
        return float(manual) * dist
    if (pool_use_default.get(pool, False)
            and admin_default
            and (base_rate is None or float(base_rate or 0) == 0)):
        return float(admin_default) * dist
    return 0.0


def _resolve_mgmt_adj_total(pool, pool_use_default, mgmt_adj_by_pool,
                             admin_default, base_rate=None):
    """Pool-total resolver for non-risk-rated pools (no per-grade row).

    Same precedence as :func:`_resolve_mgmt_adj_grade` minus the
    grade-distribution step (NRR pools have no grades). The admin
    default is gated on ``base_rate == 0`` per the Migration-model
    rule.
    """
    manual = mgmt_adj_by_pool.get(pool, 0) or 0
    if manual:
        return float(manual)
    if (pool_use_default.get(pool, False)
            and admin_default
            and (base_rate is None or float(base_rate or 0) == 0)):
        return float(admin_default)
    return 0.0


def _snap_display(snap):
    """Format snap date for display like '12/31/2025'."""
    try:
        dt = pd.to_datetime(snap)
        return dt.strftime('%#m/%#d/%Y') if os.name == 'nt' else dt.strftime('%-m/%-d/%Y')
    except Exception:
        return snap


def _set_margins(ws, t=0.25, b=0.25, l=0.25, r=0.25, header=0.3, footer=0.05):
    ws.page_margins = PageMargins(
        top=t, bottom=b, left=l, right=r, header=header, footer=footer,
    )


def _landscape(ws):
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    _set_margins(ws)


def _portrait(ws):
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    _set_margins(ws)


def _footer_cells(ws, cu, snap, row=None):
    """Write footer in cells (not Excel header/footer) – matches reference."""
    yr = datetime.now().year
    if row is None:
        row = ws.max_row + 3
    ws.cell(row=row, column=2,
            value=f"\u00a9 {yr} TCT Risk Solutions").font = FNT_T8
    ws.cell(row=row, column=5,
            value=f"For Period Ending {_snap_display(snap)}").font = FNT_T8
    ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=row + 1, column=2, value="P.O. Box 2210").font = FNT_T8
    ws.cell(row=row + 1, column=5, value=cu).font = FNT_T8
    ws.cell(row=row + 1, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=row + 2, column=2, value="Eagle, ID 83616").font = FNT_T8
    ws.cell(row=row + 3, column=2,
            value="Voice (208) 939-8366 - Fax (208) 938-6276").font = FNT_T8
    ws.cell(row=row + 4, column=2,
            value="E-Mail: RThompson@tctrisk.com or Office@tctrisk.com").font = FNT_T8


def _footer_copyright(ws, cu, snap, row):
    """Simple copyright + confidential line for body sheets."""
    yr = datetime.now().year
    ws.cell(row=row, column=2,
            value=f"\u00a9 {yr} TCT Risk Solutions").font = FNT_T8
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
    ws.cell(row=row, column=12, value="Confidential").font = FNT_A10
    ws.cell(row=row, column=12).alignment = Alignment(horizontal='right')


def _intro_col_widths(ws):
    """Set 14-column layout widths matching reference Introduction/Exec sheets."""
    widths = {'A': 3.44, 'B': 3.33, 'C': 15.44, 'D': 10.11, 'E': 8.89,
              'F': 13.0, 'G': 13.0, 'H': 13.0, 'I': 13.0, 'J': 13.0,
              'K': 13.0, 'L': 9.89, 'M': 3.33, 'N': 3.44}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _write_3row_header(ws, cu, subtitle, snap):
    """Standard 3-row header: CU name (B2:M2), subtitle (B3:M3), date (B4:M4)."""
    ws.merge_cells('B2:M2')
    ws.merge_cells('B3:M3')
    ws.merge_cells('B4:M4')
    ws['B2'] = cu
    ws['B2'].font = FNT_A16B
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['B3'] = subtitle
    ws['B3'].font = FNT_A12B
    ws['B3'].alignment = Alignment(horizontal='center')
    ws['B4'] = f"For Period Ending {_snap_display(snap)}"
    ws['B4'].font = FNT_A12
    ws['B4'].alignment = Alignment(horizontal='center')


def _ncc(df, grades, config):
    """Net credit change from migration matrix (matches Risk Change sheets).
    Returns (improved_pct, deteriorated_pct, net_pct)."""
    total = df['current_balance'].sum()
    if total == 0:
        return 0, 0, 0
    no_score = config.get('no_score_label', 'Not Reported')
    gl = _all_grades(grades, no_score)
    matrix = risk_change_matrix(df, grades, no_score)
    n_top = config.get('top_grades_double_drop', 3)
    imp_bal = 0
    det_bal = 0
    for j, og in enumerate(gl):
        for i, g in enumerate(gl):
            v = _matrix_val(matrix, g, og)
            if i > j:
                if j < n_top and (i - j) < 2:
                    pass  # unchanged – small drop within top grades
                else:
                    det_bal += v
            elif i < j:
                imp_bal += v
    return imp_bal / total, det_bal / total, (imp_bal - det_bal) / total


def _econ_stress(config):
    ed = config.get('economic_data', {})
    unemp = ed.get('unemployment_rate', 0) * 100
    pop = ed.get('population', 1)
    bk = (ed.get('bankruptcies', 0) / pop) * 100 if pop else 0
    fc = (ed.get('foreclosures', 0) / pop) * 100 if pop else 0
    return unemp + bk + fc


# ══════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ══════════════════════════════════════════════════════════════════

# ── Cover ─────────────────────────────────────────────────────────
def _sheet_cover(wb, cu, snap):
    ws = wb.active
    ws.title = "Cover"
    _portrait(ws)
    _set_margins(ws, t=0.25, b=0.25, l=0.2, r=0.2, header=0.05, footer=0.05)

    # Column widths
    for col, w in {'A': 2.66, 'B': 8.66, 'C': 20.89,
                   'D': 51.33, 'E': 8.66, 'F': 2.66}.items():
        ws.column_dimensions[col].width = w

    # Row heights
    for rn, h in {5: 43.8, 6: 35.4, 7: 22.8, 8: 22.8, 9: 22.8,
                  10: 22.8, 11: 22.8, 12: 22.8, 13: 35.25,
                  15: 45.0, 16: 45.0, 17: 21.0, 18: 24.0}.items():
        ws.row_dimensions[rn].height = h

    center = Alignment(horizontal='center')
    center_v = Alignment(horizontal='center', vertical='center')
    center_vw = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Merge A:F for all content rows
    for rn in [5, 6, 7, 8, 9, 10, 11, 12, 14, 15, 16, 17, 18, 19, 20]:
        ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=6)

    # Title block
    ws['A5'] = "RISK BASED PRICING"
    ws['A5'].font = FNT_A35B
    ws['A5'].alignment = center

    ws['A6'] = "ACL/Credit Migration Report"
    ws['A6'].font = FNT_A28B
    ws['A6'].alignment = center

    items = [
        "CECL Compliant", "Risk Change by Type",
        "Improved/Deteriorated Loan Analysis", "Environmental Factor",
        "Allowance for Credit Loss (ACL)", "Summary of Deteriorated Loans",
    ]
    for i, item in enumerate(items):
        cell = ws.cell(row=7 + i, column=1, value=item)
        cell.font = FNT_A18
        cell.alignment = center

    # Prepared For
    ws['A14'] = " Prepared For:"
    ws['A14'].font = FNT_A10
    ws['A14'].alignment = center_v

    ws['A15'] = cu
    ws['A15'].font = FNT_A26B
    ws['A15'].alignment = center_vw

    # Period
    ws['A17'] = "For Period Ending"
    ws['A17'].font = FNT_A14B
    ws['A17'].alignment = center_v

    try:
        dt = pd.to_datetime(snap)
        ws['A18'] = dt
        ws['A18'].number_format = 'mm-dd-yy'
    except Exception:
        ws['A18'] = snap
    ws['A18'].font = FNT_A14B
    ws['A18'].alignment = center_v

    # Presented by
    ws['A20'] = "Presented by:"
    ws['A20'].font = FNT_A10
    ws['A20'].alignment = center

    # Logo – centered in white space between content and footer
    logo_path = os.path.join(os.path.dirname(__file__), 'logos', 'tct_risk_solutions.png')
    if os.path.exists(logo_path):
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.utils.units import pixels_to_EMU
        logo = XlImage(logo_path)
        logo_w_px = 480   # ~5 inches at 96 DPI
        logo_h_px = int(logo_w_px * 690 / 3508)  # maintain aspect ratio
        logo.width = logo_w_px
        logo.height = logo_h_px
        # Center horizontally: total col widths A-F ≈ 95 char units ≈ 685 px
        # Column A=2.66, B=8.66 → first 11.32 chars ≈ 82 px before col C
        # Offset into col C to center: (685 - 480) / 2 - 82 ≈ 20 px
        col_c_offset_emu = pixels_to_EMU(20)
        anc = AnchorMarker(col=2, colOff=col_c_offset_emu, row=23, rowOff=0)
        logo.anchor = OneCellAnchor(_from=anc, ext=None)
        logo.anchor.ext.cx = pixels_to_EMU(logo_w_px)
        logo.anchor.ext.cy = pixels_to_EMU(logo_h_px)
        ws.add_image(logo)

    # Footer block in cells (row 36+)
    _footer_cells(ws, cu, snap, row=36)


# ── Introduction ──────────────────────────────────────────────────
def _sheet_intro(wb, cu, snap):
    ws = wb.create_sheet("Introduction")
    _portrait(ws)
    _set_margins(ws, header=0.3, footer=0.05)
    _intro_col_widths(ws)

    _write_3row_header(ws, cu, "ACL/Credit Migration Report", snap)

    # Section header
    ws.merge_cells('B8:M8')
    ws['B8'] = "Introduction and Overview"
    ws['B8'].font = FNT_A12B
    ws['B8'].alignment = Alignment(horizontal='center')

    body_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    paragraphs = [
        ("Credit Migration may be defined as a measurement of changes in credit scores and risk "
         "for individual loans in the loan portfolio of the credit union. The composite of these "
         "changes provides a valid measure of the current risk inherent in the total loan portfolio."),
        ("Migration is measured by the improvement or deterioration of risk, measured by the credit "
         "score, from the date of loan funding to the most recent data pull. New credit scores are "
         "typically pulled twice per year. Migration may still be measured on a quarterly basis to "
         "take into account new loans and changing loan balances."),
        ("Over the life of a loan borrowers\u2019 credit risk rating can change.  Borrowers, who "
         "experience financial changes such as disruptions of income, or severe unexpected expenses, "
         "will see their ability to pay decrease which will reduce their capacity to make payments. "
         "On the reverse side, borrowers may experience improved income or a reduction in expenses "
         "which will increase their ability to service their debts. These changes can have an impact "
         "on the quality of your loan portfolios. The net effect of improved and impaired credit is "
         "a direct and valid measure of the risk and hence the quality of the credit union\u2019s loans "
         "portfolio."),
        ("Credit agencies continually monitor multiple risk indicators to calculate credit scores "
         "including; payment history, amount of credit, available credit, employment history, "
         "repossessions, bankruptcies, foreclosures and others. Each of these variables is dynamic "
         "and may change at any time. Changes in variables may impact credit worthiness resulting in "
         "a change credit score. Changes in the credit score is a key indicator changes in the risk "
         "associated with the loan. Changes in risk may affect member performance, either positively "
         "or negatively, based on the either improvement or impairment of the credit score."),
        ("From the beginning of the movement, credit unions have been a primary source of financial "
         "help and support to members from all credit ranges. Providing credit to members who need "
         "help is an important reason that credit unions exist. Understanding the impacts of credit "
         "changes through the measurement and analysis of changing credit scores, both deteriorated "
         "and improved, will increase a credit unions ability to help all of its members."),
    ]
    para_rows = [9, 11, 13, 15, 17]
    para_heights = [60.75, 57.75, 104.25, 106.5, 81.75]
    for idx, (para, rn, rh) in enumerate(zip(paragraphs, para_rows, para_heights)):
        ws.merge_cells(start_row=rn, start_column=3, end_row=rn, end_column=12)
        cell = ws.cell(row=rn, column=3, value=para)
        cell.font = FNT_T12
        cell.alignment = body_align
        ws.row_dimensions[rn].height = rh

    _footer_copyright(ws, cu, snap, row=31)


# ── Executive Summary (3 sheets) ─────────────────────────────────
def _sheet_exec_summary(wb, cu, snap, df, grades, config):
    no_score = config.get('no_score_label', 'Not Reported')
    gl = _all_grades(grades, no_score)

    total = df['current_balance'].sum()
    imp_pct, det_pct, ncc_pct = _ncc(df, grades, config)
    imp_bal = imp_pct * total
    det_bal = det_pct * total

    # Composite Migration Impact calculation
    nrr = set(config.get('not_risk_rated', []))
    pools = sorted(df['loan_pool'].unique())
    pool_impacts = []
    for pool in pools:
        pdf = df[df['loan_pool'] == pool]
        pt = pdf['current_balance'].sum()
        if pt == 0:
            continue
        if pool in nrr:
            pool_net = 0.0
        else:
            _, _, pool_net = _ncc(pdf, grades, config)
        pool_impacts.append(abs(pool_net))
    composite_impact = sum(pool_impacts) / len(pool_impacts) if pool_impacts else 0

    if composite_impact < 0.15:
        impact_status = "No Significant Risk Change"
    elif composite_impact < 0.35:
        impact_status = "Moderate Risk Change"
    else:
        impact_status = "Significant Risk Change"

    body_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ═══ Page 1 ═══
    ws = wb.create_sheet("Executive Summary")
    _portrait(ws)
    _set_margins(ws, header=0.3, footer=0.05)
    _intro_col_widths(ws)
    # Extra cols for Composite Migration Impact
    for col, w in {'O': 8.89, 'P': 13.0, 'Q': 13.0, 'R': 13.0,
                   'S': 25.89, 'T': 8.89, 'U': 13.0}.items():
        ws.column_dimensions[col].width = w

    _write_3row_header(ws, cu, "Executive Summary", snap)

    ws.merge_cells('B8:M8')
    ws['B8'] = " Executive Overview"
    ws['B8'].font = FNT_A12B
    ws['B8'].alignment = Alignment(horizontal='center')

    # Overview paragraph
    ws.merge_cells('C9:L9')
    ws['C9'] = (
        "The Credit Migration Summary from TCT, Inc. presents a comprehensive picture of the "
        "changing nature of risk in the credit union\u2019s loan portfolio. Credit migration is measured "
        "by the improvement or deterioration of risk, measured by the credit score, from the date "
        "of loan funding to the most recent data pull. New credit scores are typically pulled twice "
        "per year. Migration may still be measured on a quarterly basis to take into account new "
        "loans and changing loan balances."
    )
    ws['C9'].font = FNT_T12
    ws['C9'].alignment = body_align
    ws.row_dimensions[9].height = 87.0

    # Reports summary
    ws.merge_cells('C11:L11')
    ws['C11'] = (
        "A set of seven reports examines the changing nature of the risk inherent in the credit "
        "union\u2019s loan portfolio. The reports are:"
    )
    ws['C11'].font = FNT_T12
    ws['C11'].alignment = body_align
    ws.row_dimensions[11].height = 35.25

    # Section 1
    ws['C13'] = "Section 1 Credit Migration"
    ws['C13'].font = FNT_A12B
    ws.merge_cells('C14:L14')
    ws['C14'] = "Credit change matrix\nNet Credit Change\nPool and Grade Tracking Report\nDelinquency Breakout"
    ws['C14'].font = FNT_T12
    ws['C14'].alignment = body_align
    ws.row_dimensions[14].height = 69.75

    # Section 2
    ws['C16'] = "Section 2 Improved/Impaired Loan Listings"
    ws['C16'].font = FNT_A12B
    ws['C16'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('C17:L17')
    ws['C17'] = "Improved Loans Breakout\nDeteriorated Loans Breakout"
    ws['C17'].font = FNT_T12
    ws['C17'].alignment = body_align
    ws.row_dimensions[17].height = 34.5

    # Section 3
    ws['C19'] = "Section 3 Allowance for Credit Loss Calculation"
    ws['C19'].font = FNT_A12B
    ws['C19'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('C20:L20')
    ws['C20'] = "Risk Based Allowance for Credit Loss Report"
    ws['C20'].font = FNT_T12
    ws['C20'].alignment = body_align
    ws.row_dimensions[20].height = 20.25

    # Credit Change Matrix section
    ws.merge_cells('B22:M22')
    ws['B22'] = "Credit Change Matrix"
    ws['B22'].font = FNT_A12B
    ws['B22'].alignment = Alignment(horizontal='center')

    ws.merge_cells('C23:L23')
    ws['C23'] = (
        "This report shows the disposition of loans, by grade, resulting from credit change. "
        "Loans are grouped in columns based on the original credit score at loan inception, or "
        "date of funding. Each range, or column, of loans are then divided into rows that show "
        "the dollars in each credit range based on the most recent credit score."
    )
    ws['C23'].font = FNT_T12
    ws['C23'].alignment = body_align
    ws.row_dimensions[23].height = 57.0

    ws.merge_cells('C25:L25')
    ws['C25'] = (
        "Loan balances shown in light red cells have an deteriorated credit score.  Loan balances "
        "shown in white cells have an unchanged credit score.  Loan balances shown in green cells "
        "have an improved credit score."
    )
    ws['C25'].font = FNT_T12
    ws['C25'].alignment = body_align
    ws.row_dimensions[25].height = 36.0

    ws.merge_cells('C27:L27')
    ws['C27'] = (
        "The report also provides a table at the bottom of the page showing the percentage of "
        "loans in each column (representing the original credit score) that is in each range as "
        "measured with the current credit score."
    )
    ws['C27'].font = FNT_T12
    ws['C27'].alignment = body_align
    ws.row_dimensions[27].height = 36.0

    # Composite Migration Impact Index
    ws.merge_cells('B29:M29')
    ws['B29'] = "Composite Migration Impact Index (by Initial Grade)"
    ws['B29'].font = FNT_A12B
    ws['B29'].alignment = Alignment(horizontal='center')

    ws.merge_cells('C30:L30')
    ws['C30'] = (
        f"With a Composite Migration Impact Index of {composite_impact:.2%} {cu} has {impact_status}. "
        f"The composite Migration impact measures the level of credit change."
    )
    ws['C30'].font = FNT_A10
    ws['C30'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[30].height = 28.5

    # Score thresholds in cols S-V
    ws['S30'] = f"With a composite Migration Impact of "
    ws['S30'].font = FNT_A10
    ws['T30'] = composite_impact
    ws['T30'].font = FNT_A10
    ws['T30'].number_format = PCT
    ws['U30'] = ", "
    ws['U30'].font = FNT_A10

    ws['S31'] = "No Significant Risk Change "
    ws['S31'].font = FNT_T12
    ws['T31'] = 0
    ws['T31'].font = FNT_A10
    ws['S32'] = "Moderate Risk Change"
    ws['S32'].font = FNT_T12
    ws['T32'] = 0.15
    ws['T32'].font = FNT_A10
    ws['T32'].number_format = PCT0
    ws['S33'] = "Significant Risk Change"
    ws['S33'].font = FNT_T12
    ws['T33'] = 0.35
    ws['T33'].font = FNT_A10
    ws['T33'].number_format = PCT0

    _footer_copyright(ws, cu, snap, row=36)
    ws.print_area = 'A1:N36'

    # ═══ Page 2 - Executive Summary (2) ═══
    ws2 = wb.create_sheet("Executive Summary (2)")
    _portrait(ws2)
    _set_margins(ws2, header=0.3, footer=0.05)
    _intro_col_widths(ws2)
    for col, w in {'O': 8.89, 'P': 13.0, 'Q': 13.0, 'R': 13.0,
                   'S': 13.0, 'T': 25.89, 'U': 8.89, 'V': 13.0}.items():
        ws2.column_dimensions[col].width = w

    _write_3row_header(ws2, cu, "Executive Summary (Continued)", snap)

    # Net Credit Change section
    ws2.merge_cells('B8:M8')
    ws2['B8'] = "Net Credit Change"
    ws2['B8'].font = FNT_A12B
    ws2['B8'].alignment = Alignment(horizontal='center')

    ws2.merge_cells('C9:L9')
    ncc_pct_display = ncc_pct * 100
    ncc_score_val = _score(ncc_pct_display, NCC_RANGES)
    ws2['C9'] = (
        f"The Net Credit change score presents a quantitative measure of the composite change "
        f"in risk in the loan portfolio. Loan balances with impaired and improved credit scores "
        f"are extracted from the Credit Change Matrix to create this report."
    )
    ws2['C9'].font = FNT_T12
    ws2['C9'].alignment = body_align
    ws2.row_dimensions[9].height = 54.0

    ws2.merge_cells('C11:L11')
    ws2['C11'] = (
        f"Balances of deteriorated and improved loans are divided by the total loan balance to "
        f"calculate the percent of portfolio that are impaired and improved. Deteriorated loan "
        f"percentage is then subtracted from the improved loan percentage to arrive at a Net "
        f"Credit Change score. A positive Net Credit Change score indicated more improving "
        f"balances than deteriorating scores."
    )
    ws2['C11'].font = FNT_T12
    ws2['C11'].alignment = body_align
    ws2.row_dimensions[11].height = 85.5

    ws2.merge_cells('C13:L13')
    ws2['C13'] = (
        f"The Net Credit Change for {cu} is {ncc_pct:.2%} for the period ending "
        f"{_snap_display(snap)}. This results in a Net Credit Score Index of "
        f"{ncc_score_val:.2f}%."
    )
    ws2['C13'].font = FNT_T12
    ws2['C13'].alignment = body_align
    ws2.row_dimensions[13].height = 36.0

    ws2.merge_cells('C15:L15')
    ws2['C15'] = (
        "The Net Credit Score Index is one of the three components employed to calculate the "
        "Environmental Factor."
    )
    ws2['C15'].font = FNT_T12
    ws2['C15'].alignment = body_align
    ws2.row_dimensions[15].height = 52.5

    # Pool and Grade Tracking Report
    ws2.merge_cells('B17:M17')
    ws2['B17'] = "Pool and Grade Tracking Report"
    ws2['B17'].font = FNT_A12B
    ws2['B17'].alignment = Alignment(horizontal='center')

    ws2.merge_cells('C18:L18')
    ws2['C18'] = (
        "Concentrations of loans in pools and grades are important indicators of risk. The "
        "dynamic nature of credit scores means that grade concentration may change consistently "
        "from quarter to quarter."
    )
    ws2['C18'].font = FNT_T12
    ws2['C18'].alignment = body_align
    ws2.row_dimensions[18].height = 54.0

    # Delinquency Breakout
    ws2.merge_cells('B20:M20')
    ws2['B20'] = "Delinquency Breakout"
    ws2['B20'].font = FNT_A12B
    ws2['B20'].alignment = Alignment(horizontal='center')

    ws2.merge_cells('C21:L21')
    ws2['C21'] = (
        "Loan delinquency is the single most valid predictor of impending loan losses. Because "
        "loan losses exert a major influence on financial viability, monitoring delinquency is an "
        "important aspect of loan management. By controlling delinquency management may also "
        "control losses."
    )
    ws2['C21'].font = FNT_T12
    ws2['C21'].alignment = body_align
    ws2.row_dimensions[21].height = 57.0

    # Net Credit Change score/status table in cols Q-V
    ws2['Q17'] = "Net Credit Score"
    ws2['Q17'].font = FNT_A10B
    ws2['R17'] = ncc_pct
    ws2['R17'].font = FNT_A10
    ws2['R17'].number_format = PCT

    ws2['T18'] = "Status"
    ws2['T18'].font = FNT_A10B
    ws2['T19'] = impact_status
    ws2['T19'].font = FNT_T12

    _footer_copyright(ws2, cu, snap, row=36)
    ws2.print_area = 'A1:N36'

    # ═══ Page 3 - Executive Summary (3) ═══
    ws3 = wb.create_sheet("Executive Summary (3)")
    _portrait(ws3)
    _set_margins(ws3, header=0.3, footer=0.05)
    _intro_col_widths(ws3)
    # Widen column D so dollar amounts are readable when printed
    ws3.column_dimensions['D'].width = 18.78

    _write_3row_header(ws3, cu, "Executive Summary (Continued)", snap)

    snap_disp = _snap_display(snap)

    # Compute grade-level improved/deteriorated from migration matrix
    visible_grades = [g['label'] for g in grades] + [no_score]
    matrix = risk_change_matrix(df, grades, no_score)
    n_top = config.get('top_grades_double_drop', 3)
    grade_imp = {}
    grade_det = {}
    for g in visible_grades:
        grade_imp[g] = 0
        grade_det[g] = 0
    for j, og in enumerate(gl):
        for i, cg in enumerate(gl):
            v = _matrix_val(matrix, cg, og)
            if i > j:
                if j < n_top and (i - j) < 2:
                    pass
                else:
                    if og in grade_det:
                        grade_det[og] += v
            elif i < j:
                if og in grade_imp:
                    grade_imp[og] += v

    # ── Improved Loans Breakout ──
    ws3.merge_cells('B8:N8')
    ws3['B8'] = "Improved Loans Breakout"
    ws3['B8'].font = FNT_A12B
    ws3['B8'].alignment = Alignment(horizontal='center')

    ws3.merge_cells('C9:N9')
    ws3.cell(row=9, column=3, value=(
        "Individual loans with improved credit scores are listed in this report "
        "by account number.  This list provides an excellent source for targeted "
        "marketing to expand use of credit union products and services."
    )).font = FNT_T12
    ws3.cell(row=9, column=3).alignment = body_align
    ws3.row_dimensions[9].height = 37.5

    r = 11
    ws3.merge_cells(f'C{r}:N{r}')
    ws3.cell(row=r, column=3, value=f"Improved Loans Summary as of {snap_disp}").font = FNT_A10B
    ws3.cell(row=r, column=3).alignment = Alignment(wrap_text=True)
    r += 1
    ws3.cell(row=r, column=3, value="Grade").font = FNT_A10B
    ws3.cell(row=r, column=4, value="Balance").font = FNT_A10B
    r += 1

    # Exclude top grade (first visible grade) and Not Reported row
    summary_grades = [g for g in visible_grades[1:] if g != no_score]
    for g in summary_grades:
        bal = grade_imp.get(g, 0)
        ws3.cell(row=r, column=3, value=g).font = FNT_A10
        ws3.cell(row=r, column=4, value=bal).font = FNT_A10
        ws3.cell(row=r, column=4).number_format = ACCT_FMT
        r += 1
    ws3.cell(row=r, column=3, value="Total Improved").font = FNT_A10B
    ws3.cell(row=r, column=4, value=imp_bal).font = FNT_A10B
    ws3.cell(row=r, column=4).number_format = ACCT_FMT
    r += 2

    # ── Deteriorated Loans Breakout ──
    ws3.merge_cells(f'B{r}:N{r}')
    ws3.cell(row=r, column=2, value="Deteriorated Loans Breakout").font = FNT_A12B
    ws3.cell(row=r, column=2).alignment = Alignment(horizontal='center')
    r += 1

    ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=14)
    ws3.cell(row=r, column=3, value=(
        "Individual loans with deteriorated credit scores are listed in this report "
        "by account number. For unsecured loans a cell is provided for each loan to "
        "insert the current credit limit attached to each loan. For secured loans "
        "cells are provided to insert the value of collateral.  The report then "
        "automatically calculates LTV."
    )).font = FNT_T12
    ws3.cell(row=r, column=3).alignment = body_align
    ws3.row_dimensions[r].height = 57.0
    r += 2

    ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=14)
    ws3.cell(row=r, column=3, value=(
        "This report provides a starting point to identify loans that require review "
        "and specific action.  In some cases lines of credit may need to be reduced, "
        "in other lines may require closure.  In all cases these loans require greater "
        "attention."
    )).font = FNT_T12
    ws3.cell(row=r, column=3).alignment = body_align
    ws3.row_dimensions[r].height = 46.5
    r += 2

    ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=14)
    ws3.cell(row=r, column=3, value=(
        "Since loans with deteriorated credit scores are the single greatest source "
        "of delinquency and charge-off early detection of impairment and management "
        "of the line will be an effective strategy to reduce losses."
    )).font = FNT_T12
    ws3.cell(row=r, column=3).alignment = body_align
    ws3.row_dimensions[r].height = 38.25
    r += 2

    ws3.merge_cells(f'C{r}:N{r}')
    ws3.cell(row=r, column=3, value=f"Deteriorated Loans Summary as of {snap_disp}").font = FNT_A10B
    ws3.cell(row=r, column=3).alignment = Alignment(wrap_text=True)
    r += 1
    ws3.cell(row=r, column=3, value="Grade").font = FNT_A10B
    ws3.cell(row=r, column=4, value="Balance").font = FNT_A10B
    r += 1
    # Deteriorated table: include top grade, exclude lowest grade and Not Reported
    grade_only = [g for g in visible_grades if g != no_score]
    det_grades = grade_only[:-1] if len(grade_only) > 1 else grade_only
    for g in det_grades:
        bal = grade_det.get(g, 0)
        ws3.cell(row=r, column=3, value=g).font = FNT_A10
        ws3.cell(row=r, column=4, value=bal).font = FNT_A10
        ws3.cell(row=r, column=4).number_format = ACCT_FMT
        r += 1
    ws3.cell(row=r, column=3, value="Total Impaired").font = FNT_A10B
    ws3.cell(row=r, column=4, value=det_bal).font = FNT_A10B
    ws3.cell(row=r, column=4).number_format = ACCT_FMT
    r += 2

    # ACL description
    ws3.merge_cells(f'B{r}:N{r}')
    ws3.cell(row=r, column=2, value="Risk Based Allowance for Loan Loss Calculation").font = FNT_A12B
    ws3.cell(row=r, column=2).alignment = Alignment(horizontal='center')
    r += 1
    ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=14)
    ws3.cell(row=r, column=3, value=(
        "This report utilizes the outputs from the Credit Migration Summary and Net Credit Change "
        "Matrix to calculate the Allowance for Loan Loss required by the credit union. Included in "
        "the calculation is an empirically calculated Environmental Factor.\n\n"
        "A detailed description of the methodology and individual calculations in the process is "
        "included with the ACL report."
    )).font = FNT_T12
    ws3.cell(row=r, column=3).alignment = body_align
    ws3.row_dimensions[r].height = 47.25
    r += 3

    _footer_copyright(ws3, cu, snap, row=max(r + 2, 36))
    ws3.print_area = f'A1:N{max(r + 2, 36)}'


# ── Risk Change by Credit Score ───────────────────────────────────
def _sheet_risk_change_all(wb, cu, snap, df, grades, config, hist=None):
    ws = wb.create_sheet("Risk Change by Credit Score")
    _landscape(ws)
    no_score = config.get('no_score_label', 'Not Reported')
    all_gl = _all_grades(grades, no_score)
    gl = [g for g in all_gl if not _is_hidden(g)]
    matrix = risk_change_matrix(df, grades, no_score)
    total = df['current_balance'].sum()
    grade_rngs = _grade_ranges(grades, no_score)
    n_grades = len(gl)
    gt_col = 3 + n_grades  # Grand Total column

    # Column widths – fill landscape page
    col_widths = {}
    ncols = gt_col
    col_w = round(226.0 / ncols, 2)
    for idx in range(ncols):
        col_widths[chr(ord('A') + idx)] = col_w
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Row 1: CU name
    last_col_letter = chr(ord('C') + n_grades)
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = cu
    ws['A1'].font = FNT_A14B

    # Row 2: Title
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = "Risk Change By Credit Score"
    ws['A2'].font = FNT_A12B

    # Row 3: Date
    ws.merge_cells(f'A3:{last_col_letter}3')
    ws['A3'] = f"For Period Ending {_snap_display(snap)}"
    ws['A3'].font = FNT_A10B

    # ─── Dollar Matrix ───
    # Row 4: "Dollar" header
    grade_last = chr(ord('C') + n_grades - 1)
    ws.merge_cells('A4:B4')
    ws['A4'] = "Dollar"
    ws['A4'].font = RC_HDR_FNT
    ws['A4'].fill = RC_HDR_FILL
    ws['B4'].fill = RC_HDR_FILL
    ws.merge_cells(f'C4:{grade_last}4')
    ws['C4'] = "Original Credit Grade"
    ws['C4'].font = RC_HDR_FNT
    ws['C4'].fill = RC_HDR_FILL
    ws['C4'].alignment = Alignment(horizontal='center')
    ws.merge_cells(f'{last_col_letter}4:{last_col_letter}5')
    ws[f'{last_col_letter}4'] = "Grand Total"
    ws[f'{last_col_letter}4'].font = RC_HDR_FNT2
    ws[f'{last_col_letter}4'].fill = RC_HDR_FILL
    ws[f'{last_col_letter}4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Row 5: Grade labels
    ws.merge_cells('A5:B5')
    ws['A5'] = "Current Credit Grade"
    ws['A5'].font = RC_HDR_FNT2
    ws['A5'].fill = RC_HDR_FILL
    ws['B5'].fill = RC_HDR_FILL
    for j, g in enumerate(gl):
        cell = ws.cell(row=5, column=3 + j, value=g)
        cell.font = RC_HDR_FNT
        cell.fill = RC_HDR_FILL
        cell.alignment = Alignment(horizontal='center')

    n_top = config.get('top_grades_double_drop', 3)

    # Data rows
    for i, g in enumerate(gl):
        r = 6 + i
        ws.cell(row=r, column=1, value=g).font = FNT_A12BW
        ws.cell(row=r, column=1).fill = RC_HDR_FILL
        ws.cell(row=r, column=1).border = RC_BDR
        ws.cell(row=r, column=2, value=grade_rngs.get(g, '')).font = FNT_A12BW
        ws.cell(row=r, column=2).fill = RC_HDR_FILL
        ws.cell(row=r, column=2).border = RC_BDR

        rtotal = 0
        for j, og in enumerate(gl):
            v = _matrix_val(matrix, g, og)
            cell = ws.cell(row=r, column=3 + j, value=v)
            cell.number_format = ACCT_FMT
            cell.font = FNT_A12
            cell.border = RC_BDR
            rtotal += v
            if g == no_score or og == no_score:
                pass  # Not Reported → always unchanged, no fill
            elif i > j:
                if j < n_top and (i - j) < 2:
                    pass  # unchanged
                else:
                    cell.fill = FILL_DET
                    cell.font = FNT_A12W
            elif i < j:
                cell.fill = FILL_IMP
                cell.font = FNT_A12W
        ws.cell(row=r, column=gt_col, value=rtotal).number_format = ACCT_FMT
        ws.cell(row=r, column=gt_col).font = FNT_A12B
        ws.cell(row=r, column=gt_col).border = RC_BDR

    # Total row
    total_row = 6 + n_grades
    ws.cell(row=total_row, column=1, value="Total").font = RC_TOT_FNT
    ws.cell(row=total_row, column=1).fill = RC_TOT_FILL
    ws.cell(row=total_row, column=1).border = RC_BDR
    ws.cell(row=total_row, column=2).fill = RC_TOT_FILL
    ws.cell(row=total_row, column=2).border = RC_BDR
    grand = 0
    for j, og in enumerate(gl):
        col_total = sum(_matrix_val(matrix, g2, og) for g2 in gl)
        ws.cell(row=total_row, column=3 + j, value=col_total).number_format = ACCT_FMT
        ws.cell(row=total_row, column=3 + j).font = RC_TOT_FNT
        ws.cell(row=total_row, column=3 + j).fill = RC_TOT_FILL
        ws.cell(row=total_row, column=3 + j).border = RC_BDR
        grand += col_total
    ws.cell(row=total_row, column=gt_col, value=grand).number_format = ACCT_FMT
    ws.cell(row=total_row, column=gt_col).font = RC_TOT_FNT
    ws.cell(row=total_row, column=gt_col).fill = RC_TOT_FILL
    ws.cell(row=total_row, column=gt_col).border = RC_BDR

    # Adjustments (from WARM balance adjustment data)
    _imp = hist.get('impaired', {}) if hist else {}
    bal_adj = _imp.get('total_balance_adjustment', 0.0) or 0.0
    adj_row = total_row + 1
    ws.merge_cells(start_row=adj_row, start_column=1, end_row=adj_row, end_column=gt_col - 1)
    ws.cell(row=adj_row, column=1, value="Loans Not Risk Rated and Adjustments").font = RC_TOT_FNT
    ws.cell(row=adj_row, column=gt_col, value=bal_adj).number_format = ACCT_FMT
    ws.cell(row=adj_row, column=gt_col).font = RC_TOT_FNT
    for c in range(1, gt_col + 1):
        ws.cell(row=adj_row, column=c).fill = RC_TOT_FILL
        ws.cell(row=adj_row, column=c).border = RC_BDR

    # Total in Portfolio
    tip_row = adj_row + 1
    total_in_portfolio = grand + bal_adj
    tip = _imp.get('total_in_portfolio', total_in_portfolio)
    if tip:
        total_in_portfolio = tip
    ws.merge_cells(start_row=tip_row, start_column=1, end_row=tip_row, end_column=gt_col - 1)
    ws.cell(row=tip_row, column=1, value="Total in Portfolio").font = RC_TOT_FNT
    ws.cell(row=tip_row, column=gt_col, value=total_in_portfolio).number_format = ACCT_FMT
    ws.cell(row=tip_row, column=gt_col).font = RC_TOT_FNT
    for c in range(1, gt_col + 1):
        ws.cell(row=tip_row, column=c).fill = RC_TOT_FILL
        ws.cell(row=tip_row, column=c).border = RC_BDR

    # ─── Percent Matrix ───
    pct_start = tip_row + 2
    ws.merge_cells(f'A{pct_start}:B{pct_start}')
    ws[f'A{pct_start}'] = "Percent"
    ws[f'A{pct_start}'].font = RC_HDR_FNT
    ws[f'A{pct_start}'].fill = RC_HDR_FILL
    ws[f'B{pct_start}'].fill = RC_HDR_FILL
    ws.merge_cells(f'C{pct_start}:{grade_last}{pct_start}')
    ws[f'C{pct_start}'] = "Original Credit Grade"
    ws[f'C{pct_start}'].font = RC_HDR_FNT
    ws[f'C{pct_start}'].fill = RC_HDR_FILL
    ws[f'C{pct_start}'].alignment = Alignment(horizontal='center')
    ws.merge_cells(f'{last_col_letter}{pct_start}:{last_col_letter}{pct_start + 1}')
    ws[f'{last_col_letter}{pct_start}'] = "Grand Total"
    ws[f'{last_col_letter}{pct_start}'].font = RC_HDR_FNT2
    ws[f'{last_col_letter}{pct_start}'].fill = RC_HDR_FILL
    ws[f'{last_col_letter}{pct_start}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Grade headers
    hdr_row = pct_start + 1
    ws.merge_cells(f'A{hdr_row}:B{hdr_row}')
    ws[f'A{hdr_row}'] = "Current Credit Grade"
    ws[f'A{hdr_row}'].font = RC_HDR_FNT2
    ws[f'A{hdr_row}'].fill = RC_HDR_FILL
    ws[f'B{hdr_row}'].fill = RC_HDR_FILL
    for j, g in enumerate(gl):
        cell = ws.cell(row=hdr_row, column=3 + j, value=g)
        cell.font = RC_HDR_FNT
        cell.fill = RC_HDR_FILL
        cell.alignment = Alignment(horizontal='center')

    # Percent data rows
    for i, g in enumerate(gl):
        r = hdr_row + 1 + i
        ws.cell(row=r, column=1, value=g).font = FNT_A12BW
        ws.cell(row=r, column=1).fill = RC_HDR_FILL
        ws.cell(row=r, column=1).border = RC_BDR
        ws.cell(row=r, column=2, value=grade_rngs.get(g, '')).font = FNT_A12BW
        ws.cell(row=r, column=2).fill = RC_HDR_FILL
        ws.cell(row=r, column=2).border = RC_BDR

        rtotal_val = 0
        for j, og in enumerate(gl):
            v = _matrix_val(matrix, g, og)
            col_total = sum(_matrix_val(matrix, g2, og) for g2 in gl)
            pct = v / col_total if col_total else 0
            cell = ws.cell(row=r, column=3 + j, value=pct)
            cell.number_format = PCT
            cell.font = FNT_A12
            cell.border = RC_BDR
            rtotal_val += v
            if g == no_score or og == no_score:
                pass  # Not Reported → always unchanged, no fill
            elif i > j:
                if j < n_top and (i - j) < 2:
                    pass  # unchanged
                else:
                    cell.fill = FILL_DET
                    cell.font = FNT_A12W
            elif i < j:
                cell.fill = FILL_IMP
                cell.font = FNT_A12W

        gt_pct = rtotal_val / total if total else 0
        ws.cell(row=r, column=gt_col, value=gt_pct).number_format = PCT
        ws.cell(row=r, column=gt_col).font = FNT_A12B
        ws.cell(row=r, column=gt_col).border = RC_BDR

    # Percent Total row
    pct_total_row = hdr_row + 1 + n_grades
    ws.cell(row=pct_total_row, column=1, value="Total").font = RC_TOT_FNT
    ws.cell(row=pct_total_row, column=1).fill = RC_TOT_FILL
    ws.cell(row=pct_total_row, column=1).border = RC_BDR
    ws.cell(row=pct_total_row, column=2).fill = RC_TOT_FILL
    ws.cell(row=pct_total_row, column=2).border = RC_BDR
    for j, og in enumerate(gl):
        col_total = sum(_matrix_val(matrix, g2, og) for g2 in gl)
        cell = ws.cell(row=pct_total_row, column=3 + j, value=1.0 if col_total > 0 else 0)
        cell.number_format = PCT0
        cell.font = RC_TOT_FNT
        cell.fill = RC_TOT_FILL
        cell.border = RC_BDR
    ws.cell(row=pct_total_row, column=gt_col, value=1.0).number_format = PCT0
    ws.cell(row=pct_total_row, column=gt_col).font = RC_TOT_FNT
    ws.cell(row=pct_total_row, column=gt_col).fill = RC_TOT_FILL
    ws.cell(row=pct_total_row, column=gt_col).border = RC_BDR

    # ─── KEY Section ───
    kr = pct_total_row + 2
    ws.merge_cells(start_row=kr, start_column=1, end_row=kr, end_column=gt_col)
    ws.cell(row=kr, column=1, value="KEY").font = FNT_A12B
    kr += 1
    key_lines = [
        "These tables show the dollar amount of deteriorated and improved loans",
        ("Deteriorated loans are those whose FICO scores have dropped two or more grades "
         "or have dropped to a Grade lower than C."),
        "Improved loans are those whose FICO scores have moved up at least one grade",
        "",
        "Deteriorated loans are shown in the red area (lower left) of the dollar matrix",
        "Improved loans are shown in the green area (upper right) of the dollar matrix",
        "",
        "Unchanged loans are shown in the yellow diagonal of the dollar matrix",
        "",
        ("The percent matrix table at the bottom shows the percent of dollars in each column "
         "(original credit grade at the time the loan was funded) that have migrated to a "
         "different credit grade as of the current date."),
    ]
    for line in key_lines:
        ws.merge_cells(start_row=kr, start_column=1, end_row=kr, end_column=gt_col)
        ws.cell(row=kr, column=1, value=line).font = FNT_A10
        ws.cell(row=kr, column=1).alignment = Alignment(wrap_text=True)
        if len(line) > 80:
            ws.row_dimensions[kr].height = 30
        kr += 1


# ── Improved Deteriorated Summary ─────────────────────────────────
def _sheet_impdet_summary(wb, cu, snap, df, pools, grades, config, hist=None):
    ws = wb.create_sheet("Improved Deteriorated Summary")
    _portrait(ws)
    ws.page_setup.fitToHeight = 0  # fit width only, allow multiple pages tall

    # Use WARM pool order if available; include WARM-only pools too so NRR
    # pools (no DB rows) still appear (with zero impaired/deteriorated %).
    warm_order = (hist or {}).get('impaired', {}).get('pool_order', [])
    _imp_isum = (hist or {}).get('impaired', {}) if hist else {}
    extra_pools_isum = set((_imp_isum.get('hist_bal_data') or {}).keys()) \
                      | set((_imp_isum.get('pool_bal_detail') or {}).keys())
    pool_set = set(pools) | extra_pools_isum
    if warm_order:
        pools = [p for p in warm_order
                 if p in pool_set
                 and not str(p).upper().startswith('HIDE')
                 and str(p) != 'Exclude'
                 and str(p).strip().lower() not in ('grand total','total','excluded')]
        for p in sorted(pool_set - set(pools)):
            if str(p).upper().startswith('HIDE') or str(p) == 'Exclude':
                continue
            if str(p).strip().lower() in ('grand total','total','excluded'):
                continue
            pools.append(p)
    else:
        pools = sorted(p for p in pool_set
                       if not str(p).upper().startswith('HIDE')
                       and str(p) != 'Exclude'
                       and str(p).strip().lower() not in ('grand total','total','excluded'))

    # Restrict df to pools that have DB rows so the Grand Total still matches
    df_pools = set(df['loan_pool'].unique())
    df = df[df['loan_pool'].isin([p for p in pools if p in df_pools])]

    # Column widths – portrait, matching WARM reference
    for col, w in {'A': 30.44, 'B': 30.44, 'C': 30.44, 'D': 12.33,
                   'E': 11.89, 'F': 11.89, 'G': 19.44}.items():
        ws.column_dimensions[col].width = w

    # 3-row header
    ws.merge_cells('A1:G1')
    ws['A1'] = cu
    ws['A1'].font = FNT_A14B

    ws.merge_cells('A2:G2')
    ws['A2'] = "Analysis of Improved/Deteriorated Summary"
    ws['A2'].font = FNT_A12B

    ws.merge_cells('A3:G3')
    ws['A3'] = f"For Period Ending {_snap_display(snap)}"
    ws['A3'].font = FNT_A10B

    r = 5
    chart_count = 0
    nrr = set(config.get('not_risk_rated', []))
    for pool in pools:
        pdf = df[df['loan_pool'] == pool]
        ptotal = pdf['current_balance'].sum()
        if pool in nrr:
            p_imp_pct, p_det_pct = 0.0, 0.0
        else:
            p_imp_pct, p_det_pct, _ = _ncc(pdf, grades, config)
        imp = p_imp_pct * ptotal
        det = p_det_pct * ptotal
        unc = ptotal - imp - det

        # Pool name – dark slate header
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        c1 = ws.cell(row=r, column=1, value=pool)
        c1.font = RC_HDR_FNT
        c1.fill = RC_HDR_FILL
        c1.alignment = Alignment(horizontal='center')
        for c in range(2, 8):
            ws.cell(row=r, column=c).fill = RC_HDR_FILL
        r += 1

        # Headers – dark slate
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.fill = RC_HDR_FILL
            cell.font = RC_HDR_FNT2
            cell.border = RC_BDR
            cell.alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2).value = "$"
        ws.cell(row=r, column=3).value = "%"
        r += 1
        data_start = r

        for lbl, val in [("Total-Improved", imp), ("Total-Deteriorated", det),
                         ("Total Unchanged", unc)]:
            ws.cell(row=r, column=1, value=lbl).font = FNT_A10
            ws.cell(row=r, column=2, value=val).number_format = COMMA_PARENS
            ws.cell(row=r, column=2).font = FNT_A10
            ws.cell(row=r, column=3, value=val / ptotal if ptotal else 0).number_format = PCT
            ws.cell(row=r, column=3).font = FNT_A10
            for c in range(1, 4):
                ws.cell(row=r, column=c).border = RC_BDR
            r += 1

        # Total In Pool – dark slate
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.fill = RC_TOT_FILL
            cell.font = RC_TOT_FNT
            cell.border = RC_BDR
        ws.cell(row=r, column=1).value = "Total In Pool"
        ws.cell(row=r, column=2, value=ptotal).number_format = COMMA_PARENS
        ws.cell(row=r, column=3, value=1.0).number_format = PCT
        r += 1

        # Net Change – dark slate
        net = imp - det
        r += 1
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.fill = RC_TOT_FILL
            cell.font = RC_TOT_FNT
            cell.border = RC_BDR
        ws.cell(row=r, column=1).value = "Net Change"
        ws.cell(row=r, column=2, value=net).number_format = NET_CHG_FMT
        ws.cell(row=r, column=3, value=net / ptotal if ptotal else 0).number_format = PCT

        # Pie chart
        pie = PieChart()
        pie.title = None
        pie.style = 10
        pie.width = 10.5
        pie.height = 3.5
        # Remove chart-area border
        pie.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))

        pie_col = 27  # AA – keep chart data outside print area
        ws.cell(row=data_start, column=pie_col, value=imp / ptotal if ptotal else 0)
        ws.cell(row=data_start + 1, column=pie_col, value=det / ptotal if ptotal else 0)
        ws.cell(row=data_start + 2, column=pie_col, value=unc / ptotal if ptotal else 0)
        ws.cell(row=data_start, column=pie_col + 1, value="Improved")
        ws.cell(row=data_start + 1, column=pie_col + 1, value="Deteriorated")
        ws.cell(row=data_start + 2, column=pie_col + 1, value="Unchanged")

        data_ref = Reference(ws, min_col=pie_col, min_row=data_start, max_row=data_start + 2)
        cats = Reference(ws, min_col=pie_col + 1, min_row=data_start, max_row=data_start + 2)
        pie.add_data(data_ref)
        pie.set_categories(cats)

        s = pie.series[0]
        pt0 = DataPoint(idx=0); pt0.graphicalProperties.solidFill = "2E7D32"
        pt1 = DataPoint(idx=1); pt1.graphicalProperties.solidFill = "C62828"
        pt2 = DataPoint(idx=2); pt2.graphicalProperties.solidFill = "9E9E9E"
        s.data_points = [pt0, pt1, pt2]
        s.dLbls = DataLabelList()
        s.dLbls.showPercent = True
        s.dLbls.showVal = False
        s.dLbls.showSerName = False
        s.dLbls.showCatName = False
        s.dLbls.showLegendKey = False
        pie.legend = None

        ws.add_chart(pie, f"D{data_start - 1}")
        chart_count += 1
        r += 3

    # Grand Total – dark slate header
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c1 = ws.cell(row=r, column=1, value="Grand Total")
    c1.font = RC_HDR_FNT
    c1.fill = RC_HDR_FILL
    c1.alignment = Alignment(horizontal='center')
    for c in range(2, 8):
        ws.cell(row=r, column=c).fill = RC_HDR_FILL
    r += 1

    total = df['current_balance'].sum()
    g_imp_pct, g_det_pct, _ = _ncc(df, grades, config)
    imp = g_imp_pct * total
    det = g_det_pct * total
    unc = total - imp - det

    # Headers – dark slate
    for c in range(1, 4):
        cell = ws.cell(row=r, column=c)
        cell.fill = RC_HDR_FILL
        cell.font = RC_HDR_FNT2
        cell.border = RC_BDR
        cell.alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=2).value = "$"
    ws.cell(row=r, column=3).value = "%"
    r += 1
    gt_start = r
    for lbl, val in [("Total-Improved", imp), ("Total-Deteriorated", det),
                     ("Total Unchanged", unc)]:
        ws.cell(row=r, column=1, value=lbl).font = FNT_A10
        ws.cell(row=r, column=2, value=val).number_format = COMMA_PARENS
        ws.cell(row=r, column=3, value=val / total if total else 0).number_format = PCT
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = RC_BDR
        r += 1

    # Total In Portfolio – dark slate
    for c in range(1, 4):
        cell = ws.cell(row=r, column=c)
        cell.fill = RC_TOT_FILL
        cell.font = RC_TOT_FNT
        cell.border = RC_BDR
    ws.cell(row=r, column=1).value = "Total In Portfolio"
    ws.cell(row=r, column=2, value=total).number_format = COMMA_PARENS
    ws.cell(row=r, column=3, value=1.0).number_format = PCT
    r += 2

    # Net Change – dark slate
    net = imp - det
    for c in range(1, 4):
        cell = ws.cell(row=r, column=c)
        cell.fill = RC_TOT_FILL
        cell.font = RC_TOT_FNT
        cell.border = RC_BDR
    ws.cell(row=r, column=1).value = "Net Change"
    ws.cell(row=r, column=2, value=net).number_format = NET_CHG_FMT
    ws.cell(row=r, column=3, value=net / total if total else 0).number_format = PCT

    # Grand total pie
    pie = PieChart()
    pie.style = 10; pie.width = 10.5; pie.height = 3.5
    # Remove chart-area border
    pie.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    pie_col = 27  # AA – keep chart data outside print area
    ws.cell(row=gt_start, column=pie_col, value=imp / total if total else 0)
    ws.cell(row=gt_start + 1, column=pie_col, value=det / total if total else 0)
    ws.cell(row=gt_start + 2, column=pie_col, value=unc / total if total else 0)
    ws.cell(row=gt_start, column=pie_col + 1, value="Improved")
    ws.cell(row=gt_start + 1, column=pie_col + 1, value="Deteriorated")
    ws.cell(row=gt_start + 2, column=pie_col + 1, value="Unchanged")
    data_ref = Reference(ws, min_col=pie_col, min_row=gt_start, max_row=gt_start + 2)
    cats = Reference(ws, min_col=pie_col + 1, min_row=gt_start, max_row=gt_start + 2)
    pie.add_data(data_ref)
    pie.set_categories(cats)
    s = pie.series[0]
    pt0 = DataPoint(idx=0); pt0.graphicalProperties.solidFill = "2E7D32"
    pt1 = DataPoint(idx=1); pt1.graphicalProperties.solidFill = "C62828"
    pt2 = DataPoint(idx=2); pt2.graphicalProperties.solidFill = "9E9E9E"
    s.data_points = [pt0, pt1, pt2]
    s.dLbls = DataLabelList()
    s.dLbls.showPercent = True
    s.dLbls.showVal = False
    s.dLbls.showSerName = False
    s.dLbls.showCatName = False
    s.dLbls.showLegendKey = False
    pie.legend = None
    ws.add_chart(pie, f"D{gt_start - 1}")

    ws.print_area = f'A1:G{r}'


# ── Historical Trends Balance ─────────────────────────────────────
def _sheet_hist_balance_charts(wb, cu, snap, df, grades, config, hist):
    ws = wb.create_sheet("Historical Trends Balance")

    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g['label'] for g in grades] + [no_score]

    # ── Title rows (no merge – matches Vizo supplemental) ──
    ws['A1'] = cu
    ws['A1'].font = FNT_A14B

    ws['A2'] = "Historical Loan Balances by Most Recent Credit Score"
    ws['A2'].font = FNT_A12B

    ws['A3'] = f"For Period Ending {_snap_display(snap)}"
    ws['A3'].font = FNT_A10B

    impaired = hist.get('impaired', {}) if hist else {}
    hist_bal = impaired.get('hist_bal_data', {})
    rr_map = impaired.get('risk_rated', {})
    nrr_cfg = set(config.get('not_risk_rated', []))
    acl_months = impaired.get('acl_months', {})
    pools = _sort_pools(df['loan_pool'].unique(), config)

    # Use the wizard's pool_order (config) when available so the user's
    # reordering on Step 2 (Loan Pools) drives the report. Fall back to
    # the WARM-derived order. NRR pools missing from either list are
    # still appended by the merge helper.
    warm_order = config.get('pool_order') or impaired.get('pool_order', [])
    pools = _merge_pool_orders(pools, warm_order)

    if not hist_bal:
        ws.cell(row=5, column=1, value="No historical balance data available.").font = FNT_A10
        return

    # Build visible grade list (exclude hidden)
    visible_gl = [g for g in gl if not _is_hidden(g)]

    # ── Write data tables far right (col 30+) so they don't print ──
    DATA_COL = 30  # col AD
    pool_chart_info = []
    data_row = 5

    for pool in pools:
        # Skip non-risk-rated pools (WARM detection or YAML override)
        if not rr_map.get(pool, True) or pool in nrr_cfg:
            continue
        pdata = hist_bal.get(pool)
        if not pdata:
            continue

        dates = pdata.get('dates', [])
        grade_vals = pdata.get('grades', {})
        if not dates or not grade_vals:
            continue

        # Trim to life-of-loan window (last N months)
        lol = acl_months.get(pool)
        if lol and lol < len(dates):
            dates = dates[-lol:]
            grade_vals = {g: v[-lol:] for g, v in grade_vals.items()}

        # Header row with date labels (month + year only)
        for di, dt in enumerate(dates):
            lbl = dt.strftime('%b %Y') if hasattr(dt, 'strftime') else str(dt)
            ws.cell(row=data_row, column=DATA_COL + di, value=lbl)
        ncol_end = DATA_COL + len(dates) - 1
        data_row += 1
        grade_start = data_row

        # One row per visible grade
        n_written = 0
        for g in visible_gl:
            vals = grade_vals.get(g, [0] * len(dates))
            ws.cell(row=data_row, column=DATA_COL - 1, value=g)
            for di, v in enumerate(vals):
                ws.cell(row=data_row, column=DATA_COL + di, value=v)
            data_row += 1
            n_written += 1

        if n_written == 0:
            continue

        pool_chart_info.append({
            'pool': pool,
            'date_row': grade_start - 1,
            'grade_start': grade_start,
            'grade_end': data_row - 1,
            'min_col': DATA_COL,
            'max_col': ncol_end,
        })
        data_row += 1  # separator row

    if not pool_chart_info:
        ws.cell(row=5, column=1, value="No historical grade-level balance data available.")
        ws['A5'].font = Font(name='Arial', italic=True, size=10, color='888888')
        return

    # ── Build Grand Total across all risk-rated pools (no lol trim) ──
    gt_agg = {}  # date -> {grade -> sum}
    for pool in pools:
        if not rr_map.get(pool, True) or pool in nrr_cfg:
            continue
        pdata = hist_bal.get(pool)
        if not pdata:
            continue
        dates = pdata.get('dates', [])
        grade_vals = pdata.get('grades', {})
        if not dates or not grade_vals:
            continue
        for di, dt in enumerate(dates):
            if dt not in gt_agg:
                gt_agg[dt] = {g: 0.0 for g in visible_gl}
            for g in visible_gl:
                v = grade_vals.get(g, [])
                if di < len(v) and v[di] is not None:
                    try:
                        gt_agg[dt][g] += float(v[di])
                    except (TypeError, ValueError):
                        pass

    if gt_agg:
        gt_dates = sorted(gt_agg.keys())
        for di, dt in enumerate(gt_dates):
            lbl = dt.strftime('%b %Y') if hasattr(dt, 'strftime') else str(dt)
            ws.cell(row=data_row, column=DATA_COL + di, value=lbl)
        ncol_end = DATA_COL + len(gt_dates) - 1
        data_row += 1
        grade_start = data_row
        for g in visible_gl:
            ws.cell(row=data_row, column=DATA_COL - 1, value=g)
            for di, dt in enumerate(gt_dates):
                ws.cell(row=data_row, column=DATA_COL + di, value=gt_agg[dt][g])
            data_row += 1
        # Append so the Grand Total chart appears last
        pool_chart_info.append({
            'pool': 'Grand Total',
            'date_row': grade_start - 1,
            'grade_start': grade_start,
            'grade_end': data_row - 1,
            'min_col': DATA_COL,
            'max_col': ncol_end,
        })
        data_row += 1

    # ── Create one line chart per pool (Vizo layout) ──
    CHART_ROWS = 16
    ACCENT_HEX = ['2E7D32', 'C62828', '2C3E50', '9E9E9E', '34495E', 'FFC107']
    chart_row = 5

    for idx, pb in enumerate(pool_chart_info):
        chart = LineChart()
        chart.width = 19.5
        chart.height = 8.0
        chart.style = 10

        # Remove chart-area and plot-area borders
        chart.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
        chart.plot_area.graphicalProperties = GraphicalProperties(
            ln=LineProperties(noFill=True)
        )

        # Title: 20pt Cambria to match Vizo template
        chart.title = pb['pool']
        chart.title.tx = Text(
            rich=RichText(
                p=[DrawParagraph(
                    pPr=ParagraphProperties(
                        defRPr=CharacterProperties(
                            sz=2000, b=False,
                            latin=DrawingFont(typeface='Cambria'),
                        )
                    ),
                    r=[RegularTextRun(
                        rPr=CharacterProperties(
                            sz=2000, b=False,
                            latin=DrawingFont(typeface='Cambria'),
                        ),
                        t=pb['pool'],
                    )],
                )]
            )
        )

        # Y-axis: accounting format, min=0
        chart.y_axis.numFmt = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"_);_(@_)'
        chart.y_axis.delete = False
        chart.y_axis.scaling.min = 0
        chart.x_axis.delete = False
        chart.x_axis.txPr = RichText(
            bodyPr=RichTextProperties(rot=-2400000),
            p=[DrawParagraph(
                pPr=ParagraphProperties(
                    defRPr=CharacterProperties()
                ),
                r=[],
            )]
        )

        # Categories: quarter labels
        cats = Reference(ws,
                         min_col=pb['min_col'], max_col=pb['max_col'],
                         min_row=pb['date_row'])

        # Data: one row per grade
        data = Reference(ws,
                         min_col=pb['min_col'], max_col=pb['max_col'],
                         min_row=pb['grade_start'],
                         max_row=pb['grade_end'])
        chart.add_data(data, from_rows=True, titles_from_data=False)
        chart.set_categories(cats)

        # Series labels + Vizo accent colors
        for gi, s in enumerate(chart.series):
            grade_name = ws.cell(pb['grade_start'] + gi, pb['min_col'] - 1).value or ''
            s.tx = SeriesLabel(v=grade_name)
            hex_color = ACCENT_HEX[gi % len(ACCENT_HEX)]
            s.graphicalProperties.line = LineProperties(
                w=38100, cap='rnd', prstDash='solid',
                solidFill=hex_color, round=True,
            )

        # Legend below title
        chart.legend.position = 't'
        chart.legend.layout = Layout(
            manualLayout=ManualLayout(
                xMode='edge', yMode='edge',
                x=0.10, y=0.14,
                w=0.80, h=0.06,
            )
        )
        # Plot area
        chart.layout = Layout(
            manualLayout=ManualLayout(
                xMode='edge', yMode='edge',
                x=0.05, y=0.22,
                w=0.93, h=0.72,
            )
        )

        ws.add_chart(chart, f"A{chart_row}")
        chart_row += CHART_ROWS

        # Page break after every 3rd chart
        if (idx + 1) % 3 == 0 and idx + 1 < len(pool_chart_info):
            ws.row_breaks.append(Break(id=chart_row - 1))

    # ── Page setup: portrait, 3 charts per page, repeating headers ──
    last_data_row = chart_row - 1 if pool_chart_info else 5
    ws.print_area = f'A1:L{last_data_row}'
    ws.print_title_rows = '1:3'
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25


# ── Risk Change per Pool ──────────────────────────────────────────
def _sheet_pool_risk_change(wb, cu, snap, pool_df, pool_name, pool_idx, grades, config, hist):
    """One landscape sheet per pool with dollar/percent matrix + charts."""
    no_score = config.get('no_score_label', 'Not Reported')
    all_gl = _all_grades(grades, no_score)
    gl = [g for g in all_gl if not _is_hidden(g)]
    matrix = risk_change_matrix(pool_df, grades, no_score)
    pool_total = pool_df['current_balance'].sum()
    grade_rngs = _grade_ranges(grades, no_score)
    n_grades = len(gl)
    gt_col = 3 + n_grades  # Grand Total column
    last_col_letter = chr(ord('C') + n_grades)
    grade_last = chr(ord('C') + n_grades - 1)

    # Sheet name: "Risk Change Auto", "Risk Change Real Estate", etc.
    sheet_name = f"Risk Change {pool_name}"
    # Sanitize characters invalid in Excel sheet names
    for ch in ['/', '\\', '?', '*', '[', ']', ':']:
        sheet_name = sheet_name.replace(ch, '-')
    # Excel sheet names max 31 chars
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    ws = wb.create_sheet(sheet_name)
    _landscape(ws)

    col_widths = {}
    ncols = gt_col  # total columns A through last_col_letter
    col_w = round(226.0 / ncols, 2)  # fill landscape letter page
    for idx in range(ncols):
        col_widths[chr(ord('A') + idx)] = col_w
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Row 1: CU name
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = cu
    ws['A1'].font = FNT_A14B

    # Row 2: Title
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = "Risk Change By Credit Score"
    ws['A2'].font = FNT_A12B

    # Row 3: Date
    ws.merge_cells(f'A3:{last_col_letter}3')
    ws['A3'] = f"For Period Ending {_snap_display(snap)}"
    ws['A3'].font = FNT_A10B

    # Row 4: Pool name
    ws.merge_cells(f'A4:{last_col_letter}4')
    ws['A4'] = pool_name
    ws['A4'].font = FNT_A18B

    # ─── Dollar Matrix ─── (starts row 5)
    ws.merge_cells('A5:B5')
    ws['A5'] = "Dollar"
    ws['A5'].font = RC_HDR_FNT
    ws['A5'].fill = RC_HDR_FILL
    ws['B5'].fill = RC_HDR_FILL
    ws.merge_cells(f'C5:{grade_last}5')
    ws['C5'] = "Original Grade"
    ws['C5'].font = RC_HDR_FNT
    ws['C5'].fill = RC_HDR_FILL
    ws['C5'].alignment = Alignment(horizontal='center')
    ws.merge_cells(f'{last_col_letter}5:{last_col_letter}6')
    ws[f'{last_col_letter}5'] = "Grand Total"
    ws[f'{last_col_letter}5'].font = RC_HDR_FNT2
    ws[f'{last_col_letter}5'].fill = RC_HDR_FILL
    ws[f'{last_col_letter}5'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Row 6: Grade labels
    ws.merge_cells('A6:B6')
    ws['A6'] = "Current Grade"
    ws['A6'].font = RC_HDR_FNT2
    ws['A6'].fill = RC_HDR_FILL
    ws['B6'].fill = RC_HDR_FILL
    for j, g in enumerate(gl):
        cell = ws.cell(row=6, column=3 + j, value=g)
        cell.font = RC_HDR_FNT
        cell.fill = RC_HDR_FILL
        cell.alignment = Alignment(horizontal='center')

    n_top = config.get('top_grades_double_drop', 3)

    # Data rows
    for i, g in enumerate(gl):
        r = 7 + i
        ws.cell(row=r, column=1, value=g).font = FNT_A12BW
        ws.cell(row=r, column=1).fill = RC_HDR_FILL
        ws.cell(row=r, column=1).border = RC_BDR
        ws.cell(row=r, column=2, value=grade_rngs.get(g, '')).font = FNT_A12BW
        ws.cell(row=r, column=2).fill = RC_HDR_FILL
        ws.cell(row=r, column=2).border = RC_BDR

        rt = 0
        for j, og in enumerate(gl):
            v = _matrix_val(matrix, g, og)
            cell = ws.cell(row=r, column=3 + j, value=v)
            cell.number_format = ACCT_FMT
            cell.font = FNT_A12
            cell.border = RC_BDR
            rt += v
            if g == no_score or og == no_score:
                pass  # Not Reported → always unchanged, no fill
            elif i > j:
                if j < n_top and (i - j) < 2:
                    pass  # unchanged
                else:
                    cell.fill = FILL_DET
                    cell.font = FNT_A12W
            elif i < j:
                cell.fill = FILL_IMP
                cell.font = FNT_A12W
        ws.cell(row=r, column=gt_col, value=rt).number_format = ACCT_FMT
        ws.cell(row=r, column=gt_col).font = FNT_A12B
        ws.cell(row=r, column=gt_col).border = RC_BDR

    # Total row
    total_row = 7 + n_grades
    ws.cell(row=total_row, column=1, value="Grand Total").font = RC_TOT_FNT
    ws.cell(row=total_row, column=1).fill = RC_TOT_FILL
    ws.cell(row=total_row, column=1).border = RC_BDR
    ws.cell(row=total_row, column=2).fill = RC_TOT_FILL
    ws.cell(row=total_row, column=2).border = RC_BDR
    matrix_total = 0
    for j, og in enumerate(gl):
        ct = sum(_matrix_val(matrix, g2, og) for g2 in gl)
        ws.cell(row=total_row, column=3 + j, value=ct).number_format = ACCT_FMT
        ws.cell(row=total_row, column=3 + j).font = RC_TOT_FNT
        ws.cell(row=total_row, column=3 + j).fill = RC_TOT_FILL
        ws.cell(row=total_row, column=3 + j).border = RC_BDR
        matrix_total += ct
    ws.cell(row=total_row, column=gt_col, value=matrix_total).number_format = ACCT_FMT
    ws.cell(row=total_row, column=gt_col).font = RC_TOT_FNT
    ws.cell(row=total_row, column=gt_col).fill = RC_TOT_FILL
    ws.cell(row=total_row, column=gt_col).border = RC_BDR

    # Adjustments (from WARM balance adjustment data)
    _imp = hist.get('impaired', {}) if hist else {}
    if pool_name == "Total Loans":
        bal_adj = _imp.get('total_balance_adjustment', 0.0) or 0.0
    else:
        _bal_adj_map = _imp.get('balance_adjustments', {})
        _pool_lc = pool_name.strip().lower() if pool_name else ''
        bal_adj = next((v for k, v in _bal_adj_map.items() if k.strip().lower() == _pool_lc), 0.0)
    adj_row = total_row + 1
    ws.merge_cells(start_row=adj_row, start_column=1, end_row=adj_row, end_column=gt_col - 1)
    ws.cell(row=adj_row, column=1, value="Loans Not Risk Rated and Adjustments").font = RC_TOT_FNT
    ws.cell(row=adj_row, column=gt_col, value=bal_adj).number_format = ACCT_FMT
    ws.cell(row=adj_row, column=gt_col).font = RC_TOT_FNT
    for c in range(1, gt_col + 1):
        ws.cell(row=adj_row, column=c).fill = RC_TOT_FILL
        ws.cell(row=adj_row, column=c).border = RC_BDR

    tip_row = adj_row + 1
    total_in_pool = matrix_total + bal_adj
    if pool_name == "Total Loans":
        tip = _imp.get('total_in_portfolio', total_in_pool)
        if tip:
            total_in_pool = tip
    ws.merge_cells(start_row=tip_row, start_column=1, end_row=tip_row, end_column=gt_col - 1)
    ws.cell(row=tip_row, column=1, value="Total in Pool").font = RC_TOT_FNT
    ws.cell(row=tip_row, column=gt_col, value=total_in_pool).number_format = ACCT_FMT
    ws.cell(row=tip_row, column=gt_col).font = RC_TOT_FNT
    for c in range(1, gt_col + 1):
        ws.cell(row=tip_row, column=c).fill = RC_TOT_FILL
        ws.cell(row=tip_row, column=c).border = RC_BDR

    # ─── Percent Matrix ───
    pct_start = tip_row + 2
    ws.merge_cells(f'A{pct_start}:B{pct_start}')
    ws[f'A{pct_start}'] = "Percent"
    ws[f'A{pct_start}'].font = RC_HDR_FNT
    ws[f'A{pct_start}'].fill = RC_HDR_FILL
    ws[f'B{pct_start}'].fill = RC_HDR_FILL
    ws.merge_cells(f'C{pct_start}:{grade_last}{pct_start}')
    ws[f'C{pct_start}'] = "Original Grade"
    ws[f'C{pct_start}'].font = RC_HDR_FNT
    ws[f'C{pct_start}'].fill = RC_HDR_FILL
    ws[f'C{pct_start}'].alignment = Alignment(horizontal='center')
    ws.merge_cells(f'{last_col_letter}{pct_start}:{last_col_letter}{pct_start + 1}')
    ws[f'{last_col_letter}{pct_start}'] = "Grand Total"
    ws[f'{last_col_letter}{pct_start}'].font = RC_HDR_FNT2
    ws[f'{last_col_letter}{pct_start}'].fill = RC_HDR_FILL
    ws[f'{last_col_letter}{pct_start}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    pct_hdr = pct_start + 1
    ws.merge_cells(f'A{pct_hdr}:B{pct_hdr}')
    ws[f'A{pct_hdr}'] = "Current Grade"
    ws[f'A{pct_hdr}'].font = RC_HDR_FNT2
    ws[f'A{pct_hdr}'].fill = RC_HDR_FILL
    ws[f'B{pct_hdr}'].fill = RC_HDR_FILL
    for j, g in enumerate(gl):
        cell = ws.cell(row=pct_hdr, column=3 + j, value=g)
        cell.font = RC_HDR_FNT
        cell.fill = RC_HDR_FILL
        cell.alignment = Alignment(horizontal='center')

    for i, g in enumerate(gl):
        r = pct_hdr + 1 + i
        ws.cell(row=r, column=1, value=g).font = FNT_A12BW
        ws.cell(row=r, column=1).fill = RC_HDR_FILL
        ws.cell(row=r, column=1).border = RC_BDR
        ws.cell(row=r, column=2, value=grade_rngs.get(g, '')).font = FNT_A12BW
        ws.cell(row=r, column=2).fill = RC_HDR_FILL
        ws.cell(row=r, column=2).border = RC_BDR

        rt = 0
        for j, og in enumerate(gl):
            v = _matrix_val(matrix, g, og)
            ct = sum(_matrix_val(matrix, g2, og) for g2 in gl)
            cell = ws.cell(row=r, column=3 + j, value=v / ct if ct else 0)
            cell.number_format = PCT
            cell.font = FNT_A12
            cell.border = RC_BDR
            rt += v
            if g == no_score or og == no_score:
                pass  # Not Reported → always unchanged, no fill
            elif i > j:
                if j < n_top and (i - j) < 2:
                    pass  # unchanged
                else:
                    cell.fill = FILL_DET
                    cell.font = FNT_A12W
            elif i < j:
                cell.fill = FILL_IMP
                cell.font = FNT_A12W

        gt_pct = rt / pool_total if pool_total else 0
        ws.cell(row=r, column=gt_col, value=gt_pct).number_format = PCT
        ws.cell(row=r, column=gt_col).font = FNT_A12B
        ws.cell(row=r, column=gt_col).border = RC_BDR

    # Percent total row
    pct_total_row = pct_hdr + 1 + n_grades
    ws.cell(row=pct_total_row, column=1, value="Grand Total").font = RC_TOT_FNT
    ws.cell(row=pct_total_row, column=1).fill = RC_TOT_FILL
    ws.cell(row=pct_total_row, column=1).border = RC_BDR
    ws.cell(row=pct_total_row, column=2).fill = RC_TOT_FILL
    ws.cell(row=pct_total_row, column=2).border = RC_BDR
    for j, og in enumerate(gl):
        ct = sum(_matrix_val(matrix, g2, og) for g2 in gl)
        cell = ws.cell(row=pct_total_row, column=3 + j, value=1.0 if ct > 0 else 0)
        cell.number_format = PCT0
        cell.font = RC_TOT_FNT
        cell.fill = RC_TOT_FILL
        cell.border = RC_BDR
    ws.cell(row=pct_total_row, column=gt_col, value=1.0).number_format = PCT0
    ws.cell(row=pct_total_row, column=gt_col).font = RC_TOT_FNT
    ws.cell(row=pct_total_row, column=gt_col).fill = RC_TOT_FILL
    ws.cell(row=pct_total_row, column=gt_col).border = RC_BDR

    # ─── Compute improved / deteriorated / unchanged totals ───
    n_top = config.get('top_grades_double_drop', 3)
    grand_imp = 0
    grand_det = 0
    # Also compute per-grade breakdowns for the bar chart
    grade_labels = [g for g in gl if g != no_score]
    det_by_grade = []
    imp_by_grade = []
    for j, og in enumerate(gl):
        det_g = 0
        imp_g = 0
        for i, g in enumerate(gl):
            v = _matrix_val(matrix, g, og)
            if g == no_score or og == no_score:
                pass
            elif i > j:
                if j < n_top and (i - j) < 2:
                    pass
                else:
                    grand_det += v
                    det_g += v
            elif i < j:
                grand_imp += v
                imp_g += v
        if og != no_score:
            det_by_grade.append(det_g)
            imp_by_grade.append(imp_g)
    grand_unc = pool_total - grand_imp - grand_det
    net = grand_imp - grand_det
    total = pool_total

    # ─── Row anchoring (relative to pct_total_row = "Grand Total" of % matrix) ───
    r_pgt = pct_total_row
    r_nc = r_pgt + 2

    # ─── "Net Credit Change" label ───
    ws.merge_cells(start_row=r_nc, start_column=2, end_row=r_nc + 1, end_column=4)
    ws.cell(row=r_nc, column=2, value="Net Credit Change").font = FNT_A18B
    ws.cell(row=r_nc, column=2).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r_nc].height = 15.6
    ws.row_dimensions[r_nc + 1].height = 18.0

    # ─── Summary Table (cols B-D) ───
    r_sum = r_nc + 6
    WHITE_BOLD12 = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    summary_items = [
        ("Improved",     grand_imp, grand_imp / total if total else 0, FILL_IMP),
        ("Deteriorated", grand_det, grand_det / total if total else 0, FILL_DET),
        ("Unchanged",    grand_unc, grand_unc / total if total else 0, RC_HDR_FILL),
        ("Portfolio",    total,     1.0,                               None),
        ("Net Change",   net,       net / total if total else 0,       None),
    ]
    for si, (lbl, bal, pct, fill) in enumerate(summary_items):
        r = r_sum + si
        ws.cell(row=r, column=2, value=lbl).font = FNT_A12B
        c_bal = ws.cell(row=r, column=3, value=bal)
        c_bal.number_format = '#,##0'
        c_bal.font = FNT_A12B
        c_bal.alignment = Alignment(horizontal='center')
        c_pct = ws.cell(row=r, column=4, value=pct)
        c_pct.number_format = '0.0%'
        c_pct.font = FNT_A12B
        c_pct.alignment = Alignment(horizontal='center')
        if fill:
            c_bal.fill = fill
            c_pct.fill = fill
            c_bal.font = WHITE_BOLD12
            c_pct.font = WHITE_BOLD12

    # ─── Doughnut Chart (Improved / Deteriorated / Unchanged) ───
    dc = DoughnutChart()
    dc.title = None
    dc.style = 10
    cats_dn = Reference(ws, min_col=2, min_row=r_sum, max_row=r_sum + 2)
    vals_dn = Reference(ws, min_col=4, min_row=r_sum, max_row=r_sum + 2)
    dc.add_data(vals_dn)
    dc.set_categories(cats_dn)
    s_dn = dc.series[0]
    s_dn.explosion = 16
    dn_pts = []
    for dp_idx, dp_color in enumerate(['2E7D32', 'C62828', '2C3E50']):
        dp = DataPoint(idx=dp_idx)
        dp.graphicalProperties = GraphicalProperties()
        dp.graphicalProperties.noFill = True
        dp.graphicalProperties.line = LineProperties(solidFill=dp_color, w=38100)
        dn_pts.append(dp)
    s_dn.data_points = dn_pts
    dc.innerRadius = 50
    dc.legend = None
    dc.graphical_properties = GraphicalProperties()
    dc.graphical_properties.noFill = True
    dc.graphical_properties.line = LineProperties(noFill=True)
    dc.width = 12
    dc.height = 7.5
    anc_dc = TwoCellAnchor()
    anc_dc._from = AnchorMarker(col=2, colOff=500000, row=r_nc + 1, rowOff=0)
    anc_dc.to = AnchorMarker(col=6, colOff=500000, row=r_nc + 13, rowOff=0)
    ws.add_chart(dc, anc_dc)

    # ─── Bar Chart Data (det/imp per grade in hidden cols P-R) ───
    pcol_start = 16  # column P
    ws.cell(row=6, column=pcol_start, value="Deteriorated").font = FNT_A10B
    ws.cell(row=6, column=pcol_start + 1, value="Improved").font = FNT_A10B
    for gi, g in enumerate(grade_labels):
        r = 7 + gi
        ws.cell(row=r, column=pcol_start, value=det_by_grade[gi]).number_format = ACCT_FMT
        ws.cell(row=r, column=pcol_start + 1, value=imp_by_grade[gi]).number_format = ACCT_FMT

    # ─── Bar Chart "Risk Change by Grade" ───
    rc_bar = BarChart()
    rc_bar.type = 'col'
    rc_bar.grouping = 'clustered'
    rc_bar.title = 'Risk Change by Grade'
    rc_bar.y_axis.delete = True
    rc_bar.y_axis.numFmt = ACCT_FMT
    rc_bar.x_axis.delete = False
    rc_bar.x_axis.tickLblPos = 'nextTo'
    rc_bar.x_axis.majorTickMark = 'out'
    rc_bar.x_axis.minorTickMark = 'none'
    rc_bar.gapWidth = 100
    rc_last = 7 + len(grade_labels) - 1
    cats_rc = Reference(ws, min_col=1, min_row=7, max_row=rc_last)
    det_ref = Reference(ws, min_col=pcol_start, min_row=6, max_row=rc_last)
    imp_ref = Reference(ws, min_col=pcol_start + 1, min_row=6, max_row=rc_last)
    rc_bar.add_data(det_ref, titles_from_data=True)
    rc_bar.add_data(imp_ref, titles_from_data=True)
    rc_bar.set_categories(cats_rc)
    rc_bar.series[0].graphicalProperties.noFill = True
    rc_bar.series[0].graphicalProperties.line = LineProperties(solidFill='C62828', w=38100)
    rc_bar.series[1].graphicalProperties.noFill = True
    rc_bar.series[1].graphicalProperties.line = LineProperties(solidFill='2E7D32', w=38100)
    rc_bar.legend = ChartLegend()
    rc_bar.legend.position = 't'
    rc_bar.legend.overlay = False
    rc_bar.legend.layout = Layout(
        manualLayout=ManualLayout(
            xMode='edge', yMode='edge',
            x=0.3, y=0.10, w=0.4, h=0.06,
        )
    )
    rc_bar.y_axis.majorGridlines = None
    rc_bar.width = 20
    rc_bar.height = 7.5
    # Remove chart-area border
    rc_bar.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    anc_rc = TwoCellAnchor()
    anc_rc._from = AnchorMarker(col=5, colOff=0, row=r_nc + 2, rowOff=0)
    anc_rc.to = AnchorMarker(col=9, colOff=0, row=r_nc + 13, rowOff=0)
    ws.add_chart(rc_bar, anc_rc)

    # ─── DQ / CO data from WARM ───
    _imp = hist.get('impaired', {}) if hist else {}
    if pool_name == "Total Loans":
        _dq_data = _imp.get('dq_by_status', {})
        _co_data = _imp.get('co_by_status', {})
    else:
        _dq_pool = _imp.get('dq_by_pool', {})
        _co_pool = _imp.get('co_by_pool', {})
        _pool_lc = pool_name.strip().lower() if pool_name else ''
        _dq_data = next((v for k, v in _dq_pool.items() if k.strip().lower() == _pool_lc), {})
        _co_data = next((v for k, v in _co_pool.items() if k.strip().lower() == _pool_lc), {})

    # ─── DQ Data Table (cols P-R) ───
    r_dq = r_nc + 16
    ws.cell(row=r_dq, column=pcol_start, value="Loan Status").font = FNT_A12B
    ws.cell(row=r_dq, column=pcol_start + 1, value="DQ Balance").font = FNT_A12B
    ws.cell(row=r_dq, column=pcol_start + 2, value="% of Total").font = FNT_A12B
    for di, lbl in enumerate(["Improved", "Deteriorated", "Unchanged", "Not Reported"]):
        r = r_dq + 1 + di
        dq_entry = _dq_data.get(lbl, {})
        ws.cell(row=r, column=pcol_start, value=lbl).font = FNT_A10
        ws.cell(row=r, column=pcol_start + 1, value=dq_entry.get('balance', 0)).number_format = '#,##0'
        ws.cell(row=r, column=pcol_start + 2, value=dq_entry.get('pct', 0)).number_format = '0.0%'

    # ─── DQ Pie Chart "Delinquency by Credit Grade Migration" ───
    dq_pie = PieChart()
    dq_pie.title = 'Delinquency by Credit Grade Migration'
    cats_dq = Reference(ws, min_col=pcol_start, min_row=r_dq + 1, max_row=r_dq + 4)
    vals_dq = Reference(ws, min_col=pcol_start + 2, min_row=r_dq + 1, max_row=r_dq + 4)
    dq_pie.add_data(vals_dq)
    dq_pie.set_categories(cats_dq)
    s_dq = dq_pie.series[0]
    s_dq.explosion = 21
    s_dq.graphicalProperties = GraphicalProperties()
    s_dq.graphicalProperties.noFill = True
    s_dq.graphicalProperties.line = LineProperties(solidFill='2C3E50', w=38100)
    dq_pts = []
    for dp_idx, dp_color in enumerate(['2E7D32', 'C62828', '2C3E50', '9E9E9E']):
        dp = DataPoint(idx=dp_idx)
        dp.graphicalProperties = GraphicalProperties()
        dp.graphicalProperties.noFill = True
        dp.graphicalProperties.line = LineProperties(solidFill=dp_color, w=38100)
        dq_pts.append(dp)
    s_dq.data_points = dq_pts
    s_dq.dLbls = DataLabelList()
    s_dq.dLbls.showVal = True
    s_dq.dLbls.showLegendKey = True
    s_dq.dLbls.showCatName = False
    s_dq.dLbls.showSerName = False
    s_dq.dLbls.showPercent = False
    s_dq.dLbls.showLeaderLines = True
    s_dq.dLbls.numFmt = '0.0%'
    s_dq.dLbls.dLblPos = 'outEnd'
    dq_pie.legend = ChartLegend()
    dq_pie.legend.position = 'b'
    dq_pie.legend.overlay = False
    dq_pie.layout = Layout(
        manualLayout=ManualLayout(
            xMode='edge', yMode='edge',
            x=0.32, y=0.22, w=0.36, h=0.48,
        )
    )
    dq_pie.graphical_properties = GraphicalProperties()
    dq_pie.graphical_properties.solidFill = 'FFFFFF'
    dq_pie.graphical_properties.line = LineProperties(noFill=True)
    dq_pie.width = 15
    dq_pie.height = 7.5
    anc_dq = TwoCellAnchor()
    anc_dq._from = AnchorMarker(col=0, colOff=0, row=r_nc + 15, rowOff=0)
    anc_dq.to = AnchorMarker(col=5, colOff=0, row=r_nc + 28, rowOff=0)
    ws.add_chart(dq_pie, anc_dq)

    # ─── CO Data Table (cols P-R) ───
    r_co = r_nc + 22
    ws.cell(row=r_co, column=pcol_start, value="Loan Status").font = FNT_A12B
    ws.cell(row=r_co, column=pcol_start + 1, value="CO Balance").font = FNT_A12B
    ws.cell(row=r_co, column=pcol_start + 2, value="% of Total").font = FNT_A12B
    for ci, lbl in enumerate(["Improved", "Deteriorated", "Unchanged", "Not Reported"]):
        r = r_co + 1 + ci
        co_entry = _co_data.get(lbl, {})
        ws.cell(row=r, column=pcol_start, value=lbl).font = FNT_A10
        ws.cell(row=r, column=pcol_start + 1, value=co_entry.get('balance', 0)).number_format = '#,##0'
        ws.cell(row=r, column=pcol_start + 2, value=co_entry.get('pct', 0)).number_format = '0.0%'

    # ─── CO Bar Chart "Charge off by Credit Grade Migration" ───
    co_bar = BarChart()
    co_bar.type = 'bar'
    co_bar.grouping = 'clustered'
    co_bar.title = 'Charge off by Credit Grade Migration'
    co_bar.y_axis.delete = True
    co_bar.x_axis.delete = False
    co_bar.x_axis.tickLblPos = 'nextTo'
    co_bar.x_axis.majorTickMark = 'out'
    co_bar.x_axis.minorTickMark = 'none'
    co_bar.gapWidth = 100
    cats_co = Reference(ws, min_col=pcol_start, min_row=r_co + 1, max_row=r_co + 4)
    vals_co = Reference(ws, min_col=pcol_start + 2, min_row=r_co + 1, max_row=r_co + 4)
    co_bar.add_data(vals_co)
    co_bar.set_categories(cats_co)
    s_co = co_bar.series[0]
    s_co.graphicalProperties = GraphicalProperties()
    s_co.graphicalProperties.noFill = True
    s_co.graphicalProperties.line = LineProperties(solidFill='9E9E9E', w=50800)
    co_pts = []
    for dp_idx, dp_color in enumerate(['2E7D32', 'C62828', '2C3E50', '9E9E9E']):
        dp = DataPoint(idx=dp_idx)
        dp.graphicalProperties = GraphicalProperties()
        dp.graphicalProperties.noFill = True
        dp.graphicalProperties.line = LineProperties(solidFill=dp_color, w=50800)
        co_pts.append(dp)
    s_co.data_points = co_pts
    s_co.dLbls = DataLabelList()
    s_co.dLbls.showVal = True
    s_co.dLbls.showLegendKey = False
    s_co.dLbls.showCatName = False
    s_co.dLbls.showSerName = False
    s_co.dLbls.showPercent = False
    s_co.dLbls.numFmt = '0.0%'
    s_co.dLbls.dLblPos = 'inEnd'
    co_bar.y_axis.majorGridlines = None
    co_bar.x_axis.majorGridlines = None
    co_bar.legend = ChartLegend()
    co_bar.legend.position = 'b'
    co_bar.legend.overlay = False
    co_bar.graphical_properties = GraphicalProperties()
    co_bar.graphical_properties.solidFill = 'FFFFFF'
    co_bar.graphical_properties.line = LineProperties(noFill=True)
    co_bar.width = 20
    co_bar.height = 7.5
    anc_co = TwoCellAnchor()
    anc_co._from = AnchorMarker(col=5, colOff=0, row=r_nc + 16, rowOff=0)
    anc_co.to = AnchorMarker(col=9, colOff=0, row=r_nc + 28, rowOff=0)
    ws.add_chart(co_bar, anc_co)

    # ─── Footnotes ───
    r_fn = r_nc + 29
    ws.cell(row=r_fn, column=2, value=(
        "Deteriorated loans are those whose FICO scores have dropped "
        "two or more grades or have dropped to a Grade lower than D."
    )).font = FNT_A10
    ws.cell(row=r_fn + 1, column=2, value=(
        "Improved loans are those whose FICO scores have moved up "
        "at least one grade."
    )).font = FNT_A10

    # ─── Page Setup (match Vizo) ───
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25, header=0, footer=0)
    last_row = r_fn + 1
    ws.print_area = f'A1:{last_col_letter}{max(last_row, 66)}'
def _sheet_env_factor(wb, cu, snap, df, grades, config, hist):
    ws = wb.create_sheet("Env Factor by Pool")
    ed = config.get('economic_data', {})
    # Prefer economic data from the WARM file when available
    _imp = hist.get('impaired', {}) if hist else {}
    if _imp.get('economic_data'):
        ed = _imp['economic_data']
    econ_stress = _econ_stress(config)

    # Per-pool env ranges from WARM when available
    er = _imp.get('env_ranges', {})
    _ncc_r = er.get('ncc') or NCC_RANGES
    _dq_r  = er.get('dq')  or DQ_RANGES
    _es_r  = er.get('es')  or ES_RANGES

    # Pool ordering & DQ variance (same logic as Vizo). Pools not present
    # in WARM (e.g. non-risk-rated pools) are appended at the end.
    pools = _sort_pools(df['loan_pool'].unique(), config)
    warm_order = config.get('pool_order') or _imp.get('pool_order', [])
    pools = _merge_pool_orders(pools, warm_order)
    dq_pct = hist.get('dq_pct', {}) if hist else {}
    if not dq_pct:
        dq_pct = _imp.get('warm_dq_pct', {})
    dq_all_years = sorted(dq_pct.keys())
    acl_months_map = _imp.get('acl_months', {})
    snap_year = int(snap[:4])
    snap_month = int(snap[5:7])
    dq_var_map = {}
    for pool in pools:
        pool_acl = acl_months_map.get(pool, 36)
        abs_first = (snap_year * 12 + snap_month) - pool_acl + 1
        earliest_year = (abs_first - 1) // 12
        rates = [dq_pct.get(y, {}).get(pool, 0) for y in dq_all_years if y >= earliest_year]
        if len(rates) >= 2:
            avg = sum(rates) / len(rates)
            dq_var_map[pool] = rates[-1] - avg
        else:
            dq_var_map[pool] = 0

    risk_rated_map = _imp.get('risk_rated', {})
    ACCT = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'

    def _env_score(value, ranges):
        """Score without TCT _score's auto *100 conversion."""
        for lo, hi, s in ranges:
            if lo <= value < hi:
                return s
        return 0

    # Column widths (landscape, 8 columns matching Vizo layout)
    for ci, w in enumerate([22, 15.6, 18.7, 14.7, 15.3, 20.7, 11.7, 15.6], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Header rows ──
    ws['A1'] = cu
    ws['A1'].font = FNT_A14B
    ws['A2'] = "Environmental Factor for PLL"
    ws['A2'].font = FNT_A12B
    ws['A3'] = f"For Period Ending {_snap_display(snap)}"
    ws['A3'].font = FNT_A12B

    # ── Economic Stress Index section ── light slate fill
    ESI_FILL = PatternFill('solid', fgColor='D5DBDB')  # light slate
    ESI_HDR  = Font(name='Arial', bold=True, size=11)
    ESI_DATA = Font(name='Arial', size=11)
    ESI_DATA10 = Font(name='Arial', size=10)
    _esi_align_hdr = Alignment(horizontal='left', vertical='center', wrap_text=True)
    _esi_align_val = Alignment(horizontal='right', vertical='center')
    _esi_align_lbl = Alignment(horizontal='left', vertical='center')

    r = 5
    ws.cell(row=r, column=1, value="Economic Stress Index Calculation").font = FNT_A18B
    for ci in range(1, 6):
        ws.cell(row=r, column=ci).fill = ESI_FILL
    ws.row_dimensions[r].height = 23.25

    r = 6
    for ci, lbl in enumerate(["State", "Unemployment Rate", "Foreclosures per Person",
                               "Bankruptcies", "Population"], start=1):
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.font = ESI_HDR
        cell.fill = ESI_FILL
        cell.alignment = _esi_align_hdr
    ws.row_dimensions[r].height = 30.0

    r = 7
    pop = ed.get('population', 1)
    ws.cell(row=r, column=1, value=ed.get('state', '')).font = ESI_DATA10
    ws.cell(row=r, column=1).alignment = _esi_align_lbl
    ws.cell(row=r, column=2, value=ed.get('unemployment_rate', 0)).number_format = PCT
    ws.cell(row=r, column=2).font = ESI_DATA
    ws.cell(row=r, column=2).alignment = _esi_align_val
    ws.cell(row=r, column=3, value=ed.get('foreclosures', 0)).number_format = ACCT
    ws.cell(row=r, column=3).font = ESI_DATA
    ws.cell(row=r, column=3).alignment = _esi_align_val
    ws.cell(row=r, column=4, value=ed.get('bankruptcies', 0)).number_format = ACCT
    ws.cell(row=r, column=4).font = ESI_DATA
    ws.cell(row=r, column=4).alignment = _esi_align_val
    ws.cell(row=r, column=5, value=pop).number_format = ACCT
    ws.cell(row=r, column=5).font = ESI_DATA
    ws.cell(row=r, column=5).alignment = _esi_align_val
    for ci in range(1, 6):
        ws.cell(row=r, column=ci).fill = ESI_FILL

    r = 8
    for ci, lbl in enumerate(["County", "Unemployment Rate", "Bankruptcy Percentage",
                               "Foreclosure Percentage", "Economic Stress Index"], start=1):
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.font = ESI_HDR
        cell.fill = ESI_FILL
        cell.alignment = _esi_align_hdr
    ws.row_dimensions[r].height = 30.0

    r = 9
    ws.cell(row=r, column=1, value=ed.get('county', '')).font = ESI_DATA10
    ws.cell(row=r, column=1).alignment = _esi_align_lbl
    ws.cell(row=r, column=2, value=ed.get('unemployment_rate', 0)).number_format = PCT
    ws.cell(row=r, column=2).font = ESI_DATA
    ws.cell(row=r, column=2).alignment = _esi_align_val
    bk_pct = ed.get('bankruptcies', 0) / pop if pop else 0
    fc_pct = ed.get('foreclosures', 0) / pop if pop else 0
    ws.cell(row=r, column=3, value=bk_pct).number_format = PCT
    ws.cell(row=r, column=3).font = ESI_DATA
    ws.cell(row=r, column=3).alignment = _esi_align_val
    ws.cell(row=r, column=4, value=fc_pct).number_format = PCT
    ws.cell(row=r, column=4).font = ESI_DATA
    ws.cell(row=r, column=4).alignment = _esi_align_val
    ws.cell(row=r, column=5, value=econ_stress / 100).number_format = PCT
    ws.cell(row=r, column=5).font = ESI_DATA
    ws.cell(row=r, column=5).alignment = _esi_align_val
    for ci in range(1, 6):
        ws.cell(row=r, column=ci).fill = ESI_FILL

    # ── Per-pool environmental factors table ──
    r = 12
    pool_headers = ["Portfolio Segment", "Net Credit\nChange", "Net Credit\nScore",
                    "Delinquency\nVariance from Ave.", "Delinquency\nScore",
                    "Economic Stress\nActual", "Economic Stress\nScore",
                    "Environmental\nFactor"]
    for ci, lbl in enumerate(pool_headers, start=1):
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.font = RC_HDR_FNT
        cell.fill = RC_HDR_FILL
        cell.alignment = Alignment(horizontal='left', wrap_text=True)
        cell.border = THIN
    ws.row_dimensions[r].height = 47.25

    env_results = {}
    nrr = set(config.get('not_risk_rated', []))
    for pool in pools:
        r += 1
        pdf = df[df['loan_pool'] == pool]
        is_rr = risk_rated_map.get(pool, True) if risk_rated_map else (pool not in nrr)
        if is_rr:
            _, _, ncc_pct = _ncc(pdf, grades, config)
        else:
            ncc_pct = 0.0
        ncc_score = _env_score(ncc_pct * 100, _ncc_r) / 100.0
        dq_v = dq_var_map.get(pool, 0)
        dq_score_val = _env_score(dq_v * 100, _dq_r) / 100.0
        es_score = _env_score(econ_stress, _es_r) / 100.0
        env_f = ncc_score + dq_score_val + es_score
        env_results[pool] = env_f

        ws.cell(row=r, column=1, value=pool).font = FNT_A10
        ws.cell(row=r, column=2, value=ncc_pct).number_format = PCT
        ws.cell(row=r, column=2).font = FNT_A11
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3, value=ncc_score).number_format = PCT
        ws.cell(row=r, column=3).font = FNT_A11
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4, value=dq_v).number_format = PCT
        ws.cell(row=r, column=4).font = FNT_A11
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=5, value=dq_score_val).number_format = PCT
        ws.cell(row=r, column=5).font = FNT_A11
        ws.cell(row=r, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=6, value=econ_stress / 100).number_format = PCT
        ws.cell(row=r, column=6).font = FNT_A11
        ws.cell(row=r, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=7, value=es_score).number_format = PCT
        ws.cell(row=r, column=7).font = FNT_A11
        ws.cell(row=r, column=7).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=8, value=env_f).number_format = PCT
        ws.cell(row=r, column=8).font = FNT_A11
        ws.cell(row=r, column=8).alignment = Alignment(horizontal='center')

    # ── Footnotes: data sources ──
    sources = ed.get('_sources', {})
    if sources:
        fn_font = Font(name='Arial', size=8, italic=True, color='555555')
        r += 2
        ws.cell(row=r, column=1, value="Data Sources:").font = Font(
            name='Arial', size=8, bold=True, italic=True, color='555555')
        for field_label, source_key in [
            ("Unemployment Rate", "unemployment_rate"),
            ("Population", "population"),
            ("Bankruptcies", "bankruptcies"),
            ("Foreclosures", "foreclosures"),
        ]:
            src = sources.get(source_key)
            if src:
                r += 1
                ws.cell(row=r, column=1,
                        value=f"  {field_label}: {src}").font = fn_font

    # ── Page setup: landscape, fit to one page ──
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = 'landscape'
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25,
                                  header=0, footer=0)

    return env_results


# ── ACL Env by Pool Mgmt Adj ─────────────────────────────────────
def _sheet_acl_reserve(wb, cu, snap, df, grades, config, hist, env_results, spec_id_data=None):
    """ACL Env by Pool Mgmt Adj sheet matching Vizo template layout.

    Reads per-pool per-grade data from the WARM file when available
    (hist['impaired']['acl_pools']).  Falls back to computed values when
    WARM data is absent.  Hidden grades (Hide-*) are always excluded.
    """
    ws = wb.create_sheet("ACL Env by Pool Mgmt Adj")
    no_score = config.get('no_score_label', 'Not Reported')
    mgmt_adj_by_pool = config.get('mgmt_adj_by_pool', {})
    # Per-pool opt-in to the firm-wide default mgmt adj (wizard Step 16)
    # plus the firm-wide default value itself (Admin page).
    pool_use_default = _build_pool_use_default_map(config)
    admin_default_mgmt_adj = _load_admin_default_mgmt_adj()
    gl = _all_grades(grades, no_score)
    visible_gl = [g for g in gl if not _is_hidden(g)]

    # WARM-sourced ACL data
    _imp = hist.get('impaired', {}) if hist else {}
    econ_stress = _econ_stress(config)
    er = _imp.get('env_ranges', {})
    _ncc_r = er.get('ncc') or NCC_RANGES
    _dq_r  = er.get('dq')  or DQ_RANGES
    _es_r  = er.get('es')  or ES_RANGES

    def _env_score(value, ranges):
        for lo, hi, s in ranges:
            if lo <= value < hi:
                return s
        return 0

    pools = _sort_pools(df['loan_pool'].unique(), config)

    # Life loss per pool – matches Display Hist Bal formula
    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    avg_bals = hist.get('avg_balances', {}) if hist else {}
    years = hist.get('years', []) if hist else []
    acl_months_map = _imp.get('acl_months', {})
    snap_year = int(snap[:4])
    snap_month = int(snap[5:7])
    warm_net_co = _imp.get('warm_net_co', {})
    hbd = _imp.get('hist_bal_data', {})
    annual_grade_avg = {}
    for _pk, pdata in hbd.items():
        _dates = pdata.get('dates', [])
        _grades_data = pdata.get('grades', {})
        annual_grade_avg[_pk] = {}
        for _gk, _vals in _grades_data.items():
            if _gk.upper().startswith('HIDE'):
                continue
            yr_sums = {}
            yr_cnts = {}
            for _i, _d in enumerate(_dates):
                if _i < len(_vals) and _vals[_i] > 0:
                    yr_sums[_d.year] = yr_sums.get(_d.year, 0) + _vals[_i]
                    yr_cnts[_d.year] = yr_cnts.get(_d.year, 0) + 1
            for _y in yr_sums:
                annual_grade_avg[_pk].setdefault(_y, {})
                annual_grade_avg[_pk][_y][_gk] = yr_sums[_y] / yr_cnts[_y]

    life_loss = {}
    for pool in pools:
        pool_acl = acl_months_map.get(pool, 36)
        abs_first = (snap_year * 12 + snap_month) - pool_acl + 1
        pe = (abs_first - 1) // 12
        pa = annual_grade_avg.get(pool, {})
        yr_tots = []
        for y in years:
            if y < pe:
                continue
            yt = sum(pa.get(y, {}).values())
            if not yt:
                yt = avg_bals.get(y, {}).get(pool, 0)
            if yt:
                yr_tots.append(yt)
        avg_tot = sum(yr_tots) / len(yr_tots) if yr_tots else 0
        pool_stripped = pool.strip()
        net_co_match = warm_net_co.get(pool_stripped, warm_net_co.get(pool, None))
        if net_co_match is not None:
            total_net = net_co_match
        else:
            total_net = 0
            for y in years:
                if y < pe:
                    continue
                total_net += co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
        life_loss[pool] = total_net / avg_tot if avg_tot > 0 else 0

    # DQ variance per pool
    dq_pct = hist.get('dq_pct', {}) if hist else {}
    if not dq_pct:
        dq_pct = _imp.get('warm_dq_pct', {})
    dq_all_years = sorted(dq_pct.keys())
    acl_months_map = _imp.get('acl_months', {})
    snap_year = int(snap[:4])
    snap_month = int(snap[5:7])
    dq_var = {}
    for pool in pools:
        pool_acl = acl_months_map.get(pool, 36)
        abs_first = (snap_year * 12 + snap_month) - pool_acl + 1
        earliest_year = (abs_first - 1) // 12
        rates = [dq_pct.get(y, {}).get(pool, 0) for y in dq_all_years if y >= earliest_year]
        if len(rates) >= 2:
            avg = sum(rates) / len(rates)
            dq_var[pool] = rates[-1] - avg
        else:
            dq_var[pool] = 0

    acl_pools_data = _imp.get('acl_pools', {})
    acl_impaired = _imp.get('acl_impaired', {})
    acl_summary = _imp.get('acl_summary', {})
    spec_id_by_pool = _imp.get('spec_id_by_pool', {})
    risk_rated_flags = _imp.get('risk_rated', {})
    prior_mgmt_adj = _imp.get('prior_mgmt_adj', {})
    prior_env_factor = _imp.get('prior_env_factor', {})
    nrr = set(config.get('not_risk_rated', []))

    ACCT  = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
    ACCT2 = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'

    def _tct_grade_font(label):
        if _is_hidden(label):
            return FNT_A12R
        return FNT_A12

    # Column widths (landscape, 11 columns matching Vizo)
    for ci, w in enumerate([22, 18, 16, 18, 16, 16, 16, 18, 14, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws['A1'] = cu
    ws['A1'].font = FNT_A14B
    ws['A2'] = "Allowance & Provision for Credit Loss Reserve Analysis"
    ws['A2'].font = FNT_A12B
    ws['A3'] = f"For Period Ending {_snap_display(snap)}"
    ws['A3'].font = FNT_A12B

    headers = ["Current Grade", "Balance", "Specific\nIdentification",
               "Loan Loss Calc\nBalance", "ACL Base\nLoss Rate",
               "Mgmt\nAdj", "Allowance\nFactor",
               "Allowance before\nEnv", "Env\nFactor",
               "Env\n Allowance", "Total\n Allowance"]
    nhdr = len(headers)

    # Write column headers once at row 5
    r = 5
    for hi, h in enumerate(headers):
        cell = ws.cell(row=r, column=1 + hi, value=h)
        cell.font = RC_HDR_FNT
        cell.fill = RC_HDR_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN
    r += 1

    grand_allowance = 0
    grand_allow_before = 0
    grand_env_allow = 0
    pool_starts = []
    _bal_detail = _imp.get('pool_bal_detail', {})

    # Build unified pool list in WARM order, including WARM-only pools
    # AND any DB pools missing from WARM (e.g. non-risk-rated pools that
    # have no entry in the WARM ACL section but do appear in pool_order
    # or hist_bal_data). Wizard-edited config['pool_order'] wins over
    # WARM-derived order so user reordering / NRR pools always render.
    warm_order = config.get('pool_order') or _imp.get('pool_order', [])
    db_pools = set(df['loan_pool'].unique())
    hist_bal_pools = list((_imp.get('hist_bal_data') or {}).keys())
    all_acl_pools = _merge_pool_orders(
        list(db_pools),
        warm_order,
        extra=[list(acl_pools_data.keys()), hist_bal_pools],
    )
    if not all_acl_pools:
        all_acl_pools = pools

    for pool in all_acl_pools:
        pdf = df[df['loan_pool'] == pool]
        pool_total = pdf['current_balance'].sum()
        has_db_data = pool in db_pools

        # WARM data for this pool
        _pool_lc = pool.strip().lower()
        warm_pool = next((v for k, v in acl_pools_data.items()
                          if k.strip().lower() == _pool_lc), None)
        warm_grades = warm_pool['grades'] if warm_pool else {}
        warm_total = warm_pool['total'] if warm_pool else {}

        is_rr = risk_rated_flags.get(pool, has_db_data) if risk_rated_flags else (pool not in nrr)

        # Compute env factor
        if has_db_data:
            if is_rr:
                _, _, ncc_pct = _ncc(pdf, grades, config)
            else:
                ncc_pct = 0.0
            dq_v = dq_var.get(pool, 0)
            ncc_score = _env_score(ncc_pct * 100, _ncc_r)
            dq_score = _env_score(dq_v * 100, _dq_r)
            es_score = _env_score(econ_stress, _es_r)
            env_factor_calc = (ncc_score + dq_score + es_score) / 100.0
        else:
            env_factor_calc = 0
        # Use prior report's env factor if available; otherwise computed
        env_factor = prior_env_factor.get(pool, env_factor_calc)

        pool_ll = life_loss.get(pool, 0)

        # Pool header
        pool_start_row = r
        ws.cell(row=r, column=1, value=pool).font = FNT_A12B
        r += 1
        pool_starts.append(pool_start_row)

        if is_rr:
            # ── Risk-rated pool: show per-grade detail ──
            pool_allow_before = 0
            for gi, g in enumerate(visible_gl):
                fnt = _tct_grade_font(g)

                wg = warm_grades.get(g, {})
                # Prefer Pool_Balance Adjust source (balance_sheet_total) so
                # the two sheets always agree on the per-grade balance.
                _pd = _bal_detail.get(pool, {})
                _gd = _pd.get(g, {})
                bst = _gd.get('balance_sheet_total') if _gd else None
                if wg:
                    balance = bst if bst is not None else wg.get('balance', 0)
                    specific_id = wg.get('spec_id', 0)
                    if specific_id == 0 and pool in spec_id_by_pool:
                        specific_id = spec_id_by_pool[pool].get(g, 0)
                    calc_bal = balance - specific_id
                    # ACL Base Loss Rate is always the pure calculated value
                    # (pool life loss * distribution factor) so it matches
                    # column M on the Display Hist Bal tab — even if zero.
                    dist = _dist_factor(len(DIST_FACTORS) - 1) if g == no_score else _dist_factor(gi)
                    base_rate = max(0, pool_ll * dist)
                    # Mgmt adj is resolved from wizard config (Step 16
                    # overlay + per-pool 'Use Default' checkbox + Admin
                    # firm-wide default). The WARM workbook's baked-in
                    # mgmt_adj is intentionally ignored so unchecked
                    # pools do NOT receive the default. Admin default is
                    # only applied when base_rate==0 (no historical data).
                    mgmt_adj = _resolve_mgmt_adj_grade(
                        pool, g, gi, no_score,
                        pool_use_default, mgmt_adj_by_pool,
                        admin_default_mgmt_adj, prior_mgmt_adj,
                        base_rate=base_rate,
                    )
                    factor = base_rate + mgmt_adj
                    allow_before = calc_bal * factor
                elif has_db_data:
                    _pd = _bal_detail.get(pool, {})
                    _gd = _pd.get(g, {})
                    if _gd and _gd.get('balance_sheet_total', 0):
                        balance = _gd['balance_sheet_total']
                    else:
                        g_df = pdf[pdf['current_grade'] == g]
                        balance = g_df['current_balance'].sum()
                    specific_id = spec_id_by_pool.get(pool, {}).get(g, 0)
                    calc_bal = balance - specific_id
                    # ACL Base Loss Rate is always the pure calculated value
                    # (pool life loss * distribution factor) so it matches
                    # column M on the Display Hist Bal tab — even if zero.
                    dist = _dist_factor(len(DIST_FACTORS) - 1) if g == no_score else _dist_factor(gi)
                    base_rate = max(0, pool_ll * dist)
                    mgmt_adj = _resolve_mgmt_adj_grade(
                        pool, g, gi, no_score,
                        pool_use_default, mgmt_adj_by_pool,
                        admin_default_mgmt_adj, prior_mgmt_adj,
                        base_rate=base_rate,
                    )
                    factor = base_rate + mgmt_adj
                    allow_before = calc_bal * factor
                else:
                    balance = specific_id = calc_bal = 0
                    base_rate = mgmt_adj = factor = allow_before = 0
                pool_allow_before += allow_before

                ws.cell(row=r, column=1, value=g).font = fnt
                ws.cell(row=r, column=2, value=balance).number_format = ACCT
                ws.cell(row=r, column=2).font = fnt
                ws.cell(row=r, column=3, value=specific_id).number_format = ACCT2
                ws.cell(row=r, column=3).font = fnt
                ws.cell(row=r, column=4, value=calc_bal).number_format = ACCT
                ws.cell(row=r, column=4).font = fnt
                ws.cell(row=r, column=5, value=base_rate).number_format = PCT4
                ws.cell(row=r, column=5).font = fnt
                ws.cell(row=r, column=6, value=mgmt_adj).number_format = PCT4
                ws.cell(row=r, column=6).font = fnt
                ws.cell(row=r, column=7, value=factor).number_format = PCT4
                ws.cell(row=r, column=7).font = fnt
                ws.cell(row=r, column=8, value=allow_before).number_format = ACCT
                ws.cell(row=r, column=8).font = fnt
                for ci in range(1, nhdr + 1):
                    ws.cell(row=r, column=ci).border = THIN
                r += 1

            # Pool total row
            _ptd = _bal_detail.get(pool, {}).get('Total', {})
            if _ptd and _ptd.get('balance_sheet_total'):
                total_balance = _ptd['balance_sheet_total']
            elif warm_total:
                total_balance = warm_total.get('balance', pool_total)
            else:
                total_balance = pool_total
            pool_allow_before_out = pool_allow_before
            env_allow = pool_allow_before_out * env_factor
            total_allow = pool_allow_before_out + env_allow
            grand_allowance += total_allow
            grand_allow_before += pool_allow_before_out
            grand_env_allow += env_allow

            total_spec_id = warm_total.get('spec_id', 0) if warm_total else 0
            if total_spec_id == 0 and pool in spec_id_by_pool:
                total_spec_id = sum(spec_id_by_pool[pool].values())
            total_calc_bal = total_balance - total_spec_id

            ws.cell(row=r, column=1, value="Total").font = FNT_A12B
            ws.cell(row=r, column=2, value=total_balance).number_format = ACCT
            ws.cell(row=r, column=2).font = FNT_A12B
            ws.cell(row=r, column=3, value=total_spec_id).number_format = ACCT2
            ws.cell(row=r, column=3).font = FNT_A12B
            ws.cell(row=r, column=4, value=total_calc_bal).number_format = ACCT
            ws.cell(row=r, column=4).font = FNT_A12B
            ws.cell(row=r, column=8, value=pool_allow_before_out).number_format = ACCT
            ws.cell(row=r, column=8).font = FNT_A12B
            ws.cell(row=r, column=9, value=env_factor).number_format = PCT
            ws.cell(row=r, column=9).font = FNT_A12B
            ws.cell(row=r, column=10, value=env_allow).number_format = ACCT
            ws.cell(row=r, column=10).font = FNT_A12B
            ws.cell(row=r, column=11, value=total_allow).number_format = ACCT
            ws.cell(row=r, column=11).font = FNT_A12B
            for ci in range(1, nhdr + 1):
                ws.cell(row=r, column=ci).border = THIN
            r += 2
        else:
            # ── Non-risk-rated pool: show only Total row with rate columns ──
            _ptd_nrr = _bal_detail.get(pool, {}).get('Total', {})
            if _ptd_nrr and _ptd_nrr.get('balance_sheet_total') is not None:
                nrr_balance = _ptd_nrr['balance_sheet_total']
            else:
                nrr_balance = warm_total.get('balance', 0)
            nrr_spec_id = warm_total.get('spec_id', 0)
            if nrr_spec_id == 0 and pool in spec_id_by_pool:
                nrr_spec_id = sum(spec_id_by_pool[pool].values())
            nrr_calc_bal = nrr_balance - nrr_spec_id
            nrr_base_rate = warm_total.get('base_rate', 0)
            # NRR pool mgmt adj: same resolver as RR pools but pool-level
            # (no per-grade distribution since NRR pools have no grades).
            # Admin default only applied when nrr_base_rate==0.
            nrr_mgmt_adj = _resolve_mgmt_adj_total(
                pool, pool_use_default, mgmt_adj_by_pool,
                admin_default_mgmt_adj,
                base_rate=nrr_base_rate,
            )
            nrr_factor = nrr_base_rate + nrr_mgmt_adj
            # Recompute allow_before so it reflects the resolver's
            # mgmt_adj instead of the WARM workbook's baked-in value.
            nrr_allow_before = nrr_calc_bal * nrr_factor
            nrr_env_allow = nrr_allow_before * env_factor
            nrr_total_allow = nrr_allow_before + nrr_env_allow
            grand_allowance += nrr_total_allow
            grand_allow_before += nrr_allow_before
            grand_env_allow += nrr_env_allow

            ws.cell(row=r, column=1, value="Total").font = FNT_A12B
            ws.cell(row=r, column=2, value=nrr_balance).number_format = ACCT
            ws.cell(row=r, column=2).font = FNT_A12B
            ws.cell(row=r, column=3, value=nrr_spec_id).number_format = ACCT2
            ws.cell(row=r, column=3).font = FNT_A12B
            ws.cell(row=r, column=4, value=nrr_calc_bal).number_format = ACCT
            ws.cell(row=r, column=4).font = FNT_A12B
            ws.cell(row=r, column=5, value=nrr_base_rate).number_format = PCT4
            ws.cell(row=r, column=5).font = FNT_A12B
            ws.cell(row=r, column=6, value=nrr_mgmt_adj).number_format = PCT4
            ws.cell(row=r, column=6).font = FNT_A12B
            ws.cell(row=r, column=7, value=nrr_factor).number_format = PCT4
            ws.cell(row=r, column=7).font = FNT_A12B
            ws.cell(row=r, column=8, value=nrr_allow_before).number_format = ACCT
            ws.cell(row=r, column=8).font = FNT_A12B
            ws.cell(row=r, column=9, value=env_factor).number_format = PCT
            ws.cell(row=r, column=9).font = FNT_A12B
            ws.cell(row=r, column=10, value=nrr_env_allow).number_format = ACCT
            ws.cell(row=r, column=10).font = FNT_A12B
            ws.cell(row=r, column=11, value=nrr_total_allow).number_format = ACCT
            ws.cell(row=r, column=11).font = FNT_A12B
            for ci in range(1, nhdr + 1):
                ws.cell(row=r, column=ci).border = THIN
            r += 2

    # Grand totals
    pooled_balance = acl_summary.get('pooled_balance', df['current_balance'].sum())
    pooled_total_allow = grand_allowance

    pooled_spec_id = acl_summary.get('pooled_spec_id', 0)
    if pooled_spec_id == 0 and spec_id_by_pool:
        pooled_spec_id = sum(sum(g.values()) for g in spec_id_by_pool.values())
    pooled_calc_bal = pooled_balance - pooled_spec_id

    ws.cell(row=r, column=1, value="Pooled Totals").font = FNT_A12B
    ws.cell(row=r, column=2, value=pooled_balance).number_format = ACCT
    ws.cell(row=r, column=2).font = FNT_A12B
    ws.cell(row=r, column=3, value=pooled_spec_id).number_format = ACCT2
    ws.cell(row=r, column=3).font = FNT_A12B
    ws.cell(row=r, column=4, value=pooled_calc_bal).number_format = ACCT
    ws.cell(row=r, column=4).font = FNT_A12B
    ws.cell(row=r, column=8, value=grand_allow_before).number_format = ACCT
    ws.cell(row=r, column=8).font = FNT_A12B
    ws.cell(row=r, column=10, value=grand_env_allow).number_format = ACCT
    ws.cell(row=r, column=10).font = FNT_A12B
    ws.cell(row=r, column=11, value=pooled_total_allow).number_format = ACCT
    ws.cell(row=r, column=11).font = FNT_A12B

    r += 2
    ws.cell(row=r, column=1, value="Impaired Loans").font = FNT_A12B
    ws.cell(row=r, column=10, value="Allowance").font = FNT_A12B
    for lbl in ["Delinquent Loans", "Known Losses", "Repossessions",
                "Foreclosed Real Estate", "Deceased", "Bankruptcy"]:
        imp_val = acl_impaired.get(lbl, 0)
        if lbl.upper().startswith('HIDE'):
            continue
        r += 1
        ws.cell(row=r, column=1, value=lbl).font = FNT_A12
        ws.cell(row=r, column=11, value=imp_val).number_format = ACCT
    total_spec_allow = acl_summary.get('total_spec_allow', sum(acl_impaired.values()))
    total_allow_needed = pooled_total_allow + total_spec_allow
    acl_bal = acl_summary.get('acl_balance', config.get('acl_balance', 0))
    adjustment = total_allow_needed - acl_bal

    r += 1
    ws.cell(row=r, column=1, value="Total Specifically Identified Allowance").font = FNT_A12B
    ws.cell(row=r, column=11, value=total_spec_allow).number_format = ACCT
    ws.cell(row=r, column=11).font = FNT_A12B
    r += 1
    ws.cell(row=r, column=1, value="Total Allowance Needed").font = FNT_A12B
    ws.cell(row=r, column=11, value=total_allow_needed).number_format = ACCT
    ws.cell(row=r, column=11).font = FNT_A12B
    r += 1
    ws.cell(row=r, column=1, value=f"Allowance for Credit Loss Balance as of {_snap_display(snap)}").font = FNT_A12
    ws.cell(row=r, column=11, value=acl_bal).number_format = ACCT
    r += 1
    ws.cell(row=r, column=1, value="Adjustment (Overfunded)").font = FNT_A12B
    ws.cell(row=r, column=11, value=adjustment).number_format = ACCT
    ws.cell(row=r, column=11).font = FNT_A12B

    # ─── Page Setup ───
    rows_per_first_page = 77
    rows_per_next_page = 53

    pool_blocks = []
    for i, ps in enumerate(pool_starts):
        if i + 1 < len(pool_starts):
            end = pool_starts[i + 1] - 2
        else:
            end = pool_starts[-1] + len(visible_gl) + 1
        pool_blocks.append((ps, end))

    page_end = rows_per_first_page
    for ps, pe in pool_blocks:
        if pe <= page_end:
            continue
        # Pool overflows current page – break just before it
        ws.row_breaks.append(Break(id=ps - 1))
        page_end = ps - 1 + rows_per_next_page

    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25,
                                  header=0, footer=0)
    ws.print_area = f'A1:K{r}'
    ws.print_title_rows = '1:5'

    return grand_allowance


# ── Pool_Balance Adjust (was FAS 114 & Balance Adj) ───────────────
def _sheet_pool_balance_adjust(wb, cu, snap, df, grades, config, hist=None):
    """Pool_Balance Adjust – portrait, TCT color scheme, no hidden grades."""
    ws = wb.create_sheet("Pool_Balance Adjust")
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g for g in _all_grades(grades, no_score) if not _is_hidden(g)]
    pools = _sort_pools(df['loan_pool'].unique(), config)
    warm_order = config.get('pool_order') or (hist or {}).get('impaired', {}).get('pool_order', [])
    pools = _merge_pool_orders(pools, warm_order)

    # Column widths (4 columns: Grade, Loan Report Balance, Bal Adjustment, Balance Sheet Total)
    for ci, w in enumerate([23, 32, 21, 19], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.merge_cells('A2:D2')
    ws['A2'] = cu
    ws['A2'].font = FNT_A12B
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:D3')
    ws['A3'] = f"For Period Ending {_snap_display(snap)}"
    ws['A3'].font = FNT_A12B
    ws['A3'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A4:D4')
    ws['A4'] = "Balance Adjustment Detail"
    ws['A4'].font = FNT_A10B
    ws['A4'].alignment = Alignment(horizontal='center')

    # Column headers at row 6
    r = 6
    for ci, lbl in [(1, "Current Grade"), (2, "Loan Report Balance"),
                    (3, "Bal Adjustment"), (4, "Balance Sheet Total")]:
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.font = FNT_A10B
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN

    ACCT = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'

    # WARM balance-adjustment data
    _imp = (hist or {}).get('impaired', {})
    # Try hist-based pool adjustments first, then config
    bal_adj_pool = _imp.get('balance_adjustments', {})
    bal_adj_cfg = config.get('balance_adjustments', {})
    _detail = _imp.get('pool_bal_detail', {})

    r = 7
    pool_start_rows = []

    for pool in pools:
        pdf = df[df['loan_pool'] == pool]
        pdata = _detail.get(pool, {})
        pool_start_rows.append(r)

        # Pool header row with dark slate fill
        ws.cell(row=r, column=1, value=pool).font = RC_HDR_FNT
        for ci in range(1, 5):
            ws.cell(row=r, column=ci).fill = RC_HDR_FILL
        r += 1

        pool_loan_bal = 0
        pool_adj = 0
        pool_bst = 0

        for g in gl:
            gd = pdata.get(g, {})
            if gd:
                loan_bal = gd.get('loan_report_bal', 0)
                g_adj = gd.get('bal_adj', 0)
                bst = gd.get('balance_sheet_total', 0)
            else:
                g_df = pdf[pdf['current_grade'] == g]
                loan_bal = g_df['current_balance'].sum() if not g_df.empty else 0
                g_adj = 0
                bst = loan_bal

            fnt = FNT_A10
            ws.cell(row=r, column=1, value=g).font = fnt
            ws.cell(row=r, column=2, value=loan_bal).number_format = ACCT
            ws.cell(row=r, column=2).font = fnt
            ws.cell(row=r, column=3, value=g_adj).number_format = ACCT
            ws.cell(row=r, column=3).font = fnt
            ws.cell(row=r, column=4, value=bst).number_format = ACCT
            ws.cell(row=r, column=4).font = fnt

            pool_loan_bal += loan_bal
            pool_adj += g_adj
            pool_bst += bst
            r += 1

        # Use pool-level total from detail if available
        td = pdata.get('Total', {})
        if td:
            pool_loan_bal = td.get('loan_report_bal', pool_loan_bal)
            pool_adj = td.get('bal_adj', pool_adj)
            pool_bst = td.get('balance_sheet_total', pool_bst)

        # Pool total row
        ws.cell(row=r, column=1, value="Total").font = FNT_A10B
        ws.cell(row=r, column=2, value=pool_loan_bal).number_format = ACCT
        ws.cell(row=r, column=2).font = FNT_A10B
        ws.cell(row=r, column=3, value=pool_adj).number_format = ACCT
        ws.cell(row=r, column=3).font = FNT_A10B
        ws.cell(row=r, column=4, value=pool_bst).number_format = ACCT
        ws.cell(row=r, column=4).font = FNT_A10B
        r += 2

    # ─── Page Setup ───
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5,
                                  header=0.3, footer=0.05)

    # Smart page breaks between pool groupings
    ROWS_PER_PAGE = 48
    last_data_row = r - 2
    if last_data_row > ROWS_PER_PAGE:
        page_end = ROWS_PER_PAGE
        while page_end < last_data_row:
            best = None
            for ps in pool_start_rows:
                if ps - 1 <= 1:
                    continue
                if ps - 1 <= page_end:
                    best = ps - 1
            if best and best > 1:
                ws.row_breaks.append(Break(id=best))
                page_end = best + ROWS_PER_PAGE
            else:
                break


# ── Envir Fact Ranges ─────────────────────────────────────────────
def _sheet_env_ranges(wb, cu, snap, hist):
    """Environmental Factor Ranges – portrait, TCT color scheme, WARM-sourced ranges."""
    ws = wb.create_sheet("Envir Fact Ranges")

    # Column widths: 3 pairs of (Range, Score) across 6 used columns
    for ci, w in enumerate([19, 10, 19, 10, 19, 10], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.merge_cells('A1:F1')
    ws['A1'] = cu
    ws['A1'].font = FNT_A14B
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = "Environmental Factor Ranges"
    ws['A2'].font = FNT_A12B
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:F3')
    ws['A3'] = f"For Period Ending {_snap_display(snap)}"
    ws['A3'].font = FNT_A10B
    ws['A3'].alignment = Alignment(horizontal='center')

    # Section headers with dark slate fill
    r = 5
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1, value="Net Credit Change").font = RC_HDR_FNT
    ws.cell(row=r, column=1).fill = RC_HDR_FILL
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=2).fill = RC_HDR_FILL

    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(row=r, column=3, value="Delinquency").font = RC_HDR_FNT
    ws.cell(row=r, column=3).fill = RC_HDR_FILL
    ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=4).fill = RC_HDR_FILL

    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    ws.cell(row=r, column=5, value="Economic Stress Score").font = RC_HDR_FNT
    ws.cell(row=r, column=5).fill = RC_HDR_FILL
    ws.cell(row=r, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=6).fill = RC_HDR_FILL

    r = 6
    for ci, lbl in [(1, "Range"), (2, "Score"), (3, "Range"), (4, "Score"),
                    (5, "Range"), (6, "Score")]:
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.font = RC_HDR_FNT
        cell.fill = RC_HDR_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN

    # Get ranges from WARM data when available
    _imp = hist.get('impaired', {}) if hist else {}
    er = _imp.get('env_ranges', {})
    ncc_ranges = er.get('ncc') or NCC_RANGES
    dq_ranges  = er.get('dq')  or DQ_RANGES
    es_ranges  = er.get('es')  or ES_RANGES
    ncc_labels = er.get('ncc_labels', [])
    dq_labels  = er.get('dq_labels', [])
    es_labels  = er.get('es_labels', [])

    def _range_label(lo, hi):
        if lo <= -9999:
            return f"<{hi:.2f}%"
        if hi >= 9999:
            return f">{lo:.2f}%"
        return f"{lo:.2f}% to {hi:.2f}%"

    def _build_display(ranges, labels):
        result = []
        for idx, (lo, hi, s) in enumerate(ranges):
            lbl = labels[idx] if idx < len(labels) else _range_label(lo, hi)
            result.append((lbl, s / 100.0))
        return result

    ncc_data = _build_display(ncc_ranges, ncc_labels)
    dq_data  = _build_display(dq_ranges, dq_labels)
    es_data  = _build_display(es_ranges, es_labels)

    r = 7
    for i, (rng_str, sc) in enumerate(ncc_data):
        ws.cell(row=r + i, column=1, value=rng_str).font = FNT_A10
        ws.cell(row=r + i, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=r + i, column=2, value=sc).number_format = PCT
        ws.cell(row=r + i, column=2).font = FNT_A10
        ws.cell(row=r + i, column=2).alignment = Alignment(horizontal='center')
    for i, (rng_str, sc) in enumerate(dq_data):
        ws.cell(row=r + i, column=3, value=rng_str).font = FNT_A10
        ws.cell(row=r + i, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r + i, column=4, value=sc).number_format = PCT
        ws.cell(row=r + i, column=4).font = FNT_A10
        ws.cell(row=r + i, column=4).alignment = Alignment(horizontal='center')
    for i, (rng_str, sc) in enumerate(es_data):
        ws.cell(row=r + i, column=5, value=rng_str).font = FNT_A10
        ws.cell(row=r + i, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=r + i, column=6, value=sc).number_format = PCT
        ws.cell(row=r + i, column=6).font = FNT_A10
        ws.cell(row=r + i, column=6).alignment = Alignment(horizontal='center')

    # Description section
    desc_r = r + max(len(ncc_data), len(dq_data), len(es_data)) + 2
    ws.cell(row=desc_r, column=1, value="Description").font = FNT_A12B
    desc_r += 2
    ws.merge_cells(start_row=desc_r, start_column=1, end_row=desc_r, end_column=6)
    ws.cell(row=desc_r, column=1).value = (
        "The Environmental Factor combines three distinct data sets to measure the overall "
        "impact of migration trends, delinquency patterns, and local economic conditions on "
        "the credit union's expected credit losses."
    )
    ws.cell(row=desc_r, column=1).font = FNT_A10
    ws.cell(row=desc_r, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[desc_r].height = 93.0

    # Page setup: portrait, fit to one page
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25,
                                  header=0, footer=0)


# ── Grade Ranges & Loan Codes ────────────────────────────────────
def _sheet_grade_config(wb, grades, config):
    """Grade Ranges & Loan Codes – portrait, TCT color scheme, no hidden grades."""
    ws = wb.create_sheet("Grade Ranges & Loan Codes")
    no_score = config.get('no_score_label', 'Not Reported')

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 14

    ws.merge_cells('A1:C1')
    ws['A1'] = "Credit Grade Configuration"
    ws['A1'].font = FNT_A14B
    ws['A1'].alignment = Alignment(horizontal='center')

    r = 3
    for hi, h in enumerate(["Grade", "Score Range", "Reserve Rate"]):
        cell = ws.cell(row=r, column=1 + hi, value=h)
        cell.font = RC_HDR_FNT
        cell.fill = RC_HDR_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN
    r += 1
    for g in grades:
        ws.cell(row=r, column=1, value=g['label']).font = FNT_A10
        ws.cell(row=r, column=2, value=f"{g['min_score']}-{g['max_score']}").font = FNT_A10
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3, value=g['reserve_rate']).number_format = PCT
        ws.cell(row=r, column=3).font = FNT_A10
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN
        r += 1

    # No-score row
    ns_label = no_score
    ns_grade = next((g for g in grades if g['label'] == ns_label), None)
    if ns_grade:
        pass  # already included above
    # Hidden grades intentionally excluded

    r += 2
    ws.cell(row=r, column=1, value="Loan Type Codes").font = FNT_A14B
    r += 1

    # Build sorted list of (code, pool) pairs
    codes = sorted(config.get('pool_map', {}).items(), key=lambda x: (x[1], x[0]))
    n = len(codes)

    # Determine number of column-groups: each group is (Code, Pool) = 2 cols
    # Target ~20 rows per column; minimum 1 group, max 3
    import math
    max_rows = 20
    num_groups = min(3, max(1, math.ceil(n / max_rows)))
    rows_per_group = math.ceil(n / num_groups)

    # Set column widths for each group (pairs at cols 1-2, 4-5, 7-8 with gap cols 3, 6)
    for gi in range(num_groups):
        base = 1 + gi * 3
        ws.column_dimensions[get_column_letter(base)].width = 10
        ws.column_dimensions[get_column_letter(base + 1)].width = 18
        if gi < num_groups - 1:
            ws.column_dimensions[get_column_letter(base + 2)].width = 3

    # Write header row for each group
    hdr_row = r
    for gi in range(num_groups):
        base = 1 + gi * 3
        for hi, h in enumerate(["Code", "Loan Pool"]):
            cell = ws.cell(row=hdr_row, column=base + hi, value=h)
            cell.font = RC_HDR_FNT
            cell.fill = RC_HDR_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN
    r += 1

    # Write data in column-groups
    for gi in range(num_groups):
        base = 1 + gi * 3
        start_idx = gi * rows_per_group
        end_idx = min(start_idx + rows_per_group, n)
        for i, (code, pool) in enumerate(codes[start_idx:end_idx]):
            row = r + i
            ws.cell(row=row, column=base, value=str(code)).font = FNT_A10
            ws.cell(row=row, column=base).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=base).border = THIN
            ws.cell(row=row, column=base + 1, value=pool).font = FNT_A10
            ws.cell(row=row, column=base + 1).border = THIN

    # ── Non-Risk-Rated Pools (no loan codes – sourced from WARM/external) ──
    nrr_pools = list(config.get('not_risk_rated', []) or [])
    mapped_pools = {p for _, p in codes}
    nrr_pools = [p for p in nrr_pools if p not in mapped_pools]
    if nrr_pools:
        # Place below the code table
        r2 = r + rows_per_group + 2
        ws.cell(row=r2, column=1, value="Non-Risk-Rated Pools (no loan codes)").font = FNT_A12B
        r2 += 1
        for p in nrr_pools:
            ws.cell(row=r2, column=1, value=p).font = FNT_A10
            ws.cell(row=r2, column=1).border = THIN
            r2 += 1

    # Page setup: portrait, fit to one page
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25,
                                  header=0, footer=0)


# ══════════════════════════════════════════════════════════════════
# DISPLAY HIST BAL – Loss Factor Calculation
# ══════════════════════════════════════════════════════════════════

def _sheet_display_hist_bal(wb, cu, snap, df, grades, config, hist):
    """Display HIst Bal – Loss Factor Calculation (TCT colour scheme)."""
    from openpyxl.worksheet.properties import PageSetupProperties

    ws = wb.create_sheet("Display HIst Bal")
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g for g in _all_grades(grades, no_score) if not _is_hidden(g)]

    # Use WARM pool order (same as ACL tab), fallback to config sort.
    # Include pools known to WARM (hist_bal_data / pool_bal_detail) even
    # if they have no DB rows so NRR pools render alongside RR pools.
    _imp = hist.get('impaired', {}) if hist else {}
    warm_order = config.get('pool_order') or _imp.get('pool_order', [])
    db_pools = set(df['loan_pool'].unique())
    extra_pools = set((_imp.get('hist_bal_data') or {}).keys()) \
                  | set((_imp.get('pool_bal_detail') or {}).keys())
    all_known = db_pools | extra_pools
    if warm_order:
        pools = [p for p in warm_order
                 if not p.upper().startswith('HIDE') and p != 'Exclude'
                 and str(p).strip().lower() not in ('grand total','total','excluded')
                 and p in all_known]
        # Append any WARM-known pool not in warm_order
        for p in sorted(all_known - set(pools)):
            if str(p).strip().lower() in ('grand total','total','excluded','exclude'):
                continue
            if str(p).upper().startswith('HIDE'):
                continue
            pools.append(p)
    else:
        pools = _sort_pools(all_known, config)

    co_data  = hist.get('chargeoffs', {}) if hist else {}
    rc_data  = hist.get('recoveries', {}) if hist else {}
    avg_bals = hist.get('avg_balances', {}) if hist else {}
    years    = hist.get('years', []) if hist else []

    acl_months_map = (hist.get('impaired', {}).get('acl_months', {})
                      if hist else {})
    snap_year  = int(snap[:4])
    snap_month = int(snap[5:7])

    # Trim leading years to the deepest pool's Life-of-Loan window.
    # Years older than the longest LoL across all pools aren't used by
    # any pool's calculation and just clutter the year axis.
    if pools and years:
        _max_lol = max(acl_months_map.get(p, 36) for p in pools)
        _abs_first = (snap_year * 12 + snap_month) - _max_lol + 1
        _cutoff_year = (_abs_first - 1) // 12
        years = [y for y in years if y >= _cutoff_year]

    # ── Per-grade annual averages from WARM hist_bal_data ──
    hbd = hist.get('impaired', {}).get('hist_bal_data', {}) if hist else {}
    annual_grade_avg = {}
    for _pk, pdata in hbd.items():
        _dates = pdata.get('dates', [])
        _grades_data = pdata.get('grades', {})
        annual_grade_avg[_pk] = {}
        for _gk, _vals in _grades_data.items():
            if _gk.upper().startswith('HIDE'):
                continue
            yr_sums, yr_cnts = {}, {}
            for _i, _d in enumerate(_dates):
                if _i < len(_vals) and _vals[_i] > 0:
                    yr_sums[_d.year] = yr_sums.get(_d.year, 0) + _vals[_i]
                    yr_cnts[_d.year] = yr_cnts.get(_d.year, 0) + 1
            for _y in yr_sums:
                annual_grade_avg[_pk].setdefault(_y, {})
                annual_grade_avg[_pk][_y][_gk] = yr_sums[_y] / yr_cnts[_y]

    risk_rated_map = (hist.get('impaired', {}).get('risk_rated', {})
                      if hist else {})

    def _pool_earliest_year(pool):
        pool_acl = acl_months_map.get(pool, 36)
        abs_first = (snap_year * 12 + snap_month) - pool_acl + 1
        return (abs_first - 1) // 12

    # ── Pre-compute per-pool Life Loss Rate (matches WARM formula) ──
    # WARM: Life Loss Rate = Total Net Chargeoffs / Average of yearly pool totals
    # Prefer WARM's own net CO totals when available (from Display CO-Recov -DQ)
    warm_net_co = _imp.get('warm_net_co', {})
    pool_life_rates = {}
    pool_avg_totals = {}
    for pool in pools:
        pe = _pool_earliest_year(pool)
        pa = annual_grade_avg.get(pool, {})
        yr_tots = []
        for y in years:
            if y < pe:
                continue
            yt = sum(pa.get(y, {}).values())
            if not yt:
                yt = avg_bals.get(y, {}).get(pool, 0)
            if yt:
                yr_tots.append(yt)
        avg_tot = sum(yr_tots) / len(yr_tots) if yr_tots else 0
        pool_avg_totals[pool] = avg_tot
        # Use WARM's net CO total if available, else compute from loaded data
        pool_stripped = pool.strip()
        net_co_match = warm_net_co.get(pool_stripped, warm_net_co.get(pool, None))
        if net_co_match is not None:
            total_net = net_co_match
        else:
            total_net = 0
            for y in years:
                if y < pe:
                    continue
                total_net += co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
        pool_life_rates[pool] = total_net / avg_tot if avg_tot > 0 else 0

    # ── Title block ──
    ws['A1'] = cu
    ws['A1'].font = FNT_A14B
    ws.row_dimensions[1].height = 18.75
    ws['A2'] = "Loss Factor Calculation"
    ws['A2'].font = FNT_A12B
    ws['A3'] = f"For Quarter Ending {_snap_display(snap)}"
    ws['A3'].font = FNT_A12B

    year_strs  = [str(y) for y in years]
    num_years  = len(years)
    year_start = 2                      # column B
    avg_col    = year_start + num_years  # after year columns
    right_start = avg_col + 1            # right-side block

    # ── Column widths ──
    # A=Grade/Pool, B..=years, next=Avg Bal, then rates+WARM
    widths = [22] + [15] * num_years + [16, 12, 13, 14, 11, 10]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Header row (row 5) ──
    r = 5
    last_col = right_start + 4
    for c in range(1, last_col + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = RC_HDR_FNT
        cell.fill = RC_HDR_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN

    ws.cell(row=r, column=1, value="Current Grade")
    for yi, ys in enumerate(year_strs):
        lbl = f"YTD {ys}" if yi == num_years - 1 else ys
        ws.cell(row=r, column=year_start + yi, value=lbl)
    ws.cell(row=r, column=avg_col, value="Average Balance")

    ws.cell(row=r, column=right_start, value="Life Loss Rate")
    ws.cell(row=r, column=right_start + 1, value="Distribution Factor")
    ws.cell(row=r, column=right_start + 2, value="ACL Base Loss Rate")
    ws.cell(row=r, column=right_start + 3, value="% of Loans")
    ws.cell(row=r, column=right_start + 4, value="WARM Months")
    ws.row_dimensions[r].height = 36.0

    pool_ranges = []

    for pool in pools:
        r += 1
        pool_start_r = r
        ws.cell(row=r, column=1, value=pool).font = FNT_A14B
        ws.row_dimensions[r].height = 18.75

        pdf = df[df['loan_pool'] == pool]
        pool_total = pdf['current_balance'].sum()
        pool_ll = pool_life_rates.get(pool, 0)
        pool_earliest = _pool_earliest_year(pool)
        is_rr = risk_rated_map.get(pool, True)

        if not is_rr:
            # Non-risk-rated: total row only
            r += 1
            ws.cell(row=r, column=1, value="Total").font = FNT_A12B
            pool_annual = annual_grade_avg.get(pool, {})
            for yi in range(num_years):
                if years[yi] < pool_earliest:
                    continue
                yr_total = sum(pool_annual.get(years[yi], {}).values())
                if not yr_total:
                    yr_total = avg_bals.get(years[yi], {}).get(pool, 0)
                if yr_total:
                    c = ws.cell(row=r, column=year_start + yi, value=yr_total)
                    c.number_format = ACCT_FMT
                    c.font = FNT_A12B
            ws.cell(row=r, column=avg_col, value=pool_avg_totals.get(pool, 0)).number_format = ACCT_FMT
            ws.cell(row=r, column=avg_col).font = FNT_A12B

            ws.cell(row=r, column=right_start, value=pool_ll).number_format = PCT
            ws.cell(row=r, column=right_start).font = FNT_A12B
            ws.cell(row=r, column=right_start + 3, value=1.0).number_format = PCT
            ws.cell(row=r, column=right_start + 3).font = FNT_A12B
            warm = acl_months_map.get(pool,
                       config.get('warm_months', {}).get(pool, 36))
            ws.cell(row=r, column=right_start + 4, value=warm).font = FNT_A12
            ws.cell(row=r, column=right_start + 4).alignment = Alignment(horizontal='center')
            pool_ranges.append((pool_start_r, r))
            r += 1
            ws.row_dimensions[r].height = 6.0
            continue

        # Risk-rated: per-grade rows
        for gi, g in enumerate(gl):
            r += 1
            fnt = FNT_A12
            g_df = pdf[pdf['current_grade'] == g]
            balance = g_df['current_balance'].sum()

            ws.cell(row=r, column=1, value=g).font = fnt

            pool_annual = annual_grade_avg.get(pool, {})
            yr_vals = []
            for yi in range(num_years):
                if years[yi] < pool_earliest:
                    continue
                grade_avg = pool_annual.get(years[yi], {}).get(g, 0)
                if not grade_avg:
                    avg = avg_bals.get(years[yi], {}).get(pool, 0)
                    grade_avg = avg * (balance / pool_total) if pool_total and avg else 0
                if grade_avg:
                    yr_vals.append(grade_avg)
                    c = ws.cell(row=r, column=year_start + yi, value=grade_avg)
                    c.number_format = ACCT_FMT
                    c.font = fnt

            avg_bal = sum(yr_vals) / len(yr_vals) if yr_vals else 0
            ws.cell(row=r, column=avg_col, value=avg_bal).number_format = ACCT_FMT
            ws.cell(row=r, column=avg_col).font = fnt

            # Not Reported uses last DIST_FACTORS entry (skipping hidden grades)
            dist = _dist_factor(len(DIST_FACTORS) - 1) if g == no_score else _dist_factor(gi)
            base_rate = max(0, pool_ll * dist)
            pct_pool = balance / pool_total if pool_total else 0

            ws.cell(row=r, column=right_start, value=pool_ll).number_format = PCT
            ws.cell(row=r, column=right_start).font = fnt
            ws.cell(row=r, column=right_start + 1, value=dist).number_format = PCT
            ws.cell(row=r, column=right_start + 1).font = fnt
            ws.cell(row=r, column=right_start + 2, value=base_rate).number_format = PCT
            ws.cell(row=r, column=right_start + 2).font = fnt
            ws.cell(row=r, column=right_start + 3, value=pct_pool).number_format = PCT
            ws.cell(row=r, column=right_start + 3).font = fnt

            if gi == 0:
                warm = acl_months_map.get(pool,
                           config.get('warm_months', {}).get(pool, 36))
                ws.cell(row=r, column=right_start + 4, value=warm).font = FNT_A12
                ws.cell(row=r, column=right_start + 4).alignment = Alignment(horizontal='center')

        # Pool total row
        r += 1
        ws.cell(row=r, column=1, value="Total").font = FNT_A12B
        pool_annual = annual_grade_avg.get(pool, {})
        for yi in range(num_years):
            if years[yi] < pool_earliest:
                continue
            yr_total = sum(pool_annual.get(years[yi], {}).values())
            if yr_total:
                c = ws.cell(row=r, column=year_start + yi, value=yr_total)
                c.number_format = ACCT_FMT
                c.font = FNT_A12B
        ws.cell(row=r, column=avg_col, value=pool_avg_totals.get(pool, 0)).number_format = ACCT_FMT
        ws.cell(row=r, column=avg_col).font = FNT_A12B
        ws.cell(row=r, column=right_start, value=pool_ll).number_format = PCT
        ws.cell(row=r, column=right_start).font = FNT_A12B
        ws.cell(row=r, column=right_start + 3, value=1.0).number_format = PCT
        ws.cell(row=r, column=right_start + 3).font = FNT_A12B
        pool_ranges.append((pool_start_r, r))

        r += 1
        ws.row_dimensions[r].height = 6.0

    # ── Page Setup ──
    def _rh(row):
        h = ws.row_dimensions[row].height
        return h if h else 15.0

    PAGE_PT = 800.0    # pre-scaling; keeps pools from splitting across pages
    title_ht = sum(_rh(i) for i in range(1, 6))
    page_ht = title_ht

    for ps, pe in pool_ranges:
        spacer_row = pe + 1
        block_ht = sum(_rh(i) for i in range(ps, spacer_row + 1))
        if page_ht + block_ht > PAGE_PT and page_ht > title_ht:
            ws.row_breaks.append(Break(id=ps - 1))
            page_ht = title_ht + block_ht
        else:
            page_ht += block_ht

    last_col_ltr = get_column_letter(last_col)
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25,
                                  header=0.3, footer=0.3)
    ws.print_area = f'A1:{last_col_ltr}{r}'
    ws.print_title_rows = '1:5'


# ══════════════════════════════════════════════════════════════════
# DISPLAY CO-RECOV-DQ
# ══════════════════════════════════════════════════════════════════

def _sheet_co_recov_dq(wb, cu, snap, df, config, hist):
    """Display CO-Recov-DQ sheet (TCT colour scheme)."""
    from openpyxl.worksheet.properties import PageSetupProperties

    ws = wb.create_sheet("Display CO-Recov-DQ")

    # Use WARM pool order (same as ACL tab), fallback to config sort.
    # Include WARM-only pools so NRR pools render here too.
    _imp = hist.get('impaired', {}) if hist else {}
    warm_order = config.get('pool_order') or _imp.get('pool_order', [])
    db_pools = set(df['loan_pool'].unique())
    extra_pools = set((_imp.get('hist_bal_data') or {}).keys()) \
                  | set((_imp.get('pool_bal_detail') or {}).keys())
    all_known = db_pools | extra_pools
    if warm_order:
        pools = [p for p in warm_order
                 if not p.upper().startswith('HIDE') and p != 'Exclude'
                 and str(p).strip().lower() not in ('grand total','total','excluded')
                 and p in all_known]
        for p in sorted(all_known - set(pools)):
            if str(p).strip().lower() in ('grand total','total','excluded','exclude'):
                continue
            if str(p).upper().startswith('HIDE'):
                continue
            pools.append(p)
    else:
        pools = _sort_pools(all_known, config)
    co_data  = hist.get('chargeoffs', {}) if hist else {}
    rc_data  = hist.get('recoveries', {}) if hist else {}
    dq_pct   = hist.get('dq_pct', {}) if hist else {}
    years    = hist.get('years', []) if hist else []
    if not years:
        years = list(range(2019, int(snap[:4]) + 1))

    acl_months_map = (hist.get('impaired', {}).get('acl_months', {})
                      if hist else {})
    snap_year  = int(snap[:4])
    snap_month = int(snap[5:7])

    # Trim leading years to the deepest pool's Life-of-Loan window.
    # Years older than the longest LoL across all pools aren't used by
    # any pool's calculation and just clutter the year axis.
    if pools and years:
        _max_lol = max(acl_months_map.get(p, 36) for p in pools)
        _abs_first = (snap_year * 12 + snap_month) - _max_lol + 1
        _cutoff_year = (_abs_first - 1) // 12
        years = [y for y in years if y >= _cutoff_year]

    year_strs  = [str(y) for y in years]
    num_years  = len(years)

    # Prefer WARM-sourced CO/RC/Net data when available
    warm_co  = _imp.get('warm_co', {})
    warm_rc  = _imp.get('warm_rc', {})
    warm_net = _imp.get('warm_net', {})
    warm_co_totals = _imp.get('warm_co_totals', {})
    warm_rc_totals = _imp.get('warm_rc_totals', {})
    use_warm_co_rc = bool(warm_co)

    # Monthly CO/RC for partial-year windowing
    warm_co_monthly = _imp.get('warm_co_monthly', {})
    warm_rc_monthly = _imp.get('warm_rc_monthly', {})
    if not warm_co_monthly:
        warm_co_monthly = hist.get('co_monthly', {}) if hist else {}
    if not warm_rc_monthly:
        warm_rc_monthly = hist.get('rc_monthly', {}) if hist else {}

    def _pool_window_start(pool):
        """Return (earliest_year, earliest_month) for the WARM window."""
        pool_acl = acl_months_map.get(pool, 36)
        abs_first = (snap_year * 12 + snap_month) - pool_acl + 1
        ey = (abs_first - 1) // 12
        em = abs_first - ey * 12
        return ey, em

    def _pool_earliest_year(pool):
        return _pool_window_start(pool)[0]

    def _windowed_year_val(monthly_data, yearly_data, pool, year,
                           earliest_year, earliest_month):
        """Return the value for *year* trimmed to the WARM window."""
        if year != earliest_year:
            return yearly_data.get(year, {}).get(pool, 0)
        partial = 0
        has_monthly = False
        for m in range(earliest_month, 13):
            v = monthly_data.get((year, m), {}).get(pool, 0)
            if v:
                has_monthly = True
            partial += v
        if has_monthly:
            full_year = yearly_data.get(year, {}).get(pool, 0)
            if full_year and (full_year > 0) != (partial > 0):
                partial = -partial
            return partial
        full = yearly_data.get(year, {}).get(pool, 0)
        months_in_window = 12 - earliest_month + 1
        return full * months_in_window / 12 if full else 0

    # ── Column widths ──
    ws.column_dimensions['A'].width = 30.7
    for ci in range(2, num_years + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 16.4
    ws.column_dimensions[get_column_letter(num_years + 2)].width = 17.4
    ws.column_dimensions[get_column_letter(num_years + 3)].width = 11.7

    # ── Title block ──
    ws['A1'] = cu
    ws['A1'].font = FNT_A14B
    ws.row_dimensions[1].height = 18.75
    ws['A2'] = "Delinquency Calculation"
    ws['A2'].font = FNT_A12B
    ws['A3'] = f"For Quarter Ending {_snap_display(snap)}"
    ws['A3'].font = FNT_A12B

    ncol = num_years + 3

    def _section_hdr(row, labels):
        for ci, lbl in enumerate(labels, start=1):
            cell = ws.cell(row=row, column=ci, value=lbl)
            cell.font = RC_HDR_FNT
            cell.fill = RC_HDR_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = THIN

    section_ranges = []

    # ─── Charge offs ───
    r = 5
    co_start = r
    co_hdrs = ["Charge offs"] + year_strs[:]
    if year_strs:
        co_hdrs[-1] = f"YTD {year_strs[-1]}"
    co_hdrs += ["ACL Charge offs", "WARM Months"]
    _section_hdr(r, co_hdrs)
    ws.row_dimensions[r].height = 31.5

    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool).font = FNT_A12B
        earliest, earliest_mo = _pool_window_start(pool)
        acl_total = 0
        for yi, y in enumerate(years):
            if y < earliest:
                continue
            if use_warm_co_rc:
                val = _windowed_year_val(warm_co_monthly, warm_co,
                                         pool, y, earliest, earliest_mo)
            else:
                val = _windowed_year_val(hist.get('co_monthly', {}),
                                         co_data, pool, y, earliest, earliest_mo)
            ws.cell(row=r, column=2 + yi, value=val).number_format = ACCT_FMT
            ws.cell(row=r, column=2 + yi).font = FNT_A10B
            acl_total += val
        ws.cell(row=r, column=num_years + 2, value=acl_total).number_format = ACCT_FMT
        ws.cell(row=r, column=num_years + 2).font = FNT_A10B
        warm = acl_months_map.get(pool,
                   config.get('warm_months', {}).get(pool, 36))
        ws.cell(row=r, column=num_years + 3, value=warm).font = FNT_A10B
        ws.cell(row=r, column=num_years + 3).alignment = Alignment(horizontal='center')
        ws.row_dimensions[r].height = 15.75
    section_ranges.append((co_start, r))

    # ─── Recoveries ───
    r += 3
    rc_start = r
    rc_hdrs = ["Recoveries"] + year_strs[:]
    if year_strs:
        rc_hdrs[-1] = f"YTD {year_strs[-1]}"
    rc_hdrs += ["ACL Recoveries", "WARM Months"]
    _section_hdr(r, rc_hdrs)
    ws.row_dimensions[r].height = 32.25

    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool).font = FNT_A12B
        earliest, earliest_mo = _pool_window_start(pool)
        acl_total = 0
        for yi, y in enumerate(years):
            if y < earliest:
                continue
            if use_warm_co_rc:
                val = _windowed_year_val(warm_rc_monthly, warm_rc,
                                         pool, y, earliest, earliest_mo)
            else:
                val = _windowed_year_val(hist.get('rc_monthly', {}),
                                         rc_data, pool, y, earliest, earliest_mo)
            ws.cell(row=r, column=2 + yi, value=val).number_format = ACCT_FMT
            ws.cell(row=r, column=2 + yi).font = FNT_A10B
            acl_total += val
        ws.cell(row=r, column=num_years + 2, value=acl_total).number_format = ACCT_FMT
        ws.cell(row=r, column=num_years + 2).font = FNT_A10B
        warm = acl_months_map.get(pool,
                   config.get('warm_months', {}).get(pool, 36))
        ws.cell(row=r, column=num_years + 3, value=warm).font = FNT_A10B
        ws.cell(row=r, column=num_years + 3).alignment = Alignment(horizontal='center')
        ws.row_dimensions[r].height = 15.75
    section_ranges.append((rc_start, r))

    # ─── Net Loss ───
    r += 3
    nl_start = r
    nl_hdrs = ["Net Charge offs"] + year_strs[:]
    if year_strs:
        nl_hdrs[-1] = f"YTD {year_strs[-1]}"
    nl_hdrs += ["Net Charge offs", "WARM Months"]
    _section_hdr(r, nl_hdrs)

    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool).font = FNT_A12B
        earliest, earliest_mo = _pool_window_start(pool)
        acl_total = 0
        for yi, y in enumerate(years):
            if y < earliest:
                continue
            if use_warm_co_rc:
                co_val = _windowed_year_val(warm_co_monthly, warm_co,
                                            pool, y, earliest, earliest_mo)
                rc_val = _windowed_year_val(warm_rc_monthly, warm_rc,
                                            pool, y, earliest, earliest_mo)
                net = co_val - rc_val
            else:
                co_val = _windowed_year_val(hist.get('co_monthly', {}),
                                            co_data, pool, y, earliest, earliest_mo)
                rc_val = _windowed_year_val(hist.get('rc_monthly', {}),
                                            rc_data, pool, y, earliest, earliest_mo)
                net = co_val - rc_val
            ws.cell(row=r, column=2 + yi, value=net).number_format = ACCT_FMT
            ws.cell(row=r, column=2 + yi).font = FNT_A10B
            acl_total += net
        ws.cell(row=r, column=num_years + 2, value=acl_total).number_format = ACCT_FMT
        ws.cell(row=r, column=num_years + 2).font = FNT_A10B
        warm = acl_months_map.get(pool,
                   config.get('warm_months', {}).get(pool, 36))
        ws.cell(row=r, column=num_years + 3, value=warm).font = FNT_A10B
        ws.cell(row=r, column=num_years + 3).alignment = Alignment(horizontal='center')
        ws.row_dimensions[r].height = 15.75
    section_ranges.append((nl_start, r))

    # ─── Delinquency ───
    r += 3
    dq_start = r
    warm_dq = hist.get('impaired', {}).get('warm_dq_pct', {}) if hist else {}
    use_dq = warm_dq if warm_dq else dq_pct
    dq_hdrs = ["DQ %"] + year_strs[:] + ["Average", "Variance"]
    if year_strs:
        dq_hdrs[len(year_strs)] = f"YTD {year_strs[-1]}"
    _section_hdr(r, dq_hdrs)

    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool).font = FNT_A12B
        earliest = _pool_earliest_year(pool)
        rates = []
        for yi, y in enumerate(years):
            if y < earliest:
                continue
            val = use_dq.get(y, {}).get(pool, 0)
            ws.cell(row=r, column=2 + yi, value=val).number_format = PCT
            ws.cell(row=r, column=2 + yi).font = FNT_A10B
            rates.append(val)
        avg = sum(rates) / len(rates) if rates else 0
        ws.cell(row=r, column=num_years + 2, value=avg).number_format = PCT
        ws.cell(row=r, column=num_years + 2).font = FNT_A10B
        var = rates[-1] - avg if len(rates) > 1 else 0
        ws.cell(row=r, column=num_years + 3, value=var).number_format = PCT
        ws.cell(row=r, column=num_years + 3).font = FNT_A10B
    section_ranges.append((dq_start, r))

    # ── Page Setup ──
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25,
                                  header=0.3, footer=0.3)
    ws.print_title_rows = '1:3'

    # ── Row breaks between sections ──
    def _rh(row):
        h = ws.row_dimensions[row].height
        return h if h else 15.0

    PAGE_PT = 500.0
    title_ht = sum(_rh(i) for i in range(1, 4))
    page_ht = title_ht
    prev_end = 0

    for ss, se in section_ranges:
        gap_start = (section_ranges[0][0] if ss == section_ranges[0][0]
                     else prev_end + 1)
        block_ht = sum(_rh(i) for i in range(gap_start, se + 1))
        if page_ht + block_ht > PAGE_PT and page_ht > title_ht:
            ws.row_breaks.append(Break(id=gap_start - 1))
            page_ht = title_ht + block_ht
        else:
            page_ht += block_ht
        prev_end = se


# ──────────────────────────────────────────────────────────────────
# > Detail_HIst Balances
# ──────────────────────────────────────────────────────────────────

def _sheet_detail_hist_bal(wb, cu, snap, df, grades, config, hist):
    """Detail Hist Balances – monthly breakdown by pool by grade (TCT style).

    Uses grade-level data from the WARM's 'HIst Bal Data' tab when available.
    Only shows months within the life-of-loan (ACL months) for each pool.
    """
    from openpyxl.worksheet.properties import PageSetupProperties

    ws = wb.create_sheet("> Detail_HIst Balances")
    no_score = config.get('no_score_label', 'Not Reported')
    all_gl = [g for g in _all_grades(grades, no_score) if not _is_hidden(g)]

    # 8pt Arial fonts for compact layout
    F8B = Font(name='Arial', bold=True, size=8)
    F8  = Font(name='Arial', size=8)

    # Fills – TCT dark-slate scheme
    ALT_FILL = PatternFill('solid', fgColor='D5DADC')
    TOT_FILL = PatternFill('solid', fgColor='A9B2B5')

    def _grade_font8(label):
        if _is_hidden(label):
            return Font(name='Arial', size=8, color='FF0000')
        return F8

    impaired = hist.get('impaired', {}) if hist else {}
    hbd = impaired.get('hist_bal_data', {})
    warm_order = impaired.get('pool_order', [])
    acl_months_map = impaired.get('acl_months', {})
    risk_rated = impaired.get('risk_rated', {})

    db_pools = set(df['loan_pool'].unique())
    extra_pools = set((impaired.get('hist_bal_data') or {}).keys()) \
                  | set((impaired.get('pool_bal_detail') or {}).keys())
    all_known = db_pools | extra_pools
    if warm_order:
        pools = [p for p in warm_order
                 if not p.upper().startswith('HIDE') and p != 'Exclude'
                 and str(p).strip().lower() not in ('grand total','total','excluded')
                 and p in all_known]
        for p in sorted(all_known - set(pools)):
            if str(p).strip().lower() in ('grand total','total','excluded','exclude'):
                continue
            if str(p).upper().startswith('HIDE'):
                continue
            pools.append(p)
    else:
        pools = _sort_pools(all_known, config)

    all_labels = list(pools) + all_gl + ['Current Grade', 'Total']
    max_len = max((len(str(s)) for s in all_labels), default=20)
    col_a_width = max_len * 1.1 + 2
    ws.column_dimensions['A'].width = col_a_width

    ws['A1'] = cu
    ws['A1'].font = FNT_A14B
    ws['A2'] = "Loss Factor Historical Detail"
    ws['A2'].font = FNT_A12B
    ws['A3'] = f"For Quarter Ending {_snap_display(snap)}"
    ws['A3'].font = FNT_A12B
    ws.row_dimensions[4].height = 5.0

    if not hbd:
        pools_fb = pools
        r = 5
        for pool in pools_fb:
            ws.cell(row=r, column=1, value=pool).font = Font(name='Arial', bold=True, size=9)
            r += 1
            ws.cell(row=r, column=1, value="Current Grade").font = F8B
            ws.cell(row=r, column=2, value=snap).font = F8B
            r += 1
            pdf = df[df['loan_pool'] == pool]
            for g in all_gl:
                fnt = _grade_font8(g)
                ws.cell(row=r, column=1, value=g).font = fnt
                bal = pdf[pdf['current_grade'] == g]['current_balance'].sum()
                ws.cell(row=r, column=2, value=bal).number_format = ACCT_FMT
                ws.cell(row=r, column=2).font = fnt
                r += 1
            ws.cell(row=r, column=1, value="Total").font = F8B
            ws.cell(row=r, column=2, value=pdf['current_balance'].sum()).number_format = ACCT_FMT
            ws.cell(row=r, column=2).font = F8B
            r += 2
        return

    # ── Main path: use WARM hist_bal_data ──
    first_pool_data = next(iter(hbd.values()), {})
    all_dates = first_pool_data.get('dates', [])
    DATE_COL_START = 2

    r = 5
    pool_boundaries = []
    for pool in pools:
        pdata = hbd.get(pool, {})
        pdates = pdata.get('dates', all_dates)
        pgrades = pdata.get('grades', {})
        ptotal = pdata.get('total', [])

        pool_acl = acl_months_map.get(pool, len(pdates))
        if pool_acl < len(pdates):
            start_idx = len(pdates) - pool_acl
            pdates = pdates[start_idx:]
            pgrades = {g: v[start_idx:] for g, v in pgrades.items()}
            ptotal = ptotal[start_idx:]
        nmonths = len(pdates)

        for di in range(nmonths):
            col_letter = get_column_letter(DATE_COL_START + di)
            if ws.column_dimensions[col_letter].width < 9.71:
                ws.column_dimensions[col_letter].width = 9.71

        pct_col = DATE_COL_START + nmonths
        warm_col = pct_col + 1
        ws.column_dimensions[get_column_letter(pct_col)].width = 9.71
        ws.column_dimensions[get_column_letter(warm_col)].width = 9.71

        pool_last_total = ptotal[-1] if ptotal else 0

        ws.cell(row=r, column=1, value=pool).font = Font(name='Arial', bold=True, size=9)
        ws.row_dimensions[r].height = 13.5
        r += 1

        is_rr = risk_rated.get(pool, True)

        if not is_rr:
            ws.cell(row=r, column=1, value="Balance").font = F8B
            for di, dt in enumerate(pdates):
                c = ws.cell(row=r, column=DATE_COL_START + di, value=dt)
                c.number_format = 'mmm\\-yy'
                c.font = F8B
                c.alignment = Alignment(horizontal='center')
            c_warm = ws.cell(row=r, column=warm_col, value="WARM\nMonths")
            c_warm.font = F8B
            c_warm.alignment = Alignment(horizontal='center', wrap_text=True)
            r += 1

            ws.cell(row=r, column=1, value="Total").font = F8B
            ws.cell(row=r, column=1).fill = TOT_FILL
            for di in range(nmonths):
                v = ptotal[di] if di < len(ptotal) else 0
                c = ws.cell(row=r, column=DATE_COL_START + di, value=v)
                c.number_format = ACCT_FMT
                c.font = F8B
                c.fill = TOT_FILL
            pool_acl_val = acl_months_map.get(pool, '')
            ws.cell(row=r, column=warm_col, value=pool_acl_val).font = F8
            ws.cell(row=r, column=warm_col).alignment = Alignment(horizontal='center')
            r += 1

            ws.row_dimensions[r].height = 5.0
            pool_boundaries.append(r)
            r += 1
            continue

        # ── Risk-rated pool: full grade breakdown ──
        ws.cell(row=r, column=1, value="Current Grade").font = F8B
        for di, dt in enumerate(pdates):
            c = ws.cell(row=r, column=DATE_COL_START + di, value=dt)
            c.number_format = 'mmm\\-yy'
            c.font = F8B
            c.alignment = Alignment(horizontal='center')
        c_pct = ws.cell(row=r, column=pct_col, value="% of Loans")
        c_pct.font = F8B
        c_pct.alignment = Alignment(horizontal='center')
        c_warm = ws.cell(row=r, column=warm_col, value="WARM\nMonths")
        c_warm.font = F8B
        c_warm.alignment = Alignment(horizontal='center', wrap_text=True)
        r += 1

        grade_start = r
        for gi, g in enumerate(all_gl):
            fnt = _grade_font8(g)
            ws.cell(row=r, column=1, value=g).font = fnt
            row_fill = ALT_FILL if gi % 2 == 0 else None
            if row_fill:
                ws.cell(row=r, column=1).fill = row_fill
            vals = pgrades.get(g, [])
            for di in range(nmonths):
                v = vals[di] if di < len(vals) else 0
                c = ws.cell(row=r, column=DATE_COL_START + di, value=v)
                c.number_format = ACCT_FMT
                c.font = fnt
                if row_fill:
                    c.fill = row_fill
            last_val = vals[-1] if vals else 0
            pct_val = last_val / pool_last_total if pool_last_total else 0
            c_pct = ws.cell(row=r, column=pct_col, value=pct_val)
            c_pct.number_format = PCT
            c_pct.font = fnt
            if row_fill:
                c_pct.fill = row_fill
            r += 1

        pool_acl_val = acl_months_map.get(pool, '')
        ws.cell(row=grade_start, column=warm_col, value=pool_acl_val).font = F8
        ws.cell(row=grade_start, column=warm_col).alignment = Alignment(
            horizontal='center', vertical='center')
        if len(all_gl) > 0:
            ws.merge_cells(
                start_row=grade_start, start_column=warm_col,
                end_row=grade_start + len(all_gl), end_column=warm_col)

        ws.cell(row=r, column=1, value="Total").font = F8B
        ws.cell(row=r, column=1).fill = TOT_FILL
        for di in range(nmonths):
            v = ptotal[di] if di < len(ptotal) else 0
            c = ws.cell(row=r, column=DATE_COL_START + di, value=v)
            c.number_format = ACCT_FMT
            c.font = F8B
            c.fill = TOT_FILL
        c_pct = ws.cell(row=r, column=pct_col, value=1.0)
        c_pct.number_format = PCT
        c_pct.font = F8B
        c_pct.fill = TOT_FILL
        r += 1

        ws.row_dimensions[r].height = 5.0
        pool_boundaries.append(r)
        r += 1

    # Page breaks at pool boundaries
    FIT_PAGES_WIDE = 5
    PRINTABLE_W_PT = (11.0 - 0.25 - 0.25) * 72
    PRINTABLE_H_PT = (8.5 - 0.25 - 0.25) * 72
    CHAR_TO_PT = 7.0 * 72.0 / 96.0
    total_col_width_pt = 0
    for ci in range(1, ws.max_column + 1):
        letter = ws.cell(1, ci).column_letter
        cd = ws.column_dimensions.get(letter)
        w = cd.width if cd and cd.width else 8.43
        total_col_width_pt += w * CHAR_TO_PT
    pages_needed = total_col_width_pt / PRINTABLE_W_PT
    scale = min(1.0, FIT_PAGES_WIDE / pages_needed) if pages_needed > 0 else 1.0
    eff_page_ht = PRINTABLE_H_PT / scale
    DEFAULT_ROW_HT = 15.0

    page_start = 1
    last_data_row = r - 1
    while page_start <= last_data_row:
        cumulative = 0.0
        trigger_row = None
        for row_num in range(page_start, last_data_row + 1):
            dim = ws.row_dimensions.get(row_num)
            if dim and dim.height is not None:
                cumulative += dim.height
            else:
                cumulative += DEFAULT_ROW_HT
            if cumulative >= eff_page_ht:
                trigger_row = row_num
                break
        if trigger_row is None:
            break
        best = None
        for b in pool_boundaries:
            if page_start < b <= trigger_row:
                best = b
        if best is None:
            break
        ws.row_breaks.append(Break(id=best))
        page_start = best + 1

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 5
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.25
    ws.page_margins.footer = 0.25
    ws.print_title_cols = 'A:A'


# ──────────────────────────────────────────────────────────────────
# >Detail_Charge off Hist
# ──────────────────────────────────────────────────────────────────

def _sheet_detail_co_hist(wb, cu, snap, config, hist):
    """Detail Charge off Hist – monthly chargeoff/recovery data by pool (TCT style)."""
    from openpyxl.worksheet.properties import PageSetupProperties

    ws = wb.create_sheet(">Detail_Charge off Hist")

    F8  = Font(name='Arial', size=8)
    F8B = Font(name='Arial', bold=True, size=8)

    ALT_FILL = PatternFill('solid', fgColor='D5DADC')
    TOT_FILL = PatternFill('solid', fgColor='A9B2B5')

    impaired = hist.get('impaired', {}) if hist else {}
    hbd = impaired.get('hist_bal_data', {})
    warm_order = impaired.get('pool_order', [])
    acl_months_map = impaired.get('acl_months', {})

    # Prefer WARM "Charge off History" monthly data over file-parsed data
    co_monthly = impaired.get('warm_co_monthly', {}) or (hist.get('co_monthly', {}) if hist else {})
    rc_monthly = impaired.get('warm_rc_monthly', {}) or (hist.get('rc_monthly', {}) if hist else {})

    pools = warm_order if warm_order else sorted(set(config.get('pool_map', {}).values()))

    first_pool_data = next(iter(hbd.values()), {}) if hbd else {}
    all_dates = first_pool_data.get('dates', [])

    if not all_dates:
        all_ym = sorted(set(list(co_monthly.keys()) + list(rc_monthly.keys())))
        import datetime
        all_dates = [datetime.datetime(y, m, 1) for y, m in all_ym]

    all_labels = list(pools) + ['Charge offs', 'Recoveries', 'Net Loss',
                                'Total Charge offs', 'Total Recoveries', 'Total Net Loss']
    max_len = max((len(str(s)) for s in all_labels), default=20)
    col_a_width = max_len * 1.1 + 2
    ws.column_dimensions['A'].width = col_a_width

    DATE_COL_START = 2

    ws['A1'] = cu
    ws['A1'].font = FNT_A14B
    ws['A2'] = "Charge off and Recoveries Historical Detail"
    ws['A2'].font = FNT_A12B
    ws['A3'] = f"For Quarter Ending {_snap_display(snap)}"
    ws['A3'].font = FNT_A12B
    ws.row_dimensions[4].height = 5.0

    def _write_section(ws, start_row, section_label, monthly_data, pools,
                       all_dates, acl_months_map):
        r = start_row

        ws.cell(row=r, column=1, value=section_label).font = Font(
            name='Arial', bold=True, size=9)

        # Per-pool window slice based on Life-of-Loan (acl_months).
        # Missing entries default to 36 months — NOT len(all_dates), which
        # would force the global window to the full available history and
        # defeat the LoL trim for the deepest pool.
        max_months = 0
        pool_date_slices = {}
        for pool in pools:
            n = acl_months_map.get(pool, 36)
            n = min(n, len(all_dates))
            pool_date_slices[pool] = all_dates[-n:] if n > 0 else []
            if n > max_months:
                max_months = n

        if all_dates:
            header_dates = all_dates[-max_months:] if max_months > 0 else all_dates
        else:
            header_dates = []
        nmonths = len(header_dates)

        for di in range(nmonths):
            col_letter = get_column_letter(DATE_COL_START + di)
            ws.column_dimensions[col_letter].width = 9.71

        for di, dt in enumerate(header_dates):
            c = ws.cell(row=r, column=DATE_COL_START + di, value=dt)
            c.number_format = 'mmm\\-yy'
            c.font = F8B
            c.alignment = Alignment(horizontal='center')

        acl_col = DATE_COL_START + nmonths
        ws.cell(row=r, column=acl_col, value="ACL\nTotal").font = F8B
        ws.cell(row=r, column=acl_col).alignment = Alignment(
            horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(acl_col)].width = 9.71
        r += 1

        pool_values = {}
        for pi, pool in enumerate(pools):
            pdates = pool_date_slices[pool]
            pn = len(pdates)
            offset = nmonths - pn

            fnt = F8
            ws.cell(row=r, column=1, value=pool).font = fnt
            row_fill = ALT_FILL if pi % 2 == 0 else None
            if row_fill:
                ws.cell(row=r, column=1).fill = row_fill

            vals = []
            for di, dt in enumerate(pdates):
                ym = (dt.year, dt.month) if hasattr(dt, 'year') else (
                    dt.year(), dt.month())
                v = monthly_data.get(ym, {}).get(pool, 0) or 0
                vals.append(v)
                # Inside the pool's LoL window — always show a value
                # (including 0) so empty months are visible.
                c = ws.cell(row=r, column=DATE_COL_START + offset + di,
                            value=v)
                c.number_format = ACCT_FMT
                c.font = fnt
                if row_fill:
                    c.fill = row_fill

            pool_values[pool] = vals

            acl_total = sum(v for v in vals if isinstance(v, (int, float)))
            c = ws.cell(row=r, column=acl_col, value=acl_total if acl_total else '')
            if acl_total:
                c.number_format = ACCT_FMT
            c.font = fnt
            if row_fill:
                c.fill = row_fill

            # Pre-LoL cells (shorter-history pool): leave blank to
            # indicate these months are outside this pool's window.
            if row_fill:
                for di in range(offset):
                    ws.cell(row=r, column=DATE_COL_START + di).fill = row_fill

            r += 1

        ws.cell(row=r, column=1, value=f"Total {section_label}").font = F8B
        ws.cell(row=r, column=1).fill = TOT_FILL
        grand_total = 0
        for di in range(nmonths):
            dt = header_dates[di]
            ym = (dt.year, dt.month) if hasattr(dt, 'year') else (
                dt.year(), dt.month())
            # Sum only pools whose LoL window includes this month.
            ttl = 0
            for p in pools:
                pn = len(pool_date_slices[p])
                offset = nmonths - pn
                if di >= offset:
                    ttl += monthly_data.get(ym, {}).get(p, 0) or 0
            c = ws.cell(row=r, column=DATE_COL_START + di, value=ttl)
            c.number_format = ACCT_FMT
            c.font = F8B
            c.fill = TOT_FILL
            grand_total += ttl
        c = ws.cell(row=r, column=acl_col, value=grand_total if grand_total else '')
        if grand_total:
            c.number_format = ACCT_FMT
        c.font = F8B
        c.fill = TOT_FILL
        r += 1

        return r, pool_values, nmonths

    r = 5
    r, co_vals, nmonths = _write_section(
        ws, r, "Charge offs", co_monthly, pools,
        all_dates, acl_months_map)

    r += 1

    r, rc_vals, _ = _write_section(
        ws, r, "Recoveries", rc_monthly, pools,
        all_dates, acl_months_map)

    r += 1

    net_monthly = {}
    all_ym_keys = set(list(co_monthly.keys()) + list(rc_monthly.keys()))
    for ym in all_ym_keys:
        net_monthly[ym] = {}
        co_pools = co_monthly.get(ym, {})
        rc_pools = rc_monthly.get(ym, {})
        all_pool_keys = set(list(co_pools.keys()) + list(rc_pools.keys()))
        for p in all_pool_keys:
            net_monthly[ym][p] = co_pools.get(p, 0) - rc_pools.get(p, 0)

    r, _, _ = _write_section(
        ws, r, "Net Loss", net_monthly, pools,
        all_dates, acl_months_map)

    FIT_PAGES_WIDE = 5
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = FIT_PAGES_WIDE
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.25
    ws.page_margins.footer = 0.25
    ws.print_title_cols = 'A:A'


# ══════════════════════════════════════════════════════════════════
# MAIN COMPOSER
# ══════════════════════════════════════════════════════════════════

def compose_tct(client, snap, df, config, grades, hist=None):
    """
    Build complete TCT-format CECL-Migration-WARM workbook.

    Returns (workbook, filename).
    """
    cu = config['credit_union']
    no_score = config.get('no_score_label', 'Not Reported')
    pools = _sort_pools(df['loan_pool'].unique(), config)

    # Use WARM pool order when available; keep any pools missing from WARM
    # (e.g. non-risk-rated pools) appended at the end.
    _imp = (hist or {}).get('impaired', {})
    warm_order = config.get('pool_order') or _imp.get('pool_order', [])
    pools = _merge_pool_orders(pools, warm_order)

    wb = Workbook()

    # Cover
    _sheet_cover(wb, cu, snap)

    # Introduction
    _sheet_intro(wb, cu, snap)

    # Executive Summary (3 sheets)
    _sheet_exec_summary(wb, cu, snap, df, grades, config)

    # Risk Change by Credit Score (grand total matrix)
    _sheet_risk_change_all(wb, cu, snap, df, grades, config, hist)

    # Improved Deteriorated Summary
    _sheet_impdet_summary(wb, cu, snap, df, pools, grades, config, hist)

    # Historical Trends Balance
    _sheet_hist_balance_charts(wb, cu, snap, df, grades, config, hist)

    # Risk ChangeType per pool (numbered 01, 02, ...)
    # Skip non-risk-rated / single-line pools (no per-grade breakdown in
    # WARM) — honor BOTH the WARM-derived map AND any YAML override.
    rr_map = _imp.get('risk_rated', {})
    nrr_cfg = set(config.get('not_risk_rated', []))
    rc_pools = [p for p in pools
                if rr_map.get(p, True) and p not in nrr_cfg]
    for idx, pool in enumerate(rc_pools, start=1):
        pdf = df[df['loan_pool'] == pool]
        _sheet_pool_risk_change(wb, cu, snap, pdf, pool, idx, grades, config, hist)

    # Risk ChangeType Total Loans
    _sheet_pool_risk_change(wb, cu, snap, df, "Total Loans", 0, grades, config, hist)

    # Env Factor by Pool
    env_results = _sheet_env_factor(wb, cu, snap, df, grades, config, hist)

    # ACL Env by Pool Mgmt Adj
    _sheet_acl_reserve(wb, cu, snap, df, grades, config, hist, env_results)

    # Pool_Balance Adjust
    _sheet_pool_balance_adjust(wb, cu, snap, df, grades, config, hist)

    # Display HIst Bal
    _sheet_display_hist_bal(wb, cu, snap, df, grades, config, hist)

    # Display CO-Recov-DQ
    _sheet_co_recov_dq(wb, cu, snap, df, config, hist)

    # > Detail_HIst Balances
    _sheet_detail_hist_bal(wb, cu, snap, df, grades, config, hist)

    # >Detail_Charge off Hist
    _sheet_detail_co_hist(wb, cu, snap, config, hist)

    # Envir Fact Ranges
    _sheet_env_ranges(wb, cu, snap, hist)

    # Grade Ranges & Loan Codes
    _sheet_grade_config(wb, grades, config)

    safe_cu = cu.replace(' ', '_').replace('/', '-')
    fname = f"{snap}_CECL_Migration_{safe_cu}_TCT_Model.xlsx"
    return wb, fname
