"""
Vizo Report Builder – Produces main + supplemental Excel workbooks
matching the Vizo CECL Credit Migration template format.

Main Report tabs:
  Vizo Cover / Report Index / Introduction / Exec Summary /
  Impr Deter / Risk Change Total / Risk Change per-pool /
  ACL Env by Pool / Env Factor by Pool / Display Hist Bal /
  Display CO-Recov-DQ / Envir Fact Ranges

Supplemental Report tabs:
  Vizo Cover (2) / Report Index (2) / Historical Trends Balance /
  Detail Hist Balances / Detail Charge off Hist /
  Pool Balance Adjust / Appendix Supplemental
"""
import os, re
from io import BytesIO
from datetime import datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.chart import BarChart, DoughnutChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.series import DataPoint, Series, SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.title import Title
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RegularTextRun, Font as DrawingFont
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker

from cecl_engine import risk_change_matrix

# ══════════════════════════════════════════════════════════════════
# CONSTANTS & STYLING
# ══════════════════════════════════════════════════════════════════

_BASE = os.path.dirname(os.path.abspath(__file__))
# _WORKSPACE_BASE holds analyst-owned data (Sample Reports templates, etc.)
# which may live on a shared drive while the code lives in a local clone.
# Falls back to the code dir when CECL_WORKSPACE_ROOT is unset.
_WORKSPACE_BASE = os.environ.get('CECL_WORKSPACE_ROOT') or _BASE

# Path to the Vizo template containing the Vizo Color Theme 1 (theme1.xml)
VIZO_TEMPLATE_PATH = os.path.join(
    _WORKSPACE_BASE, 'Sample Reports',
    'YYYY-MM CECL-Migration-WARM - Template Credit Union with Vizo.xlsx')

_VIZO_THEME_BYTES = None

def _apply_vizo_theme(wb):
    """Replace the workbook's theme with the Vizo Color Theme 1.

    Without this, openpyxl writes Office's default theme so theme-color
    references (Accent 4, etc.) render in Office colors (purple) instead of
    the Vizo palette (teal).
    """
    global _VIZO_THEME_BYTES
    if _VIZO_THEME_BYTES is None:
        try:
            import zipfile
            with zipfile.ZipFile(VIZO_TEMPLATE_PATH) as z:
                _VIZO_THEME_BYTES = z.read('xl/theme/theme1.xml')
        except Exception:
            _VIZO_THEME_BYTES = b''
    if _VIZO_THEME_BYTES:
        wb.loaded_theme = _VIZO_THEME_BYTES

LOGO_VIZO = os.path.join(_BASE, 'logos', 'vizo_financial.png')
LOGO_TCT  = os.path.join(_BASE, 'logos', 'tct_risk_solutions.png')
ICON_INFO_DARKRED   = os.path.join(_WORKSPACE_BASE, 'Sample Reports', 'assets', 'info_darkred.png')
ICON_INFO_DARKGREEN = os.path.join(_WORKSPACE_BASE, 'Sample Reports', 'assets', 'info_darkgreen.png')

HIDDEN_GRADES = ['Hide-F', 'Hide-G', 'Hide-H', 'Hide-I']

# ── Calibri fonts (template standard) ────────────────────────────
V26   = Font(name='Calibri', size=26)
V26B  = Font(name='Calibri', bold=True, size=26)
V18B  = Font(name='Calibri', bold=True, size=18)
V14B  = Font(name='Calibri', bold=True, size=14)
V14   = Font(name='Calibri', size=14)
V12B  = Font(name='Calibri', bold=True, size=12)
V12   = Font(name='Calibri', size=12)
V11B  = Font(name='Calibri', bold=True, size=11)
V11   = Font(name='Calibri', size=11)
V10B  = Font(name='Calibri', bold=True, size=10)
V10   = Font(name='Calibri', size=10)
V8    = Font(name='Calibri', size=8)
V8B   = Font(name='Calibri', bold=True, size=8)

# ── Red fonts for hidden grades ──────────────────────────────────
V12R  = Font(name='Calibri', size=12, color='FF0000')
V12BR = Font(name='Calibri', bold=True, size=12, color='FF0000')
V10R  = Font(name='Calibri', size=10, color='FF0000')

# ── Header fills (Vizo Color Theme 1) ────────────────────────────
HDR_FILL = PatternFill('solid', fgColor='0D4D5E')   # accent1 teal
HDR_FONT = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
HDR11    = Font(name='Calibri', bold=True, size=11, color='FFFFFF')

IMP_FILL = PatternFill('solid', fgColor='829901')   # dk2 olive
DET_FILL = PatternFill('solid', fgColor='873A3A')   # lt2 maroon tint=0.25
ALT_FILL = PatternFill('solid', fgColor='DAEDEF')   # theme7 accent4 tint~0.8
TOT_FILL = PatternFill('solid', fgColor='B6DBDE')   # theme7 accent4 tint~0.6

THIN = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin'),
)

# ── Number formats ───────────────────────────────────────────────
ACCT    = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
ACCT2   = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
DOLLAR  = '"$"#,##0'
PCT     = '0.00%'
PCT4    = '0.0000%'

# ── Score range tables ───────────────────────────────────────────
NCC_RANGES = [
    (-999, -18.00, 7), (-18.00, -16.00, 6), (-16.00, -14.00, 5),
    (-14.00, -11.00, 4), (-11.00, -8.00, 3), (-8.00, -6.00, 2),
    (-6.00, -4.00, 1), (-4.00, 4.00, 0), (4.00, 6.00, -1),
    (6.00, 8.00, -2), (8.00, 9.00, -3), (9.00, 11.00, -4),
    (11.00, 13.00, -5), (13.00, 15.00, -6), (15.00, 999, -7),
]
DQ_RANGES = [
    (5.00, 999, 20), (4.00, 5.00, 17), (3.00, 4.00, 12),
    (2.50, 3.00, 8), (2.00, 2.50, 4), (1.50, 2.00, 2.5),
    (1.00, 1.50, 1.5), (0.50, 1.00, 0.75), (-0.50, 0.50, 0),
    (-1.00, -0.50, -0.75), (-1.50, -1.00, -1.5), (-2.00, -1.50, -2.5),
    (-2.50, -2.00, -4), (-3.00, -2.50, -8), (-4.00, -3.00, -12),
    (-5.00, -4.00, -17), (-999, -5.00, -20),
]
ES_RANGES = [
    (25.00, 999, 10), (24.00, 25.00, 8), (22.00, 24.00, 7),
    (20.00, 22.00, 6), (18.00, 20.00, 5), (16.00, 18.00, 4),
    (14.00, 16.00, 3.5), (12.00, 14.00, 3), (10.00, 12.00, 2),
    (8.00, 10.00, 1), (6.00, 8.00, 0), (4.00, 6.00, 0),
    (2.00, 4.00, -1), (0.00, 2.00, -2),
]
DIST_FACTORS = [10.52, 22.93, 45.15, 116.10, 141.17, 152.04, 160.21]


def _env_ranges(hist):
    """Return (ncc, dq, es) range tables from WARM data or hardcoded defaults."""
    _imp = hist.get('impaired', {}) if hist else {}
    er = _imp.get('env_ranges', {})
    return (
        er.get('ncc') or NCC_RANGES,
        er.get('dq') or DQ_RANGES,
        er.get('es') or ES_RANGES,
    )


def _ordered_pools(df, hist):
    """Return loan pools in WARM file order, falling back to alphabetical.
    Non-risk-rated pools are always placed at the end.

    Includes pools that exist in any data source: the loan DataFrame,
    WARM hist_bal_data, or pool_bal_detail (balance-sheet-total source).
    NRR pools that have no DB rows but DO appear in WARM/monthly-balance
    data are therefore preserved.
    """
    _imp = hist.get('impaired', {}) if hist else {}
    warm_order = _imp.get('pool_order', [])
    risk_rated = _imp.get('risk_rated', {})
    hbd_pools = set((_imp.get('hist_bal_data') or {}).keys())
    pbd_pools = set((_imp.get('pool_bal_detail') or {}).keys())
    df_pools = set(df['loan_pool'].unique()) if df is not None and len(df) else set()
    all_pools = df_pools | hbd_pools | pbd_pools

    def _ok(p):
        if not p:
            return False
        s = str(p).strip()
        if not s or s == 'Exclude' or s.upper().startswith('HIDE'):
            return False
        if s.lower() in ('grand total', 'total', 'excluded'):
            return False
        return True

    all_pools = {p for p in all_pools if _ok(p)}
    if warm_order:
        ordered = [p for p in warm_order if p in all_pools and _ok(p)]
        remainder = sorted(all_pools - set(ordered))
        ordered = ordered + remainder
    else:
        ordered = sorted(all_pools)
    # Push non-risk-rated pools to the end
    def _is_rr(p):
        return risk_rated.get(p, True)
    rr = [p for p in ordered if _is_rr(p)]
    nrr = [p for p in ordered if not _is_rr(p)]
    return rr + nrr


# ══════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════

def _all_grades(grades, no_score):
    """Return complete 11-grade list: visible + hidden + no_score."""
    visible = {g['label'] for g in grades}
    hidden = [h for h in HIDDEN_GRADES if h.replace('Hide-', '') not in visible]
    n_hidden = max(0, len(HIDDEN_GRADES) - (len(grades) - 5))
    return [g['label'] for g in grades] + hidden[:n_hidden] + [no_score]


def _is_hidden(label):
    return label.lower().startswith('hide') if label else False


def _grade_font(label, bold=False):
    """Return red font for hidden grades, normal otherwise."""
    if _is_hidden(label):
        return V12BR if bold else V12R
    return V12B if bold else V12


def _grade_font10(label, bold=False):
    """Return red font for hidden grades at 10pt."""
    if _is_hidden(label):
        return Font(name='Calibri', bold=bold, size=10, color='FF0000')
    return V10B if bold else V10


def _score(value, ranges):
    for lo, hi, s in ranges:
        if lo <= value < hi:
            return s
    return 0


def _dist_factor(idx):
    return DIST_FACTORS[min(idx, len(DIST_FACTORS) - 1)] / 100.0


# ── Admin-default + per-pool management-adjustment resolver ──────────
def _load_admin_default_mgmt_adj():
    """Read the firm-wide default management adjustment from
    ``admin_defaults.yaml``. Returns ``0.0`` on any error."""
    try:
        import yaml as _yaml
        from pathlib import Path as _Path
        p = _Path(__file__).resolve().parent / 'admin_defaults.yaml'
        if not p.exists():
            return 0.0
        data = _yaml.safe_load(p.read_text(encoding='utf-8')) or {}
        return float(data.get('default_mgmt_adj', 0.0) or 0.0)
    except Exception:
        return 0.0


def _build_pool_use_default_map(config):
    """Return ``{pool_name: bool}`` from ``config['pools']``."""
    out = {}
    for p in (config.get('pools') or []):
        name = (p.get('name') or '').strip()
        if name:
            out[name] = bool(p.get('use_default_mgmt_adj'))
    return out


def _resolve_mgmt_adj_grade(pool, grade_label, grade_idx, no_score_label,
                             pool_use_default, mgmt_adj_by_pool,
                             admin_default, prior_mgmt_adj_map,
                             base_rate=None):
    """Per-(pool, grade) mgmt adj resolver — mirrors report_tct.py.

    Precedence: prior report value > manual overlay×dist > admin
    default×dist (only when use_default AND no manual AND
    base_rate==0) > 0.
    """
    pm = prior_mgmt_adj_map.get(pool, {}) if prior_mgmt_adj_map else {}
    if grade_label in pm:
        return pm[grade_label]
    dist = (_dist_factor(len(DIST_FACTORS) - 1)
            if grade_label == no_score_label
            else _dist_factor(grade_idx))
    manual = mgmt_adj_by_pool.get(pool, 0) or 0
    if manual:
        return float(manual) * dist
    if (pool_use_default.get(pool, False)
            and admin_default
            and (base_rate is None or float(base_rate or 0) == 0)):
        return float(admin_default) * dist
    return 0.0


def _resolve_mgmt_adj_total(pool, pool_use_default, mgmt_adj_by_pool,
                             admin_default, base_rate=None):
    """Pool-level resolver for non-risk-rated pools."""
    manual = mgmt_adj_by_pool.get(pool, 0) or 0
    if manual:
        return float(manual)
    if (pool_use_default.get(pool, False)
            and admin_default
            and (base_rate is None or float(base_rate or 0) == 0)):
        return float(admin_default)
    return 0.0


def _snap_display(snap):
    try:
        dt = pd.to_datetime(snap)
        return dt.strftime('%m/%d/%Y')
    except Exception:
        return snap


def _snap_date(snap):
    try:
        return pd.to_datetime(snap)
    except Exception:
        return snap


def _ncc(pool_df, grades, config):
    """Net credit change from migration matrix (matches Risk Change sheets).
    Returns (improved_pct, deteriorated_pct, net_pct)."""
    total = pool_df['current_balance'].sum()
    if total == 0:
        return 0, 0, 0
    no_score = config.get('no_score_label', 'Not Reported')
    gl = _all_grades(grades, no_score)
    matrix = risk_change_matrix(pool_df, grades, no_score)
    n_top = config.get('top_grades_double_drop', 3)
    imp_bal = 0
    det_bal = 0
    for j, og in enumerate(gl):
        for i, g in enumerate(gl):
            v = _matrix_val(matrix, g, og)
            if i > j:
                if j < n_top and (i - j) < 2:
                    pass  # unchanged – small drop within top grades
                else:
                    det_bal += v
            elif i < j:
                imp_bal += v
    return imp_bal / total, det_bal / total, (imp_bal - det_bal) / total


def _eco_stress(config, ed_override=None):
    ed = ed_override if ed_override else config.get('economic_data', {})
    unemp = ed.get('unemployment_rate', 0) * 100
    pop = ed.get('population', 1)
    bk = (ed.get('bankruptcies', 0) / pop) * 100 if pop else 0
    fc = (ed.get('foreclosures', 0) / pop) * 100 if pop else 0
    return unemp + bk + fc


def _pool_life_loss(pools, hist):
    """Compute life loss rate per pool from historical data."""
    co = hist.get('chargeoffs', {}) if hist else {}
    rc = hist.get('recoveries', {}) if hist else {}
    ab = hist.get('avg_balances', {}) if hist else {}
    years = hist.get('years', []) if hist else []
    result = {}
    for pool in pools:
        rates = []
        for y in years:
            net = co.get(y, {}).get(pool, 0) - rc.get(y, {}).get(pool, 0)
            avg = ab.get(y, {}).get(pool, 0)
            if avg > 0:
                rates.append(net / avg)
        result[pool] = sum(rates) / len(rates) if rates else 0
    return result


def _pool_dq_variance(pools, hist, snap='2025-12-31'):
    """Compute delinquency variance per pool.

    Uses the same year-filtering as Display CO-Recov-DQ: each pool's
    DQ rates are limited to years within its ACL months window.
    """
    dq = hist.get('dq_pct', {}) if hist else {}
    if not dq:
        _imp = hist.get('impaired', {}) if hist else {}
        dq = _imp.get('warm_dq_pct', {})
    all_years = sorted(dq.keys())

    # Determine per-pool earliest year from ACL months
    _imp = hist.get('impaired', {}) if hist else {}
    acl_months_map = _imp.get('acl_months', {})
    snap_year = int(snap[:4])
    snap_month = int(snap[5:7])

    result = {}
    for pool in pools:
        pool_acl = acl_months_map.get(pool, 36)
        abs_first = (snap_year * 12 + snap_month) - pool_acl + 1
        earliest_year = (abs_first - 1) // 12
        rates = [dq.get(y, {}).get(pool, 0) for y in all_years if y >= earliest_year]
        if len(rates) >= 2:
            avg = sum(rates) / len(rates)
            result[pool] = rates[-1] - avg
        else:
            result[pool] = 0
    return result


def _hdr_row(ws, row, ncols, font=HDR_FONT, fill=HDR_FILL):
    """Apply header styling to a row."""
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN


def _grade_ranges(grades, no_score):
    """Build grade label -> score range string map."""
    rng = {}
    for g in grades:
        if g['max_score'] >= 900:
            rng[g['label']] = f"{g['min_score']}+"
        elif g['min_score'] <= 1:
            rng[g['label']] = f"{g['max_score']} or less"
        else:
            rng[g['label']] = f"{g['min_score']}-{g['max_score']}"
    for h in HIDDEN_GRADES:
        rng[h] = 'N/A'
    rng[no_score] = ''
    return rng


def _matrix_val(matrix, cur, orig):
    if _is_hidden(cur) or _is_hidden(orig):
        return 0
    if cur in matrix.index and orig in matrix.columns:
        return matrix.loc[cur, orig]
    return 0


def _compute_acl_totals(df, grades, config, hist, snap=''):
    """Compute the 4 CECL Adjustment summary values.

    Pulls Pooled Total Allowance, Specifically Identified Allowance, and
    ACL Balance from the existing working file data in hist['impaired'].
    Falls back to computing from loan data if working-file values unavailable.

    Returns dict with keys:
      spec_id_allowance, total_allowance_needed, acl_balance, adjustment
    and label keys:
      spec_id_label, needed_label, balance_label, adjustment_label
    """
    imp = hist.get('impaired', {}) if hist else {}

    # Specifically Identified Allowance — match the ACL Env by Pool Mgmt Adj
    # sheet exactly (report_vizo._sheet_acl_reserve / report_tct._sheet_acl_reserve):
    # prefer the parsed "Total Specifically Identified Allowance" cell from the
    # prior WARM workbook's ACL Env tab, then fall back to the sum of the
    # acl_impaired dict, then finally to total_spec_id from the Impaired Loans
    # tab parser. Keeping the same precedence here guarantees that the
    # "Total Allowance Needed" line on Impr Deter matches the same line on
    # ACL Env by Pool Mgmt Adj.
    _acl_sum = imp.get('acl_summary') or {}
    _acl_imp = imp.get('acl_impaired') or {}
    if 'total_spec_allow' in _acl_sum:
        spec_id = _acl_sum.get('total_spec_allow', 0)
    elif _acl_imp:
        spec_id = sum(_acl_imp.values())
    else:
        spec_id = imp.get('total_spec_id', 0)

    # Pooled Total Allowance — prefer the value the ACL Env by Pool Mgmt Adj
    # sheet just computed and stashed on hist['impaired'] (key set inside
    # _sheet_acl_reserve). Falls back to the prior WARM-parsed value, then
    # to a fresh inline computation. compose_vizo_main is responsible for
    # building the ACL Env sheet BEFORE the Impr Deter sheet so the stash
    # is populated; tab order is restored afterward via wb.move_sheet.
    if '_computed_pooled_total_allow' in imp:
        grand_allowance = imp['_computed_pooled_total_allow']
    elif 'pooled_total_allow' in _acl_sum:
        grand_allowance = _acl_sum['pooled_total_allow']
    elif 'pooled_total_allowance' in imp:
        grand_allowance = imp['pooled_total_allowance']
    else:
        # Fallback: compute from loan data
        no_score = config.get('no_score_label', 'Not Reported')
        gl = _all_grades(grades, no_score)
        _imp_ed = imp.get('economic_data') if imp else None
        econ_stress = _eco_stress(config, ed_override=_imp_ed)
        _ncc_r, _dq_r, _es_r = _env_ranges(hist)
        pools = _ordered_pools(df, hist)
        life_loss = _pool_life_loss(pools, hist)
        dq_var = _pool_dq_variance(pools, hist, snap)

        grand_allowance = 0
        _rr_map = imp.get('risk_rated', {}) if imp else {}
        for pool in pools:
            pdf = df[df['loan_pool'] == pool]
            if _rr_map.get(pool, True):
                _, _, ncc_pct = _ncc(pdf, grades, config)
            else:
                ncc_pct = 0.0
            dq_v = dq_var.get(pool, 0)
            ncc_score = _score(ncc_pct * 100, _ncc_r)
            dq_score = _score(dq_v * 100, _dq_r)
            es_score = _score(econ_stress, _es_r)
            env_factor = (ncc_score + dq_score + es_score) / 100.0
            pool_ll = life_loss.get(pool, 0)

            pool_allow_before = 0
            for gi, g in enumerate(gl):
                balance = pdf[pdf['current_grade'] == g]['current_balance'].sum() if not _is_hidden(g) else 0
                dist = _dist_factor(gi)
                base_rate = pool_ll * dist
                if base_rate == 0 and not _is_hidden(g):
                    base_rate = next((gr['reserve_rate'] for gr in grades if gr['label'] == g), 0.005)
                    if g == no_score:
                        base_rate = np.median([gr['reserve_rate'] for gr in grades])
                factor = base_rate  # mgmt_adj=0
                pool_allow_before += balance * factor

            env_allow = pool_allow_before * env_factor
            grand_allowance += pool_allow_before + env_allow

    total_needed = grand_allowance + spec_id

    # ACL Balance (from ACL Env by Pool Mgmt Adj tab, or config fallback)
    acl_balance = imp.get('acl_balance', config.get('acl_balance', 0))

    # Format the snap date for the balance label
    try:
        snap_str = _snap_display(snap)
    except Exception:
        snap_str = str(snap)
    return {
        'spec_id_allowance': spec_id,
        'total_allowance_needed': total_needed,
        'acl_balance': acl_balance,
        'adjustment': total_needed - acl_balance,
        'spec_id_label': 'Total Specifically Identified Allowance',
        'needed_label': 'Total Allowance Needed',
        'balance_label': f'Allowance for Credit Loss Balance as of {snap_str}',
        'adjustment_label': 'Adjustment (Overfunded)',
    }


# ══════════════════════════════════════════════════════════════════
# SHEET BUILDERS – MAIN REPORT
# ══════════════════════════════════════════════════════════════════

def _sheet_cover(wb, cu, snap, supplemental=False):
    """Cover page matching Vizo template with logos and fit-to-page."""
    ws = wb.active
    ws.title = "Vizo Cover" if not supplemental else "Vizo Cover (2)"

    # Column widths from template
    for col, w in [('A', 2.7), ('B', 8.7), ('C', 20.9), ('D', 51.3),
                   ('E', 8.7), ('F', 2.7)]:
        ws.column_dimensions[col].width = w

    # Row heights matching template
    ws.row_dimensions[5].height = 33.6
    ws.row_dimensions[6].height = 33.6
    ws.row_dimensions[13].height = 35.25
    ws.row_dimensions[14].height = 33.6
    ws.row_dimensions[15].height = 33.6
    ws.row_dimensions[16].height = 45.0
    ws.row_dimensions[17].height = 24.0
    ws.row_dimensions[39].height = 12.75

    # ── Vizo Financial logo (top centre, rows 3-12 in column D) ──
    if os.path.isfile(LOGO_VIZO):
        # Trim shadow on all sides of the source logo.
        if PILImage is not None:
            with PILImage.open(LOGO_VIZO) as _img:
                w, h = _img.size
                crop_left = max(1, int(w * 0.02))
                crop_top = max(1, int(h * 0.02))
                crop_right = max(1, int(w * 0.05))
                crop_bottom = max(1, int(h * 0.08))
                crop_box = (crop_left, crop_top,
                            max(crop_left + 1, w - crop_right),
                            max(crop_top + 1, h - crop_bottom))
                _cropped = _img.crop(crop_box)
                _buf = BytesIO()
                _cropped.save(_buf, format='PNG')
                _buf.seek(0)
                vizo_img = XlImage(_buf)
        else:
            vizo_img = XlImage(LOGO_VIZO)
        # Template anchor: from col=3(D) colOff=169545 row=3 rowOff=9525
        #                  to   col=3(D) colOff=1998345 row=12 rowOff=215265
        vizo_img.anchor = TwoCellAnchor(
            _from=AnchorMarker(col=3, colOff=169545, row=3, rowOff=9525),
            to=AnchorMarker(col=3, colOff=1998345, row=12, rowOff=215265),
        )
        ws.add_image(vizo_img)

    # ── TCT Risk Solutions logo (bottom, rows 40-42) ──
    if os.path.isfile(LOGO_TCT):
        tct_img = XlImage(LOGO_TCT)
        if supplemental:
            # Supplemental template: from col=0 row=38 to col=3 row=42
            tct_img.anchor = TwoCellAnchor(
                _from=AnchorMarker(col=0, colOff=1, row=38, rowOff=68582),
                to=AnchorMarker(col=3, colOff=1228726, row=42, rowOff=100698),
            )
        else:
            # Main template: from col=1 row=40 to col=2 row=42
            tct_img.anchor = TwoCellAnchor(
                _from=AnchorMarker(col=1, colOff=1, row=40, rowOff=68582),
                to=AnchorMarker(col=2, colOff=1228726, row=42, rowOff=100698),
            )
        ws.add_image(tct_img)

    # Row 14: "CECL Credit Migration Report" (merged A14:F14, s=26)
    ws.merge_cells('A14:F14')
    c = ws['A14']
    c.value = "CECL Credit Migration Report"
    c.font = V26
    c.alignment = Alignment(horizontal='center', wrap_text=True)

    if supplemental:
        ws.merge_cells('A15:F15')
        c2 = ws['A15']
        c2.value = "Supplemental Reports"
        c2.font = V26
        c2.alignment = Alignment(horizontal='center')

    # Row 16: CU name (merged A16:F16, bold s=26)
    ws.merge_cells('A16:F16')
    c = ws['A16']
    c.value = cu
    c.font = V26B
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[16].height = 45.0

    # Row 17: Date (merged A17:F17, bold s=14, date format)
    ws.merge_cells('A17:F17')
    c = ws['A17']
    c.value = _snap_date(snap)
    c.font = V14B
    c.number_format = 'mm-dd-yy'
    c.alignment = Alignment(horizontal='center', vertical='center')

    # Row 21: Disclaimer (merged B21:E31)
    ws.merge_cells('B21:E31')
    ws['B21'].value = (
        "The following analysis and all parts thereof (\u2018analysis\u2019) are based upon "
        "information obtained by Vizo Financial Corporate Credit Union (Vizo Financial) from "
        "the credit union that is subject of this analysis and other sources that Vizo Financial "
        "believes to be reliable and utilized in models using methods and assumptions which Vizo "
        "Financial believes to be reasonable.  However, actual performance compared to estimated "
        "performance of the subject credit union may be different and cannot be guaranteed.  This "
        "analysis is for informational purposes only and is intended only for the use of the "
        "subject credit union.  The analysis does not constitute either legal or tax advice. \n"
        "All reports are confidential."
    )
    ws['B21'].font = V10
    ws['B21'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Row 44: Copyright
    ws['B44'] = "\u00a9 {} TCT Risk Solutions".format(datetime.now().year)
    ws['B44'].font = V8

    # ── Page Setup: portrait, fit to one page ──
    ws.page_setup.orientation = 'portrait'
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    if supplemental:
        # Symmetric left/right margins so the cover content centers on the page.
        ws.page_margins = PageMargins(
            left=0.75, right=0.75, top=0.25, bottom=0.25, header=0.05, footer=0.05
        )
    else:
        ws.page_margins = PageMargins(
            left=0.75, right=0.2, top=0.25, bottom=0.25, header=0.05, footer=0.05
        )


def _sheet_report_index(wb, cu, snap, supplemental=False):
    """Report Index page matching Vizo template exactly."""
    tab_name = "Report Index" if not supplemental else "Report Index (2)"
    ws = wb.create_sheet(tab_name)

    # Theme-colored fonts matching template (theme=4 = accent blue, theme=1 = dark text)
    theme4_14b = Font(name='Calibri', bold=True, size=14, color='1B4F72')
    theme1_12  = Font(name='Calibri', size=12, color='000000')
    theme4_12b = Font(name='Calibri', bold=True, size=12, color='1B4F72')

    if not supplemental:
        # ── Column widths ──
        ws.column_dimensions['A'].width = 8.86

        # ── Row heights (all rows from template) ──
        ws.row_dimensions[1].height = 15.75
        ws.row_dimensions[2].height = 18.75
        ws.row_dimensions[3].height = 106.9
        ws.row_dimensions[4].height = 18.75
        ws.row_dimensions[5].height = 9.6
        ws.row_dimensions[6].height = 15.75
        ws.row_dimensions[7].height = 15.75
        ws.row_dimensions[8].height = 15.75
        ws.row_dimensions[9].height = 7.15
        ws.row_dimensions[10].height = 15.75
        ws.row_dimensions[11].height = 107.45
        ws.row_dimensions[12].height = 23.25
        ws.row_dimensions[13].height = 13.9

        # ── Merged ranges ──
        ws.merge_cells('A1:K1')
        ws.merge_cells('A3:J3')
        ws.merge_cells('A4:J4')
        ws.merge_cells('A6:J6')
        ws.merge_cells('A7:J7')
        ws.merge_cells('A8:J8')
        ws.merge_cells('A10:J10')
        ws.merge_cells('A11:J11')

        # ── Cell content ──
        # Row 2: "Report Overview"
        ws['A2'] = "Report Overview"
        ws['A2'].font = theme4_14b

        # Row 3: description paragraph
        ws['A3'] = (
            "The CECL Credit Migration Reports from TCT, Inc. presents a comprehensive "
            "picture of the changing nature of risk in the credit union\u2019s loan portfolio. "
            "Credit migration is measured by the improvement or deterioration of risk, "
            "measured by the credit score, from the date of loan funding to the most recent "
            "data pull.  New credit scores are typically pulled twice per year.  Migration "
            "may still be measured on a quarterly basis to take into account new loans and "
            "changing loan balances."
        )
        ws['A3'].font = theme1_12
        ws['A3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Row 4: "Report Index:"
        ws['A4'] = "Report Index:"
        ws['A4'].font = theme4_14b

        # Row 6: "Executive Summary" (indent=1)
        ws['A6'] = "Executive Summary"
        ws['A6'].font = theme4_12b
        ws['A6'].alignment = Alignment(horizontal='left', indent=1)

        # Row 7: sub-item (indent=2)
        ws['A7'] = "CECL Adjustment  & Improved/Deteriorated "
        ws['A7'].font = theme1_12
        ws['A7'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)

        # Row 8: sub-item (indent=2)
        ws['A8'] = "Improved & Deteriorated Loans Risk Change By Credit Score"
        ws['A8'].font = theme1_12
        ws['A8'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)

        # Row 10: "Detailed Reporting" (indent=1)
        ws['A10'] = "Detailed Reporting"
        ws['A10'].font = theme4_12b
        ws['A10'].alignment = Alignment(horizontal='left', indent=1)

        # Row 11: multi-line detail items (indent=2)
        ws['A11'] = (
            "Allowance & Provision for Credit Loss Reserve Analysis\n"
            "Risk Change by Credit Score - Total Loans\n"
            "Risk Change by Credit Score - Loan Pools\n"
            "Environmental Factor Provision for Loan Loss\n"
            "Loss Factor Calculation\n"
            "Delinquency Calculation"
        )
        ws['A11'].font = theme1_12
        ws['A11'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)

        # Row 12: supplemental reference
        ws['A12'] = "Additional detailed reporting located in the Supplemental Reporting Package"
        ws['A12'].font = theme1_12
        ws['A12'].alignment = Alignment(horizontal='left', vertical='center', indent=2)

        # ── Page setup: portrait, print area A1:J13 ──
        ws.page_setup.orientation = 'portrait'
        ws.print_area = 'A1:J13'
    else:
        # ══ Supplemental Report Index ══
        ws.column_dimensions['A'].width = 87.86
        ws.column_dimensions['B'].width = 8.86

        # Row heights
        ws.row_dimensions[1].height = 15.75
        ws.row_dimensions[2].height = 18.75
        ws.row_dimensions[3].height = 15.75
        ws.row_dimensions[4].height = 31.5
        ws.row_dimensions[5].height = 15.75
        ws.row_dimensions[6].height = 15.6
        ws.row_dimensions[7].height = 23.45
        ws.row_dimensions[8].height = 58.15
        ws.row_dimensions[9].height = 13.9

        # Row 2: "Report Overview"
        ws['A2'] = "Report Overview"
        ws['A2'].font = theme4_14b

        # Row 4: description
        ws['A4'] = (
            "The CECL Credit Migration Supplemental Reports from TCT, Inc. presents the "
            "historical details of the changing nature of risk in the credit union\u2019s "
            "loan portfolio."
        )
        ws['A4'].font = theme1_12
        ws['A4'].alignment = Alignment(vertical='center', wrap_text=True)

        # Row 6: "Report Index:"
        ws['A6'] = "Report Index:"
        ws['A6'].font = theme4_14b

        # Row 7: section header
        ws['A7'] = " Supplemental Reporting Package: "
        ws['A7'].font = theme4_12b
        ws['A7'].alignment = Alignment(horizontal='left')

        # Row 8: list items (indent=1)
        ws['A8'] = (
            "Historical Loan Balances by Credit Score\n"
            "Loss Factor Historical Detail\n"
            "Charge off and Recoveries Historical Detail\n"
            "Balance Adjustment Detail"
        )
        ws['A8'].font = theme1_12
        ws['A8'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

        # Page setup: portrait, fit everything on one page
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1


def _sheet_introduction(wb, cu, snap):
    """Introduction / Appendix page with methodology description."""
    ws = wb.create_sheet("Introduction")
    ws.column_dimensions['A'].width = 118.0

    ws['A1'] = "Appendix"
    ws['A1'].font = V12B

    ws['A2'] = "Credit Migration"
    ws['A2'].font = V12B

    ws.merge_cells('A3:J3')
    ws['A3'] = (
        "Credit Migration describes the movement of individual loans through the credit "
        "scoring system. Each loan is assigned a risk grade based on the borrower's credit "
        "score at origination and the most recent credit score. When the current score differs "
        "from the original score, the loan has \"migrated\" - either improving (higher score) "
        "or deteriorating (lower score). This migration forms the basis for assessing changes "
        "in portfolio risk."
    )
    ws['A3'].font = V12
    ws['A3'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 78.0

    ws['A5'] = "CECL Methodology"
    ws['A5'].font = V12B

    ws.merge_cells('A6:J6')
    ws['A6'] = (
        "Under the Current Expected Credit Losses (CECL) standard, institutions must estimate "
        "lifetime expected credit losses on financial assets measured at amortized cost. "
        "The Credit Migration methodology uses the Weighted Average Remaining Maturity (WARM) "
        "approach to estimate these losses, incorporating historical loss experience, current "
        "conditions, and reasonable and supportable forecasts."
    )
    ws['A6'].font = V12
    ws['A6'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[6].height = 63.0


def _sheet_exec_summary(wb, cu, snap):
    """Executive Summary - text-only appendix matching Vizo template."""
    ws = wb.create_sheet("Exec Summary")
    ws.column_dimensions['A'].width = 118.0

    ws['A1'] = "Appendix"
    ws['A1'].font = V12B

    ws['A2'] = "Executive Summary"
    ws['A2'].font = V12B

    ws.merge_cells('A3:J3')
    ws['A3'] = (
        "The Executive Summary provides an overview of the credit union's current portfolio "
        "risk position. It includes the CECL Adjustment calculation showing the relationship "
        "between pooled allowance, specifically identified allowance, total allowance needed, "
        "and the current ACL balance. The summary also presents improved and deteriorated loan "
        "totals by portfolio segment."
    )
    ws['A3'].font = V12
    ws['A3'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 63.0


def _sheet_impdet(wb, cu, snap, df, grades, config, hist=None):
    """Improved/Deteriorated summary matching Impr Deter-Vizo template.

    Layout:
      Rows 1-4   : merged title block (A:J)
      D6-G10     : CECL Adjustment box (4 line items, thick-border box)
      2 charts   : Improved / Deteriorated by grade (rows 11-25)
      2 charts   : Improved-Deteriorated stacked bar & Net Change by pool (rows 26-40)
      Row 45+    : data table (red font, outside print area) feeding the charts
    """
    ws = wb.create_sheet("Impr Deter")
    ws.sheet_view.showGridLines = False
    no_score = config.get('no_score_label', 'Not Reported')
    gl = _all_grades(grades, no_score)

    # Filter: exclude anything with "Hide" in the name. Include WARM-only
    # pools (NRR pools that have no DB rows but appear in WARM/monthly bal).
    visible_grades = [g for g in gl if 'hide' not in g.lower()]
    _imp_for_pools = (hist or {}).get('impaired', {}) if hist else {}
    extra = set((_imp_for_pools.get('hist_bal_data') or {}).keys()) \
            | set((_imp_for_pools.get('pool_bal_detail') or {}).keys())
    pools_set = set(df['loan_pool'].unique()) | extra
    pools = sorted(p for p in pools_set
                   if p and 'hide' not in str(p).lower()
                   and str(p).strip().lower() not in ('grand total','total','excluded','exclude'))

    # ── Column widths (from template) ─────────────────────────────
    ws.column_dimensions['A'].width = 16.71
    ws.column_dimensions['F'].width = 43.29
    ws.column_dimensions['G'].width = 16.71
    ws.column_dimensions['K'].width = 12.0
    ws.column_dimensions['L'].width = 9.71
    ws.column_dimensions['N'].width = 11.29

    # ── Rows 1-4  Title block (merged A1:J1 … A4:J4) ─────────────
    for r in range(1, 5):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws['A1'] = cu
    ws['A1'].font = V14B
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 18.75

    ws['A2'] = 'Executive Summary '
    ws['A2'].font = V12B
    ws['A2'].alignment = Alignment(horizontal='center')

    ws['A3'] = 'CECL Adjustment  & Improved/Deteriorated '
    ws['A3'].font = V12B
    ws['A3'].alignment = Alignment(horizontal='center')

    ws['A4'] = pd.to_datetime(snap)
    ws['A4'].font = V12B
    ws['A4'].alignment = Alignment(horizontal='center')
    ws['A4'].number_format = 'mm-dd-yy'

    # Row 5: spacer
    ws.row_dimensions[5].height = 9.6

    # ── D6-G10  CECL Adjustment box (thick-border box, template rows 6-10) ─
    # Compute ACL summary values (same calc as ACL Env by Pool Mgmt Adj sheet)
    acl = _compute_acl_totals(df, grades, config, hist, snap)

    # D6: title with top+left thick borders
    ws['D6'] = 'CECL Adjustment'
    ws['D6'].font = V14
    ws.row_dimensions[6].height = 19.5

    # The 4 data rows (7-10) reference the ACL reserve summary
    adj_items = [
        (acl['spec_id_label'],    acl['spec_id_allowance']),
        (acl['needed_label'],     acl['total_allowance_needed']),
        (acl['balance_label'],    acl['acl_balance']),
        (acl['adjustment_label'], acl['adjustment']),
    ]
    for i, (lbl, val) in enumerate(adj_items):
        r = 7 + i
        c_d = ws.cell(row=r, column=4, value=lbl)
        c_d.font = V12
        c_d.alignment = Alignment(horizontal='left')
        c_g = ws.cell(row=r, column=7, value=val)
        c_g.font = V12
        c_g.alignment = Alignment(horizontal='right')
        c_g.number_format = ACCT

    # Thick dark-teal border box from D6:G10 (matches Vizo TEAL accent color)
    _thick = Side(style='thick', color='FF0D4D5E')
    for r in range(6, 11):
        is_top = (r == 6)
        is_bot = (r == 10)
        for c in range(4, 8):   # D=4 .. G=7
            is_left = (c == 4)
            is_right = (c == 7)
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                top=_thick if is_top else None,
                bottom=_thick if is_bot else None,
                left=_thick if is_left else None,
                right=_thick if is_right else None,
            )
    ws.row_dimensions[10].height = 16.5
    ws.row_dimensions[11].height = 16.5

    # ══════════════════════════════════════════════════════════════
    #  DATA TABLE  (red font, outside print area)
    # ══════════════════════════════════════════════════════════════
    RED12 = Font(name='Calibri', size=12, color='FF0000')

    # ── Loan-type data  (cols F-I, rows 45+)  feeds Charts 0 & 3 ─
    ws.cell(row=45, column=6, value='Loan Type').font = RED12
    ws.cell(row=45, column=7, value='Improved').font = RED12
    ws.cell(row=45, column=8, value='Deteriorated').font = RED12
    ws.cell(row=45, column=9, value='Net Change').font = RED12

    _rr_map_id = hist.get('impaired', {}).get('risk_rated', {}) if hist else {}
    for pi, pool in enumerate(pools):
        r = 46 + pi
        pdf = df[df['loan_pool'] == pool]
        if _rr_map_id.get(pool, True):
            imp_pct, det_pct, net_pct = _ncc(pdf, grades, config)
        else:
            imp_pct, det_pct, net_pct = 0.0, 0.0, 0.0

        ws.cell(row=r, column=6, value=pool).font = RED12
        ws.cell(row=r, column=6).alignment = Alignment(horizontal='left')
        ws.cell(row=r, column=7, value=imp_pct).font = RED12
        ws.cell(row=r, column=7).number_format = '0%'
        ws.cell(row=r, column=8, value=-det_pct).font = RED12
        ws.cell(row=r, column=8).number_format = '0%'
        ws.cell(row=r, column=9, value=net_pct).font = RED12
        ws.cell(row=r, column=9).number_format = '0.0%'

    pool_last_row = 45 + len(pools)   # last data row for pools

    # ── Grade-level data  (cols C-E, row 70+)  feeds Charts 1 & 2 ─
    # Use Executive Summary (3) data from WARM file if available
    es3 = (hist.get('impaired', {}).get('exec_summary_3', {})
           if hist else {})
    GR_START = 70
    ws.cell(row=GR_START, column=3, value='Grade').font = RED12
    ws.cell(row=GR_START, column=4, value=' Balance ').font = RED12

    # Grades to chart (exclude "Not Reported" / no-score label)
    chart_grades = [g for g in visible_grades if g != no_score]

    # Compute grade-level improved/deteriorated from migration matrix
    if not es3:
        matrix = risk_change_matrix(df, grades, no_score)
        n_top = config.get('top_grades_double_drop', 3)
        grade_imp = {g: 0 for g in chart_grades}
        grade_det = {g: 0 for g in chart_grades}
        for j, og in enumerate(gl):
            for i, cg in enumerate(gl):
                v = _matrix_val(matrix, cg, og)
                if i > j:
                    if j < n_top and (i - j) < 2:
                        pass  # unchanged – small drop within top grades
                    else:
                        if og in grade_det:
                            grade_det[og] += v
                elif i < j:
                    if og in grade_imp:
                        grade_imp[og] += v

    for gi, grade in enumerate(chart_grades):
        r = GR_START + 1 + gi
        if es3:
            imp_bal = es3.get('improved', {}).get(grade, 0)
            det_bal = es3.get('deteriorated', {}).get(grade, 0)
        else:
            imp_bal = grade_imp.get(grade, 0)
            det_bal = grade_det.get(grade, 0)
        ws.cell(row=r, column=3, value=grade).font = RED12
        ws.cell(row=r, column=4, value=imp_bal).font = RED12
        ws.cell(row=r, column=4).number_format = ACCT
        ws.cell(row=r, column=5, value=det_bal).font = RED12
        ws.cell(row=r, column=5).number_format = ACCT

    gr_first = GR_START + 1
    gr_last  = GR_START + len(chart_grades)

    # Helper percentage columns (F=improved %, G=deteriorated %) drive the
    # Improved / Deteriorated grade charts so data labels render as percentages.
    imp_total = sum(ws.cell(row=r, column=4).value or 0
                    for r in range(gr_first, gr_last + 1))
    det_total = sum(ws.cell(row=r, column=5).value or 0
                    for r in range(gr_first, gr_last + 1))
    ws.cell(row=GR_START, column=6, value='Improved %').font = RED12
    ws.cell(row=GR_START, column=7, value='Deteriorated %').font = RED12
    for r in range(gr_first, gr_last + 1):
        imp_v = ws.cell(row=r, column=4).value or 0
        det_v = ws.cell(row=r, column=5).value or 0
        ws.cell(row=r, column=6,
                value=(imp_v / imp_total) if imp_total else 0).font = RED12
        ws.cell(row=r, column=6).number_format = '0%'
        ws.cell(row=r, column=7,
                value=(det_v / det_total) if det_total else 0).font = RED12
        ws.cell(row=r, column=7).number_format = '0%'

    # ══════════════════════════════════════════════════════════════
    #  CHARTS  (exact match to Impr Deter-Vizo template)
    # ══════════════════════════════════════════════════════════════
    ACCT_FMT = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"_);_(@_)'

    # Vizo Color Theme 1 hex values
    TEAL   = '0D4D5E'   # accent1 – Improved
    MAROON = '3D1A1A'   # lt2/bg2 – Deteriorated
    OLIVE  = '829901'   # accent2/dk2 – Net Change
    WHITE  = 'FFFFFF'   # lt1/bg1 – data label font

    from openpyxl.drawing.text import RichTextProperties

    def _set_series_fill(series, hex_color, line_noFill=True):
        """Set bar fill color and optionally remove outline."""
        series.graphicalProperties.solidFill = hex_color
        if line_noFill:
            series.graphicalProperties.line = LineProperties(noFill=True)

    def _remove_chart_borders(chart):
        """Remove chart-area and plot-area outlines to match borderless template."""
        if getattr(chart, 'spPr', None) is None:
            chart.spPr = GraphicalProperties()
        if getattr(chart, 'graphical_properties', None) is not None:
            chart.graphical_properties.line = LineProperties(noFill=True)

        pa = chart.plot_area
        if getattr(pa, 'spPr', None) is None:
            pa.spPr = GraphicalProperties()
        pa_gp = getattr(pa, 'graphicalProperties', None)
        if pa_gp is None:
            pa_gp = getattr(pa, 'graphical_properties', None)
        if pa_gp is not None:
            pa_gp.line = LineProperties(noFill=True)

    def _remove_axis_lines(chart):
        """Remove visible axis lines on both category and value axes."""
        for ax in (getattr(chart, 'x_axis', None), getattr(chart, 'y_axis', None)):
            if ax is None:
                continue
            if getattr(ax, 'spPr', None) is None:
                ax.spPr = GraphicalProperties()
            ax.spPr.line = LineProperties(noFill=True)

    def _dlbl_txpr(bold=False, rot=None, fill_color=WHITE):
        """Build RichText for data labels: 9pt, specified font color, optional rotation."""
        rpr = CharacterProperties(
            sz=900, b=bold,
            solidFill=fill_color,
            latin=DrawingFont(typeface='+mn-lt'),
            ea=DrawingFont(typeface='+mn-ea'),
            cs=DrawingFont(typeface='+mn-cs'),
        )
        body_kw = {'spcFirstLastPara': True, 'vertOverflow': 'ellipsis',
                    'wrap': 'square', 'anchor': 'ctr', 'anchorCtr': True}
        if rot is not None:
            body_kw['rot'] = rot
        else:
            body_kw['rot'] = 0
        return RichText(
            bodyPr=RichTextProperties(**body_kw),
            p=[Paragraph(pPr=ParagraphProperties(defRPr=rpr), endParaRPr=None)],
        )

    def _cat_ax_txpr(bold=False, fill_color=None):
        """Build RichText for category axis tick labels."""
        kw = {'sz': 900, 'b': bold,
              'latin': DrawingFont(typeface='+mn-lt'),
              'ea': DrawingFont(typeface='+mn-ea'),
              'cs': DrawingFont(typeface='+mn-cs')}
        if fill_color:
            kw['solidFill'] = fill_color
        rpr = CharacterProperties(**kw)
        return RichText(
            bodyPr=RichTextProperties(rot=-60000000, spcFirstLastPara=True,
                                      vertOverflow='ellipsis', vert='horz',
                                      wrap='square', anchor='ctr', anchorCtr=True),
            p=[Paragraph(pPr=ParagraphProperties(defRPr=rpr), endParaRPr=None)],
        )

    def _calibri18_title(text):
        """Build a chart Title styled as Calibri 18pt (bold)."""
        rpr = CharacterProperties(
            sz=1800, b=True,
            latin=DrawingFont(typeface='Calibri'),
            ea=DrawingFont(typeface='Calibri'),
            cs=DrawingFont(typeface='Calibri'),
        )
        para = Paragraph(
            pPr=ParagraphProperties(defRPr=rpr),
            r=[RegularTextRun(rPr=rpr, t=text)],
        )
        return Title(tx=Text(rich=RichText(p=[para])), overlay=False)

    # ── Chart 1 : "Improved Loans" – clustered column by grade (TOP-LEFT) ──
    c1 = BarChart()
    c1.type = 'col'
    c1.grouping = 'clustered'
    c1.title = _calibri18_title('Improved Loans')
    c1.legend = None
    c1.y_axis.delete = True           # valAx: delete=1
    c1.y_axis.numFmt = ACCT_FMT
    c1.y_axis.number_format = ACCT_FMT
    c1.x_axis.delete = False          # catAx: delete=0, axPos=b
    c1.x_axis.majorTickMark = 'out'
    c1.x_axis.minorTickMark = 'none'
    c1.x_axis.tickLblPos = 'nextTo'
    c1.x_axis.txPr = _cat_ax_txpr(bold=True)
    c1.gapWidth = 100
    c1.width = 15
    c1.height = 7.5
    # Exclude top grade (always 0 for Improved – cannot improve beyond best grade)
    imp_first = gr_first + 1 if gr_last > gr_first else gr_first
    d1 = Reference(ws, min_col=6, min_row=imp_first, max_row=gr_last)
    cat1 = Reference(ws, min_col=3, min_row=imp_first, max_row=gr_last)
    c1.add_data(d1)
    c1.set_categories(cat1)
    _remove_chart_borders(c1)
    _remove_axis_lines(c1)
    _set_series_fill(c1.series[0], TEAL)
    c1.series[0].dLbls = DataLabelList()
    c1.series[0].dLbls.showVal = True
    c1.series[0].dLbls.numFmt = '0%'
    c1.series[0].dLbls.dLblPos = 'inEnd'
    c1.series[0].dLbls.txPr = _dlbl_txpr(rot=-5400000)
    # Anchor offsets balance chart widths: column F is 43.29 wide vs others ~8.43,
    # so split the chart strip down the middle of column F (~17.43 width units ≈ 1209675 EMU)
    # to give all four charts equal outer width.
    IMPDET_MID_OFF = 1209675
    anc1 = TwoCellAnchor()
    anc1._from = AnchorMarker(col=0, colOff=0, row=11, rowOff=0)
    anc1.to = AnchorMarker(col=5, colOff=IMPDET_MID_OFF, row=25, rowOff=0)
    ws.add_chart(c1, anc1)

    # ── Chart 2 : "Deteriorated Loans" – clustered column by grade (TOP-RIGHT) ──
    c2 = BarChart()
    c2.type = 'col'
    c2.grouping = 'clustered'
    c2.title = _calibri18_title('Deteriorated Loans')
    c2.legend = None
    c2.y_axis.delete = True           # valAx: delete=1
    c2.y_axis.numFmt = ACCT_FMT
    c2.y_axis.number_format = ACCT_FMT
    c2.x_axis.delete = False          # catAx: delete=0, axPos=b
    c2.x_axis.majorTickMark = 'out'
    c2.x_axis.minorTickMark = 'none'
    c2.x_axis.tickLblPos = 'nextTo'
    c2.x_axis.txPr = _cat_ax_txpr(bold=True)
    c2.gapWidth = 100
    c2.width = 15
    c2.height = 7.5
    # Exclude bottom grade (always 0 for Deteriorated – cannot deteriorate beyond worst grade)
    det_last = gr_last - 1 if gr_last > gr_first else gr_last
    d2 = Reference(ws, min_col=7, min_row=gr_first, max_row=det_last)
    cat2 = Reference(ws, min_col=3, min_row=gr_first, max_row=det_last)
    c2.add_data(d2)
    c2.set_categories(cat2)
    _remove_chart_borders(c2)
    _remove_axis_lines(c2)
    _set_series_fill(c2.series[0], MAROON)
    c2.series[0].graphicalProperties.line = LineProperties(w=19050, noFill=True)
    c2.series[0].dLbls = DataLabelList()
    c2.series[0].dLbls.showVal = True
    c2.series[0].dLbls.numFmt = '0%'
    c2.series[0].dLbls.dLblPos = 'inEnd'
    c2.series[0].dLbls.txPr = _dlbl_txpr(bold=True, rot=-5400000)
    anc2 = TwoCellAnchor()
    anc2._from = AnchorMarker(col=5, colOff=IMPDET_MID_OFF, row=11, rowOff=0)
    anc2.to = AnchorMarker(col=10, colOff=0, row=25, rowOff=0)
    ws.add_chart(c2, anc2)

    # ── Chart 3 : "Improved/Deteriorated Loans" – stacked bar by pool (BOTTOM-LEFT) ──
    # For type='bar': x_axis = catAx (right), y_axis = valAx (bottom)
    c0 = BarChart()
    c0.type = 'bar'
    c0.grouping = 'stacked'
    c0.title = _calibri18_title('Improved/Deteriorated Loans ')
    c0.overlap = 100
    c0.gapWidth = 10
    c0.layout = Layout(
        manualLayout=ManualLayout(
            xMode='edge', yMode='edge',
            x=0.039320822162645222, y=0.12380952380952381,
            w=0.92135835567470958, h=0.72857142857142854,
        )
    )
    # catAx (x_axis): orientation=minMax, axPos=r, tickLblPos=high
    c0.x_axis.delete = False
    c0.x_axis.scaling.orientation = 'minMax'
    c0.x_axis.majorTickMark = 'out'
    c0.x_axis.minorTickMark = 'none'
    c0.x_axis.tickLblPos = 'high'
    c0.x_axis.numFmt = '0%'
    c0.x_axis.number_format = '0%'
    # valAx (y_axis): orientation=maxMin, labels hidden
    c0.y_axis.delete = False
    c0.y_axis.scaling.orientation = 'maxMin'
    c0.y_axis.numFmt = '0%'
    c0.y_axis.tickLblPos = 'none'
    c0.y_axis.majorTickMark = 'none'
    c0.y_axis.minorTickMark = 'none'
    c0.y_axis.majorGridlines = None  # will be set via spPr below
    c0.width = 15
    c0.height = 7.5
    imp_ref = Reference(ws, min_col=7, min_row=45, max_row=pool_last_row)
    det_ref = Reference(ws, min_col=8, min_row=45, max_row=pool_last_row)
    cat0    = Reference(ws, min_col=6, min_row=46, max_row=pool_last_row)
    c0.add_data(imp_ref, titles_from_data=True)
    c0.add_data(det_ref, titles_from_data=True)
    c0.set_categories(cat0)
    _remove_chart_borders(c0)
    _remove_axis_lines(c0)
    # Series 0 = Improved → MAROON fill (per Brian's edit)
    _set_series_fill(c0.series[0], MAROON)
    c0.series[0].dLbls = DataLabelList()
    c0.series[0].dLbls.showVal = True
    c0.series[0].dLbls.dLblPos = 'inBase'
    c0.series[0].dLbls.txPr = _dlbl_txpr()
    # Series 1 = Deteriorated → TEAL fill (per Brian's edit)
    _set_series_fill(c0.series[1], TEAL)
    c0.series[1].dLbls = DataLabelList()
    c0.series[1].dLbls.showVal = True
    c0.series[1].dLbls.dLblPos = 'inBase'
    c0.series[1].dLbls.txPr = _dlbl_txpr()
    if c0.legend:
        c0.legend.position = 'b'
    anc0 = TwoCellAnchor()
    anc0._from = AnchorMarker(col=0, colOff=0, row=25, rowOff=0)
    anc0.to = AnchorMarker(col=5, colOff=IMPDET_MID_OFF, row=39, rowOff=0)
    ws.add_chart(c0, anc0)

    # ── Chart 4 : "Net Change" – clustered bar by loan type (BOTTOM-RIGHT) ──
    # For type='bar': x_axis = catAx (right), y_axis = valAx (bottom)
    c3 = BarChart()
    c3.type = 'bar'
    c3.grouping = 'clustered'
    c3.title = _calibri18_title('Net Change')
    c3.overlap = 100
    c3.gapWidth = 10
    c3.legend = None
    c3.layout = Layout(
        manualLayout=ManualLayout(
            xMode='edge', yMode='edge',
            x=0.039320822162645222, y=0.12380952380952381,
            w=0.92135835567470958, h=0.72857142857142854,
        )
    )
    # catAx (x_axis): orientation=minMax, axPos=r, tickLblPos=high
    c3.x_axis.delete = False
    c3.x_axis.scaling.orientation = 'minMax'
    c3.x_axis.majorTickMark = 'out'
    c3.x_axis.minorTickMark = 'none'
    c3.x_axis.tickLblPos = 'high'
    c3.x_axis.numFmt = '0.0%'
    c3.x_axis.number_format = '0.0%'
    # valAx (y_axis): orientation=maxMin, labels hidden
    c3.y_axis.delete = False
    c3.y_axis.scaling.orientation = 'maxMin'
    c3.y_axis.numFmt = '0.0%'
    c3.y_axis.tickLblPos = 'none'
    c3.y_axis.majorTickMark = 'out'
    c3.y_axis.minorTickMark = 'none'
    c3.y_axis.majorGridlines = None  # will be set via spPr below
    c3.width = 15
    c3.height = 7.5
    net_ref = Reference(ws, min_col=9, min_row=45, max_row=pool_last_row)
    cat3    = Reference(ws, min_col=6, min_row=46, max_row=pool_last_row)
    c3.add_data(net_ref, titles_from_data=True)
    c3.set_categories(cat3)
    _remove_chart_borders(c3)
    _remove_axis_lines(c3)
    _set_series_fill(c3.series[0], OLIVE)
    c3.series[0].dLbls = DataLabelList()
    c3.series[0].dLbls.showVal = True
    c3.series[0].dLbls.numFmt = '0.0%'
    c3.series[0].dLbls.dLblPos = 'outEnd'
    c3.series[0].dLbls.txPr = _dlbl_txpr(fill_color='000000')
    anc3 = TwoCellAnchor()
    anc3._from = AnchorMarker(col=5, colOff=IMPDET_MID_OFF, row=25, rowOff=0)
    anc3.to = AnchorMarker(col=10, colOff=0, row=39, rowOff=0)
    ws.add_chart(c3, anc3)

    # ── Page setup ────────────────────────────────────────────────
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.25, bottom=0.25,
        header=0, footer=0,
    )
    ws.print_area = 'A1:J39'


def _sheet_risk_change(wb, cu, snap, data_df, grades, config, pool_name=None, hist=None):
    """Risk Change matrix sheet – used for both total and per-pool."""
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g for g in _all_grades(grades, no_score) if not _is_hidden(g)]
    matrix = risk_change_matrix(data_df, grades, no_score)
    rng = _grade_ranges(grades, no_score)
    total = data_df['current_balance'].sum()

    # Resolve per-pool or grand-total DQ/CO data
    _imp = hist.get('impaired', {}) if hist else {}
    if pool_name:
        # Match dataframe pool name to WARM-file pool key (may differ by trailing spaces)
        _dq_pool = _imp.get('dq_by_pool', {})
        _co_pool = _imp.get('co_by_pool', {})
        _pool_lc = pool_name.strip().lower()
        _dq_data = next((v for k, v in _dq_pool.items() if k.strip().lower() == _pool_lc), {})
        _co_data = next((v for k, v in _co_pool.items() if k.strip().lower() == _pool_lc), {})
    else:
        _dq_data = _imp.get('dq_by_status', {})
        _co_data = _imp.get('co_by_status', {})

    if pool_name:
        safe = re.sub(r'[^\w\s-]', '', pool_name)[:20]
        ws = wb.create_sheet(f"Risk Chg {safe}")
    else:
        ws = wb.create_sheet("Risk Change Total")
    ws.sheet_view.showGridLines = False

    ncol = 3 + len(gl)  # A=Current Grade, B=Score Range, C..M=grade cols, N=Grand Total
    # Columns P-S for Deteriorated/Improved/Unchanged summary
    pcol_start = ncol + 2  # col P = ncol + 2
    hdr_center = Alignment(horizontal='center', wrap_text=True)
    side_left = Alignment(horizontal='left')

    # Column widths matching template
    for ci in range(1, 10):  # A through I = 25.57 (184 px)
        ws.column_dimensions[get_column_letter(ci)].width = 25.57
    for ci in range(10, ncol):
        ws.column_dimensions[get_column_letter(ci)].width = 17.3
    ws.column_dimensions[get_column_letter(ncol)].width = 21.7  # Grand Total

    # ── Title rows (merged A1:N, rows 1-4) ──
    for ri in range(1, 5):
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=ncol)
    ws['A1'] = cu
    ws['A1'].font = V14B
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'] = "Executive Summary Total Loans" if not pool_name else "Risk Change By Credit Score"
    ws['A2'].font = V12B
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A3'] = "Risk Change By Credit Score" if not pool_name else pool_name
    ws['A3'].font = V12B
    ws['A3'].alignment = Alignment(horizontal='center')
    ws['A4'] = f"For Quarter Ending {_snap_display(snap)}"
    ws['A4'].font = V12B
    ws['A4'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 21.0
    for ri in range(2, 5):
        ws.row_dimensions[ri].height = 21.6

    # ─── Dollar Matrix Header (rows 5-6) ───
    # Row 5: A5:B5 merged (blank header), C5:M5 merged "Original Grade", N5:N6 merged "Grand Total"
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)
    ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=ncol - 1)
    ws.merge_cells(start_row=5, start_column=ncol, end_row=6, end_column=ncol)
    c5ab = ws.cell(row=5, column=1)
    c5ab.fill = HDR_FILL
    c5ab.font = HDR_FONT
    ws.cell(row=5, column=2).fill = HDR_FILL
    c5og = ws.cell(row=5, column=3, value="Original Grade")
    c5og.font = HDR_FONT
    c5og.fill = HDR_FILL
    c5og.alignment = hdr_center
    c5og.number_format = DOLLAR
    for ci in range(4, ncol):
        ws.cell(row=5, column=ci).fill = HDR_FILL
    c5gt = ws.cell(row=5, column=ncol, value="Grand Total")
    c5gt.font = HDR_FONT
    c5gt.fill = HDR_FILL
    c5gt.alignment = hdr_center
    c5gt.number_format = DOLLAR

    # Row 6: A6:B6 merged "$ Current Grade", C6-M6 grade labels
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=2)
    c6ab = ws.cell(row=6, column=1, value="$ Current Grade")
    c6ab.font = HDR_FONT
    c6ab.fill = HDR_FILL
    c6ab.alignment = hdr_center
    c6ab.number_format = DOLLAR
    ws.cell(row=6, column=2).fill = HDR_FILL
    for j, g in enumerate(gl):
        cell = ws.cell(row=6, column=3 + j, value=g)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = hdr_center
    ws.cell(row=6, column=ncol).fill = HDR_FILL  # part of N5:N6 merge

    # ─── Dollar Data rows ───
    # WARM rule: top N original grades require 2+ grade drop for deterioration;
    # remaining grades require only 1. Improved is always 1+ grade improvement.
    n_top = config.get('top_grades_double_drop', 3)
    for i, g in enumerate(gl):
        r = 7 + i
        # Side headers (cols A-B): header fill + white font
        fnt_side = V12BR if _is_hidden(g) else Font(name='Calibri', bold=True, size=12, color='FFFFFF')
        ws.cell(row=r, column=1, value=g).font = fnt_side
        ws.cell(row=r, column=1).fill = HDR_FILL
        ws.cell(row=r, column=1).alignment = side_left
        ws.cell(row=r, column=2, value=rng.get(g, '')).font = fnt_side
        ws.cell(row=r, column=2).fill = HDR_FILL
        ws.cell(row=r, column=2).alignment = side_left

        rtotal = 0
        for j, og in enumerate(gl):
            v = _matrix_val(matrix, g, og)
            cell = ws.cell(row=r, column=3 + j, value=v)
            cell.number_format = ACCT
            cell.font = V12R if _is_hidden(g) else V12
            rtotal += v
            if g == no_score or og == no_score:
                pass  # Not Reported → always unchanged, no fill
            elif i > j:
                # Potential deterioration – apply top-grade exception
                if j < n_top and (i - j) < 2:
                    pass  # unchanged
                else:
                    cell.fill = DET_FILL
            elif i < j:
                cell.fill = IMP_FILL
        gt_cell = ws.cell(row=r, column=ncol, value=rtotal)
        gt_cell.number_format = ACCT
        gt_cell.font = V12BR if _is_hidden(g) else V12B

    # P-S: per-original-grade Deteriorated/Improved/Unchanged (column perspective)
    # WARM computes: for each original grade j, sum deteriorated/improved across current grades i
    grand_det = 0
    grand_imp = 0
    grand_unc = 0
    for j, og in enumerate(gl):
        r = 7 + j
        det_from = 0
        imp_from = 0
        unc_from = 0
        for i, g in enumerate(gl):
            v = _matrix_val(matrix, g, og)
            if g == no_score or og == no_score:
                unc_from += v  # Not Reported → always unchanged
            elif i > j:
                # Current grade worse than original → potential deterioration
                if j < n_top and (i - j) < 2:
                    unc_from += v
                else:
                    det_from += v
            elif i < j:
                imp_from += v
            else:
                unc_from += v
        fnt_ps = _grade_font(og)
        ws.cell(row=r, column=pcol_start, value=det_from).number_format = ACCT
        ws.cell(row=r, column=pcol_start).font = fnt_ps
        ws.cell(row=r, column=pcol_start + 1, value=imp_from).number_format = ACCT
        ws.cell(row=r, column=pcol_start + 1).font = fnt_ps
        ws.cell(row=r, column=pcol_start + 2, value=unc_from).number_format = ACCT
        ws.cell(row=r, column=pcol_start + 2).font = fnt_ps
        grand_det += det_from
        grand_imp += imp_from
        grand_unc += unc_from

    # P-S column headers (on row 6)
    for c, lbl in [(pcol_start, "Deteriorated"), (pcol_start + 1, "Improved"),
                   (pcol_start + 2, "Unchanged")]:
        cell = ws.cell(row=6, column=c, value=lbl)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = hdr_center
        ws.column_dimensions[get_column_letter(c)].width = 17.0

    # ─── Grand Total row ───
    r_gt = 7 + len(gl)
    ws.cell(row=r_gt, column=1, value="Grand Total").font = HDR_FONT
    ws.cell(row=r_gt, column=1).fill = HDR_FILL
    ws.cell(row=r_gt, column=2).fill = HDR_FILL
    ws.cell(row=r_gt, column=2).font = HDR_FONT
    for j, og in enumerate(gl):
        ct = sum(_matrix_val(matrix, g2, og) for g2 in gl)
        c = ws.cell(row=r_gt, column=3 + j, value=ct)
        c.number_format = ACCT
        c.font = V12B
    ws.cell(row=r_gt, column=ncol, value=total).number_format = ACCT
    ws.cell(row=r_gt, column=ncol).font = V12B

    # Balance Adj row
    r_ba = r_gt + 1
    _imp = hist.get('impaired', {}) if hist else {}
    if pool_name:
        _bal_adj_map = _imp.get('balance_adjustments', {})
        _pool_lc = pool_name.strip().lower()
        bal_adj = next((v for k, v in _bal_adj_map.items() if k.strip().lower() == _pool_lc), 0.0)
    else:
        bal_adj = _imp.get('total_balance_adjustment', 0.0)
    ws.cell(row=r_ba, column=1, value="Balance Adjustment").font = HDR_FONT
    ws.cell(row=r_ba, column=1).fill = HDR_FILL
    ws.cell(row=r_ba, column=2).fill = HDR_FILL
    ws.cell(row=r_ba, column=ncol, value=bal_adj).number_format = ACCT
    ws.cell(row=r_ba, column=ncol).font = V12

    # Total in Portfolio row
    r_tp = r_ba + 1
    total_in_portfolio = total + bal_adj
    if not pool_name:
        tip = _imp.get('total_in_portfolio', total_in_portfolio)
        if tip:
            total_in_portfolio = tip
    ws.cell(row=r_tp, column=1, value="Total in Portfolio").font = HDR_FONT
    ws.cell(row=r_tp, column=1).fill = HDR_FILL
    ws.cell(row=r_tp, column=2).fill = HDR_FILL
    ws.cell(row=r_tp, column=ncol, value=total_in_portfolio).number_format = ACCT
    ws.cell(row=r_tp, column=ncol).font = V12B

    # ─── Percent Matrix Header (2 rows below Total in Portfolio) ───
    r_ph1 = r_tp + 2  # row 22 equivalent
    r_ph2 = r_ph1 + 1  # row 23 equivalent

    # Row ph1: A:B merged (blank header), C:M merged "Original Grade", N:N+1 merged "Grand Total"
    ws.merge_cells(start_row=r_ph1, start_column=1, end_row=r_ph1, end_column=2)
    ws.merge_cells(start_row=r_ph1, start_column=3, end_row=r_ph1, end_column=ncol - 1)
    ws.merge_cells(start_row=r_ph1, start_column=ncol, end_row=r_ph2, end_column=ncol)
    ws.cell(row=r_ph1, column=1).fill = HDR_FILL
    ws.cell(row=r_ph1, column=1).font = HDR_FONT
    ws.cell(row=r_ph1, column=2).fill = HDR_FILL
    c_pog = ws.cell(row=r_ph1, column=3, value="Original Grade")
    c_pog.font = HDR_FONT
    c_pog.fill = HDR_FILL
    c_pog.alignment = hdr_center
    c_pog.number_format = DOLLAR
    for ci in range(4, ncol):
        ws.cell(row=r_ph1, column=ci).fill = HDR_FILL
    c_pgt = ws.cell(row=r_ph1, column=ncol, value="Grand Total")
    c_pgt.font = HDR_FONT
    c_pgt.fill = HDR_FILL
    c_pgt.alignment = hdr_center
    c_pgt.number_format = DOLLAR

    # Row ph2: A:B merged "% Current Grade", grade labels
    ws.merge_cells(start_row=r_ph2, start_column=1, end_row=r_ph2, end_column=2)
    c_pcg = ws.cell(row=r_ph2, column=1, value="% Current Grade")
    c_pcg.font = HDR_FONT
    c_pcg.fill = HDR_FILL
    c_pcg.alignment = hdr_center
    c_pcg.number_format = DOLLAR
    ws.cell(row=r_ph2, column=2).fill = HDR_FILL
    for j, g in enumerate(gl):
        cell = ws.cell(row=r_ph2, column=3 + j, value=g)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = hdr_center
    ws.cell(row=r_ph2, column=ncol).fill = HDR_FILL  # part of merge

    # ─── Percent Data rows ───
    for i, g in enumerate(gl):
        r = r_ph2 + 1 + i
        fnt_side = V12BR if _is_hidden(g) else Font(name='Calibri', bold=True, size=12, color='FFFFFF')
        ws.cell(row=r, column=1, value=g).font = fnt_side
        ws.cell(row=r, column=1).fill = HDR_FILL
        ws.cell(row=r, column=1).alignment = side_left
        ws.cell(row=r, column=2, value=rng.get(g, '')).font = fnt_side
        ws.cell(row=r, column=2).fill = HDR_FILL
        ws.cell(row=r, column=2).alignment = side_left

        rtotal = 0
        for j, og in enumerate(gl):
            v = _matrix_val(matrix, g, og)
            col_total = sum(_matrix_val(matrix, g2, og) for g2 in gl)
            pct = v / col_total if col_total else 0
            cell = ws.cell(row=r, column=3 + j, value=pct)
            cell.number_format = PCT
            cell.font = Font(name='Calibri', size=11, color='FF0000') if _is_hidden(g) else V11
            if g == no_score or og == no_score:
                pass  # Not Reported → always unchanged, no fill
            elif i > j:
                if j < n_top and (i - j) < 2:
                    pass  # unchanged
                else:
                    cell.fill = DET_FILL
            elif i < j:
                cell.fill = IMP_FILL
            rtotal += v
        gt_pct = rtotal / total if total else 0
        ws.cell(row=r, column=ncol, value=gt_pct).number_format = PCT
        ws.cell(row=r, column=ncol).font = V11

    # Percent Grand Total row
    r_pgt = r_ph2 + 1 + len(gl)
    ws.cell(row=r_pgt, column=1, value="Grand Total").font = HDR_FONT
    ws.cell(row=r_pgt, column=1).fill = HDR_FILL
    ws.cell(row=r_pgt, column=2).fill = HDR_FILL
    ws.cell(row=r_pgt, column=2).font = HDR_FONT
    for j in range(len(gl)):
        c = ws.cell(row=r_pgt, column=3 + j, value=1.0)
        c.number_format = '0%'
        c.font = V11B
    ws.cell(row=r_pgt, column=ncol, value=1.0).number_format = '0%'
    ws.cell(row=r_pgt, column=ncol).font = V11B

    # ─── Net Credit Change (large font, merged) ───
    r_nc = r_pgt + 2
    imp_bal = grand_imp
    det_bal = grand_det
    net = imp_bal - det_bal
    ws.merge_cells(start_row=r_nc, start_column=2, end_row=r_nc + 1, end_column=4)
    ws.cell(row=r_nc, column=2, value="Net Credit Change").font = Font(name='Calibri', bold=False, size=18)
    ws.cell(row=r_nc, column=2).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r_nc].height = 15.6
    ws.row_dimensions[r_nc + 1].height = 18.0

    # ─── Summary Table (Improved / Deteriorated / Unchanged / Portfolio / Net Change) ───
    unc_bal = total - imp_bal - det_bal
    r_sum = r_nc + 6
    WHITE_BOLD12 = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    V12_REG = Font(name='Calibri', bold=False, size=12)
    WHITE_REG12 = Font(name='Calibri', bold=False, size=12, color='FFFFFF')
    summary_items = [
        ("Improved",     imp_bal, imp_bal / total if total else 0, IMP_FILL),
        ("Deteriorated", det_bal, det_bal / total if total else 0, DET_FILL),
        ("Unchanged",    unc_bal, unc_bal / total if total else 0, HDR_FILL),
        ("Portfolio",    total,   1.0,                             None),
        ("Net Change",   net,     net / total if total else 0,     None),
    ]
    for idx, (lbl, bal, pct, fill) in enumerate(summary_items):
        r = r_sum + idx
        ws.cell(row=r, column=2, value=lbl).font = V12_REG
        c_bal = ws.cell(row=r, column=3, value=bal)
        c_bal.number_format = '#,##0'
        c_bal.font = V12_REG
        c_bal.alignment = Alignment(horizontal='center')
        c_pct = ws.cell(row=r, column=4, value=pct)
        c_pct.number_format = '0.0%'
        c_pct.font = V12_REG
        c_pct.alignment = Alignment(horizontal='center')
        if fill:
            c_bal.fill = fill
            c_pct.fill = fill
            # Keep label (col B) black; only the numeric cells use white-on-fill
            ws.cell(row=r, column=2).font = V12_REG
            c_bal.font = WHITE_REG12
            c_pct.font = WHITE_REG12

    # ─── Doughnut Chart (Improved / Deteriorated / Unchanged) ───
    dc = DoughnutChart()
    dc.title = None
    dc.style = 10
    cats_dn = Reference(ws, min_col=2, min_row=r_sum, max_row=r_sum + 2)
    vals_dn = Reference(ws, min_col=4, min_row=r_sum, max_row=r_sum + 2)
    dc.add_data(vals_dn)
    dc.set_categories(cats_dn)
    s_dn = dc.series[0]
    s_dn.explosion = 16
    dn_pts = []
    for dp_idx, dp_color in enumerate(['829901', '873A3A', '0D4D5E']):
        dp = DataPoint(idx=dp_idx)
        dp.graphicalProperties = GraphicalProperties()
        dp.graphicalProperties.noFill = True
        dp.graphicalProperties.line = LineProperties(solidFill=dp_color, w=38100)
        dn_pts.append(dp)
    s_dn.data_points = dn_pts
    dc.innerRadius = 50
    dc.legend = None
    dc.graphical_properties = GraphicalProperties()
    dc.graphical_properties.noFill = True
    dc.graphical_properties.line = LineProperties(noFill=True)
    dc.width = 12
    dc.height = 7.5
    anc_dc = TwoCellAnchor()
    anc_dc._from = AnchorMarker(col=2, colOff=500000, row=r_nc + 1, rowOff=0)
    anc_dc.to = AnchorMarker(col=6, colOff=500000, row=r_nc + 13, rowOff=0)
    ws.add_chart(dc, anc_dc)

    # ─── Bar Chart "Risk Change by Grade" ───
    rc_bar = BarChart()
    rc_bar.type = 'col'
    rc_bar.grouping = 'clustered'
    rc_bar.title = 'Risk Change by Grade'
    rc_bar.y_axis.delete = True
    rc_bar.y_axis.numFmt = ACCT
    rc_bar.y_axis.number_format = ACCT
    rc_bar.x_axis.delete = False
    rc_bar.x_axis.tickLblPos = 'nextTo'
    rc_bar.x_axis.majorTickMark = 'out'
    rc_bar.x_axis.minorTickMark = 'none'
    rc_bar.gapWidth = 100
    rc_bar.overlap = -10
    # Exclude last grade ("Not Reported") from chart
    rc_last = 7 + len(gl) - 2  # skip Not Reported row
    cats_rc = Reference(ws, min_col=1, min_row=7, max_row=rc_last)
    det_ref = Reference(ws, min_col=pcol_start, min_row=6, max_row=rc_last)
    imp_ref = Reference(ws, min_col=pcol_start + 1, min_row=6, max_row=rc_last)
    rc_bar.add_data(det_ref, titles_from_data=True)
    rc_bar.add_data(imp_ref, titles_from_data=True)
    rc_bar.set_categories(cats_rc)
    rc_bar.series[0].graphicalProperties.noFill = True
    rc_bar.series[0].graphicalProperties.line = LineProperties(solidFill='873A3A', w=38100)
    rc_bar.series[1].graphicalProperties.noFill = True
    rc_bar.series[1].graphicalProperties.line = LineProperties(solidFill='829901', w=38100)
    from openpyxl.chart.legend import Legend
    rc_bar.legend = Legend()
    rc_bar.legend.position = 't'
    rc_bar.legend.overlay = False
    rc_bar.legend.layout = Layout(
        manualLayout=ManualLayout(
            xMode='edge', yMode='edge',
            x=0.3, y=0.20,
            w=0.4, h=0.06,
        )
    )
    rc_bar.y_axis.majorGridlines = None
    rc_bar.width = 20
    rc_bar.height = 7.5
    anc_rc = TwoCellAnchor()
    anc_rc._from = AnchorMarker(col=5, colOff=0, row=r_nc - 2, rowOff=180000)
    anc_rc.to = AnchorMarker(col=9, colOff=0, row=r_nc + 13, rowOff=0)
    ws.add_chart(rc_bar, anc_rc)

    # ─── DQ Data Table (cols P-R, supporting the pie chart) ───
    r_dq = r_nc + 16
    ws.cell(row=r_dq, column=pcol_start, value="Loan Status").font = V12B
    ws.cell(row=r_dq, column=pcol_start + 1, value="DQ Balance").font = V12B
    ws.cell(row=r_dq, column=pcol_start + 2, value="% of Total").font = V12B
    dq_status = _dq_data
    for di, lbl in enumerate(["Improved", "Deteriorated", "Unchanged", "Not Reported"]):
        r = r_dq + 1 + di
        dq_entry = dq_status.get(lbl, {})
        ws.cell(row=r, column=pcol_start, value=lbl).font = V12
        ws.cell(row=r, column=pcol_start + 1, value=dq_entry.get('balance', 0)).number_format = '#,##0'
        ws.cell(row=r, column=pcol_start + 2, value=dq_entry.get('pct', 0)).number_format = '0.0%'

    # ─── DQ Pie Chart "Delinquency by Credit Grade Migration" ───
    dq_pie = PieChart()
    dq_pie.title = 'Delinquency by Credit Grade Migration'
    cats_dq = Reference(ws, min_col=pcol_start, min_row=r_dq + 1, max_row=r_dq + 4)
    vals_dq = Reference(ws, min_col=pcol_start + 2, min_row=r_dq + 1, max_row=r_dq + 4)
    dq_pie.add_data(vals_dq)
    dq_pie.set_categories(cats_dq)
    s_dq = dq_pie.series[0]
    s_dq.explosion = 21
    s_dq.graphicalProperties = GraphicalProperties()
    s_dq.graphicalProperties.noFill = True
    s_dq.graphicalProperties.line = LineProperties(solidFill='0D4D5E', w=38100)
    dq_pts = []
    for dp_idx, dp_color in enumerate(['829901', '873A3A', '0D4D5E', 'FFC000']):
        dp = DataPoint(idx=dp_idx)
        dp.graphicalProperties = GraphicalProperties()
        dp.graphicalProperties.noFill = True
        dp.graphicalProperties.line = LineProperties(solidFill=dp_color, w=38100)
        dq_pts.append(dp)
    s_dq.data_points = dq_pts
    # Data labels: DataLabelList defaults only.  openpyxl's DataLabel class
    # does not support the <delete> element, so zero-slice labels are hidden
    # by the post-processing step patch_dq_pie_zero_labels().
    s_dq.dLbls = DataLabelList()
    s_dq.dLbls.showVal = True
    s_dq.dLbls.showLegendKey = True
    s_dq.dLbls.showCatName = False
    s_dq.dLbls.showSerName = False
    s_dq.dLbls.showPercent = False
    s_dq.dLbls.showLeaderLines = True
    s_dq.dLbls.numFmt = '0.0%'
    s_dq.dLbls.dLblPos = 'outEnd'
    # Legend at bottom
    from openpyxl.chart.legend import Legend as Lgnd
    dq_pie.legend = Lgnd()
    dq_pie.legend.position = 'b'
    dq_pie.legend.overlay = False
    # Plot area layout – keep pie small to leave room for outEnd labels
    dq_pie.layout = Layout(
        manualLayout=ManualLayout(
            xMode='edge', yMode='edge',
            x=0.32, y=0.22,
            w=0.36, h=0.48,
        )
    )
    # Chart area: white fill, no border
    dq_pie.graphical_properties = GraphicalProperties()
    dq_pie.graphical_properties.solidFill = 'FFFFFF'
    dq_pie.graphical_properties.line = LineProperties(noFill=True)
    dq_pie.width = 15
    dq_pie.height = 7.5
    anc_dq = TwoCellAnchor()
    anc_dq._from = AnchorMarker(col=0, colOff=0, row=r_nc + 15, rowOff=0)
    anc_dq.to = AnchorMarker(col=5, colOff=0, row=r_nc + 28, rowOff=0)
    ws.add_chart(dq_pie, anc_dq)

    # ─── CO Data Table (cols P-R, supporting the bar chart) ───
    r_co = r_nc + 22
    ws.cell(row=r_co, column=pcol_start, value="Loan Status").font = V12B
    ws.cell(row=r_co, column=pcol_start + 1, value="CO Balance").font = V12B
    ws.cell(row=r_co, column=pcol_start + 2, value="% of Total").font = V12B
    co_status = _co_data
    for ci, lbl in enumerate(["Improved", "Deteriorated", "Unchanged", "Not Reported"]):
        r = r_co + 1 + ci
        co_entry = co_status.get(lbl, {})
        ws.cell(row=r, column=pcol_start, value=lbl).font = V12
        ws.cell(row=r, column=pcol_start + 1, value=co_entry.get('balance', 0)).number_format = '#,##0'
        ws.cell(row=r, column=pcol_start + 2, value=co_entry.get('pct', 0)).number_format = '0.0%'

    # ─── CO Bar Chart "Charge off by Credit Grade Migration" ───
    co_bar = BarChart()
    co_bar.type = 'bar'  # horizontal bars per template
    co_bar.grouping = 'clustered'
    co_bar.title = 'Charge off by Credit Grade Migration'
    co_bar.y_axis.delete = True  # hide value axis
    co_bar.x_axis.delete = False
    co_bar.x_axis.tickLblPos = 'nextTo'
    co_bar.x_axis.majorTickMark = 'out'
    co_bar.x_axis.minorTickMark = 'none'
    co_bar.gapWidth = 100
    cats_co = Reference(ws, min_col=pcol_start, min_row=r_co + 1, max_row=r_co + 4)
    vals_co = Reference(ws, min_col=pcol_start + 2, min_row=r_co + 1, max_row=r_co + 4)
    co_bar.add_data(vals_co)
    co_bar.set_categories(cats_co)
    s_co = co_bar.series[0]
    # Outline-only bars with 4pt borders
    s_co.graphicalProperties = GraphicalProperties()
    s_co.graphicalProperties.noFill = True
    s_co.graphicalProperties.line = LineProperties(solidFill='FFC000', w=50800)
    co_pts = []
    for dp_idx, dp_color in enumerate(['829901', '873A3A', '0D4D5E', 'FFC000']):
        dp = DataPoint(idx=dp_idx)
        dp.graphicalProperties = GraphicalProperties()
        dp.graphicalProperties.noFill = True
        dp.graphicalProperties.line = LineProperties(solidFill=dp_color, w=50800)
        co_pts.append(dp)
    s_co.data_points = co_pts
    # Data labels: show values, position at inner end
    s_co.dLbls = DataLabelList()
    s_co.dLbls.showVal = True
    s_co.dLbls.showLegendKey = False
    s_co.dLbls.showCatName = False
    s_co.dLbls.showSerName = False
    s_co.dLbls.showPercent = False
    s_co.dLbls.numFmt = '0.0%'
    s_co.dLbls.dLblPos = 'inEnd'
    # Remove major gridlines
    co_bar.y_axis.majorGridlines = None
    co_bar.x_axis.majorGridlines = None
    # Legend at bottom
    from openpyxl.chart.legend import Legend as Lgnd2
    co_bar.legend = Lgnd2()
    co_bar.legend.position = 'b'
    co_bar.legend.overlay = False
    # Shrink plot area so title sits above graph
    co_bar.layout = Layout(
        manualLayout=ManualLayout(
            layoutTarget='inner',
            xMode='edge', yMode='edge',
            x=0.18, y=0.20,
            w=0.78, h=0.62,
        )
    )
    # Chart area: white fill, no border
    co_bar.graphical_properties = GraphicalProperties()
    co_bar.graphical_properties.solidFill = 'FFFFFF'
    co_bar.graphical_properties.line = LineProperties(noFill=True)
    co_bar.width = 20
    co_bar.height = 7.5
    anc_co = TwoCellAnchor()
    anc_co._from = AnchorMarker(col=5, colOff=0, row=r_nc + 15, rowOff=0)
    anc_co.to = AnchorMarker(col=9, colOff=0, row=r_nc + 28, rowOff=0)
    ws.add_chart(co_bar, anc_co)

    # ─── Footnotes ───
    r_fn = r_nc + 29
    ws.cell(row=r_fn, column=2, value=(
        "Deteriorated loans are those whose FICO scores have dropped "
        "two or more grades or have dropped to a Grade lower than D."
    )).font = V10
    ws.cell(row=r_fn + 1, column=2, value=(
        "Improved loans are those whose FICO scores have moved up "
        "at least one grade."
    )).font = V10

    # Info icons next to footnotes (right-justified in col A)
    # Col A = 25.57 chars ≈ 184 px; icon = 20 px → colOff ≈ 164 px = 1,562,100 EMU
    _ico_off = 1562100
    if os.path.isfile(ICON_INFO_DARKRED):
        ico1 = XlImage(ICON_INFO_DARKRED)
        ico1.width = 20
        ico1.height = 20
        ico1.anchor = TwoCellAnchor()
        ico1.anchor._from = AnchorMarker(col=0, colOff=_ico_off, row=r_fn - 1, rowOff=0)
        ico1.anchor.to = AnchorMarker(col=1, colOff=0, row=r_fn, rowOff=0)
        ws.add_image(ico1)
    if os.path.isfile(ICON_INFO_DARKGREEN):
        ico2 = XlImage(ICON_INFO_DARKGREEN)
        ico2.width = 20
        ico2.height = 20
        ico2.anchor = TwoCellAnchor()
        ico2.anchor._from = AnchorMarker(col=0, colOff=_ico_off, row=r_fn, rowOff=0)
        ico2.anchor.to = AnchorMarker(col=1, colOff=0, row=r_fn + 1, rowOff=0)
        ws.add_image(ico2)

    # ─── Page Setup ───
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25, header=0, footer=0)
    last_row = r_fn + 1
    ws.print_area = f'A1:{get_column_letter(ncol)}{max(last_row, 66)}'


def _sheet_acl_reserve(wb, cu, snap, df, grades, config, hist):
    """ACL Env by Pool Mgmt Adj sheet matching Vizo template.

    Reads per-pool per-grade data from the WARM file when available
    (hist['impaired']['acl_pools']).  Falls back to computed values when
    WARM data is absent.  Hidden grades (Hide-*) are always excluded.
    """
    ws = wb.create_sheet("ACL Env by Pool Mgmt Adj")
    ws.sheet_view.showGridLines = False
    no_score = config.get('no_score_label', 'Not Reported')
    mgmt_adj_by_pool = config.get('mgmt_adj_by_pool', {})
    pool_use_default = _build_pool_use_default_map(config)
    admin_default_mgmt_adj = _load_admin_default_mgmt_adj()
    gl = _all_grades(grades, no_score)
    # Filter out hidden grades for display
    visible_gl = [g for g in gl if not _is_hidden(g)]

    # WARM-sourced ACL data (if available)
    _imp = hist.get('impaired', {}) if hist else {}
    econ_stress = _eco_stress(config, ed_override=_imp.get('economic_data'))
    _ncc_r, _dq_r, _es_r = _env_ranges(hist)

    pools = _ordered_pools(df, hist)
    dq_var = _pool_dq_variance(pools, hist, snap)
    acl_pools_data = _imp.get('acl_pools', {})
    acl_impaired = _imp.get('acl_impaired', {})
    acl_summary = _imp.get('acl_summary', {})
    prior_mgmt_adj = _imp.get('prior_mgmt_adj', {})
    prior_env_factor = _imp.get('prior_env_factor', {})
    spec_id_by_pool = _imp.get('spec_id_by_pool', {})  # from Impaired Loans detail

    # ── Compute per-pool Life Loss Rate matching Display Hist Bal formula ──
    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    avg_bals = hist.get('avg_balances', {}) if hist else {}
    years = hist.get('years', []) if hist else []
    acl_months_map = _imp.get('acl_months', {})
    snap_year = int(snap[:4])
    snap_month = int(snap[5:7])
    warm_net_co = _imp.get('warm_net_co', {})
    hbd = _imp.get('hist_bal_data', {})
    annual_grade_avg = {}
    for _pk, pdata in hbd.items():
        _dates = pdata.get('dates', [])
        _grades_data = pdata.get('grades', {})
        annual_grade_avg[_pk] = {}
        for _gk, _vals in _grades_data.items():
            if _gk.upper().startswith('HIDE'):
                continue
            yr_sums = {}
            yr_cnts = {}
            for _i, _d in enumerate(_dates):
                if _i < len(_vals) and _vals[_i] > 0:
                    yr_sums[_d.year] = yr_sums.get(_d.year, 0) + _vals[_i]
                    yr_cnts[_d.year] = yr_cnts.get(_d.year, 0) + 1
            for _y in yr_sums:
                annual_grade_avg[_pk].setdefault(_y, {})
                annual_grade_avg[_pk][_y][_gk] = yr_sums[_y] / yr_cnts[_y]

    life_loss = {}
    for pool in pools:
        pool_acl = acl_months_map.get(pool, 36)
        abs_first = (snap_year * 12 + snap_month) - pool_acl + 1
        pe = (abs_first - 1) // 12
        pa = annual_grade_avg.get(pool, {})
        yr_tots = []
        for y in years:
            if y < pe:
                continue
            yt = sum(pa.get(y, {}).values())
            if not yt:
                yt = avg_bals.get(y, {}).get(pool, 0)
            if yt:
                yr_tots.append(yt)
        avg_tot = sum(yr_tots) / len(yr_tots) if yr_tots else 0
        pool_stripped = pool.strip()
        net_co_match = warm_net_co.get(pool_stripped, warm_net_co.get(pool, None))
        if net_co_match is not None:
            total_net = net_co_match
        else:
            total_net = 0
            for y in years:
                if y < pe:
                    continue
                total_net += co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
        life_loss[pool] = total_net / avg_tot if avg_tot > 0 else 0

    # Column widths
    for ci, w in enumerate([22, 18, 16, 18, 16, 16, 16, 18, 14, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws['A1'] = cu
    ws['A1'].font = V14B
    ws['A2'] = "Allowance & Provision for Credit Loss Reserve Analysis"
    ws['A2'].font = V12B
    ws['A3'] = f"For Quarter Ending {_snap_display(snap)}"
    ws['A3'].font = V12B

    headers = ["Current Grade", "Balance", "Specific\nIdentification",
               "Loan Loss Calc\nBalance", "ACL Base\nLoss Rate",
               "Mgmt\nAdj", "Allowance\nFactor",
               "Allowance before\nEnv", "Env\nFactor",
               "Env\n Allowance", "Total\n Allowance"]

    # Write column headers once at row 5
    r = 5
    for hi, h in enumerate(headers):
        cell = ws.cell(row=r, column=1 + hi, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    r += 1

    grand_allowance = 0
    grand_allow_before = 0
    grand_env_allow = 0
    pool_starts = []
    pool_ends = []
    _bal_detail = _imp.get('pool_bal_detail', {})

    # Build unified pool list in WARM order, including WARM-only pools
    risk_rated_flags = _imp.get('risk_rated', {})
    warm_order = _imp.get('pool_order', [])
    db_pools = set(df['loan_pool'].unique())
    # Fall back to ordered DB pools when WARM file is unavailable
    if not warm_order:
        warm_order = list(pools)
    all_acl_pools = []
    seen = set()
    # Include pools coming from any source: WARM order, WARM ACL data, DB,
    # Pool_Balance Adjust detail, and WARM hist_bal_data. This ensures NRR
    # pools (Loan Participation, Repo/Foreclosed, etc.) that have no WARM
    # ACL section but DO have a balance from the monthly-balance file still
    # render on the Vizo ACL tab.
    extra_pools = list((_imp.get('hist_bal_data') or {}).keys()) \
                  + list(_bal_detail.keys()) \
                  + list(acl_pools_data.keys())
    candidates = list(warm_order) + list(pools) + extra_pools
    nrr_set = set(config.get('not_risk_rated', []) or [])
    for p in candidates:
        if not p:
            continue
        s = str(p).strip()
        if not s or s == 'Exclude' or s.upper().startswith('HIDE'):
            continue
        if s.lower() in ('total', 'grand total', 'excluded'):
            continue
        if p in seen:
            continue
        seen.add(p)
        all_acl_pools.append(p)

    for pool in all_acl_pools:
        pdf = df[df['loan_pool'] == pool]
        pool_total = pdf['current_balance'].sum()
        has_db_data = pool in db_pools

        # Try to find WARM data for this pool (fuzzy-match on stripped lowercase)
        _pool_lc = pool.strip().lower()
        warm_pool = next((v for k, v in acl_pools_data.items()
                          if k.strip().lower() == _pool_lc), None)
        warm_grades = warm_pool['grades'] if warm_pool else {}
        warm_total = warm_pool['total'] if warm_pool else {}

        # Determine if pool is risk-rated (default True for DB pools)
        if pool in nrr_set:
            is_rr = False
        else:
            is_rr = risk_rated_flags.get(pool, has_db_data)

        # Compute env factor (fallback if WARM total not available)
        if has_db_data:
            if is_rr:
                _, _, ncc_pct = _ncc(pdf, grades, config)
            else:
                ncc_pct = 0.0
            dq_v = dq_var.get(pool, 0)
            ncc_score = _score(ncc_pct * 100, _ncc_r)
            dq_score = _score(dq_v * 100, _dq_r)
            es_score = _score(econ_stress, _es_r)
            env_factor_calc = (ncc_score + dq_score + es_score) / 100.0
        else:
            env_factor_calc = 0
        # Use prior report's env factor if available; otherwise computed
        env_factor = prior_env_factor.get(pool, env_factor_calc)

        pool_ll = life_loss.get(pool, 0)

        # Pool header – bold black, no fill
        pool_start_row = r
        ws.cell(row=r, column=1, value=pool).font = V12B
        r += 1

        # Track pool start rows for page break placement
        pool_starts.append(pool_start_row)

        if is_rr:
            # ── Risk-rated pool: show per-grade detail ──
            pool_allow_before = 0
            for gi, g in enumerate(visible_gl):
                fnt = _grade_font(g)

                # Try WARM data first
                wg = warm_grades.get(g, {})
                # Prefer Pool_Balance Adjust source (balance_sheet_total) so
                # ACL Balance column matches that sheet's column D.
                _pd = _bal_detail.get(pool, {})
                _gd = _pd.get(g, {})
                bst = _gd.get('balance_sheet_total') if _gd else None
                if wg:
                    balance = bst if bst is not None else wg.get('balance', 0)
                    specific_id = wg.get('spec_id', 0)
                    # Override from Impaired Loans detail if WARM had 0
                    if specific_id == 0 and pool in spec_id_by_pool:
                        specific_id = spec_id_by_pool[pool].get(g, 0)
                    calc_bal = balance - specific_id
                    # ACL Base Loss Rate is always the pure calculated value
                    # (pool life loss * distribution factor) so it matches
                    # column M on the Display Hist Bal tab — even if zero.
                    dist = _dist_factor(len(DIST_FACTORS) - 1) if g == no_score else _dist_factor(gi)
                    base_rate = max(0, pool_ll * dist)
                    # Mgmt adj resolved via wizard Step 16 overlay + per-pool
                    # 'Use Default' checkbox + Admin firm-wide default.
                    # WARM workbook's baked-in mgmt_adj is intentionally
                    # ignored. Admin default only fires when base_rate==0.
                    mgmt_adj = _resolve_mgmt_adj_grade(
                        pool, g, gi, no_score,
                        pool_use_default, mgmt_adj_by_pool,
                        admin_default_mgmt_adj, prior_mgmt_adj,
                        base_rate=base_rate,
                    )
                    factor = base_rate + mgmt_adj
                    allow_before = calc_bal * factor
                elif has_db_data:
                    # Fallback: compute from data, using adjusted balance if available
                    if _gd and _gd.get('balance_sheet_total', 0):
                        balance = _gd['balance_sheet_total']
                    else:
                        g_df = pdf[pdf['current_grade'] == g]
                        balance = g_df['current_balance'].sum()
                    specific_id = spec_id_by_pool.get(pool, {}).get(g, 0)
                    calc_bal = balance - specific_id
                    # ACL Base Loss Rate is always the pure calculated value
                    # (pool life loss * distribution factor) so it matches
                    # column M on the Display Hist Bal tab — even if zero.
                    dist = _dist_factor(len(DIST_FACTORS) - 1) if g == no_score else _dist_factor(gi)
                    base_rate = max(0, pool_ll * dist)
                    mgmt_adj = _resolve_mgmt_adj_grade(
                        pool, g, gi, no_score,
                        pool_use_default, mgmt_adj_by_pool,
                        admin_default_mgmt_adj, prior_mgmt_adj,
                        base_rate=base_rate,
                    )
                    factor = base_rate + mgmt_adj
                    allow_before = calc_bal * factor
                else:
                    balance = specific_id = calc_bal = 0
                    base_rate = mgmt_adj = factor = allow_before = 0
                pool_allow_before += allow_before

                ws.cell(row=r, column=1, value=g).font = fnt
                ws.cell(row=r, column=2, value=balance).number_format = ACCT
                ws.cell(row=r, column=2).font = fnt
                ws.cell(row=r, column=3, value=specific_id).number_format = ACCT2
                ws.cell(row=r, column=3).font = fnt
                ws.cell(row=r, column=4, value=calc_bal).number_format = ACCT
                ws.cell(row=r, column=4).font = fnt
                ws.cell(row=r, column=5, value=base_rate).number_format = PCT4
                ws.cell(row=r, column=5).font = fnt
                ws.cell(row=r, column=6, value=mgmt_adj).number_format = PCT4
                ws.cell(row=r, column=6).font = fnt
                ws.cell(row=r, column=7, value=factor).number_format = PCT4
                ws.cell(row=r, column=7).font = fnt
                ws.cell(row=r, column=8, value=allow_before).number_format = ACCT
                ws.cell(row=r, column=8).font = fnt
                r += 1

            # Pool total row – use sum of per-grade allowances we computed above
            _ptd = _bal_detail.get(pool, {}).get('Total', {})
            if _ptd and _ptd.get('balance_sheet_total'):
                total_balance = _ptd['balance_sheet_total']
            elif warm_total:
                total_balance = warm_total.get('balance', pool_total)
            else:
                total_balance = pool_total
            pool_allow_before_out = pool_allow_before
            env_allow = pool_allow_before_out * env_factor
            total_allow = pool_allow_before_out + env_allow
            grand_allowance += total_allow
            grand_allow_before += pool_allow_before_out
            grand_env_allow += env_allow

            total_spec_id = warm_total.get('spec_id', 0) if warm_total else 0
            # Override from Impaired Loans detail if WARM had 0
            if total_spec_id == 0 and pool in spec_id_by_pool:
                total_spec_id = sum(spec_id_by_pool[pool].values())
            total_calc_bal = total_balance - total_spec_id

            ws.cell(row=r, column=1, value="Total").font = V12B
            ws.cell(row=r, column=2, value=total_balance).number_format = ACCT
            ws.cell(row=r, column=2).font = V12B
            ws.cell(row=r, column=3, value=total_spec_id).number_format = ACCT2
            ws.cell(row=r, column=3).font = V12B
            ws.cell(row=r, column=4, value=total_calc_bal).number_format = ACCT
            ws.cell(row=r, column=4).font = V12B
            ws.cell(row=r, column=8, value=pool_allow_before_out).number_format = ACCT
            ws.cell(row=r, column=8).font = V12B
            ws.cell(row=r, column=9, value=env_factor).number_format = PCT
            ws.cell(row=r, column=9).font = V12B
            ws.cell(row=r, column=10, value=env_allow).number_format = ACCT
            ws.cell(row=r, column=10).font = V12B
            ws.cell(row=r, column=11, value=total_allow).number_format = ACCT
            ws.cell(row=r, column=11).font = V12B
            pool_ends.append(r)   # last printed row of this pool block
            r += 2
        else:
            # ── Non-risk-rated pool: show only Total row with rate columns ──
            _ptd_nrr = _bal_detail.get(pool, {}).get('Total', {})
            if _ptd_nrr and _ptd_nrr.get('balance_sheet_total') is not None:
                nrr_balance = _ptd_nrr['balance_sheet_total']
            else:
                nrr_balance = warm_total.get('balance', 0)
            nrr_spec_id = warm_total.get('spec_id', 0)
            if nrr_spec_id == 0 and pool in spec_id_by_pool:
                nrr_spec_id = sum(spec_id_by_pool[pool].values())
            nrr_calc_bal = nrr_balance - nrr_spec_id
            nrr_base_rate = warm_total.get('base_rate', 0)
            # NRR pool mgmt adj: resolver + admin default (gated on
            # nrr_base_rate==0). Recompute factor/allow_before so they
            # reflect the resolver's value rather than WARM's baked-in.
            nrr_mgmt_adj = _resolve_mgmt_adj_total(
                pool, pool_use_default, mgmt_adj_by_pool,
                admin_default_mgmt_adj,
                base_rate=nrr_base_rate,
            )
            nrr_factor = nrr_base_rate + nrr_mgmt_adj
            nrr_allow_before = nrr_calc_bal * nrr_factor
            nrr_env_factor = env_factor  # use computed value to match Env Factor by Pool
            nrr_env_allow = nrr_allow_before * nrr_env_factor
            nrr_total_allow = nrr_allow_before + nrr_env_allow
            grand_allowance += nrr_total_allow
            grand_allow_before += nrr_allow_before
            grand_env_allow += nrr_env_allow

            ws.cell(row=r, column=1, value="Total").font = V12B
            ws.cell(row=r, column=2, value=nrr_balance).number_format = ACCT
            ws.cell(row=r, column=2).font = V12B
            ws.cell(row=r, column=3, value=nrr_spec_id).number_format = ACCT2
            ws.cell(row=r, column=3).font = V12B
            ws.cell(row=r, column=4, value=nrr_calc_bal).number_format = ACCT
            ws.cell(row=r, column=4).font = V12B
            ws.cell(row=r, column=5, value=nrr_base_rate).number_format = PCT4
            ws.cell(row=r, column=5).font = V12B
            ws.cell(row=r, column=6, value=nrr_mgmt_adj).number_format = PCT4
            ws.cell(row=r, column=6).font = V12B
            ws.cell(row=r, column=7, value=nrr_factor).number_format = PCT4
            ws.cell(row=r, column=7).font = V12B
            ws.cell(row=r, column=8, value=nrr_allow_before).number_format = ACCT
            ws.cell(row=r, column=8).font = V12B
            ws.cell(row=r, column=9, value=nrr_env_factor).number_format = PCT
            ws.cell(row=r, column=9).font = V12B
            ws.cell(row=r, column=10, value=nrr_env_allow).number_format = ACCT
            ws.cell(row=r, column=10).font = V12B
            ws.cell(row=r, column=11, value=nrr_total_allow).number_format = ACCT
            ws.cell(row=r, column=11).font = V12B
            pool_ends.append(r)   # last printed row of this pool block
            r += 2

    # Grand totals — use computed sums across all pools
    pooled_balance = acl_summary.get('pooled_balance', df['current_balance'].sum())
    pooled_total_allow = grand_allowance

    # Stash the computed pooled total back onto hist['impaired'] so the
    # Impr Deter sheet (built after this one in compose_vizo_main) can reuse
    # the exact same number via _compute_acl_totals. Without this, Impr Deter
    # would fall back to imp['pooled_total_allowance'] (the prior WARM-parsed
    # value) and the two tabs' Total Allowance Needed lines would disagree.
    if isinstance(_imp, dict):
        _imp['_computed_pooled_total_allow'] = pooled_total_allow
        _imp['_computed_grand_allow_before'] = grand_allow_before
        _imp['_computed_grand_env_allow'] = grand_env_allow

    pooled_spec_id = acl_summary.get('pooled_spec_id', 0)
    # Override from Impaired Loans detail if WARM had 0
    if pooled_spec_id == 0 and spec_id_by_pool:
        pooled_spec_id = sum(sum(g.values()) for g in spec_id_by_pool.values())
    pooled_calc_bal = pooled_balance - pooled_spec_id

    ws.cell(row=r, column=1, value="Pooled Totals").font = V12B
    ws.cell(row=r, column=2, value=pooled_balance).number_format = ACCT
    ws.cell(row=r, column=2).font = V12B
    ws.cell(row=r, column=3, value=pooled_spec_id).number_format = ACCT2
    ws.cell(row=r, column=3).font = V12B
    ws.cell(row=r, column=4, value=pooled_calc_bal).number_format = ACCT
    ws.cell(row=r, column=4).font = V12B
    ws.cell(row=r, column=8, value=grand_allow_before).number_format = ACCT
    ws.cell(row=r, column=8).font = V12B
    ws.cell(row=r, column=10, value=grand_env_allow).number_format = ACCT
    ws.cell(row=r, column=10).font = V12B
    ws.cell(row=r, column=11, value=pooled_total_allow).number_format = ACCT
    ws.cell(row=r, column=11).font = V12B

    r += 2
    ws.cell(row=r, column=1, value="Impaired Loans").font = V12B
    ws.cell(row=r, column=10, value="Allowance").font = V12B
    for lbl in ["Delinquent Loans", "Known Losses", "Repossessions",
                "Foreclosed Real Estate", "Deceased", "Bankruptcy"]:
        imp_val = acl_impaired.get(lbl, 0)
        if lbl.upper().startswith('HIDE'):
            continue
        r += 1
        ws.cell(row=r, column=1, value=lbl).font = V12
        ws.cell(row=r, column=11, value=imp_val).number_format = ACCT
    total_spec_allow = acl_summary.get('total_spec_allow', sum(acl_impaired.values()))
    total_allow_needed = pooled_total_allow + total_spec_allow
    acl_bal = acl_summary.get('acl_balance', config.get('acl_balance', 0))
    adjustment = total_allow_needed - acl_bal

    r += 1
    ws.cell(row=r, column=1, value="Total Specifically Identified Allowance").font = V12B
    ws.cell(row=r, column=11, value=total_spec_allow).number_format = ACCT
    ws.cell(row=r, column=11).font = V12B
    r += 1
    ws.cell(row=r, column=1, value="Total Allowance Needed").font = V12B
    ws.cell(row=r, column=11, value=total_allow_needed).number_format = ACCT
    ws.cell(row=r, column=11).font = V12B
    r += 1
    ws.cell(row=r, column=1, value=f"Allowance for Credit Loss Balance as of {snap}").font = V12
    ws.cell(row=r, column=11, value=acl_bal).number_format = ACCT
    r += 1
    ws.cell(row=r, column=1, value="Adjustment (Overfunded)").font = V12B
    ws.cell(row=r, column=11, value=adjustment).number_format = ACCT
    ws.cell(row=r, column=11).font = V12B

    # ─── Page Setup ───
    # Greedy bin-packing: fit as many complete pool blocks per page as the
    # printable area allows, never splitting a pool across pages.
    # Landscape Letter @ 0.25" margins with the standard default row height
    # comfortably prints ~55 rows on page 1; pages 2+ repeat title rows
    # 1:5 via print_title_rows, leaving ~50 content rows.
    PAGE1_ROWS  = 45
    OTHER_ROWS  = 40  # = PAGE1_ROWS - 5 repeated title rows

    # Row budget is the index of the last row that may appear on the
    # current page.  Start with all of page 1 available.
    page_bottom = PAGE1_ROWS
    for ps, pe in zip(pool_starts, pool_ends):
        block_end = pe + 1   # include the trailing blank row
        if block_end > page_bottom:
            # This pool would spill onto the next page – break before it.
            ws.row_breaks.append(Break(id=ps - 1))
            page_bottom = ps + OTHER_ROWS - 1

    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25,
                                  header=0.3, footer=0.3)
    ws.print_area = f'A1:K{r}'
    ws.print_title_rows = '1:5'


def _sheet_env_factor(wb, cu, snap, df, grades, config, hist):
    """Environmental Factor by Pool matching Vizo template."""
    ws = wb.create_sheet("Env Factor by Pool")
    ed = config.get('economic_data', {})
    # Prefer economic data from the WARM file when available
    _imp = hist.get('impaired', {}) if hist else {}
    if _imp.get('economic_data'):
        ed = _imp['economic_data']
    econ_stress = _eco_stress(config, ed_override=ed)
    _ncc_r, _dq_r, _es_r = _env_ranges(hist)
    no_score = config.get('no_score_label', 'Not Reported')

    pools = _ordered_pools(df, hist)
    dq_var = _pool_dq_variance(pools, hist, snap)

    # Column widths
    for ci, w in enumerate([22, 15.6, 18.7, 14.7, 15.3, 20.7, 11.7, 15.6], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws['A1'] = cu
    ws['A1'].font = V14B
    ws['A2'] = "Environmental Factor for PLL"
    ws['A2'].font = V12B
    ws['A3'] = f"For Quarter Ending {_snap_display(snap)}"
    ws['A3'].font = V12B

    # Economic Stress Index section ── light gold fill, black text, no borders
    ESI_FILL = PatternFill('solid', fgColor='CFDBDF')  # Teal Accent4 80% lighter
    ESI_HDR  = Font(name='Calibri', bold=True, size=11)
    ESI_DATA = Font(name='Calibri', size=11)
    ESI_DATA10 = Font(name='Calibri', size=10)
    _esi_align_hdr = Alignment(horizontal='left', vertical='center', wrap_text=True)
    _esi_align_val = Alignment(horizontal='right', vertical='center')
    _esi_align_lbl = Alignment(horizontal='left', vertical='center')

    r = 5
    ws.cell(row=r, column=1, value="Economic Stress Index Calculation").font = V18B
    for ci in range(1, 6):
        ws.cell(row=r, column=ci).fill = ESI_FILL
    ws.row_dimensions[r].height = 23.25

    r = 6
    for ci, lbl in enumerate(["State", "Unemployment Rate", "Foreclosures per Person",
                               "Bankruptcies", "Population"], start=1):
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.font = ESI_HDR
        cell.fill = ESI_FILL
        cell.alignment = _esi_align_hdr
    ws.row_dimensions[r].height = 30.0

    r = 7
    pop = ed.get('population', 1)
    ws.cell(row=r, column=1, value=ed.get('state', '')).font = ESI_DATA10
    ws.cell(row=r, column=1).alignment = _esi_align_lbl
    ws.cell(row=r, column=2, value=ed.get('unemployment_rate', 0)).number_format = PCT
    ws.cell(row=r, column=2).font = ESI_DATA
    ws.cell(row=r, column=2).alignment = _esi_align_val
    ws.cell(row=r, column=3, value=ed.get('foreclosures', 0)).number_format = ACCT
    ws.cell(row=r, column=3).font = ESI_DATA
    ws.cell(row=r, column=3).alignment = _esi_align_val
    ws.cell(row=r, column=4, value=ed.get('bankruptcies', 0)).number_format = ACCT
    ws.cell(row=r, column=4).font = ESI_DATA
    ws.cell(row=r, column=4).alignment = _esi_align_val
    ws.cell(row=r, column=5, value=pop).number_format = ACCT
    ws.cell(row=r, column=5).font = ESI_DATA
    ws.cell(row=r, column=5).alignment = _esi_align_val
    for ci in range(1, 6):
        ws.cell(row=r, column=ci).fill = ESI_FILL

    r = 8
    for ci, lbl in enumerate(["County", "Unemployment Rate", "Bankruptcy %",
                               "Foreclosure %", "Economic Stress Index"], start=1):
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.font = ESI_HDR
        cell.fill = ESI_FILL
        cell.alignment = _esi_align_hdr
    ws.row_dimensions[r].height = 30.0

    r = 9
    ws.cell(row=r, column=1, value=ed.get('county', '')).font = ESI_DATA10
    ws.cell(row=r, column=1).alignment = _esi_align_lbl
    ws.cell(row=r, column=2, value=ed.get('unemployment_rate', 0)).number_format = PCT
    ws.cell(row=r, column=2).font = ESI_DATA
    ws.cell(row=r, column=2).alignment = _esi_align_val
    bk_pct = ed.get('bankruptcies', 0) / pop if pop else 0
    fc_pct = ed.get('foreclosures', 0) / pop if pop else 0
    ws.cell(row=r, column=3, value=bk_pct).number_format = PCT
    ws.cell(row=r, column=3).font = ESI_DATA
    ws.cell(row=r, column=3).alignment = _esi_align_val
    ws.cell(row=r, column=4, value=fc_pct).number_format = PCT
    ws.cell(row=r, column=4).font = ESI_DATA
    ws.cell(row=r, column=4).alignment = _esi_align_val
    ws.cell(row=r, column=5, value=econ_stress / 100).number_format = PCT
    ws.cell(row=r, column=5).font = ESI_DATA
    ws.cell(row=r, column=5).alignment = _esi_align_val
    for ci in range(1, 6):
        ws.cell(row=r, column=ci).fill = ESI_FILL

    # Per-pool environmental factors
    r = 12
    pool_headers = ["Portfolio Segment", "Net Credit\nChange", "Net Credit\nScore",
                    "Delinquency\nVariance from Ave.", "Delinquency\nScore",
                    "Economic Stress\nActual", "Economic Stress\nScore",
                    "Environmental\nFactor"]
    for ci, lbl in enumerate(pool_headers, start=1):
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='left', wrap_text=True)
    ws.row_dimensions[r].height = 47.25

    risk_rated_map = _imp.get('risk_rated', {})
    for pool in pools:
        r += 1
        pdf = df[df['loan_pool'] == pool]
        is_rr = risk_rated_map.get(pool, True)
        if is_rr:
            _, _, ncc_pct = _ncc(pdf, grades, config)
        else:
            ncc_pct = 0.0
        ncc_score = _score(ncc_pct * 100, _ncc_r) / 100.0
        dq_v = dq_var.get(pool, 0)
        dq_score_val = _score(dq_v * 100, _dq_r) / 100.0
        es_score = _score(econ_stress, _es_r) / 100.0
        env_f = ncc_score + dq_score_val + es_score

        ws.cell(row=r, column=1, value=pool).font = V10
        ws.cell(row=r, column=2, value=ncc_pct).number_format = PCT
        ws.cell(row=r, column=2).font = V11
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3, value=ncc_score).number_format = PCT
        ws.cell(row=r, column=3).font = V11
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4, value=dq_v).number_format = PCT
        ws.cell(row=r, column=4).font = V11
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=5, value=dq_score_val).number_format = PCT
        ws.cell(row=r, column=5).font = V11
        ws.cell(row=r, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=6, value=econ_stress / 100).number_format = PCT
        ws.cell(row=r, column=6).font = V11
        ws.cell(row=r, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=7, value=es_score).number_format = PCT
        ws.cell(row=r, column=7).font = V11
        ws.cell(row=r, column=7).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=8, value=env_f).number_format = PCT
        ws.cell(row=r, column=8).font = V11
        ws.cell(row=r, column=8).alignment = Alignment(horizontal='center')

    # ── Footnotes: data sources ──────────────────────────────────
    sources = ed.get('_sources', {})
    if sources:
        fn_font = Font(name='Calibri', size=8, italic=True, color='555555')
        r += 2  # skip a blank row
        ws.cell(row=r, column=1, value="Data Sources:").font = Font(
            name='Calibri', size=8, bold=True, italic=True, color='555555')
        for field_label, source_keys in [
            ("Unemployment Rate", "unemployment_rate"),
            ("Population", "population"),
            ("Bankruptcies", "bankruptcies"),
            ("Foreclosures", "foreclosures"),
        ]:
            src = sources.get(source_keys)
            if src:
                r += 1
                ws.cell(row=r, column=1,
                        value=f"  {field_label}: {src}").font = fn_font

    # Page setup: landscape, fit all columns on one page, narrow margins
    ws.page_setup.orientation = 'landscape'
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25,
                                  header=0, footer=0)


def _sheet_loss_factor(wb, cu, snap, df, grades, config, hist):
    """Display Hist Bal – Loss Factor Calculation matching Vizo template."""
    ws = wb.create_sheet("Display HIst Bal")
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g for g in _all_grades(grades, no_score) if not _is_hidden(g)]

    pools = _ordered_pools(df, hist)

    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    avg_bals = hist.get('avg_balances', {}) if hist else {}
    years = hist.get('years', []) if hist else []

    # ACL months (life of loan) per pool
    _imp = hist.get('impaired', {}) if hist else {}
    acl_months_map = _imp.get('acl_months', {})
    snap_year = int(snap[:4])
    snap_month = int(snap[5:7])

    # Trim leading years to the deepest pool's Life-of-Loan window.
    # Years older than the longest LoL across all pools aren't used by
    # any pool's calculation and just clutter the year axis.
    if pools and years:
        _max_lol = max(acl_months_map.get(p, 36) for p in pools)
        _abs_first = (snap_year * 12 + snap_month) - _max_lol + 1
        _cutoff_year = (_abs_first - 1) // 12
        years = [y for y in years if y >= _cutoff_year]

    # ── Per-grade annual averages from WARM hist_bal_data ──
    hbd = hist.get('impaired', {}).get('hist_bal_data', {}) if hist else {}
    annual_grade_avg = {}  # {pool: {year: {grade: avg_bal}}}
    for _pk, pdata in hbd.items():
        _dates = pdata.get('dates', [])
        _grades_data = pdata.get('grades', {})
        annual_grade_avg[_pk] = {}
        for _gk, _vals in _grades_data.items():
            if _gk.upper().startswith('HIDE'):
                continue
            yr_sums = {}
            yr_cnts = {}
            for _i, _d in enumerate(_dates):
                if _i < len(_vals) and _vals[_i] > 0:
                    yr_sums[_d.year] = yr_sums.get(_d.year, 0) + _vals[_i]
                    yr_cnts[_d.year] = yr_cnts.get(_d.year, 0) + 1
            for _y in yr_sums:
                annual_grade_avg[_pk].setdefault(_y, {})
                annual_grade_avg[_pk][_y][_gk] = yr_sums[_y] / yr_cnts[_y]

    def _pool_earliest_year(pool):
        """Return earliest year with data for this pool based on ACL months."""
        pool_acl = acl_months_map.get(pool, 36)
        abs_first = (snap_year * 12 + snap_month) - pool_acl + 1
        return (abs_first - 1) // 12

    # ── Pre-compute per-pool Life Loss Rate (matches WARM formula) ──
    # WARM: Life Loss Rate = Total Net Chargeoffs / Average of yearly pool totals
    warm_net_co = _imp.get('warm_net_co', {})
    pool_life_rates = {}
    pool_avg_totals = {}
    for pool in pools:
        pe = _pool_earliest_year(pool)
        pa = annual_grade_avg.get(pool, {})
        yr_tots = []
        for y in years:
            if y < pe:
                continue
            yt = sum(pa.get(y, {}).values())
            if not yt:
                yt = avg_bals.get(y, {}).get(pool, 0)
            if yt:
                yr_tots.append(yt)
        avg_tot = sum(yr_tots) / len(yr_tots) if yr_tots else 0
        pool_avg_totals[pool] = avg_tot
        pool_stripped = pool.strip()
        net_co_match = warm_net_co.get(pool_stripped, warm_net_co.get(pool, None))
        if net_co_match is not None:
            total_net = net_co_match
        else:
            total_net = 0
            for y in years:
                if y < pe:
                    continue
                total_net += co_data.get(y, {}).get(pool, 0) - rc_data.get(y, {}).get(pool, 0)
        pool_life_rates[pool] = total_net / avg_tot if avg_tot > 0 else 0

    # Column widths: A wide for grade labels, B-P uniform 12.33, Q+ defaults
    ws.column_dimensions['A'].width = 26
    for _ci in range(2, 17):  # columns B (2) through P (16)
        ws.column_dimensions[get_column_letter(_ci)].width = 12.33
    # Column J (right-side "Current Grade" label column) matches column A width
    ws.column_dimensions['J'].width = 26

    # Title block
    ws['A1'] = cu
    ws['A1'].font = V14B
    ws.row_dimensions[1].height = 18.75
    ws['A2'] = "Loss Factor Calculation"
    ws['A2'].font = V12B
    ws['A3'] = f"For Quarter Ending {_snap_display(snap)}"
    ws['A3'].font = V12B

    year_strs = [str(y) for y in years]
    # Determine column range for year data
    # Left side: A=Grade, B..I=years (up to 8), J=Average Balance
    # Right side: K=Grade, L=Avg Bal, M=Life Loss Rate, N=Dist Factor, O=ACL Base, P=% of Loans, Q=WARM
    num_years = len(years)
    year_start_col = 2  # Column B
    avg_col = year_start_col + num_years  # After years
    # Right-side columns
    right_start = avg_col + 1

    r = 5
    # Header row
    _hdr_row(ws, r, right_start + 6)
    ws.cell(row=r, column=1, value="Current Grade").font = HDR_FONT
    ws.cell(row=r, column=1).fill = HDR_FILL
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', wrap_text=True)
    for yi, ys in enumerate(year_strs):
        lbl = f"YTD {ys}" if yi == num_years - 1 else ys
        ws.cell(row=r, column=year_start_col + yi, value=lbl).font = HDR_FONT
        ws.cell(row=r, column=year_start_col + yi).fill = HDR_FILL
        ws.cell(row=r, column=year_start_col + yi).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=avg_col, value="Average Balance").font = HDR_FONT
    ws.cell(row=r, column=avg_col).fill = HDR_FILL
    ws.cell(row=r, column=avg_col).alignment = Alignment(horizontal='center', wrap_text=True)

    ws.cell(row=r, column=right_start, value="Current Grade").font = HDR_FONT
    ws.cell(row=r, column=right_start).fill = HDR_FILL
    ws.cell(row=r, column=right_start + 1, value="Average Balance").font = HDR_FONT
    ws.cell(row=r, column=right_start + 1).fill = HDR_FILL
    ws.cell(row=r, column=right_start + 2, value="Life Loss Rate").font = HDR_FONT
    ws.cell(row=r, column=right_start + 2).fill = HDR_FILL
    ws.cell(row=r, column=right_start + 3, value="Distribution Factor").font = HDR_FONT
    ws.cell(row=r, column=right_start + 3).fill = HDR_FILL
    ws.cell(row=r, column=right_start + 4, value="ACL Base Loss Rate").font = HDR_FONT
    ws.cell(row=r, column=right_start + 4).fill = HDR_FILL
    ws.cell(row=r, column=right_start + 5, value="% of Loans").font = HDR_FONT
    ws.cell(row=r, column=right_start + 5).fill = HDR_FILL
    ws.cell(row=r, column=right_start + 6, value="WARM Months").font = HDR_FONT
    ws.cell(row=r, column=right_start + 6).fill = HDR_FILL
    # Remove cell borders from the Current Grade header row per design spec
    from openpyxl.styles import Border as _NoBorder
    _no_border = _NoBorder()
    for _c in range(1, right_start + 7):
        ws.cell(row=r, column=_c).border = _no_border
    ws.row_dimensions[r].height = 36.0

    pool_ranges = []  # (start_row, end_row) for each pool block
    risk_rated_map = hist.get('impaired', {}).get('risk_rated', {}) if hist else {}
    for pool in pools:
        r += 1
        pool_start_r = r
        # Pool header
        ws.cell(row=r, column=1, value=pool).font = V14B
        ws.cell(row=r, column=right_start, value=pool).font = V14B
        ws.row_dimensions[r].height = 18.75

        pdf = df[df['loan_pool'] == pool]
        pool_total = pdf['current_balance'].sum()
        pool_ll = pool_life_rates.get(pool, 0)
        pool_earliest = _pool_earliest_year(pool)
        is_rr = risk_rated_map.get(pool, True)

        if not is_rr:
            # Non-risk-rated pool: total row only (no grade breakdown)
            r += 1
            ws.cell(row=r, column=1, value="Total").font = V12B
            ws.cell(row=r, column=1).number_format = DOLLAR
            pool_annual = annual_grade_avg.get(pool, {})
            for yi in range(num_years):
                if years[yi] < pool_earliest:
                    continue
                yr_total = sum(pool_annual.get(years[yi], {}).values())
                if not yr_total:
                    yr_total = avg_bals.get(years[yi], {}).get(pool, 0)
                if yr_total:
                    ws.cell(row=r, column=year_start_col + yi, value=yr_total).number_format = ACCT
                    ws.cell(row=r, column=year_start_col + yi).font = V12B
            nrr_avg = pool_avg_totals.get(pool, 0)
            ws.cell(row=r, column=avg_col, value=nrr_avg).number_format = ACCT
            ws.cell(row=r, column=avg_col).font = V12B
            ws.cell(row=r, column=right_start, value="Total").font = V12B
            ws.cell(row=r, column=right_start).number_format = DOLLAR
            ws.cell(row=r, column=right_start + 1, value=nrr_avg).number_format = ACCT
            ws.cell(row=r, column=right_start + 1).font = V12B
            ws.cell(row=r, column=right_start + 2, value=pool_ll).number_format = PCT
            ws.cell(row=r, column=right_start + 2).font = V12B
            ws.cell(row=r, column=right_start + 5, value=1.0).number_format = PCT
            ws.cell(row=r, column=right_start + 5).font = V12B
            warm = acl_months_map.get(pool, config.get('warm_months', {}).get(pool, 36))
            ws.cell(row=r, column=right_start + 6, value=warm).font = V12
            ws.cell(row=r, column=right_start + 6).alignment = Alignment(horizontal='center')
            pool_ranges.append((pool_start_r, r))
            # Small spacer row
            r += 1
            ws.row_dimensions[r].height = 6.0
            continue

        for gi, g in enumerate(gl):
            r += 1
            fnt = _grade_font(g)
            g_df = pdf[pdf['current_grade'] == g]
            balance = g_df['current_balance'].sum()

            ws.cell(row=r, column=1, value=g).font = fnt
            ws.cell(row=r, column=1).number_format = DOLLAR

            # Year balance columns – use actual WARM hist_bal_data per-grade averages
            pool_annual = annual_grade_avg.get(pool, {})
            yr_vals = []
            for yi in range(num_years):
                if years[yi] < pool_earliest:
                    continue  # skip years outside life of loan
                grade_avg = pool_annual.get(years[yi], {}).get(g, 0)
                if not grade_avg:
                    # Fallback: approximate from pool-level avg_balances
                    avg = avg_bals.get(years[yi], {}).get(pool, 0)
                    grade_avg = avg * (balance / pool_total) if pool_total and avg else 0
                if grade_avg:
                    yr_vals.append(grade_avg)
                    ws.cell(row=r, column=year_start_col + yi, value=grade_avg).number_format = ACCT
                    ws.cell(row=r, column=year_start_col + yi).font = fnt

            # Average balance = mean of yearly values (WARM formula)
            avg_bal = sum(yr_vals) / len(yr_vals) if yr_vals else 0
            ws.cell(row=r, column=avg_col, value=avg_bal).number_format = ACCT
            ws.cell(row=r, column=avg_col).font = fnt

            # Right side
            ws.cell(row=r, column=right_start, value=g).font = fnt
            ws.cell(row=r, column=right_start).number_format = DOLLAR
            ws.cell(row=r, column=right_start + 1, value=avg_bal).number_format = ACCT
            ws.cell(row=r, column=right_start + 1).font = V12

            # Not Reported uses last DIST_FACTORS entry (skipping hidden grades)
            dist = _dist_factor(len(DIST_FACTORS) - 1) if g == no_score else _dist_factor(gi)
            base_rate = max(0, pool_ll * dist)
            pct_pool = balance / pool_total if pool_total else 0

            ws.cell(row=r, column=right_start + 2, value=pool_ll).number_format = PCT
            ws.cell(row=r, column=right_start + 2).font = fnt
            ws.cell(row=r, column=right_start + 3, value=dist).number_format = PCT
            ws.cell(row=r, column=right_start + 3).font = fnt
            ws.cell(row=r, column=right_start + 4, value=base_rate).number_format = PCT
            ws.cell(row=r, column=right_start + 4).font = fnt
            ws.cell(row=r, column=right_start + 5, value=pct_pool).number_format = PCT
            ws.cell(row=r, column=right_start + 5).font = fnt

            # WARM months - only for first real grade per pool
            if gi == 0:
                warm = acl_months_map.get(pool, config.get('warm_months', {}).get(pool, 36))
                ws.cell(row=r, column=right_start + 6, value=warm).font = V12
                ws.cell(row=r, column=right_start + 6).alignment = Alignment(horizontal='center')

        # Pool total row
        r += 1
        ws.cell(row=r, column=1, value="Total").font = V12B
        ws.cell(row=r, column=1).number_format = DOLLAR
        # Year total columns – sum grade-level annual averages
        pool_annual = annual_grade_avg.get(pool, {})
        for yi in range(num_years):
            if years[yi] < pool_earliest:
                continue
            yr_total = sum(pool_annual.get(years[yi], {}).values())
            if yr_total:
                ws.cell(row=r, column=year_start_col + yi, value=yr_total).number_format = ACCT
                ws.cell(row=r, column=year_start_col + yi).font = V12B
        rr_avg = pool_avg_totals.get(pool, 0)
        ws.cell(row=r, column=avg_col, value=rr_avg).number_format = ACCT
        ws.cell(row=r, column=avg_col).font = V12B
        ws.cell(row=r, column=right_start, value="Total").font = V12B
        ws.cell(row=r, column=right_start).number_format = DOLLAR
        ws.cell(row=r, column=right_start + 1, value=rr_avg).number_format = ACCT
        ws.cell(row=r, column=right_start + 1).font = V12B
        ws.cell(row=r, column=right_start + 2, value=pool_ll).number_format = PCT
        ws.cell(row=r, column=right_start + 2).font = V12B
        ws.cell(row=r, column=right_start + 5, value=1.0).number_format = PCT
        ws.cell(row=r, column=right_start + 5).font = V12B
        pool_ranges.append((pool_start_r, r))

        # Small spacer row
        r += 1
        ws.row_dimensions[r].height = 6.0

    # ── Grand Total row across all pools ──
    r += 1
    gt_row = r
    ws.cell(row=r, column=1, value="Grand Total").font = V12B
    # Yearly grand totals = sum of per-pool annual averages across all pools
    for yi in range(num_years):
        y = years[yi]
        ytot = 0.0
        for pool in pools:
            ytot += sum(annual_grade_avg.get(pool, {}).get(y, {}).values())
        if ytot:
            c = ws.cell(row=r, column=year_start_col + yi, value=ytot)
            c.number_format = ACCT
            c.font = V12B
    grand_avg = sum(pool_avg_totals.get(p, 0) for p in pools)
    cga = ws.cell(row=r, column=avg_col, value=grand_avg)
    cga.number_format = ACCT
    cga.font = V12B
    # Right-side Grand Total label & average
    ws.cell(row=r, column=right_start, value="Grand Total").font = V12B
    cgr = ws.cell(row=r, column=right_start + 1, value=grand_avg)
    cgr.number_format = ACCT
    cgr.font = V12B
    ws.cell(row=r, column=right_start + 5, value=1.0).number_format = PCT
    ws.cell(row=r, column=right_start + 5).font = V12B

    # ── Ensure all numeric data cells remain visible at the narrow column
    # widths set above by enabling shrink_to_fit (so values like large dollar
    # totals never render as ##### in columns B-P).  Text cells are skipped so
    # long labels (e.g. pool titles like "Participation Loans") can still
    # overflow into adjacent empty cells normally.
    for _row in range(6, r + 1):
        for _col in range(2, 17):  # B (2) through P (16)
            _cell = ws.cell(row=_row, column=_col)
            if not isinstance(_cell.value, (int, float)) or isinstance(_cell.value, bool):
                continue
            _existing = _cell.alignment
            _cell.alignment = Alignment(
                horizontal=_existing.horizontal,
                vertical=_existing.vertical,
                wrap_text=_existing.wrap_text,
                shrink_to_fit=True,
            )

    # ── Page Setup ──
    # Limit output to 2 pages wide x 2 pages tall.  Force a vertical page break
    # between columns I and J so the left-side balance grid prints on page 1
    # and the right-side rate summary prints on page 2.
    ws.col_breaks.append(Break(id=9))
    last_col = get_column_letter(right_start + 6)
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 0
    ws.page_setup.fitToHeight = 2
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25,
                                  header=0.3, footer=0.3)
    ws.print_area = f'A1:{last_col}{r}'
    ws.print_title_rows = '1:5'


def _sheet_co_recov_dq(wb, cu, snap, df, config, hist):
    """Display CO-Recov-DQ sheet matching Vizo template."""
    ws = wb.create_sheet("Display CO-Recov-DQ")

    pools = _ordered_pools(df, hist)
    co_data = hist.get('chargeoffs', {}) if hist else {}
    rc_data = hist.get('recoveries', {}) if hist else {}
    dq_pct = hist.get('dq_pct', {}) if hist else {}
    years = hist.get('years', []) if hist else []
    if not years:
        years = list(range(2019, int(snap[:4]) + 1))

    # ACL months (life of loan) per pool
    _imp = hist.get('impaired', {}) if hist else {}
    acl_months_map = _imp.get('acl_months', {})

    snap_year = int(snap[:4])
    snap_month = int(snap[5:7])

    # Trim leading years to the deepest pool's Life-of-Loan window.
    # Years older than the longest LoL across all pools aren't used by
    # any pool's calculation and just clutter the year axis.
    if pools and years:
        _max_lol = max(acl_months_map.get(p, 36) for p in pools)
        _abs_first = (snap_year * 12 + snap_month) - _max_lol + 1
        _cutoff_year = (_abs_first - 1) // 12
        years = [y for y in years if y >= _cutoff_year]

    year_strs = [str(y) for y in years]
    num_years = len(years)

    # Prefer WARM-sourced CO/RC/Net data when available
    warm_co  = _imp.get('warm_co', {})
    warm_rc  = _imp.get('warm_rc', {})
    warm_net = _imp.get('warm_net', {})
    warm_co_totals = _imp.get('warm_co_totals', {})
    warm_rc_totals = _imp.get('warm_rc_totals', {})
    use_warm_co_rc = bool(warm_co)

    # Monthly CO/RC for partial-year windowing
    warm_co_monthly = _imp.get('warm_co_monthly', {})
    warm_rc_monthly = _imp.get('warm_rc_monthly', {})
    if not warm_co_monthly:
        warm_co_monthly = hist.get('co_monthly', {}) if hist else {}
    if not warm_rc_monthly:
        warm_rc_monthly = hist.get('rc_monthly', {}) if hist else {}

    def _pool_window_start(pool):
        """Return (earliest_year, earliest_month) for the WARM window."""
        pool_acl = acl_months_map.get(pool, 36)
        abs_first = (snap_year * 12 + snap_month) - pool_acl + 1
        ey = (abs_first - 1) // 12
        em = abs_first - ey * 12
        return ey, em

    def _pool_earliest_year(pool):
        return _pool_window_start(pool)[0]

    def _windowed_year_val(monthly_data, yearly_data, pool, year,
                           earliest_year, earliest_month):
        """Return the value for *year* trimmed to the WARM window.

        For the earliest year, sum only months >= earliest_month from monthly
        data.  For all other years use the full yearly total.
        """
        if year != earliest_year:
            return yearly_data.get(year, {}).get(pool, 0)
        # Partial year – sum monthly data from earliest_month onward
        partial = 0
        has_monthly = False
        for m in range(earliest_month, 13):
            v = monthly_data.get((year, m), {}).get(pool, 0)
            if v:
                has_monthly = True
            partial += v
        if has_monthly:
            # Monthly recovery data may be stored negative; align sign
            # with the yearly convention
            full_year = yearly_data.get(year, {}).get(pool, 0)
            if full_year and (full_year > 0) != (partial > 0):
                partial = -partial
            return partial
        # Fallback: prorate the yearly total
        full = yearly_data.get(year, {}).get(pool, 0)
        months_in_window = 12 - earliest_month + 1
        return full * months_in_window / 12 if full else 0

    # Column widths
    ws.column_dimensions['A'].width = 30.7
    for ci in range(2, num_years + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 16.4
    ws.column_dimensions[get_column_letter(num_years + 2)].width = 17.4  # ACL column
    ws.column_dimensions[get_column_letter(num_years + 3)].width = 11.7  # WARM

    # Title
    ws['A1'] = cu
    ws['A1'].font = V14B
    ws.row_dimensions[1].height = 18.75
    ws['A2'] = "Delinquency Calculation"
    ws['A2'].font = V12B
    ws['A3'] = f"For Quarter Ending {_snap_display(snap)}"
    ws['A3'].font = V12B

    ncol = num_years + 3  # A + years + ACL total + WARM

    section_ranges = []  # (start_row, end_row) for each section block

    # ─── Charge offs ───
    r = 5
    co_start = r
    co_headers = ["Charge offs"] + year_strs[:num_years]
    # Add "YTD" prefix to last year
    if year_strs:
        co_headers[-1] = f"YTD {year_strs[-1]}"
    co_headers += ["ACL Charge offs", "WARM Months"]

    for ci, lbl in enumerate(co_headers, start=1):
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[r].height = 31.5

    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool).font = V12B
        earliest, earliest_mo = _pool_window_start(pool)
        acl_total = 0
        for yi, y in enumerate(years):
            if y < earliest:
                continue
            if use_warm_co_rc:
                val = _windowed_year_val(warm_co_monthly, warm_co,
                                         pool, y, earliest, earliest_mo)
            else:
                val = _windowed_year_val(hist.get('co_monthly', {}),
                                         co_data, pool, y, earliest, earliest_mo)
            ws.cell(row=r, column=2 + yi, value=val).number_format = ACCT
            ws.cell(row=r, column=2 + yi).font = V10B
            acl_total += val
        ws.cell(row=r, column=num_years + 2, value=acl_total).number_format = ACCT
        ws.cell(row=r, column=num_years + 2).font = V10B
        warm = acl_months_map.get(pool, config.get('warm_months', {}).get(pool, 36))
        ws.cell(row=r, column=num_years + 3, value=warm).font = V10B
        ws.cell(row=r, column=num_years + 3).alignment = Alignment(horizontal='center')
        ws.row_dimensions[r].height = 15.75
    section_ranges.append((co_start, r))

    # ─── Recoveries ───
    r += 3
    rc_start = r
    rc_headers = ["Recoveries"] + year_strs[:]
    if year_strs:
        rc_headers[-1] = f"YTD {year_strs[-1]}"
    rc_headers += ["ACL Recoveries", "WARM Months"]

    for ci, lbl in enumerate(rc_headers, start=1):
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[r].height = 32.25

    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool).font = V12B
        earliest, earliest_mo = _pool_window_start(pool)
        acl_total = 0
        for yi, y in enumerate(years):
            if y < earliest:
                continue
            if use_warm_co_rc:
                val = _windowed_year_val(warm_rc_monthly, warm_rc,
                                         pool, y, earliest, earliest_mo)
            else:
                val = _windowed_year_val(hist.get('rc_monthly', {}),
                                         rc_data, pool, y, earliest, earliest_mo)
            ws.cell(row=r, column=2 + yi, value=val).number_format = ACCT
            ws.cell(row=r, column=2 + yi).font = V10B
            acl_total += val
        ws.cell(row=r, column=num_years + 2, value=acl_total).number_format = ACCT
        ws.cell(row=r, column=num_years + 2).font = V10B
        warm = acl_months_map.get(pool, config.get('warm_months', {}).get(pool, 36))
        ws.cell(row=r, column=num_years + 3, value=warm).font = V10B
        ws.cell(row=r, column=num_years + 3).alignment = Alignment(horizontal='center')
        ws.row_dimensions[r].height = 15.75
    section_ranges.append((rc_start, r))

    # ─── Net Charge offs ───
    r += 3
    nl_start = r
    nl_headers = ["Net Charge offs"] + year_strs[:]
    if year_strs:
        nl_headers[-1] = f"YTD {year_strs[-1]}"
    nl_headers += ["Net Charge offs", "WARM Months"]

    for ci, lbl in enumerate(nl_headers, start=1):
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool).font = V12B
        earliest, earliest_mo = _pool_window_start(pool)
        acl_total = 0
        for yi, y in enumerate(years):
            if y < earliest:
                continue
            if use_warm_co_rc:
                co_val = _windowed_year_val(warm_co_monthly, warm_co,
                                            pool, y, earliest, earliest_mo)
                rc_val = _windowed_year_val(warm_rc_monthly, warm_rc,
                                            pool, y, earliest, earliest_mo)
                net = co_val - rc_val
            else:
                co_val = _windowed_year_val(hist.get('co_monthly', {}),
                                            co_data, pool, y, earliest, earliest_mo)
                rc_val = _windowed_year_val(hist.get('rc_monthly', {}),
                                            rc_data, pool, y, earliest, earliest_mo)
                net = co_val - rc_val
            ws.cell(row=r, column=2 + yi, value=net).number_format = ACCT
            ws.cell(row=r, column=2 + yi).font = V10B
            acl_total += net
        ws.cell(row=r, column=num_years + 2, value=acl_total).number_format = ACCT
        ws.cell(row=r, column=num_years + 2).font = V10B
        warm = acl_months_map.get(pool, config.get('warm_months', {}).get(pool, 36))
        ws.cell(row=r, column=num_years + 3, value=warm).font = V10B
        ws.cell(row=r, column=num_years + 3).alignment = Alignment(horizontal='center')
        ws.row_dimensions[r].height = 15.75
    section_ranges.append((nl_start, r))

    # ─── Delinquency ───
    r += 3
    dq_start = r
    # Use WARM-sourced DQ% if available, otherwise fall back to computed dq_pct
    warm_dq = hist.get('impaired', {}).get('warm_dq_pct', {}) if hist else {}
    use_dq = warm_dq if warm_dq else dq_pct
    dq_headers = ["DQ %"] + year_strs[:] + ["Average", "Variance"]
    if year_strs:
        dq_headers[len(year_strs)] = f"YTD {year_strs[-1]}"
    for ci, lbl in enumerate(dq_headers, start=1):
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for pool in pools:
        r += 1
        ws.cell(row=r, column=1, value=pool).font = V12B
        earliest = _pool_earliest_year(pool)
        rates = []
        for yi, y in enumerate(years):
            if y < earliest:
                continue
            val = use_dq.get(y, {}).get(pool, 0)
            ws.cell(row=r, column=2 + yi, value=val).number_format = PCT
            ws.cell(row=r, column=2 + yi).font = V10B
            rates.append(val)
        avg = sum(rates) / len(rates) if rates else 0
        ws.cell(row=r, column=num_years + 2, value=avg).number_format = PCT
        ws.cell(row=r, column=num_years + 2).font = V10B
        var = rates[-1] - avg if len(rates) > 1 else 0
        ws.cell(row=r, column=num_years + 3, value=var).number_format = PCT
        ws.cell(row=r, column=num_years + 3).font = V10B
    section_ranges.append((dq_start, r))

    # ── Page Setup ──
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25,
                                  header=0.3, footer=0.3)
    ws.print_title_rows = '1:3'
    ws.print_area = f'A1:J{r}'

    # Force a single manual page break before the Net Charge offs section so
    # page 1 = Charge offs + Recoveries, page 2 = Net Charge offs + DQ %.
    if len(section_ranges) >= 3:
        nl_start = section_ranges[2][0]
        ws.row_breaks.append(Break(id=nl_start - 1))


def _range_label(lo, hi):
    """Reconstruct a display label from (lo, hi) range boundaries."""
    if lo <= -999:
        return f"<{hi:.2f}%"
    if hi >= 999:
        return f">{lo:.2f}%"
    return f"{lo:.2f}% to {hi - 0.01:.2f}%"


def _sheet_env_ranges(wb, cu, snap, hist):
    """Environmental Factor Ranges reference sheet matching Vizo template."""
    ws = wb.create_sheet(">Envir Fact Ranges")

    # Column widths
    for ci, w in enumerate([3.4, 18.6, 16, 18.4, 17.6, 18.1, 16.9, 3.4], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Title block (merged B2:G4)
    ws.merge_cells('B2:G2')
    ws['B2'] = cu
    ws['B2'].font = V14B
    ws['B2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('B3:G3')
    ws['B3'] = "Environmental Factor Ranges"
    ws['B3'].font = V12B
    ws['B3'].alignment = Alignment(horizontal='center')

    ws.merge_cells('B4:G4')
    ws['B4'] = f"For Quarter Ending {_snap_display(snap)}"
    ws['B4'].font = V10B
    ws['B4'].alignment = Alignment(horizontal='center')

    # Column headers (row 6-7)
    # Section fills use theme accent 4 (theme index 7) with Lighter 40% / 60% / 80% tints
    SEC_MED = PatternFill('solid', fgColor=Color(theme=7, tint=0.3999450666829432))
    SEC_DQ = PatternFill('solid', fgColor=Color(theme=7, tint=0.5999938962981048))
    SEC_LITE = PatternFill('solid', fgColor=Color(theme=7, tint=0.7999816888943144))
    HDR_BLACK = Font(name='Calibri', bold=True, size=12, color='000000')

    ws.merge_cells('B6:C6')
    ws['B6'] = "Net Credit Change"
    ws['B6'].font = V12B
    ws['B6'].fill = SEC_MED
    ws['B6'].font = HDR_BLACK
    ws['B6'].alignment = Alignment(horizontal='center')
    ws['C6'].fill = SEC_MED

    ws.merge_cells('D6:E6')
    ws['D6'] = "Delinquency"
    ws['D6'].font = HDR_BLACK
    ws['D6'].fill = SEC_DQ
    ws['D6'].alignment = Alignment(horizontal='center')
    ws['E6'].fill = SEC_DQ

    ws.merge_cells('F6:G6')
    ws['F6'] = "Economic Stress Score"
    ws['F6'].font = HDR_BLACK
    ws['F6'].fill = SEC_LITE
    ws['F6'].alignment = Alignment(horizontal='center')
    ws['G6'].fill = SEC_LITE

    r = 7
    section_fills = {2: SEC_MED, 3: SEC_MED, 4: SEC_DQ, 5: SEC_DQ,
                     6: SEC_LITE, 7: SEC_LITE}
    for ci, lbl in [(2, "Range"), (3, "Score"), (4, "Range"), (5, "Score"),
                    (6, "Range"), (7, "Score")]:
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.font = HDR_BLACK
        cell.fill = section_fills[ci]
        cell.alignment = Alignment(horizontal='center')

    # Get ranges from WARM (same source as env factor calculations)
    _ncc_r, _dq_r, _es_r = _env_ranges(hist)
    _imp = hist.get('impaired', {}) if hist else {}
    er = _imp.get('env_ranges', {})
    ncc_labels = er.get('ncc_labels', [])
    dq_labels  = er.get('dq_labels', [])
    es_labels  = er.get('es_labels', [])

    def _build_display(ranges, labels):
        """Build (label, score_decimal) pairs from range tuples + optional labels."""
        result = []
        for idx, (lo, hi, s) in enumerate(ranges):
            lbl = labels[idx] if idx < len(labels) else _range_label(lo, hi)
            result.append((lbl, s / 100.0))
        return result

    ncc_data = _build_display(_ncc_r, ncc_labels)
    dq_data  = _build_display(_dq_r, dq_labels)
    es_data  = _build_display(_es_r, es_labels)

    r = 8
    for i, (rng_str, sc) in enumerate(ncc_data):
        c2 = ws.cell(row=r + i, column=2, value=rng_str)
        c2.font = V12; c2.alignment = Alignment(horizontal='center'); c2.fill = SEC_MED
        c3 = ws.cell(row=r + i, column=3, value=sc)
        c3.number_format = PCT; c3.font = V12
        c3.alignment = Alignment(horizontal='center'); c3.fill = SEC_MED
    for i, (rng_str, sc) in enumerate(dq_data):
        c4 = ws.cell(row=r + i, column=4, value=rng_str)
        c4.font = V12; c4.alignment = Alignment(horizontal='center'); c4.fill = SEC_DQ
        c5 = ws.cell(row=r + i, column=5, value=sc)
        c5.number_format = PCT; c5.font = V12
        c5.alignment = Alignment(horizontal='center'); c5.fill = SEC_DQ
    for i, (rng_str, sc) in enumerate(es_data):
        c6 = ws.cell(row=r + i, column=6, value=rng_str)
        c6.font = V12; c6.alignment = Alignment(horizontal='center'); c6.fill = SEC_LITE
        c7 = ws.cell(row=r + i, column=7, value=sc)
        c7.number_format = PCT; c7.font = V12
        c7.alignment = Alignment(horizontal='center'); c7.fill = SEC_LITE

    # Fill any remaining (shorter section) rows so each column block has uniform height
    max_len = max(len(ncc_data), len(dq_data), len(es_data))
    for i in range(len(ncc_data), max_len):
        ws.cell(row=r + i, column=2).fill = SEC_MED
        ws.cell(row=r + i, column=3).fill = SEC_MED
    for i in range(len(dq_data), max_len):
        ws.cell(row=r + i, column=4).fill = SEC_DQ
        ws.cell(row=r + i, column=5).fill = SEC_DQ
    for i in range(len(es_data), max_len):
        ws.cell(row=r + i, column=6).fill = SEC_LITE
        ws.cell(row=r + i, column=7).fill = SEC_LITE

    # Description section – matches reference layout exactly (rows 27-41)
    # Row 27: "Description" label
    # Row 28: Combined Q&E description (merged B:G, wrap, height 93)
    # Row 29: spacer (1.8)
    # Row 30: GAAP quote (merged B:G, wrap, height 52.2)
    # Row 31: spacer (1.8)
    # Row 32: Methodology paragraph (merged B:G, wrap, height 127.8)
    # Row 33: spacer (1.8)
    # Row 34: "Net Credit Change:" label
    # Row 35: NCC description (merged, wrap, h 48, left/top)
    # Row 36: spacer (1.8)
    # Row 37: "Delinquency:" label
    # Row 38: DQ description (merged, wrap, h 46.8)
    # Row 39: spacer (1.8)
    # Row 40: "Economic Stress Score:" label
    # Row 41: ES description (merged, wrap, h 65.4, left)

    ws.cell(row=27, column=2, value="Description").font = V12B
    ws.row_dimensions[27].height = 15.6

    ws.merge_cells('B28:G28')
    ws['B28'] = (
        "The Environmental Factor combines three distinct data sets to calculate the "
        "likely variance between the historical loss rate and the anticipated loss rate.  "
        "As these three data sets improve the likelihood of loss decreases and as they "
        "deteriorate the likelihood of loss increases, hence the need to adjust the pool "
        "provision. Statistical tests including regression, MANOVA and Pearson 'R "
        "(correlation) were employed to validate the causal relationship of each factor "
        "and then establish the appropriate ranges."
    )
    ws['B28'].font = V12
    ws['B28'].alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[28].height = 93.0

    ws.row_dimensions[29].height = 1.8

    ws.merge_cells('B30:G30')
    ws['B30'] = (
        "GAAP states that \u201cwhen estimating credit losses on each group of loans with "
        "similar risk characteristics, an institution should consider its historical loss "
        "experience on the group, adjusted for changes in trends, conditions, and other "
        "relevant factors that affect repayment of the loans as of the evaluation date.\u201d"
    )
    ws['B30'].font = V12
    ws['B30'].alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[30].height = 52.2

    ws.row_dimensions[31].height = 1.8

    ws.merge_cells('B32:G32')
    ws['B32'] = (
        "In this methodology the Environmental Factor, or Q&E, is used to adjust the "
        "allowance for each loan pool to assure it is representative of current risk over "
        "the life of loans in that pool. Three measures, identified in the Comptrollers "
        "Handbook, (e.g., net credit change in the portfolio, delinquency and economic "
        "stress factor) are correlated to the charge-off history of the credit union, to "
        "determine if adjustments to each pool\u2019s risk factors are indicated. The three "
        "measures are applied to minimize the potential for skewing that could result from "
        "a single measure. Using accepted statistical techniques, (e.g., regression, ANOVA "
        "and correlation) the extent of the relationship is measured, and the adjustment "
        "is applied to the Pooled allowance amount for each grade to arrive at the "
        "adjusted allowance amount.    "
    )
    ws['B32'].font = V12
    ws['B32'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[32].height = 127.8

    ws.row_dimensions[33].height = 1.8

    ws.cell(row=34, column=2, value="Net Credit Change:").font = V12B
    ws.row_dimensions[34].height = 15.6

    ws.merge_cells('B35:G35')
    ws['B35'] = (
        "Derived from the credit migration for each pool in the Credit Union's loan "
        "portfolio. (Refer to the Analysis of Impaired/Improved Loans Report). A positive "
        "net credit change indicates reduced risk while a negative credit change indicates "
        "increased risk of loan losses. "
    )
    ws['B35'].font = V12
    ws['B35'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[35].height = 48.0

    ws.row_dimensions[36].height = 1.8

    ws.cell(row=37, column=2, value="Delinquency:").font = V12B
    ws.row_dimensions[37].height = 15.6

    ws.merge_cells('B38:G38')
    ws['B38'] = (
        "Reported Delinquency percentage trends for pool based the Credit Union\u2019s life "
        "of loans, in each distinctive pool.  Increasing delinquent percentages are "
        "predictive of increasing losses, while decreasing percentages suggest decreasing "
        "losses."
    )
    ws['B38'].font = V12
    ws['B38'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[38].height = 46.8

    ws.row_dimensions[39].height = 1.8

    ws.cell(row=40, column=2, value="Economic Stress Score:").font = V12B
    ws.row_dimensions[40].height = 15.6

    ws.merge_cells('B41:G41')
    ws['B41'] = (
        "Measures the relative impact of the local economy on the historical loss rate.  "
        "Calculated using Unemployment Rate, Bankruptcy Percentage, Foreclosure Rates and "
        "Population statistics for the State in which the Credit Union is based. The "
        "sources of these figures are: The Bureau of Labor Statistics, Realty Trac, and "
        "Electronic Access to Federal Court Records. "
    )
    ws['B41'].font = V12
    ws['B41'].alignment = Alignment(horizontal='left', wrap_text=True)
    ws.row_dimensions[41].height = 65.4

    # Page setup: portrait, fit all rows and columns on one page, narrow margins
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'portrait'
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25,
                                  header=0, footer=0)


# ══════════════════════════════════════════════════════════════════
# SHEET BUILDERS – SUPPLEMENTAL REPORT
# ══════════════════════════════════════════════════════════════════

def _sheet_hist_trends_bal(wb, cu, snap, df, grades, config, hist):
    """Historical Trends Balance – per-pool line charts for supplemental.

    Creates one line chart per risk-rated pool, with a series for each grade.
    Data is referenced directly from the '> Detail_HIst Balances' tab.
    Matches template '> Historical Trends Balance-Viz' layout.
    """
    ws = wb.create_sheet("> Historical Trends Balance")

    # ── Title rows ──
    ws['A1'] = cu
    ws['A1'].font = V14B

    ws['A2'] = "Historical Loan Balances by Most Recent Credit Score"
    ws['A2'].font = V12B

    ws['A3'] = f"For Period Ending {_snap_display(snap)}"
    ws['A3'].font = V10B

    # Must already be created before this sheet
    detail_ws = wb["> Detail_HIst Balances"]

    impaired = hist.get('impaired', {}) if hist else {}
    risk_rated_map = impaired.get('risk_rated', {})

    # ── Scan detail sheet to find pool blocks ──
    pool_blocks = []
    max_row = detail_ws.max_row
    r = 5
    while r <= max_row:
        cell = detail_ws.cell(r, 1)
        val = cell.value
        font = cell.font
        # Pool header: bold, sz >= 9
        if val and font and font.sz and float(font.sz) >= 9 and font.bold:
            pool_name = val
            is_rr = risk_rated_map.get(pool_name, True)
            hdr_row = r + 1  # date header row

            # Count only date columns (datetime values, skip text like "% of Loans")
            n_dates = 0
            for c in range(2, detail_ws.max_column + 1):
                cv = detail_ws.cell(hdr_row, c).value
                if isinstance(cv, datetime):
                    n_dates += 1
                elif cv is not None:
                    break  # hit a text header (% of Loans, WARM Months)

            # Find Total row
            tr = hdr_row + 1
            while tr <= max_row and detail_ws.cell(tr, 1).value != 'Total':
                tr += 1
            n_grades = tr - (hdr_row + 1)

            if is_rr and n_grades > 0 and n_dates > 0:
                pool_blocks.append({
                    'pool': pool_name,
                    'date_row': hdr_row,
                    'grade_start': hdr_row + 1,
                    'grade_end': tr - 1,
                    'n_dates': n_dates,
                })
            r = tr + 2  # skip Total + separator
        else:
            r += 1

    if not pool_blocks:
        ws['A5'] = "No historical grade-level balance data available."
        ws['A5'].font = Font(name='Calibri', italic=True, size=10, color='888888')
        return

    # ── Create one line chart per pool ──
    # Landscape page: 11" wide × 8.5" tall, 0.25" margins → 10.5" × 8.0" printable.
    # Repeated header rows ~0.5" → ~7.5" for charts → 3.75" each chart.
    # 18 rows × 15pt ≈ 3.75" → CHART_ROWS = 18.
    CHART_ROWS = 18
    chart_row = 5

    for idx, pb in enumerate(pool_blocks):
        chart = LineChart()
        chart.width = 26.5        # ~10.4 inches – fills landscape printable width
        chart.height = 9.5        # ~3.74 inches – 2 charts fill a landscape page
        chart.style = 10

        # Remove chart area border
        chart.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))

        # Title: 20pt Calibri to match template
        chart.title = pb['pool']
        chart.title.tx = Text(
            rich=RichText(
                p=[Paragraph(
                    pPr=ParagraphProperties(
                        defRPr=CharacterProperties(
                            sz=2000, b=False,
                            latin=DrawingFont(typeface='Calibri'),
                        )
                    ),
                    r=[RegularTextRun(
                        rPr=CharacterProperties(
                            sz=2000, b=False,
                            latin=DrawingFont(typeface='Calibri'),
                        ),
                        t=pb['pool'],
                    )],
                )]
            )
        )

        # Y-axis (value axis): accounting number format, minimum = 0
        chart.y_axis.numFmt = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"_);_(@_)'
        chart.y_axis.delete = False
        chart.y_axis.scaling.min = 0
        chart.x_axis.delete = False

        # Remove vertical (Y) axis line
        chart.y_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))
        # Remove horizontal (X) axis line
        chart.x_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))

        # Major gridlines: gray (default light gray "BFBFBF" matches Excel default)
        from openpyxl.chart.axis import ChartLines
        chart.y_axis.majorGridlines = ChartLines(
            spPr=GraphicalProperties(
                ln=LineProperties(solidFill='BFBFBF')
            )
        )

        # X-axis: default formatting matching the Risk Chg [Pool] charts in the
        # user's reference Vizo file (General number format, default tick marks,
        # nextTo label position, no rotation, no custom font).
        chart.x_axis.majorTickMark = 'out'
        chart.x_axis.minorTickMark = 'none'
        chart.x_axis.tickLblPos = 'nextTo'
        # Custom -45 degree label rotation (OOXML uses 60000 units per degree)
        from openpyxl.drawing.text import RichTextProperties as _RTP
        chart.x_axis.txPr = RichText(
            bodyPr=_RTP(rot=-2700000, spcFirstLastPara=True,
                        vertOverflow='ellipsis', wrap='square',
                        anchor='ctr', anchorCtr=True),
            p=[Paragraph(pPr=ParagraphProperties(
                defRPr=CharacterProperties()))],
        )

        nd = pb['n_dates']

        # Categories: dates from the header row (col 2 through 2+nd-1)
        cats = Reference(detail_ws,
                         min_col=2, max_col=1 + nd,
                         min_row=pb['date_row'])

        # Data: grade rows (each row = one series)
        data = Reference(detail_ws,
                         min_col=2, max_col=1 + nd,
                         min_row=pb['grade_start'],
                         max_row=pb['grade_end'])
        chart.add_data(data, from_rows=True, titles_from_data=False)
        chart.set_categories(cats)

        # Label each series with its grade name and apply template colors
        # Colors match theme accent1-6 from template, cycling for >6 series
        ACCENT_HEX = ['0D4D5E', '829901', '3D1A1A', '48A5AD', '5F5F5F', 'FFC000']
        for gi, s in enumerate(chart.series):
            grade_name = detail_ws.cell(pb['grade_start'] + gi, 1).value or ''
            s.tx = SeriesLabel(v=grade_name)
            hex_color = ACCENT_HEX[gi % len(ACCENT_HEX)]
            s.graphicalProperties.line = LineProperties(
                w=38100, cap='rnd', prstDash='solid',
                solidFill=hex_color, round=True,
            )

        # Legend positioned below title
        chart.legend.position = 't'
        chart.legend.layout = Layout(
            manualLayout=ManualLayout(
                xMode='edge', yMode='edge',
                x=0.10, y=0.14,
                w=0.80, h=0.06,
            )
        )
        # Plot area pushed down to clear title + legend
        chart.layout = Layout(
            manualLayout=ManualLayout(
                xMode='edge', yMode='edge',
                x=0.05, y=0.22,
                w=0.93, h=0.72,
            )
        )

        ws.add_chart(chart, f"A{chart_row}")
        chart_row += CHART_ROWS

        # Force page break after every 2nd chart so none get cut
        if (idx + 1) % 2 == 0 and idx + 1 < len(pool_blocks):
            ws.row_breaks.append(Break(id=chart_row - 1))

    # ── Page setup: landscape, 2 charts per page, headers repeat ──
    ws.print_title_rows = '1:3'
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25


def _sheet_detail_hist_bal(wb, cu, snap, df, grades, config, hist):
    """Detail Hist Balances – monthly breakdown by pool by grade.

    Uses grade-level data from the WARM's 'HIst Bal Data' tab when available.
    Only shows months within the life-of-loan (ACL months) for each pool.
    Formatted to match template '> Detail_HIst Balances-Vizo'.
    """
    ws = wb.create_sheet("> Detail_HIst Balances")
    no_score = config.get('no_score_label', 'Not Reported')
    all_gl = [g for g in _all_grades(grades, no_score) if not _is_hidden(g)]

    # 8pt fonts for compact layout (12 months per printed page)
    F8B = V8B
    F8  = V8

    def _grade_font8(label):
        if _is_hidden(label):
            return Font(name='Calibri', size=8, color='FF0000')
        return F8

    # Get WARM hist data (grade-level per pool per month)
    impaired = hist.get('impaired', {}) if hist else {}
    hbd = impaired.get('hist_bal_data', {})
    warm_order = impaired.get('pool_order', [])
    acl_months = impaired.get('acl_months', {})
    risk_rated = impaired.get('risk_rated', {})

    # Use WARM pool order; fall back to _ordered_pools
    pools = warm_order if warm_order else _ordered_pools(df, hist)

    # Compute column A width from longest pool name / grade label
    all_labels = list(pools) + all_gl + ['Current Grade', 'Total']
    max_len = max((len(str(s)) for s in all_labels), default=20)
    # Approximate: 1 char ≈ 1.1 Excel width units at 8pt Calibri
    col_a_width = max_len * 1.1 + 2

    ws.column_dimensions['A'].width = col_a_width

    # Row 1: CU name
    ws['A1'] = cu
    ws['A1'].font = Font(name='Calibri', bold=True, size=11)

    # Row 2: subtitle
    ws['A2'] = "Loss Factor Historical Detail"
    ws['A2'].font = Font(name='Calibri', bold=True, size=10)

    # Row 3: quarter ending
    ws['A3'] = f"For Quarter Ending {_snap_display(snap)}"
    ws['A3'].font = Font(name='Calibri', bold=True, size=10)

    # Row 4: spacer
    ws.row_dimensions[4].height = 5.0

    if not hbd:
        # Fallback: just show current data from DB
        pools_fb = _ordered_pools(df, hist)
        r = 5
        for pool in pools_fb:
            ws.cell(row=r, column=1, value=pool).font = Font(name='Calibri', bold=True, size=9)
            r += 1
            ws.cell(row=r, column=1, value="Current Grade").font = F8B
            ws.cell(row=r, column=2, value=snap).font = F8B
            r += 1
            pdf = df[df['loan_pool'] == pool]
            vgl = [g for g in all_gl if not _is_hidden(g)]
            for g in vgl:
                fnt = _grade_font8(g)
                ws.cell(row=r, column=1, value=g).font = fnt
                bal = pdf[pdf['current_grade'] == g]['current_balance'].sum()
                ws.cell(row=r, column=2, value=bal).number_format = ACCT
                ws.cell(row=r, column=2).font = fnt
                r += 1
            ws.cell(row=r, column=1, value="Total").font = F8B
            ws.cell(row=r, column=2, value=pdf['current_balance'].sum()).number_format = ACCT
            ws.cell(row=r, column=2).font = F8B
            r += 2
        return

    # ── Main path: use WARM hist_bal_data ──
    # Get canonical date list from first pool
    first_pool_data = next(iter(hbd.values()), {})
    all_dates = first_pool_data.get('dates', [])

    # Date columns start at column 2 (no blank column B)
    DATE_COL_START = 2

    r = 5
    pool_boundaries = []  # separator row numbers between pools
    for pool in pools:
        pdata = hbd.get(pool, {})
        pdates = pdata.get('dates', all_dates)
        pgrades = pdata.get('grades', {})
        ptotal = pdata.get('total', [])

        # Determine how many months to show based on life-of-loan
        pool_acl = acl_months.get(pool, len(pdates))  # default: show all
        # Take only the last pool_acl months
        if pool_acl < len(pdates):
            start_idx = len(pdates) - pool_acl
            pdates = pdates[start_idx:]
            pgrades = {g: v[start_idx:] for g, v in pgrades.items()}
            ptotal = ptotal[start_idx:]
        nmonths = len(pdates)

        # Set column widths for date columns (~9.7 units for 12 per page)
        for di in range(nmonths):
            col_letter = get_column_letter(DATE_COL_START + di)
            if ws.column_dimensions[col_letter].width < 9.71:
                ws.column_dimensions[col_letter].width = 9.71

        # Columns for "% of Loans" and "WARM Months"
        pct_col = DATE_COL_START + nmonths
        warm_col = pct_col + 1
        ws.column_dimensions[get_column_letter(pct_col)].width = 9.71
        ws.column_dimensions[get_column_letter(warm_col)].width = 9.71

        # Latest total for % of Loans calculation
        pool_last_total = ptotal[-1] if ptotal else 0

        # Pool header row
        ws.cell(row=r, column=1, value=pool).font = Font(name='Calibri', bold=True, size=9)
        ws.row_dimensions[r].height = 13.5
        r += 1

        is_rr = risk_rated.get(pool, True)

        if not is_rr:
            # ── Non-risk-rated pool: total line only ──
            # Date header row with dates
            ws.cell(row=r, column=1, value="Balance").font = F8B
            for di, dt in enumerate(pdates):
                c = ws.cell(row=r, column=DATE_COL_START + di, value=dt)
                c.number_format = 'mmm\\-yy'
                c.font = F8B
                c.alignment = Alignment(horizontal='center')
            c_warm = ws.cell(row=r, column=warm_col, value="WARM\nMonths")
            c_warm.font = F8B
            c_warm.alignment = Alignment(horizontal='center', wrap_text=True)
            r += 1

            # Total row
            ws.cell(row=r, column=1, value="Total").font = F8B
            ws.cell(row=r, column=1).fill = TOT_FILL
            for di in range(nmonths):
                v = ptotal[di] if di < len(ptotal) else 0
                c = ws.cell(row=r, column=DATE_COL_START + di, value=v)
                c.number_format = ACCT
                c.font = F8B
                c.fill = TOT_FILL
            # WARM Months
            pool_acl_val = acl_months.get(pool, '')
            ws.cell(row=r, column=warm_col, value=pool_acl_val).font = F8
            ws.cell(row=r, column=warm_col).alignment = Alignment(horizontal='center')
            r += 1

            # Blank separator row
            ws.row_dimensions[r].height = 5.0
            pool_boundaries.append(r)
            r += 1
            continue

        # ── Risk-rated pool: full grade breakdown ──
        # Date header row: "Current Grade" + date values + summary headers
        ws.cell(row=r, column=1, value="Current Grade").font = F8B
        for di, dt in enumerate(pdates):
            c = ws.cell(row=r, column=DATE_COL_START + di, value=dt)
            c.number_format = 'mmm\\-yy'
            c.font = F8B
            c.alignment = Alignment(horizontal='center')
        c_pct = ws.cell(row=r, column=pct_col, value="% of Loans")
        c_pct.font = F8B
        c_pct.alignment = Alignment(horizontal='center')
        c_warm = ws.cell(row=r, column=warm_col, value="WARM\nMonths")
        c_warm.font = F8B
        c_warm.alignment = Alignment(horizontal='center', wrap_text=True)
        r += 1

        # Grade rows with alternating fill
        grade_start = r
        for gi, g in enumerate(all_gl):
            fnt = _grade_font8(g)
            ws.cell(row=r, column=1, value=g).font = fnt
            # Alternating row fill
            row_fill = ALT_FILL if gi % 2 == 0 else None
            if row_fill:
                ws.cell(row=r, column=1).fill = row_fill
            vals = pgrades.get(g, [])
            for di in range(nmonths):
                v = vals[di] if di < len(vals) else 0
                c = ws.cell(row=r, column=DATE_COL_START + di, value=v)
                c.number_format = ACCT
                c.font = fnt
                if row_fill:
                    c.fill = row_fill
            # % of Loans
            last_val = vals[-1] if vals else 0
            pct_val = last_val / pool_last_total if pool_last_total else 0
            c_pct = ws.cell(row=r, column=pct_col, value=pct_val)
            c_pct.number_format = PCT
            c_pct.font = fnt
            if row_fill:
                c_pct.fill = row_fill
            r += 1

        # WARM Months: merge across all grade rows + total, value on first grade row
        pool_acl_val = acl_months.get(pool, '')
        ws.cell(row=grade_start, column=warm_col, value=pool_acl_val).font = F8
        ws.cell(row=grade_start, column=warm_col).alignment = Alignment(
            horizontal='center', vertical='center')
        if len(all_gl) > 0:
            ws.merge_cells(
                start_row=grade_start, start_column=warm_col,
                end_row=grade_start + len(all_gl), end_column=warm_col)

        # Total row (always shaded – darker)
        ws.cell(row=r, column=1, value="Total").font = F8B
        ws.cell(row=r, column=1).fill = TOT_FILL
        for di in range(nmonths):
            v = ptotal[di] if di < len(ptotal) else 0
            c = ws.cell(row=r, column=DATE_COL_START + di, value=v)
            c.number_format = ACCT
            c.font = F8B
            c.fill = TOT_FILL
        # Total % of Loans = 100%
        c_pct = ws.cell(row=r, column=pct_col, value=1.0)
        c_pct.number_format = PCT
        c_pct.font = F8B
        c_pct.fill = TOT_FILL
        r += 1

        # Blank separator row (small height)
        ws.row_dimensions[r].height = 5.0
        pool_boundaries.append(r)
        r += 1

    # Insert page breaks only at pool boundaries where a natural break would land.
    # With fitToWidth=5, Excel scales content down, so more rows fit per page.
    # Calculate scale factor from total column width vs 5 pages of printable width.
    FIT_PAGES_WIDE = 5
    PRINTABLE_W_PT = (11.0 - 0.25 - 0.25) * 72   # landscape 11" - L/R margins
    PRINTABLE_H_PT = (8.5 - 0.25 - 0.25) * 72     # landscape short edge - T/B margins
    # Excel column-width chars → physical points: 1 char-width ≈ 7px at 96DPI
    CHAR_TO_PT = 7.0 * 72.0 / 96.0                 # = 5.25
    total_col_width_pt = 0
    for ci in range(1, ws.max_column + 1):
        letter = ws.cell(1, ci).column_letter
        cd = ws.column_dimensions.get(letter)
        w = cd.width if cd and cd.width else 8.43
        total_col_width_pt += w * CHAR_TO_PT
    pages_needed = total_col_width_pt / PRINTABLE_W_PT
    scale = min(1.0, FIT_PAGES_WIDE / pages_needed) if pages_needed > 0 else 1.0
    # Effective page height in unscaled points (more rows fit when scaled down)
    eff_page_ht = PRINTABLE_H_PT / scale
    DEFAULT_ROW_HT = 15.0  # Excel default row height

    # Walk rows, placing breaks at pool boundaries before the page overflows
    page_start = 1
    last_data_row = r - 1
    while page_start <= last_data_row:
        cumulative = 0.0
        trigger_row = None
        for row_num in range(page_start, last_data_row + 1):
            dim = ws.row_dimensions.get(row_num)
            if dim and dim.height is not None:
                cumulative += dim.height
            else:
                cumulative += DEFAULT_ROW_HT
            if cumulative >= eff_page_ht:
                trigger_row = row_num
                break
        if trigger_row is None:
            break  # remaining content fits on current page
        # Find the last pool boundary between page_start and trigger_row
        best = None
        for b in pool_boundaries:
            if page_start < b <= trigger_row:
                best = b
        if best is None:
            break  # no suitable boundary; let Excel handle it
        ws.row_breaks.append(Break(id=best))
        page_start = best + 1

    # Page setup: landscape, 0.25" margins, fit to 5 pages wide
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 5
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.25
    ws.page_margins.footer = 0.25
    ws.print_title_cols = 'A:A'


def _sheet_detail_co_hist(wb, cu, snap, config, hist):
    """Detail Charge off Hist – monthly chargeoff/recovery data by pool.

    Uses monthly CO/recovery data and aligns to hist_bal_data dates,
    filtered to life-of-loan (ACL months) per pool.
    Matches template '>Detail_Charge off Hist-Vizo'.
    """
    ws = wb.create_sheet(">Detail_Charge off Hist")

    F8  = V8
    F8B = V8B

    impaired = hist.get('impaired', {}) if hist else {}
    hbd = impaired.get('hist_bal_data', {})
    warm_order = impaired.get('pool_order', [])
    acl_months_map = impaired.get('acl_months', {})

    # Prefer WARM "Charge off History" monthly data over file-parsed data
    co_monthly = impaired.get('warm_co_monthly', {}) or (hist.get('co_monthly', {}) if hist else {})
    rc_monthly = impaired.get('warm_rc_monthly', {}) or (hist.get('rc_monthly', {}) if hist else {})

    # Pool order from WARM; fallback to sorted pool_map values
    pools = warm_order if warm_order else sorted(set(config.get('pool_map', {}).values()))

    # Get canonical date list from hist_bal_data (same as Detail_HIst Balances)
    first_pool_data = next(iter(hbd.values()), {}) if hbd else {}
    all_dates = first_pool_data.get('dates', [])

    # If no hist_bal_data dates, build from co_monthly + rc_monthly keys
    if not all_dates:
        all_ym = sorted(set(list(co_monthly.keys()) + list(rc_monthly.keys())))
        import datetime
        all_dates = [datetime.datetime(y, m, 1) for y, m in all_ym]

    # Column A width from longest pool name
    all_labels = list(pools) + ['Charge offs', 'Recoveries', 'Net Loss',
                                'Total Charge offs', 'Total Recoveries', 'Total Net Loss']
    max_len = max((len(str(s)) for s in all_labels), default=20)
    col_a_width = max_len * 1.1 + 2
    ws.column_dimensions['A'].width = col_a_width

    DATE_COL_START = 2  # dates start at column B (no blank column)

    # ── Title rows ──
    ws['A1'] = cu
    ws['A1'].font = Font(name='Calibri', bold=True, size=11)
    ws['A2'] = "Charge off and Recoveries Historical Detail"
    ws['A2'].font = Font(name='Calibri', bold=True, size=10)
    ws['A3'] = f"For Quarter Ending {_snap_display(snap)}"
    ws['A3'].font = Font(name='Calibri', bold=True, size=10)
    ws.row_dimensions[4].height = 5.0

    def _write_section(ws, start_row, section_label, monthly_data, pools,
                       all_dates, acl_months_map):
        """Write one section (Chargeoffs / Recoveries / Net).
        Returns (next_row, {pool: [monthly_values]})."""
        r = start_row

        # Section header row: label + date headers
        ws.cell(row=r, column=1, value=section_label).font = Font(
            name='Calibri', bold=True, size=9)

        # Determine max months across all pools for date columns
        max_months = 0
        pool_date_slices = {}
        for pool in pools:
            n = acl_months_map.get(pool, len(all_dates))
            n = min(n, len(all_dates))
            pool_date_slices[pool] = all_dates[-n:] if n > 0 else all_dates
            if n > max_months:
                max_months = n

        # Use the widest date range for column headers
        if all_dates:
            header_dates = all_dates[-max_months:] if max_months > 0 else all_dates
        else:
            header_dates = []
        nmonths = len(header_dates)

        # Set date column widths
        for di in range(nmonths):
            col_letter = get_column_letter(DATE_COL_START + di)
            ws.column_dimensions[col_letter].width = 9.71

        # Date header values
        for di, dt in enumerate(header_dates):
            c = ws.cell(row=r, column=DATE_COL_START + di, value=dt)
            c.number_format = 'mmm\\-yy'
            c.font = F8B
            c.alignment = Alignment(horizontal='center')

        # ACL total column
        acl_col = DATE_COL_START + nmonths
        ws.cell(row=r, column=acl_col, value="ACL\nTotal").font = F8B
        ws.cell(row=r, column=acl_col).alignment = Alignment(
            horizontal='center', wrap_text=True)
        col_letter = get_column_letter(acl_col)
        ws.column_dimensions[col_letter].width = 9.71
        r += 1

        # Pool rows
        pool_values = {}  # {pool: [monthly_values]}
        for pi, pool in enumerate(pools):
            pdates = pool_date_slices[pool]
            pn = len(pdates)
            offset = nmonths - pn  # blank columns before pool data starts

            fnt = F8
            ws.cell(row=r, column=1, value=pool).font = fnt
            row_fill = ALT_FILL if pi % 2 == 0 else None
            if row_fill:
                ws.cell(row=r, column=1).fill = row_fill

            vals = []
            for di, dt in enumerate(pdates):
                ym = (dt.year, dt.month) if hasattr(dt, 'year') else (
                    dt.year(), dt.month())
                v = monthly_data.get(ym, {}).get(pool, 0)
                vals.append(v)
                c = ws.cell(row=r, column=DATE_COL_START + offset + di,
                            value=v if v else '')
                if v:
                    c.number_format = ACCT
                c.font = fnt
                if row_fill:
                    c.fill = row_fill

            pool_values[pool] = vals

            # ACL total = sum of pool values
            acl_total = sum(v for v in vals if isinstance(v, (int, float)))
            c = ws.cell(row=r, column=acl_col, value=acl_total if acl_total else '')
            if acl_total:
                c.number_format = ACCT
            c.font = fnt
            if row_fill:
                c.fill = row_fill

            # Fill remaining cells in blank columns with fill
            if row_fill:
                for di in range(offset):
                    ws.cell(row=r, column=DATE_COL_START + di).fill = row_fill

            r += 1

        # Total row
        ws.cell(row=r, column=1, value=f"Total {section_label}").font = F8B
        ws.cell(row=r, column=1).fill = TOT_FILL
        grand_total = 0
        for di in range(nmonths):
            dt = header_dates[di]
            ym = (dt.year, dt.month) if hasattr(dt, 'year') else (
                dt.year(), dt.month())
            ttl = sum(monthly_data.get(ym, {}).get(p, 0) for p in pools)
            c = ws.cell(row=r, column=DATE_COL_START + di, value=ttl if ttl else '')
            if ttl:
                c.number_format = ACCT
            c.font = F8B
            c.fill = TOT_FILL
            grand_total += ttl
        c = ws.cell(row=r, column=acl_col, value=grand_total if grand_total else '')
        if grand_total:
            c.number_format = ACCT
        c.font = F8B
        c.fill = TOT_FILL
        r += 1

        return r, pool_values, nmonths

    # ── Charge offs section ──
    r = 5
    r, co_vals, nmonths = _write_section(
        ws, r, "Charge offs", co_monthly, pools,
        all_dates, acl_months_map)

    # Blank separator rows
    r += 1

    # ── Recoveries section ──
    r, rc_vals, _ = _write_section(
        ws, r, "Recoveries", rc_monthly, pools,
        all_dates, acl_months_map)

    # Blank separator rows
    r += 1

    # ── Net Loss section ──
    # Build net monthly data = chargeoffs - recoveries (COs are positive losses)
    net_monthly = {}
    all_ym_keys = set(list(co_monthly.keys()) + list(rc_monthly.keys()))
    for ym in all_ym_keys:
        net_monthly[ym] = {}
        co_pools = co_monthly.get(ym, {})
        rc_pools = rc_monthly.get(ym, {})
        all_pool_keys = set(list(co_pools.keys()) + list(rc_pools.keys()))
        for p in all_pool_keys:
            net_monthly[ym][p] = co_pools.get(p, 0) - rc_pools.get(p, 0)

    r, _, _ = _write_section(
        ws, r, "Net Loss", net_monthly, pools,
        all_dates, acl_months_map)

    # ── Page setup ──
    FIT_PAGES_WIDE = 5
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = FIT_PAGES_WIDE
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.25
    ws.page_margins.footer = 0.25
    ws.print_title_cols = 'A:A'


def _sheet_bal_adjust(wb, cu, snap, df, grades, config, hist=None):
    """Pool Balance Adjust – data sourced from WARM Risk Change Data Entry tab."""
    ws = wb.create_sheet("Pool_Balance Adjust")
    no_score = config.get('no_score_label', 'Not Reported')
    gl = [g for g in _all_grades(grades, no_score) if not _is_hidden(g)]

    # Columns: A=Grade/Pool label, B=Loan Report Balance, C=Bal Adjustment, D=Balance Sheet Total
    for ci, w in enumerate([23.4, 31.9, 21.1, 18.9], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.merge_cells('A2:D2')
    ws['A2'] = cu
    ws['A2'].font = V12B
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:D3')
    ws['A3'] = f"For Quarter Ending {_snap_display(snap)}"
    ws['A3'].font = V12B
    ws['A3'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A4:D4')
    ws['A4'] = "Balance Adjustment Detail"
    ws['A4'].font = V10B
    ws['A4'].alignment = Alignment(horizontal='center')

    # Column headers
    r = 6
    for ci, lbl in [(1, "Current Grade"), (2, "Loan Report Balance"),
                    (3, "Bal Adjustment"), (4, "Balance Sheet Total")]:
        cell = ws.cell(row=r, column=ci, value=lbl)
        cell.font = V10B
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Build pool list: WARM order, including non-risk-rated pools not in df
    _imp = hist.get('impaired', {}) if hist else {}
    _detail = _imp.get('pool_bal_detail', {})
    _risk_rated = _imp.get('risk_rated', {})
    warm_order = _imp.get('pool_order', [])
    df_pools = set(df['loan_pool'].unique())
    if warm_order:
        pools = [p for p in warm_order
                 if p in df_pools or p in _detail]
    else:
        pools = sorted(df_pools)

    def _lookup_pool_detail(pool_name):
        """Case-insensitive lookup into pool_bal_detail."""
        plc = pool_name.strip().lower()
        for k, v in _detail.items():
            if k.strip().lower() == plc:
                return v
        return {}

    r = 7
    pool_start_rows = []          # track first row of each pool group
    grand_loan = 0.0
    grand_adj = 0.0
    grand_bst = 0.0
    for pool in pools:
        pdata = _lookup_pool_detail(pool)
        is_rr = _risk_rated.get(pool, True)

        pool_start_rows.append(r)  # pool header row

        # Pool header
        ws.cell(row=r, column=1, value=pool).font = V12B
        ws.cell(row=r, column=1).fill = HDR_FILL
        ws.cell(row=r, column=1).font = HDR_FONT
        for ci in range(2, 5):
            ws.cell(row=r, column=ci).fill = HDR_FILL
        r += 1

        if is_rr:
            for g in gl:
                fnt = _grade_font10(g)
                gd = pdata.get(g, {})
                loan_bal = gd.get('loan_report_bal', 0.0)
                adj = gd.get('bal_adj', 0.0)
                bst = gd.get('balance_sheet_total', 0.0)

                ws.cell(row=r, column=1, value=g).font = fnt
                ws.cell(row=r, column=1).alignment = Alignment(horizontal='left')
                ws.cell(row=r, column=2, value=loan_bal).number_format = ACCT
                ws.cell(row=r, column=2).font = fnt
                ws.cell(row=r, column=3, value=adj).number_format = ACCT
                ws.cell(row=r, column=3).font = fnt
                ws.cell(row=r, column=4, value=bst).number_format = ACCT
                ws.cell(row=r, column=4).font = fnt
                r += 1

        # Pool total from RCDE
        td = pdata.get('Total', {})
        t_loan = td.get('loan_report_bal', 0.0)
        t_adj = td.get('bal_adj', 0.0)
        t_bst = td.get('balance_sheet_total', 0.0)

        ws.cell(row=r, column=1, value="Total").font = V10B
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='left')
        ws.cell(row=r, column=2, value=t_loan).number_format = ACCT
        ws.cell(row=r, column=2).font = V10B
        ws.cell(row=r, column=3, value=t_adj).number_format = ACCT
        ws.cell(row=r, column=3).font = V10B
        ws.cell(row=r, column=4, value=t_bst).number_format = ACCT
        ws.cell(row=r, column=4).font = V10B
        r += 2

        grand_loan += t_loan
        grand_adj += t_adj
        grand_bst += t_bst

    # Grand Totals row
    ws.cell(row=r, column=1, value="Grand Totals").font = V12B
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='left')
    ws.cell(row=r, column=2, value=grand_loan).number_format = ACCT
    ws.cell(row=r, column=2).font = V12B
    ws.cell(row=r, column=3, value=grand_adj).number_format = ACCT
    ws.cell(row=r, column=3).font = V12B
    ws.cell(row=r, column=4, value=grand_bst).number_format = ACCT
    ws.cell(row=r, column=4).font = V12B
    r += 1

    # Repeat header row 6 on every printed page
    ws.print_title_rows = '6:6'

    # ── Page setup: fit all columns on one page width ──
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.05

    # ── Smart page breaks between pool groupings ──
    # Estimate rows per page: ~10" usable height @ ~15pt rows ≈ 48 rows
    ROWS_PER_PAGE = 48
    last_data_row = r - 2  # last Total row
    if last_data_row > ROWS_PER_PAGE:
        page_end = ROWS_PER_PAGE
        while page_end < last_data_row:
            # Find the pool start row closest to (but not past) the natural break
            best = None
            for ps in pool_start_rows:
                # break goes on the blank row before the pool header (ps - 1)
                if ps - 1 <= 1:
                    continue
                if ps - 1 <= page_end:
                    best = ps - 1
            if best and best > 1:
                ws.row_breaks.append(Break(id=best))
                page_end = best + ROWS_PER_PAGE
            else:
                break  # no suitable break point found


def _sheet_appendix_supp(wb):
    """Appendix Supplemental text page – matches template exactly."""
    ws = wb.create_sheet("Appendix_Supplemental")
    ws.column_dimensions['A'].width = 118.0

    ws['A1'] = "Appendix"
    ws['A1'].font = V12B

    ws['A2'] = "Historical Loan Balances by Most Recent Credit Score"
    ws['A2'].font = V12B

    ws['A3'] = (
        "Concentrations of loans in pools and grades are important indicators of risk.  "
        "The dynamic nature of credit scores means that grade concentration may change "
        "consistently from quarter to quarter.  The deterioration of loans may lead to "
        "higher concentrations of loans in lower credit ranges without any additional "
        "funding of loans in those ranges.  Improvement of scores may lead to lower "
        "concentrations in ranges indicating opportunities for loan growth."
    )
    ws['A3'].font = V12
    ws['A3'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 63.0

    ws['A5'] = (
        "This report is presented as a line graph to track the concentration of loans "
        "in each pool by grade over time.  The trend lines in this report show the "
        "changing makeup of loans in the portfolio and the accompanying changes in risk."
    )
    ws['A5'].font = V12
    ws['A5'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[5].height = 31.5

    ws['A7'] = "Loss Factor Historical Detail"
    ws['A7'].font = V12B
    ws['A8'] = ">insert report details"
    ws['A8'].font = V12
    ws['A8'].alignment = Alignment(wrap_text=True)

    ws['A9'] = "Charge off and Recoveries Historical Detail"
    ws['A9'].font = V12B
    ws['A10'] = ">insert report details"
    ws['A10'].font = V12
    ws['A10'].alignment = Alignment(wrap_text=True)

    ws['A12'] = "Balance Adjustment Detail"
    ws['A12'].font = V12B
    ws['A13'] = ">insert report details"
    ws['A13'].font = V12
    ws['A13'].alignment = Alignment(wrap_text=True)

    # Page setup: portrait, scale 88%, 0.25" margins
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.scale = 88
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.05


# ══════════════════════════════════════════════════════════════════
# CHART POST-PROCESSING – patches chart XML to match template exactly
# ══════════════════════════════════════════════════════════════════

import zipfile, shutil, tempfile
import xml.etree.ElementTree as ET

_C_NS  = 'http://schemas.openxmlformats.org/drawingml/2006/chart'
_A_NS  = 'http://schemas.openxmlformats.org/drawingml/2006/main'
_R_NS  = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
_XDR_NS = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'


def _find_elem(parent, tag_name, ns=_C_NS):
    """Find a child element by local tag name."""
    for child in parent:
        ltag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if ltag == tag_name:
            return child
    return None


def _find_all(parent, tag_name, ns=_C_NS):
    """Find all child elements by local tag name."""
    results = []
    for child in parent:
        ltag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if ltag == tag_name:
            results.append(child)
    return results


def _set_or_add(parent, tag_name, attrib=None, ns=_C_NS, position=None):
    """Set attributes on existing element or create it."""
    elem = _find_elem(parent, tag_name)
    if elem is None:
        elem = ET.SubElement(parent, f'{{{ns}}}{tag_name}')
        if position is not None:
            # Move to correct position
            parent.remove(elem)
            parent.insert(position, elem)
    if attrib:
        elem.attrib.update(attrib)
    return elem


def _remove_elem(parent, tag_name):
    """Remove a child element by local tag name if it exists."""
    elem = _find_elem(parent, tag_name)
    if elem is not None:
        parent.remove(elem)


def _ensure_gridlines_nofill(ax_elem):
    """Ensure majorGridlines has spPr > ln > noFill."""
    mg = _find_elem(ax_elem, 'majorGridlines')
    if mg is None:
        mg = ET.SubElement(ax_elem, f'{{{_C_NS}}}majorGridlines')
    sp = _find_elem(mg, 'spPr', _C_NS)
    if sp is None:
        sp = ET.SubElement(mg, f'{{{_C_NS}}}spPr')
    ln = _find_elem(sp, 'ln', _A_NS)
    if ln is None:
        ln = ET.SubElement(sp, f'{{{_A_NS}}}ln')
    nf = _find_elem(ln, 'noFill', _A_NS)
    if nf is None:
        ET.SubElement(ln, f'{{{_A_NS}}}noFill')


def _ensure_shape_nofill(shape_parent):
    """Ensure c:spPr exists with noFill and line noFill for borderless shapes.

    Replaces any existing fill/line children with explicit noFill so a previously
    set solidFill or sized outline is removed (not merely supplemented).
    """
    sp = _find_elem(shape_parent, 'spPr', _C_NS)
    if sp is None:
        sp = ET.SubElement(shape_parent, f'{{{_C_NS}}}spPr')

    # Force fill = noFill (strip any existing fill children)
    for tag in ('noFill', 'solidFill', 'gradFill', 'blipFill', 'pattFill'):
        for e in sp.findall(f'{{{_A_NS}}}{tag}'):
            sp.remove(e)
    ET.SubElement(sp, f'{{{_A_NS}}}noFill')

    # Force line = noFill (replace any existing <a:ln> contents and attrs)
    for ln_old in sp.findall(f'{{{_A_NS}}}ln'):
        sp.remove(ln_old)
    ln = ET.SubElement(sp, f'{{{_A_NS}}}ln')
    ET.SubElement(ln, f'{{{_A_NS}}}noFill')


def _ensure_axis_line_nofill(ax_elem):
    """Ensure axis line itself is hidden (not just gridlines)."""
    sp = _find_elem(ax_elem, 'spPr', _C_NS)
    if sp is None:
        sp = ET.SubElement(ax_elem, f'{{{_C_NS}}}spPr')

    # Force line = noFill (replace any existing <a:ln> contents and attrs)
    for ln_old in sp.findall(f'{{{_A_NS}}}ln'):
        sp.remove(ln_old)
    ln = ET.SubElement(sp, f'{{{_A_NS}}}ln')
    ET.SubElement(ln, f'{{{_A_NS}}}noFill')


def _apply_graduated_transparency(bc_elem, base_color, step):
    """Add per-data-point fills to bars in a barChart with graduated alpha.

    The first data point is fully opaque; each subsequent point's alpha is
    reduced by `step` (OOXML alpha units; 100000 = opaque, 0 = clear).

    base_color: 6-char hex (RGB) used for solidFill srgbClr value.
    """
    for ser in _find_all(bc_elem, 'ser'):
        val = _find_elem(ser, 'val')
        n_pts = 0
        if val is not None:
            num_ref = _find_elem(val, 'numRef')
            if num_ref is not None:
                cache = _find_elem(num_ref, 'numCache')
                if cache is not None:
                    pt_count = _find_elem(cache, 'ptCount')
                    if pt_count is not None and pt_count.get('val'):
                        try:
                            n_pts = int(pt_count.get('val'))
                        except ValueError:
                            n_pts = 0
                    if not n_pts:
                        n_pts = len(_find_all(cache, 'pt'))
                if not n_pts:
                    # Fall back to parsing the <c:f> range, e.g. "Sheet!$F$72:$F$76"
                    f_elem = _find_elem(num_ref, 'f')
                    if f_elem is not None and f_elem.text:
                        m = re.search(r'\$?[A-Z]+\$?(\d+):\$?[A-Z]+\$?(\d+)',
                                      f_elem.text)
                        if m:
                            n_pts = int(m.group(2)) - int(m.group(1)) + 1
        if n_pts <= 0:
            continue

        # Remove any pre-existing dPt elements so we control the full set.
        for old in ser.findall(f'{{{_C_NS}}}dPt'):
            ser.remove(old)

        # Per CT_BarSer schema: dPt comes after idx/order/tx/spPr/invertIfNegative
        # /pictureOptions and before dLbls/cat/val.
        insert_at = 0
        for idx, child in enumerate(list(ser)):
            tag = child.tag.split('}')[-1]
            if tag in ('idx', 'order', 'tx', 'spPr', 'invertIfNegative',
                       'pictureOptions'):
                insert_at = idx + 1
            else:
                break

        for i in range(n_pts):
            alpha = max(0, 100000 - step * i)
            dpt = ET.Element(f'{{{_C_NS}}}dPt')
            ET.SubElement(dpt, f'{{{_C_NS}}}idx', {'val': str(i)})
            ET.SubElement(dpt, f'{{{_C_NS}}}invertIfNegative', {'val': '0'})
            ET.SubElement(dpt, f'{{{_C_NS}}}bubble3D', {'val': '0'})
            sp = ET.SubElement(dpt, f'{{{_C_NS}}}spPr')
            sf = ET.SubElement(sp, f'{{{_A_NS}}}solidFill')
            clr = ET.SubElement(sf, f'{{{_A_NS}}}srgbClr', {'val': base_color})
            ET.SubElement(clr, f'{{{_A_NS}}}alpha', {'val': str(alpha)})
            ln = ET.SubElement(sp, f'{{{_A_NS}}}ln')
            ET.SubElement(ln, f'{{{_A_NS}}}noFill')
            ser.insert(insert_at, dpt)
            insert_at += 1


def _normalize_impdet_anchor(anchor):
    """Force Impr Deter charts into a fixed 2x2 layout with identical extents.

    The Impr Deter sheet has an unusually wide column F (43.29 vs ~8.43 default),
    so a plain 5-col left + 5-col right split produces unequal chart widths.
    We split the strip down the middle of column F using an EMU offset so all
    four charts end up the same outer width.
    """
    frm = anchor.find(f'{{{_XDR_NS}}}from')
    to = anchor.find(f'{{{_XDR_NS}}}to')
    if frm is None or to is None:
        return False, None

    from_col = int(frm.find(f'{{{_XDR_NS}}}col').text)
    from_row = int(frm.find(f'{{{_XDR_NS}}}row').text)

    is_left = from_col < 5
    top_row = 11 if from_row < 25 else 25
    bottom_row = top_row + 14
    mid_off = 1209675  # ~17.43 width units into col F (Calibri 11)

    if is_left:
        left_col, left_off = 0, 0
        right_col, right_off = 5, mid_off
    else:
        left_col, left_off = 5, mid_off
        right_col, right_off = 10, 0

    def _set(parent, tag, value):
        e = parent.find(f'{{{_XDR_NS}}}{tag}')
        if e is None:
            e = ET.SubElement(parent, f'{{{_XDR_NS}}}{tag}')
        old = e.text
        e.text = str(value)
        return old != e.text

    changed = False
    changed |= _set(frm, 'col', left_col)
    changed |= _set(frm, 'colOff', left_off)
    changed |= _set(frm, 'row', top_row)
    changed |= _set(frm, 'rowOff', 0)

    changed |= _set(to, 'col', right_col)
    changed |= _set(to, 'colOff', right_off)
    changed |= _set(to, 'row', bottom_row)
    changed |= _set(to, 'rowOff', 0)

    return changed, (left_col, top_row, right_col, bottom_row)


def _set_chart_meta(root):
    """Set chartSpace-level metadata: roundedCorners, style, etc."""
    # Ensure these come right after any existing lang element
    _set_or_add(root, 'roundedCorners', {'val': '1'})
    _set_or_add(root, 'style', {'val': '2'})


def _fix_title(chart_elem):
    """Fix title: add overlay, autoTitleDeleted."""
    title = _find_elem(chart_elem, 'title')
    if title is not None:
        _set_or_add(title, 'overlay', {'val': '1'})
    _set_or_add(chart_elem, 'autoTitleDeleted', {'val': '0'})


def _fix_series_common(ser_elem):
    """Add invertIfNegative and fix dLbls spPr for a series."""
    _set_or_add(ser_elem, 'invertIfNegative', {'val': '0'})
    dlbls = _find_elem(ser_elem, 'dLbls')
    if dlbls is not None:
        # Add spPr with noFill, line noFill, effectLst if not present
        sp = _find_elem(dlbls, 'spPr', _C_NS)
        if sp is None:
            sp = ET.Element(f'{{{_C_NS}}}spPr')
            dlbls.insert(0, sp)
            ET.SubElement(sp, f'{{{_A_NS}}}noFill')
            ln = ET.SubElement(sp, f'{{{_A_NS}}}ln')
            ET.SubElement(ln, f'{{{_A_NS}}}noFill')
            ET.SubElement(sp, f'{{{_A_NS}}}effectLst')
        # Ensure showLegendKey, showCatName etc.
        _set_or_add(dlbls, 'showLegendKey', {'val': '0'})
        _set_or_add(dlbls, 'showCatName', {'val': '0'})
        _set_or_add(dlbls, 'showSerName', {'val': '0'})
        _set_or_add(dlbls, 'showPercent', {'val': '1'})
        _set_or_add(dlbls, 'showBubbleSize', {'val': '1'})
        _set_or_add(dlbls, 'showLeaderLines', {'val': '0'})


def _fix_barchart_defaults(bc_elem):
    """Add varyColors and chart-level dLbls."""
    _set_or_add(bc_elem, 'varyColors', {'val': '1'})
    # Add chart-level dLbls with all show* = 0
    chart_dlbls = _find_elem(bc_elem, 'dLbls')
    if chart_dlbls is None:
        chart_dlbls = ET.SubElement(bc_elem, f'{{{_C_NS}}}dLbls')
    for attr in ['showLegendKey', 'showVal', 'showCatName', 'showSerName',
                 'showPercent', 'showBubbleSize']:
        _set_or_add(chart_dlbls, attr, {'val': '0'})


def _fix_catax_col(catax):
    """Fix catAx for column charts: axPos=b, numFmt General, noMultiLvlLbl."""
    _set_or_add(catax, 'axPos', {'val': 'b'})
    _set_or_add(catax, 'numFmt', {'formatCode': 'General', 'sourceLinked': '1'})
    _set_or_add(catax, 'crosses', {'val': 'autoZero'})
    _set_or_add(catax, 'auto', {'val': '1'})
    _set_or_add(catax, 'lblAlgn', {'val': 'ctr'})
    _set_or_add(catax, 'noMultiLvlLbl', {'val': '1'})


def _fix_catax_bar(catax, numfmt_code):
    """Fix catAx for bar charts: axPos=r, tickLblPos=high, noMultiLvlLbl."""
    _set_or_add(catax, 'axPos', {'val': 'r'})
    _set_or_add(catax, 'numFmt', {'formatCode': numfmt_code, 'sourceLinked': '0'})
    _set_or_add(catax, 'tickLblPos', {'val': 'high'})
    _set_or_add(catax, 'crosses', {'val': 'autoZero'})
    _set_or_add(catax, 'auto', {'val': '1'})
    _set_or_add(catax, 'lblAlgn', {'val': 'ctr'})
    _set_or_add(catax, 'noMultiLvlLbl', {'val': '1'})


def _fix_valax_col(valax, numfmt_code):
    """Fix valAx for column charts: majorGridlines with noFill."""
    _ensure_gridlines_nofill(valax)
    _set_or_add(valax, 'numFmt', {'formatCode': numfmt_code, 'sourceLinked': '0'})
    _set_or_add(valax, 'crosses', {'val': 'autoZero'})
    _set_or_add(valax, 'crossBetween', {'val': 'between'})


def _fix_valax_bar(valax, numfmt_code):
    """Fix valAx for bar charts: axPos=b, orientation=maxMin, majorGridlines with noFill."""
    _set_or_add(valax, 'axPos', {'val': 'b'})
    _ensure_gridlines_nofill(valax)
    _set_or_add(valax, 'tickLblPos', {'val': 'none'})
    _set_or_add(valax, 'numFmt', {'formatCode': numfmt_code, 'sourceLinked': '1'})
    _set_or_add(valax, 'crosses', {'val': 'autoZero'})
    _set_or_add(valax, 'crossBetween', {'val': 'between'})


def _fix_manual_layout(layout_elem):
    """Fix ManualLayout: add layoutTarget, remove wMode/hMode."""
    ml = _find_elem(layout_elem, 'manualLayout')
    if ml is not None:
        _set_or_add(ml, 'layoutTarget', {'val': 'inner'})
        _remove_elem(ml, 'wMode')
        _remove_elem(ml, 'hMode')


def _fix_txpr_paragraphs(root_elem):
    """Replace <r><t/></r> with <endParaRPr lang='en-US'/> in all txPr blocks,
    and add <lstStyle/> after bodyPr."""
    for elem in root_elem.iter():
        ltag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        if ltag == 'txPr':
            # Add lstStyle after bodyPr if missing
            body_pr = _find_elem(elem, 'bodyPr', _A_NS)
            lst_style = _find_elem(elem, 'lstStyle', _A_NS)
            if body_pr is not None and lst_style is None:
                idx = list(elem).index(body_pr) + 1
                ls = ET.Element(f'{{{_A_NS}}}lstStyle')
                elem.insert(idx, ls)
            # Fix paragraphs: replace <r><t/></r> with <endParaRPr>
            for p in _find_all(elem, 'p', _A_NS):
                r_elem = _find_elem(p, 'r', _A_NS)
                if r_elem is not None:
                    t_elem = _find_elem(r_elem, 't', _A_NS)
                    if t_elem is not None and (t_elem.text is None or t_elem.text.strip() == ''):
                        p.remove(r_elem)
                        epr = ET.SubElement(p, f'{{{_A_NS}}}endParaRPr')
                        epr.set('lang', 'en-US')
        # Fix title rich text: add rPr lang and lstStyle
        if ltag == 'rich':
            body_pr = _find_elem(elem, 'bodyPr', _A_NS)
            lst_style = _find_elem(elem, 'lstStyle', _A_NS)
            if body_pr is not None and lst_style is None:
                idx = list(elem).index(body_pr) + 1
                ls = ET.Element(f'{{{_A_NS}}}lstStyle')
                elem.insert(idx, ls)
            for p in _find_all(elem, 'p', _A_NS):
                for r in _find_all(p, 'r', _A_NS):
                    rpr = _find_elem(r, 'rPr', _A_NS)
                    if rpr is None:
                        rpr = ET.Element(f'{{{_A_NS}}}rPr')
                        rpr.set('lang', 'en-US')
                        r.insert(0, rpr)


def _add_plot_vis_defaults(chart_elem):
    """Add plotVisOnly, dispBlanksAs, showDLblsOverMax."""
    _set_or_add(chart_elem, 'plotVisOnly', {'val': '1'})
    _set_or_add(chart_elem, 'dispBlanksAs', {'val': 'gap'})
    _set_or_add(chart_elem, 'showDLblsOverMax', {'val': '1'})


def patch_dq_pie_zero_labels(xlsx_path):
    """Post-process DQ pie charts to inject <delete val='1'/> for zero-value slices.

    openpyxl's DataLabel class does not support the <c:delete> element, so
    this must be done via raw XML manipulation after the workbook is saved.
    Cell values are read via openpyxl, then chart XML is patched via zipfile.
    """
    from openpyxl import load_workbook as _lwb
    from openpyxl.utils import range_boundaries

    wb = _lwb(xlsx_path, data_only=True)

    # Build a lookup: (sheet_name, col, row) -> value from all Risk Chg sheets
    cell_vals = {}
    for ws_name in wb.sheetnames:
        if 'Risk Chg' not in ws_name and 'Risk Change' not in ws_name:
            continue
        ws = wb[ws_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell_vals[(ws_name, cell.column, cell.row)] = cell.value
    wb.close()

    ET.register_namespace('', _C_NS)
    ET.register_namespace('a', _A_NS)
    ET.register_namespace('r', _R_NS)

    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(tmp_fd)

    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            patched_charts = {}

            for name in zin.namelist():
                if not (name.startswith('xl/charts/chart') and name.endswith('.xml')):
                    continue
                raw = zin.read(name).decode('utf-8')
                if 'pieChart' not in raw:
                    continue

                root = ET.fromstring(raw)
                chart_elem = _find_elem(root, 'chart')
                if chart_elem is None:
                    continue
                plot_area = _find_elem(chart_elem, 'plotArea')
                if plot_area is None:
                    continue
                pie = _find_elem(plot_area, 'pieChart')
                if pie is None:
                    continue

                changed = False
                for ser in _find_all(pie, 'ser'):
                    val_elem = _find_elem(ser, 'val')
                    if val_elem is None:
                        continue
                    num_ref = _find_elem(val_elem, 'numRef')
                    if num_ref is None:
                        continue
                    f_elem = _find_elem(num_ref, 'f')
                    if f_elem is None or not f_elem.text:
                        continue

                    # Parse reference like "'Risk Chg Visa'!$M$44:$M$47"
                    ref_text = f_elem.text
                    if '!' not in ref_text:
                        continue
                    sheet_part, cell_part = ref_text.rsplit('!', 1)
                    sheet_name = sheet_part.strip("'")
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(cell_part)
                    except Exception:
                        continue

                    # Find zero-value slice indices
                    zero_indices = set()
                    for i, r in enumerate(range(min_row, max_row + 1)):
                        val = cell_vals.get((sheet_name, min_col, r), 0)
                        try:
                            fval = float(val) if val else 0.0
                        except (ValueError, TypeError):
                            fval = 0.0
                        if fval <= 0.001:  # 0.1% threshold
                            zero_indices.add(i)

                    if not zero_indices:
                        continue

                    dlbls = _find_elem(ser, 'dLbls')
                    if dlbls is None:
                        continue

                    # Remove any existing <dLbl> entries
                    for existing_dl in list(_find_all(dlbls, 'dLbl')):
                        dlbls.remove(existing_dl)

                    # Insert <dLbl><idx val="N"/><delete val="1"/></dLbl>
                    for zi in sorted(zero_indices):
                        dl_elem = ET.Element(f'{{{_C_NS}}}dLbl')
                        idx_elem = ET.SubElement(dl_elem, f'{{{_C_NS}}}idx')
                        idx_elem.set('val', str(zi))
                        del_elem = ET.SubElement(dl_elem, f'{{{_C_NS}}}delete')
                        del_elem.set('val', '1')
                        dlbls.insert(0, dl_elem)
                    changed = True

                if changed:
                    ET.indent(root, space='')
                    patched_xml = ET.tostring(root, encoding='unicode',
                                              xml_declaration=True)
                    patched_charts[name] = patched_xml

            if not patched_charts:
                return

            with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if item.filename in patched_charts:
                        zout.writestr(item, patched_charts[item.filename])
                    else:
                        zout.writestr(item, zin.read(item.filename))

        shutil.move(tmp_path, xlsx_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise
    finally:
        # Restore pristine namespace bindings so subsequent openpyxl saves
        # don't emit duplicate xmlns on <wsDr> drawing roots.
        ET.register_namespace('xdr', _XDR_NS)
        ET.register_namespace('c', _C_NS)
        ET.register_namespace('a', _A_NS)
        ET.register_namespace('r', _R_NS)


def patch_impdet_charts(xlsx_path):
    """Post-process chart XML inside the saved xlsx to match Brian's template."""
    ET.register_namespace('', _C_NS)
    ET.register_namespace('a', _A_NS)
    ET.register_namespace('r', _R_NS)
    # NOTE: We restore the spreadsheetDrawing default namespace at the end
    # to avoid polluting subsequent openpyxl saves (drawing XML).

    ACCT_FMT = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"_);_(@_)'

    # Work with a temp file
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(tmp_fd)

    try:
        # Find the Impr Deter sheet and its charts
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            wb_xml = zin.read('xl/workbook.xml').decode()
            wb_root = ET.fromstring(wb_xml)
            wb_ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            sheets = wb_root.findall('.//s:sheet', wb_ns)

            target_rid = None
            for s in sheets:
                name = s.get('name', '')
                if 'impr' in name.lower() and 'deter' in name.lower():
                    target_rid = s.get(f'{{{_R_NS}}}id')
                    break
            if not target_rid:
                return  # No Impr Deter sheet found

            rels_xml = zin.read('xl/_rels/workbook.xml.rels').decode()
            rels_root = ET.fromstring(rels_xml)
            target_sheet = None
            for rel in rels_root.iter():
                if rel.get('Id', '') == target_rid:
                    target_sheet = rel.get('Target')
            if not target_sheet:
                return

            sheet_num = re.search(r'sheet(\d+)', target_sheet).group(1)
            sheet_rels_path = f'xl/worksheets/_rels/sheet{sheet_num}.xml.rels'
            drawing_path = None
            if sheet_rels_path in zin.namelist():
                sr = zin.read(sheet_rels_path).decode()
                sr_root = ET.fromstring(sr)
                for rel in sr_root.iter():
                    t = rel.get('Target', '')
                    if 'drawing' in t.lower():
                        if t.startswith('/'):
                            drawing_path = t.lstrip('/')
                        elif t.startswith('..'):
                            drawing_path = 'xl/' + t.replace('../', '')
                        else:
                            drawing_path = 'xl/worksheets/' + t
            if not drawing_path:
                return

            # Read drawing to find chart anchors
            drawing_xml = zin.read(drawing_path).decode()
            drawing_root = ET.fromstring(drawing_xml)

            chart_refs = []
            drawing_changed = False
            for anchor in drawing_root.findall(f'{{{_XDR_NS}}}twoCellAnchor'):
                anchor_changed, coords = _normalize_impdet_anchor(anchor)
                if anchor_changed:
                    drawing_changed = True

                if coords is None:
                    continue
                from_col, from_row, to_col, to_row = coords
                for elem in anchor.iter():
                    if 'chart' in elem.tag and elem.get(f'{{{_R_NS}}}id'):
                        rid = elem.get(f'{{{_R_NS}}}id')
                        chart_refs.append((rid, from_col, from_row, to_col, to_row))

            patched_drawings = {}
            if drawing_changed:
                ET.indent(drawing_root, space='')
                patched_drawings[drawing_path] = ET.tostring(
                    drawing_root, encoding='unicode', xml_declaration=True
                )

            # Resolve rIds to chart file paths
            draw_rels_path = drawing_path.replace('drawings/', 'drawings/_rels/').replace('.xml', '.xml.rels')
            dr_xml = zin.read(draw_rels_path).decode()
            dr_root = ET.fromstring(dr_xml)
            rid_to_file = {}
            for rel in dr_root.iter():
                rid = rel.get('Id', '')
                target = rel.get('Target', '')
                if rid and target:
                    if target.startswith('/'):
                        rid_to_file[rid] = target.lstrip('/')
                    elif target.startswith('..'):
                        rid_to_file[rid] = 'xl/' + target.replace('../', '')
                    else:
                        rid_to_file[rid] = 'xl/charts/' + target

            # Classify charts by position and patch
            patched_charts = {}  # chart_path -> new_xml
            for rid, fc, fr, tc, tr in chart_refs:
                chart_path = rid_to_file.get(rid)
                if not chart_path:
                    continue

                raw = zin.read(chart_path).decode()
                root = ET.fromstring(raw)
                chart_elem = _find_elem(root, 'chart')
                if chart_elem is None:
                    continue
                plot_area = _find_elem(chart_elem, 'plotArea')
                if plot_area is None:
                    continue
                bc = _find_elem(plot_area, 'barChart')
                if bc is None:
                    continue

                bar_dir = _find_elem(bc, 'barDir')
                grouping = _find_elem(bc, 'grouping')
                is_col = bar_dir is not None and bar_dir.get('val') == 'col'
                is_bar = bar_dir is not None and bar_dir.get('val') == 'bar'
                is_stacked = grouping is not None and grouping.get('val') == 'stacked'

                catax = _find_elem(plot_area, 'catAx')
                valax = _find_elem(plot_area, 'valAx')

                # --- Apply common fixes ---
                _set_chart_meta(root)
                _fix_title(chart_elem)
                _fix_barchart_defaults(bc)
                _fix_txpr_paragraphs(root)
                _add_plot_vis_defaults(chart_elem)
                # Chart-area border lives on the chartSpace (root), not on c:chart
                _ensure_shape_nofill(root)
                _ensure_shape_nofill(plot_area)

                for ser in _find_all(bc, 'ser'):
                    _fix_series_common(ser)

                if is_col:
                    # Column charts (Improved / Deteriorated)
                    if catax is not None:
                        _fix_catax_col(catax)
                        _ensure_axis_line_nofill(catax)
                    if valax is not None:
                        _fix_valax_col(valax, ACCT_FMT)
                        _ensure_axis_line_nofill(valax)

                    # Improved chart (left column, fc==0): graduated transparency
                    # on bars — top grade in chart at 0% (alpha 100000), each
                    # subsequent grade +15% transparency (alpha -15000).
                    if fc == 0:
                        _apply_graduated_transparency(bc, base_color='0D4D5E',
                                                     step=15000)
                    # Deteriorated chart (right column, fc==5): same graduated
                    # transparency in MAROON.
                    elif fc == 5:
                        _apply_graduated_transparency(bc, base_color='3D1A1A',
                                                     step=15000)

                elif is_bar:
                    # Bar charts (Impr/Deter stacked, Net Change)
                    if is_stacked:
                        numfmt = '0%'
                    else:
                        numfmt = '0.0%'

                    if catax is not None:
                        _fix_catax_bar(catax, numfmt)
                        _ensure_axis_line_nofill(catax)
                    if valax is not None:
                        _fix_valax_bar(valax, numfmt)
                        _ensure_axis_line_nofill(valax)

                    # Fix ManualLayout
                    layout = _find_elem(plot_area, 'layout')
                    if layout is not None:
                        _fix_manual_layout(layout)

                    # Net Change (clustered bar): no data labels
                    if not is_stacked:
                        chart_dlbls = _find_elem(bc, 'dLbls')
                        if chart_dlbls is not None:
                            _set_or_add(chart_dlbls, 'dLblPos', {'val': 'inBase'})
                            for attr in ['showLegendKey', 'showVal', 'showCatName',
                                         'showSerName', 'showPercent', 'showBubbleSize']:
                                _set_or_add(chart_dlbls, attr, {'val': '0'})

                # Serialize back
                ET.indent(root, space='')
                patched_xml = ET.tostring(root, encoding='unicode', xml_declaration=True)
                patched_charts[chart_path] = patched_xml

            # Rewrite the xlsx with patched charts
            with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if item.filename in patched_charts:
                        zout.writestr(item, patched_charts[item.filename])
                    elif item.filename in patched_drawings:
                        zout.writestr(item, patched_drawings[item.filename])
                    else:
                        zout.writestr(item, zin.read(item.filename))

        # Replace original with patched version
        shutil.move(tmp_path, xlsx_path)

    except Exception:
        # Clean up temp file on error
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise
    finally:
        # Restore pristine namespace bindings so subsequent openpyxl saves
        # don't emit duplicate xmlns on <wsDr> drawing roots.
        ET.register_namespace('xdr', _XDR_NS)
        ET.register_namespace('c', _C_NS)
        ET.register_namespace('a', _A_NS)
        ET.register_namespace('r', _R_NS)


def patch_drawing_onecell_to_twocell(xlsx_path):
    """Rewrite all <oneCellAnchor> to <twoCellAnchor> in drawing XML for Excel compatibility.
    Also fixes duplicate default namespace declarations that can occur from
    ET.register_namespace pollution."""
    import zipfile, shutil, tempfile, xml.etree.ElementTree as ET
    import re as _re
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(tmp_fd)
    patched = False

    # The correct default namespace for drawing XML
    XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'

    def _fix_drawing_xml(raw_bytes):
        """Parse drawing XML, fixing duplicate default namespace if needed.
        Returns (root, ns_was_fixed) where ns_was_fixed is True if a stale
        default xmlns was stripped (meaning the file MUST be rewritten)."""
        text = raw_bytes.decode('utf-8') if isinstance(raw_bytes, bytes) else raw_bytes
        # Remove any stale default namespace (e.g. chart ns) that isn't spreadsheetDrawing
        # Pattern: xmlns="<something-other-than-spreadsheetDrawing>"
        new_text = _re.sub(
            r'\s+xmlns="(?!http://schemas\.openxmlformats\.org/drawingml/2006/spreadsheetDrawing)[^"]*"',
            '', text, count=1)
        return ET.fromstring(new_text), (new_text != text)

    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename.startswith('xl/drawings/') and item.filename.endswith('.xml'):
                        try:
                            root, ns_fixed = _fix_drawing_xml(data)
                        except ET.ParseError:
                            zout.writestr(item, data)
                            continue
                        changed = ns_fixed  # Rewrite if we stripped a bogus default xmlns
                        for anchor in list(root):
                            if anchor.tag.endswith('oneCellAnchor'):
                                anchor.tag = anchor.tag.replace('oneCellAnchor', 'twoCellAnchor')
                                # Read <from><row> to compute <to><row>
                                from_el = anchor.find('{*}from')
                                from_row = int(from_el.find('{*}row').text) if from_el is not None else 0
                                # Remove <ext> (only valid for oneCellAnchor)
                                ext_el = anchor.find('{*}ext')
                                if ext_el is not None:
                                    anchor.remove(ext_el)
                                # Add a <to> element if missing
                                if anchor.find('{*}to') is None:
                                    to = ET.Element('to')
                                    for tag, val in zip(['col', 'colOff', 'row', 'rowOff'],
                                                        ['5', '0', str(from_row + 11), '0']):
                                        e = ET.Element(tag)
                                        e.text = val
                                        to.append(e)
                                    anchor.insert(1, to)
                                changed = True
                        if changed:
                            patched = True
                            # Restore canonical drawing namespaces before serializing so
                            # we don't emit a stale xmlns="...chart" on the wsDr root.
                            ET.register_namespace('', XDR)
                            ET.register_namespace('a', 'http://schemas.openxmlformats.org/drawingml/2006/main')
                            ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
                            ET.register_namespace('c', 'http://schemas.openxmlformats.org/drawingml/2006/chart')
                            ET.indent(root, space='')
                            data = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                        zout.writestr(item, data)
                    else:
                        zout.writestr(item, data)
        if patched:
            shutil.move(tmp_path, xlsx_path)
        else:
            os.remove(tmp_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise


def patch_remove_chart_borders_and_axis_lines(xlsx_path):
    """Sweep every xl/charts/chart*.xml and:
       • strip the chart-area border (chartSpace → spPr → line = noFill)
       • hide every axis line (catAx/valAx/dateAx/serAx → spPr → line = noFill)
       • normalize every chart title to Calibri (size 18 by default;
         charts on the Impr Deter sheet use size 10), not bold
    """
    import zipfile, shutil, tempfile
    ET.register_namespace('', _C_NS)
    ET.register_namespace('a', _A_NS)
    ET.register_namespace('r', _R_NS)

    def _normalize_title(title_elem, sz):
        # Find the rich text under c:title/c:tx/c:rich
        tx = title_elem.find(f'{{{_C_NS}}}tx')
        if tx is None:
            return
        rich = tx.find(f'{{{_C_NS}}}rich')
        if rich is None:
            return
        for rpr_tag in ('defRPr', 'rPr', 'endParaRPr'):
            for rpr in rich.iter(f'{{{_A_NS}}}{rpr_tag}'):
                rpr.set('sz', str(sz))
                rpr.set('b', '0')
                # remove any existing latin typeface, then add Calibri
                for lt in rpr.findall(f'{{{_A_NS}}}latin'):
                    rpr.remove(lt)
                latin = ET.SubElement(rpr, f'{{{_A_NS}}}latin')
                latin.set('typeface', 'Calibri')

    # Identify which chart files are on the Impr Deter sheet so we can
    # use a smaller title font there.
    impdet_charts = set()
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            wb_xml = zin.read('xl/workbook.xml').decode()
            wb_root = ET.fromstring(wb_xml)
            wb_ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            target_rid = None
            for s in wb_root.findall('.//s:sheet', wb_ns):
                if 'impr' in s.get('name', '').lower() and 'deter' in s.get('name', '').lower():
                    target_rid = s.get(f'{{{_R_NS}}}id')
                    break
            if target_rid:
                rels_root = ET.fromstring(zin.read('xl/_rels/workbook.xml.rels').decode())
                target_sheet = None
                for rel in rels_root.iter():
                    if rel.get('Id', '') == target_rid:
                        target_sheet = rel.get('Target')
                if target_sheet:
                    m = re.search(r'sheet(\d+)', target_sheet)
                    if m:
                        sheet_rels_path = f'xl/worksheets/_rels/sheet{m.group(1)}.xml.rels'
                        if sheet_rels_path in zin.namelist():
                            sr = ET.fromstring(zin.read(sheet_rels_path).decode())
                            drawing_path = None
                            for rel in sr.iter():
                                t = rel.get('Target', '')
                                if 'drawing' in t.lower():
                                    if t.startswith('/'):
                                        drawing_path = t.lstrip('/')
                                    elif t.startswith('..'):
                                        drawing_path = 'xl/' + t.replace('../', '')
                                    else:
                                        drawing_path = 'xl/worksheets/' + t
                            if drawing_path:
                                drawing_rels_name = drawing_path.rsplit('/', 1)[0] + '/_rels/' + drawing_path.rsplit('/', 1)[1] + '.rels'
                                if drawing_rels_name in zin.namelist():
                                    dr = ET.fromstring(zin.read(drawing_rels_name).decode())
                                    for rel in dr.iter():
                                        t = rel.get('Target', '')
                                        if 'chart' in t.lower() and t.endswith('.xml'):
                                            cf = t.lstrip('/').replace('../', '')
                                            if not cf.startswith('xl/'):
                                                cf = 'xl/' + cf
                                            impdet_charts.add(cf)
    except Exception:
        impdet_charts = set()

    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(tmp_fd)
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename.startswith('xl/charts/chart') and item.filename.endswith('.xml'):
                        try:
                            root = ET.fromstring(data)
                            # 1) chart-area border off (chartSpace is root)
                            _ensure_shape_nofill(root)
                            # 2) axis lines off for every axis
                            for ax_tag in ('catAx', 'valAx', 'dateAx', 'serAx'):
                                for ax in root.iter(f'{{{_C_NS}}}{ax_tag}'):
                                    _ensure_axis_line_nofill(ax)
                            # 3) normalize all chart titles
                            sz = 1800 if item.filename in impdet_charts else 1800
                            for title in root.iter(f'{{{_C_NS}}}title'):
                                _normalize_title(title, sz)
                            data = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                        except ET.ParseError:
                            pass
                    zout.writestr(item, data)
        shutil.move(tmp_path, xlsx_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise
    finally:
        # Restore canonical namespaces so subsequent openpyxl saves don't
        # emit a stale xmlns="...chart" on the <wsDr> root of drawing files.
        ET.register_namespace('', 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing')
        ET.register_namespace('xdr', 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing')
        ET.register_namespace('c', _C_NS)
        ET.register_namespace('a', _A_NS)
        ET.register_namespace('r', _R_NS)


# ══════════════════════════════════════════════════════════════════
# COMPOSERS
# ══════════════════════════════════════════════════════════════════

def compose_vizo_main(client, snap, df, config, grades, hist=None):
    """Build complete Vizo-format main CECL Credit Migration workbook."""
    cu = config['credit_union']
    pools = _ordered_pools(df, hist)
    wb = Workbook()
    _apply_vizo_theme(wb)

    _sheet_cover(wb, cu, snap, supplemental=False)
    _sheet_report_index(wb, cu, snap, supplemental=False)
    # ACL Env by Pool Mgmt Adj is built FIRST so it can compute and stash
    # the canonical pooled total allowance on hist['impaired']. Impr Deter
    # then consumes that stash via _compute_acl_totals, guaranteeing both
    # tabs show the same "Total Allowance Needed". Tab display order is
    # restored below with wb.move_sheet so the final workbook still shows
    # Impr Deter / Risk Change / ACL Env in the original sequence.
    _sheet_acl_reserve(wb, cu, snap, df, grades, config, hist)
    _sheet_impdet(wb, cu, snap, df, grades, config, hist)
    _sheet_risk_change(wb, cu, snap, df, grades, config, pool_name=None, hist=hist)
    # Restore intended tab order: push ACL Env past Impr Deter + Risk Change.
    try:
        wb.move_sheet("ACL Env by Pool Mgmt Adj", offset=2)
    except (KeyError, ValueError):
        pass

    # Per-pool Risk Change sheets: use WARM order, risk-rated pools only
    _imp = hist.get('impaired', {}) if hist else {}
    warm_order = _imp.get('pool_order', [])
    risk_rated = _imp.get('risk_rated', {})
    # Normalized lookup so trailing whitespace / case differences don't slip
    # through and create tabs for non-risk-rated pools (e.g. Participation Loans)
    _rr_norm = {k.strip().lower(): v for k, v in risk_rated.items()}

    def _is_rr(pool):
        flag = _rr_norm.get(pool.strip().lower())
        if flag is not None:
            return bool(flag)
        # Fallback: treat as risk-rated only if the pool's loans span >1 grade
        sub = df[df['loan_pool'] == pool]
        if sub.empty:
            return False
        return sub['current_grade'].nunique() > 1

    if warm_order:
        rr_pools = [p for p in warm_order
                    if _is_rr(p) and not p.upper().startswith('HIDE')]
    else:
        rr_pools = [p for p in pools
                    if _is_rr(p) and not p.upper().startswith('HIDE')]
    for pool in rr_pools:
        pdf = df[df['loan_pool'] == pool]
        _sheet_risk_change(wb, cu, snap, pdf, grades, config, pool_name=pool, hist=hist)

    # Insert new tabs after 'Display CO-Recov-DQ' and before '>Envir Fact Ranges'
    _sheet_env_factor(wb, cu, snap, df, grades, config, hist)
    _sheet_loss_factor(wb, cu, snap, df, grades, config, hist)
    _sheet_co_recov_dq(wb, cu, snap, df, config, hist)

    # Insert Introduction-Vizo and Executive Summary-Vizo tabs from template
    from openpyxl import load_workbook
    template_path = os.path.join(_WORKSPACE_BASE, 'Sample Reports', 'YYYY-MM CECL-Migration-WARM - Template Credit Union with Vizo.xlsx')
    if os.path.exists(template_path):
        template_wb = load_workbook(template_path)
        for tab_name in ["Introduction-Vizo", "Executive Summary-Vizo"]:
            if tab_name in template_wb.sheetnames:
                tmpl_ws = template_wb[tab_name]
                # Create new sheet and copy all cells, styles, merges, row/col dims, page setup
                new_ws = wb.create_sheet(tab_name)
                from openpyxl.cell.cell import MergedCell
                from copy import copy
                for row in tmpl_ws.iter_rows():
                    for cell in row:
                        if isinstance(cell, MergedCell):
                            continue
                        new_cell = new_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = cell.number_format
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
                # Copy merged cells
                for merged in tmpl_ws.merged_cells.ranges:
                    new_ws.merge_cells(str(merged))
                # Copy column widths
                for col_letter, dim in tmpl_ws.column_dimensions.items():
                    new_ws.column_dimensions[col_letter].width = dim.width
                # Copy row heights
                for row_idx, dim in tmpl_ws.row_dimensions.items():
                    new_ws.row_dimensions[row_idx].height = dim.height
                # Copy page margins (safe individual values)
                new_ws.page_margins.left = tmpl_ws.page_margins.left
                new_ws.page_margins.right = tmpl_ws.page_margins.right
                new_ws.page_margins.top = tmpl_ws.page_margins.top
                new_ws.page_margins.bottom = tmpl_ws.page_margins.bottom
                new_ws.page_margins.header = tmpl_ws.page_margins.header
                new_ws.page_margins.footer = tmpl_ws.page_margins.footer
                # Copy page setup (orientation, paper size)
                new_ws.page_setup.orientation = tmpl_ws.page_setup.orientation
                new_ws.page_setup.paperSize = tmpl_ws.page_setup.paperSize
                new_ws.page_setup.fitToWidth = tmpl_ws.page_setup.fitToWidth
                new_ws.page_setup.fitToHeight = tmpl_ws.page_setup.fitToHeight
                # Set print area to actual content rows to avoid blank trailing pages
                last_content_row = max(
                    (cell.row for row in new_ws.iter_rows() for cell in row if cell.value),
                    default=new_ws.max_row,
                )
                last_col_letter = get_column_letter(new_ws.max_column)
                new_ws.print_area = f'A1:{last_col_letter}{last_content_row}'
                # Force Introduction-Vizo to fit on a single page, column A only
                if tab_name == "Introduction-Vizo":
                    new_ws.print_area = f'A1:A{last_content_row}'
                    new_ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
                    new_ws.page_setup.fitToWidth = 1
                    new_ws.page_setup.fitToHeight = 1
                # Force Executive Summary-Vizo to fit on a single page, column A only
                if tab_name == "Executive Summary-Vizo":
                    new_ws.print_area = f'A1:A{last_content_row}'
                    new_ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
                    new_ws.page_setup.fitToWidth = 1
                    new_ws.page_setup.fitToHeight = 1
        # Move the new tabs to just after 'Display CO-Recov-DQ'
        try:
            target_idx = wb.sheetnames.index('Display CO-Recov-DQ')
            for offset, tab_name in enumerate(["Introduction-Vizo", "Executive Summary-Vizo"], start=1):
                if tab_name in wb.sheetnames:
                    current_idx = wb.sheetnames.index(tab_name)
                    desired_idx = target_idx + offset
                    wb.move_sheet(tab_name, offset=desired_idx - current_idx)
        except ValueError:
            pass

    _sheet_env_ranges(wb, cu, snap, hist)

    safe_cu = cu.replace(' ', '_').replace('/', '-')
    fname = f"{snap}_CECL_Migration_{safe_cu}_Vizo_Model.xlsx"
    return wb, fname


def compose_vizo_supp(client, snap, df, config, grades, hist=None):
    """Build complete Vizo-format supplemental workbook."""
    cu = config['credit_union']
    wb = Workbook()
    _apply_vizo_theme(wb)

    _sheet_cover(wb, cu, snap, supplemental=True)
    _sheet_report_index(wb, cu, snap, supplemental=True)
    _sheet_detail_hist_bal(wb, cu, snap, df, grades, config, hist)
    _sheet_hist_trends_bal(wb, cu, snap, df, grades, config, hist)
    _sheet_detail_co_hist(wb, cu, snap, config, hist)
    _sheet_bal_adjust(wb, cu, snap, df, grades, config, hist)
    _sheet_appendix_supp(wb)

    # Move Historical Trends Balance tab to after Report Index (2)
    trend_idx = wb.sheetnames.index("> Historical Trends Balance")
    idx_idx = wb.sheetnames.index("Report Index (2)")
    wb.move_sheet("> Historical Trends Balance", offset=idx_idx + 1 - trend_idx)

    safe_cu = cu.replace(' ', '_').replace('/', '-')
    fname = f"{snap}_CECL_Supplemental_{safe_cu}_Vizo_Model.xlsx"
    return wb, fname
