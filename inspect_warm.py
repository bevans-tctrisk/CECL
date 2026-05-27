"""Fast BS Data layout dump. Writes to inspect_warm.out so PowerShell
piping doesn't strip the warning lines."""
import glob, os, sys, warnings, shutil, tempfile
warnings.filterwarnings("ignore")
import openpyxl

OUT = open("inspect_warm.out", "w", encoding="utf-8")
def p(*a):
    print(*a, file=OUT)

paths = []
paths += glob.glob(r"Reports\_warm_baselines\*.xlsx")
paths += glob.glob(r"Reports\*WARM*.xlsx")
paths += glob.glob(os.path.join(os.environ.get("TEMP",""), "cecl_ui_uploads", "**", "*WARM*.xlsx"), recursive=True)
if not paths:
    p("No WARM workbook found."); OUT.close(); sys.exit(1)
paths.sort(key=lambda x: os.path.getmtime(x), reverse=True)
src = paths[0]
p(f"OPENED_FILE: {src}")

# Copy to local temp first (network drive is slow)
local = os.path.join(tempfile.gettempdir(), "_inspect_warm_copy.xlsx")
shutil.copy2(src, local)
p(f"LOCAL_COPY: {local}")

wb = openpyxl.load_workbook(local, data_only=True, read_only=True)
p(f"SHEETS: {wb.sheetnames}")
if "BS Data" not in wb.sheetnames:
    OUT.close(); sys.exit(0)
ws = wb["BS Data"]

rows = []
for row in ws.iter_rows(min_row=1, max_row=80, max_col=25, values_only=True):
    rows.append(row)

def s(v):
    if v is None: return ""
    t = str(v).strip().replace("\n", " ")
    return t[:32]

po = lp = None
for r, row in enumerate(rows):
    for c, v in enumerate(row):
        if isinstance(v, str):
            sv = v.strip().lower()
            if sv == "pool order" and po is None: po = (r, c)
            if sv == "loan pools" and lp is None: lp = (r, c)

p(f"\nPool Order at row={po[0]+1 if po else None}, col={po[1]+1 if po else None}")
p(f"Loan Pools at row={lp[0]+1 if lp else None}, col={lp[1]+1 if lp else None}")

p("\n=== First 30 rows x 20 cols ===")
for r in range(min(30, len(rows))):
    p(f"R{r+1:02d}: " + " | ".join(s(v) for v in rows[r][:20]))

if po:
    pr = po[0]
    p(f"\n=== Rows {max(1,pr-10)}..{pr+15} (around Pool Order) ===")
    for r in range(max(0, pr-10), min(len(rows), pr+16)):
        p(f"R{r+1:02d}: " + " | ".join(s(v) for v in rows[r][:25]))

OUT.close()
print("DONE - see inspect_warm.out")
