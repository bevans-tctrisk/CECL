"""
create_sample_cu.py  —  Generate a complete synthetic "Sample Credit Union"
for marketing / demo purposes.

Steps:
  1. Create client_configs/sample.yaml
  2. Create Sample CU/ data directory with quarterly folders
  3. Generate synthetic loan-level data → insert into DB
  4. Create charge-off / recovery Excel files in quarterly folders
  5. Create monthly balance Excel file
  6. Run generate_report to create TCT (→ WARM working file)
  7. Patch the WARM file with BS CO DQ Data Enter + Display CO-Recov -DQ tabs
  8. Run generate_report to create Vizo + Vizo Supp

Usage:
    python create_sample_cu.py
"""

import os, sys, random, math, shutil
from datetime import datetime, date, timedelta
import numpy as np
import pandas as pd
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from sqlalchemy import create_engine, text
from dotenv import load_dotenv

load_dotenv()
BASE = os.path.dirname(os.path.abspath(__file__))
CFG_DIR = os.path.join(BASE, 'client_configs')
RPT_DIR = os.path.join(BASE, 'Reports')
DATA_DIR = os.path.join(BASE, 'Sample CU')

CU_NAME = "Sample Credit Union"
SNAPSHOT = "2025-12-31"
SNAP_PREFIX = "2025-12"

np.random.seed(42)
random.seed(42)

# ── Impaired Loan Provision Amounts (by category) ────────────────
IMPAIRED_ITEMS = {
    'Delinquent Loans': 12950.00,
    'Known Losses': 40900.00,
    'Bankruptcy': 15200.00,
    'Repossessions': 11300.00,
    'Foreclosed Real Estate': 0,
}

# ── Pool Definitions ─────────────────────────────────────────────
POOLS = [
    {
        'name': 'New Auto',
        'acl_months': 36,
        'risk_rated': True,
        'target_balance': 25_000_000,
        'avg_loan': 28_000,
        'fico_mean': 700,
        'fico_std': 70,
        'drift_mean': -1,
        'drift_std': 70,
        'ncc_pct': 0.04,
        'co_pct': 0.048,
        'rc_pct': 0.008,
        'dq_pct_by_year': {2019: 0.030, 2020: 0.035, 2021: 0.028, 2022: 0.025, 2023: 0.020, 2024: 0.025, 2025: 0.045},
        'codes': ['NW', 'NV'],
    },
    {
        'name': 'Used Auto',
        'acl_months': 36,
        'risk_rated': True,
        'target_balance': 30_000_000,
        'avg_loan': 18_000,
        'fico_mean': 680,
        'fico_std': 75,
        'drift_mean': -3,
        'drift_std': 70,
        'ncc_pct': 0.035,
        'co_pct': 0.042,
        'rc_pct': 0.007,
        'dq_pct_by_year': {2019: 0.038, 2020: 0.042, 2021: 0.035, 2022: 0.030, 2023: 0.024, 2024: 0.030, 2025: 0.042},
        'codes': ['UA', 'UV'],
    },
    {
        'name': 'Indirect New Auto',
        'acl_months': 36,
        'risk_rated': True,
        'target_balance': 18_000_000,
        'avg_loan': 30_000,
        'fico_mean': 690,
        'fico_std': 70,
        'drift_mean': -2,
        'drift_std': 70,
        'ncc_pct': 0.06,
        'co_pct': 0.072,
        'rc_pct': 0.012,
        'dq_pct_by_year': {2019: 0.042, 2020: 0.048, 2021: 0.038, 2022: 0.035, 2023: 0.030, 2024: 0.033, 2025: 0.039},
        'codes': ['IN'],
    },
    {
        'name': 'Indirect Used Auto',
        'acl_months': 36,
        'risk_rated': True,
        'target_balance': 15_000_000,
        'avg_loan': 17_000,
        'fico_mean': 675,
        'fico_std': 75,
        'drift_mean': -3,
        'drift_std': 70,
        'ncc_pct': 0.045,
        'co_pct': 0.052,
        'rc_pct': 0.007,
        'dq_pct_by_year': {2019: 0.040, 2020: 0.045, 2021: 0.036, 2022: 0.033, 2023: 0.035, 2024: 0.032, 2025: 0.026},
        'codes': ['IU'],
    },
    {
        'name': 'Pay Day Loans',
        'acl_months': 12,
        'risk_rated': True,
        'target_balance': 3_000_000,
        'avg_loan': 1_200,
        'fico_mean': 620,
        'fico_std': 80,
        'drift_mean': -23,
        'drift_std': 70,
        'ncc_pct': 0.05,
        'co_pct': 0.065,
        'rc_pct': 0.015,
        'dq_pct_by_year': {2019: 0.058, 2020: 0.062, 2021: 0.055, 2022: 0.050, 2023: 0.045, 2024: 0.042, 2025: 0.038},
        'codes': ['PD', 'PY'],
    },
    {
        'name': 'Personal Loans',
        'acl_months': 36,
        'risk_rated': True,
        'target_balance': 12_000_000,
        'avg_loan': 7_000,
        'fico_mean': 695,
        'fico_std': 70,
        'drift_mean': -8,
        'drift_std': 70,
        'ncc_pct': 0.007,
        'co_pct': 0.010,
        'rc_pct': 0.003,
        'dq_pct_by_year': {2019: 0.024, 2020: 0.028, 2021: 0.022, 2022: 0.020, 2023: 0.026, 2024: 0.022, 2025: 0.009},
        'codes': ['PL', 'SL'],
    },
    {
        'name': 'Credit Card',
        'acl_months': 36,
        'risk_rated': True,
        'target_balance': 15_000_000,
        'avg_loan': 4_500,
        'fico_mean': 720,
        'fico_std': 65,
        'drift_mean': -18,
        'drift_std': 70,
        'ncc_pct': 0.008,
        'co_pct': 0.012,
        'rc_pct': 0.004,
        'dq_pct_by_year': {2019: 0.025, 2020: 0.030, 2021: 0.022, 2022: 0.020, 2023: 0.035, 2024: 0.030, 2025: 0.010},
        'codes': ['CC', 'VC'],
    },
    {
        'name': '1st Mortgage',
        'acl_months': 84,
        'risk_rated': True,
        'target_balance': 55_000_000,
        'avg_loan': 185_000,
        'fico_mean': 745,
        'fico_std': 55,
        'drift_mean': -27,
        'drift_std': 70,
        'ncc_pct': 0.003,
        'co_pct': 0.005,
        'rc_pct': 0.002,
        'dq_pct_by_year': {2019: 0.016, 2020: 0.020, 2021: 0.015, 2022: 0.013, 2023: 0.011, 2024: 0.009, 2025: 0.005},
        'codes': ['1M', 'HA'],
    },
    {
        'name': 'HELOC',
        'acl_months': 84,
        'risk_rated': True,
        'target_balance': 21_000_000,
        'avg_loan': 42_000,
        'fico_mean': 735,
        'fico_std': 60,
        'drift_mean': -5,
        'drift_std': 70,
        'ncc_pct': 0.005,
        'co_pct': 0.008,
        'rc_pct': 0.003,
        'dq_pct_by_year': {2019: 0.012, 2020: 0.015, 2021: 0.011, 2022: 0.009, 2023: 0.008, 2024: 0.007, 2025: 0.019},
        'codes': ['HE', 'HL'],
    },
    {
        'name': 'Participation Loans',
        'acl_months': 36,
        'risk_rated': False,
        'target_balance': 6_000_000,
        'avg_loan': 250_000,
        'fico_mean': 730,
        'fico_std': 50,
        'drift_mean': 0,
        'drift_std': 70,
        'ncc_pct': 0.0,
        'co_pct': 0.003,
        'rc_pct': 0.003,
        'dq_pct_by_year': {2019: 0.012, 2020: 0.014, 2021: 0.011, 2022: 0.009, 2023: 0.008, 2024: 0.009, 2025: 0.013},
        'codes': ['LP'],
    },
]

CREDIT_GRADES = [
    {'label': 'A+', 'min_score': 720, 'max_score': 900, 'reserve_rate': 0.0011},
    {'label': 'A',  'min_score': 660, 'max_score': 719, 'reserve_rate': 0.0025},
    {'label': 'B',  'min_score': 620, 'max_score': 659, 'reserve_rate': 0.0050},
    {'label': 'C',  'min_score': 580, 'max_score': 619, 'reserve_rate': 0.0116},
    {'label': 'D',  'min_score': 520, 'max_score': 579, 'reserve_rate': 0.0250},
    {'label': 'E',  'min_score': 0,   'max_score': 519, 'reserve_rate': 0.0500},
]

ECONOMIC_DATA = {
    'state': 'New York',
    'county': 'Suffolk',
    'unemployment_rate': 0.038,
    'foreclosures': 5200,
    'bankruptcies': 4100,
    'population': 4500000,
}

# Quarterly folders (7 years of history to cover 84-month ACL) — must be YYYY-MM format
QUARTERS = [
    '2019-03', '2019-06', '2019-09', '2019-12',
    '2020-03', '2020-06', '2020-09', '2020-12',
    '2021-03', '2021-06', '2021-09', '2021-12',
    '2022-03', '2022-06', '2022-09', '2022-12',
    '2023-03', '2023-06', '2023-09', '2023-12',
    '2024-03', '2024-06', '2024-09', '2024-12',
    '2025-03', '2025-06', '2025-09', '2025-12',
]


# ── 1. Create Config YAML ───────────────────────────────────────
def create_config():
    pool_map = {}
    for p in POOLS:
        for code in p['codes']:
            pool_map[code] = p['name']

    warm_months = {p['name']: p['acl_months'] for p in POOLS}

    config = {
        'credit_union': CU_NAME,
        'data_directory': 'Sample CU',
        'file_pattern': r'loans.*\.(xlsx|xls|csv)$',
        'date_pattern': r'(\d{2})(\d{2})(\d{2})',
        'date_format': 'MMDDYY',
        'has_header': False,
        'account_suffix_length': 3,
        'column_mappings': {
            'member_number': 1,
            'current_balance': 2,
            'original_fico_score': 3,
            'loan_pool_code': 4,
        },
        'credit_pull': {
            'file_pattern': None,
            'fallback_report_pattern': r'CECL-Migration.*\.xlsx$',
            'fallback_report_folder': 'Sample CU',
            'fallback_sheet_pattern': 'Credit Pull',
            'fallback_member_col': 0,
            'fallback_score_col': 1,
        },
        'balance_format': {
            'remove_chars': ['$', ','],
            'accounting_negatives': True,
        },
        'pool_code_split': None,
        'pool_map': pool_map,
        'default_pool': 'Other/Uncategorized',
        'credit_grades': CREDIT_GRADES,
        'no_score_label': 'Not Reported',
        'warm_months': warm_months,
        'not_risk_rated': [p['name'] for p in POOLS if not p['risk_rated']],
        'pool_order': [p['name'] for p in POOLS],
        'reports': {
            'tct': True,
            'vizo': True,
            'vizo_supp': True,
        },
        'economic_data': ECONOMIC_DATA,
        'mgmt_adj': {
            'ltv_baseline': 0.9,
            'probability_factor': 0.35,
        },
        'mgmt_adj_by_pool': {
            'New Auto': 0.005,
        },
        'impaired_items': IMPAIRED_ITEMS,
        'acl_balance': 1955326.84,
    }

    path = os.path.join(CFG_DIR, 'sample.yaml')
    with open(path, 'w', encoding='utf-8') as f:
        f.write('# Sample Credit Union - Synthetic Demo Data\n')
        f.write('# Auto-generated by create_sample_cu.py\n\n')
        yaml.dump(config, f, default_flow_style=False, sort_keys=False, allow_unicode=True)
    print(f"  Created {path}")
    return config


# ── 2. Create Folder Structure ───────────────────────────────────
def create_folders():
    os.makedirs(DATA_DIR, exist_ok=True)
    for q in QUARTERS:
        os.makedirs(os.path.join(DATA_DIR, q), exist_ok=True)
    print(f"  Created folder structure: Sample CU/ with {len(QUARTERS)} quarter folders")


# ── 3. Generate Loan Data and Insert to DB ───────────────────────
def generate_loan_data():
    """Create synthetic loan-level records. Returns DataFrame."""
    records = []
    member_counter = 100000

    for pool_idx, p in enumerate(POOLS):
        # Per-pool RNG so changes to one pool don't shift others
        pool_rng = np.random.RandomState(42 + pool_idx)
        pool_py_rng = random.Random(42 + pool_idx)

        n_loans = max(5, round(p['target_balance'] / p['avg_loan']))
        # Generate balances with lognormal distribution around avg
        raw = pool_rng.lognormal(
            mean=np.log(p['avg_loan']),
            sigma=0.4,
            size=n_loans
        )
        # Scale to hit target balance
        scale = p['target_balance'] / raw.sum()
        balances = raw * scale
        balances = np.round(balances, 2)
        balances = np.clip(balances, 100, p['avg_loan'] * 10)  # sanity bounds

        for i in range(n_loans):
            member_counter += 1
            member_num = f"{member_counter}001"  # 3-digit suffix

            orig_fico = int(np.clip(
                pool_rng.normal(p['fico_mean'], p['fico_std']),
                300, 850
            ))

            # Migration: current FICO drifts from original
            drift = pool_rng.normal(p.get('drift_mean', 5), p.get('drift_std', 70))
            cur_fico = int(np.clip(orig_fico + drift, 300, 850))

            # 1-5% of loans per pool start as "Not Reported" (orig FICO = 0)
            nr_pct = 0.01 + (sum(ord(c) for c in p['name']) % 400) / 10000.0  # 1% – 5%
            if pool_py_rng.random() < nr_pct:
                orig_fico = 0
                # 80-90% of originally NR loans migrate to a real grade
                migrate_pct = 0.80 + (sum(ord(c) for c in p['name'] + '_nr') % 100) / 1000.0
                if pool_py_rng.random() < migrate_pct:
                    # Assign a current FICO (migrated out of NR)
                    cur_fico = int(np.clip(
                        pool_rng.normal(p['fico_mean'], p['fico_std']), 300, 850))
                else:
                    cur_fico = 0  # still Not Reported

            code = pool_py_rng.choice(p['codes'])
            records.append({
                'credit_union': CU_NAME,
                'snapshot_date': SNAPSHOT,
                'member_number': member_num,
                'current_balance': float(balances[i]),
                'current_fico_score': cur_fico,
                'original_fico_score': orig_fico,
                'loan_pool': p['name'],
            })

    df = pd.DataFrame(records)
    total = df['current_balance'].sum()
    print(f"  Generated {len(df)} loans, total balance ${total:,.2f}")
    for p_name, grp in df.groupby('loan_pool'):
        print(f"    {p_name:25s}: {len(grp):5d} loans, ${grp['current_balance'].sum():>14,.2f}")

    return df


def insert_loan_data(df):
    """Insert loan DataFrame into the DB."""
    db_url = os.getenv('DATABASE_URL')
    if not db_url:
        raise RuntimeError("DATABASE_URL not set in .env")
    engine = create_engine(db_url)

    with engine.begin() as conn:
        # Delete any existing Sample CU data
        conn.execute(text("DELETE FROM monthly_loan_data WHERE credit_union = :cu"),
                     {"cu": CU_NAME})
    df.to_sql('monthly_loan_data', engine, if_exists='append', index=False)
    print(f"  Inserted {len(df)} records into monthly_loan_data")


# ── 4. Create Charge-Off / Recovery Files ────────────────────────
def create_co_recovery_files():
    """Create chargeoff and recovery Excel files in quarterly folders."""

    # Map quarters to year
    q_to_year = {}
    for q in QUARTERS:
        q_to_year[q] = int(q[:4])

    # Build annual CO/recovery targets
    # Distribute evenly across quarters within each year
    for q in QUARTERS:
        yr = q_to_year[q]
        # Count quarters in this year
        qs_in_year = [qq for qq in QUARTERS if q_to_year[qq] == yr]
        n_qs = len(qs_in_year)

        # Determine month for dates in this quarter
        q_month = int(q[5:7])

        co_rows = []
        rc_rows = []
        acct_counter = 50000

        for p in POOLS:
            annual_co = p['target_balance'] * p['co_pct']
            annual_rc = p['target_balance'] * p['rc_pct']
            qtr_co = annual_co / n_qs
            qtr_rc = annual_rc / n_qs

            # Add some year-over-year variation (±15%)
            year_factor = 1.0 + (yr - 2024) * 0.05 + random.uniform(-0.10, 0.10)
            qtr_co *= year_factor
            qtr_rc *= year_factor

            code = p['codes'][0]

            # Split into individual transactions
            n_co_txns = max(1, int(qtr_co / (p['avg_loan'] * 0.3)))
            n_co_txns = min(n_co_txns, 50)
            if qtr_co > 0:
                co_amounts = np.random.dirichlet(np.ones(n_co_txns)) * qtr_co
                for amt in co_amounts:
                    acct_counter += 1
                    day = random.randint(1, 28)
                    dt = datetime(yr, q_month, day)
                    co_rows.append([acct_counter, '001', code, round(amt, 2), dt])

            n_rc_txns = max(1, int(qtr_rc / (p['avg_loan'] * 0.2)))
            n_rc_txns = min(n_rc_txns, 50)
            if qtr_rc > 0:
                rc_amounts = np.random.dirichlet(np.ones(n_rc_txns)) * qtr_rc
                for amt in rc_amounts:
                    acct_counter += 1
                    day = random.randint(1, 28)
                    dt = datetime(yr, q_month, day)
                    rc_rows.append([acct_counter, '001', code, round(amt, 2), dt])

        # Write chargeoff file
        q_folder = os.path.join(DATA_DIR, q)
        if co_rows:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Chargeoffs'
            ws.append(['Account', 'Suffix', 'Security Code', 'Amount', 'Date'])
            for row in co_rows:
                ws.append(row)
            co_path = os.path.join(q_folder, f'chargeoff_{q}.xlsx')
            wb.save(co_path)

        if rc_rows:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Recoveries'
            ws.append(['Account', 'Suffix', 'Security Code', 'Recovery Amount', 'Date'])
            for row in rc_rows:
                ws.append(row)
            rc_path = os.path.join(q_folder, f'recovery_{q}.xlsx')
            wb.save(rc_path)

    print(f"  Created charge-off and recovery files in {len(QUARTERS)} quarter folders")


# ── 5. Create Monthly Balance File ──────────────────────────────
def create_monthly_balance_file():
    """Create a monthly balance Excel file with pools × dates grid."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Loan Balances by Pool'

    # Generate monthly dates from Jan 2019 to Dec 2025 (84 months for full LoL coverage)
    dates = []
    for yr in range(2019, 2026):
        for mo in range(1, 13):
            # Last day of month
            if mo == 12:
                d = datetime(yr, 12, 31)
            else:
                d = datetime(yr, mo + 1, 1) - timedelta(days=1)
            dates.append(d)

    # Header row: Pool | date1 | date2 | ...
    ws.cell(row=1, column=1, value='Pool')
    for i, d in enumerate(dates):
        ws.cell(row=1, column=2 + i, value=d)

    # Pool rows with trending balances
    for ri, p in enumerate(POOLS):
        ws.cell(row=2 + ri, column=1, value=p['name'])
        base = p['target_balance']
        for i, d in enumerate(dates):
            # Slight growth trend + seasonal noise
            months_from_start = (d.year - 2019) * 12 + d.month - 1
            growth = 1 + months_from_start * 0.002  # ~2.4%/year growth
            seasonal = 1 + 0.02 * math.sin(months_from_start * math.pi / 6)
            noise = random.uniform(0.97, 1.03)
            bal = base * growth * seasonal * noise
            ws.cell(row=2 + ri, column=2 + i, value=round(bal, 2))

    path = os.path.join(DATA_DIR, 'monthly_loan_balances.xlsx')
    wb.save(path)
    print(f"  Created monthly balance file with {len(POOLS)} pools × {len(dates)} months")


# ── 6. Patch WARM File with Additional Tabs ──────────────────────
def patch_warm_file():
    """Find the WARM file in the data directory and add missing tabs."""
    # Find the WARM file in the data directory (where we copied it)
    warm_path = None
    for root, dirs, files in os.walk(DATA_DIR):
        for f in files:
            if SNAP_PREFIX in f and 'CECL-Migration-WARM' in f and f.endswith('.xlsx') and not f.startswith('~$'):
                warm_path = os.path.join(root, f)
                break
        if warm_path:
            break

    if not warm_path:
        print("  ERROR: WARM file not found in data directory")
        return False

    print(f"  Patching WARM: {os.path.basename(warm_path)}")
    wb = load_workbook(warm_path)

    # ── Add "Impaired Loans" tab ──
    ws_imp = wb.create_sheet("Impaired Loans")
    cu_name = CU_NAME
    ws_imp.cell(row=1, column=1, value=cu_name).font = Font(bold=True, size=12)
    ws_imp.cell(row=2, column=1, value="Impaired Loans").font = Font(bold=True, size=12)
    ws_imp.cell(row=3, column=1, value="Report for Period Ending ")
    ws_imp.cell(row=3, column=2, value=datetime(2025, 12, 31))

    # Summary section headers (row 4)
    ws_imp.cell(row=4, column=1, value="Impairment Type")
    ws_imp.cell(row=4, column=2, value="Provision Percentage (Collateral Value Calculation Only)")
    ws_imp.cell(row=4, column=12, value="Impairment Type")
    ws_imp.cell(row=4, column=14, value="Sum of Loss Given Default (LGD)")
    ws_imp.cell(row=4, column=16, value="Sum of Provision Amount")
    ws_imp.cell(row=4, column=17, value="Sum of Balance Removed from Homogeneous Pools")

    # Impaired loan detail items
    impaired_detail = [
        # (type, prov_pct, member, suffix, loan_type, balance, days_delq, bal_other, collateral, allowance, notes, pool, grade)
        ('Delinquent Loans', 0.35, '100234', '01', 'Used Auto',   32500.00, 62,  0, 18000, 0, '', 'Used Auto', 'D'),
        ('Delinquent Loans', 0.35, '100891', '01', '1st Mortgage', 48000.00, 45,  0, 32000, 0, '', '1st Mortgage', 'C'),
        ('Delinquent Loans', 0.35, '101456', '02', 'HELOC',        18500.00, 91,  0, 12000, 0, '', 'HELOC', 'D'),
        ('Known Losses',     1.00, '102789', '01', 'Personal',     18200.00, 180, 0,     0, 0, 'Charged off Q3', 'Personal Loans', 'E'),
        ('Known Losses',     1.00, '103012', '01', 'Credit Card',  12400.00, 210, 0,     0, 0, 'Charged off Q4', 'Credit Card', 'E'),
        ('Known Losses',     1.00, '103567', '02', 'New Auto',     14800.00, 150, 0,  4500, 0, '', 'New Auto', 'E'),
        ('Bankruptcy',       1.00, '104200', '01', 'Personal',      8200.00, 120, 0,     0, 0, 'Ch 7 filed', 'Personal Loans', 'E'),
        ('Bankruptcy',       1.00, '104789', '01', '1st Mortgage', 22000.00, 90,  0, 15000, 0, 'Ch 13 filed', '1st Mortgage', 'D'),
        ('Repossessions',    1.00, '105100', '01', 'New Auto',      9800.00, 95,  0,  3200, 0, 'Repo in progress', 'New Auto', 'E'),
        ('Repossessions',    1.00, '105432', '01', 'Used Auto',     7500.00, 110, 0,  2800, 0, 'Repo complete', 'Used Auto', 'E'),
    ]

    # Compute LGD and provision for each loan
    prov_pct_map = {'Delinquent Loans': 0.35, 'Known Losses': 1.00,
                    'Bankruptcy': 1.00, 'Repossessions': 1.00}
    detail_rows = []
    for item in impaired_detail:
        (imp_type, prov_pct, member, suffix, loan_type, balance, days_delq,
         bal_other, collateral, allowance, notes, pool, grade) = item
        total_loans = balance  # single loan per row
        if collateral > 0:
            ltv = (balance + bal_other) / collateral if collateral else 0
            lgd = max(0, balance + bal_other - collateral)
        else:
            lgd = balance  # unsecured → full balance is at risk
        provision = round(lgd * prov_pct, 2)
        bal_removed = balance
        detail_rows.append({
            'type': imp_type, 'prov_pct': prov_pct, 'member': member,
            'suffix': suffix, 'loan_type': loan_type, 'balance': balance,
            'days_delq': days_delq, 'bal_other': bal_other,
            'collateral': collateral, 'allowance': allowance, 'notes': notes,
            'total_loans': total_loans, 'ltv': ltv if collateral > 0 else 0,
            'lgd': lgd, 'provision': provision, 'bal_removed': bal_removed,
            'pool': pool, 'grade': grade,
        })

    # Build summary by category
    cat_order = ['Delinquent Loans', 'Known Losses', 'Bankruptcy',
                 'Repossessions', 'Foreclosed Real Estate']
    cat_lgd = {}
    cat_prov = {}
    cat_bal_removed = {}
    for dr in detail_rows:
        t = dr['type']
        cat_lgd[t] = cat_lgd.get(t, 0) + dr['lgd']
        cat_prov[t] = cat_prov.get(t, 0) + dr['provision']
        cat_bal_removed[t] = cat_bal_removed.get(t, 0) + dr['bal_removed']

    # Write summary rows (rows 5-9+)
    for i, cat in enumerate(cat_order):
        r = 5 + i
        ws_imp.cell(row=r, column=1, value=cat)
        ws_imp.cell(row=r, column=2, value=prov_pct_map.get(cat, 0))
        ws_imp.cell(row=r, column=12, value=cat)
        ws_imp.cell(row=r, column=14, value=round(cat_lgd.get(cat, 0), 2))
        ws_imp.cell(row=r, column=16, value=round(cat_prov.get(cat, 0), 2))
        ws_imp.cell(row=r, column=17, value=round(cat_bal_removed.get(cat, 0), 2))

    # HIDE rows (10-19)
    for i in range(10, 20):
        ws_imp.cell(row=i, column=1, value='HIDE')
        ws_imp.cell(row=i, column=2, value=0)
        ws_imp.cell(row=i, column=12, value='HIDE')
        ws_imp.cell(row=i, column=14, value=0)
        ws_imp.cell(row=i, column=16, value=0)
        ws_imp.cell(row=i, column=17, value=0)

    # Total row (row 20)
    total_lgd = sum(cat_lgd.values())
    total_prov = sum(cat_prov.values())
    total_bal_rem = sum(cat_bal_removed.values())
    ws_imp.cell(row=20, column=12, value='Total')
    ws_imp.cell(row=20, column=14, value=round(total_lgd, 2))
    ws_imp.cell(row=20, column=16, value=round(total_prov, 2))
    ws_imp.cell(row=20, column=17, value=round(total_bal_rem, 2))

    # Detail header rows (32-33) — start after row 30 to stay outside loader range
    ws_imp.cell(row=32, column=1, value='Data Entry')
    for c in range(2, 12):
        ws_imp.cell(row=32, column=c, value='Data Entry')
    for c in range(11, 18):
        ws_imp.cell(row=32, column=c, value='Calculation ')
    ws_imp.cell(row=32, column=18, value='TCT ')
    ws_imp.cell(row=32, column=19, value='TCT')

    detail_headers = ['Impairment Type', 'Member #', 'Loan Suffix', 'Loan Type',
                      'Current Balance', 'Days Delinquent',
                      "Balance at Other Lender (2nd Mortgage and HELOC's)",
                      'Collateral Value (for unsecured loans, leave blank)',
                      f'Allowance Provided by {cu_name} (if Cell is blank, provision will be calculated)',
                      'Notes: ', 'Member #-Suffix', 'Total Loans', 'LTV',
                      'Loss Given Default (LGD)',
                      'Percent of Amount at Risk applied to Provision',
                      'Provision Amount', 'Balance Removed from Homogeneous Pools',
                      'Loan Pool', 'Current Credit Grade']
    for ci, h in enumerate(detail_headers):
        ws_imp.cell(row=33, column=1 + ci, value=h)

    # Detail data rows (34+)
    for di, dr in enumerate(detail_rows):
        r = 34 + di
        ws_imp.cell(row=r, column=1, value=dr['type'])
        ws_imp.cell(row=r, column=2, value=dr['member'])
        ws_imp.cell(row=r, column=3, value=dr['suffix'])
        ws_imp.cell(row=r, column=4, value=dr['loan_type'])
        ws_imp.cell(row=r, column=5, value=dr['balance'])
        ws_imp.cell(row=r, column=6, value=dr['days_delq'])
        ws_imp.cell(row=r, column=7, value=dr['bal_other'])
        ws_imp.cell(row=r, column=8, value=dr['collateral'])
        ws_imp.cell(row=r, column=9, value=dr['allowance'])
        ws_imp.cell(row=r, column=10, value=dr['notes'])
        ws_imp.cell(row=r, column=11, value=f"{dr['member']}-{dr['suffix']}")
        ws_imp.cell(row=r, column=12, value=dr['total_loans'])
        ws_imp.cell(row=r, column=13, value=round(dr['ltv'], 4) if dr['ltv'] else 0)
        ws_imp.cell(row=r, column=14, value=round(dr['lgd'], 2))
        ws_imp.cell(row=r, column=15, value=dr['prov_pct'])
        ws_imp.cell(row=r, column=16, value=dr['provision'])
        ws_imp.cell(row=r, column=17, value=dr['bal_removed'])
        ws_imp.cell(row=r, column=18, value=dr['pool'])
        ws_imp.cell(row=r, column=19, value=dr['grade'])

    # ── Add "BS CO DQ Data Enter" tab ──
    ws = wb.create_sheet("BS CO DQ Data Enter")
    # Row 1-3: headers
    ws.cell(row=1, column=1, value="Balance Sheet CO DQ Data Entry").font = Font(bold=True, size=12)
    ws.cell(row=4, column=1, value="Loan Pool").font = Font(bold=True)
    ws.cell(row=4, column=2, value="Risk Rated").font = Font(bold=True)
    ws.cell(row=4, column=7, value="ACL Months").font = Font(bold=True)

    # Pool rows starting at row 5 (0-indexed = 4)
    for i, p in enumerate(POOLS):
        r = 5 + i
        ws.cell(row=r, column=1, value=p['name'])
        ws.cell(row=r, column=2, value='Yes' if p['risk_rated'] else 'No')
        ws.cell(row=r, column=7, value=p['acl_months'])

    # Exclude sentinel
    ws.cell(row=5 + len(POOLS), column=1, value='Exclude')

    # Economic data at row 7 (0-indexed = 6), cols L-Q (12-17 in 1-indexed)
    ws.cell(row=6, column=12, value="State").font = Font(bold=True)
    ws.cell(row=6, column=13, value="County").font = Font(bold=True)
    ws.cell(row=6, column=14, value="Unemp%").font = Font(bold=True)
    ws.cell(row=6, column=15, value="Foreclosures").font = Font(bold=True)
    ws.cell(row=6, column=16, value="Bankruptcies").font = Font(bold=True)
    ws.cell(row=6, column=17, value="Population").font = Font(bold=True)
    ws.cell(row=7, column=12, value=ECONOMIC_DATA['state'])
    ws.cell(row=7, column=13, value=ECONOMIC_DATA['county'])
    ws.cell(row=7, column=14, value=ECONOMIC_DATA['unemployment_rate'])
    ws.cell(row=7, column=15, value=ECONOMIC_DATA['foreclosures'])
    ws.cell(row=7, column=16, value=ECONOMIC_DATA['bankruptcies'])
    ws.cell(row=7, column=17, value=ECONOMIC_DATA['population'])

    # ── Add "HIst Bal Data" tab (grade-level monthly balances) ──
    ws_hbd = wb.create_sheet("HIst Bal Data")

    # Generate 84 monthly dates (Jan 2019 – Dec 2025)
    hbd_dates = []
    for yr in range(2019, 2026):
        for mo in range(1, 13):
            if mo == 12:
                d = datetime(yr, 12, 31)
            else:
                d = datetime(yr, mo + 1, 1) - timedelta(days=1)
            hbd_dates.append(d)

    # Row 1: year numbers in cols C+ (matching real WARM format)
    for ci, d in enumerate(hbd_dates):
        ws_hbd.cell(row=1, column=3 + ci, value=d.year)
    # Row 2-4: CU name, title, period
    ws_hbd.cell(row=2, column=1, value=CU_NAME)
    ws_hbd.cell(row=3, column=1, value="Loss Factor Calculation Data")
    ws_hbd.cell(row=4, column=1, value=f"For Period Ending {SNAPSHOT}")
    # Row 5: dates in cols C+
    for ci, d in enumerate(hbd_dates):
        ws_hbd.cell(row=5, column=3 + ci, value=d)

    grade_labels = [g['label'] for g in CREDIT_GRADES] + ['Not Reported']

    # Distribute pool balance across grades based on FICO distribution
    def _grade_fractions(fico_mean, fico_std, pool_name=None):
        """Estimate what fraction of a pool falls into each grade based on FICO dist.
        Includes 1-5% Not Reported (per-pool seeded)."""
        def _norm_cdf(x, mu, sigma):
            """Normal CDF via error function (no scipy needed)."""
            return 0.5 * (1 + math.erf((x - mu) / (sigma * math.sqrt(2))))

        fracs = {}
        for g in CREDIT_GRADES:
            lo, hi = g['min_score'], g['max_score']
            frac = _norm_cdf(hi + 0.5, fico_mean, fico_std) - _norm_cdf(lo - 0.5, fico_mean, fico_std)
            fracs[g['label']] = max(frac, 0.0)
        # Not Reported: 1-5% per pool (seeded by pool name)
        if pool_name:
            nr_pct = 0.01 + (hash(pool_name) % 400) / 10000.0  # 1% – 5%
        else:
            nr_pct = 0.03  # default 3%
        fracs['Not Reported'] = nr_pct
        # Scale other grades down so everything sums to 1.0
        scored_total = sum(v for k, v in fracs.items() if k != 'Not Reported')
        if scored_total > 0:
            scale = (1.0 - nr_pct) / scored_total
            for k in fracs:
                if k != 'Not Reported':
                    fracs[k] *= scale
        return fracs

    hbd_row = 6  # start pool blocks at row 6
    for p in POOLS:
        base = p['target_balance']
        gfracs = _grade_fractions(p['fico_mean'], p['fico_std'], p['name'])

        # Pool name row
        ws_hbd.cell(row=hbd_row, column=1, value=p['name'])
        hbd_row += 1
        # "Current Grade" header + dates
        ws_hbd.cell(row=hbd_row, column=1, value="Current Grade")
        for ci, d in enumerate(hbd_dates):
            ws_hbd.cell(row=hbd_row, column=3 + ci, value=d)
        hbd_row += 1

        # Grade rows
        for glabel in grade_labels:
            ws_hbd.cell(row=hbd_row, column=1, value=glabel)
            gfrac = gfracs.get(glabel, 0.0)
            for ci, d in enumerate(hbd_dates):
                months_from_start = (d.year - 2019) * 12 + d.month - 1
                growth = 1 + months_from_start * 0.002
                seasonal = 1 + 0.02 * math.sin(months_from_start * math.pi / 6)
                bal = base * gfrac * growth * seasonal
                # Add slight grade-level noise (deterministic per pool+grade+month)
                noise_seed = hash((p['name'], glabel, ci)) % 10000
                noise = 0.97 + (noise_seed / 10000) * 0.06  # 0.97–1.03
                ws_hbd.cell(row=hbd_row, column=3 + ci, value=round(bal * noise, 2))
            hbd_row += 1

        # Total row
        ws_hbd.cell(row=hbd_row, column=1, value="Total")
        for ci, d in enumerate(hbd_dates):
            months_from_start = (d.year - 2019) * 12 + d.month - 1
            growth = 1 + months_from_start * 0.002
            seasonal = 1 + 0.02 * math.sin(months_from_start * math.pi / 6)
            ws_hbd.cell(row=hbd_row, column=3 + ci, value=round(base * growth * seasonal, 2))
        hbd_row += 1

        # Blank separator row
        hbd_row += 1

    print(f"  HIst Bal Data: {len(POOLS)} pools, {len(hbd_dates)} months, {len(grade_labels)} grades")

    # ── Add "Display CO-Recov -DQ" tab ──
    ws2 = wb.create_sheet("Display CO-Recov -DQ")
    years = list(range(2019, 2026))  # 2019-2025: 7 years for full LoL coverage
    last_yr = years[-1]

    # Row 1: title
    ws2.cell(row=1, column=1, value="Display CO-Recov -DQ").font = Font(bold=True, size=12)

    # Row 3 (0-indexed row 2): year headers in cols B, C, D, ...
    for ci, yr in enumerate(years):
        ws2.cell(row=3, column=2 + ci, value=yr)

    # Charge Offs section
    r = 5
    ws2.cell(row=r, column=1, value="Charge Offs").font = Font(bold=True)
    r += 1
    for p in POOLS:
        ws2.cell(row=r, column=1, value=p['name'])
        for ci, yr in enumerate(years):
            co_amt = p['target_balance'] * p['co_pct']
            year_factor = 1.0 + (yr - 2024) * 0.05
            ws2.cell(row=r, column=2 + ci, value=round(co_amt * year_factor, 2))
        r += 1

    # Blank row
    r += 1

    # Recoveries section
    ws2.cell(row=r, column=1, value="Recoveries").font = Font(bold=True)
    r += 1
    for p in POOLS:
        ws2.cell(row=r, column=1, value=p['name'])
        for ci, yr in enumerate(years):
            rc_amt = p['target_balance'] * p['rc_pct']
            year_factor = 1.0 + (yr - 2024) * 0.05
            ws2.cell(row=r, column=2 + ci, value=round(rc_amt * year_factor, 2))
        r += 1

    # Blank row
    r += 1

    # Net Loss section
    ws2.cell(row=r, column=1, value="Net Loss").font = Font(bold=True)
    r += 1
    for p in POOLS:
        ws2.cell(row=r, column=1, value=p['name'])
        for ci, yr in enumerate(years):
            net = p['target_balance'] * (p['co_pct'] - p['rc_pct'])
            year_factor = 1.0 + (yr - 2024) * 0.05
            ws2.cell(row=r, column=2 + ci, value=round(net * year_factor, 2))
        r += 1

    # Blank row
    r += 1

    # WARM Months section
    ws2.cell(row=r, column=1, value="WARM Months").font = Font(bold=True)
    r += 1
    for p in POOLS:
        ws2.cell(row=r, column=1, value=p['name'])
        ws2.cell(row=r, column=2, value=p['acl_months'])
        r += 1

    # Blank row
    r += 1

    # DQ % section
    ws2.cell(row=r, column=1, value="DQ %").font = Font(bold=True)
    r += 1
    for p in POOLS:
        ws2.cell(row=r, column=1, value=p['name'])
        for ci, yr in enumerate(years):
            dq_val = p['dq_pct_by_year'].get(yr, 0.02)
            ws2.cell(row=r, column=2 + ci, value=round(dq_val, 6))
        r += 1

    # ── Add "Risk Change Data Entry" tab ──
    ws_rc = wb.create_sheet("Risk Change Data Entry")
    rc_grade_labels = [g['label'] for g in CREDIT_GRADES]
    rc_hidden = [h for h in ['Hide-E', 'Hide-F', 'Hide-G', 'Hide-H', 'Hide-I']
                 if h not in [g['label'] for g in CREDIT_GRADES]]
    # Trim hidden grades so visible + hidden + NR = 11
    n_hidden_needed = max(0, 11 - len(rc_grade_labels) - 1)
    rc_hidden = rc_hidden[:n_hidden_needed]
    rc_all_grades = rc_grade_labels + rc_hidden + ['Not Reported']

    ws_rc.cell(row=3, column=1, value="Dollar")
    ws_rc.cell(row=4, column=14, value="Loan Pool")
    ws_rc.cell(row=4, column=15, value="Pool order")

    rc_row = 5
    for pi, p in enumerate(POOLS):
        gfracs = _grade_fractions(p['fico_mean'], p['fico_std'], p['name'])
        pool_bal = p['target_balance']

        # Pool header row
        ws_rc.cell(row=rc_row, column=1, value=p['name'])
        ws_rc.cell(row=rc_row, column=13, value="CG")
        ws_rc.cell(row=rc_row, column=14, value=p['name'])
        ws_rc.cell(row=rc_row, column=15, value=pi + 1)
        rc_row += 1

        # Column headers
        ws_rc.cell(row=rc_row, column=1, value="Current Grade")
        ws_rc.cell(row=rc_row, column=13, value="Loan Report Balance")
        ws_rc.cell(row=rc_row, column=14, value="% of Loan Balance")
        ws_rc.cell(row=rc_row, column=15, value="Balance Sheet Total")
        ws_rc.cell(row=rc_row, column=16, value="Bal Adjustment")
        ws_rc.cell(row=rc_row, column=17, value="Specific Identification")
        rc_row += 1

        # Grade rows
        pool_total_bal = 0.0
        pool_total_spec = 0.0
        for g in rc_all_grades:
            is_hidden = g.startswith('Hide')
            grade_bal = pool_bal * gfracs.get(g, 0.0) if not is_hidden else 0.0
            grade_pct = gfracs.get(g, 0.0) if not is_hidden else 0.0
            # Specific identification: small % of D and E grade balances
            spec_id = 0.0
            if g in ('D', 'E') and grade_bal > 0:
                spec_pct = 0.03 + (hash((p['name'], g)) % 50) / 1000.0  # 3-8%
                spec_id = round(grade_bal * spec_pct, 2)

            ws_rc.cell(row=rc_row, column=1, value=g)
            ws_rc.cell(row=rc_row, column=13, value=round(grade_bal, 2))
            ws_rc.cell(row=rc_row, column=14, value=round(grade_pct, 6))
            ws_rc.cell(row=rc_row, column=15, value=round(grade_bal, 2))  # bal sheet = loan report (no adj)
            ws_rc.cell(row=rc_row, column=16, value=0)  # no balance adjustment
            ws_rc.cell(row=rc_row, column=17, value=round(spec_id, 2))
            pool_total_bal += grade_bal
            pool_total_spec += spec_id
            rc_row += 1

        # Total row
        ws_rc.cell(row=rc_row, column=1, value="Total")
        ws_rc.cell(row=rc_row, column=13, value=round(pool_total_bal, 2))
        ws_rc.cell(row=rc_row, column=14, value=1)
        ws_rc.cell(row=rc_row, column=15, value=round(pool_total_bal, 2))
        ws_rc.cell(row=rc_row, column=16, value=0)
        ws_rc.cell(row=rc_row, column=17, value=round(pool_total_spec, 2))
        rc_row += 1

        # Blank separator
        rc_row += 1

    print(f"  Risk Change Data Entry: {len(POOLS)} pools, {len(rc_all_grades)} grades")

    # ── Add "DQ Data Entry" and "CO Data Entry" tabs ──
    grade_labels_full = ['A+', 'A', 'B', 'C', 'D', 'E',
                         'Hide-F', 'Hide-G', 'Hide-H', 'Hide-I',
                         'Not Reported']
    active_grades = ['A+', 'A', 'B', 'C', 'D', 'E']

    def _write_migration_block(ws, start_row, pool_name, header_tag, bal_label,
                               pct_label, status_balances, grade_fracs, total_balance):
        """Write one pool migration block (14 rows). Returns next row."""
        r = start_row
        # Row 0: header
        ws.cell(row=r, column=1, value=pool_name)
        ws.cell(row=r, column=2, value='Original Grade')
        ws.cell(row=r, column=13, value=header_tag)
        ws.cell(row=r, column=16, value='Loan Status')
        ws.cell(row=r, column=17, value='total')
        ws.cell(row=r, column=18, value=pct_label)
        r += 1

        # Row 1: column headers + Improved
        ws.cell(row=r, column=1, value='Current Grade')
        for gi, gl in enumerate(grade_labels_full):
            ws.cell(row=r, column=2 + gi, value=gl)
        ws.cell(row=r, column=13, value=bal_label)
        ws.cell(row=r, column=14, value='% of ' + bal_label)
        status_names = ['Improved', 'Deteriorated', 'Unchanged', 'Not Reported']
        ws.cell(row=r, column=16, value=status_names[0])
        ws.cell(row=r, column=17, value=round(status_balances[0], 2))
        ws.cell(row=r, column=18, value=round(status_balances[0] / max(total_balance, 1), 6))
        r += 1

        # Build a synthetic grade migration matrix
        # Transition probabilities: diagonal dominant, some off-diagonal
        n_active = len(active_grades)
        matrix = {}
        col_totals = {gl: 0.0 for gl in grade_labels_full}
        for ci, curr_g in enumerate(active_grades):
            curr_bal = total_balance * grade_fracs.get(curr_g, 0.0)
            row_vals = {}
            for oi, orig_g in enumerate(active_grades):
                if ci == oi:
                    row_vals[orig_g] = curr_bal * 0.65
                elif ci == oi - 1:
                    row_vals[orig_g] = curr_bal * 0.15  # improved from worse
                elif ci == oi + 1:
                    row_vals[orig_g] = curr_bal * 0.12  # deteriorated from better
                elif abs(ci - oi) == 2:
                    row_vals[orig_g] = curr_bal * 0.03
                else:
                    row_vals[orig_g] = curr_bal * 0.01
            # Normalize so row sums to curr_bal
            rsum = sum(row_vals.values())
            if rsum > 0:
                for k in row_vals:
                    row_vals[k] = row_vals[k] / rsum * curr_bal
            for hg in ['Hide-F', 'Hide-G', 'Hide-H', 'Hide-I']:
                row_vals[hg] = 0.0
            # NR-original column: loans that were originally NR but migrated to this grade
            # Distribute proportionally to each grade's share of total balance
            nr_bal = total_balance * grade_fracs.get('Not Reported', 0.0)
            _stay_pct = 0.10 + (hash((pool_name, 'nr_stay')) % 100) / 1000.0
            nr_migrated = nr_bal * (1.0 - _stay_pct)
            row_vals['Not Reported'] = nr_migrated * grade_fracs.get(curr_g, 0.0) / (1.0 - grade_fracs.get('Not Reported', 0.0)) if grade_fracs.get('Not Reported', 0.0) < 1.0 else 0.0
            matrix[curr_g] = row_vals
            for gl in grade_labels_full:
                col_totals[gl] += row_vals.get(gl, 0.0)

        # Add Not Reported row: 80-90% migrate to real grades, 10-20% stay NR
        nr_total = total_balance * grade_fracs.get('Not Reported', 0.0)
        nr_row = {gl: 0.0 for gl in grade_labels_full}
        stay_nr_pct = 0.10 + (hash((pool_name, 'nr_stay')) % 100) / 1000.0  # 10%-20%
        nr_row['Not Reported'] = nr_total * stay_nr_pct
        # Distribute the migrated portion across real grades proportionally
        migrated = nr_total * (1.0 - stay_nr_pct)
        scored_fracs = {g: grade_fracs.get(g, 0.0) for g in active_grades}
        scored_sum = sum(scored_fracs.values())
        if scored_sum > 0:
            for g in active_grades:
                nr_row[g] = migrated * scored_fracs[g] / scored_sum
        matrix['Not Reported'] = nr_row
        for gl in grade_labels_full:
            col_totals[gl] += nr_row.get(gl, 0.0)

        # Hide grades rows
        for hg in ['Hide-F', 'Hide-G', 'Hide-H', 'Hide-I']:
            matrix[hg] = {gl: 0.0 for gl in grade_labels_full}

        # Write grade rows (A+, A, B, C, D, E + statuses on first 3 rows)
        all_row_grades = active_grades + ['Hide-F', 'Hide-G', 'Hide-H', 'Hide-I', 'Not Reported']
        for ri, curr_g in enumerate(all_row_grades):
            ws.cell(row=r, column=1, value=curr_g)
            row_total = 0.0
            for gi, gl in enumerate(grade_labels_full):
                val = round(matrix[curr_g].get(gl, 0.0), 2)
                ws.cell(row=r, column=2 + gi, value=val)
                row_total += val
            ws.cell(row=r, column=13, value=round(row_total, 2))
            # Status data on rows offset 2-4 (A+, A, B rows)
            si = ri + 1  # status index: 1=Deteriorated, 2=Unchanged, 3=Not Reported
            if si < len(status_names):
                ws.cell(row=r, column=16, value=status_names[si])
                ws.cell(row=r, column=17, value=round(status_balances[si], 2))
                ws.cell(row=r, column=18, value=round(status_balances[si] / max(total_balance, 1), 6))
            r += 1

        # Total row
        ws.cell(row=r, column=1, value='Total')
        grand = 0.0
        for gi, gl in enumerate(grade_labels_full):
            val = round(col_totals[gl], 2)
            ws.cell(row=r, column=2 + gi, value=val)
            grand += val
        ws.cell(row=r, column=13, value=round(grand, 2))
        ws.cell(row=r, column=14, value=round(grand / max(total_balance, 1), 6) if total_balance > 0 else 0)
        r += 1

        # Blank separator
        r += 1
        return r

    for tab_name, bal_label, pct_label, amount_key in [
        ('DQ Data Entry', 'DQ Balance', '% of Total DQ', 'dq'),
        ('CO Data Entry', 'CO Amount', '% of Total CO', 'co'),
    ]:
        ws_tab = wb.create_sheet(tab_name)
        row = 1

        # Status split ratios:  [Improved, Deteriorated, Unchanged, Not Reported]
        if amount_key == 'dq':
            # DQ: 60-80% Deteriorated, 10-20% Unchanged, 5-10% Not Reported, <1% Improved
            base_splits = [0.005, 0.72, 0.18, 0.095]
        else:
            pass  # CO splits computed per-pool below

        grand_status = [0.0, 0.0, 0.0, 0.0]
        grand_total = 0.0

        for p in POOLS:
            gfracs = _grade_fractions(p['fico_mean'], p['fico_std'], p['name'])
            if amount_key == 'dq':
                pool_total = p['target_balance'] * p['dq_pct_by_year'].get(2025, 0.02)
            else:
                pool_total = p['target_balance'] * p['co_pct']

            if amount_key == 'dq':
                # Per-pool variation within realistic DQ ranges
                # Deteriorated 60-80%, Unchanged 10-20%, Not Reported 5-10%, Improved <1%
                seed_val = hash((p['name'], 'dq_split')) % 1000
                t = seed_val / 1000.0  # 0.0 – 1.0
                improved = 0.003 + t * 0.005              # 0.3% – 0.8%
                not_reported = 0.05 + (1.0 - t) * 0.04    # 5% – 9%
                unchanged = 0.14 + (1.0 - t) * 0.04       # 14% – 18%
                deteriorated = 1.0 - improved - not_reported - unchanged  # ~73% – 80%
                status_bals = [pool_total * improved,
                               pool_total * deteriorated,
                               pool_total * unchanged,
                               pool_total * not_reported]
            else:
                # Per-pool variation within realistic CO ranges
                # Deteriorated 70-90%, Unchanged 15-30%, Not Reported 5-10%, Improved <0.1%
                seed_val = hash((p['name'], 'co_split')) % 1000
                t = seed_val / 1000.0  # 0.0 – 1.0
                improved = 0.0003 + t * 0.0006           # 0.03% – 0.09%
                not_reported = 0.05 + (1.0 - t) * 0.03   # 5% – 8%
                unchanged = 0.15 + (1.0 - t) * 0.07      # 15% – 22%
                deteriorated = 1.0 - improved - not_reported - unchanged  # ~70% – 80%
                status_bals = [pool_total * improved,
                               pool_total * deteriorated,
                               pool_total * unchanged,
                               pool_total * not_reported]

            for i in range(4):
                grand_status[i] += status_bals[i]
            grand_total += pool_total

            row = _write_migration_block(ws_tab, row, p['name'], 'CG', bal_label,
                                         pct_label, status_bals, gfracs, pool_total)

        # Risk Rated Loans block (all zeros, with numeric grade IDs)
        ws_tab.cell(row=row, column=1, value='Risk Rated Loans')
        ws_tab.cell(row=row, column=2, value='Original Grade')
        ws_tab.cell(row=row, column=13, value='RR')
        ws_tab.cell(row=row, column=16, value='Loan Status')
        ws_tab.cell(row=row, column=17, value='total')
        ws_tab.cell(row=row, column=18, value=pct_label)
        row += 1
        ws_tab.cell(row=row, column=1, value='Current Grade')
        for gi in range(11):
            ws_tab.cell(row=row, column=2 + gi, value=2 + gi if gi < 10 else 'Not Reported')
        ws_tab.cell(row=row, column=13, value=bal_label)
        ws_tab.cell(row=row, column=14, value='% of ' + bal_label)
        ws_tab.cell(row=row, column=16, value='Improved')
        ws_tab.cell(row=row, column=17, value=0)
        ws_tab.cell(row=row, column=18, value=0)
        row += 1
        rr_statuses = ['Deteriorated', 'Unchanged', 'Not Reported']
        for ri in range(12):
            if ri < 10:
                ws_tab.cell(row=row, column=1, value=2 + ri)
            elif ri == 10:
                ws_tab.cell(row=row, column=1, value='Not Reported')
            else:
                ws_tab.cell(row=row, column=1, value='Total')
            for gi in range(11):
                ws_tab.cell(row=row, column=2 + gi, value=0)
            ws_tab.cell(row=row, column=13, value=0)
            ws_tab.cell(row=row, column=14, value=0)
            if ri < len(rr_statuses):
                ws_tab.cell(row=row, column=16, value=rr_statuses[ri])
                ws_tab.cell(row=row, column=17, value=0)
                ws_tab.cell(row=row, column=18, value=0)
            row += 1
        row += 1  # blank separator

        # Grand Total block
        grand_gfracs = {}
        for gl in active_grades + ['Not Reported']:
            total_frac = 0.0
            total_weight = 0.0
            for p in POOLS:
                gf = _grade_fractions(p['fico_mean'], p['fico_std'], p['name'])
                total_frac += gf.get(gl, 0.0) * p['target_balance']
                total_weight += p['target_balance']
            grand_gfracs[gl] = total_frac / total_weight if total_weight > 0 else 0.0

        row = _write_migration_block(ws_tab, row, 'Grand Total ', 'CG', bal_label,
                                     pct_label, grand_status, grand_gfracs, grand_total)

        print('  {}: {} pools + Grand Total'.format(tab_name, len(POOLS)))

    wb.save(warm_path)
    print("  Patched WARM file with DQ/CO Data Entry + BS CO DQ Data Enter + Display CO-Recov -DQ tabs")
    return True


# ── Main ─────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  Creating Sample Credit Union")
    print("=" * 60)

    print("\n[1/7] Creating config YAML...")
    create_config()

    print("\n[2/7] Creating folder structure...")
    create_folders()

    print("\n[3/7] Generating loan data...")
    loan_df = generate_loan_data()

    print("\n[4/7] Inserting loan data into database...")
    insert_loan_data(loan_df)

    print("\n[5/7] Creating historical data files...")
    create_co_recovery_files()
    create_monthly_balance_file()

    print("\n[6/7] Generating TCT (WARM) report...")
    # Import and run the report generator
    from generate_report import generate_report
    tct_files = generate_report('sample', SNAPSHOT, reports=['tct'])
    if not tct_files:
        print("  ERROR: TCT report generation failed!")
        return

    # Copy WARM file to data directory so load_impaired_data finds it
    # (avoids picking up another client's WARM file from Reports/)
    warm_src = tct_files[0]
    warm_dst_name = f"{SNAP_PREFIX} CECL-Migration-WARM - {CU_NAME}.xlsx"
    warm_dst = os.path.join(DATA_DIR, SNAP_PREFIX, warm_dst_name)
    os.makedirs(os.path.join(DATA_DIR, SNAP_PREFIX), exist_ok=True)
    shutil.copy2(warm_src, warm_dst)
    print(f"  Copied WARM to {warm_dst}")

    print("\n[7/7] Patching WARM file and generating Vizo reports...")
    if not patch_warm_file():
        return

    # Now generate Vizo reports (these will read the patched WARM file)
    vizo_files = generate_report('sample', SNAPSHOT, reports=['vizo'])
    supp_files = generate_report('sample', SNAPSHOT, reports=['vizo_supp'])

    all_files = tct_files + vizo_files + supp_files
    print("\n" + "=" * 60)
    print(f"  DONE! Generated {len(all_files)} report(s):")
    for f in all_files:
        print(f"    {os.path.basename(f)}")
    print("=" * 60)


if __name__ == '__main__':
    main()
