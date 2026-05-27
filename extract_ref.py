"""Extract formatting specs from reference Excel file - outputs to text files."""
import os, warnings, sys, time, json
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

REF = os.path.join(os.environ['TEMP'], 'ref_report.xlsx')

print(f"Loading {REF}...", flush=True)
t0 = time.time()
wb = load_workbook(REF, data_only=True)
print(f"Loaded in {time.time()-t0:.0f}s  ({len(wb.sheetnames)} sheets)", flush=True)

# Only analyze the report/output sheets (first 48)
REPORT_SHEETS = wb.sheetnames[:48]

def safe_bg(fill):
    """Get background color string safely."""
    try:
        if fill and fill.fgColor:
            rgb = fill.fgColor.rgb
            if rgb and rgb not in ('00000000', '0'):
                return str(rgb)
    except:
        pass
    try:
        if fill and fill.patternType and fill.patternType != 'none':
            return f"pattern={fill.patternType}"
    except:
        pass
    return ""

def dump_sheet_compact(name, out, max_rows=100, max_cols=25):
    ws = wb[name]
    out.write(f"\n{'='*80}\n")
    out.write(f"SHEET: {name}\n")
    out.write(f"  Dims: {ws.max_row}r x {ws.max_column}c\n")
    out.write(f"  Orientation: {ws.page_setup.orientation}\n")
    out.write(f"  Paper: {ws.page_setup.paperSize}\n")
    out.write(f"  Charts: {len(ws._charts)}\n")
    out.write(f"  Merges: {len(ws.merged_cells.ranges)}\n")
    
    # Page margins
    pm = ws.page_margins
    if pm:
        out.write(f"  Margins: T={pm.top} B={pm.bottom} L={pm.left} R={pm.right} H={pm.header} F={pm.footer}\n")
    
    # Print setup
    ps = ws.page_setup
    if ps:
        out.write(f"  FitToWidth={ps.fitToWidth} FitToHeight={ps.fitToHeight}\n")
    
    # Column widths
    cw = {}
    for c in range(1, min(ws.max_column + 1, max_cols)):
        lt = get_column_letter(c)
        w = ws.column_dimensions[lt].width
        if w:
            cw[lt] = round(w, 2)
    out.write(f"  Col widths: {cw}\n")
    
    # Row heights
    rh = {}
    for r in range(1, min(ws.max_row + 1, max_rows + 1)):
        h = ws.row_dimensions[r].height
        if h and h != 15:
            rh[r] = h
    if rh:
        out.write(f"  Row heights: {rh}\n")
    
    # Merged ranges (all of them)
    if ws.merged_cells.ranges:
        merges = sorted([str(m) for m in ws.merged_cells.ranges])
        out.write(f"  Merges:\n")
        for m in merges:
            out.write(f"    {m}\n")
    
    # Cell content with formatting
    limit = min(ws.max_row, max_rows)
    climit = min(ws.max_column, max_cols)
    for r in range(1, limit + 1):
        for c in range(1, climit + 1):
            cell = ws.cell(r, c)
            if cell.value is not None:
                font = cell.font
                fill = cell.fill
                align = cell.alignment
                
                parts = [f"R{r}C{c}"]
                val = repr(cell.value)
                if len(val) > 80:
                    val = val[:80] + "..."
                parts.append(val)
                
                if font:
                    fb = "B" if font.bold else ""
                    fi = "I" if font.italic else ""
                    fu = "U" if font.underline else ""
                    parts.append(f"f={font.name}/{font.size}{fb}{fi}{fu}")
                    if font.color and font.color.rgb:
                        fc = str(font.color.rgb)
                        if fc not in ('00000000', '0'):
                            parts.append(f"fc={fc}")
                
                bg = safe_bg(fill)
                if bg:
                    parts.append(f"bg={bg}")
                
                if cell.number_format and cell.number_format != 'General':
                    parts.append(f"fmt={cell.number_format}")
                
                if align:
                    ha = align.horizontal or ""
                    va = align.vertical or ""
                    wrap = "W" if align.wrap_text else ""
                    if ha or va or wrap:
                        parts.append(f"align={ha}/{va}/{wrap}")
                
                out.write("  " + " | ".join(parts) + "\n")
    
    # Chart info
    for ci, chart in enumerate(ws._charts):
        out.write(f"  CHART {ci}: type={type(chart).__name__} title={chart.title} w={chart.width} h={chart.height}\n")

# Dump all report sheets
outpath = os.path.join(os.environ['TEMP'], 'ref_format_all.txt')
print(f"Dumping to {outpath}...", flush=True)
with open(outpath, 'w', encoding='utf-8') as f:
    f.write(f"Reference file: {REF}\n")
    f.write(f"Total sheets: {len(wb.sheetnames)}\n")
    f.write(f"Report sheets (first 48): {REPORT_SHEETS}\n\n")
    
    for i, sname in enumerate(REPORT_SHEETS):
        print(f"  [{i+1}/{len(REPORT_SHEETS)}] {sname}", flush=True)
        try:
            dump_sheet_compact(sname, f, max_rows=100, max_cols=25)
        except Exception as e:
            f.write(f"\n  ERROR on {sname}: {e}\n")

print(f"\nDone! Output at {outpath}", flush=True)
